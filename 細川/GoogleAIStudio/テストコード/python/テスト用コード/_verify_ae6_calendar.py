# -*- coding: utf-8 -*-
"""3月カレンダー AE6 の IsFactoryWorkingCell / 日付列 相当を検証（出勤簿.txt ロジック再現）"""
from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

def _planning_repo_root() -> Path:
    here = Path(__file__).resolve().parent
    parent = here.parent
    if (parent / "planning_core.py").is_file() or (parent / "planning_core" / "__init__.py").is_file():
        return parent
    return here


REPO = _planning_repo_root()
MASTER = REPO / "master.xlsm"
CALENDAR_LAST_DATA_COL = 33  # AG
AE_COL = 31  # AE


def _sheet_3月カレンダー(wb) -> str | None:
    for name in wb.sheetnames:
        if "カレンダー" in name and not name.startswith("結果_"):
            if name.startswith("3") or name[0] == "3":
                return name
    for name in wb.sheetnames:
        if re.match(r"^3\s*月", name) or name.startswith("3月"):
            if "カレンダー" in name:
                return name
    return None


def _rgb_from_cell_fill(cell) -> tuple[int | None, str]:
    """openpyxl の fill から (R,G,B) とメモを返す。取得できなければ (None, reason)"""
    f = cell.fill
    if f is None or not isinstance(f, PatternFill):
        return None, "fill=None or not PatternFill"
    pt = getattr(f, "patternType", None)
    if pt is None or str(pt) == "none":
        return None, "patternType none -> VBA xlPatternNone 相当（稼働日 True）"

    # fgColor / start_color
    for attr in ("fgColor", "start_color"):
        col = getattr(f, attr, None)
        if col is None:
            continue
        rgb = getattr(col, "rgb", None)
        if rgb and isinstance(rgb, str) and len(rgb) >= 6:
            s = rgb[-6:] if len(rgb) > 6 else rgb
            r = int(s[0:2], 16)
            g = int(s[2:4], 16)
            b = int(s[4:6], 16)
            return (r, g, b), f"{attr}.rgb={rgb}"
        if getattr(col, "type", None) == "theme":
            return None, f"{attr} theme={getattr(col, 'theme', None)} (RGB はブック依存のため要 Excel 実測)"
    return None, "RGB 未取得"


def is_factory_working_cell_rrggbb(rr: int, gg: int, bb: int) -> bool:
    """出勤簿.txt IsFactoryWorkingCell の黄判定部分（塗りあり・白以外のあと）"""
    if rr >= 235 and gg >= 220 and bb <= 190:
        return True
    return False


def is_factory_working_vba_style(
    pattern_none: bool, rgb255: bool, rr: int, gg: int, bb: int
) -> bool:
    if pattern_none:
        return True
    if rgb255:
        return True
    if is_factory_working_cell_rrggbb(rr, gg, bb):
        return True
    return False


def main() -> int:
    if not MASTER.exists():
        print(f"ファイルがありません: {MASTER}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(MASTER, data_only=False, read_only=True, keep_links=False)
    sheet_name = _sheet_3月カレンダー(wb)
    if not sheet_name:
        wb.close()
        print("3月カレンダー シートが見つかりません", file=sys.stderr)
        return 1

    ws = wb[sheet_name]
    c4 = ws.cell(row=4, column=AE_COL)
    c6 = ws.cell(row=6, column=AE_COL)
    a1 = ws.cell(row=1, column=1)

    out: list[str] = []
    out.append(f"ブック: {MASTER.name}")
    out.append(f"シート: {sheet_name}")
    out.append(f"A1: {a1.value!r}")
    out.append(f"列 AE = {AE_COL} (有効上限 AG={CALENDAR_LAST_DATA_COL} 内: {AE_COL <= CALENDAR_LAST_DATA_COL})")
    out.append("")
    out.append("--- 4行目 AE（日付行） ---")
    out.append(f"値: {c4.value!r}")
    out.append("")
    out.append("--- 6行目 AE（メンバー行2想定） ---")
    out.append(f"値: {c6.value!r}")

    f = c6.fill
    pat = getattr(f, "patternType", None) if f else None
    pattern_none = f is None or pat is None or str(pat) == "none"
    out.append(f"fill.patternType: {pat!r} -> xlPatternNone 相当: {pattern_none}")

    rr, gg, bb = -1, -1, -1
    got_rgb = False
    tup, memo = _rgb_from_cell_fill(c6)
    if isinstance(tup, tuple) and len(tup) == 3:
        rr, gg, bb = tup
        got_rgb = True
        out.append(f"RGB 推定: R={rr} G={gg} B={bb} ({memo})")
    else:
        out.append(f"RGB: {memo}")

    if got_rgb:
        rgb255 = rr == 255 and gg == 255 and bb == 255
        out.append(f"RGB(255,255,255) 一致: {rgb255}")
        out.append(f"黄系条件 R>=235 & G>=220 & B<=190: {rr >= 235 and gg >= 220 and bb <= 190}")
        working = is_factory_working_vba_style(pattern_none, rgb255, rr, gg, bb)
    else:
        working = pattern_none
        out.append("※ RGB 未取得のため、patternType のみで近似（無塗り=True、それ以外は不明）")

    out.append("")
    out.append("=== 結論（出勤簿.txt と同じ分岐） ===")
    if pattern_none:
        out.append("IsFactoryWorkingCell -> True（塗りつぶし無し）")
    elif got_rgb:
        if rr == 255 and gg == 255 and bb == 255:
            out.append("IsFactoryWorkingCell -> True（白）")
        elif is_factory_working_cell_rrggbb(rr, gg, bb):
            out.append("IsFactoryWorkingCell -> True（黄系）")
        else:
            out.append("IsFactoryWorkingCell -> False（上記以外の色）→ 出勤簿は note='休'")
    else:
        out.append("openpyxl では塗り RGB を確定できず。Excel 上で Interior.Color を確認してください。")

    out.append(f"合成 isWorkingDay（上記ロジック）: {working}")

    wb.close()

    report = "\n".join(out)
    print(report)
    outp = REPO / "output" / "verify_ae6_calendar.txt"
    outp.parent.mkdir(parents=True, exist_ok=True)
    outp.write_text(report, encoding="utf-8")
    print(f"\n保存: {outp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
