"""
planning_core — 工場向け「配台計画」シミュレーションの中核（単一モジュール）

【このファイルの役割】
    VBA / マクロ実行ブックから環境変数で渡されたパスを読み、加工計画データ・勤怠・
    スキル need を統合して、設備・担当者への割付結果（Excel）を生成する。

【外部から直接使う入口（他 .py から import される想定）】
    - ``TASKS_INPUT_WORKBOOK`` … マクロブックのパス（環境変数 TASK_INPUT_WORKBOOK）
    - ``run_stage1_extract()`` … 段階1。`task_extract_stage1.py` から呼ぶ
    - ``generate_plan()`` … 段階2。`plan_simulation_stage2.py` から呼ぶ
    - ``apply_result_task_column_layout_only()`` … 取り込み済み「結果_タスク一覧」の列順・非表示を
      「列設定_結果_タスク一覧」に合わせる（`apply_result_task_column_layout.py` / VBA ボタン用・xlwings）

【処理の流れ（ざっくり）】
    1. 段階1: 「加工計画DATA」→ 中間 `output/plan_input_tasks.xlsx`、
       マクロブック内「設定_配台不要工程」の行同期・D列→E列(Gemini)・保存（既定は xlwings）
    2. 段階2: master の skills / 勤怠 / 配台計画シートを読み、日付ループで割付、
       `output/` 直下に `production_plan_multi_day_*.xlsx` と `member_schedule_*.xlsx` を出力（常に最新1組のみ・実行前に同名パターンを全削除）

【ソース上の構成（=#= 見出しでスクロール検索可能）】
    - 先頭 … ログ・パス・レガシーファイル掃除
    - 「【設定】APIキー / 基本ルール / ファイル名」… 列名定数・パス
    - 配台計画シート列・参照列ヘルパ
    - 結果ガント・タスク一覧の整形
    - 実績 DATA・特別指定備考・Gemini 連携
    - 配台不要（分割自動・設定シート・openpyxl 保存＋ xlwings 同期フォールバック）
    - ``run_stage1_extract`` … 段階1本体
    - 「1. コア計算ロジック」… 時刻・休憩を踏んだ実働分計算
    - 「2. マスタデータ・出勤簿 と AI解析」… skills/need/勤怠
    - ``generate_plan`` … 段階2本体（メインループ）

【命名】
    - 先頭 ``_`` … モジュール内専用ヘルパ（外部から呼ばない想定）
    - ``PLAN_*`` / ``TASK_*`` … Excel 見出しと一致させる定数

【依存】
    pandas, openpyxl, xlwings（Excel 起動中の保存・設定シート同期）, google.genai, cryptography（暗号化認証 JSON 利用時）
    API キーはマクロブック「設定」B1 の JSON パスからのみ（環境変数 GEMINI_API_KEY は使わない）。
    暗号化 JSON の復号は ``planning_core`` 内の定数のみ使用（値はソースにあり、社内共有資料やログには書かないこと）。ログにキーは出さない。

【初回環境】 テストコード直下で ``py -3 python/setup_environment.py`` またはマクロ「環境構築」（requirements.txt 一括）
"""

import pandas as pd
from datetime import datetime, timedelta, time, date
from collections import Counter, defaultdict
import itertools
import csv
import json
import copy
import re
import traceback
import base64
import hashlib
import unicodedata
import time as time_module
from google import genai
import logging
import calendar
import math
import os
import fnmatch
import shutil
import sys
import ctypes
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================
# 【重要】カレントを「テストコード直下」に設定（master.xlsm・output・log・json・API_Payment と同じ階層）
# planning_core.py が python\ 配下にあるときは親フォルダへ上がる。単体配置のときは自フォルダ。
# （VBA は …\output\plan_input_tasks.xlsx を参照するため output は直下必須）
# =========================================================
_planning_core_dir = os.path.dirname(os.path.abspath(__file__))
if os.path.basename(_planning_core_dir).lower() == "python":
    os.chdir(os.path.dirname(_planning_core_dir))
else:
    os.chdir(_planning_core_dir)

# cmd で chcp 65001 時にログの日本語をコンソールへ出しやすくする（段階2でリダイレクト無し実行時向け）
if os.name == "nt" and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ---------------------------------------------------------
# コマンドプロンプト画面を「最前面（トップレベル）」にする処理
# ---------------------------------------------------------
if os.name == 'nt':
    hwnd = ctypes.windll.kernel32.GetConsoleWindow()
    if hwnd:
        # HWND_TOPMOST = -1, SWP_NOMOVE = 0x0002, SWP_NOSIZE = 0x0001
        ctypes.windll.user32.SetWindowPos(hwnd, -1, 0, 0, 0, 0, 3)

# ---------------------------------------------------------
# ログを「画面（コンソール）」と「一時ファイル」の両方に出力する
# ---------------------------------------------------------
logger = logging.getLogger()
logger.setLevel(logging.INFO)
# 既存のハンドラをクリア
if logger.hasHandlers():
    logger.handlers.clear()

formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# 1. 画面(コンソール)用ハンドラ
stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)

# 2. 成果物用 output / 実行ログ用 log（いずれも常に最新1ファイルのみ上書き）
output_dir = os.path.join(os.getcwd(), 'output')
os.makedirs(output_dir, exist_ok=True)
log_dir = os.path.join(os.getcwd(), 'log')
os.makedirs(log_dir, exist_ok=True)
# 段階2がユーザー検証で中断したときの1行メッセージ（VBA が log から読み MsgBox 用）
STAGE2_BLOCKING_MESSAGE_FILE = "stage2_blocking_message.txt"
# 計画基準納期当日の完了目安時刻（業務ルール）。順次配台・超過再試行ロジックへの組み込みは段階的に拡張可能。
PLAN_DUE_DAY_COMPLETION_TIME = time(16, 0)
# Gemini 利用・推定料金の累計 JSON（log ではなく専用フォルダ）
api_payment_dir = os.path.join(os.getcwd(), 'API_Payment')
os.makedirs(api_payment_dir, exist_ok=True)
# 上書き保存するアプリ用 JSON（API 料金累計は API_Payment のみ）
json_data_dir = os.path.join(os.getcwd(), 'json')
os.makedirs(json_data_dir, exist_ok=True)
# 旧仕様: ログが output 直下にあった名残を削除（常に log 側の1ファイルのみが最新）
for _legacy_name in (
    "execution_log.txt",
    "ai_task_special_remark_last.txt",
    "ai_task_special_last_prompt.txt",
    "planning_conflict_highlight.tsv",
    "cmd_stage2.log",
):
    try:
        os.remove(os.path.join(output_dir, _legacy_name))
    except OSError:
        pass


def _remove_prior_stage2_workbooks_and_prune_empty_dirs(output_root: str) -> None:
    """
    ``production_plan_multi_day_*.xlsx`` / ``member_schedule_*.xlsx`` を output 配下からすべて削除し、
    空になったサブフォルダを削除する（日付階層の旧出力を含む）。
    段階2の直前に呼び、常に最新1組の成果物だけを残す土台にする。
    """
    if not output_root or not os.path.isdir(output_root):
        return
    patterns = ("production_plan_multi_day_*.xlsx", "member_schedule_*.xlsx")
    root_abs = os.path.normcase(os.path.abspath(output_root))
    removed = 0
    for dirpath, _dirnames, filenames in os.walk(output_root, topdown=False):
        for name in filenames:
            if name.startswith("~$"):
                continue
            for pat in patterns:
                if fnmatch.fnmatch(name, pat):
                    fp = os.path.join(dirpath, name)
                    try:
                        os.remove(fp)
                        removed += 1
                    except OSError as ex:
                        logging.warning("段階2旧出力の削除に失敗: %s (%s)", fp, ex)
                    break
        dir_abs = os.path.normcase(os.path.abspath(dirpath))
        if dir_abs == root_abs:
            continue
        try:
            if not os.listdir(dirpath):
                os.rmdir(dirpath)
        except OSError:
            pass
    if removed:
        logging.info(
            "段階2出力の整理: production_plan_multi_day_*.xlsx / member_schedule_*.xlsx を %s 件削除しました。",
            removed,
        )


# 3. ファイル用ハンドラ（VBAで後から読み取るため UTF-8 で保存）
log_file_path = os.path.join(log_dir, 'execution_log.txt')
# BOM 付き UTF-8（Excel / VBA の ADODB.Stream が文字化けしにくい）
file_handler = logging.FileHandler(log_file_path, encoding='utf-8-sig')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

stage2_blocking_message_path = os.path.join(log_dir, STAGE2_BLOCKING_MESSAGE_FILE)


class PlanningValidationError(Exception):
    """配台計画シートの検証エラー（段階2を続行しない）。メッセージは log にも書く。"""


def _clear_stage2_blocking_message_file() -> None:
    try:
        if os.path.isfile(stage2_blocking_message_path):
            os.remove(stage2_blocking_message_path)
    except OSError:
        pass


def _write_stage2_blocking_message(message: str) -> None:
    s = (message or "").strip()
    if not s:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(stage2_blocking_message_path, "w", encoding="utf-8", newline="\n") as f:
            f.write(s)
    except OSError as ex:
        logging.warning("stage2_blocking_message の書き込みに失敗: %s", ex)

# AI 備考・配台不能ロジック D→E の TTL キャッシュ（旧 output/ から json/ へ移行）
_ai_remarks_cache_name = "ai_remarks_cache.json"
_ai_cache_legacy = os.path.join(output_dir, _ai_remarks_cache_name)
_ai_cache_new = os.path.join(json_data_dir, _ai_remarks_cache_name)
if os.path.isfile(_ai_cache_legacy) and not os.path.isfile(_ai_cache_new):
    try:
        shutil.move(_ai_cache_legacy, _ai_cache_new)
    except OSError:
        pass
ai_cache_path = _ai_cache_new
# 「設定_配台不要工程」シート作成・保存の成否デバッグ（execution_log と併用）
exclude_rules_sheet_debug_log_path = os.path.join(log_dir, "exclude_rules_sheet_debug.txt")
# 保存失敗時に E 列（ロジック式）だけを退避し、次回 run_exclude_rules_sheet_maintenance で自動適用する（json フォルダ）
EXCLUDE_RULES_E_SIDECAR_FILENAME = "exclude_rules_e_column_pending.json"
# openpyxl 保存失敗時に VBA が E 列へ書き込むための UTF-8 TSV（Base64）。
EXCLUDE_RULES_E_VBA_TSV_FILENAME = "exclude_rules_e_column_vba.tsv"
# openpyxl 保存失敗時に VBA が A〜E を一括反映する UTF-8 TSV（行ごとに 5 セル分 Base64）。
EXCLUDE_RULES_MATRIX_VBA_FILENAME = "exclude_rules_matrix_vba.tsv"
# VBA がメイン P 列へ書き込むための UTF-8 テキスト（Excel 開いたまま save できない問題の回避）
GEMINI_USAGE_SUMMARY_FOR_MAIN_FILE = "gemini_usage_summary_for_main.txt"
# 全実行を通した Gemini 利用・推定料金の累計（API 応答ごとに更新。保存先は API_Payment フォルダ）
GEMINI_USAGE_CUMULATIVE_JSON_FILE = "gemini_usage_cumulative.json"
# 期間別バケットをフラット化した CSV（Excel の折れ線・棒グラフ用）
GEMINI_USAGE_BUCKETS_CSV_FILE = "gemini_usage_buckets_for_chart.csv"
# テスト: EXCLUDE_RULES_TEST_E1234=1 で EXCLUDE_RULES_SHEET_NAME（「設定_配台不要工程」）の E 列に "1234" を書く（保存経路の確認用）。
# TASK_INPUT_WORKBOOK は「加工計画DATA」シート付きブック（例: 生産管理_AI配台テスト.xlsm）を指定すること。
# 行は EXCLUDE_RULES_TEST_E1234_ROW（既定 9、2 未満は 9 に丸める）。

# =========================================================
# 【設定】APIキー / 基本ルール / ファイル名
# =========================================================
# Gemini API キーは TASK_INPUT_WORKBOOK 確定後、下記「設定」B1 の JSON から解決（平文または format_version 2 の暗号化）。
# 未設定時のみ移行用に環境変数 GEMINI_API_KEY を参照。

GEMINI_MODEL_FLASH = "gemini-2.5-flash"
# 推定料金: USD / 1M tokens（入力, 出力）。公式の最新単価に合わせて更新すること。
# 環境変数 GEMINI_PRICE_USD_IN_PER_M / GEMINI_PRICE_USD_OUT_PER_M で上書き可（Flash 向け）。
_GEMINI_FLASH_IN_PER_M = float(
    os.environ.get("GEMINI_PRICE_USD_IN_PER_M", "0.075") or 0.075
)
_GEMINI_FLASH_OUT_PER_M = float(
    os.environ.get("GEMINI_PRICE_USD_OUT_PER_M", "0.30") or 0.30
)
GEMINI_JPY_PER_USD = float(os.environ.get("GEMINI_JPY_PER_USD", "150") or 150)

# ---------------------------------------------------------------------------
# 以降の定数ブロックは「Excel 列見出し」と 1:1 で対応させる。
# 列名を変える場合は VBA・マクロ側シートと同時に直すこと。
# ---------------------------------------------------------------------------

MASTER_FILE = "master.xlsm" # skillsとattendance(およびtasks)を統合したファイル
# VBA「master_組み合わせ表を更新」で作るシート（工程+機械キーとメンバー編成）
MASTER_SHEET_TEAM_COMBINATIONS = "組み合わせ表"
# メンバー別勤怠シート: master.xlsm では「休暇区分」と「備考」が別列。
# 勤怠AIの入力は備考のみ。ただし reason（表示・中抜け補正・個人シートの休憩/休暇文言）は、備考が空のとき休暇区分を引き継ぐ。
# master カレンダー／出勤簿.txt 準拠: 前休=午前年休・12:45～17:00（午後休憩14:45～15:00）／後休=8:45～12:00・午後年休／国=他拠点勤務。
# 備考列・休暇区分は勤怠 AI で構造化（配台不参加・is_holiday・中抜け等）。備考が空でも休暇区分のみの行は AI に渡す。
ATT_COL_LEAVE_TYPE = "休暇区分"
ATT_COL_REMARK = "備考"
# 勤怠備考 AI の JSON スキーマを変えたら更新し、キャッシュキーを無効化する
ATTENDANCE_REMARK_AI_SCHEMA_ID = "v2_haitai_fuka"
# need シート: 「基本必要人数」行（A列に「必要人数」を含む）＋ その直下の「配台時追加人数／余力時追加人数」等
# （Excel 上は概ね 5 行目付近。余剰時に増やせる人数上限・工程×機械列）
# ＋ 行「特別指定1」～「特別指定99」（必要人数の上書き・1～99）
NEED_COL_CONDITION = "依頼NO条件"
NEED_COL_NOTE = "備考"
# need「配台時追加人数」を満枠使っても、単位あたり加工時間が短くなるのは最大でこの割合（例: 0.05 ≒ 5%）
SURPLUS_TEAM_MAX_SPEEDUP_RATIO = 0.05
# タスクは tasks.xlsx を使わず、VBA から渡す TASK_INPUT_WORKBOOK の「加工計画DATA」のみ
TASKS_INPUT_WORKBOOK = os.environ.get("TASK_INPUT_WORKBOOK", "").strip()
TASKS_SHEET_NAME = "加工計画DATA"

# このシート名を含むブックは openpyxl が読み書きに失敗することがあるため、load_workbook を試行しない
OPENPYXL_INCOMPATIBLE_SHEET_MARKER = "配台_配台不要工程"


def _ooxml_workbook_sheet_names(wb_path: str) -> list[str] | None:
    """ZIP 内 xl/workbook.xml からシート名一覧を取る（openpyxl を使わない）。"""
    import zipfile
    import xml.etree.ElementTree as ET

    if not wb_path or not os.path.isfile(wb_path):
        return None
    low = wb_path.lower()
    if not low.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return None
    try:
        with zipfile.ZipFile(wb_path, "r") as zf:
            if "xl/workbook.xml" not in zf.namelist():
                return None
            raw = zf.read("xl/workbook.xml")
    except (OSError, zipfile.BadZipFile, KeyError):
        return None
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return None
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    tag_sheet = "{%s}sheet" % ns_main
    names: list[str] = []
    for el in root.iter():
        if el.tag == tag_sheet or el.tag.endswith("}sheet"):
            n = el.get("name")
            if n:
                names.append(n)
    return names


def _workbook_should_skip_openpyxl_io(wb_path: str) -> bool:
    """当該パスが OOXML でシート「配台_配台不要工程」を含むとき True（openpyxl 利用を避ける）。"""
    p = (wb_path or "").strip()
    if not p:
        return False
    names = _ooxml_workbook_sheet_names(p)
    if not names:
        return False
    return OPENPYXL_INCOMPATIBLE_SHEET_MARKER in names


# マクロブック「設定」B1: 社内共有上の Gemini 認証 JSON のパス
APP_CONFIG_SHEET_NAME = "設定"
# 暗号化認証 JSON（format_version 2）の復号は常にこの定数のみ（社内手順のパスフレーズと一致させる。ログ・UI に出さない）。
_GEMINI_CREDENTIALS_PASSPHRASE_FIXED = "nagaoka1234"
_GEMINI_CREDENTIALS_PBKDF2_ITERATIONS_DEFAULT = 480_000


def _config_cell_text(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def _resolve_path_relative_to_workbook(wb_path: str, user_path: str) -> str:
    p = (user_path or "").strip().strip('"')
    if not p:
        return ""
    if os.path.isabs(p):
        return os.path.normpath(p)
    base = os.path.dirname(os.path.abspath(wb_path))
    return os.path.normpath(os.path.join(base, p))


def _read_gemini_credentials_json_path_from_workbook(wb_path: str) -> str | None:
    """「設定」シート B1 から認証 JSON ファイルパスを読む。無ければ None。"""
    if not wb_path or not os.path.isfile(wb_path):
        return None
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.debug(
            "Gemini: ブックに「%s」があるため openpyxl で「%s」!B1 を読みません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            APP_CONFIG_SHEET_NAME,
        )
        return None
    try:
        keep_vba = str(wb_path).lower().endswith(".xlsm")
        wb = load_workbook(
            wb_path, read_only=True, data_only=True, keep_vba=keep_vba
        )
        try:
            if APP_CONFIG_SHEET_NAME not in wb.sheetnames:
                return None
            ws = wb[APP_CONFIG_SHEET_NAME]
            b1 = _config_cell_text(ws.cell(row=1, column=2).value)
        finally:
            wb.close()
    except Exception as ex:
        logging.debug(
            "Gemini: マクロブック「%s」の「%s」!B1 を読めません: %s",
            wb_path,
            APP_CONFIG_SHEET_NAME,
            ex,
        )
        return None
    if not b1:
        return None
    return _resolve_path_relative_to_workbook(wb_path, b1) or None


def _read_output_book_font_prefs_from_workbook(wb_path: str) -> tuple[str | None, int | None]:
    """
    「設定」B4=配台結果 xlsx に埋め込むフォント名。空なら書体名・既定サイズをセルに埋め込まない
    （マクロ「全シートフォント」や取り込み先ブックの既定書体を維持しやすい）。
    B5=ポイント（B4 に名前があるときのみ有効、空なら 11）。
    マクロブックが無い・読めないときは従来互換で BIZ UDPゴシック 11。
    """
    fallback_name = "BIZ UDPゴシック"
    fallback_size = 11
    if not wb_path or not os.path.isfile(wb_path):
        return fallback_name, fallback_size
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.debug(
            "結果ブック用フォント: ブックに「%s」があるため openpyxl で設定シートを読みません（既定フォント）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return fallback_name, fallback_size
    try:
        keep_vba = str(wb_path).lower().endswith(".xlsm")
        wb = load_workbook(wb_path, read_only=True, data_only=True, keep_vba=keep_vba)
        try:
            if APP_CONFIG_SHEET_NAME not in wb.sheetnames:
                return fallback_name, fallback_size
            ws = wb[APP_CONFIG_SHEET_NAME]
            b4 = _config_cell_text(ws.cell(row=4, column=2).value)
            b5_raw = ws.cell(row=5, column=2).value
        finally:
            wb.close()
    except Exception as ex:
        logging.debug(
            "結果ブック用フォント: マクロブック「%s」の「%s」!B4/B5 を読めません: %s",
            wb_path,
            APP_CONFIG_SHEET_NAME,
            ex,
        )
        return fallback_name, fallback_size

    if not b4:
        return None, None

    sz = fallback_size
    if b5_raw is not None and str(b5_raw).strip() != "":
        try:
            sz = int(float(b5_raw))
        except (TypeError, ValueError):
            pass
    return b4, sz


def _read_task_ids_from_config_sheet_column(
    wb_path: str,
    column_index: int,
    log_label: str,
    column_letter_desc: str,
    *,
    openpyxl_skip_hint: str | None = None,
) -> list[str]:
    """
    マクロブック「設定」シートの指定列（1=A, 2=B）3 行目以降から依頼NOを読む。
    空セルはスキップ。連続 30 セル空で打ち切り。最大 500 行。カンマ区切りで複数可。
    """
    out: list[str] = []
    if not wb_path or not os.path.isfile(wb_path):
        return out
    if _workbook_should_skip_openpyxl_io(wb_path):
        msg = (
            f"{log_label}: ブックに「{OPENPYXL_INCOMPATIBLE_SHEET_MARKER}」があるため"
            f"「{APP_CONFIG_SHEET_NAME}」!{column_letter_desc}3 以降は openpyxl で読めません。"
        )
        if openpyxl_skip_hint:
            msg += " " + openpyxl_skip_hint.strip()
        logging.info(msg)
        return out
    try:
        keep_vba = str(wb_path).lower().endswith(".xlsm")
        wb = load_workbook(
            wb_path, read_only=True, data_only=True, keep_vba=keep_vba
        )
        try:
            if APP_CONFIG_SHEET_NAME not in wb.sheetnames:
                return out
            ws = wb[APP_CONFIG_SHEET_NAME]
            consecutive_empty = 0
            for r in range(3, 3 + 500):
                t = _config_cell_text(ws.cell(row=r, column=column_index).value)
                if not t:
                    consecutive_empty += 1
                    if consecutive_empty >= 30:
                        break
                    continue
                consecutive_empty = 0
                if "," in t:
                    for part in t.split(","):
                        p = part.strip()
                        if p:
                            out.append(p)
                else:
                    out.append(t)
        finally:
            wb.close()
    except Exception as ex:
        logging.warning(
            "%s: 「%s」!%s3 以降の依頼NOを読めません（無視）: %s",
            log_label,
            APP_CONFIG_SHEET_NAME,
            column_letter_desc,
            ex,
        )
        return []
    return out


def _read_trace_schedule_task_ids_from_config_sheet(wb_path: str) -> list[str]:
    """
    マクロブック「設定」シート A 列の 3 行目以降を、配台トレース対象の依頼NOとして読む。
    空セルはスキップ。連続 30 セル空なら打ち切り。最大 500 行まで走査。
    """
    return _read_task_ids_from_config_sheet_column(
        wb_path,
        1,
        "配台トレース",
        "A",
        openpyxl_skip_hint="配台トレースは「設定」シート A 列を openpyxl で読めないため無効です。",
    )


def _read_dispatch_debug_only_task_ids_from_config_sheet_b(wb_path: str) -> list[str]:
    """
    マクロブック「設定」シート B 列 3 行目以降に依頼NOがある場合、段階2はその依頼NOの行だけ配台する。
    （空なら全件。トレース用 A 列とは独立。）
    """
    return _read_task_ids_from_config_sheet_column(
        wb_path,
        2,
        "デバッグ配台",
        "B",
        openpyxl_skip_hint="限定配台は環境変数 DISPATCH_DEBUG_ONLY_TASK_IDS（カンマ区切り）でも指定できます。",
    )


def _extract_gemini_api_key_from_plain_dict(data: dict, json_path: str) -> str | None:
    key = data.get("gemini_api_key")
    if key is None or (isinstance(key, str) and not key.strip()):
        key = data.get("GEMINI_API_KEY")
    if key is None:
        logging.warning(
            "Gemini: 認証データに gemini_api_key（または GEMINI_API_KEY）がありません（%s）。",
            json_path,
        )
        return None
    s = str(key).strip()
    return s or None


def _derive_fernet_key_from_passphrase(
    passphrase: str, salt: bytes, iterations: int
) -> bytes:
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives import hashes

    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=iterations,
        backend=default_backend(),
    )
    return base64.urlsafe_b64encode(kdf.derive(passphrase.encode("utf-8")))


def _credentials_json_is_encrypted_v2(data: dict) -> bool:
    if data.get("format_version") == 2:
        return True
    return bool(
        data.get("kdf") == "pbkdf2_sha256" and (data.get("fernet_ciphertext") or "").strip()
    )


def _decrypt_gemini_credentials_v2(
    data: dict, passphrase: str, json_path: str
) -> dict | None:
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        logging.warning(
            "Gemini: 暗号化認証 JSON には cryptography が必要です（pip install cryptography）。"
        )
        return None
    token_s = (data.get("fernet_ciphertext") or "").strip()
    if not token_s:
        logging.warning(
            "Gemini: 暗号化認証 JSON に fernet_ciphertext がありません（%s）。",
            json_path,
        )
        return None
    salt_b64 = (data.get("salt_b64") or "").strip()
    if not salt_b64:
        logging.warning(
            "Gemini: 暗号化認証 JSON に salt_b64 がありません（%s）。",
            json_path,
        )
        return None
    try:
        salt = base64.standard_b64decode(salt_b64)
    except Exception as ex:
        logging.warning("Gemini: salt_b64 の解釈に失敗しました（%s）: %s", json_path, ex)
        return None
    iterations = int(data.get("iterations") or _GEMINI_CREDENTIALS_PBKDF2_ITERATIONS_DEFAULT)
    kdf_name = (data.get("kdf") or "pbkdf2_sha256").strip()
    if kdf_name != "pbkdf2_sha256":
        logging.warning("Gemini: 未対応の kdf（%s）: %s", kdf_name, json_path)
        return None
    try:
        fkey = _derive_fernet_key_from_passphrase(passphrase, salt, iterations)
        plain = Fernet(fkey).decrypt(token_s.encode("ascii"))
    except Exception:
        logging.debug("Gemini: 暗号化認証の復号処理に失敗しました（%s）。", json_path)
        return None
    try:
        inner = json.loads(plain.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as ex:
        logging.warning(
            "Gemini: 復号後の JSON が不正です（%s）: %s",
            json_path,
            ex,
        )
        return None
    if not isinstance(inner, dict):
        logging.warning("Gemini: 復号後の JSON はオブジェクトである必要があります（%s）。", json_path)
        return None
    return inner


def _load_gemini_api_key_from_credentials_json(
    json_path: str, workbook_path: str | None = None
) -> tuple[str | None, bool]:
    """戻り値: (api_key または None, 暗号化形式だったか)。暗号化時は _GEMINI_CREDENTIALS_PASSPHRASE_FIXED のみで復号。"""
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
    except OSError as ex:
        logging.warning("Gemini: 認証 JSON を開けません: %s (%s)", json_path, ex)
        return None, False
    except json.JSONDecodeError as ex:
        logging.warning("Gemini: 認証 JSON の形式が不正です: %s (%s)", json_path, ex)
        return None, False
    if not isinstance(data, dict):
        logging.warning("Gemini: 認証 JSON はオブジェクト形式である必要があります: %s", json_path)
        return None, False
    if _credentials_json_is_encrypted_v2(data):
        inner = _decrypt_gemini_credentials_v2(
            data, _GEMINI_CREDENTIALS_PASSPHRASE_FIXED, json_path
        )
        if inner is None:
            return None, True
        return _extract_gemini_api_key_from_plain_dict(inner, json_path), True
    return _extract_gemini_api_key_from_plain_dict(data, json_path), False


API_KEY = None
_cred_path = _read_gemini_credentials_json_path_from_workbook(TASKS_INPUT_WORKBOOK)
_used_encrypted_credentials = False
if _cred_path and os.path.isfile(_cred_path):
    API_KEY, _used_encrypted_credentials = _load_gemini_api_key_from_credentials_json(
        _cred_path, workbook_path=TASKS_INPUT_WORKBOOK
    )
    if API_KEY:
        if _used_encrypted_credentials:
            logging.info("Gemini API キー: 暗号化認証 JSON から読み込みました。")
        else:
            logging.info(
                "Gemini API キー: マクロブック「%s」B1 のパスから読み込みました。",
                APP_CONFIG_SHEET_NAME,
            )
elif _cred_path:
    logging.warning(
        "Gemini: 「%s」B1 で指定された認証 JSON が見つかりません。",
        APP_CONFIG_SHEET_NAME,
    )

# B1 が暗号化 JSON なのにキーが取れない（平文 JSON でキー欠落との区別）。原因の特定はログに書かず汎用メッセージのみ。
_encrypted_json_missing_key = (
    bool(_cred_path)
    and os.path.isfile(_cred_path)
    and _used_encrypted_credentials
    and not API_KEY
)
if _encrypted_json_missing_key:
    logging.error(
        "Gemini: 「%s」B1 の認証ファイルから API キーを利用できません。"
        " 社内手順に従い認証を再設定するか、管理者に問い合わせてください。",
        APP_CONFIG_SHEET_NAME,
    )

if not API_KEY:
    logging.warning(
        "Gemini API キーが未設定です。マクロブックに「%s」シートを用意し B1 に認証 JSON のフルパスを書いてください。"
        " 備考の AI 解析等はスキップされます。"
        " ひな型: gemini_credentials.example.json / encrypt_gemini_credentials.py（暗号化）。",
        APP_CONFIG_SHEET_NAME,
    )

# 計画結果 xlsx のセルフォント（「設定」B4/B5。B4 空=書体名を openpyxl で付けず取り込み先に任せる）
OUTPUT_BOOK_FONT_NAME, OUTPUT_BOOK_FONT_SIZE = _read_output_book_font_prefs_from_workbook(
    TASKS_INPUT_WORKBOOK
)
if OUTPUT_BOOK_FONT_NAME is None:
    logging.info(
        "結果ブック: 「%s」B4 が空のため、セルに書体名を埋め込みません（全シートフォント統一後の取り込みで上書きされにくくなります）。"
        " Python 出力に明示フォントを付けたい場合は B4 にフォント名、必要なら B5 にポイントを入力してください。",
        APP_CONFIG_SHEET_NAME,
    )
RESULT_SHEET_GANTT_NAME = "結果_設備ガント"

# タスク列名（マクロ実行ブック「加工計画DATA」）
TASK_COL_TASK_ID = "依頼NO"
TASK_COL_MACHINE = "工程名"
TASK_COL_MACHINE_NAME = "機械名"
TASK_COL_QTY = "換算数量"
TASK_COL_ORDER_QTY = "受注数"
TASK_COL_SPEED = "加工速度"
TASK_COL_PRODUCT = "製品名"
TASK_COL_ANSWER_DUE = "回答納期"
TASK_COL_SPECIFIED_DUE = "指定納期"
TASK_COL_RAW_INPUT_DATE = "原反投入日"
# 同一依頼NOの工程順（カンマ区切りの工程名）。加工計画DATA／配台計画_タスク入力。
TASK_COL_PROCESS_CONTENT = "加工内容"
# 投入可能日の目安は「回答納期」、未入力時は「指定納期」（前日基準・当日/遅れは最優先）。「加工開始日」列は参照しない。
# 完了判定・進捗（加工計画DATA）
TASK_COL_COMPLETION_FLAG = "加工完了区分"
TASK_COL_ACTUAL_DONE = "実加工数"   # 旧互換（直接の加工済数量）
TASK_COL_ACTUAL_OUTPUT = "実出来高"  # 完成品数量（換算に使う）
TASK_COL_DATA_EXTRACTION_DT = "データ抽出日"
AI_CACHE_TTL_SECONDS = 6 * 60 * 60  # 6時間
# json/ai_remarks_cache.json 内のキー接頭辞（設定_配台不要工程・配台不能ロジック D→E）
AI_CACHE_KEY_PREFIX_EXCLUDE_RULE_DE = "exclude_rule_de_v1"

# マクロブック「加工実績DATA」（Power Query 等で取り込み想定）
ACTUALS_SHEET_NAME = "加工実績DATA"
ACT_COL_TASK_ID = "依頼NO"
ACT_COL_PROCESS = "工程名"
ACT_COL_OPERATOR = "担当者"
ACT_COL_START_DT = "開始日時"
ACT_COL_END_DT = "終了日時"
ACT_COL_START_ALT = "実績開始"
ACT_COL_END_ALT = "実績終了"
ACT_COL_DAY = "日付"
ACT_COL_TIME_START = "開始時刻"
ACT_COL_TIME_END = "終了時刻"
ACTUAL_HEADER_CANONICAL = (
    ACT_COL_TASK_ID,
    ACT_COL_PROCESS,
    ACT_COL_OPERATOR,
    ACT_COL_START_DT,
    ACT_COL_END_DT,
    ACT_COL_START_ALT,
    ACT_COL_END_ALT,
    ACT_COL_DAY,
    ACT_COL_TIME_START,
    ACT_COL_TIME_END,
)

# --- 2段階処理: 段階1抽出 → ブック「配台計画_タスク入力」編集 → 段階2計画 ---
STAGE1_OUTPUT_FILENAME = "plan_input_tasks.xlsx"
PLAN_INPUT_SHEET_NAME = os.environ.get("TASK_PLAN_SHEET", "").strip() or "配台計画_タスク入力"
PLAN_COL_REQUIRED_OP = "必要人数"
# 上書き列「必要OP人数」の左隣の参照列見出し（「（元）必要OP人数」から改称）
PLAN_REF_REQUIRED_OP = "（元）必要人数"
PLAN_COL_SPEED_OVERRIDE = "加工速度_上書き"
PLAN_COL_TASK_EFFICIENCY = "タスク加工効率"
PLAN_COL_PRIORITY = "優先度"
PLAN_COL_SPECIFIED_DUE_OVERRIDE = "指定納期_上書き"
PLAN_COL_START_DATE_OVERRIDE = "加工開始日_指定"
PLAN_COL_START_TIME_OVERRIDE = "加工開始時刻_指定"
PLAN_COL_PREFERRED_OP = "担当OP_指定"
PLAN_COL_SPECIAL_REMARK = "特別指定_備考"
# 参照列「（元）配台不要」は置かない（元データに相当するマスタ列が無いため）。
# セル値の例（配台から外す）: Excel の TRUE / 数値 1 / 文字列「はい」「yes」「true」「○」「〇」「●」等。
# 空・FALSE・0・「いいえ」等は配台対象。詳細は _plan_row_exclude_from_assignment。
PLAN_COL_EXCLUDE_FROM_ASSIGNMENT = "配台不要"
PLAN_COL_AI_PARSE = "AI特別指定_解析"
PLAN_COL_PROCESS_FACTOR = "加工工程の決定プロセスの因子"
DEBUG_TASK_ID = os.environ.get("DEBUG_TASK_ID", "Y3-26").strip()
# 例: set TRACE_TEAM_ASSIGN_TASK_ID=W3-14 … 配台ループで「人数別の最良候補」と採用理由を INFO ログに出す
TRACE_TEAM_ASSIGN_TASK_ID = os.environ.get("TRACE_TEAM_ASSIGN_TASK_ID", "").strip()
# 配台トレース対象はマクロブック「設定」シート A 列 3 行目以降のみ（generate_plan 冒頭で確定）。環境変数は使わない。
TRACE_SCHEDULE_TASK_IDS: frozenset[str] = frozenset()
# デバッグ限定配台: 「設定」B3 以降が優先。空なら環境変数 DISPATCH_DEBUG_ONLY_TASK_IDS（カンマ区切り）。
_DISPATCH_DEBUG_ONLY_TASK_IDS_RAW = os.environ.get("DISPATCH_DEBUG_ONLY_TASK_IDS", "").strip()
DISPATCH_DEBUG_ONLY_TASK_IDS: frozenset[str] = frozenset()


def _trace_schedule_task_enabled(task_id) -> bool:
    if not TRACE_SCHEDULE_TASK_IDS:
        return False
    return str(task_id or "").strip() in TRACE_SCHEDULE_TASK_IDS


def _sanitize_dispatch_trace_filename_part(task_id: str) -> str:
    """依頼NOを log ファイル名に使うための簡易サニタイズ（Windows 禁止文字を避ける）。"""
    s = "".join(
        c if (c.isalnum() or c in "-_.") else "_"
        for c in str(task_id or "").strip()
    )
    return s[:120] if s else "task"


def _reset_dispatch_trace_per_task_logfiles() -> None:
    """
    配台トレース対象ごとに log/dispatch_trace_<依頼NO>.txt を新規作成（当該段階2実行の冒頭で1回）。
    実行前に log 内の dispatch_trace_*.txt をすべて削除し、過去実行の残骸を残さない。
    execution_log.txt とは別ファイル。内容は [配台トレース task=…] 行を _log_dispatch_trace_schedule で追記
    （日次残・ロール確定の余剰有無・余力追記・終了時サマリ等）。
    """
    if not TRACE_SCHEDULE_TASK_IDS:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
    except OSError:
        return
    try:
        for _name in os.listdir(log_dir):
            if not (
                str(_name).startswith("dispatch_trace_") and str(_name).endswith(".txt")
            ):
                continue
            _p = os.path.join(log_dir, _name)
            try:
                os.unlink(_p)
            except OSError:
                pass
    except OSError:
        pass
    for tid in TRACE_SCHEDULE_TASK_IDS:
        t = str(tid or "").strip()
        if not t:
            continue
        safe = _sanitize_dispatch_trace_filename_part(t)
        path = os.path.join(log_dir, f"dispatch_trace_{safe}.txt")
        try:
            with open(path, "w", encoding="utf-8", newline="\n") as f:
                f.write(
                    "# 配台トレース（依頼NOごと）。同一行は log/execution_log.txt にも出力されます。\n"
                    f"# task_id={t}\n\n"
                )
        except OSError as ex:
            logging.warning("dispatch_trace ログの初期化に失敗: %s (%s)", path, ex)


def _log_dispatch_trace_schedule(task_id, msg: str, *args) -> None:
    """[配台トレース task=…] を execution_log に出しつつ、対象依頼NO専用ファイルにも追記する。"""
    logging.info(msg, *args)
    t = str(task_id or "").strip()
    if not t or t not in TRACE_SCHEDULE_TASK_IDS:
        return
    safe = _sanitize_dispatch_trace_filename_part(t)
    path = os.path.join(log_dir, f"dispatch_trace_{safe}.txt")
    try:
        body = msg % args if args else msg
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
        line = f"{ts} - INFO - {body}\n"
        with open(path, "a", encoding="utf-8", newline="\n") as f:
            f.write(line)
    except OSError as ex:
        try:
            logging.warning("dispatch_trace 側ファイルへの追記に失敗: %s (%s)", path, ex)
        except Exception:
            pass


# True: 従来の「人数最優先」タプル (-人数, 開始, -単位数, 優先度合計)。False のとき下記スラック分と組み合わせ
TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF = os.environ.get(
    "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF", "0"
).strip().lower() not in ("0", "false", "no", "off", "いいえ")


def _team_assign_start_slack_wait_minutes() -> int:
    """全日候補の最早開始からこの分以内の遅れなら、開始より人数を優先（分）。0 で無効。"""
    raw = os.environ.get("TEAM_ASSIGN_START_SLACK_WAIT_MINUTES", "60").strip()
    try:
        v = int(raw)
    except ValueError:
        v = 60
    return max(0, v)


TEAM_ASSIGN_START_SLACK_WAIT_MINUTES = _team_assign_start_slack_wait_minutes()

# True のとき need シート「配台時追加人数」行を無視し、チーム人数は基本必要人数（req_num）のみ試行し、メイン後追記もしない。
TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW = (
    os.environ.get("TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)

# True: 従来どおりメイン割付の組み合わせ探索で req_num〜req_num+追加人数上限まで試す。
# False（既定）: メインは req_num のみ。追加人数上限は全シミュレーション完了後、当該ブロック時間に
#     他タスクへ未割当（時間重なりなし）かつ skills 適合の者をサブとして追記（append_surplus_staff_after_main_dispatch）。
TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS = (
    os.environ.get("TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS", "")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)

# True（既定）: メイン配台の必要人数は need（基本必要人数＋特別指定）のみ。
# 計画シート「必要人数」は headcount に使わない（参照列の表示用に残る）。
TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY = (
    os.environ.get("TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)
# True（既定）: master「組み合わせ表」に該当行がある工程+機械は、組合せ優先度の昇順で
# 最初に成立したメンバー編成を採用。すべて不可なら従来の itertools 組合せ探索。
TEAM_ASSIGN_USE_MASTER_COMBO_SHEET = (
    os.environ.get("TEAM_ASSIGN_USE_MASTER_COMBO_SHEET", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)

# §B-2 熱融着検査を同一設備（工程列キー）で「開始済み1件に残ロールがある間は他依頼の検査を試さない」か。
# 0 / false / no / off で無効にすると設備時間割上で依頼が混在し得るが、占有による長期ブロック（例: W3-14 型）を避けられる。
PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE = (
    os.environ.get("PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ", "無効")
)

# マクロブック「設定_配台不要工程」: 既定では openpyxl save を試さず xlwings 同期→Save（Excel 占有時は openpyxl が実質失敗するため）。失敗時は TSV→VBA 反映。
# コマンド等で openpyxl を試す場合は EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1。
EXCLUDE_RULES_SHEET_NAME = "設定_配台不要工程"
EXCLUDE_RULES_SKIP_OPENPYXL_SAVE = os.environ.get(
    "EXCLUDE_RULES_TRY_OPENPYXL_SAVE", ""
).strip().lower() not in ("1", "true", "yes", "on")
EXCLUDE_RULE_COL_PROCESS = "工程名"
EXCLUDE_RULE_COL_MACHINE = "機械名"
EXCLUDE_RULE_COL_FLAG = "配台不要"
EXCLUDE_RULE_COL_LOGIC_JA = "配台不能ロジック"
EXCLUDE_RULE_COL_LOGIC_JSON = "ロジック式"
# 元ブックがロックされ別名保存した場合、同一プロセス内のルール読込はこのパスを優先
_exclude_rules_effective_read_path: str | None = None
# 直後の apply_exclude_rules（同一プロセス）用: VBA 反映前でも E 列付きルールを使う
_exclude_rules_rules_snapshot: list | None = None
_exclude_rules_snapshot_wb: str | None = None
# ルール JSON の conditions で参照可能な列（AI プロンプトと評価器を一致させる）
EXCLUDE_RULE_ALLOWED_COLUMNS = frozenset(
    {
        TASK_COL_TASK_ID,
        TASK_COL_MACHINE,
        TASK_COL_MACHINE_NAME,
        TASK_COL_QTY,
        TASK_COL_ORDER_QTY,
        TASK_COL_SPEED,
        TASK_COL_PRODUCT,
        TASK_COL_ANSWER_DUE,
        TASK_COL_SPECIFIED_DUE,
        TASK_COL_RAW_INPUT_DATE,
        TASK_COL_PROCESS_CONTENT,
        TASK_COL_COMPLETION_FLAG,
        TASK_COL_ACTUAL_DONE,
        TASK_COL_ACTUAL_OUTPUT,
        TASK_COL_DATA_EXTRACTION_DT,
        PLAN_COL_REQUIRED_OP,
        PLAN_COL_SPEED_OVERRIDE,
        PLAN_COL_TASK_EFFICIENCY,
        PLAN_COL_PRIORITY,
        PLAN_COL_SPECIFIED_DUE_OVERRIDE,
        PLAN_COL_START_DATE_OVERRIDE,
        PLAN_COL_START_TIME_OVERRIDE,
        PLAN_COL_PREFERRED_OP,
        PLAN_COL_SPECIAL_REMARK,
        PLAN_COL_PROCESS_FACTOR,
    }
)

# 計画結果ブック「結果_タスク一覧」の列順・表示（マクロ実行ブックの同名シートで上書き可）
RESULT_TASK_SHEET_NAME = "結果_タスク一覧"
RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME = "結果_設備毎の時間割"
# 配台シミュレーション開始前（初回 task_queue.sort 後）のキュー順。1 始まり・全日程で不変
RESULT_TASK_COL_DISPATCH_TRIAL_ORDER = "配台試行順番"
# 配完_加工終了が「回答納期+16:00」または「指定納期+16:00」（回答が空のとき）以前かを表示
RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16 = "配完_回答指定16時まで"
# マスタ skills の工程+機械列ごとの OP/AS 割当参考順（優先度値・氏名順）とチーム採用ルールの説明
RESULT_MEMBER_PRIORITY_SHEET_NAME = "結果_人員配台優先順"
COLUMN_CONFIG_SHEET_NAME = "列設定_結果_タスク一覧"
COLUMN_CONFIG_HEADER_COL = "列名"
COLUMN_CONFIG_VISIBLE_COL = "表示"
# 結果_タスク一覧の日付系（yyyy/mm/dd 文字列）に付けるフォント色。履歴列の【日付】と揃える
RESULT_TASK_DATE_STYLE_HEADERS = frozenset(
    {
        "回答納期",
        "指定納期",
        "計画基準納期",
        TASK_COL_RAW_INPUT_DATE,
        "加工開始日",
        "配完_加工開始",
        "配完_加工終了",
    }
)

SOURCE_BASE_COLUMNS = [
    TASK_COL_TASK_ID, TASK_COL_MACHINE, TASK_COL_MACHINE_NAME, TASK_COL_QTY, TASK_COL_ORDER_QTY, TASK_COL_SPEED, TASK_COL_PRODUCT,
    TASK_COL_ANSWER_DUE, TASK_COL_SPECIFIED_DUE, TASK_COL_RAW_INPUT_DATE,
    TASK_COL_PROCESS_CONTENT,
    TASK_COL_COMPLETION_FLAG, TASK_COL_ACTUAL_DONE, TASK_COL_ACTUAL_OUTPUT,
]
PLAN_OVERRIDE_COLUMNS = [
    PLAN_COL_EXCLUDE_FROM_ASSIGNMENT,
    PLAN_COL_REQUIRED_OP, PLAN_COL_SPEED_OVERRIDE, PLAN_COL_TASK_EFFICIENCY,
    PLAN_COL_PRIORITY, PLAN_COL_SPECIFIED_DUE_OVERRIDE, PLAN_COL_START_DATE_OVERRIDE, PLAN_COL_START_TIME_OVERRIDE,
    PLAN_COL_PREFERRED_OP,
    PLAN_COL_SPECIAL_REMARK,
    PLAN_COL_AI_PARSE,
]
# 矛盾検出でリセット対象にする列（見出し行の文言と一致すること）
PLAN_CONFLICT_STYLABLE_COLS = tuple(PLAN_OVERRIDE_COLUMNS)
# 段階1再抽出時、既存「配台計画_タスク入力」から継承する列（AIの解析結果列は毎回空に戻す）
PLAN_STAGE1_MERGE_COLUMNS = tuple(c for c in PLAN_OVERRIDE_COLUMNS if c != PLAN_COL_AI_PARSE)
# openpyxl 保存がブックロックで失敗したとき、VBA が開いているブックへ書式適用するための指示ファイル
PLANNING_CONFLICT_SIDECAR = "planning_conflict_highlight.tsv"
# 配台計画_タスク入力へ「グローバルコメント解析」を書く列（表の右端より外側。1行目から縦にラベル／値）
# ★ 参照表示のみ: load_planning_tasks_df 等は本列を一切読まない。配台適用は常にメイン「グローバルコメント」1経路のため二重適用にならない。
PLAN_SHEET_GLOBAL_PARSE_LABEL_COL = 50  # AX
PLAN_SHEET_GLOBAL_PARSE_VALUE_COL = 51  # AY
PLAN_SHEET_GLOBAL_PARSE_MAX_ROWS = 42


def plan_reference_column_name(override_col: str) -> str:
    """上書き列の左隣に置く参照列の見出し（セル値は括弧付きで元データを表示）。"""
    if override_col == PLAN_COL_REQUIRED_OP:
        return PLAN_REF_REQUIRED_OP
    return f"（元）{override_col}"


def plan_input_sheet_column_order():
    """
    配台計画_タスク入力の列順（段階1出力・段階2読込で共通）。

    1. 配台不要（参照列なし）
    2. 加工計画DATA 由来（SOURCE_BASE_COLUMNS）… 依頼NO〜実出来高まで
    3. 加工工程の決定プロセスの因子
    4. 上書き列… 各列の直前に「（元）…」参照列。AI特別指定_解析のみ参照列なし。

    global_speed_rules 等で変わる実効速度はシート列では持たず、配台内部のみで反映する。
    """
    cols = [PLAN_COL_EXCLUDE_FROM_ASSIGNMENT]
    cols.extend(SOURCE_BASE_COLUMNS)
    cols.append(PLAN_COL_PROCESS_FACTOR)
    for c in PLAN_OVERRIDE_COLUMNS:
        if c == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
            continue
        if c == PLAN_COL_AI_PARSE:
            cols.append(c)
        else:
            cols.append(plan_reference_column_name(c))
            cols.append(c)
    return cols


def _format_paren_ref_scalar(val):
    """参照表示用: 空は（―）、日付・その他は（値）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "（―）"
    if isinstance(val, datetime):
        d = val.date() if hasattr(val, "date") else val
        if isinstance(d, date):
            return f"（{d.year}/{d.month}/{d.day}）"
    if isinstance(val, date):
        return f"（{val.year}/{val.month}/{val.day}）"
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return "（―）"
    return f"（{s}）"


def _reference_text_for_override_row(row, override_col: str, req_map: dict, need_rules: list) -> str:
    """1行分の上書き列に対応する参照文言（括弧付き）。"""
    mach = str(row.get(TASK_COL_MACHINE, "") or "").strip()
    mname = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()

    if override_col == PLAN_COL_REQUIRED_OP:
        try:
            n = resolve_need_required_op(mach, mname, planning_task_id_str_from_plan_row(row), req_map, need_rules)
            return f"（{n}）"
        except Exception:
            return "（―）"
    if override_col == PLAN_COL_SPEED_OVERRIDE:
        v = row.get(TASK_COL_SPEED)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "（―）"
        try:
            x = float(v)
            if abs(x - round(x)) < 1e-9:
                return f"（{int(round(x))}）"
            return f"（{x}）"
        except (TypeError, ValueError):
            return _format_paren_ref_scalar(v)
    if override_col in (
        PLAN_COL_TASK_EFFICIENCY,
        PLAN_COL_PRIORITY,
        PLAN_COL_START_TIME_OVERRIDE,
        PLAN_COL_PREFERRED_OP,
        PLAN_COL_SPECIAL_REMARK,
    ):
        return "（―）"
    if override_col == PLAN_COL_SPECIFIED_DUE_OVERRIDE:
        sd = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_SPECIFIED_DUE))
        if sd is not None:
            return _format_paren_ref_scalar(sd)
        ad = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_ANSWER_DUE))
        if ad is not None:
            return _format_paren_ref_scalar(ad)
        return "（―）"
    if override_col == PLAN_COL_START_DATE_OVERRIDE:
        return _format_paren_ref_scalar(
            parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE))
        )
    return "（―）"


def _refresh_plan_reference_columns(df, req_map: dict, need_rules: list):
    """加工計画DATA／need に基づき「（元）…」列を再計算（マージ後に必ず呼ぶ）。"""
    if df is None or df.empty:
        return df
    need_rules = need_rules or []
    req_map = req_map or {}
    for i in df.index:
        row = df.loc[i]
        for oc in PLAN_OVERRIDE_COLUMNS:
            if oc == PLAN_COL_AI_PARSE:
                continue
            if oc == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
                continue
            ref_col = plan_reference_column_name(oc)
            if ref_col not in df.columns:
                continue
            df.at[i, ref_col] = _reference_text_for_override_row(row, oc, req_map, need_rules)
    return df


def _apply_plan_input_visual_format(path: str, sheet_name: str = "タスク一覧"):
    """上書き入力列に薄い黄色を付与（参照列は未着色。AI解析列は除外）。"""
    # 見出し文字の表記ゆれで列名検索に失敗しがちなため、段階1の列順（plan_input_sheet_column_order）の
    # 1-based 列番号で塗る（to_excel の列順と一致させる）。
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    order = plan_input_sheet_column_order()
    col_1based = {name: i + 1 for i, name in enumerate(order)}
    if _workbook_should_skip_openpyxl_io(path):
        logging.info(
            "配台計画の視覚整形: ブックに「%s」があるため openpyxl での着色をスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    wb = load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        last_row = ws.max_row or 1
        if last_row < 2:
            return
        for oc in PLAN_OVERRIDE_COLUMNS:
            if oc == PLAN_COL_AI_PARSE:
                continue
            ci = col_1based.get(oc)
            if not ci:
                continue
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=ci).fill = fill_yellow
        wb.save(path)
    finally:
        wb.close()


def _planning_conflict_sidecar_path():
    return os.path.join(log_dir, PLANNING_CONFLICT_SIDECAR)


def _remove_planning_conflict_sidecar_safe():
    try:
        os.remove(_planning_conflict_sidecar_path())
    except OSError:
        pass


def write_planning_conflict_highlight_sidecar(sheet_name, num_data_rows, conflicts_by_row):
    """
    Excel がブックを開いたままのとき保存できない場合に、VBA 用の TSV を log に書く。
    形式: V1 / シート名 / データ行数 / クリア列をタブ結合 / 以降 行番号\\t列名
    """
    path = _planning_conflict_sidecar_path()
    clear_cols = "\t".join(PLAN_CONFLICT_STYLABLE_COLS)
    lines = ["V1", sheet_name, str(int(num_data_rows)), clear_cols]
    for r in sorted(conflicts_by_row.keys()):
        for name in sorted(conflicts_by_row[r]):
            lines.append(f"{int(r)}\t{name}")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")

# 段階1出力・ブック内の日付列を Excel 上「日付のみ」(時刻なし表示) に整える
STAGE1_SHEET_DATEONLY_HEADERS = frozenset(
    {
        TASK_COL_ANSWER_DUE,
        TASK_COL_SPECIFIED_DUE,
        TASK_COL_RAW_INPUT_DATE,
        PLAN_COL_SPECIFIED_DUE_OVERRIDE,
        PLAN_COL_START_DATE_OVERRIDE,
    }
)


def _result_font(**kwargs):
    """結果ブック用 Font。OUTPUT_BOOK_FONT_* があれば付与（kwargs の name/size が優先）。"""
    if OUTPUT_BOOK_FONT_NAME and "name" not in kwargs:
        kwargs["name"] = OUTPUT_BOOK_FONT_NAME
    if OUTPUT_BOOK_FONT_SIZE is not None and "size" not in kwargs:
        kwargs["size"] = OUTPUT_BOOK_FONT_SIZE
    return Font(**kwargs)


def _output_book_font(bold=False):
    return _result_font(bold=bold)


def _apply_output_font_to_result_sheet(ws):
    """結果_* のうちガント以外向け: 既定フォント・1行目太字のみ（列幅は VBA AutoFit）。"""
    base = _output_book_font(bold=False)
    hdr = _output_book_font(bold=True)
    mr, mc = ws.max_row or 1, ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
        for cell in row:
            cell.font = base
    for cell in ws[1]:
        cell.font = hdr


def _apply_excel_date_columns_date_only_display(path, sheet_name, header_names=None):
    """openpyxl: 指定ヘッダー列を yyyy/mm/dd の日付表示にする（時刻を表示しない）。"""
    from openpyxl import load_workbook

    headers = header_names or STAGE1_SHEET_DATEONLY_HEADERS
    if _workbook_should_skip_openpyxl_io(path):
        logging.info(
            "日付列表示整形: ブックに「%s」があるため openpyxl での処理をスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    wb = load_workbook(path)
    try:
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[int(sheet_name)]
        cmap = {}
        for cell in ws[1]:
            if cell.value is not None:
                cmap[str(cell.value).strip()] = cell.column
        fmt = "yyyy/mm/dd"
        for h in headers:
            col = cmap.get(h)
            if not col:
                continue
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                v = c.value
                if v is None:
                    continue
                if isinstance(v, datetime):
                    c.value = v.date()
                elif isinstance(v, date):
                    pass
                else:
                    try:
                        d0 = pd.to_datetime(v, errors="coerce")
                        if pd.isna(d0):
                            continue
                        c.value = d0.date()
                    except Exception:
                        continue
                c.number_format = fmt
        wb.save(path)
    finally:
        wb.close()


def _extract_data_extraction_datetime():
    """
    `加工計画DATA` シートの `データ抽出日` から datetime を取得する。
    """
    try:
        if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
            return None
        df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=TASKS_SHEET_NAME)
        df.columns = df.columns.str.strip()
        if TASK_COL_DATA_EXTRACTION_DT not in df.columns:
            return None
        s = df[TASK_COL_DATA_EXTRACTION_DT]
        first = None
        for v in s:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            first = v
            break
        if first is None:
            return None
        dt = pd.to_datetime(first, errors="coerce")
        if pd.isna(dt):
            return None
        if isinstance(dt, pd.Timestamp):
            return dt.to_pydatetime()
        return dt if isinstance(dt, datetime) else None
    except Exception:
        return None


def _extract_data_extraction_datetime_str():
    """
    `加工計画DATA` シートの `データ抽出日` から「データ吸出し日時」を取得して文字列化する。
    """
    try:
        dt = _extract_data_extraction_datetime()
        if dt is None:
            return "—"
        return dt.strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        return "—"


def _weekday_jp(d):
    return "月火水木金土日"[d.weekday()]


# ガントの作業バー：いずれも明るい地色＋黒文字が読めるトーン（モノクロ印刷でも濃淡で識別しやすい）
_GANTT_BAR_FILLS_PRINT_SAFE = (
    "E8E8E8",
    "D8E4EF",
    "E6E2DB",
    "DEEADF",
    "E8E0E8",
    "EAE8D8",
    "DDE6EA",
    "E5DCE5",
)

# 実績バー用（計画と並べてもモノクロで区別しやすいトーン）
_GANTT_BAR_FILLS_ACTUAL = (
    "D4E4D4",
    "C9DDE8",
    "DED8CC",
    "D2E5CD",
    "DAD2D9",
    "E0DCCF",
    "CDE2E8",
    "DCD2DC",
)


def _gantt_bar_fill_for_task_id(task_id):
    """依頼NOごとに上記パレットから1色（RRGGBB）。濃色＋白文字の組み合わせは使わない。"""
    h = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    i = int(h[:8], 16) % len(_GANTT_BAR_FILLS_PRINT_SAFE)
    return _GANTT_BAR_FILLS_PRINT_SAFE[i]


def _gantt_bar_fill_actual_for_task_id(task_id):
    h = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    i = int(h[:8], 16) % len(_GANTT_BAR_FILLS_ACTUAL)
    return _GANTT_BAR_FILLS_ACTUAL[i]


def _gantt_slot_state_tuple(evlist, slot_mid, task_fill_fn=None):
    """スロット中央時刻における1マス分の状態。('idle',) | ('break',) | ('task', tid, fill_hex, pct)"""
    fill_fn = task_fill_fn or _gantt_bar_fill_for_task_id
    active = None
    for e in evlist:
        if e["start_dt"] <= slot_mid < e["end_dt"]:
            active = e
            break
    if active is None:
        return ("idle",)
    if any(b_s <= slot_mid < b_e for b_s, b_e in active.get("breaks") or ()):
        return ("break",)
    tid = str(active["task_id"])
    gh = fill_fn(active["task_id"])
    pct = None
    try:
        # 「マクロ実行時点」の完了率を優先（pct_macro を timeline_event に持たせる）
        if active.get("pct_macro") is not None:
            pct = int(round(parse_float_safe(active.get("pct_macro"), 0.0)))
            pct = max(0, min(100, pct))
        else:
            # フェイルセーフ（従来の擬似進捗計算）
            tot = parse_float_safe(active.get("total_units"), 0.0)
            done = parse_float_safe(active.get("already_done_units"), 0.0) + parse_float_safe(
                active.get("units_done"), 0.0
            )
            if tot > 0:
                pct = max(0, min(100, int(round((done / tot) * 100))))
    except Exception:
        pct = None
    return ("task", tid, gh, pct)


def _gantt_merge_key(st):
    if st[0] == "idle":
        return ("idle",)
    if st[0] == "break":
        return ("break",)
    return ("task", st[1])


def _paint_gantt_timeline_row_merged(
    ws,
    row,
    n_fixed,
    slots,
    slot_mins,
    evlist,
    idle_fill,
    break_fill,
    gantt_label_font,
    grid_border,
    task_fill_fn=None,
    label_font=None,
):
    """
    時間軸を塗り分けたうえで、同一状態が連続するセルを横結合し帯状のバーにする。
    （細マス単体の塗りではなく15分刻み＋同一状態のセル結合で、帯状のバーとして表現する）
    """
    bar_label_font = label_font or gantt_label_font
    n_slots = len(slots)
    states = []
    for slot_start in slots:
        mid = slot_start + timedelta(minutes=slot_mins / 2)
        states.append(_gantt_slot_state_tuple(evlist, mid, task_fill_fn))
    tcol0 = n_fixed + 1
    i = 0
    while i < n_slots:
        mk = _gantt_merge_key(states[i])
        j = i + 1
        while j < n_slots and _gantt_merge_key(states[j]) == mk:
            j += 1
        col_s = tcol0 + i
        col_e = tcol0 + j - 1
        st0 = states[i]
        if col_s < col_e:
            ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
        c = ws.cell(row=row, column=col_s)
        c.border = grid_border
        # 左寄せ・折り返しなし・縮小なし（長いラベルはセル外へはみ出し表示しやすくする）
        c.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False,
            shrink_to_fit=False,
            indent=1,
        )
        if st0[0] == "idle":
            c.fill = idle_fill
        elif st0[0] == "break":
            c.fill = break_fill
        else:
            _, tid, gh, pct = st0
            c.fill = PatternFill(fill_type="solid", start_color=gh, end_color=gh)
            c.value = f"{tid[:9]} {pct}%" if pct is not None else tid[:9]
            c.font = bar_label_font
        i = j


def _write_results_equipment_gantt_sheet(
    writer,
    timeline_events,
    equipment_list,
    sorted_dates,
    attendance_data,
    data_extract_dt_str,
    base_now_dt=None,
    actual_timeline_events=None,
):
    """
    結果_設備毎の時間割と同一データ源（timeline_events）に基づき、
    設備×横軸時間のガンチャート風シートを追加する。
    横軸は15分刻み。連続する同一タスク／休憩／空きはセル結合して帯状に表示する。
    actual_timeline_events があれば設備ごとに「実績」行を計画行の下へ追加する。
    """
    wb = writer.book
    try:
        insert_at = wb.sheetnames.index("結果_設備毎の時間割") + 1
    except ValueError:
        insert_at = len(wb.sheetnames)
    ws = wb.create_sheet("結果_設備ガント", insert_at)
    try:
        ws.sheet_properties.tabColor = "7F7F7F"
    except Exception:
        pass

    events_by_date = defaultdict(list)
    for e in timeline_events:
        events_by_date[e["date"]].append(e)

    by_dm = defaultdict(lambda: defaultdict(list))
    for e in timeline_events:
        by_dm[e["date"]][e["machine"]].append(e)
    for d0 in by_dm:
        for mk in by_dm[d0]:
            by_dm[d0][mk].sort(key=lambda x: x["start_dt"])

    by_dm_actual = defaultdict(lambda: defaultdict(list))
    show_actual_rows = bool(actual_timeline_events)
    actual_events_by_date = defaultdict(list)
    if show_actual_rows:
        for e in actual_timeline_events:
            actual_events_by_date[e["date"]].append(e)
            by_dm_actual[e["date"]][e["machine"]].append(e)
        for d0 in by_dm_actual:
            for mk in by_dm_actual[d0]:
                by_dm_actual[d0][mk].sort(key=lambda x: x["start_dt"])

    slot_mins = 15
    eq_display = _equipment_schedule_header_labels(equipment_list)
    hdr_font = _result_font(bold=True, color="000000", size=10)
    hdr_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    hdr_time_font = _result_font(bold=True, color="000000", size=9)
    title_font = _result_font(bold=True, size=20, color="1A1A1A")
    title_fill = PatternFill(fill_type="solid", start_color="DDDDDD", end_color="DDDDDD")
    meta_font = _result_font(size=9, color="333333")
    meta_fill = PatternFill(fill_type="solid", start_color="F3F3F3", end_color="F3F3F3")
    day_banner_font = _result_font(bold=True, size=11, color="1A1A1A")
    day_banner_fill = PatternFill(fill_type="solid", start_color="D0D0D0", end_color="D0D0D0")
    accent_left = Side(style="thick", color="2B2B2B")
    banner_sep = Side(style="thin", color="7A7A7A")
    thin = Side(style="thin", color="666666")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    idle_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    break_fill = PatternFill(fill_type="solid", start_color="B8B8B8", end_color="B8B8B8")
    gantt_label_font = _result_font(size=8, bold=True, color="000000")
    gantt_label_font_actual = _result_font(size=8, bold=True, color="000000", italic=True)

    # 横軸(10分刻み)は日付で共通のため、slot_times を先に確定
    base_dt = base_now_dt if isinstance(base_now_dt, datetime) else datetime.now()
    dummy_d = sorted_dates[0] if sorted_dates else base_dt.date()
    d_start0 = datetime.combine(dummy_d, DEFAULT_START_TIME)
    d_end0 = datetime.combine(dummy_d, DEFAULT_END_TIME)
    slot_times = []
    t0 = d_start0
    while t0 < d_end0:
        slot_times.append(t0.time())
        t0 += timedelta(minutes=slot_mins)

    n_slots = len(slot_times)
    n_fixed = 3  # 設備 / タスク概要 / 主担当（日付は日ブロック見出しのみ）
    last_col = n_fixed + n_slots

    # タイトル＆日時（ページ上部）
    create_ts = base_dt.strftime("%Y/%m/%d %H:%M:%S")
    master_path = os.path.join(os.getcwd(), MASTER_FILE) if MASTER_FILE else ""

    def _fmt_mtime(p):
        try:
            if p and os.path.exists(p):
                return datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y/%m/%d %H:%M:%S")
        except Exception:
            pass
        return "—"

    master_mtime = _fmt_mtime(master_path)

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    tcell = ws.cell(row=row, column=1, value="湖南工場 加工計画")
    tcell.font = title_font
    tcell.fill = title_fill
    # 結合セルでも左端から表示（縮小・折り返しなし）
    tcell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=False,
        indent=1,
    )
    tcell.border = Border(left=accent_left, bottom=banner_sep)
    ws.row_dimensions[row].height = 34
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    meta_line = (
        f"作成　{create_ts}"
        f"　・　データ吸出し　{data_extract_dt_str or '—'}"
        f"　・　マスタ（master.xlsm）　{master_mtime}"
    )
    mtop = ws.cell(row=row, column=1, value=meta_line)
    mtop.font = meta_font
    mtop.fill = meta_fill
    mtop.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=1,
        wrap_text=False,
        shrink_to_fit=False,
    )
    mtop.border = Border(left=accent_left, bottom=banner_sep)
    ws.row_dimensions[row].height = 22
    row += 1

    first_freeze_set = False

    for d in sorted_dates:
        evs = events_by_date.get(d, [])
        a_evs_day = actual_events_by_date.get(d, []) if show_actual_rows else []
        is_anyone_working = any(
            attendance_data[d][mm]["is_working"] for mm in attendance_data[d] if mm in attendance_data[d]
        )
        if not evs and not a_evs_day and not is_anyone_working:
            continue

        slots = [datetime.combine(d, tm) for tm in slot_times]

        # 日付見出し（左の固定列幅に合わせて A〜C に表示）
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_fixed)
        ban = ws.cell(row=row, column=1, value=f"▶ {d.strftime('%Y/%m/%d')}　{_weekday_jp(d)}")
        ban.font = day_banner_font
        ban.fill = day_banner_fill
        ban.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ban.border = Border(left=accent_left, bottom=thin)
        ws.row_dimensions[row].height = 22
        row += 1

        fixed_hdr = ["設備", "タスク概要", "主担当"]
        for ci, h in enumerate(fixed_hdr, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        for si, st in enumerate(slots):
            c = ws.cell(row=row, column=n_fixed + 1 + si, value=st.strftime("%H:%M"))
            c.font = hdr_time_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="bottom", textRotation=90)
        ws.row_dimensions[row].height = 38
        row += 1

        data_row0 = row
        if not first_freeze_set:
            ws.freeze_panes = f"{get_column_letter(n_fixed + 1)}{data_row0}"
            first_freeze_set = True

        zebra = False
        for eq, disp in zip(equipment_list, eq_display):
            zebra = not zebra
            evlist = by_dm[d].get(eq, [])
            if evlist:
                parts = []
                ops = []
                for e in evlist:
                    cum = e["already_done_units"] + e["units_done"]
                    tot = e["total_units"]
                    parts.append(f"[{e['task_id']}] {cum}/{tot}R")
                    ops.append(str(e["op"]))
                task_sum = " ｜ ".join(parts)[:120]
                op_disp = ops[0] if len(set(ops)) == 1 else ",".join(dict.fromkeys(ops))
            else:
                task_sum = "—"
                op_disp = "—"

            lab_fill = PatternFill(fill_type="solid", start_color="FAFAFA", end_color="FAFAFA")
            if zebra:
                lab_fill = PatternFill(fill_type="solid", start_color="F0F4FA", end_color="F0F4FA")

            c1 = ws.cell(row=row, column=1, value=disp)
            c2 = ws.cell(row=row, column=2, value=task_sum)
            c3 = ws.cell(row=row, column=3, value=op_disp)
            for c in (c1, c2, c3):
                c.font = _result_font(size=10, color="000000")
                c.fill = lab_fill
                c.border = grid_border
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            c1.font = _result_font(size=10, bold=True, color="000000")

            _paint_gantt_timeline_row_merged(
                ws,
                row,
                n_fixed,
                slots,
                slot_mins,
                evlist,
                idle_fill,
                break_fill,
                gantt_label_font,
                grid_border,
            )

            ws.row_dimensions[row].height = 22
            row += 1

            if show_actual_rows:
                evlist_a = by_dm_actual[d].get(eq, [])
                if evlist_a:
                    parts_a = [f"[{e_a['task_id']}]" for e_a in evlist_a]
                    task_sum_a = " ".join(dict.fromkeys(parts_a))[:120]
                    ops_a = [
                        str(e_a.get("op") or "").strip()
                        for e_a in evlist_a
                        if str(e_a.get("op") or "").strip()
                    ]
                    op_disp_a = (
                        ops_a[0]
                        if len(set(ops_a)) == 1
                        else ",".join(dict.fromkeys(ops_a))
                    )
                else:
                    task_sum_a = "—"
                    op_disp_a = "—"

                lab_fill_a = PatternFill(
                    fill_type="solid", start_color="EEF6EE", end_color="EEF6EE"
                )
                if zebra:
                    lab_fill_a = PatternFill(
                        fill_type="solid", start_color="E0EEE0", end_color="E0EEE0"
                    )

                ca1 = ws.cell(row=row, column=1, value=f"{disp}（実績）")
                ca2 = ws.cell(row=row, column=2, value=task_sum_a)
                ca3 = ws.cell(row=row, column=3, value=op_disp_a)
                for c in (ca1, ca2, ca3):
                    c.font = _result_font(size=10, color="000000")
                    c.fill = lab_fill_a
                    c.border = grid_border
                    c.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=False
                    )
                ca1.font = _result_font(size=10, bold=True, color="000000", italic=True)

                _paint_gantt_timeline_row_merged(
                    ws,
                    row,
                    n_fixed,
                    slots,
                    slot_mins,
                    evlist_a,
                    idle_fill,
                    break_fill,
                    gantt_label_font_actual,
                    grid_border,
                    task_fill_fn=_gantt_bar_fill_actual_for_task_id,
                    label_font=gantt_label_font_actual,
                )

                ws.row_dimensions[row].height = 22
                row += 1
        # 凡例は高さ確保のため省略（モノクロ印刷は色の濃淡/セルの枠で識別）
    # 列幅は VBA 取り込み時（結果_設備ガント_列幅を設定）で設定

    try:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1
        # A3（openpyxl 上で paperSize=8 が A3 相当）
        ws.page_setup.paperSize = 8
        # 余白を狭めて横1ページに収まりやすくする（単位: インチ）
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.2
        # タイトル・表をページ左基準に（レポート風）
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False
        ws.print_options.gridLines = False
    except Exception:
        pass


def row_has_completion_keyword(row):
    """加工完了区分に「完了」の文字が含まれる場合はタスク完了とみなす。"""
    v = row.get(TASK_COL_COMPLETION_FLAG)
    if v is None or pd.isna(v):
        return False
    return "完了" in str(v)


def _plan_row_exclude_from_assignment(row) -> bool:
    """
    「配台不要」列がオンなら、その行は配台キューへ入れず、特別指定_備考の AI 解析行からも除く。

    配台から外す（真）: 論理値 True、数値 1、文字列（NFKC 後・小文字）
      true / 1 / yes / on / y / t / はい / ○ / 〇 / ●
    配台対象（偽）: 空、None、False、0、no / off / false / いいえ / 否 等
    上記以外の文字列は偽（配台する）。チェックボックス連動セルは通常 TRUE/FALSE または 1/0。
    """
    v = row.get(PLAN_COL_EXCLUDE_FROM_ASSIGNMENT)
    if v is True:
        return True
    if v is False:
        return False
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            iv = int(v)
            if iv == 1:
                return True
            if iv == 0:
                return False
        except (TypeError, ValueError):
            pass
    s = unicodedata.normalize("NFKC", str(v).strip()).lower()
    if not s or s in ("nan", "none", "false", "0", "no", "off", "いいえ", "否"):
        return False
    if s in ("true", "1", "yes", "on", "はい", "y", "t", "○", "〇", "●"):
        return True
    return False


def _coerce_plan_exclude_column_value_for_storage(v):
    """
    「配台不要」列へ書き込む値を、StringDtype 列でも代入エラーにならない形にそろえる。
    Excel 取り込みの True / 1 / False / 0 と文字列を保持し、_plan_row_exclude_from_assignment と整合する。
    """
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if v is True:
        return "yes"
    if v is False:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            iv = int(v)
            if iv == 1:
                return "yes"
            if iv == 0:
                return ""
        except (TypeError, ValueError):
            pass
    return str(v).strip()


def parse_float_safe(val, default=0.0):
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def calc_done_qty_equivalent_from_row(row):
    """
    加工済数量（工程投入量換算）を返す。

    基本式:
      実出来高 ÷ (受注数 ÷ 換算数量)
    = 実出来高 * 換算数量 / 受注数

    受注数が無い/不正な場合は、旧列「実加工数」を互換フォールバックとして使う。
    """
    qty_total = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
    order_qty = parse_float_safe(row.get(TASK_COL_ORDER_QTY), 0.0)
    actual_output = parse_float_safe(row.get(TASK_COL_ACTUAL_OUTPUT), 0.0)
    legacy_done = parse_float_safe(row.get(TASK_COL_ACTUAL_DONE), 0.0)

    if qty_total <= 0:
        return max(0.0, legacy_done)

    if order_qty > 0 and actual_output >= 0:
        done_qty = actual_output * qty_total / order_qty
        return max(0.0, done_qty)

    return max(0.0, legacy_done)


def parse_optional_int(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    try:
        return int(round(float(s)))
    except (TypeError, ValueError):
        return None


def parse_optional_date(val):
    if val is None or pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def _planning_df_cell_scalar(row, col_name):
    """
    iterrows() 1行分から列値を取る。同一見出しの重複列があると row.get は Series になり、
    str→to_datetime で誤った日付になることがあるため、先頭の非欠損スカラーを返す。
    """
    v = row.get(col_name) if hasattr(row, "get") else None
    if isinstance(v, pd.Series):
        for x in v:
            if x is None or (isinstance(x, float) and pd.isna(x)):
                continue
            return x
        return None
    return v


def _planning_row_plan_priority_cell(row):
    """
    配台計画シート「優先度」の実セル値を1つに決める。
    pandas.read_excel が重複見出しを 優先度.1 のように付けた列や、同一ラベル重複で Series になる場合を吸収。
    （左の 優先度 が空で .1 にだけ数値がある行でも検出・配台で同じ値を使う）
    """
    v = _planning_df_cell_scalar(row, PLAN_COL_PRIORITY)
    if v is not None and not (isinstance(v, float) and pd.isna(v)):
        if isinstance(v, str) and not str(v).strip():
            v = None
        else:
            return v
    idx = getattr(row, "index", None)
    if idx is None:
        return None
    for suf in (".1", ".2", ".3"):
        alt = f"{PLAN_COL_PRIORITY}{suf}"
        if alt not in idx:
            continue
        try:
            vx = row[alt]
        except (KeyError, TypeError):
            vx = row.get(alt) if hasattr(row, "get") else None
        if vx is None or (isinstance(vx, float) and pd.isna(vx)):
            continue
        if isinstance(vx, str) and not str(vx).strip():
            continue
        return vx
    return None


def load_ai_cache():
    try:
        if os.path.exists(ai_cache_path):
            with open(ai_cache_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    # 期限切れエントリを除去（6時間）
                    now_ts = time_module.time()
                    cleaned = {}
                    expired_count = 0
                    for k, v in data.items():
                        # 新形式: {"ts": epoch_seconds, "data": {...}}
                        if isinstance(v, dict) and "ts" in v and "data" in v:
                            ts = parse_float_safe(v.get("ts"), 0.0)
                            if ts > 0 and (now_ts - ts) <= AI_CACHE_TTL_SECONDS:
                                cleaned[k] = v
                            else:
                                expired_count += 1
                        # 旧形式: 値が直接AI結果dict（互換で読み取り、即時に新形式へ再保存される）
                        else:
                            cleaned[k] = {"ts": now_ts, "data": v}
                    if expired_count > 0:
                        logging.info(f"AIキャッシュ期限切れを削除: {expired_count}件")
                    return cleaned
    except Exception as e:
        logging.warning(f"AIキャッシュ読み込み失敗: {e}")
    return {}

def save_ai_cache(cache_obj):
    try:
        with open(ai_cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_obj, f, ensure_ascii=False)
    except Exception as e:
        logging.warning(f"AIキャッシュ保存失敗: {e}")

def get_cached_ai_result(cache_obj, cache_key, content_key=None):
    """
    content_key: オプション。保存時と同一の文字列でないヒットは無効化する（特別指定・照合用の二次チェック）。
    旧エントリに content_key が無い場合は SHA256 キー一致のみで従来どおりヒットとみなす。
    """
    entry = cache_obj.get(cache_key)
    if not isinstance(entry, dict):
        return None
    ts = parse_float_safe(entry.get("ts"), 0.0)
    if ts <= 0:
        return None
    if (time_module.time() - ts) > AI_CACHE_TTL_SECONDS:
        return None
    if content_key is not None:
        stored_ck = entry.get("content_key")
        if stored_ck is not None and stored_ck != content_key:
            logging.info(
                "AIキャッシュ: キーは一致しますが content_key が現行入力と異なるため無効化します。"
            )
            return None
    data = entry.get("data")
    if isinstance(data, dict):
        return data
    return None

def put_cached_ai_result(cache_obj, cache_key, parsed_obj, content_key=None):
    payload = {"ts": time_module.time(), "data": parsed_obj}
    if content_key is not None:
        payload["content_key"] = content_key
    cache_obj[cache_key] = payload

def extract_retry_seconds(err_text):
    # 例: "Please retry in 57.089735313s."
    m = re.search(r"retry in ([0-9]+(?:\.[0-9]+)?)s", err_text, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    # 例: "'retryDelay': '57s'"
    m = re.search(r"retryDelay'\s*:\s*'([0-9]+)s'", err_text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


def infer_unit_m_from_product_name(product_name, fallback_unit):
    """
    製品名文字列から加工単位(m)を推定する暫定ルール。
    例: 15020-JX5R- 770X300F-A   R -> 300
    ※ バリエーションが多い前提のため、ここを都度調整できるよう関数化している。
    """
    if product_name is None or pd.isna(product_name):
        return fallback_unit
    s = str(product_name)
    # "770X300..." のようなパターンから X の後の数値を拾う（最後に見つかったXを優先）
    matches = re.findall(r"[xX]\s*(\d{2,6})", s)
    if matches:
        try:
            v = int(matches[-1])
            if v > 0:
                return v
        except ValueError:
            pass
    return fallback_unit

def load_tasks_df():
    """
    タスク入力を取得する（tasks.xlsx は使用しない）。
    必須: 環境変数 TASK_INPUT_WORKBOOK にマクロ実行ブックのフルパス（VBA が設定）
         シート「加工計画DATA」を読み込む（投入目安は「回答納期」、未入力時は「指定納期」）。
    """
    if not TASKS_INPUT_WORKBOOK:
        raise FileNotFoundError(
            "TASK_INPUT_WORKBOOK が未設定です。VBA の RunPython でマクロ実行ブックのパスを渡してください。"
        )
    if not os.path.exists(TASKS_INPUT_WORKBOOK):
        raise FileNotFoundError(f"TASK_INPUT_WORKBOOK が存在しません: {TASKS_INPUT_WORKBOOK}")
    df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=TASKS_SHEET_NAME)
    df.columns = df.columns.str.strip()
    logging.info(f"タスク入力: '{TASKS_INPUT_WORKBOOK}' の '{TASKS_SHEET_NAME}' を読み込みました。")
    return df


def _nfkc_column_aliases(canonical_name):
    """見出しの表記ゆれ（全角記号・互換文字）を吸収するための比較キー。"""
    return unicodedata.normalize("NFKC", str(canonical_name).strip())


def _rename_legacy_plan_input_ref_columns(df):
    """旧参照列見出し「（元）必要OP人数」を現行「（元）必要人数」へ寄せる。"""
    if df is None or df.empty:
        return df
    legacy = unicodedata.normalize("NFKC", "（元）必要OP人数")
    renames = {}
    for c in df.columns:
        if _nfkc_column_aliases(c) == legacy:
            renames[c] = PLAN_REF_REQUIRED_OP
    if renames:
        df = df.rename(columns=renames)
    return df


def _align_dataframe_headers_to_canonical(df, canonical_names):
    """列名を NFKC 一致で canonical に寄せる（Excel 側が全角 '_' 等でも読めるように）。"""
    key_to_canonical = {_nfkc_column_aliases(c): c for c in canonical_names}
    rename_map = {}
    for col in df.columns:
        k = _nfkc_column_aliases(col)
        if k in key_to_canonical:
            target = key_to_canonical[k]
            if col != target:
                rename_map[col] = target
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def _normalize_equipment_match_key(val):
    """
    工程名（設備名）の照合用キー。
    NFKC・前後空白・連続空白・NBSP/全角スペース・ゼロ幅文字を正規化する。
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    t = unicodedata.normalize("NFKC", str(val))
    t = t.replace("\u00a0", " ").replace("\u3000", " ")
    t = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _equipment_lookup_normalized_to_canonical(equipment_list):
    """正規化キー → master スキルシート上の列名（canonical 表記）。"""
    lookup = {}
    for eq in equipment_list:
        k = _normalize_equipment_match_key(eq)
        if k and k not in lookup:
            lookup[k] = eq
    # 工程名のみの照合（加工実績DATA等）: 同一工程の先頭列（工程+機械）へ寄せる
    for eq in equipment_list:
        s = str(eq).strip()
        if "+" not in s:
            continue
        p, _rest = s.split("+", 1)
        pk = _normalize_equipment_match_key(p)
        if pk and pk not in lookup:
            lookup[pk] = eq
    return lookup


def _equipment_schedule_header_labels(equipment_list: list) -> list:
    """
    結果_設備毎の時間割・結果_設備ガントの行／列見出し用。
    内部キーが「工程+機械」のときは機械名を表示し、機械名の重複時のみ工程を括弧で補う。
    """
    raw = []
    for eq in equipment_list:
        s = str(eq).strip()
        if "+" in s:
            mpart = s.split("+", 1)[1].strip()
            raw.append(mpart if mpart else s)
        else:
            raw.append(s)
    counts = {}
    for r in raw:
        counts[r] = counts.get(r, 0) + 1
    out = []
    for eq, r in zip(equipment_list, raw):
        if counts.get(r, 0) > 1:
            s = str(eq).strip()
            if "+" in s:
                p = s.split("+", 1)[0].strip()
                out.append(f"{r}（{p}）" if p else r)
            else:
                out.append(r)
        else:
            out.append(r)
    return out


def _resolve_equipment_line_key_for_task(task: dict, equipment_list: list | None) -> str:
    """
    設備時間割・設備専有空きの列キー（skills / need と同じ「工程+機械」を基本とする）。
    機械名が空でマスタに当該工程の列が1つだけならその複合キーへ寄せる。
    """
    p = str(task.get("machine") or "").strip()
    mn = str(task.get("machine_name") or "").strip()
    cand = f"{p}+{mn}" if (p and mn) else (p or mn)
    elist = [str(x).strip() for x in (equipment_list or []) if str(x).strip()]
    if cand in elist:
        return cand
    if mn:
        return cand
    if not p:
        return cand
    exact_p = [x for x in elist if x == p]
    if len(exact_p) == 1:
        return exact_p[0]
    prefixed = [x for x in elist if x.startswith(p + "+")]
    if len(prefixed) == 1:
        return prefixed[0]
    return p


def load_planning_tasks_df():
    """
    2段階目用: マクロブック上の「配台計画_タスク入力」シートを読み込む。

    「担当OP_指定」列または特別指定備考の AI 出力 preferred_operator で主担当 OP を指名できる（skills のメンバー名とあいまい一致）。
    メイン「再優先特別記載」の task_preferred_operators は generate_plan 側で最優先マージされる。
    「配台不要」がオン（TRUE/1/はい 等）の行は配台対象外。
    読み込み後、同一依頼NO・重複機械名があるグループの工程「分割」行へ空なら「配台不要」=yes（段階1と同じ）。
    「設定_配台不要工程」で工程+機械の組を同期し、C/D/E に基づき配台不要を反映する（シート作成は VBA）。
    """
    if not TASKS_INPUT_WORKBOOK:
        raise FileNotFoundError(
            "TASK_INPUT_WORKBOOK が未設定です。VBA の RunPython でマクロ実行ブックのパスを渡してください。"
        )
    if not os.path.exists(TASKS_INPUT_WORKBOOK):
        raise FileNotFoundError(f"TASK_INPUT_WORKBOOK が存在しません: {TASKS_INPUT_WORKBOOK}")
    df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=PLAN_INPUT_SHEET_NAME)
    df.columns = df.columns.str.strip()
    df = _rename_legacy_plan_input_ref_columns(df)
    df = _align_dataframe_headers_to_canonical(
        df, plan_input_sheet_column_order()
    )
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""
    try:
        _pairs_lr = []
        _seen_lr = set()
        for _, _row_lr in df.iterrows():
            _p = str(_row_lr.get(TASK_COL_MACHINE, "") or "").strip()
            _m = str(_row_lr.get(TASK_COL_MACHINE_NAME, "") or "").strip()
            if not _p:
                continue
            _k = (
                _normalize_process_name_for_rule_match(_p),
                _normalize_equipment_match_key(_m),
            )
            if _k in _seen_lr:
                continue
            _seen_lr.add(_k)
            _pairs_lr.append((_p, _m))
        run_exclude_rules_sheet_maintenance(TASKS_INPUT_WORKBOOK, _pairs_lr, "配台シート読込")
    except Exception:
        logging.exception("配台シート読込: 設定_配台不要工程の保守で例外（続行）")
    try:
        _apply_auto_exclude_bunkatsu_duplicate_machine(df, log_prefix="配台シート読込")
    except Exception as ex:
        logging.warning("配台シート読込: 分割行の配台不要自動設定で例外（続行）: %s", ex)
    try:
        df = apply_exclude_rules_config_to_plan_df(df, TASKS_INPUT_WORKBOOK, "配台シート読込")
    except Exception as ex:
        logging.warning("配台シート読込: 設定シートによる配台不要適用で例外（続行）: %s", ex)
    logging.info(
        f"計画タスク入力: '{TASKS_INPUT_WORKBOOK}' の '{PLAN_INPUT_SHEET_NAME}' を読み込みました。"
    )
    return df


def _main_sheet_cell_is_global_comment_label(val) -> bool:
    """メインシート上「グローバルコメント」見出しセルか（表記ゆれ許容）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    s = unicodedata.normalize("NFKC", str(val).strip())
    if not s:
        return False
    if _nfkc_column_aliases(s) == _nfkc_column_aliases("グローバルコメント"):
        return True
    if "グローバル" in s and "コメント" in s:
        return True
    return False


def load_main_sheet_global_priority_override_text() -> str:
    """
    TASK_INPUT_WORKBOOK のメインシートで「グローバルコメント」と書かれたセルの **直下** を読む。
    シート名: 「メイン」「Main」、または名前に「メイン」を含む（VBA GetMainWorksheet と同趣旨）。

    内容は **Gemini で一括解釈**（`analyze_global_priority_override_comment`）。工場休業日・再優先フラグ・未実装指示のメモを JSON 化する。
    API キーが無い場合のみ、工場休業日はルールベースの `parse_factory_closure_dates_from_global_comment` で補完する。
    """
    wb_path = TASKS_INPUT_WORKBOOK.strip() if TASKS_INPUT_WORKBOOK else ""
    if not wb_path or not os.path.exists(wb_path):
        return ""
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "メイン再優先特記: ブックに「%s」があるため openpyxl でグローバルコメントを読みません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return ""
    try:
        wb = load_workbook(wb_path, data_only=True, read_only=False)
    except Exception as e:
        logging.warning("メイン再優先特記: ブックを開けませんでした: %s", e)
        return ""
    try:
        ws = None
        for name in ("メイン", "Main"):
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            for sn in wb.sheetnames:
                if "メイン" in sn:
                    ws = wb[sn]
                    break
        if ws is None:
            return ""
        max_r = min(ws.max_row or 0, 400)
        max_c = min(ws.max_column or 0, 40)
        if max_r < 1 or max_c < 1:
            return ""
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if not _main_sheet_cell_is_global_comment_label(cell.value):
                    continue
                below = ws.cell(row=r + 1, column=c).value
                if below is None or (isinstance(below, float) and pd.isna(below)):
                    return ""
                return str(below).strip()
        return ""
    finally:
        pass


def _global_comment_chunk_implies_factory_closure(chunk: str) -> bool:
    """
    メイン「グローバルコメント」の断片が、工場単位の休業・非稼働を意味するか（個人休みだけを誤検出しない）。
    """
    c = unicodedata.normalize("NFKC", str(chunk or ""))
    if not c.strip():
        return False
    if re.search(r"臨時\s*休業", c):
        return True
    if "休場" in c:
        return True
    if re.search(r"工場", c) and re.search(r"休|休業|休み|停止|お休み", c):
        return True
    if re.search(r"(?:全社|全館|全工場).{0,15}(?:休|休業|停止)", c):
        return True
    if re.search(r"(?:稼働|生産|ライン).{0,12}(?:停止|なし|無し)", c):
        return True
    if re.search(r"加工.{0,15}(?:しない|無し|なし|お休み)", c):
        return True
    if "休業" in c and re.search(
        r"(?:工場|全社|本社|当日|弊社|当社|全員|社全体)", c
    ):
        return True
    return False


def _md_slash_is_likely_fraction_not_date(t: str, start: int, end: int, mo: int, day: int) -> bool:
    """
    「加工速度は1/3とします」の 1/3 を 1月3日 と誤認しない。
    「4/1は工場を休み」の 4/1 は日付のまま（直後が「は」なら分数扱いにしない）。
    """
    if mo <= 0 or day <= 0:
        return True
    before = t[max(0, start - 32) : start]
    after = t[end : min(len(t), end + 14)]
    after_st = after.lstrip()
    if after_st.startswith("は"):
        return False
    if re.search(
        r"(?:加工速度|加工\s*スピード|速度|倍率|スピード|効率|割合)(?:\s*は)?\s*$",
        before,
    ):
        return True
    # 1/2・1/3・2/3 等 + 「とします」「倍」… は分数・比率寄り（「3/1です」等の日付を誤スキップしないよう です/である は含めない）
    frac_pat = re.compile(
        r"^(?:とします?|とする|倍|割合|にする|に設定|くらい|程度|に固定|に変更)"
    )
    if mo <= 12 and day <= 12 and frac_pat.match(after_st):
        if mo <= 2 or (mo == 3 and day <= 3):
            return True
    # 「1/2です」「1/10です」のような分母表現（先頭が 1/ のみ）
    if (
        mo == 1
        and 2 <= day <= 12
        and re.match(r"^です|である\b", after_st)
    ):
        return True
    return False


def _extract_calendar_dates_from_text(s: str, default_year: int) -> list[date]:
    """グローバルコメント内の日付表記を date に変換（基準年は計画の基準年）。"""
    t = unicodedata.normalize("NFKC", str(s or ""))
    found: list[date] = []
    seen: set[date] = set()

    def add(y: int, mo: int, d: int) -> None:
        try:
            dd = date(y, mo, d)
        except ValueError:
            return
        if dd not in seen:
            seen.add(dd)
            found.append(dd)

    for m in re.finditer(
        r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?",
        t,
    ):
        add(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for m in re.finditer(
        r"(\d{4})\s*[/\-\.／]\s*(\d{1,2})\s*[/\-\.／]\s*(\d{1,2})",
        t,
    ):
        add(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for m in re.finditer(r"(\d{1,2})\s*月\s*(\d{1,2})\s*日", t):
        add(int(default_year), int(m.group(1)), int(m.group(2)))
    for m in re.finditer(
        r"(?<!\d)(\d{1,2})\s*[/／]\s*(\d{1,2})(?!\d)",
        t,
    ):
        mo_i, d_i = int(m.group(1)), int(m.group(2))
        if _md_slash_is_likely_fraction_not_date(t, m.start(), m.end(), mo_i, d_i):
            continue
        add(int(default_year), mo_i, d_i)
    return found


def _split_global_comment_into_chunks(blob: str) -> list[str]:
    """
    グローバルコメントを「独立した指示」の塊に分ける。
    改行（Excel の Alt+Enter・Unicode 改行含む）で必ず分割し、同一行内は 。;； で続けて分割。
    """
    t = unicodedata.normalize("NFKC", str(blob or "").strip())
    if not t:
        return []
    lines = [ln.strip() for ln in re.split(r"[\n\r\v\f\u2028\u2029]+", t) if ln.strip()]
    if not lines:
        return []
    chunks: list[str] = []
    for line in lines:
        subs = [c.strip() for c in re.split(r"[。;；]+", line) if c.strip()]
        if subs:
            chunks.extend(subs)
        else:
            chunks.append(line)
    return chunks


def parse_factory_closure_dates_from_global_comment(
    text: str, default_year: int
) -> set[date]:
    """
    メインシート「グローバルコメント」に、工場臨時休業などと日付が書かれている場合に
    その日を工場休み（全員非稼働・配台で加工しない）として扱う日付集合を返す。
    """
    blob = unicodedata.normalize("NFKC", str(text or "").strip())
    if not blob:
        return set()
    chunks = _split_global_comment_into_chunks(blob)
    if not chunks:
        chunks = [blob]
    out: set[date] = set()
    y0 = int(default_year)
    for ch in chunks:
        if not _global_comment_chunk_implies_factory_closure(ch):
            continue
        for d in _extract_calendar_dates_from_text(ch, y0):
            out.add(d)
    if not out and _global_comment_chunk_implies_factory_closure(blob):
        for d in _extract_calendar_dates_from_text(blob, y0):
            out.add(d)
    return out


def apply_factory_closure_dates_to_attendance(
    attendance_data: dict, members: list, closure_dates: set[date]
) -> None:
    """工場休業日: 勤怠上は全員 is_working=False とし、その日は設備割付を行わない。"""
    if not closure_dates or not attendance_data:
        return
    tag = "工場休業（メイン・グローバルコメント）"
    for d in sorted(closure_dates):
        if d not in attendance_data:
            logging.warning(
                "グローバルコメントの工場休業日 %s はマスタ勤怠に行がありません。"
                " その日は計画ループに含まれない場合、配台上の効果が限定的です。",
                d,
            )
            continue
        day = attendance_data[d]
        for m in members:
            if m not in day:
                continue
            ent = day[m]
            ent["is_working"] = False
            ent["eligible_for_assignment"] = False
            prev = str(ent.get("reason") or "").strip()
            ent["reason"] = f"{tag} {prev}".strip() if prev else tag


def _apply_global_priority_abolish_heuristic(blob: str, coerced: dict) -> dict:
    """
    「制限撤廃」「あらゆる条件」等: 設備専有・時刻ガードまで含め配台制約を緩める（abolish_all_scheduling_limits）。
    """
    b = unicodedata.normalize("NFKC", str(blob or ""))
    strong = (
        "制限撤廃",
        "制限を撤廃",
        "すべての制限",
        "全ての制限",
        "あらゆる制限",
        "あらゆる条件",
        "すべての条件",
        "全ての条件",
        "撤廃して",
        "撤廃し",
    )
    if any(k in b for k in strong):
        out = dict(coerced)
        out["abolish_all_scheduling_limits"] = True
        out["ignore_skill_requirements"] = True
        out["ignore_need_minimum"] = True
        logging.warning(
            "メイン再優先特記: 制限撤廃キーワードを検出。設備専有・時刻ガードを含め配台上の制約を緩めます。"
        )
        return out
    return coerced


def _maybe_fill_global_speed_rules_from_scheduler_notes(coerced: dict) -> dict:
    """
    AI が global_speed_rules を空にしたが scheduler_notes に具体パターンがある場合の補完。
    広く推測しない（熱融着＋検査＋1/3 系のみ）。
    """
    if not isinstance(coerced, dict):
        return coerced
    if coerced.get("global_speed_rules"):
        return coerced
    sn = str(coerced.get("scheduler_notes_ja") or "")
    t = unicodedata.normalize("NFKC", sn)
    if "熱融着" not in t or "検査" not in t:
        return coerced
    if not re.search(r"(?:1\s*/\s*3|１\s*/\s*3|三分の一|3\s*分の\s*1)", t):
        return coerced
    out = dict(coerced)
    out["global_speed_rules"] = [
        {
            "process_contains": "熱融着",
            "machine_contains": "検査",
            "speed_multiplier": 1.0 / 3.0,
        }
    ]
    logging.info(
        "メイン再優先特記: scheduler_notes_ja から global_speed_rules を補完（熱融着・検査・1/3）"
    )
    return out


def _finalize_global_priority_override(blob: str, coerced: dict) -> dict:
    """ソロ補正の後、abolish が true ならスキル・人数も強制オン。"""
    coerced = _maybe_fill_global_speed_rules_from_scheduler_notes(dict(coerced))
    coerced = _apply_global_priority_solo_heuristic(blob, coerced)
    coerced = _apply_global_priority_abolish_heuristic(blob, coerced)
    if coerced.get("abolish_all_scheduling_limits"):
        out = dict(coerced)
        out["ignore_skill_requirements"] = True
        out["ignore_need_minimum"] = True
        return out
    return coerced


def _apply_global_priority_solo_heuristic(blob: str, coerced: dict) -> dict:
    """
    「一人で担当」「単独」等で人数だけ緩めても、指名メンバーがスキル非該当だと配台されない。
    その場合はスキル無視を同時に立てる。
    """
    if not coerced.get("ignore_need_minimum") or coerced.get("ignore_skill_requirements"):
        return coerced
    b = unicodedata.normalize("NFKC", str(blob or ""))
    solo_kw = ("一人", "ひとり", "単独", "１人", "1人", "独自", "単身")
    if any(k in b for k in solo_kw):
        out = dict(coerced)
        out["ignore_skill_requirements"] = True
        logging.info(
            "メイン再優先特記: 単独系キーワードのため ignore_skill_requirements を補助的に true にしました。"
        )
        return out
    return coerced


def _coerce_task_preferred_operators_dict(raw_val) -> dict:
    """AI の task_preferred_operators を {依頼NO: 氏名} に正規化。"""
    out = {}
    if not isinstance(raw_val, dict):
        return out
    for k, v in raw_val.items():
        ks = str(k).strip()
        if not ks:
            continue
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        vs = str(v).strip()
        if vs and vs.lower() not in ("nan", "none", "null"):
            out[ks] = vs
    return out


def _normalize_factory_closure_dates_iso_list(val, default_year: int) -> list[str]:
    """
    AI またはフォールバックの日付リストを YYYY-MM-DD 文字列の昇順ユニークに正規化。
    要素は ISO 文字列・Excel 日付・「4/1」程度の短文でも可。
    """
    y0 = int(default_year)
    seen: set[str] = set()
    out: list[str] = []
    if not isinstance(val, list):
        return out
    for item in val:
        if item is None or (isinstance(item, float) and pd.isna(item)):
            continue
        d = parse_optional_date(item)
        if d is not None:
            iso = d.isoformat()
            if iso not in seen:
                seen.add(iso)
                out.append(iso)
            continue
        s = unicodedata.normalize("NFKC", str(item).strip())
        if not s:
            continue
        for d2 in _extract_calendar_dates_from_text(s, y0):
            iso = d2.isoformat()
            if iso not in seen:
                seen.add(iso)
                out.append(iso)
    return sorted(out)


def _coerce_global_speed_rules(raw_val) -> list[dict]:
    """
    Gemini の global_speed_rules を正規化。
    各要素: process_contains / machine_contains（いずれか必須・部分一致用）, speed_multiplier（既存速度に乗算、0超〜10以下）。
    """
    out: list[dict] = []
    if not isinstance(raw_val, list):
        return out
    for item in raw_val:
        if not isinstance(item, dict):
            continue
        sm = item.get("speed_multiplier")
        if sm is None:
            sm = item.get("relative_speed")
        try:
            mult = float(sm)
        except (TypeError, ValueError):
            continue
        if mult <= 0 or mult > 10.0:
            continue
        pps = unicodedata.normalize("NFKC", str(item.get("process_contains") or "")).strip()
        mms = unicodedata.normalize("NFKC", str(item.get("machine_contains") or "")).strip()
        if not pps and not mms:
            continue
        out.append(
            {
                "process_contains": pps,
                "machine_contains": mms,
                "speed_multiplier": mult,
            }
        )
    return out


def _global_speed_rule_substring_matches_row(pnorm: str, mnorm: str, sub_nfkc: str) -> bool:
    """sub が空でなければ、工程名または機械名のいずれかに部分一致すれば True。"""
    if not sub_nfkc:
        return True
    return sub_nfkc in pnorm or sub_nfkc in mnorm


def _global_speed_multiplier_for_row(process_name: str, machine_name: str, rules: list) -> float:
    """
    工程名・機械名に一致するルールの speed_multiplier を掛け合わせる（一致なしは 1.0）。

    process_contains / machine_contains はそれぞれ **工程名または機械名のどちらか** に含まれればよい。
    両方指定時は AND（例: 「熱融着」と「検査」が、列の組み合わせで両方現れる行にマッチ。
    マスタ上で工程=検査・機械=熱融着機 のようにキーワードが逆側の列にあっても同じルールで効く。
    """
    if not rules:
        return 1.0
    pnorm = unicodedata.normalize("NFKC", str(process_name or "")).strip()
    mnorm = unicodedata.normalize("NFKC", str(machine_name or "")).strip()
    combined = 1.0
    for r in rules:
        if not isinstance(r, dict):
            continue
        pc = unicodedata.normalize(
            "NFKC", str(r.get("process_contains") or "").strip()
        )
        mc = unicodedata.normalize(
            "NFKC", str(r.get("machine_contains") or "").strip()
        )
        if not pc and not mc:
            continue
        if pc and not _global_speed_rule_substring_matches_row(pnorm, mnorm, pc):
            continue
        if mc and not _global_speed_rule_substring_matches_row(pnorm, mnorm, mc):
            continue
        try:
            m = float(r.get("speed_multiplier", 1.0))
        except (TypeError, ValueError):
            continue
        if m <= 0:
            continue
        combined *= m
    if combined <= 0:
        return 1.0
    return combined


def _coerce_task_planning_basis_dates(raw) -> dict[str, str]:
    """グローバルコメント AI の 依頼NO→計画基準納期（YYYY-MM-DD 文字列）。"""
    out: dict[str, str] = {}
    if not isinstance(raw, dict):
        return out
    for k, v in raw.items():
        ks = unicodedata.normalize("NFKC", str(k).strip())
        if not ks or ks.lower() in ("nan", "none", "null"):
            continue
        d = parse_optional_date(v)
        if d is not None:
            out[ks] = d.isoformat()
    return out


def _coerce_global_priority_override_dict(raw, reference_year: int | None = None) -> dict:
    """Gemini 戻りを配台用フラグ・工場休業日リストに正規化。"""
    y0 = int(reference_year) if reference_year is not None else date.today().year

    def as_bool(v):
        if v is True:
            return True
        if v is False:
            return False
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return False
        s = unicodedata.normalize("NFKC", str(v).strip()).lower()
        return s in ("true", "1", "yes", "はい", "on")

    base = {
        "ignore_skill_requirements": False,
        "ignore_need_minimum": False,
        "abolish_all_scheduling_limits": False,
        "task_preferred_operators": {},
        "interpretation_ja": "",
        "factory_closure_dates": [],
        "scheduler_notes_ja": "",
        "global_speed_rules": [],
        "task_planning_basis_dates": {},
    }
    if not isinstance(raw, dict):
        return base
    base["ignore_skill_requirements"] = as_bool(raw.get("ignore_skill_requirements"))
    base["ignore_need_minimum"] = as_bool(raw.get("ignore_need_minimum"))
    base["abolish_all_scheduling_limits"] = as_bool(
        raw.get("abolish_all_scheduling_limits")
    )
    base["task_preferred_operators"] = _coerce_task_preferred_operators_dict(
        raw.get("task_preferred_operators")
    )
    ij = raw.get("interpretation_ja")
    if ij is not None and not (isinstance(ij, float) and pd.isna(ij)):
        base["interpretation_ja"] = str(ij).strip()[:800]
    base["factory_closure_dates"] = _normalize_factory_closure_dates_iso_list(
        raw.get("factory_closure_dates"), y0
    )
    sn = raw.get("scheduler_notes_ja")
    if sn is not None and not (isinstance(sn, float) and pd.isna(sn)):
        base["scheduler_notes_ja"] = str(sn).strip()[:600]
    base["global_speed_rules"] = _coerce_global_speed_rules(raw.get("global_speed_rules"))
    base["task_planning_basis_dates"] = _coerce_task_planning_basis_dates(
        raw.get("task_planning_basis_dates")
    )
    return base


def _parse_global_priority_override_gemini_response(res):
    """Gemini 応答から JSON オブジェクト1つを取り出す（```json フェンス付きでも可）。"""
    raw = (_gemini_result_text(res) or "").strip()
    if not raw:
        return None
    candidate = None
    fence = re.search(
        r"```(?:json)?\s*(\{.*\})\s*```",
        raw,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        candidate = fence.group(1).strip()
    elif raw.startswith("{"):
        candidate = raw
    else:
        loose = re.search(r"\{.*\}", raw, re.DOTALL)
        candidate = loose.group(0).strip() if loose else None
    if not candidate:
        return None
    try:
        parsed = json.loads(candidate)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _apply_regex_factory_closure_fallback(coerced: dict, blob: str, ref_y: int) -> dict:
    """Gemini 未使用・応答解釈失敗時: ルールベースで工場休業日だけ補完（従来互換）。"""
    out = dict(coerced)
    rx = parse_factory_closure_dates_from_global_comment(blob, ref_y)
    out["factory_closure_dates"] = sorted({d.isoformat() for d in rx})
    return out


def analyze_global_priority_override_comment(
    text: str, members: list, reference_year: int, ai_sheet_sink: dict | None = None
) -> dict:
    """
    メインシート「グローバルコメント」（UI 上の自由記述）を **Gemini で一括解釈**し、配台に効く JSON に落とす。
    自然言語の文脈切り分け・改行の別指示解釈は AI に任せ、戻り値のキーだけシステムが機械適用する。

    - factory_closure_dates: **工場全体**で稼働しない日（全員非稼働扱い）の YYYY-MM-DD 文字列の配列。該当なしは []。
    - ignore_skill_requirements / ignore_need_minimum / abolish_all_scheduling_limits / task_preferred_operators: 従来どおり。
    - global_speed_rules: **工程名・機械名**への部分一致（各キーワードは **どちらの列にあっても可**）で、既存の加工速度（シート／上書き後）に **乗算**するルールの配列。該当なしは []。
    - task_planning_basis_dates: 依頼NO→計画基準納期（YYYY-MM-DD）。配台の納期基準として **特別指定_備考AIより優先**して適用する。該当なしは {{}}。
    - scheduler_notes_ja: 上記に落としきれない補足や運用メモ（速度は可能なら global_speed_rules も併記）。

    API キー無し・JSON 解釈失敗時: 上記ブール・指名は既定値、工場休業日のみ従来のルールベース解析で補完。
    """
    ref_y = int(reference_year) if reference_year is not None else date.today().year
    empty = _coerce_global_priority_override_dict({}, ref_y)
    if not text or not str(text).strip():
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "スキップ（メイン原文なし）"
        return empty
    blob = str(text).strip()
    mem_sig = ",".join(sorted(str(m).strip() for m in (members or []) if m))
    cache_fingerprint = f"{GLOBAL_PRIORITY_OVERRIDE_CACHE_PREFIX}{ref_y}\n{blob}\n{mem_sig}"
    cache_key = hashlib.sha256(cache_fingerprint.encode("utf-8")).hexdigest()
    ai_cache = load_ai_cache()
    cached = get_cached_ai_result(ai_cache, cache_key, content_key=cache_fingerprint)
    if cached is not None:
        logging.info("メイン再優先特記: キャッシュヒット（Gemini は呼びません）。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "なし（キャッシュ使用）"
        return _finalize_global_priority_override(
            blob, _coerce_global_priority_override_dict(cached, ref_y)
        )

    if not API_KEY:
        logging.info("GEMINI_API_KEY 未設定のためメイン再優先特記の AI 解析をスキップしました。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "なし（APIキー未設定・工場休業のみルール補完）"
        coerced = _apply_regex_factory_closure_fallback(
            _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
        )
        return _finalize_global_priority_override(blob, coerced)

    member_sample = ", ".join(str(m) for m in (members or [])[:80])
    if len(members or []) > 80:
        member_sample += " …"

    prompt = f"""あなたは工場の配台計画システム用アシスタントです。
Excel メインシートの **「グローバルコメント」**（自由記述・自然言語）の **全文** を読み、次のキーだけを持つ JSON を1つ返してください。

【役割】
ユーザーは改行や句点で複数の指示を書くことがあります。**文脈を読み分け**、配台システムが **機械的に適用できる値** に落とし込んでください。
推測でブールを true にしないこと。根拠が明確なときだけ true。

【最優先】
この欄の内容はマスタ・スキル・need・タスク行・特別指定_備考の AI 指名より優先される例外指示として扱われます。

【改行・複数行】
各行・各文は **原則として独立した指示** です。行をまたいで1つにまとめたり、**割合表現（例 1/3）を日付と結び付けたりしない**こと。

【キー別ルール】

A) **factory_closure_dates** （配列・必須）
   - **工場全体**が稼働しない日（臨時休業・全工場休み・その日は加工しない等）の日付を **YYYY-MM-DD** の文字列で列挙。
   - **個人の休み・特定ラインだけ**の停止はここに **含めない**（[]）。
   - 該当がなければ **空配列 []**（キー省略不可）。
   - 年が省略されていれば西暦 {ref_y} 年として解釈。

B) **ignore_skill_requirements** / **ignore_need_minimum** / **abolish_all_scheduling_limits** / **task_preferred_operators**
   - 従来どおり（配台のスキル無視・人数1固定・制限撤廃・依頼NO→主担当OP指名）。該当なければ false または {{}}。

C) **global_speed_rules** （配列・必須）
   - 特定の **工程名**（Excel「工程名」列）や **機械名**（「機械名」列）に対し、**既存の加工速度に掛ける倍率** を指定するオブジェクトのリスト。
   - 各オブジェクトのキー:
     - "process_contains": 文字列（省略可）。**工程名または機械名のいずれか**に **部分一致**（NFKC 想定）。
     - "machine_contains": 文字列（省略可）。**工程名または機械名のいずれか**に **部分一致**。
     - "speed_multiplier": 正の数。**1/3 の速度**なら約 **0.333333**（既存速度 × この値）。**2倍速**なら 2.0。
   - **両方指定時は AND**（2つのキーワードが、**両方とも**「工程名・機械名のどちらか」に現れる行）。例: 工程=検査・機械=熱融着機 でも、工程=熱融着・機械=検査用設備 でもマッチしうる。
   - どちらか一方だけ指定すれば、そのキーワードが工程名または機械名のどちらかにあればマッチ。
   - 該当指示がなければ **空配列 []**。
   - 例: 「熱融着を使う検査の加工速度は1/3」→
     [{{"process_contains":"熱融着","machine_contains":"検査","speed_multiplier":0.333333}}]
     （「熱融着」と「検査」が工程名・機械名の組み合わせで揃うタスクの速度が約1/3になる）

D) **scheduler_notes_ja** （文字列・必須）
   - 上記キーに落としきれない補足。速度は **global_speed_rules で構造化できるときは必ずそちらにも出す**（ここは人間向け要約でもよい）。無ければ ""。

E) **interpretation_ja** （文字列・必須）
   - 原文の要約を1文（200文字以内）。

F) **task_planning_basis_dates** （オブジェクト・必須）
   - グローバルコメントから読み取れる **依頼NO（Excel「依頼NO」列と同一表記）** ごとの **計画基準納期** を **YYYY-MM-DD** の文字列で格納。
   - 例: {{"Y3-12": "2026-04-10", "W1-5": "2026-04-15"}}
   - 依頼NOごとの納期指示が無ければ **空オブジェクト {{}}**（キー省略不可）。

【返答形式】
先頭が {{ で終わりが }} の **JSON オブジェクト1つのみ**（説明文・マークダウン禁止）。

必須キー一覧:
- "factory_closure_dates": string の配列（YYYY-MM-DD）
- "ignore_skill_requirements": true または false
- "ignore_need_minimum": true または false
- "abolish_all_scheduling_limits": true または false
- "task_preferred_operators": オブジェクト（該当なしは {{}}）
- "global_speed_rules": オブジェクトの配列（該当なしは []）
- "task_planning_basis_dates": オブジェクト（依頼NO→YYYY-MM-DD、該当なしは {{}}）
- "scheduler_notes_ja": 文字列
- "interpretation_ja": 文字列

【基準年】 日付言及があれば西暦 {ref_y} 年として解釈してよい。

【登録メンバー名の参考】（照合用。JSON キーには含めない）
{member_sample}

【グローバルコメント・原文】
{blob}
"""
    try:
        ppath = os.path.join(log_dir, "ai_global_priority_override_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("メイン再優先特記: プロンプト全文 → %s", ppath)
    except OSError as ex:
        logging.warning("メイン再優先特記: プロンプト保存失敗: %s", ex)

    client = genai.Client(api_key=API_KEY)
    try:
        res = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
        record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
        parsed = _parse_global_priority_override_gemini_response(res)
        if parsed is None:
            logging.warning(
                "メイン再優先特記: AI 応答から JSON を解釈できませんでした。キャッシュせず、次回再試行されます。"
            )
            try:
                rpath = os.path.join(log_dir, "ai_global_priority_override_last_response.txt")
                with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                    rf.write(_gemini_result_text(res) or "")
            except OSError:
                pass
            if ai_sheet_sink is not None:
                ai_sheet_sink["メイン再優先特記_AI_API"] = "あり（JSON解釈失敗・工場休業はルール補完）"
            coerced = _apply_regex_factory_closure_fallback(
                _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
            )
            return _finalize_global_priority_override(blob, coerced)
        coerced = _coerce_global_priority_override_dict(parsed, ref_y)
        coerced = _finalize_global_priority_override(blob, coerced)
        try:
            rpath = os.path.join(log_dir, "ai_global_priority_override_last_response.txt")
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(_gemini_result_text(res) or "")
        except OSError:
            pass
        put_cached_ai_result(ai_cache, cache_key, coerced, content_key=cache_fingerprint)
        save_ai_cache(ai_cache)
        _tpo = coerced.get("task_preferred_operators") or {}
        _fcd = coerced.get("factory_closure_dates") or []
        _gsr = coerced.get("global_speed_rules") or []
        logging.info(
            "メイン再優先特記: AI 解釈 factory休業=%s日 速度ルール=%s件 skill=%s need1=%s abolish=%s task_pref=%s件 — %s",
            len(_fcd),
            len(_gsr),
            coerced["ignore_skill_requirements"],
            coerced["ignore_need_minimum"],
            coerced.get("abolish_all_scheduling_limits"),
            len(_tpo),
            coerced.get("interpretation_ja", "")[:100],
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "あり"
        return coerced
    except Exception as e:
        logging.warning("メイン再優先特記: Gemini 呼び出し失敗: %s", e)
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = f"失敗: {e}"[:500]
        coerced = _apply_regex_factory_closure_fallback(
            _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
        )
        return _finalize_global_priority_override(blob, coerced)


def default_result_task_sheet_column_order(max_history_len: int) -> list:
    """結果_タスク一覧の既定列順（履歴列数は実行時に決まる）。"""
    hist = [f"履歴{i+1}" for i in range(max_history_len)]
    return [
        "ステータス",
        "タスクID",
        "工程名",
        "機械名",
        "優先度",
        RESULT_TASK_COL_DISPATCH_TRIAL_ORDER,
        *hist,
        "必要OP(上書)",
        "タスク効率",
        "加工途中",
        "特別指定あり",
        "担当OP指名",
        "回答納期",
        "指定納期",
        "計画基準納期",
        TASK_COL_RAW_INPUT_DATE,
        "納期緊急",
        "加工開始日",
        "配完_加工開始",
        "配完_加工終了",
        RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16,
        "総加工量",
        "残加工量",
        "完了率(実行時点)",
        "特別指定_AI",
    ]


def _task_date_key_for_result_sheet_sort(val):
    """結果_タスク一覧の並べ替え用。欠損・解釈不能は最後（date.max）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return date.max
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        ts = pd.Timestamp(val)
        if pd.isna(ts):
            return date.max
        return ts.date()
    except Exception:
        return date.max


def _coerce_planning_date_for_deadline(d) -> date | None:
    """回答納期・指定納期などを date に正規化（欠損は None）。"""
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return None


def _result_task_plan_end_within_answer_or_spec_16_label(
    plan_window: list | None, answer_due, specified_due
) -> str:
    """
    結果_タスク一覧用: 「配完_加工終了」相当の最終終了が、
    回答納期の日付 + PLAN_DUE_DAY_COMPLETION_TIME（既定 16:00）以下かを判定。
    回答納期が無い行は指定納期の日付 + 16:00 で判定。
    両方無い場合は「納期なし」。
    """
    if not plan_window or len(plan_window) < 2:
        return "未割当"
    _pe = plan_window[1]
    if _pe is None:
        return "未割当"
    dd = _coerce_planning_date_for_deadline(answer_due)
    if dd is None:
        dd = _coerce_planning_date_for_deadline(specified_due)
    if dd is None:
        return "納期なし"
    try:
        deadline_dt = datetime.combine(dd, PLAN_DUE_DAY_COMPLETION_TIME)
        if _pe <= deadline_dt:
            return "はい"
        return "いいえ"
    except Exception:
        return "判定不能"


def _result_task_sheet_sort_key(t: dict):
    """
    結果_タスク一覧の表示順。①配台試行順番（generate_plan 冒頭でキュー順に付与した 1..n）昇順。
    欠損・非数は最後。同一試行順内は依頼NO・機械名、続けて加工開始日・納期で安定化。
    """
    _dto = t.get("dispatch_trial_order")
    try:
        trial_k = int(_dto) if _dto is not None else 10**9
    except (TypeError, ValueError):
        trial_k = 10**9
    return (
        trial_k,
        str(t.get("task_id", "")).strip(),
        str(t.get("machine", "")).strip(),
        _task_date_key_for_result_sheet_sort(t.get("start_date_req")),
        _task_date_key_for_result_sheet_sort(t.get("answer_due_date")),
        _task_date_key_for_result_sheet_sort(t.get("specified_due_date")),
    )


def _is_result_task_history_expand_token(cell_val) -> bool:
    """列設定シートで「履歴」1行を置くと履歴1～n をその位置に展開する。"""
    if cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val)):
        return False
    s = unicodedata.normalize("NFKC", str(cell_val).strip())
    return s in ("履歴", "履歴*")


def _result_task_column_alias_map(df_columns) -> dict:
    """見出しの NFKC 正規化キー → DataFrame 上の実列名。"""
    m = {}
    for c in df_columns:
        m[_nfkc_column_aliases(str(c).strip())] = c
    return m


def _resolve_result_task_column_label(label, col_by_norm: dict):
    if label is None or (isinstance(label, float) and pd.isna(label)):
        return None
    s = unicodedata.normalize("NFKC", str(label).strip())
    if not s or s.lower() in ("nan", "none"):
        return None
    nk = _nfkc_column_aliases(s)
    resolved = col_by_norm.get(nk)
    if resolved is not None:
        return resolved
    # 旧列名（計画基準納期ベース）→ 配完_回答指定16時まで
    if nk == _nfkc_column_aliases("配完_基準16時まで"):
        return col_by_norm.get(
            _nfkc_column_aliases(RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16)
        )
    return None


def _parse_column_visible_cell(val) -> bool:
    """表示列: 空・未記入は True（表示）。FALSE/0/いいえ 等で非表示。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return True
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if val == 0:
            return False
        if val == 1:
            return True
    s = unicodedata.normalize("NFKC", str(val).strip()).lower()
    if s in ("", "true", "1", "はい", "yes", "on", "表示", "○"):
        return True
    if s in ("false", "flase", "0", "いいえ", "no", "off", "非表示", "隠す", "×"):
        return False
    return True


def parse_result_task_column_config_dataframe(
    df_cfg: pd.DataFrame | None, max_history_len: int
) -> list | None:
    """
    「列設定_結果_タスク一覧」相当の DataFrame から (列ラベル, 表示) を上から読む。
    見出し「列名」と「表示」（無い場合は表示はすべて True）。
    「履歴」「履歴*」の1行は履歴1～履歴n に展開し、同一行の表示フラグを共有する。
    同一列名（NFKC・別名正規化後）が複数行ある場合は先頭行のみ採用し、以降はログに出して捨てる。
    """
    if df_cfg is None or df_cfg.empty:
        return None
    df_cfg = df_cfg.dropna(how="all")
    if df_cfg.empty:
        return None

    name_col = None
    for c in df_cfg.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases(COLUMN_CONFIG_HEADER_COL):
            name_col = c
            break
    if name_col is None:
        name_col = df_cfg.columns[0]

    vis_col = None
    for c in df_cfg.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases(COLUMN_CONFIG_VISIBLE_COL):
            vis_col = c
            break

    seen_norm: set[str] = set()
    out: list[tuple[str, bool]] = []

    def _try_add(label: str, vis: bool) -> None:
        lab = str(label).strip()
        if not lab:
            return
        nk = _nfkc_column_aliases(unicodedata.normalize("NFKC", lab))
        if nk in seen_norm:
            logging.warning(
                "列設定「%s」: 重複列名「%s」をスキップしました（上の行を優先）。",
                COLUMN_CONFIG_SHEET_NAME,
                lab,
            )
            return
        seen_norm.add(nk)
        out.append((lab, vis))

    for i in range(len(df_cfg)):
        raw = df_cfg[name_col].iloc[i]
        vis = _parse_column_visible_cell(df_cfg[vis_col].iloc[i] if vis_col is not None else None)
        if _is_result_task_history_expand_token(raw):
            for j in range(max_history_len):
                _try_add(f"履歴{j+1}", vis)
            continue
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        s = unicodedata.normalize("NFKC", str(raw).strip())
        if not s or s.lower() in ("nan", "none"):
            continue
        _try_add(s, vis)
    return out or None


def _xlwings_write_column_config_sheet_ab(xw_sheet, rows: list[tuple[str, bool]]) -> None:
    """列設定シートの A:B を 列名・表示 のみで上書き（1行目見出し＋データ）。"""
    mat = [[COLUMN_CONFIG_HEADER_COL, COLUMN_CONFIG_VISIBLE_COL]]
    for lab, vis in rows:
        mat.append([lab, bool(vis)])
    n_r = len(mat)
    try:
        ur = xw_sheet.used_range
        lim_r = max(ur.row + ur.rows.count - 1, n_r, 2)
        xw_sheet.range((1, 1), (lim_r, 2)).clear_contents()
    except Exception:
        try:
            xw_sheet.range((1, 1)).resize(max(n_r, 50), 2).clear_contents()
        except Exception:
            pass
    xw_sheet.range((1, 1)).resize(n_r, 2).value = mat


def load_result_task_column_rows_from_input_workbook(max_history_len: int) -> list | None:
    """
    TASK_INPUT_WORKBOOK の「列設定_結果_タスク一覧」シートから (列ラベル, 表示) を上から読む。
    """
    wb = TASKS_INPUT_WORKBOOK
    if not wb or not os.path.exists(wb):
        return None
    if _workbook_should_skip_openpyxl_io(wb):
        logging.info(
            "列設定: ブックに「%s」があるため pandas(openpyxl) での「%s」読込をスキップ（既定列順を使います）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            COLUMN_CONFIG_SHEET_NAME,
        )
        return None
    try:
        df_cfg = pd.read_excel(wb, sheet_name=COLUMN_CONFIG_SHEET_NAME, header=0)
    except ValueError:
        return None
    except Exception as e:
        logging.warning(
            "シート「%s」: 読み込みに失敗したため既定の列順を使います (%s)",
            COLUMN_CONFIG_SHEET_NAME,
            e,
        )
        return None
    return parse_result_task_column_config_dataframe(df_cfg, max_history_len)


def apply_result_task_sheet_column_order(
    df: pd.DataFrame,
    max_history_len: int,
    *,
    config_dataframe: pd.DataFrame | None = None,
):
    """
    列設定シートがあればその順・表示を優先し、無い列は既定順で後ろに追記（表示は True）。
    config_dataframe を渡した場合はファイルを読まずその内容を列設定とみなす（xlwings 実行時用）。
    戻り値: (並べ替え後 DataFrame, 実際の列名リスト, 設定ソース説明文字列, 列名→表示bool)
    """
    default_order = default_result_task_sheet_column_order(max_history_len)
    if config_dataframe is not None:
        user_rows = parse_result_task_column_config_dataframe(config_dataframe, max_history_len)
    else:
        user_rows = load_result_task_column_rows_from_input_workbook(max_history_len)
    if user_rows:
        primary = user_rows
        source = (
            f"マクロブック「{COLUMN_CONFIG_SHEET_NAME}」"
            if config_dataframe is None
            else f"シート「{COLUMN_CONFIG_SHEET_NAME}」（実行中ブック）"
        )
    else:
        primary = [(n, True) for n in default_order]
        source = "既定"

    actual = list(df.columns)
    actual_set = set(actual)
    col_by_norm = _result_task_column_alias_map(actual)
    vis_map = {c: True for c in actual}

    seen = set()
    ordered = []
    unknown = []

    for item, vis in primary:
        resolved = _resolve_result_task_column_label(item, col_by_norm)
        if resolved and resolved not in seen:
            ordered.append(resolved)
            seen.add(resolved)
            vis_map[resolved] = vis
        elif not resolved:
            if item is not None and not (isinstance(item, float) and pd.isna(item)):
                lab = str(item).strip()
                if lab and lab not in unknown:
                    unknown.append(lab)

    for name in default_order:
        if name in actual_set and name not in seen:
            ordered.append(name)
            seen.add(name)
    for name in actual:
        if name not in seen:
            ordered.append(name)
            seen.add(name)

    if unknown:
        logging.warning(
            "列設定: 結果に無い列名を無視しました（最大20件）: %s",
            ", ".join(unknown[:20]) + (" …" if len(unknown) > 20 else ""),
        )
    logging.info("結果_タスク一覧の列順ソース: %s（%s 列）", source, len(ordered))
    if not user_rows and config_dataframe is None:
        logging.info(
            "列順・表示のカスタマイズ: マクロ実行ブックにシート「%s」を追加。"
            " 見出し「%s」「%s」… 表示が FALSE の列は結果シートで非表示。"
            " 1行「履歴」で履歴1～n を挿入。VBA の「列設定_結果_タスク一覧_チェックボックスを配置」でチェックボックスを表示列に連動可能。",
            COLUMN_CONFIG_SHEET_NAME,
            COLUMN_CONFIG_HEADER_COL,
            COLUMN_CONFIG_VISIBLE_COL,
        )
    return df[ordered], ordered, source, vis_map


def _xlwings_sheet_to_matrix(sheet) -> list:
    """xlwings Sheet の UsedRange を矩形の list[list] にする（1行のみでも2次元）。"""
    ur = sheet.used_range
    if ur is None:
        return []
    raw = ur.options(ndim=2).value
    if raw is None:
        return []
    if not isinstance(raw, list):
        return [[raw]]
    if len(raw) == 0:
        return []
    if not isinstance(raw[0], list):
        return [raw]
    return raw


def _matrix_to_dataframe_header_first(matrix: list) -> pd.DataFrame | None:
    """1行目を列名とみなし DataFrame を返す。空なら None。"""
    if not matrix or not matrix[0]:
        return None
    header = []
    for x in matrix[0]:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            header.append("")
        else:
            header.append(str(x).strip())
    if not any(h for h in header):
        return None
    body = matrix[1:] if len(matrix) > 1 else []
    return pd.DataFrame(body, columns=header)


def _max_history_len_from_result_task_df_columns(columns) -> int:
    """結果_タスク一覧の「履歴n」列から n の最大を返す（無ければ 1）。"""
    imax = 0
    for c in columns:
        m = re.match(r"^履歴(\d+)$", str(c).strip())
        if m:
            imax = max(imax, int(m.group(1)))
    return max(imax, 1)


def apply_result_task_column_layout_via_xlwings(workbook_path: str | None = None) -> bool:
    """
    Excel で開いているマクロブックについて、
    「列設定_結果_タスク一覧」の内容に合わせて「結果_タスク一覧」の列順と列非表示を更新する。
    ブックは事前に保存し、本処理中も Excel 上で開いたままにすること（xlwings が接続する）。
    """
    path = (workbook_path or "").strip() or TASKS_INPUT_WORKBOOK.strip()
    if not path:
        logging.error("結果_タスク一覧 列適用: ブックパスが空です（TASK_INPUT_WORKBOOK を設定してください）。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("結果_タスク一覧 列適用: xlwings が import できません。pip install xlwings を確認してください。")
        return False

    try:
        wb = xw.Book(path)
    except Exception as e:
        logging.error("結果_タスク一覧 列適用: ブックに接続できません: %s", e)
        return False

    try:
        ws_res = wb.sheets[RESULT_TASK_SHEET_NAME]
        ws_cfg = wb.sheets[COLUMN_CONFIG_SHEET_NAME]
    except Exception as e:
        logging.error("結果_タスク一覧 列適用: 必要シートが見つかりません: %s", e)
        return False

    mat_res = _xlwings_sheet_to_matrix(ws_res)
    mat_cfg = _xlwings_sheet_to_matrix(ws_cfg)
    df_res = _matrix_to_dataframe_header_first(mat_res)
    df_cfg = _matrix_to_dataframe_header_first(mat_cfg)
    if df_res is None or df_res.empty:
        logging.error("結果_タスク一覧 列適用: 「%s」にデータがありません。", RESULT_TASK_SHEET_NAME)
        return False
    if df_cfg is None:
        logging.error("結果_タスク一覧 列適用: 「%s」の見出しを読めません。", COLUMN_CONFIG_SHEET_NAME)
        return False

    max_h = _max_history_len_from_result_task_df_columns(df_res.columns)
    rows_cfg = parse_result_task_column_config_dataframe(df_cfg, max_h)
    if not rows_cfg:
        logging.error(
            "結果_タスク一覧 列適用: 「%s」に有効な列名行がありません。",
            COLUMN_CONFIG_SHEET_NAME,
        )
        return False
    _xlwings_write_column_config_sheet_ab(ws_cfg, rows_cfg)
    df_cfg_clean = pd.DataFrame(
        rows_cfg, columns=[COLUMN_CONFIG_HEADER_COL, COLUMN_CONFIG_VISIBLE_COL]
    )
    df_out, ordered, source, vis_map = apply_result_task_sheet_column_order(
        df_res, max_h, config_dataframe=df_cfg_clean
    )

    df_write = df_out.astype(object).where(pd.notna(df_out), None)
    headers = [str(h) for h in df_write.columns.tolist()]
    body = df_write.values.tolist()
    out_matrix = [headers] + body
    nrows = len(out_matrix)
    ncols = len(headers)
    if ncols == 0:
        return False

    try:
        ur_old = ws_res.used_range
        if ur_old is not None:
            ws_res.range((ur_old.row, ur_old.column)).resize(
                ur_old.rows.count, ur_old.columns.count
            ).clear_contents()
    except Exception:
        try:
            ws_res.used_range.clear_contents()
        except Exception:
            pass

    ws_res.range((1, 1)).resize(nrows, ncols).value = out_matrix

    for ci in range(1, ncols + 1):
        try:
            ws_res.range((1, ci)).api.EntireColumn.Hidden = False
        except Exception:
            pass

    for ci, col_name in enumerate(ordered, 1):
        if not vis_map.get(col_name, True):
            try:
                ws_res.range((1, ci)).api.EntireColumn.Hidden = True
            except Exception as e:
                logging.warning("列非表示に失敗（列%s %s）: %s", ci, col_name, e)

    try:
        wb.save()
    except Exception as e:
        logging.warning("結果_タスク一覧 列適用: 保存で警告（データはシート上は更新済みの可能性）: %s", e)

    logging.info(
        "結果_タスク一覧 列適用完了: %s（%s 列、非表示=%s）",
        source,
        len(ordered),
        sum(1 for c in ordered if not vis_map.get(c, True)),
    )
    return True


def apply_result_task_column_layout_only() -> bool:
    """環境変数 TASK_INPUT_WORKBOOK のブックに対し列設定を適用する（VBA ボタン用）。"""
    p = os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK
    return apply_result_task_column_layout_via_xlwings(p)


def dedupe_result_task_column_config_sheet_via_xlwings(workbook_path: str | None = None) -> bool:
    """
    「列設定_結果_タスク一覧」の A:B だけを、重複列名を除いた一覧で書き直す（先の行を優先）。
    「結果_タスク一覧」があれば履歴列数の解釈に使う。結果シートは変更しない。
    """
    path = (workbook_path or "").strip() or TASKS_INPUT_WORKBOOK.strip()
    if not path:
        logging.error("列設定 重複整理: ブックパスが空です。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("列設定 重複整理: xlwings が import できません。")
        return False
    try:
        wb = xw.Book(path)
        ws_cfg = wb.sheets[COLUMN_CONFIG_SHEET_NAME]
    except Exception as e:
        logging.error("列設定 重複整理: 接続またはシート取得に失敗: %s", e)
        return False

    max_h = 1
    try:
        ws_res = wb.sheets[RESULT_TASK_SHEET_NAME]
        df_r = _matrix_to_dataframe_header_first(_xlwings_sheet_to_matrix(ws_res))
        if df_r is not None and not df_r.empty:
            max_h = _max_history_len_from_result_task_df_columns(df_r.columns)
    except Exception:
        pass

    df_cfg = _matrix_to_dataframe_header_first(_xlwings_sheet_to_matrix(ws_cfg))
    if df_cfg is None:
        logging.error("列設定 重複整理: 「%s」の見出しを読めません。", COLUMN_CONFIG_SHEET_NAME)
        return False
    rows = parse_result_task_column_config_dataframe(df_cfg, max_h)
    if not rows:
        logging.warning("列設定 重複整理: 有効なデータ行がありません。")
        return False
    _xlwings_write_column_config_sheet_ab(ws_cfg, rows)
    try:
        wb.save()
    except Exception as e:
        logging.warning("列設定 重複整理: 保存警告: %s", e)
    logging.info(
        "列設定「%s」を重複除去済みで %s 行に整理しました（履歴展開後の行数）。",
        COLUMN_CONFIG_SHEET_NAME,
        len(rows),
    )
    return True


def dedupe_result_task_column_config_sheet_only() -> bool:
    """環境変数 TASK_INPUT_WORKBOOK のブックの列設定シートだけ重複整理（VBA 用）。"""
    p = os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK
    return dedupe_result_task_column_config_sheet_via_xlwings(p)


def _apply_result_task_sheet_column_visibility(worksheet, column_names: list, vis_map: dict):
    """結果_タスク一覧で、vis_map が False の列を非表示にする。"""
    for idx, col_name in enumerate(column_names, 1):
        if not vis_map.get(col_name, True):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True


def _norm_history_member_label(name: str) -> str:
    """履歴の担当名比較用（全角空白を半角1個化・前後trim・連続空白の圧縮）。"""
    t = str(name or "").replace("\u3000", " ").strip()
    return " ".join(t.split())


def _history_team_text_main_assignment_only(h: dict) -> str:
    """
    結果シート「担当」欄用: メイン割付確定時点の名前（余力追記サブは含めない）。
    append_surplus 後の h['team'] から post_dispatch_surplus_names を除外する。
    """
    raw = (h.get("team") or "").strip()
    if not raw:
        return ""
    ps = h.get("post_dispatch_surplus_names") or []
    if not ps:
        return raw
    ps_set = {
        _norm_history_member_label(x)
        for x in ps
        if x and str(x).strip()
    }
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    kept = [p for p in parts if _norm_history_member_label(p) not in ps_set]
    return ", ".join(kept) if kept else raw


def _format_result_task_history_cell(task: dict, h: dict) -> str:
    """結果_タスク一覧の履歴セル文字列（組合せ表の採用行ID・メイン追加人数・余力追記の明示を含む）。"""
    um = task.get("unit_m") or 0
    try:
        done_r = int(h["done_m"] / um) if um else 0
    except (TypeError, ValueError, ZeroDivisionError):
        done_r = 0
    dm = h.get("done_m", 0)
    parts_out: list[str] = [f"・【{h.get('date', '')}】：{done_r}R ({dm}m)"]
    cid = h.get("combo_sheet_row_id")
    if cid is not None:
        try:
            parts_out.append(f"組合せ表#{int(cid)}")
        except (TypeError, ValueError):
            parts_out.append(f"組合せ表#{cid}")
    parts_out.append(f"担当[{_history_team_text_main_assignment_only(h)}]")
    sm = h.get("surplus_member_names") or []
    if sm:
        parts_out.append(f"追加[{','.join(str(x) for x in sm)}]")
    ps = h.get("post_dispatch_surplus_names") or []
    if ps:
        parts_out.append(f"余力追記[{','.join(str(x) for x in ps)}]")
    return " ".join(parts_out)


_RESULT_TASK_HISTORY_RICH_HEAD_RE = re.compile(r"^・(【[^】]*】)(.*)$", re.DOTALL)


def _apply_result_task_history_rich_text(worksheet, column_names: list):
    """
    履歴列: 「・【日付】：…」の日付括弧部分を青色リッチテキストにする。
    openpyxl 3.1 未満ではスキップ（文字列の【】のみ）。
    """
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles.colors import Color
    except ImportError:
        return

    hist_cols = [
        i + 1 for i, c in enumerate(column_names) if str(c).startswith("履歴")
    ]
    if not hist_cols:
        return

    _plain_kw: dict = {}
    _blue_kw: dict = {"color": Color(rgb="FF0070C0")}
    if OUTPUT_BOOK_FONT_NAME:
        _plain_kw["rFont"] = OUTPUT_BOOK_FONT_NAME
        _blue_kw["rFont"] = OUTPUT_BOOK_FONT_NAME
    if OUTPUT_BOOK_FONT_SIZE is not None:
        _plain_kw["sz"] = OUTPUT_BOOK_FONT_SIZE
        _blue_kw["sz"] = OUTPUT_BOOK_FONT_SIZE
    plain_if = InlineFont(**_plain_kw)
    blue_if = InlineFont(**_blue_kw)
    top = Alignment(wrap_text=False, vertical="top")

    for r in range(2, worksheet.max_row + 1):
        for ci in hist_cols:
            cell = worksheet.cell(row=r, column=ci)
            v = cell.value
            if not isinstance(v, str) or not v.startswith("・【"):
                continue
            m = _RESULT_TASK_HISTORY_RICH_HEAD_RE.match(v)
            if not m:
                continue
            bracketed, rest = m.group(1), m.group(2)
            cell.value = CellRichText(
                TextBlock(plain_if, "・"),
                TextBlock(blue_if, bracketed),
                TextBlock(plain_if, rest),
            )
            cell.alignment = top


def _apply_result_task_date_columns_blue_font(worksheet, column_names: list):
    """
    結果_タスク一覧: 回答納期・指定納期・計画基準納期・原反投入日・加工開始日のセルを青色にする。
    （履歴列の【日付】は _apply_result_task_history_rich_text 側。色は 0070C0 で統一）
    """
    blue = _result_font(color="0070C0")
    top = Alignment(wrap_text=False, vertical="top")
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) not in RESULT_TASK_DATE_STYLE_HEADERS:
            continue
        for r in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=r, column=col_idx)
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not str(v).strip():
                continue
            cell.font = blue
            cell.alignment = top


def _apply_result_task_history_need_surplus_highlight(
    worksheet, column_names: list, sorted_tasks: list
):
    """
    need「配台時追加人数」相当で基本必要人数を超えて採用したブロック、または
    メイン完了後の余力追記でサブが増えたブロックに対応する「履歴n」セルを薄黄に塗る。
    """
    hist_cols: list[tuple[int, int]] = []
    for col_idx, col_name in enumerate(column_names, 1):
        m = re.match(r"^履歴(\d+)$", str(col_name).strip())
        if m:
            hist_cols.append((int(m.group(1)), col_idx))
    hist_cols.sort(key=lambda x: x[0])
    if not hist_cols or worksheet.max_row < 2:
        return
    fill_surplus = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    n_tasks = len(sorted_tasks)
    for r in range(2, worksheet.max_row + 1):
        ti = r - 2
        if ti < 0 or ti >= n_tasks:
            continue
        ah = sorted_tasks[ti].get("assigned_history") or []
        for ord1, cidx in hist_cols:
            i = ord1 - 1
            if i < 0 or i >= len(ah):
                continue
            if not ah[i].get("need_surplus_assigned"):
                continue
            worksheet.cell(row=r, column=cidx).fill = fill_surplus


def _apply_result_task_task_id_content_mismatch_highlight(
    worksheet, column_names: list, sorted_tasks: list
):
    """
    加工内容に工程名が含まれない行の「タスクID」セルを赤背景・白文字にする（元データ不整合の視認用）。
    """
    task_id_col_idx = None
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) == "タスクID":
            task_id_col_idx = col_idx
            break
    if task_id_col_idx is None or worksheet.max_row < 2:
        return
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white = _result_font(color="FFFFFF")
    top = Alignment(wrap_text=False, vertical="top")
    n_data = worksheet.max_row - 1
    for i in range(min(len(sorted_tasks), n_data)):
        if not sorted_tasks[i].get("process_content_mismatch"):
            continue
        cell = worksheet.cell(row=i + 2, column=task_id_col_idx)
        cell.fill = fill_red
        cell.font = font_white
        cell.alignment = top


def _apply_result_task_id_hyperlinks_to_equipment_schedule(
    worksheet_tasks,
    column_names: list,
    sorted_tasks_for_row_order: list,
    task_id_to_schedule_cell: dict[str, str],
    schedule_sheet_name: str,
) -> None:
    """
    結果_タスク一覧の「タスクID」セルに、結果_設備毎の時間割で当該タスクが最初に現れるセルへの内部ハイパーリンクを付与する。
    時間割に現れないタスク（未割当のみ等）はリンクなし。
    """
    if not task_id_to_schedule_cell or worksheet_tasks.max_row < 2:
        return
    task_id_col_idx = None
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) == "タスクID":
            task_id_col_idx = col_idx
            break
    if task_id_col_idx is None:
        return
    esc = schedule_sheet_name.replace("'", "''")
    loc_prefix = f"#'{esc}'!"
    font_link = Font(color="0563C1", underline="single")
    font_link_on_red = Font(color="FFFFFF", underline="single")
    top = Alignment(wrap_text=False, vertical="top")
    n_tasks = len(sorted_tasks_for_row_order)
    for r in range(2, worksheet_tasks.max_row + 1):
        cell = worksheet_tasks.cell(row=r, column=task_id_col_idx)
        raw = cell.value
        if raw is None:
            continue
        tid = str(raw).strip()
        if not tid:
            continue
        addr = task_id_to_schedule_cell.get(tid)
        if not addr:
            continue
        cell.hyperlink = loc_prefix + addr
        row_i = r - 2
        mismatch = (
            row_i < n_tasks
            and bool(sorted_tasks_for_row_order[row_i].get("process_content_mismatch"))
        )
        cell.font = font_link_on_red if mismatch else font_link
        cell.alignment = top


def _add_column_config_sheet_helpers(ws_cfg, num_data_rows: int):
    """表示列に TRUE/FALSE リスト（チェックの代わりにプルダウン）を付与。"""
    last_r = max(num_data_rows + 1, 2)
    cap = max(last_r + 50, 500)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws_cfg.add_data_validation(dv)
    dv.add(f"B2:B{cap}")


def _coerce_actual_sheet_datetime(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, time(0, 0))
    try:
        ts = pd.to_datetime(val, errors="coerce")
        if pd.isna(ts) or ts is pd.NaT:
            return None
        if isinstance(ts, pd.Timestamp):
            return ts.to_pydatetime()
        return ts if isinstance(ts, datetime) else None
    except Exception:
        return None


def _actual_row_time_bounds(row):
    """加工実績DATA の1行から (開始, 終了) を得る。解けなければ (None, None)。"""
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_START_DT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_END_DT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_START_ALT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_END_ALT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt

    d_date = parse_optional_date(row.get(ACT_COL_DAY))
    if not d_date:
        cd = _coerce_actual_sheet_datetime(row.get(ACT_COL_DAY))
        if isinstance(cd, datetime):
            d_date = cd.date()
        elif isinstance(cd, date):
            d_date = cd
    if not d_date:
        return None, None

    ts_s = row.get(ACT_COL_TIME_START)
    ts_e = row.get(ACT_COL_TIME_END)
    if ts_s is None or pd.isna(ts_s) or ts_e is None or pd.isna(ts_e):
        return None, None

    if isinstance(ts_s, time):
        t0 = ts_s
    elif isinstance(ts_s, datetime):
        t0 = ts_s.time()
    else:
        t0 = parse_time_str(ts_s, None)

    if isinstance(ts_e, time):
        t1 = ts_e
    elif isinstance(ts_e, datetime):
        t1 = ts_e.time()
    else:
        t1 = parse_time_str(ts_e, None)

    if t0 is None or t1 is None or t0 >= t1:
        return None, None
    return datetime.combine(d_date, t0), datetime.combine(d_date, t1)


def load_machining_actuals_df():
    """
    マクロブックの「加工実績DATA」を読む（無ければ空 DataFrame）。
    Power Query 等で用意したシートを想定。
    """
    if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
        return pd.DataFrame()
    try:
        df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=ACTUALS_SHEET_NAME)
    except ValueError:
        logging.info(
            f"シート「{ACTUALS_SHEET_NAME}」が無いため、ガントの実績行は出力しません。"
        )
        return pd.DataFrame()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, ACTUAL_HEADER_CANONICAL)
    logging.info(
        f"加工実績: '{TASKS_INPUT_WORKBOOK}' の '{ACTUALS_SHEET_NAME}' を {len(df)} 行読み込み。"
    )
    return df


def build_actual_timeline_events(df, equipment_list, sorted_dates):
    """
    実績シートの各行をガント用イベントへ変換。
    計画表示日（sorted_dates）かつ設備マスタに一致する「工程名」だけ対象。
    工程名は NFKC・空白正規化後にマスタ列名へマッピングする。
    時刻は DEFAULT_START_TIME / DEFAULT_END_TIME の枠内にクリップ。
    """
    if df is None or len(df) == 0:
        return []
    equip_lookup = _equipment_lookup_normalized_to_canonical(equipment_list)
    date_ok = set(sorted_dates)
    events = []
    bad_eq = 0
    bad_time = 0
    no_plan_overlap = 0
    mismatch_norm_samples = []

    for _, row in df.iterrows():
        tid = row.get(ACT_COL_TASK_ID)
        if tid is None or pd.isna(tid):
            continue
        tid_s = str(tid).strip()
        if not tid_s:
            continue
        proc = row.get(ACT_COL_PROCESS)
        if proc is None or pd.isna(proc):
            continue
        proc_key = _normalize_equipment_match_key(proc)
        mach = equip_lookup.get(proc_key)
        if not mach:
            bad_eq += 1
            if len(mismatch_norm_samples) < 12 and proc_key:
                if proc_key not in mismatch_norm_samples:
                    mismatch_norm_samples.append(proc_key)
            continue
        start_dt, end_dt = _actual_row_time_bounds(row)
        if not start_dt or not end_dt or start_dt >= end_dt:
            bad_time += 1
            continue
        op_val = row.get(ACT_COL_OPERATOR)
        op_s = ""
        if op_val is not None and not pd.isna(op_val):
            op_s = str(op_val).strip()

        before = len(events)
        for d in sorted_dates:
            if d not in date_ok:
                continue
            day_start = datetime.combine(d, DEFAULT_START_TIME)
            day_end = datetime.combine(d, DEFAULT_END_TIME)
            if end_dt <= day_start or start_dt >= day_end:
                continue
            s_clip = max(start_dt, day_start)
            e_clip = min(end_dt, day_end)
            if s_clip >= e_clip:
                continue
            events.append(
                {
                    "date": d,
                    "task_id": tid_s,
                    "machine": mach,
                    "op": op_s,
                    "sub": "",
                    "start_dt": s_clip,
                    "end_dt": e_clip,
                    "breaks": [],
                    "units_done": 0,
                    "already_done_units": 0,
                    "total_units": 0,
                    "eff_time_per_unit": 0.0,
                    "unit_m": 0.0,
                }
            )
        if len(events) == before:
            no_plan_overlap += 1

    if bad_eq:
        logging.warning(
            f"加工実績DATA: 工程名がマスタ設備と一致しない行を {bad_eq} 件スキップしました（空白等は正規化済み）。"
        )
        if mismatch_norm_samples:
            logging.info(
                "  不一致となった工程名の正規化後サンプル: "
                + " | ".join(mismatch_norm_samples[:12])
            )
    if bad_time:
        logging.info(
            f"加工実績DATA: 開始/終了日時が解釈できない行を {bad_time} 件スキップしました。"
        )
    if no_plan_overlap and sorted_dates:
        logging.info(
            f"加工実績DATA: 設備・日時は有効だが、計画対象日（当日以降の勤怠日×{DEFAULT_START_TIME}～{DEFAULT_END_TIME}）と重ならない行が {no_plan_overlap} 件ありました。"
        )
    if not events and len(df) > 0:
        logging.info(
            "加工実績DATA: ガント用セグメントが0件です。過去日の実績のみの場合、計画の表示日（sorted_dates）に含まれないため描画されません。"
        )
    logging.info(f"加工実績DATA からガント用セグメント {len(events)} 件を生成しました。")
    return events


TASK_SPECIAL_AI_LAST_RESPONSE_FILE = "ai_task_special_remark_last.txt"
# 勤怠備考キャッシュとキー空間を分離（同一SHA衝突を避ける）。指紋に基準年を含め日付解釈のズレを防ぐ。
TASK_SPECIAL_CACHE_KEY_PREFIX = "TASK_SPECIAL_v3|"
# メインシート「グローバルコメント」下の自由記述 → Gemini 解釈（配台の最優先オーバーライド）
GLOBAL_PRIORITY_OVERRIDE_CACHE_PREFIX = "GLOBAL_PRIO_v7|"


def _normalize_special_task_id_for_ai(val):
    """
    依頼NOをキャッシュキー・プロンプト行で一貫させる。
    Excel の数値セルは float になりがちなので 12345.0 → \"12345\" に揃える。
    文字列は NFKC（全角英数字など）で表記ゆれを吸収（同一実体の再API呼び出しを減らす）。
    """
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except TypeError:
        pass
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if math.isnan(val):
            return None
        if val.is_integer():
            return str(int(val))
        s = str(val).strip()
        if not s or s.lower() in ("nan", "none", "null"):
            return None
        return unicodedata.normalize("NFKC", s).strip() or None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    s = unicodedata.normalize("NFKC", s).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    # 文字列としての "20010.0" 等（Excel・CSV）を整数表記の依頼NOに寄せる
    if re.fullmatch(r"-?\d+\.0+", s):
        try:
            return str(int(float(s)))
        except ValueError:
            pass
    return s or None


def planning_task_id_str_from_scalar(val) -> str:
    """配台・段階1マージ・キュー構築で用いる依頼NOの安定文字列（空なら \"\"）。"""
    return _normalize_special_task_id_for_ai(val) or ""


def planning_task_id_str_from_plan_row(row) -> str:
    """重複見出し列でも先頭スカラーを拾い、依頼NOを planning_task_id_str_from_scalar に渡す。"""
    return planning_task_id_str_from_scalar(_planning_df_cell_scalar(row, TASK_COL_TASK_ID))


def _cell_text_task_special_remark(val):
    """
    特別指定_備考をプロンプト用に取り出す。仕様どおり **strip のみ**
    （先頭末尾の空白・Excel の偽空白を除き、文中の改行・スペースは保持。数値セルは表記を固定）。
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except TypeError:
        pass
    if isinstance(val, bool):
        s = str(val)
    elif isinstance(val, float):
        if math.isnan(val):
            return ""
        # 備考列に数値だけ入っている場合の表記ゆれを減らす
        if val.is_integer():
            s = str(int(val))
        else:
            s = str(val)
    elif isinstance(val, int):
        s = str(val)
    else:
        s = str(val)
    s = s.strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _task_special_prompt_lines(tasks_df):
    """プロンプトに載せる行リスト（ソート前）。正規化は上記ヘルパーに統一。"""
    lines = []
    for _, row in tasks_df.iterrows():
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
        if not tid or not rem:
            continue
        proc = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        macn = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        proc_disp = proc if proc else "（空）"
        macn_disp = macn if macn else "（空）"
        lines.append(
            f"- 依頼NO【{tid}】| 工程名「{proc_disp}」 | 機械名「{macn_disp}」 | 備考本文: {rem}"
        )
    return lines


def _repair_task_special_ai_wrong_top_level_keys(parsed: dict, tasks_df) -> dict:
    """
    備考が品番・原反コード（例: 20010 で始まる数字列）で始まると、モデルがその列を JSON トップキーに
    誤用することがある。依頼NO【…】と一致しない数字のみのキーを、当該備考を持つ行の依頼NOへ付け替える。
    """
    if not isinstance(parsed, dict) or not parsed or tasks_df is None or getattr(tasks_df, "empty", True):
        return parsed
    valid_tids: set[str] = set()
    remark_by_tid: dict[str, list[str]] = {}
    for _, row in tasks_df.iterrows():
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
        if not tid or not rem:
            continue
        valid_tids.add(tid)
        remark_by_tid.setdefault(tid, []).append(rem)

    for bad_key in list(parsed.keys()):
        sk = str(bad_key).strip()
        if sk in valid_tids:
            continue
        if not re.fullmatch(r"\d{4,}", sk):
            continue
        hits = [
            tid
            for tid, rems in remark_by_tid.items()
            if any(
                r.startswith(sk)
                or r.startswith(sk + " ")
                or r.startswith(sk + "\u3000")
                or r.startswith(sk + "-")
                or r.startswith(sk + "ー")
                for r in rems
            )
        ]
        if len(hits) != 1:
            continue
        target = hits[0]
        val = parsed.pop(bad_key, None)
        if val is None:
            continue
        if target not in parsed:
            parsed[target] = val
            logging.info(
                "タスク特別指定: JSON トップキー誤りを修復（%r は依頼NOではない → %r）",
                bad_key,
                target,
            )
        else:
            parsed[bad_key] = val
    return parsed


def _normalize_task_special_scope_str(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return unicodedata.normalize("NFKC", str(s).strip())


def _task_special_scope_matches_row_field(row_val, restrict_val) -> bool:
    """
    restrict が無い・空なら制限なし（True）。
    非空なら Excel 側の値とあいまい一致（部分一致可）。
    """
    if restrict_val is None:
        return True
    r = _normalize_task_special_scope_str(restrict_val)
    if not r:
        return True
    v = _normalize_task_special_scope_str(row_val)
    if not v:
        return False
    if v == r:
        return True
    if r in v or v in r:
        return True
    return False


def _ai_remark_entry_applies_to_row(entry: dict, row) -> bool:
    """restrict_to_* が無いときは同一依頼NOの全行に適用。"""
    if not isinstance(entry, dict):
        return False
    rp = row.get(TASK_COL_MACHINE, "")
    rm = row.get(TASK_COL_MACHINE_NAME, "")
    if not _task_special_scope_matches_row_field(rp, entry.get("restrict_to_process_name")):
        return False
    if not _task_special_scope_matches_row_field(rm, entry.get("restrict_to_machine_name")):
        return False
    return True


def _row_matches_remark_source_row(entry: dict, row) -> bool:
    """
    JSON の process_name / machine_name が、当該 Excel 行の工程名・機械名と一致するか。
    （プロンプトで渡した「備考があった行」と対応づける。片方だけ一致でも可）
    """
    if not isinstance(entry, dict):
        return False
    rp = _normalize_task_special_scope_str(row.get(TASK_COL_MACHINE))
    rm = _normalize_task_special_scope_str(row.get(TASK_COL_MACHINE_NAME))
    ep = _normalize_task_special_scope_str(entry.get("process_name"))
    em = _normalize_task_special_scope_str(entry.get("machine_name"))
    proc_ok = (not ep) or (not rp) or ep == rp or ep in rp or rp in ep
    mac_ok = (not em) or (not rm) or em == rm or em in rm or rm in em
    return proc_ok and mac_ok


def _entry_is_global_task_special_scope(entry: dict) -> bool:
    """restrict_to_* が無い・空＝同一依頼NOの全工程行に効かせる指定。"""
    if not isinstance(entry, dict):
        return False
    a = _normalize_task_special_scope_str(entry.get("restrict_to_process_name"))
    b = _normalize_task_special_scope_str(entry.get("restrict_to_machine_name"))
    return not a and not b


def _select_ai_task_special_entry_for_tid_value(val, row):
    """1依頼NOに対する値が dict または dict の配列のどちらでも行に合う要素を返す。"""
    if val is None:
        return None
    if isinstance(val, list):
        for item in val:
            if (
                isinstance(item, dict)
                and _ai_remark_entry_applies_to_row(item, row)
                and _row_matches_remark_source_row(item, row)
            ):
                return item
        for item in val:
            if (
                isinstance(item, dict)
                and _ai_remark_entry_applies_to_row(item, row)
                and _entry_is_global_task_special_scope(item)
            ):
                return item
        for item in val:
            if isinstance(item, dict) and _ai_remark_entry_applies_to_row(item, row):
                return item
        return None
    if isinstance(val, dict):
        if _ai_remark_entry_applies_to_row(val, row):
            return val
        return None
    return None


def _ai_task_special_entry_for_row(ai_by_tid, row):
    """
    analyze_task_special_remarks の戻りから当該行のエントリを取る。
    プロンプトキーは正規化済み依頼NOなので、Excel が 12345.0 でもヒットする。
    restrict_to_process_name / restrict_to_machine_name が無い・空のときは
    同一依頼NOの工程・機械が異なる全行に同じ指示を適用する。
    """
    if not isinstance(ai_by_tid, dict) or not ai_by_tid:
        return None
    tid_norm = planning_task_id_str_from_plan_row(row)
    tid_raw = str(_planning_df_cell_scalar(row, TASK_COL_TASK_ID) or "").strip()

    def try_val(v):
        return _select_ai_task_special_entry_for_tid_value(v, row)

    if tid_norm and tid_norm in ai_by_tid:
        hit = try_val(ai_by_tid[tid_norm])
        if hit is not None:
            return hit
    if tid_raw:
        for key in (tid_raw, str(tid_raw)):
            if key in ai_by_tid:
                hit = try_val(ai_by_tid[key])
                if hit is not None:
                    return hit
    if tid_norm:
        for k, v in ai_by_tid.items():
            if str(k).strip() == tid_norm:
                hit = try_val(v)
                if hit is not None:
                    return hit
    if tid_raw:
        for k, v in ai_by_tid.items():
            if str(k).strip() == tid_raw:
                hit = try_val(v)
                if hit is not None:
                    return hit
    return None


def _gemini_result_text(res):
    try:
        return (res.text or "").strip()
    except Exception:
        return ""


# 1 回の Python 実行（段階1 または 段階2）単位でリセットする
_gemini_usage_session: dict[str, dict[str, int]] = {}


def reset_gemini_usage_tracker() -> None:
    global _gemini_usage_session
    _gemini_usage_session = {}


def _gemini_cumulative_json_path() -> str:
    path = os.path.join(api_payment_dir, GEMINI_USAGE_CUMULATIVE_JSON_FILE)
    legacy = os.path.join(log_dir, GEMINI_USAGE_CUMULATIVE_JSON_FILE)
    if os.path.isfile(legacy) and not os.path.isfile(path):
        try:
            shutil.move(legacy, path)
        except OSError:
            pass
    return path


def _load_gemini_cumulative_payload() -> dict:
    """API_Payment 内の累計 JSON を読む。無い・壊れていれば初期形を返す。"""
    path = _gemini_cumulative_json_path()
    default: dict = {
        "version": 1,
        "updated_at": None,
        "calls_total": 0,
        "prompt_total": 0,
        "candidates_total": 0,
        "thoughts_total": 0,
        "total_tokens_reported": 0,
        "estimated_cost_usd_total": 0.0,
        "by_model": {},
    }
    if not os.path.isfile(path):
        _gemini_buckets_ensure_structure(default)
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict) or int(data.get("version") or 0) != 1:
            _gemini_buckets_ensure_structure(default)
            return default
        data.setdefault("by_model", {})
        _gemini_buckets_ensure_structure(data)
        for k in (
            "calls_total",
            "prompt_total",
            "candidates_total",
            "thoughts_total",
            "total_tokens_reported",
        ):
            data[k] = int(data.get(k) or 0)
        data["estimated_cost_usd_total"] = float(data.get("estimated_cost_usd_total") or 0.0)
        return data
    except Exception:
        _gemini_buckets_ensure_structure(default)
        return default


def _save_gemini_cumulative_payload(data: dict) -> None:
    path = _gemini_cumulative_json_path()
    try:
        os.makedirs(api_payment_dir, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8", newline="\n") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except OSError as ex:
        logging.debug("Gemini 累計 JSON の保存に失敗: %s", ex)


def _gemini_buckets_ensure_structure(data: dict) -> None:
    """累計 JSON に期間別バケット用の辞書を用意する（既存 v1 ファイルもマージ）。"""
    b = data.setdefault("buckets", {})
    if not isinstance(b, dict):
        b = {}
        data["buckets"] = b
    for sub in ("by_year", "by_month", "by_week", "by_day", "by_hour"):
        x = b.setdefault(sub, {})
        if not isinstance(x, dict):
            b[sub] = {}
    b.setdefault(
        "timezone_note",
        "period_key は PC ローカル時刻（datetime.now）で付与。他 PC との集計は混ぜないでください。",
    )


def _gemini_time_bucket_keys(dt: datetime) -> tuple[str, str, str, str, str]:
    """年・月・ISO週・日・時 のキー（文字列ソートで時系列比較しやすい形）。"""
    iy, iw, _ = dt.isocalendar()
    y = dt.strftime("%Y")
    ym = dt.strftime("%Y-%m")
    wk = f"{iy}-W{iw:02d}"
    d = dt.strftime("%Y-%m-%d")
    h = dt.strftime("%Y-%m-%dT%H")
    return y, ym, wk, d, h


def _gemini_bucket_add_one_call(
    buckets_root: dict,
    pt: int,
    ct: int,
    th: int,
    tt: int,
    inc_usd: float | None,
    *,
    when: datetime | None = None,
) -> None:
    """1 回の API 呼出しを年・月・週・日・時の各バケットに加算する。"""
    dt = when or datetime.now()
    y, ym, wk, d, h = _gemini_time_bucket_keys(dt)
    pairs = (
        ("by_year", y),
        ("by_month", ym),
        ("by_week", wk),
        ("by_day", d),
        ("by_hour", h),
    )
    for sub, pk in pairs:
        subd = buckets_root.setdefault(sub, {})
        ent = subd.setdefault(
            pk,
            {
                "calls": 0,
                "prompt": 0,
                "candidates": 0,
                "thoughts": 0,
                "total_tokens": 0,
                "estimated_cost_usd": 0.0,
            },
        )
        ent["calls"] = int(ent.get("calls") or 0) + 1
        ent["prompt"] = int(ent.get("prompt") or 0) + pt
        ent["candidates"] = int(ent.get("candidates") or 0) + ct
        ent["thoughts"] = int(ent.get("thoughts") or 0) + th
        ent["total_tokens"] = int(ent.get("total_tokens") or 0) + tt
        if inc_usd is not None:
            ent["estimated_cost_usd"] = float(
                ent.get("estimated_cost_usd") or 0.0
            ) + float(inc_usd)


def _append_gemini_cumulative_one_call(
    model_id: str, pt: int, ct: int, th: int, tt: int
) -> None:
    """1 回の API 応答分を累計 JSON に加算する（ログに単発料金は出さない）。"""
    mid = str(model_id).strip()
    data = _load_gemini_cumulative_payload()
    data["calls_total"] = int(data["calls_total"]) + 1
    data["prompt_total"] = int(data["prompt_total"]) + pt
    data["candidates_total"] = int(data["candidates_total"]) + ct
    data["thoughts_total"] = int(data["thoughts_total"]) + th
    data["total_tokens_reported"] = int(data["total_tokens_reported"]) + tt
    bm: dict = data["by_model"]
    if mid not in bm or not isinstance(bm[mid], dict):
        bm[mid] = {
            "calls": 0,
            "prompt": 0,
            "candidates": 0,
            "thoughts": 0,
            "total": 0,
            "estimated_cost_usd": 0.0,
        }
    m = bm[mid]
    m["calls"] = int(m.get("calls") or 0) + 1
    m["prompt"] = int(m.get("prompt") or 0) + pt
    m["candidates"] = int(m.get("candidates") or 0) + ct
    m["thoughts"] = int(m.get("thoughts") or 0) + th
    m["total"] = int(m.get("total") or 0) + tt
    inc_usd = _gemini_estimate_cost_usd(mid, pt, ct, th)
    if inc_usd is not None:
        m["estimated_cost_usd"] = float(m.get("estimated_cost_usd") or 0.0) + float(inc_usd)
        data["estimated_cost_usd_total"] = float(
            data.get("estimated_cost_usd_total") or 0.0
        ) + float(inc_usd)
    _gemini_buckets_ensure_structure(data)
    _gemini_bucket_add_one_call(
        data["buckets"], pt, ct, th, tt, inc_usd, when=datetime.now()
    )
    data["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_gemini_cumulative_payload(data)


def record_gemini_response_usage(res, model_id: str) -> None:
    """generate_content の応答から usage_metadata を集計する（セッション＋累計 JSON）。"""
    global _gemini_usage_session
    if res is None or not str(model_id or "").strip():
        return
    um = getattr(res, "usage_metadata", None)
    if um is None:
        return

    def _iv(name: str) -> int:
        v = getattr(um, name, None)
        try:
            return int(v) if v is not None else 0
        except (TypeError, ValueError):
            return 0

    pt = _iv("prompt_token_count")
    ct = _iv("candidates_token_count")
    tt = _iv("total_token_count")
    th = _iv("thoughts_token_count")
    if tt <= 0 and (pt > 0 or ct > 0 or th > 0):
        tt = pt + ct + th
    mid = str(model_id).strip()
    b = _gemini_usage_session.setdefault(
        mid,
        {"prompt": 0, "candidates": 0, "total": 0, "thoughts": 0, "calls": 0},
    )
    b["prompt"] += pt
    b["candidates"] += ct
    b["total"] += tt
    b["thoughts"] += th
    b["calls"] += 1
    try:
        _append_gemini_cumulative_one_call(mid, pt, ct, th, tt)
    except Exception as ex:
        logging.debug("Gemini 累計の更新で例外（続行）: %s", ex)


def _gemini_estimate_cost_usd(
    model_id: str, prompt_tok: int, cand_tok: int, thoughts_tok: int
) -> float | None:
    m = str(model_id).strip().lower()
    rin, rout = None, None
    if "flash" in m:
        rin, rout = _GEMINI_FLASH_IN_PER_M, _GEMINI_FLASH_OUT_PER_M
    elif "pro" in m:
        # 目安（未使用モデル向けフォールバック）
        rin, rout = 1.25, 5.0
    if rin is None:
        return None
    out_equiv = cand_tok + thoughts_tok
    return (prompt_tok / 1_000_000.0) * rin + (out_equiv / 1_000_000.0) * rout


def _gemini_bucket_entry_line(period_key: str, ent: dict) -> str:
    calls = int(ent.get("calls") or 0)
    pt = int(ent.get("prompt") or 0)
    cc = int(ent.get("candidates") or 0)
    th = int(ent.get("thoughts") or 0)
    usd = float(ent.get("estimated_cost_usd") or 0.0)
    parts = [
        f"  {period_key}",
        f"呼出し {calls:,} 回",
        f"入力 {pt:,}",
        f"出力 {cc:,}",
    ]
    if th:
        parts.append(f"思考 {th:,}")
    if usd > 0:
        parts.append(f"推定USD ${usd:.6f}")
        parts.append(f"推定JPY ¥{usd * GEMINI_JPY_PER_USD:.2f}")
    return "  ".join(parts)


def _gemini_time_bucket_summary_lines(cum: dict) -> list[str]:
    b = cum.get("buckets")
    if not isinstance(b, dict):
        return []
    lines: list[str] = []
    note = b.get("timezone_note")
    lines.append("【期間別集計】（トレンドは log の CSV を Excel でグラフ化）")
    if note:
        lines.append(f"  （{note}）")

    def emit_block(title: str, sub: str, limit: int | None) -> None:
        subd = b.get(sub)
        if not isinstance(subd, dict) or not subd:
            return
        keys = sorted(subd.keys(), reverse=True)
        if limit is not None:
            keys = keys[:limit]
        lines.append(f"  [{title}]")
        for pk in keys:
            ent = subd.get(pk)
            if isinstance(ent, dict):
                lines.append(_gemini_bucket_entry_line(pk, ent))

    emit_block("年", "by_year", None)
    emit_block("月（新しい順・最大12）", "by_month", 12)
    emit_block("週 ISO（新しい順・最大8）", "by_week", 8)
    emit_block("日（新しい順・最大14）", "by_day", 14)
    emit_block("時間（新しい順・最大48）", "by_hour", 48)
    lines.append(
        f"  グラフ用: log\\{GEMINI_USAGE_BUCKETS_CSV_FILE}"
    )
    return lines


def _export_gemini_buckets_csv_for_charts(cum: dict) -> None:
    """Excel 折れ線・棒グラフ向けに長形式 CSV を log に書き出す。"""
    b = cum.get("buckets")
    if not isinstance(b, dict):
        return
    mapping = (
        ("year", "by_year"),
        ("month", "by_month"),
        ("week_iso", "by_week"),
        ("day", "by_day"),
        ("hour", "by_hour"),
    )
    rows_out: list[dict[str, object]] = []
    for gran_label, sub in mapping:
        subd = b.get(sub)
        if not isinstance(subd, dict):
            continue
        for pk in sorted(subd.keys()):
            ent = subd.get(pk)
            if not isinstance(ent, dict):
                continue
            calls = int(ent.get("calls") or 0)
            pt = int(ent.get("prompt") or 0)
            cc = int(ent.get("candidates") or 0)
            th = int(ent.get("thoughts") or 0)
            tt = int(ent.get("total_tokens") or 0)
            usd = float(ent.get("estimated_cost_usd") or 0.0)
            rows_out.append(
                {
                    "granularity": gran_label,
                    "period_key": pk,
                    "calls": calls,
                    "prompt_tokens": pt,
                    "candidates_tokens": cc,
                    "thoughts_tokens": th,
                    "total_tokens": tt,
                    "estimated_cost_usd": round(usd, 8),
                    "estimated_cost_jpy": round(usd * GEMINI_JPY_PER_USD, 4),
                }
            )
    if not rows_out:
        return
    path = os.path.join(log_dir, GEMINI_USAGE_BUCKETS_CSV_FILE)
    fieldnames = [
        "granularity",
        "period_key",
        "calls",
        "prompt_tokens",
        "candidates_tokens",
        "thoughts_tokens",
        "total_tokens",
        "estimated_cost_usd",
        "estimated_cost_jpy",
    ]
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows_out)
    except OSError as ex:
        logging.debug("Gemini バケット CSV の保存に失敗: %s", ex)


def build_gemini_usage_summary_text() -> str:
    """メイン表示・結果ログ用の複数行テキスト（この実行分＋累計 JSON）。"""
    cum = _load_gemini_cumulative_payload()
    ct_tot = int(cum.get("calls_total") or 0)
    if not _gemini_usage_session and ct_tot <= 0:
        return ""

    lines: list[str] = []
    ts = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    if _gemini_usage_session:
        lines.append(f"集計時刻: {ts}（この実行での Gemini API 合計）")
        tot_calls = sum(b["calls"] for b in _gemini_usage_session.values())
        tot_p = sum(b["prompt"] for b in _gemini_usage_session.values())
        tot_c = sum(b["candidates"] for b in _gemini_usage_session.values())
        tot_th = sum(b["thoughts"] for b in _gemini_usage_session.values())
        tot_t = sum(b["total"] for b in _gemini_usage_session.values())
        lines.append(
            f"【合計】呼出し {tot_calls} 回 / 入力 {tot_p:,} / 出力 {tot_c:,}"
            + (f" / 思考 {tot_th:,}" if tot_th else "")
            + f" / total 報告 {tot_t:,}"
        )
        grand_usd = 0.0
        any_price = False
        for mid in sorted(_gemini_usage_session.keys()):
            b = _gemini_usage_session[mid]
            lines.append(f"--- モデル: {mid} ---")
            lines.append(f"  呼出し: {b['calls']} 回")
            lines.append(f"  入力トークン: {b['prompt']:,}")
            lines.append(f"  出力トークン(candidates): {b['candidates']:,}")
            if b.get("thoughts", 0):
                lines.append(f"  思考トークン(thoughts): {b['thoughts']:,}")
            lines.append(f"  total_token_count 合計: {b['total']:,}")
            est = _gemini_estimate_cost_usd(
                mid, b["prompt"], b["candidates"], b.get("thoughts", 0)
            )
            if est is not None:
                any_price = True
                grand_usd += est
                lines.append(f"  推定料金(USD): ${est:.6f}")
                lines.append(
                    f"  推定料金(JPY・{GEMINI_JPY_PER_USD:.0f}円/USD): ¥{est * GEMINI_JPY_PER_USD:.2f}"
                )
            else:
                lines.append("  推定料金: （単価未登録モデル）")
        if any_price:
            lines.append(f"【推定料金 合計(USD)】${grand_usd:.6f}")
            lines.append(
                f"【推定料金 合計(JPY)】¥{grand_usd * GEMINI_JPY_PER_USD:.2f}"
            )
    else:
        lines.append(f"集計時刻: {ts}")
        lines.append("（この実行での Gemini API 呼出しはありません）")
    lines.append("※ トークンは API の usage_metadata に基づきます。")
    lines.append(
        "※ USD 単価はコード／環境変数の目安です。実課金は Google の請求を参照してください。"
    )
    lines.append(
        "※ 各 API 呼出しごとの推定料金はコンソールに出さず、下記累計 JSON にのみ積み上げます。"
    )

    if ct_tot > 0:
        lines.append("")
        lines.append(
            f"【累計】{GEMINI_USAGE_CUMULATIVE_JSON_FILE}（全実行・推定値・ファイル: API_Payment フォルダ）"
        )
        lines.append(f"  最終更新: {cum.get('updated_at') or '—'}")
        pt0 = int(cum.get("prompt_total") or 0)
        cc0 = int(cum.get("candidates_total") or 0)
        th0 = int(cum.get("thoughts_total") or 0)
        tt0 = int(cum.get("total_tokens_reported") or 0)
        lines.append(
            f"  呼出し {ct_tot:,} 回 / 入力 {pt0:,} / 出力 {cc0:,}"
            + (f" / 思考 {th0:,}" if th0 else "")
            + f" / total 報告 {tt0:,}"
        )
        usd_all = float(cum.get("estimated_cost_usd_total") or 0.0)
        if usd_all > 0:
            lines.append(f"  推定料金累計(USD): ${usd_all:.6f}")
            lines.append(
                f"  推定料金累計(JPY・{GEMINI_JPY_PER_USD:.0f}円/USD): ¥{usd_all * GEMINI_JPY_PER_USD:.2f}"
            )
        bm = cum.get("by_model") or {}
        if isinstance(bm, dict) and bm:
            for mid in sorted(bm.keys()):
                m = bm[mid]
                if not isinstance(m, dict):
                    continue
                lines.append(f"  --- 累計 モデル: {mid} ---")
                lines.append(f"    呼出し: {int(m.get('calls') or 0):,} 回")
                lines.append(f"    入力: {int(m.get('prompt') or 0):,} / 出力: {int(m.get('candidates') or 0):,}")
                if int(m.get("thoughts") or 0):
                    lines.append(f"    思考: {int(m.get('thoughts') or 0):,}")
                mud = float(m.get("estimated_cost_usd") or 0.0)
                if mud > 0:
                    lines.append(f"    推定料金累計(USD): ${mud:.6f}")
                    lines.append(
                        f"    推定料金累計(JPY): ¥{mud * GEMINI_JPY_PER_USD:.2f}"
                    )
        lines.extend(_gemini_time_bucket_summary_lines(cum))
    return "\n".join(lines)


def _write_main_sheet_gemini_usage_via_openpyxl(
    macro_wb_path: str, text: str, log_prefix: str
) -> bool:
    """メインシート P 列 16 行目以降に Gemini サマリを openpyxl で書き save する（ブックが閉じているとき）。"""
    if _workbook_should_skip_openpyxl_io(macro_wb_path):
        logging.info(
            "%s: ブックに「%s」があるため openpyxl でメイン P 列へ書き込みません。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    keep_vba = str(macro_wb_path).lower().endswith(".xlsm")
    start_r, col_p, clear_n = 16, 16, 120
    wb = None
    try:
        wb = load_workbook(
            macro_wb_path, keep_vba=keep_vba, read_only=False, data_only=False
        )
    except Exception as ex:
        logging.info(
            "%s: メイン P 列への openpyxl 書込のためブックを開けません（Excel で開きっぱなし等）: %s",
            log_prefix,
            ex,
        )
        return False
    try:
        ws_main = None
        for name in ("メイン", "Main"):
            if name in wb.sheetnames:
                ws_main = wb[name]
                break
        if ws_main is None:
            for sn in wb.sheetnames:
                if "メイン" in str(sn):
                    ws_main = wb[sn]
                    break
        if ws_main is None:
            logging.info(
                "%s: メインシートが無いため openpyxl での AI サマリをスキップしました。",
                log_prefix,
            )
            return False
        last_clear = start_r + clear_n - 1
        for i in range(clear_n):
            ws_main.cell(row=start_r + i, column=col_p, value=None)
        lines = text.split("\n") if (text or "").strip() else []
        for i, line in enumerate(lines):
            if i >= clear_n:
                break
            cell = ws_main.cell(row=start_r + i, column=col_p, value=line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(macro_wb_path)
        logging.info(
            "%s: メインシート P%d 以降に AI 利用サマリを openpyxl で保存しました。",
            log_prefix,
            start_r,
        )
        return True
    except Exception as ex:
        logging.warning(
            "%s: メイン AI サマリの openpyxl 保存に失敗: %s", log_prefix, ex
        )
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def write_main_sheet_gemini_usage_summary(wb_path: str, log_prefix: str) -> None:
    """Gemini 利用サマリを log に書き、可能なら openpyxl でメイン P 列へ保存。開いたまま保存できないときは VBA 用テキストのみ。"""
    text = build_gemini_usage_summary_text()
    path = os.path.join(log_dir, GEMINI_USAGE_SUMMARY_FOR_MAIN_FILE)
    disk_ok = False
    if wb_path and os.path.isfile(wb_path):
        try:
            disk_ok = _write_main_sheet_gemini_usage_via_openpyxl(
                wb_path, text, log_prefix
            )
        except Exception as ex:
            logging.warning("%s: AI サマリの openpyxl 書き込みで例外: %s", log_prefix, ex)
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
    except OSError:
        pass
    try:
        cum2 = _load_gemini_cumulative_payload()
        if int(cum2.get("calls_total") or 0) > 0:
            _export_gemini_buckets_csv_for_charts(cum2)
    except Exception as ex:
        logging.debug("Gemini バケット CSV 出力で例外（続行）: %s", ex)
    if disk_ok:
        return
    if text.strip():
        logging.info(
            "%s: メイン P 列は openpyxl で保存できませんでした（ブックが Excel で開いている可能性）。"
            " %s に出力済み → マクロ「メインシート_Gemini利用サマリをP列に反映」で反映してください。",
            log_prefix,
            path,
        )
    else:
        logging.info(
            "%s: Gemini 未使用: サマリを空で %s に出力。",
            log_prefix,
            path,
        )


def _try_write_main_sheet_gemini_usage_summary(phase: str) -> None:
    try:
        write_main_sheet_gemini_usage_summary(TASKS_INPUT_WORKBOOK, phase)
    except Exception as ex:
        logging.warning(
            "%s: メインシートへの AI 利用サマリ書き込みで例外（続行）: %s", phase, ex
        )


def write_plan_sheet_global_comment_parse_block(
    wb_path: str,
    sheet_name: str,
    global_priority_override: dict,
    *,
    when_str: str,
    log_prefix: str = "段階2",
) -> bool:
    """
    「配台計画_タスク入力」シートの右端付近（AX:AY）に、グローバルコメントの解析結果を書き込む。
    メイン原文はここに転記しない（メイン欄との重複・誤解を避ける）。本列は再読込されず参照専用。
    Excel でブックを開いたままだと保存に失敗することがある（他の openpyxl 書込と同様）。
    """
    if not wb_path or not os.path.isfile(wb_path):
        return False
    gpo = global_priority_override or {}
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: ブックに「%s」があるため openpyxl でグローバルコメント解析を配台シートへ書き込みません。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    try:
        wb = load_workbook(
            wb_path, keep_vba=keep_vba, read_only=False, data_only=False
        )
    except Exception as ex:
        logging.info(
            "%s: グローバルコメント解析の配台シート書込のためブックを開けません: %s",
            log_prefix,
            ex,
        )
        return False
    try:
        if sheet_name not in wb.sheetnames:
            logging.info(
                "%s: シート '%s' が無いためグローバルコメント解析の反映をスキップ。",
                log_prefix,
                sheet_name,
            )
            return False
        ws = wb[sheet_name]
        lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
        vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
        max_r = PLAN_SHEET_GLOBAL_PARSE_MAX_ROWS
        for i in range(max_r):
            ws.cell(row=1 + i, column=lc, value=None)
            ws.cell(row=1 + i, column=vc, value=None)
        align_top = Alignment(wrap_text=True, vertical="top")
        pairs: list[tuple[str, str]] = [
            ("【グローバルコメント解析】", "参照用・段階2で自動記録"),
            (
                "※二重適用について",
                "配台への反映はメインシート「グローバルコメント」からのみ行われます。"
                "このAX〜AY列は読み取られません。編集しても次回実行まで配台に効きません。"
                "原文はメイン欄を参照してください。",
            ),
            ("計画基準日時", (when_str or "").strip() or "―"),
            (
                "工場休業日",
                ", ".join(str(x) for x in (gpo.get("factory_closure_dates") or []))
                if gpo.get("factory_closure_dates")
                else "（なし）",
            ),
            (
                "スキル要件を無視",
                "はい" if gpo.get("ignore_skill_requirements") else "いいえ",
            ),
            (
                "need人数1固定",
                "はい" if gpo.get("ignore_need_minimum") else "いいえ",
            ),
            (
                "配台制限の撤廃",
                "はい" if gpo.get("abolish_all_scheduling_limits") else "いいえ",
            ),
            (
                "グローバルOP指名",
                json.dumps(gpo.get("task_preferred_operators") or {}, ensure_ascii=False)
                if gpo.get("task_preferred_operators")
                else "（なし）",
            ),
            (
                "グローバル速度ルール",
                json.dumps(gpo.get("global_speed_rules") or [], ensure_ascii=False)
                if gpo.get("global_speed_rules")
                else "（なし）",
            ),
            (
                "未適用メモ(AI)",
                str(gpo.get("scheduler_notes_ja") or "").strip() or "（なし）",
            ),
            (
                "AI要約",
                str(gpo.get("interpretation_ja") or "").strip() or "（なし）",
            ),
        ]
        for i, (lab, val) in enumerate(pairs):
            if i >= max_r:
                break
            c1 = ws.cell(row=1 + i, column=lc, value=lab)
            c2 = ws.cell(row=1 + i, column=vc, value=val)
            c1.alignment = align_top
            c2.alignment = align_top
        wb.save(wb_path)
        logging.info(
            "%s: 「%s」%s:%s 列にグローバルコメント解析を保存しました。",
            log_prefix,
            sheet_name,
            get_column_letter(lc),
            get_column_letter(vc),
        )
        return True
    except OSError as ex:
        logging.warning(
            "%s: グローバルコメント解析を配台シートへ保存できませんでした（Excel で開いたまま等）: %s",
            log_prefix,
            ex,
        )
        return False
    except Exception as ex:
        logging.warning(
            "%s: グローバルコメント解析の配台シート書込で例外: %s", log_prefix, ex
        )
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _try_write_plan_sheet_global_comment_parse_block(
    global_priority_override: dict,
    when_str: str,
) -> None:
    try:
        write_plan_sheet_global_comment_parse_block(
            TASKS_INPUT_WORKBOOK,
            PLAN_INPUT_SHEET_NAME,
            global_priority_override,
            when_str=when_str,
            log_prefix="段階2",
        )
    except Exception as ex:
        logging.warning(
            "段階2: 配台シートへのグローバルコメント解析書き込みで例外（続行）: %s",
            ex,
        )


def _log_task_special_ai_response(raw_text, parsed, extracted_json_str, prompt_text=None):
    """特別指定_備考向け Gemini のプロンプト・生テキスト・抽出JSON・パース結果を1ファイルに残す。"""
    path = os.path.join(log_dir, TASK_SPECIAL_AI_LAST_RESPONSE_FILE)
    try:
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            if prompt_text is not None and str(prompt_text).strip():
                f.write("=== Gemini へ送信したプロンプト（全文） ===\n")
                f.write(str(prompt_text).strip())
                f.write("\n\n")
            f.write("=== Gemini 返却テキスト（モデル出力そのまま） ===\n")
            f.write(raw_text or "")
            f.write(
                "\n\n=== AI が返したテキストからクライアントが切り出した JSON 文字列 ===\n"
                "（※ユーザー特別指定の解析に正規表現は使っていません。モデル応答のパース用です）\n"
            )
            f.write(extracted_json_str if extracted_json_str else "(抽出なし)")
            f.write("\n\n=== json.loads 後（依頼NOキー） ===\n")
            if isinstance(parsed, dict):
                f.write(json.dumps(parsed, ensure_ascii=False, indent=2))
            else:
                f.write("(パースできず)")
        logging.info(
            "タスク特別指定: プロンプト＋AI応答の詳細 → %s",
            path,
        )
    except OSError as ex:
        logging.warning("タスク特別指定: AI応答ファイル保存に失敗: %s", ex)
    if isinstance(parsed, dict) and parsed:
        logging.info(
            "タスク特別指定: 解析された依頼NO: %s",
            ", ".join(sorted(parsed.keys(), key=lambda x: str(x))),
        )
        for tid_k in sorted(parsed.keys(), key=lambda x: str(x)):
            logging.info(
                "  依頼NO [%s] AI解析フィールド: %s",
                tid_k,
                json.dumps(parsed[tid_k], ensure_ascii=False),
            )


def _parse_and_log_task_special_gemini_response(res, prompt_text=None):
    """
    API レスポンスを JSON 化しログ／ファイルへ記録。失敗時は None。
    ユーザーの特別指定文言には触れず、モデル出力から JSON ブロックを取り出す処理のみ。
    """
    raw = _gemini_result_text(res)
    if raw:
        stripped = raw.strip()
        if stripped.startswith("{"):
            try:
                trial = json.loads(stripped)
                if isinstance(trial, dict):
                    _log_task_special_ai_response(raw, trial, stripped, prompt_text)
                    return trial
            except json.JSONDecodeError:
                pass
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        _log_task_special_ai_response(raw, {}, None, prompt_text)
        logging.warning(
            "タスク特別指定: AI応答から JSON を抽出できませんでした。生テキスト先頭 3000 文字:\n%s",
            (raw[:3000] if raw else "(空)"),
        )
        return None
    extracted = match.group(0)
    try:
        parsed = json.loads(extracted)
    except json.JSONDecodeError as je:
        _log_task_special_ai_response(raw, None, extracted, prompt_text)
        logging.warning("タスク特別指定: JSON パース失敗: %s", je)
        return None
    if not isinstance(parsed, dict):
        _log_task_special_ai_response(raw, None, extracted, prompt_text)
        logging.warning("タスク特別指定: トップレベルが JSON オブジェクトではありません。")
        return None
    _log_task_special_ai_response(raw, parsed, extracted, prompt_text)
    return parsed


def analyze_task_special_remarks(tasks_df, reference_year=None, ai_sheet_sink: dict | None = None):
    """
    「配台計画_タスク入力」の「特別指定_備考」を AI で構造化（セルに値がある項目は後段でセルを優先）。
    「配台不要」がオンな行はプロンプトに載せない（API 節約・当該行は配台しないため）。
    担当OP指名はプロンプトの返却契約でモデルに preferred_operator を出力させる（備考を正規表現で切り出す処理は行わない）。
    json/ai_remarks_cache.json に TTL AI_CACHE_TTL_SECONDS でキャッシュ（同一入力・同一基準年なら API を呼ばない）。
    依頼NOは数値表記・全角などを正規化してキーを安定化し、基準年は指紋に含めて日付解釈の変化とキャッシュの食い違いを防ぐ。

    戻り値の例: 依頼NO -> オブジェクト、または同一依頼NOに備考行が複数ある場合はオブジェクトの配列。
      process_name, machine_name … 当該備考セルがある行の工程名・機械名（プロンプトの行と一致）
      restrict_to_process_name, restrict_to_machine_name … 省略または空なら同一依頼NOの全工程・全機械行に適用。
      その他 required_op, speed_override, task_efficiency, priority, start_date, start_time,
      target_completion_date, ship_by_date, preferred_operator など。
    """
    lines = _task_special_prompt_lines(tasks_df)
    if not lines:
        n_rows = len(tasks_df)
        n_rem_only = 0
        n_tid_raw = 0
        for _, row in tasks_df.iterrows():
            tid = planning_task_id_str_from_plan_row(row)
            rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
            if tid:
                n_tid_raw += 1
            if rem:
                n_rem_only += 1
        miss_col = PLAN_COL_SPECIAL_REMARK not in tasks_df.columns
        logging.warning(
            "タスク特別指定: AI 解析対象がありません（「%s」列は%s）。"
            "总行数=%s、依頼NOのある行=%s、備考が入っている行=%s。"
            "段階2実行前にブックを保存し、本当に「%s」列へ入力しているか確認してください。",
            PLAN_COL_SPECIAL_REMARK,
            "見つかりません" if miss_col else "空の可能性があります",
            n_rows,
            n_tid_raw,
            n_rem_only,
            PLAN_COL_SPECIAL_REMARK,
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "スキップ（対象行なし）"
        return {}

    blob = "\n".join(sorted(lines))
    ref_y = int(reference_year) if reference_year is not None else date.today().year
    cache_fingerprint = f"{ref_y}\n{blob}"
    cache_key_input = f"{TASK_SPECIAL_CACHE_KEY_PREFIX}{cache_fingerprint}"
    cache_key = hashlib.sha256(cache_key_input.encode("utf-8")).hexdigest()
    ai_cache = load_ai_cache()
    cached_parsed = get_cached_ai_result(
        ai_cache, cache_key, content_key=cache_fingerprint
    )
    if cached_parsed is not None:
        logging.info(
            "タスク特別指定: キャッシュヒット（%s 件・基準年=%s）。Gemini は呼びません。",
            len(lines),
            ref_y,
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "なし（キャッシュ使用）"
        out = copy.deepcopy(cached_parsed)
        if isinstance(out, dict):
            _repair_task_special_ai_wrong_top_level_keys(out, tasks_df)
        return out

    logging.info(
        "タスク特別指定: キャッシュなし。Gemini で %s 件の備考を解析します（基準年=%s）。",
        len(lines),
        ref_y,
    )

    if not API_KEY:
        logging.info("GEMINI_API_KEY 未設定のためタスク特別指定のAI解析をスキップしました。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "なし（APIキー未設定）"
        return {}

    prompt = f"""
あなたは工場の配台計画向けに、Excel「特別指定_備考」欄への自由記述を読み、配台ロジックが使えるフィールドだけに落とし込むアシスタントです。

【最重要】
1) 【特別指定原文】の各行は、ユーザーがセルに入力した文字列を **改変・要約・断ち切りはしておらず**（先頭末尾の空白のみ除去）、そのまま渡しています。**原文の事実や意図を別の文言に置き換えないでください。**
2) あなたの応答は **1個の JSON オブジェクトのみ**（先頭が {{ 、末尾が }} ）。説明文・マークダウン・コードフェンスは禁止。
3) JSON のトップレベルキーは、各行の **依頼NO【と】の間の文字列のみ** と **完全一致** させること。**備考本文**に書かれた品番・原反名・製品コード（例: 20010 で始まる番号列）をキーにしてはならない。備考がそのような番号で始まっていても、キーは必ず【】内の依頼NOだけとする。

【返却JSONの契約（この節どおりに出力すること）】
■ トップレベル
- キー: 上記【特別指定原文】の **依頼NO【…】の括弧内** の文字列と **完全一致**（表記・ハイフン・英大文字小文字を原文どおり）。備考本文中の数字列をキーにしない。
- 値: 次のいずれか。
  (A) **JSONオブジェクト1つ** … 当該依頼NOの備考がプロンプト上 **1行だけ** のとき。
  (B) **JSON配列**（要素はオブジェクト）… 同一依頼NOで工程名・機械名が異なる備考行が **複数** あるとき。要素の順はプロンプトの行順と対応させる。

■ process_name（文字列）・machine_name（文字列）— **必須**
- 当該備考に対応するプロンプト行の **工程名「…」**・**機械名「…」** の値と **一致** させる（「（空）」のときは空文字列 ""）。
- ログ・トレース用。省略不可。

■ restrict_to_process_name（文字列）・restrict_to_machine_name（文字列）— **任意**
- **原文が「特定の工程だけ」「この機械だけ」など、適用範囲を絞っているときだけ** 出力する。
- **原文に工程名・機械名の限定が無い**（依頼全体・全行程への指示）ときは **両方とも省略** するか **空文字列 ""** とする。
- その場合、配台ロジックは **同一依頼NOの別行（例: エンボス行と分割行）にも同じ指示を適用** する。
- 絞る場合は、原文で示された識別名を入れる（Excel の工程名・機械名と照合しやすい表記）。

■ preferred_operator（文字列）— 条件付き**必須**
- **必要条件**: 当該依頼の原文を読み、「**誰がこの加工・作業の主担当（OP）として割り当てたいか**」が **意味として** 読み取れるとき。
  例: 特定の人にやってもらう／その人に任せる／担当はあの人／OPは〜／〜さん（氏名）に依頼、など。**表現の型に依存せず**、文の意味で判断する。
- **満たしたときの出力義務**: 上記の意味が成立すると判断したオブジェクトでは、**必ず** キー `preferred_operator` を含め、値は **空でない文字列** とする。併せて **process_name / machine_name は必須**（例: `{{"process_name":"…","machine_name":"…","preferred_operator":"…"}}`）。
- **値の形式**: 原文で示された **担当者の識別名を1名分**（姓・名・ニックネーム等、原文に現れた表記を維持）。末尾の敬称（さん・君・氏）のみ除去。例:「森岡さんにやってもらいます」→ `"森岡"`。
- **出力してはいけないとき**: 原文に担当者の指意が **一切ない** と判断した依頼NOでは `preferred_operator` キー自体を **省略** する（空文字列も付けない）。

■ その他フィールド（required_op, speed_override, task_efficiency, priority, start_date, start_time, target_completion_date, ship_by_date）
- 原文から **明確に** 読み取れる場合のみ出力。読み取れない数値・日付は **省略**（推測で埋めない）。

【同一依頼NO・複数工程の例】
依頼NO Y4-2 に「エンボス」と「分割」の行があり、備考が「4/5までに終わらせる」のみで工程の限定が無い場合:
- process_name / machine_name は **備考が書かれた行** の値を入れる。
- restrict_to_* は **出さないか空** にし、**エンボス行・分割行の両方** に同じ優先度・日付等が効くようにする。

【基準年（年なし日付用）】
「4/5」「4/5に出荷」のように **年が無い** 日付は原則 **西暦 {ref_y} 年** とし、YYYY-MM-DD で出力。

【フィールド一覧（型の参考）】
- process_name, machine_name: 文字列（必須。プロンプト行と一致）
- restrict_to_process_name, restrict_to_machine_name: 文字列（任意。限定なら）
- preferred_operator: 文字列（上記契約に従う）
- required_op: 正の整数
- speed_override: 正の数（m/分）
- task_efficiency: 0〜1
- priority: 整数（小さいほど先に割付）
- start_date: YYYY-MM-DD / start_time: HH:MM
- target_completion_date, ship_by_date: YYYY-MM-DD

【解釈の指針】
- 「間に合うように」「繰り上げる」→ priority を上げる（数値を下げる）。日付が文中にあれば target_completion_date または ship_by_date に入れる。
- 担当者指名は **意味理解** で preferred_operator を決める（特定のキーワード列挙に頼らない）。
- 数値・日付は推測で補わない。
- **備考が特定の工程・機械にだけ言及していない限り**、restrict_to_* は空にし、同一依頼NOの他行にも適用される形にする。

【出力直前の自己検証（必ず実行してから JSON を閉じる）】
- 【特別指定原文】の **各行** について、対応するオブジェクトに **process_name** と **machine_name** があるか。
- 同一依頼NOが複数行あるときは **配列** で各行に1オブジェクト、または適切にマージした単一オブジェクト＋restrict の運用を一貫させる。
- 「主担当OPの指意」がある行では **非空の preferred_operator** を付ける。

【出力形式の例】（依頼NO・値は実データに合わせ替えること）
{{
  "W3-14": {{
    "process_name": "検査",
    "machine_name": "ラインA",
    "preferred_operator": "森岡"
  }},
  "Y3-26": {{
    "process_name": "コーティング",
    "machine_name": "",
    "priority": 1,
    "ship_by_date": "{ref_y}-04-05",
    "target_completion_date": "{ref_y}-04-05"
  }},
  "Y4-2": {{
    "process_name": "エンボス",
    "machine_name": "E1",
    "priority": 2,
    "restrict_to_process_name": "",
    "restrict_to_machine_name": ""
  }}
}}

【特別指定原文】（Excel からそのまま。1行＝依頼NOと備考のペア）
{blob}
"""
    try:
        ppath = os.path.join(log_dir, "ai_task_special_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("タスク特別指定: 今回 Gemini に渡したプロンプト全文 → %s", ppath)
    except OSError as ex:
        logging.warning("タスク特別指定: プロンプト保存失敗: %s", ex)

    client = genai.Client(api_key=API_KEY)
    try:
        res = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
        record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
        parsed = _parse_and_log_task_special_gemini_response(res, prompt_text=prompt)
        if parsed is not None:
            _repair_task_special_ai_wrong_top_level_keys(parsed, tasks_df)
            put_cached_ai_result(
                ai_cache, cache_key, parsed, content_key=cache_fingerprint
            )
            save_ai_cache(ai_cache)
            logging.info("タスク特別指定: AI解析が完了しました。")
            if ai_sheet_sink is not None:
                ai_sheet_sink["特別指定備考_AI_API"] = "あり"
            return parsed
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "あり（JSON解釈失敗）"
        return {}
    except Exception as e:
        err_text = str(e)
        is_quota = ("429" in err_text) or ("RESOURCE_EXHAUSTED" in err_text)
        is_unavailable = ("503" in err_text) or ("UNAVAILABLE" in err_text)
        retry_sec = extract_retry_seconds(err_text) if is_quota else None
        if is_quota and retry_sec is not None:
            wait_sec = min(max(retry_sec, 1.0), 90.0)
            logging.warning(f"タスク特別指定 AI 429。{wait_sec:.1f}秒待機して再試行します。")
            time_module.sleep(wait_sec)
            try:
                res = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
                record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
                parsed = _parse_and_log_task_special_gemini_response(res, prompt_text=prompt)
                if parsed is not None:
                    _repair_task_special_ai_wrong_top_level_keys(parsed, tasks_df)
                    put_cached_ai_result(
                        ai_cache, cache_key, parsed, content_key=cache_fingerprint
                    )
                    save_ai_cache(ai_cache)
                    if ai_sheet_sink is not None:
                        ai_sheet_sink["特別指定備考_AI_API"] = "あり（429再試行後）"
                    return parsed
            except Exception as e2:
                logging.warning(f"タスク特別指定 AI 再試行失敗: {e2}")
        elif is_unavailable:
            wait_sec = 8.0
            logging.warning(
                f"タスク特別指定 AI 503/UNAVAILABLE。{wait_sec:.1f}秒待機して再試行します。"
            )
            time_module.sleep(wait_sec)
            try:
                res = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
                record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
                parsed = _parse_and_log_task_special_gemini_response(res, prompt_text=prompt)
                if parsed is not None:
                    _repair_task_special_ai_wrong_top_level_keys(parsed, tasks_df)
                    put_cached_ai_result(
                        ai_cache, cache_key, parsed, content_key=cache_fingerprint
                    )
                    save_ai_cache(ai_cache)
                    logging.info("タスク特別指定: AI再試行で解析が完了しました。")
                    if ai_sheet_sink is not None:
                        ai_sheet_sink["特別指定備考_AI_API"] = "あり（503再試行後）"
                    return parsed
                logging.warning("タスク特別指定 AI 503再試行: JSON 抽出に失敗しました。")
            except Exception as e2:
                logging.warning(f"タスク特別指定 AI 503再試行失敗: {e2}")
        else:
            logging.warning(f"タスク特別指定 AI エラー: {e}")
        logging.warning(
            "タスク特別指定: AI解析結果を取得できなかったため、特別指定_備考の開始日/優先指示は反映されません。"
            " 必要なら『加工開始日_指定』または『指定納期_上書き』を入力してください。"
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = f"失敗: {e}"[:500]
        return {}


def _merge_preferred_operator_cell_and_ai(row, ai_for_tid):
    """Excel「担当OP_指定」を優先し、空なら AI の preferred_operator。"""
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}
    v = row.get(PLAN_COL_PREFERRED_OP)
    if v is not None and not (isinstance(v, float) and pd.isna(v)):
        s = str(v).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
    a = ai.get("preferred_operator")
    if a is not None:
        s = str(a).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
    return ""


def _global_override_preferred_operator_for_task(tpref, task_id) -> str | None:
    """
    メイン「再優先特別記載」の task_preferred_operators。
    キーは依頼NO（大文字・小文字の差は無視）。
    """
    if not isinstance(tpref, dict) or not tpref:
        return None
    tid = str(task_id).strip()
    if not tid:
        return None
    tlo = tid.lower()
    for k, v in tpref.items():
        if str(k).strip().lower() != tlo:
            continue
        s = str(v).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
        return None
    return None


def _merge_task_row_with_ai(
    row, ai_for_tid, *, allow_ai_dispatch_priority_from_remark: bool = True
):
    """
    セル優先で上書き値を決定。ai_for_tid は analyze_task_special_remarks の1エントリ。
    allow_ai_dispatch_priority_from_remark が False のとき、AI の priority / start_date / start_time は
    セルが空でも採用しない（備考に納期系文言が無い行向け）。
    """
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}

    def first_int(cell, ai_key):
        if cell == PLAN_COL_PRIORITY:
            v = parse_optional_int(_planning_row_plan_priority_cell(row))
        else:
            v = parse_optional_int(row.get(cell))
        if v is not None:
            return v
        return parse_optional_int(ai.get(ai_key))

    def first_float_pos(cell, ai_key):
        v = parse_float_safe(row.get(cell), None)
        if v is not None and (not isinstance(v, float) or not pd.isna(v)) and float(v) > 0:
            return float(v)
        a = ai.get(ai_key)
        try:
            if a is not None and float(a) > 0:
                return float(a)
        except (TypeError, ValueError):
            pass
        return None

    req_op = first_int(PLAN_COL_REQUIRED_OP, "required_op")
    if req_op is not None and req_op < 1:
        req_op = None

    te = first_float_pos(PLAN_COL_TASK_EFFICIENCY, "task_efficiency")
    if te is None or te <= 0:
        te = 1.0

    if allow_ai_dispatch_priority_from_remark:
        pri = first_int(PLAN_COL_PRIORITY, "priority")
    else:
        pv = parse_optional_int(_planning_row_plan_priority_cell(row))
        if pv is not None:
            pri = pv
        else:
            pri = parse_optional_int(row.get(PLAN_COL_PRIORITY))
    if pri is None:
        pri = 999

    st_date = parse_optional_date(row.get(PLAN_COL_START_DATE_OVERRIDE))
    if st_date is None and allow_ai_dispatch_priority_from_remark and ai.get("start_date"):
        st_date = parse_optional_date(ai.get("start_date"))

    st_time = parse_time_str(row.get(PLAN_COL_START_TIME_OVERRIDE), None)
    if st_time is None and allow_ai_dispatch_priority_from_remark and ai.get("start_time"):
        st_time = parse_time_str(str(ai.get("start_time")), None)

    speed_ov = first_float_pos(PLAN_COL_SPEED_OVERRIDE, "speed_override")

    return req_op, speed_ov, te, pri, st_date, st_time, ai


def _plan_row_cell_nonempty(row, col_name):
    v = row.get(col_name)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none"):
        return False
    return True


def _ai_int_for_conflict(ai, key):
    if not ai or ai.get(key) is None:
        return None
    return parse_optional_int(ai.get(key))


def _ai_float_for_conflict(ai, key):
    if not ai or ai.get(key) is None:
        return None
    try:
        f = float(ai.get(key))
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def detect_planning_remark_ai_conflicts(row, ai_for_tid):
    """
    特別指定_備考に依る AI 解析結果と、明示セルの両方に値があり食い違う列を返す。
    備考・AIいずれか欠ける場合は空集合。
    """
    remark = str(row.get(PLAN_COL_SPECIAL_REMARK, "") or "").strip()
    if not remark or remark.lower() in ("nan", "none"):
        return set()
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}
    if not ai:
        return set()
    out = set()

    if _plan_row_cell_nonempty(row, PLAN_COL_REQUIRED_OP):
        cv = parse_optional_int(row.get(PLAN_COL_REQUIRED_OP))
        av = _ai_int_for_conflict(ai, "required_op")
        if cv is not None and av is not None and cv != av:
            out.add(PLAN_COL_REQUIRED_OP)

    if _plan_row_cell_nonempty(row, PLAN_COL_SPEED_OVERRIDE):
        cv = parse_float_safe(row.get(PLAN_COL_SPEED_OVERRIDE), None)
        if cv is not None and cv > 0:
            av = _ai_float_for_conflict(ai, "speed_override")
            if av is not None and abs(cv - av) > 1e-5:
                out.add(PLAN_COL_SPEED_OVERRIDE)

    if _plan_row_cell_nonempty(row, PLAN_COL_TASK_EFFICIENCY):
        cv = parse_float_safe(row.get(PLAN_COL_TASK_EFFICIENCY), None)
        if cv is not None and cv > 0:
            av = _ai_float_for_conflict(ai, "task_efficiency")
            if av is not None and abs(cv - av) > 1e-5:
                out.add(PLAN_COL_TASK_EFFICIENCY)

    _pri_cell = parse_optional_int(_planning_row_plan_priority_cell(row))
    if _pri_cell is not None:
        cv = _pri_cell
        av = _ai_int_for_conflict(ai, "priority")
        if cv is not None and av is not None and cv != av:
            out.add(PLAN_COL_PRIORITY)

    if _plan_row_cell_nonempty(row, PLAN_COL_START_DATE_OVERRIDE):
        cv = parse_optional_date(row.get(PLAN_COL_START_DATE_OVERRIDE))
        av = parse_optional_date(ai.get("start_date")) if ai.get("start_date") else None
        if cv is not None and av is not None and cv != av:
            out.add(PLAN_COL_START_DATE_OVERRIDE)

    if _plan_row_cell_nonempty(row, PLAN_COL_START_TIME_OVERRIDE):
        cv = parse_time_str(row.get(PLAN_COL_START_TIME_OVERRIDE), None)
        av = parse_time_str(str(ai.get("start_time")), None) if ai.get("start_time") else None
        if cv is not None and av is not None and cv != av:
            out.add(PLAN_COL_START_TIME_OVERRIDE)

    if _plan_row_cell_nonempty(row, PLAN_COL_PREFERRED_OP):
        cv = _normalize_person_name_for_match(row.get(PLAN_COL_PREFERRED_OP))
        av = _normalize_person_name_for_match(ai.get("preferred_operator"))
        if cv and av and cv != av:
            out.add(PLAN_COL_PREFERRED_OP)

    if out:
        out.add(PLAN_COL_SPECIAL_REMARK)
    return out


def collect_planning_conflicts_by_excel_row(tasks_df, ai_by_tid):
    """Excel 行番号(1始まり・ヘッダー=1行目) -> 矛盾があった列名の集合"""
    res = {}
    for i, (_, row) in enumerate(tasks_df.iterrows()):
        if _plan_row_exclude_from_assignment(row):
            continue
        ai_one = _ai_task_special_entry_for_row(ai_by_tid, row)
        cset = detect_planning_remark_ai_conflicts(row, ai_one)
        if cset:
            res[i + 2] = cset
    return res


def apply_planning_sheet_conflict_styles(wb_path, sheet_name, num_data_rows, conflicts_by_row):
    """
    配台計画_タスク入力シートのデータ行を、矛盾列のみ赤地・白太字にする。
    事前パスでは上書き入力列を段階1と同じ薄黄色に戻し、フォントは変更しない（体裁維持）。
    AI解析列は着色しない（段階1の仕様に合わせる）。
    .xlsm は keep_vba=True で保存する。
    """
    if not wb_path or not os.path.exists(wb_path):
        return
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "矛盾書式: ブックに「%s」があるため openpyxl でのハイライトをスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = load_workbook(wb_path, keep_vba=keep_vba)
    try:
        if sheet_name not in wb.sheetnames:
            logging.warning(f"矛盾書式: シート '{sheet_name}' が見つかりません。")
            return
        ws = wb[sheet_name]
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(1, col_idx).value
            if v is not None:
                header_map[str(v).strip()] = col_idx

        last_row = max(2, 1 + int(num_data_rows))
        clear_fill = PatternFill(fill_type=None)
        yellow_input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        conflict_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        conflict_font = Font(color="FFFFFF", bold=True)

        for r in range(2, last_row + 1):
            for name in PLAN_CONFLICT_STYLABLE_COLS:
                ci = header_map.get(name)
                if not ci:
                    continue
                cell = ws.cell(row=r, column=ci)
                if name == PLAN_COL_AI_PARSE:
                    cell.fill = clear_fill
                else:
                    cell.fill = yellow_input_fill
                # フォントは上書きしない（ブック既定・ユーザー設定を維持）

        for r, colnames in conflicts_by_row.items():
            if r < 2:
                continue
            for name in colnames:
                ci = header_map.get(name)
                if not ci:
                    continue
                cell = ws.cell(row=r, column=ci)
                cell.fill = conflict_fill
                cell.font = conflict_font

        try:
            wb.save(wb_path)
        except OSError as e:
            write_planning_conflict_highlight_sidecar(sheet_name, num_data_rows, conflicts_by_row)
            logging.warning(
                "配台シートへの矛盾ハイライトをファイル保存できませんでした（Excel でブックを開いたまま等）。"
                " '%s' に指示を書き出しました。マクロがシート上に直接適用します。 (%s)",
                _planning_conflict_sidecar_path(),
                e,
            )
        else:
            _remove_planning_conflict_sidecar_safe()
            if conflicts_by_row:
                logging.info(
                    f"特別指定_備考と列の矛盾: {len(conflicts_by_row)} 行を '{sheet_name}' でハイライトしました。"
                )
    finally:
        wb.close()


def _ai_planning_target_due_date(ai_dict):
    """AI JSON の完了・出荷目標日から、配台の目標日1つを決める（複数あれば最も早い日＝厳しい方）。"""
    if not isinstance(ai_dict, dict):
        return None
    dates = []
    for k in ("target_completion_date", "ship_by_date", "latest_ship_date", "due_date"):
        d = parse_optional_date(ai_dict.get(k))
        if d is not None:
            dates.append(d)
    if not dates:
        return None
    return min(dates)


def _special_remark_implies_due_related_dispatch_priority(remark_raw: str) -> bool:
    """
    特別指定_備考に、納期・期限・最優先など「配台試行を前に出す」意図の文言があるとき True。
    備考が記入されているだけでは True にしない（AI 由来の目標日・開始日・優先度は使わない）。
    """
    if not remark_raw:
        return False
    s = str(remark_raw).strip()
    if not s or s.lower() in ("nan", "none"):
        return False
    n = unicodedata.normalize("NFKC", s)
    n_lower = n.casefold()
    needles = (
        "納期",
        "指定納期",
        "回答納期",
        "計画基準",
        "期日",
        "締切",
        "締め切り",
        "期限",
        "最優先",
        "至急",
        "急ぎ",
        "直ちに",
        "早急",
        "出荷",
        "納入",
        "必着",
        "deadline",
        "デッドライン",
        "前倒し",
        "早めに",
        "厳守",
        "までに",
        "間に合わ",
        "間に合い",
        "遅れない",
        "遅延不可",
        "優先配台",
        "先に配台",
        "完了予定",
        "本納期",
        "回答期限",
    )
    return any(w.casefold() in n_lower for w in needles)


def _global_planning_basis_date_from_override(gpo: dict | None, task_id: str):
    """メイン・グローバルコメント AI の task_planning_basis_dates から計画基準納期を取得。"""
    gpo = gpo or {}
    m = gpo.get("task_planning_basis_dates")
    if not isinstance(m, dict):
        return None
    tid = str(task_id or "").strip()
    if not tid:
        return None
    raw_v = m.get(tid)
    if raw_v is None:
        for k, val in m.items():
            if str(k).strip() == tid:
                raw_v = val
                break
    return parse_optional_date(raw_v)


def _task_id_same_machine_due_tiebreak_key(task_id) -> tuple:
    """
    計画基準納期・機械名が同じ帯での試行順。
    Y3-24 は末尾の数値。Y4-1-1 のようにハイフンが2つ以上あるときは「最初の - の直後」の数値部を採用。
    """
    s = str(task_id or "").strip()
    if not s:
        return (2, 10**9, "")
    parts = s.split("-", 1)
    if len(parts) < 2:
        return (2, 10**9, s)
    rest = parts[1]
    if "-" in rest:
        first_seg = rest.split("-", 1)[0]
        try:
            return (0, int(first_seg), s)
        except ValueError:
            return (1, 10**9, s)
    tail = rest.strip()
    try:
        return (0, int(tail), s)
    except ValueError:
        return (1, 10**9, s)


def validate_no_duplicate_explicit_plan_priorities(tasks_df) -> None:
    """
    配台ルール: 「優先度」列に同一の整数（明示入力）が2行以上あると段階2を中止。
    999 は未入力扱いのため重複チェックから除外。
    """
    from collections import Counter

    if tasks_df is None or getattr(tasks_df, "empty", True):
        return
    c = Counter()
    for _, row in tasks_df.iterrows():
        if row_has_completion_keyword(row):
            continue
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        if not tid:
            continue
        pv = _planning_row_plan_priority_cell(row)
        p = parse_optional_int(pv)
        if p is None:
            continue
        if p == 999:
            continue
        c[p] += 1
    dupes = sorted([k for k, v in c.items() if v > 1])
    if dupes:
        msg = (
            "【配台中止】配台計画_タスク入力の「優先度」が重複しています（同じ数値が複数行）。"
            "重複している値: "
            + ", ".join(str(x) for x in dupes)
            + "。優先度を一意に直してから段階2を再実行してください。"
        )
        _write_stage2_blocking_message(msg)
        raise PlanningValidationError(msg)


# ---------------------------------------------------------------------------
# 配台用タスクキュー
#   配台計画 DataFrame 1行 → 割付アルゴリズム用 dict への変換（優先度・納期・AI 上書きを集約）
# ---------------------------------------------------------------------------
def build_task_queue_from_planning_df(
    tasks_df,
    run_date,
    req_map,
    ai_by_tid=None,
    global_priority_override=None,
    equipment_list=None,
):
    """
    ``generate_plan`` 内で呼ばれる。完了済み・配台不要行を除き、残りを task_queue に積む。
    ai_by_tid が None のときだけ内部で analyze_task_special_remarks を実行する。
    """
    if ai_by_tid is None:
        ai_by_tid = analyze_task_special_remarks(tasks_df, reference_year=run_date.year)
    task_queue = []
    n_exclude_plan = 0
    seq_by_tid = _collect_process_content_order_by_task_id(tasks_df)
    same_tid_line_seq = defaultdict(int)
    # 依頼NO直列配台の順序用: iterrows の読み込み順（0 始まり）。task_queue.sort 後も不変。
    planning_sheet_row_seq = 0

    for _, row in tasks_df.iterrows():
        if row_has_completion_keyword(row):
            continue
        if _plan_row_exclude_from_assignment(row):
            n_exclude_plan += 1
            continue

        task_id = planning_task_id_str_from_plan_row(row)
        machine = str(row.get(TASK_COL_MACHINE, "")).strip()
        machine_name = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        qty_total = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
        done_qty = calc_done_qty_equivalent_from_row(row)
        speed_raw = row.get(TASK_COL_SPEED, 1)
        product_name = row.get(TASK_COL_PRODUCT, None)
        answer_due = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_ANSWER_DUE))
        specified_due = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_SPECIFIED_DUE))
        specified_due_ov = parse_optional_date(
            _planning_df_cell_scalar(row, PLAN_COL_SPECIFIED_DUE_OVERRIDE)
        )
        # 計画基準納期（due_basis）の採用順（配台ルール改定）:
        # 0) メイン・グローバルコメント AI の task_planning_basis_dates（依頼NO別）
        # 1) 特別指定_備考 AI の完了・出荷目標日 … 備考に納期・期限・最優先等の文言がある行のみ
        # 2) 特別指定_備考 AI の start_date（目標日が無い場合の締め）… 同上
        # 3) 指定納期_上書き（空白は無視済み）
        # 4) 原反投入日（原反がある行は計画基準納期＝原反投入日）
        # 5) 回答納期
        # 6) 指定納期
        due_basis = None
        due_source = "none"
        due_source_rank = 9
        raw_input_date = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE))

        qty = max(0.0, qty_total - done_qty)
        speed = parse_float_safe(speed_raw, 1.0)
        if speed <= 0:
            speed = 1.0

        if qty <= 0 or not machine or not task_id:
            continue

        _line_seq = same_tid_line_seq[task_id]
        same_tid_line_seq[task_id] += 1

        remark_raw = str(row.get(PLAN_COL_SPECIAL_REMARK, "") or "").strip()
        has_special_remark = bool(remark_raw) and remark_raw.lower() not in ("nan", "none")
        remark_implies_due_dispatch_priority = (
            _special_remark_implies_due_related_dispatch_priority(remark_raw)
        )
        in_progress = done_qty > 0.0
        has_done_deadline_override = False

        ai_one = _ai_task_special_entry_for_row(ai_by_tid, row)
        req_op, speed_ov, task_eff_factor, priority, start_date_ov, start_time_ov, ai_used = _merge_task_row_with_ai(
            row,
            ai_one,
            allow_ai_dispatch_priority_from_remark=remark_implies_due_dispatch_priority,
        )
        preferred_operator_raw = _merge_preferred_operator_cell_and_ai(row, ai_one)
        gpo = global_priority_override or {}
        gop_name = _global_override_preferred_operator_for_task(
            gpo.get("task_preferred_operators"), task_id
        )
        if gop_name is not None:
            preferred_operator_raw = gop_name
            logging.info(
                "メイン再優先特記: 依頼NO=%s の担当OPをグローバル指名で上書き %r（セル・特別指定備考AIより優先）",
                task_id,
                gop_name,
            )

        ai_target_due = _ai_planning_target_due_date(ai_used)
        ai_start_date = None
        if isinstance(ai_used, dict) and ai_used.get("start_date") is not None:
            ai_start_date = parse_optional_date(ai_used.get("start_date"))

        g_basis = _global_planning_basis_date_from_override(gpo, task_id)
        if g_basis is not None:
            due_basis = g_basis
            due_source = "global_comment_ai_due"
            due_source_rank = 0
            has_done_deadline_override = True
        elif remark_implies_due_dispatch_priority and ai_target_due is not None:
            due_basis = ai_target_due
            due_source = "ai_target_due"
            due_source_rank = 1
            has_done_deadline_override = True
        elif remark_implies_due_dispatch_priority and ai_start_date is not None:
            due_basis = ai_start_date
            due_source = "ai_start_date"
            due_source_rank = 2
            has_done_deadline_override = True
        elif specified_due_ov is not None:
            due_basis = specified_due_ov
            due_source = "specified_due_override"
            due_source_rank = 3
            has_done_deadline_override = True
        elif raw_input_date is not None:
            due_basis = raw_input_date
            due_source = "raw_input_same_as_plan_basis"
            due_source_rank = 4
        elif answer_due is not None:
            due_basis = answer_due
            due_source = "answer_due"
            due_source_rank = 5
        elif specified_due is not None:
            due_basis = specified_due
            due_source = "specified_due"
            due_source_rank = 6

        if speed_ov is not None:
            speed = speed_ov
        if speed <= 0:
            speed = 1.0

        gsm = _global_speed_multiplier_for_row(
            machine, machine_name, gpo.get("global_speed_rules") or []
        )
        if abs(gsm - 1.0) > 1e-12:
            speed_before_g = speed
            speed = speed * gsm
            if speed <= 0:
                speed = 1.0
            logging.info(
                "メイングローバル: 依頼NO=%s 工程=%r 機械名=%r に speed_multiplier 累積=%s を適用（速度 %s → %s）",
                task_id,
                machine,
                machine_name,
                gsm,
                speed_before_g,
                speed,
            )

        unit = infer_unit_m_from_product_name(product_name, fallback_unit=qty_total if qty_total > 0 else qty)
        try:
            unit = float(unit)
        except Exception:
            unit = qty
        if unit <= 0:
            unit = qty

        # 納期は優先順位・緊急度には使うが、開始日の下限には使わない（余力があれば前倒し開始するため）。
        if due_basis is None:
            due_urgent = False
        else:
            due_urgent = due_basis <= run_date

        # 開始日ルール:
        # 1) 原反投入日があるときは「原反当日から」可（同日開始は 13:00 以降。same_day_raw_start_limit）
        #    原反が無いときは run_date
        # 2) 特別指定（セル/AI）の開始日がある場合はそれを優先
        if raw_input_date:
            effective_start_date = max(run_date, raw_input_date)
        else:
            effective_start_date = run_date
        if start_date_ov is not None:
            effective_start_date = start_date_ov
            if raw_input_date and start_date_ov <= raw_input_date:
                logging.info(
                    "開始日上書きを優先: 依頼NO=%s 指定開始日=%s 原反投入日=%s（当日開始を許容）",
                    task_id,
                    start_date_ov,
                    raw_input_date,
                )

        same_day_raw_start_limit = (
            time(13, 0)
            if (raw_input_date and start_date_ov is None and effective_start_date == raw_input_date)
            else None
        )

        calc_time_val = qty * speed
        ai_note = ""
        if ai_used:
            try:
                ai_note = json.dumps(ai_used, ensure_ascii=False)[:500]
            except Exception:
                ai_note = str(ai_used)[:500]

        _order_list = seq_by_tid.get(task_id) or []
        _p_rank = _process_sequence_rank_for_machine(machine, _order_list)
        _init_rem = float(qty / unit if unit else 0.0)
        _process_content_mismatch = bool(_order_list) and not _process_name_matches_kakou_content_tokens(
            machine, _order_list
        )

        task_queue.append(
            {
                "task_id": task_id,
                "machine": machine,
                "machine_name": machine_name,
                "equipment_line_key": _resolve_equipment_line_key_for_task(
                    {"machine": machine, "machine_name": machine_name},
                    equipment_list,
                ),
                "start_date_req": effective_start_date,
                "answer_due_date": answer_due,
                "specified_due_date": specified_due,
                "specified_due_override": specified_due_ov,
                "due_basis_date": due_basis,
                # 納期後ろ倒し再試行で due_basis_date を内部 +1 しても、結果_タスク一覧の計画基準納期はこの値のまま
                "due_basis_date_result_sheet": due_basis,
                "due_source": due_source,
                "due_source_rank": due_source_rank,
                "due_urgent": due_urgent,
                "raw_input_date": raw_input_date,
                "same_day_raw_start_limit": same_day_raw_start_limit,
                "total_qty_m": int(qty_total),
                "unit_m": int(unit),
                "remaining_units": qty / unit if unit else 0,
                "base_time_per_unit": (qty / speed) / (qty / unit) if unit and speed and qty else 0,
                "assigned_history": [],
                "calc_time_value": calc_time_val,
                "required_op": req_op,
                "task_eff_factor": task_eff_factor,
                "priority": priority,
                "earliest_start_time": start_time_ov,
                "preferred_operator_raw": preferred_operator_raw,
                "task_special_ai_note": ai_note,
                "in_progress": in_progress,
                "has_special_remark": has_special_remark,
                "has_done_deadline_override": has_done_deadline_override,
                "done_qty_reported": done_qty,
                "process_sequence_rank": _p_rank,
                "same_request_line_seq": _line_seq,
                "initial_remaining_units": _init_rem,
                "roll_pipeline_ec": _row_matches_roll_pipeline_ec(machine, machine_name),
                "roll_pipeline_inspection": _row_matches_roll_pipeline_inspection(
                    machine, machine_name
                ),
                "process_content_mismatch": _process_content_mismatch,
                "planning_sheet_row_seq": planning_sheet_row_seq,
            }
        )
        planning_sheet_row_seq += 1

    logging.info(
        "task_queue 構築完了: total=%s（配台不要によりスキップ %s 行）",
        len(task_queue),
        n_exclude_plan,
    )
    return task_queue


def _task_id_priority_key(task_id):
    """
    依頼NOの同条件タイブレーク用キー。
    例: Y3-24, Y3-34 のような場合はハイフン後半の数値が小さい方を優先。
    """
    s = str(task_id or "").strip()
    if not s:
        return ("", 10**9, "")
    parts = s.rsplit("-", 1)
    if len(parts) == 2:
        head = parts[0].strip()
        tail = parts[1].strip()
        if re.match(r"^\d+$", tail):
            return (head, int(tail), s)
    return (s, 10**9, s)


def _serial_dispatch_order_task_ids(task_queue) -> list:
    """
    依頼NO直列配台の処理順。配台計画 DataFrame の読み込み行順（各依頼NOについて最も早い行の
    planning_sheet_row_seq が小さいほど先）で完走させる。初回 task_queue.sort（納期・優先度）
    の並びとは独立。
    """
    first_seq_by_tid: dict = {}
    for t in task_queue:
        tid = str(t.get("task_id", "") or "").strip()
        if not tid:
            continue
        seq = t.get("planning_sheet_row_seq")
        seq = int(seq) if seq is not None else 10**9
        prev = first_seq_by_tid.get(tid)
        if prev is None or seq < prev:
            first_seq_by_tid[tid] = seq
    return sorted(
        first_seq_by_tid.keys(),
        key=lambda tid: (first_seq_by_tid[tid], _task_id_priority_key(tid)),
    )


def _excel_scalar_to_plan_string_cell(v):
    """
    既存シート（read_excel）由来のスカラーを、配台計画 DataFrame の文字列列（StringDtype）へ
    代入できる str に正規化する。Excel が数値として保持した優先度 1 → \"1\" など。
    """
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, bool):
        return str(v).lower()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and math.isfinite(v) and float(int(v)) == v:
            return str(int(v))
        if isinstance(v, float) and math.isfinite(v):
            s = str(v)
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return s
        return str(int(v))
    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return ""
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.strftime("%Y/%m/%d")
        return v.strftime("%Y/%m/%d %H:%M")
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.date().strftime("%Y/%m/%d")
        return v.strftime("%Y/%m/%d %H:%M")
    if isinstance(v, date):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _merge_plan_sheet_user_overrides(out_df):
    """
    ブック内の「配台計画_タスク入力」にユーザーが入力した上書き列を、
    段階1の抽出結果へ (依頼NO, 工程名) 単位で引き継ぐ。
    空のセルはマージしない（新規抽出側の空のまま）。
    """
    if out_df is None or out_df.empty:
        return out_df
    if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
        return out_df
    try:
        df_old = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=PLAN_INPUT_SHEET_NAME)
    except Exception as e:
        logging.info("段階1: 既存の配台シートを読めないため上書き継承をスキップ (%s)", e)
        return out_df
    df_old.columns = df_old.columns.str.strip()
    df_old = _rename_legacy_plan_input_ref_columns(df_old)
    df_old = _align_dataframe_headers_to_canonical(
        df_old,
        plan_input_sheet_column_order(),
    )
    if TASK_COL_TASK_ID not in df_old.columns or TASK_COL_MACHINE not in df_old.columns:
        return out_df

    lookup = {}
    for _, r in df_old.iterrows():
        tid = planning_task_id_str_from_plan_row(r)
        mach = str(r.get(TASK_COL_MACHINE, "") or "").strip()
        if not tid or not mach:
            continue
        key = (tid, mach)
        bucket = lookup.setdefault(key, {})
        for c in PLAN_STAGE1_MERGE_COLUMNS:
            if c not in df_old.columns or c not in out_df.columns:
                continue
            v = r.get(c)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            if isinstance(v, str):
                s = v.strip()
                if not s or s.lower() in ("nan", "none"):
                    continue
            bucket[c] = v

    if not lookup:
        return out_df

    merged_rows = 0
    for i, row in out_df.iterrows():
        tid = planning_task_id_str_from_plan_row(row)
        mach = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        bucket = lookup.get((tid, mach))
        if not bucket:
            continue
        merged_rows += 1
        for c, v in bucket.items():
            if c == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
                v = _coerce_plan_exclude_column_value_for_storage(v)
            elif c in out_df.columns and pd.api.types.is_string_dtype(out_df[c].dtype):
                v = _excel_scalar_to_plan_string_cell(v)
            out_df.at[i, c] = v

    if merged_rows:
        logging.info(
            "段階1: 既存シートから上書き列を %s 行へ引き継ぎました（キー: 依頼NO+工程名）。",
            merged_rows,
        )
    return out_df


# ---------------------------------------------------------------------------
# 配台不要（2系統）
#   (A) DataFrame 上のルール … 同一依頼NO×同一機械で「分割」行に yes（手入力は上書きしない）
#   (B) マクロブック「設定_配台不要工程」… 工程+機械ごとの C/D/E 列、Gemini で D→E、
#       保存ロック時は xlwings で A:E 同期→Save のフォールバックあり
#   いずれも apply_exclude_rules_config_to_plan_df で計画 DataFrame に反映される。
# ---------------------------------------------------------------------------

def _auto_exclude_cell_empty_for_autofill(v) -> bool:
    """配台不要セルが未入力のときだけ自動で yes を書き込む。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return True
    if isinstance(v, str):
        s = str(v).strip()
        return not s or s.lower() in ("nan", "none")
    return False


def _normalize_task_id_for_dup_grouping(raw) -> str:
    """同一依頼NOのグルーピング用（表記ゆれ・英字の大小を寄せる）。"""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    if isinstance(raw, float) and raw == int(raw):
        s = str(int(raw))
    else:
        s = unicodedata.normalize("NFKC", str(raw).strip())
    s = s.strip()
    if not s or s.lower() == "nan":
        return ""
    return s.upper()


def _process_name_is_bunkatsu_for_auto_exclude(raw) -> bool:
    """工程名が「分割」（空白除去・NFKC 後）。"""
    t = unicodedata.normalize("NFKC", str(raw or "").strip())
    t = re.sub(r"[\s　]+", "", t)
    return t == "分割"


def _apply_auto_exclude_bunkatsu_duplicate_machine(
    df: pd.DataFrame, log_prefix: str = "段階1"
) -> pd.DataFrame:
    """
    同一依頼NOが2行以上あり、かつ空でない同一機械名が2行以上あるグループでは、
    工程名が「分割」の行の「配台不要」に yes を入れる（セルが空のときのみ）。
    機械名は _normalize_equipment_match_key で重複判定。
    """
    if df is None or df.empty:
        return df
    need_cols = (TASK_COL_TASK_ID, TASK_COL_MACHINE, TASK_COL_MACHINE_NAME)
    for c in need_cols:
        if c not in df.columns:
            return df
    if PLAN_COL_EXCLUDE_FROM_ASSIGNMENT not in df.columns:
        df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = ""
    # read_excel 等で StringDtype になると数値・真偽の .at 代入で TypeError になるため object に寄せる
    df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT].astype(object)

    by_tid = defaultdict(list)
    for i in df.index:
        tid = _normalize_task_id_for_dup_grouping(df.at[i, TASK_COL_TASK_ID])
        if not tid:
            continue
        by_tid[tid].append(i)

    n_set = 0
    for _tid_key, idx_list in by_tid.items():
        if len(idx_list) < 2:
            continue
        counts = defaultdict(int)
        for i in idx_list:
            mn_key = _normalize_equipment_match_key(df.at[i, TASK_COL_MACHINE_NAME])
            if not mn_key:
                continue
            counts[mn_key] += 1
        if not any(c >= 2 for c in counts.values()):
            continue
        for i in idx_list:
            if not _process_name_is_bunkatsu_for_auto_exclude(df.at[i, TASK_COL_MACHINE]):
                continue
            if not _auto_exclude_cell_empty_for_autofill(
                df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT]
            ):
                continue
            # 列が StringDtype のとき int 代入で TypeError になるため文字列にする（_plan_row_exclude_from_assignment は yes を真とみなす）
            df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = "yes"
            n_set += 1

    if n_set:
        logging.info(
            "%s: 同一依頼NOかつ同一機械名が複数行あるグループで、工程名「分割」の行 %s 件に「配台不要」=yes を自動設定しました。",
            log_prefix,
            n_set,
        )
    return df


def _normalize_process_name_for_rule_match(raw) -> str:
    """工程名のルール照合（NFKC・空白除去）。"""
    t = unicodedata.normalize("NFKC", str(raw or "").strip())
    t = re.sub(r"[\s　]+", "", t)
    return t


def _exclude_rules_sheet_header_map(ws) -> dict:
    """1行目見出し → 列番号(1始まり)。
    openpyxl は新規シート直後に max_column が 0 のままのことがあり、見出しが読めず保存前に return してしまう。
    そのため最低 A～E 列は必ず走査する。
    """
    h = {}
    last_col = max(5, int(ws.max_column or 0))
    for col in range(1, last_col + 1):
        v = ws.cell(1, col).value
        if v is not None:
            h[str(v).strip()] = col
    return h


def _ensure_exclude_rules_sheet_headers_and_columns(ws, log_prefix: str) -> tuple[int, int, int, int, int]:
    """
    1行目に標準見出し（工程名・機械名・配台不要・配台不能ロジック・ロジック式）があることを保証する。
    手動で空シートだけ追加した場合は A1:E1 が空のため、ここで書き込んで列番号を返す。
    """
    headers = (
        EXCLUDE_RULE_COL_PROCESS,
        EXCLUDE_RULE_COL_MACHINE,
        EXCLUDE_RULE_COL_FLAG,
        EXCLUDE_RULE_COL_LOGIC_JA,
        EXCLUDE_RULE_COL_LOGIC_JSON,
    )
    hm = _exclude_rules_sheet_header_map(ws)
    if all(hm.get(x) for x in headers):
        return tuple(hm[x] for x in headers)
    for i, name in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=name)
    logging.info(
        "%s: 「%s」の見出しが無い／列名が一致しないため、標準の1行目（A1:E1）を設定しました。",
        log_prefix,
        EXCLUDE_RULES_SHEET_NAME,
    )
    return (1, 2, 3, 4, 5)


def _compact_exclude_rules_data_rows(
    ws,
    c_proc: int,
    c_mach: int,
    c_flag: int,
    c_d: int,
    c_e: int,
    log_prefix: str,
) -> tuple[int, int]:
    """
    2 行目以降から「空行」を除いて上に詰める（元の並びは維持、ソートしない）。
    空行: 工程名が空、または A～E 相当の5セルがすべて空白相当。
    Returns (残したデータ行数, 削除した行数).
    """
    max_r = int(ws.max_row or 1)
    if max_r < 2:
        return 0, 0

    old_body = max_r - 1
    cols = (c_proc, c_mach, c_flag, c_d, c_e)
    rows: list[tuple[str, str, object, object, object]] = []
    for r in range(2, max_r + 1):
        pv = ws.cell(row=r, column=c_proc).value
        mv = ws.cell(row=r, column=c_mach).value
        cv = ws.cell(row=r, column=c_flag).value
        dv = ws.cell(row=r, column=c_d).value
        ev = ws.cell(row=r, column=c_e).value
        all_blank = all(
            _cell_is_blank_for_rule(ws.cell(row=r, column=c).value) for c in cols
        )
        p = str(pv).strip() if pv is not None and not (isinstance(pv, float) and pd.isna(pv)) else ""
        m = str(mv).strip() if mv is not None and not (isinstance(mv, float) and pd.isna(mv)) else ""
        if all_blank or not p:
            continue
        rows.append((p, m, cv, dv, ev))

    n_skip = old_body - len(rows)

    if not rows:
        ws.delete_rows(2, old_body)
        if old_body > 0:
            logging.info(
                "%s: 「%s」は有効なデータ行が無かったため、データ行 %s 行を削除しました。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
                old_body,
            )
        return 0, n_skip

    ws.delete_rows(2, old_body)
    for i, (p, m, cv, dv, ev) in enumerate(rows, start=2):
        ws.cell(row=i, column=c_proc, value=p)
        ws.cell(row=i, column=c_mach, value=m)
        ws.cell(row=i, column=c_flag, value=cv)
        ws.cell(row=i, column=c_d, value=dv)
        ws.cell(row=i, column=c_e, value=ev)

    if n_skip:
        logging.info(
            "%s: 「%s」から空行を %s 件削除し、%s 行に詰めました（並び順は維持）。",
            log_prefix,
            EXCLUDE_RULES_SHEET_NAME,
            n_skip,
            len(rows),
        )
    return len(rows), n_skip


def _cell_is_blank_for_rule(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    s = str(v).strip()
    return not s or s.lower() in ("nan", "none", "null")


def _exclude_rule_c_column_is_yes(v) -> bool:
    """C列「配台不要」がオン（この工程+機械パターンは常に配台不要）。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return int(v) == 1
        except (TypeError, ValueError):
            pass
    s = unicodedata.normalize("NFKC", str(v).strip()).lower()
    return s in ("yes", "true", "1", "y", "はい", "○", "〇", "●")


def _task_row_matches_exclude_rule_target(
    task_proc: str, task_mach: str, rule_proc: str, rule_mach: str
) -> bool:
    if _normalize_process_name_for_rule_match(task_proc) != _normalize_process_name_for_rule_match(
        rule_proc
    ):
        return False
    rm = str(rule_mach or "").strip()
    if not rm:
        return True
    return _normalize_equipment_match_key(task_mach) == _normalize_equipment_match_key(rm)


def _collect_process_machine_pairs_for_exclude_rules(df_src: pd.DataFrame) -> list[tuple[str, str]]:
    """加工計画DATA から、段階1と同じ抽出条件で (工程名, 機械名) の一覧（重複除く・順序維持）。"""
    out: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for _, row in df_src.iterrows():
        if row_has_completion_keyword(row):
            continue
        task_id = planning_task_id_str_from_scalar(row.get(TASK_COL_TASK_ID))
        machine = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        machine_name = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        qty_total = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
        done_qty = calc_done_qty_equivalent_from_row(row)
        qty = max(0.0, qty_total - done_qty)
        if qty <= 0 or not machine or not task_id:
            continue
        key = (
            _normalize_process_name_for_rule_match(machine),
            _normalize_equipment_match_key(machine_name),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append((machine, machine_name))
    return out


def _parse_exclude_rule_json_cell(raw) -> dict | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).strip()
    if not s:
        return None
    fence = re.search(
        r"```(?:json)?\s*(\{.*\})\s*```",
        s,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        s = fence.group(1).strip()
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, dict) else None


def _validate_exclude_rule_parsed_dict(o: object) -> dict | None:
    """Gemini／E列から得た dict が配台不要ルールとして有効か。"""
    if not isinstance(o, dict):
        return None
    if int(o.get("version") or 0) != 1:
        return None
    mode = str(o.get("mode") or "").strip().lower()
    if mode not in ("always_exclude", "conditions"):
        return None
    return o


def _exclude_rule_de_cache_key(stripped_blob: str) -> str:
    """「配台不能ロジック」文言（正規化済み）に対する ai_remarks_cache 用キー。"""
    h = hashlib.sha256(stripped_blob.encode("utf-8")).hexdigest()
    return f"{AI_CACHE_KEY_PREFIX_EXCLUDE_RULE_DE}:{h}"


def _cache_get_exclude_rule_de_parsed(cache_obj: dict, blob: str) -> dict | None:
    s = str(blob or "").strip()
    if not s:
        return None
    data = get_cached_ai_result(
        cache_obj, _exclude_rule_de_cache_key(s), content_key=s
    )
    if not isinstance(data, dict):
        return None
    return _validate_exclude_rule_parsed_dict(data)


def _cache_put_exclude_rule_de_parsed(
    cache_obj: dict, blob: str, parsed: dict | None
) -> None:
    if parsed is None:
        return
    s = str(blob or "").strip()
    if not s:
        return
    put_cached_ai_result(
        cache_obj, _exclude_rule_de_cache_key(s), parsed, content_key=s
    )


def _exclude_rule_logic_gemini_schema_instructions() -> str:
    allowed = ", ".join(sorted(EXCLUDE_RULE_ALLOWED_COLUMNS))
    return (
        "【スキーマ version は必ず 1】\n"
        "1) 常に配台不要（説明が条件なしで外す意味）のとき:\n"
        '{"version":1,"mode":"always_exclude"}\n\n'
        "2) 列の条件で配台不要とするとき:\n"
        '{"version":1,"mode":"conditions","require_all": true または false,"conditions":[ ... ]}\n\n'
        "conditions の各要素:\n"
        "- {\"column\":\"列名\",\"op\":\"empty\"} … セルが空\n"
        "- {\"column\":\"列名\",\"op\":\"not_empty\"}\n"
        "- {\"column\":\"列名\",\"op\":\"eq\",\"value\":\"文字列\"} / ne / contains / not_contains / regex（正規表現）\n"
        "- {\"column\":\"列名\",\"op\":\"gt\"|\"gte\"|\"lt\"|\"lte\",\"value\":数値} … 数値比較（列は数として解釈）\n\n"
        f"【使用可能な列名のみ】（これ以外は使わない）:\n{allowed}\n"
    )


def _parse_exclude_rule_json_array_response(text: str) -> list | None:
    """モデル応答から JSON 配列を取り出す（```json フェンス付き可）。"""
    s = (text or "").strip()
    if not s:
        return None
    fence = re.search(
        r"```(?:json)?\s*(\[.*\])\s*```",
        s,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        s = fence.group(1).strip()
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, list) else None


def _row_scalar_for_exclude_rule(row, col_name: str):
    try:
        return row.get(col_name)
    except Exception:
        return None


def _evaluate_exclude_rule_one_condition(cond: dict, row) -> bool:
    if not isinstance(cond, dict):
        return False
    col = cond.get("column")
    if col not in EXCLUDE_RULE_ALLOWED_COLUMNS:
        logging.warning("配台不要ルール: 未対応の列名をスキップしました: %s", col)
        return False
    op = str(cond.get("op") or "").strip().lower()
    val = _row_scalar_for_exclude_rule(row, col)
    val_s = "" if val is None or (isinstance(val, float) and pd.isna(val)) else str(val).strip()
    val_s_lower = val_s.lower()

    if op == "empty":
        return val_s == ""
    if op == "not_empty":
        return val_s != ""

    if op in ("contains", "not_contains", "regex", "eq", "ne"):
        rhs = cond.get("value", "")
        pat = "" if rhs is None else str(rhs)
        if op == "contains":
            return pat in val_s
        if op == "not_contains":
            return pat not in val_s
        if op == "regex":
            try:
                return re.search(pat, val_s) is not None
            except re.error:
                return False
        if op == "eq":
            return val_s == pat
        if op == "ne":
            return val_s != pat

    def _num(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    nv = _num(val)
    cv = _num(cond.get("value"))
    if nv is None or cv is None:
        return False
    if op == "gt":
        return nv > cv
    if op == "gte":
        return nv >= cv
    if op == "lt":
        return nv < cv
    if op == "lte":
        return nv <= cv
    return False


def evaluate_exclude_rule_json_for_row(rule: dict, row) -> bool:
    """
    E列の JSON（version=1）を評価し、当該タスク行を配台不要とすべきなら True。
    mode: always_exclude | conditions
    """
    if not isinstance(rule, dict) or int(rule.get("version") or 0) != 1:
        return False
    mode = str(rule.get("mode") or "").strip().lower()
    if mode == "always_exclude":
        return True
    if mode != "conditions":
        return False
    conds = rule.get("conditions")
    if not isinstance(conds, list) or not conds:
        return False
    require_all = bool(rule.get("require_all", True))
    checks = []
    for c in conds:
        if isinstance(c, dict) and c.get("column") in EXCLUDE_RULE_ALLOWED_COLUMNS:
            checks.append(_evaluate_exclude_rule_one_condition(c, row))
    if not checks:
        return False
    return all(checks) if require_all else any(checks)


def _ai_compile_exclude_rule_logic_to_json(natural_language: str) -> dict | None:
    """
    D列の自然言語を Gemini で JSON ルールに変換。失敗時 None。
    json/ai_remarks_cache.json に TTL 付きでキャッシュ（同一文言なら API を呼ばない）。
    """
    blob = str(natural_language or "").strip()
    if not blob:
        return None
    ai_cache = load_ai_cache()
    hit = _cache_get_exclude_rule_de_parsed(ai_cache, blob)
    if hit is not None:
        logging.info("配台不要ルール: AIキャッシュヒット（配台不能ロジック→JSON）")
        return hit
    if not API_KEY:
        return None
    schema = _exclude_rule_logic_gemini_schema_instructions()
    prompt = (
        "あなたは工場の配台システム用です。次の「配台不能の説明」を、タスク1行を判定する機械可読ルールに変換してください。\n\n"
        "【出力】先頭が { で終わりが } の JSON オブジェクト1つのみ（説明・マークダウン禁止）。\n\n"
        f"{schema}\n"
        f"【説明文】\n{blob}\n"
    )
    try:
        ppath = os.path.join(log_dir, "ai_exclude_rule_logic_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("配台不要ルール: プロンプト → %s", ppath)
    except OSError as ex:
        logging.warning("配台不要ルール: プロンプト保存失敗: %s", ex)
    try:
        client = genai.Client(api_key=API_KEY)
        res = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
        record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
        raw = (_gemini_result_text(res) or "").strip()
        rpath = os.path.join(log_dir, "ai_exclude_rule_logic_last_response.txt")
        try:
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(raw)
        except OSError:
            pass
        parsed = _validate_exclude_rule_parsed_dict(_parse_exclude_rule_json_cell(raw))
        if parsed:
            _cache_put_exclude_rule_de_parsed(ai_cache, blob, parsed)
            save_ai_cache(ai_cache)
        return parsed
    except Exception as e:
        logging.warning("配台不要ルール: Gemini 変換失敗: %s", e)
        return None


def _ai_compile_exclude_rule_logics_batch(blobs: list[str]) -> list[dict | None]:
    """
    複数の D 列文言を 1 回の Gemini 呼び出しで JSON 化。失敗・要素数不一致時は 1 件ずつにフォールバック。
    json/ai_remarks_cache.json にヒットした文言は API を呼ばない。
    """
    n = len(blobs)
    if n == 0:
        return []
    ai_cache = load_ai_cache()
    out: list[dict | None] = [None] * n
    pend_i: list[int] = []
    pend_b: list[str] = []
    for i, b in enumerate(blobs):
        s = str(b).strip()
        hit = _cache_get_exclude_rule_de_parsed(ai_cache, s) if s else None
        if hit is not None:
            out[i] = hit
        else:
            pend_i.append(i)
            pend_b.append(s)
    if not pend_b:
        logging.info(
            "配台不要ルール: AIキャッシュのみで D→E バッチ %s 件を完結（API 呼び出しなし）。",
            n,
        )
        return out
    if not API_KEY:
        return out
    m = len(pend_b)
    if m == 1:
        out[pend_i[0]] = _ai_compile_exclude_rule_logic_to_json(pend_b[0])
        return out

    schema = _exclude_rule_logic_gemini_schema_instructions()
    numbered = "\n".join(f"[{i + 1}] {str(b).strip()}" for i, b in enumerate(pend_b))
    prompt = (
        "あなたは工場の配台システム用です。以下の N 個の「配台不能の説明」を、与えた順序でそれぞれ JSON ルールに変換してください。\n\n"
        f"【出力】JSON 配列のみ。先頭が [ で終わりが ] 。要素数は必ず {m}（Markdown・説明禁止）。\n"
        f"配列の先頭要素が [1]、2 番目が [2] … に対応します。\n\n"
        f"{schema}\n"
        f"【説明文】\n{numbered}\n"
    )
    try:
        ppath = os.path.join(log_dir, "ai_exclude_rule_logic_batch_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("配台不要ルール(バッチ): プロンプト → %s", ppath)
    except OSError as ex:
        logging.warning("配台不要ルール(バッチ): プロンプト保存失敗: %s", ex)
    try:
        client = genai.Client(api_key=API_KEY)
        res = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
        record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
        raw = (_gemini_result_text(res) or "").strip()
        rpath = os.path.join(log_dir, "ai_exclude_rule_logic_batch_last_response.txt")
        try:
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(raw)
        except OSError:
            pass
        arr = _parse_exclude_rule_json_array_response(raw)
        if not isinstance(arr, list) or len(arr) != m:
            logging.warning(
                "配台不要ルール: バッチ応答が不正（要素数 %s、期待 %s）。1 件ずつ再試行します。",
                len(arr) if isinstance(arr, list) else None,
                m,
            )
            for j, idx in enumerate(pend_i):
                out[idx] = _ai_compile_exclude_rule_logic_to_json(pend_b[j])
            return out
        cache_dirty = False
        for j, item in enumerate(arr):
            parsed = _validate_exclude_rule_parsed_dict(item)
            out[pend_i[j]] = parsed
            if parsed:
                _cache_put_exclude_rule_de_parsed(ai_cache, pend_b[j], parsed)
                cache_dirty = True
        if cache_dirty:
            save_ai_cache(ai_cache)
        return out
    except Exception as e:
        logging.warning("配台不要ルール: バッチ Gemini 失敗、単発にフォールバック: %s", e)
        for j, idx in enumerate(pend_i):
            out[idx] = _ai_compile_exclude_rule_logic_to_json(pend_b[j])
        return out


def _log_exclude_rules_sheet_debug(
    event: str,
    log_prefix: str,
    summary: str,
    details: str = "",
    exc: BaseException | None = None,
) -> None:
    """
    「設定_配台不要工程」の保守処理のイベントログ。

    設定シート処理の成否を log/exclude_rules_sheet_debug.txt に追記し、execution_log にもタグ付きで出力する。
    event 例: START, OPEN_OK, OPEN_RETRY, OPEN_FAIL, HEADER_FIX, SYNC_ROWS, OPENPYXL_SAVE_OK, OPENPYXL_SAVE_FAIL,
    OPENPYXL_SAVE_SKIPPED_EXCLUDE_RULES_POLICY, OPENPYXL_RETRY_WAIT, OPENPYXL_VBA_FALLBACK, MATRIX_TSV_WRITTEN,
    XLWINGS_UNAVAILABLE, XLWINGS_ATTACH_FAIL, XLWINGS_SYNC_SKIP, XLWINGS_SYNC_OK, XLWINGS_SYNC_FAIL,
    E_SIDECAR_WRITTEN, E_SIDECAR_APPLIED, FALLBACK_FAIL,
    SKIP_NO_PATH, SKIP_NO_FILE, SKIP_NO_SHEET, DATA_COMPACT
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"--- {ts} ---",
        f"event={event}",
        f"phase={log_prefix}",
        f"summary={summary}",
    ]
    if details:
        lines.append(f"details={details}")
    if exc is not None:
        lines.append(f"exception={type(exc).__name__}: {exc}")
        lines.append(traceback.format_exc().rstrip())
    block = "\n".join(lines) + "\n\n"
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(exclude_rules_sheet_debug_log_path, "a", encoding="utf-8", newline="\n") as df:
            df.write(block)
    except OSError as wex:
        logging.warning("exclude_rules_sheet_debug.txt へ書けません: %s", wex)

    tag = "[設定_配台不要工程]"
    msg = f"{tag} {event} | {log_prefix} | {summary}"
    if details:
        msg += f" | {details}"
    if event in (
        "OPEN_FAIL",
        "SAVE_FAIL",
        "COM_MERGE_FAIL",
        "FALLBACK_FAIL",
        "SKIP_NO_PATH",
        "SKIP_NO_FILE",
        "SKIP_NO_SHEET",
        "FATAL",
    ):
        logging.error(msg)
    elif event in (
        "OPEN_RETRY",
        "SAVE_FAIL_HINT",
        "SAVE_RETRY",
        "COM_SYNC_UNAVAILABLE",
        "COM_ATTACH_OPEN_FAIL",
        "XLWINGS_UNAVAILABLE",
        "XLWINGS_ATTACH_FAIL",
        "XLWINGS_SYNC_SKIP",
        "XLWINGS_SYNC_FAIL",
        "E_SIDECAR_WRITTEN",
        "OPENPYXL_SAVE_FAIL",
        "OPENPYXL_VBA_FALLBACK",
    ):
        logging.warning(msg)
    elif event in (
        "COM_MERGE_SKIP",
        "MATRIX_TSV_WRITTEN",
        "OPENPYXL_SAVE_OK",
        "OPENPYXL_RETRY_WAIT",
        "XLWINGS_SYNC_OK",
    ):
        logging.info(msg)
    else:
        logging.info(msg)


def _xlwings_paths_equivalent(disk_path: str, book_fullname: str) -> bool:
    """ディスクパスと xlwings Book.full_name が同一ファイルを指すか（表記ゆれを多少吸収）。"""
    try:
        fn = str(book_fullname).strip()
    except Exception:
        return False

    def _norm(p: str) -> str:
        p = os.path.normpath(str(p).strip().replace("/", "\\"))
        return os.path.normcase(os.path.abspath(p))

    try:
        if _norm(disk_path) == _norm(fn):
            return True
    except Exception:
        pass
    try:
        return os.path.samefile(disk_path, fn)
    except Exception:
        pass
    try:
        import win32api  # type: ignore

        a = _norm(win32api.GetLongPathName(disk_path))
        b = _norm(win32api.GetLongPathName(fn))
        if a == b:
            return True
    except Exception:
        pass
    try:
        if os.path.basename(_norm(disk_path)).lower() == os.path.basename(_norm(fn)).lower():
            if _norm(os.path.dirname(disk_path)) == _norm(os.path.dirname(fn)):
                return True
    except Exception:
        pass
    return False


def _xlwings_book_matches_path(book, disk_path: str) -> bool:
    try:
        fn = book.full_name
    except Exception:
        return False
    return _xlwings_paths_equivalent(disk_path, fn)


def _xlwings_find_book_on_running_instances(abs_path: str):
    """起動中の Excel からパス一致する xlwings Book を返す。無ければ None。"""
    try:
        import xlwings as xw
    except ImportError:
        return None
    target = os.path.abspath(abs_path)
    try:
        for app in list(xw.apps):
            try:
                for book in app.books:
                    try:
                        if _xlwings_book_matches_path(book, target):
                            return book
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception:
        return None
    return None


def _xlwings_try_open_in_running_apps(abs_path: str):
    """既存の Excel.App で Workbooks.Open を試す。成功時 Book、失敗時 None。"""
    try:
        import xlwings as xw
    except ImportError:
        return None
    path = os.path.abspath(abs_path)
    for app in list(xw.apps):
        try:
            return app.books.open(path, update_links=False)
        except Exception:
            continue
    return None


def _xlwings_release_book_after_mutation(xw_book, info: dict, mutation_ok: bool) -> None:
    """専用起動した Excel は終了する。実行中 Excel でだけ Open したブックは失敗時のみ閉じる。"""
    if xw_book is None:
        return
    mode = info.get("mode", "keep")
    opened_here = bool(info.get("opened_wb_here"))
    if mode == "quit_excel":
        try:
            xw_book.close()
        except Exception:
            pass
        try:
            xw_book.app.quit()
        except Exception:
            pass
        return
    if opened_here and not mutation_ok:
        try:
            xw_book.close()
        except Exception:
            pass


def _xlwings_attach_open_macro_workbook(macro_wb_path: str, log_prefix: str):
    """
    マクロブックを xlwings で取得する（本番・テスト共通）。
    戻り値: (Book, release_info) / 失敗時 None。
    release_info: mode が keep または quit_excel、opened_wb_here が bool。
    """
    try:
        import xlwings as xw  # noqa: F401
    except ImportError:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_UNAVAILABLE",
            log_prefix,
            "xlwings が import できません（pip install xlwings を確認）。",
        )
        return None

    abs_path = os.path.abspath(macro_wb_path)

    book = _xlwings_find_book_on_running_instances(abs_path)
    if book is not None:
        return book, {"mode": "keep", "opened_wb_here": False}

    book = _xlwings_try_open_in_running_apps(abs_path)
    if book is not None:
        return book, {"mode": "keep", "opened_wb_here": True}

    try:
        import xlwings as xw

        app = xw.App(visible=False, add_book=False)
        try:
            app.display_alerts = False
        except Exception:
            pass
        book = app.books.open(abs_path, update_links=False)
        return book, {"mode": "quit_excel", "opened_wb_here": True}
    except Exception as ex:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_ATTACH_FAIL",
            log_prefix,
            "xlwings でブックを開けませんでした。",
            details=f"path={abs_path}",
            exc=ex,
        )
        return None


def _xlwings_attach_workbook_for_tests(
    book_path: str,
    label: str,
    *,
    allow_dispatch_open: bool = False,
):
    """
    検証スクリプト用: 起動中ブックを優先し、必要なら表示付き Excel で開く。
    戻り値: (Book, info, 説明文字列) または None。
    """
    abs_path = os.path.abspath(book_path)
    book = _xlwings_find_book_on_running_instances(abs_path)
    if book is not None:
        return book, {"mode": "keep", "opened_wb_here": False}, f"{label}:既存インスタンス"
    if not allow_dispatch_open:
        return None
    try:
        import xlwings as xw

        app = xw.App(visible=True, add_book=False)
        try:
            app.display_alerts = False
        except Exception:
            pass
        book = app.books.open(abs_path, update_links=False)
        return book, {"mode": "keep", "opened_wb_here": True}, f"{label}:dispatch-open"
    except Exception:
        return None


def _xlwings_sync_exclude_rules_sheet_from_openpyxl(
    wb_path: str, ws_oxl, log_prefix: str
) -> bool:
    """
    openpyxl で保存できないとき、xlwings で「設定_配台不要工程」A:E をメモリ上の値で上書きし Save。
    """
    global _exclude_rules_effective_read_path

    attached = _xlwings_attach_open_macro_workbook(wb_path, log_prefix)
    if attached is None:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_SYNC_SKIP",
            log_prefix,
            "xlwings でブックに接続できず A:E 同期をスキップ。",
            details=f"path={wb_path}",
        )
        return False

    xw_book, info = attached
    ok = False
    try:
        try:
            xw_book.app.display_alerts = False
        except Exception:
            pass
        try:
            sheet_names = [s.name for s in xw_book.sheets]
        except Exception:
            sheet_names = []
        if EXCLUDE_RULES_SHEET_NAME not in sheet_names:
            _log_exclude_rules_sheet_debug(
                "XLWINGS_SYNC_SKIP",
                log_prefix,
                f"xlwings 側にシート「{EXCLUDE_RULES_SHEET_NAME}」がありません。",
                details=f"path={wb_path}",
            )
            return False

        sht = xw_book.sheets[EXCLUDE_RULES_SHEET_NAME]
        max_r = max(1, int(ws_oxl.max_row or 1))
        ncols = EXCLUDE_RULES_SHEET_COM_SYNC_MAX_COL
        data = [
            [ws_oxl.cell(row=r, column=c).value for c in range(1, ncols + 1)]
            for r in range(1, max_r + 1)
        ]
        sht.range((1, 1)).resize(len(data), ncols).value = data
        xw_book.save()
        ok = True
        _exclude_rules_effective_read_path = wb_path
        _clear_exclude_rules_e_apply_files()
        _log_exclude_rules_sheet_debug(
            "XLWINGS_SYNC_OK",
            log_prefix,
            "xlwings 経由で設定シート A〜E を同期しブックを保存しました。",
            details=f"path={wb_path} rows={max_r}",
        )
        logging.info(
            "%s: 設定シートを xlwings でマクロブックに保存しました（A〜E）。",
            log_prefix,
        )
        return True
    except Exception as ex:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_SYNC_FAIL",
            log_prefix,
            "xlwings での A:E 同期または Save に失敗しました。",
            details=f"path={wb_path}",
            exc=ex,
        )
        return False
    finally:
        _xlwings_release_book_after_mutation(xw_book, info, ok)


# 設定シートの列範囲（A〜E）。xlwings 同期・VBA 行列 TSV 出力でも使用。
EXCLUDE_RULES_SHEET_COM_SYNC_MAX_COL = 5
EXCLUDE_RULES_MATRIX_CLIP_MAX_COL = 5


def _persist_exclude_rules_workbook(_wb, wb_path: str, ws, log_prefix: str) -> bool:
    """
    設定シートのディスク反映。既定は xlwings で A:E 同期→Save（EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1 のときのみ openpyxl save を試行）。
    保存できないときは log に行列 TSV を出し、VBA「設定_配台不要工程_AからE_TSVから反映」で反映する。

    _wb … 編集済み openpyxl ブック（openpyxl 経路時のみ save に使用）。
    """
    global _exclude_rules_effective_read_path

    def _openpyxl_persist_ok(which: str) -> bool:
        try:
            _wb.save(wb_path)
        except Exception as ex:
            _log_exclude_rules_sheet_debug(
                "OPENPYXL_SAVE_FAIL",
                log_prefix,
                f"openpyxl での .xlsm 保存に失敗しました {which}（Excel で開きっぱなし・ロックの可能性）。",
                details=f"path={wb_path}",
                exc=ex,
            )
            return False
        _exclude_rules_effective_read_path = wb_path
        _clear_exclude_rules_e_apply_files()
        _log_exclude_rules_sheet_debug(
            "OPENPYXL_SAVE_OK",
            log_prefix,
            "openpyxl で設定シートを含むブックを保存しました（A〜E）。",
            details=f"path={wb_path} {which}",
        )
        logging.info(
            "%s: 設定シートを openpyxl でマクロブックに保存しました。%s",
            log_prefix,
            which,
        )
        return True

    saved_openpyxl = False
    if EXCLUDE_RULES_SKIP_OPENPYXL_SAVE:
        _log_exclude_rules_sheet_debug(
            "OPENPYXL_SAVE_SKIPPED_EXCLUDE_RULES_POLICY",
            log_prefix,
            "設定_配台不要工程の保存では openpyxl save を試行しません（xlwings 同期を先行。再試行する場合は EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1）。",
            details=f"path={wb_path}",
        )
        logging.info(
            "%s: 設定_配台不要工程は openpyxl を試さず xlwings 同期→Save を試みます（不可なら VBA 用行列 TSV）。",
            log_prefix,
        )
    elif not _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: 設定_配台不要工程は openpyxl で保存します（不可のときは xlwings 同期→Save、それも不可なら VBA 用行列 TSV）。",
            log_prefix,
        )
        labels = ("(1/4)", "(2/4)", "(3/4)", "(4/4)")
        for i, label in enumerate(labels):
            if i:
                _log_exclude_rules_sheet_debug(
                    "OPENPYXL_RETRY_WAIT",
                    log_prefix,
                    f"openpyxl 再保存まで 2 秒待ちます {label}。",
                    details=f"path={wb_path}",
                )
                time_module.sleep(2.0)
            if _openpyxl_persist_ok(label):
                saved_openpyxl = True
                break
    else:
        _log_exclude_rules_sheet_debug(
            "OPENPYXL_SAVE_SKIPPED_INCOMPATIBLE_SHEET",
            log_prefix,
            f"ブックに「{OPENPYXL_INCOMPATIBLE_SHEET_MARKER}」があるため openpyxl での保存を試みません。",
            details=f"path={wb_path}",
        )
        logging.info(
            "%s: ブックに「%s」があるため openpyxl save をスキップし、xlwings または行列 TSV に切り替えます。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )

    if saved_openpyxl:
        return True

    if _xlwings_sync_exclude_rules_sheet_from_openpyxl(wb_path, ws, log_prefix):
        return True

    if _write_exclude_rules_matrix_vba_tsv(wb_path, ws, log_prefix):
        logging.warning(
            "%s: 設定シートを log\\%s に出力しました。"
            " Excel でマクロ「設定_配台不要工程_AからE_TSVから反映」を実行してください。",
            log_prefix,
            EXCLUDE_RULES_MATRIX_VBA_FILENAME,
        )

    _log_exclude_rules_sheet_debug(
        "OPENPYXL_VBA_FALLBACK",
        log_prefix,
        "openpyxl 保存に失敗したため VBA 用行列 TSV を出力しました（ブックは Excel 上で手動反映が必要な場合があります）。",
        details=f"path={wb_path}",
    )
    return False


def _exclude_rules_e_sidecar_path() -> str:
    path = os.path.join(json_data_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME)
    legacy = os.path.join(log_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME)
    if os.path.isfile(legacy) and not os.path.isfile(path):
        try:
            shutil.move(legacy, path)
        except OSError:
            pass
    return path


def _exclude_rules_e_vba_tsv_path() -> str:
    return os.path.join(log_dir, EXCLUDE_RULES_E_VBA_TSV_FILENAME)


def _exclude_rules_matrix_vba_path() -> str:
    return os.path.join(log_dir, EXCLUDE_RULES_MATRIX_VBA_FILENAME)


def _serialize_cell_for_matrix_tsv(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    return str(val)


def _write_exclude_rules_matrix_vba_tsv(
    wb_path: str, ws, log_prefix: str
) -> bool:
    """VBA 用: 設定シート 1 行目〜 max_row の A〜E を Base64(UTF-8) 付き TSV で出力する。"""
    max_r = max(1, int(ws.max_row or 1))
    lines = [
        "v1",
        "workbook\t" + os.path.abspath(wb_path),
        "sheet\t" + EXCLUDE_RULES_SHEET_NAME,
        "ncols\t5",
        "---",
    ]
    for r in range(1, max_r + 1):
        parts: list[str] = [str(r)]
        for c in range(1, 6):
            s = _serialize_cell_for_matrix_tsv(ws.cell(row=r, column=c).value)
            parts.append(base64.b64encode(s.encode("utf-8")).decode("ascii"))
        lines.append("\t".join(parts))
    path = _exclude_rules_matrix_vba_path()
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lines) + "\n")
        _log_exclude_rules_sheet_debug(
            "MATRIX_TSV_WRITTEN",
            log_prefix,
            "設定シート A〜E を VBA 反映用 TSV に書き出しました（openpyxl 保存不可時）。",
            details=f"path={path} rows={max_r}",
        )
        return True
    except OSError as ex:
        logging.warning("%s: 行列 VBA 用 TSV を書けません: %s", log_prefix, ex)
        return False


def _build_exclude_rules_list_from_openpyxl_ws(
    ws, c_proc: int, c_mach: int, c_flag: int, c_e: int
) -> list[dict]:
    """openpyxl 上の設定シートから _load_exclude_rules_from_workbook と同形のリストを構築。"""
    rules: list[dict] = []
    max_r = int(ws.max_row or 1)
    for r in range(2, max_r + 1):
        pv = ws.cell(row=r, column=c_proc).value
        proc = (
            ""
            if pv is None or (isinstance(pv, float) and pd.isna(pv))
            else str(pv).strip()
        )
        if not proc:
            continue
        mv = ws.cell(row=r, column=c_mach).value
        mach = (
            ""
            if mv is None or (isinstance(mv, float) and pd.isna(mv))
            else str(mv).strip()
        )
        cv = ws.cell(row=r, column=c_flag).value
        ev = ws.cell(row=r, column=c_e).value
        parsed = _parse_exclude_rule_json_cell(ev)
        rules.append(
            {"proc": proc, "mach": mach, "c_val": cv, "parsed": parsed}
        )
    return rules


def _set_exclude_rules_snapshot_from_ws(
    wb_path: str, ws, c_proc: int, c_mach: int, c_flag: int, c_e: int
) -> None:
    global _exclude_rules_rules_snapshot, _exclude_rules_snapshot_wb
    _exclude_rules_rules_snapshot = _build_exclude_rules_list_from_openpyxl_ws(
        ws, c_proc, c_mach, c_flag, c_e
    )
    _exclude_rules_snapshot_wb = os.path.normcase(os.path.abspath(wb_path))


def _clear_exclude_rules_e_apply_files() -> None:
    for p in (
        os.path.join(json_data_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME),
        os.path.join(log_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME),
    ):
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass
    for rel in (EXCLUDE_RULES_E_VBA_TSV_FILENAME, EXCLUDE_RULES_MATRIX_VBA_FILENAME):
        p = os.path.join(log_dir, rel)
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass


def _write_exclude_rules_e_vba_tsv_from_cells(
    wb_path: str, c_e: int, cells: dict[str, str], log_prefix: str
) -> None:
    """VBA 用: 行番号と Base64(UTF-8) セル文字列の TSV。"""
    lines = [
        "v1",
        "workbook\t" + os.path.abspath(wb_path),
        "sheet\t" + EXCLUDE_RULES_SHEET_NAME,
        "column_e\t" + str(int(c_e)),
        "---",
    ]
    for rk in sorted(cells.keys(), key=lambda x: int(x)):
        s = cells[rk]
        b64 = base64.b64encode(s.encode("utf-8")).decode("ascii")
        lines.append(rk + "\t" + b64)
    path_tsv = _exclude_rules_e_vba_tsv_path()
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path_tsv, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lines) + "\n")
        _log_exclude_rules_sheet_debug(
            "E_VBA_TSV_WRITTEN",
            log_prefix,
            "E 列を VBA 反映用 TSV に書き出しました（保存失敗時のフォールバック用）。",
            details=f"path={path_tsv} cells={len(cells)}",
        )
    except OSError as ex:
        logging.warning("%s: E 列 VBA 用 TSV を書けません: %s", log_prefix, ex)


def _write_exclude_rules_e_apply_artifacts(
    wb_path: str, ws, c_e: int, log_prefix: str
) -> None:
    """
    E 列（非空）を JSON サイドカードと VBA 用 TSV に書く。空なら両ファイルを削除。
    Python 次回起動時の E 復元用 JSON と、マクロからの E 書込み用 TSV。
    """
    cells: dict[str, str] = {}
    max_r = int(ws.max_row or 1)
    for r in range(2, max_r + 1):
        ev = ws.cell(row=r, column=c_e).value
        if _cell_is_blank_for_rule(ev):
            continue
        s = str(ev).strip() if ev is not None else ""
        if not s:
            continue
        cells[str(r)] = s
    if not cells:
        _clear_exclude_rules_e_apply_files()
        return
    payload = {
        "version": 1,
        "workbook": os.path.abspath(wb_path),
        "sheet": EXCLUDE_RULES_SHEET_NAME,
        "column_e": c_e,
        "cells": cells,
    }
    path_sc = _exclude_rules_e_sidecar_path()
    try:
        os.makedirs(json_data_dir, exist_ok=True)
        with open(path_sc, "w", encoding="utf-8", newline="\n") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except OSError as ex:
        logging.warning("%s: E 列 JSON を書けません: %s", log_prefix, ex)
    _write_exclude_rules_e_vba_tsv_from_cells(wb_path, c_e, cells, log_prefix)
    _log_exclude_rules_sheet_debug(
        "E_APPLY_FILES_WRITTEN",
        log_prefix,
        "E 列を JSON と VBA 用 TSV に書き出しました（マクロで E 列を反映後、ファイル削除）。",
        details=f"cells={len(cells)}",
    )


def _try_apply_pending_exclude_rules_e_column(
    wb_path: str, ws, c_e: int, log_prefix: str
) -> int:
    """
    前回保存に失敗したとき書き出した JSON から E 列を復元する。
    ブックパスが一致しなければ何もしない。適用後はサイドカードを削除する。
    """
    path_sc = _exclude_rules_e_sidecar_path()
    if not os.path.isfile(path_sc):
        return 0
    try:
        with open(path_sc, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return 0
    if int(payload.get("version") or 0) != 1:
        return 0
    target = os.path.normcase(os.path.abspath(wb_path))
    if os.path.normcase(str(payload.get("workbook") or "")) != target:
        return 0
    if str(payload.get("sheet") or "") != EXCLUDE_RULES_SHEET_NAME:
        return 0
    cells = payload.get("cells")
    if not isinstance(cells, dict):
        return 0
    n = 0
    for rk, val in cells.items():
        try:
            ri = int(rk)
        except (TypeError, ValueError):
            continue
        if ri < 2:
            continue
        if isinstance(val, dict):
            sval = json.dumps(val, ensure_ascii=False)
        else:
            sval = "" if val is None else str(val).strip()
        if not sval:
            continue
        ws.cell(row=ri, column=c_e, value=sval)
        n += 1
    try:
        os.remove(path_sc)
    except OSError:
        pass
    if n:
        _log_exclude_rules_sheet_debug(
            "E_SIDECAR_APPLIED",
            log_prefix,
            f"未保存だった E 列をサイドカードから {n} セル復元しました。",
            details=path_sc,
        )
        logging.info(
            "%s: %s の内容をシートのロジック式列へ適用しました（続けて保存を試みます）。",
            log_prefix,
            path_sc,
        )
    return n


def _read_exclude_rules_d_cells_data_only_for_rows(
    wb_path: str, rows: list[int], c_d: int
) -> dict[int, object]:
    """
    D 列が数式のとき、openpyxl の通常読込では '=...' しか取れない。
    data_only=True でキャッシュ値を読む（Excel が一度でも保存・計算済みのブックで有効）。
    """
    out: dict[int, object] = {}
    if not rows or not os.path.isfile(wb_path):
        return out
    if _workbook_should_skip_openpyxl_io(wb_path):
        return out
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wbro = None
    try:
        wbro = load_workbook(
            wb_path,
            read_only=True,
            data_only=True,
            keep_vba=keep_vba,
        )
    except Exception:
        return out
    try:
        if EXCLUDE_RULES_SHEET_NAME not in wbro.sheetnames:
            return out
        wsro = wbro[EXCLUDE_RULES_SHEET_NAME]
        for r in rows:
            if r < 2:
                continue
            try:
                out[r] = wsro.cell(row=r, column=c_d).value
            except Exception:
                pass
    finally:
        if wbro is not None:
            try:
                wbro.close()
            except Exception:
                pass
    return out


def run_exclude_rules_sheet_maintenance(
    wb_path: str, pairs: list[tuple[str, str]], log_prefix: str
) -> None:
    """
    「設定_配台不要工程」の行同期・D→E の AI 補完・ディスク反映（既定は xlwings で A〜E 同期→Save。``EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1`` のとき openpyxl save を試行）。

    xlwings でも保存できないときは ``log/exclude_rules_matrix_vba.tsv`` を残し、マクロ
    ``設定_配台不要工程_AからE_TSVから反映`` で A〜E を反映する。
    併せて従来どおり E 列のみの ``exclude_rules_e_column_vba.tsv`` も出力され得る（行列 TSV 優先で反映後は削除）。
    保存成功時は TSV/JSON は削除される。

    ``json/exclude_rules_e_column_pending.json`` は Python 次回起動時の E 列復元用。
    シートの新規作成と 1 行目見出しは VBA「設定_配台不要工程_シートを確保」。
    """
    if not wb_path:
        _log_exclude_rules_sheet_debug(
            "SKIP_NO_PATH",
            log_prefix,
            "TASK_INPUT_WORKBOOK が空のため設定シート処理をしません。",
        )
        return
    if not os.path.exists(wb_path):
        _log_exclude_rules_sheet_debug(
            "SKIP_NO_FILE",
            log_prefix,
            "ブックが存在しません。",
            details=f"path={wb_path}",
        )
        return

    _log_exclude_rules_sheet_debug(
        "START",
        log_prefix,
        "設定シート保守開始",
        details=f"path={wb_path} pairs={len(pairs)}",
    )
    global _exclude_rules_effective_read_path
    _exclude_rules_effective_read_path = None

    if _workbook_should_skip_openpyxl_io(wb_path):
        _log_exclude_rules_sheet_debug(
            "SKIP_OPENPYXL_INCOMPATIBLE_BOOK",
            log_prefix,
            f"ブックに「{OPENPYXL_INCOMPATIBLE_SHEET_MARKER}」が含まれるため、openpyxl による設定シート保守は行いません。",
            details=f"path={wb_path}",
        )
        logging.warning(
            "%s: 「%s」含有のため「%s」の openpyxl 保守をスキップしました（Excel／xlwings で編集してください）。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            EXCLUDE_RULES_SHEET_NAME,
        )
        return

    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(wb_path, keep_vba=keep_vba, read_only=False, data_only=False)
    except Exception as e1:
        if keep_vba:
            _log_exclude_rules_sheet_debug(
                "OPEN_RETRY",
                log_prefix,
                "keep_vba=True でブックを開けず keep_vba=False で再試行します（マクロが失われる可能性）。",
                exc=e1,
            )
            try:
                wb = load_workbook(wb_path, keep_vba=False, read_only=False, data_only=False)
            except Exception as e2:
                _log_exclude_rules_sheet_debug(
                    "OPEN_FAIL",
                    log_prefix,
                    "ブックを開けません。シートは作成・保存されません。",
                    details=f"path={wb_path}",
                    exc=e2,
                )
                return
        else:
            _log_exclude_rules_sheet_debug(
                "OPEN_FAIL",
                log_prefix,
                "ブックを開けません。シートは作成・保存されません。",
                details=f"path={wb_path}",
                exc=e1,
            )
            return

    _log_exclude_rules_sheet_debug(
        "OPEN_OK",
        log_prefix,
        "ブックを開きました。",
        details=f"keep_vba={keep_vba} sheets={len(wb.sheetnames)}",
    )

    try:
        if EXCLUDE_RULES_SHEET_NAME not in wb.sheetnames:
            _log_exclude_rules_sheet_debug(
                "SKIP_NO_SHEET",
                log_prefix,
                "シートがありません。VBA の「設定_配台不要工程_シートを確保」を実行するか、段階1/2 をマクロから起動してください。",
                details=f"path={wb_path}",
            )
            logging.error(
                "%s: 「%s」がありません。Python ではシートを作成しません。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
            )
            return

        ws = wb[EXCLUDE_RULES_SHEET_NAME]
        hm_before = _exclude_rules_sheet_header_map(ws)
        c_proc, c_mach, c_flag, c_d, c_e = _ensure_exclude_rules_sheet_headers_and_columns(
            ws, log_prefix
        )
        hm_after = _exclude_rules_sheet_header_map(ws)
        if tuple(hm_before.get(x) for x in (
            EXCLUDE_RULE_COL_PROCESS,
            EXCLUDE_RULE_COL_MACHINE,
            EXCLUDE_RULE_COL_FLAG,
            EXCLUDE_RULE_COL_LOGIC_JA,
            EXCLUDE_RULE_COL_LOGIC_JSON,
        )) != tuple(hm_after.get(x) for x in (
            EXCLUDE_RULE_COL_PROCESS,
            EXCLUDE_RULE_COL_MACHINE,
            EXCLUDE_RULE_COL_FLAG,
            EXCLUDE_RULE_COL_LOGIC_JA,
            EXCLUDE_RULE_COL_LOGIC_JSON,
        )):
            _log_exclude_rules_sheet_debug(
                "HEADER_FIX",
                log_prefix,
                "1行目に標準見出しを書き込みました（空シート・列名不一致の補正）。",
                details=f"cols=({c_proc},{c_mach},{c_flag},{c_d},{c_e})",
            )

        # 前回ブック保存に失敗したとき退避した E 列を、先にワークシートへ戻す（続く保存でディスクへ載る）
        _try_apply_pending_exclude_rules_e_column(wb_path, ws, c_e, log_prefix)

        existing_keys: set[tuple[str, str]] = set()
        max_r = max(2, int(ws.max_row or 2))
        for r in range(2, max_r + 1):
            pv = ws.cell(row=r, column=c_proc).value
            mv = ws.cell(row=r, column=c_mach).value
            p = str(pv).strip() if pv is not None and not (isinstance(pv, float) and pd.isna(pv)) else ""
            m = str(mv).strip() if mv is not None and not (isinstance(mv, float) and pd.isna(mv)) else ""
            if not p:
                continue
            existing_keys.add(
                (_normalize_process_name_for_rule_match(p), _normalize_equipment_match_key(m))
            )

        added = 0
        for p, m in pairs:
            key = (_normalize_process_name_for_rule_match(p), _normalize_equipment_match_key(m))
            if key in existing_keys:
                continue
            ws.append([p, m, None, None, None])
            existing_keys.add(key)
            added += 1
        if added:
            _log_exclude_rules_sheet_debug(
                "SYNC_ROWS",
                log_prefix,
                f"工程+機械の行を {added} 件追加しました。",
            )
            logging.info(
                "%s: 「%s」に工程+機械の組み合わせを %s 行追加しました。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
                added,
            )

        # 加工計画からペアが1件も取れず、シートにもデータ行が無いときは例行のみ（従来の新規シート相当）
        if added == 0 and not existing_keys:
            ws.append(["梱包", "", "yes", "", ""])
            existing_keys.add(
                (_normalize_process_name_for_rule_match("梱包"), _normalize_equipment_match_key(""))
            )
            _log_exclude_rules_sheet_debug(
                "EXAMPLE_ROW",
                log_prefix,
                "データ行が無かったため例（梱包=yes）を1行追加。",
            )
            logging.info(
                "%s: 「%s」にデータ行が無かったため、例（梱包=yes）を1行追加しました。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
            )

        # 空行詰めは AI より先に行う（後から詰めると、書き込んだ行番号と画面上の行がずれる）
        n_kept, n_removed_empty = _compact_exclude_rules_data_rows(
            ws, c_proc, c_mach, c_flag, c_d, c_e, log_prefix
        )
        if n_removed_empty:
            _log_exclude_rules_sheet_debug(
                "DATA_COMPACT",
                log_prefix,
                "空行を削除してデータ行を詰めました（並び順は維持）。AI 補完より前。",
                details=f"rows={n_kept} removed_empty={n_removed_empty}",
            )

        max_r = int(ws.max_row or 1)
        pending_rows: list[int] = []
        for r in range(2, max_r + 1):
            dv = ws.cell(row=r, column=c_d).value
            ev = ws.cell(row=r, column=c_e).value
            # C 列の有無に関係なく、D に説明があり E が空なら D→E を試す
            if _cell_is_blank_for_rule(dv):
                continue
            if not _cell_is_blank_for_rule(ev):
                continue
            pending_rows.append(r)

        # D が数式のときは通常読込では '=...' だけ取れる。data_only でキャッシュ表示値を補う。
        formula_rows = [
            r
            for r in pending_rows
            if isinstance(ws.cell(row=r, column=c_d).value, str)
            and str(ws.cell(row=r, column=c_d).value).strip().startswith("=")
        ]
        d_cached = (
            _read_exclude_rules_d_cells_data_only_for_rows(wb_path, formula_rows, c_d)
            if formula_rows
            else {}
        )
        pending_texts: list[str] = []
        filtered_rows: list[int] = []
        for r in pending_rows:
            dv = ws.cell(row=r, column=c_d).value
            blob = (
                ""
                if dv is None or (isinstance(dv, float) and pd.isna(dv))
                else str(dv).strip()
            )
            if blob.startswith("="):
                alt = d_cached.get(r)
                if alt is not None and not (isinstance(alt, float) and pd.isna(alt)):
                    blob = str(alt).strip()
                else:
                    logging.warning(
                        "%s: 「%s」%s 行目の D 列が数式で、キャッシュ値を読めませんでした（Excel で一度保存するか D を値にしてください）。",
                        log_prefix,
                        EXCLUDE_RULES_SHEET_NAME,
                        r,
                    )
                    continue
            if _cell_is_blank_for_rule(blob):
                continue
            filtered_rows.append(r)
            pending_texts.append(blob)
        pending_rows = filtered_rows

        ai_filled = 0
        ai_e_cell_addrs: list[str] = []
        if pending_texts:
            parsed_list = _ai_compile_exclude_rule_logics_batch(pending_texts)
            for r, parsed in zip(pending_rows, parsed_list):
                if not parsed:
                    logging.warning(
                        "%s: 「%s」%s 行目の D 列を JSON にできませんでした（APIキー・応答を確認）。",
                        log_prefix,
                        EXCLUDE_RULES_SHEET_NAME,
                        r,
                    )
                    continue
                jstr = json.dumps(parsed, ensure_ascii=False)
                ws.cell(row=r, column=c_e, value=jstr)
                cell_addr = f"{get_column_letter(c_e)}{r}"
                ai_e_cell_addrs.append(cell_addr)
                preview = jstr if len(jstr) <= 160 else (jstr[:160] + "…")
                logging.info(
                    "%s: 「%s」ロジック式列「%s」セル %s に JSON を書き込み: %s",
                    log_prefix,
                    EXCLUDE_RULES_SHEET_NAME,
                    EXCLUDE_RULE_COL_LOGIC_JSON,
                    cell_addr,
                    preview,
                )
                ai_filled += 1
        if ai_filled:
            _log_exclude_rules_sheet_debug(
                "AI_E_FILLED",
                log_prefix,
                f"D→E の AI 補完を {ai_filled} 行実施。",
                details="cells=" + ",".join(ai_e_cell_addrs),
            )
            logging.info(
                "%s: 「%s」で D→E の AI 補完を %s 行（セル: %s）。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
                ai_filled,
                ",".join(ai_e_cell_addrs),
            )

        _er_test = os.environ.get("EXCLUDE_RULES_TEST_E1234", "").strip().lower()
        if _er_test in ("1", "yes", "true"):
            try:
                _er_row = int(os.environ.get("EXCLUDE_RULES_TEST_E1234_ROW", "9") or "9")
            except ValueError:
                _er_row = 9
            if _er_row < 2:
                _er_row = 9
            ws.cell(row=_er_row, column=c_e, value="1234")
            _e_addr = f"{get_column_letter(c_e)}{_er_row}"
            _log_exclude_rules_sheet_debug(
                "TEST_E1234",
                log_prefix,
                f'E列 {_e_addr} にテストで "1234" を書き込み',
                details=f"row={_er_row}",
            )
            logging.warning(
                '%s: 【テスト】%s に "1234" を書き込み（EXCLUDE_RULES_TEST_E1234）。',
                log_prefix,
                _e_addr,
            )

        _set_exclude_rules_snapshot_from_ws(
            wb_path, ws, c_proc, c_mach, c_flag, c_e
        )
        _write_exclude_rules_e_apply_artifacts(wb_path, ws, c_e, log_prefix)
        persisted = _persist_exclude_rules_workbook(wb, wb_path, ws, log_prefix)
        if not persisted:
            logging.warning(
                "%s: 設定シートの openpyxl 保存に失敗しました。"
                " log の行列 TSV をマクロ「設定_配台不要工程_AからE_TSVから反映」、"
                "または E 列のみ「設定_配台不要工程_E列_TSVから反映」で反映してください。",
                log_prefix,
            )
    except Exception as ex:
        _log_exclude_rules_sheet_debug(
            "FATAL",
            log_prefix,
            "設定シート処理中に未捕捉例外が発生しました。",
            exc=ex,
        )
        logging.exception("%s: 設定_配台不要工程の処理で例外", log_prefix)
    finally:
        if wb is not None:
            wb.close()
            _log_exclude_rules_sheet_debug("CLOSED", log_prefix, "ブックをクローズしました。")


def _resolve_exclude_rules_workbook_path_for_read(wb_path: str) -> str:
    """直前の保守で実効パスが変わったとき（通常は保存成功後の元ブック）にそれを使う。"""
    p = _exclude_rules_effective_read_path
    if p and os.path.exists(p):
        return p
    return wb_path


def _load_exclude_rules_from_workbook(wb_path: str) -> list[dict]:
    """シートからルール行を読み、評価用リストを返す。"""
    if not wb_path:
        return []
    global _exclude_rules_rules_snapshot, _exclude_rules_snapshot_wb
    ap_arg = os.path.normcase(os.path.abspath(wb_path))
    if (
        _exclude_rules_rules_snapshot is not None
        and _exclude_rules_snapshot_wb == ap_arg
    ):
        snap = list(_exclude_rules_rules_snapshot)
        _exclude_rules_rules_snapshot = None
        _exclude_rules_snapshot_wb = None
        return snap
    path = _resolve_exclude_rules_workbook_path_for_read(wb_path)
    if not os.path.exists(path):
        return []
    if _workbook_should_skip_openpyxl_io(path):
        logging.warning(
            "配台不要ルール: ブックに「%s」があるため pandas(openpyxl) での「%s」読込をスキップしました（ルールは未適用）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            EXCLUDE_RULES_SHEET_NAME,
        )
        return []
    try:
        df = pd.read_excel(path, sheet_name=EXCLUDE_RULES_SHEET_NAME)
    except Exception:
        return []
    df.columns = df.columns.str.strip()
    need = [EXCLUDE_RULE_COL_PROCESS, EXCLUDE_RULE_COL_MACHINE]
    for c in need:
        if c not in df.columns:
            return []
    rules = []
    for _, row in df.iterrows():
        proc = str(row.get(EXCLUDE_RULE_COL_PROCESS, "") or "").strip()
        if not proc:
            continue
        mach = str(row.get(EXCLUDE_RULE_COL_MACHINE, "") or "").strip()
        c_val = row.get(EXCLUDE_RULE_COL_FLAG)
        e_raw = row.get(EXCLUDE_RULE_COL_LOGIC_JSON)
        parsed = _parse_exclude_rule_json_cell(e_raw)
        rules.append(
            {
                "proc": proc,
                "mach": mach,
                "c_val": c_val,
                "parsed": parsed,
            }
        )
    return rules


def apply_exclude_rules_config_to_plan_df(
    df: pd.DataFrame, wb_path: str, log_prefix: str
) -> pd.DataFrame:
    """設定シートに基づき「配台不要」を設定（C=yes または E の JSON が真）。"""
    if df is None or df.empty:
        return df
    if TASK_COL_MACHINE not in df.columns or PLAN_COL_EXCLUDE_FROM_ASSIGNMENT not in df.columns:
        return df
    rules = _load_exclude_rules_from_workbook(wb_path)
    if not rules:
        return df
    df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT].astype(object)
    n = 0
    for i in df.index:
        try:
            row = df.loc[i]
        except Exception:
            continue
        tp = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        tm = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        if not tp:
            continue
        for ru in rules:
            if not _task_row_matches_exclude_rule_target(tp, tm, ru["proc"], ru["mach"]):
                continue
            if _exclude_rule_c_column_is_yes(ru["c_val"]):
                df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = "yes"
                n += 1
                break
            if ru.get("parsed") and evaluate_exclude_rule_json_for_row(ru["parsed"], row):
                df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = "yes"
                n += 1
                break
    if n:
        logging.info("%s: 設定「%s」により配台不要=yes を %s 行に設定しました。", log_prefix, EXCLUDE_RULES_SHEET_NAME, n)
    return df


# =============================================================================
# 段階1エントリ（task_extract_stage1.py → run_stage1_extract）
#   加工計画DATA 読取 → 配台不要自動処理 → 設定シート保守 → plan_input_tasks.xlsx 出力
# =============================================================================
def run_stage1_extract():
    """
    段階1: 加工計画DATA から配台用タスク一覧を抽出し output/plan_input_tasks.xlsx へ出力。
    同一依頼NOで同一機械名が複数行あるとき、工程名「分割」行の空の「配台不要」に yes を自動設定する。
    マクロブックの「設定_配台不要工程」で工程+機械ごとの配台不要・条件式（AI）を管理する（シート作成は VBA）。
    """
    if not TASKS_INPUT_WORKBOOK:
        logging.error("TASK_INPUT_WORKBOOK が未設定です。")
        return False
    if not os.path.exists(TASKS_INPUT_WORKBOOK):
        logging.error(f"TASK_INPUT_WORKBOOK が存在しません: {TASKS_INPUT_WORKBOOK}")
        return False
    reset_gemini_usage_tracker()
    df_src = load_tasks_df()
    try:
        _pm_pairs = _collect_process_machine_pairs_for_exclude_rules(df_src)
        run_exclude_rules_sheet_maintenance(TASKS_INPUT_WORKBOOK, _pm_pairs, "段階1")
    except Exception:
        logging.exception("段階1: 設定_配台不要工程の保守で例外（続行）")
    records = []
    for _, row in df_src.iterrows():
        if row_has_completion_keyword(row):
            continue
        task_id = planning_task_id_str_from_scalar(row.get(TASK_COL_TASK_ID))
        machine = str(row.get(TASK_COL_MACHINE, "")).strip()
        machine_name = str(row.get(TASK_COL_MACHINE_NAME, "")).strip()
        qty_total = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
        done_qty = calc_done_qty_equivalent_from_row(row)
        qty = max(0.0, qty_total - done_qty)
        if qty <= 0 or not machine or not task_id:
            continue
        rec = {c: row.get(c) for c in SOURCE_BASE_COLUMNS}
        rec[TASK_COL_TASK_ID] = task_id
        # 工程名 + 機械名 を“因子”として表示用に追加（後段は計算キーにも使用）
        if machine_name:
            rec[PLAN_COL_PROCESS_FACTOR] = f"{machine}+{machine_name}"
        else:
            rec[PLAN_COL_PROCESS_FACTOR] = f"{machine}+"
        rec[PLAN_COL_REQUIRED_OP] = ""
        rec[PLAN_COL_SPEED_OVERRIDE] = ""
        rec[PLAN_COL_TASK_EFFICIENCY] = ""
        rec[PLAN_COL_PRIORITY] = ""
        rec[PLAN_COL_SPECIFIED_DUE_OVERRIDE] = ""
        rec[PLAN_COL_START_DATE_OVERRIDE] = ""
        rec[PLAN_COL_START_TIME_OVERRIDE] = ""
        rec[PLAN_COL_PREFERRED_OP] = ""
        rec[PLAN_COL_SPECIAL_REMARK] = ""
        rec[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = ""
        rec[PLAN_COL_AI_PARSE] = ""
        records.append(rec)
    if not records:
        logging.warning("段階1: 抽出対象タスクがありません。")
    order = plan_input_sheet_column_order()
    out_df = pd.DataFrame(records)
    if out_df.empty:
        out_df = pd.DataFrame(columns=order)
    else:
        out_df = out_df.reindex(columns=order).fillna("")
    if PLAN_COL_EXCLUDE_FROM_ASSIGNMENT in out_df.columns:
        out_df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = out_df[
            PLAN_COL_EXCLUDE_FROM_ASSIGNMENT
        ].astype(object)
    try:
        _, _, _, req_map, need_rules, _ = load_skills_and_needs()
    except PlanningValidationError:
        logging.error("段階1を中断: マスタ skills の検証エラー（優先度の数値重複など）。")
        raise
    except Exception as e:
        logging.info("段階1: マスタ need を読めず元列は need なしで埋めます (%s)", e)
        req_map, need_rules = {}, []
    out_df = _merge_plan_sheet_user_overrides(out_df)
    _refresh_plan_reference_columns(out_df, req_map, need_rules)
    try:
        _apply_auto_exclude_bunkatsu_duplicate_machine(out_df, log_prefix="段階1")
    except Exception as ex:
        logging.exception("段階1: 分割行の配台不要自動設定で例外（出力は続行）: %s", ex)
    try:
        out_df = apply_exclude_rules_config_to_plan_df(out_df, TASKS_INPUT_WORKBOOK, "段階1")
    except Exception as ex:
        logging.warning("段階1: 設定シートによる配台不要適用で例外（続行）: %s", ex)
    out_path = os.path.join(output_dir, STAGE1_OUTPUT_FILENAME)
    out_df.to_excel(out_path, sheet_name="タスク一覧", index=False)
    _apply_excel_date_columns_date_only_display(out_path, "タスク一覧")
    _apply_plan_input_visual_format(out_path, "タスク一覧")
    logging.info(f"段階1完了: '{out_path}' を出力しました。マクロで '{PLAN_INPUT_SHEET_NAME}' に取り込んでください。")
    _try_write_main_sheet_gemini_usage_summary("段階1")
    return True


# 稼働ルール（デフォルト値・2026年3月基準）
TARGET_YEAR = 2026
TARGET_MONTH = 3
DEFAULT_START_TIME = time(8, 45)
DEFAULT_END_TIME = time(17, 0)
DEFAULT_BREAKS = [
    (time(12, 0), time(12, 50)),
    (time(14, 45), time(15, 0))
]
# 休憩直前・（任意で）終業直前に「まだロールが残る」タスクを詰めない。
# ASSIGN_DEFER_MIN_REMAINING_ROLLS … 既定3。環境変数 ASSIGN_DEFER_MIN_REMAINING_ROLLS で上書き。0 で本ブロック全体スキップ。
# ASSIGN_END_OF_DAY_DEFER_MINUTES … team_end_limit までの残りがこの分数以下なら当日開始しない（None）。既定0=無効。
#   事業所標準終業は DEFAULT_END_TIME(17:00) だが、残業・早出・昼休憩ずらしは未実装のため終業直前デファーは既定OFF。
#   将来はメンバー別の実勤務終了（最大21:00想定など）と ASSIGN_END_OF_DAY_DEFER_MINUTES を組み合わせて有効化する想定。
ASSIGN_DEFER_MIN_REMAINING_ROLLS = max(
    0, int(os.environ.get("ASSIGN_DEFER_MIN_REMAINING_ROLLS", "3").strip() or 0)
)
ASSIGN_PRE_BREAK_DEFER_GAP_MINUTES = max(
    0,
    int(os.environ.get("ASSIGN_PRE_BREAK_DEFER_GAP_MINUTES", "15").strip() or 0),
)
ASSIGN_END_OF_DAY_DEFER_MINUTES = max(
    0,
    int(os.environ.get("ASSIGN_END_OF_DAY_DEFER_MINUTES", "0").strip() or 0),
)

# 配台デバッグ: 1 ロールごとに JSONL へ追記（DISPATCH_ROLL_TRACE_JSONL）、
# または DISPATCH_DEBUG_STOP_AFTER_ROLLS 到達でシミュレーションを打ち切り（部分タイムラインのまま結果出力へ）。
_DISPATCH_ROLL_TRACE_SEQ = 0
_DISPATCH_ROLL_TRACE_PATH: str | None = None
_DISPATCH_DEBUG_STOP_AFTER_ROLLS: int | None = None
_DISPATCH_DEBUG_STOP_FLAG = False


def _dispatch_debug_should_stop_early() -> bool:
    return _DISPATCH_DEBUG_STOP_FLAG


def _dispatch_debug_reset_roll_trace(workbook_path: str | None) -> None:
    """generate_plan のスケジューリング開始直前に呼ぶ。JSONL を空にし、ロール计数・停止フラグをリセット。"""
    global _DISPATCH_ROLL_TRACE_SEQ, _DISPATCH_ROLL_TRACE_PATH, _DISPATCH_DEBUG_STOP_AFTER_ROLLS
    global _DISPATCH_DEBUG_STOP_FLAG
    _DISPATCH_DEBUG_STOP_FLAG = False
    _DISPATCH_ROLL_TRACE_SEQ = 0
    raw = (os.environ.get("DISPATCH_ROLL_TRACE_JSONL") or "").strip()
    stop_raw = (os.environ.get("DISPATCH_DEBUG_STOP_AFTER_ROLLS") or "").strip()
    if stop_raw.isdigit():
        _DISPATCH_DEBUG_STOP_AFTER_ROLLS = max(1, int(stop_raw))
        if not raw:
            raw = "log/dispatch_roll_trace.jsonl"
    else:
        _DISPATCH_DEBUG_STOP_AFTER_ROLLS = None
    if not raw:
        _DISPATCH_ROLL_TRACE_PATH = None
        return
    if workbook_path:
        _DISPATCH_ROLL_TRACE_PATH = _resolve_path_relative_to_workbook(workbook_path, raw)
    else:
        _DISPATCH_ROLL_TRACE_PATH = os.path.abspath(raw)
    _d = os.path.dirname(_DISPATCH_ROLL_TRACE_PATH)
    if _d:
        os.makedirs(_d, exist_ok=True)
    with open(_DISPATCH_ROLL_TRACE_PATH, "w", encoding="utf-8") as f:
        f.write("")
    if _DISPATCH_ROLL_TRACE_PATH or _DISPATCH_DEBUG_STOP_AFTER_ROLLS is not None:
        logging.info(
            "配台デバッグ: ロールトレース path=%s STOP_AFTER=%s",
            _DISPATCH_ROLL_TRACE_PATH or "（カウントのみ・ファイルなし）",
            _DISPATCH_DEBUG_STOP_AFTER_ROLLS,
        )


def _dispatch_roll_trace_after_roll(
    current_date: date,
    task: dict,
    eq_line: str,
    start_dt,
    end_dt,
    units_done: float,
    rem_rolls_after: int,
    lead_op: str,
    sub_members: list,
    *,
    roll_surplus_meta: dict | None = None,
) -> None:
    """タイムラインに 1 ロール追記した直後に呼ぶ。JSONL・早期停止判定。"""
    global _DISPATCH_ROLL_TRACE_SEQ, _DISPATCH_DEBUG_STOP_FLAG
    if _DISPATCH_ROLL_TRACE_PATH is None and _DISPATCH_DEBUG_STOP_AFTER_ROLLS is None:
        return
    p = _DISPATCH_ROLL_TRACE_PATH
    _DISPATCH_ROLL_TRACE_SEQ += 1
    seq = _DISPATCH_ROLL_TRACE_SEQ
    st_s = start_dt.isoformat(sep=" ") if hasattr(start_dt, "isoformat") else str(start_dt)
    ed_s = end_dt.isoformat(sep=" ") if hasattr(end_dt, "isoformat") else str(end_dt)
    rec = {
        "seq": seq,
        "date": str(current_date),
        "task_id": str(task.get("task_id") or ""),
        "machine_line": str(eq_line),
        "dispatch_trial_order": task.get("dispatch_trial_order"),
        "start_dt": st_s,
        "end_dt": ed_s,
        "units_done": float(units_done),
        "remaining_rolls_after": int(rem_rolls_after),
        "lead_op": str(lead_op or "").strip(),
        "subs": ",".join(str(x).strip() for x in (sub_members or []) if x and str(x).strip()),
    }
    if roll_surplus_meta:
        rec.update(roll_surplus_meta)
    if p:
        with open(p, "a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    lim = _DISPATCH_DEBUG_STOP_AFTER_ROLLS
    if lim is not None and seq >= lim:
        _DISPATCH_DEBUG_STOP_FLAG = True
        logging.info(
            "配台デバッグ: DISPATCH_DEBUG_STOP_AFTER_ROLLS=%s に達したためこの時点で割付シミュレーションを打ち切ります（以降は未割当のまま結果シートへ）。",
            lim,
        )


# =========================================================
# 1. コア計算ロジック (日時ベース)
#    休憩帯を挟んだ「実働分」換算・終了時刻の繰り上げ。割付ループの下回り。
# =========================================================
def merge_time_intervals(intervals):
    """時刻区間のリストをソートし、重なる区間を結合して返す。"""
    if not intervals:
        return []
    intervals.sort(key=lambda x: x[0])
    merged = [intervals[0]]
    for current_start, current_end in intervals[1:]:
        last_start, last_end = merged[-1]
        if current_start <= last_end:
            merged[-1] = (last_start, max(last_end, current_end))
        else:
            merged.append((current_start, current_end))
    return merged


def _defer_team_start_past_prebreak_and_end_of_day(
    task: dict,
    team: tuple,
    team_start: datetime,
    team_end_limit: datetime,
    team_breaks: list,
    refloor_fn,
) -> datetime | None:
    """
    残ロールが ASSIGN_DEFER_MIN_REMAINING_ROLLS 以上のとき:
    - 次の休憩開始までが ASSIGN_PRE_BREAK_DEFER_GAP_MINUTES 分以内なら開始を休憩終了後へ繰り下げ（refloor_fn を再適用）
    - ASSIGN_END_OF_DAY_DEFER_MINUTES > 0 のとき、(team_end_limit - 試行開始) がその分数以下なら当日は不可（None）
      team_end_limit は呼び出し元の勤務上限（現状は主に標準終業に整合。残業連動は未実装で既定では終業直前デファーもOFF）
    """
    if ASSIGN_DEFER_MIN_REMAINING_ROLLS <= 0:
        return team_start
    rem_ceil = math.ceil(float(task.get("remaining_units") or 0))
    if rem_ceil < ASSIGN_DEFER_MIN_REMAINING_ROLLS:
        return team_start

    gap_pre = float(ASSIGN_PRE_BREAK_DEFER_GAP_MINUTES)
    gap_end = ASSIGN_END_OF_DAY_DEFER_MINUTES

    ts = refloor_fn(team_start)
    for _ in range(32):
        if ts >= team_end_limit:
            return None
        if gap_end > 0 and (team_end_limit - ts) <= timedelta(minutes=gap_end):
            return None
        progressed = False
        for bs, be in team_breaks:
            if be <= ts:
                continue
            if bs <= ts < be:
                ts = refloor_fn(be)
                progressed = True
                break
            if bs > ts:
                if gap_pre > 0 and (bs - ts).total_seconds() / 60.0 <= gap_pre:
                    ts = refloor_fn(be)
                    progressed = True
                break
        if not progressed:
            break
    if ts >= team_end_limit:
        return None
    if gap_end > 0 and (team_end_limit - ts) <= timedelta(minutes=gap_end):
        return None
    return ts


def _expand_timeline_events_for_equipment_grid(timeline_events: list) -> list:
    """
    設備毎の時間割・メンバー日程・稼働率用インデックス向け。
    1 本のイベントが日をまたぐ場合、e["date"] だけ当日に載せると翌朝セグメントが欠けるため、
    start_dt〜end_dt を各就業日 DEFAULT_START_TIME〜DEFAULT_END_TIME にクリップした複製へ展開する。
    """
    expanded: list = []
    for e in timeline_events:
        sd = e.get("start_dt")
        ed = e.get("end_dt")
        if not isinstance(sd, datetime) or not isinstance(ed, datetime):
            expanded.append(e)
            continue
        if ed <= sd:
            expanded.append(e)
            continue
        segments: list = []
        cal = sd.date()
        last_d = ed.date()
        while cal <= last_d:
            w0 = datetime.combine(cal, DEFAULT_START_TIME)
            w1 = datetime.combine(cal, DEFAULT_END_TIME)
            a = max(sd, w0)
            b = min(ed, w1)
            if a < b:
                ne = dict(e)
                ne["date"] = cal
                ne["start_dt"] = a
                ne["end_dt"] = b
                segments.append(ne)
            cal += timedelta(days=1)
        if segments:
            expanded.extend(segments)
        else:
            expanded.append(e)
    return expanded


def get_actual_work_minutes(start_dt, end_dt, breaks_dt):
    """
    start_dt から end_dt までの「休憩を除いた実働分数」。
    breaks_dt … (区間開始, 区間終了) の列（datetime または time。呼び出し元の勤怠イベントと整合）。
    """
    current = start_dt
    actual_mins = 0
    while current < end_dt:
        next_event = end_dt
        in_break = False
        b_end_time = None
        for b_s, b_e in breaks_dt:
            if b_s <= current < b_e:
                in_break = True
                b_end_time = b_e
                break
            elif current < b_s < next_event:
                next_event = b_s
        
        if in_break:
            current = b_end_time
        else:
            actual_mins += int((next_event - current).total_seconds() / 60)
            current = next_event
    return actual_mins


def calculate_end_time(start_dt, duration_minutes, breaks_dt, end_limit_dt):
    """
    start_dt から実働 duration_minutes 分進めた終了 datetime を求める（休憩はスキップ）。
    end_limit_dt を超えないよう打ち切り。戻り値: (終了時刻, 実際に進めた実働分, 残り未消化分)
    """
    current = start_dt
    remaining_work = duration_minutes
    actual_work_time = 0 

    while current < end_limit_dt and remaining_work > 0:
        next_event = end_limit_dt
        in_break = False
        break_end = None
        for b_start, b_end in breaks_dt:
            if b_start <= current < b_end:
                in_break = True
                break_end = b_end
                break
            elif current < b_start < next_event:
                next_event = b_start
                
        if in_break:
            current = break_end
            continue
            
        block_mins = int((next_event - current).total_seconds() / 60)
        if remaining_work <= block_mins:
            actual_work_time += remaining_work
            current += timedelta(minutes=remaining_work)
            remaining_work = 0
        else:
            actual_work_time += block_mins
            remaining_work -= block_mins
            current = next_event

    end_dt = min(current, end_limit_dt)
    return end_dt, actual_work_time, remaining_work

def match_need_sheet_condition(condition_raw: str, task_id: str) -> bool:
    """
    need シート「依頼NO条件」欄の解釈。
    空・*・全件 → 常にマッチ。
    prefix:ABC / 接頭辞:ABC → 依頼NO がその文字列で始まる
    regex:... / 正規表現:... → 正規表現（部分一致）
    それ以外の短文は接頭辞として扱う。従来の日本語例「依頼NOがJRで…」は JR を検出したら接頭辞JR扱い。
    """
    cond = (condition_raw or "").strip()
    tid = str(task_id).strip()
    if not cond or cond in ("*", "全件", "全て", "any", "ANY"):
        return True
    low = cond.lower()
    cn = cond.replace("：", ":")
    if low.startswith("prefix:") or low.startswith("接頭辞:"):
        pref = cn.split(":", 1)[1].strip() if ":" in cn else ""
        return bool(pref) and tid.startswith(pref)
    if low.startswith("regex:") or low.startswith("正規表現:"):
        pat = cn.split(":", 1)[1].strip() if ":" in cn else ""
        if not pat:
            return False
        try:
            return re.search(pat, tid) is not None
        except re.error:
            logging.warning(f"need 依頼NO条件の正規表現が無効です: {pat}")
            return False
    if "依頼" in cond and "JR" in cond.upper():
        return tid.upper().startswith("JR")
    return tid.startswith(cond)


def parse_need_sheet_special_rules(needs_df, label_col, equipment_list, cond_col):
    """特別指定1～99 行から、設備別の必要人数上書き（1～99）を抽出（先に定義された番号が優先）。"""
    rules = []
    for _, row in needs_df.iterrows():
        lab = str(row.get(label_col, "") or "").strip()
        m = re.match(r"特別指定\s*(\d+)", lab)
        if not m:
            continue
        order = int(m.group(1))
        if order < 1 or order > 99:
            continue
        cond = ""
        if cond_col is not None:
            v = row.get(cond_col)
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                cond = str(v).strip()
        overrides = {}
        for eq in equipment_list:
            v = row.get(eq)
            n = parse_optional_int(v)
            if n is not None and 1 <= n <= 99:
                overrides[str(eq).strip()] = n
        if not overrides:
            continue
        rules.append({"order": order, "condition": cond, "overrides": overrides})
    rules.sort(key=lambda r: r["order"])
    return rules


def resolve_need_required_op(process: str, machine_name: str, task_id: str, req_map: dict, need_rules: list) -> int:
    """
    need シートの「工程名 + 機械名」で必要OP人数を解決（特別指定1〜99は order が小さいほど優先）。

    req_map は
      - f\"{process}+{machine_name}\"（厳密キー）
      - machine_name（機械だけのフォールバック）
      - process（工程だけのフォールバック）
    のいずれかで base を引ける前提。
    need_rules の overrides も同様にキーを持つ。
    """
    p = str(process).strip()
    m = str(machine_name).strip()

    combo_key = f"{p}+{m}" if p and m else None

    base = None
    if combo_key and combo_key in req_map:
        base = req_map.get(combo_key)
    if base is None and m:
        base = req_map.get(m)
    if base is None and p:
        base = req_map.get(p)
    if base is None:
        base = 1

    for rule in need_rules:
        if not match_need_sheet_condition(rule["condition"], task_id):
            continue

        if combo_key and combo_key in rule["overrides"]:
            return int(rule["overrides"][combo_key])
        if m and m in rule["overrides"]:
            return int(rule["overrides"][m])
        if p and p in rule["overrides"]:
            return int(rule["overrides"][p])

    return int(base)


def resolve_need_required_op_explain(
    process: str, machine_name: str, task_id: str, req_map: dict, need_rules: list
) -> tuple[int, str]:
    """
    resolve_need_required_op と同値を返しつつ、ログ用に参照元の説明文字列を付ける。
    """
    p = str(process).strip()
    m = str(machine_name).strip()
    combo_key = f"{p}+{m}" if p and m else None
    base = None
    base_src = ""
    if combo_key and combo_key in req_map:
        base = req_map.get(combo_key)
        base_src = f"req_map[{combo_key!r}]={base}"
    elif m and m in req_map:
        base = req_map[m]
        base_src = f"req_map[機械名のみ {m!r}]={base}（複合キー不在）"
    elif p and p in req_map:
        base = req_map[p]
        base_src = f"req_map[工程名のみ {p!r}]={base}（複合・機械キー不在）"
    else:
        base = 1
        base_src = "req_map該当なし→既定1"
    for rule in need_rules:
        if not match_need_sheet_condition(rule["condition"], task_id):
            continue
        order = rule.get("order", "?")
        if combo_key and combo_key in rule["overrides"]:
            v = int(rule["overrides"][combo_key])
            return v, f"need特別指定{order} [{combo_key!r}]={v}"
        if m and m in rule["overrides"]:
            v = int(rule["overrides"][m])
            return v, f"need特別指定{order} [機械名{m!r}]={v}"
        if p and p in rule["overrides"]:
            v = int(rule["overrides"][p])
            return v, f"need特別指定{order} [工程名{p!r}]={v}"
    return int(base), base_src


def _need_row_label_hints_surplus_add(label_a0: str) -> bool:
    """need シート A列: 基本必要人数の直下にある「配台結果で余剰が出たときの追加増員上限」行か。"""
    s = unicodedata.normalize("NFKC", str(label_a0 or "").strip())
    if not s or s.startswith("特別指定"):
        return False
    if "依頼" in s and "条件" in s:
        return False
    if "追加" in s and ("人数" in s or "人員" in s or "増員" in s):
        return True
    if "増員" in s or "余剰" in s:
        return True
    if "配台" in s and ("追加" in s or "増" in s or "余剰" in s):
        return True
    return False


def _find_need_surplus_add_row_index(
    needs_raw, base_row: int, col0: int, pm_cols: list
) -> int | None:
    """基本必要人数行の次行を優先。ラベルまたは数値で追加人数行と判定。"""
    r = base_row + 1
    if r >= needs_raw.shape[0]:
        return None
    v0 = needs_raw.iat[r, col0]
    s0 = "" if pd.isna(v0) else str(v0).strip()
    if s0.startswith("特別指定"):
        return None
    if _need_row_label_hints_surplus_add(s0):
        return r
    nz = 0
    for col_idx, _, _ in pm_cols:
        if parse_optional_int(needs_raw.iat[r, col_idx]) is not None:
            nz += 1
    if nz > 0 and not unicodedata.normalize("NFKC", s0).startswith("特別"):
        return r
    return None


def resolve_need_surplus_extra_max(
    process: str,
    machine_name: str,
    task_id: str,
    surplus_map: dict,
    need_rules: list,
) -> int:
    """
    need シート「配台時追加人数」行（工程×機械列）の値＝必要人数を満たしたうえで
    さらに割り当て可能な人数の上限（0 なら従来どおり必要人数ちょうどのみ）。
    need_rules は現状この行を上書きしない（将来拡張用に task_id を受け取る）。
    """
    _ = (task_id, need_rules)
    if not surplus_map:
        return 0
    p = str(process).strip()
    m = str(machine_name).strip()
    combo_key = f"{p}+{m}" if p and m else None
    v = None
    if combo_key and combo_key in surplus_map:
        v = surplus_map[combo_key]
    elif m and m in surplus_map:
        v = surplus_map[m]
    elif p and p in surplus_map:
        v = surplus_map[p]
    if v is None:
        return 0
    try:
        n = int(v)
    except (TypeError, ValueError):
        return 0
    return max(0, min(n, 50))


def resolve_need_surplus_extra_max_explain(
    process: str,
    machine_name: str,
    task_id: str,
    surplus_map: dict,
    need_rules: list,
) -> tuple[int, str]:
    """resolve_need_surplus_extra_max と同値＋参照元説明（ログ用）。"""
    val = resolve_need_surplus_extra_max(
        process, machine_name, task_id, surplus_map, need_rules
    )
    _ = need_rules
    if not surplus_map:
        return val, "surplus_map空（配台時追加人数行なし）"
    p = str(process).strip()
    m = str(machine_name).strip()
    combo_key = f"{p}+{m}" if p and m else None
    if combo_key and combo_key in surplus_map:
        raw = surplus_map[combo_key]
        return val, f"surplus_map[{combo_key!r}]={raw}"
    if m and m in surplus_map:
        raw = surplus_map[m]
        return val, f"surplus_map[機械名のみ {m!r}]={raw}（複合キー不在）"
    if p and p in surplus_map:
        raw = surplus_map[p]
        return val, f"surplus_map[工程名のみ {p!r}]={raw}（複合キー不在）"
    return val, "surplus当キーなし→0"


def _surplus_team_time_factor(
    rq_base: int, team_len: int, extra_max_allowed: int
) -> float:
    """
    必要人数を超えて入れたメンバーによる単位時間への係数（1.0＝短縮なし）。
    追加枠（extra_max_allowed）を使い切ったときでも、短縮は SURPLUS_TEAM_MAX_SPEEDUP_RATIO を上限とする線形モデル。
    """
    rq = max(1, int(rq_base))
    tl = int(team_len)
    em = max(0, int(extra_max_allowed))
    extra = max(0, tl - rq)
    if extra <= 0 or em <= 0:
        return 1.0
    frac = min(1.0, extra / float(em))
    return 1.0 - SURPLUS_TEAM_MAX_SPEEDUP_RATIO * frac


def _team_assign_trace_tuple_label() -> str:
    if TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF:
        return "(-人数, 開始, -単位数, 優先度合計)"
    if TEAM_ASSIGN_START_SLACK_WAIT_MINUTES <= 0:
        return "(開始, -単位数, 優先度合計)"
    return (
        f"最早開始から{TEAM_ASSIGN_START_SLACK_WAIT_MINUTES}分以内は"
        "(0,-人数,開始,-単位数,優先度)、超過は(1,開始,-人数,-単位数,優先度)"
    )


def _team_assignment_sort_tuple(
    team: tuple,
    team_start: datetime,
    units_today: int,
    team_prio_sum: int,
    t_min: datetime | None = None,
) -> tuple:
    """
    チーム候補の優劣用タプル（辞書式で小さい方が採用）。
    - TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF: (-人数, 開始, -単位数, 優先度合計)
    - それ以外かつ TEAM_ASSIGN_START_SLACK_WAIT_MINUTES>0 かつ t_min あり:
        最早開始からスラック以内 → (0, -人数, 開始, -単位数, 優先度) … 遅れても人数を厚く
        スラック超 → (1, 開始, -人数, -単位数, 優先度) … 開始を優先
    - 上記以外: (開始, -単位数, 優先度合計)
    """
    n = len(team)
    if TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF:
        return (-n, team_start, -units_today, team_prio_sum)
    sm = TEAM_ASSIGN_START_SLACK_WAIT_MINUTES
    if sm <= 0 or t_min is None:
        return (team_start, -units_today, team_prio_sum)
    sl = timedelta(minutes=sm)
    if team_start - t_min <= sl:
        return (0, -n, team_start, -units_today, team_prio_sum)
    return (1, team_start, -n, -units_today, team_prio_sum)


# skills セル: OP / AS + 任意の優先度整数（例 OP1, AS 3）。数値が小さいほど割当で先に選ばれる。
_SKILL_OP_AS_CELL_RE = re.compile(r"^(OP|AS)(\d*)$", re.IGNORECASE)


def parse_op_as_skill_cell(cell_val):
    """
    master.xlsm「skills」のセル1つを解釈する。
    - 「OP」または「AS」の直後に優先度用の整数（空白は除去して解釈）。例: OP, OP1, AS3, AS 12
    - 優先度は小さいほど高優先（同一条件のチーム候補から先に選ばれる）。数字省略時は 1。
    - OP/AS で始まらない・空はスキルなし。
    """
    if cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val)):
        return None, 10**9
    s = str(cell_val).strip()
    if not s:
        return None, 10**9
    compact = re.sub(r"\s+", "", s).upper()
    m = _SKILL_OP_AS_CELL_RE.match(compact)
    if not m:
        return None, 10**9
    role = m.group(1).upper()
    tail = m.group(2) or ""
    if tail == "":
        pr = 1
    else:
        try:
            pr = int(tail)
        except ValueError:
            return None, 10**9
    if pr < 0:
        pr = 0
    return role, pr


def _validate_skills_op_as_priority_numbers_unique(
    skills_dict: dict, column_keys: list
) -> None:
    """
    master「skills」の各列（工程+機械キー等）について、OP/AS の割当優先度の**数値**が
    メンバー間で重複していないか検証する。重複時は PlanningValidationError。
    （OP1 と AS1 のようにロールが異なっても同一数値なら重複とみなす）
    """
    errors: list[str] = []
    for combo in column_keys:
        ck = str(combo or "").strip()
        if not ck:
            continue
        pr_to_entries: dict[int, list[str]] = defaultdict(list)
        for mem, row in (skills_dict or {}).items():
            mnm = str(mem or "").strip()
            if not mnm or not isinstance(row, dict):
                continue
            raw = row.get(ck)
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                continue
            sval = str(raw).strip()
            if not sval or sval.lower() in ("nan", "none", "null"):
                continue
            role, pr = parse_op_as_skill_cell(sval)
            if role not in ("OP", "AS"):
                continue
            pr_to_entries[int(pr)].append(f"{mnm}({role})")
        for pr, entries in sorted(pr_to_entries.items()):
            if len(entries) > 1:
                errors.append(f'列「{ck}」: 優先度 {pr} が重複 → ' + "、".join(entries))
    if errors:
        cap = 50
        tail = errors[:cap]
        msg = (
            "マスタ「skills」で、同一列の OP/AS 優先度の数値が重複しています。"
            " 列ごとに数値は1人につき1種類にしてください。\n"
            + "\n".join(tail)
        )
        if len(errors) > cap:
            msg += f"\n…他 {len(errors) - cap} 件"
        raise PlanningValidationError(msg)


def build_member_assignment_priority_reference(
    skills_dict: dict,
    members: list | None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    結果ブック用: マスタ skills の「工程名+機械名」列ごとに、割当アルゴリズムと同じ
    (優先度値昇順, メンバー名昇順) で並べた参考表と、ルール説明の表を返す。
    当日の出勤・設備空き・同一依頼の工程順・チーム人数は反映しない（あくまでマスタ上の順序）。
    """
    mem_list = list(members) if members else list((skills_dict or {}).keys())
    mem_list = [str(m).strip() for m in mem_list if m and str(m).strip()]

    surplus_on = bool(TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF)
    slack_m = TEAM_ASSIGN_START_SLACK_WAIT_MINUTES
    if surplus_on:
        team_rule = (
            "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF=有効: "
            "(-人数, 開始, -単位数, 優先度合計) の辞書式（人数最優先・従来）。"
        )
    elif slack_m > 0:
        team_rule = (
            f"既定: その日の成立候補全体の「最早開始」を基準に、"
            f"開始がその{slack_m}分以内の遅れなら人数を厚く優先（0,-人数,開始,-単位数,優先度）、"
            f"それより遅い候補は開始を優先（1,開始,-人数,-単位数,優先度）。"
            f"環境変数 TEAM_ASSIGN_START_SLACK_WAIT_MINUTES=0 で無効化。"
        )
    else:
        team_rule = (
            "TEAM_ASSIGN_START_SLACK_WAIT_MINUTES=0: "
            "(開始, -単位数, 優先度合計) のみ（開始最優先）。"
        )

    legend_rows = [
        {
            "区分": "スキル列の並び",
            "内容": "各「工程名+機械名」列について、セルが OP/AS（+優先度整数）のメンバーのみ対象。"
            " 数値が小さいほど高優先。省略時は優先度 1（parse_op_as_skill_cell と同一）。"
            " 同一列では優先度の数値はメンバー間で重複不可（マスタ読込時に検証）。",
        },
        {
            "区分": "当日との差",
            "内容": "実際の配台は、この順のうちその日出勤かつ AS/OP 要件を満たす者だけが候補。"
            " 設備の空き・同一依頼NOの工程順・必要人数・増員枠・指名OPで変わります。",
        },
        {
            "区分": "チーム候補の比較",
            "内容": team_rule,
        },
        {
            "区分": "指名・グローバル上書き",
            "内容": "担当OP_指定・メイン「再優先特別記載」の OP 指名は本表より優先されます。",
        },
        {
            "区分": "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF",
            "内容": "1/有効（人数最優先・従来）" if surplus_on else "0/無効（既定）",
        },
        {
            "区分": "TEAM_ASSIGN_START_SLACK_WAIT_MINUTES",
            "内容": str(slack_m),
        },
    ]
    df_legend = pd.DataFrame(legend_rows)

    combo_keys: set[str] = set()
    for m in mem_list:
        row = (skills_dict or {}).get(m) or {}
        for k in row:
            ks = str(k).strip()
            if "+" in ks:
                combo_keys.add(ks)
    sorted_combos = sorted(combo_keys)

    out: list[dict] = []
    for combo in sorted_combos:
        parts = combo.split("+", 1)
        proc = parts[0].strip()
        mach = parts[1].strip() if len(parts) > 1 else ""
        ranked: list[tuple[int, str, str, str]] = []
        for m in sorted(mem_list):
            cell = ((skills_dict or {}).get(m) or {}).get(combo)
            if cell is None or (isinstance(cell, float) and pd.isna(cell)):
                cell_s = ""
            else:
                cell_s = str(cell).strip()
            role, pr = parse_op_as_skill_cell(cell_s if cell_s else None)
            if role in ("OP", "AS"):
                ranked.append((pr, m, role, cell_s))
        ranked.sort(key=lambda x: (x[0], x[1]))
        if not ranked:
            out.append(
                {
                    "工程名": proc,
                    "機械名": mach,
                    "スキル列キー": combo,
                    "優先順位": "",
                    "メンバー": "（なし）",
                    "ロール": "",
                    "優先度値_小さいほど先": "",
                    "skillsセル値": "",
                    "備考": "この列に OP/AS の資格セルがあるメンバーがいません",
                }
            )
            continue
        for i, (pr, m, role, cell_s) in enumerate(ranked, start=1):
            out.append(
                {
                    "工程名": proc,
                    "機械名": mach,
                    "スキル列キー": combo,
                    "優先順位": i,
                    "メンバー": m,
                    "ロール": role,
                    "優先度値_小さいほど先": pr,
                    "skillsセル値": cell_s,
                    "備考": "",
                }
            )

    df_tbl = pd.DataFrame(out)
    return df_legend, df_tbl


def _normalize_person_name_for_match(s):
    """担当者指名のあいまい一致用（NFKC・富田/冨田の表記寄せ・空白除去・末尾敬称のみ除去）。"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    if "富田" in t:
        t = t.replace("富田", "冨田")
    t = re.sub(r"[\s　]+", "", t)
    t = re.sub(r"(さん|様|氏)$", "", t)
    return t


def _split_person_sei_mei(s) -> tuple[str, str]:
    """
    氏名を姓・名に分ける。最初の半角／全角空白の手前を姓、以降を名とする。
    空白が無い場合は (全体, '')（名なし扱い）。
    末尾の さん／様／氏 は分割前に除去する。
    """
    if s is None:
        return "", ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    if not t or t.lower() in ("nan", "none", "null"):
        return "", ""
    t = re.sub(r"(さん|様|氏)$", "", t)
    for i, ch in enumerate(t):
        if ch in " \u3000":
            sei = t[:i].strip()
            rest = t[i + 1 :]
            mei = re.sub(r"[\s　]+", "", rest.strip())
            return sei, mei
    return t.strip(), ""


def _normalize_sei_for_match(sei: str) -> str:
    """姓のみ正規化。表記ゆれは許容しない前提で、NFKC・富田/冨田寄せ・空白除去。"""
    if not sei:
        return ""
    t = unicodedata.normalize("NFKC", str(sei).strip())
    if "富田" in t:
        t = t.replace("富田", "冨田")
    t = re.sub(r"[\s　]+", "", t)
    return t


def _normalize_mei_for_match(mei: str) -> str:
    """名の正規化（ゆれ許容の前処理）。NFKC・空白除去。姓用の富田置換は行わない。"""
    if not mei:
        return ""
    t = unicodedata.normalize("NFKC", str(mei).strip())
    t = re.sub(r"[\s　]+", "", t)
    return t


def _has_duplicate_surname_among_members(member_names) -> bool:
    """skills メンバー一覧に、正規化後同一の姓が2人以上いるか。"""
    cnt = Counter()
    for name in member_names or []:
        if name is None or (isinstance(name, float) and pd.isna(name)):
            continue
        s = str(name).strip()
        if not s:
            continue
        sei, _mei = _split_person_sei_mei(s)
        key = _normalize_sei_for_match(sei)
        if key:
            cnt[key] += 1
    return any(n >= 2 for n in cnt.values())


def _mei_matches_with_fuzzy_allowed(r_mei_n: str, m_mei_n: str) -> bool:
    """同一姓がロスターで重複しないときのみ使う名のゆれ許容。"""
    if not r_mei_n and not m_mei_n:
        return True
    if not r_mei_n or not m_mei_n:
        return False
    if r_mei_n == m_mei_n:
        return True
    return r_mei_n in m_mei_n or m_mei_n in r_mei_n


def _resolve_preferred_op_to_member(raw, op_candidates, roster_member_names=None):
    """
    自由記述の指名を、当日スキル上OPのメンバー名（skills シートの行キー）に解決する。
    op_candidates: その設備でOPのメンバー名リスト。
    roster_member_names: skills の全メンバー名（省略時は op_candidates）。同一姓の重複判定に使用。

    名前の表記ゆれ:
    - 姓は正規化後に完全一致のみ（ゆれ許容しない。富田/冨田のみ従来どおり寄せ）。
    - roster に同一姓が2人以上いないときだけ、名は部分一致（どちらかが他方を含む）または完全一致を許容。
    - 同一姓がロスターにいる間は名も完全一致必須。
    - 姓のみの入力で名ゆれモードのとき、姓が一致するOPが複数いれば解決不可（None）。
    """
    if not raw or not op_candidates:
        return None
    r0 = unicodedata.normalize("NFKC", str(raw).strip())
    r = _normalize_person_name_for_match(r0)
    if not r:
        return None
    for m in op_candidates:
        if _normalize_person_name_for_match(m) == r:
            return m
        if unicodedata.normalize("NFKC", str(m).strip()) == r0.strip():
            return m

    roster = list(roster_member_names) if roster_member_names is not None else list(op_candidates)
    allow_mei_fuzzy = not _has_duplicate_surname_among_members(roster)

    r_sei, r_mei = _split_person_sei_mei(raw)
    r_sei_n = _normalize_sei_for_match(r_sei)
    r_mei_n = _normalize_mei_for_match(r_mei)
    if not r_sei_n:
        return None

    matches = []
    for m in op_candidates:
        m_sei, m_mei = _split_person_sei_mei(m)
        m_sei_n = _normalize_sei_for_match(m_sei)
        m_mei_n = _normalize_mei_for_match(m_mei)
        if r_sei_n != m_sei_n:
            continue
        if allow_mei_fuzzy:
            if r_mei_n:
                if _mei_matches_with_fuzzy_allowed(r_mei_n, m_mei_n):
                    matches.append(m)
            else:
                matches.append(m)
        else:
            if r_mei_n == m_mei_n:
                matches.append(m)

    if len(matches) == 1:
        return matches[0]
    return None


# =========================================================
# 2. マスタデータ・出勤簿(カレンダー) と AI解析
#    master.xlsm の skills / need / 各メンバー勤怠シートを読み、
#    備考・休暇区分は必要に応じて Gemini で構造化する。
# =========================================================
def load_skills_and_needs():
    """
    統合ファイル(MASTER_FILE)からスキルと need を動的に読み込みます。

    今回の need は（Excel上で）
      工程名行・機械名行のあと「基本必要人数」行（A列に「必要人数」を含む）
      その直下: 配台で余剰人員があるときに追加で入れられる人数（工程×機械ごと。未設定は 0）
      以降: 特別指定1〜99
    という構造のため、必要OPは「工程名+機械名」で解決する。

    skills 交差セルは OP/AS の後に優先度整数（例 OP1, AS3）。数値が小さいほど当該工程への割当で優先。
    数字省略の OP/AS は優先度 1。
    同一列（同一工程×機械）では優先度の数値はメンバー間で重複不可（重複時は PlanningValidationError）。
    """
    try:
        # skills は新仕様:
        #   1行目: 工程名
        #   2行目: 機械名
        #   A3以降: メンバー名
        #   交差セル: OP または AS の後に割当優先度の整数（例 OP1, AS3）。数値が小さいほど当該工程へ優先割当。
        #             数字省略の OP/AS は優先度 1（従来どおり最優先扱い）。
        # を基本としつつ、旧仕様（1行ヘッダ）にもフォールバック対応する。
        skills_raw = pd.read_excel(MASTER_FILE, sheet_name="skills", header=None)
        skills_dict = {}
        equipment_list = []
        members = []

        use_two_header = False
        if skills_raw.shape[0] >= 3 and skills_raw.shape[1] >= 2:
            non_empty_pm = 0
            for c in range(1, skills_raw.shape[1]):
                p = skills_raw.iat[0, c]
                m = skills_raw.iat[1, c]
                if pd.isna(p) or pd.isna(m):
                    continue
                p_s = str(p).strip()
                m_s = str(m).strip()
                if p_s and m_s and p_s.lower() != "nan" and m_s.lower() != "nan":
                    non_empty_pm += 1
            use_two_header = non_empty_pm > 0

        if use_two_header:
            pm_cols = []
            seen_combo = set()
            for c in range(1, skills_raw.shape[1]):
                p = skills_raw.iat[0, c]
                m = skills_raw.iat[1, c]
                if pd.isna(p) or pd.isna(m):
                    continue
                p_s = str(p).strip()
                m_s = str(m).strip()
                if not p_s or not m_s or p_s.lower() == "nan" or m_s.lower() == "nan":
                    continue
                combo = f"{p_s}+{m_s}"
                pm_cols.append((c, p_s, m_s, combo))
                if combo not in seen_combo:
                    seen_combo.add(combo)
                    equipment_list.append(combo)

            for r in range(2, skills_raw.shape[0]):
                m_name_raw = skills_raw.iat[r, 0]
                if pd.isna(m_name_raw):
                    continue
                m_name = str(m_name_raw).strip()
                if not m_name or m_name.lower() in ("nan", "none", "null"):
                    continue
                row_skills = {}
                for c, p_s, m_s, combo in pm_cols:
                    v = skills_raw.iat[r, c] if c < skills_raw.shape[1] else None
                    sval = "" if pd.isna(v) else str(v).strip()
                    if not sval or sval.lower() in ("nan", "none", "null"):
                        continue
                    row_skills[combo] = sval
                    if m_s not in row_skills:
                        row_skills[m_s] = sval
                    if p_s not in row_skills:
                        row_skills[p_s] = sval
                skills_dict[m_name] = row_skills
            members = list(skills_dict.keys())
            logging.info(
                "skillsシート: 2段ヘッダ形式で読み込みました（工程+機械=%s列, メンバー=%s人）。",
                len(pm_cols),
                len(members),
            )
        else:
            skills_df = pd.read_excel(MASTER_FILE, sheet_name="skills")
            skills_df.columns = skills_df.columns.str.strip()
            skill_cols = [
                str(c).strip()
                for c in skills_df.columns
                if not str(c).startswith("Unnamed")
            ]

            member_col = None
            for c in skill_cols:
                if c in ("メンバー", "担当者", "氏名", "作業者"):
                    member_col = c
                    break
            if member_col is None and skill_cols:
                member_col = skill_cols[0]
                logging.warning(
                    "skillsシート: メンバー列名が標準と一致しないため、先頭列 '%s' をメンバー列として扱います。",
                    member_col,
                )

            seen_eq = set()
            for c in skill_cols:
                if c == member_col:
                    continue
                cid = str(c).strip()
                if not cid or cid.lower() in ("nan", "none", "null"):
                    continue
                if cid not in seen_eq:
                    seen_eq.add(cid)
                    equipment_list.append(cid)

            for _, row in skills_df.iterrows():
                m_name = str(row.get(member_col, "")).strip() if member_col else ""
                if not m_name or m_name.lower() == "nan":
                    continue
                row_skills = {}
                for c in skill_cols:
                    if c == member_col:
                        continue
                    sval = str(row.get(c, "")).strip()
                    if not sval or sval.lower() in ("nan", "none", "null"):
                        continue
                    row_skills[c] = sval
                    if "+" in c:
                        p, m = c.split("+", 1)
                        p = p.strip()
                        m = m.strip()
                        if m and m not in row_skills:
                            row_skills[m] = sval
                        if p and p not in row_skills:
                            row_skills[p] = sval
                skills_dict[m_name] = row_skills
            members = list(skills_dict.keys())
            logging.info(
                "skillsシート: 1行ヘッダ形式（旧互換）で読み込みました（メンバー=%s人）。",
                len(members),
            )

        if not members:
            logging.error("skillsシートからメンバーを読み込めませんでした。")
        else:
            _validate_skills_op_as_priority_numbers_unique(skills_dict, equipment_list)

        # need は header=None で読み、先頭の複数行を“見出し行”として解釈
        needs_raw = pd.read_excel(MASTER_FILE, sheet_name="need", header=None)
        col0 = 0
        process_header_row = None
        machine_header_row = None
        base_row = None

        for r in range(needs_raw.shape[0]):
            v0 = needs_raw.iat[r, col0]
            if pd.isna(v0):
                continue
            s0 = str(v0).strip()
            if process_header_row is None and s0 == "工程名":
                process_header_row = r
            elif machine_header_row is None and s0 == "機械名":
                machine_header_row = r
            if base_row is None and "必要人数" in s0 and not s0.startswith("特別指定"):
                base_row = r
            if process_header_row is not None and machine_header_row is not None and base_row is not None:
                break

        if process_header_row is None or machine_header_row is None or base_row is None:
            raise ValueError("need シートのヘッダー行（工程名/機械名/基本必要人数）が見つかりません。")

        # 「依頼NO条件」列位置（デフォルトは 1列目）
        cond_col_idx = 1
        for r in range(needs_raw.shape[0]):
            c1 = needs_raw.iat[r, 1] if needs_raw.shape[1] > 1 else None
            c2 = needs_raw.iat[r, 2] if needs_raw.shape[1] > 2 else None
            if pd.isna(c1) or pd.isna(c2):
                continue
            if str(c1).strip() == NEED_COL_CONDITION and str(c2).strip() == NEED_COL_NOTE:
                cond_col_idx = 1
                break

        # 工程名×機械名 の列一覧（列番号は Excel上の実列を保持）
        pm_cols = []
        for col_idx in range(needs_raw.shape[1]):
            if col_idx < 3:
                continue
            p = needs_raw.iat[process_header_row, col_idx]
            m = needs_raw.iat[machine_header_row, col_idx]
            if pd.isna(p) or pd.isna(m):
                continue
            p_s = str(p).strip()
            m_s = str(m).strip()
            if not p_s or not m_s or p_s.lower() == "nan" or m_s.lower() == "nan":
                continue
            pm_cols.append((col_idx, p_s, m_s))

        req_map = {}
        # need_rules: [{'order': int, 'condition': str, 'overrides': {combo_key/machine/process: int}}]
        need_rules = []

        # 基本必要人数
        for col_idx, p_s, m_s in pm_cols:
            n = parse_optional_int(needs_raw.iat[base_row, col_idx])
            if n is None or n < 1:
                n = 1
            combo_key = f"{p_s}+{m_s}"
            req_map[combo_key] = n
            # フォールバック用（機械名 or 工程名だけで引けるようにする）
            if p_s not in req_map:
                req_map[p_s] = n
            if m_s not in req_map:
                req_map[m_s] = n

        surplus_map: dict[str, int] = {}
        surplus_row = _find_need_surplus_add_row_index(
            needs_raw, base_row, col0, pm_cols
        )
        if surplus_row is not None:
            for col_idx, p_s, m_s in pm_cols:
                raw_ex = parse_optional_int(needs_raw.iat[surplus_row, col_idx])
                ex = int(raw_ex) if raw_ex is not None and raw_ex >= 0 else 0
                ex = max(0, min(ex, 50))
                combo_key = f"{p_s}+{m_s}"
                surplus_map[combo_key] = ex
                if p_s not in surplus_map:
                    surplus_map[p_s] = ex
                if m_s not in surplus_map:
                    surplus_map[m_s] = ex
            logging.info(
                "need シート: 配台時追加人数行を検出（Excel行≈%s）。列ごとの上限を読み込みました。",
                surplus_row + 1,
            )
        else:
            logging.info(
                "need シート: 基本必要人数の直下に配台時追加人数行を検出できませんでした（省略可）。"
            )

        if TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
            logging.info(
                "TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW が有効: 配台時追加人数は読み込んでも常に 0 扱い（チームは基本必要人数のみ試行）。"
            )

        logging.info(
            "need人数マスタ: %s を都度 read_excel（need シート専用のディスクキャッシュは無し・AI json キャッシュとは無関係）。",
            os.path.abspath(MASTER_FILE),
        )
        for _ci, _ps, _ms in pm_cols:
            _ck = f"{_ps}+{_ms}"
            _bn = req_map.get(_ck)
            _sx = surplus_map.get(_ck, 0) if surplus_map else 0
            logging.info(
                "need列サマリ combo=%r 基本必要人数=%s 配台時追加人数上限=%s",
                _ck,
                _bn,
                _sx,
            )

        # 特別指定
        for r in range(needs_raw.shape[0]):
            v0 = needs_raw.iat[r, col0]
            if pd.isna(v0):
                continue
            lab = str(v0).strip()
            m = re.match(r"特別指定\s*(\d+)", lab)
            if not m:
                continue
            order = int(m.group(1))
            if order < 1 or order > 99:
                continue

            cond_raw = needs_raw.iat[r, cond_col_idx] if needs_raw.shape[1] > cond_col_idx else None
            cond = "" if pd.isna(cond_raw) else str(cond_raw).strip()

            overrides = {}
            for col_idx, p_s, m_s in pm_cols:
                v = needs_raw.iat[r, col_idx]
                n = parse_optional_int(v)
                if n is not None and 1 <= n <= 99:
                    combo_key = f"{p_s}+{m_s}"
                    overrides[combo_key] = n
                    # フォールバック用
                    overrides[p_s] = n
                    overrides[m_s] = n

            if overrides:
                need_rules.append({"order": order, "condition": cond, "overrides": overrides})

        need_rules.sort(key=lambda rr: rr["order"])
        logging.info(f"need 特別指定ルール: {len(need_rules)} 件（工程名+機械名キー）。")

        logging.info(f"『{MASTER_FILE}』からスキルと設備要件(need)を読み込みました。")
        return skills_dict, members, equipment_list, req_map, need_rules, surplus_map

    except PlanningValidationError:
        raise
    except Exception as e:
        logging.error(f"マスタファイル({MASTER_FILE})のスキル/need読み込みエラー: {e}")
        return {}, [], [], {}, [], {}


def load_team_combination_presets_from_master() -> dict[
    str, list[tuple[int, int | None, tuple[str, ...], int | None]]
]:
    """
    master.xlsm「組み合わせ表」を読み、工程+機械キーごとに
    [(組合せ優先度, 必要人数またはNone, メンバータプル, 組合せ行IDまたはNone), ...] を返す。
    同一キー内は優先度昇順、同順位はシート上の行順。
    「必要人数」列は配台時に need 基本人数より優先する（メンバー列人数と一致すること）。
    A 列「組合せ行ID」が無い／空の旧シートでは ID は None。
    """
    if not TEAM_ASSIGN_USE_MASTER_COMBO_SHEET:
        return {}
    path = MASTER_FILE
    if not os.path.isfile(path):
        return {}
    try:
        df = pd.read_excel(path, sheet_name=MASTER_SHEET_TEAM_COMBINATIONS, header=0)
    except Exception as e:
        logging.info("組み合わせ表シートの読込をスキップします: %s", e)
        return {}
    if df is None or df.empty:
        return {}

    def norm_cell(x) -> str:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        return str(x).strip()

    colmap = {norm_cell(c): c for c in df.columns if norm_cell(c)}
    id_c = colmap.get("組合せ行ID") or colmap.get("インデックス")
    proc_c = colmap.get("工程名")
    mach_c = colmap.get("機械名")
    combo_c = colmap.get("工程+機械")
    prio_c = colmap.get("組合せ優先度")
    req_c = colmap.get("必要人数")

    def mem_col_order(c) -> int:
        m = re.search(r"メンバー\s*(\d+)", norm_cell(c))
        return int(m.group(1)) if m else 9999

    mem_keys = sorted(
        [c for c in df.columns if norm_cell(str(c)).startswith("メンバー")],
        key=mem_col_order,
    )
    buckets: dict[
        str,
        list[tuple[int, int, int | None, tuple[str, ...], int | None]],
    ] = defaultdict(list)
    for row_i, (_, row) in enumerate(df.iterrows()):
        proc = norm_cell(row.get(proc_c)) if proc_c else ""
        mach = norm_cell(row.get(mach_c)) if mach_c else ""
        combo_cell = norm_cell(row.get(combo_c)) if combo_c else ""
        if proc and mach:
            key = f"{proc}+{mach}"
        elif combo_cell:
            key = combo_cell
        else:
            continue
        pr = parse_optional_int(row.get(prio_c)) if prio_c else None
        if pr is None:
            pr = 10**9
        sheet_req: int | None = None
        if req_c:
            sheet_req = parse_optional_int(row.get(req_c))
            if sheet_req is not None and sheet_req < 1:
                sheet_req = None
        sheet_combo_id: int | None = None
        if id_c:
            sheet_combo_id = parse_optional_int(row.get(id_c))
            if sheet_combo_id is not None and sheet_combo_id < 1:
                sheet_combo_id = None
        team: list[str] = []
        for mc in mem_keys:
            s = norm_cell(row.get(mc))
            if not s or s.lower() in ("nan", "none", "null"):
                continue
            team.append(s)
        if not team:
            continue
        buckets[key].append(
            (pr, row_i, sheet_req, tuple(team), sheet_combo_id)
        )

    out: dict[
        str, list[tuple[int, int | None, tuple[str, ...], int | None]]
    ] = {}
    for key, lst in buckets.items():
        lst.sort(key=lambda x: (x[0], x[1]))
        out[key] = [(t[0], t[2], t[3], t[4]) for t in lst]
    return out


def generate_default_calendar_dates(year, month):
    cal = calendar.Calendar()
    return [d for d in cal.itermonthdates(year, month) if d.year == year and d.month == month and d.weekday() < 5]

def parse_time_str(time_str, default_time):
    if time_str is None or pd.isna(time_str) or not str(time_str).strip() or str(time_str).strip().lower() == 'null':
        return default_time
    try:
        if isinstance(time_str, time): return time_str
        if isinstance(time_str, datetime): return time_str.time()
        time_str = str(time_str).strip()
        if len(time_str.split(':')) == 3:
            return datetime.strptime(time_str, "%H:%M:%S").time()
        return datetime.strptime(time_str, "%H:%M").time()
    except:
        return default_time

def infer_mid_break_from_reason(reason_text, start_t, end_t):
    """
    備考から中抜け時間を推定するローカル補正。
    AIが中抜けを返さない場合のフェイルセーフとして使う。
    master.xlsm カレンダー由来の休暇区分: 前休=午前年休・午後のみ勤務、後休=午後年休・午前のみ勤務（出勤簿.txt と同義）。
    """
    if reason_text is None:
        return None, None
    txt = str(reason_text).strip()
    if not txt or txt.lower() in ("nan", "none", "null", "通常"):
        return None, None

    noon_end = time(12, 0)
    afternoon_start = time(13, 0)
    mae_kyuu_day_start = time(12, 45)  # 出勤簿.txt 前休の所定出勤（午前は年休）
    # カレンダー記号と一致させる（シフト時刻が誤っている場合の補完用。正しい行では区間が空になり追加されない）
    if txt == "前休":
        # 正しい行は出勤 12:45 以降で補完不要。全日シフトの誤入力時は 12:45 までを中抜け（午前年休相当）
        if start_t and start_t < mae_kyuu_day_start:
            return start_t, mae_kyuu_day_start
        return None, None
    if txt == "後休":
        if end_t and afternoon_start < end_t:
            return afternoon_start, end_t
        return None, None

    # 1) 明示的な時刻範囲（例: 11:00-14:00 / 11:00～14:00）
    m = re.search(r"(\d{1,2}[:：]\d{2})\s*[~〜\-－ー]\s*(\d{1,2}[:：]\d{2})", txt)
    if m:
        s = parse_time_str(m.group(1).replace("：", ":"), None)
        e = parse_time_str(m.group(2).replace("：", ":"), None)
        if s and e and s < e:
            return s, e

    # 2) あいまい語（午前/午後/終日） + 現場離脱・休暇系キーワード
    # 「午後休みです」等は「午後」を含むが、旧ロジックは「抜け」等のみ見ており中抜け推定に到達しなかった
    leave_keywords = (
        "事務所", "会議", "教育", "研修", "外出", "離れ", "抜け", "中抜け", "打合せ",
        "休み", "休暇", "欠勤",
    )
    has_leave_hint = any(k in txt for k in leave_keywords)
    if not has_leave_hint:
        return None, None

    if ("終日" in txt) or ("1日" in txt and "通常" not in txt):
        return start_t, end_t
    if ("午前中" in txt) or ("午前" in txt):
        return start_t, noon_end
    if ("午後" in txt):
        return afternoon_start, end_t

    return None, None


# 結果_カレンダー(出勤簿) の退勤表示。VBA 出勤簿「後休」（午後年休）と同様に実質 12:00 終了とみなす。
_AFTERNOON_OFF_DISPLAY_END = time(12, 0)


def _reason_is_afternoon_off(reason: str) -> bool:
    """後休（午後年休・午前のみ勤務）または備考の午後休系。"""
    r = str(reason or "")
    return ("午後" in r and ("休" in r or "休み" in r)) or ("後休" in r)


def _reason_is_morning_off(reason: str) -> bool:
    """前休（午前年休・午後のみ勤務）。カレンダー由来の略号のみ明示扱い（事務所勤務などと混同しない）。"""
    return "前休" in str(reason or "")


def _calendar_display_clock_out_for_calendar_sheet(entry: dict, day_date: date):
    """
    配台は breaks_dt の午後中抜けで正しくなる一方、end_dt が 17:00 のままだと結果カレンダーの退勤列だけ誤る。
    後休（午後年休）または備考が午後休み系で、定時まで続く午後の中抜けがあるときだけ退勤表示を 12:00 に揃える（end_dt 本体は変更しない）。
    """
    if not entry.get("is_working"):
        return None
    end_dt = entry.get("end_dt")
    if end_dt is None:
        return None
    reason = str(entry.get("reason") or "")
    afternoon_off = _reason_is_afternoon_off(reason)
    if not afternoon_off:
        return None
    breaks_dt = entry.get("breaks_dt") or []
    for b_s, b_e in breaks_dt:
        if b_s is None or b_e is None:
            continue
        bs = b_s.time() if isinstance(b_s, datetime) else b_s
        if isinstance(bs, datetime):
            bs = bs.time()
        if bs < time(12, 0):
            continue
        if isinstance(b_e, datetime):
            be_cmp = b_e
        elif isinstance(b_e, time):
            be_cmp = datetime.combine(day_date, b_e)
        else:
            continue
        if be_cmp >= end_dt - timedelta(seconds=1):
            return datetime.combine(day_date, _AFTERNOON_OFF_DISPLAY_END)
    return None


def _member_schedule_break_cell_label(grid_mid_dt, breaks_dt, shift_end_dt, reason):
    """
    個人_* スケジュールの10分枠が休憩帯に入るときの文言。
    昼食など通常休憩は「休憩」。後休（午後年休）で定時まで工場にいない午後帯は「休暇」。
    前休（午前年休）で午前の欠勤区間が休憩帯として入っている場合は「休暇」。
    """
    reason = str(reason or "")
    afternoon_off = _reason_is_afternoon_off(reason)
    morning_off = _reason_is_morning_off(reason)
    for b_s, b_e in breaks_dt:
        if b_s is None or b_e is None:
            continue
        if not (b_s <= grid_mid_dt < b_e):
            continue
        if isinstance(b_e, datetime) and isinstance(shift_end_dt, datetime):
            bs = b_s.time() if isinstance(b_s, datetime) else b_s
            if isinstance(bs, datetime):
                bs = bs.time()
            if afternoon_off and bs >= time(12, 0) and b_e >= shift_end_dt - timedelta(seconds=2):
                return "休暇"
            if morning_off and bs < time(12, 0):
                be_t = b_e.time() if isinstance(b_e, datetime) else b_e
                if be_t <= time(13, 0):
                    return "休暇"
        return "休憩"
    return None


def _member_schedule_off_shift_label(
    day_date: date,
    grid_mid_dt: datetime,
    d_start_dt: datetime,
    d_end_dt: datetime,
    reason: str,
) -> str:
    """
    個人_* シートで所定出退勤の外側の10分枠。
    前休の午前（工場日の所定開始～午後出勤まで）は年休、後休の午後は年休。それ以外のシフト外は勤務外。
    """
    r = str(reason or "")
    day_start = datetime.combine(day_date, DEFAULT_START_TIME)
    day_end = datetime.combine(day_date, DEFAULT_END_TIME)
    if grid_mid_dt < d_start_dt:
        if _reason_is_morning_off(r) and grid_mid_dt >= day_start:
            return "年休"
        return "勤務外"
    if grid_mid_dt >= d_end_dt:
        if _reason_is_afternoon_off(r) and grid_mid_dt < day_end:
            return "年休"
        return "勤務外"
    return "勤務外"


def _member_schedule_full_day_off_label(entry) -> str:
    """
    全日非勤務（is_working=False）の個人シート列の表示。
    休暇区分が年休（カレンダー *）のときは『年休』。工場休日などは『休』。
    """
    if not entry:
        return "休"
    r = str(entry.get("reason") or "").strip()
    if r == "年休" or r.startswith("年休 "):
        return "年休"
    return "休"


def _attendance_remark_text(row) -> str:
    """
    勤怠1行から「備考」列のテキストのみ取得する。
    勤怠AIの解析リストへの投入はこの列のみ。reason 文字列は load_attendance で備考と休暇区分を合成する。
    """
    if row is None:
        return ""
    try:
        v = row.get(ATT_COL_REMARK)
    except Exception:
        return ""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _attendance_leave_type_text(row) -> str:
    """勤怠1行から「休暇区分」列（カレンダー由来の 前休/後休 等）。"""
    if row is None:
        return ""
    try:
        v = row.get(ATT_COL_LEAVE_TYPE)
    except Exception:
        return ""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _ai_json_bool(v, default: bool = False) -> bool:
    """勤怠備考 AI の真偽値（bool / 数値 / 文字列の揺れを吸収）。"""
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v != 0
    if isinstance(v, float):
        if pd.isna(v):
            return default
        return v != 0.0
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "はい", "真", "on"):
        return True
    if s in ("false", "0", "no", "n", "いいえ", "偽", "off", ""):
        return False
    return default


def load_attendance_and_analyze(members):
    attendance_data = {}
    # ※「勤怠備考」は master 各メンバーシートの「備考」列のみ。メイン再優先・特別指定_備考は別API（generate_plan 側で追記）。
    ai_log = {
        "（注）このシートの見方": "先頭2行は勤怠「備考」の出退勤AIのみ。メイン再優先・特別指定は下段のJSONと「_*_AI_API」行。",
        "勤怠備考_AI_API": "なし",
        "勤怠備考_AI_詳細": "解析対象の備考行なし",
    }
    
    # 1. メンバー別シートからの読み込み
    all_records = []
    try:
        xls = pd.ExcelFile(MASTER_FILE)
        for sheet_name in xls.sheet_names:
            if "カレンダー" in sheet_name or sheet_name.lower() in ['skills', 'need', 'tasks']:
                continue 
                
            m_name = sheet_name.strip()
            if m_name not in members:
                continue 
                
            df_sheet = pd.read_excel(xls, sheet_name=sheet_name)
            df_sheet.columns = df_sheet.columns.str.strip()
            df_sheet['メンバー'] = m_name 
            all_records.append(df_sheet)
            
        if all_records:
            df = pd.concat(all_records, ignore_index=True)
            df['日付'] = pd.to_datetime(df['日付'], errors='coerce').dt.date
            df = df.dropna(subset=['日付'])
            logging.info(f"『{MASTER_FILE}』の各メンバーの勤怠シートを読み込みました。")
            _cols = {str(c).strip() for c in df.columns}
            if ATT_COL_REMARK in _cols and ATT_COL_LEAVE_TYPE in _cols:
                logging.info(
                    "勤怠列: AI 入力は「%s」のみ。備考が空の日は「%s」（前休・後休・他拠点勤務など）を reason に反映します。",
                    ATT_COL_REMARK,
                    ATT_COL_LEAVE_TYPE,
                )
            elif ATT_COL_REMARK not in _cols:
                logging.warning(
                    "勤怠データに「%s」列がありません。備考ベースの AI 解析は空扱いになります。",
                    ATT_COL_REMARK,
                )
        else:
            raise FileNotFoundError("有効なメンバー別勤怠シートが見つかりません。")
            
    except Exception as e:
        logging.warning(f"勤怠シート読み込みエラー: {e} デフォルトカレンダーを生成します。")
        default_dates = generate_default_calendar_dates(TARGET_YEAR, TARGET_MONTH)
        records = []
        for d in default_dates:
            for m in members: records.append({'日付': d, 'メンバー': m, '備考': '通常'})
        df = pd.DataFrame(records)

    # 2. AI による勤怠文脈の解析（備考が空でも休暇区分のみの行は AI に渡し、表記揺れはモデルに解釈させる）
    remarks_to_analyze = []
    for _, row in df.iterrows():
        m = str(row.get('メンバー', '')).strip()
        if m not in members:
            continue
        rem = _attendance_remark_text(row)
        lt = _attendance_leave_type_text(row)
        d_str = row['日付'].strftime("%Y-%m-%d") if pd.notna(row['日付']) else ""
        if rem:
            remarks_to_analyze.append(f"{d_str}_{m} の備考: {rem}")
        elif lt and lt not in ("通常", ""):
            remarks_to_analyze.append(f"{d_str}_{m} の休暇区分（備考は空）: {lt}")

    if remarks_to_analyze:
        remarks_blob = "\n".join(remarks_to_analyze)
        cache_key = hashlib.sha256(
            (remarks_blob + "\n" + ATTENDANCE_REMARK_AI_SCHEMA_ID).encode("utf-8")
        ).hexdigest()
        ai_cache = load_ai_cache()

        # 同一備考セットはキャッシュを優先利用し、APIコールを節約
        cached_data = get_cached_ai_result(ai_cache, cache_key)
        if cached_data is not None:
            ai_parsed = cached_data
            ai_log["勤怠備考_AI_API"] = "なし(キャッシュ使用)"
            ai_log["勤怠備考_AI_詳細"] = "キャッシュヒット"
        elif not API_KEY:
            ai_parsed = {}
            ai_log["勤怠備考_AI_API"] = "なし"
            ai_log["勤怠備考_AI_詳細"] = "GEMINI_API_KEY未設定のため勤怠備考AIをスキップ"
            logging.info("GEMINI_API_KEY 未設定のため備考AI解析をスキップしました。")
        else:
            logging.info("■ AIが複数日の特記事項を解析中...")
            ai_log["勤怠備考_AI_API"] = "あり"
            
            prompt = f"""
            以下の各日・メンバーの備考を読み取り、出退勤時刻の変更や中抜け、休日の判定を行い、JSON形式で出力してください。
            マークダウン記号(``` 等)は一切含めず、純粋なJSON文字列のみを返してください。

            【JSONの出力形式（キー名を厳密に守ること）】
            {{
              "YYYY-MM-DD_メンバー名": {{
                "出勤時刻": "HH:MM", 
                "退勤時刻": "HH:MM", 
                "中抜け開始": "HH:MM",
                "中抜け終了": "HH:MM",
                "作業効率": 1.0,     
                "is_holiday": false,
                "配台不参加": false
              }}
            }}
            ・キー名は上記の日本語キーをそのまま使う（英語キーに置き換えない）
            ・出勤時刻/退勤時刻: 当該行の「備考」または「休暇区分（備考は空）」の文脈から推測。不明や変更なしなら null
            ・中抜け開始/終了: 一時的な離脱（中抜け・事務所・会議など）がある場合、その開始・終了。ない場合は null
            ・曖昧語の解釈例:
              - 「午前中は事務所で作業」=> 中抜け開始 "08:45", 中抜け終了 "12:00"
              - 「午後は会議」=> 中抜け開始 "13:00", 中抜け終了 "17:00"
            ・is_holiday: その日が会社に来ない・終日休暇・欠勤など **勤務自体がない** と判断できる場合のみ true。午前休・午後休など部分的な休みは false（中抜けや時刻で表現）
            ・配台不参加: 勤務はあるが **加工ラインへの配台（OP/AS の割当）に載せてはいけない** と読み取れる場合は true。表記は問わず意味で判断すること。
              例: 「配台不可」「配台ＮＧ」「ラインに乗らない」「月次点検のみ」「点検で一日」「事務のみ」「教育で現場不可」「手配なし」「アサイン不要」などの揺れや婉曲表現も含む。
              通常勤務で特に制限が読み取れない場合は false
            ・作業効率: 0.0〜1.0の数値
            
            【特記事項リスト】
            {chr(10).join(remarks_to_analyze)}
            """
            try:
                client = genai.Client(api_key=API_KEY)
                res = client.models.generate_content(model='gemini-2.5-flash', contents=prompt)
                record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
                match = re.search(r'\{.*\}', res.text, re.DOTALL)
                if match:
                    ai_parsed = json.loads(match.group(0))
                    put_cached_ai_result(ai_cache, cache_key, ai_parsed)
                    save_ai_cache(ai_cache)
                    ai_log["勤怠備考_AI_詳細"] = "解析成功"
                else:
                    ai_parsed = {}
                    ai_log["勤怠備考_AI_詳細"] = "JSONパース失敗"
            except Exception as e:
                err_text = str(e)
                is_quota_or_rate = ("429" in err_text) or ("RESOURCE_EXHAUSTED" in err_text)
                retry_sec = extract_retry_seconds(err_text) if is_quota_or_rate else None

                if is_quota_or_rate and retry_sec is not None:
                    wait_sec = min(max(retry_sec, 1.0), 90.0)
                    logging.warning(f"AI通信 429/RESOURCE_EXHAUSTED。{wait_sec:.1f}秒待機して1回だけ再試行します。")
                    time_module.sleep(wait_sec)
                    try:
                        res = client.models.generate_content(model='gemini-2.5-flash', contents=prompt)
                        record_gemini_response_usage(res, GEMINI_MODEL_FLASH)
                        match = re.search(r'\{.*\}', res.text, re.DOTALL)
                        if match:
                            ai_parsed = json.loads(match.group(0))
                            put_cached_ai_result(ai_cache, cache_key, ai_parsed)
                            save_ai_cache(ai_cache)
                            ai_log["勤怠備考_AI_詳細"] = "再試行で解析成功"
                        else:
                            ai_parsed = {}
                            ai_log["勤怠備考_AI_詳細"] = "再試行後JSONパース失敗"
                    except Exception as e2:
                        ai_parsed = {}
                        logging.warning(f"AI再試行エラー: {e2}")
                        ai_log["勤怠備考_AI_詳細"] = f"429後再試行失敗: {e2}"
                else:
                    ai_parsed = {}
                    logging.warning(f"AI通信エラー: {e}")
                    ai_log["勤怠備考_AI_詳細"] = str(e)
    else:
        ai_parsed = {}

    # 3. 日付ごとの制約辞書を構築
    for _, row in df.iterrows():
        if pd.isna(row['日付']): continue
        curr_date = row['日付']
        m = str(row.get('メンバー', '')).strip()
        if m not in members: continue

        if curr_date not in attendance_data:
            attendance_data[curr_date] = {}

        original_reason = _attendance_remark_text(row)
        leave_type = _attendance_leave_type_text(row)

        key = f"{curr_date.strftime('%Y-%m-%d')}_{m}"
        ai_info = ai_parsed.get(key, {})

        is_empty_shift = pd.isna(row.get('出勤時間')) and pd.isna(row.get('退勤時間')) and not ai_info
        is_holiday = _ai_json_bool(ai_info.get("is_holiday"), False) or is_empty_shift
        exclude_from_line = _ai_json_bool(ai_info.get("配台不参加"), False)

        ai_eff = ai_info.get("作業効率")
        excel_eff = row.get('作業効率')
        
        if ai_eff is not None:
            eff_val = ai_eff
        elif excel_eff is not None and not pd.isna(excel_eff):
            eff_val = excel_eff
        else:
            eff_val = 1.0
            
        try:
            efficiency = float(eff_val)
        except (ValueError, TypeError):
            efficiency = 1.0

        if original_reason:
            if (
                leave_type
                and leave_type not in ("通常", "")
                and leave_type not in original_reason
            ):
                reason = f"{leave_type} {original_reason}"
            else:
                reason = original_reason
        elif leave_type and leave_type not in ("通常", ""):
            reason = leave_type
        else:
            reason = '通常' if not is_empty_shift else '休日シフト'

        start_t = parse_time_str(ai_info.get("出勤時刻") or row.get('出勤時間'), DEFAULT_START_TIME)
        end_t = parse_time_str(ai_info.get("退勤時刻") or row.get('退勤時間'), DEFAULT_END_TIME)
        
        b1_s = parse_time_str(row.get('休憩時間1_開始'), DEFAULT_BREAKS[0][0])
        b1_e = parse_time_str(row.get('休憩時間1_終了'), DEFAULT_BREAKS[0][1])
        b2_s = parse_time_str(row.get('休憩時間2_開始'), DEFAULT_BREAKS[1][0])
        b2_e = parse_time_str(row.get('休憩時間2_終了'), DEFAULT_BREAKS[1][1])

        # ★追加: AIから中抜け時間を取得
        mid_break_s = parse_time_str(ai_info.get("中抜け開始"), None)
        mid_break_e = parse_time_str(ai_info.get("中抜け終了"), None)
        # AIが中抜けを返さなかった場合は、備考文言からローカル推定で補完
        if not (mid_break_s and mid_break_e):
            fb_s, fb_e = infer_mid_break_from_reason(reason, start_t, end_t)
            if fb_s and fb_e:
                mid_break_s, mid_break_e = fb_s, fb_e

        def combine_dt(t): return datetime.combine(curr_date, t) if t else None
        
        start_dt = combine_dt(start_t)
        end_dt = combine_dt(end_t)
        breaks_dt = []
        
        # 通常の休憩を追加
        if b1_s and b1_e: breaks_dt.append((combine_dt(b1_s), combine_dt(b1_e)))
        if b2_s and b2_e: breaks_dt.append((combine_dt(b2_s), combine_dt(b2_e)))
        
        # ★追加: 中抜け時間がある場合は、特別な「休憩」としてスケジュール計算に追加
        if mid_break_s and mid_break_e: breaks_dt.append((combine_dt(mid_break_s), combine_dt(mid_break_e)))
        
        is_working = not is_holiday
        attendance_data[curr_date][m] = {
            "is_working": is_working,
            "eligible_for_assignment": is_working and (not exclude_from_line),
            "start_dt": start_dt,
            "end_dt": end_dt,
            "breaks_dt": merge_time_intervals(breaks_dt),
            "efficiency": efficiency,
            "reason": reason,
        }

    return attendance_data, ai_log


# ---------------------------------------------------------------------------
# 全依頼共通: 加工内容列の工程順序 / 個別: EC→検査ロールパイプライン
# ---------------------------------------------------------------------------
ROLL_PIPELINE_EC_PROCESS = "EC"
ROLL_PIPELINE_EC_MACHINE = "EC機　湖南"
ROLL_PIPELINE_INSP_PROCESS = "検査"
ROLL_PIPELINE_INSP_MACHINE = "熱融着機　湖南"
ROLL_PIPELINE_INITIAL_BUFFER_ROLLS = 2
# 検査の割当上限 min に使う。同一依頼に EC 行が無いときは need・スキルに従い通常配台する（ec_done=0 固定で永久スキップしない）。
ROLL_PIPELINE_INSP_UNCAPPED_ROOM = 1.0e18

# 勤怠に載っている最終日までで割付が終わらないとき、最終日と同じシフト型で日付を延長する（オプション）。
# False のとき段階2はマスタ勤怠の日付範囲のみで割付し、残りは配台残・配台不可のままとする。
STAGE2_EXTEND_ATTENDANCE_CALENDAR = False
SCHEDULE_EXTEND_MAX_EXTRA_DAYS = 366

# 計画基準納期日を過ぎても当該依頼に残量があるとき、**その依頼NOだけ** due_basis を +1 し、
# 当該依頼の割当・タイムラインを巻き戻して**カレンダー先頭から**再シミュレーションする（他依頼の割当は維持）。
# マスタ勤怠の最終日を超えて後ろ倒しできない依頼は「配台残(勤務カレンダー不足)」とする。各再試行前に勤怠拡張分はマスタ日付へ戻す。
STAGE2_RETRY_SHIFT_DUE_ON_PARTIAL_REMAINING = True
# 計画基準納期の +1 日による巻き戻し再シミュは依頼NOごとに最大この回数（41 回目以降は当該依頼のみシフトせず、未完了行に納期見直し必要を付与し得る）。
STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS = 40

# True のとき、配台計画シートの読み込み行順（各依頼NOの初出行が早いほど先）で 1 依頼だけを
# 当日候補に残し、完走してから次依頼へ進む。**他依頼は一切その日配台されない**ため、
# アクティブ依頼の1行でも詰まると全体が配台不可に見える（ログ「依頼NO直列配台 直列後=1」）。
# 既定 False。厳密な依頼NO直列が必要なときだけ STAGE2_SERIAL_DISPATCH_BY_TASK_ID=1 を設定する。
STAGE2_SERIAL_DISPATCH_BY_TASK_ID = (
    os.environ.get("STAGE2_SERIAL_DISPATCH_BY_TASK_ID", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)

# True: ①残タスクのうち配台試行順が最小の1タスクだけを選び、1ロールずつ割付。
# ②原反投入日と同一日に開始する場合は 13:00 以降（same_day_raw_start_limit も 13:00）。
# ③④設備空きを max で繰り上げ（日内。翌日は日付ループでタイムラインシード）。
# ⑤⑥⑦⑧人の空きでチームを決め、ロールごとに avail を更新（同日は前ロールと同一チームを優先）。
# 無効化: 環境変数 STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST=0
STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST = os.environ.get(
    "STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")


def _clone_attendance_day_shifted(source_day: dict, old_date: date, new_date: date) -> dict:
    """メンバー別勤怠ブロックを new_date にシフトした浅いコピーを返す。"""
    delta_days = (new_date - old_date).days
    if delta_days == 0:
        return {m: dict(st) for m, st in source_day.items()}
    delta = timedelta(days=delta_days)
    out: dict = {}
    for m, st in source_day.items():
        new_st = dict(st)
        sd = st.get("start_dt")
        ed = st.get("end_dt")
        new_st["start_dt"] = sd + delta if sd else None
        new_st["end_dt"] = ed + delta if ed else None
        nb = []
        for pair in st.get("breaks_dt") or []:
            if len(pair) >= 2:
                a, b = pair[0], pair[1]
                if a is not None and b is not None:
                    nb.append((a + delta, b + delta))
        new_st["breaks_dt"] = merge_time_intervals(nb)
        out[m] = new_st
    return out


def _pick_extension_template_date(attendance_data: dict, plan_dates: list):
    """配台可能なメンバーが1人でもいる直近の日をテンプレに採用（最終日が全休でも有効な型を使う）。"""
    for i in range(len(plan_dates) - 1, -1, -1):
        d = plan_dates[i]
        day = attendance_data.get(d)
        if not day:
            continue
        if any(
            v.get("eligible_for_assignment", v.get("is_working", False))
            for v in day.values()
        ):
            return d
    return plan_dates[-1] if plan_dates else None


def _extend_attendance_one_calendar_day(
    attendance_data: dict,
    plan_dates: list,
) -> bool:
    """カレンダー上1日先を plan_dates に追加し、テンプレ日のシフト複製で attendance を埋める。失敗時 False。"""
    if not plan_dates:
        return False
    last_d = plan_dates[-1]
    next_d = last_d + timedelta(days=1)
    tmpl_d = _pick_extension_template_date(attendance_data, plan_dates)
    if tmpl_d is None:
        return False
    template = attendance_data.get(tmpl_d)
    if not template:
        return False
    attendance_data[next_d] = _clone_attendance_day_shifted(template, tmpl_d, next_d)
    plan_dates.append(next_d)
    logging.info(
        "配台完了まで勤怠を自動拡張: %s を追加（テンプレ=%s、メンバー数=%s）",
        next_d,
        tmpl_d,
        len(attendance_data[next_d]),
    )
    return True


def _iter_plan_dates_extending(
    plan_dates: list,
    attendance_data: dict,
    task_queue: list,
):
    """
    plan_dates を先頭から順に yield。末尾まで来ても残タスクがあれば勤怠を1日ずつ拡張して継続。
    plan_dates / attendance_data はインプレース更新される。
    """
    si = 0
    ext_used = 0
    while True:
        while si < len(plan_dates):
            yield plan_dates[si]
            si += 1
        pending = any(float(t.get("remaining_units") or 0) > 1e-12 for t in task_queue)
        if not pending:
            return
        if ext_used >= SCHEDULE_EXTEND_MAX_EXTRA_DAYS:
            logging.warning(
                "残タスクがありますが勤怠の自動拡張が上限（%s 日）に達しました。配台残・配台不可が残る可能性があります。",
                SCHEDULE_EXTEND_MAX_EXTRA_DAYS,
            )
            return
        if not _extend_attendance_one_calendar_day(attendance_data, plan_dates):
            logging.warning(
                "勤怠を1日拡張できませんでした（テンプレ日のデータ欠落）。残タスクは未割当のままです。"
            )
            return
        ext_used += 1


def _parse_process_content_tokens(val) -> list[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    s = unicodedata.normalize("NFKC", str(val).strip())
    if not s or s.lower() in ("nan", "none", "null"):
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def _collect_process_content_order_by_task_id(tasks_df) -> dict[str, list[str]]:
    """依頼NO → 加工内容の工程名リスト（表の上の方で最初に現れた非空の行を採用）。"""
    out: dict[str, list[str]] = {}
    if tasks_df is None or tasks_df.empty:
        return out
    for _, row in tasks_df.iterrows():
        tid = planning_task_id_str_from_plan_row(row)
        if not tid:
            continue
        parts = _parse_process_content_tokens(row.get(TASK_COL_PROCESS_CONTENT))
        if not parts:
            continue
        if tid not in out:
            out[tid] = parts
    return out


def _process_name_matches_kakou_content_tokens(
    process_name: str, content_tokens: list[str]
) -> bool:
    """
    工程名（配台計画の「工程名」列）が、元データの「加工内容」カンマ区切りトークンのいずれかと
    正規化一致するか。トークンが無い（加工内容未記入の依頼）は照合対象外として True。
    """
    if not content_tokens:
        return True
    proc = _normalize_process_name_for_rule_match(process_name)
    if not proc:
        return False
    for tok in content_tokens:
        if _normalize_process_name_for_rule_match(tok) == proc:
            return True
    return False


def _process_sequence_rank_for_machine(proc, order_list: list[str]):
    if not order_list:
        return None
    pn = _normalize_process_name_for_rule_match(proc)
    for i, token in enumerate(order_list):
        if _normalize_process_name_for_rule_match(token) == pn:
            return i
    return None


def _task_rank_int_or_none(task) -> int | None:
    r = task.get("process_sequence_rank")
    if r is None:
        return None
    try:
        return int(r)
    except (TypeError, ValueError):
        return None


def _plan_sheet_priority_sort_value(t: dict) -> int:
    """配台計画シートの「優先度」。小さいほど先。未入力・不正は 999。"""
    p = t.get("priority", 999)
    try:
        return int(p)
    except (TypeError, ValueError):
        return 999


def _task_blocked_by_same_request_dependency(task, task_queue) -> bool:
    """
    同一依頼NOの異なる工程を同時刻に回さない（配台ルール §A-1・§A-2）。
    - 両行に加工内容由来の rank があるときは rank のみで前後（§A-1）。
    - どちらかに rank が無いときは、配台計画シートの行順 same_request_line_seq で前後（§A-2）。
    §B-2: ``roll_pipeline_inspection`` 行が ``roll_pipeline_ec`` 先行により §A-1 で止まる場合、
    ``_roll_pipeline_inspection_assign_room`` > 0 なら当該ペアだけブロックしない（ロール並行）。
    """
    tid = str(task.get("task_id", "") or "").strip()
    if not tid:
        return False
    try:
        my_seq = int(task.get("same_request_line_seq", 0))
    except (TypeError, ValueError):
        my_seq = 0
    my_r = _task_rank_int_or_none(task)

    for t2 in task_queue:
        if str(t2.get("task_id", "") or "").strip() != tid:
            continue
        if float(t2.get("remaining_units") or 0) <= 1e-9:
            continue
        r2 = _task_rank_int_or_none(t2)
        try:
            s2 = int(t2.get("same_request_line_seq", 0))
        except (TypeError, ValueError):
            s2 = 0

        if my_r is not None and r2 is not None:
            precedes = r2 < my_r
        elif my_r is None and r2 is None:
            precedes = s2 < my_seq
        else:
            precedes = s2 < my_seq

        if precedes:
            if (
                task.get("roll_pipeline_inspection")
                and t2.get("roll_pipeline_ec")
                and _roll_pipeline_inspection_assign_room(task_queue, tid) > 1e-12
            ):
                continue
            return True
    return False


def _row_matches_roll_pipeline_ec(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_EC_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_EC_MACHINE)
    )


def _row_matches_roll_pipeline_inspection(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_INSP_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_INSP_MACHINE)
    )


def _pipeline_ec_roll_done_units(task_queue, tid: str) -> float:
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s


def _pipeline_inspection_roll_done_units(task_queue, tid: str) -> float:
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_inspection"):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s


def _task_queue_has_roll_pipeline_ec_for_tid(task_queue, task_id: str) -> bool:
    """同一依頼NOに EC（ロールパイプライン先行）タスクがキューに含まれるか。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if t.get("roll_pipeline_ec"):
            return True
    return False


def _pipeline_ec_fully_done_for_tid(task_queue, task_id: str) -> bool:
    """同一依頼NOの EC ロールパイプライン行がすべて残量ゼロ（完走）か。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    found = False
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        found = True
        if float(t.get("remaining_units") or 0) > 1e-9:
            return False
    return found


def _roll_pipeline_inspection_assign_room(task_queue, task_id: str) -> float:
    tid = str(task_id or "").strip()
    if not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
        return float(ROLL_PIPELINE_INSP_UNCAPPED_ROOM)
    ec_done = _pipeline_ec_roll_done_units(task_queue, task_id)
    insp_done = _pipeline_inspection_roll_done_units(task_queue, task_id)
    max_insp = max(0.0, ec_done - float(ROLL_PIPELINE_INITIAL_BUFFER_ROLLS) + 1.0)
    # 先行バッファ式は ec_done と max_insp が 1 ずれるため、EC 完走直後に検査が 1 ロール足りなくなる。
    # EC が全ロール終了した後は検査も同数まで進められるよう上限を ec_done に合わせる。
    if _pipeline_ec_fully_done_for_tid(task_queue, task_id):
        max_insp = max(max_insp, ec_done)
    return max(0.0, max_insp - insp_done)


def _roll_pipeline_inspection_task_row_for_tid(
    task_queue: list, task_id: str
) -> dict | None:
    """同一依頼NOの §B-2 検査行（roll_pipeline_inspection）を1件返す。無ければ None。"""
    tid = str(task_id or "").strip()
    if not tid:
        return None
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if t.get("roll_pipeline_inspection"):
            return t
    return None


def _pipeline_b2_ec_roll_end_datetimes_sorted(
    task_queue: list, task_id: str
) -> list[datetime]:
    """同一依頼の EC ロール確定ごとの終了時刻を時系列で返す（assigned_history の end_dt）。"""
    tid = str(task_id or "").strip()
    ends: list[datetime] = []
    if not tid:
        return ends
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        for h in t.get("assigned_history") or []:
            ed = h.get("end_dt")
            if isinstance(ed, datetime):
                ends.append(ed)
    ends.sort()
    return ends


def _roll_pipeline_b2_inspection_ec_completion_floor_dt(
    task_queue: list, task_id: str
) -> datetime | None:
    """
    次の検査ロールを開始してよい最早時刻。
    累計検査完了ロール数を K、バッファを B（=ROLL_PIPELINE_INITIAL_BUFFER_ROLLS）とすると、
    EC 完了ロールが時系列で (K+B) 本目に到達した時刻（そのロールの end_dt）未満には開始しない。
    （業務ルール: 任意の時点で EC_RollEndCount - KENSA_RollEndCount >= B を満たすまで検査を進めない、
    の「ロール終了時刻基準」の実装。）
    """
    tid = str(task_id or "").strip()
    if not tid or not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
        return None
    insp_done = int(
        math.floor(float(_pipeline_inspection_roll_done_units(task_queue, tid)))
    )
    need_n = insp_done + int(ROLL_PIPELINE_INITIAL_BUFFER_ROLLS)
    ends = _pipeline_b2_ec_roll_end_datetimes_sorted(task_queue, tid)
    if need_n < 1 or len(ends) < need_n:
        return None
    return ends[need_n - 1]


def _pipeline_b2_team_history_names(team_cell) -> set[str]:
    """assigned_history の team 文字列（主・補を「,」「、」区切り）から担当者名を抽出（NFKC）。"""
    if team_cell is None:
        return set()
    s = str(team_cell).strip()
    if not s:
        return set()
    out: set[str] = set()
    for part in re.split(r"[,、]", s):
        t = part.strip()
        if t:
            out.add(unicodedata.normalize("NFKC", t))
    return out


def _pipeline_b2_assigned_member_names_nfkc_for_side(
    task_queue: list, task_id: str, *, ec_side: bool
) -> set[str]:
    """同一依頼の EC 行または検査行の assigned_history に出た担当者名（NFKC 集合）。"""
    tid = str(task_id or "").strip()
    if not tid:
        return set()
    names: set[str] = set()
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if ec_side:
            if not t.get("roll_pipeline_ec"):
                continue
        else:
            if not t.get("roll_pipeline_inspection"):
                continue
        for h in t.get("assigned_history") or []:
            names |= _pipeline_b2_team_history_names(h.get("team"))
    return names


def _b2_ec_insp_pair_in_queue(task_queue: list, task_id: str) -> bool:
    """同一依頼NOに §B-2 の EC 行と検査行の両方がキューにあるか。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    return bool(
        _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid)
        and _roll_pipeline_inspection_task_row_for_tid(task_queue, tid) is not None
    )


def _filter_capable_members_b2_disjoint_teams(
    task: dict, task_queue: list, capable_members: list
) -> list:
    """
    §B-2 同一依頼では、EC 行に一度でも入った者は検査の候補から外し、検査に入った者は EC の候補から外す。
    （社内ルール: 担当者集合を必ず分ける）
    """
    if not capable_members:
        return capable_members
    tid = str(task.get("task_id") or "").strip()
    if not tid or not _b2_ec_insp_pair_in_queue(task_queue, tid):
        return capable_members
    is_ec = bool(task.get("roll_pipeline_ec"))
    is_insp = bool(task.get("roll_pipeline_inspection"))
    if not is_ec and not is_insp:
        return capable_members
    if is_ec:
        excl = _pipeline_b2_assigned_member_names_nfkc_for_side(
            task_queue, tid, ec_side=False
        )
    else:
        excl = _pipeline_b2_assigned_member_names_nfkc_for_side(
            task_queue, tid, ec_side=True
        )
    if not excl:
        return capable_members
    return [
        m
        for m in capable_members
        if unicodedata.normalize("NFKC", str(m).strip()) not in excl
    ]


def _exclusive_b1_inspection_holder_for_machine(task_queue, line_key: str):
    """
    同一設備列（equipment_line_key／工程+機械）上で、§B-2 熱融着検査が **既に割付を開始** し残ロールが残る行があれば
    そのタスク dict を1件返す（なければ None）。

    パイプライン枠で検査を数ロールずつしか入れない設計のため、枠ゼロの隙間に **別依頼の検査** が同じ設備に入り、
    結果_設備毎の時間割でタスク表示が途中で切り替わる事象を防ぐ。占有中は当該設備列では他タスクを試行しない
    （別設備列は従来どおり）。
    """
    m = str(line_key or "").strip()
    if not m:
        return None
    holders: list = []
    for t in task_queue:
        lk = str(t.get("equipment_line_key") or t.get("machine") or "").strip()
        if lk != m:
            continue
        if not t.get("roll_pipeline_inspection"):
            continue
        rem = float(t.get("remaining_units") or 0)
        if rem <= 1e-9:
            continue
        init = float(t.get("initial_remaining_units") or 0)
        started = (init - rem) > 1e-9 or bool(t.get("assigned_history"))
        if not started:
            continue
        holders.append(t)
    if not holders:
        return None
    return min(
        holders,
        key=lambda t: (
            int(t.get("dispatch_trial_order") or 10**9),
            str(t.get("task_id") or ""),
            int(t.get("same_request_line_seq") or 0),
        ),
    )


def _generate_plan_task_queue_sort_key(task: dict, req_map: dict, need_rules: list) -> tuple:
    """
    generate_plan 冒頭および納期シフト再試行時の task_queue.sort 用キー。
    §B を §A・一般キーより先に効かせるため、B-1（roll_pipeline_inspection かつ in_progress＋試行順土台）
    → B-2 帯（同一帯内は **EC（roll_pipeline_ec）を未着手検査より先**・§B-2）
    → その他の加工途中 → 以降は納期・優先度等。
    """
    rp = bool(task.get("roll_pipeline_inspection"))
    ip = bool(task.get("in_progress"))
    ec = bool(task.get("roll_pipeline_ec"))
    if rp and ip:
        b_tier = 0  # §B-1
    elif ec or (rp and not ip):
        # §B-2: EC と未着手検査を同じ b_tier にまとめ、b2_queue_sub で EC 先行
        b_tier = 1
    else:
        b_tier = 2
    if b_tier == 1:
        if ec:
            b2_queue_sub = 0
        elif rp and not ip:
            b2_queue_sub = 1
        else:
            b2_queue_sub = 2
    else:
        b2_queue_sub = 0
    try:
        pss = int(task.get("planning_sheet_row_seq") or 0)
    except (TypeError, ValueError):
        pss = 10**9
    b1_trial = (0, pss) if (rp and ip) else (1, 0)
    gen_ip = 0 if (not rp and ip) else 1
    return (
        task.get("due_source_rank", 9),
        0 if task.get("has_done_deadline_override") else 1,
        b_tier,
        b2_queue_sub,
        b1_trial,
        gen_ip,
        task["priority"],
        0 if task["due_urgent"] else 1,
        task["due_basis_date"] or date.max,
        _normalize_equipment_match_key(task.get("machine_name") or ""),
        _task_id_same_machine_due_tiebreak_key(task.get("task_id")),
        task["start_date_req"],
        _task_id_priority_key(task.get("task_id")),
        -resolve_need_required_op(
            task["machine"],
            task.get("machine_name", ""),
            task["task_id"],
            req_map,
            need_rules,
        ),
    )


def _reorder_task_queue_b2_ec_inspection_consecutive(task_queue: list) -> None:
    """
    §B-2: 同一 task_id の `roll_pipeline_ec` 行の直後に、対応する未着手 `roll_pipeline_inspection` を隣接させる。

    `_generate_plan_task_queue_sort_key` では b_tier=1 帯で EC を未着手検査より先にまとめるが、
    帯内の他キー（納期・優先度等）により「全 EC が先・全検査が後」に見えることがある。
    結果シートの「配台試行順番」は依頼ごとに EC→検査の連番に揃える。
    """
    if len(task_queue) < 2:
        return
    moved_tids: list[str] = []
    n_rounds = 0
    max_rounds = max(len(task_queue) * 4, 8)
    while n_rounds < max_rounds:
        n_rounds += 1
        by_tid: dict = {}
        for t in task_queue:
            tid = str(t.get("task_id") or "").strip()
            if not tid:
                continue
            if t.get("roll_pipeline_ec"):
                by_tid.setdefault(tid, {})["ec"] = t
            if t.get("roll_pipeline_inspection") and not t.get("in_progress"):
                by_tid.setdefault(tid, {})["insp"] = t
        pairs = []
        for tid, d in by_tid.items():
            ec_t, insp_t = d.get("ec"), d.get("insp")
            if ec_t is not None and insp_t is not None:
                pairs.append((tid, ec_t, insp_t))
        if not pairs:
            break
        pairs.sort(key=lambda x: task_queue.index(x[1]))
        moved = False
        for tid, ec_task, insp_task in pairs:
            try:
                ie = task_queue.index(ec_task)
                ii = task_queue.index(insp_task)
            except ValueError:
                continue
            if ii == ie + 1:
                continue
            if ii > ie:
                task_queue.pop(ii)
                ie = task_queue.index(ec_task)
                task_queue.insert(ie + 1, insp_task)
            else:
                task_queue.pop(ie)
                ii = task_queue.index(insp_task)
                task_queue.insert(ii, ec_task)
            moved_tids.append(tid)
            moved = True
            break
        if not moved:
            break
    if moved_tids:
        logging.info(
            "§B-2 配台試行順: EC と未着手検査を隣接した依頼NO（配台試行順番を依頼内連番に揃える）: %s",
            ",".join(moved_tids),
        )


def _day_schedule_task_sort_key(task: dict, task_queue: list | None = None):
    """
    同一日内の割付試行順。
    計画基準納期→機械名→依頼NOタイブレーク（同日同機械）のあと、
    §B-1（検査+熱融着機湖南かつ加工途中）は配台試行順 dispatch_trial_order を加工内容 rank より先に効かせる。
    続けて r、行順、dto、§B-2 帯では **EC（roll_pipeline_ec）を未着手検査より先**（b2_roll_pipeline_stage）、優先度、結果用ソートキー。
    同一設備列の隙間割り込みは _equipment_line_lower_dispatch_trial_still_pending で試行順を強制する。
    """
    dbk = task.get("due_basis_date")
    if not isinstance(dbk, date):
        dbk = date.max
    mk = _normalize_equipment_match_key(task.get("machine_name") or "")
    tb = _task_id_same_machine_due_tiebreak_key(task.get("task_id"))
    raw_r = task.get("process_sequence_rank")
    if raw_r is None:
        r = 10**9
    else:
        r = int(raw_r)
    try:
        line_seq = int(task.get("same_request_line_seq", 0))
    except (TypeError, ValueError):
        line_seq = 0
    try:
        dto = int(task.get("dispatch_trial_order") or 10**9)
    except (TypeError, ValueError):
        dto = 10**9
    rp = bool(task.get("roll_pipeline_inspection"))
    ip = bool(task.get("in_progress"))
    ec = bool(task.get("roll_pipeline_ec"))
    if ec:
        b2_roll_pipeline_stage = 0
    elif rp and not ip:
        b2_roll_pipeline_stage = 1
    else:
        b2_roll_pipeline_stage = 2
    b1_trial_early = (0, dto) if (rp and ip) else (1, 0)
    return (
        (
            dbk,
            mk,
            tb,
            b1_trial_early,
            r,
            line_seq,
            dto,
            b2_roll_pipeline_stage,
            _plan_sheet_priority_sort_value(task),
        )
        + _result_task_sheet_sort_key(task)
    )


def _equipment_line_lower_dispatch_trial_still_pending(
    task_queue: list, eq_line: str, my_dispatch_order: int, current_date: date
) -> bool:
    """
    同一 equipment_line_key（工程+機械の設備列）で、より小さい配台試行順の行がまだ残量を持つか。
    machine_avail_dt はチャンク間の隙間に後続試行順が入り込めるため、ここで順序を強制する。

    キュー先頭に残量があるだけではブロックしない。tasks_today と同様に
    start_date_req <= current_date の行だけを「先試行順の競合」とみなす。
    （まだ開始日に達していない行が全日ブロッカーになり、後続がほぼ配台不可になるのを防ぐ。）
    """
    line = (eq_line or "").strip()
    if not line:
        return False
    try:
        my_o = int(my_dispatch_order)
    except (TypeError, ValueError):
        my_o = 10**9
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        _sdr = t.get("start_date_req")
        if not isinstance(_sdr, date) or _sdr > current_date:
            continue
        t_line = str(
            t.get("equipment_line_key") or t.get("machine") or ""
        ).strip()
        if t_line != line:
            continue
        try:
            o = int(t.get("dispatch_trial_order") or 10**9)
        except (TypeError, ValueError):
            o = 10**9
        if o < my_o:
            return True
    return False


def _purge_attendance_days_not_in_set(attendance_data: dict, keep_dates: frozenset) -> None:
    """勤怠辞書からマスタに無い日付キーを削除する（自動拡張分の巻き戻し）。"""
    for dk in list(attendance_data.keys()):
        if dk not in keep_dates:
            del attendance_data[dk]


def _partial_task_id_due_shift_outcome(
    task_queue: list, task_id: str, calendar_last: date
) -> tuple[bool, bool]:
    """
    配台残の依頼NOについて納期+1日リトライの分類。
    戻り値: (shift_ok, calendar_shortfall)
    - shift_ok: 計画基準納期を持つ行があり、それらすべてで +1 日がマスタ最終計画日以下
    - calendar_shortfall: 計画基準納期を持つ行があり、いずれかで +1 日がマスタ最終計画日を超える
    基準納期が一行も無い依頼は (False, False)（通常の配台残のまま）。
    """
    tid = (task_id or "").strip()
    if not tid:
        return False, False
    rows = [t for t in task_queue if str(t.get("task_id", "") or "").strip() == tid]
    if not rows:
        return False, False
    basis_rows = [t for t in rows if t.get("due_basis_date") is not None]
    if not basis_rows:
        return False, False
    for t in basis_rows:
        db = t["due_basis_date"]
        if db + timedelta(days=1) > calendar_last:
            return False, True
    return True, False


def _shift_task_due_calendar_fields_one_day(task: dict, run_date: date) -> None:
    """
    配台残リトライ用: **内部の計画基準納期（due_basis_date）だけ**を +1 日する。
    結果_タスク一覧用の ``due_basis_date_result_sheet`` は変更しない（+1 前の日付を保持）。
    回答納期・指定納期・指定納期_上書きも配台計画シート由来のまま。
    due_urgent はずらした due_basis_date で再計算する。
    """
    if task.get("due_basis_date") is not None:
        task["due_basis_date"] = task["due_basis_date"] + timedelta(days=1)
    db = task.get("due_basis_date")
    if db is not None:
        task["due_urgent"] = db <= run_date


def _seed_avail_from_timeline_for_date(
    timeline_events: list,
    current_date: date,
    machine_avail_dt: dict,
    avail_dt: dict,
    machine_day_start: datetime,
) -> None:
    """同一日内の既存 timeline から設備空き・メンバー空きの下限を反映する（部分再配台用）。"""
    for e in timeline_events:
        if e.get("date") != current_date:
            continue
        end_dt = e.get("end_dt")
        if end_dt is None or not hasattr(end_dt, "replace"):
            continue
        mach = e.get("machine")
        if mach:
            prev = machine_avail_dt.get(mach, machine_day_start)
            if end_dt > prev:
                machine_avail_dt[mach] = end_dt
        op = str(e.get("op") or "").strip()
        if op and op in avail_dt:
            prev_m = avail_dt[op]
            if end_dt > prev_m:
                avail_dt[op] = end_dt
        sub_raw = e.get("sub") or ""
        for sn in str(sub_raw).split(","):
            sm = sn.strip()
            if sm and sm in avail_dt:
                prev_s = avail_dt[sm]
                if end_dt > prev_s:
                    avail_dt[sm] = end_dt


def _collect_task_ids_missed_deadline_after_day(task_queue: list, current_date: date) -> set:
    """
    当該日の終了時点で、計画基準納期日（当日含む）以前なのに残量が残る依頼NO。
    「納期日内に完遂できなかった」= 後ろ倒し再試行の候補。
    """
    out = set()
    eps = 1e-9
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= eps:
            continue
        db = t.get("due_basis_date")
        if db is None:
            continue
        sdr = t.get("start_date_req")
        if isinstance(sdr, date) and sdr > current_date:
            continue
        if current_date >= db:
            tid = str(t.get("task_id", "") or "").strip()
            if tid:
                out.add(tid)
    return out


def _normalize_timeline_task_id(ev: dict) -> str:
    return str(ev.get("task_id", "") or "").strip()


def _trial_order_flow_day_start_floor(
    task: dict, current_date: date, macro_run_date: date, macro_now_dt: datetime
) -> datetime:
    """原反投入日を起点に、その日の加工開始の下限時刻（同日は 13:00 以降を含む）。"""
    floor = datetime.combine(current_date, DEFAULT_START_TIME)
    # §B-2 検査（roll_pipeline_inspection）は EC 完了を待って開始できるため、
    # 原反投入日（=同日13:00以降）の制約をそのまま適用すると検査が不必要に後ろへ倒れる。
    # EC完了時刻下限（_roll_pipeline_b2_inspection_ec_completion_floor_dt）で整合を取る。
    is_b2_inspection = bool(task.get("roll_pipeline_inspection"))
    rid = task.get("raw_input_date")
    if not is_b2_inspection and isinstance(rid, date) and rid == current_date:
        floor = max(floor, datetime.combine(current_date, time(13, 0)))
    sdl = task.get("same_day_raw_start_limit")
    s_req = task.get("start_date_req")
    if (
        (not is_b2_inspection)
        and sdl
        and isinstance(s_req, date)
        and current_date == s_req
        and isinstance(sdl, time)
    ):
        floor = max(floor, datetime.combine(current_date, sdl))
    est = task.get("earliest_start_time")
    if (not is_b2_inspection) and isinstance(s_req, date) and current_date == s_req and est:
        if isinstance(est, time):
            floor = max(floor, datetime.combine(current_date, est))
    if current_date == macro_run_date and floor < macro_now_dt:
        floor = macro_now_dt
    return floor


def _trial_order_flow_eligible_tasks(
    tasks_today: list, task_queue: list, current_date: date
) -> list:
    out = []
    for task in tasks_today:
        if float(task.get("remaining_units") or 0) <= 1e-12:
            continue
        if _task_blocked_by_same_request_dependency(task, task_queue):
            continue
        if task.get("roll_pipeline_inspection") and (
            _roll_pipeline_inspection_assign_room(
                task_queue, str(task.get("task_id", "") or "").strip()
            )
            <= 1e-12
        ):
            continue
        if PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE:
            _b1_holder = _exclusive_b1_inspection_holder_for_machine(
                task_queue,
                str(
                    task.get("equipment_line_key") or task.get("machine") or ""
                ).strip(),
            )
            if _b1_holder is not None and _b1_holder is not task:
                continue
        machine = task["machine"]
        eq_line = str(
            task.get("equipment_line_key") or machine or ""
        ).strip() or machine
        try:
            _my_dispatch_ord = int(task.get("dispatch_trial_order") or 10**9)
        except (TypeError, ValueError):
            _my_dispatch_ord = 10**9
        if _equipment_line_lower_dispatch_trial_still_pending(
            task_queue, eq_line, _my_dispatch_ord, current_date
        ):
            continue
        out.append(task)
    return out


def _combo_preset_team_size_bounds(
    preset_team: tuple,
    sheet_req_n: int | None,
    max_team_size_need: int,
) -> tuple[int, int] | None:
    """
    組み合わせ表プリセット1行の人数範囲 (lo, hi)。need の基本人数よりシート側を優先する。
    - 必要人数列が正のときはメンバー列の人数と一致すること。
    - hi は need の上限と実人数の大きい方（プリセットが need より少人数でも採用可能）。
    """
    nmem = len(preset_team)
    if nmem < 1:
        return None
    if sheet_req_n is not None and sheet_req_n >= 1:
        if nmem != sheet_req_n:
            return None
        lo = sheet_req_n
    else:
        lo = nmem
    hi = max(max_team_size_need, nmem)
    if not (lo <= nmem <= hi):
        return None
    return lo, hi


def _plan_sheet_required_op_optional(task: dict) -> int | None:
    """加工計画の必要人数列が正の整数ならその値。無効なら None。"""
    ro = task.get("required_op")
    if ro is None or (isinstance(ro, float) and pd.isna(ro)):
        return None
    try:
        n = int(ro)
    except (TypeError, ValueError):
        return None
    return n if n >= 1 else None


def _append_legacy_dispatch_candidate_for_team(
    task: dict,
    team: tuple,
    avail_dt: dict,
    machine_avail_dt: dict,
    daily_status: dict,
    current_date: date,
    macro_run_date: date,
    macro_now_dt: datetime,
    skill_role_priority,
    eq_line: str,
    rq_base: int,
    extra_max: int,
    global_priority_override: dict,
    team_candidates: list,
    *,
    combo_sheet_row_id: int | None = None,
    combo_preset_team: tuple[str, ...] | None = None,
) -> bool:
    """レガシー日次配台ループ用: 単一チームが成立すれば team_candidates に 1 件追加して True。"""
    _gpo = global_priority_override or {}
    op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
    if not op_list:
        return False
    team_start = max(avail_dt[m] for m in team)
    if not _gpo.get("abolish_all_scheduling_limits"):
        machine_free_dt = machine_avail_dt.get(
            eq_line, datetime.combine(current_date, DEFAULT_START_TIME)
        )
        if team_start < machine_free_dt:
            team_start = machine_free_dt
        if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
            min_start_dt = datetime.combine(
                current_date, task["same_day_raw_start_limit"]
            )
            if team_start < min_start_dt:
                team_start = min_start_dt
        if current_date == task["start_date_req"] and task.get("earliest_start_time"):
            min_user_t = datetime.combine(
                current_date, task["earliest_start_time"]
            )
            if team_start < min_user_t:
                team_start = min_user_t
        if current_date == macro_run_date and team_start < macro_now_dt:
            team_start = macro_now_dt
    team_end_limit = min(daily_status[m]["end_dt"] for m in team)
    if team_start >= team_end_limit:
        return False
    team_breaks = []
    for m in team:
        team_breaks.extend(daily_status[m]["breaks_dt"])
    team_breaks = merge_time_intervals(team_breaks)

    def _refloor_legacy_roll(ts: datetime) -> datetime:
        ts = max(ts, max(avail_dt[m] for m in team))
        if not _gpo.get("abolish_all_scheduling_limits"):
            machine_free_dt = machine_avail_dt.get(
                eq_line, datetime.combine(current_date, DEFAULT_START_TIME)
            )
            if ts < machine_free_dt:
                ts = machine_free_dt
            if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
                min_start_dt = datetime.combine(
                    current_date, task["same_day_raw_start_limit"]
                )
                if ts < min_start_dt:
                    ts = min_start_dt
            if current_date == task["start_date_req"] and task.get("earliest_start_time"):
                min_user_t = datetime.combine(
                    current_date, task["earliest_start_time"]
                )
                if ts < min_user_t:
                    ts = min_user_t
            if current_date == macro_run_date and ts < macro_now_dt:
                ts = macro_now_dt
        return ts

    team_start_adj = _defer_team_start_past_prebreak_and_end_of_day(
        task,
        team,
        team_start,
        team_end_limit,
        team_breaks,
        _refloor_legacy_roll,
    )
    if team_start_adj is None:
        return False
    team_start = team_start_adj
    if team_start >= team_end_limit:
        return False

    avg_eff = sum(daily_status[m]["efficiency"] for m in team) / len(team)
    if avg_eff <= 0:
        avg_eff = 0.01
    t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
    if t_eff <= 0:
        t_eff = 1.0
    eff_time_per_unit = (
        task["base_time_per_unit"]
        / avg_eff
        / t_eff
        * _surplus_team_time_factor(rq_base, len(team), extra_max)
    )
    _, avail_mins, _ = calculate_end_time(team_start, 9999, team_breaks, team_end_limit)
    units_can_do = int(avail_mins / eff_time_per_unit)
    if units_can_do == 0:
        return False
    units_today = min(units_can_do, math.ceil(task["remaining_units"]))
    work_mins_needed = int(units_today * eff_time_per_unit)
    actual_end_dt, _, _ = calculate_end_time(
        team_start, work_mins_needed, team_breaks, team_end_limit
    )
    team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
    team_candidates.append(
        {
            "team": team,
            "team_start": team_start,
            "actual_end_dt": actual_end_dt,
            "units_today": units_today,
            "team_breaks": team_breaks,
            "avg_eff": avg_eff,
            "prio_sum": team_prio_sum,
            "op_list": op_list,
            "eff_time_per_unit": eff_time_per_unit,
            "combo_sheet_row_id": combo_sheet_row_id,
            "combo_preset_team": combo_preset_team,
        }
    )
    return True


def _assign_one_roll_trial_order_flow(
    task: dict,
    current_date: date,
    daily_status: dict,
    avail_dt: dict,
    machine_avail_dt: dict,
    task_queue: list,
    skills_dict: dict,
    members: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict,
    macro_run_date: date,
    macro_now_dt: datetime,
    preferred_team: tuple | None,
    _need_headcount_logged_orders: set,
    team_combo_presets: dict | None = None,
) -> dict | None:
    """
    1ロール分の最良チームを決定する。設備空き・日開始下限を team_start に織り込む。
    preferred_team が与えられ同一日内で成立すれば、組合せ探索より優先して採用する。
    戻り値: team(tuple), start_dt, end_dt, breaks, eff, op, eff_time_per_unit, extra_max, rq_base, need_src_line, extra_src_line, machine, machine_name, eq_line, req_num, max_team_size
    """
    machine = task["machine"]
    machine_name = str(task.get("machine_name", "") or "").strip()
    machine_proc = str(machine or "").strip()
    eq_line = str(task.get("equipment_line_key") or machine or "").strip() or machine
    _gpo = global_priority_override or {}

    plan_ro = _plan_sheet_required_op_optional(task)
    need_src_line = ""
    if TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY:
        req_num, need_src_line = resolve_need_required_op_explain(
            machine,
            machine_name,
            task["task_id"],
            req_map,
            need_rules,
        )
        if plan_ro is not None and plan_ro != req_num:
            need_src_line = (need_src_line + "；") if need_src_line else ""
            need_src_line += f"計画シート必要人数{plan_ro}は未使用（need基準={req_num}）"
    else:
        if plan_ro is not None:
            req_num = plan_ro
            need_src_line = f"計画シート「必要OP(上書)」={req_num}"
        else:
            req_num, need_src_line = resolve_need_required_op_explain(
                machine,
                machine_name,
                task["task_id"],
                req_map,
                need_rules,
            )
    if _gpo.get("ignore_need_minimum"):
        req_num = 1
        need_src_line = (
            (need_src_line + " → ") if need_src_line else ""
        ) + "メイン上書ignore_need_minimumでreq=1"

    skill_meta_cache: dict = {}

    def skill_role_priority(mem):
        if _gpo.get("ignore_skill_requirements"):
            return ("OP", 100)
        if mem not in skill_meta_cache:
            srow = skills_dict.get(mem, {})
            v = ""
            if machine_proc and machine_name:
                v = srow.get(f"{machine_proc}+{machine_name}", "")
            elif machine_name:
                v = srow.get(machine_name, "")
            elif machine_proc:
                v = srow.get(machine_proc, "")
            skill_meta_cache[mem] = parse_op_as_skill_cell(v)
        return skill_meta_cache[mem]

    capable_members = [m for m in avail_dt if skill_role_priority(m)[0] in ("OP", "AS")]
    capable_members.sort(key=lambda mm: (skill_role_priority(mm)[1], mm))
    capable_members = _filter_capable_members_b2_disjoint_teams(
        task, task_queue, capable_members
    )

    pref_raw = str(task.get("preferred_operator_raw") or "").strip()
    op_today = [m for m in capable_members if skill_role_priority(m)[0] == "OP"]
    pref_mem = (
        _resolve_preferred_op_to_member(pref_raw, op_today, members)
        if pref_raw
        else None
    )

    extra_max_sheet, extra_src_line = resolve_need_surplus_extra_max_explain(
        machine,
        machine_name,
        task["task_id"],
        surplus_map,
        need_rules,
    )
    if TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
        extra_max_sheet = 0
        extra_src_line = (
            (extra_src_line + " → ") if extra_src_line else ""
        ) + "TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROWで0"
    extra_max = (
        extra_max_sheet if TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS else 0
    )
    if (
        extra_max_sheet > 0
        and not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
        and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
    ):
        extra_src_line = (
            (extra_src_line + " → ") if extra_src_line else ""
        ) + "メインは基本人数のみ（余力枠は全配台後に未割当×スキルで追記）"
    max_team_size = min(req_num + extra_max, len(capable_members))
    if max_team_size < req_num:
        max_team_size = req_num
    rq_base = max(1, int(req_num))

    _dto_head = task.get("dispatch_trial_order")
    if _dto_head is not None and _dto_head not in _need_headcount_logged_orders:
        _need_headcount_logged_orders.add(_dto_head)
        logging.info(
            "need人数(試行順優先フロー) order=%s task=%s 工程/機械=%s/%s "
            "req_num=%s [%s] extra_max=%s [%s] max_team候補=%s capable=%s人",
            _dto_head,
            task["task_id"],
            machine,
            machine_name,
            req_num,
            need_src_line,
            extra_max,
            extra_src_line,
            max_team_size,
            len(capable_members),
        )

    day_floor = _trial_order_flow_day_start_floor(
        task, current_date, macro_run_date, macro_now_dt
    )
    machine_day_floor = datetime.combine(current_date, DEFAULT_START_TIME)
    b2_insp_ec_floor: datetime | None = None
    _tid_assign = str(task.get("task_id") or "").strip()
    if task.get("roll_pipeline_inspection") and _task_queue_has_roll_pipeline_ec_for_tid(
        task_queue, _tid_assign
    ):
        b2_insp_ec_floor = _roll_pipeline_b2_inspection_ec_completion_floor_dt(
            task_queue, _tid_assign
        )

    def _one_roll_from_team(
        team: tuple,
        min_n: int | None = None,
        max_n: int | None = None,
    ) -> dict | None:
        lo = req_num if min_n is None else min_n
        hi = max_team_size if max_n is None else max_n
        if len(team) < lo or len(team) > hi:
            return None
        op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
        if not op_list:
            return None
        if not all(m in daily_status for m in team):
            return None
        team_start = max(avail_dt[m] for m in team)
        if not _gpo.get("abolish_all_scheduling_limits"):
            machine_free_dt = machine_avail_dt.get(eq_line, machine_day_floor)
            if team_start < machine_free_dt:
                team_start = machine_free_dt
            if team_start < day_floor:
                team_start = day_floor
        if b2_insp_ec_floor is not None and team_start < b2_insp_ec_floor:
            team_start = b2_insp_ec_floor
        team_end_limit = min(daily_status[m]["end_dt"] for m in team)
        if team_start >= team_end_limit:
            return None
        team_breaks = []
        for m in team:
            team_breaks.extend(daily_status[m]["breaks_dt"])
        team_breaks = merge_time_intervals(team_breaks)

        def _refloor_trial_roll(ts: datetime) -> datetime:
            ts = max(ts, max(avail_dt[m] for m in team))
            if not _gpo.get("abolish_all_scheduling_limits"):
                mf = machine_avail_dt.get(eq_line, machine_day_floor)
                if ts < mf:
                    ts = mf
                if ts < day_floor:
                    ts = day_floor
            if b2_insp_ec_floor is not None and ts < b2_insp_ec_floor:
                ts = b2_insp_ec_floor
            return ts

        team_start_d = _defer_team_start_past_prebreak_and_end_of_day(
            task,
            team,
            team_start,
            team_end_limit,
            team_breaks,
            _refloor_trial_roll,
        )
        if team_start_d is None:
            return None
        team_start = team_start_d
        if team_start >= team_end_limit:
            return None

        avg_eff = sum(daily_status[m]["efficiency"] for m in team) / len(team)
        if avg_eff <= 0:
            avg_eff = 0.01
        t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
        if t_eff <= 0:
            t_eff = 1.0
        eff_time_per_unit = (
            task["base_time_per_unit"]
            / avg_eff
            / t_eff
            * _surplus_team_time_factor(rq_base, len(team), extra_max)
        )
        _, avail_mins, _ = calculate_end_time(
            team_start, 9999, team_breaks, team_end_limit
        )
        if int(avail_mins / eff_time_per_unit) < 1:
            return None
        work_mins_needed = int(eff_time_per_unit)
        actual_end_dt, _, _ = calculate_end_time(
            team_start, work_mins_needed, team_breaks, team_end_limit
        )
        if pref_mem and pref_mem in op_list:
            lead_op = pref_mem
        else:
            lead_op = min(op_list, key=lambda mm: (skill_role_priority(mm)[1], mm))
        team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
        return {
            "team": team,
            "team_start": team_start,
            "actual_end_dt": actual_end_dt,
            "team_breaks": team_breaks,
            "avg_eff": avg_eff,
            "prio_sum": team_prio_sum,
            "op_list": op_list,
            "eff_time_per_unit": eff_time_per_unit,
            "lead_op": lead_op,
        }

    # 特別指定: 同一日・連続ロールは前回チームを優先
    if preferred_team:
        pt = tuple(preferred_team)
        if all(m in capable_members and m in avail_dt for m in pt):
            got = _one_roll_from_team(pt)
            if got is not None:
                return {
                    **got,
                    "extra_max": extra_max,
                    "rq_base": rq_base,
                    "need_src_line": need_src_line,
                    "extra_src_line": extra_src_line,
                    "machine": machine,
                    "machine_name": machine_name,
                    "eq_line": eq_line,
                    "req_num": req_num,
                    "max_team_size": max_team_size,
                    "combo_sheet_row_id": None,
                    "combo_preset_team": None,
                }

    combo_key = (
        f"{machine_proc}+{machine_name}"
        if machine_proc and machine_name
        else ""
    )
    preset_rows = (
        (team_combo_presets or {}).get(combo_key)
        if (team_combo_presets and combo_key)
        else None
    )
    if preset_rows:
        for _prio, sheet_rs, preset_team, combo_row_id in preset_rows:
            bounds = _combo_preset_team_size_bounds(
                tuple(preset_team), sheet_rs, max_team_size
            )
            if bounds is None:
                continue
            lo_pt, hi_pt = bounds
            if pref_mem is not None and pref_mem not in preset_team:
                continue
            if not all(m in capable_members for m in preset_team):
                continue
            if sum(1 for m in preset_team if skill_role_priority(m)[0] == "OP") < 1:
                continue
            got = _one_roll_from_team(
                tuple(preset_team), min_n=lo_pt, max_n=hi_pt
            )
            if got is not None:
                return {
                    **got,
                    "extra_max": extra_max,
                    "rq_base": rq_base,
                    "need_src_line": need_src_line,
                    "extra_src_line": extra_src_line,
                    "machine": machine,
                    "machine_name": machine_name,
                    "eq_line": eq_line,
                    "req_num": req_num,
                    "max_team_size": max_team_size,
                    "combo_sheet_row_id": combo_row_id,
                    "combo_preset_team": tuple(preset_team),
                }

    team_candidates: list[dict] = []
    for tsize in range(req_num, max_team_size + 1):
        if (
            pref_mem is not None
            and pref_mem in capable_members
            and skill_role_priority(pref_mem)[0] == "OP"
        ):
            others = [m for m in capable_members if m != pref_mem]
            if tsize == 1:
                teams_iter = [(pref_mem,)]
            elif len(others) >= tsize - 1:
                teams_iter = [
                    tuple([pref_mem] + list(rest))
                    for rest in itertools.combinations(others, tsize - 1)
                ]
            else:
                teams_iter = itertools.combinations(capable_members, tsize)
        else:
            teams_iter = itertools.combinations(capable_members, tsize)

        for team in teams_iter:
            got = _one_roll_from_team(team)
            if got is not None:
                team_candidates.append(
                    {
                        **got,
                        "combo_sheet_row_id": None,
                        "combo_preset_team": None,
                    }
                )

    if not team_candidates:
        return None
    t_min = min(c["team_start"] for c in team_candidates)

    def _team_cand_key(c):
        return _team_assignment_sort_tuple(
            c["team"],
            c["team_start"],
            1,
            c["prio_sum"],
            t_min,
        )

    best_c = min(team_candidates, key=_team_cand_key)
    return {
        **best_c,
        "extra_max": extra_max,
        "rq_base": rq_base,
        "need_src_line": need_src_line,
        "extra_src_line": extra_src_line,
        "machine": machine,
        "machine_name": machine_name,
        "eq_line": eq_line,
        "req_num": req_num,
        "max_team_size": max_team_size,
        "combo_sheet_row_id": best_c.get("combo_sheet_row_id"),
        "combo_preset_team": best_c.get("combo_preset_team"),
    }


def _trial_order_first_schedule_pass(
    current_date: date,
    tasks_today: list,
    task_queue: list,
    daily_status: dict,
    machine_avail_dt: dict,
    avail_dt: dict,
    timeline_events: list,
    skills_dict: dict,
    members: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict,
    macro_run_date: date,
    macro_now_dt: datetime,
    _need_headcount_logged_orders: set,
    team_combo_presets: dict | None = None,
) -> bool:
    """
    ①当日候補を配台試行順の昇順に並べる（1 パス分）。
    **完全二相（§B-2）**: **フェーズ1**で熱融着**検査行を除く**候補（EC・他依頼・他工程）を試行順どおり
    **`_drain_rolls_for_task`** し、**フェーズ2**で §B-2 検査行だけを同順で `_drain_rolls_for_task` する。
    EC と検査を **交互に 1 ロールずつ試すと**同一担当者が途中で検査へ回り **EC がブロック**されるため、
    同一パス内では検査を試さず先に EC 等を詰める。
    検査の各ロールは `_roll_pipeline_inspection_assign_room` に加え
    `_roll_pipeline_b2_inspection_ec_completion_floor_dt` で **EC ロール終了時刻**の下限を満たす。
    試行順最小の行だけが当日入らない場合でも、**同じフェーズ内で次の試行順へ進み**他設備を埋める。
    機械・人の空きはロールごとに更新する（⑦⑧）。
    """
    eligible = _trial_order_flow_eligible_tasks(
        tasks_today, task_queue, current_date
    )
    if not eligible:
        return False
    eligible_sorted = sorted(
        eligible,
        key=lambda t: int(t.get("dispatch_trial_order") or 10**9),
    )
    _gpo = global_priority_override or {}

    def _drain_rolls_for_task(task: dict) -> bool:
        preferred_team: tuple | None = None
        made_local = False
        while float(task.get("remaining_units") or 0) > 1e-12:
            res = _assign_one_roll_trial_order_flow(
                task,
                current_date,
                daily_status,
                avail_dt,
                machine_avail_dt,
                task_queue,
                skills_dict,
                members,
                req_map,
                need_rules,
                surplus_map,
                global_priority_override,
                macro_run_date,
                macro_now_dt,
                preferred_team,
                _need_headcount_logged_orders,
                team_combo_presets,
            )
            if res is None:
                break
            done_units = 1
            if task.get("roll_pipeline_inspection"):
                _rp_room = _roll_pipeline_inspection_assign_room(
                    task_queue, str(task.get("task_id", "") or "").strip()
                )
                if _rp_room <= 1e-12:
                    break
                done_units = min(
                    1, int(min(_rp_room, math.ceil(task["remaining_units"])))
                )
            if done_units <= 0:
                break
            best_team = tuple(res["team"])
            lead_op = res["lead_op"]
            sub_members = [m for m in best_team if m != lead_op]
            best_start = res["team_start"]
            best_end = res["actual_end_dt"]
            best_breaks = res["team_breaks"]
            best_eff = res["avg_eff"]
            rq_base = res["rq_base"]
            extra_max = res["extra_max"]
            eq_line = res["eq_line"]
            _te_disp = parse_float_safe(task.get("task_eff_factor"), 1.0)
            if _te_disp <= 0:
                _te_disp = 1.0

            total_u = (
                math.ceil(task["total_qty_m"] / task["unit_m"]) if task["unit_m"] else 0
            )
            rem_u_before = math.ceil(task["remaining_units"])
            already_done = total_u - rem_u_before
            try:
                tot_qty = parse_float_safe(task.get("total_qty_m"), 0.0)
                done_qty = parse_float_safe(task.get("done_qty_reported"), 0.0)
                if tot_qty > 0:
                    pct_macro = max(
                        0, min(100, int(round((done_qty / tot_qty) * 100)))
                    )
                else:
                    pct_macro = 0
            except Exception:
                pct_macro = 0

            timeline_events.append(
                {
                    "date": current_date,
                    "task_id": task["task_id"],
                    "machine": eq_line,
                    "op": lead_op,
                    "sub": ", ".join(sub_members),
                    "start_dt": best_start,
                    "end_dt": best_end,
                    "breaks": best_breaks,
                    "units_done": done_units,
                    "already_done_units": already_done,
                    "total_units": total_u,
                    "pct_macro": pct_macro,
                    "eff_time_per_unit": task["base_time_per_unit"]
                    / best_eff
                    / _te_disp
                    * _surplus_team_time_factor(
                        rq_base, len(best_team), extra_max
                    ),
                    "unit_m": task["unit_m"],
                }
            )
            task["remaining_units"] -= float(done_units)
            op_main = (lead_op or "").strip()
            subs_part = ",".join(
                s.strip() for s in sub_members if s and str(s).strip()
            )
            team_s = f"{op_main}, {subs_part}" if subs_part else op_main
            req_num_run = int(res.get("req_num") or 0)
            extra_max_run = int(res.get("extra_max") or 0)
            need_surplus_assigned = (
                TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                and extra_max_run > 0
                and len(best_team) > req_num_run
            )
            names_ordered: list[str] = []
            if op_main:
                names_ordered.append(op_main)
            for _m in sub_members:
                if _m and str(_m).strip():
                    names_ordered.append(str(_m).strip())
            surplus_member_names = (
                names_ordered[req_num_run:]
                if need_surplus_assigned
                and len(names_ordered) > req_num_run
                else []
            )
            task["assigned_history"].append(
                {
                    "date": current_date.strftime("%m/%d"),
                    "team": team_s,
                    "done_m": int(done_units * task["unit_m"]),
                    "start_dt": best_start,
                    "end_dt": best_end,
                    "need_surplus_assigned": need_surplus_assigned,
                    "combo_sheet_row_id": res.get("combo_sheet_row_id"),
                    "surplus_member_names": surplus_member_names,
                }
            )
            for m in best_team:
                avail_dt[m] = best_end
            if not _gpo.get("abolish_all_scheduling_limits"):
                machine_avail_dt[eq_line] = best_end
            rem_after = int(math.ceil(float(task.get("remaining_units") or 0)))
            _dispatch_roll_trace_after_roll(
                current_date,
                task,
                eq_line,
                best_start,
                best_end,
                float(done_units),
                rem_after,
                lead_op,
                list(sub_members),
                roll_surplus_meta={
                    "surplus_phase": "main",
                    "req_num": req_num_run,
                    "team_size": len(best_team),
                    "extra_max_main_pass": extra_max_run,
                    "need_surplus_assigned": need_surplus_assigned,
                    "team_summary": team_s,
                },
            )
            if _trace_schedule_task_enabled(task.get("task_id")):
                _log_dispatch_trace_schedule(
                    task.get("task_id"),
                    "[配台トレース task=%s] ロール確定 メイン day=%s machine=%s machine_name=%s "
                    "start=%s end=%s 採用人数=%s req_num=%s メイン探索extra_max=%s "
                    "余剰人数適用(メイン)=%s team=%s",
                    task.get("task_id"),
                    current_date,
                    eq_line,
                    str(task.get("machine_name") or "").strip(),
                    best_start,
                    best_end,
                    len(best_team),
                    req_num_run,
                    extra_max_run,
                    need_surplus_assigned,
                    team_s,
                )
            if _dispatch_debug_should_stop_early():
                made_local = True
                return made_local
            preferred_team = best_team
            made_local = True
        return made_local

    def _is_b2_inspection_eligible_row(t: dict) -> bool:
        _tid = str(t.get("task_id") or "").strip()
        return bool(
            t.get("roll_pipeline_inspection")
            and _task_queue_has_roll_pipeline_ec_for_tid(task_queue, _tid)
        )

    phase1_tasks = [t for t in eligible_sorted if not _is_b2_inspection_eligible_row(t)]
    phase2_tasks = [t for t in eligible_sorted if _is_b2_inspection_eligible_row(t)]

    pass_made = False
    for task in phase1_tasks:
        if _drain_rolls_for_task(task):
            pass_made = True
    for task in phase2_tasks:
        if _drain_rolls_for_task(task):
            pass_made = True
    return pass_made


def _timeline_event_team_names_set(ev: dict) -> set:
    names: set = set()
    op = str(ev.get("op") or "").strip()
    if op:
        names.add(op)
    sub = str(ev.get("sub") or "").strip()
    if sub:
        for s in sub.split(","):
            t = s.strip()
            if t:
                names.add(t)
    return names


def _task_dict_for_timeline_event(ev: dict, task_queue: list) -> dict | None:
    tid = str(ev.get("task_id") or "").strip()
    if not tid:
        return None
    eq = str(ev.get("machine") or "").strip()
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        t_eq = str(t.get("equipment_line_key") or t.get("machine") or "").strip()
        if t_eq == eq:
            return t
    for t in task_queue:
        if str(t.get("task_id") or "").strip() == tid:
            return t
    return None


def _member_overlaps_busy(
    busy_map: dict, member: str, st: datetime, ed: datetime
) -> bool:
    for bs, be in busy_map.get(member, ()):
        if st < be and bs < ed:
            return True
    return False


def append_surplus_staff_after_main_dispatch(
    timeline_events: list,
    attendance_data: dict,
    skills_dict: dict,
    members: list,
    task_queue: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict | None,
) -> int:
    """
    need「配台時追加人数／余力時追加人数」行の上限まで、メイン割付で採用しきれなかった枠を追記する。
    各タイムラインブロックについて、その時間帯に他ブロックへ未参加（区間重なりなし）で
    eligible かつ OP/AS スキルの者をサブに追加する。
    """
    gpo = global_priority_override or {}
    if not surplus_map or TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
        return 0

    busy: dict[str, list[tuple[datetime, datetime]]] = defaultdict(list)
    for e in timeline_events:
        st = e.get("start_dt")
        ed = e.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime):
            continue
        for name in _timeline_event_team_names_set(e):
            busy[name].append((st, ed))

    appended_total = 0
    sorted_evs = sorted(
        (
            e
            for e in timeline_events
            if isinstance(e.get("start_dt"), datetime)
            and isinstance(e.get("end_dt"), datetime)
        ),
        key=lambda x: (x.get("date"), x.get("start_dt") or datetime.min),
    )

    for ev in sorted_evs:
        d = ev.get("date")
        if d is None or d not in attendance_data:
            continue
        daily_status = attendance_data[d]
        task = _task_dict_for_timeline_event(ev, task_queue)
        if task is None:
            continue
        machine = task.get("machine")
        machine_name = str(task.get("machine_name") or "").strip()
        tid = str(task.get("task_id") or "").strip()

        if TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY:
            req_num = resolve_need_required_op(
                str(machine or "").strip(),
                machine_name,
                tid,
                req_map,
                need_rules,
            )
        else:
            ro = task.get("required_op")
            if ro is not None:
                try:
                    riv = int(ro)
                    if riv >= 1:
                        req_num = riv
                    else:
                        req_num = resolve_need_required_op(
                            str(machine or "").strip(),
                            machine_name,
                            tid,
                            req_map,
                            need_rules,
                        )
                except (TypeError, ValueError):
                    req_num = resolve_need_required_op(
                        str(machine or "").strip(),
                        machine_name,
                        tid,
                        req_map,
                        need_rules,
                    )
            else:
                req_num = resolve_need_required_op(
                    str(machine or "").strip(),
                    machine_name,
                    tid,
                    req_map,
                    need_rules,
                )
        if gpo.get("ignore_need_minimum"):
            req_num = 1

        extra_max_sheet = resolve_need_surplus_extra_max(
            str(machine or "").strip(),
            machine_name,
            tid,
            surplus_map,
            need_rules,
        )
        if extra_max_sheet <= 0:
            continue

        names = _timeline_event_team_names_set(ev)
        team_size = len(names)
        cap_add = req_num + extra_max_sheet - team_size
        if cap_add <= 0:
            continue

        skill_meta_cache: dict = {}

        def skill_role_priority(mem):
            if gpo.get("ignore_skill_requirements"):
                return ("OP", 100)
            if mem not in skill_meta_cache:
                srow = skills_dict.get(mem, {})
                machine_proc = str(machine or "").strip()
                v = ""
                if machine_proc and machine_name:
                    v = srow.get(f"{machine_proc}+{machine_name}", "")
                elif machine_name:
                    v = srow.get(machine_name, "")
                elif machine_proc:
                    v = srow.get(machine_proc, "")
                skill_meta_cache[mem] = parse_op_as_skill_cell(v)
            return skill_meta_cache[mem]

        capable = []
        for mem in members:
            if mem not in daily_status:
                continue
            st_ent = daily_status[mem]
            if not st_ent.get(
                "eligible_for_assignment", st_ent.get("is_working", False)
            ):
                continue
            if skill_role_priority(mem)[0] not in ("OP", "AS"):
                continue
            capable.append(mem)
        capable.sort(key=lambda mm: (skill_role_priority(mm)[1], mm))

        st = ev["start_dt"]
        ed = ev["end_dt"]
        candidates = [
            m
            for m in capable
            if m not in names and not _member_overlaps_busy(busy, m, st, ed)
        ]
        candidates.sort(
            key=lambda mm: (
                0 if skill_role_priority(mm)[0] == "AS" else 1,
                skill_role_priority(mm)[1],
                mm,
            )
        )

        chosen = candidates[:cap_add]
        if not chosen:
            continue

        team_size_before = team_size
        final_team_size = team_size_before + len(chosen)
        highlight_surplus = final_team_size > req_num

        old_sub = str(ev.get("sub") or "").strip()
        parts = [s.strip() for s in old_sub.split(",") if s.strip()]
        parts.extend(chosen)
        ev["sub"] = ", ".join(parts)
        for m in chosen:
            busy[m].append((st, ed))
        appended_total += len(chosen)

        op_sync = str(ev.get("op") or "").strip()
        subs_sync = ",".join(
            s.strip()
            for s in str(ev.get("sub") or "").split(",")
            if s.strip()
        )
        team_sync = f"{op_sync}, {subs_sync}" if subs_sync else op_sync

        _hist = task.get("assigned_history")
        if _hist:
            for h in _hist:
                if (
                    h.get("start_dt") == st
                    and h.get("end_dt") == ed
                ):
                    if highlight_surplus:
                        h["need_surplus_assigned"] = True
                    h["team"] = team_sync
                    prev_pd = h.get("post_dispatch_surplus_names") or []
                    h["post_dispatch_surplus_names"] = prev_pd + [
                        str(x) for x in chosen
                    ]
                    break

        if _trace_schedule_task_enabled(tid):
            _log_dispatch_trace_schedule(
                tid,
                "[配台トレース task=%s] 余力追記(メイン完了後) day=%s machine=%s machine_name=%s "
                "start=%s end=%s 追記人数=%s 追記前人数=%s 追記後人数=%s req_num=%s "
                "need追加枠(シート)=%s 履歴黄(余剰人数超過)=%s 追記メンバー=%s",
                tid,
                d,
                str(machine or "").strip(),
                machine_name,
                st,
                ed,
                len(chosen),
                team_size_before,
                final_team_size,
                req_num,
                extra_max_sheet,
                highlight_surplus,
                ",".join(chosen),
            )

    return appended_total


# =========================================================
# 3. メイン計画生成 (日毎ループ・持ち越し対応)
#    段階2の本体。plan_simulation_stage2 からのみ呼ばれる想定。
#    配台計画シート読込 → タスクキュー → 日付ごとに設備・OP割付 → 結果ブック出力。
# =========================================================
def generate_plan():
    """
    段階2のメイン処理。戻り値なし（ログ・Excel 出力で完結）。

    前提: 環境変数 TASK_INPUT_WORKBOOK、カレントディレクトリがスクリプトフォルダ。
    出力: ``output_dir`` 直下の ``production_plan_multi_day_*.xlsx`` / ``member_schedule_*.xlsx``（最新1組のみ）、および log/execution_log.txt。
    """
    # ロールトレース JSONL は日次ループより前に早期 return しうるため、ここで初期化する
    # （DISPATCH_ROLL_TRACE_JSONL 未設定・空ならファイルは作らない）。
    _dispatch_debug_reset_roll_trace(
        (os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK)
    )
    # 配台トレース（設定シート A3 以降のみ）は、メンバー0人等で早期 return しても
    # execution_log に残るよう skills 読込より前で確定・ログする。
    global TRACE_SCHEDULE_TASK_IDS
    _wb_trace = (os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK)
    _ids_from_sheet = _read_trace_schedule_task_ids_from_config_sheet(_wb_trace)
    TRACE_SCHEDULE_TASK_IDS = frozenset(
        str(x).strip() for x in _ids_from_sheet if str(x).strip()
    )
    if _ids_from_sheet:
        _preview = _ids_from_sheet[:25]
        _suffix = " …" if len(_ids_from_sheet) > 25 else ""
        logging.info(
            "設定シート「%s」A3 以降: トレース用依頼NOを %s 件読み込み（%s%s）",
            APP_CONFIG_SHEET_NAME,
            len(_ids_from_sheet),
            ", ".join(_preview),
            _suffix,
        )
    else:
        logging.info(
            "設定シート「%s」A3 以降: トレース用依頼NOは無し（空またはシート無し）",
            APP_CONFIG_SHEET_NAME,
        )
    if TRACE_SCHEDULE_TASK_IDS:
        logging.info(
            "配台トレース: 有効 task_id = %s（設定シート A3 以降）",
            ", ".join(sorted(TRACE_SCHEDULE_TASK_IDS)),
        )
    else:
        logging.info(
            "配台トレース: 対象なし（[配台トレース …] ログは出ません）"
        )
    if TRACE_TEAM_ASSIGN_TASK_ID:
        logging.info(
            "環境変数 TRACE_TEAM_ASSIGN_TASK_ID=%r → チーム割当トレース有効",
            TRACE_TEAM_ASSIGN_TASK_ID,
        )

    global DISPATCH_DEBUG_ONLY_TASK_IDS
    _ids_debug_b = _read_dispatch_debug_only_task_ids_from_config_sheet_b(_wb_trace)
    DISPATCH_DEBUG_ONLY_TASK_IDS = frozenset()
    if _ids_debug_b:
        DISPATCH_DEBUG_ONLY_TASK_IDS = frozenset(
            planning_task_id_str_from_scalar(x) for x in _ids_debug_b
        )
        DISPATCH_DEBUG_ONLY_TASK_IDS = frozenset(x for x in DISPATCH_DEBUG_ONLY_TASK_IDS if x)
    if not DISPATCH_DEBUG_ONLY_TASK_IDS:
        DISPATCH_DEBUG_ONLY_TASK_IDS = frozenset(
            planning_task_id_str_from_scalar(x.strip())
            for x in _DISPATCH_DEBUG_ONLY_TASK_IDS_RAW.split(",")
            if x.strip()
        )
        DISPATCH_DEBUG_ONLY_TASK_IDS = frozenset(x for x in DISPATCH_DEBUG_ONLY_TASK_IDS if x)
        if _ids_debug_b and not DISPATCH_DEBUG_ONLY_TASK_IDS:
            logging.warning(
                "デバッグ限定配台: 設定シート「%s」B 列に値はありますが、依頼NOとして正規化できるものがありません。"
                " 環境変数 DISPATCH_DEBUG_ONLY_TASK_IDS も解釈できませんでした。この実行は全依頼を対象にします。",
                APP_CONFIG_SHEET_NAME,
            )
    if DISPATCH_DEBUG_ONLY_TASK_IDS:
        logging.warning(
            "デバッグ限定配台: 次の依頼NOのみ「配台計画_タスク入力」から配台します → %s",
            ", ".join(sorted(DISPATCH_DEBUG_ONLY_TASK_IDS)),
        )
    elif not _ids_debug_b:
        logging.info("デバッグ限定配台: 未指定（全依頼を対象）")

    _reset_dispatch_trace_per_task_logfiles()

    skills_dict, members, equipment_list, req_map, need_rules, surplus_map = (
        load_skills_and_needs()
    )
    team_combo_presets = load_team_combination_presets_from_master()
    if team_combo_presets:
        _nrules = sum(len(v) for v in team_combo_presets.values())
        logging.info(
            "組み合わせ表: 工程+機械キー %s 種類・編成行 %s を配台プリセットとして読み込みました。",
            len(team_combo_presets),
            _nrules,
        )
    elif TEAM_ASSIGN_USE_MASTER_COMBO_SHEET:
        logging.info(
            "組み合わせ表: プリセット無し（シート欠如・空・または読込失敗）。従来のチーム探索のみ。"
        )
    if not members:
        master_abs = os.path.abspath(MASTER_FILE)
        logging.error(
            "段階2を中断しました: メンバーが0人です（マスタの skills が空、または読み込み失敗）。"
            " 期待パス: %s （カレント: %s）。テストコード直下に master.xlsm を置き、"
            "planning_core のカレントがそのフォルダになるよう python\\ 配置を確認してください。"
            " この状態では production_plan / member_schedule は出力されません。",
            master_abs,
            os.getcwd(),
        )
        return
    reset_gemini_usage_tracker()
    _clear_stage2_blocking_message_file()
    if (
        not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
        and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
    ):
        logging.info(
            "need配台時追加人数: メイン割付は基本必要人数のみ。"
            "余力は全シミュレーション後、時間重なりのない未割当かつスキル適合者をサブに追記します。"
            "（メインで増員探索する従来挙動: TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS=1）"
        )

    # 段階2の基準日時は「マクロ実行時刻」ではなく「データ抽出日」を使用
    data_extract_dt = _extract_data_extraction_datetime()
    base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
    run_date = base_now_dt.date()
    data_extract_dt_str = (
        base_now_dt.strftime("%Y/%m/%d %H:%M:%S") if data_extract_dt is not None else "—"
    )
    logging.info(
        "計画基準日時: %s（%s）",
        base_now_dt.strftime("%Y/%m/%d %H:%M:%S"),
        "データ抽出日" if data_extract_dt is not None else "現在時刻フォールバック",
    )

    attendance_data, ai_log_data = load_attendance_and_analyze(members)
    global_priority_raw = load_main_sheet_global_priority_override_text()
    global_priority_override = analyze_global_priority_override_comment(
        global_priority_raw, members, run_date.year, ai_sheet_sink=ai_log_data
    )
    _factory_closure_dates: set[date] = set()
    for _iso in global_priority_override.get("factory_closure_dates") or []:
        _d = parse_optional_date(_iso)
        if _d is not None:
            _factory_closure_dates.add(_d)
    if _factory_closure_dates:
        apply_factory_closure_dates_to_attendance(
            attendance_data, members, _factory_closure_dates
        )
        logging.info(
            "メイン・グローバルコメント: 工場休業扱いの日付 → %s",
            ", ".join(str(x) for x in sorted(_factory_closure_dates)),
        )
    ai_log_data["メイン_グローバル_工場休業日(解析)"] = (
        ", ".join(str(x) for x in sorted(_factory_closure_dates))
        if _factory_closure_dates
        else "（なし）"
    )
    _sn = str(global_priority_override.get("scheduler_notes_ja") or "").strip()
    if _sn:
        ai_log_data["メイン_グローバル_未適用メモ(AI)"] = _sn[:2000]

    sorted_dates = sorted(list(attendance_data.keys()))
    # 結果シートは「基準日（データ抽出日）」以降のみ表示・計画対象とする
    sorted_dates = [d for d in sorted_dates if d >= run_date]
    if not sorted_dates:
        logging.error("当日以降の処理対象日付がありません。")
        _try_write_main_sheet_gemini_usage_summary("段階2")
        return

    # タスク入力: ブック内「配台計画_タスク入力」（段階1で出力→取り込み後に編集）
    try:
        tasks_df = load_planning_tasks_df()
    except Exception as e:
        logging.error(f"配台計画タスクシート読み込みエラー: {e}")
        _try_write_main_sheet_gemini_usage_summary("段階2")
        return

    if DISPATCH_DEBUG_ONLY_TASK_IDS:
        _n_tasks_before = len(tasks_df)
        _tid_ok = tasks_df[TASK_COL_TASK_ID].map(
            lambda v: planning_task_id_str_from_scalar(v) in DISPATCH_DEBUG_ONLY_TASK_IDS
        )
        tasks_df = tasks_df.loc[_tid_ok].copy()
        logging.info(
            "デバッグ限定配台: タスク入力 %s 行 → フィルタ後 %s 行",
            _n_tasks_before,
            len(tasks_df),
        )
        if tasks_df.empty:
            logging.error(
                "デバッグ限定配台: 指定依頼NOに該当するタスク行がありません（%s）。段階2を中断します。",
                ", ".join(sorted(DISPATCH_DEBUG_ONLY_TASK_IDS)),
            )
            _try_write_main_sheet_gemini_usage_summary("段階2")
            return

    try:
        validate_no_duplicate_explicit_plan_priorities(tasks_df)
    except PlanningValidationError as e:
        logging.error("%s", e)
        raise

    if global_priority_raw.strip():
        snip = global_priority_raw[:2500]
        if len(global_priority_raw) > 2500:
            snip += "…"
        ai_log_data["メイン_再優先特別記載(原文)"] = snip
    else:
        ai_log_data["メイン_再優先特別記載(原文)"] = (
            "（空、またはメインシートに「グローバルコメント」見出しが見つかりません）"
        )
    ai_log_data["メイン_再優先特別記載(AI)"] = json.dumps(
        global_priority_override, ensure_ascii=False
    )
    if global_priority_override.get("ignore_skill_requirements"):
        logging.warning(
            "メイン再優先特記: スキル要件を無視して配台します。%s",
            global_priority_override.get("interpretation_ja", ""),
        )
    if global_priority_override.get("ignore_need_minimum"):
        logging.warning(
            "メイン再優先特記: チーム人数を1名に固定します（need・行の必要OP上書きより優先）。%s",
            global_priority_override.get("interpretation_ja", ""),
        )
    if global_priority_override.get("abolish_all_scheduling_limits"):
        logging.warning(
            "メイン再優先特記: 設備専有・原反同日開始・指定開始時刻・マクロ実行時刻下限を適用しません。%s",
            global_priority_override.get("interpretation_ja", ""),
        )

    _try_write_plan_sheet_global_comment_parse_block(
        global_priority_override,
        data_extract_dt_str,
    )

    # 「当日」判定と最早開始時刻には基準日時（データ抽出日）を使う
    macro_now_dt = base_now_dt
    macro_run_date = macro_now_dt.date()
    ai_task_by_tid = analyze_task_special_remarks(
        tasks_df, reference_year=run_date.year, ai_sheet_sink=ai_log_data
    )
    task_queue = build_task_queue_from_planning_df(
        tasks_df,
        run_date,
        req_map,
        ai_task_by_tid,
        global_priority_override,
        equipment_list,
    )
    # 開始日が非稼働日の場合は、直前の稼働日へ補正（例: 4/4, 4/5 が非稼働なら 4/3 へ）
    working_days = [
        d for d in sorted_dates
        if any(attendance_data[d][m]["is_working"] for m in attendance_data[d])
    ]
    if working_days:
        for t in task_queue:
            req_d = t.get("start_date_req")
            if not isinstance(req_d, date):
                continue
            if req_d in working_days:
                continue
            prev_work = None
            for wd in working_days:
                if wd <= req_d:
                    prev_work = wd
                else:
                    break
            if prev_work is not None:
                if str(t.get("task_id", "")).strip() == DEBUG_TASK_ID:
                    logging.info(
                        "DEBUG[task=%s] start_date_req を非稼働日補正: %s -> %s",
                        DEBUG_TASK_ID,
                        req_d,
                        prev_work,
                    )
                t["start_date_req"] = prev_work
    conflict_rows = collect_planning_conflicts_by_excel_row(tasks_df, ai_task_by_tid)
    try:
        apply_planning_sheet_conflict_styles(
            TASKS_INPUT_WORKBOOK,
            PLAN_INPUT_SHEET_NAME,
            len(tasks_df),
            conflict_rows,
        )
    except Exception as e:
        logging.warning(f"配台シート矛盾ハイライト適用をスキップ: {e}")

    if not task_queue:
        logging.warning(
            f"有効なタスクがありません。「{PLAN_INPUT_SHEET_NAME}」の「依頼NO」「工程名」「換算数量」、"
            "または完了区分・実出来高換算により残量が無い行のみの可能性があります。"
        )

    # §B-1（加工途中の熱融着検査＋試行順）→ §B-2（同パイプライン・未着手）→ その他加工途中 → 一般キー
    task_queue.sort(
        key=lambda x: _generate_plan_task_queue_sort_key(x, req_map, need_rules)
    )
    _reorder_task_queue_b2_ec_inspection_consecutive(task_queue)
    # 配台試行順: 日次ループ前・ソート済みキューの並び（初回割当の成否とは無関係）
    for _trial_ord, _tq in enumerate(task_queue, start=1):
        _tq["dispatch_trial_order"] = _trial_ord
    if DEBUG_TASK_ID:
        dbg_items = [t for t in task_queue if str(t.get("task_id", "")).strip() == DEBUG_TASK_ID]
        if dbg_items:
            t0 = dbg_items[0]
            logging.info(
                "DEBUG[task=%s] queue基準: start_date_req=%s due_basis=%s answer_due=%s specified_due=%s specified_due_ov=%s due_source=%s priority=%s in_progress=%s remark=%s",
                DEBUG_TASK_ID,
                t0.get("start_date_req"),
                t0.get("due_basis_date"),
                t0.get("answer_due_date"),
                t0.get("specified_due_date"),
                t0.get("specified_due_override"),
                t0.get("due_source"),
                t0.get("priority"),
                t0.get("in_progress"),
                t0.get("has_special_remark"),
            )
        else:
            logging.info("DEBUG[task=%s] task_queueに存在しません（完了/残量0/依頼NO不一致の可能性）。", DEBUG_TASK_ID)
    timeline_events = []

    # ---------------------------------------------------------
    # 日毎のスケジューリングループ
    # STAGE2_EXTEND_ATTENDANCE_CALENDAR が True のときのみ、残タスクがあれば勤怠を日付複製で拡張。
    # 計画基準納期を過ぎても残がある依頼のみ due_basis +1 し当該依頼の割当を戻して先頭から再実行（STAGE2_RETRY_*）。
    # 各再試行前に勤怠の自動拡張分はマスタ日付へ巻き戻す。
    # ---------------------------------------------------------
    _master_attendance_date_set = frozenset(attendance_data.keys())
    _master_plan_dates_template = list(sorted_dates)
    _calendar_last_plan_day = _master_plan_dates_template[-1]

    for t in task_queue:
        t["remaining_units"] = float(t.get("initial_remaining_units") or 0)
        t["assigned_history"].clear()
    timeline_events.clear()

    if STAGE2_SERIAL_DISPATCH_BY_TASK_ID:
        logging.info(
            "依頼NO直列配台: 有効（STAGE2_SERIAL_DISPATCH_BY_TASK_ID）。"
            " 各日はアクティブな依頼NOの行だけが候補のため、当該依頼が詰まると他依頼は一切進みません。"
        )
    else:
        logging.info(
            "依頼NO直列配台: 無効。start_date を満たす全行が当日候補になり、配台試行順・設備ルールで順序付けします。"
        )

    _due_shift_retry_count_by_request: dict[str, int] = {}
    _due_shift_exhausted_requests: set[str] = set()
    _due_shift_cap_warned_tids: set[str] = set()
    _outer_retry_round = 0
    while True:
        _need_headcount_logged_orders: set = set()
        if _outer_retry_round > 0:
            _purge_attendance_days_not_in_set(
                attendance_data, _master_attendance_date_set
            )
            sorted_dates[:] = list(_master_plan_dates_template)

        for t in task_queue:
            t.pop("_partial_retry_calendar_blocked", None)

        if STAGE2_SERIAL_DISPATCH_BY_TASK_ID:
            _serial_order_tids = _serial_dispatch_order_task_ids(task_queue)
        else:
            _serial_order_tids = []

        _plan_day_iter = (
            _iter_plan_dates_extending(sorted_dates, attendance_data, task_queue)
            if STAGE2_EXTEND_ATTENDANCE_CALENDAR
            else sorted_dates
        )
        _full_calendar_without_deadline_restart = True
        for current_date in _plan_day_iter:
            daily_status = attendance_data[current_date]
            # 設備ごとの空き時刻（同一設備の同時並行割当を防止）
            machine_avail_dt = {}
            
            avail_dt = {}
            for m in members:
                if m not in daily_status:
                    continue
                st = daily_status[m]
                if st.get("eligible_for_assignment", st.get("is_working", False)):
                    avail_dt[m] = st["start_dt"]

            _machine_day_start = datetime.combine(current_date, DEFAULT_START_TIME)
            if avail_dt:
                _seed_avail_from_timeline_for_date(
                    timeline_events,
                    current_date,
                    machine_avail_dt,
                    avail_dt,
                    _machine_day_start,
                )

            if not avail_dt:
                logging.info("DEBUG[day=%s] 稼働メンバー0のため割付スキップ", current_date)
                continue
    
            tasks_today = [t for t in task_queue if t['remaining_units'] > 0 and t['start_date_req'] <= current_date]
            if STAGE2_SERIAL_DISPATCH_BY_TASK_ID and _serial_order_tids:
                _tasks_today_before_serial = len(tasks_today)
                _active_serial_tid = None
                for _tid in _serial_order_tids:
                    if any(
                        float(x.get("remaining_units") or 0) > 1e-12
                        for x in task_queue
                        if str(x.get("task_id", "") or "").strip() == _tid
                    ):
                        _active_serial_tid = _tid
                        break
                if _active_serial_tid is not None:
                    tasks_today = [
                        t
                        for t in tasks_today
                        if str(t.get("task_id", "") or "").strip() == _active_serial_tid
                    ]
                _serial_pos = (
                    _serial_order_tids.index(_active_serial_tid) + 1
                    if _active_serial_tid in _serial_order_tids
                    else 0
                )
                _pending_rows = sum(1 for t in task_queue if t["remaining_units"] > 0)
                logging.info(
                    "依頼NO直列配台 day=%s アクティブ依頼NO=%s 直列リスト位置=%s/%s "
                    "当日候補行数(直列前)=%s 直列後=%s キュー残行(全日)=%s",
                    current_date,
                    _active_serial_tid if _active_serial_tid is not None else "—",
                    _serial_pos if _serial_pos else "—",
                    len(_serial_order_tids),
                    _tasks_today_before_serial,
                    len(tasks_today),
                    _pending_rows,
                )
            pending_total = sum(1 for t in task_queue if t["remaining_units"] > 0)
            if not tasks_today:
                earliest_wait = min(
                    [t["start_date_req"] for t in task_queue if t["remaining_units"] > 0],
                    default=None,
                )
                logging.info(
                    "DEBUG[day=%s] 割付対象タスク0件 pending_total=%s earliest_start_date_req=%s",
                    current_date,
                    pending_total,
                    earliest_wait,
                )
            elif DEBUG_TASK_ID:
                has_dbg_today = any(str(t.get("task_id", "")).strip() == DEBUG_TASK_ID for t in tasks_today)
                if current_date.isoformat() == "2026-04-03" or has_dbg_today:
                    logging.info(
                        "DEBUG[day=%s] avail_members=%s tasks_today=%s (task=%s 含む=%s)",
                        current_date,
                        len(avail_dt),
                        len(tasks_today),
                        DEBUG_TASK_ID,
                        has_dbg_today,
                    )
            
            _sched_max_passes = max(96, max(1, len(tasks_today)) * 15)
            _sched_pi = 0
            while _sched_pi < _sched_max_passes:
                _sched_pi += 1
                _sched_made_progress = False
                if STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST:
                    _sched_made_progress = _trial_order_first_schedule_pass(
                        current_date,
                        tasks_today,
                        task_queue,
                        daily_status,
                        machine_avail_dt,
                        avail_dt,
                        timeline_events,
                        skills_dict,
                        members,
                        req_map,
                        need_rules,
                        surplus_map,
                        global_priority_override,
                        macro_run_date,
                        macro_now_dt,
                        _need_headcount_logged_orders,
                        team_combo_presets,
                    )
                    if _dispatch_debug_should_stop_early():
                        _sched_made_progress = True
                        break
                if not STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST:
                    for task in sorted(
                        [t for t in tasks_today if float(t.get("remaining_units") or 0) > 1e-12],
                        key=lambda t: _day_schedule_task_sort_key(t, task_queue),
                    ):
                        if _task_blocked_by_same_request_dependency(task, task_queue):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] スキップ: 同一依頼NOの先行工程待ち day=%s machine=%s rem=%.4f",
                                    task.get("task_id"),
                                    current_date,
                                    task.get("machine"),
                                    float(task.get("remaining_units") or 0),
                                )
                            continue
                        if task.get("roll_pipeline_inspection") and (
                            _roll_pipeline_inspection_assign_room(
                                task_queue, str(task.get("task_id", "")).strip()
                            )
                            <= 1e-12
                        ):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _tid_tr = str(task.get("task_id", "") or "").strip()
                                _ec_d = _pipeline_ec_roll_done_units(task_queue, _tid_tr)
                                _in_d = _pipeline_inspection_roll_done_units(
                                    task_queue, _tid_tr
                                )
                                _log_dispatch_trace_schedule(
                                    _tid_tr,
                                    "[配台トレース task=%s] スキップ: §B-2 検査ロール枠ゼロ day=%s machine=%s "
                                    "ec累計完了R=%.4f insp累計完了R=%.4f rem_insp=%.4f",
                                    _tid_tr,
                                    current_date,
                                    task.get("machine"),
                                    _ec_d,
                                    _in_d,
                                    float(task.get("remaining_units") or 0),
                                )
                            continue
                        if PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE:
                            _b1_holder = _exclusive_b1_inspection_holder_for_machine(
                                task_queue,
                                str(
                                    task.get("equipment_line_key")
                                    or task.get("machine")
                                    or ""
                                ).strip(),
                            )
                            if _b1_holder is not None and _b1_holder is not task:
                                if _trace_schedule_task_enabled(task.get("task_id")):
                                    _log_dispatch_trace_schedule(
                                        task.get("task_id"),
                                        "[配台トレース task=%s] スキップ: 同一設備の検査占有中 day=%s "
                                        "占有者依頼NO=%s 占有者試行順=%s",
                                        task.get("task_id"),
                                        current_date,
                                        _b1_holder.get("task_id"),
                                        _b1_holder.get("dispatch_trial_order"),
                                    )
                                continue
                        if DEBUG_TASK_ID and str(task.get("task_id", "")).strip() == DEBUG_TASK_ID:
                            logging.info(
                                "DEBUG[task=%s] day=%s 開始判定: start_date_req=%s remaining_units=%s machine=%s",
                                DEBUG_TASK_ID,
                                current_date,
                                task.get("start_date_req"),
                                task.get("remaining_units"),
                                task.get("machine"),
                            )
                        if task.get("has_done_deadline_override"):
                            logging.info(
                                "DEBUG[完了日指定] 依頼NO=%s 日付=%s start_date_req=%s due_basis=%s 指定納期(上書き)=%s 進捗=%s/%s",
                                task.get("task_id"),
                                current_date,
                                task.get("start_date_req"),
                                task.get("due_basis_date"),
                                task.get("specified_due_override"),
                                task.get("done_qty_reported"),
                                task.get("total_qty_m"),
                            )
    
                        machine = task['machine']
                        eq_line = str(
                            task.get("equipment_line_key") or machine or ""
                        ).strip() or machine
                        try:
                            _my_dispatch_ord = int(
                                task.get("dispatch_trial_order") or 10**9
                            )
                        except (TypeError, ValueError):
                            _my_dispatch_ord = 10**9
                        if _equipment_line_lower_dispatch_trial_still_pending(
                            task_queue, eq_line, _my_dispatch_ord, current_date
                        ):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] スキップ: 同一設備で配台試行順が先の行が未完了 "
                                    "day=%s eq_line=%s my_order=%s",
                                    task.get("task_id"),
                                    current_date,
                                    eq_line,
                                    _my_dispatch_ord,
                                )
                            continue
                        machine_name = str(task.get("machine_name", "") or "").strip()
                        machine_proc = str(machine or "").strip()
                        plan_ro = _plan_sheet_required_op_optional(task)
                        need_src_line = ""
                        if TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY:
                            req_num, need_src_line = resolve_need_required_op_explain(
                                machine,
                                machine_name,
                                task["task_id"],
                                req_map,
                                need_rules,
                            )
                            if plan_ro is not None and plan_ro != req_num:
                                need_src_line = (
                                    (need_src_line + "；") if need_src_line else ""
                                )
                                need_src_line += (
                                    f"計画シート必要人数{plan_ro}は未使用（need基準={req_num}）"
                                )
                        else:
                            if plan_ro is not None:
                                req_num = plan_ro
                                need_src_line = f"計画シート「必要OP(上書)」={req_num}"
                            else:
                                req_num, need_src_line = resolve_need_required_op_explain(
                                    machine,
                                    machine_name,
                                    task["task_id"],
                                    req_map,
                                    need_rules,
                                )
                        if global_priority_override.get("ignore_need_minimum"):
                            req_num = 1
                            need_src_line = (
                                (need_src_line + " → ")
                                if need_src_line
                                else ""
                            ) + "メイン上書ignore_need_minimumでreq=1"
    
                        # メンバー×設備スキル（parse_op_as_skill_cell: 小さい優先度ほど先にチーム候補へ採用）
                        # skills 読込時に「機械名」単独キーへエイリアスするため、工程名+機械名が両方ある行では
                        # 複合キー「工程名+機械名」のみを見る（別工程の同名機械の OP が流れ込まないようにする）。
                        skill_meta_cache = {}
                        _gpo = global_priority_override
    
                        def skill_role_priority(mem):
                            if _gpo.get("ignore_skill_requirements"):
                                return ("OP", 100)
                            if mem not in skill_meta_cache:
                                srow = skills_dict.get(mem, {})
                                v = ""
                                if machine_proc and machine_name:
                                    v = srow.get(f"{machine_proc}+{machine_name}", "")
                                elif machine_name:
                                    v = srow.get(machine_name, "")
                                elif machine_proc:
                                    v = srow.get(machine_proc, "")
                                skill_meta_cache[mem] = parse_op_as_skill_cell(v)
                            return skill_meta_cache[mem]
    
                        capable_members = [m for m in avail_dt.keys() if skill_role_priority(m)[0] in ("OP", "AS")]
                        capable_members.sort(key=lambda mm: (skill_role_priority(mm)[1], mm))
                        capable_members = _filter_capable_members_b2_disjoint_teams(
                            task, task_queue, capable_members
                        )
                        if task.get("has_done_deadline_override"):
                            machine_free_dbg = machine_avail_dt.get(
                                eq_line, datetime.combine(current_date, DEFAULT_START_TIME)
                            )
                            logging.info(
                                "DEBUG[完了日指定] 依頼NO=%s 設備=%s req_num=%s capable_members=%s machine_free=%s",
                                task.get("task_id"),
                                eq_line,
                                req_num,
                                len(capable_members),
                                machine_free_dbg,
                            )
    
                        pref_raw = str(task.get("preferred_operator_raw") or "").strip()
                        op_today = [m for m in capable_members if skill_role_priority(m)[0] == "OP"]
                        pref_mem = (
                            _resolve_preferred_op_to_member(pref_raw, op_today, members)
                            if pref_raw
                            else None
                        )
                        if pref_raw and pref_mem is None and op_today:
                            logging.info(
                                "担当OP指名: 当日のOP候補に一致せず制約なし task=%s raw=%r",
                                task.get("task_id"),
                                pref_raw,
                            )
    
                        extra_max_sheet, extra_src_line = resolve_need_surplus_extra_max_explain(
                            machine,
                            machine_name,
                            task["task_id"],
                            surplus_map,
                            need_rules,
                        )
                        if TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
                            extra_max_sheet = 0
                            extra_src_line = (
                                (extra_src_line + " → ")
                                if extra_src_line
                                else ""
                            ) + "TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROWで0"
                        extra_max = (
                            extra_max_sheet
                            if TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                            else 0
                        )
                        if (
                            extra_max_sheet > 0
                            and not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                            and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
                        ):
                            extra_src_line = (
                                (extra_src_line + " → ")
                                if extra_src_line
                                else ""
                            ) + "メインは基本人数のみ（余力枠は全配台後に未割当×スキルで追記）"
                        max_team_size = min(req_num + extra_max, len(capable_members))
                        if max_team_size < req_num:
                            max_team_size = req_num
                        rq_base = max(1, int(req_num))
    
                        _dto_head = task.get("dispatch_trial_order")
                        if (
                            _dto_head is not None
                            and _dto_head not in _need_headcount_logged_orders
                        ):
                            _need_headcount_logged_orders.add(_dto_head)
                            logging.info(
                                "need人数(配台試行順初回) order=%s task=%s 工程/機械=%s/%s "
                                "req_num=%s [%s] extra_max=%s [%s] max_team候補=%s capable=%s人",
                                _dto_head,
                                task["task_id"],
                                machine,
                                machine_name,
                                req_num,
                                need_src_line,
                                extra_max,
                                extra_src_line,
                                max_team_size,
                                len(capable_members),
                            )
    
                        trace_assign = bool(TRACE_TEAM_ASSIGN_TASK_ID) and (
                            str(task.get("task_id", "")).strip() == TRACE_TEAM_ASSIGN_TASK_ID
                        )
                        if trace_assign:
                            logging.info(
                                "TRACE配台[%s] %s 工程/機械=%s / %s req_num=%s extra_max=%s → max_team=%s "
                                "capable(n=%s)=%s ignore_need1=%s ignore_skill=%s abolish=%s 担当OP指定=%r→%s",
                                task["task_id"],
                                current_date,
                                machine,
                                machine_name,
                                req_num,
                                extra_max,
                                max_team_size,
                                len(capable_members),
                                capable_members,
                                global_priority_override.get("ignore_need_minimum"),
                                global_priority_override.get("ignore_skill_requirements"),
                                global_priority_override.get("abolish_all_scheduling_limits"),
                                pref_raw,
                                pref_mem,
                            )
    
                        team_candidates: list[dict] = []
                        combo_key = (
                            f"{machine_proc}+{machine_name}"
                            if machine_proc and machine_name
                            else ""
                        )
                        preset_rows = (
                            (team_combo_presets or {}).get(combo_key)
                            if (team_combo_presets and combo_key)
                            else None
                        )
                        preset_matched = False
                        if preset_rows:
                            for _prio, sheet_rs, preset_team, combo_row_id in preset_rows:
                                pteam = tuple(preset_team)
                                bounds = _combo_preset_team_size_bounds(
                                    pteam, sheet_rs, max_team_size
                                )
                                if bounds is None:
                                    continue
                                if pref_mem is not None and pref_mem not in pteam:
                                    continue
                                if not all(m in capable_members for m in pteam):
                                    continue
                                if _append_legacy_dispatch_candidate_for_team(
                                    task,
                                    pteam,
                                    avail_dt,
                                    machine_avail_dt,
                                    daily_status,
                                    current_date,
                                    macro_run_date,
                                    macro_now_dt,
                                    skill_role_priority,
                                    eq_line,
                                    rq_base,
                                    extra_max,
                                    global_priority_override,
                                    team_candidates,
                                    combo_sheet_row_id=combo_row_id,
                                    combo_preset_team=pteam,
                                ):
                                    preset_matched = True
                                    break
    
                        for tsize in range(req_num, max_team_size + 1):
                            if preset_matched:
                                break
                            if (
                                pref_mem is not None
                                and pref_mem in capable_members
                                and skill_role_priority(pref_mem)[0] == "OP"
                            ):
                                others = [m for m in capable_members if m != pref_mem]
                                if tsize == 1:
                                    teams_iter = [(pref_mem,)]
                                elif len(others) >= tsize - 1:
                                    teams_iter = [
                                        tuple([pref_mem] + list(rest))
                                        for rest in itertools.combinations(others, tsize - 1)
                                    ]
                                else:
                                    logging.info(
                                        "担当OP指名: チーム人数を満たせないため指名を無視 task=%s size=%s raw=%r",
                                        task.get("task_id"),
                                        tsize,
                                        pref_raw,
                                    )
                                    teams_iter = itertools.combinations(capable_members, tsize)
                            else:
                                teams_iter = itertools.combinations(capable_members, tsize)
    
                            for team in teams_iter:
                                op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
                                if not op_list:
                                    continue
    
                                team_start = max(avail_dt[m] for m in team)
                                if not _gpo.get("abolish_all_scheduling_limits"):
                                    # 同一設備は1時点で1タスクのみ（設備空き時刻を反映）
                                    machine_free_dt = machine_avail_dt.get(
                                        eq_line, datetime.combine(current_date, DEFAULT_START_TIME)
                                    )
                                    if team_start < machine_free_dt:
                                        team_start = machine_free_dt
                                    # 原反投入日と同日の開始は 13:00 以降（試行順優先フローと一致）
                                    if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
                                        min_start_dt = datetime.combine(
                                            current_date, task["same_day_raw_start_limit"]
                                        )
                                        if team_start < min_start_dt:
                                            team_start = min_start_dt
                                    if current_date == task["start_date_req"] and task.get("earliest_start_time"):
                                        min_user_t = datetime.combine(
                                            current_date, task["earliest_start_time"]
                                        )
                                        if team_start < min_user_t:
                                            team_start = min_user_t
                                    # 当日は「マクロ実行した時刻」より前に開始できない
                                    if current_date == macro_run_date and team_start < macro_now_dt:
                                        team_start = macro_now_dt
                                team_end_limit = min(daily_status[m]['end_dt'] for m in team)
    
                                if team_start >= team_end_limit:
                                    continue
    
                                team_breaks = []
                                for m in team:
                                    team_breaks.extend(daily_status[m]['breaks_dt'])
                                team_breaks = merge_time_intervals(team_breaks)
    
                                def _refloor_legacy_inline(ts):
                                    ts = max(ts, max(avail_dt[m] for m in team))
                                    if not _gpo.get("abolish_all_scheduling_limits"):
                                        _mfd = machine_avail_dt.get(
                                            eq_line,
                                            datetime.combine(
                                                current_date, DEFAULT_START_TIME
                                            ),
                                        )
                                        if ts < _mfd:
                                            ts = _mfd
                                        if task.get(
                                            "same_day_raw_start_limit"
                                        ) and current_date == task["start_date_req"]:
                                            _msd = datetime.combine(
                                                current_date,
                                                task["same_day_raw_start_limit"],
                                            )
                                            if ts < _msd:
                                                ts = _msd
                                        if current_date == task[
                                            "start_date_req"
                                        ] and task.get("earliest_start_time"):
                                            _mut = datetime.combine(
                                                current_date,
                                                task["earliest_start_time"],
                                            )
                                            if ts < _mut:
                                                ts = _mut
                                        if (
                                            current_date == macro_run_date
                                            and ts < macro_now_dt
                                        ):
                                            ts = macro_now_dt
                                    return ts
    
                                _ts_adj = _defer_team_start_past_prebreak_and_end_of_day(
                                    task,
                                    tuple(team),
                                    team_start,
                                    team_end_limit,
                                    team_breaks,
                                    _refloor_legacy_inline,
                                )
                                if _ts_adj is None:
                                    continue
                                team_start = _ts_adj
                                if team_start >= team_end_limit:
                                    continue
    
                                avg_eff = sum(daily_status[m]['efficiency'] for m in team) / len(team)
                                if avg_eff <= 0:
                                    avg_eff = 0.01
                                t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
                                if t_eff <= 0:
                                    t_eff = 1.0
                                # 追加増員による短縮は最大でも SURPLUS_TEAM_MAX_SPEEDUP_RATIO 程度（線形）
                                eff_time_per_unit = (
                                    task["base_time_per_unit"]
                                    / avg_eff
                                    / t_eff
                                    * _surplus_team_time_factor(rq_base, len(team), extra_max)
                                )
    
                                _, avail_mins, _ = calculate_end_time(team_start, 9999, team_breaks, team_end_limit)
    
                                units_can_do = int(avail_mins / eff_time_per_unit)
                                if units_can_do == 0:
                                    continue
    
                                units_today = min(units_can_do, math.ceil(task['remaining_units']))
                                work_mins_needed = int(units_today * eff_time_per_unit)
                                actual_end_dt, _, _ = calculate_end_time(team_start, work_mins_needed, team_breaks, team_end_limit)
    
                                team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
                                team_candidates.append(
                                    {
                                        "team": team,
                                        "team_start": team_start,
                                        "actual_end_dt": actual_end_dt,
                                        "units_today": units_today,
                                        "team_breaks": team_breaks,
                                        "avg_eff": avg_eff,
                                        "prio_sum": team_prio_sum,
                                        "op_list": op_list,
                                        "eff_time_per_unit": eff_time_per_unit,
                                        "combo_sheet_row_id": None,
                                        "combo_preset_team": None,
                                    }
                                )
    
                        best_team = None
                        best_info = {
                            "start_dt": datetime.max,
                            "units_today": 0,
                            "prio_sum": 10**9,
                        }
                        t_min = (
                            min(c["team_start"] for c in team_candidates)
                            if team_candidates
                            else None
                        )
    
                        def _team_cand_key(c):
                            return _team_assignment_sort_tuple(
                                c["team"],
                                c["team_start"],
                                c["units_today"],
                                c["prio_sum"],
                                t_min,
                            )
    
                        if team_candidates:
                            best_c = min(team_candidates, key=_team_cand_key)
                            if pref_mem and pref_mem in best_c["op_list"]:
                                lead_op = pref_mem
                            else:
                                lead_op = min(
                                    best_c["op_list"],
                                    key=lambda mm: (skill_role_priority(mm)[1], mm),
                                )
                            best_team = best_c["team"]
                            best_info = {
                                "start_dt": best_c["team_start"],
                                "end_dt": best_c["actual_end_dt"],
                                "op": lead_op,
                                "units_today": best_c["units_today"],
                                "breaks": best_c["team_breaks"],
                                "eff": best_c["avg_eff"],
                                "prio_sum": best_c["prio_sum"],
                            }
    
                        if trace_assign:
                            _tk = _team_assign_trace_tuple_label()
                            tid = task["task_id"]
                            for tsize in range(req_num, max_team_size + 1):
                                sub = [c for c in team_candidates if len(c["team"]) == tsize]
                                if not sub:
                                    logging.info(
                                        "TRACE配台[%s] %s tsize=%s → この人数で成立するチームなし",
                                        tid,
                                        current_date,
                                        tsize,
                                    )
                                else:
                                    sm = min(sub, key=_team_cand_key)
                                    logging.info(
                                        "TRACE配台[%s] %s tsize=%s 人数内最良: members=%s "
                                        "start=%s units_today=%s prio_sum=%s eff_t/unit=%.6f "
                                        "比較ルール=%s ※全日最早開始=%s を基準に辞書式で小さい方が採用",
                                        tid,
                                        current_date,
                                        tsize,
                                        sm["team"],
                                        sm["team_start"],
                                        sm["units_today"],
                                        sm["prio_sum"],
                                        sm["eff_time_per_unit"],
                                        _tk,
                                        t_min.isoformat(sep=" ") if t_min else "—",
                                    )
    
                        if trace_assign and best_team is not None:
                            logging.info(
                                "TRACE配台[%s] %s ★採用 n=%s members=%s start=%s units_today=%s prio_sum=%s",
                                task["task_id"],
                                current_date,
                                len(best_team),
                                best_team,
                                best_info["start_dt"],
                                best_info["units_today"],
                                best_info["prio_sum"],
                            )
                            if len(best_team) == 1 and max_team_size > req_num:
                                if TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF:
                                    logging.info(
                                        "TRACE配台[%s] %s 1人採用（TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF）: "
                                        "より大きい人数で有効なチームなし（OP不足・0単位・開始>=終了等）。",
                                        task["task_id"],
                                        current_date,
                                    )
                                else:
                                    logging.info(
                                        "TRACE配台[%s] %s 1人採用: 人数を増やすと開始が遅れ、"
                                        "スラック外では開始優先で1人が選ばれた可能性。"
                                        "TEAM_ASSIGN_START_SLACK_WAIT_MINUTES=%s、または従来の人数最優先は環境変数参照。",
                                        task["task_id"],
                                        current_date,
                                        TEAM_ASSIGN_START_SLACK_WAIT_MINUTES,
                                    )
    
                        if best_team:
                            if len(best_team) > req_num:
                                logging.info(
                                    "配台採用人数>req_num task=%s day=%s order=%s 工程/機械=%s/%s "
                                    "採用=%s人 req_num=%s extra_max=%s max_team=%s [%s] [%s]",
                                    task["task_id"],
                                    current_date,
                                    task.get("dispatch_trial_order"),
                                    machine,
                                    machine_name,
                                    len(best_team),
                                    req_num,
                                    extra_max,
                                    max_team_size,
                                    need_src_line,
                                    extra_src_line,
                                )
                            sub_members = [m for m in best_team if m != best_info["op"]]
                            done_units = best_info["units_today"]
                            if task.get("roll_pipeline_inspection"):
                                _rp_room = _roll_pipeline_inspection_assign_room(
                                    task_queue, str(task.get("task_id", "")).strip()
                                )
                                done_units = min(
                                    int(done_units),
                                    int(min(_rp_room, math.ceil(task["remaining_units"]))),
                                )
                            else:
                                done_units = int(done_units)
                            if done_units <= 0:
                                if _trace_schedule_task_enabled(task.get("task_id")):
                                    _rp_log = None
                                    if task.get("roll_pipeline_inspection"):
                                        _rp_log = _roll_pipeline_inspection_assign_room(
                                            task_queue,
                                            str(task.get("task_id", "") or "").strip(),
                                        )
                                    _log_dispatch_trace_schedule(
                                        task.get("task_id"),
                                        "[配台トレース task=%s] スキップ: チーム採用後の実効ユニット0 "
                                        "day=%s machine=%s best_units_today=%s rp_room=%s rem=%.4f",
                                        task.get("task_id"),
                                        current_date,
                                        machine,
                                        best_info.get("units_today"),
                                        _rp_log,
                                        float(task.get("remaining_units") or 0),
                                    )
                                continue
                            if done_units < best_info["units_today"]:
                                team_end_limit = min(
                                    daily_status[m]["end_dt"] for m in best_team
                                )
                                _teff = parse_float_safe(task.get("task_eff_factor"), 1.0)
                                if _teff <= 0:
                                    _teff = 1.0
                                _eff_t = (
                                    task["base_time_per_unit"]
                                    / best_info["eff"]
                                    / _teff
                                    * _surplus_team_time_factor(rq_base, len(best_team), extra_max)
                                )
                                _wm = int(done_units * _eff_t)
                                _end_dt, _, _ = calculate_end_time(
                                    best_info["start_dt"],
                                    _wm,
                                    best_info["breaks"],
                                    team_end_limit,
                                )
                                best_info = dict(best_info)
                                best_info["end_dt"] = _end_dt
                                best_info["units_today"] = done_units
    
                            total_u = math.ceil(task['total_qty_m'] / task['unit_m']) if task['unit_m'] else 0
                            rem_u_before = math.ceil(task['remaining_units'])
                            already_done = total_u - rem_u_before
                            
                            # 「マクロ実行時点」の完了率（予定の進捗ではなく、実加工数ベース）
                            try:
                                tot_qty = parse_float_safe(task.get('total_qty_m'), 0.0)
                                done_qty = parse_float_safe(task.get('done_qty_reported'), 0.0)
                                if tot_qty > 0:
                                    pct_macro = max(0, min(100, int(round((done_qty / tot_qty) * 100))))
                                else:
                                    pct_macro = 0
                            except Exception:
                                pct_macro = 0
                            
                            _te_disp = parse_float_safe(task.get("task_eff_factor"), 1.0)
                            if _te_disp <= 0:
                                _te_disp = 1.0
                            timeline_events.append({
                                "date": current_date, "task_id": task['task_id'], "machine": eq_line,
                                "op": best_info["op"], "sub": ", ".join(sub_members),
                                "start_dt": best_info["start_dt"], "end_dt": best_info["end_dt"],
                                "breaks": best_info["breaks"], "units_done": done_units,
                                "already_done_units": already_done,
                                "total_units": total_u,
                                "pct_macro": pct_macro,
                                "eff_time_per_unit": task["base_time_per_unit"]
                                / best_info["eff"]
                                / _te_disp
                                * _surplus_team_time_factor(rq_base, len(best_team), extra_max),
                                "unit_m": task['unit_m']
                            })
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _rp_tr = None
                                if task.get("roll_pipeline_inspection"):
                                    _rp_tr = _roll_pipeline_inspection_assign_room(
                                        task_queue,
                                        str(task.get("task_id", "") or "").strip(),
                                    )
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] タイムライン追記 chunk day=%s machine=%s "
                                    "done_units=%s already_done=%s total_u=%s rem_after=%.4f "
                                    "start=%s end=%s eff_t/unit=%.4f rp_room(当時)=%s",
                                    task.get("task_id"),
                                    current_date,
                                    eq_line,
                                    done_units,
                                    already_done,
                                    total_u,
                                    float(task.get("remaining_units") or 0)
                                    - float(done_units),
                                    best_info["start_dt"],
                                    best_info["end_dt"],
                                    float(
                                        task["base_time_per_unit"]
                                        / best_info["eff"]
                                        / _te_disp
                                        * _surplus_team_time_factor(
                                            rq_base, len(best_team), extra_max
                                        )
                                    ),
                                    _rp_tr,
                                )

                            task['remaining_units'] -= done_units
                            op_main = (best_info.get("op") or "").strip()
                            subs_part = ",".join(
                                s.strip() for s in sub_members if s and str(s).strip()
                            )
                            team_s = f"{op_main}, {subs_part}" if subs_part else op_main
                            need_surplus_assigned = (
                                TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                                and extra_max > 0
                                and len(best_team) > req_num
                            )
                            _lo = (best_info.get("op") or "").strip()
                            _subs_legacy = [
                                str(s).strip()
                                for s in sub_members
                                if s and str(s).strip()
                            ]
                            _names_ord = ([] if not _lo else [_lo]) + _subs_legacy
                            _surplus_names = (
                                _names_ord[int(req_num) :]
                                if need_surplus_assigned
                                and len(_names_ord) > int(req_num)
                                else []
                            )
                            task["assigned_history"].append(
                                {
                                    "date": current_date.strftime("%m/%d"),
                                    "team": team_s,
                                    "done_m": int(done_units * task["unit_m"]),
                                    "start_dt": best_info["start_dt"],
                                    "end_dt": best_info["end_dt"],
                                    "need_surplus_assigned": need_surplus_assigned,
                                    "combo_sheet_row_id": best_c.get(
                                        "combo_sheet_row_id"
                                    ),
                                    "surplus_member_names": _surplus_names,
                                }
                            )

                            for m in best_team:
                                avail_dt[m] = best_info["end_dt"]
                            if not _gpo.get("abolish_all_scheduling_limits"):
                                machine_avail_dt[eq_line] = best_info["end_dt"]
                            rem_after = int(math.ceil(float(task.get("remaining_units") or 0)))
                            _dispatch_roll_trace_after_roll(
                                current_date,
                                task,
                                eq_line,
                                best_info["start_dt"],
                                best_info["end_dt"],
                                float(done_units),
                                rem_after,
                                str(best_info.get("op") or ""),
                                list(sub_members),
                                roll_surplus_meta={
                                    "surplus_phase": "main",
                                    "req_num": int(req_num),
                                    "team_size": len(best_team),
                                    "extra_max_main_pass": int(extra_max),
                                    "need_surplus_assigned": need_surplus_assigned,
                                    "team_summary": team_s,
                                },
                            )
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] ロール確定 メイン day=%s machine=%s machine_name=%s "
                                    "start=%s end=%s 採用人数=%s req_num=%s メイン探索extra_max=%s "
                                    "余剰人数適用(メイン)=%s team=%s",
                                    task.get("task_id"),
                                    current_date,
                                    eq_line,
                                    str(machine_name or "").strip(),
                                    best_info["start_dt"],
                                    best_info["end_dt"],
                                    len(best_team),
                                    int(req_num),
                                    int(extra_max),
                                    need_surplus_assigned,
                                    team_s,
                                )
                            _sched_made_progress = True
                            if _dispatch_debug_should_stop_early():
                                break
                        else:
                            if task.get("has_done_deadline_override"):
                                logging.info(
                                    "DEBUG[完了日指定] 依頼NO=%s 日付=%s は割当不可（要員/設備空き条件でチーム不成立）。remaining_units=%s",
                                    task.get("task_id"),
                                    current_date,
                                    task.get("remaining_units"),
                                )
    
                if _dispatch_debug_should_stop_early():
                    break
                if not _sched_made_progress:
                    break

            if TRACE_SCHEDULE_TASK_IDS:
                for _tt in TRACE_SCHEDULE_TASK_IDS:
                    for _t in task_queue:
                        if str(_t.get("task_id", "")).strip() != _tt:
                            continue
                        _rem_tr = float(_t.get("remaining_units") or 0)
                        if _rem_tr <= 1e-9:
                            continue
                        _log_dispatch_trace_schedule(
                            _tt,
                            "[配台トレース task=%s] 日次終了時点の残 day=%s machine=%s "
                            "machine_name=%s rem=%.4f roll_insp=%s 試行順=%s",
                            _tt,
                            current_date,
                            _t.get("machine"),
                            _t.get("machine_name"),
                            _rem_tr,
                            bool(_t.get("roll_pipeline_inspection")),
                            _t.get("dispatch_trial_order"),
                        )

            if STAGE2_RETRY_SHIFT_DUE_ON_PARTIAL_REMAINING:
                missed_tids = _collect_task_ids_missed_deadline_after_day(
                    task_queue, current_date
                )
                if missed_tids:
                    blocked_tids = set()
                    shift_tid_list = []
                    for _ptid in sorted(missed_tids):
                        _do_shift, _cal_short = _partial_task_id_due_shift_outcome(
                            task_queue, _ptid, _calendar_last_plan_day
                        )
                        if _cal_short:
                            blocked_tids.add(_ptid)
                        if _do_shift:
                            shift_tid_list.append(_ptid)
                    for t in task_queue:
                        _tid = str(t.get("task_id", "") or "").strip()
                        if _tid in blocked_tids:
                            t["_partial_retry_calendar_blocked"] = True
                    if shift_tid_list:
                        allowed_shift_tids = [
                            tid
                            for tid in shift_tid_list
                            if _due_shift_retry_count_by_request.get(tid, 0)
                            < STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS
                        ]
                        for tid in shift_tid_list:
                            if tid not in allowed_shift_tids:
                                _due_shift_exhausted_requests.add(tid)
                        if allowed_shift_tids:
                            _outer_retry_round += 1
                            for tid in allowed_shift_tids:
                                _due_shift_retry_count_by_request[tid] = (
                                    _due_shift_retry_count_by_request.get(tid, 0) + 1
                                )
                            shift_set = set(allowed_shift_tids)
                            for t in task_queue:
                                if str(t.get("task_id", "") or "").strip() in shift_set:
                                    _shift_task_due_calendar_fields_one_day(t, run_date)
                            timeline_events[:] = [
                                e
                                for e in timeline_events
                                if _normalize_timeline_task_id(e) not in shift_set
                            ]
                            for t in task_queue:
                                if str(t.get("task_id", "") or "").strip() in shift_set:
                                    t["remaining_units"] = float(
                                        t.get("initial_remaining_units") or 0
                                    )
                                    t["assigned_history"].clear()
                            task_queue.sort(
                                key=lambda x: _generate_plan_task_queue_sort_key(
                                    x, req_map, need_rules
                                )
                            )
                            _reorder_task_queue_b2_ec_inspection_consecutive(task_queue)
                            _trials_detail = ",".join(
                                f"{tid}:{_due_shift_retry_count_by_request[tid]}"
                                for tid in sorted(allowed_shift_tids)
                            )
                            logging.info(
                                "納期超過リトライ: 計画基準+1日して当該依頼のみ再配台（検出日=%s 依頼NO=%s 当該依頼の累計試行=%s）",
                                current_date.isoformat(),
                                ",".join(sorted(allowed_shift_tids)),
                                _trials_detail,
                            )
                            _full_calendar_without_deadline_restart = False
                            break
                        else:
                            # 依頼ごと上限でシフトできないだけのときは日付ループを継続する（break すると未処理日が残り配台不可が大量発生する）。
                            _cap_tids = sorted(
                                tid
                                for tid in shift_tid_list
                                if tid not in allowed_shift_tids
                            )
                            _first_cap_warn = [
                                tid for tid in _cap_tids if tid not in _due_shift_cap_warned_tids
                            ]
                            for tid in _first_cap_warn:
                                _due_shift_cap_warned_tids.add(tid)
                            if _first_cap_warn:
                                logging.warning(
                                    "納期後ろ倒し再配台: 次の依頼NOは依頼ごとの上限（各 %s 回）のためこの検出では +1 しません。"
                                    " カレンダーは継続します（未完了は終了時に納期見直し必要を付け得ます）: %s",
                                    STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS,
                                    ",".join(_cap_tids),
                                )

        if _dispatch_debug_should_stop_early():
            break
        if _full_calendar_without_deadline_restart:
            break

    if TRACE_SCHEDULE_TASK_IDS:
        for _tt in TRACE_SCHEDULE_TASK_IDS:
            for _t in task_queue:
                if str(_t.get("task_id", "")).strip() != _tt:
                    continue
                _log_dispatch_trace_schedule(
                    _tt,
                    "[配台トレース task=%s] シミュレーション終了時 machine=%s machine_name=%s "
                    "rem=%.4f initial=%.4f roll_insp=%s",
                    _tt,
                    _t.get("machine"),
                    _t.get("machine_name"),
                    float(_t.get("remaining_units") or 0),
                    float(_t.get("initial_remaining_units") or 0),
                    bool(_t.get("roll_pipeline_inspection")),
                )
            _evs_tr = sorted(
                (
                    e
                    for e in timeline_events
                    if str(e.get("task_id", "")).strip() == _tt
                ),
                key=lambda e: (e.get("date"), e.get("start_dt") or datetime.min),
            )
            _last_ev_by_machine: dict = {}
            for _e in _evs_tr:
                _last_ev_by_machine[str(_e.get("machine") or "")] = _e
            for _mk, _ev in sorted(_last_ev_by_machine.items()):
                _ad = int(_ev.get("already_done_units") or 0)
                _ud = int(_ev.get("units_done") or 0)
                _log_dispatch_trace_schedule(
                    _tt,
                    "[配台トレース task=%s] タイムライン最終塊(工程列ごと) machine=%s "
                    "already_done+units_done=%s+%s=%s total_units=%s end_dt=%s",
                    _tt,
                    _mk,
                    _ad,
                    _ud,
                    _ad + _ud,
                    _ev.get("total_units"),
                    _ev.get("end_dt"),
                )

    # need「配台時追加人数」: メイン割付後に、未参加×スキル適合者をサブへ追記（既定）
    if (
        not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
        and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
        and surplus_map
        and timeline_events
    ):
        _n_sur = append_surplus_staff_after_main_dispatch(
            timeline_events,
            attendance_data,
            skills_dict,
            members,
            task_queue,
            req_map,
            need_rules,
            surplus_map,
            global_priority_override,
        )
        if _n_sur:
            logging.info(
                "need余力: メイン割付完了後にサブ %s 名を追記（未割当×スキル・時間重なりなし）",
                _n_sur,
            )

    # タイムラインを日付別にインデックス化し、サブメンバー一覧を事前解析（以降の出力ループを高速化）
    for e in timeline_events:
        e["subs_list"] = [s.strip() for s in e["sub"].split(",")] if e.get("sub") else []
    timeline_for_eq_grid = _expand_timeline_events_for_equipment_grid(timeline_events)
    events_by_date = defaultdict(list)
    for e in timeline_for_eq_grid:
        events_by_date[e["date"]].append(e)

    # =========================================================
    # 4. Excel出力 (メイン計画)
    # =========================================================
    _remove_prior_stage2_workbooks_and_prune_empty_dirs(output_dir)
    output_filename = os.path.join(
        output_dir, f'production_plan_multi_day_{base_now_dt.strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    all_eq_rows = []
    # タスクID → 結果_設備毎の時間割で当該タスクが最初に現れるセル（例 B12）。結果_タスク一覧のリンク用。
    first_eq_schedule_cell_by_task_id: dict[str, str] = {}

    eq_empty_cols = {}
    for eq in equipment_list:
        eq_empty_cols[eq] = ""
        eq_empty_cols[f"{eq}進度"] = ""
    
    for d in sorted_dates:
        d_start = datetime.combine(d, DEFAULT_START_TIME)
        d_end = datetime.combine(d, DEFAULT_END_TIME)
        events_today = events_by_date[d]
        machine_to_events = defaultdict(list)
        for ev in events_today:
            machine_to_events[ev["machine"]].append(ev)
        for _eq_k, _evs in machine_to_events.items():
            _evs.sort(key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or "")))

        is_anyone_working = any(daily_status['is_working'] for daily_status in attendance_data[d].values())
        if not events_today and not is_anyone_working: continue
        
        all_eq_rows.append({"日時帯": f"■ {d.strftime('%Y/%m/%d (%a)')} ■", **eq_empty_cols})
        
        curr_grid = d_start
        while curr_grid < d_end:
            next_grid = curr_grid + timedelta(minutes=10)
            if next_grid > d_end: next_grid = d_end
            
            mid_t = curr_grid + (next_grid - curr_grid) / 2
            row_data = {"日時帯": f"{curr_grid.strftime('%H:%M')}-{next_grid.strftime('%H:%M')}"}
            
            for eq in equipment_list:
                eq_text = ""
                progress_text = ""
                active_ev = None
                for ev in machine_to_events.get(eq, ()):
                    if ev["start_dt"] <= mid_t < ev["end_dt"]:
                        active_ev = ev
                        break
                
                if active_ev:
                    if any(b_s <= mid_t < b_e for b_s, b_e in active_ev['breaks']):
                        eq_text = "休憩"
                    else:
                        elapsed = get_actual_work_minutes(active_ev['start_dt'], min(next_grid, active_ev['end_dt']), active_ev['breaks'])
                        block_done_now = min(int(elapsed / active_ev['eff_time_per_unit']), active_ev['units_done'])
                        
                        cumulative_done = active_ev['already_done_units'] + block_done_now
                        total_u = active_ev['total_units']
                        
                        sub_text = f" 補:{active_ev['sub']}" if active_ev['sub'] else ""
                        eq_text = f"[{active_ev['task_id']}] 主:{active_ev['op']}{sub_text}"
                        progress_text = f"{cumulative_done}/{total_u}R"
                        _tid_sched = str(active_ev.get("task_id") or "").strip()
                        if _tid_sched and _tid_sched not in first_eq_schedule_cell_by_task_id:
                            _row_ex = len(all_eq_rows) + 2
                            _ci = 2 + 2 * equipment_list.index(eq)
                            first_eq_schedule_cell_by_task_id[_tid_sched] = (
                                f"{get_column_letter(_ci)}{_row_ex}"
                            )

                row_data[eq] = eq_text
                row_data[f"{eq}進度"] = progress_text
            
            all_eq_rows.append(row_data)
            curr_grid = next_grid
        all_eq_rows.append({"日時帯": "", **eq_empty_cols})
        
    df_eq_schedule = pd.DataFrame(all_eq_rows)
    _eq_hdr = _equipment_schedule_header_labels(equipment_list)
    _eq_rename = {}
    for _eq, _lab in zip(equipment_list, _eq_hdr):
        if _eq in df_eq_schedule.columns:
            _eq_rename[_eq] = _lab
        _pqc = f"{_eq}進度"
        if _pqc in df_eq_schedule.columns:
            _eq_rename[_pqc] = f"{_lab}進度"
    if _eq_rename:
        df_eq_schedule = df_eq_schedule.rename(columns=_eq_rename)

    # 結果_タスク一覧用: シミュレーション上の当該タスクの最早開始・最遅終了（timeline_events 集約）
    plan_window_by_task_id: dict = {}
    for _ev in timeline_events:
        tid = _ev.get("task_id")
        if tid is None:
            continue
        sd = _ev.get("start_dt")
        ed = _ev.get("end_dt")
        if sd is None or ed is None:
            continue
        if tid not in plan_window_by_task_id:
            plan_window_by_task_id[tid] = [sd, ed]
        else:
            w = plan_window_by_task_id[tid]
            if sd < w[0]:
                w[0] = sd
            if ed > w[1]:
                w[1] = ed

    # 結果_タスク一覧の「回答納期」「指定納期」「原反投入日」は配台計画_タスク入力の当該行セルのみ（計画基準納期と混同しない）
    _result_sheet_answer_spec_by_line = {}
    _result_sheet_raw_input_by_line: dict = {}
    if tasks_df is not None and not getattr(tasks_df, "empty", True):
        for _, _r in tasks_df.iterrows():
            if _plan_row_exclude_from_assignment(_r):
                continue
            _tid = str(_planning_df_cell_scalar(_r, TASK_COL_TASK_ID) or "").strip()
            _mach = str(_planning_df_cell_scalar(_r, TASK_COL_MACHINE) or "").strip()
            if not _tid or not _mach:
                continue
            _ad = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_ANSWER_DUE))
            _sd = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_SPECIFIED_DUE))
            _rid = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_RAW_INPUT_DATE))
            _result_sheet_answer_spec_by_line[(_tid, _mach)] = (_ad, _sd)
            _result_sheet_raw_input_by_line[(_tid, _mach)] = _rid

    task_results = []
    max_history_len = max([len(t['assigned_history']) for t in task_queue] + [0])
    
    # ステータス（配台の可否・残）：完了相当=配台可／未割当=配台不可／一部のみ=配台残
    # 計画基準+1 の再試行が依頼NOごとの上限に達した依頼の未完了行には（納期見直し必要）を付与する。
    sorted_tasks_for_result = sorted(task_queue, key=_result_task_sheet_sort_key)
    for t in sorted_tasks_for_result:
        rem_u = float(t.get("remaining_units") or 0)
        hist = bool(t.get("assigned_history"))
        if rem_u <= 1e-9:
            status = "配台可"
        elif hist and t.get("_partial_retry_calendar_blocked"):
            status = "配台残(勤務カレンダー不足)"
        elif not hist and rem_u > 1e-9:
            status = "配台不可"
        else:
            status = "配台残"
        _tid_res = str(t.get("task_id", "") or "").strip()
        if (
            _tid_res in _due_shift_exhausted_requests
            and rem_u > 1e-9
            and "納期見直し必要" not in status
        ):
            status = f"{status}（納期見直し必要）"
        
        total_r = int(t['total_qty_m'] / t['unit_m']) if t['unit_m'] else 0
        rem_r = int(t['remaining_units'])
        
        _line_key = (str(t.get("task_id", "") or "").strip(), str(t.get("machine", "") or "").strip())
        _sheet_pair = _result_sheet_answer_spec_by_line.get(_line_key)
        if _sheet_pair is not None:
            _ans_d, _spec_d = _sheet_pair
            ans_s = _ans_d.strftime("%Y/%m/%d") if _ans_d else ""
            spec_s = _spec_d.strftime("%Y/%m/%d") if _spec_d else ""
        else:
            _ans_d = t.get("answer_due_date")
            _spec_d = t.get("specified_due_date")
            ans_s = _ans_d.strftime("%Y/%m/%d") if _ans_d else ""
            spec_s = _spec_d.strftime("%Y/%m/%d") if _spec_d else ""
        _basis_for_sheet = t.get("due_basis_date_result_sheet")
        if _basis_for_sheet is None:
            _basis_for_sheet = t.get("due_basis_date")
        basis_s = (
            _basis_for_sheet.strftime("%Y/%m/%d")
            if _basis_for_sheet is not None and hasattr(_basis_for_sheet, "strftime")
            else ""
        )
        if _line_key in _result_sheet_raw_input_by_line:
            _rid_d = _result_sheet_raw_input_by_line[_line_key]
            kenhan_s = _rid_d.strftime("%Y/%m/%d") if _rid_d else ""
        else:
            _rid_t = t.get("raw_input_date")
            kenhan_s = (
                _rid_t.strftime("%Y/%m/%d")
                if _rid_t is not None and hasattr(_rid_t, "strftime")
                else ""
            )
        start_req = t["start_date_req"]
        start_req_s = start_req.strftime("%Y/%m/%d") if hasattr(start_req, "strftime") else str(start_req)
        rov = t.get("required_op")
        # 列順: A=ステータス → タスクID/工程/機械/優先度 → 履歴1..n → その他 → 最後に特別指定_AI
        row_status = {"ステータス": status}
        _dto = t.get("dispatch_trial_order")
        row_core = {
            "タスクID": t['task_id'],
            "工程名": t['machine'],
            "機械名": t.get("machine_name", ""),
            "優先度": t.get("priority", 999),
            RESULT_TASK_COL_DISPATCH_TRIAL_ORDER: _dto if _dto is not None else "",
        }
        row_history = {}
        for i in range(max_history_len):
            if i < len(t['assigned_history']):
                h = t['assigned_history'][i]
                row_history[f"履歴{i+1}"] = _format_result_task_history_cell(t, h)
            else:
                row_history[f"履歴{i+1}"] = ""

        try:
            tot_qty = parse_float_safe(t.get("total_qty_m"), 0.0)
            done_qty = parse_float_safe(t.get("done_qty_reported"), 0.0)
            pct_macro = max(0, min(100, int(round((done_qty / tot_qty) * 100)))) if tot_qty > 0 else 0
        except Exception:
            pct_macro = 0

        _pw = plan_window_by_task_id.get(t["task_id"])
        if _pw:
            _ps, _pe = _pw[0], _pw[1]
            plan_assign_start_s = (
                _ps.strftime("%Y/%m/%d %H:%M") if hasattr(_ps, "strftime") else ""
            )
            plan_assign_end_s = (
                _pe.strftime("%Y/%m/%d %H:%M") if hasattr(_pe, "strftime") else ""
            )
        else:
            plan_assign_start_s = ""
            plan_assign_end_s = ""

        _plan_end_ans_spec16 = _result_task_plan_end_within_answer_or_spec_16_label(
            _pw, _ans_d, _spec_d
        )

        row_tail = {
            "必要OP(上書)": rov if rov is not None else "",
            "タスク効率": parse_float_safe(t.get("task_eff_factor"), 1.0),
            "加工途中": "はい" if t.get("in_progress") else "いいえ",
            "特別指定あり": "はい" if t.get("has_special_remark") else "いいえ",
            "担当OP指名": (t.get("preferred_operator_raw") or "")[:120],
            "回答納期": ans_s,
            "指定納期": spec_s,
            "計画基準納期": basis_s,
            TASK_COL_RAW_INPUT_DATE: kenhan_s,
            "納期緊急": "はい" if t.get("due_urgent") else "いいえ",
            "加工開始日": start_req_s,
            "配完_加工開始": plan_assign_start_s,
            "配完_加工終了": plan_assign_end_s,
            RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16: _plan_end_ans_spec16,
            "総加工量": f"{total_r}R ({t['total_qty_m']}m)",
            "残加工量": f"{rem_r}R ({int(t['remaining_units'] * t['unit_m'])}m)",
            "完了率(実行時点)": f"{pct_macro}%",
        }
        row_ai_last = {"特別指定_AI": (t.get("task_special_ai_note") or "")[:300]}
        row_data = {**row_status, **row_core, **row_history, **row_tail, **row_ai_last}
        task_results.append(row_data)
        
    cal_rows = []
    for d in sorted_dates:
        for m in members:
            if m in attendance_data[d]:
                data = attendance_data[d][m]
                if data['is_working']:
                    cal_end = _calendar_display_clock_out_for_calendar_sheet(data, d)
                    end_disp = cal_end if cal_end is not None else data['end_dt']
                    clock_out_s = end_disp.strftime("%H:%M")
                else:
                    clock_out_s = "休"
                cal_rows.append({
                    "日付": d,
                    "メンバー": m,
                    "出勤": data['start_dt'].strftime("%H:%M") if data['is_working'] else "休",
                    "退勤": clock_out_s,
                    "効率": data['efficiency'],
                    "備考": data['reason'],
                })

    utilization_data = []
    for d in sorted_dates:
        row_data = {"年月日": d.strftime("%Y/%m/%d (%a)")}
        # その日のイベントからメンバー別作業分を一括集計（全メンバー×全イベントの二重ループを避ける）
        member_worked_mins = defaultdict(int)
        for ev in events_by_date[d]:
            mins = get_actual_work_minutes(ev["start_dt"], ev["end_dt"], ev["breaks"])
            member_worked_mins[ev["op"]] += mins
            for s in ev["subs_list"]:
                if s:
                    member_worked_mins[s] += mins
        for m in members:
            if m in attendance_data[d] and attendance_data[d][m]['is_working']:
                default_start = datetime.combine(d, DEFAULT_START_TIME)
                default_end = datetime.combine(d, DEFAULT_END_TIME)
                
                actual_start = attendance_data[d][m]['start_dt']
                actual_end = attendance_data[d][m]['end_dt']
                clip_start = max(actual_start, default_start)
                clip_end = min(actual_end, default_end)
                
                if clip_start >= clip_end:
                    total_avail_mins = 0
                else:
                    breaks_dt = attendance_data[d][m]['breaks_dt']
                    total_avail_mins = get_actual_work_minutes(clip_start, clip_end, breaks_dt)
                
                if total_avail_mins <= 0:
                    row_data[m] = "0.0%"
                    continue
                
                worked_mins = member_worked_mins.get(m, 0)
                ratio = (worked_mins / total_avail_mins) * 100
                row_data[m] = f"{ratio:.1f}% ({worked_mins}/{total_avail_mins}分)"
            else:
                row_data[m] = "休"
        utilization_data.append(row_data)
        
    df_utilization = pd.DataFrame(utilization_data)

    df_mprio_legend, df_mprio_tbl = build_member_assignment_priority_reference(
        skills_dict, members
    )
    if df_mprio_tbl.empty:
        df_mprio_tbl = pd.DataFrame(
            [
                {
                    "工程名": "",
                    "機械名": "",
                    "スキル列キー": "",
                    "優先順位": "",
                    "メンバー": "",
                    "ロール": "",
                    "優先度値_小さいほど先": "",
                    "skillsセル値": "",
                    "備考": "マスタ skills に「工程名+機械名」形式の列が見つからないか、データがありません。",
                }
            ]
        )

    _usage_txt = build_gemini_usage_summary_text()
    if _usage_txt:
        ai_log_data["Gemini_トークン・料金サマリ"] = _usage_txt[:50000]

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_eq_schedule.to_excel(
            writer, sheet_name=RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME, index=False
        )
        pd.DataFrame(cal_rows).to_excel(writer, sheet_name='結果_カレンダー(出勤簿)', index=False)
        df_utilization.to_excel(writer, sheet_name='結果_メンバー別作業割合', index=False)
        df_tasks = pd.DataFrame(task_results)
        df_tasks, task_column_order, _, vis_map = apply_result_task_sheet_column_order(
            df_tasks, max_history_len
        )
        seen_tc: set[str] = set()
        task_column_order_dedup: list = []
        vis_list_dedup: list = []
        for c in task_column_order:
            if c in seen_tc:
                continue
            seen_tc.add(c)
            task_column_order_dedup.append(c)
            vis_list_dedup.append(bool(vis_map.get(c, True)))
        pd.DataFrame(
            {
                "列名": task_column_order_dedup,
                "表示": vis_list_dedup,
            }
        ).to_excel(writer, sheet_name=COLUMN_CONFIG_SHEET_NAME, index=False)
        df_tasks.to_excel(writer, sheet_name=RESULT_TASK_SHEET_NAME, index=False)
        pd.DataFrame(list(ai_log_data.items()), columns=["項目", "内容"]).to_excel(writer, sheet_name='結果_AIログ', index=False)

        _mprio_sheet = RESULT_MEMBER_PRIORITY_SHEET_NAME
        df_mprio_legend.to_excel(writer, sheet_name=_mprio_sheet, index=False)
        _mprio_gap = len(df_mprio_legend) + 2
        df_mprio_tbl.to_excel(
            writer, sheet_name=_mprio_sheet, index=False, startrow=_mprio_gap
        )

        _write_results_equipment_gantt_sheet(
            writer,
            timeline_events,
            equipment_list,
            sorted_dates,
            attendance_data,
            data_extract_dt_str,
            base_now_dt,
        )

        for sheet_name, ws_out in writer.sheets.items():
            if sheet_name == RESULT_SHEET_GANTT_NAME:
                continue
            _apply_output_font_to_result_sheet(ws_out)

        ws_cfg = writer.sheets[COLUMN_CONFIG_SHEET_NAME]
        _add_column_config_sheet_helpers(ws_cfg, len(task_column_order_dedup))

        worksheet_tasks = writer.sheets[RESULT_TASK_SHEET_NAME]
        max_col = worksheet_tasks.max_column
        for row in worksheet_tasks.iter_rows(min_row=1, max_row=worksheet_tasks.max_row, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="top")

        _apply_result_task_sheet_column_visibility(
            worksheet_tasks, list(df_tasks.columns), vis_map
        )

        _apply_result_task_history_rich_text(worksheet_tasks, list(df_tasks.columns))
        _apply_result_task_date_columns_blue_font(worksheet_tasks, list(df_tasks.columns))

        # 未スケジュール行（配台不可・配台残）を目立たせる
        status_col_idx = None
        for col_idx, col_name in enumerate(df_tasks.columns, 1):
            if str(col_name) == "ステータス":
                status_col_idx = col_idx
                break
        if status_col_idx is not None:
            unscheduled_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for r in range(2, worksheet_tasks.max_row + 1):
                st_val = worksheet_tasks.cell(row=r, column=status_col_idx).value
                st = str(st_val).strip() if st_val is not None else ""
                if st in ("配台不可", "配台残"):
                    for c in range(1, max_col + 1):
                        worksheet_tasks.cell(row=r, column=c).fill = unscheduled_fill

        _apply_result_task_history_need_surplus_highlight(
            worksheet_tasks, list(df_tasks.columns), sorted_tasks_for_result
        )

        _apply_result_task_task_id_content_mismatch_highlight(
            worksheet_tasks, list(df_tasks.columns), sorted_tasks_for_result
        )
        _apply_result_task_id_hyperlinks_to_equipment_schedule(
            worksheet_tasks,
            list(df_tasks.columns),
            sorted_tasks_for_result,
            first_eq_schedule_cell_by_task_id,
            RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME,
        )

    try:
        _apply_excel_date_columns_date_only_display(
            output_filename, "結果_カレンダー(出勤簿)", frozenset({"日付"})
        )
    except Exception as e:
        logging.warning(f"結果_カレンダー(出勤簿)の日付列表示整形: {e}")

    logging.info(f"完了: '{output_filename}' を生成しました。")

    # =========================================================
    # 5. ★追加: メンバー毎の行動スケジュール (別ファイル) 出力
    # =========================================================
    member_output_filename = os.path.join(
        output_dir, f'member_schedule_{base_now_dt.strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    
    # 時間帯は全メンバー共通で1回だけ生成（メンバー数分の重複計算を避ける）
    time_labels = []
    time_grids = []
    curr_dt = datetime.combine(run_date, DEFAULT_START_TIME)
    end_dt_grid = datetime.combine(run_date, DEFAULT_END_TIME)
    while curr_dt < end_dt_grid:
        next_dt = curr_dt + timedelta(minutes=10)
        if next_dt > end_dt_grid:
            next_dt = end_dt_grid
        time_labels.append(f"{curr_dt.strftime('%H:%M')}-{next_dt.strftime('%H:%M')}")
        time_grids.append((curr_dt.time(), next_dt.time()))
        curr_dt = next_dt
    
    with pd.ExcelWriter(member_output_filename, engine='openpyxl') as member_writer:
        for m in members:
            # 各行の辞書を初期化
            m_schedule = {t_label: {"時間帯": t_label} for t_label in time_labels}
            
            # 各日付のスケジュールを列として埋めていく
            for d in sorted_dates:
                d_str = d.strftime("%m/%d (%a)")
                
                # 全日非勤務: 年休（カレンダー *）は『年休』、工場休日などは『休』
                if m not in attendance_data[d] or not attendance_data[d][m]['is_working']:
                    off_label = _member_schedule_full_day_off_label(
                        attendance_data[d].get(m) if m in attendance_data[d] else None
                    )
                    for t_label in time_labels:
                        m_schedule[t_label][d_str] = off_label
                    continue
                
                daily_info = attendance_data[d][m]
                d_start_dt = daily_info['start_dt']
                d_end_dt = daily_info['end_dt']
                breaks_dt = daily_info['breaks_dt']
                
                events_today = events_by_date[d]
                
                for i, (t_start, t_end) in enumerate(time_grids):
                    t_label = time_labels[i]
                    
                    # 判定用の中間時刻を計算
                    grid_start_dt = datetime.combine(d, t_start)
                    grid_end_dt = datetime.combine(d, t_end)
                    grid_mid_dt = grid_start_dt + (grid_end_dt - grid_start_dt) / 2
                    
                    text = ""
                    if grid_mid_dt < d_start_dt or grid_mid_dt >= d_end_dt:
                        text = _member_schedule_off_shift_label(
                            d, grid_mid_dt, d_start_dt, d_end_dt, daily_info.get("reason")
                        )
                    else:
                        br_txt = _member_schedule_break_cell_label(
                            grid_mid_dt, breaks_dt, d_end_dt, daily_info.get("reason")
                        )
                        if br_txt is not None:
                            text = br_txt
                    if text == "":
                        # 該当するタスクを探す（subs_list は事前解析済み）
                        active_ev = next((e for e in events_today if e['start_dt'] <= grid_mid_dt < e['end_dt'] and (e['op'] == m or m in e.get('subs_list', []))), None)
                        if active_ev:
                            role = "主" if active_ev['op'] == m else "補"
                            text = f"[{active_ev['task_id']}] {active_ev['machine']}({role})"
                        else:
                            text = "" # 何も割り当てられていない空き時間
                    
                    m_schedule[t_label][d_str] = text
                    
            # データフレーム化してシートに書き込み
            df_m = pd.DataFrame(list(m_schedule.values()))
            cols = ["時間帯"] + [d.strftime("%m/%d (%a)") for d in sorted_dates]
            df_m = df_m[[c for c in cols if c in df_m.columns]]
            df_m.to_excel(member_writer, sheet_name=m, index=False)
            
            # --- 既定フォント・罫線・見出し背景（列幅は VBA 取り込み時の AutoFit） ---
            worksheet = member_writer.sheets[m]
            _apply_output_font_to_result_sheet(worksheet)
            header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = header_fill
                
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    
    logging.info(f"完了: 個人別スケジュールを '{member_output_filename}' に出力しました。")
    _try_write_main_sheet_gemini_usage_summary("段階2")