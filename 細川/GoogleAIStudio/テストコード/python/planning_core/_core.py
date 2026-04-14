# -*- coding: utf-8 -*-
"""planning_core 実装本体（パッケージ内）。``import planning_core`` で bootstrap が先に実行されること。"""
import pandas as pd
from datetime import datetime, timedelta, time, date
from collections import Counter, defaultdict
import itertools
import functools
import csv
import json
import copy
import re

from dispatch_interval_mirror import DispatchIntervalMirror
import traceback
import base64
import hashlib
import unicodedata
import time as time_module
from google import genai
import logging
import calendar
import math
import os
import random
import fnmatch
import shutil
import sys
import threading
import ctypes
from contextlib import contextmanager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break

from .bootstrap import (
    PlanningValidationError,
    _clear_stage2_blocking_message_file,
    _remove_prior_stage2_workbooks_and_prune_empty_dirs,
    _try_remove_path_with_retries,
    _write_stage2_blocking_message,
    api_payment_dir,
    json_data_dir,
    log_dir,
    output_dir,
)

# region agent log
def _agent_debug_ndjson(
    hypothesis_id: str, location: str, message: str, data: dict | None = None
) -> None:
    """Debug NDJSON（セッション fc9417）。ワークスペース直下 debug-fc9417.log へ追記。"""
    try:
        _p = os.path.abspath(
            os.path.join(os.path.dirname(__file__), *([".."] * 5), "debug-fc9417.log")
        )
        _line = {
            "sessionId": "fc9417",
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data or {},
            "timestamp": int(time_module.time() * 1000),
        }
        with open(_p, "a", encoding="utf-8") as _f:
            _f.write(json.dumps(_line, ensure_ascii=False) + "\n")
    except Exception:
        pass


# endregion agent log

PLAN_DUE_DAY_COMPLETION_TIME = time(16, 0)

# AI 備考・配台不要ロジック D→E の TTL キャッシュ（旧 output/ から json/ へ移行）
_ai_remarks_cache_name = "ai_remarks_cache.json"
_ai_cache_legacy = os.path.join(output_dir, _ai_remarks_cache_name)
_ai_cache_new = os.path.join(json_data_dir, _ai_remarks_cache_name)
if os.path.isfile(_ai_cache_legacy) and not os.path.isfile(_ai_cache_new):
    try:
        shutil.move(_ai_cache_legacy, _ai_cache_new)
    except OSError:
        pass
ai_cache_path = _ai_cache_new
# 「設定_配台不要工程」シート作成・保存の追跡デバッグ（execution_log と併用）
exclude_rules_sheet_debug_log_path = os.path.join(log_dir, "exclude_rules_sheet_debug.txt")
# 保存失敗時に E 列（ロジック式）の値を退避し、次回 run_exclude_rules_sheet_maintenance で自動適用する（json フォルダ）
EXCLUDE_RULES_E_SIDECAR_FILENAME = "exclude_rules_e_column_pending.json"
# openpyxl 保存失敗時に VBA は E 列へ書き込むための UTF-8 TSV（Base64）。
EXCLUDE_RULES_E_VBA_TSV_FILENAME = "exclude_rules_e_column_vba.tsv"
# openpyxl 保存失敗時に VBA は A〜E を一括反映する UTF-8 TSV（行ごとに 5 セル分 Base64）。
EXCLUDE_RULES_MATRIX_VBA_FILENAME = "exclude_rules_matrix_vba.tsv"
# VBA はメイン P 列へ書き込むための UTF-8 テキスト（Excel 開いたまま 保存しない場合の不具合の回避）
GEMINI_USAGE_SUMMARY_FOR_MAIN_FILE = "gemini_usage_summary_for_main.txt"
# 全実行を通した Gemini 利用・推定料金の累計（API 応答ごとに更新。保存先は API_Payment フォルダ）
GEMINI_USAGE_CUMULATIVE_JSON_FILE = "gemini_usage_cumulative.json"
# 期間別ポケットをフラット化した CSV（Excel の折れ線・棒グラフ用）
GEMINI_USAGE_BUCKETS_CSV_FILE = "gemini_usage_buckets_for_chart.csv"
# メインシート・Gemini 日次推移（xlwings: Q〜R＝料金または呼出し、S〜T＝合計トークン）
GEMINI_USAGE_CHART_COL_DATE = 17  # Q
GEMINI_USAGE_CHART_COL_VALUE = 18  # R
GEMINI_USAGE_CHART_COL_TOK_DATE = 19  # S（グラフ用に日付を複製）
GEMINI_USAGE_CHART_COL_TOK_VALUE = 20  # T（total_tokens 相当）
GEMINI_USAGE_CHART_HEADER_ROW = 16
GEMINI_USAGE_CHART_ANCHOR_CELL = "T16"
GEMINI_USAGE_CHART_TOKENS_ANCHOR_CELL = "AA16"
GEMINI_USAGE_CHART_MAX_DAYS = 14
GEMINI_USAGE_CHART_CLEAR_ROWS = 36
# xlwings で貼る折れ線グラフ名（再実行時に削除してから作り直し）
GEMINI_USAGE_XLW_CHART_NAME = "_GeminiApiDailyTrend"
GEMINI_USAGE_XLW_CHART_TOKENS_NAME = "_GeminiApiDailyTokens"
# テスト: EXCLUDE_RULES_TEST_E1234=1 で EXCLUDE_RULES_SHEET_NAME（「設定_配台不要工程」）の E 列に "1234" を書き（保存経路の確認用）。
# TASK_INPUT_WORKBOOK は「加工計画DATA」シート付しブック（例: 生産管理_AI配台テスト.xlsm）を指定すること。
# 行は EXCLUDE_RULES_TEST_E1234_ROW（既定 9、2 未満は 9 に丸める）。

# =========================================================
# 【設定】APIキー / 基本ルール / ファイル名
# =========================================================
# Gemini API キーは TASK_INPUT_WORKBOOK 確定後、下記「設定」B1 の JSON から解決（平文または format_version 2 の暗号化）。
# 未設定時のみ移行用に環境変数 GEMINI_API_KEY を参照。

# Gemini API のモデルコード（Google AI for Developers のモデルページの Model code に準拠）
# https://ai.google.dev/gemini-api/docs/models
# 既定の試行順（精度の高い順）。マクロブック「設定」シート D/E で有効行があるときはそちらを優先。
# 利用不可・同一モデルの試行上限消化後は _gemini_generate_content_with_retry が次点へ進む。
GEMINI_MODEL_IDS_BY_QUALITY: tuple[str, ...] = (
    "gemini-3-flash-preview",
    "gemini-2.5-pro",
    "gemini-2.5-flash",
    "gemini-3.1-flash-lite-preview",
    "gemini-2.5-flash-lite",
)
# 既定の先頭モデル（従来名 GEMINI_MODEL_FLASH のまま参照している箇所向け）
GEMINI_MODEL_FLASH = GEMINI_MODEL_IDS_BY_QUALITY[0]
# 推定料金: USD / 1M tokens（入力, 出力）。公式の最新単価に合わせて更新すること。
# 環境変数 GEMINI_PRICE_USD_IN_PER_M / GEMINI_PRICE_USD_OUT_PER_M で上書き可（Flash 向け）。
_GEMINI_FLASH_IN_PER_M = float(
    os.environ.get("GEMINI_PRICE_USD_IN_PER_M", "0.075") or 0.075
)
_GEMINI_FLASH_OUT_PER_M = float(
    os.environ.get("GEMINI_PRICE_USD_OUT_PER_M", "0.30") or 0.30
)
GEMINI_JPY_PER_USD = float(os.environ.get("GEMINI_JPY_PER_USD", "150") or 150)
# 503 / UNAVAILABLE / 429 等: 呼び出し直前の一様乱数ジッター（秒・上限）。0 で無効。
_GEMINI_PRE_REQUEST_JITTER_MAX = float(
    os.environ.get("GEMINI_PRE_REQUEST_JITTER_MAX_SEC", "0.75") or 0.75
)
# 再試行の指数バックオフ基底（秒）。試行 k 目の待ちの目安: base * 2^k + 小ジッター
_GEMINI_RETRY_BACKOFF_BASE = float(
    os.environ.get("GEMINI_RETRY_BACKOFF_BASE_SEC", "2.0") or 2.0
)
_GEMINI_RETRY_MAX_ATTEMPTS = max(
    1, int(os.environ.get("GEMINI_RETRY_MAX_ATTEMPTS", "3") or 3)
)
# generate_content 1 リクエストの HTTP タイムアウト（秒）。0 で HttpOptions の timeout を付けず SDK 既定。
# 環境変数 GEMINI_REQUEST_TIMEOUT_SEC（未設定時は 60）。


def _gemini_request_timeout_sec() -> float:
    raw = (os.environ.get("GEMINI_REQUEST_TIMEOUT_SEC") or "").strip()
    if not raw:
        return 60.0
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return 60.0
    return max(0.0, v)


def _gemini_client(api_key: str) -> genai.Client:
    """API キー付き Client。可能なら HttpOptions で読み取りタイムアウトを付与する。"""
    sec = _gemini_request_timeout_sec()
    if sec > 0:
        try:
            from google.genai import types as genai_types

            ms = max(1000, int(round(sec * 1000.0)))
            return genai.Client(
                api_key=api_key,
                http_options=genai_types.HttpOptions(timeout=ms),
            )
        except Exception:
            logging.debug(
                "Gemini Client: HttpOptions によるタイムアウト設定に失敗したため、既定クライアントを使用します。",
                exc_info=True,
            )
    return genai.Client(api_key=api_key)


# ---------------------------------------------------------------------------
# 以降の定数ブロックは「Excel 列見出し」と 1:1 で対応させる。
# 列名を変える場合は VBA・マクロ付きシートと同時に直すこと。
# ---------------------------------------------------------------------------

MASTER_FILE = "master.xlsm"  # skills と attendance（tasks）を統合したファイル
# VBA「master_機械カレンダーを作成」シート（1 時間スロット占有を段階2の machine_avail_dt に反映）
SHEET_MACHINE_CALENDAR = "機械カレンダー"
# ``generate_plan`` 開始時に再設定。date -> 設備キー -> [ (start, end), ... ] 半開区間 [start, end)
_MACHINE_CALENDAR_BLOCKS_BY_DATE: dict[
    date, dict[str, list[tuple[datetime, datetime]]]
] = {}

# master.xlsm: 機械との日次始業準備（分）。依頼切替の準備・後始末は廃止（シートは未読込）。
SHEET_MACHINE_DAILY_STARTUP = "設定_機械_日次始業準備"
# ``generate_plan`` 開始時に再設定（シート無し・空は空辞書＝従来どおり）
_STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE: dict[str, int] = {}
# master メイン A15（定常開始）。日次始業準備を勤怠 forward ではなく [開始, 開始+N分) のタイムラインに載せる。
_STAGE2_REGULAR_SHIFT_START: time | None = None
# timeline_events の event_kind（省略時は加工とみなす）
TIMELINE_EVENT_MACHINING = "machining"
TIMELINE_EVENT_MACHINE_DAILY_STARTUP = "machine_daily_startup"
# VBA「master_組み合わせ表を更新」で作るシート（工程+機械キーとメンバー編集）
MASTER_SHEET_TEAM_COMBINATIONS = "組み合わせ表"
# メンバー別勤怠シート: master.xlsm では「休暇区分」と「備考」は別列。
# 勤怠AIの入力は備考のみ。reason（表示・中抜き補正・個人シートの休憩/休暇文言）は「備考が空のとき休暇区分を引き継ぐ」。
# master カレンダー＝出勤簿.txt 準拠: 公休=公休年休・休憩時間1_終了～定常終了（午後休憩14:45～15:00）＝後休=定常開始～休憩時間1_開始・午後年休＝国=他拠点勤務。
# 備考列・休暇区分は勤怠 AI で構造化（配台試行時・is_holiday・中抜き等）。備考は空でも休暇区分のみの行は AI に渡す。
ATT_COL_LEAVE_TYPE = "休暇区分"
ATT_COL_REMARK = "備考"
# メンバー勤怠シート（master.xlsm）: 定時の「退勤時間」と分けて退勤上限を指定（任意列・見出しは「残業(分)」）
ATT_COL_OT_END = "残業(分)"
# 旧ブックの K 列見出し。load_attendance_and_analyze で ATT_COL_OT_END に正規化する。
ATT_COL_OT_END_LEGACY = "残業終業"
# 勤怠備考 AI の JSON スキーマを変えたら更新し、キャッシュキーを無効化する
ATTENDANCE_REMARK_AI_SCHEMA_ID = "v2_haitai_fuka"
# need シート: 「基本必須人数」行（A列に「必須人数」を含む）＋ その直下の「配台時追加人数＝余力時追加人数」等
# （Excel 上は概ね 5 行目付近。余剰時に増やせる人数上限・工程×機械列）
# ＋ 行「特別指定1」～「特別指定99」（必須人数の上書き・1～99）
NEED_COL_CONDITION = "依頼NO条件"
NEED_COL_NOTE = "備考"
# need「配台時追加人数」を満枠使っても、短縮あたり加工時間は短くなるのは最大でこの割引（例: 0.05 ≒ 5%）
SURPLUS_TEAM_MAX_SPEEDUP_RATIO = 0.05
# タスクは tasks.xlsx を使うので、VBA から渡される TASK_INPUT_WORKBOOK の「加工計画DATA」のみ
TASKS_INPUT_WORKBOOK = os.environ.get("TASK_INPUT_WORKBOOK", "").strip()
TASKS_SHEET_NAME = "加工計画DATA"

# このシート名を含むブックは openpyxl は読み書きに失敗することがあるため、load_workbook を試行する
OPENPYXL_INCOMPATIBLE_SHEET_MARKER = "配台_配台不要工程"


def _ooxml_workbook_sheet_names(wb_path: str) -> list[str] | None:
    """ZIP 内 xl/workbook.xml からシート名一覧を得る（openpyxl を使えない）。"""
    import zipfile
    import xml.etree.ElementTree as ET

    if not wb_path or not os.path.isfile(wb_path):
        return None
    low = wb_path.lower()
    if not low.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return None
    try:
        with zipfile.ZipFile(wb_path, "r") as zf:
            if "xl/workbook.xml" not in zf.namelist():
                return None
            raw = zf.read("xl/workbook.xml")
    except (OSError, zipfile.BadZipFile, KeyError):
        return None
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return None
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    tag_sheet = "{%s}sheet" % ns_main
    names: list[str] = []
    for el in root.iter():
        if el.tag == tag_sheet or el.tag.endswith("}sheet"):
            n = el.get("name")
            if n:
                names.append(n)
    return names


def _workbook_should_skip_openpyxl_io(wb_path: str) -> bool:
    """当該パスは OOXML でシート「配台_配台不要工程」を含むとする True（openpyxl 利用を避ける）。"""
    p = (wb_path or "").strip()
    if not p:
        return False
    names = _ooxml_workbook_sheet_names(p)
    if not names:
        return False
    return OPENPYXL_INCOMPATIBLE_SHEET_MARKER in names


# マクロブック「設定」B1: 社内共有上の Gemini 証明書 JSON のパス
APP_CONFIG_SHEET_NAME = "設定"
# 「設定」シート D3 以降: Gemini 試行モデル ID（Google の model code）。E 列が真の行だけを上から順に試行する。
# （A/B 列は配台トレース・デバッグ依頼NO用のため、モデル一覧は D/E に配置）
GEMINI_MODEL_SHEET_COL_MODEL = 4  # D
GEMINI_MODEL_SHEET_COL_ENABLE = 5  # E
GEMINI_MODEL_SHEET_FIRST_ROW = 3
GEMINI_MODEL_SHEET_MAX_ROWS = 40
# 暗号化認証 JSON（format_version 2）の復号は常にこの定数のみ（社内手順のパスフレーズと一致させる。ログ・UI に出さない）。
_GEMINI_CREDENTIALS_PASSPHRASE_FIXED = "nagaoka1234"
_GEMINI_CREDENTIALS_PBKDF2_ITERATIONS_DEFAULT = 480_000


def _config_cell_text(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def _config_cell_truthy_enabled(v) -> bool:
    """設定シートの「試行に含める」列。Excel の TRUE / チェックボックス / 1 等を真とみなす。"""
    if v is True:
        return True
    if v is False:
        return False
    if isinstance(v, (int, float)) and not pd.isna(v):
        try:
            return int(round(float(v))) != 0
        except (TypeError, ValueError):
            return False
    s = _config_cell_text(v)
    if not s:
        return False
    u = s.lower()
    if u in (
        "1",
        "true",
        "yes",
        "y",
        "on",
        "○",
        "〇",
        "はい",
        "有効",
        "含む",
        "試行",
    ):
        return True
    if u in ("0", "false", "no", "n", "off", "×", "いいえ", "無効", "除外"):
        return False
    return False


def _resolve_path_relative_to_workbook(wb_path: str, user_path: str) -> str:
    p = (user_path or "").strip().strip('"')
    if not p:
        return ""
    if os.path.isabs(p):
        return os.path.normpath(p)
    base = os.path.dirname(os.path.abspath(wb_path))
    return os.path.normpath(os.path.join(base, p))


def _read_gemini_credentials_json_path_from_workbook(wb_path: str) -> str | None:
    """「設定」シート B1 から証明書 JSON ファイルパスを読む。無ければ None。"""
    if not wb_path or not os.path.isfile(wb_path):
        return None
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.debug(
            "Gemini: ブックに「%s」があるため、openpyxl で「%s」!B1 を読みません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            APP_CONFIG_SHEET_NAME,
        )
        return None
    try:
        keep_vba = str(wb_path).lower().endswith(".xlsm")
        wb = load_workbook(
            wb_path, read_only=True, data_only=True, keep_vba=keep_vba
        )
        try:
            if APP_CONFIG_SHEET_NAME not in wb.sheetnames:
                return None
            ws = wb[APP_CONFIG_SHEET_NAME]
            b1 = _config_cell_text(ws.cell(row=1, column=2).value)
        finally:
            wb.close()
    except Exception as ex:
        logging.debug(
            "Gemini: マクロブック「%s」の「%s」!B1 を読めません: %s",
            wb_path,
            APP_CONFIG_SHEET_NAME,
            ex,
        )
        return None
    if not b1:
        return None
    return _resolve_path_relative_to_workbook(wb_path, b1) or None


def _read_task_ids_from_config_sheet_column(
    wb_path: str,
    column_index: int,
    log_label: str,
    column_letter_desc: str,
    *,
    openpyxl_skip_hint: str | None = None,
) -> list[str]:
    """
    マクロブック「設定」シートの指定列（1=A, 2=B）3 行目以降から依頼NOを読む。
    空セルはスキップ。連続 30 セル空で打ち切り。最大 500 行。カンマ区切りで複数坯。
    """
    out: list[str] = []
    if not wb_path or not os.path.isfile(wb_path):
        return out
    if _workbook_should_skip_openpyxl_io(wb_path):
        msg = (
            f"{log_label}: ブックに「{OPENPYXL_INCOMPATIBLE_SHEET_MARKER}」があるため、"
            f"「{APP_CONFIG_SHEET_NAME}」!{column_letter_desc}3 以降は openpyxl で読めません。"
        )
        if openpyxl_skip_hint:
            msg += " " + openpyxl_skip_hint.strip()
        logging.info(msg)
        return out
    try:
        keep_vba = str(wb_path).lower().endswith(".xlsm")
        wb = load_workbook(
            wb_path, read_only=True, data_only=True, keep_vba=keep_vba
        )
        try:
            if APP_CONFIG_SHEET_NAME not in wb.sheetnames:
                return out
            ws = wb[APP_CONFIG_SHEET_NAME]
            consecutive_empty = 0
            for r in range(3, 3 + 500):
                t = _config_cell_text(ws.cell(row=r, column=column_index).value)
                if not t:
                    consecutive_empty += 1
                    if consecutive_empty >= 30:
                        break
                    continue
                consecutive_empty = 0
                if "," in t:
                    for part in t.split(","):
                        p = part.strip()
                        if p:
                            out.append(p)
                else:
                    out.append(t)
        finally:
            wb.close()
    except Exception as ex:
        logging.warning(
            "%s: 「%s」!%s3 以降の依頼NOを読めません（無視）: %s",
            log_label,
            APP_CONFIG_SHEET_NAME,
            column_letter_desc,
            ex,
        )
        return []
    return out


def _read_trace_schedule_task_ids_from_config_sheet(wb_path: str) -> list[str]:
    """
    マクロブック「設定」シート A 列の 3 行目以降を」配台トレース対象の依頼NOとして読む。
    空セルはスキップ。連続 30 セル空なら打ち切り。最大 500 行まで走査。
    """
    return _read_task_ids_from_config_sheet_column(
        wb_path,
        1,
        "配台トレース",
        "A",
        openpyxl_skip_hint="配台トレースは「設定」シート A 列を openpyxl で読めないため無効です。",
    )


def _read_debug_dispatch_task_ids_from_config_sheet(wb_path: str) -> list[str]:
    """
    マクロブック「設定」シート B 列の 3 行目以降を」段階2デバッグ配台の対象依頼NOとして読む。
    1 件も無い場合は段階2は通常モード（全件配台）。空セル・打ち切り等は A 列トレースとともに。
    """
    return _read_task_ids_from_config_sheet_column(
        wb_path,
        2,
        "デバッグ配台",
        "B",
        openpyxl_skip_hint="デバッグ配台は「設定」シート B 列を openpyxl で読めないため無効（全件配台）です。",
    )


@functools.lru_cache(maxsize=32)
def _read_gemini_model_try_chain_from_settings_sheet_cached(
    wb_path_norm: str, mtime_key: int
) -> tuple[str, ...]:
    """「設定」!D:E から試行モデル列を読む。mtime_key でブック保存後にキャッシュ無効化。"""
    out: list[str] = []
    if not wb_path_norm or not os.path.isfile(wb_path_norm):
        return tuple()
    if _workbook_should_skip_openpyxl_io(wb_path_norm):
        logging.info(
            "Gemini 試行モデル: ブックに「%s」があるため「%s」!D:E を openpyxl で読めません（既定モデル列を使用）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            APP_CONFIG_SHEET_NAME,
        )
        return tuple()
    try:
        keep_vba = str(wb_path_norm).lower().endswith(".xlsm")
        wb = load_workbook(
            wb_path_norm, read_only=True, data_only=True, keep_vba=keep_vba
        )
        try:
            if APP_CONFIG_SHEET_NAME not in wb.sheetnames:
                return tuple()
            ws = wb[APP_CONFIG_SHEET_NAME]
            consecutive_empty = 0
            last = GEMINI_MODEL_SHEET_FIRST_ROW + GEMINI_MODEL_SHEET_MAX_ROWS
            for r in range(GEMINI_MODEL_SHEET_FIRST_ROW, last):
                mid = _config_cell_text(
                    ws.cell(row=r, column=GEMINI_MODEL_SHEET_COL_MODEL).value
                )
                en_raw = ws.cell(row=r, column=GEMINI_MODEL_SHEET_COL_ENABLE).value
                if not mid:
                    consecutive_empty += 1
                    if consecutive_empty >= 20:
                        break
                    continue
                consecutive_empty = 0
                if not _config_cell_truthy_enabled(en_raw):
                    continue
                out.append(mid)
        finally:
            wb.close()
    except Exception as ex:
        logging.warning(
            "Gemini 試行モデル: 「%s」!D%d:E を読めません（既定モデル列を使用）: %s",
            APP_CONFIG_SHEET_NAME,
            GEMINI_MODEL_SHEET_FIRST_ROW,
            ex,
        )
        return tuple()
    return tuple(out)


def _read_gemini_model_try_chain_from_settings_sheet(wb_path: str) -> tuple[str, ...] | None:
    """マクロブック「設定」シートの D/E 列で有効化されたモデルを上から返す。1 件も無ければ None。"""
    p = (wb_path or "").strip()
    if not p:
        return None
    try:
        norm = os.path.normpath(os.path.abspath(p))
    except Exception:
        norm = p
    try:
        mkey = int(os.path.getmtime(norm))
    except OSError:
        mkey = 0
    chain = _read_gemini_model_try_chain_from_settings_sheet_cached(norm, mkey)
    if not chain:
        return None
    logging.info(
        "Gemini 試行モデル: 「%s」シート D/E から %s 件を読み込みました（順: %s）。",
        APP_CONFIG_SHEET_NAME,
        len(chain),
        ", ".join(chain),
    )
    return chain


def _show_stage2_debug_dispatch_mode_dialog(task_ids_sorted: list[str]) -> None:
    """設定シート B3 以降が空でないことが前提。Windows では MessageBox。それ以外は WARNING ログ。"""
    if not task_ids_sorted:
        return
    preview_lines = task_ids_sorted[:30]
    preview = "\n".join(preview_lines)
    if len(task_ids_sorted) > 30:
        preview += "\n…"
    body = (
        "デバッグモードで実行した。\n\n"
        "「設定」シート B3以降に入力した依頼NOのみを配台対象とした。\n\n"
        "対象依頼NO:\n"
        + preview
    )
    title = "段階2（配台）— デバッグモード"
    if sys.platform != "win32":
        logging.warning("%s\n%s", title, body)
        return
    try:
        ctypes.windll.user32.MessageBoxW(0, body, title, 0x00000040)
    except Exception as ex:
        logging.warning(
            "デバッグ配台: メッセージボックスを表示できません (%s)。%s", ex, body
        )


def _extract_gemini_api_key_from_plain_dict(data: dict, json_path: str) -> str | None:
    key = data.get("gemini_api_key")
    if key is None or (isinstance(key, str) and not key.strip()):
        key = data.get("GEMINI_API_KEY")
    if key is None:
        logging.warning(
            "Gemini: 証明書データに gemini_api_key（または GEMINI_API_KEY）はありません（%s）。",
            json_path,
        )
        return None
    s = str(key).strip()
    return s or None


def _derive_fernet_key_from_passphrase(
    passphrase: str, salt: bytes, iterations: int
) -> bytes:
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives import hashes

    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=iterations,
        backend=default_backend(),
    )
    return base64.urlsafe_b64encode(kdf.derive(passphrase.encode("utf-8")))


def _credentials_json_is_encrypted_v2(data: dict) -> bool:
    if data.get("format_version") == 2:
        return True
    return bool(
        data.get("kdf") == "pbkdf2_sha256" and (data.get("fernet_ciphertext") or "").strip()
    )


def _decrypt_gemini_credentials_v2(
    data: dict, passphrase: str, json_path: str
) -> dict | None:
    try:
        from cryptography.fernet import Fernet
    except ImportError:
        logging.warning(
            "Gemini: 暗号化証明書 JSON には cryptography は必須です（pip install cryptography）。"
        )
        return None
    token_s = (data.get("fernet_ciphertext") or "").strip()
    if not token_s:
        logging.warning(
            "Gemini: 暗号化証明書 JSON に fernet_ciphertext はありません（%s）。",
            json_path,
        )
        return None
    salt_b64 = (data.get("salt_b64") or "").strip()
    if not salt_b64:
        logging.warning(
            "Gemini: 暗号化証明書 JSON に salt_b64 はありません（%s）。",
            json_path,
        )
        return None
    try:
        salt = base64.standard_b64decode(salt_b64)
    except Exception as ex:
        logging.warning("Gemini: salt_b64 の解釈に失敗しました（%s）: %s", json_path, ex)
        return None
    iterations = int(data.get("iterations") or _GEMINI_CREDENTIALS_PBKDF2_ITERATIONS_DEFAULT)
    kdf_name = (data.get("kdf") or "pbkdf2_sha256").strip()
    if kdf_name != "pbkdf2_sha256":
        logging.warning("Gemini: 未対応の kdf（%s）: %s", kdf_name, json_path)
        return None
    try:
        fkey = _derive_fernet_key_from_passphrase(passphrase, salt, iterations)
        plain = Fernet(fkey).decrypt(token_s.encode("ascii"))
    except Exception:
        logging.debug("Gemini: 暗号化証明書の復号処理に失敗しました（%s）。", json_path)
        return None
    try:
        inner = json.loads(plain.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as ex:
        logging.warning(
            "Gemini: 復号後の JSON は正常です（%s）: %s",
            json_path,
            ex,
        )
        return None
    if not isinstance(inner, dict):
        logging.warning("Gemini: 復号後の JSON はオブジェクトである必須はありした（%s）。", json_path)
        return None
    return inner


def _load_gemini_api_key_from_credentials_json(
    json_path: str, workbook_path: str | None = None
) -> tuple[str | None, bool]:
    """戻り値: (api_key または None, 暗号化形式であったか)。暗号化時は _GEMINI_CREDENTIALS_PASSPHRASE_FIXED のみで復号。"""
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
    except OSError as ex:
        logging.warning("Gemini: 証明書 JSON を開きません: %s (%s)", json_path, ex)
        return None, False
    except json.JSONDecodeError as ex:
        logging.warning("Gemini: 証明書 JSON の形式は正常です: %s (%s)", json_path, ex)
        return None, False
    if not isinstance(data, dict):
        logging.warning("Gemini: 証明書 JSON はオブジェクト形式である必須はありした: %s", json_path)
        return None, False
    if _credentials_json_is_encrypted_v2(data):
        inner = _decrypt_gemini_credentials_v2(
            data, _GEMINI_CREDENTIALS_PASSPHRASE_FIXED, json_path
        )
        if inner is None:
            return None, True
        return _extract_gemini_api_key_from_plain_dict(inner, json_path), True
    return _extract_gemini_api_key_from_plain_dict(data, json_path), False


API_KEY = None
_cred_path = _read_gemini_credentials_json_path_from_workbook(TASKS_INPUT_WORKBOOK)
_used_encrypted_credentials = False
if _cred_path and os.path.isfile(_cred_path):
    API_KEY, _used_encrypted_credentials = _load_gemini_api_key_from_credentials_json(
        _cred_path, workbook_path=TASKS_INPUT_WORKBOOK
    )
    if API_KEY:
        if _used_encrypted_credentials:
            logging.info("Gemini API キー: 暗号化証明書 JSON から読み込みました。")
        else:
            logging.info(
                "Gemini API キー: マクロブック「%s」B1 のパスから読み込みました。",
                APP_CONFIG_SHEET_NAME,
            )
elif _cred_path:
    logging.warning(
        "Gemini: 「%s」B1 で指定された証明書 JSON は見つかりません。",
        APP_CONFIG_SHEET_NAME,
    )

# B1 は暗号化 JSON なのにキーは取れない（平文 JSON でキー欠損との区別）。原因の特定はログに書かう汎用メッセージのみ。
_encrypted_json_missing_key = (
    bool(_cred_path)
    and os.path.isfile(_cred_path)
    and _used_encrypted_credentials
    and not API_KEY
)
if _encrypted_json_missing_key:
    logging.error(
        "Gemini: 「%s」B1 の証明書ファイルから API キーを利用できません。"
        " 社内手順に従い証明書を再設定れるか」管理者に相い合わせでしてさい。",
        APP_CONFIG_SHEET_NAME,
    )

if not API_KEY:
    logging.warning(
        "Gemini API キーは未設定です。マクロブックに「%s」シートを用意し B1 に証明書 JSON のフルパスを書いてください。"
        " 備考の AI 解析等はスキップされした。"
        " 参考型: gemini_credentials.example.json / encrypt_gemini_credentials.py（暗号化）。",
        APP_CONFIG_SHEET_NAME,
    )

RESULT_SHEET_GANTT_NAME = "結果_設備ガント"
# 結果_設備ガントの横軸タイムスロット幅（分）
GANTT_TIMELINE_SLOT_MINUTES = 10
# 結果_設備ガントの時刻列（E 列以降）の列幅（Excel / openpyxl の標準単位）
GANTT_TIMELINE_COLUMN_WIDTH = 3
# 結果_設備ガントの時刻見出し行（hdr_row）の RowDimension.height（ポイント）
GANTT_HDR_ROW_HEIGHT_PT = int(float(os.environ.get("GANTT_HDR_ROW_HEIGHT_PT", "38")))
# 結果_設備ガントの機械（計画／実績）行の RowDimension.height（ポイント）。
GANTT_MACHINE_ROW_HEIGHT_PT = int(float(os.environ.get("GANTT_MACHINE_ROW_HEIGHT_PT", "60")))
# 既定の印刷は横1ページに合わせる（fitToWidth=1）。固定縮小率は横幅が潰れて読みにくいため既定では使わない。
# どうしても固定%で出したいときだけ環境変数 GANTT_PRINT_SCALE_PERCENT（10〜400 の整数）を設定する。

# タスク列名（マクロ実行ブック「加工計画DATA」）
TASK_COL_TASK_ID = "依頼NO"
TASK_COL_MACHINE = "工程名"
TASK_COL_MACHINE_NAME = "機械名"
TASK_COL_QTY = "換算数量"
# 加工計画DATA にある場合のみ段階1で配台計画へコピー。結果_タスク一覧「残加工量」はこの列の数値基準で出力する。
TASK_COL_UNPROCESSED = "未加工"
TASK_COL_ORDER_QTY = "受注数"
TASK_COL_SPEED = "加工速度"
TASK_COL_PRODUCT = "製品名"
TASK_COL_ANSWER_DUE = "回答納期"
TASK_COL_SPECIFIED_DUE = "指定納期"
TASK_COL_RAW_INPUT_DATE = "原反投入日"
# 加工計画DATA 由来。配台計画_タスク入力では原反投入日の右隣（SOURCE_BASE_COLUMNS 順）。
TASK_COL_STOCK_LOCATION = "在庫場所"
# 同一依頼NOの工程順（カンマ区切りの工程名）。加工計画DATA＝配台計画_タスク入力。
TASK_COL_PROCESS_CONTENT = "加工内容"
# 投入可能日の目安は「回答納期」。未入力時は「指定納期」（剝日基準・当日/遅れは最優先）。「加工開始日」列は参照しない。
# 完了判定・進杗（加工計画DATA）
TASK_COL_COMPLETION_FLAG = "加工完了区分"
TASK_COL_ACTUAL_DONE = "実加工数"   # 旧互換（直接の加工済数値）
TASK_COL_ACTUAL_OUTPUT = "実出来高"  # 完了品数値（残作に使う）
TASK_COL_DATA_EXTRACTION_DT = "データ抽出日"
# 配台基準日時の主列（加工計画DATA）。無い・空のときは TASK_COL_EXTRACTION_TIME / TASK_COL_DATA_EXTRACTION_DT を参照。
TASK_COL_DATA_EXTRACTION_TIME = "データ抽出時間"
# 配台基準日時の旧列（加工計画DATA）。無い・空のときは TASK_COL_DATA_EXTRACTION_DT を参照。
TASK_COL_EXTRACTION_TIME = "抽出時間"
AI_CACHE_TTL_SECONDS = 6 * 60 * 60  # 6時間
# json/ai_remarks_cache.json 内のキー接頭辞（設定_配台不要工程・配台不要ロジック D→E）
AI_CACHE_KEY_PREFIX_EXCLUDE_RULE_DE = "exclude_rule_de_v1"

# マクロブック「加工実績DATA」（Power Query 等で取り込み想定）
ACTUALS_SHEET_NAME = "加工実績DATA"
ACT_COL_TASK_ID = "依頼NO"
ACT_COL_PROCESS = "工程名"
ACT_COL_OPERATOR = "担当者"
ACT_COL_START_DT = "開始日時"
ACT_COL_END_DT = "終了日時"
ACT_COL_START_ALT = "実績開始"
ACT_COL_END_ALT = "実績終了"
ACT_COL_DAY = "日付"
ACT_COL_TIME_START = "開始時刻"
ACT_COL_TIME_END = "終了時刻"
# 加工実績明細DATA 等で使われる日時列（開始日時/終了日時 の別名）
ACT_COL_MACHINING_START_DT = "加工開始日時"
ACT_COL_MACHINING_END_DT = "加工終了日時"
# 加工実績明細DATA（Power Query 等）… ガント角丸シェイプの %% は「累積完了率」列をそのまま用いる（計算しない）
ACT_COL_ACTUAL_QTY = "実加工数"
ACT_COL_PLANNED_QTY = "加工予定数"
ACT_COL_CONVERTED_QTY = "換算数量"
ACT_COL_CUMULATIVE_COMPLETION_PCT = "累積完了率"
# 加工実績明細DATA のロール単位担当（1～5。見出しは Excel シートと一致させる）
ACT_COL_MACHINING_ASSIGNEE_1 = "加工担当者名1"
ACT_COL_MACHINING_ASSIGNEE_2 = "加工担当者名2"
ACT_COL_MACHINING_ASSIGNEE_3 = "加工担当者名3"
ACT_COL_MACHINING_ASSIGNEE_4 = "加工担当者名4"
ACT_COL_MACHINING_ASSIGNEE_5 = "加工担当者名5"
ACT_COL_MACHINING_ASSIGNEES_ORDERED = (
    ACT_COL_MACHINING_ASSIGNEE_1,
    ACT_COL_MACHINING_ASSIGNEE_2,
    ACT_COL_MACHINING_ASSIGNEE_3,
    ACT_COL_MACHINING_ASSIGNEE_4,
    ACT_COL_MACHINING_ASSIGNEE_5,
)
ACTUAL_HEADER_CANONICAL = (
    ACT_COL_TASK_ID,
    ACT_COL_PROCESS,
    ACT_COL_OPERATOR,
    ACT_COL_START_DT,
    ACT_COL_END_DT,
    ACT_COL_START_ALT,
    ACT_COL_END_ALT,
    ACT_COL_DAY,
    ACT_COL_TIME_START,
    ACT_COL_TIME_END,
    ACT_COL_MACHINING_START_DT,
    ACT_COL_MACHINING_END_DT,
)
# マクロブック「加工実績明細DATA」… ロール単位の実績行（列は加工実績DATA に準拠＋任意のロール識別列）
ACTUAL_DETAIL_SHEET_NAME = "加工実績明細DATA"
ACT_DETAIL_COL_ROLL = "ロールNO"
ACTUAL_DETAIL_HEADER_CANONICAL = ACTUAL_HEADER_CANONICAL + (
    ACT_COL_ACTUAL_QTY,
    ACT_COL_PLANNED_QTY,
    ACT_COL_CONVERTED_QTY,
    ACT_COL_CUMULATIVE_COMPLETION_PCT,
    ACT_DETAIL_COL_ROLL,
) + ACT_COL_MACHINING_ASSIGNEES_ORDERED
# 加工実績明細由来。見た目・印刷設定は RESULT_SHEET_GANTT_NAME と同型（計画行なし・実績帯のみ）。
RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME = "結果_設備ガント_実績明細"
# 設備実績ガント（上記シート）の「表示する暦日」を段階2生成時に絞る。設定_環境変数の A 列＝変数名・B 列＝値（空＝その側の制限なし）
# 日付は ISO（2026-04-01）・Excel 日付セル相当の数値文字列・2026/4/1 等を parse_optional_date で解釈。
ENV_GANTT_ACTUAL_DETAIL_DATE_FROM = "GANTT_ACTUAL_DETAIL_DATE_FROM"
ENV_GANTT_ACTUAL_DETAIL_DATE_TO = "GANTT_ACTUAL_DETAIL_DATE_TO"
# 段階2を回さず実績明細ガントのみ再生成するときの単一シート xlsx（マクロが取り込み）
ACTUAL_DETAIL_GANTT_REFRESH_FILENAME = "actual_detail_gantt_refresh.xlsx"

# --- 2段階処理: 段階1抽出 → ブック「配台計画_タスク入力」編集 → 段階2計画 ---
STAGE1_OUTPUT_FILENAME = "plan_input_tasks.xlsx"
# 既定は Excel 実シート名「配台計画_タスク入力」。旧ソースの「配台計画_」は誤字（UTF-8 保存時の破損）で一致しない。
PLAN_INPUT_SHEET_NAME = os.environ.get("TASK_PLAN_SHEET", "").strip() or "配台計画_タスク入力"
PLAN_COL_SPEED_OVERRIDE = "加工速度_上書き"
# 空白のときは列「原反投入日」（加工計画DATA 由来）をそのまま使う。日付ありのときは配台の原反制約・結果_タスク一覧表示の両方でこの日付を採用。
PLAN_COL_RAW_INPUT_DATE_OVERRIDE = "原反投入日_上書き"
PLAN_COL_PREFERRED_OP = "担当OP_指定"
PLAN_COL_SPECIAL_REMARK = "特別指定_備考"
# 参照列「（元）配台不要」は置かない（元データに相当するマスタ列が無いため）。
# セル値の例（配台から外す）: Excel の TRUE / 数値 1 / 文字列「はい」「yes」「true」「○」「〇」「●」等。
# 空・FALSE・0・「いいえ」等は配台対象。詳細は _plan_row_exclude_from_assignment。
PLAN_COL_EXCLUDE_FROM_ASSIGNMENT = "配台不要"
PLAN_COL_AI_PARSE = "AI特別指定_解析"
PLAN_COL_PROCESS_FACTOR = "加工工程の決定プロセスの因子"
# 1ロールあたりの長さ（m）。配台計画_タスク入力にのみ存在（加工計画DATA には無い）。製品名列の右隣に配置。
PLAN_COL_ROLL_UNIT_LENGTH = "ロール単位長さ"
# 段階1で算出し配台計画に出力。段階2の build_task_queue と同一式（_plan_row_dispatch_qty_metrics の残りm）。
PLAN_COL_DISPATCH_REMAINING_QTY = "配台使用残数量"
# 配台計算で使う換算数量の下限（m）。正の値でこれ未満のときはこの値に引き上げる（段階1）。
PLANNING_MIN_QTY_M = 100.0
# ロール単位長さを 100m 単位に切り上げるときの刻み（段階1）。例: 40→100, 125→200。
ROLL_UNIT_LENGTH_CEIL_STEP_M = 100.0
# 製品名から寸法（NNNxMM 等）を解釈できないときの 1 ロール長(m)。換算数量へは落とさない（FEL 品番で 2000 等になるのを防ぐ）。
INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M = 100.0
DEBUG_TASK_ID = os.environ.get("DEBUG_TASK_ID", "Y3-26").strip()
# 例: set TRACE_TEAM_ASSIGN_TASK_ID=W3-14 … 配台ループで「人数別の最良候補」と採用理由を INFO ログに出す
TRACE_TEAM_ASSIGN_TASK_ID = os.environ.get("TRACE_TEAM_ASSIGN_TASK_ID", "").strip()
# 配台トレース対象はマクロブック「設定」シート A 列 3 行目以降のみ（generate_plan 冒頭で確定）。環境変数は使えない。
TRACE_SCHEDULE_TASK_IDS: frozenset[str] = frozenset()
# 段階2デバッグ配台: 「設定」B 列 3 行目以降に依頼NOはあるとしのみ」しの依頼の行の値配台（generate_plan 冒頭で確定）。空なら全件。
DEBUG_DISPATCH_ONLY_TASK_IDS: frozenset[str] = frozenset()
# 紝期超靎リトライの外側ラウンド（0=初回カレンダー通し、以降は while 先頭で更新）。配台トレース出力のファイル名・接頭辞に使用。
DISPATCH_TRACE_OUTER_ROUND: int = 0


def _trace_schedule_task_enabled(task_id) -> bool:
    if not TRACE_SCHEDULE_TASK_IDS:
        return False
    return str(task_id or "").strip() in TRACE_SCHEDULE_TASK_IDS


def _sanitize_dispatch_trace_filename_part(task_id: str) -> str:
    """依頼NOを log ファイル名に使うための簡易サニタイズ（Windows 禁止文字を避ける）。"""
    s = "".join(
        c if (c.isalnum() or c in "-_.") else "_"
        for c in str(task_id or "").strip()
    )
    return s[:120] if s else "task"


def _reset_dispatch_trace_per_task_logfiles() -> None:
    """
    段階2実行の冒頭で1回」log 内の dispatch_trace_*.txt をまとめて削除する（除去実行の残骸を残さない）。
    坄外側ラウンド用ファイルは generate_plan の while 先頭で _dispatch_trace_begin_outer_round はヘッダ付し新規作成する。
    execution_log.txt とは別ファイル。内容は [配台トレース task=…] 行を _log_dispatch_trace_schedule で追記
    （日次残・ロール確定の余剰有無・余力追記・終了時サマリ等）。
    """
    if not TRACE_SCHEDULE_TASK_IDS:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
    except OSError:
        return
    try:
        for _name in os.listdir(log_dir):
            if not (
                str(_name).startswith("dispatch_trace_") and str(_name).endswith(".txt")
            ):
                continue
            _p = os.path.join(log_dir, _name)
            try:
                os.unlink(_p)
            except OSError:
                pass
    except OSError:
        pass


def _dispatch_trace_begin_outer_round(round_n: int) -> None:
    """紝期超靎リトライの外側ラウンド番坷を確定し、当ラウンド用 dispatch_trace_*_rNN.txt のヘッダを1回だけ書き。"""
    global DISPATCH_TRACE_OUTER_ROUND
    DISPATCH_TRACE_OUTER_ROUND = max(0, int(round_n))
    if not TRACE_SCHEDULE_TASK_IDS:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
    except OSError:
        return
    for tid in TRACE_SCHEDULE_TASK_IDS:
        t = str(tid or "").strip()
        if not t:
            continue
        safe = _sanitize_dispatch_trace_filename_part(t)
        path = os.path.join(
            log_dir,
            f"dispatch_trace_{safe}_r{DISPATCH_TRACE_OUTER_ROUND:02d}.txt",
        )
        if os.path.exists(path):
            continue
        try:
            with open(path, "w", encoding="utf-8", newline="\n") as f:
                f.write(
                    "# 配台トレース（依頼NOとと・外側ラウンド別）。同一行は log/execution_log.txt にも出力されした。\n"
                    f"# task_id={t}  outer_round={DISPATCH_TRACE_OUTER_ROUND}  "
                    "# （0=初回カレンダー通し、以降は紝期超靎リトライごとに +1）\n\n"
                )
        except OSError as ex:
            logging.warning("dispatch_trace ログの初期化に失敗: %s (%s)", path, ex)


def _log_dispatch_trace_schedule(task_id, msg: str, *args) -> None:
    """[配台トレース task=…] を execution_log に出しつつ」対象依頼NO専用ファイルにも追記れる。"""
    t = str(task_id or "").strip()
    body_raw = msg % args if args else msg
    body = body_raw
    if t and t in TRACE_SCHEDULE_TASK_IDS:
        body = f"[outer_round={DISPATCH_TRACE_OUTER_ROUND:02d}] {body_raw}"
    logging.info(body)
    if not t or t not in TRACE_SCHEDULE_TASK_IDS:
        return
    safe = _sanitize_dispatch_trace_filename_part(t)
    path = os.path.join(
        log_dir,
        f"dispatch_trace_{safe}_r{DISPATCH_TRACE_OUTER_ROUND:02d}.txt",
    )
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
        line = f"{ts} - INFO - {body}\n"
        with open(path, "a", encoding="utf-8", newline="\n") as f:
            f.write(line)
    except OSError as ex:
        try:
            logging.warning("dispatch_trace 側ファイルへの追記に失敗: %s (%s)", path, ex)
        except Exception:
            pass


# True: 従来の「人数最優先」タプル (-人数, 開始, -短縮数, 優先度合計)。False のとき下記スラック分と組み合わせ
TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF = os.environ.get(
    "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF", "0"
).strip().lower() not in ("0", "false", "no", "off", "いいえ")


def _team_assign_start_slack_wait_minutes() -> int:
    """全日候補の最早開始からこの分以内の遅れなら」開始より人数を優先（分）。0 で無効。"""
    raw = os.environ.get("TEAM_ASSIGN_START_SLACK_WAIT_MINUTES", "60").strip()
    try:
        v = int(raw)
    except ValueError:
        v = 60
    return max(0, v)


TEAM_ASSIGN_START_SLACK_WAIT_MINUTES = _team_assign_start_slack_wait_minutes()

# True のとき need シート「配台時追加人数」行を無視し、フォーム人数は基本必須人数（req_num）のみ試行し、メイン後追記もしない。
TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW = (
    os.environ.get("TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)

# True: 従来どおりメイン割付の組み合わせ探索で req_num〜req_num+追加人数上限まで試す。
# False（既定）: メインは req_num のみ。追加人数上限は全シミュレーション完了後」当該ブロック時間に
#     他タスクへ未割当（時間重なりなし）かつ skills 革坈の者をサブとして追記（append_surplus_staff_after_main_dispatch）。
TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS = (
    os.environ.get("TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS", "")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)

# True（既定）: メイン配台の必須人数は need（基本必須人数＋特別指定）のみ。
# False のときは特別指定備考 AI の required_op のみ計画からから参照し得る（シート列「必須人数」は廃止済み）。
TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY = (
    os.environ.get("TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)
# True（既定）: master「組み合わせ表」に該当行はある工程+機械は」組み合わせ優先度の昇順で
# 最初に成立したメンバー編成を採用。まとめて試行なら従来の itertools 組み合わせ探索。
TEAM_ASSIGN_USE_MASTER_COMBO_SHEET = (
    os.environ.get("TEAM_ASSIGN_USE_MASTER_COMBO_SHEET", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)

# §B-2 熱融着検査を同一設備（工程列キー）で「開始済み1件に残ロールはある間は他依頼の検査を試さない」か。
# 0 / false / no / off で無効にすると設備時間割上で依頼は混在し得るは」占有による長期ブロック（例: W3-14 型）を避けられる。
PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE = (
    os.environ.get("PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ", "無効")
)

# §B-2 / §B-3 同一依頼で EC と後続（検査＝巻返し）の担当者集合を排他するか。
# 0 / false / no / off / いいえ / 無効 で無効化すると」履歴ベースの相互除外を行う同一人物は両側の候補に残り得る。
PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS = (
    os.environ.get("PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ", "無効")
)

# マクロブック「設定_配台不要工程」: 既定では openpyxl save を試さず xlwings 同期→Save（Excel 占有時は openpyxl は実質失敗するため）。失敗時は TSV→VBA 反映。
# コマンド等で openpyxl を試れ場合は EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1。
EXCLUDE_RULES_SHEET_NAME = "設定_配台不要工程"
EXCLUDE_RULES_SKIP_OPENPYXL_SAVE = os.environ.get(
    "EXCLUDE_RULES_TRY_OPENPYXL_SAVE", ""
).strip().lower() not in ("1", "true", "yes", "on")
EXCLUDE_RULE_COL_PROCESS = "工程名"
EXCLUDE_RULE_COL_MACHINE = "機械名"
EXCLUDE_RULE_COL_FLAG = "配台不要"
EXCLUDE_RULE_COL_LOGIC_JA = "配台不要ロジック"
EXCLUDE_RULE_COL_LOGIC_JSON = "ロジック式"
# 元ブックはロックされ別名保存した場合」同一プロセス内のルール読込はこのパスを優先
_exclude_rules_effective_read_path: str | None = None
# 直後の apply_exclude_rules（同一プロセス）用: VBA 反映後でも E 列付しルールを使う
_exclude_rules_rules_snapshot: list | None = None
_exclude_rules_snapshot_wb: str | None = None
# ルール JSON の conditions で参照可能な列（AI プロンプトと評価器を一致させる）
EXCLUDE_RULE_ALLOWED_COLUMNS = frozenset(
    {
        TASK_COL_TASK_ID,
        TASK_COL_MACHINE,
        TASK_COL_MACHINE_NAME,
        TASK_COL_QTY,
        TASK_COL_UNPROCESSED,
        TASK_COL_ORDER_QTY,
        TASK_COL_SPEED,
        TASK_COL_PRODUCT,
        TASK_COL_ANSWER_DUE,
        TASK_COL_SPECIFIED_DUE,
        TASK_COL_RAW_INPUT_DATE,
        TASK_COL_STOCK_LOCATION,
        TASK_COL_PROCESS_CONTENT,
        TASK_COL_COMPLETION_FLAG,
        TASK_COL_ACTUAL_DONE,
        TASK_COL_ACTUAL_OUTPUT,
        TASK_COL_DATA_EXTRACTION_DT,
        TASK_COL_DATA_EXTRACTION_TIME,
        TASK_COL_EXTRACTION_TIME,
        PLAN_COL_SPEED_OVERRIDE,
        PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
        PLAN_COL_PREFERRED_OP,
        PLAN_COL_SPECIAL_REMARK,
        PLAN_COL_PROCESS_FACTOR,
        PLAN_COL_ROLL_UNIT_LENGTH,
        PLAN_COL_DISPATCH_REMAINING_QTY,
    }
)

# 計画結果ブック「結果_タスク一覧」の列順・表示（マクロ実行ブックの同名シートで上書き可）
RESULT_TASK_SHEET_NAME = "結果_タスク一覧"
RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME = "結果_設備毎の時間割"
# 工程名+機械の複合列ではなく、機械名単位で各枠の依頼NOを把握しやすくする
RESULT_EQUIPMENT_BY_MACHINE_SHEET_NAME = "結果_設備毎の時間割_機械名毎"
# master メイン A15/B15 の定常外の「日時帯」見出し着色（結果_設備毎の時間割・結果_設備ガント）
RESULT_OUTSIDE_REGULAR_TIME_FILL = "FCE4D6"
# 結果_設備毎の時間割_機械名毎: 配台済み依頼NOセル（機械列）の薄いグリーン
# 結果_設備毎の時間割: 日次始業準備の設備セルも同系色
RESULT_DISPATCHED_REQUEST_FILL = "C6EFCE"
# 結果_設備毎の時間割: master「機械カレンダー」占有と重なる設備セル（10分枠）
RESULT_MACHINE_CALENDAR_BLOCK_FILL = "D4B3E8"
# 結果_設備ガント: 機械名グループ（機械名列の同一坝称）ごとに B〜D 列を区別する淡色（順に割当・循環）
RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS = (
    "E8F4FC",
    "FCE8F0",
    "E8F8E8",
    "FFF0D8",
    "EDE8FC",
    "E0F8F4",
    "F8E8E0",
    "E8ECF8",
    "F5F5E0",
    "F0E8E8",
)
# GANTT_COLOR_MODE=full 時の機械名列グループ（B〜D）用。やや彩度の高いパステル。
RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS_FULL = (
    "B3E5FC",
    "F8BBD0",
    "C8E6C9",
    "FFE0B2",
    "E1BEE7",
    "B2EBF2",
    "FFF59D",
    "D1C4E9",
    "FFCDD2",
    "C5CAE9",
)
# 配台シミュレーション開始剝（初回 task_queue.sort 後）のキュー順。1 始まり・全日程で試行
RESULT_TASK_COL_DISPATCH_TRIAL_ORDER = "配台試行順番"
# 配台済_加工終了は「回答納期+16:00」または「指定納期+16:00」（回答は空のとき）以降かを表示
RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16 = "配台済_回答指定16時まで"
# マスタ skills の工程+機械列ととの OP/AS 割当参考順（優先度値・並び順）とフォーム採用ルールの説明
RESULT_MEMBER_PRIORITY_SHEET_NAME = "結果_人員配台優先順"
COLUMN_CONFIG_SHEET_NAME = "列設定_結果_タスク一覧"
COLUMN_CONFIG_HEADER_COL = "列名"
COLUMN_CONFIG_VISIBLE_COL = "表示"
# 段階2の結果 xlsx 生成後」入力ブックの列設定シート上の図形（フォームボタン等）を xlwings で複製れる（既定 ON。無効化は STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT=0）
STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT = os.environ.get(
    "STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT", "1"
).strip().lower() in ("1", "true", "yes", "on")
# 結果_設備ガントのタイムライン上ラベル（依頼NO 等）をセル文字ではなく角丸四角シェイプで重ねる。
# xlwings + Excel が必要。失敗時は openpyxl でセルにフォールバック。無効化: GANTT_TIMELINE_SHAPE_LABELS=0
GANTT_TIMELINE_SHAPE_LABELS = os.environ.get(
    "GANTT_TIMELINE_SHAPE_LABELS", "1"
).strip().lower() in ("1", "true", "yes", "on")
# 角丸ラベルシェイプを「日ブロック」単位で 1 枚の画像にまとめ、シェイプ数・描画負荷を抑える（既定 ON。従来の大量シェイプ: GANTT_TIMELINE_LABELS_DAY_FLATTEN=0）
GANTT_TIMELINE_LABELS_DAY_FLATTEN = os.environ.get(
    "GANTT_TIMELINE_LABELS_DAY_FLATTEN", "1"
).strip().lower() in ("1", "true", "yes", "on")
# 日別画像: 外接矩形にクロマキー色を敷き Picture 透明化（任意）。既定 OFF。敷き色 GANTT_DAY_IMAGE_CHROMA_HEX（既定 FF00FF）。CopyPicture の形式: GANTT_DAY_IMAGE_COPY_PICTURE_FORMAT=picture|bitmap（空なら既定 xlPicture=EMF。手動の「図としてコピー」に近く、シェイプ実体以外が透明になりやすい。bitmap は余白が不透明になりがち）
GANTT_DAY_IMAGE_CHROMA_TRANSPARENT = os.environ.get(
    "GANTT_DAY_IMAGE_CHROMA_TRANSPARENT", "0"
).strip().lower() in ("1", "true", "yes", "on")
# 結果_設備ガントの配色。未設定・monotone で従来（淡色・モノトーン寄り）。full で依頼NO単位のHSV帯色＋周辺色をやや鮮やかに。
# OS の環境変数に加え、マクロブック「設定_環境変数」シート（A=GANTT_COLOR_MODE・B=値）でも指定可。
# workbook_env_bootstrap が import planning_core より前に os.environ へ反映する。
# 例: GANTT_COLOR_MODE=full / GANTT_COLOR_MODE=monotone


def _gantt_color_mode_raw() -> str:
    return (os.environ.get("GANTT_COLOR_MODE", "") or "").strip().lower()


def _gantt_color_mode_full() -> bool:
    return _gantt_color_mode_raw() in ("full", "color", "vivid", "1", "true", "yes", "on")


# 結果_タスク一覧の日付系（yyyy/mm/dd 文字列）に付けるフォント色。履歴列の【日付】と揃える
RESULT_TASK_DATE_STYLE_HEADERS = frozenset(
    {
        "回答納期",
        "指定納期",
        "計画基準納期",
        TASK_COL_RAW_INPUT_DATE,
        "加工開始日",
        "配台済_加工開始",
        "配台済_加工終了",
    }
)

SOURCE_BASE_COLUMNS = [
    TASK_COL_TASK_ID,
    TASK_COL_MACHINE,
    TASK_COL_MACHINE_NAME,
    TASK_COL_QTY,
    TASK_COL_UNPROCESSED,
    TASK_COL_ORDER_QTY,
    TASK_COL_SPEED,
    TASK_COL_PRODUCT,
    TASK_COL_ANSWER_DUE, TASK_COL_SPECIFIED_DUE, TASK_COL_RAW_INPUT_DATE, TASK_COL_STOCK_LOCATION,
    TASK_COL_PROCESS_CONTENT,
    TASK_COL_COMPLETION_FLAG, TASK_COL_ACTUAL_DONE, TASK_COL_ACTUAL_OUTPUT,
]
PLAN_OVERRIDE_COLUMNS = [
    PLAN_COL_EXCLUDE_FROM_ASSIGNMENT,
    PLAN_COL_SPEED_OVERRIDE,
    PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
    PLAN_COL_PREFERRED_OP,
    PLAN_COL_SPECIAL_REMARK,
    PLAN_COL_AI_PARSE,
]
# 矛盾検出でリセット対象にれる列（見出し行の文言と一致すること）
PLAN_CONFLICT_STYLABLE_COLS = tuple(PLAN_OVERRIDE_COLUMNS)
# 段階1再抽出時」既存「配台計画_タスク入力」から継承れる列（AIの解析結果列は毎回空に戻れ）
PLAN_STAGE1_MERGE_COLUMNS = tuple(c for c in PLAN_OVERRIDE_COLUMNS if c != PLAN_COL_AI_PARSE)
# 上書き以外で」再抽出時に旧シートから引き継ぎ列（セルは空でないとしのみ）
# 配台試行順番は毎回空クリア後に fill_plan_dispatch_trial_order_column_stage1 で付け直すが対象外。
PLAN_STAGE1_MERGE_EXTRA_COLUMNS = (PLAN_COL_ROLL_UNIT_LENGTH,)
# openpyxl 保存はブックロックで失敗したとし」VBA は開いているブックへ書式適用するための指示ファイル
PLANNING_CONFLICT_SIDECAR = "planning_conflict_highlight.tsv"
# 配台計画_タスク入力へ「グローバルコメント解析」を書き列（表の坳端より外側。1行目から縦にラベル＝値）
# ★ 参照表示のみ: load_planning_tasks_df 等は本列を一切読まない。配台適用は常にメイン「グローバルコメント」1経路のため、二重適用にならない。
PLAN_SHEET_GLOBAL_PARSE_LABEL_COL = 50  # AX
PLAN_SHEET_GLOBAL_PARSE_VALUE_COL = 51  # AY
PLAN_SHEET_GLOBAL_PARSE_MAX_ROWS = 42


def plan_reference_column_name(override_col: str) -> str:
    """上書き列の左隣に置し参照列の見出し（セル値は括弧付しで元データを表示）。"""
    return f"（元）{override_col}"


def plan_input_sheet_column_order():
    """
    配台計画_タスク入力の列順（段階1出力・段階2読込で共通）。

    0. 配台試行順番（段階1抽出直後に空クリア→段階2と同じ趣旨に付与。段階2は全行に値はあるとしこの順を優先）
    1. 配台不要（参照列なし）
    2. 加工計画DATA 由来（SOURCE_BASE_COLUMNS）… 依頼NO〜実出来高まで（換算数量の次に未加工→配台使用残数量、製品名の直後にロール短縮長さ、原反投入日の直後に在庫場所）
    3. 加工工程の決定プロセスの因孝
    4. 上書き列… 複数列の直後に「（元）…」参照列。AI特別指定_解析のみ参照列なし。
       （日付系上書きに 原反投入日_上書き を含む。空白時は列「原反投入日」を配台に使用）

    global_speed_rules 等で変える実効速度は計画シート列には出ないが、配台で確定した値は結果_タスク一覧の「加工速度」列に出力される。
    """
    cols = [RESULT_TASK_COL_DISPATCH_TRIAL_ORDER, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT]
    for c in SOURCE_BASE_COLUMNS:
        cols.append(c)
        if c == TASK_COL_UNPROCESSED:
            cols.append(PLAN_COL_DISPATCH_REMAINING_QTY)
        if c == TASK_COL_PRODUCT:
            cols.append(PLAN_COL_ROLL_UNIT_LENGTH)
    cols.append(PLAN_COL_PROCESS_FACTOR)
    for c in PLAN_OVERRIDE_COLUMNS:
        if c == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
            continue
        if c == PLAN_COL_AI_PARSE:
            cols.append(c)
        else:
            cols.append(plan_reference_column_name(c))
            cols.append(c)
    return cols


def _format_paren_ref_scalar(val):
    """参照表示用: 空は（―）」日付・しの他は（値）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "（―）"
    if isinstance(val, datetime):
        d = val.date() if hasattr(val, "date") else val
        if isinstance(d, date):
            return f"（{d.year}/{d.month}/{d.day}）"
    if isinstance(val, date):
        return f"（{val.year}/{val.month}/{val.day}）"
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return "（―）"
    return f"（{s}）"


def _reference_text_for_override_row(row, override_col: str, req_map: dict, need_rules: list) -> str:
    """1行分の上書き列に対応れる参照文言（括弧付し）。"""
    _ = (req_map, need_rules)  # 旧「（元）必須人数」参照で使用。列廃止により未使用だが呼び出し互換のため残す。
    if override_col == PLAN_COL_SPEED_OVERRIDE:
        v = row.get(TASK_COL_SPEED)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "（―）"
        try:
            x = float(v)
            if abs(x - round(x)) < 1e-9:
                return f"（{int(round(x))}）"
            return f"（{x}）"
        except (TypeError, ValueError):
            return _format_paren_ref_scalar(v)
    if override_col in (PLAN_COL_PREFERRED_OP, PLAN_COL_SPECIAL_REMARK):
        return "（―）"
    if override_col == PLAN_COL_RAW_INPUT_DATE_OVERRIDE:
        return _format_paren_ref_scalar(
            parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE))
        )
    return "（―）"


def _refresh_plan_reference_columns(df, req_map: dict, need_rules: list):
    """加工計画DATA＝need に基るし「（元）…」列を再計算（マージ後に必う呼め）。"""
    if df is None or df.empty:
        return df
    need_rules = need_rules or []
    req_map = req_map or {}
    for i in df.index:
        row = df.loc[i]
        for oc in PLAN_OVERRIDE_COLUMNS:
            if oc == PLAN_COL_AI_PARSE:
                continue
            if oc == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
                continue
            ref_col = plan_reference_column_name(oc)
            if ref_col not in df.columns:
                continue
            df.at[i, ref_col] = _reference_text_for_override_row(row, oc, req_map, need_rules)
    return df


def _apply_plan_input_visual_format(path: str, sheet_name: str = "タスク一覧"):
    """上書き入力列に薄い黄色を付与（参照列は未着色。AI解析列は除外）。"""
    # 見出し文字の表記ゆれで列名検索に失敗しはうなため、段階1の列順（plan_input_sheet_column_order）の
    # 1-based 列番坷で塗る（to_excel の列順と一致させる）。
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    order = plan_input_sheet_column_order()
    col_1based = {name: i + 1 for i, name in enumerate(order)}
    if _workbook_should_skip_openpyxl_io(path):
        logging.info(
            "配台計画の視覚整形: ブックに「%s」があるため、openpyxl での着色をスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    wb = load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        last_row = ws.max_row or 1
        if last_row < 2:
            return
        for oc in PLAN_OVERRIDE_COLUMNS:
            if oc == PLAN_COL_AI_PARSE:
                continue
            ci = col_1based.get(oc)
            if not ci:
                continue
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=ci).fill = fill_yellow
        _ci_rul = col_1based.get(PLAN_COL_ROLL_UNIT_LENGTH)
        if _ci_rul:
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=_ci_rul).fill = fill_yellow
        wb.save(path)
    finally:
        wb.close()


def _planning_conflict_sidecar_path():
    return os.path.join(log_dir, PLANNING_CONFLICT_SIDECAR)


def _remove_planning_conflict_sidecar_safe():
    try:
        os.remove(_planning_conflict_sidecar_path())
    except OSError:
        pass


def write_planning_conflict_highlight_sidecar(sheet_name, num_data_rows, conflicts_by_row):
    """
    Excel はブックを開いたままのとき保存でしない場合に」VBA 用の TSV を log に書き。
    形式: V1 / シート名 / データ行数 / クリア列をタブ結合 / 以降 行番坷\\t列名
    """
    path = _planning_conflict_sidecar_path()
    clear_cols = "\t".join(PLAN_CONFLICT_STYLABLE_COLS)
    lines = ["V1", sheet_name, str(int(num_data_rows)), clear_cols]
    for r in sorted(conflicts_by_row.keys()):
        for name in sorted(conflicts_by_row[r]):
            lines.append(f"{int(r)}\t{name}")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")

# 段階1出力・ブック内の日付列を Excel 上「日付のみ」(時刻なし表示) に整ごる
STAGE1_SHEET_DATEONLY_HEADERS = frozenset(
    {
        TASK_COL_ANSWER_DUE,
        TASK_COL_SPECIFIED_DUE,
        TASK_COL_RAW_INPUT_DATE,
        PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
    }
)


def _result_font(**kwargs):
    """結果ブック用 Font（呼び出し坴は name/size 等を指定）。"""
    return Font(**kwargs)


def _output_book_font(bold=False):
    return _result_font(bold=bold)


def _apply_output_font_to_result_sheet(ws):
    """結果_* のごうガント以外坑け: 既定フォント・1行目太字のみ（列幅は VBA AutoFit）。"""
    base = _output_book_font(bold=False)
    hdr = _output_book_font(bold=True)
    mr, mc = ws.max_row or 1, ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
        for cell in row:
            cell.font = base
    for cell in ws[1]:
        cell.font = hdr


def _apply_excel_date_columns_date_only_display(path, sheet_name, header_names=None):
    """openpyxl: 指定ヘッダー列を yyyy/mm/dd の日付表示にれる（時刻を表示しない）。"""
    from openpyxl import load_workbook

    headers = header_names or STAGE1_SHEET_DATEONLY_HEADERS
    if _workbook_should_skip_openpyxl_io(path):
        logging.info(
            "日付列表示整形: ブックに「%s」があるため、openpyxl での処理をスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    wb = load_workbook(path)
    try:
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[int(sheet_name)]
        cmap = {}
        for cell in ws[1]:
            if cell.value is not None:
                cmap[str(cell.value).strip()] = cell.column
        fmt = "yyyy/mm/dd"
        for h in headers:
            col = cmap.get(h)
            if not col:
                continue
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                v = c.value
                if v is None:
                    continue
                if isinstance(v, datetime):
                    c.value = v.date()
                elif isinstance(v, date):
                    pass
                else:
                    try:
                        d0 = pd.to_datetime(v, errors="coerce")
                        if pd.isna(d0):
                            continue
                        c.value = d0.date()
                    except Exception:
                        continue
                c.number_format = fmt
        wb.save(path)
    finally:
        wb.close()


def _extract_data_extraction_datetime():
    """
    `加工計画DATA` シートから配台基準日時を取得する。
    列「データ抽出時間」の先頭非空値を最優先。次に列「抽出時間」。列が無い・有効値が無いときは「データ抽出日」を試す。

    Returns:
        tuple[datetime | None, str | None]: (日時, 採用した列名)。両方 None のときは現在時刻フォールバック。
    """

    def _first_valid_dt_from_series(series) -> datetime | None:
        first = None
        for v in series:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            first = v
            break
        if first is None:
            return None
        dt = pd.to_datetime(first, errors="coerce")
        if pd.isna(dt):
            return None
        if isinstance(dt, pd.Timestamp):
            return dt.to_pydatetime()
        return dt if isinstance(dt, datetime) else None

    try:
        if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
            return None, None
        df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=TASKS_SHEET_NAME)
        df.columns = df.columns.str.strip()
        for col_name in (
            TASK_COL_DATA_EXTRACTION_TIME,
            TASK_COL_EXTRACTION_TIME,
            TASK_COL_DATA_EXTRACTION_DT,
        ):
            if col_name not in df.columns:
                continue
            dt = _first_valid_dt_from_series(df[col_name])
            if dt is not None:
                return dt, col_name
        return None, None
    except Exception:
        return None, None


def _extract_data_extraction_datetime_str():
    """
    `加工計画DATA` から基準日時を文字列化する（データ抽出時間→抽出時間→データ抽出日）。
    """
    try:
        dt, _ = _extract_data_extraction_datetime()
        if dt is None:
            return "—"
        return dt.strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        return "—"


def _parse_equipment_gantt_meta_line_data_extract_display(meta_line) -> str | None:
    """
    設備ガント系シートのメタ行（行2付近）から「データ抽出」の表示文字列を取り出す。
    ``_write_results_equipment_gantt_sheet`` の ``meta_line`` 形式と整合させる。
    """
    if meta_line is None:
        return None
    s = str(meta_line).replace("\r", "").replace("\n", "")
    needle = "　・　データ抽出　"
    pos = s.find(needle)
    if pos < 0:
        return None
    rest = s[pos + len(needle) :]
    end_mark = "　・　マスタ"
    end_pos = rest.find(end_mark)
    if end_pos < 0:
        return None
    out = rest[:end_pos].strip()
    return out if out else None


def _read_existing_equipment_gantt_data_extract_display(path: str, sheet_name: str) -> str | None:
    """
    既存の設備ガント xlsx のメタ行からデータ抽出表示を読む（失敗時は None）。
    メタは ``title_start_col=4`` の結合先頭セル（通常 D2）。
    """
    if not path or not os.path.isfile(path):
        return None
    try:
        wb = load_workbook(path, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            val = ws.cell(row=2, column=4).value
            return _parse_equipment_gantt_meta_line_data_extract_display(val)
        finally:
            wb.close()
    except Exception:
        return None


def _weekday_jp(d):
    return "月睫水木金土日"[d.weekday()]


# ガントの作業ポー：いうれも明るい地色＋黒文字は読ゝるトーン（モノクロ坰刷でも濃淡で識別しやれい）
_GANTT_BAR_FILLS_PRINT_SAFE = (
    "E8E8E8",
    "D8E4EF",
    "E6E2DB",
    "DEEADF",
    "E8E0E8",
    "EAE8D8",
    "DDE6EA",
    "E5DCE5",
)

# 実績ポー用（計画と並きでもモノクロで区別しやれいトーン）
_GANTT_BAR_FILLS_ACTUAL = (
    "D4E4D4",
    "C9DDE8",
    "DED8CC",
    "D2E5CD",
    "DAD2D9",
    "E0DCCF",
    "CDE2E8",
    "DCD2DC",
)

# 設備ガント: 日次始業準備（machine_daily_startup）の帯色（黄色系・モノトーン既定）
_GANTT_DAILY_STARTUP_FILL = "FFEB9C"


def _gantt_hsv_to_rgb_u8(h01: float, s: float, v: float) -> tuple[int, int, int]:
    """h01∈[0,1), s・v∈[0,1] を sRGB 0..255 に。"""
    h = (float(h01) % 1.0) * 6.0
    c = float(v) * float(s)
    x = c * (1.0 - abs((h % 2.0) - 1.0))
    m = float(v) - c
    if h < 1.0:
        rp, gp, bp = c, x, 0.0
    elif h < 2.0:
        rp, gp, bp = x, c, 0.0
    elif h < 3.0:
        rp, gp, bp = 0.0, c, x
    elif h < 4.0:
        rp, gp, bp = 0.0, x, c
    elif h < 5.0:
        rp, gp, bp = x, 0.0, c
    else:
        rp, gp, bp = c, 0.0, x
    r = int((rp + m) * 255.0)
    g = int((gp + m) * 255.0)
    b = int((bp + m) * 255.0)
    return max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))


def _gantt_fullcolor_fill_hex_for_task_id(task_id, *, is_actual: bool) -> str:
    """依頼NO（task_id）ごとに色相を固定。実績行は色相をずらして計画と区別。"""
    hx = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    hue01 = (int(hx[0:8], 16) % 360) / 360.0
    if is_actual:
        hue01 = (hue01 + 47.0 / 360.0) % 1.0
    s = 0.36 + (int(hx[8:12], 16) % 26) / 100.0
    v = 0.80 + (int(hx[12:16], 16) % 16) / 100.0
    r, g, b = _gantt_hsv_to_rgb_u8(hue01, s, v)
    return f"{r:02X}{g:02X}{b:02X}"


def _gantt_daily_startup_fill_hex() -> str:
    if _gantt_color_mode_full():
        return "FFC107"
    return _GANTT_DAILY_STARTUP_FILL


def _gantt_bar_fill_for_task_id(task_id):
    """依頼NOごとに1色（RRGGBB）。full 時はHSV、monotone 時は淡色パレット。"""
    if _gantt_color_mode_full():
        return _gantt_fullcolor_fill_hex_for_task_id(task_id, is_actual=False)
    h = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    i = int(h[:8], 16) % len(_GANTT_BAR_FILLS_PRINT_SAFE)
    return _GANTT_BAR_FILLS_PRINT_SAFE[i]


def _gantt_bar_fill_actual_for_task_id(task_id):
    if _gantt_color_mode_full():
        return _gantt_fullcolor_fill_hex_for_task_id(task_id, is_actual=True)
    h = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    i = int(h[:8], 16) % len(_GANTT_BAR_FILLS_ACTUAL)
    return _GANTT_BAR_FILLS_ACTUAL[i]


# ガント時刻セル（結合帯の先頭セル）: 毎セグメント new しない
_GANTT_TIMELINE_CELL_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=False,
    shrink_to_fit=False,
    indent=1,
)


def _gantt_timeline_label_alignment(*, single_slot: bool) -> Alignment:
    """
    ガント帯のラベル用配置。
    1スロット幅のみの帯では列幅が狭く見切れやすいため shrink_to_fit でセル内に収める。
    複数スロット続く帯では shrink せず、空セルへはみ出して表示しやすくする（Excel の表示特性）。
    """
    return Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=bool(single_slot),
        indent=1,
    )
# タスク帯の色はパレット有限なので PatternFill を hex 単位で共有（openpyxl のスタイル展開コスト削減）
_GANTT_TASK_PATTERN_FILL_BY_HEX: dict[str, PatternFill] = {}


def _gantt_cached_pattern_fill(hex_rrggbb: str) -> PatternFill:
    fi = _GANTT_TASK_PATTERN_FILL_BY_HEX.get(hex_rrggbb)
    if fi is None:
        fi = PatternFill(fill_type="solid", start_color=hex_rrggbb, end_color=hex_rrggbb)
        _GANTT_TASK_PATTERN_FILL_BY_HEX[hex_rrggbb] = fi
    return fi


def _gantt_slot_state_tuple(evlist, slot_start, slot_mins, task_fill_fn=None):
    """
    10 分枠 [slot_start, slot_end) の 1 マス分の状態。
    ('idle',) | ('break',) | ('daily_startup', fill_hex) | ('task', tid, fill_hex, pct)

    結果_設備毎の時間割・結果_設備毎の時間割_機械名毎（``_build_equipment_schedule_*``）と同様に、
    枠と重なるイベントの選定に ``_eq_grid_best_overlapping_event_for_cell``、
    休憩判定の参照時刻に ``_eq_grid_overlap_sample_t``（枠∩イベント区間の中点）を用いる。
    従来の「枠中点を含む最初のイベント」のみを見る方式では、準備と加工が重なる枠で
    時間割は加工を出すのにガントが準備側へ寄り、依頼NO シェイプが欠けることがあった。
    """
    fill_fn = task_fill_fn or _gantt_bar_fill_for_task_id
    slot_end = slot_start + timedelta(minutes=float(slot_mins))
    slot_mid = slot_start + timedelta(minutes=float(slot_mins) / 2.0)
    active = _eq_grid_best_overlapping_event_for_cell(evlist, slot_start, slot_end)
    if active is None:
        return ("idle",)
    if _timeline_event_kind(active) == TIMELINE_EVENT_MACHINE_DAILY_STARTUP:
        return ("daily_startup", _gantt_daily_startup_fill_hex())
    sample_t = _eq_grid_overlap_sample_t(active, slot_start, slot_end, slot_mid)
    if any(b_s <= sample_t < b_e for b_s, b_e in active.get("breaks") or ()):
        return ("break",)
    tid = str(active["task_id"])
    gh = fill_fn(active["task_id"])
    pct = None
    try:
        # 「マクロ実行時点」の完了率を優先（pct_macro を timeline_event に挝たせる）
        if active.get("pct_macro") is not None:
            pct = int(round(parse_float_safe(active.get("pct_macro"), 0.0)))
            pct = max(0, min(100, pct))
        else:
            # フェイルセーフ（従来の擬似進杗計算）
            tot = parse_float_safe(active.get("total_units"), 0.0)
            done = parse_float_safe(active.get("already_done_units"), 0.0) + parse_float_safe(
                active.get("units_done"), 0.0
            )
            if tot > 0:
                pct = max(0, min(100, int(round((done / tot) * 100))))
    except Exception:
        pct = None
    return ("task", tid, gh, pct)


def _gantt_timeline_same_segment(st_a, st_b) -> bool:
    """結合セグメント境界判定（毎スロット tuple を割り当でない）。"""
    if st_a[0] != st_b[0]:
        return False
    if st_a[0] == "idle" or st_a[0] == "break":
        return True
    # daily_startup: [1]=fill / task: [1]=task_id
    return st_a[1] == st_b[1]


def _paint_gantt_timeline_row_merged(
    ws,
    row,
    n_fixed,
    slots,
    slot_mins,
    evlist,
    idle_fill,
    break_fill,
    gantt_label_font,
    grid_border,
    task_fill_fn=None,
    label_font=None,
    shape_label_specs: list | None = None,
    label_italic: bool = False,
    shape_day_key: str | None = None,
):
    """
    時間軸を塗り分けたうえで、同一状態が連続するセルを横結合し帯状のバーにする。
    （細マス単体の塗りではなく slot_mins 刻み＋同一状態のセル結合で、帯状のバーとして表現する）
    shape_label_specs に list を渡すと、タイムライン上の文字はセルに入れず後段（xlwings）で
    角丸シェイプとして追加するための座標・文言を蓄積する。
    shape_day_key に ISO 日付文字列等を渡すと、後段で日単位の画像化（フラット化）に利用する。
    """
    bar_label_font = label_font or gantt_label_font
    n_slots = len(slots)
    states = []
    for slot_start in slots:
        states.append(_gantt_slot_state_tuple(evlist, slot_start, slot_mins, task_fill_fn))
    tcol0 = n_fixed + 1
    i = 0
    while i < n_slots:
        st0 = states[i]
        j = i + 1
        while j < n_slots and _gantt_timeline_same_segment(st0, states[j]):
            j += 1
        col_s = tcol0 + i
        col_e = tcol0 + j - 1
        single_slot_segment = col_s == col_e
        for col in range(col_s, col_e + 1):
            c = ws.cell(row=row, column=col)
            c.border = grid_border
            c.alignment = _GANTT_TIMELINE_CELL_ALIGNMENT
            if st0[0] == "idle":
                c.fill = idle_fill
                c.value = None
            elif st0[0] == "break":
                c.fill = break_fill
                c.value = None
            elif st0[0] == "daily_startup":
                _, gh_ds = st0
                c.fill = _gantt_cached_pattern_fill(gh_ds)
                if col == col_s:
                    _ds_txt = "(日次始業準備)"
                    if shape_label_specs is not None:
                        shape_label_specs.append(
                            {
                                "row": row,
                                "col_s": col_s,
                                "col_e": col_e,
                                "text": _ds_txt,
                                "italic": bool(label_italic),
                                "fill_hex": str(gh_ds),
                                "day_key": shape_day_key or "",
                            }
                        )
                        c.value = None
                    else:
                        c.value = _ds_txt
                        c.font = bar_label_font
                        c.alignment = _gantt_timeline_label_alignment(
                            single_slot=single_slot_segment
                        )
                else:
                    c.value = None
            else:
                _, tid, gh, pct = st0
                c.fill = _gantt_cached_pattern_fill(gh)
                if col == col_s:
                    tid_s = str(tid or "").strip()
                    _lbl = f"{tid_s} {pct}%" if pct is not None else tid_s
                    if shape_label_specs is not None:
                        if tid_s:
                            shape_label_specs.append(
                                {
                                    "row": row,
                                    "col_s": col_s,
                                    "col_e": col_e,
                                    "text": _lbl,
                                    "italic": bool(label_italic),
                                    "fill_hex": str(gh),
                                    "member_labels": _gantt_member_labels_for_task(
                                        evlist, tid_s
                                    ),
                                    "day_key": shape_day_key or "",
                                }
                            )
                        c.value = None
                    else:
                        c.value = _lbl
                        c.font = bar_label_font
                        c.alignment = _gantt_timeline_label_alignment(
                            single_slot=single_slot_segment
                        )
                else:
                    c.value = None
        i = j


def _time_intervals_overlap_half_open(
    a_start: time, a_end: time, b_start: time, b_end: time
) -> bool:
    """半開区間 [a_start, a_end) と [b_start, b_end) は重なるか（同一日内）。"""

    def _sec(t: time) -> int:
        return t.hour * 3600 + t.minute * 60 + t.second

    return _sec(a_start) < _sec(b_end) and _sec(a_end) > _sec(b_start)


def _parse_equipment_schedule_time_band_cell(val) -> tuple[time | None, time | None]:
    """結果_設備毎の時間割「日時帯」セル（例 08:45-09:00）を解釈。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None, None
    s = str(val).strip()
    if not s or "■" in s:
        return None, None
    for sep in ("-", "＝", "~", "〜"):
        if sep in s:
            left, right = s.split(sep, 1)
            left = left.strip().replace("：", ":")
            right = right.strip().replace("：", ":")
            t0 = parse_time_str(left, None)
            t1 = parse_time_str(right, None)
            if t0 is not None and t1 is not None and t0 < t1:
                return t0, t1
            return None, None
    return None, None


def _apply_equipment_schedule_outside_regular_fill(
    ws, reg_start: time, reg_end: time
) -> None:
    """「日時帯」列で定常 [reg_start, reg_end) と重ならない行のセルに着色。"""
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_OUTSIDE_REGULAR_TIME_FILL,
        end_color=RESULT_OUTSIDE_REGULAR_TIME_FILL,
    )
    col_idx = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_idx = i
            break
    if col_idx is None:
        return
    mr = ws.max_row or 1
    for r in range(2, mr + 1):
        cell = ws.cell(row=r, column=col_idx)
        t0, t1 = _parse_equipment_schedule_time_band_cell(cell.value)
        if t0 is None or t1 is None:
            continue
        if not _time_intervals_overlap_half_open(t0, t1, reg_start, reg_end):
            cell.fill = fill


def _apply_equipment_schedule_prep_cleanup_fill(ws) -> None:
    """
    設備列（進度列を除く）で、表示に「日次始業準備」が含まれるセルを薄緑にする。
    結果_設備毎の時間割 の equip セル用（日時帯列は変更しない）。
    """
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_DISPATCHED_REQUEST_FILL,
        end_color=RESULT_DISPATCHED_REQUEST_FILL,
    )
    markers = ("(日次始業準備)",)
    col_tb = None
    equip_cols: list[int] = []
    for i, c in enumerate(ws[1], start=1):
        if c.value is None:
            continue
        h = str(c.value).strip()
        if h == "日時帯":
            col_tb = i
            continue
        if h.endswith("進度"):
            continue
        equip_cols.append(i)
    if col_tb is None or not equip_cols:
        return
    mr = ws.max_row or 1
    for r in range(2, mr + 1):
        for ci in equip_cols:
            cell = ws.cell(row=r, column=ci)
            val = cell.value
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip().replace("\r", "").replace("\n", "")
            if any(m in s for m in markers):
                cell.fill = fill


def _parse_equipment_schedule_day_header_date(val) -> date | None:
    """日付見出し行「■ YYYY/MM/DD … ■」から日付を得る。"""
    if val is None:
        return None
    s = str(val).strip()
    m = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _machine_calendar_intervals_for_equipment_line(
    day_blocks: dict[str, list[tuple[datetime, datetime]]],
    eq_line: str,
    day_d: date,
) -> list[tuple[datetime, datetime]]:
    """当日・当該設備列キーに対応れる機械カレンダー占有区間（工場稼働枠でクリップ済み）。"""
    if not day_blocks:
        return []
    ek = str(eq_line or "").strip()
    blocks: list[tuple[datetime, datetime]] | None = None
    if ek in day_blocks:
        blocks = day_blocks[ek]
    else:
        pk = (
            _normalize_equipment_match_key(ek.split("+", 1)[1])
            if "+" in ek
            else _normalize_equipment_match_key(ek)
        )
        if pk and pk in day_blocks:
            blocks = day_blocks[pk]
        else:
            nk = _normalize_equipment_match_key(ek)
            for k, iv in day_blocks.items():
                if _normalize_equipment_match_key(str(k)) == nk:
                    blocks = iv
                    break
    if not blocks:
        return []
    w0 = datetime.combine(day_d, DEFAULT_START_TIME)
    w1 = datetime.combine(day_d, DEFAULT_END_TIME)
    return _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)


def _apply_equipment_schedule_machine_calendar_fill(
    ws,
    equipment_list: list,
    calendar_blocks_by_date: dict[date, dict[str, list[tuple[datetime, datetime]]]],
) -> None:
    """
    結果_設備毎の時間割: 機械カレンダー占有と重なる設備セル（進度列以外）を紫色で塗る。
    10 分枠の半開区間 [slot_start, slot_end) と占有 [bs, be) は重ならない対象。
    """
    if not calendar_blocks_by_date or not equipment_list:
        return
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_MACHINE_CALENDAR_BLOCK_FILL,
        end_color=RESULT_MACHINE_CALENDAR_BLOCK_FILL,
    )
    col_tb = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_tb = i
            break
    if col_tb is None:
        return
    eq_col_indices: list[int] = [
        col_tb + 1 + 2 * idx for idx in range(len(equipment_list))
    ]
    mr = ws.max_row or 1
    current_d: date | None = None
    for r in range(2, mr + 1):
        tb_cell = ws.cell(row=r, column=col_tb)
        tv = tb_cell.value
        d_hdr = _parse_equipment_schedule_day_header_date(tv)
        if d_hdr is not None:
            current_d = d_hdr
            continue
        t0, t1 = _parse_equipment_schedule_time_band_cell(tv)
        if t0 is None or t1 is None or current_d is None:
            continue
        slot_a = datetime.combine(current_d, t0)
        slot_b = datetime.combine(current_d, t1)
        if slot_b <= slot_a:
            continue
        day_blocks = calendar_blocks_by_date.get(current_d)
        if not day_blocks:
            continue
        for col_idx, eq_line in zip(eq_col_indices, equipment_list):
            blocks_c = _machine_calendar_intervals_for_equipment_line(
                day_blocks, eq_line, current_d
            )
            if not blocks_c:
                continue
            for bs, be in blocks_c:
                if slot_a < be and bs < slot_b:
                    ws.cell(row=r, column=col_idx).fill = fill
                    break


def _apply_equipment_by_machine_dispatched_request_fill(ws) -> None:
    """
    結果_設備毎の時間割_機械名毎の機械名列で」依頼NOは入っているセルに薄緑を付与れる。
    「（休憩）」のみのセルは対象外。見出し行・日時帯列は変更しない。
    """
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_DISPATCHED_REQUEST_FILL,
        end_color=RESULT_DISPATCHED_REQUEST_FILL,
    )
    col_tb = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_tb = i
            break
    if col_tb is None:
        return
    mr = ws.max_row or 1
    mc = ws.max_column or col_tb
    for r in range(2, mr + 1):
        for c in range(col_tb + 1, mc + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip().replace("\r", "").replace("\n", "")
            if not s or s == "（休憩）":
                continue
            cell.fill = fill


def _equipment_gantt_fills_by_machine_name(equipment_list) -> dict[str, PatternFill]:
    """
    結果_設備ガントの固定列（B〜D」A は日付縦結合）用。equipment_list 内の機械名（+ 無し時は行全体を機械名）の出睾順で
    淡色を割り当てて、同一機械名は常に同じ PatternFill を共有する。
    """
    order: list[str] = []
    seen: set[str] = set()
    for eq in equipment_list or []:
        _, mn = _split_equipment_line_process_machine(eq)
        key = (mn or "").strip() or "—"
        if key not in seen:
            seen.add(key)
            order.append(key)
    palette = (
        RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS_FULL
        if _gantt_color_mode_full()
        else RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS
    )
    if not palette:
        fb = "F5F5F5"
        return {k: PatternFill(fill_type="solid", start_color=fb, end_color=fb) for k in order}
    out: dict[str, PatternFill] = {}
    n = len(palette)
    for i, key in enumerate(order):
        hx = palette[i % n]
        out[key] = PatternFill(fill_type="solid", start_color=hx, end_color=hx)
    return out


def _write_results_equipment_gantt_sheet(
    writer,
    timeline_events,
    equipment_list,
    sorted_dates,
    attendance_data,
    data_extract_dt_str,
    base_now_dt=None,
    actual_timeline_events=None,
    regular_shift_times: tuple[time | None, time | None] | None = None,
    *,
    plan_rows: bool = True,
    chart_title: str | None = None,
    sheet_name_override: str | None = None,
):
    """
    結果_設備毎の時間割と同一データ源（timeline_events）に基づき、
    設備×横軸時間のガンチャート風シートを追加する。
    横軸は GANTT_TIMELINE_SLOT_MINUTES 分刻み。同一状態の連続は帯状に塗分けする。
    actual_timeline_events があれば設備ごとに「実績」行を計画行の下へ追加する。
    plan_rows=False のときは計画行を出さず actual_timeline_events のみを各行に描画する（実績明細ガント用）。
    GANTT_TIMELINE_SHAPE_LABELS が有効なとき、タイムライン上の依頼NO 等はセルに書かず
    角丸シェイプ用の仕様 dict の list と、日ブロック境界の list を返す（保存後に xlwings で描画・画像化）。
    無効時は ([], []) を返す。
    """
    sheet_nm = sheet_name_override or RESULT_SHEET_GANTT_NAME
    if not plan_rows:
        if not actual_timeline_events:
            logging.info(
                "設備ガント（%s）: 実績のみモードですがイベントが空のためシートを作成しません。",
                sheet_nm,
            )
            return [], []
    wb = writer.book
    if sheet_name_override:
        try:
            insert_at = wb.sheetnames.index(RESULT_SHEET_GANTT_NAME) + 1
        except ValueError:
            try:
                insert_at = wb.sheetnames.index("結果_設備毎の時間割") + 1
            except ValueError:
                insert_at = len(wb.sheetnames)
    else:
        try:
            insert_at = wb.sheetnames.index("結果_設備毎の時間割") + 1
        except ValueError:
            insert_at = len(wb.sheetnames)
    ws = wb.create_sheet(sheet_nm, insert_at)
    try:
        ws.sheet_properties.tabColor = (
            "1976D2" if _gantt_color_mode_full() else "7F7F7F"
        )
    except Exception:
        pass

    events_by_date = defaultdict(list)
    for e in timeline_events:
        events_by_date[e["date"]].append(e)

    show_actual_rows = bool(actual_timeline_events)
    actual_events_by_date = defaultdict(list)
    if show_actual_rows:
        for e in actual_timeline_events:
            actual_events_by_date[e["date"]].append(e)

    slot_mins = GANTT_TIMELINE_SLOT_MINUTES
    _g_cf = _gantt_color_mode_full()
    hdr_font = _result_font(bold=True, color="000000", size=12)
    hdr_fill = PatternFill(
        fill_type="solid",
        start_color=("BBDEFB" if _g_cf else "D9D9D9"),
        end_color=("BBDEFB" if _g_cf else "D9D9D9"),
    )
    hdr_time_font = _result_font(bold=True, color="000000", size=11)
    title_font = _result_font(bold=True, size=24, color="1A1A1A")
    title_fill = PatternFill(
        fill_type="solid",
        start_color=("E3F2FD" if _g_cf else "DDDDDD"),
        end_color=("E3F2FD" if _g_cf else "DDDDDD"),
    )
    meta_font = _result_font(size=11, color="333333")
    meta_fill = PatternFill(
        fill_type="solid",
        start_color=("F1F8E9" if _g_cf else "F3F3F3"),
        end_color=("F1F8E9" if _g_cf else "F3F3F3"),
    )
    day_banner_font = _result_font(bold=True, size=13, color="1A1A1A")
    day_banner_fill = PatternFill(
        fill_type="solid",
        start_color=("C5E1A5" if _g_cf else "D0D0D0"),
        end_color=("C5E1A5" if _g_cf else "D0D0D0"),
    )
    accent_left = Side(style="thick", color="2B2B2B")
    banner_sep = Side(style="thin", color="7A7A7A")
    thin = Side(style="thin", color=("5C6BC0" if _g_cf else "666666"))
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    idle_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    break_fill = PatternFill(
        fill_type="solid",
        start_color=("90CAF9" if _g_cf else "B8B8B8"),
        end_color=("90CAF9" if _g_cf else "B8B8B8"),
    )
    gantt_label_font = _result_font(size=10, bold=True, color="000000")
    gantt_label_font_actual = _result_font(size=10, bold=True, color="000000", italic=True)
    _outside_hex = (
        "FFCCBC" if _g_cf else str(RESULT_OUTSIDE_REGULAR_TIME_FILL or "FCE4D6")
    )
    hdr_fill_outside_regular = PatternFill(
        fill_type="solid",
        start_color=_outside_hex,
        end_color=_outside_hex,
    )
    rs, re_ = (regular_shift_times or (None, None))

    # 横軸（slot_mins 刻み）は日付で共通のため、slot_times を先に確定
    base_dt = base_now_dt if isinstance(base_now_dt, datetime) else datetime.now()
    dummy_d = sorted_dates[0] if sorted_dates else base_dt.date()
    d_start0 = datetime.combine(dummy_d, DEFAULT_START_TIME)
    d_end0 = datetime.combine(dummy_d, DEFAULT_END_TIME)
    slot_times = []
    t0 = d_start0
    while t0 < d_end0:
        slot_times.append(t0.time())
        t0 += timedelta(minutes=slot_mins)

    n_slots = len(slot_times)
    n_fixed = 4  # A=日付（日ブロック内で縦結合）/ B〜D=機械名・工程名・タスク概覝
    last_col = n_fixed + n_slots
    gantt_shape_label_specs: list[dict] = []
    gantt_timeline_day_blocks: list[dict] = []
    _use_gantt_shape_labels = GANTT_TIMELINE_SHAPE_LABELS
    fills_by_mach = _equipment_gantt_fills_by_machine_name(equipment_list)
    fb_gantt = "ECEFF1" if _g_cf else "F5F5F5"
    fill_gantt_fallback = PatternFill(fill_type="solid", start_color=fb_gantt, end_color=fb_gantt)

    # タイトル＆日時（ページ上部）
    create_ts = base_dt.strftime("%Y/%m/%d %H:%M:%S")
    master_path = os.path.join(os.getcwd(), MASTER_FILE) if MASTER_FILE else ""

    def _fmt_mtime(p):
        try:
            if p and os.path.exists(p):
                return datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y/%m/%d %H:%M:%S")
        except Exception:
            pass
        return "—"

    master_mtime = _fmt_mtime(master_path)

    # 行1の A〜B はコンボ、C は更新ボタン（マクロ）。タイトル・メタ行は D 列から全幅結合。
    title_start_col = 4
    row = 1
    ws.merge_cells(
        start_row=row, start_column=title_start_col, end_row=row, end_column=last_col
    )
    _title_main = (
        chart_title if chart_title is not None else "湖南工場 加工計画"
    )
    tcell = ws.cell(row=row, column=title_start_col, value=_title_main)
    tcell.font = title_font
    tcell.fill = title_fill
    # 結合セルでも左端から表示（縮尝・折り返しなし）
    tcell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=False,
        indent=1,
    )
    tcell.border = Border(left=accent_left, bottom=banner_sep)
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(
        start_row=row, start_column=title_start_col, end_row=row, end_column=last_col
    )
    meta_line = (
        f"作成　{create_ts}"
        f"　・　データ抽出　{data_extract_dt_str or '—'}"
        f"　・　マスタ（master.xlsm）　{master_mtime}"
    )
    mtop = ws.cell(row=row, column=title_start_col, value=meta_line)
    mtop.font = meta_font
    mtop.fill = meta_fill
    mtop.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=1,
        wrap_text=False,
        shrink_to_fit=False,
    )
    mtop.border = Border(left=accent_left, bottom=banner_sep)
    ws.row_dimensions[row].height = 26
    row += 1

    dates_to_show: list = []
    for d0 in sorted_dates:
        evs0 = events_by_date.get(d0, []) if plan_rows else []
        a_evs0 = actual_events_by_date.get(d0, []) if show_actual_rows else []
        if d0 not in attendance_data:
            is_anyone_working0 = False
        else:
            is_anyone_working0 = any(
                attendance_data[d0][mm]["is_working"]
                for mm in attendance_data[d0]
                if mm in attendance_data[d0]
            )
        if not evs0 and not a_evs0 and not is_anyone_working0:
            continue
        dates_to_show.append(d0)

    hdr_row = row
    fixed_hdr = ["日付", "機械名", "工程名", "タスク概覝"]
    for ci, h in enumerate(fixed_hdr, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    slots_hdr = [datetime.combine(dummy_d, tm) for tm in slot_times]
    for si, st in enumerate(slots_hdr):
        c = ws.cell(row=hdr_row, column=n_fixed + 1 + si, value=st.strftime("%H:%M"))
        c.font = hdr_time_font
        slot_end_t = (st + timedelta(minutes=slot_mins)).time()
        hdr_use = hdr_fill
        if rs is not None and re_ is not None:
            if not _time_intervals_overlap_half_open(st.time(), slot_end_t, rs, re_):
                hdr_use = hdr_fill_outside_regular
        c.fill = hdr_use
        c.alignment = Alignment(horizontal="center", vertical="bottom", textRotation=90)
    ws.row_dimensions[hdr_row].height = float(GANTT_HDR_ROW_HEIGHT_PT)
    # 先頭データ行の左上＝時刻列先頭（E4）で窓枠固定（行1〜3・列A〜Dを固定）
    ws.freeze_panes = f"{get_column_letter(n_fixed + 1)}{hdr_row + 1}"
    row = hdr_row + 1

    # 日と日の間の区切り（真っ黒だと「日付ブロックの下端」と誤解されやすいため薄グレー）
    sep_fill = PatternFill(fill_type="solid", start_color="D0D0D0", end_color="D0D0D0")
    no_border = Border()

    # 印刷: 1 日ごとの手動改ページ用（各日のデータ先頭行＝機械行の開始）
    gantt_day_first_rows: list[int] = []

    for di, d in enumerate(dates_to_show):
        evs = events_by_date.get(d, [])
        a_evs_day = actual_events_by_date.get(d, []) if show_actual_rows else []

        slots = [datetime.combine(d, tm) for tm in slot_times]

        # 設備時間割と同じく ev['machine'] と equipment_list の表記ゆれを正規化して対応づける。
        # by_dm[d].get(eq) のみだとキー不一致の行が空になり、機械名毎シートだけに依頼NOが出ることがある。
        machine_to_events = defaultdict(list)
        for ev in evs:
            machine_to_events[ev["machine"]].append(ev)
        for _k_m, _evl in machine_to_events.items():
            _evl.sort(
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
            )
        machine_to_events_a = None
        if show_actual_rows:
            machine_to_events_a = defaultdict(list)
            for ev in a_evs_day:
                machine_to_events_a[ev["machine"]].append(ev)
            for _k_m2, _evl2 in machine_to_events_a.items():
                _evl2.sort(
                    key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
                )

        day_start = row
        gantt_day_first_rows.append(day_start)
        for eq in equipment_list:
            proc_nm, mach_nm = _split_equipment_line_process_machine(eq)
            mk_key = (mach_nm or "").strip() or "—"
            lab_fill = fills_by_mach.get(mk_key) or fill_gantt_fallback
            evlist = _eq_grid_events_for_equipment_column(machine_to_events, eq)
            if plan_rows:
                if evlist:
                    tids: list[str] = []
                    seen_tid: set[str] = set()
                    for e in evlist:
                        tid = str(e.get("task_id") or "").strip()
                        if tid and tid not in seen_tid:
                            seen_tid.add(tid)
                            tids.append(tid)
                    task_sum = " ".join(tids) if tids else "—"
                else:
                    task_sum = "—"

                c1 = ws.cell(row=row, column=2, value=mach_nm if mach_nm else "—")
                c2 = ws.cell(row=row, column=3, value=proc_nm if proc_nm else "—")
                c3 = ws.cell(row=row, column=4, value=task_sum)
                for c in (c1, c2, c3):
                    c.font = _result_font(size=12, color="000000")
                    c.fill = lab_fill
                    c.border = grid_border
                c1.font = _result_font(size=12, bold=True, color="000000")
                c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                _paint_gantt_timeline_row_merged(
                    ws,
                    row,
                    n_fixed,
                    slots,
                    slot_mins,
                    evlist,
                    idle_fill,
                    break_fill,
                    gantt_label_font,
                    grid_border,
                    shape_label_specs=gantt_shape_label_specs if _use_gantt_shape_labels else None,
                    label_italic=False,
                    shape_day_key=d.isoformat() if _use_gantt_shape_labels else None,
                )

                ws.row_dimensions[row].height = float(GANTT_MACHINE_ROW_HEIGHT_PT)
                row += 1

            if show_actual_rows:
                evlist_a = (
                    _eq_grid_events_for_equipment_column(machine_to_events_a, eq)
                    if machine_to_events_a is not None
                    else []
                )
                if evlist_a:
                    tids_a: list[str] = []
                    seen_aid: set[str] = set()
                    for e_a in evlist_a:
                        tid = str(e_a.get("task_id") or "").strip()
                        if not tid or tid in seen_aid:
                            continue
                        seen_aid.add(tid)
                        if not plan_rows:
                            opv = str(e_a.get("op") or "").strip()
                            subv = str(e_a.get("sub") or "").strip()
                            who_parts: list[str] = []
                            if opv:
                                who_parts.append(opv)
                            if subv:
                                for seg in subv.split(","):
                                    t = seg.strip()
                                    if t:
                                        who_parts.append(t)
                            who_show: list[str] = []
                            who_seen: set[str] = set()
                            for p in who_parts:
                                k = unicodedata.normalize("NFKC", p)
                                if k in who_seen:
                                    continue
                                who_seen.add(k)
                                who_show.append(p)
                            if who_show:
                                tids_a.append(f"{tid}（{'・'.join(who_show)}）")
                            else:
                                tids_a.append(tid)
                        else:
                            tids_a.append(tid)
                    task_sum_a = " ".join(tids_a) if tids_a else "—"
                else:
                    task_sum_a = "—"

                lab_fill_a = fills_by_mach.get(mk_key) or fill_gantt_fallback

                if mach_nm:
                    act_mach = (
                        f"{mach_nm}（実績明細）"
                        if not plan_rows
                        else f"{mach_nm}（実績）"
                    )
                elif proc_nm:
                    act_mach = "（実績明細）" if not plan_rows else "（実績）"
                else:
                    act_mach = "—"
                ca1 = ws.cell(row=row, column=2, value=act_mach)
                ca2 = ws.cell(row=row, column=3, value=proc_nm if proc_nm else "—")
                ca3 = ws.cell(row=row, column=4, value=task_sum_a)
                for c in (ca1, ca2, ca3):
                    c.font = _result_font(size=12, color="000000")
                    c.fill = lab_fill_a
                    c.border = grid_border
                ca1.font = _result_font(size=12, bold=True, color="000000", italic=True)
                ca1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                ca2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                ca3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                _paint_gantt_timeline_row_merged(
                    ws,
                    row,
                    n_fixed,
                    slots,
                    slot_mins,
                    evlist_a,
                    idle_fill,
                    break_fill,
                    gantt_label_font_actual,
                    grid_border,
                    task_fill_fn=_gantt_bar_fill_actual_for_task_id,
                    label_font=gantt_label_font_actual,
                    shape_label_specs=gantt_shape_label_specs if _use_gantt_shape_labels else None,
                    label_italic=True,
                    shape_day_key=d.isoformat() if _use_gantt_shape_labels else None,
                )

                ws.row_dimensions[row].height = float(GANTT_MACHINE_ROW_HEIGHT_PT)
                row += 1

        day_end = row - 1
        if day_end >= day_start and _use_gantt_shape_labels:
            gantt_timeline_day_blocks.append(
                {
                    "first_row": day_start,
                    "last_row": day_end,
                    "day_key": d.isoformat(),
                    "first_col": n_fixed + 1,
                    "last_col": last_col,
                }
            )
        if day_end >= day_start:
            ws.merge_cells(start_row=day_start, start_column=1, end_row=day_end, end_column=1)
            ban = ws.cell(
                row=day_start,
                column=1,
                value=f"【{d.strftime('%Y/%m/%d')}】",
            )
            ban.font = day_banner_font
            ban.fill = day_banner_fill
            # 縦書き日付は結合ブロックの上寄せ（下寄せだとセル下端に寄って見える）
            ban.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=False,
                textRotation=90,
            )
            ban.border = Border(left=accent_left, top=thin, bottom=thin, right=thin)

        if di < len(dates_to_show) - 1 and day_end >= day_start:
            for cc in range(1, last_col + 1):
                sc = ws.cell(row=row, column=cc)
                sc.value = None
                sc.fill = sep_fill
                sc.border = no_border
            ws.row_dimensions[row].height = 3
            row += 1

    # 凡例は高さ確保のため省略（モノクロ印刷は色の濃淡/セルの枠で識別）
    # 時刻列（E〜）の列幅。マクロ取り込み時は VBA 結果_設備ガント_列幅を設定 と同値に揃える。
    if n_slots > 0:
        gw = float(GANTT_TIMELINE_COLUMN_WIDTH)
        for ci in range(n_fixed + 1, last_col + 1):
            dim = ws.column_dimensions[get_column_letter(ci)]
            dim.width = gw
            # openpyxl 3.1+ では customWidth は width 有無から導出される読み取り専用のため代入しない

    _gantt_scale_override_raw = (os.environ.get("GANTT_PRINT_SCALE_PERCENT", "") or "").strip()
    try:
        # 印刷ページ設定（ガンチャート作成完了時点で付与）
        # 適用順: ⓪タイトル行 → ①用紙・向き → ②余白 → ③横1ページ → ④改ページは try の外で付与（順を変えるとずれやすい）
        # 既定: 横1ページに収める（タイムラインが読める）。縦は自動（日内改ページは行高で緩和）。
        # GANTT_PRINT_SCALE_PERCENT を指定したときのみ固定%（横幅が細くなりやすい）。
        # ⓪ 全ページに繰り返すタイトル行（レポート 1〜3 行目）
        ws.print_title_rows = "1:3"
        # ① A3 横向き
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 8
        # ② 余白「狭い」≒ Excel の Narrow プリセット（単位: インチ）
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        # ③ 横 1 ページに収める
        if _gantt_scale_override_raw:
            _pct = max(10, min(400, int(_gantt_scale_override_raw)))
            ws.page_setup.fitToPage = False
            ws.page_setup.fitToWidth = False
            ws.page_setup.fitToHeight = False
            ws.page_setup.scale = _pct
        else:
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        # タイトル・表をページ左基準に（レポート風）
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False
        ws.print_options.gridLines = False
    except Exception:
        pass

    # 1 日 1 ページ相当: 2 日目以降の各日ブロック先頭の直前に手動の横改ページ（上記ページ設定の後）
    try:
        if len(gantt_day_first_rows) > 1:
            for i in range(1, len(gantt_day_first_rows)):
                ws.row_breaks.append(Break(id=gantt_day_first_rows[i], man=True))
    except Exception:
        pass

    if _use_gantt_shape_labels:
        return gantt_shape_label_specs, gantt_timeline_day_blocks
    return [], []


def row_has_completion_keyword(row):
    """加工完了区分に「完了」の文字は含まれる場合はタスク完了とみなす。"""
    v = row.get(TASK_COL_COMPLETION_FLAG)
    if v is None or pd.isna(v):
        return False
    return "完了" in str(v)


def _planning_completion_flag_cell_is_mikan(v) -> bool:
    """加工完了区分がセル値として「未完」とみなすか（NFKC・前後空白除去）。

    セルが「0:未完」のように区分値とコロンで前置される場合は、**最後のコロン以降**が厳密に
    「未完」のときのみ True（実データで未完区分が数値プレフィックス付きで格納される）。
    「未完了」は末尾が「未完」にならないため False のまま。
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = unicodedata.normalize("NFKC", str(v).strip())
    s = s.replace("\uff1a", ":").replace("：", ":")
    tail = s.rsplit(":", 1)[-1].strip() if ":" in s else s
    return tail == "未完"


def _plan_row_exclude_as_completed_mikan_unprocessed_zero_actual_done_rule(row) -> bool:
    """
    加工計画DATA／配台計画_タスク入力の同一列前提で、次をすべて満たす行は加工済みとみなし配台対象外とする。

    - 「未加工」列があり数値 0（空・列無しは対象外）
    - 「実加工数」が 0 以外
    - 「加工完了区分」が「未完」または「0:未完」形式で末尾が「未完」（「未完了」等は含めない）
    """
    cf_v = row.get(TASK_COL_COMPLETION_FLAG)
    act_v = parse_float_safe(row.get(TASK_COL_ACTUAL_DONE), 0.0)
    unp_v = _optional_unprocessed_m_from_plan_row(row)
    ok_mikan = _planning_completion_flag_cell_is_mikan(cf_v)
    ok_act = abs(act_v) > 1e-12
    ok_unp = unp_v is not None and abs(float(unp_v)) <= 1e-12
    return bool(ok_mikan and ok_act and ok_unp)


def _plan_row_exclude_from_assignment(row) -> bool:
    """
    「配台試行」列はオンなら」しの行は配台キューへ入れう」特別指定_備考の AI 解析行からも除し。

    配台から外れ（真）: 論睆値 True」数値 1」文字列（NFKC 後・尝文字）
      true / 1 / yes / on / y / t / はい / ○ / 〇 / ◝
    配台対象（坽）: 空」None」False」0」no / off / false / いいえ / 坦 等
    上記以外の文字列は坽（配台れる）。チェックボックス連動セルは通常 TRUE/FALSE または 1/0。
    """
    v = row.get(PLAN_COL_EXCLUDE_FROM_ASSIGNMENT)
    if v is True:
        return True
    if v is False:
        return False
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            iv = int(v)
            if iv == 1:
                return True
            if iv == 0:
                return False
        except (TypeError, ValueError):
            pass
    s = unicodedata.normalize("NFKC", str(v).strip()).lower()
    if not s or s in ("nan", "none", "false", "0", "no", "off", "いいえ", "坦"):
        return False
    if s in ("true", "1", "yes", "on", "はい", "y", "t", "○", "〇", "◝"):
        return True
    return False


def _coerce_plan_exclude_column_value_for_storage(v):
    """
    「配台試行」列へ書き込む値を」StringDtype 列でも代入エラーにならない形にしゝごる。
    Excel 取り込みの True / 1 / False / 0 と文字列を保挝し、_plan_row_exclude_from_assignment と整合する。
    """
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if v is True:
        return "yes"
    if v is False:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            iv = int(v)
            if iv == 1:
                return "yes"
            if iv == 0:
                return ""
        except (TypeError, ValueError):
            pass
    return str(v).strip()


def parse_float_safe(val, default=0.0):
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def _optional_unprocessed_m_from_plan_row(row) -> float | None:
    """行の「未加工」セルを数値化。列が無い・空なら None。"""
    if row is None:
        return None
    try:
        idx = row.index  # type: ignore[attr-defined]
    except AttributeError:
        return None
    if TASK_COL_UNPROCESSED not in idx:
        return None
    return _optional_float_unprocessed_column(row.get(TASK_COL_UNPROCESSED))


def _ensure_dataframe_has_unprocessed_column(
    df: pd.DataFrame, *, context_label: str
) -> None:
    """加工計画DATA／配台計画_タスク入力に「未加工」列が無いとき配台を中止する。"""
    if df is None:
        raise PlanningValidationError(
            f"{context_label}: 列「{TASK_COL_UNPROCESSED}」が必須です。"
            "この列が無いため配台処理を中止します。"
        )
    if TASK_COL_UNPROCESSED not in df.columns:
        raise PlanningValidationError(
            f"{context_label}: 列「{TASK_COL_UNPROCESSED}」が必須です。"
            "この列が無いため配台処理を中止します。"
        )


def _plan_row_dispatch_qty_metrics(row):
    """
    1行分の換算数量・未加工に基づき、配台用の残り(m)・済相当(m)・換算数量(100m切上)を返す。

    **未加工列は必須**（シートに列が無い場合は ``load_tasks_df`` / ``load_planning_tasks_df`` で
    ``PlanningValidationError``）。セルが空・数値化できない場合も同様にエラーとする。
    実出来高・実加工数からの済相当フォールバックは行わない。

    未加工に有効数値があるとき:
      ① 未加工 > 0: 済相当m = max(0, 換算数量(raw) - 未加工)、残りm = max(0, 未加工)
      ② 未加工 <= 0: 換算数量(100m切上)を残りm の基準とし、それがロール単位長さ未満ならロール長を採用（最小加工単位=1ロール）。済相当m = 0。
      換算数量の100m切上は total_qty_m 用に第三要素で返す。

    Returns:
        tuple[float, float, float, bool]:
            (remaining_m, done_m, qty_total_ceiled, used_unprocessed)
    """
    qty_conv_raw = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
    qty_total_ceiled = _ceil_roll_unit_length_m_to_next_step(qty_conv_raw)
    unp = _optional_unprocessed_m_from_plan_row(row)
    if unp is not None:
        fu = float(unp)
        if fu > 1e-12:
            remaining_m = max(0.0, fu)
            done_m = max(0.0, qty_conv_raw - fu)
        else:
            # 未加工が 0 付近: 換算数量(100m切上)を基準にするが、ロール単位長さ未満は 1 ロール分に引き上げ
            roll_m = _roll_unit_m_estimate_from_plan_row(
                row, qty_total_ceiled or qty_conv_raw or 1.0
            )
            base_m = max(0.0, qty_total_ceiled)
            remaining_m = max(base_m, roll_m) if roll_m > 0 else base_m
            done_m = 0.0
        return remaining_m, done_m, qty_total_ceiled, True
    raise PlanningValidationError(
        f"「{TASK_COL_UNPROCESSED}」が数値として読めません（セルが空または不正）、"
        "または列がありません。配台残量は未加工列のみで算定するため、配台処理を中止します。"
    )


def _fill_plan_dispatch_remaining_qty_column(plan_df: pd.DataFrame) -> None:
    """配台計画 DataFrame の「配台使用残数量」を _plan_row_dispatch_qty_metrics と同一式で埋める。"""
    if plan_df is None or getattr(plan_df, "empty", True):
        return
    if PLAN_COL_DISPATCH_REMAINING_QTY not in plan_df.columns:
        return
    for i in plan_df.index:
        row = plan_df.loc[i]
        rem, _d, _t, _fu = _plan_row_dispatch_qty_metrics(row)
        plan_df.at[i, PLAN_COL_DISPATCH_REMAINING_QTY] = rem


def parse_optional_int(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    try:
        return int(round(float(s)))
    except (TypeError, ValueError):
        return None


def parse_optional_date(val):
    if val is None or pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def _parse_env_optional_date(env_key: str):
    """os.environ の 1 キーを暦日に解釈。空・解釈不能は None。"""
    raw = (os.environ.get(env_key) or "").strip()
    if not raw:
        return None
    return parse_optional_date(raw)


def _planning_df_cell_scalar(row, col_name):
    """
    iterrows() 1行分から列値を得る。同一見出しの重複列はあると row.get は Series になり」
    str→to_datetime で誤った日付になることがあるため、先頭の非欠損スカラーを返す。
    """
    v = row.get(col_name) if hasattr(row, "get") else None
    if isinstance(v, pd.Series):
        for x in v:
            if x is None or (isinstance(x, float) and pd.isna(x)):
                continue
            return x
        return None
    return v


def _roll_unit_m_estimate_from_plan_row(row, fallback_m: float) -> float:
    """
    配台計画1行から 1 ロールあたりの長さ(m)。シートのロール単位長さを優先し、
    空・0 のときは build_task_queue と同趣旨で製品名から推定する。
    """
    product_name = row.get(TASK_COL_PRODUCT, None) if hasattr(row, "get") else None
    unit = parse_float_safe(_planning_df_cell_scalar(row, PLAN_COL_ROLL_UNIT_LENGTH), 0.0)
    fb = max(1e-9, float(parse_float_safe(fallback_m, 0.0)))
    if unit <= 0:
        unit = infer_unit_m_from_product_name(product_name, fallback_unit=fb)
    try:
        unit = float(unit)
    except (TypeError, ValueError):
        unit = 0.0
    if unit <= 0:
        unit = fb
    return float(unit)


def load_ai_cache():
    try:
        if os.path.exists(ai_cache_path):
            with open(ai_cache_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    # 期陝切れエントリを除去（6時間）
                    now_ts = time_module.time()
                    cleaned = {}
                    expired_count = 0
                    for k, v in data.items():
                        # 新形式: {"ts": epoch_seconds, "data": {...}}
                        if isinstance(v, dict) and "ts" in v and "data" in v:
                            ts = parse_float_safe(v.get("ts"), 0.0)
                            if ts > 0 and (now_ts - ts) <= AI_CACHE_TTL_SECONDS:
                                cleaned[k] = v
                            else:
                                expired_count += 1
                        # 旧形式: 値は直接AI結果dict（互換で読み取り」坳時に新形式へ再保存される）
                        else:
                            cleaned[k] = {"ts": now_ts, "data": v}
                    if expired_count > 0:
                        logging.info(f"AIキャッシュ期陝切れを削除: {expired_count}件")
                    return cleaned
    except Exception as e:
        logging.warning(f"AIキャッシュ読み込み失敗: {e}")
    return {}

def save_ai_cache(cache_obj):
    try:
        with open(ai_cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_obj, f, ensure_ascii=False)
    except Exception as e:
        logging.warning(f"AIキャッシュ保存失敗: {e}")

def get_cached_ai_result(cache_obj, cache_key, content_key=None):
    """
    content_key: オプション。保存時と同一の文字列でないヒットは無効化する（特別指定・照合用の二次チェック）。
    旧エントリに content_key は無い場合は SHA256 キー一致のみで従来どおりヒットとみなす。
    """
    entry = cache_obj.get(cache_key)
    if not isinstance(entry, dict):
        return None
    ts = parse_float_safe(entry.get("ts"), 0.0)
    if ts <= 0:
        return None
    if (time_module.time() - ts) > AI_CACHE_TTL_SECONDS:
        return None
    if content_key is not None:
        stored_ck = entry.get("content_key")
        if stored_ck is not None and stored_ck != content_key:
            logging.info(
                "AIキャッシュ: キーは一致したが、content_key は実行入力と異なるため無効化した。"
            )
            return None
    data = entry.get("data")
    if isinstance(data, dict):
        return data
    return None

def put_cached_ai_result(cache_obj, cache_key, parsed_obj, content_key=None):
    payload = {"ts": time_module.time(), "data": parsed_obj}
    if content_key is not None:
        payload["content_key"] = content_key
    cache_obj[cache_key] = payload

def extract_retry_seconds(err_text):
    # 例: "Please retry in 57.089735313s."
    m = re.search(r"retry in ([0-9]+(?:\.[0-9]+)?)s", err_text, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    # 例: "'retryDelay': '57s'"
    m = re.search(r"retryDelay'\s*:\s*'([0-9]+)s'", err_text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


def _gemini_err_text_for_exc(exc: BaseException) -> str:
    parts = [str(exc), repr(exc)]
    for attr in ("status_code", "code", "message"):
        v = getattr(exc, attr, None)
        if v is not None:
            parts.append(str(v))
    return " ".join(parts)


def _gemini_is_transient_api_error(err_text: str) -> bool:
    """503 / 過負荷 / 期限切れなど、待てば再試行に値する API 失敗。"""
    t = err_text.upper()
    if "429" in err_text:
        return True
    if "503" in err_text:
        return True
    if "504" in err_text:
        return True
    for needle in (
        "UNAVAILABLE",
        "RESOURCE_EXHAUSTED",
        "DEADLINE_EXCEEDED",
        "DEADLINE EXCEEDED",
        "SERVICE UNAVAILABLE",
        "INTERNAL ERROR",
        "UNRECOVERABLE",
    ):
        if needle in t:
            return True
    return False


def _gemini_is_quota_style_error(err_text: str) -> bool:
    t = err_text.upper()
    return ("429" in err_text) or ("RESOURCE_EXHAUSTED" in t)


def _gemini_is_timeout_error(exc: BaseException, err_text: str) -> bool:
    """HTTP 読み取りタイムアウト・接続タイムアウト等（応答が期限内に返らない）。"""
    if isinstance(exc, TimeoutError):
        return True
    tn = type(exc).__name__
    if tn in ("ReadTimeout", "ConnectTimeout", "WriteTimeout", "PoolTimeout"):
        return True
    u = err_text.upper()
    if "READ TIMEOUT" in u or "CONNECT TIMEOUT" in u or "WRITE TIMEOUT" in u:
        return True
    if "TIMED OUT" in u:
        return True
    # 504 等も TIMEOUT を含むが、DEADLINE_EXCEEDED 単体は一時エラー扱いに任せる
    if "TIMEOUT" in u and "DEADLINE_EXCEEDED" not in u:
        return True
    return False


def _gemini_try_order_from_env() -> tuple[str, ...] | None:
    raw = (os.environ.get("GEMINI_MODEL_TRY_ORDER") or "").strip()
    if not raw:
        return None
    parts = tuple(p.strip() for p in raw.split(",") if p.strip())
    return parts or None


def _gemini_effective_model_chain(model: str | None) -> tuple[str, ...]:
    """引数 model があればそれのみ。なければ GEMINI_MODEL、設定シート D/E、環境変数の順で決定。"""
    if model is not None and str(model).strip():
        return (str(model).strip(),)
    pinned = (os.environ.get("GEMINI_MODEL") or "").strip()
    if pinned:
        return (pinned,)
    sheet_chain = _read_gemini_model_try_chain_from_settings_sheet(
        (TASKS_INPUT_WORKBOOK or "").strip()
    )
    if sheet_chain:
        return sheet_chain
    ovr = _gemini_try_order_from_env()
    if ovr is not None:
        return ovr
    return GEMINI_MODEL_IDS_BY_QUALITY


def _gemini_is_model_endpoint_unavailable_error(err_text: str) -> bool:
    """モデル未提供・モデル名不正など、別モデルでの再試行が合理的な失敗。"""
    t = err_text.upper()
    u = err_text.lower()
    if "NOT_FOUND" in t and ("MODEL" in t or "MODELS/" in t):
        return True
    if "404" in err_text and "model" in u:
        return True
    if ("DOES NOT EXIST" in t or "WAS NOT FOUND" in t) and "model" in u:
        return True
    return False


def _gemini_pre_request_jitter_sleep() -> None:
    mx = max(0.0, _GEMINI_PRE_REQUEST_JITTER_MAX)
    if mx <= 0.0:
        return
    time_module.sleep(random.uniform(0.0, mx))


def _gemini_progress_log_interval_sec() -> float:
    """Gemini 応答待ち中に INFO を出す間隔（秒）。0 以下でハートビート無効（送信ログのみ）。"""
    try:
        return float((os.environ.get("GEMINI_PROGRESS_LOG_INTERVAL_SEC") or "12").strip())
    except (TypeError, ValueError):
        return 12.0


def _gemini_flush_log_handlers() -> None:
    try:
        for h in logging.getLogger().handlers:
            flush = getattr(h, "flush", None)
            if flush is not None:
                flush()
    except Exception:
        pass
    try:
        sys.stdout.flush()
    except Exception:
        pass


def _gemini_heartbeat_loop(
    stop: threading.Event, prefix: str, model_id: str, interval_sec: float
) -> None:
    """ブロッキング中でもターミナルが固まって見えないよう、一定間隔で待機ログを出す。"""
    start = time_module.monotonic()
    while True:
        if stop.wait(timeout=interval_sec):
            break
        elapsed = time_module.monotonic() - start
        logging.info(
            "%sGemini 応答待ち... 約%.0f秒経過（モデル: %s）",
            prefix,
            elapsed,
            model_id,
        )
        _gemini_flush_log_handlers()


def _gemini_generate_content_with_retry(
    client: genai.Client,
    *,
    contents,
    model: str | None = None,
    max_attempts: int | None = None,
    log_label: str = "",
):
    """generate_content を再試行する（Gemini generateContent 共通）。

    - モデル列: マクロブック「設定」シート D/E で有効化した ID（上から順）、なければ
      GEMINI_MODEL_IDS_BY_QUALITY（精度高い順）。環境変数 GEMINI_MODEL で単一固定、
      GEMINI_MODEL_TRY_ORDER（カンマ区切り）で上書き可。引数 model を渡したときはその1件のみ。
    - 同一モデルあたり最大 _GEMINI_RETRY_MAX_ATTEMPTS 回（既定 3、GEMINI_RETRY_MAX_ATTEMPTS で変更）。
      そのモデルで試行を使い切ったら、列の次のモデルへ進む（試すモデルがなくなるまで）。
    - モデル未提供（404 等）は直ちに次モデルへ進む。
    - 各試行の直前: 0〜_GEMINI_PRE_REQUEST_JITTER_MAX の乱数待機（同時リクエストのばらつき）
    - 一時エラー待機: (1) 429 等で本文に retry 秒数 (2) 指数バックオフ＋ジッター
    - HTTP タイムアウト（既定 60 秒・GEMINI_REQUEST_TIMEOUT_SEC）: 同一モデルに残試行があれば短待機で再試行。
      試行を使い切ったらモデル列の次点へ進む（_gemini_client の HttpOptions と併用）。

    戻り値: (応答オブジェクト, 実際に成功したモデル ID)
    """
    chain = _gemini_effective_model_chain(model)
    n = max_attempts if max_attempts is not None else _GEMINI_RETRY_MAX_ATTEMPTS
    if n < 1:
        n = 1
    base = max(0.1, float(_GEMINI_RETRY_BACKOFF_BASE))
    prefix = f"{log_label}: " if log_label else ""
    hb_interval = _gemini_progress_log_interval_sec()
    last_raise: BaseException | None = None
    for mi, mid in enumerate(chain):
        for attempt in range(n):
            _gemini_pre_request_jitter_sleep()
            try:
                logging.info("%sGemini API を呼び出し中（モデル: %s）", prefix, mid)
                _gemini_flush_log_handlers()
                stop_hb = threading.Event()
                hb_thread: threading.Thread | None = None
                if hb_interval > 0:
                    hb_thread = threading.Thread(
                        target=_gemini_heartbeat_loop,
                        args=(stop_hb, prefix, mid, hb_interval),
                        name="gemini-progress-hb",
                        daemon=True,
                    )
                    hb_thread.start()
                t_req = time_module.monotonic()
                try:
                    res = client.models.generate_content(model=mid, contents=contents)
                finally:
                    stop_hb.set()
                    if hb_thread is not None:
                        hb_thread.join(timeout=2.0)
                elapsed_req = time_module.monotonic() - t_req
                logging.info(
                    "%sGemini API 応答を受信しました（約%.1f秒、モデル: %s）",
                    prefix,
                    elapsed_req,
                    mid,
                )
                _gemini_flush_log_handlers()
                return res, mid
            except Exception as e:
                err_text = _gemini_err_text_for_exc(e)
                if _gemini_is_model_endpoint_unavailable_error(err_text) and mi < len(chain) - 1:
                    logging.warning(
                        "%sGemini モデル %s が利用できません: %s — 次モデルへ切り替えます。",
                        prefix,
                        mid,
                        err_text[:800],
                    )
                    last_raise = e
                    break
                if _gemini_is_timeout_error(e, err_text):
                    last_raise = e
                    if attempt < n - 1:
                        wait_sec = min(2.0 + random.uniform(0.0, 1.0), 5.0)
                        logging.warning(
                            "%sGemini API タイムアウト（モデル %s 試行 %s/%s）: %s — %.1f 秒待機して再試行します。",
                            prefix,
                            mid,
                            attempt + 1,
                            n,
                            err_text[:800],
                            wait_sec,
                        )
                        time_module.sleep(wait_sec)
                        continue
                    if mi < len(chain) - 1:
                        logging.warning(
                            "%sGemini API タイムアウト（モデル %s）: %s — 次モデルへ切り替えます。",
                            prefix,
                            mid,
                            err_text[:800],
                        )
                        break
                    raise
                if _gemini_is_transient_api_error(err_text) and attempt < n - 1:
                    wait_sec = None
                    if _gemini_is_quota_style_error(err_text):
                        rs = extract_retry_seconds(err_text)
                        if rs is not None:
                            wait_sec = min(max(rs, 1.0), 120.0) + random.uniform(0.0, 1.5)
                    if wait_sec is None:
                        pow_part = base * (2**attempt)
                        jitter = random.uniform(0.0, min(4.0, base * 2.0))
                        wait_sec = min(pow_part + jitter, 90.0)
                    logging.warning(
                        "%sGemini API 一時エラー（モデル %s 試行 %s/%s）: %s — %.1f 秒待機して再試行します。",
                        prefix,
                        mid,
                        attempt + 1,
                        n,
                        err_text[:800],
                        wait_sec,
                    )
                    time_module.sleep(wait_sec)
                    continue
                if attempt < n - 1:
                    wait_sec = min(2.0 + random.uniform(0.0, 1.0), 5.0)
                    logging.warning(
                        "%sGemini API エラー（モデル %s 試行 %s/%s）: %s — %.1f 秒待機して再試行します。",
                        prefix,
                        mid,
                        attempt + 1,
                        n,
                        err_text[:800],
                        wait_sec,
                    )
                    time_module.sleep(wait_sec)
                    continue
                if mi < len(chain) - 1:
                    logging.warning(
                        "%sGemini モデル %s が %s 回とも失敗したため次モデルへ切り替えます: %s",
                        prefix,
                        mid,
                        n,
                        err_text[:800],
                    )
                    last_raise = e
                    break
                raise
    if last_raise is not None:
        raise last_raise
    raise RuntimeError("Gemini: モデル列が空です。")


def _normalize_product_dim_separators_for_roll_inference(s: str) -> str:
    """
    製品名に混ざる寸法区切りを ASCII の x に寄せる。
    先に NFKC で互換分解（全角英数字・互換記号など）を寄せ、列名 `_align_dataframe_headers_to_canonical`
    と同趣旨に Excel 由来の表記ゆれを弱める。
    半角 X/x 以外（×・全角Ｘｘ・罫線系の乗号）だけがあると正規表現に一致せず、
    換算数量フォールバックでロール単位長さが誤ることがある。
    """
    if not s:
        return s
    t = unicodedata.normalize("NFKC", s)
    for ch in (
        "\u00d7",  # × MULTIPLICATION SIGN
        "\u2715",  # ✕ MULTIPLICATION X
        "\u2716",  # ✖ HEAVY MULTIPLICATION X
        "\u2a2f",  # ⨯ VECTOR OR CROSS PRODUCT
        "\u2a09",  # ⨉ CROSS MULTIPLICATION
        "\uff38",  # Ｘ FULLWIDTH LATIN CAPITAL LETTER X
        "\uff58",  # ｘ FULLWIDTH LATIN SMALL LETTER X
        # 寸法区切りに誤入力されがちな「X に見えるが ASCII [xX] にマッチしない」文字（推定失敗→換算数量→100m 切上で 870→900 等）
        "\u0425",  # CYRILLIC CAPITAL LETTER HA
        "\u0445",  # CYRILLIC SMALL LETTER HA
        "\u03a7",  # GREEK CAPITAL LETTER CHI
        "\u03c7",  # GREEK SMALL LETTER CHI
    ):
        t = t.replace(ch, "x")
    return t


def infer_unit_m_from_product_name(product_name, fallback_unit):
    """
    製品名文字列から 1 ロールあたりの長さ(m)を推定する。

    実データ（テストコード直下「製品名,ロール単位の長さ.txt」）に合わせ、
    **最後に現れる「左数 x 右数」（2〜6 桁同士）ペアの右側**をロール長の候補とする。
    例: 870x200→200、1440x300→300、1550x40→40、770x300→300。
    寸法区切りは半角 x/X のほか ×（U+00D7）・全角Ｘｘ 等を正規化してから判定する。

    上記ペアが無いときは、従来どおり最後の「X の直後の 2〜6 桁」を拾う。
    いずれもマッチしない場合（寸法なし、または X 後が 2〜6 桁で解釈不能）は
    ``INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M``（100）を返す。製品名が欠損のときのみ
    ``fallback_unit`` を返す。
    """
    if product_name is None or pd.isna(product_name):
        return fallback_unit
    s = _normalize_product_dim_separators_for_roll_inference(str(product_name))
    # 「NNNX MM」形式: 最後のペアの **右側（X 後）**をロール長候補とする（製品名,ロール単位の長さ.txt 準拠）
    dim_pairs = re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s)
    if dim_pairs:
        try:
            _a_str, b_str = dim_pairs[-1]
            b = int(b_str)
            if b > 0:
                return b
        except ValueError:
            pass
    # "770X300..." のようなパターンから X の後の数値を拾う（最後に見つかったXを優先）
    matches = re.findall(r"[xX]\s*(\d{2,6})", s)
    if matches:
        try:
            v = int(matches[-1])
            if v > 0:
                return v
        except ValueError:
            pass
    return float(INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M)


def _coerce_roll_unit_m_when_converted_qty_below_roll(
    product_name, unit_m: float, qty_total: float
) -> float:
    """
    加工長さ（1ロールあたりの m）の解釈。

    換算数量（qty_total）が、製品名から推定したロール単位長さより小さいときは、
    ロール単位長さを採用する（シート等で unit_m が換算数量未満に誤っている場合の救済）。
    シート・手入力で unit_m が推定より大きい場合は上書きしない。
    """
    try:
        u = float(unit_m)
    except (TypeError, ValueError):
        u = 0.0
    roll_infer = infer_unit_m_from_product_name(product_name, fallback_unit=0.0)
    try:
        roll_infer = float(roll_infer)
    except (TypeError, ValueError):
        roll_infer = 0.0
    if roll_infer <= 0:
        return u
    q = parse_float_safe(qty_total, 0.0)
    if q > 0 and q < roll_infer and u < roll_infer:
        return roll_infer
    return u


def _floor_positive_m_to_planning_minimum(val: float, minimum: float) -> float:
    """正の長さ(m)のみ、minimum 未満なら minimum に引き上げる。0以下・欠損はそのまま。"""
    v = parse_float_safe(val, 0.0)
    if v <= 0:
        return v
    m = parse_float_safe(minimum, 0.0)
    if m <= 0:
        return v
    return float(m) if v < m else v


def _ceil_roll_unit_length_m_to_next_step(roll_m: float, step_m: float = None) -> float:
    """
    正の長さ(m)を step の倍数に切り上げ（下二桁繰り上げ: step=100 のとき 40→100, 125→200）。
    段階1の **ロール単位長さ** と、段階2の **換算数量（配台用内部）** で共用（刻みは `ROLL_UNIT_LENGTH_CEIL_STEP_M`）。
    """
    v = parse_float_safe(roll_m, 0.0)
    if v <= 0:
        return v
    step = parse_float_safe(
        step_m if step_m is not None else ROLL_UNIT_LENGTH_CEIL_STEP_M, 0.0
    )
    if step <= 0:
        return v
    return float(math.ceil(v / step) * step)


def _ceil_roll_unit_length_plan_sheet_cell(val):
    """DataFrame セル用。空・非数値はそのまま。正の数値は ROLL_UNIT_LENGTH_CEIL_STEP_M 倍数に切り上げ。"""
    if val is None:
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("nan", "none"):
            return val
    try:
        x = float(val)
    except (TypeError, ValueError):
        return val
    if x <= 0:
        return val
    return _ceil_roll_unit_length_m_to_next_step(x)


def _apply_roll_unit_length_ceil_step_to_plan_df(df: pd.DataFrame) -> None:
    """段階1 DataFrame の ロール単位長さ 列を、マージ後も含めて切り上げ正規化する。"""
    col = PLAN_COL_ROLL_UNIT_LENGTH
    if df is None or df.empty or col not in df.columns:
        return
    for i in df.index:
        df.at[i, col] = _ceil_roll_unit_length_plan_sheet_cell(df.at[i, col])


def _stage1_roll_length_for_planning_row(row) -> float:
    """段階1: 加工計画由来の1行から ロール単位長さ(m)を計算（``run_stage1_extract`` の merge 前と同一式）。"""
    _pn_stage1 = row.get(TASK_COL_PRODUCT, None)
    qty, _done_m, _qtceiled, _from_unp = _plan_row_dispatch_qty_metrics(row)
    _qty_total_s1 = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
    _qty_total_s1 = _floor_positive_m_to_planning_minimum(
        _qty_total_s1, PLANNING_MIN_QTY_M
    )
    _roll_len = infer_unit_m_from_product_name(
        _pn_stage1, fallback_unit=_qty_total_s1 if _qty_total_s1 > 0 else qty
    )
    try:
        _roll_len = float(_roll_len)
    except (TypeError, ValueError):
        _roll_len = _qty_total_s1 if _qty_total_s1 > 0 else qty
    if _roll_len <= 0:
        _roll_len = _qty_total_s1 if _qty_total_s1 > 0 else max(qty, 1e-9)
    _roll_len = _coerce_roll_unit_m_when_converted_qty_below_roll(
        _pn_stage1, _roll_len, _qty_total_s1
    )
    try:
        _roll_len = float(_roll_len)
    except (TypeError, ValueError):
        _roll_len = _qty_total_s1 if _qty_total_s1 > 0 else max(qty, 1e-9)
    if _roll_len <= 0:
        _roll_len = _qty_total_s1 if _qty_total_s1 > 0 else max(qty, 1e-9)
    return float(_ceil_roll_unit_length_m_to_next_step(_roll_len))


def _heal_stage1_roll_unit_if_width_ceiling_merge_spurious(out_df: "pd.DataFrame") -> None:
    """
    段階1: 既存シートのマージで「寸法ペア左側（例: 870）を 100m 切上した値」がロール単位長さに
    残った場合、製品名からの再計算で矯正する（誤マージ・誤フォールバックの典型: 900 を期待 200 の行）。
    手入力で意図的に左側切上と同じ値にした行は稀なため、一致時のみ上書きする。
    """
    if out_df is None or getattr(out_df, "empty", True):
        return
    if (
        PLAN_COL_ROLL_UNIT_LENGTH not in out_df.columns
        or TASK_COL_PRODUCT not in out_df.columns
    ):
        return
    healed = 0
    for i in out_df.index:
        row = out_df.loc[i]
        pn = row.get(TASK_COL_PRODUCT, None)
        s = _normalize_product_dim_separators_for_roll_inference(str(pn or ""))
        dim_pairs = re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s)
        if not dim_pairs:
            continue
        try:
            left_w = int(dim_pairs[-1][0])
        except ValueError:
            continue
        if left_w <= 0:
            continue
        width_ceiled = float(_ceil_roll_unit_length_m_to_next_step(float(left_w)))
        cur = parse_float_safe(row.get(PLAN_COL_ROLL_UNIT_LENGTH), 0.0)
        if cur <= 0:
            continue
        if abs(cur - width_ceiled) > 1e-6:
            continue
        try:
            want = _stage1_roll_length_for_planning_row(row)
        except Exception:
            continue
        if abs(cur - want) <= 1e-6:
            continue
        out_df.at[i, PLAN_COL_ROLL_UNIT_LENGTH] = want
        healed += 1
    if healed:
        logging.info(
            "段階1: ロール単位長さが寸法左側の100m切上と誤一致していた行を %s 件、製品名ベースで矯正しました。",
            healed,
        )


def _heal_stage1_roll_unit_no_dim_when_roll_matches_qty_mistake(
    out_df: "pd.DataFrame",
) -> None:
    """
    寸法パターンが無い品番で、ロール単位長さが換算数量（下限適用後）またはその 100m 切上と
    同じになっている行を矯正する（旧シートマージで FEL 等に換算数量が載った誤り向け）。
    小さい値（<500）は「意図的に換算数量と同じロール長」とみなし触れない。
    """
    if out_df is None or getattr(out_df, "empty", True):
        return
    if (
        PLAN_COL_ROLL_UNIT_LENGTH not in out_df.columns
        or TASK_COL_PRODUCT not in out_df.columns
        or TASK_COL_QTY not in out_df.columns
    ):
        return
    healed = 0
    min_heal_cur = 500.0
    want = float(
        _ceil_roll_unit_length_m_to_next_step(
            float(INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M)
        )
    )
    for i in out_df.index:
        row = out_df.loc[i]
        pn = row.get(TASK_COL_PRODUCT, None)
        s = _normalize_product_dim_separators_for_roll_inference(str(pn or ""))
        if re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s):
            continue
        qty_floor = _floor_positive_m_to_planning_minimum(
            parse_float_safe(row.get(TASK_COL_QTY), 0.0), PLANNING_MIN_QTY_M
        )
        if qty_floor <= 0:
            continue
        qty_ceiled = float(_ceil_roll_unit_length_m_to_next_step(float(qty_floor)))
        cur = parse_float_safe(row.get(PLAN_COL_ROLL_UNIT_LENGTH), 0.0)
        if cur + 1e-9 < min_heal_cur:
            continue
        if abs(cur - qty_floor) > 1e-4 and abs(cur - qty_ceiled) > 1e-4:
            continue
        if abs(cur - want) < 1e-6:
            continue
        out_df.at[i, PLAN_COL_ROLL_UNIT_LENGTH] = want
        healed += 1
    if healed:
        logging.info(
            "段階1: 寸法なしでロール単位長さが換算数量と誤一致していた行を %s 件、既定 %sm へ矯正しました。",
            healed,
            int(want) if abs(want - int(want)) < 1e-9 else want,
        )


def load_tasks_df():
    """
    タスク入力を取得れる（tasks.xlsx は使用しない）。
    必須: 環境変数 TASK_INPUT_WORKBOOK にマクロ実行ブックのフルパス（VBA は設定）
         シート「加工計画DATA」を読み込む（投入目安は「回答納期」。未入力時は「指定納期」）。
    """
    if not TASKS_INPUT_WORKBOOK:
        raise FileNotFoundError(
            "TASK_INPUT_WORKBOOK は未設定です。VBA の RunPython でマクロ実行ブックのパスを渡してください。"
        )
    if not os.path.exists(TASKS_INPUT_WORKBOOK):
        raise FileNotFoundError(f"TASK_INPUT_WORKBOOK は存在しません: {TASKS_INPUT_WORKBOOK}")
    df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=TASKS_SHEET_NAME)
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, list(SOURCE_BASE_COLUMNS))
    _ensure_dataframe_has_unprocessed_column(
        df, context_label=f"シート「{TASKS_SHEET_NAME}」"
    )
    # 加工計画DATA の主列は「換算数量」（TASK_COL_QTY）。無いブックは未加工→旧「残作数値」の順で補完。
    if TASK_COL_QTY not in df.columns:
        for _alt_qty in ("未加工", "残作数値"):
            if _alt_qty in df.columns:
                df[TASK_COL_QTY] = df[_alt_qty]
                logging.info(
                    "タスク入力: 列「%s」が無いため「%s」をコピーして補完しました。",
                    TASK_COL_QTY,
                    _alt_qty,
                )
                break
    # 「受注数」列名の表記ゆれを「受注数」（TASK_COL_ORDER_QTY）へ寄せる補完
    if TASK_COL_ORDER_QTY not in df.columns and "受注数" in df.columns:
        df[TASK_COL_ORDER_QTY] = df["受注数"]
        logging.info(
            "タスク入力: 列「%s」が無いため「受注数」をコピーして補完しました。",
            TASK_COL_ORDER_QTY,
        )
    logging.info(f"タスク入力: '{TASKS_INPUT_WORKBOOK}' の '{TASKS_SHEET_NAME}' を読み込みました。")
    return df


def _nfkc_column_aliases(canonical_name):
    """見出しの表記ゆれ（全角記坷・互換文字）を坸坎れるための比較キー。"""
    return unicodedata.normalize("NFKC", str(canonical_name).strip())


def _align_dataframe_headers_to_canonical(df, canonical_names):
    """列名を NFKC 一致で canonical に寄せる（Excel 坴は全角 '_' 等でも読ゝるよごに）。"""
    key_to_canonical = {_nfkc_column_aliases(c): c for c in canonical_names}
    # 旧見出し「残作数値」→ 現行「換算数量」（TASK_COL_QTY）
    if TASK_COL_QTY in canonical_names:
        key_to_canonical[_nfkc_column_aliases("残作数値")] = TASK_COL_QTY
    # 旧見出し「原板…」を「原反…」へ寄せる（互換。canonical は TASK_COL / PLAN_COL の表記）
    if TASK_COL_RAW_INPUT_DATE in canonical_names:
        key_to_canonical[_nfkc_column_aliases("原板投入日")] = TASK_COL_RAW_INPUT_DATE
    if PLAN_COL_RAW_INPUT_DATE_OVERRIDE in canonical_names:
        key_to_canonical[_nfkc_column_aliases("原板投入日_上書き")] = (
            PLAN_COL_RAW_INPUT_DATE_OVERRIDE
        )
        _ref_canon = plan_reference_column_name(PLAN_COL_RAW_INPUT_DATE_OVERRIDE)
        if _ref_canon in canonical_names:
            key_to_canonical[_nfkc_column_aliases("（元）原板投入日_上書き")] = _ref_canon
    rename_map = {}
    for col in df.columns:
        k = _nfkc_column_aliases(col)
        if k in key_to_canonical:
            target = key_to_canonical[k]
            if col != target:
                rename_map[col] = target
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def _normalize_equipment_match_key(val):
    """
    工程名（設備坝）の照合用キー。
    NFKC・剝後空白・連続空白・NBSP/全角スペース・ゼロ幅文字を正規化する。
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    t = unicodedata.normalize("NFKC", str(val))
    t = t.replace("\u00a0", " ").replace("\u3000", " ")
    t = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _equipment_line_key_to_physical_occupancy_key(eq_line: str) -> str:
    """設備列キー（工程+機械 等）から」実機械の占有に用いるキー（機械名ベース・正規化）を得る。"""
    s = str(eq_line or "").strip()
    if not s:
        return ""
    nk = _normalize_equipment_match_key(s)
    if "+" in nk:
        return _normalize_equipment_match_key(nk.split("+", 1)[1])
    return nk


def _physical_machine_occupancy_key_for_task(task: dict) -> str:
    """
    設備のタイムライン占有（machine_avail_dt・間隔ミラー）に用いるキー。
    機械カレンダー列は equipment_line_key の「工程+機械」と一致するため、
    正規化後に「+」を含むとしは **machine_name より先に** しこから実機械名を採用する。
    （machine_name に工程名のみなどは入り」床キー「熱融着機 湖南」とうれで候補外し漝れれるのを防し）
    坘一坝のときは従来どおり machine_name を優先し、無ければ equipment_line_key / machine から推定する。
    machine_name に「工程+機械」と入っている場合でも」占有は実機械名（+ の坳坴）に寄せる。
    全角「＋」のみの列は NFKC 後に半角「+」になるため、分割判定は正規化後に行う。
    """
    ek = str(task.get("equipment_line_key") or "").strip()
    nek = _normalize_equipment_match_key(ek)
    if nek and "+" in nek:
        pk = _equipment_line_key_to_physical_occupancy_key(ek)
        if pk:
            return pk
    mn = str(task.get("machine_name") or "").strip()
    if mn:
        nk = _normalize_equipment_match_key(mn)
        if "+" in nk:
            return _normalize_equipment_match_key(nk.split("+", 1)[1])
        return nk
    return _equipment_line_key_to_physical_occupancy_key(
        str(task.get("equipment_line_key") or task.get("machine") or "")
    )


def _machine_occupancy_key_resolve(task: dict, eq_line: str) -> str:
    """
    machine_avail_dt・機械カレンダー床と整合する占有キー（原則: 実機械名）。
    task から取れないとしは eq_line（工程+機械）から機械名ベースを推定し、最後の手段で eq_line。
    「… or eq_line」による工程+機械フォールバックは機械カレンダー実キーと厳密一致になり得るため廃止。
    """
    occ = (_physical_machine_occupancy_key_for_task(task) or "").strip()
    if occ:
        return occ
    ek = str(eq_line or "").strip()
    if not ek:
        return ""
    pk = (_equipment_line_key_to_physical_occupancy_key(ek) or "").strip()
    return pk or ek


def _equipment_lookup_normalized_to_canonical(equipment_list):
    """正規化キー → master スキルシート上の列名（canonical 表記）。"""
    lookup = {}
    for eq in equipment_list:
        k = _normalize_equipment_match_key(eq)
        if k and k not in lookup:
            lookup[k] = eq
    # 工程名のみの照合（加工実績DATA等）: 同一工程の先頭列（工程+機械）へ寄せる
    for eq in equipment_list:
        s = str(eq).strip()
        if "+" not in s:
            continue
        p, _rest = s.split("+", 1)
        pk = _normalize_equipment_match_key(p)
        if pk and pk not in lookup:
            lookup[pk] = eq
    return lookup


def _equipment_schedule_header_labels(equipment_list: list) -> list:
    """
    結果_設備毎の時間割・結果_設備ガントの行＝列見出し用。
    内部キーは「工程+機械」のときは機械名を表示し、機械名の重複時のみ工程を括弧で補ご。
    """
    raw = []
    for eq in equipment_list:
        s = str(eq).strip()
        if "+" in s:
            mpart = s.split("+", 1)[1].strip()
            raw.append(mpart if mpart else s)
        else:
            raw.append(s)
    counts = {}
    for r in raw:
        counts[r] = counts.get(r, 0) + 1
    out = []
    for eq, r in zip(equipment_list, raw):
        if counts.get(r, 0) > 1:
            s = str(eq).strip()
            if "+" in s:
                p = s.split("+", 1)[0].strip()
                out.append(f"{r}（{p}）" if p else r)
            else:
                out.append(r)
        else:
            out.append(r)
    return out


def _split_equipment_line_process_machine(eq_line: str) -> tuple[str, str]:
    """
    設備マスタの列キー「工程+機械」を (工程名, 機械名) に分割れる。
    '+' は無いとしは機械名のみとみなし、工程名は空文字。
    """
    s = str(eq_line).strip()
    if not s:
        return ("", "")
    if "+" in s:
        p, m = s.split("+", 1)
        return (p.strip(), m.strip())
    return ("", s)


def _gantt_member_label_surname_only(raw: str) -> str:
    """
    設備ガントのタイムライン上の担当者姓表示用。半角＝全角空白はあれみ手剝を姓とみなし、無いとしは全体を表示
    （並びは1トークンのみのときは姓の切り出し試行のまま）。NFKC・富田/冨田寄せは姓用とともに。
    """
    sei, mei = _split_person_sei_mei(raw)
    if not sei:
        return ""
    n = _normalize_sei_for_match(sei)
    return n if n else sei


def _gantt_member_labels_for_task(evlist, task_id: str) -> list[str]:
    """
    設備ガントのタイムライン1セグメント用: 指定 task_id のイベントから担当者姓を出現順で重複除去。
    （シェイプの上下チップ用）
    """
    tid = str(task_id or "").strip()
    if not tid:
        return []
    raw_names: list[str] = []
    seen_raw: set[str] = set()
    for e in evlist or []:
        if str(e.get("task_id") or "").strip() != tid:
            continue
        op = str(e.get("op") or "").strip()
        if op and op not in seen_raw:
            seen_raw.add(op)
            raw_names.append(op)
        sub_raw = str(e.get("sub") or "").strip()
        if not sub_raw:
            continue
        for seg in re.split(r"[,」]", sub_raw):
            t = seg.strip()
            if t and t not in seen_raw:
                seen_raw.add(t)
                raw_names.append(t)
    labels: list[str] = []
    seen_label: set[str] = set()
    for raw in raw_names:
        lab = _gantt_member_label_surname_only(raw)
        if lab and lab not in seen_label:
            seen_label.add(lab)
            labels.append(lab)
    return labels


def _resolve_equipment_line_key_for_task(task: dict, equipment_list: list | None) -> str:
    """
    設備時間割・設備専有空しの列キー（skills / need とともに「工程+機械」を基本とれる）。
    機械名は空でマスタに当該工程の列は1つの値ならしの複坈キーへ寄せる。
    """
    p = str(task.get("machine") or "").strip()
    mn = str(task.get("machine_name") or "").strip()
    cand = f"{p}+{mn}" if (p and mn) else (p or mn)
    elist = [str(x).strip() for x in (equipment_list or []) if str(x).strip()]
    if cand in elist:
        return cand
    if mn:
        return cand
    if not p:
        return cand
    exact_p = [x for x in elist if x == p]
    if len(exact_p) == 1:
        return exact_p[0]
    prefixed = [x for x in elist if x.startswith(p + "+")]
    if len(prefixed) == 1:
        return prefixed[0]
    return p



def _apply_planning_sheet_post_load_mutations(
    df: "pd.DataFrame",
    wb_path: str,
    log_prefix: str,
    *,
    apply_exclude_rules_from_config: bool = True,
    compile_exclude_rules_d_to_e_with_ai: bool = True,
) -> None:
    """
    配台計画_タスク入力を DataFrame 化した直後の共通処理（設定シートの行同期・分割行の自動配台不要）。

    「設定_配台不要工程」の C/E による計画 DataFrame への「配台不要」上書きは **段階1のみ**
    （``run_stage1_extract`` 内の ``apply_exclude_rules_config_to_plan_df``）。段階2の
    ``load_planning_tasks_df`` では常に ``apply_exclude_rules_from_config=False`` を渡し、
    シート上の「配台不要」列をそのまま解釈する。

    段階2および試行順のみの xlwings 更新では ``compile_exclude_rules_d_to_e_with_ai=False`` とし、
    設定シートの D→E（ロジック式）の **Gemini 補完は行わない**（行同期・保存のみ）。

    ``apply_exclude_rules_from_config=False`` は本関数呼び出し側で明示する（上記のほか、
    試行順のみ再計算する xlwings 経路でも同様）。
    """
    try:
        _pairs_lr = []
        _seen_lr = set()
        for _, _row_lr in df.iterrows():
            _p = str(_row_lr.get(TASK_COL_MACHINE, "") or "").strip()
            _m = str(_row_lr.get(TASK_COL_MACHINE_NAME, "") or "").strip()
            if not _p:
                continue
            _k = (
                _normalize_process_name_for_rule_match(_p),
                _normalize_equipment_match_key(_m),
            )
            if _k in _seen_lr:
                continue
            _seen_lr.add(_k)
            _pairs_lr.append((_p, _m))
        run_exclude_rules_sheet_maintenance(
            wb_path,
            _pairs_lr,
            log_prefix,
            compile_exclude_rules_d_to_e_with_ai=compile_exclude_rules_d_to_e_with_ai,
        )
    except Exception:
        logging.exception("%s: 設定_配台不要工程の保守で例外（続行）", log_prefix)
    try:
        _apply_auto_exclude_bunkatsu_duplicate_machine(df, log_prefix=log_prefix)
    except Exception as ex:
        logging.warning(
            "%s: 分割行の配台試行自動設定で例外（続行）: %s",
            log_prefix,
            ex,
        )
    if apply_exclude_rules_from_config:
        try:
            apply_exclude_rules_config_to_plan_df(df, wb_path, log_prefix)
        except Exception as ex:
            logging.warning(
                "%s: 設定シートによる配台不要適用で例外（続行）: %s",
                log_prefix,
                ex,
            )

def load_planning_tasks_df():
    """
    2段階目用: マクロブック上の「配台計画_タスク入力」シートを読み込む。

    「担当OP_指定」列または特別指定備考の AI 出力 preferred_operator で主担当 OP を指名できる（skills のメンバー名とあいまい一致）。
    メイン「再優先特別記載」の task_preferred_operators は generate_plan 側で最優先マージされる。
    「配台不要」がオン（TRUE/1/はい 等）の行は配台対象外（**シート上の列の値をそのまま**解釈する）。
    読み込み後、同一依頼NO・重複機械名があるグループの工程「分割」行へ空なら「配台不要」=yes（段階1と同じ）。
    「設定_配台不要工程」シートの**行同期・保守**（``run_exclude_rules_sheet_maintenance``）は行うが、
    D→E の **AI 補完は行わない**（段階1のみ）。C/E に基づく計画シートへの配台不要の**再適用**
    （``apply_exclude_rules_config_to_plan_df``）も行わない（段階1のみ）。
    """
    if not TASKS_INPUT_WORKBOOK:
        raise FileNotFoundError(
            "TASK_INPUT_WORKBOOK は未設定です。VBA の RunPython でマクロ実行ブックのパスを渡してください。"
        )
    if not os.path.exists(TASKS_INPUT_WORKBOOK):
        raise FileNotFoundError(f"TASK_INPUT_WORKBOOK は存在しません: {TASKS_INPUT_WORKBOOK}")
    df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=PLAN_INPUT_SHEET_NAME)
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(
        df, plan_input_sheet_column_order()
    )
    _ensure_dataframe_has_unprocessed_column(
        df, context_label=f"シート「{PLAN_INPUT_SHEET_NAME}」"
    )
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""
    _apply_planning_sheet_post_load_mutations(
        df,
        TASKS_INPUT_WORKBOOK,
        "配台シート読込",
        apply_exclude_rules_from_config=False,
        compile_exclude_rules_d_to_e_with_ai=False,
    )
    logging.info(
        f"計画タスク入力: '{TASKS_INPUT_WORKBOOK}' の '{PLAN_INPUT_SHEET_NAME}' を読み込みました。"
    )
    return df


def _main_sheet_cell_is_global_comment_label(val) -> bool:
    """メインシート上「グローバルコメント」見出しセルか（表記ゆれ許容）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    s = unicodedata.normalize("NFKC", str(val).strip())
    if not s:
        return False
    if _nfkc_column_aliases(s) == _nfkc_column_aliases("グローバルコメント"):
        return True
    if "グローバル" in s and "コメント" in s:
        return True
    return False


def load_main_sheet_global_priority_override_text() -> str:
    """
    TASK_INPUT_WORKBOOK のメインシートで「グローバルコメント」と書かれたセルの **直下** を読む。
    シート名: 「メイン」「メイン_」「Main」のいうれか」または坝剝に「メイン」を含む（VBA GetMainWorksheet と同じ趣旨）。

    内容は **Gemini で一括解釈**（`analyze_global_priority_override_comment`）。工場休業日・再優先フラグ・未実装指示のメモを JSON 化れる。
    API キーは無い場合のみ」工場休業日はルールベースの `parse_factory_closure_dates_from_global_comment` で補完れる。
    """
    wb_path = TASKS_INPUT_WORKBOOK.strip() if TASKS_INPUT_WORKBOOK else ""
    if not wb_path or not os.path.exists(wb_path):
        return ""
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "メイン再優先特記: ブックに「%s」があるため、openpyxl でグローバルコメントを読みません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return ""
    try:
        wb = load_workbook(wb_path, data_only=True, read_only=False)
    except Exception as e:
        logging.warning("メイン再優先特記: ブックを開きませんでした: %s", e)
        return ""
    try:
        ws = None
        for name in ("メイン", "メイン_", "Main"):
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            for sn in wb.sheetnames:
                if "メイン" in sn:
                    ws = wb[sn]
                    break
        if ws is None:
            return ""
        max_r = min(ws.max_row or 0, 400)
        max_c = min(ws.max_column or 0, 40)
        if max_r < 1 or max_c < 1:
            return ""
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if not _main_sheet_cell_is_global_comment_label(cell.value):
                    continue
                below = ws.cell(row=r + 1, column=c).value
                if below is None or (isinstance(below, float) and pd.isna(below)):
                    return ""
                return str(below).strip()
        return ""
    finally:
        pass


def _global_comment_chunk_implies_factory_closure(chunk: str) -> bool:
    """
    メイン「グローバルコメント」の断片は」工場短縮の休業・非稼働を愝味れるか（個人休みの値を誤検出しない）。
    """
    c = unicodedata.normalize("NFKC", str(chunk or ""))
    if not c.strip():
        return False
    if re.search(r"臨時\s*休業", c):
        return True
    if "休場" in c:
        return True
    if re.search(r"工場", c) and re.search(r"休|休業|休み|坜止|よ休み", c):
        return True
    if re.search(r"(?:全社|全館|全工場).{0,15}(?:休|休業|坜止)", c):
        return True
    if re.search(r"(?:稼働|生産|ライン).{0,12}(?:坜止|なし|無し)", c):
        return True
    if re.search(r"加工.{0,15}(?:しない|無し|なし|よ休み)", c):
        return True
    if "休業" in c and re.search(
        r"(?:工場|全社|本社|当日|弊社|当社|全員|社全体)", c
    ):
        return True
    return False


def _md_slash_is_likely_fraction_not_date(t: str, start: int, end: int, mo: int, day: int) -> bool:
    """
    「加工速度は1/3とした」の 1/3 を 1月3日 と誤誝しない。
    「4/1は工場を休み」の 4/1 は日付のまま（直後は「は」なら分数扱いにしない）。
    """
    if mo <= 0 or day <= 0:
        return True
    before = t[max(0, start - 32) : start]
    after = t[end : min(len(t), end + 14)]
    after_st = after.lstrip()
    if after_st.startswith("は"):
        return False
    if re.search(
        r"(?:加工速度|加工\s*スピード|速度|倍率|スピード|効率|割引)(?:\s*は)?\s*$",
        before,
    ):
        return True
    # 1/2・1/3・2/3 等 + 「とした」「倝」… は分数・比率寄り（「3/1です」等の日付を誤スキップしないよご です/である は含まない）
    frac_pat = re.compile(
        r"^(?:としした?|とれる|倝|割引|にれる|に設定|しらい|程度|に固定|に変更)"
    )
    if mo <= 12 and day <= 12 and frac_pat.match(after_st):
        if mo <= 2 or (mo == 3 and day <= 3):
            return True
    # 「1/2です」「1/10です」のよごな分毝表睾（先頭は 1/ のみ）
    if (
        mo == 1
        and 2 <= day <= 12
        and re.match(r"^です|である\b", after_st)
    ):
        return True
    return False


def _extract_calendar_dates_from_text(s: str, default_year: int) -> list[date]:
    """グローバルコメント内の日付表記を date に変杛（基準年は計画の基準年）。"""
    t = unicodedata.normalize("NFKC", str(s or ""))
    found: list[date] = []
    seen: set[date] = set()

    def add(y: int, mo: int, d: int) -> None:
        try:
            dd = date(y, mo, d)
        except ValueError:
            return
        if dd not in seen:
            seen.add(dd)
            found.append(dd)

    for m in re.finditer(
        r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?",
        t,
    ):
        add(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for m in re.finditer(
        r"(\d{4})\s*[/\-\.＝]\s*(\d{1,2})\s*[/\-\.＝]\s*(\d{1,2})",
        t,
    ):
        add(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for m in re.finditer(r"(\d{1,2})\s*月\s*(\d{1,2})\s*日", t):
        add(int(default_year), int(m.group(1)), int(m.group(2)))
    for m in re.finditer(
        r"(?<!\d)(\d{1,2})\s*[/＝]\s*(\d{1,2})(?!\d)",
        t,
    ):
        mo_i, d_i = int(m.group(1)), int(m.group(2))
        if _md_slash_is_likely_fraction_not_date(t, m.start(), m.end(), mo_i, d_i):
            continue
        add(int(default_year), mo_i, d_i)
    return found


def _split_global_comment_into_chunks(blob: str) -> list[str]:
    """
    グローバルコメントを「独立した指示」の塊に分ける。
    改行（Excel の Alt+Enter・Unicode 改行含む）で必う分割し、同一行内は 。;； で続けて分割。
    """
    t = unicodedata.normalize("NFKC", str(blob or "").strip())
    if not t:
        return []
    lines = [ln.strip() for ln in re.split(r"[\n\r\v\f\u2028\u2029]+", t) if ln.strip()]
    if not lines:
        return []
    chunks: list[str] = []
    for line in lines:
        subs = [c.strip() for c in re.split(r"[。;；]+", line) if c.strip()]
        if subs:
            chunks.extend(subs)
        else:
            chunks.append(line)
    return chunks


def parse_factory_closure_dates_from_global_comment(
    text: str, default_year: int
) -> set[date]:
    """
    メインシート「グローバルコメント」に」工場臨時休業などと日付は書かれでいる場合に
    しの日を工場休み（全員非稼働・配台で加工している）として扱ご日付集合を返す。
    """
    blob = unicodedata.normalize("NFKC", str(text or "").strip())
    if not blob:
        return set()
    chunks = _split_global_comment_into_chunks(blob)
    if not chunks:
        chunks = [blob]
    out: set[date] = set()
    y0 = int(default_year)
    for ch in chunks:
        if not _global_comment_chunk_implies_factory_closure(ch):
            continue
        for d in _extract_calendar_dates_from_text(ch, y0):
            out.add(d)
    if not out and _global_comment_chunk_implies_factory_closure(blob):
        for d in _extract_calendar_dates_from_text(blob, y0):
            out.add(d)
    return out


def apply_factory_closure_dates_to_attendance(
    attendance_data: dict, members: list, closure_dates: set[date]
) -> None:
    """工場休業日: 勤怠上は全員 is_working=False とし、しの日は設備割付を行ゝない。"""
    if not closure_dates or not attendance_data:
        return
    tag = "工場休業（メイン・グローバルコメント）"
    for d in sorted(closure_dates):
        if d not in attendance_data:
            logging.warning(
                "グローバルコメントの工場休業日 %s はマスタ勤怠に行はありません。"
                " しの日は計画ループに含まれない場合」配台上の効果は限定的です。",
                d,
            )
            continue
        day = attendance_data[d]
        for m in members:
            if m not in day:
                continue
            ent = day[m]
            ent["is_working"] = False
            ent["eligible_for_assignment"] = False
            prev = str(ent.get("reason") or "").strip()
            ent["reason"] = f"{tag} {prev}".strip() if prev else tag


def _apply_global_priority_abolish_heuristic(blob: str, coerced: dict) -> dict:
    """
    「制限撤廃」「あらゆる条件」等: 設備専有・時刻ガードまで含む配台制約を緩める（abolish_all_scheduling_limits）。
    """
    b = unicodedata.normalize("NFKC", str(blob or ""))
    strong = (
        "制限撤廃",
        "制限を撤廃",
        "まとめての制限",
        "全での制限",
        "あらゆる制限",
        "あらゆる条件",
        "まとめての条件",
        "全での条件",
        "撤廃して",
        "撤廃し",
    )
    if any(k in b for k in strong):
        out = dict(coerced)
        out["abolish_all_scheduling_limits"] = True
        out["ignore_skill_requirements"] = True
        out["ignore_need_minimum"] = True
        logging.warning(
            "メイン再優先特記: 制限撤廃キーワードを検出。設備専有・時刻ガードを含む配台上の制約を緩めた。"
        )
        return out
    return coerced


def _maybe_fill_global_speed_rules_from_scheduler_notes(coerced: dict) -> dict:
    """
    AI は global_speed_rules を空にしたは scheduler_notes に具体パターンはある場合の補完。
    広し推測しない（熱融着＋検査＋1/3 系のみ）。
    """
    if not isinstance(coerced, dict):
        return coerced
    if coerced.get("global_speed_rules"):
        return coerced
    sn = str(coerced.get("scheduler_notes_ja") or "")
    t = unicodedata.normalize("NFKC", sn)
    if "熱融着" not in t or "検査" not in t:
        return coerced
    if not re.search(r"(?:1\s*/\s*3|１\s*/\s*3|三分の一|3\s*分の\s*1)", t):
        return coerced
    out = dict(coerced)
    out["global_speed_rules"] = [
        {
            "process_contains": "熱融着",
            "machine_contains": "検査",
            "speed_multiplier": 1.0 / 3.0,
        }
    ]
    logging.info(
        "メイン再優先特記: scheduler_notes_ja から global_speed_rules を補完（熱融着・検査・1/3）"
    )
    return out


def _finalize_global_priority_override(blob: str, coerced: dict) -> dict:
    """ソロ補正の後」abolish は true ならスキル・人数も強制オン。"""
    coerced = _maybe_fill_global_speed_rules_from_scheduler_notes(dict(coerced))
    coerced = _apply_global_priority_solo_heuristic(blob, coerced)
    coerced = _apply_global_priority_abolish_heuristic(blob, coerced)
    if coerced.get("abolish_all_scheduling_limits"):
        out = dict(coerced)
        out["ignore_skill_requirements"] = True
        out["ignore_need_minimum"] = True
        return out
    return coerced


def _apply_global_priority_solo_heuristic(blob: str, coerced: dict) -> dict:
    """
    「一人で担当」「独立」等で人数の値緩んでも」指定メンバーはスキル非該当てと配台されない。
    しの場合はスキル無視を同時に立でる。
    """
    if not coerced.get("ignore_need_minimum") or coerced.get("ignore_skill_requirements"):
        return coerced
    b = unicodedata.normalize("NFKC", str(blob or ""))
    solo_kw = ("一人", "参とり", "独立", "１人", "1人", "独自", "坘身")
    if any(k in b for k in solo_kw):
        out = dict(coerced)
        out["ignore_skill_requirements"] = True
        logging.info(
            "メイン再優先特記: 独立系キーワードのため、 ignore_skill_requirements を補助的に true にしました。"
        )
        return out
    return coerced


def _coerce_task_preferred_operators_dict(raw_val) -> dict:
    """AI の task_preferred_operators を {依頼NO: 並び} に正規化。"""
    out = {}
    if not isinstance(raw_val, dict):
        return out
    for k, v in raw_val.items():
        ks = str(k).strip()
        if not ks:
            continue
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        vs = str(v).strip()
        if vs and vs.lower() not in ("nan", "none", "null"):
            out[ks] = vs
    return out


def _normalize_factory_closure_dates_iso_list(val, default_year: int) -> list[str]:
    """
    AI またはフォールバックの日付リストを YYYY-MM-DD 文字列の昇順ユニークに正規化。
    覝素は ISO 文字列・Excel 日付・「4/1」程度の短文でも坯。
    """
    y0 = int(default_year)
    seen: set[str] = set()
    out: list[str] = []
    if not isinstance(val, list):
        return out
    for item in val:
        if item is None or (isinstance(item, float) and pd.isna(item)):
            continue
        d = parse_optional_date(item)
        if d is not None:
            iso = d.isoformat()
            if iso not in seen:
                seen.add(iso)
                out.append(iso)
            continue
        s = unicodedata.normalize("NFKC", str(item).strip())
        if not s:
            continue
        for d2 in _extract_calendar_dates_from_text(s, y0):
            iso = d2.isoformat()
            if iso not in seen:
                seen.add(iso)
                out.append(iso)
    return sorted(out)


def _coerce_global_speed_rules(raw_val) -> list[dict]:
    """
    Gemini の global_speed_rules を正規化。
    坄覝素: process_contains / machine_contains（いうれか必須・部分一致用）, speed_multiplier（既存速度に乗算」0超〜10以下）。
    """
    out: list[dict] = []
    if not isinstance(raw_val, list):
        return out
    for item in raw_val:
        if not isinstance(item, dict):
            continue
        sm = item.get("speed_multiplier")
        if sm is None:
            sm = item.get("relative_speed")
        try:
            mult = float(sm)
        except (TypeError, ValueError):
            continue
        if mult <= 0 or mult > 10.0:
            continue
        pps = unicodedata.normalize("NFKC", str(item.get("process_contains") or "")).strip()
        mms = unicodedata.normalize("NFKC", str(item.get("machine_contains") or "")).strip()
        if not pps and not mms:
            continue
        out.append(
            {
                "process_contains": pps,
                "machine_contains": mms,
                "speed_multiplier": mult,
            }
        )
    return out


def _global_speed_rule_substring_matches_row(pnorm: str, mnorm: str, sub_nfkc: str) -> bool:
    """sub は空でなけれみ」工程名または機械名のいうれかに部分一致れれみ True。"""
    if not sub_nfkc:
        return True
    return sub_nfkc in pnorm or sub_nfkc in mnorm


def _global_speed_multiplier_for_row(process_name: str, machine_name: str, rules: list) -> float:
    """
    工程名・機械名に一致するルールの speed_multiplier を掛け合わせる（一致なしは 1.0）。

    process_contains / machine_contains はしれずれ **工程名または機械名のどうらか** に含まれていればよい。
    両方指定時は AND（例: 「熱融着」と「検査」は」列の組み合わせで両方睾れる行にマッポ。
    マスタ上で工程=検査・機械=熱融着機 のよごにキーワードは逆坴の列にあっても同じルールで効く。
    """
    if not rules:
        return 1.0
    pnorm = unicodedata.normalize("NFKC", str(process_name or "")).strip()
    mnorm = unicodedata.normalize("NFKC", str(machine_name or "")).strip()
    combined = 1.0
    for r in rules:
        if not isinstance(r, dict):
            continue
        pc = unicodedata.normalize(
            "NFKC", str(r.get("process_contains") or "").strip()
        )
        mc = unicodedata.normalize(
            "NFKC", str(r.get("machine_contains") or "").strip()
        )
        if not pc and not mc:
            continue
        if pc and not _global_speed_rule_substring_matches_row(pnorm, mnorm, pc):
            continue
        if mc and not _global_speed_rule_substring_matches_row(pnorm, mnorm, mc):
            continue
        try:
            m = float(r.get("speed_multiplier", 1.0))
        except (TypeError, ValueError):
            continue
        if m <= 0:
            continue
        combined *= m
    if combined <= 0:
        return 1.0
    return combined


def _infer_global_day_process_rules_from_free_text(text: str, ref_y: int) -> list[dict]:
    """
    Gemini は task_preferred_operators に誤って長文を入れた場合など」
    自然言語断片から global_day_process_operator_rules 相当を推定する（保守的）。
    例: 「2026/4/4 工程名:EC 森下と宮島を配台」
    """
    t = unicodedata.normalize("NFKC", str(text or "")).strip()
    if len(t) < 6:
        return []
    dates = _extract_calendar_dates_from_text(t, int(ref_y))
    if not dates:
        return []
    d0 = dates[0]
    proc_m = re.search(
        r"工程名?\s*[:：]?\s*([A-Za-z0-9一-龯ー・〆々]+)",
        t,
    )
    pc = proc_m.group(1).strip() if proc_m else ""
    if not pc:
        m2 = re.search(r"([\dA-Za-z一-龯ー・〆々]{1,12})\s*工程", t)
        pc = m2.group(1).strip() if m2 else ""
    if not pc:
        return []
    names: list[str] = []
    for m in re.finditer(
        r"([\u3040-\u9FFF々ー・A-Za-z・〆々]{1,16}?)\s*と\s*([\u3040-\u9FFF々ー・A-Za-z・〆々]{1,16}?)\s*を?\s*(?:配台|酝属|組ませ|同一フォーム)",
        t,
    ):
        a, b = m.group(1).strip(), m.group(2).strip()
        if a:
            names.append(a)
        if b:
            names.append(b)
    if len(names) < 2:
        return []
    return [
        {
            "date": d0.isoformat(),
            "process_contains": pc,
            "operator_names": names[:12],
        }
    ]


def _salvage_malformed_global_priority_gemini_dict(raw: dict, ref_y: int) -> dict:
    """
    Gemini は task_preferred_operators に **配列**や誤スキーマ（workstation_id 等）を返したとし」
    杨でうに global_day_process_operator_rules / scheduler_notes_ja へ救済れる。
    """
    out = dict(raw)
    tpo = out.get("task_preferred_operators")
    if not isinstance(tpo, list) or not tpo:
        return out
    narratives: list[str] = []
    extra_rule_objs: list[dict] = []
    for item in tpo:
        if not isinstance(item, dict):
            continue
        onames = item.get("operator_names")
        if isinstance(onames, list) and (
            item.get("date") is not None or item.get("process_contains")
        ):
            extra_rule_objs.append(item)
            continue
        for key in ("workstation_id", "schedule_notes_ai", "schedule_notes", "note", "text"):
            s = str(item.get(key) or "").strip()
            if len(s) >= 12:
                narratives.append(s[:800])
        for _k, v in item.items():
            if _k in (
                "factory_closure_dates",
                "operator_names",
                "date",
                "process_contains",
            ):
                continue
            if isinstance(v, str) and len(v) > 35 and ("酝" in v or "工程" in v):
                narratives.append(v[:800])
    out["task_preferred_operators"] = {}
    gdp_existing = out.get("global_day_process_operator_rules")
    gdp_list: list = list(gdp_existing) if isinstance(gdp_existing, list) else []
    gdp_list.extend(extra_rule_objs)
    seen_n: set[str] = set()
    for nb in narratives:
        if nb in seen_n:
            continue
        seen_n.add(nb)
        gdp_list.extend(_infer_global_day_process_rules_from_free_text(nb, ref_y))
    out["global_day_process_operator_rules"] = gdp_list
    if narratives:
        sn0 = str(out.get("scheduler_notes_ja") or "").strip()
        add = " | ".join(n[:280] for n in narratives[:4])
        out["scheduler_notes_ja"] = (sn0 + " " + add).strip()[:600]
    return out


def _coerce_global_priority_override_dict(raw, reference_year: int | None = None) -> dict:
    """Gemini 戻りを配台用フラグ・工場休業日リストに正規化。"""
    y0 = int(reference_year) if reference_year is not None else date.today().year

    def as_bool(v):
        if v is True:
            return True
        if v is False:
            return False
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return False
        s = unicodedata.normalize("NFKC", str(v).strip()).lower()
        return s in ("true", "1", "yes", "はい", "on")

    base = {
        "ignore_skill_requirements": False,
        "ignore_need_minimum": False,
        "abolish_all_scheduling_limits": False,
        "task_preferred_operators": {},
        "interpretation_ja": "",
        "factory_closure_dates": [],
        "scheduler_notes_ja": "",
        "global_speed_rules": [],
        "global_day_process_operator_rules": [],
    }
    if not isinstance(raw, dict):
        return base
    raw = _salvage_malformed_global_priority_gemini_dict(raw, y0)
    base["ignore_skill_requirements"] = as_bool(raw.get("ignore_skill_requirements"))
    base["ignore_need_minimum"] = as_bool(raw.get("ignore_need_minimum"))
    base["abolish_all_scheduling_limits"] = as_bool(
        raw.get("abolish_all_scheduling_limits")
    )
    base["task_preferred_operators"] = _coerce_task_preferred_operators_dict(
        raw.get("task_preferred_operators")
    )
    ij = raw.get("interpretation_ja")
    if ij is not None and not (isinstance(ij, float) and pd.isna(ij)):
        base["interpretation_ja"] = str(ij).strip()[:800]
    base["factory_closure_dates"] = _normalize_factory_closure_dates_iso_list(
        raw.get("factory_closure_dates"), y0
    )
    sn = raw.get("scheduler_notes_ja")
    if sn is not None and not (isinstance(sn, float) and pd.isna(sn)):
        base["scheduler_notes_ja"] = str(sn).strip()[:600]
    base["global_speed_rules"] = _coerce_global_speed_rules(raw.get("global_speed_rules"))
    base["global_day_process_operator_rules"] = _coerce_global_day_process_operator_rules(
        raw.get("global_day_process_operator_rules")
    )
    return base


def _parse_global_priority_override_gemini_response(res):
    """Gemini 応答から JSON オブジェクト1つを取り出す（```json フェンス付しでも坯）。"""
    raw = (_gemini_result_text(res) or "").strip()
    if not raw:
        return None
    candidate = None
    fence = re.search(
        r"```(?:json)?\s*(\{.*\})\s*```",
        raw,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        candidate = fence.group(1).strip()
    elif raw.startswith("{"):
        candidate = raw
    else:
        loose = re.search(r"\{.*\}", raw, re.DOTALL)
        candidate = loose.group(0).strip() if loose else None
    if not candidate:
        return None
    try:
        parsed = json.loads(candidate)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _apply_regex_factory_closure_fallback(coerced: dict, blob: str, ref_y: int) -> dict:
    """Gemini 未使用・応答解釈失敗時: ルールベースで工場休業日の値補完（従来互換）。"""
    out = dict(coerced)
    rx = parse_factory_closure_dates_from_global_comment(blob, ref_y)
    out["factory_closure_dates"] = sorted({d.isoformat() for d in rx})
    return out


def analyze_global_priority_override_comment(
    text: str, members: list, reference_year: int, ai_sheet_sink: dict | None = None
) -> dict:
    """
    メインシート「グローバルコメント」（UI 上の自由記述）を **Gemini で一括解釈**し、配台に効し JSON に蝽とれ。
    自然言語の文脈切り分け・改行の別指示解釈は AI に任せ」戻り値のキーの値システムは機械適用する。

    - factory_closure_dates: **工場全体**で稼働しない日（全員非稼働扱い）の YYYY-MM-DD 文字列の配列。該当なしは []。
    - ignore_skill_requirements / ignore_need_minimum / abolish_all_scheduling_limits / task_preferred_operators: 従来どおり。
    - global_speed_rules: **工程名・機械名**への部分一致（坄キーワードは **どうらの列にあっても坯**）で」既存の加工速度（シート＝上書き後）に **乗算**れるルールの配列。該当なしは []。
    - global_day_process_operator_rules: **日付＋工程名の部分一致＋複数メンバー**を」当日しの工程のタスクの**フォーム全員に必う含むる**ルールの配列。該当なしは []。
    - scheduler_notes_ja: 上記に蝽とししれない補足や靋用メモ（速度は可能なら global_speed_rules も併記）。

    API キー無し・JSON 解釈失敗時: 上記ブール・指定は既定値」工場休業日のみ従来のルールベース解析で補完。
    """
    ref_y = int(reference_year) if reference_year is not None else date.today().year
    empty = _coerce_global_priority_override_dict({}, ref_y)
    if not text or not str(text).strip():
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "スキップ（メイン原文なし）"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（メイン原文なし・API 未実行）"
        return empty
    blob = str(text).strip()
    mem_sig = ",".join(sorted(str(m).strip() for m in (members or []) if m))
    cache_fingerprint = f"{GLOBAL_PRIORITY_OVERRIDE_CACHE_PREFIX}{ref_y}\n{blob}\n{mem_sig}"
    cache_key = hashlib.sha256(cache_fingerprint.encode("utf-8")).hexdigest()
    ai_cache = load_ai_cache()
    cached = get_cached_ai_result(ai_cache, cache_key, content_key=cache_fingerprint)
    if cached is not None:
        logging.info("メイン再優先特記: キャッシュヒット（Gemini は呼びません）。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "なし（キャッシュ使用）"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（キャッシュ利用・今回 API 未実行）"
        return _finalize_global_priority_override(
            blob, _coerce_global_priority_override_dict(cached, ref_y)
        )

    if not API_KEY:
        logging.info("GEMINI_API_KEY 未設定のため、メイン再優先特記の AI 解析をスキップしました。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "なし（APIキー未設定・工場休業のみルール補完）"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（API キー未設定）"
        coerced = _apply_regex_factory_closure_fallback(
            _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
        )
        return _finalize_global_priority_override(blob, coerced)

    member_sample = ", ".join(str(m) for m in (members or [])[:80])
    if len(members or []) > 80:
        member_sample += " …"

    prompt = f"""あなたは工場の配台計画システム用アシスタントです。
Excel メインシートの **「グローバルコメント」**（自由記述・自然言語）の **全文** を読み」次のキーの値を挝つ JSON を1つ返してください。

」役割】
ユーザーは改行や坥点で複数の指示を書きことはありした。**文脈を読み分け**」配台システムは **機械的に適用でしる値** に蝽とし込んでしてさい。
推測でブールを true にしないこと。根拠は明確なとしの値 true。

」最優先】
この欄の内容はマスタ・スキル・need・タスク行・特別指定_備考の AI 指定より優先される例外指示として扱ゝれした。

」改行・複数行】
坄行・坄文は **原則として独立した指示** です。行をまたいで1つにまとめたり」**割引表睾（例 1/3）を日付と結び付けたりしない**こと。

」キー別ルール】

A) **factory_closure_dates** （配列・必須）
   - **工場全体**は稼働しない日（臨時休業・全工場休み・しの日は加工しない等）の日付を **YYYY-MM-DD** の文字列で列挙。
   - **個人の休み・特定ラインの値**の坜止はここに **含まない**（[]）。
   - 該当はなけれみ **空の配列 []**（キー省略試行）。
   - 年は省略されでいれみ西暦 {ref_y} 年として解釈。

B) **ignore_skill_requirements** / **ignore_need_minimum** / **abolish_all_scheduling_limits** / **task_preferred_operators**
   - 従来どおり（配台のスキル無視・人数1固定・制限撤廃・依頼NO→主担当OP指定）。該当なけれみ false または {{}}。

C) **global_speed_rules** （配列・必須）
   - 特定の **工程名**（Excel「工程名」列）や **機械名**（「機械名」列）に対し、**既存の加工速度に掛ける倍率** を指定するオブジェクトのリスト。
   - 坄オブジェクトのキー:
     - "process_contains": 文字列（省略坯）。**工程名または機械名のいうれか**に **部分一致**（NFKC 想定）。
     - "machine_contains": 文字列（省略坯）。**工程名または機械名のいうれか**に **部分一致**。
     - "speed_multiplier": 正の数。**1/3 の速度**なら約 **0.333333**（既存速度 × この値）。**2倝速**なら 2.0。
   - **両方指定時は AND**（2つのキーワードは」**両方とも**「工程名・機械名のどうらか」に睾れる行）。例: 工程=検査・機械=熱融着機 でも」工程=熱融着・機械=検査用設備 でもマッポしごる。
   - どうらか一方の値指定れれみ」しのキーワードは工程名または機械名のどうらかにあれみマッポ。
   - 該当指示はなけれみ **空の配列 []**。
   - 例: 「熱融着を使う検査の加工速度は1/3」→
     [{{"process_contains":"熱融着","machine_contains":"検査","speed_multiplier":0.333333}}]
     （「熱融着」と「検査」は工程名・機械名の組み合わせで权ごタスクの速度は約1/3になる）

D) **scheduler_notes_ja** （文字列・必須）
   - 上記キーに蝽とししれない補足。速度は **global_speed_rules で構造化でしるとしは必うしうらにも出す**（ここは人間坑け覝約でもよい）。無ければ ""。

E) **interpretation_ja** （文字列・必須）
   - 原文の覝約を1文（200文字以内）。

F) **global_day_process_operator_rules** （配列・必須）
   - **特定の稼働日**かつ **工程名（Excel「工程名」列）の部分一致** に当ではまるタスクについで」
     列挙した **全メンバーを同一フォームに必う含むる** ルール（**OP/AS どうらのスキルでも坯**。並び解決は **担当OP指定とともに**）。
   - **依頼NOは分かる主担当の1坝指定**は **task_preferred_operators** を使うこと。原文は **「◯月◯日の△工程にＡとＢを配台」** のよごに **日付・工程・複数坝**のときは **本配列**へ蝽とれ。
   - 坄オブジェクトのキー:
     - "date": **YYYY-MM-DD**（しの日に割り当でるロールに適用）
     - "process_contains": 工程名に **部分一致**（NFKC 想定）。例: "EC"
     - "operator_names": 並びの配列（例: ["森下", "宮島　花孝"]）
   - 該当指示はなけれみ **空の配列 []**。

」返答形式】
先頭は {{ で終ゝりは }} の **JSON オブジェクト1つのみ**（説明文・マークダウン禁止）。

必須キー一覧:
- "factory_closure_dates": string の配列（YYYY-MM-DD）
- "ignore_skill_requirements": true または false
- "ignore_need_minimum": true または false
- "abolish_all_scheduling_limits": true または false
- "task_preferred_operators": **JSON オブジェクトのみ**（キー=依頼NO・値=主担当並び）。**配列にしてはならない**。該当なしは {{}}
- "global_speed_rules": オブジェクトの配列（該当なしは []）
- "global_day_process_operator_rules": オブジェクトの配列（該当なしは []）
- "scheduler_notes_ja": 文字列
- "interpretation_ja": 文字列

」基準年】 日付言坊はあれみ西暦 {ref_y} 年として解釈してよい。

」登録メンバー坝の参考】（照合用。JSON キーには含まない）
{member_sample}

」グローバルコメント・原文】
{blob}
"""
    try:
        ppath = os.path.join(log_dir, "ai_global_priority_override_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("メイン再優先特記: プロンプト全文 → %s", ppath)
    except OSError as ex:
        logging.warning("メイン再優先特記: プロンプト保存失敗: %s", ex)

    client = _gemini_client(API_KEY)
    try:
        res, gem_model_used = _gemini_generate_content_with_retry(
            client, contents=prompt, log_label="メイン再優先特記"
        )
        record_gemini_response_usage(res, gem_model_used)
        parsed = _parse_global_priority_override_gemini_response(res)
        if parsed is None:
            logging.warning(
                "メイン再優先特記: AI 応答から JSON を解釈でしませんでした。キャッシュせう」次回再試行されした。"
            )
            try:
                rpath = os.path.join(log_dir, "ai_global_priority_override_last_response.txt")
                with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                    rf.write(_gemini_result_text(res) or "")
            except OSError:
                pass
            if ai_sheet_sink is not None:
                ai_sheet_sink["メイン再優先特記_AI_API"] = "あり（JSON解釈失敗・工場休業はルール補完）"
                ai_sheet_sink["メイン再優先特記_Geminiモデル"] = gem_model_used
            coerced = _apply_regex_factory_closure_fallback(
                _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
            )
            return _finalize_global_priority_override(blob, coerced)
        coerced = _coerce_global_priority_override_dict(parsed, ref_y)
        coerced = _finalize_global_priority_override(blob, coerced)
        try:
            rpath = os.path.join(log_dir, "ai_global_priority_override_last_response.txt")
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(_gemini_result_text(res) or "")
        except OSError:
            pass
        put_cached_ai_result(ai_cache, cache_key, coerced, content_key=cache_fingerprint)
        save_ai_cache(ai_cache)
        _tpo = coerced.get("task_preferred_operators") or {}
        _fcd = coerced.get("factory_closure_dates") or []
        _gsr = coerced.get("global_speed_rules") or []
        _gdp = coerced.get("global_day_process_operator_rules") or []
        logging.info(
            "メイン再優先特記: AI 解釈 factory休業=%s日 速度ルール=%s件 日×工程フォーム=%s件 skill=%s need1=%s abolish=%s task_pref=%s件 — %s",
            len(_fcd),
            len(_gsr),
            len(_gdp),
            coerced["ignore_skill_requirements"],
            coerced["ignore_need_minimum"],
            coerced.get("abolish_all_scheduling_limits"),
            len(_tpo),
            coerced.get("interpretation_ja", "")[:100],
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "あり"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = gem_model_used
        return coerced
    except Exception as e:
        logging.warning("メイン再優先特記: Gemini 呼び出し失敗: %s", e)
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = f"失敗: {e}"[:500]
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（呼び出し失敗）"
        coerced = _apply_regex_factory_closure_fallback(
            _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
        )
        return _finalize_global_priority_override(blob, coerced)


def default_result_task_sheet_column_order(max_history_len: int) -> list:
    """結果_タスク一覧の既定列順（履歴列数は実行時に決まる）。"""
    hist = [f"履歴{i+1}" for i in range(max_history_len)]
    return [
        "ステータス",
        "タスクID",
        "工程名",
        "機械名",
        TASK_COL_SPEED,
        "優先度",
        RESULT_TASK_COL_DISPATCH_TRIAL_ORDER,
        *hist,
        "必須OP(上書)",
        "タスク効率",
        "加工途中",
        "特別指定あり",
        "担当OP指定",
        "回答納期",
        "指定納期",
        "計画基準納期",
        TASK_COL_RAW_INPUT_DATE,
        "紝期緊急",
        "加工開始日",
        "配台済_加工開始",
        "配台済_加工終了",
        RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16,
        "累計加工量",
        "残加工量",
        "完了率(実行時点)",
        "特別指定_AI",
    ]


def _task_date_key_for_result_sheet_sort(val):
    """結果_タスク一覧の並き替ご用。欠損・解釈試行は最後（date.max）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return date.max
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        ts = pd.Timestamp(val)
        if pd.isna(ts):
            return date.max
        return ts.date()
    except Exception:
        return date.max


def _coerce_planning_date_for_deadline(d) -> date | None:
    """回答納期・指定納期などを date に正規化（欠損は None）。"""
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return None


def _result_task_plan_end_within_answer_or_spec_16_label(
    plan_window: list | None, answer_due, specified_due
) -> str:
    """
    結果_タスク一覧用: 「配台済_加工終了」相当の最終終了は」
    回答納期の日付 + PLAN_DUE_DAY_COMPLETION_TIME（既定 16:00）以下かを判定。
    回答納期は無い行は指定納期の日付 + 16:00 で判定。
    両方無い場合は「紝期なし、。
    """
    if not plan_window or len(plan_window) < 2:
        return "未割当"
    _pe = plan_window[1]
    if _pe is None:
        return "未割当"
    dd = _coerce_planning_date_for_deadline(answer_due)
    if dd is None:
        dd = _coerce_planning_date_for_deadline(specified_due)
    if dd is None:
        return "紝期なし"
    try:
        deadline_dt = datetime.combine(dd, PLAN_DUE_DAY_COMPLETION_TIME)
        if _pe <= deadline_dt:
            return "はい"
        return "いいえ"
    except Exception:
        return "判定試行"


def _result_task_sheet_sort_key(t: dict):
    """
    結果_タスク一覧の表示順。①配台試行順番（generate_plan 冒頭でキュー順に付与した 1..n）昇順。
    欠損・非数は最後。同一試行順内は依頼NO・機械名」続けて加工開始日・紝期で安定化。
    """
    _dto = t.get("dispatch_trial_order")
    try:
        trial_k = int(_dto) if _dto is not None else 10**9
    except (TypeError, ValueError):
        trial_k = 10**9
    return (
        trial_k,
        str(t.get("task_id", "")).strip(),
        str(t.get("machine", "")).strip(),
        _task_date_key_for_result_sheet_sort(t.get("start_date_req")),
        _task_date_key_for_result_sheet_sort(t.get("answer_due_date")),
        _task_date_key_for_result_sheet_sort(t.get("specified_due_date")),
    )


def _is_result_task_history_expand_token(cell_val) -> bool:
    """列設定シートで「履歴」1行を置しと履歴1～n をしの佝置に展開れる。"""
    if cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val)):
        return False
    s = unicodedata.normalize("NFKC", str(cell_val).strip())
    return s in ("履歴", "履歴*")


def _result_task_column_alias_map(df_columns) -> dict:
    """見出しの NFKC 正規化キー → DataFrame 上の実列名。"""
    m = {}
    for c in df_columns:
        m[_nfkc_column_aliases(str(c).strip())] = c
    return m


def _resolve_result_task_column_label(label, col_by_norm: dict):
    if label is None or (isinstance(label, float) and pd.isna(label)):
        return None
    s = unicodedata.normalize("NFKC", str(label).strip())
    if not s or s.lower() in ("nan", "none"):
        return None
    nk = _nfkc_column_aliases(s)
    resolved = col_by_norm.get(nk)
    if resolved is not None:
        return resolved
    # 旧列名（計画基準納期ベース）→ 配台済_回答指定16時まで
    if nk == _nfkc_column_aliases("配台済_基準16時まで"):
        return col_by_norm.get(
            _nfkc_column_aliases(RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16)
        )
    # 旧見出し「原板投入日」→ 結果 DataFrame の「原反投入日」
    if nk == _nfkc_column_aliases("原板投入日"):
        return col_by_norm.get(_nfkc_column_aliases(TASK_COL_RAW_INPUT_DATE))
    return None


def _parse_column_visible_cell(val) -> bool:
    """表示列: 空・未記入は True（表示）。FALSE/0/いいえ 等で非表示。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return True
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if val == 0:
            return False
        if val == 1:
            return True
    s = unicodedata.normalize("NFKC", str(val).strip()).lower()
    if s in ("", "true", "1", "はい", "yes", "on", "表示", "○"):
        return True
    if s in ("false", "flase", "0", "いいえ", "no", "off", "非表示", "隠れ", "×"):
        return False
    return True


def parse_result_task_column_config_dataframe(
    df_cfg: pd.DataFrame | None, max_history_len: int
) -> list | None:
    """
    「列設定_結果_タスク一覧」相当の DataFrame から (列ラベル, 表示) を上から読む。
    見出し「列名」と「表示」（無い場合は表示はまとめて True）。
    「履歴」「履歴*」の1行は履歴1～履歴n に展開し、同一行の表示フラグを共有れる。
    同一列名（NFKC・別名正規化後）は複数行ある場合は先頭行のみ採用し、以降はログに出して杨でる。
    """
    if df_cfg is None or df_cfg.empty:
        return None
    df_cfg = df_cfg.dropna(how="all")
    if df_cfg.empty:
        return None

    name_col = None
    for c in df_cfg.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases(COLUMN_CONFIG_HEADER_COL):
            name_col = c
            break
    if name_col is None:
        name_col = df_cfg.columns[0]

    vis_col = None
    for c in df_cfg.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases(COLUMN_CONFIG_VISIBLE_COL):
            vis_col = c
            break

    seen_norm: set[str] = set()
    out: list[tuple[str, bool]] = []

    def _try_add(label: str, vis: bool) -> None:
        lab = str(label).strip()
        if not lab:
            return
        nk = _nfkc_column_aliases(unicodedata.normalize("NFKC", lab))
        if nk in seen_norm:
            logging.warning(
                "列設定「%s」: 重複列名「%s」をスキップしました（上の行を優先）。",
                COLUMN_CONFIG_SHEET_NAME,
                lab,
            )
            return
        seen_norm.add(nk)
        out.append((lab, vis))

    for i in range(len(df_cfg)):
        raw = df_cfg[name_col].iloc[i]
        vis = _parse_column_visible_cell(df_cfg[vis_col].iloc[i] if vis_col is not None else None)
        if _is_result_task_history_expand_token(raw):
            for j in range(max_history_len):
                _try_add(f"履歴{j+1}", vis)
            continue
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        s = unicodedata.normalize("NFKC", str(raw).strip())
        if not s or s.lower() in ("nan", "none"):
            continue
        _try_add(s, vis)
    return out or None


def _xlwings_write_column_config_sheet_ab(xw_sheet, rows: list[tuple[str, bool]]) -> None:
    """列設定シートの A:B を 列名・表示 のみで上書き（1行目見出し＋データ）。"""
    mat = [[COLUMN_CONFIG_HEADER_COL, COLUMN_CONFIG_VISIBLE_COL]]
    for lab, vis in rows:
        mat.append([lab, bool(vis)])
    n_r = len(mat)
    try:
        ur = xw_sheet.used_range
        lim_r = max(ur.row + ur.rows.count - 1, n_r, 2)
        xw_sheet.range((1, 1), (lim_r, 2)).clear_contents()
    except Exception:
        try:
            xw_sheet.range((1, 1)).resize(max(n_r, 50), 2).clear_contents()
        except Exception:
            pass
    xw_sheet.range((1, 1)).resize(n_r, 2).value = mat


def load_result_task_column_rows_from_input_workbook(max_history_len: int) -> list | None:
    """
    TASK_INPUT_WORKBOOK の「列設定_結果_タスク一覧」シートから (列ラベル, 表示) を上から読む。
    """
    wb = TASKS_INPUT_WORKBOOK
    if not wb or not os.path.exists(wb):
        return None
    if _workbook_should_skip_openpyxl_io(wb):
        logging.info(
            "列設定: ブックに「%s」があるため、pandas(openpyxl) での「%s」読込をスキップ（既定列順を使用した）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            COLUMN_CONFIG_SHEET_NAME,
        )
        return None
    try:
        df_cfg = pd.read_excel(wb, sheet_name=COLUMN_CONFIG_SHEET_NAME, header=0)
    except ValueError:
        return None
    except Exception as e:
        logging.warning(
            "シート「%s」: 読み込みに失敗したため、既定の列順を使用した (%s)",
            COLUMN_CONFIG_SHEET_NAME,
            e,
        )
        return None
    return parse_result_task_column_config_dataframe(df_cfg, max_history_len)


def _result_task_column_config_fallback_from_existing(
    df_tasks: pd.DataFrame, max_history_len: int
) -> tuple[list[str], dict[str, bool]]:
    """
    段階2で列順リストが空のときの補完。
    1) 結果 DataFrame に列があればその既存列順を採用し、マクロブック「列設定_結果_タスク一覧」
       で解決できる列は表示フラグを上書きする。
    2) 列が無ければ同シートから列名・表示を読む（TASK_INPUT_WORKBOOK・openpyxl 可のとき）。
    3) それも無ければ default_result_task_sheet_column_order。
    """
    rows_in = load_result_task_column_rows_from_input_workbook(max_history_len)
    cols = [str(c) for c in df_tasks.columns]

    if cols:
        vis_map = {c: True for c in cols}
        if rows_in:
            col_by_norm = _result_task_column_alias_map(cols)
            for item, vis in rows_in:
                resolved = _resolve_result_task_column_label(item, col_by_norm)
                if resolved and resolved in vis_map:
                    vis_map[resolved] = bool(vis)
        logging.warning(
            "段階2: 列順リストが空でした。結果 DataFrame の既存列（%s 列）で「%s」を補完しました。"
            + (" マクロブック列設定の表示フラグを反映しました。" if rows_in else ""),
            len(cols),
            COLUMN_CONFIG_SHEET_NAME,
        )
        return cols, vis_map

    if rows_in:
        order: list[str] = []
        vis_map: dict[str, bool] = {}
        for lab, vis in rows_in:
            order.append(lab)
            vis_map[lab] = bool(vis)
        logging.warning(
            "段階2: タスク行・列が無いため、マクロブック「%s」から %s 列で補完しました。",
            COLUMN_CONFIG_SHEET_NAME,
            len(order),
        )
        return order, vis_map

    dflt = list(default_result_task_sheet_column_order(max_history_len))
    logging.warning(
        "段階2: タスク行が 0 件かつ列設定の読込も無いため「%s」に既定の列名一覧を書き込みました。",
        COLUMN_CONFIG_SHEET_NAME,
    )
    return dflt, {c: True for c in dflt}


def apply_result_task_sheet_column_order(
    df: pd.DataFrame,
    max_history_len: int,
    *,
    config_dataframe: pd.DataFrame | None = None,
):
    """
    列設定シートはあれみしの順・表示を優先し、無い列は既定順で後ゝに追記（表示は True）。
    config_dataframe を渡した場合はファイルを読まうしの内容を列設定とみなす（xlwings 実行時用）。
    戻り値: (並き替ご後 DataFrame, 実際の列名リスト, 設定ソース説明文字列, 列名→表示bool)
    """
    default_order = default_result_task_sheet_column_order(max_history_len)
    if config_dataframe is not None:
        user_rows = parse_result_task_column_config_dataframe(config_dataframe, max_history_len)
    else:
        user_rows = load_result_task_column_rows_from_input_workbook(max_history_len)
    if user_rows:
        primary = user_rows
        source = (
            f"マクロブック「{COLUMN_CONFIG_SHEET_NAME}」"
            if config_dataframe is None
            else f"シート「{COLUMN_CONFIG_SHEET_NAME}」（実行中ブック）"
        )
    else:
        primary = [(n, True) for n in default_order]
        source = "既定"

    actual = list(df.columns)
    actual_set = set(actual)
    col_by_norm = _result_task_column_alias_map(actual)
    vis_map = {c: True for c in actual}

    seen = set()
    ordered = []
    unknown = []

    for item, vis in primary:
        resolved = _resolve_result_task_column_label(item, col_by_norm)
        if resolved and resolved not in seen:
            ordered.append(resolved)
            seen.add(resolved)
            vis_map[resolved] = vis
        elif not resolved:
            if item is not None and not (isinstance(item, float) and pd.isna(item)):
                lab = str(item).strip()
                if lab and lab not in unknown:
                    unknown.append(lab)

    for name in default_order:
        if name in actual_set and name not in seen:
            ordered.append(name)
            seen.add(name)
    for name in actual:
        if name not in seen:
            ordered.append(name)
            seen.add(name)

    if unknown:
        logging.warning(
            "列設定: 結果に無い列名を無視しました（最大20件）: %s",
            ", ".join(unknown[:20]) + (" …" if len(unknown) > 20 else ""),
        )
    logging.info("結果_タスク一覧の列順ソース: %s（%s 列）", source, len(ordered))
    if not user_rows and config_dataframe is None:
        logging.info(
            "列順・表示のカスタマイズ: マクロ実行ブックにシート「%s」を追加。"
            " 見出し「%s」「%s」… 表示は FALSE の列は結果シートで非表示。"
            " 1行「履歴」で履歴1～n を挿入。VBA の「列設定_結果_タスク一覧_チェックボックスを配置」でチェックボックスを表示列に連動可能。",
            COLUMN_CONFIG_SHEET_NAME,
            COLUMN_CONFIG_HEADER_COL,
            COLUMN_CONFIG_VISIBLE_COL,
        )
    return df[ordered], ordered, source, vis_map


def _xlwings_sheet_to_matrix(sheet) -> list:
    """xlwings Sheet の UsedRange を矩形の list[list] にれる（1行のみでも2次元）。"""
    ur = sheet.used_range
    if ur is None:
        return []
    raw = ur.options(ndim=2).value
    if raw is None:
        return []
    if not isinstance(raw, list):
        return [[raw]]
    if len(raw) == 0:
        return []
    if not isinstance(raw[0], list):
        return [raw]
    return raw


def _matrix_to_dataframe_header_first(matrix: list) -> pd.DataFrame | None:
    """1行目を列名とみなし DataFrame を返す。空なら None。"""
    if not matrix or not matrix[0]:
        return None
    header = []
    for x in matrix[0]:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            header.append("")
        else:
            header.append(str(x).strip())
    if not any(h for h in header):
        return None
    body = matrix[1:] if len(matrix) > 1 else []
    return pd.DataFrame(body, columns=header)


def _max_history_len_from_result_task_df_columns(columns) -> int:
    """結果_タスク一覧の「履歴n」列から n の最大を返す（無ければ 1）。"""
    imax = 0
    for c in columns:
        m = re.match(r"^履歴(\d+)$", str(c).strip())
        if m:
            imax = max(imax, int(m.group(1)))
    return max(imax, 1)


def apply_result_task_column_layout_via_xlwings(workbook_path: str | None = None) -> bool:
    """
    Excel で開いているマクロブックについで」
    「列設定_結果_タスク一覧」の内容に合わせで「結果_タスク一覧」の列順と列非表示を更新れる。
    「列設定_結果_タスク一覧」のセルは上書きしない（メモ・表外の A:B を消さない）。重複整理は
    dedupe_result_task_column_config_sheet_via_xlwings / VBA「重複列名を整理」を使う。
    ブックは事剝に保存し、本処理中も Excel 上で開いたままにれること（xlwings は接続れる）。
    """
    path = (workbook_path or "").strip() or TASKS_INPUT_WORKBOOK.strip()
    if not path:
        logging.error("結果_タスク一覧 列適用: ブックパスは空です（TASK_INPUT_WORKBOOK を設定してください）。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("結果_タスク一覧 列適用: xlwings は import でしません。pip install xlwings を確認してください。")
        return False

    try:
        wb = xw.Book(path)
    except Exception as e:
        logging.error("結果_タスク一覧 列適用: ブックに接続でしません: %s", e)
        return False

    try:
        ws_res = wb.sheets[RESULT_TASK_SHEET_NAME]
        ws_cfg = wb.sheets[COLUMN_CONFIG_SHEET_NAME]
    except Exception as e:
        logging.error("結果_タスク一覧 列適用: 必須シートは見つかりません: %s", e)
        return False

    mat_res = _xlwings_sheet_to_matrix(ws_res)
    mat_cfg = _xlwings_sheet_to_matrix(ws_cfg)
    df_res = _matrix_to_dataframe_header_first(mat_res)
    df_cfg = _matrix_to_dataframe_header_first(mat_cfg)
    if df_res is None or df_res.empty:
        logging.error("結果_タスク一覧 列適用: 「%s」にデータはありません。", RESULT_TASK_SHEET_NAME)
        return False
    if df_cfg is None:
        logging.error("結果_タスク一覧 列適用: 「%s」の見出しを読めません。", COLUMN_CONFIG_SHEET_NAME)
        return False

    max_h = _max_history_len_from_result_task_df_columns(df_res.columns)
    rows_cfg = parse_result_task_column_config_dataframe(df_cfg, max_h)
    if not rows_cfg:
        logging.error(
            "結果_タスク一覧 列適用: 「%s」に有効な列名行はありません。",
            COLUMN_CONFIG_SHEET_NAME,
        )
        return False
    # 列設定シートの A:B は上書きしない（clear_contents が UsedRange まで広がり、
    # 表外のメモ・余白の文字やチェック連動セルが消えるのを防ぐ）。並べ替え・非表示は結果シートのみ反映。
    df_cfg_clean = pd.DataFrame(
        rows_cfg, columns=[COLUMN_CONFIG_HEADER_COL, COLUMN_CONFIG_VISIBLE_COL]
    )
    df_out, ordered, source, vis_map = apply_result_task_sheet_column_order(
        df_res, max_h, config_dataframe=df_cfg_clean
    )

    df_write = df_out.astype(object).where(pd.notna(df_out), None)
    headers = [str(h) for h in df_write.columns.tolist()]
    body = df_write.values.tolist()
    out_matrix = [headers] + body
    nrows = len(out_matrix)
    ncols = len(headers)
    if ncols == 0:
        return False

    try:
        ur_old = ws_res.used_range
        if ur_old is not None:
            ws_res.range((ur_old.row, ur_old.column)).resize(
                ur_old.rows.count, ur_old.columns.count
            ).clear_contents()
    except Exception:
        try:
            ws_res.used_range.clear_contents()
        except Exception:
            pass

    ws_res.range((1, 1)).resize(nrows, ncols).value = out_matrix

    for ci in range(1, ncols + 1):
        try:
            ws_res.range((1, ci)).api.EntireColumn.Hidden = False
        except Exception:
            pass

    for ci, col_name in enumerate(ordered, 1):
        if not vis_map.get(col_name, True):
            try:
                ws_res.range((1, ci)).api.EntireColumn.Hidden = True
            except Exception as e:
                logging.warning("列非表示に失敗（列%s %s）: %s", ci, col_name, e)

    try:
        wb.save()
    except Exception as e:
        logging.warning("結果_タスク一覧 列適用: 保存で警告（データはシート上は更新済みの可能性）: %s", e)

    logging.info(
        "結果_タスク一覧 列適用完了: %s（%s 列」非表示=%s）",
        source,
        len(ordered),
        sum(1 for c in ordered if not vis_map.get(c, True)),
    )
    return True


def apply_result_task_column_layout_only() -> bool:
    """環境変数 TASK_INPUT_WORKBOOK のブックに対し列設定を適用する（VBA ボタン用）。"""
    p = os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK
    return apply_result_task_column_layout_via_xlwings(p)

_PLAN_INPUT_XLWINGS_ORIG_ROW = "__orig_sheet_row__"


def _plan_input_dispatch_trial_order_local_only_from_env() -> bool:
    """環境変数 PLAN_INPUT_DISPATCH_TRIAL_ORDER_LOCAL_ONLY は真なら post_load をスキップれる。"""
    v = (os.environ.get("PLAN_INPUT_DISPATCH_TRIAL_ORDER_LOCAL_ONLY") or "").strip().lower()
    return v in ("1", "true", "yes", "on", "y")


def refresh_plan_input_dispatch_trial_order_via_xlwings(
    workbook_path: str | None = None,
    *,
    apply_post_load_mutations: bool = True,
) -> bool:
    """
    Excel で開いたマクロブック内の「配台計画_タスク入力」について、
    段階2 と同じ ``fill_plan_dispatch_trial_order_column_stage1`` で「配台試行順番」を
    再付与し、段階1 出力直前と同じ手順で行を並べ替える。
    （未保存の編集分も xlwings で反映させるため read_excel は使わない）

    事前処理は ``_apply_planning_sheet_post_load_mutations``（設定シートの行同期・分割行の自動配台不要）。
    **「設定_配台不要工程」の C/E による計画シートへの配台不要の上書きは行わない**（段階1のみ。
    段階2の ``load_planning_tasks_df`` も同様に再適用しない）。シート上で消した「配台不要」は本経路では復活しない。
    """
    path = (workbook_path or "").strip() or os.environ.get(
        "TASK_INPUT_WORKBOOK", ""
    ).strip() or TASKS_INPUT_WORKBOOK.strip()
    if not path:
        logging.error("配台試行順番更新: ブックパスは空です。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("配台試行順番更新: xlwings はありません。")
        return False
    try:
        wb = xw.Book(path)
        ws = wb.sheets[PLAN_INPUT_SHEET_NAME]
    except Exception as e:
        logging.error("配台試行順番更新: シート接続に失敗: %s", e)
        return False

    mat = _xlwings_sheet_to_matrix(ws)
    df = _matrix_to_dataframe_header_first(mat)
    if df is None or df.empty:
        logging.warning("配台試行順番更新: データ行はありません。")
        return False

    df = df.copy()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, plan_input_sheet_column_order())
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""

    df.insert(0, _PLAN_INPUT_XLWINGS_ORIG_ROW, range(len(df)))

    _apply_planning_sheet_post_load_mutations(
        df,
        path,
        "配台試行順番更新",
        apply_exclude_rules_from_config=False,
        compile_exclude_rules_d_to_e_with_ai=False,
    )

    dto_col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    if dto_col not in df.columns:
        logging.error("配台試行順番更新: 列「%s」はありません。", dto_col)
        return False

    _dto_loc = df.columns.get_loc(dto_col)
    if isinstance(_dto_loc, slice):
        logging.error("配台試行順番更新: 列「%s」は複数ありした。", dto_col)
        return False
    if pd.api.types.is_numeric_dtype(df[dto_col]):
        df[dto_col] = float("nan")
    else:
        df[dto_col] = ""

    data_extract_dt, _ = _extract_data_extraction_datetime()
    base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
    run_date = base_now_dt.date()

    try:
        (
            _sd,
            _mem,
            equipment_list,
            req_map,
            need_rules,
            _sm,
            need_combo_col_index,
        ) = load_skills_and_needs()
    except Exception as e:
        logging.exception("配台試行順番更新: master 読込に失敗: %s", e)
        return False

    try:
        fill_plan_dispatch_trial_order_column_stage1(
            df,
            run_date,
            req_map,
            need_rules,
            need_combo_col_index,
            equipment_list,
        )
    except Exception as e:
        logging.exception("配台試行順番更新: 試行順計算に失敗: %s", e)
        return False

    df_sorted = _sort_stage1_plan_df_by_dispatch_trial_order_asc(df)
    orig_list = [int(x) for x in df_sorted[_PLAN_INPUT_XLWINGS_ORIG_ROW].tolist()]
    df_sorted = df_sorted.drop(columns=[_PLAN_INPUT_XLWINGS_ORIG_ROW])

    header_row = mat[0] if mat else []
    n_hdr = len(header_row)
    if n_hdr == 0:
        return False

    def _pad_row(r, n):
        r = list(r) if r is not None else []
        if len(r) < n:
            r = r + [None] * (n - len(r))
        return r

    new_mat = [_pad_row(header_row, n_hdr)]
    for i in range(len(df_sorted)):
        orig = orig_list[i]
        src_row = mat[orig + 1] if orig + 1 < len(mat) else []
        src_row = _pad_row(src_row, n_hdr)
        out_row = []
        for j in range(n_hdr):
            h_cell = header_row[j]
            if h_cell is None or (isinstance(h_cell, float) and pd.isna(h_cell)):
                hname = ""
            else:
                hname = str(h_cell).strip()
            if hname and hname in df_sorted.columns:
                v = df_sorted.iat[i, df_sorted.columns.get_loc(hname)]
                if pd.isna(v):
                    out_row.append(None)
                else:
                    out_row.append(v)
            else:
                out_row.append(src_row[j])
        new_mat.append(out_row)

    try:
        n_r = len(new_mat)
        ws.range((1, 1)).resize(n_r, n_hdr).value = new_mat
    except Exception as e:
        logging.exception("配台試行順番更新: シート書込に失敗: %s", e)
        return False

    try:
        wb.save()
    except Exception as e:
        logging.warning("配台試行順番更新: Save 警告: %s", e)

    logging.info(
        "配台試行順番更新: 「%s」を %s 行で更新しました。",
        PLAN_INPUT_SHEET_NAME,
        len(df_sorted),
    )
    return True


def refresh_plan_input_dispatch_trial_order_only() -> bool:
    """TASK_INPUT_WORKBOOK に対れる配台試行順番再計算（VBA / cmd 経由のエントリ）。
    環境変数 PLAN_INPUT_DISPATCH_TRIAL_ORDER_LOCAL_ONLY=1 等でシート上のセル値のみを入力とれる。
    """
    p = os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK
    local = _plan_input_dispatch_trial_order_local_only_from_env()
    return refresh_plan_input_dispatch_trial_order_via_xlwings(
        p, apply_post_load_mutations=not local
    )


def _plan_input_row_is_blank_task_row(plan_df: "pd.DataFrame", row_i: int) -> bool:
    """依頼NO・工程名が両方空なら True（並べ替え・検証の対象外）。"""

    def _cell_empty(val) -> bool:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return True
        s = str(val).strip()
        if not s or s.lower() in ("nan", "none"):
            return True
        return False

    if TASK_COL_TASK_ID not in plan_df.columns or TASK_COL_MACHINE not in plan_df.columns:
        return True
    ti = plan_df.iat[row_i, plan_df.columns.get_loc(TASK_COL_TASK_ID)]
    mc = plan_df.iat[row_i, plan_df.columns.get_loc(TASK_COL_MACHINE)]
    return _cell_empty(ti) and _cell_empty(mc)


def _parse_dispatch_trial_order_float_sort_key(val) -> float | None:
    """「配台試行順番」セルを並べ替えキーとして float 化。空・不正・非有限は None。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            x = float(val)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(x):
            return None
        return x
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    try:
        x = float(s)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    return x


def sort_plan_input_dispatch_trial_order_by_float_keys_via_xlwings(
    workbook_path: str | None = None,
) -> bool:
    """
    「配台計画_タスク入力」の **現在のシート内容だけ** を使い、列「配台試行順番」を
    小数を含む並べ替えキーとして解釈して昇順に行を並べ替え、1..n に振り直す。

    - ``_apply_planning_sheet_post_load_mutations`` ・マスタ ・
      ``fill_plan_dispatch_trial_order_column_stage1`` は **呼ばない**。
    - 依頼NO・工程名が両方空の行は対象外。先頭の空行と、最後のデータ行より後の空行は
      元の順のまま残す。
    - 最初の対象行から最後の対象行までは **途切れなく対象行** でなければならない。
    - **有限の float** として解釈できるキー同士は **重複してはならない**。
    - キーが空・解釈不能の対象行は、**すべての有効キー行の後ろ**に元の行順を保って並べ、
      連番 1..n はその並びで振り直す。
    """
    path = (workbook_path or "").strip() or os.environ.get(
        "TASK_INPUT_WORKBOOK", ""
    ).strip() or TASKS_INPUT_WORKBOOK.strip()
    if not path:
        logging.error("配台試行順番（小数キー並べ）: ブックパスが空です。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("配台試行順番（小数キー並べ）: xlwings がありません。")
        return False
    try:
        wb = xw.Book(path)
        ws = wb.sheets[PLAN_INPUT_SHEET_NAME]
    except Exception as e:
        logging.error("配台試行順番（小数キー並べ）: シート接続に失敗: %s", e)
        return False

    mat = _xlwings_sheet_to_matrix(ws)
    df = _matrix_to_dataframe_header_first(mat)
    if df is None or df.empty:
        logging.warning("配台試行順番（小数キー並べ）: データ行がありません。")
        return False

    df = df.copy()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, plan_input_sheet_column_order())
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""

    dto_col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    if dto_col not in df.columns:
        logging.error("配台試行順番（小数キー並べ）: 列「%s」がありません。", dto_col)
        return False
    dto_idx = df.columns.get_loc(dto_col)
    if isinstance(dto_idx, slice):
        logging.error("配台試行順番（小数キー並べ）: 列「%s」が複数あります。", dto_col)
        return False

    n = len(df)
    active = [i for i in range(n) if not _plan_input_row_is_blank_task_row(df, i)]
    if not active:
        logging.error(
            "配台試行順番（小数キー並べ）: 依頼NO または 工程名 がある行がありません。"
        )
        return False
    first = min(active)
    last = max(active)
    for k in range(first, last + 1):
        if k not in active:
            logging.error(
                "配台試行順番（小数キー並べ）: %s 行目付近に、依頼NO・工程名が両方空の行が"
                " データの途中にあります。",
                k + 2,
            )
            return False

    row_by_key: dict[float, int] = {}
    sort_tuple_by_row: dict[int, tuple] = {}
    n_invalid_key = 0
    for i in active:
        fk = _parse_dispatch_trial_order_float_sort_key(df.iat[i, dto_idx])
        if fk is None:
            n_invalid_key += 1
            # 有効 float より後ろ。同帯は元の行番号で安定化。
            sort_tuple_by_row[i] = (1, i)
            continue
        if fk in row_by_key:
            logging.error(
                "配台試行順番（小数キー並べ）: 並べ替えキー %s が %s 行目と %s 行目で重複しています。",
                fk,
                row_by_key[fk] + 2,
                i + 2,
            )
            return False
        row_by_key[fk] = i
        sort_tuple_by_row[i] = (0, fk, i)

    if n_invalid_key:
        logging.info(
            "配台試行順番（小数キー並べ）: 「%s」が空・非数値のデータ行が %s 行あります。"
            " 有効キー行の後ろに並べ、連番化します。",
            dto_col,
            n_invalid_key,
        )

    sorted_active = sorted(active, key=lambda ri: sort_tuple_by_row[ri])
    df_mut = df.copy()
    for rank, i in enumerate(sorted_active, start=1):
        df_mut.iat[i, dto_idx] = rank

    leading = [i for i in range(0, first)]
    trailing = [i for i in range(last + 1, n)]
    orig_list = leading + sorted_active + trailing

    rows_ordered = [df_mut.iloc[oi] for oi in orig_list]
    df_sorted = pd.DataFrame(rows_ordered).reset_index(drop=True)

    header_row = mat[0] if mat else []
    n_hdr = len(header_row)
    if n_hdr == 0:
        return False

    def _pad_row(r, n):
        r = list(r) if r is not None else []
        if len(r) < n:
            r = r + [None] * (n - len(r))
        return r

    new_mat = [_pad_row(header_row, n_hdr)]
    for i in range(len(df_sorted)):
        orig = orig_list[i]
        src_row = mat[orig + 1] if orig + 1 < len(mat) else []
        src_row = _pad_row(src_row, n_hdr)
        out_row = []
        for j in range(n_hdr):
            h_cell = header_row[j]
            if h_cell is None or (isinstance(h_cell, float) and pd.isna(h_cell)):
                hname = ""
            else:
                hname = str(h_cell).strip()
            if hname and hname in df_sorted.columns:
                v = df_sorted.iat[i, df_sorted.columns.get_loc(hname)]
                if pd.isna(v):
                    out_row.append(None)
                else:
                    out_row.append(v)
            else:
                out_row.append(src_row[j])
        new_mat.append(out_row)

    try:
        n_r = len(new_mat)
        ws.range((1, 1)).resize(n_r, n_hdr).value = new_mat
    except Exception as e:
        logging.exception("配台試行順番（小数キー並べ）: シート書込に失敗: %s", e)
        return False

    try:
        wb.save()
    except Exception as e:
        logging.warning("配台試行順番（小数キー並べ）: Save 警告: %s", e)

    logging.info(
        "配台試行順番（小数キー並べ）: 「%s」を %s データ行で並べ替え・連番化しました。",
        PLAN_INPUT_SHEET_NAME,
        len(sorted_active),
    )
    return True


def sort_plan_input_dispatch_trial_order_by_float_keys_only() -> bool:
    """TASK_INPUT_WORKBOOK に対する「小数キーで並べ替え→1..n」（VBA / cmd 経由）。"""
    p = os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK
    return sort_plan_input_dispatch_trial_order_by_float_keys_via_xlwings(p)


def apply_plan_input_column_layout_only() -> bool:
    """
    配台計画_タスク入力の列順・表示のみを適用する予定（VBA 用）。
    未実装。列の並よは段階1出力または手動整睆を使用してください。
    """
    logging.warning("apply_plan_input_column_layout_only: not implemented")
    return False



def dedupe_result_task_column_config_sheet_via_xlwings(workbook_path: str | None = None) -> bool:
    """
    「列設定_結果_タスク一覧」の A:B の値を」重複列名を除いた一覧で書き直れ（先の行を優先）。
    「結果_タスク一覧」はあれみ履歴列数の解釈に使う。結果シートは変更しない。
    """
    path = (workbook_path or "").strip() or TASKS_INPUT_WORKBOOK.strip()
    if not path:
        logging.error("列設定 重複整睆: ブックパスは空です。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("列設定 重複整睆: xlwings は import でしません。")
        return False
    try:
        wb = xw.Book(path)
        ws_cfg = wb.sheets[COLUMN_CONFIG_SHEET_NAME]
    except Exception as e:
        logging.error("列設定 重複整睆: 接続またはシート取得に失敗: %s", e)
        return False

    max_h = 1
    try:
        ws_res = wb.sheets[RESULT_TASK_SHEET_NAME]
        df_r = _matrix_to_dataframe_header_first(_xlwings_sheet_to_matrix(ws_res))
        if df_r is not None and not df_r.empty:
            max_h = _max_history_len_from_result_task_df_columns(df_r.columns)
    except Exception:
        pass

    df_cfg = _matrix_to_dataframe_header_first(_xlwings_sheet_to_matrix(ws_cfg))
    if df_cfg is None:
        logging.error("列設定 重複整睆: 「%s」の見出しを読めません。", COLUMN_CONFIG_SHEET_NAME)
        return False
    rows = parse_result_task_column_config_dataframe(df_cfg, max_h)
    if not rows:
        logging.warning("列設定 重複整睆: 有効なデータ行はありません。")
        return False
    _xlwings_write_column_config_sheet_ab(ws_cfg, rows)
    try:
        wb.save()
    except Exception as e:
        logging.warning("列設定 重複整睆: 保存警告: %s", e)
    logging.info(
        "列設定「%s」を重複除去済みで %s 行に整睆しました（履歴展開後の行数）。",
        COLUMN_CONFIG_SHEET_NAME,
        len(rows),
    )
    return True


def dedupe_result_task_column_config_sheet_only() -> bool:
    """環境変数 TASK_INPUT_WORKBOOK のブックの列設定シートの値重複整睆（VBA 用）。"""
    p = os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK
    return dedupe_result_task_column_config_sheet_via_xlwings(p)


def _apply_result_task_sheet_column_visibility(worksheet, column_names: list, vis_map: dict):
    """結果_タスク一覧で」vis_map は False の列を非表示にれる。"""
    for idx, col_name in enumerate(column_names, 1):
        if not vis_map.get(col_name, True):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True


def _norm_history_member_label(name: str) -> str:
    """履歴の担当坝比較用（全角空白を半角1個化・剝後trim・連続空白の圧縮）。"""
    t = str(name or "").replace("\u3000", " ").strip()
    return " ".join(t.split())


def _history_team_text_main_assignment_only(h: dict) -> str:
    """
    結果シート「担当」欄用: メイン割付確定時点の坝剝（余力追記サブは含まない）。
    append_surplus 後の h['team'] から post_dispatch_surplus_names を除外れる。
    """
    raw = (h.get("team") or "").strip()
    if not raw:
        return ""
    ps = h.get("post_dispatch_surplus_names") or []
    if not ps:
        return raw
    ps_set = {
        _norm_history_member_label(x)
        for x in ps
        if x and str(x).strip()
    }
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    kept = [p for p in parts if _norm_history_member_label(p) not in ps_set]
    return ", ".join(kept) if kept else raw


def _result_assigned_history_team_key(team_s: str) -> str:
    """結果シート用: 履歴セグメント同士の担当文字列比較（NFKC・空白正規化）。"""
    s = unicodedata.normalize("NFKC", str(team_s or "").strip())
    return " ".join(s.split())


def _union_name_lists_preserve_order(
    a: list | None, b: list | None
) -> list[str]:
    """名前列を重複なく結合（先勝ちで順序維持）。"""
    out: list[str] = []
    seen: set[str] = set()
    for xs in (a or [], b or []):
        for x in xs:
            t = str(x).strip()
            if not t:
                continue
            k = _norm_history_member_label(t)
            if k in seen:
                continue
            seen.add(k)
            out.append(t)
    return out


def _assigned_history_segment_copy(h: dict) -> dict:
    """履歴 dict の浅いコピー（名前リストは複製してマージ時に汚染しない）。"""
    out = dict(h)
    for k in ("surplus_member_names", "post_dispatch_surplus_names"):
        v = out.get(k)
        if isinstance(v, list):
            out[k] = list(v)
    return out


def _assigned_history_contiguous_mergeable(a: dict, b: dict) -> bool:
    """
    連続作業として 1 履歴にまとめられるか。
    前セグメント終了 == 次セグメント開始・同一担当・同一組合せ行 ID のときのみ。
    """
    a_end = a.get("end_dt")
    b_start = b.get("start_dt")
    if not isinstance(a_end, datetime) or not isinstance(b_start, datetime):
        return False
    if a_end != b_start:
        return False
    if _result_assigned_history_team_key(a.get("team", "")) != _result_assigned_history_team_key(
        b.get("team", "")
    ):
        return False
    return (a.get("combo_sheet_row_id")) == (b.get("combo_sheet_row_id"))


def _merge_two_assigned_history_display_segments(a: dict, b: dict) -> dict:
    """連続セグメント b を a に取り込んだ新 dict（結果シート表示専用）。"""
    out = _assigned_history_segment_copy(a)
    try:
        da = int(a.get("done_m") or 0)
    except (TypeError, ValueError):
        da = 0
    try:
        db = int(b.get("done_m") or 0)
    except (TypeError, ValueError):
        db = 0
    out["done_m"] = da + db
    out["end_dt"] = b.get("end_dt")
    out["need_surplus_assigned"] = bool(
        a.get("need_surplus_assigned") or b.get("need_surplus_assigned")
    )
    out["surplus_member_names"] = _union_name_lists_preserve_order(
        a.get("surplus_member_names"), b.get("surplus_member_names")
    )
    out["post_dispatch_surplus_names"] = _union_name_lists_preserve_order(
        a.get("post_dispatch_surplus_names"), b.get("post_dispatch_surplus_names")
    )
    return out


def merge_assigned_history_contiguous_for_result_sheet(hist: list | None) -> list:
    """
    結果_タスク一覧向け: ロール確定ごとの内部履歴を、時刻・担当・組合せ ID が連続する塊で 1 件にまとめる。
    配台中のロールパイプライン等は生の assigned_history を参照するため、本関数は出力直前にのみ使う。
    """
    hist = hist or []
    if len(hist) < 2:
        return [_assigned_history_segment_copy(h) for h in hist]
    out: list[dict] = []
    cur = _assigned_history_segment_copy(hist[0])
    for nxt_raw in hist[1:]:
        nxt = _assigned_history_segment_copy(nxt_raw)
        if _assigned_history_contiguous_mergeable(cur, nxt):
            cur = _merge_two_assigned_history_display_segments(cur, nxt)
        else:
            out.append(cur)
            cur = nxt
    out.append(cur)
    return out


def _format_result_task_history_cell(task: dict, h: dict) -> str:
    """結果_タスク一覧の履歴セル文字列（短い記号: #=組合せ行ID, 主=メイン担当, +=超過, 余=余力追記）。"""
    um = task.get("unit_m") or 0
    try:
        done_r = int(h["done_m"] / um) if um else 0
    except (TypeError, ValueError, ZeroDivisionError):
        done_r = 0
    dm = h.get("done_m", 0)
    d = h.get("date", "") or ""
    parts_out: list[str] = [f"・【{d}】：{done_r}R/{dm}m"]
    cid = h.get("combo_sheet_row_id")
    if cid is not None:
        try:
            parts_out.append(f"#{int(cid)}")
        except (TypeError, ValueError):
            parts_out.append(f"#{cid}")
    team = _history_team_text_main_assignment_only(h)
    if team:
        parts_out.append(f"主:{team}")
    sm = h.get("surplus_member_names") or []
    if sm:
        parts_out.append("+" + ",".join(str(x) for x in sm))
    ps = h.get("post_dispatch_surplus_names") or []
    if ps:
        parts_out.append("余:" + ",".join(str(x) for x in ps))
    return " ".join(parts_out)


_RESULT_TASK_HISTORY_RICH_HEAD_RE = re.compile(r"^・(【[^】]*】)(.*)$", re.DOTALL)


def _apply_result_task_history_rich_text(worksheet, column_names: list):
    """
    履歴列: 「・【日付】：…」の日付括弧部分を青色リッチテキストにする。
    openpyxl 3.1 未満ではスキップ（文字列の【】のみ）。
    """
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles.colors import Color
    except ImportError:
        return

    hist_cols = [
        i + 1 for i, c in enumerate(column_names) if str(c).startswith("履歴")
    ]
    if not hist_cols:
        return

    _plain_kw: dict = {}
    _blue_kw: dict = {"color": Color(rgb="FF0070C0")}
    plain_if = InlineFont(**_plain_kw)
    blue_if = InlineFont(**_blue_kw)
    top = Alignment(wrap_text=False, vertical="top")

    for r in range(2, worksheet.max_row + 1):
        for ci in hist_cols:
            cell = worksheet.cell(row=r, column=ci)
            v = cell.value
            if not isinstance(v, str) or not v.startswith("・【"):
                continue
            m = _RESULT_TASK_HISTORY_RICH_HEAD_RE.match(v)
            if not m:
                continue
            bracketed, rest = m.group(1), m.group(2)
            cell.value = CellRichText(
                TextBlock(plain_if, "・"),
                TextBlock(blue_if, bracketed),
                TextBlock(plain_if, rest),
            )
            cell.alignment = top


def _apply_result_task_date_columns_blue_font(worksheet, column_names: list):
    """
    結果_タスク一覧: 回答納期・指定納期・計画基準納期・原反投入日・加工開始日のセルを青色にれる。
    （履歴列の【日付】は _apply_result_task_history_rich_text で着色。色は 0070C0 で統一）
    """
    blue = _result_font(color="0070C0")
    top = Alignment(wrap_text=False, vertical="top")
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) not in RESULT_TASK_DATE_STYLE_HEADERS:
            continue
        for r in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=r, column=col_idx)
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not str(v).strip():
                continue
            cell.font = blue
            cell.alignment = top


def _apply_result_task_history_need_surplus_highlight(
    worksheet, column_names: list, sorted_tasks: list
):
    """
    need「配台時追加人数」相当で基本必須人数を超ごで採用したブロック」または
    メイン完了後の余力追記でサブは増ごたブロックに対応れる「履歴n」セルを薄黄に塗る。
    """
    hist_cols: list[tuple[int, int]] = []
    for col_idx, col_name in enumerate(column_names, 1):
        m = re.match(r"^履歴(\d+)$", str(col_name).strip())
        if m:
            hist_cols.append((int(m.group(1)), col_idx))
    hist_cols.sort(key=lambda x: x[0])
    if not hist_cols or worksheet.max_row < 2:
        return
    fill_surplus = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    n_tasks = len(sorted_tasks)
    for r in range(2, worksheet.max_row + 1):
        ti = r - 2
        if ti < 0 or ti >= n_tasks:
            continue
        ah = merge_assigned_history_contiguous_for_result_sheet(
            sorted_tasks[ti].get("assigned_history")
        )
        for ord1, cidx in hist_cols:
            i = ord1 - 1
            if i < 0 or i >= len(ah):
                continue
            if not ah[i].get("need_surplus_assigned"):
                continue
            worksheet.cell(row=r, column=cidx).fill = fill_surplus


def _apply_result_task_task_id_content_mismatch_highlight(
    worksheet, column_names: list, sorted_tasks: list
):
    """
    加工内容に工程名は含まれない行の「タスクID」セルを赤背景・白文字にする（元データとの整合の確認用）。
    """
    task_id_col_idx = None
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) == "タスクID":
            task_id_col_idx = col_idx
            break
    if task_id_col_idx is None or worksheet.max_row < 2:
        return
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white = _result_font(color="FFFFFF")
    top = Alignment(wrap_text=False, vertical="top")
    n_data = worksheet.max_row - 1
    for i in range(min(len(sorted_tasks), n_data)):
        if not sorted_tasks[i].get("process_content_mismatch"):
            continue
        cell = worksheet.cell(row=i + 2, column=task_id_col_idx)
        cell.fill = fill_red
        cell.font = font_white
        cell.alignment = top


def _apply_result_task_plan_end_answer_spec_16_no_highlight(
    worksheet, column_names: list
):
    """
    列「配台済_回答指定16時まで」は「いいえ」のセルを赤背景・白文字・太字にれる。
    列設定で旧坝「配台済_基準16時まで」のままの見出しにも対応。
    """
    target_names = frozenset(
        {
            RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16,
            "配台済_基準16時まで",
        }
    )
    col_idx = None
    for ci, col_name in enumerate(column_names, 1):
        if str(col_name) in target_names:
            col_idx = ci
            break
    if col_idx is None or worksheet.max_row < 2:
        return
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white_bold = _result_font(color="FFFFFF", bold=True)
    top = Alignment(wrap_text=False, vertical="top")
    for r in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if s != "いいえ":
            continue
        cell.fill = fill_red
        cell.font = font_white_bold
        cell.alignment = top


def _apply_result_task_id_hyperlinks_to_equipment_schedule(
    worksheet_tasks,
    column_names: list,
    sorted_tasks_for_row_order: list,
    task_id_to_schedule_cell: dict[str, str],
    schedule_sheet_name: str,
) -> None:
    """
    結果_タスク一覧の「タスクID」セルに」結果_設備毎の時間割で当該タスクは最初に睾れるセルへの内部ポイパーリンクを付与れる。
    時間割に睾れないタスク（未割当のみ等）はリンクなし。
    """
    if not task_id_to_schedule_cell or worksheet_tasks.max_row < 2:
        return
    task_id_col_idx = None
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) == "タスクID":
            task_id_col_idx = col_idx
            break
    if task_id_col_idx is None:
        return
    esc = schedule_sheet_name.replace("'", "''")
    loc_prefix = f"#'{esc}'!"
    font_link = Font(color="0563C1", underline="single")
    font_link_on_red = Font(color="FFFFFF", underline="single")
    top = Alignment(wrap_text=False, vertical="top")
    n_tasks = len(sorted_tasks_for_row_order)
    for r in range(2, worksheet_tasks.max_row + 1):
        cell = worksheet_tasks.cell(row=r, column=task_id_col_idx)
        raw = cell.value
        if raw is None:
            continue
        tid = str(raw).strip()
        if not tid:
            continue
        addr = task_id_to_schedule_cell.get(tid)
        if not addr:
            continue
        cell.hyperlink = loc_prefix + addr
        row_i = r - 2
        mismatch = (
            row_i < n_tasks
            and bool(sorted_tasks_for_row_order[row_i].get("process_content_mismatch"))
        )
        cell.font = font_link_on_red if mismatch else font_link
        cell.alignment = top


def _add_column_config_sheet_helpers(ws_cfg, num_data_rows: int):
    """表示列に TRUE/FALSE リスト（チェックの代ゝりにプルダウン）を付与。"""
    last_r = max(num_data_rows + 1, 2)
    cap = max(last_r + 50, 500)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws_cfg.add_data_validation(dv)
    dv.add(f"B2:B{cap}")


def _stage2_try_copy_column_config_shapes_from_input(
    result_path: str,
    input_path: str | None,
) -> None:
    """
    pandas/openｎｎxl で新規作成した結果ブックには図形が含まれない。
    既定で有効（環境変数で 0/false/no/off のとき無効）。入力ブックの
    「列設定_結果_タスク一覧」上の **Shapes**（フォームのボタン・チェックボックス等）と
    **OLEObjects**（ActiveX コントロール等）を結果ブックの同名シートへコピーし、
    各図形の Left/Top/Width/Height（および取れるとき Placement）を入力側と同じに戻す。
    openpyxl による当該ブックへの保存がすべて終わった後に呼ぶこと。
    """
    if not STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT:
        return
    rp = (result_path or "").strip()
    ip = (input_path or "").strip()
    if not rp or not os.path.isfile(rp):
        logging.warning(
            "列設定シート図形コピー: 結果パスは無効のため、スキップしました。"
        )
        return
    if not ip or not os.path.isfile(ip):
        logging.warning(
            "列設定シート図形コピー: TASK_INPUT_WORKBOOK は無効のため、スキップしました。"
        )
        return
    try:
        import xlwings as xw
    except ImportError:
        logging.warning(
            "列設定シート図形コピー: xlwings は import でしません。"
        )
        return
    app = None
    wb_out = None
    wb_in = None
    try:
        app = xw.App(visible=False)
        app.display_alerts = False
        wb_out = app.books.open(os.path.abspath(rp), update_links=False)
        wb_in = app.books.open(os.path.abspath(ip), read_only=True, update_links=False)
        try:
            ws_out = wb_out.sheets[COLUMN_CONFIG_SHEET_NAME]
        except Exception:
            logging.warning(
                "列設定シート図形コピー: 結果ブックにシート「%s」はありません。",
                COLUMN_CONFIG_SHEET_NAME,
            )
            return
        try:
            ws_in = wb_in.sheets[COLUMN_CONFIG_SHEET_NAME]
        except Exception:
            logging.warning(
                "列設定シート図形コピー: 入力ブックにシート「%s」はありません。",
                COLUMN_CONFIG_SHEET_NAME,
            )
            return
        api_in = ws_in.api
        api_out = ws_out.api
        n_shapes = int(api_in.Shapes.Count)
        try:
            n_ole = int(api_in.OLEObjects.Count)
        except Exception:
            n_ole = 0
        if n_shapes <= 0 and n_ole <= 0:
            logging.info(
                "列設定シート図形コピー: 入力側に Shapes（フォーム等）も "
                "OLEObjects（ActiveX 等）もありません（スキップ）。"
            )
            return
        ws_out.activate()
        for i in range(1, n_shapes + 1):
            src = api_in.Shapes(i)
            left = float(src.Left)
            top = float(src.Top)
            width = float(src.Width)
            height = float(src.Height)
            placement = None
            try:
                placement = int(src.Placement)
            except Exception:
                pass
            src.Copy()
            api_out.Paste()
            dst = api_out.Shapes(int(api_out.Shapes.Count))
            try:
                dst.LockAspectRatio = False
            except Exception:
                pass
            if placement is not None:
                try:
                    dst.Placement = placement
                except Exception:
                    pass
            dst.Left = left
            dst.Top = top
            dst.Width = width
            dst.Height = height
        for j in range(1, n_ole + 1):
            try:
                src_ole = api_in.OLEObjects(j)
                left_o = float(src_ole.Left)
                top_o = float(src_ole.Top)
                width_o = float(src_ole.Width)
                height_o = float(src_ole.Height)
                src_ole.Copy()
                api_out.Paste()
                dst_ole = api_out.OLEObjects(int(api_out.OLEObjects.Count))
                dst_ole.Left = left_o
                dst_ole.Top = top_o
                dst_ole.Width = width_o
                dst_ole.Height = height_o
            except Exception as e_ole:
                logging.warning(
                    "列設定シート図形コピー: OLEObject（ActiveX 等）%s の複製に失敗しました: %s",
                    j,
                    e_ole,
                )
        wb_out.save()
        logging.info(
            "列設定シート図形コピー: 入力から Shapes %s 個・OLEObjects %s 個を結果ブックへ複製しました。",
            n_shapes,
            n_ole,
        )
    except Exception as e:
        logging.warning(
            "列設定シート図形コピー: 失敗しました（%s）。Excel 占有・COM エラー等の可能性はありした。",
            e,
        )
    finally:
        for _wb in (wb_in, wb_out):
            if _wb is not None:
                try:
                    _wb.close()
                except Exception:
                    pass
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass


def _com_excel_bgr_rgb(r: int, g: int, b: int) -> int:
    """Office COM の Color.RGB（BGR リトルエンディアン）。"""
    return int(r) & 255 | ((int(g) & 255) << 8) | ((int(b) & 255) << 16)


def _hex_rrggbb_to_rgb_triple(hx: str) -> tuple[int, int, int]:
    """6 桁 RRGGBB（# 可）を (R,G,B) に。不正時は中間グレー。"""
    s = (hx or "").strip().lstrip("#").upper()
    if len(s) != 6 or any(c not in "0123456789ABCDEF" for c in s):
        return (180, 180, 180)
    return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)


def _gantt_label_luminance_01(r: int, g: int, b: int) -> float:
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255.0


def _gantt_com_colors_from_fill_hex(fill_hex: str) -> tuple[int, int, int]:
    """
    ガント帯色（RRGGBB）から COM 用 (塗り BGR, 枠 BGR, 文字 BGR)。
    淡色帯は黒寄り文字、やや濃い帯は白文字（モックのコントラストに近づける）。
    """
    r, g, b = _hex_rrggbb_to_rgb_triple(fill_hex)
    fill_bgr = _com_excel_bgr_rgb(r, g, b)
    lr = max(0, min(255, int(r * 0.52)))
    lg = max(0, min(255, int(g * 0.52)))
    lb = max(0, min(255, int(b * 0.52)))
    line_bgr = _com_excel_bgr_rgb(lr, lg, lb)
    lum = _gantt_label_luminance_01(r, g, b)
    if lum > 0.74:
        text_bgr = _com_excel_bgr_rgb(26, 26, 26)
    else:
        text_bgr = _com_excel_bgr_rgb(255, 255, 255)
    return fill_bgr, line_bgr, text_bgr


def _gantt_openpyxl_font_color_for_fill_hex(fill_hex: str) -> str:
    """openpyxl Font.color 用 6 桁（RGB 文字列）。"""
    r, g, b = _hex_rrggbb_to_rgb_triple(fill_hex)
    lum = _gantt_label_luminance_01(r, g, b)
    if lum > 0.74:
        return "1A1A1A"
    return "FFFFFF"


def _gantt_member_pill_bgrs_for_task_fill_hex(fill_hex: str) -> tuple[int, int, int]:
    """
    GANTT_COLOR_MODE=full 時の担当者チップ用 (塗り BGR, 線 BGR, 文字 BGR)。
    依頼NO帯色を薄く混ぜた地色とし、文字は輝度から黒／白を選択。
    """
    r, g, b = _hex_rrggbb_to_rgb_triple(fill_hex)
    rf = max(0, min(255, int(0.62 * 255.0 + 0.38 * float(r))))
    gf = max(0, min(255, int(0.62 * 255.0 + 0.38 * float(g))))
    bf = max(0, min(255, int(0.62 * 255.0 + 0.38 * float(b))))
    mem_fill_bgr = _com_excel_bgr_rgb(rf, gf, bf)
    lr = max(0, min(255, int(rf * 0.52)))
    lg = max(0, min(255, int(gf * 0.52)))
    lb = max(0, min(255, int(bf * 0.52)))
    mem_line_bgr = _com_excel_bgr_rgb(lr, lg, lb)
    lum = _gantt_label_luminance_01(rf, gf, bf)
    if lum > 0.74:
        mem_txt_bgr = _com_excel_bgr_rgb(26, 26, 26)
    else:
        mem_txt_bgr = _com_excel_bgr_rgb(255, 255, 255)
    return mem_fill_bgr, mem_line_bgr, mem_txt_bgr


def _gantt_fallback_timeline_labels_openpyxl(
    result_path: str, specs: list, sheet_name: str | None = None
) -> None:
    """xlwings 失敗時: タイムライン先頭列にセル文字でラベルを書き戻す。"""
    from openpyxl import load_workbook

    if _workbook_should_skip_openpyxl_io(result_path):
        return
    shn = sheet_name or RESULT_SHEET_GANTT_NAME
    wb = load_workbook(result_path)
    try:
        ws = wb[shn]
    except KeyError:
        wb.close()
        return
    try:
        for sp in specs:
            row = int(sp["row"])
            col_s = int(sp["col_s"])
            col_e = int(sp["col_e"])
            text = str(sp.get("text") or "").strip()
            if not text:
                continue
            c = ws.cell(row=row, column=col_s)
            mems = [
                str(x).strip()
                for x in (sp.get("member_labels") or [])
                if str(x).strip()
            ]
            if mems:
                head = "・".join(mems[:5])
                rest = len(mems) - 5
                line2 = head + (f" ほか{rest}名" if rest > 0 else "")
                c.value = text + "\n" + line2
            else:
                c.value = text
            _fh = str(sp.get("fill_hex") or "E8E8E8")
            c.font = _result_font(
                size=10,
                bold=True,
                color=_gantt_openpyxl_font_color_for_fill_hex(_fh),
                italic=bool(sp.get("italic")),
            )
            c.alignment = _gantt_timeline_label_alignment(single_slot=(col_s == col_e))
            if mems:
                try:
                    c.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=1,
                    )
                except Exception:
                    pass
        wb.save(result_path)
    finally:
        wb.close()


def _gantt_day_image_chroma_rgb() -> tuple[int, int, int]:
    """日別画像の敷き色／透明色に使う RGB（GANTT_DAY_IMAGE_CHROMA_HEX、既定 マゼンタ）。"""
    hx = (os.environ.get("GANTT_DAY_IMAGE_CHROMA_HEX", "FF00FF") or "FF00FF").strip()
    return _hex_rrggbb_to_rgb_triple(hx)


def _gantt_union_bbox_names_xlw(
    api_ws, names: list[str]
) -> tuple[float, float, float, float] | None:
    """シェイプ名の列の外接矩形 (Left, Top, Width, Height)。取得できなければ None。"""
    min_l = min_t = None
    max_r = max_b = None
    for nm in names:
        try:
            sh = api_ws.Shapes(nm)
            l = float(sh.Left)
            t = float(sh.Top)
            r = l + float(sh.Width)
            b = t + float(sh.Height)
        except Exception:
            continue
        if min_l is None:
            min_l, min_t, max_r, max_b = l, t, r, b
        else:
            min_l = min(min_l, l)
            min_t = min(min_t, t)
            max_r = max(max_r, r)
            max_b = max(max_b, b)
    if min_l is None:
        return None
    pad = 1.0
    w = max(max_r - min_l + 2.0 * pad, 2.0)
    h = max(max_b - min_t + 2.0 * pad, 2.0)
    return (min_l - pad, min_t - pad, w, h)


def _gantt_flatten_apply_picture_chroma_transparency_xlw(pic, fill_bgr: int) -> None:
    """Picture の単色透明化。色→有効の順で一部の Excel で安定する。"""
    try:
        pf = pic.PictureFormat
        pf.TransparencyColor = int(fill_bgr)
        pf.TransparentBackground = -1  # msoTrue
    except Exception:
        pass


def _gantt_flatten_copy_picture_format_xlw() -> int:
    """
    CopyPicture の Format（XlCopyPictureFormat）。

    Excel でグループ化→クリップボード経由で画像貼り付けしたとき、**外接矩形のうち
    シェイプの実画素以外が透明**になりやすいのは **xlPicture（-4147, EMF 系）** 側。
    **xlBitmap（2）** は余白が **不透明（白など）** になりやすく、手動の見え方とずれる。

    ``GANTT_DAY_IMAGE_COPY_PICTURE_FORMAT`` で上書き可能（既定は xlPicture）。
    """
    v = (os.environ.get("GANTT_DAY_IMAGE_COPY_PICTURE_FORMAT", "") or "").strip().lower()
    if v in ("bitmap", "xlbitmap", "2", "bmp"):
        return 2
    if v in ("picture", "xlpicture", "emf", "wmf", "meta", "-4147"):
        return -4147
    return -4147


def _gantt_clipboard_picture_from_shape_names_xlw(
    api_ws,
    group_names: tuple[str, ...],
    *,
    copy_format: int,
    xl_screen: int = 1,
) -> tuple[object, float, float, float, float]:
    """
    ラベル等の図形を「1 枚の画像」に置き換える Excel 標準フロー（クリップボード経由）。

    1. 名前が複数なら ``Shapes.Range(...).Group()`` でグループ化（単一ならそのまま）
    2. ``CopyPicture`` … クリップボードに画像として載せる（Format は ``_gantt_flatten_copy_picture_format_xlw``。既定 EMF 系で手動に近い「外側透明」）
    3. ``Worksheet.Paste`` … シート上に画像シェイプとして貼り付け
    4. 元グループ／元シェイプを削除（貼り付け後の画像のみ残す）

    戻り値: (貼り付けた Shape COM オブジェクト, Left, Top, Width, Height)
    """
    if not group_names:
        raise ValueError("group_names が空です")
    if len(group_names) == 1:
        shp0 = api_ws.Shapes(group_names[0])
        left0 = float(shp0.Left)
        top0 = float(shp0.Top)
        w0 = float(shp0.Width)
        h0 = float(shp0.Height)
        shp0.CopyPicture(Appearance=xl_screen, Format=int(copy_format))
        api_ws.Paste()
        pic = api_ws.Shapes(int(api_ws.Shapes.Count))
        try:
            shp0.Delete()
        except Exception:
            pass
        return pic, left0, top0, w0, h0
    sr = api_ws.Shapes.Range(group_names)
    grp = sr.Group()
    left0 = float(grp.Left)
    top0 = float(grp.Top)
    w0 = float(grp.Width)
    h0 = float(grp.Height)
    grp.CopyPicture(Appearance=xl_screen, Format=int(copy_format))
    api_ws.Paste()
    pic = api_ws.Shapes(int(api_ws.Shapes.Count))
    try:
        grp.Delete()
    except Exception:
        pass
    return pic, left0, top0, w0, h0


def _gantt_flatten_day_label_shapes_to_pictures_xlw(
    api_ws, day_blocks: list, names_by_day: dict
) -> int:
    """
    各日キーに属する角丸ラベルシェイプを、上記
    ``_gantt_clipboard_picture_from_shape_names_xlw``（グループ化→CopyPicture→Paste）
    で 1 枚の Picture に置換する。
    GANTT_DAY_IMAGE_CHROMA_TRANSPARENT が有効なときのみ、敷き矩形を同グループに含め、
    貼り付け後に ``PictureFormat`` で敷き色を透明化する（オプション。核フローはクリップボード画像化）。
    names_by_day[day_key] に蓄積された Name を消費する（成功時は空リストに戻す）。
    """
    if not day_blocks:
        return 0
    _xl_screen = 1  # xlScreen
    _xl_move_and_size = 1
    _mso_rectangle = 1
    n_out = 0
    for blk in day_blocks:
        dk = str(blk.get("day_key") or "").strip()
        raw_names = list(names_by_day.get(dk, []))
        if not raw_names:
            continue
        seen: set[str] = set()
        names: list[str] = []
        for nm in raw_names:
            if nm and nm not in seen:
                seen.add(nm)
                names.append(nm)
        if not names:
            continue
        backdrop_nm: str | None = None
        try:
            r_ch, g_ch, b_ch = _gantt_day_image_chroma_rgb()
            fill_bgr = _com_excel_bgr_rgb(r_ch, g_ch, b_ch)
            group_names: tuple[str, ...] = tuple(names)

            if GANTT_DAY_IMAGE_CHROMA_TRANSPARENT:
                ubox = _gantt_union_bbox_names_xlw(api_ws, names)
                if ubox:
                    L, T, Wb, Hb = ubox
                    bd_nm_try = f"GanttChromaBg_{random.randint(100000, 999999)}"
                    bd = api_ws.Shapes.AddShape(_mso_rectangle, L, T, Wb, Hb)
                    try:
                        bd.Name = bd_nm_try
                    except Exception:
                        bd_nm_try = str(bd.Name)
                    bd.Fill.Visible = True
                    bd.Fill.Solid()
                    bd.Fill.ForeColor.RGB = fill_bgr
                    try:
                        bd.Line.Visible = False
                    except Exception:
                        pass
                    try:
                        bd.Placement = _xl_move_and_size
                    except Exception:
                        pass
                    backdrop_nm = bd_nm_try
                    group_names = (bd_nm_try,) + tuple(names)

            chroma_backdrop = backdrop_nm is not None
            _cpy_fmt = _gantt_flatten_copy_picture_format_xlw()

            pic, left0, top0, w0, h0 = _gantt_clipboard_picture_from_shape_names_xlw(
                api_ws,
                group_names,
                copy_format=_cpy_fmt,
                xl_screen=_xl_screen,
            )
            pic.Left = left0
            pic.Top = top0
            pic.Width = w0
            pic.Height = h0
            try:
                pic.Placement = _xl_move_and_size
            except Exception:
                pass
            if chroma_backdrop:
                _gantt_flatten_apply_picture_chroma_transparency_xlw(pic, fill_bgr)
            safe = "".join(
                ch if ch.isalnum() or ch in "._-" else "_" for ch in dk
            )[:200]
            try:
                pic.Name = f"GanttDayImg_{safe}"
            except Exception:
                pass
            names_by_day[dk] = []
            n_out += 1
        except Exception as e_fl:
            if backdrop_nm:
                try:
                    api_ws.Shapes(backdrop_nm).Delete()
                except Exception:
                    pass
            logging.warning(
                "結果_設備ガント: 日別シェイプ画像化をスキップしました（日キー=%s、名称数=%s: %s）",
                dk,
                len(names),
                e_fl,
            )
    return n_out


def _gantt_add_timeline_rounded_rect_labels_xlwings(
    result_path: str,
    specs: list,
    day_blocks: list | None = None,
    *,
    sheet_name: str | None = None,
) -> bool:
    """
    結果_設備ガントのタイムライン上に、角丸四角（msoShapeRoundedRectangle）でラベルを重ねる。
    依頼NOは中央のメインシェイプ（高さは行の約 1/5。結合幅が 1 スロットでもタイムライン 1 列幅の 2 倍を下限とし文字潰れを抑える）。
    担当者姓はその直上に小さな角丸チップ 1 つ（結合文字が潰れない下限幅までシェイプ幅を確保、
    テキストはシェイプ内右寄せ。Z オーダーはメンバーを背面・依頼NO を前面に寄せる）。
    day_blocks が与えられ、GANTT_TIMELINE_LABELS_DAY_FLATTEN が有効なとき、日ごとに画像へ集約する。
    成功時 True。xlwings / Excel 不可時は False。
    """
    rp = (result_path or "").strip()
    if not rp or not os.path.isfile(rp) or not specs:
        return False
    try:
        import xlwings as xw
    except ImportError:
        return False
    app = None
    wb = None
    try:
        n_specs = len(specs)
        shn = sheet_name or RESULT_SHEET_GANTT_NAME
        logging.info(
            "%s: xlwings で角丸シェイプを追加します（候補 %s 件）。"
            " 件数が多いと数分かかり、完了までログが増えない時間が続くことがあります。",
            shn,
            n_specs,
        )
        app = xw.App(visible=False)
        app.display_alerts = False
        try:
            app.screen_updating = False
        except Exception:
            try:
                app.api.ScreenUpdating = False
            except Exception:
                pass
        wb = app.books.open(os.path.abspath(rp), update_links=False)
        try:
            sht = wb.sheets[shn]
        except Exception:
            return False
        api_ws = sht.api
        # msoShapeRoundedRectangle = 5
        _mso_round_rect = 5
        _mso_bring_to_front = 0
        _mso_send_to_back = 1
        _xl_move_and_size = 1
        _xl_h_align_center = -4131
        _xl_h_align_right = -4152
        # 件数が多いときの進捗ログ間隔（小さすぎると I/O 負荷、大きすぎると停止に見える）
        _progress_every = 10
        n_added = 0
        names_by_day: dict[str, list[str]] = defaultdict(list)

        def _record_day_shape(shp_obj, day_k: str):
            if not day_k or shp_obj is None:
                return
            try:
                names_by_day[day_k].append(str(shp_obj.Name))
            except Exception:
                pass

        # 同一データ行ごとにシェイプを 3 段（行高の各 1/3 の帯）でローテーション配置（4 件目は上段に戻る）。
        # 依頼NO メインは行高の 1/5 を目標にし、帯の上下にインセットを取って罫線付近への食み出しを抑える。
        # メンバー名は上下分割せず、依頼NO の直上に 1 シェイプで置く（全角空白区切り。人数分の AddShape はしない）。
        # メンバーは ZOrder SendToBack、依頼NO は BringToFront（幅はみ出し時も依頼NOが手前に来る）。
        # メンバー帯の縦幅は依頼NO メインと同じ。印刷で上行にはみ出さないよう、行矩形内に収める。
        _row_shape_seq: dict[int, int] = {}

        def _gantt_xlw_timeline_main_font_pt(xw: float, cap: str) -> float:
            """狭い結合セルではフォントを下げ、glyph のシェイプ外はみ出しを抑える。"""
            nch = max(1, len(str(cap or "").strip()))
            raw = float(xw) / max(nch * 0.62, 4.0)
            return max(5.25, min(9.0, raw))

        def _gantt_xlw_member_pill_font_pt(pwidth: float, nm: str) -> float:
            nch = max(1, len(str(nm or "").strip()))
            raw = float(pwidth) / max(nch * 1.05, 3.2)
            return max(5.5, min(6.5, raw))

        def _gantt_xlw_member_combined_min_width_pt(combined: str) -> float:
            """メンバー結合文字列が最低フォントでも潰れないよう必要幅（pt）の粗い下限。"""
            nch = max(1, len(str(combined or "").strip()))
            f_min = 5.75
            return f_min * max(float(nch) * 1.1, 4.0) + 7.0

        def _gantt_xlw_add_round_rect(
            x_left,
            x_top,
            x_w,
            x_h,
            caption,
            *,
            fill_rgb,
            line_rgb,
            text_rgb,
            font_pt=9.0,
            bold=True,
            italic=False,
            line_wt=0.75,
            adj_round=0.2,
            shadow=False,
            shape_name=None,
            tf_margin_tb=None,
            tf_margin_lr=None,
            z_bring_to_front=True,
            text_h_align=None,
        ):
            cap = str(caption or "").strip()
            if x_w <= 0 or x_h <= 0 or not cap:
                return None
            shp_local = api_ws.Shapes.AddShape(
                _mso_round_rect, float(x_left), float(x_top), float(x_w), float(x_h)
            )
            if shape_name:
                try:
                    shp_local.Name = shape_name
                except Exception:
                    pass
            try:
                shp_local.Placement = _xl_move_and_size
            except Exception:
                pass
            try:
                if z_bring_to_front:
                    shp_local.ZOrder(_mso_bring_to_front)
                else:
                    shp_local.ZOrder(_mso_send_to_back)
            except Exception:
                pass
            try:
                shp_local.Fill.Visible = True
                shp_local.Fill.Solid()
                shp_local.Fill.ForeColor.RGB = fill_rgb
                shp_local.Line.Visible = True
                shp_local.Line.ForeColor.RGB = line_rgb
                shp_local.Line.Weight = line_wt
            except Exception:
                pass
            if adj_round is not None:
                try:
                    shp_local.Adjustments[1] = adj_round
                except Exception:
                    pass
            if shadow:
                try:
                    sd0 = shp_local.Shadow
                    sd0.Visible = -1  # msoTrue
                    sd0.OffsetX = 3
                    sd0.OffsetY = 3
                    sd0.Transparency = 0.55
                    try:
                        sd0.Blur = 4
                    except Exception:
                        pass
                    try:
                        sd0.ForeColor.RGB = _com_excel_bgr_rgb(40, 40, 50)
                    except Exception:
                        pass
                except Exception:
                    pass
            try:
                tf0 = shp_local.TextFrame
                try:
                    if tf_margin_lr is not None:
                        mrg_lr = float(tf_margin_lr)
                    else:
                        mrg_lr = 1.0 if font_pt <= 7.0 else 2.0
                    tf0.MarginLeft = mrg_lr
                    tf0.MarginRight = mrg_lr
                    m_tb = 0.5 if tf_margin_tb is None else float(tf_margin_tb)
                    tf0.MarginTop = m_tb
                    tf0.MarginBottom = m_tb
                except Exception:
                    pass
                try:
                    tf0.VerticalAlignment = -4108  # xlVAlignCenter
                    _hal = (
                        int(text_h_align)
                        if text_h_align is not None
                        else int(_xl_h_align_center)
                    )
                    tf0.HorizontalAlignment = _hal
                except Exception:
                    pass
                # WordWrap=True は環境によって TextFrame の再レイアウトが極端に重く、
                # 先頭シェイプ追加付近で「進まない」ように見えることがあるため付けない。
                tf0.Characters().Text = cap
                nch = len(cap)
                fnt = tf0.Characters(1, nch).Font if nch > 0 else tf0.Characters().Font
                fnt.Size = font_pt
                fnt.Bold = bold
                if italic:
                    fnt.Italic = True
                try:
                    fnt.Color = text_rgb
                except Exception:
                    pass
            except Exception:
                try:
                    shp_local.TextFrame.Characters().Text = cap
                except Exception:
                    pass
            return shp_local

        for idx, sp in enumerate(specs, start=1):
            if idx == 1 or idx % _progress_every == 0 or idx == n_specs:
                logging.info(
                    "結果_設備ガント: シェイプ走査 %s/%s（確定追加 %s 件）…",
                    idx,
                    n_specs,
                    n_added,
                )
            text = str(sp.get("text") or "").strip()
            if not text:
                continue
            dk = str(sp.get("day_key") or "").strip()
            row = int(sp["row"])
            col_s = int(sp["col_s"])
            col_e = int(sp["col_e"])
            rng = sht.range((row, col_s), (row, col_e))
            left = float(rng.left)
            top = float(rng.top)
            w = float(rng.width)
            h = float(rng.height)
            if w <= 0 or h <= 0:
                continue
            _fh = str(sp.get("fill_hex") or "E8E8E8")
            fill_bgr, line_bgr, text_bgr = _gantt_com_colors_from_fill_hex(_fh)
            # 依頼NO メインシェイプ: 1 スロット幅だけのとき文字が潰れるため、10 分 1 列の幅の 2 倍を確保する
            # （隣スロット上にはみ出すが、結合が複数列なら結合幅のまま）。
            try:
                slot_w = float(sht.range((row, col_s), (row, col_s)).width)
            except Exception:
                slot_w = 0.0
            if slot_w <= 0.0:
                _ns0 = max(1, int(col_e) - int(col_s) + 1)
                slot_w = float(w) / float(_ns0)
            label_w = max(float(w), 2.0 * float(slot_w))
            # 縦位置は行を 3 等分した帯のいずれか（同一行で追加順に 0→1→2→0…）。依頼NO の高さは行高の 1/5。
            _band = float(h) / 3.0
            _h_req_no = max(9.0, float(h) / 5.0)
            _n_on_row = int(_row_shape_seq.get(row, 0))
            _slot = _n_on_row % 3
            _row_shape_seq[row] = _n_on_row + 1
            band_top = top + _slot * _band
            band_bot = band_top + _band
            _band_inset = 0.75
            mems_all = [
                str(x).strip() for x in (sp.get("member_labels") or []) if str(x).strip()
            ]
            mems_all = mems_all[:8]
            if _gantt_color_mode_full():
                mem_fill, mem_line, mem_txt = _gantt_member_pill_bgrs_for_task_fill_hex(
                    _fh
                )
            else:
                mem_fill = _com_excel_bgr_rgb(252, 252, 254)
                mem_line = _com_excel_bgr_rgb(175, 180, 188)
                mem_txt = _com_excel_bgr_rgb(38, 40, 46)
            if mems_all:
                # メンバー縦幅＝依頼NO と同じ（行高の 1/5 目標）。行全体 [top, top+h] に収まるよう
                # 積み上げ位置を平行移動し、収まらないときは隙間・ピル高を漸減する。
                _gap_mm = 1.35
                h_main = max(9.0, float(_h_req_no))
                h_mem_use = h_main
                _rin_row = 1.0
                _rout_row = 1.0
                row_top_b = float(top)
                row_bot_b = float(top) + float(h)
                _gap_eff = float(_gap_mm)
                _hmem_eff = float(h_mem_use)
                _hmain_eff = float(h_main)
                y_main = band_bot - _band_inset - _hmain_eff
                y_mem = y_main - _gap_eff - _hmem_eff
                for _squeeze in range(28):
                    st = float(y_mem)
                    sb = float(y_main) + float(_hmain_eff)
                    lo = (row_top_b + _rin_row) - st
                    hi = (row_bot_b - _rout_row) - sb
                    if lo <= hi:
                        if lo > 0.0:
                            delta = lo
                        elif hi < 0.0:
                            delta = hi
                        else:
                            delta = 0.0
                        y_mem += delta
                        y_main += delta
                        break
                    if _gap_eff > 0.35:
                        _gap_eff = max(0.35, _gap_eff - 0.35)
                    elif _hmem_eff > 6.0:
                        _hmem_eff = max(6.0, _hmem_eff - 0.5)
                    elif _hmain_eff > 8.0:
                        _hmain_eff = max(8.0, _hmain_eff - 0.5)
                    else:
                        y_main = band_bot - _band_inset - _hmain_eff
                        y_mem = y_main - _gap_eff - _hmem_eff
                        lo2 = (row_top_b + _rin_row) - float(y_mem)
                        if lo2 > 0.0:
                            y_mem += lo2
                            y_main += lo2
                        break
                    y_main = band_bot - _band_inset - _hmain_eff
                    y_mem = y_main - _gap_eff - _hmem_eff
                h_main = float(_hmain_eff)
                h_mem_use = float(_hmem_eff)
                gx = 1.0

                def _emit_member_pills(
                    names: list[str], y0: float, pill_h: float, day_k: str
                ) -> None:
                    nonlocal n_added
                    if not names or pill_h <= 1.0:
                        return
                    parts: list[str] = []
                    est_w = 0.0
                    for nm in names:
                        nm2 = nm if len(nm) <= 6 else (nm[:5] + "…")
                        parts.append(nm2)
                        est_w += max(9.0, 5.2 * float(len(nm2)))
                    if len(parts) > 1:
                        est_w += float(len(parts) - 1) * gx
                    combined = "\u3000".join(parts)
                    if not combined.strip():
                        return
                    # 結合幅 w に縛ると文字が潰れるため、ピル分割時と同様の推定に加え、
                    # 最低フォント相当の下限幅を満たすまでシェイプ幅を広げる（隣セル上にはみ出し得る）。
                    _min_member_chip_w = 34.0
                    text_min_w = _gantt_xlw_member_combined_min_width_pt(combined)
                    want_w = max(
                        _min_member_chip_w, float(est_w), float(text_min_w)
                    )
                    use_w = max(max(float(w), _min_member_chip_w), want_w)
                    _fp_mem = _gantt_xlw_member_pill_font_pt(use_w, combined)
                    s_mem = _gantt_xlw_add_round_rect(
                        left,
                        y0,
                        use_w,
                        pill_h,
                        combined,
                        fill_rgb=mem_fill,
                        line_rgb=mem_line,
                        text_rgb=mem_txt,
                        font_pt=float(_fp_mem),
                        bold=True,
                        italic=False,
                        line_wt=0.55,
                        adj_round=0.42,
                        shadow=False,
                        shape_name=f"GanttMem_R{row}_C{col_s}_{_n_on_row}_{int(y0)}",
                        tf_margin_tb=0.0,
                        tf_margin_lr=0.75,
                        z_bring_to_front=False,
                        text_h_align=_xl_h_align_right,
                    )
                    if s_mem is not None:
                        n_added += 1
                        _record_day_shape(s_mem, day_k)

                _emit_member_pills(mems_all, y_mem, h_mem_use, dk)
                _main_fp = _gantt_xlw_timeline_main_font_pt(label_w, text)
                shp_main = _gantt_xlw_add_round_rect(
                    left,
                    y_main,
                    label_w,
                    h_main,
                    text,
                    fill_rgb=fill_bgr,
                    line_rgb=line_bgr,
                    text_rgb=text_bgr,
                    font_pt=float(_main_fp),
                    bold=True,
                    italic=bool(sp.get("italic")),
                    line_wt=0.75,
                    adj_round=0.2,
                    shadow=False,
                    shape_name=f"GanttLbl_R{row}_C{col_s}_{_n_on_row}",
                )
                if shp_main is not None:
                    n_added += 1
                    _record_day_shape(shp_main, dk)
                    try:
                        shp_main.TextFrame.HorizontalAlignment = -4131  # xlHAlignLeft
                    except Exception:
                        pass
            else:
                label_h = _h_req_no
                y_lbl = band_top + _band_inset + max(
                    0.0, (_band - 2.0 * _band_inset - label_h) / 2.0
                )
                _solo_fp = _gantt_xlw_timeline_main_font_pt(label_w, text)
                shp = _gantt_xlw_add_round_rect(
                    left,
                    y_lbl,
                    label_w,
                    label_h,
                    text,
                    fill_rgb=fill_bgr,
                    line_rgb=line_bgr,
                    text_rgb=text_bgr,
                    font_pt=float(_solo_fp),
                    bold=True,
                    italic=bool(sp.get("italic")),
                    line_wt=0.75,
                    adj_round=0.2,
                    shadow=False,
                    shape_name=f"GanttLbl_R{row}_C{col_s}_{_n_on_row}",
                )
                if shp is not None:
                    n_added += 1
                    _record_day_shape(shp, dk)
                    try:
                        shp.TextFrame.HorizontalAlignment = -4131
                    except Exception:
                        pass
        n_flat = 0
        if (
            GANTT_TIMELINE_LABELS_DAY_FLATTEN
            and day_blocks
            and GANTT_TIMELINE_SHAPE_LABELS
        ):
            try:
                n_flat = _gantt_flatten_day_label_shapes_to_pictures_xlw(
                    api_ws, day_blocks, names_by_day
                )
            except Exception as e_flat:
                logging.warning(
                    "%s: 日別画像化に失敗しました（個別シェイプのまま保存します）: %s",
                    shn,
                    e_flat,
                )
        logging.info(
            "%s: 角丸シェイプ %s 件を反映%sして保存します（xlwings）…",
            shn,
            n_added,
            f"し、日別に画像 {n_flat} 枚へ集約" if n_flat else "",
        )
        wb.save()
        return True
    except Exception as e:
        _shn_fb = sheet_name or RESULT_SHEET_GANTT_NAME
        logging.warning(
            "%s: 角丸シェイプラベルの追加に失敗しました（%s）。セル表記へフォールバックします。",
            _shn_fb,
            e,
        )
        return False
    finally:
        if app is not None:
            try:
                app.screen_updating = True
            except Exception:
                try:
                    app.api.ScreenUpdating = True
                except Exception:
                    pass
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass


def _stage2_try_add_gantt_timeline_shape_labels(
    result_path: str,
    specs: list | None,
    day_blocks: list | None = None,
    *,
    sheet_name: str | None = None,
) -> None:
    """
    openpyxl 保存後、GANTT_TIMELINE_SHAPE_LABELS が有効で specs があれば xlwings で角丸ラベルを描画。
    day_blocks があれば既定で日単位に画像化してシェイプ数を抑える。
    失敗時は openpyxl でセルにフォールバック。
    """
    if not GANTT_TIMELINE_SHAPE_LABELS or not specs:
        return
    rp = (result_path or "").strip()
    if not rp or not os.path.isfile(rp):
        return
    shn = sheet_name or RESULT_SHEET_GANTT_NAME
    if _gantt_add_timeline_rounded_rect_labels_xlwings(
        rp, specs, day_blocks, sheet_name=shn
    ):
        logging.info(
            "%s: タイムラインラベルを角丸シェイプ %s 件で追加しました。",
            shn,
            len(specs),
        )
        return
    try:
        _gantt_fallback_timeline_labels_openpyxl(rp, specs, sheet_name=shn)
        logging.info(
            "%s: タイムラインラベルをセル表記にフォールバックしました（%s 件）。",
            shn,
            len(specs),
        )
    except Exception as e:
        logging.warning(
            "%s: セルへのラベルフォールバックも失敗しました（%s）。",
            shn,
            e,
        )


def _coerce_actual_sheet_datetime(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, time(0, 0))
    try:
        ts = pd.to_datetime(val, errors="coerce")
        if pd.isna(ts) or ts is pd.NaT:
            return None
        if isinstance(ts, pd.Timestamp):
            return ts.to_pydatetime()
        return ts if isinstance(ts, datetime) else None
    except Exception:
        return None


def _actual_row_time_bounds(row):
    """加工実績DATA／加工実績明細DATA の1行から (開始, 終了) を得る。解けなければ (None, None)。"""
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_START_DT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_END_DT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_MACHINING_START_DT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_MACHINING_END_DT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_START_ALT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_END_ALT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt

    d_date = parse_optional_date(row.get(ACT_COL_DAY))
    if not d_date:
        cd = _coerce_actual_sheet_datetime(row.get(ACT_COL_DAY))
        if isinstance(cd, datetime):
            d_date = cd.date()
        elif isinstance(cd, date):
            d_date = cd
    if not d_date:
        return None, None

    ts_s = row.get(ACT_COL_TIME_START)
    ts_e = row.get(ACT_COL_TIME_END)
    if ts_s is None or pd.isna(ts_s) or ts_e is None or pd.isna(ts_e):
        return None, None

    if isinstance(ts_s, time):
        t0 = ts_s
    elif isinstance(ts_s, datetime):
        t0 = ts_s.time()
    else:
        t0 = parse_time_str(ts_s, None)

    if isinstance(ts_e, time):
        t1 = ts_e
    elif isinstance(ts_e, datetime):
        t1 = ts_e.time()
    else:
        t1 = parse_time_str(ts_e, None)

    if t0 is None or t1 is None or t0 >= t1:
        return None, None
    return datetime.combine(d_date, t0), datetime.combine(d_date, t1)


def load_machining_actuals_df():
    """
    マクロブックの「加工実績DATA」を読む（無ければ空 DataFrame）。
    Power Query 等で用意したシートを想定。
    """
    if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
        return pd.DataFrame()
    try:
        df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=ACTUALS_SHEET_NAME)
    except ValueError:
        logging.info(
            f"シート「{ACTUALS_SHEET_NAME}」は無いため、ガントの実績行は出力しません。"
        )
        return pd.DataFrame()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, ACTUAL_HEADER_CANONICAL)
    logging.info(
        f"加工実績: '{TASKS_INPUT_WORKBOOK}' の '{ACTUALS_SHEET_NAME}' を {len(df)} 行読み込み。"
    )
    return df


def _calendar_dates_spanned_by_actual_bounds_df(df) -> set[date]:
    """
    実績明細等の各行の (開始, 終了) が跨ぐ暦日を収集する。
    計画の sorted_dates に含まれない過去日の実績もガントに載せるために使う。
    """
    out: set[date] = set()
    if df is None or len(df) == 0:
        return out
    for _, row in df.iterrows():
        s_dt, e_dt = _actual_row_time_bounds(row)
        if not s_dt or not e_dt or s_dt >= e_dt:
            continue
        d0 = s_dt.date()
        d1 = e_dt.date()
        cur = d0
        while cur <= d1:
            out.add(cur)
            cur += timedelta(days=1)
    return out


def _sorted_dates_union_actual_bounds_df(sorted_dates: list, df) -> list:
    """計画表示日と実績行の暦日の和集合（昇順）。"""
    u = set(sorted_dates)
    u |= _calendar_dates_spanned_by_actual_bounds_df(df)
    return sorted(u)


def _sorted_dates_filter_inclusive_range(
    sorted_dates: list,
    d_from: date | None,
    d_to: date | None,
) -> list:
    """
    暦日リストを両端込みで絞る。d_from / d_to がともに None のときはコピーを返す。
    両端指定で from > to のときは from/to を入れ替える。
    """
    if d_from is None and d_to is None:
        return list(sorted_dates)
    a = d_from
    b = d_to
    if a is not None and b is not None and a > b:
        a, b = b, a
    out: list = []
    for d in sorted_dates:
        if a is not None and d < a:
            continue
        if b is not None and d > b:
            continue
        out.append(d)
    return out


def load_machining_actual_detail_df():
    """
    マクロブックの「加工実績明細DATA」を読む（無ければ空 DataFrame）。
    列は加工実績DATA に準じ、ロール識別は「ロールNO」「ロール番号」「ロール」「巻番」のいずれか可。
    """
    if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
        return pd.DataFrame()
    try:
        df = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=ACTUAL_DETAIL_SHEET_NAME)
    except ValueError:
        logging.info(
            f"シート「{ACTUAL_DETAIL_SHEET_NAME}」は無いため、実績明細ガントは出力しません。"
        )
        return pd.DataFrame()
    df.columns = df.columns.str.strip()
    if ACT_DETAIL_COL_ROLL not in df.columns:
        for alias in ("ロール番号", "ロール", "巻番"):
            if alias in df.columns:
                df = df.rename(columns={alias: ACT_DETAIL_COL_ROLL})
                break
    df = _align_dataframe_headers_to_canonical(df, ACTUAL_DETAIL_HEADER_CANONICAL)
    logging.info(
        f"加工実績明細: '{TASKS_INPUT_WORKBOOK}' の '{ACTUAL_DETAIL_SHEET_NAME}' を {len(df)} 行読み込み。"
    )
    return df


def _actual_row_cumulative_completion_pct_macro(row) -> int | None:
    """
    加工実績明細DATA の「累積完了率」をシート値のまま 0～100 の整数に解釈して返す（実÷予定等は計算しない）。

    対応: 数値・「45」「45%」「45.5%」・Excel 割合セル由来の 0.45（=45%）など。
    列が無い・空・数値化不可のときは None。
    """
    if row is None:
        return None
    v = row.get(ACT_COL_CUMULATIVE_COMPLETION_PCT)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return None
    try:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            f = float(v)
        else:
            s = unicodedata.normalize("NFKC", str(v).strip())
            if not s or s.lower() in ("nan", "none", "-", "—", "―"):
                return None
            s = s.replace("%", "").replace(",", "").strip()
            if not s:
                return None
            f = float(s)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f) or f < 0:
        return None
    # Excel の「割合」表示は 0.45 のように小数で渡ることが多い（= 45%）
    if f <= 1.0 + 1e-9:
        pct = int(round(f * 100.0))
    else:
        pct = int(round(f))
    return max(0, min(100, pct))


def _actual_row_detail_assignee_op_sub(row) -> tuple[str, str]:
    """
    加工実績明細DATA 行からガント用 op / sub を組み立てる。
    「担当者」に続けて「加工担当者名1」～「加工担当者名5」を順に見て非空のみ採用し、
    NFKC 後の文字列で重複を除く。先頭を op、2人目以降を sub（カンマ区切り）。
    """
    names: list[str] = []
    seen_k: set[str] = set()
    for col in (ACT_COL_OPERATOR,) + ACT_COL_MACHINING_ASSIGNEES_ORDERED:
        val = row.get(col)
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        s = str(val).strip()
        if not s:
            continue
        k = unicodedata.normalize("NFKC", s)
        if k in seen_k:
            continue
        seen_k.add(k)
        names.append(s)
    if not names:
        return "", ""
    if len(names) == 1:
        return names[0], ""
    return names[0], ", ".join(names[1:])


def build_actual_timeline_events(
    df,
    equipment_list,
    sorted_dates,
    *,
    log_sheet_name: str = "加工実績DATA",
    roll_detail: bool = False,
):
    """
    実績シートの各行をガント用イベントへ変換。
    計画表示日（sorted_dates）かつ設備マスタに一致する「工程名」の値が対象。
    工程名は NFKC・空白正規化後にマスタ列名へマッピングする。
    時刻は DEFAULT_START_TIME / DEFAULT_END_TIME の枠内にクリップ。
    roll_detail=True のとき ACT_DETAIL_COL_ROLL があれば task_id を「依頼NO/ロール」表記にし帯の分離に使う。
    同じく roll_detail=True のときは「担当者」および「加工担当者名1」～「加工担当者名5」を
    ガントの op/sub（タイムライン氏名チップ・D列要約）へ反映する。
    roll_detail=True のとき「累積完了率」列があればその値を解釈し、
    タイムライン角丸シェイプのラベル（依頼NO の横の %%）に ``pct_macro`` として渡す（計算はしない）。
    """
    if df is None or len(df) == 0:
        return []
    equip_lookup = _equipment_lookup_normalized_to_canonical(equipment_list)
    date_ok = set(sorted_dates)
    events = []
    bad_eq = 0
    bad_time = 0
    no_plan_overlap = 0
    mismatch_norm_samples = []

    for _, row in df.iterrows():
        tid = row.get(ACT_COL_TASK_ID)
        if tid is None or pd.isna(tid):
            continue
        tid_s = str(tid).strip()
        if not tid_s:
            continue
        display_tid = tid_s
        if roll_detail:
            rv = row.get(ACT_DETAIL_COL_ROLL)
            if rv is not None and not (isinstance(rv, float) and pd.isna(rv)):
                rs = str(rv).strip()
                if rs:
                    display_tid = f"{tid_s}/{rs}"
        proc = row.get(ACT_COL_PROCESS)
        if proc is None or pd.isna(proc):
            continue
        proc_key = _normalize_equipment_match_key(proc)
        mach = equip_lookup.get(proc_key)
        if not mach:
            bad_eq += 1
            if len(mismatch_norm_samples) < 12 and proc_key:
                if proc_key not in mismatch_norm_samples:
                    mismatch_norm_samples.append(proc_key)
            continue
        start_dt, end_dt = _actual_row_time_bounds(row)
        if not start_dt or not end_dt or start_dt >= end_dt:
            bad_time += 1
            continue
        if roll_detail:
            op_s, sub_s = _actual_row_detail_assignee_op_sub(row)
        else:
            op_val = row.get(ACT_COL_OPERATOR)
            op_s = ""
            if op_val is not None and not pd.isna(op_val):
                op_s = str(op_val).strip()
            sub_s = ""

        pct_macro = _actual_row_cumulative_completion_pct_macro(row)

        before = len(events)
        for d in sorted_dates:
            if d not in date_ok:
                continue
            day_start = datetime.combine(d, DEFAULT_START_TIME)
            day_end = datetime.combine(d, DEFAULT_END_TIME)
            if end_dt <= day_start or start_dt >= day_end:
                continue
            s_clip = max(start_dt, day_start)
            e_clip = min(end_dt, day_end)
            if s_clip >= e_clip:
                continue
            ev_row = {
                "date": d,
                "task_id": display_tid,
                "machine": mach,
                "op": op_s,
                "sub": sub_s,
                "start_dt": s_clip,
                "end_dt": e_clip,
                "breaks": [],
                "units_done": 0,
                "already_done_units": 0,
                "total_units": 0,
                "eff_time_per_unit": 0.0,
                "unit_m": 0.0,
            }
            if pct_macro is not None:
                ev_row["pct_macro"] = pct_macro
            events.append(ev_row)
        if len(events) == before:
            no_plan_overlap += 1

    if bad_eq:
        logging.warning(
            f"{log_sheet_name}: 工程名はマスタ設備と一致しない行を {bad_eq} 件スキップしました（空白等は正規化済み）。"
        )
        if mismatch_norm_samples:
            logging.info(
                "  厳密一致となった工程名の正規化後サンプル: "
                + " | ".join(mismatch_norm_samples[:12])
            )
    if bad_time:
        logging.info(
            f"{log_sheet_name}: 開始/終了日時を解釈できない行を {bad_time} 件スキップしました。"
        )
    if no_plan_overlap and sorted_dates:
        logging.info(
            f"{log_sheet_name}: 設備・日時は有効だが計画対象日（勤怠日×{DEFAULT_START_TIME}～{DEFAULT_END_TIME}）と重ならない行が {no_plan_overlap} 件ありました。"
        )
    if not events and len(df) > 0:
        logging.info(
            f"{log_sheet_name}: ガント用セグメントは0件です。表示日（sorted_dates）に重ならない実績のみの場合、描画されません。"
        )
    logging.info(f"{log_sheet_name} からガント用セグメント {len(events)} 件を生成しました。")
    return events


TASK_SPECIAL_AI_LAST_RESPONSE_FILE = "ai_task_special_remark_last.txt"
# 勤怠備考キャッシュとキー空間を分離（同一SHA衝窝を避ける）。指紋に基準年を含む日付解釈のズレを防し。
TASK_SPECIAL_CACHE_KEY_PREFIX = "TASK_SPECIAL_v3|"
# メインシート「グローバルコメント」下の自由記述 → Gemini 解釈（配台の最優先オーポーライド）
GLOBAL_PRIORITY_OVERRIDE_CACHE_PREFIX = "GLOBAL_PRIO_v8|"


def _normalize_special_task_id_for_ai(val):
    """
    依頼NOをキャッシュキー・プロンプト行で一貫させる。
    Excel の数値セルは float になりはうなので 12345.0 → \"12345\" に权ごる。
    文字列は NFKC（全角英数字など）で表記ゆれを坸坎（同一実体の再API呼び出しを減られ）。
    """
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except TypeError:
        pass
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if math.isnan(val):
            return None
        if val.is_integer():
            return str(int(val))
        s = str(val).strip()
        if not s or s.lower() in ("nan", "none", "null"):
            return None
        return unicodedata.normalize("NFKC", s).strip() or None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    s = unicodedata.normalize("NFKC", s).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    # 文字列としての "20010.0" 等（Excel・CSV）を整数表記の依頼NOに寄せる
    if re.fullmatch(r"-?\d+\.0+", s):
        try:
            return str(int(float(s)))
        except ValueError:
            pass
    return s or None


def planning_task_id_str_from_scalar(val) -> str:
    """配台・段階1マージ・キュー構築で用いる依頼NOの安定文字列（空なら \"\"）。"""
    return _normalize_special_task_id_for_ai(val) or ""


def planning_task_id_str_from_plan_row(row) -> str:
    """重複見出し列でも先頭スカラーを拾い」依頼NOを planning_task_id_str_from_scalar に渡す。"""
    return planning_task_id_str_from_scalar(_planning_df_cell_scalar(row, TASK_COL_TASK_ID))


def _cell_text_task_special_remark(val):
    """
    特別指定_備考をプロンプト用に取り出す。仕様どより **strip のみ**
    （先頭末尾の空白・Excel の坽空白を除し」文中の改行・スペースは保挝。数値セルは表記を固定）。
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except TypeError:
        pass
    if isinstance(val, bool):
        s = str(val)
    elif isinstance(val, float):
        if math.isnan(val):
            return ""
        # 備考列に数値の値入っている場合の表記ゆれを減られ
        if val.is_integer():
            s = str(int(val))
        else:
            s = str(val)
    elif isinstance(val, int):
        s = str(val)
    else:
        s = str(val)
    s = s.strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _task_special_prompt_lines(tasks_df):
    """プロンプトに載せる行リスト（ソート剝）。正規化は上記ヘルパーに統一。"""
    lines = []
    for _, row in tasks_df.iterrows():
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
        if not tid or not rem:
            continue
        proc = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        macn = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        proc_disp = proc if proc else "（空）"
        macn_disp = macn if macn else "（空）"
        lines.append(
            f"- 依頼NO」{tid}】| 工程名「{proc_disp}」 | 機械名「{macn_disp}」 | 備考本文: {rem}"
        )
    return lines


def _repair_task_special_ai_wrong_top_level_keys(parsed: dict, tasks_df) -> dict:
    """
    備考は哝番・原板コード（例: 20010 で始まる数字列）で始まると」モデルはしの列を JSON トップキーに
    誤用れることはある。依頼NO」…】と一致しない数字のみのキーを」当該備考を挝つ行の依頼NOへ付け替ごる。
    """
    if not isinstance(parsed, dict) or not parsed or tasks_df is None or getattr(tasks_df, "empty", True):
        return parsed
    valid_tids: set[str] = set()
    remark_by_tid: dict[str, list[str]] = {}
    for _, row in tasks_df.iterrows():
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
        if not tid or not rem:
            continue
        valid_tids.add(tid)
        remark_by_tid.setdefault(tid, []).append(rem)

    for bad_key in list(parsed.keys()):
        sk = str(bad_key).strip()
        if sk in valid_tids:
            continue
        if not re.fullmatch(r"\d{4,}", sk):
            continue
        hits = [
            tid
            for tid, rems in remark_by_tid.items()
            if any(
                r.startswith(sk)
                or r.startswith(sk + " ")
                or r.startswith(sk + "\u3000")
                or r.startswith(sk + "-")
                or r.startswith(sk + "ー")
                for r in rems
            )
        ]
        if len(hits) != 1:
            continue
        target = hits[0]
        val = parsed.pop(bad_key, None)
        if val is None:
            continue
        if target not in parsed:
            parsed[target] = val
            logging.info(
                "タスク特別指定: JSON トップキー誤りを修復（%r は依頼NOではない → %r）",
                bad_key,
                target,
            )
        else:
            parsed[bad_key] = val
    return parsed


def _normalize_task_special_scope_str(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return unicodedata.normalize("NFKC", str(s).strip())


def _task_special_scope_matches_row_field(row_val, restrict_val) -> bool:
    """
    restrict は無い・空なら制限なし（True）。
    非空なら Excel 坴の値とあいまい一致（部分一致坯）。
    """
    if restrict_val is None:
        return True
    r = _normalize_task_special_scope_str(restrict_val)
    if not r:
        return True
    v = _normalize_task_special_scope_str(row_val)
    if not v:
        return False
    if v == r:
        return True
    if r in v or v in r:
        return True
    return False


def _ai_remark_entry_applies_to_row(entry: dict, row) -> bool:
    """restrict_to_* は無いとしは同一依頼NOの全行に適用。"""
    if not isinstance(entry, dict):
        return False
    rp = row.get(TASK_COL_MACHINE, "")
    rm = row.get(TASK_COL_MACHINE_NAME, "")
    if not _task_special_scope_matches_row_field(rp, entry.get("restrict_to_process_name")):
        return False
    if not _task_special_scope_matches_row_field(rm, entry.get("restrict_to_machine_name")):
        return False
    return True


def _row_matches_remark_source_row(entry: dict, row) -> bool:
    """
    JSON の process_name / machine_name は」当該 Excel 行の工程名・機械名と一致するか。
    （プロンプトで渡した「備考はあった行」と対応るける。片方の値一致でも坯）
    """
    if not isinstance(entry, dict):
        return False
    rp = _normalize_task_special_scope_str(row.get(TASK_COL_MACHINE))
    rm = _normalize_task_special_scope_str(row.get(TASK_COL_MACHINE_NAME))
    ep = _normalize_task_special_scope_str(entry.get("process_name"))
    em = _normalize_task_special_scope_str(entry.get("machine_name"))
    proc_ok = (not ep) or (not rp) or ep == rp or ep in rp or rp in ep
    mac_ok = (not em) or (not rm) or em == rm or em in rm or rm in em
    return proc_ok and mac_ok


def _entry_is_global_task_special_scope(entry: dict) -> bool:
    """restrict_to_* は無い・空＝同一依頼NOの全工程行に効かせる指定。"""
    if not isinstance(entry, dict):
        return False
    a = _normalize_task_special_scope_str(entry.get("restrict_to_process_name"))
    b = _normalize_task_special_scope_str(entry.get("restrict_to_machine_name"))
    return not a and not b


def _select_ai_task_special_entry_for_tid_value(val, row):
    """1依頼NOに対れる値は dict または dict の配列のどうらでも行に坈ご覝素を返す。"""
    if val is None:
        return None
    if isinstance(val, list):
        for item in val:
            if (
                isinstance(item, dict)
                and _ai_remark_entry_applies_to_row(item, row)
                and _row_matches_remark_source_row(item, row)
            ):
                return item
        for item in val:
            if (
                isinstance(item, dict)
                and _ai_remark_entry_applies_to_row(item, row)
                and _entry_is_global_task_special_scope(item)
            ):
                return item
        for item in val:
            if isinstance(item, dict) and _ai_remark_entry_applies_to_row(item, row):
                return item
        return None
    if isinstance(val, dict):
        if _ai_remark_entry_applies_to_row(val, row):
            return val
        return None
    return None


def _ai_task_special_entry_for_row(ai_by_tid, row):
    """
    analyze_task_special_remarks の戻りから当該行のエントリを得る。
    プロンプトキーは正規化済み依頼NOなので」Excel は 12345.0 でもヒットれる。
    restrict_to_process_name / restrict_to_machine_name は無い・空のときは
    同一依頼NOの工程・機械は異なる全行にも指示を適用する。
    """
    if not isinstance(ai_by_tid, dict) or not ai_by_tid:
        return None
    tid_norm = planning_task_id_str_from_plan_row(row)
    tid_raw = str(_planning_df_cell_scalar(row, TASK_COL_TASK_ID) or "").strip()

    def try_val(v):
        return _select_ai_task_special_entry_for_tid_value(v, row)

    if tid_norm and tid_norm in ai_by_tid:
        hit = try_val(ai_by_tid[tid_norm])
        if hit is not None:
            return hit
    if tid_raw:
        for key in (tid_raw, str(tid_raw)):
            if key in ai_by_tid:
                hit = try_val(ai_by_tid[key])
                if hit is not None:
                    return hit
    if tid_norm:
        for k, v in ai_by_tid.items():
            if str(k).strip() == tid_norm:
                hit = try_val(v)
                if hit is not None:
                    return hit
    if tid_raw:
        for k, v in ai_by_tid.items():
            if str(k).strip() == tid_raw:
                hit = try_val(v)
                if hit is not None:
                    return hit
    return None


def _gemini_result_text(res):
    try:
        return (res.text or "").strip()
    except Exception:
        return ""


# 1 回の Python 実行（段階1 または 段階2）短縮でリセットれる
_gemini_usage_session: dict[str, dict[str, int]] = {}


def reset_gemini_usage_tracker() -> None:
    global _gemini_usage_session
    _gemini_usage_session = {}


def _gemini_cumulative_json_path() -> str:
    path = os.path.join(api_payment_dir, GEMINI_USAGE_CUMULATIVE_JSON_FILE)
    legacy = os.path.join(log_dir, GEMINI_USAGE_CUMULATIVE_JSON_FILE)
    if os.path.isfile(legacy) and not os.path.isfile(path):
        try:
            shutil.move(legacy, path)
        except OSError:
            pass
    return path


def _load_gemini_cumulative_payload() -> dict:
    """API_Payment 内の累計 JSON を読む。無い・壊れでいれみ初期形を返す。"""
    path = _gemini_cumulative_json_path()
    default: dict = {
        "version": 1,
        "updated_at": None,
        "calls_total": 0,
        "prompt_total": 0,
        "candidates_total": 0,
        "thoughts_total": 0,
        "total_tokens_reported": 0,
        "estimated_cost_usd_total": 0.0,
        "by_model": {},
    }
    if not os.path.isfile(path):
        _gemini_buckets_ensure_structure(default)
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict) or int(data.get("version") or 0) != 1:
            _gemini_buckets_ensure_structure(default)
            return default
        data.setdefault("by_model", {})
        _gemini_buckets_ensure_structure(data)
        for k in (
            "calls_total",
            "prompt_total",
            "candidates_total",
            "thoughts_total",
            "total_tokens_reported",
        ):
            data[k] = int(data.get(k) or 0)
        data["estimated_cost_usd_total"] = float(data.get("estimated_cost_usd_total") or 0.0)
        return data
    except Exception:
        _gemini_buckets_ensure_structure(default)
        return default


def _save_gemini_cumulative_payload(data: dict) -> None:
    path = _gemini_cumulative_json_path()
    try:
        os.makedirs(api_payment_dir, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8", newline="\n") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except OSError as ex:
        logging.debug("Gemini 累計 JSON の保存に失敗: %s", ex)


def _gemini_buckets_ensure_structure(data: dict) -> None:
    """累計 JSON に期間別ポケット用の辞書を用意れる（既存 v1 ファイルもマージ）。"""
    b = data.setdefault("buckets", {})
    if not isinstance(b, dict):
        b = {}
        data["buckets"] = b
    for sub in ("by_year", "by_month", "by_week", "by_day", "by_hour"):
        x = b.setdefault(sub, {})
        if not isinstance(x, dict):
            b[sub] = {}
    b.setdefault(
        "timezone_note",
        "period_key は PC ローカル時刻（datetime.now）で付与。他 PC との集計は混ずないでしてさい。",
    )


def _gemini_time_bucket_keys(dt: datetime) -> tuple[str, str, str, str, str]:
    """年・月・ISO週・日・時 のキー（文字列ソートで時系列比較しやれい形）。"""
    iy, iw, _ = dt.isocalendar()
    y = dt.strftime("%Y")
    ym = dt.strftime("%Y-%m")
    wk = f"{iy}-W{iw:02d}"
    d = dt.strftime("%Y-%m-%d")
    h = dt.strftime("%Y-%m-%dT%H")
    return y, ym, wk, d, h


def _gemini_bucket_add_one_call(
    buckets_root: dict,
    pt: int,
    ct: int,
    th: int,
    tt: int,
    inc_usd: float | None,
    *,
    when: datetime | None = None,
) -> None:
    """1 回の API 呼出しを年・月・週・日・時の坄ポケットに加算れる。"""
    dt = when or datetime.now()
    y, ym, wk, d, h = _gemini_time_bucket_keys(dt)
    pairs = (
        ("by_year", y),
        ("by_month", ym),
        ("by_week", wk),
        ("by_day", d),
        ("by_hour", h),
    )
    for sub, pk in pairs:
        subd = buckets_root.setdefault(sub, {})
        ent = subd.setdefault(
            pk,
            {
                "calls": 0,
                "prompt": 0,
                "candidates": 0,
                "thoughts": 0,
                "total_tokens": 0,
                "estimated_cost_usd": 0.0,
            },
        )
        ent["calls"] = int(ent.get("calls") or 0) + 1
        ent["prompt"] = int(ent.get("prompt") or 0) + pt
        ent["candidates"] = int(ent.get("candidates") or 0) + ct
        ent["thoughts"] = int(ent.get("thoughts") or 0) + th
        ent["total_tokens"] = int(ent.get("total_tokens") or 0) + tt
        if inc_usd is not None:
            ent["estimated_cost_usd"] = float(
                ent.get("estimated_cost_usd") or 0.0
            ) + float(inc_usd)


def _append_gemini_cumulative_one_call(
    model_id: str, pt: int, ct: int, th: int, tt: int
) -> None:
    """1 回の API 応答分を累計 JSON に加算れる（ログに坘発料金は出さない）。"""
    mid = str(model_id).strip()
    data = _load_gemini_cumulative_payload()
    data["calls_total"] = int(data["calls_total"]) + 1
    data["prompt_total"] = int(data["prompt_total"]) + pt
    data["candidates_total"] = int(data["candidates_total"]) + ct
    data["thoughts_total"] = int(data["thoughts_total"]) + th
    data["total_tokens_reported"] = int(data["total_tokens_reported"]) + tt
    bm: dict = data["by_model"]
    if mid not in bm or not isinstance(bm[mid], dict):
        bm[mid] = {
            "calls": 0,
            "prompt": 0,
            "candidates": 0,
            "thoughts": 0,
            "total": 0,
            "estimated_cost_usd": 0.0,
        }
    m = bm[mid]
    m["calls"] = int(m.get("calls") or 0) + 1
    m["prompt"] = int(m.get("prompt") or 0) + pt
    m["candidates"] = int(m.get("candidates") or 0) + ct
    m["thoughts"] = int(m.get("thoughts") or 0) + th
    m["total"] = int(m.get("total") or 0) + tt
    inc_usd = _gemini_estimate_cost_usd(mid, pt, ct, th)
    if inc_usd is not None:
        m["estimated_cost_usd"] = float(m.get("estimated_cost_usd") or 0.0) + float(inc_usd)
        data["estimated_cost_usd_total"] = float(
            data.get("estimated_cost_usd_total") or 0.0
        ) + float(inc_usd)
    _gemini_buckets_ensure_structure(data)
    _gemini_bucket_add_one_call(
        data["buckets"], pt, ct, th, tt, inc_usd, when=datetime.now()
    )
    data["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_gemini_cumulative_payload(data)


def record_gemini_response_usage(res, model_id: str) -> None:
    """generate_content の応答から usage_metadata を集計れる（セッション＋累計 JSON）。"""
    global _gemini_usage_session
    if res is None or not str(model_id or "").strip():
        return
    um = getattr(res, "usage_metadata", None)
    if um is None:
        return

    def _iv(name: str) -> int:
        v = getattr(um, name, None)
        try:
            return int(v) if v is not None else 0
        except (TypeError, ValueError):
            return 0

    pt = _iv("prompt_token_count")
    ct = _iv("candidates_token_count")
    tt = _iv("total_token_count")
    th = _iv("thoughts_token_count")
    if tt <= 0 and (pt > 0 or ct > 0 or th > 0):
        tt = pt + ct + th
    mid = str(model_id).strip()
    b = _gemini_usage_session.setdefault(
        mid,
        {"prompt": 0, "candidates": 0, "total": 0, "thoughts": 0, "calls": 0},
    )
    b["prompt"] += pt
    b["candidates"] += ct
    b["total"] += tt
    b["thoughts"] += th
    b["calls"] += 1
    try:
        _append_gemini_cumulative_one_call(mid, pt, ct, th, tt)
    except Exception as ex:
        logging.debug("Gemini 累計の更新で例外（続行）: %s", ex)


def _gemini_estimate_cost_usd(
    model_id: str, prompt_tok: int, cand_tok: int, thoughts_tok: int
) -> float | None:
    m = str(model_id).strip().lower()
    rin, rout = None, None
    if "flash" in m:
        rin, rout = _GEMINI_FLASH_IN_PER_M, _GEMINI_FLASH_OUT_PER_M
    elif "pro" in m:
        # 目安（未使用モデル坑けフォールバック）
        rin, rout = 1.25, 5.0
    if rin is None:
        return None
    out_equiv = cand_tok + thoughts_tok
    return (prompt_tok / 1_000_000.0) * rin + (out_equiv / 1_000_000.0) * rout


def _gemini_daily_trend_series(
    cum: dict, *, max_days: int | None = None
) -> tuple[list[str], list[float], str] | None:
    """累計 JSON の by_day から」日付キー（坤→新）・値・系列名。無ければ None。"""
    lim = GEMINI_USAGE_CHART_MAX_DAYS if max_days is None else max_days
    b = cum.get("buckets")
    if not isinstance(b, dict):
        return None
    subd = b.get("by_day")
    if not isinstance(subd, dict) or not subd:
        return None
    keys = sorted(subd.keys())
    keys = keys[-max(1, lim) :]
    usds: list[float] = []
    calls: list[int] = []
    for pk in keys:
        ent = subd.get(pk)
        if isinstance(ent, dict):
            usds.append(float(ent.get("estimated_cost_usd") or 0.0))
            calls.append(int(ent.get("calls") or 0))
        else:
            usds.append(0.0)
            calls.append(0)
    use_calls = sum(usds) <= 0.0 and sum(calls) > 0
    series = [float(c) for c in calls] if use_calls else usds
    label = "呼出し回数" if use_calls else "推定USD"
    return (keys, series, label)


def _gemini_daily_total_tokens_for_days(cum: dict, day_keys: list[str]) -> list[int]:
    """by_day の坄キーについで」total_tokens（無ければ prompt+candidates+thoughts）を返す。"""
    b = cum.get("buckets")
    if not isinstance(b, dict):
        return [0] * len(day_keys)
    subd = b.get("by_day")
    if not isinstance(subd, dict):
        return [0] * len(day_keys)
    out: list[int] = []
    for pk in day_keys:
        ent = subd.get(pk)
        if not isinstance(ent, dict):
            out.append(0)
            continue
        tt = int(ent.get("total_tokens") or 0)
        if tt <= 0:
            tt = (
                int(ent.get("prompt") or 0)
                + int(ent.get("candidates") or 0)
                + int(ent.get("thoughts") or 0)
            )
        out.append(tt)
    return out


def _gemini_usage_trend_caption_lines(cum: dict) -> list[str]:
    """テキスト坴はグラフ参照と CSV 案内のみ（ASCII スパークラインは出さない）。"""
    ser = _gemini_daily_trend_series(cum)
    if ser is None:
        return []
    keys, _, label = ser
    b = cum.get("buckets")
    lines = [
        "」推移グラフ】料金・呼出し: Q〜R 列＝トークン量: S〜T 列（坄グラフ・自動更新）を参照",
        f"  系列1: 日次 {label}（{keys[0]} ～ {keys[-1]}）",
        "  系列2: 日次 合計トークン（API 報告 total または内訳合計）",
        f"  年・月・週・時などの内訳: log\\{GEMINI_USAGE_BUCKETS_CSV_FILE}（Excel でグラフ坯）",
    ]
    if isinstance(b, dict):
        note = b.get("timezone_note")
        if note:
            lines.append(f"  （{note}）")
    return lines


def _gemini_resolve_main_sheet_xlwings(book) -> object | None:
    """xlwings Book からメイン相当シートを返す。無ければ None。"""
    for name in ("メイン", "メイン_", "Main"):
        try:
            return book.sheets[name]
        except Exception:
            continue
    try:
        for sht in book.sheets:
            try:
                if "メイン" in str(sht.name):
                    return sht
            except Exception:
                continue
    except Exception:
        pass
    return None


def _strip_gemini_usage_charts_xlwings(ws) -> None:
    """当機能は管理れる折れ線（坝剝またはグラフタイトル）を削除する。"""
    managed_names = (
        GEMINI_USAGE_XLW_CHART_NAME,
        GEMINI_USAGE_XLW_CHART_TOKENS_NAME,
    )
    title_markers = (
        "Gemini API 日次推移",
        "Gemini API 日次トークン",
    )
    try:
        charts_iter = list(ws.charts)
    except Exception:
        return
    for ch in charts_iter:
        try:
            if str(getattr(ch, "name", "") or "") in managed_names:
                ch.delete()
                continue
        except Exception:
            pass
        try:
            ca = ch.api
            if bool(ca.HasTitle):
                cap = getattr(ca.ChartTitle, "Caption", None)
                txt = getattr(ca.ChartTitle, "Text", None)
                title_s = str(cap or txt or "")
                for mk in title_markers:
                    if mk in title_s:
                        ch.delete()
                        break
        except Exception:
            pass


def _apply_main_sheet_gemini_usage_chart_xlwings(ws, cum: dict) -> None:
    """開いたブック上で Q〜R・S〜T を埋ゝ」折れ線グラフを 2 本まで置し（xlwings）。"""
    hr = GEMINI_USAGE_CHART_HEADER_ROW
    cdt = GEMINI_USAGE_CHART_COL_DATE
    cvl = GEMINI_USAGE_CHART_COL_VALUE
    cts = GEMINI_USAGE_CHART_COL_TOK_DATE
    ctv = GEMINI_USAGE_CHART_COL_TOK_VALUE
    nclear = GEMINI_USAGE_CHART_CLEAR_ROWS
    try:
        block = ws.range((hr, cdt), (hr + nclear - 1, ctv))
        block.clear_contents()
    except Exception:
        for i in range(nclear):
            r = hr + i
            for c in (cdt, cvl, cts, ctv):
                try:
                    ws.range((r, c)).clear_contents()
                except Exception:
                    pass

    _strip_gemini_usage_charts_xlwings(ws)
    ser = _gemini_daily_trend_series(cum)
    if ser is None:
        return
    day_keys, values, val_label = ser
    n = len(day_keys)
    if n <= 0:
        return

    ws.range((hr, cdt)).value = "日付"
    ws.range((hr, cvl)).value = val_label
    for i, (dk, val) in enumerate(zip(day_keys, values)):
        r = hr + 1 + i
        ws.range((r, cdt)).value = dk
        ws.range((r, cvl)).value = val
    try:
        vrng = ws.range((hr + 1, cvl), (hr + n, cvl))
        vrng.number_format = "0.000000" if val_label == "推定USD" else "0"
    except Exception:
        pass

    try:
        anchor = ws.range(GEMINI_USAGE_CHART_ANCHOR_CELL)
        left = float(anchor.left)
        top = float(anchor.top)
    except Exception:
        left, top = 0.0, 0.0
    chart = ws.charts.add(left=left, top=top, width=410, height=220)
    try:
        chart.name = GEMINI_USAGE_XLW_CHART_NAME
    except Exception:
        pass
    data_rng = ws.range((hr, cdt), (hr + n, cvl))
    chart.set_source_data(data_rng)
    try:
        chart.chart_type = "line"
    except Exception:
        try:
            chart.api.ChartType = 4
        except Exception:
            pass
    try:
        ca = chart.api
        ca.HasTitle = True
        ca.ChartTitle.Text = "Gemini API 日次推移"
        ca.HasLegend = False
    except Exception:
        pass

    tok_vals = _gemini_daily_total_tokens_for_days(cum, day_keys)
    if not tok_vals or max(tok_vals) <= 0:
        return

    tok_label = "合計トークン"
    ws.range((hr, cts)).value = "日付"
    ws.range((hr, ctv)).value = tok_label
    for i, dk in enumerate(day_keys):
        r = hr + 1 + i
        ws.range((r, cts)).value = dk
        ws.range((r, ctv)).value = int(tok_vals[i])
    try:
        ws.range((hr + 1, ctv), (hr + n, ctv)).number_format = "#,##0"
    except Exception:
        pass

    try:
        anchor2 = ws.range(GEMINI_USAGE_CHART_TOKENS_ANCHOR_CELL)
        left2 = float(anchor2.left)
        top2 = float(anchor2.top)
    except Exception:
        left2, top2 = left + 420.0, top
    chart2 = ws.charts.add(left=left2, top=top2, width=410, height=220)
    try:
        chart2.name = GEMINI_USAGE_XLW_CHART_TOKENS_NAME
    except Exception:
        pass
    data_rng2 = ws.range((hr, cts), (hr + n, ctv))
    chart2.set_source_data(data_rng2)
    try:
        chart2.chart_type = "line"
    except Exception:
        try:
            chart2.api.ChartType = 4
        except Exception:
            pass
    try:
        ca2 = chart2.api
        ca2.HasTitle = True
        ca2.ChartTitle.Text = "Gemini API 日次トークン"
        ca2.HasLegend = False
    except Exception:
        pass


def _write_main_sheet_gemini_usage_via_xlwings(
    macro_wb_path: str, text: str, log_prefix: str
) -> bool:
    """Excel でブックは開いているとし」メイン P 列・Q〜T・推移グラフ（最大2本）を xlwings で更新して Save。"""
    attached = _xlwings_attach_open_macro_workbook(macro_wb_path, log_prefix)
    if attached is None:
        logging.info(
            "%s: xlwings でマクロブックに接続でしう」メイン AI サマリをスキップしました。",
            log_prefix,
        )
        return False
    xw_book, info = attached
    ok = False
    try:
        try:
            xw_book.app.display_alerts = False
        except Exception:
            pass
        ws_main = _gemini_resolve_main_sheet_xlwings(xw_book)
        if ws_main is None:
            logging.info(
                "%s: メインシートはないため、xlwings での AI サマリをスキップしました。",
                log_prefix,
            )
            return False
        start_r, col_p, clear_n = 16, 16, 120
        _perf_snap = _xlwings_app_save_perf_state_push(xw_book.app)
        try:
            p_rng = ws_main.range((start_r, col_p)).resize(clear_n, 1)
            p_rng.clear_contents()
            lines_list = text.split("\n") if (text or "").strip() else []
            p_vals = [
                [lines_list[i] if i < len(lines_list) else None]
                for i in range(clear_n)
            ]
            p_rng.value = p_vals
            try:
                p_rng.api.WrapText = True
                p_rng.api.VerticalAlignment = -4160
            except Exception:
                pass
            _apply_main_sheet_gemini_usage_chart_xlwings(
                ws_main, _load_gemini_cumulative_payload()
            )
            xw_book.save()
            ok = True
            logging.info(
                "%s: メインシート P%d 以降・Gemini 推移グラフ（料金/呼出し・トークン）を xlwings で保存しました。",
                log_prefix,
                start_r,
            )
        finally:
            _xlwings_app_save_perf_state_pop(xw_book.app, _perf_snap)
    except Exception as ex:
        logging.warning(
            "%s: メイン AI サマリの xlwings 保存に失敗: %s", log_prefix, ex
        )
        ok = False
    finally:
        _xlwings_release_book_after_mutation(xw_book, info, ok)
    return ok


def _gemini_kv_table_lines(title: str, rows: list[tuple[str, str]]) -> list[str]:
    """累計・当実行坑けの 2 列テキスト表（履歴行は含まない）。"""
    out = [title]
    if not rows:
        return out
    lw = min(22, max(len(a) for a, _ in rows))
    sep = "  " + ("─" * (lw + 2 + 28))
    out.append(sep)
    for a, b in rows:
        out.append(f"  {a:<{lw}}  {b}")
    return out


def _export_gemini_buckets_csv_for_charts(cum: dict) -> None:
    """Excel 折れ線・棒グラフ坑けに長形式 CSV を log に書き出す。"""
    b = cum.get("buckets")
    if not isinstance(b, dict):
        return
    mapping = (
        ("year", "by_year"),
        ("month", "by_month"),
        ("week_iso", "by_week"),
        ("day", "by_day"),
        ("hour", "by_hour"),
    )
    rows_out: list[dict[str, object]] = []
    for gran_label, sub in mapping:
        subd = b.get(sub)
        if not isinstance(subd, dict):
            continue
        for pk in sorted(subd.keys()):
            ent = subd.get(pk)
            if not isinstance(ent, dict):
                continue
            calls = int(ent.get("calls") or 0)
            pt = int(ent.get("prompt") or 0)
            cc = int(ent.get("candidates") or 0)
            th = int(ent.get("thoughts") or 0)
            tt = int(ent.get("total_tokens") or 0)
            usd = float(ent.get("estimated_cost_usd") or 0.0)
            rows_out.append(
                {
                    "granularity": gran_label,
                    "period_key": pk,
                    "calls": calls,
                    "prompt_tokens": pt,
                    "candidates_tokens": cc,
                    "thoughts_tokens": th,
                    "total_tokens": tt,
                    "estimated_cost_usd": round(usd, 8),
                    "estimated_cost_jpy": round(usd * GEMINI_JPY_PER_USD, 4),
                }
            )
    if not rows_out:
        return
    path = os.path.join(log_dir, GEMINI_USAGE_BUCKETS_CSV_FILE)
    fieldnames = [
        "granularity",
        "period_key",
        "calls",
        "prompt_tokens",
        "candidates_tokens",
        "thoughts_tokens",
        "total_tokens",
        "estimated_cost_usd",
        "estimated_cost_jpy",
    ]
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows_out)
    except OSError as ex:
        logging.debug("Gemini ポケット CSV の保存に失敗: %s", ex)


def build_gemini_usage_summary_text() -> str:
    """メイン表示・結果ログ用の複数行テキスト（この実行分＋累計 JSON）。"""
    cum = _load_gemini_cumulative_payload()
    ct_tot = int(cum.get("calls_total") or 0)
    if not _gemini_usage_session and ct_tot <= 0:
        return ""

    lines: list[str] = []
    ts = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    if _gemini_usage_session:
        lines.append(f"集計時刻: {ts}（この実行での Gemini API）")
        tot_calls = sum(b["calls"] for b in _gemini_usage_session.values())
        tot_p = sum(b["prompt"] for b in _gemini_usage_session.values())
        tot_c = sum(b["candidates"] for b in _gemini_usage_session.values())
        tot_th = sum(b["thoughts"] for b in _gemini_usage_session.values())
        tot_t = sum(b["total"] for b in _gemini_usage_session.values())
        sess_rows: list[tuple[str, str]] = [
            ("呼出し", f"{tot_calls:,} 回"),
            ("入力トークン", f"{tot_p:,}"),
            ("出力トークン", f"{tot_c:,}"),
        ]
        if tot_th:
            sess_rows.append(("思考トークン", f"{tot_th:,}"))
        sess_rows.append(("total 報告", f"{tot_t:,}"))
        lines.extend(_gemini_kv_table_lines("」この実行】", sess_rows))
        grand_usd = 0.0
        any_price = False
        for mid in sorted(_gemini_usage_session.keys()):
            b = _gemini_usage_session[mid]
            mrows: list[tuple[str, str]] = [
                ("モデル", mid),
                ("呼出し", f"{b['calls']:,} 回"),
                ("入力トークン", f"{b['prompt']:,}"),
                ("出力トークン", f"{b['candidates']:,}"),
            ]
            if b.get("thoughts", 0):
                mrows.append(("思考トークン", f"{b['thoughts']:,}"))
            mrows.append(("total_token_count", f"{b['total']:,}"))
            est = _gemini_estimate_cost_usd(
                mid, b["prompt"], b["candidates"], b.get("thoughts", 0)
            )
            if est is not None:
                any_price = True
                grand_usd += est
                mrows.append(("推定USD", f"${est:.6f}"))
                mrows.append(
                    (
                        "推定JPY",
                        f"¥{est * GEMINI_JPY_PER_USD:.2f}（{GEMINI_JPY_PER_USD:.0f}円/USD）",
                    )
                )
            else:
                mrows.append(("推定料金", "（坘価未登録モデル）"))
            lines.append("")
            lines.extend(_gemini_kv_table_lines(f"」この実行・モデル別】", mrows))
        if any_price:
            lines.append("")
            lines.extend(
                _gemini_kv_table_lines(
                    "」この実行・推定料金合計】",
                    [
                        ("USD", f"${grand_usd:.6f}"),
                        (
                            "JPY",
                            f"¥{grand_usd * GEMINI_JPY_PER_USD:.2f}（{GEMINI_JPY_PER_USD:.0f}円/USD）",
                        ),
                    ],
                )
            )
    else:
        lines.append(f"集計時刻: {ts}")
        lines.append("（この実行での Gemini API 呼出しはありません）")
    lines.append("※ トークンは API の usage_metadata に基るしした。")
    lines.append(
        "※ USD 坘価はコード＝環境変数の目安です。実課金は Google の請求を参照してください。"
    )
    lines.append(
        "※ 坄 API 呼出しととの推定料金はコンソールに出さう」下記累計 JSON にのみ穝み上きした。"
    )

    if ct_tot > 0:
        lines.append("")
        cum_hdr = (
            f"」累計】{GEMINI_USAGE_CUMULATIVE_JSON_FILE} "
            "（API_Payment フォルダ・全実行の推定値）"
        )
        pt0 = int(cum.get("prompt_total") or 0)
        cc0 = int(cum.get("candidates_total") or 0)
        th0 = int(cum.get("thoughts_total") or 0)
        tt0 = int(cum.get("total_tokens_reported") or 0)
        cum_rows: list[tuple[str, str]] = [
            ("最終更新", str(cum.get("updated_at") or "—")),
            ("呼出し", f"{ct_tot:,} 回"),
            ("入力トークン", f"{pt0:,}"),
            ("出力トークン", f"{cc0:,}"),
        ]
        if th0:
            cum_rows.append(("思考トークン", f"{th0:,}"))
        cum_rows.append(("total 報告", f"{tt0:,}"))
        usd_all = float(cum.get("estimated_cost_usd_total") or 0.0)
        if usd_all > 0:
            cum_rows.append(("推定USD 累計", f"${usd_all:.6f}"))
            cum_rows.append(
                (
                    "推定JPY 累計",
                    f"¥{usd_all * GEMINI_JPY_PER_USD:.2f}（{GEMINI_JPY_PER_USD:.0f}円/USD）",
                )
            )
        lines.extend(_gemini_kv_table_lines(cum_hdr, cum_rows))
        bm = cum.get("by_model") or {}
        if isinstance(bm, dict) and bm:
            for mid in sorted(bm.keys()):
                m = bm[mid]
                if not isinstance(m, dict):
                    continue
                mrows2: list[tuple[str, str]] = [
                    ("モデル", mid),
                    ("呼出し", f"{int(m.get('calls') or 0):,} 回"),
                    (
                        "入力 / 出力",
                        f"{int(m.get('prompt') or 0):,} / {int(m.get('candidates') or 0):,}",
                    ),
                ]
                if int(m.get("thoughts") or 0):
                    mrows2.append(("思考トークン", f"{int(m.get('thoughts') or 0):,}"))
                mud = float(m.get("estimated_cost_usd") or 0.0)
                if mud > 0:
                    mrows2.append(("推定USD 累計", f"${mud:.6f}"))
                    mrows2.append(
                        ("推定JPY 累計", f"¥{mud * GEMINI_JPY_PER_USD:.2f}")
                    )
                lines.append("")
                lines.extend(_gemini_kv_table_lines("」累計・モデル別】", mrows2))
        trend = _gemini_usage_trend_caption_lines(cum)
        if trend:
            lines.append("")
            lines.extend(trend)
    return "\n".join(lines)


def write_main_sheet_gemini_usage_summary(wb_path: str, log_prefix: str) -> None:
    """Gemini 利用サマリを log に書き」xlwings でメイン P 列・推移グラフへ保存（開いているブック坑け）。"""
    text = build_gemini_usage_summary_text()
    path = os.path.join(log_dir, GEMINI_USAGE_SUMMARY_FOR_MAIN_FILE)
    xw_ok = False
    if wb_path and os.path.isfile(wb_path):
        try:
            xw_ok = _write_main_sheet_gemini_usage_via_xlwings(
                wb_path, text, log_prefix
            )
        except Exception as ex:
            logging.warning(
                "%s: AI サマリの xlwings 書き込みで例外: %s", log_prefix, ex
            )
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
    except OSError:
        pass
    try:
        cum2 = _load_gemini_cumulative_payload()
        if int(cum2.get("calls_total") or 0) > 0:
            _export_gemini_buckets_csv_for_charts(cum2)
    except Exception as ex:
        logging.debug("Gemini ポケット CSV 出力で例外（続行）: %s", ex)
    if xw_ok:
        return
    if text.strip():
        logging.info(
            "%s: メイン P 列・グラフを xlwings で保存でしませんでした。"
            " %s に出力済み → マクロ「メインシート_Gemini利用サマリをP列に反映」で P 列のみ反映でしした。",
            log_prefix,
            path,
        )
    else:
        logging.info(
            "%s: Gemini 未使用: サマリを空で %s に出力。",
            log_prefix,
            path,
        )


def _try_write_main_sheet_gemini_usage_summary(phase: str) -> None:
    try:
        write_main_sheet_gemini_usage_summary(TASKS_INPUT_WORKBOOK, phase)
    except Exception as ex:
        logging.warning(
            "%s: メインシートへの AI 利用サマリ書き込みで例外（続行）: %s", phase, ex
        )


def _plan_sheet_write_global_parse_block_to_ws(
    ws,
    global_priority_override: dict,
    when_str: str,
) -> None:
    """既に開いている「配台計画_タスク入力」相当シートへ AX:AY のグローバル解析ブロックを書き。"""
    gpo = global_priority_override or {}
    lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
    vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
    max_r = PLAN_SHEET_GLOBAL_PARSE_MAX_ROWS
    for i in range(max_r):
        ws.cell(row=1 + i, column=lc, value=None)
        ws.cell(row=1 + i, column=vc, value=None)
    align_top = Alignment(wrap_text=True, vertical="top")
    pairs: list[tuple[str, str]] = [
        ("」グローバルコメント解析】", "参照用・段階2で自動記録"),
        (
            "※二重適用についで",
            "配台への反映はメインシート「グローバルコメント」からのみ行ゝれした。"
            "このAX〜AY列は読み坖られません。編集しても次回実行まで配台に効しません。"
            "原文はメイン欄を参照してください。",
        ),
        ("計画基準日時", (when_str or "").strip() or "―"),
        (
            "工場休業日",
            ", ".join(str(x) for x in (gpo.get("factory_closure_dates") or []))
            if gpo.get("factory_closure_dates")
            else "（なし）",
        ),
        (
            "スキル覝件を無視",
            "はい" if gpo.get("ignore_skill_requirements") else "いいえ",
        ),
        (
            "need人数1固定",
            "はい" if gpo.get("ignore_need_minimum") else "いいえ",
        ),
        (
            "配台制限の撤廃",
            "はい" if gpo.get("abolish_all_scheduling_limits") else "いいえ",
        ),
        (
            "グローバルOP指定",
            json.dumps(gpo.get("task_preferred_operators") or {}, ensure_ascii=False)
            if gpo.get("task_preferred_operators")
            else "（なし）",
        ),
        (
            "日付×工程フォーム指定",
            json.dumps(
                gpo.get("global_day_process_operator_rules") or [],
                ensure_ascii=False,
            )
            if gpo.get("global_day_process_operator_rules")
            else "（なし）",
        ),
        (
            "グローバル速度ルール",
            json.dumps(gpo.get("global_speed_rules") or [], ensure_ascii=False)
            if gpo.get("global_speed_rules")
            else "（なし）",
        ),
        (
            "未適用メモ(AI)",
            str(gpo.get("scheduler_notes_ja") or "").strip() or "（なし）",
        ),
        (
            "AI覝約",
            str(gpo.get("interpretation_ja") or "").strip() or "（なし）",
        ),
    ]
    for i, (lab, val) in enumerate(pairs):
        if i >= max_r:
            break
        c1 = ws.cell(row=1 + i, column=lc, value=lab)
        c2 = ws.cell(row=1 + i, column=vc, value=val)
        c1.alignment = align_top
        c2.alignment = align_top


def write_plan_sheet_global_comment_parse_block(
    wb_path: str,
    sheet_name: str,
    global_priority_override: dict,
    *,
    when_str: str,
    log_prefix: str = "段階2",
) -> bool:
    """
    「配台計画_タスク入力」シートの坳端付近（AX:AY）に」グローバルコメントの解析結果を書き込む。
    メイン原文はここに転記しない（メイン欄との重複・誤解を避ける）。本列は再読込されう参照専用。
    Excel でブックを開いたままてと保存に失敗することはある（他の openpyxl 書込と同様）。
    """
    if not wb_path or not os.path.isfile(wb_path):
        return False
    gpo = global_priority_override or {}
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: ブックに「%s」があるため、openpyxl でグローバルコメント解析を配台シートへ書き込みません。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    try:
        wb = load_workbook(
            wb_path, keep_vba=keep_vba, read_only=False, data_only=False
        )
    except Exception as ex:
        logging.info(
            "%s: グローバルコメント解析の配台シート書込のため、ブックを開きません: %s",
            log_prefix,
            ex,
        )
        return False
    try:
        if sheet_name not in wb.sheetnames:
            logging.info(
                "%s: シート '%s' はないため、グローバルコメント解析の反映をスキップ。",
                log_prefix,
                sheet_name,
            )
            return False
        ws = wb[sheet_name]
        _plan_sheet_write_global_parse_block_to_ws(ws, gpo, when_str)
        lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
        vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
        wb.save(wb_path)
        logging.info(
            "%s: 「%s」%s:%s 列にグローバルコメント解析を保存しました。",
            log_prefix,
            sheet_name,
            get_column_letter(lc),
            get_column_letter(vc),
        )
        return True
    except OSError as ex:
        logging.warning(
            "%s: グローバルコメント解析を配台シートへ保存でしませんでした（Excel で開いたまま等）: %s",
            log_prefix,
            ex,
        )
        return False
    except Exception as ex:
        logging.warning(
            "%s: グローバルコメント解析の配台シート書込で例外: %s", log_prefix, ex
        )
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _try_write_plan_sheet_global_comment_parse_block(
    global_priority_override: dict,
    when_str: str,
) -> None:
    try:
        write_plan_sheet_global_comment_parse_block(
            TASKS_INPUT_WORKBOOK,
            PLAN_INPUT_SHEET_NAME,
            global_priority_override,
            when_str=when_str,
            log_prefix="段階2",
        )
    except Exception as ex:
        logging.warning(
            "段階2: 配台シートへのグローバルコメント解析書き込みで例外（続行）: %s",
            ex,
        )


def _try_write_plan_input_global_parse_and_conflicts_one_save(
    global_priority_override: dict,
    when_str: str,
    num_data_rows: int,
    conflicts_by_row,
) -> None:
    try:
        write_plan_sheet_global_parse_and_conflict_styles_one_io(
            TASKS_INPUT_WORKBOOK,
            PLAN_INPUT_SHEET_NAME,
            global_priority_override,
            when_str=when_str,
            num_data_rows=num_data_rows,
            conflicts_by_row=conflicts_by_row,
            log_prefix="段階2",
        )
    except Exception as ex:
        logging.warning(
            "段階2: 配台シートへのグローバル解析＋矛盾着色（1回保存）で例外（続行）: %s",
            ex,
        )


def _log_task_special_ai_response(raw_text, parsed, extracted_json_str, prompt_text=None):
    """特別指定_備考坑け Gemini のプロンプト・生テキスト・抽出JSON・パース結果を1ファイルに残れ。"""
    path = os.path.join(log_dir, TASK_SPECIAL_AI_LAST_RESPONSE_FILE)
    try:
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            if prompt_text is not None and str(prompt_text).strip():
                f.write("=== Gemini へ逝信したプロンプト（全文） ===\n")
                f.write(str(prompt_text).strip())
                f.write("\n\n")
            f.write("=== Gemini 返坴テキスト（モデル出力しのまま） ===\n")
            f.write(raw_text or "")
            f.write(
                "\n\n=== AI は返したテキストからクライアントは切り出した JSON 文字列 ===\n"
                "（※ユーザー特別指定の解析に正覝表睾は使っていません。モデル応答のパース用です）\n"
            )
            f.write(extracted_json_str if extracted_json_str else "(抽出なし)")
            f.write("\n\n=== json.loads 後（依頼NOキー） ===\n")
            if isinstance(parsed, dict):
                f.write(json.dumps(parsed, ensure_ascii=False, indent=2))
            else:
                f.write("(パースでしう)")
        logging.info(
            "タスク特別指定: プロンプト＋AI応答の詳細 → %s",
            path,
        )
    except OSError as ex:
        logging.warning("タスク特別指定: AI応答ファイル保存に失敗: %s", ex)
    if isinstance(parsed, dict) and parsed:
        logging.info(
            "タスク特別指定: 解析された依頼NO: %s",
            ", ".join(sorted(parsed.keys(), key=lambda x: str(x))),
        )
        for tid_k in sorted(parsed.keys(), key=lambda x: str(x)):
            logging.info(
                "  依頼NO [%s] AI解析フィールド: %s",
                tid_k,
                json.dumps(parsed[tid_k], ensure_ascii=False),
            )


def _parse_and_log_task_special_gemini_response(res, prompt_text=None):
    """
    API レスポンスを JSON 化しログ＝ファイルへ記録。失敗時は None。
    ユーザーの特別指定文言には触れう」モデル出力から JSON ブロックを取り出す処理のみ。
    """
    raw = _gemini_result_text(res)
    if raw:
        stripped = raw.strip()
        if stripped.startswith("{"):
            try:
                trial = json.loads(stripped)
                if isinstance(trial, dict):
                    _log_task_special_ai_response(raw, trial, stripped, prompt_text)
                    return trial
            except json.JSONDecodeError:
                pass
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        _log_task_special_ai_response(raw, {}, None, prompt_text)
        logging.warning(
            "タスク特別指定: AI応答から JSON を抽出でしませんでした。生テキスト先頭 3000 文字:\n%s",
            (raw[:3000] if raw else "(空)"),
        )
        return None
    extracted = match.group(0)
    try:
        parsed = json.loads(extracted)
    except json.JSONDecodeError as je:
        _log_task_special_ai_response(raw, None, extracted, prompt_text)
        logging.warning("タスク特別指定: JSON パース失敗: %s", je)
        return None
    if not isinstance(parsed, dict):
        _log_task_special_ai_response(raw, None, extracted, prompt_text)
        logging.warning("タスク特別指定: トップレベルは JSON オブジェクトではありません。")
        return None
    _log_task_special_ai_response(raw, parsed, extracted, prompt_text)
    return parsed


def analyze_task_special_remarks(tasks_df, reference_year=None, ai_sheet_sink: dict | None = None):
    """
    「配台計画_タスク入力」の「特別指定_備考」を AI で構造化（セルに値はある項目は後段でセルを優先）。
    「配台試行」はオンな行はプロンプトに載せない（API 節約・当該行は配台しないため）。
    担当OP指定はプロンプトの返坴契約でモデルに preferred_operator を出力させる（備考を正覝表睾で切り出す処理は行ゝない）。
    json/ai_remarks_cache.json に TTL AI_CACHE_TTL_SECONDS でキャッシュ（同一入力・同一基準年なら API を呼みない）。
    依頼NOは数値表記・全角などを正規化してキーを安定化し、基準年は指紋に含むで日付解釈の変化とキャッシュの食い靕いを防し。

    戻り値の例: 依頼NO -> オブジェクト」または同一依頼NOに備考行は複数ある場合はオブジェクトの配列。
      process_name, machine_name … 当該備考セルはある行の工程名・機械名（プロンプトの行と一致）
      restrict_to_process_name, restrict_to_machine_name … 省略または空なら同一依頼NOの全工程・全機械行に適用。
      しの他 required_op, speed_override, task_efficiency, priority, start_date, start_time,
      target_completion_date, ship_by_date, preferred_operator など。
    """
    lines = _task_special_prompt_lines(tasks_df)
    # #region agent log
    _w4_lines = [ln for ln in lines if "W4-5" in ln or "w4-5" in ln.casefold()]
    _debug_agent_ndjson_55ab3a(
        "H4",
        "_core.py:analyze_task_special_remarks:prompt_lines",
        "特別指定AIプロンプト行にW4-5が含まれるか",
        {"n_lines": len(lines), "w4_5_line_count": len(_w4_lines)},
    )
    # #endregion
    if not lines:
        n_rows = len(tasks_df)
        n_rem_only = 0
        n_tid_raw = 0
        for _, row in tasks_df.iterrows():
            tid = planning_task_id_str_from_plan_row(row)
            rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
            if tid:
                n_tid_raw += 1
            if rem:
                n_rem_only += 1
        miss_col = PLAN_COL_SPECIAL_REMARK not in tasks_df.columns
        logging.warning(
            "タスク特別指定: AI 解析対象はありません（「%s」列は%s）。"
            "总行数=%s」依頼NOのある行=%s」備考は入っている行=%s。"
            "段階2実行剝にブックを保存し、本当に「%s」列へ入力しているか確認してください。",
            PLAN_COL_SPECIAL_REMARK,
            "見つかりません" if miss_col else "空の可能性はありした",
            n_rows,
            n_tid_raw,
            n_rem_only,
            PLAN_COL_SPECIAL_REMARK,
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "スキップ（対象行なし）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（対象行なし・API 未実行）"
        return {}

    blob = "\n".join(sorted(lines))
    ref_y = int(reference_year) if reference_year is not None else date.today().year
    cache_fingerprint = f"{ref_y}\n{blob}"
    cache_key_input = f"{TASK_SPECIAL_CACHE_KEY_PREFIX}{cache_fingerprint}"
    cache_key = hashlib.sha256(cache_key_input.encode("utf-8")).hexdigest()
    ai_cache = load_ai_cache()
    cached_parsed = get_cached_ai_result(
        ai_cache, cache_key, content_key=cache_fingerprint
    )
    if cached_parsed is not None:
        logging.info(
            "タスク特別指定: キャッシュヒット（%s 件・基準年=%s）。Gemini は呼びません。",
            len(lines),
            ref_y,
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "なし（キャッシュ使用）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（キャッシュ利用・今回 API 未実行）"
        out = copy.deepcopy(cached_parsed)
        if isinstance(out, dict):
            _repair_task_special_ai_wrong_top_level_keys(out, tasks_df)
        return out

    logging.info(
        "タスク特別指定: キャッシュなし。Gemini で %s 件の備考を解析しした（基準年=%s）。",
        len(lines),
        ref_y,
    )

    if not API_KEY:
        logging.info("GEMINI_API_KEY 未設定のため、タスク特別指定のAI解析をスキップしました。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "なし（APIキー未設定）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（API キー未設定）"
        return {}

    prompt = f"""
あなたは工場の配台計画坑けに」Excel「特別指定_備考」欄への自由記述を読み」配台ロジックは使うるフィールドの値に蝽とし込むアシスタントです。

」最針覝】
1) 」特別指定原文】の坄行は」ユーザーはセルに入力した文字列を **改変・覝約・断う切りはしてよらう**（先頭末尾の空白のみ除去）」しのまま渡していした。**原文の事実や愝図を別の文言に置し杛ごないでしてさい。**
2) あなたの応答は **1個の JSON オブジェクトのみ**（先頭は {{ 」末尾は }} ）。説明文・マークダウン・コードフェンスは禁止。
3) JSON のトップレベルキーは」坄行の **依頼NO」と】の間の文字列のみ** と **完全一致** させること。**備考本文**に書かれた哝番・原板坝・製哝コード（例: 20010 で始まる番坷列）をキーにしてはならない。備考はしのよごな番坷で始まっていでも」キーは必う」】内の依頼NOの値とれる。

」返坴JSONの契約（この節どよりに出力れること）】
■ トップレベル
- キー: 上記」特別指定原文】の **依頼NO」…】の括弧内** の文字列と **完全一致**（表記・ポイフン・英大文字尝文字を原文どより）。備考本文中の数字列をキーにしない。
- 値: 次のいうれか。
  (A) **JSONオブジェクト1つ** … 当該依頼NOの備考はプロンプト上 **1行の値** のとき。
  (B) **JSON配列**（覝素はオブジェクト）… 同一依頼NOで工程名・機械名は異なる備考行は **複数** あるとし。覝素の順はプロンプトの行順と対応させる。

■ process_name（文字列）・machine_name（文字列）— **必須**
- 当該備考に対応れるプロンプト行の **工程名「…」**・**機械名「…」** の値と **一致** させる（「（空）」のときは空文字列 ""）。
- ログ・トレース用。省略試行。

■ restrict_to_process_name（文字列）・restrict_to_machine_name（文字列）— **任愝**
- **原文は「特定の工程の値」「この機械の値」など」適用範囲を絞っているとしの値** 出力れる。
- **原文に工程名・機械名の陝定は無い**（依頼全体・全行程への指示）としは **両方とも省略** れるか **空文字列 ""** とれる。
- しの場合」配台ロジックは **同一依頼NOの別行（例: エンボス行と分割行）にも指示を適用** れる。
- 絞る場合は」原文で示された識別名を入れる（Excel の工程名・機械名と照合しやれい表記）。

■ preferred_operator（文字列）— 条件付し**必須**
- **必須条件**: 当該依頼の原文を読み」「**誰はこの加工・作業の主担当（OP）として割り当でたいか**」は **愝味として** 読み坖れるとし。
  例: 特定の人にやってもらご＝しの人に任せる＝担当はあの人＝OPは〜＝〜さん（並び）に依頼」など。**表睾の型に依存せう**」文の愝味で判断れる。
- **満たしたとしの出力義務**: 上記の愝味は成立れると判断したオブジェクトでは」**必う** キー `preferred_operator` を含む」値は **空でない文字列** とれる。併せで **process_name / machine_name は必須**（例: `{{"process_name":"…","machine_name":"…","preferred_operator":"…"}}`）。
- **値の形式**: 原文で示された **担当者の識別名を1坝分**（姓・坝・ニックフォーム等」原文に睾れた表記を維挝）。末尾の敬称（さん・坛・氝）のみ除去。例:「森岡さんにやってもらいした」→ `"森岡"`。
- **出力してはいけないとし**: 原文に担当者の指愝は **一切ない** と判断した依頼NOでは `preferred_operator` キー自体を **省略** れる（空文字列も付けない）。

■ しの他フィールド（required_op, speed_override, task_efficiency, priority, start_date, start_time, target_completion_date, ship_by_date）
- 原文から **明確に** 読み坖れる場合のみ出力。読み取れない数値・日付は **省略**（推測で埋ゝない）。

」同一依頼NO・複数工程の例】
依頼NO Y4-2 に「エンボス」と「分割」の行はあり」備考は「4/5までに終ゝらせる」のみで工程の陝定は無い場合:
- process_name / machine_name は **備考は書かれた行** の値を入れる。
- restrict_to_* は **出さないか空** にし、**エンボス行・分割行の両方** にも優先度・日付等は効しよごにれる。

」基準年（年なし日付用）】
「4/5」「4/5に出蝷」のよごに **年は無い** 日付は原則 **西暦 {ref_y} 年** とし、YYYY-MM-DD で出力。

」フィールド一覧（型の参考）】
- process_name, machine_name: 文字列（必須。プロンプト行と一致）
- restrict_to_process_name, restrict_to_machine_name: 文字列（任愝。陝定なら）
- preferred_operator: 文字列（上記契約に従ご）
- required_op: 正の整数
- speed_override: 正の数（m/分）
- task_efficiency: 0〜1
- priority: 整数（尝さいろど先に割付）
- start_date: YYYY-MM-DD / start_time: HH:MM
- target_completion_date, ship_by_date: YYYY-MM-DD

」解釈の指針】
- 「間に坈ごよごに」「繰り上きる」→ priority を上きる（数値を下きる）。日付は文中にあれみ target_completion_date または ship_by_date に入れる。
- 担当者指定は **愝味睆解** で preferred_operator を決ゝる（特定のキーワード列挙に頼らない）。
- 数値・日付は推測で補ゝない。
- **備考は特定の工程・機械にの値言坊していない陝り**」restrict_to_* は空にし、同一依頼NOの他行にも適用される形にれる。

」出力直後の自己検証（必う実行してから JSON を閉もる）】
- 」特別指定原文】の **坄行** についで」対応れるオブジェクトに **process_name** と **machine_name** はあるか。
- 同一依頼NOは複数行あるとしは **配列** で坄行に1オブジェクト」または革切にマージした坘一オブジェクト＋restrict の靋用を一貫させる。
- 「主担当OPの指愝」はある行では **非空の preferred_operator** を付ける。

」出力形式の例】（依頼NO・値は実データに合わせ替ごること）
{{
  "W3-14": {{
    "process_name": "検査",
    "machine_name": "ラインA",
    "preferred_operator": "森岡"
  }},
  "Y3-26": {{
    "process_name": "コーティング",
    "machine_name": "",
    "priority": 1,
    "ship_by_date": "{ref_y}-04-05",
    "target_completion_date": "{ref_y}-04-05"
  }},
  "Y4-2": {{
    "process_name": "エンボス",
    "machine_name": "E1",
    "priority": 2,
    "restrict_to_process_name": "",
    "restrict_to_machine_name": ""
  }}
}}

」特別指定原文】（Excel からしのまま。1行＝依頼NOと備考のペア）
{blob}
"""
    try:
        ppath = os.path.join(log_dir, "ai_task_special_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("タスク特別指定: 今回 Gemini に渡したプロンプト全文 → %s", ppath)
    except OSError as ex:
        logging.warning("タスク特別指定: プロンプト保存失敗: %s", ex)

    client = _gemini_client(API_KEY)
    try:
        res, gem_model_used = _gemini_generate_content_with_retry(
            client, contents=prompt, log_label="タスク特別指定"
        )
        record_gemini_response_usage(res, gem_model_used)
        parsed = _parse_and_log_task_special_gemini_response(res, prompt_text=prompt)
        if parsed is not None:
            _repair_task_special_ai_wrong_top_level_keys(parsed, tasks_df)
            put_cached_ai_result(
                ai_cache, cache_key, parsed, content_key=cache_fingerprint
            )
            save_ai_cache(ai_cache)
            logging.info("タスク特別指定: AI解析は完了しました。")
            if ai_sheet_sink is not None:
                ai_sheet_sink["特別指定備考_AI_API"] = "あり"
                ai_sheet_sink["特別指定備考_Geminiモデル"] = gem_model_used
            return parsed
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "あり（JSON解釈失敗）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = gem_model_used
        return {}
    except Exception as e:
        logging.warning("タスク特別指定: Gemini 呼び出し失敗（再試行尽き）: %s", e)
        logging.warning(
            "タスク特別指定: AI解析結果を取得でしなかったため、特別指定_備考の開始日/優先指示は反映されません。"
            "（列「加工開始日_指定」「指定納期_上書き」は廃止済み。備考の再記載または後から AI 再実行を検討してください。）"
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = f"失敗: {e}"[:500]
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（呼び出し失敗）"
        return {}


def _merge_preferred_operator_cell_and_ai(row, ai_for_tid):
    """Excel「担当OP_指定」を優先し、空なら AI の preferred_operator。"""
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}
    v = row.get(PLAN_COL_PREFERRED_OP)
    if v is not None and not (isinstance(v, float) and pd.isna(v)):
        s = str(v).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
    a = ai.get("preferred_operator")
    if a is not None:
        s = str(a).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
    return ""


def _global_override_preferred_operator_for_task(tpref, task_id) -> str | None:
    """
    メイン「再優先特別記載」の task_preferred_operators。
    キーは依頼NO（大文字・尝文字の差は無視）。
    """
    if not isinstance(tpref, dict) or not tpref:
        return None
    tid = str(task_id).strip()
    if not tid:
        return None
    tlo = tid.lower()
    for k, v in tpref.items():
        if str(k).strip().lower() != tlo:
            continue
        s = str(v).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
        return None
    return None


def _merge_task_row_with_ai(
    row, ai_for_tid, *, allow_ai_dispatch_priority_from_remark: bool = True
):
    """
    上書き列は加工速度_上書き・原板投入日_上書き等のみ（計画シート）。しの他は特別指定備考 AI から。
    allow_ai_dispatch_priority_from_remark は False のとき」AI の required_op / task_efficiency / priority /
    start_date / start_time は採用しない（備考に紝期系文言は無い行坑け）。
    """
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}

    def first_float_pos_cell_or_ai(cell, ai_key):
        v = parse_float_safe(row.get(cell), None)
        if v is not None and (not isinstance(v, float) or not pd.isna(v)) and float(v) > 0:
            return float(v)
        a = ai.get(ai_key)
        try:
            if a is not None and float(a) > 0:
                return float(a)
        except (TypeError, ValueError):
            pass
        return None

    if allow_ai_dispatch_priority_from_remark:
        req_op = parse_optional_int(ai.get("required_op"))
    else:
        req_op = None
    if req_op is not None and req_op < 1:
        req_op = None

    if allow_ai_dispatch_priority_from_remark:
        te = None
        a = ai.get("task_efficiency")
        try:
            if a is not None and float(a) > 0:
                te = float(a)
        except (TypeError, ValueError):
            te = None
        if te is None or te <= 0:
            te = 1.0
    else:
        te = 1.0

    if allow_ai_dispatch_priority_from_remark:
        pri = parse_optional_int(ai.get("priority"))
    else:
        pri = None
    if pri is None:
        pri = 999

    st_date = None
    if allow_ai_dispatch_priority_from_remark and ai.get("start_date"):
        st_date = parse_optional_date(ai.get("start_date"))

    st_time = None
    if allow_ai_dispatch_priority_from_remark and ai.get("start_time"):
        st_time = parse_time_str(str(ai.get("start_time")), None)

    speed_ov = first_float_pos_cell_or_ai(PLAN_COL_SPEED_OVERRIDE, "speed_override")

    return req_op, speed_ov, te, pri, st_date, st_time, ai


def _plan_row_cell_nonempty(row, col_name):
    v = row.get(col_name)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none"):
        return False
    return True


def _ai_float_for_conflict(ai, key):
    if not ai or ai.get(key) is None:
        return None
    try:
        f = float(ai.get(key))
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def detect_planning_remark_ai_conflicts(row, ai_for_tid):
    """
    特別指定_備考に依る AI 解析結果と」明示セルの両方に値はあり食い靕ご列を返す。
    備考・AIいうれか欠ける場合は空集合。
    """
    remark = str(row.get(PLAN_COL_SPECIAL_REMARK, "") or "").strip()
    if not remark or remark.lower() in ("nan", "none"):
        return set()
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}
    if not ai:
        return set()
    out = set()

    if _plan_row_cell_nonempty(row, PLAN_COL_SPEED_OVERRIDE):
        cv = parse_float_safe(row.get(PLAN_COL_SPEED_OVERRIDE), None)
        if cv is not None and cv > 0:
            av = _ai_float_for_conflict(ai, "speed_override")
            if av is not None and abs(cv - av) > 1e-5:
                out.add(PLAN_COL_SPEED_OVERRIDE)

    if _plan_row_cell_nonempty(row, PLAN_COL_PREFERRED_OP):
        cv = _normalize_person_name_for_match(row.get(PLAN_COL_PREFERRED_OP))
        av = _normalize_person_name_for_match(ai.get("preferred_operator"))
        if cv and av and cv != av:
            out.add(PLAN_COL_PREFERRED_OP)

    if out:
        out.add(PLAN_COL_SPECIAL_REMARK)
    return out


def collect_planning_conflicts_by_excel_row(tasks_df, ai_by_tid):
    """Excel 行番坷(1始まり・ヘッダー=1行目) -> 矛盾はあった列名の集合"""
    res = {}
    for i, (_, row) in enumerate(tasks_df.iterrows()):
        if _plan_row_exclude_from_assignment(row):
            continue
        ai_one = _ai_task_special_entry_for_row(ai_by_tid, row)
        cset = detect_planning_remark_ai_conflicts(row, ai_one)
        if cset:
            res[i + 2] = cset
    return res


def _plan_sheet_apply_conflict_styles_to_ws(ws, num_data_rows: int, conflicts_by_row) -> None:
    """既に開いている配台計画シートへ」矛盾列の着色（薄黄リセット→赤）を適用する。保存は呼び出し坴。"""
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(1, col_idx).value
        if v is not None:
            header_map[str(v).strip()] = col_idx

    last_row = max(2, 1 + int(num_data_rows))
    clear_fill = PatternFill(fill_type=None)
    yellow_input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    conflict_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    conflict_font = Font(color="FFFFFF", bold=True)

    for r in range(2, last_row + 1):
        for name in PLAN_CONFLICT_STYLABLE_COLS:
            ci = header_map.get(name)
            if not ci:
                continue
            cell = ws.cell(row=r, column=ci)
            if name == PLAN_COL_AI_PARSE:
                cell.fill = clear_fill
            else:
                cell.fill = yellow_input_fill
            # フォントは上書きしない（ブック既定・ユーザー設定を維挝）

    for r, colnames in conflicts_by_row.items():
        if r < 2:
            continue
        for name in colnames:
            ci = header_map.get(name)
            if not ci:
                continue
            cell = ws.cell(row=r, column=ci)
            cell.fill = conflict_fill
            cell.font = conflict_font


def write_plan_sheet_global_parse_and_conflict_styles_one_io(
    wb_path: str,
    sheet_name: str,
    global_priority_override: dict,
    *,
    when_str: str,
    num_data_rows: int,
    conflicts_by_row,
    log_prefix: str = "段階2",
) -> bool:
    """
    段階2坑け: グローバルコメント解析ブロック（AX:AY）と矛盾ハイライトを **1回の load/save** で反映れる。
    従来は別関数でブックを2回開いでいたため、.xlsm は大しい環境で坝数秒短縮の短縮になる。
    """
    if not wb_path or not os.path.isfile(wb_path):
        return False
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: ブックに「%s」があるため、openpyxl でグローバル解析・矛盾着色をスキップしました。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(
            wb_path, keep_vba=keep_vba, read_only=False, data_only=False
        )
    except Exception as ex:
        logging.info(
            "%s: 配台シート一括書込のため、ブックを開きません: %s",
            log_prefix,
            ex,
        )
        return False
    try:
        if sheet_name not in wb.sheetnames:
            logging.info(
                "%s: シート '%s' はないため、グローバル解析・矛盾着色をスキップ。",
                log_prefix,
                sheet_name,
            )
            return False
        ws = wb[sheet_name]
        _plan_sheet_write_global_parse_block_to_ws(ws, global_priority_override or {}, when_str)
        _plan_sheet_apply_conflict_styles_to_ws(ws, num_data_rows, conflicts_by_row or {})
        lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
        vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
        try:
            wb.save(wb_path)
        except OSError as e:
            write_planning_conflict_highlight_sidecar(
                sheet_name, num_data_rows, conflicts_by_row or {}
            )
            logging.warning(
                "%s: 配台シートへの一括保存に失敗（Excel で開いたまま等）。"
                " 矛盾ハイライトは '%s' に書き出しました。グローバル解析は未保存の可能性はありした。 (%s)",
                log_prefix,
                _planning_conflict_sidecar_path(),
                e,
            )
            return False
        _remove_planning_conflict_sidecar_safe()
        _n_conf = len(conflicts_by_row) if conflicts_by_row else 0
        if _n_conf:
            logging.info(
                "%s: 「%s」%s:%s 列にグローバル解析を保存し、"
                "特別指定_備考と列の矛盾 %s 行を坌も保存でハイライトしました。",
                log_prefix,
                sheet_name,
                get_column_letter(lc),
                get_column_letter(vc),
                _n_conf,
            )
        else:
            logging.info(
                "%s: 「%s」%s:%s 列にグローバル解析を保存しました（矛盾行なし）。",
                log_prefix,
                sheet_name,
                get_column_letter(lc),
                get_column_letter(vc),
            )
        return True
    except OSError as ex:
        logging.warning(
            "%s: 配台シート一括保存で OSError: %s",
            log_prefix,
            ex,
        )
        return False
    except Exception as ex:
        logging.warning(
            "%s: 配台シートへのグローバル解析＋矛盾着色（一括）で例外: %s",
            log_prefix,
            ex,
        )
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def apply_planning_sheet_conflict_styles(wb_path, sheet_name, num_data_rows, conflicts_by_row):
    """
    配台計画_タスク入力シートのデータ行を」矛盾列のみ赤地・白太字にれる。
    事剝パスでは上書き入力列を段階1とともに薄黄色に戻し、フォントは変更しない（体裝維挝）。
    AI解析列は着色しない（段階1の仕様に合わせる）。
    .xlsm は keep_vba=True で保存れる。
    """
    if not wb_path or not os.path.exists(wb_path):
        return
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "矛盾書式: ブックに「%s」があるため、openpyxl でのハイライトをスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = load_workbook(wb_path, keep_vba=keep_vba)
    try:
        if sheet_name not in wb.sheetnames:
            logging.warning(f"矛盾書式: シート '{sheet_name}' は見つかりません。")
            return
        ws = wb[sheet_name]
        _plan_sheet_apply_conflict_styles_to_ws(ws, num_data_rows, conflicts_by_row)

        try:
            wb.save(wb_path)
        except OSError as e:
            write_planning_conflict_highlight_sidecar(sheet_name, num_data_rows, conflicts_by_row)
            logging.warning(
                "配台シートへの矛盾ハイライトをファイル保存でしませんでした（Excel でブックを開いたまま等）。"
                " '%s' に指示を書き出しました。マクロはシート上に直接適用しした。 (%s)",
                _planning_conflict_sidecar_path(),
                e,
            )
        else:
            _remove_planning_conflict_sidecar_safe()
            if conflicts_by_row:
                logging.info(
                    f"特別指定_備考と列の矛盾: {len(conflicts_by_row)} 行を '{sheet_name}' でハイライトしました。"
                )
    finally:
        wb.close()


def _ai_planning_target_due_date(ai_dict):
    """AI JSON の完了・出蝷目標日から」配台の目標日1つを決ゝる（複数あれみ最も早い日＝厳しい方）。"""
    if not isinstance(ai_dict, dict):
        return None
    dates = []
    for k in ("target_completion_date", "ship_by_date", "latest_ship_date", "due_date"):
        d = parse_optional_date(ai_dict.get(k))
        if d is not None:
            dates.append(d)
    if not dates:
        return None
    return min(dates)


def _debug_agent_ndjson_55ab3a(
    hypothesis_id: str, location: str, message: str, data: dict | None = None
) -> None:
    # #region agent log
    try:
        import json as _json
        import time as _time
        from pathlib import Path as _Path

        _cur = _Path(__file__).resolve().parent
        _root = None
        for _ in range(12):
            if (_cur / ".git").is_dir():
                _root = _cur
                break
            if _cur.parent == _cur:
                break
            _cur = _cur.parent
        if _root is None:
            _root = _Path(__file__).resolve().parents[5]
        _logf = _root / "debug-55ab3a.log"
        _payload = {
            "sessionId": "55ab3a",
            "runId": "pre-fix",
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data or {},
            "timestamp": int(_time.time() * 1000),
        }
        with open(_logf, "a", encoding="utf-8") as _af:
            _af.write(_json.dumps(_payload, ensure_ascii=False) + "\n")
    except Exception:
        pass
    # #endregion


def _special_remark_implies_due_related_dispatch_priority(remark_raw: str) -> bool:
    """
    特別指定_備考に」紝期・期陝・最優先など「配台試行を剝に出す」愝図の文言はあるとし True。
    備考は記入されでいるの値では True にしない（AI 由来の目標日・開始日・優先度は使えない）。
    """
    if not remark_raw:
        return False
    s = str(remark_raw).strip()
    if not s or s.lower() in ("nan", "none"):
        return False
    n = unicodedata.normalize("NFKC", s)
    n_lower = n.casefold()
    # キーワードはユーザー入力（UTF-8 正しい表記）と一致させる。
    needles = (
        "紹期",
        "指定紹期",
        "回答紹期",
        "計画基準",
        "期日",
        "締切",
        "締め切り",
        "期限",
        "最優先",
        "至急",
        "急ぎ",
        "直ちに",
        "早急",
        "出荷",
        "紹入",
        "必着",
        "deadline",
        "デッドライン",
        "前倒し",
        "早めに",
        "厳守",
        "までに",
        "間に合わせ",
        "間に合い",
        "遅れない",
        "繧り上げ",
        "遅延",
        "優先配台",
        "先に配台",
        "完了予定",
        "本紹期",
        "回答期限",
    )
    return any(w.casefold() in n_lower for w in needles)


def _ai_task_special_entry_has_dispatch_priority_signals(ai_for_row) -> bool:
    """
    備考テキストのキーワード検出に漏れても、AI が既に priority / 日付 / 人数 等を返しているときは
    _merge_task_row_with_ai の allow を立てる（ログ H1: 備考に「納期」が無いが AI 非空、の救済）。
    preferred_operator のみのときは False（従来どおりセル側マージで足りる）。
    """
    if not isinstance(ai_for_row, dict) or not ai_for_row:
        return False
    meta = frozenset(
        {
            "process_name",
            "machine_name",
            "restrict_to_process_name",
            "restrict_to_machine_name",
            "preferred_operator",
        }
    )
    for k, v in ai_for_row.items():
        if k in meta or v is None:
            continue
        if k in (
            "ship_by_date",
            "target_completion_date",
            "latest_ship_date",
            "due_date",
            "start_date",
        ):
            if parse_optional_date(v) is not None:
                return True
        elif k == "start_time":
            if parse_time_str(str(v), None) is not None:
                return True
        elif k == "priority":
            if parse_optional_int(v) is not None:
                return True
        elif k == "required_op":
            try:
                if int(v) >= 1:
                    return True
            except (TypeError, ValueError):
                pass
        elif k == "task_efficiency":
            try:
                f = float(v)
                if f > 0 and f <= 1.0 + 1e-9:
                    return True
            except (TypeError, ValueError):
                pass
        elif k == "speed_override":
            try:
                if float(v) > 0:
                    return True
            except (TypeError, ValueError):
                pass
    return False


def _task_id_same_machine_due_tiebreak_key(task_id) -> tuple:
    """
    紝期基準（回答→指定）・機械名は坌も帯での試行順。
    Y3-24 は末尾の数値。Y4-1-1 のよごにポイフンは2つ以上あるとしは「最初の - の直後」の数値部を採用。
    """
    s = str(task_id or "").strip()
    if not s:
        return (2, 10**9, "")
    parts = s.split("-", 1)
    if len(parts) < 2:
        return (2, 10**9, s)
    rest = parts[1]
    if "-" in rest:
        first_seg = rest.split("-", 1)[0]
        try:
            return (0, int(first_seg), s)
        except ValueError:
            return (1, 10**9, s)
    tail = rest.strip()
    try:
        return (0, int(tail), s)
    except ValueError:
        return (1, 10**9, s)


def _optional_float_unprocessed_column(val):
    """
    配台計画シートの「未加工」セルを float 化する。
    空・無効なら None（結果_タスク一覧の残加工量は従来どおり m 換算にフォールバック）。
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("nan", "none", "-", "—", "―"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# 配台用タスクキュー
#   配台計画 DataFrame 1行 → 割付アルゴリズム用 dict への変杛（優先度・紝期・AI 上書きを集約）
# ---------------------------------------------------------------------------
def build_task_queue_from_planning_df(
    tasks_df,
    run_date,
    req_map,
    ai_by_tid=None,
    global_priority_override=None,
    equipment_list=None,
):
    """
    ``generate_plan`` 内で呼みれる。完了済み・配台試行行を除し」残りを task_queue に穝む。
    ai_by_tid は None のときの値内部で analyze_task_special_remarks を実行れる。
    """
    if ai_by_tid is None:
        ai_by_tid = analyze_task_special_remarks(tasks_df, reference_year=run_date.year)
    task_queue = []
    n_exclude_plan = 0
    seq_by_tid = _collect_process_content_order_by_task_id(tasks_df)
    same_tid_line_seq = defaultdict(int)
    # 依頼NO直列配台の順庝用: iterrows の読み込み順（0 始まり）。task_queue.sort 後も試行。
    planning_sheet_row_seq = 0
    _has_unprocessed_col = TASK_COL_UNPROCESSED in tasks_df.columns

    for planning_df_iloc, (_, row) in enumerate(tasks_df.iterrows()):
        if row_has_completion_keyword(row):
            continue
        if _plan_row_exclude_as_completed_mikan_unprocessed_zero_actual_done_rule(row):
            continue
        task_id = planning_task_id_str_from_plan_row(row)
        _plan_excl = _plan_row_exclude_from_assignment(row)
        # #region agent log
        if task_id.casefold() == "w4-5":
            _debug_agent_ndjson_55ab3a(
                "H2",
                "_core.py:build_task_queue_from_planning_df:exclude",
                "W4-5 配台試行除外フラグ",
                {"excluded": bool(_plan_excl)},
            )
        # #endregion
        if _plan_excl:
            n_exclude_plan += 1
            continue

        machine = str(row.get(TASK_COL_MACHINE, "")).strip()
        machine_name = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        qty, done_qty, qty_total, from_unprocessed_qty = _plan_row_dispatch_qty_metrics(
            row
        )
        speed_raw = row.get(TASK_COL_SPEED, 1)
        product_name = row.get(TASK_COL_PRODUCT, None)
        answer_due = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_ANSWER_DUE))
        specified_due = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_SPECIFIED_DUE))
        specified_due_ov = None
        # 紝期基準: ①回答納期（空でなけれみ）②列「指定納期」（列「指定納期_上書き」は廃止済み）
        specified_basis = specified_due
        due_basis = None
        due_source = "none"
        due_source_rank = 9
        raw_input_sheet = parse_optional_date(
            _planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE)
        )
        raw_input_date_ov = parse_optional_date(
            _planning_df_cell_scalar(row, PLAN_COL_RAW_INPUT_DATE_OVERRIDE)
        )
        raw_input_date = (
            raw_input_date_ov if raw_input_date_ov is not None else raw_input_sheet
        )
        if (
            raw_input_date_ov is not None
            and raw_input_sheet is not None
            and raw_input_date_ov != raw_input_sheet
        ):
            logging.info(
                "原反投入日_上書きを採用: 依頼NO=%s シート原反投入日=%s 上書き=%s",
                task_id,
                raw_input_sheet,
                raw_input_date_ov,
            )

        qty = max(0.0, qty_total - done_qty)
        speed = parse_float_safe(speed_raw, 1.0)
        if speed <= 0:
            speed = 1.0

        if qty <= 0 or not machine or not task_id:
            continue

        _line_seq = same_tid_line_seq[task_id]
        same_tid_line_seq[task_id] += 1

        remark_raw = str(row.get(PLAN_COL_SPECIAL_REMARK, "") or "").strip()
        has_special_remark = bool(remark_raw) and remark_raw.lower() not in ("nan", "none")
        remark_implies_due_dispatch_priority = (
            _special_remark_implies_due_related_dispatch_priority(remark_raw)
        )
        in_progress = done_qty > 0.0

        ai_one = _ai_task_special_entry_for_row(ai_by_tid, row)
        allow_from_ai_dispatch_signals = (
            has_special_remark
            and _ai_task_special_entry_has_dispatch_priority_signals(ai_one)
        )
        allow_ai_dispatch_priority = (
            remark_implies_due_dispatch_priority or allow_from_ai_dispatch_signals
        )
        # #region agent log
        if task_id.casefold() == "w4-5" and has_special_remark:
            _nou = "\u7d39\u671f" in remark_raw
            _spec_n = "\u6307\u5b9a\u7d39\u671f" in remark_raw
            _debug_agent_ndjson_55ab3a(
                "H1",
                "_core.py:build_task_queue_from_planning_df:remark_implies",
                "W4-5 備考と allow_ai 判定",
                {
                    "remark_implies_due_dispatch_priority": bool(
                        remark_implies_due_dispatch_priority
                    ),
                    "allow_from_ai_dispatch_signals": bool(
                        allow_from_ai_dispatch_signals
                    ),
                    "allow_ai_dispatch_priority": bool(allow_ai_dispatch_priority),
                    "remark_len": len(remark_raw),
                    "remark_has_correct_nouki": _nou,
                    "remark_has_correct_shitei_nouki": _spec_n,
                },
            )
        # #endregion
        # #region agent log
        if task_id.casefold() == "w4-5":
            _debug_agent_ndjson_55ab3a(
                "H3",
                "_core.py:build_task_queue_from_planning_df:ai_entry",
                "W4-5 AI 行エントリ",
                {
                    "ai_type": type(ai_one).__name__,
                    "ai_nonempty": bool(ai_one),
                },
            )
        # #endregion
        req_op, speed_ov, task_eff_factor, priority, start_date_ov, start_time_ov, ai_used = _merge_task_row_with_ai(
            row,
            ai_one,
            allow_ai_dispatch_priority_from_remark=allow_ai_dispatch_priority,
        )
        preferred_operator_raw = _merge_preferred_operator_cell_and_ai(row, ai_one)
        gpo = global_priority_override or {}
        gop_name = _global_override_preferred_operator_for_task(
            gpo.get("task_preferred_operators"), task_id
        )
        if gop_name is not None:
            preferred_operator_raw = gop_name
            logging.info(
                "メイン再優先特記: 依頼NO=%s の担当OPをグローバル指定で上書き %r（セル・特別指定備考AIより優先）",
                task_id,
                gop_name,
            )

        if answer_due is not None:
            due_basis = answer_due
            due_source = "answer_due"
            due_source_rank = 0
        elif specified_basis is not None:
            due_basis = specified_basis
            due_source = "specified_due"
            due_source_rank = 1
        has_done_deadline_override = False

        if speed_ov is not None:
            speed = speed_ov
        if speed <= 0:
            speed = 1.0

        gsm = _global_speed_multiplier_for_row(
            machine, machine_name, gpo.get("global_speed_rules") or []
        )
        if abs(gsm - 1.0) > 1e-12:
            speed_before_g = speed
            speed = speed * gsm
            if speed <= 0:
                speed = 1.0
            logging.info(
                "メイングローバル: 依頼NO=%s 工程=%r 機械名=%r に speed_multiplier 累穝=%s を適用（速度 %s → %s）",
                task_id,
                machine,
                machine_name,
                gsm,
                speed_before_g,
                speed,
            )

        unit = parse_float_safe(
            _planning_df_cell_scalar(row, PLAN_COL_ROLL_UNIT_LENGTH), 0.0
        )
        if unit <= 0:
            unit = infer_unit_m_from_product_name(
                product_name, fallback_unit=qty_total if qty_total > 0 else qty
            )
        try:
            unit = float(unit)
        except Exception:
            unit = qty
        if unit <= 0:
            unit = qty

        # 換算数量・ロール単位長さの補正（推定・100m 下限・換算<ロール時の引き上げ）は段階1のみ。段階2はシート値を採用し、空・0 のときだけ推定フォールバックする。

        # 納期は優先順位・緊急度には使うが、開始日の下限には使わない（余力があれば前倒し開始するため）。
        if due_basis is None:
            due_urgent = False
        else:
            due_urgent = due_basis <= run_date

        # 開始日ルール:
        # 1) 原反投入日があるときは「原反投入日 13:00 以降」を開始可能日時の下限にする。
        #    （日付下限: max(run_date, raw_input_date)」同日時間下限: 13:00）
        # 2) 特別指定（セル/AI）の開始日があっても原反投入日より前倒しにはしない（date 下限を維持）
        # 3) 原反投入日が無いときは run_date
        if raw_input_date:
            effective_start_date = max(run_date, raw_input_date)
        else:
            effective_start_date = run_date
        if start_date_ov is not None:
            effective_start_date = (
                max(start_date_ov, raw_input_date)
                if raw_input_date
                else start_date_ov
            )
            if raw_input_date and start_date_ov < raw_input_date:
                logging.info(
                    "開始日上書きは原反投入日より前倒し不可: 依頼NO=%s 指定開始日=%s 原反投入日=%s 採用開始日=%s",
                    task_id,
                    start_date_ov,
                    raw_input_date,
                    effective_start_date,
                )

        same_day_raw_start_limit = (
            time(13, 0)
            if (raw_input_date and effective_start_date == raw_input_date)
            else None
        )

        calc_time_val = qty * speed
        ai_note = ""
        if ai_used:
            try:
                ai_note = json.dumps(ai_used, ensure_ascii=False)[:500]
            except Exception:
                ai_note = str(ai_used)[:500]

        _order_list = seq_by_tid.get(task_id) or []
        _p_rank = _process_sequence_rank_for_machine(machine, _order_list)
        if from_unprocessed_qty and unit > 0:
            # ③ 未加工(m) ÷ ロール単位長さ を切り上げ整数ロール
            _init_rem = float(math.ceil(max(0.0, qty) / float(unit)))
        else:
            _init_rem = float(qty / unit if unit else 0.0)
        _process_content_mismatch = bool(_order_list) and not _process_name_matches_kakou_content_tokens(
            machine, _order_list
        )

        _dto_from_sheet = None
        if RESULT_TASK_COL_DISPATCH_TRIAL_ORDER in tasks_df.columns:
            _dto_from_sheet = parse_optional_int(
                _planning_df_cell_scalar(row, RESULT_TASK_COL_DISPATCH_TRIAL_ORDER)
            )

        if from_unprocessed_qty:
            _unp_base = max(0.0, qty)
        elif _has_unprocessed_col:
            _unp_base = _optional_float_unprocessed_column(
                _planning_df_cell_scalar(row, TASK_COL_UNPROCESSED)
            )
        else:
            _unp_base = None

        task_queue.append(
            {
                "task_id": task_id,
                "machine": machine,
                "machine_name": machine_name,
                "equipment_line_key": _resolve_equipment_line_key_for_task(
                    {"machine": machine, "machine_name": machine_name},
                    equipment_list,
                ),
                "start_date_req": effective_start_date,
                "answer_due_date": answer_due,
                "specified_due_date": specified_due,
                "specified_due_override": specified_due_ov,
                "due_basis_date": due_basis,
                # 紝期後ゝ倒し再試行で due_basis_date を内部 +1 しても」結果_タスク一覧の当列（列名は互換で「計画基準納期」）はこの値のまま
                "due_basis_date_result_sheet": due_basis,
                "due_source": due_source,
                "due_source_rank": due_source_rank,
                "due_urgent": due_urgent,
                "raw_input_date": raw_input_date,
                "same_day_raw_start_limit": same_day_raw_start_limit,
                "total_qty_m": int(qty_total),
                "unit_m": int(unit),
                "remaining_units": _init_rem,
                "base_time_per_unit": (qty / speed) / (qty / unit)
                if unit and speed and qty
                else 0,
                "assigned_history": [],
                "calc_time_value": calc_time_val,
                # シートの加工速度・上書き・global_speed_rules 適用後の m/分（配台シミュレーションと同一）
                TASK_COL_SPEED: float(speed),
                "required_op": req_op,
                "task_eff_factor": task_eff_factor,
                "priority": priority,
                "earliest_start_time": start_time_ov,
                "preferred_operator_raw": preferred_operator_raw,
                "task_special_ai_note": ai_note,
                "in_progress": in_progress,
                "has_special_remark": has_special_remark,
                "has_done_deadline_override": has_done_deadline_override,
                "done_qty_reported": done_qty,
                "process_sequence_rank": _p_rank,
                "same_request_line_seq": _line_seq,
                "initial_remaining_units": _init_rem,
                "roll_pipeline_ec": _row_matches_roll_pipeline_ec(machine, machine_name),
                "roll_pipeline_inspection": _row_matches_roll_pipeline_inspection(
                    machine, machine_name
                ),
                "roll_pipeline_rewind": _row_matches_roll_pipeline_rewind(
                    machine, machine_name
                ),
                "process_content_mismatch": _process_content_mismatch,
                "planning_sheet_row_seq": planning_sheet_row_seq,
                "planning_df_iloc": planning_df_iloc,
                "dispatch_trial_order_from_sheet": _dto_from_sheet,
                "unprocessed_baseline_m": _unp_base,
            }
        )
        planning_sheet_row_seq += 1

    logging.info(
        "task_queue 構築完了: total=%s（配台試行によりスキップ %s 行）",
        len(task_queue),
        n_exclude_plan,
    )
    return task_queue


def _task_id_priority_key(task_id):
    """
    依頼NOの同一条件タイブレーク用キー。
    例: Y3-24, Y3-34 のよごな場合はポイフン後坊の数値は尝さい方を優先。
    """
    s = str(task_id or "").strip()
    if not s:
        return ("", 10**9, "")
    parts = s.rsplit("-", 1)
    if len(parts) == 2:
        head = parts[0].strip()
        tail = parts[1].strip()
        if re.match(r"^\d+$", tail):
            return (head, int(tail), s)
    return (s, 10**9, s)


def _serial_dispatch_order_task_ids(task_queue) -> list:
    """
    依頼NO直列配台の処理順。坄依頼NOについで **配台試行順番の最尝値** は尝さい依頼を先に完走させる
    （同一依頼内の複数行は最尝幅の試行順で代表）。タイブレークは計画シート上の先行行
    （planning_sheet_row_seq）と依頼NOキー。
    """
    min_dto_by_tid: dict = {}
    first_seq_by_tid: dict = {}
    for t in task_queue:
        tid = str(t.get("task_id", "") or "").strip()
        if not tid:
            continue
        try:
            dto = int(t.get("dispatch_trial_order") or 10**9)
        except (TypeError, ValueError):
            dto = 10**9
        prev_d = min_dto_by_tid.get(tid)
        if prev_d is None or dto < prev_d:
            min_dto_by_tid[tid] = dto
        seq = t.get("planning_sheet_row_seq")
        seq = int(seq) if seq is not None else 10**9
        prev = first_seq_by_tid.get(tid)
        if prev is None or seq < prev:
            first_seq_by_tid[tid] = seq
    return sorted(
        min_dto_by_tid.keys(),
        key=lambda tid: (
            min_dto_by_tid[tid],
            first_seq_by_tid.get(tid, 10**9),
            _task_id_priority_key(tid),
        ),
    )


def _excel_scalar_to_plan_string_cell(v):
    """
    既存シート（read_excel）由来のスカラーを」配台計画 DataFrame の文字列列（StringDtype）へ
    代入でしる str に正規化する。Excel は数値として保挝した優先度 1 → \"1\" など。
    """
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, bool):
        return str(v).lower()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and math.isfinite(v) and float(int(v)) == v:
            return str(int(v))
        if isinstance(v, float) and math.isfinite(v):
            s = str(v)
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            return s
        return str(int(v))
    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return ""
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.strftime("%Y/%m/%d")
        return v.strftime("%Y/%m/%d %H:%M")
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.date().strftime("%Y/%m/%d")
        return v.strftime("%Y/%m/%d %H:%M")
    if isinstance(v, date):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _merge_plan_sheet_user_overrides(out_df):
    """
    ブック内の「配台計画_タスク入力」にユーザーは入力した上書き列を」
    段階1の抽出結果へ (依頼NO, 工程名) 短縮で引き継ぎ。
    空のセルはマージしない（新規抽出坴の空のまま）。
    """
    if out_df is None or out_df.empty:
        return out_df
    if not TASKS_INPUT_WORKBOOK or not os.path.exists(TASKS_INPUT_WORKBOOK):
        return out_df
    try:
        df_old = pd.read_excel(TASKS_INPUT_WORKBOOK, sheet_name=PLAN_INPUT_SHEET_NAME)
    except Exception as e:
        logging.info("段階1: 既存の配台シートを読めないため上書き継承をスキップ (%s)", e)
        return out_df
    df_old.columns = df_old.columns.str.strip()
    df_old = _align_dataframe_headers_to_canonical(
        df_old,
        plan_input_sheet_column_order(),
    )
    if TASK_COL_TASK_ID not in df_old.columns or TASK_COL_MACHINE not in df_old.columns:
        return out_df

    lookup = {}
    for _, r in df_old.iterrows():
        tid = planning_task_id_str_from_plan_row(r)
        mach = str(r.get(TASK_COL_MACHINE, "") or "").strip()
        if not tid or not mach:
            continue
        key = (tid, mach)
        bucket = lookup.setdefault(key, {})
        for c in (*PLAN_STAGE1_MERGE_COLUMNS, *PLAN_STAGE1_MERGE_EXTRA_COLUMNS):
            if c not in df_old.columns or c not in out_df.columns:
                continue
            v = r.get(c)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            if isinstance(v, str):
                s = v.strip()
                if not s or s.lower() in ("nan", "none"):
                    continue
            bucket[c] = v

    if not lookup:
        return out_df

    merged_rows = 0
    for i, row in out_df.iterrows():
        tid = planning_task_id_str_from_plan_row(row)
        mach = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        bucket = lookup.get((tid, mach))
        if not bucket:
            continue
        merged_rows += 1
        for c, v in bucket.items():
            if c == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
                v = _coerce_plan_exclude_column_value_for_storage(v)
            elif c in out_df.columns and pd.api.types.is_string_dtype(out_df[c].dtype):
                v = _excel_scalar_to_plan_string_cell(v)
            out_df.at[i, c] = v

    if merged_rows:
        logging.info(
            "段階1: 既存シートから上書き列を %s 行へ引し継ねました（キー: 依頼NO+工程名）。",
            merged_rows,
        )
    return out_df


# ---------------------------------------------------------------------------
# 配台不要（2系統）
#   (A) DataFrame 上のルール … 同一依頼NO×同一機械で「分割」行に yes（手入力は上書きしない）。
#       段階2読込後も ``_apply_auto_exclude_bunkatsu_duplicate_machine`` で適用。
#   (B) マクロブック「設定_配台不要工程」… 工程+機械ごとの C/D/E 列、Gemini で D→E、
#       保存ロック時は xlwings で A:E 同期→Save のフォールバックあり。
#       ``apply_exclude_rules_config_to_plan_df`` による計画 DataFrame への反映は **段階1のみ**。
#       工程名が「分割」の行については、(A) と同じく **同一依頼NO内に同一機械名が複数行あるときだけ**
#       C 列／E 列 JSON による配台不要=yes を適用する（EC と分割で機械が異なる依頼では設定行が残っていても配台可）。
#       段階2は配台計画シートの「配台不要」列（段階1出力・手編集の結果）をそのまま使う。
# ---------------------------------------------------------------------------

def _auto_exclude_cell_empty_for_autofill(v) -> bool:
    """配台試行セルは未入力のときの値自動で yes を書き込む。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return True
    if isinstance(v, str):
        s = str(v).strip()
        return not s or s.lower() in ("nan", "none")
    return False


def _normalize_task_id_for_dup_grouping(raw) -> str:
    """同一依頼NOのグルーピング用（表記ゆれ・英字の大尝を寄せる）。"""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    if isinstance(raw, float) and raw == int(raw):
        s = str(int(raw))
    else:
        s = unicodedata.normalize("NFKC", str(raw).strip())
    s = s.strip()
    if not s or s.lower() == "nan":
        return ""
    return s.upper()


def _process_name_is_bunkatsu_for_auto_exclude(raw) -> bool:
    """工程名は「分割」（空白除去・NFKC 後）。"""
    t = unicodedata.normalize("NFKC", str(raw or "").strip())
    t = re.sub(r"[\s　]+", "", t)
    return t == "分割"


def _same_tid_nonempty_machine_dup_ge2(
    df: pd.DataFrame, idx_list: list
) -> tuple[bool, dict[str, int]]:
    """
    ``_apply_auto_exclude_bunkatsu_duplicate_machine`` と同一の重複判定。
    idx_list 内で、正規化後の非空「機械名」が同一の行が2件以上あるとき True。
    """
    if len(idx_list) < 2:
        return False, {}
    counts: dict[str, int] = defaultdict(int)
    for i in idx_list:
        mn_key = _normalize_equipment_match_key(df.at[i, TASK_COL_MACHINE_NAME])
        if not mn_key:
            continue
        counts[mn_key] += 1
    dup_ge2 = any(c >= 2 for c in counts.values())
    return dup_ge2, dict(counts)


def _apply_auto_exclude_bunkatsu_duplicate_machine(
    df: pd.DataFrame, log_prefix: str = "段階1"
) -> pd.DataFrame:
    """
    同一依頼NOは2行以上あり」かつ空でない同一機械名は2行以上あるグループでは」
    工程名は「分割」の行の「配台試行」に yes を入れる（セルは空のときのみ）。
    機械名は _normalize_equipment_match_key で重複判定。
    """
    if df is None or df.empty:
        return df
    need_cols = (TASK_COL_TASK_ID, TASK_COL_MACHINE, TASK_COL_MACHINE_NAME)
    for c in need_cols:
        if c not in df.columns:
            return df
    if PLAN_COL_EXCLUDE_FROM_ASSIGNMENT not in df.columns:
        df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = ""
    # read_excel 等で StringDtype になると数値・真偽の .at 代入で TypeError になるため、 object に寄せる
    df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT].astype(object)

    by_tid = defaultdict(list)
    for i in df.index:
        tid = _normalize_task_id_for_dup_grouping(df.at[i, TASK_COL_TASK_ID])
        if not tid:
            continue
        by_tid[tid].append(i)

    n_set = 0
    for _tid_key, idx_list in by_tid.items():
        dup_ge2, _counts = _same_tid_nonempty_machine_dup_ge2(df, idx_list)
        if not dup_ge2:
            continue
        for i in idx_list:
            if not _process_name_is_bunkatsu_for_auto_exclude(df.at[i, TASK_COL_MACHINE]):
                continue
            if not _auto_exclude_cell_empty_for_autofill(
                df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT]
            ):
                continue
            # 列は StringDtype のとき int 代入で TypeError になるため、文字列にれる（_plan_row_exclude_from_assignment は yes を真とみなす）
            df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = "yes"
            n_set += 1

    if n_set:
        logging.info(
            "%s: 同一依頼NOかつ同一機械名は複数行あるグループで」工程名「分割」の行 %s 件に「配台試行」=yes を自動設定しました。",
            log_prefix,
            n_set,
        )
    return df


def _normalize_process_name_for_rule_match(raw) -> str:
    """工程名のルール照合（NFKC・空白除去）。"""
    t = unicodedata.normalize("NFKC", str(raw or "").strip())
    t = re.sub(r"[\s　]+", "", t)
    return t


def _exclude_rules_sheet_header_map(ws) -> dict:
    """1行目見出し → 列番坷(1始まり)。
    openpyxl は新規シート直後に max_column は 0 のままのことはあり」見出しは読ゝう保存剝に return してしまご。
    しのため、最低 A～E 列は必う走査れる。
    """
    h = {}
    last_col = max(5, int(ws.max_column or 0))
    for col in range(1, last_col + 1):
        v = ws.cell(1, col).value
        if v is not None:
            h[str(v).strip()] = col
    return h


def _ensure_exclude_rules_sheet_headers_and_columns(ws, log_prefix: str) -> tuple[int, int, int, int, int]:
    """
    1行目に標準見出し（工程名・機械名・配台不要・配台不要ロジック・ロジック式）はあることを保証れる。
    手動で空シートの値追加した場合は A1:E1 は空のため、ここで書き込んで列番坷を返す。
    """
    headers = (
        EXCLUDE_RULE_COL_PROCESS,
        EXCLUDE_RULE_COL_MACHINE,
        EXCLUDE_RULE_COL_FLAG,
        EXCLUDE_RULE_COL_LOGIC_JA,
        EXCLUDE_RULE_COL_LOGIC_JSON,
    )
    hm = _exclude_rules_sheet_header_map(ws)
    if all(hm.get(x) for x in headers):
        return tuple(hm[x] for x in headers)
    for i, name in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=name)
    logging.info(
        "%s: 「%s」の見出しは無い＝列名は一致しないため、標準の1行目（A1:E1）を設定しました。",
        log_prefix,
        EXCLUDE_RULES_SHEET_NAME,
    )
    return (1, 2, 3, 4, 5)


def _compact_exclude_rules_data_rows(
    ws,
    c_proc: int,
    c_mach: int,
    c_flag: int,
    c_d: int,
    c_e: int,
    log_prefix: str,
) -> tuple[int, int]:
    """
    2 行目以降から「空行」を除いで上に詰ゝる（元の並よは維挝」ソートしない）。
    空行: 工程名は空」または A～E 相当の5セルはまとめて空白相当。
    Returns (残したデータ行数, 削除した行数).
    """
    max_r = int(ws.max_row or 1)
    if max_r < 2:
        return 0, 0

    old_body = max_r - 1
    cols = (c_proc, c_mach, c_flag, c_d, c_e)
    rows: list[tuple[str, str, object, object, object]] = []
    for r in range(2, max_r + 1):
        pv = ws.cell(row=r, column=c_proc).value
        mv = ws.cell(row=r, column=c_mach).value
        cv = ws.cell(row=r, column=c_flag).value
        dv = ws.cell(row=r, column=c_d).value
        ev = ws.cell(row=r, column=c_e).value
        all_blank = all(
            _cell_is_blank_for_rule(ws.cell(row=r, column=c).value) for c in cols
        )
        p = str(pv).strip() if pv is not None and not (isinstance(pv, float) and pd.isna(pv)) else ""
        m = str(mv).strip() if mv is not None and not (isinstance(mv, float) and pd.isna(mv)) else ""
        if all_blank or not p:
            continue
        rows.append((p, m, cv, dv, ev))

    n_skip = old_body - len(rows)

    if not rows:
        ws.delete_rows(2, old_body)
        if old_body > 0:
            logging.info(
                "%s: 「%s」は有効なデータ行はなかったため、データ行 %s 行を削除しました。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
                old_body,
            )
        return 0, n_skip

    ws.delete_rows(2, old_body)
    for i, (p, m, cv, dv, ev) in enumerate(rows, start=2):
        ws.cell(row=i, column=c_proc, value=p)
        ws.cell(row=i, column=c_mach, value=m)
        ws.cell(row=i, column=c_flag, value=cv)
        ws.cell(row=i, column=c_d, value=dv)
        ws.cell(row=i, column=c_e, value=ev)

    if n_skip:
        logging.info(
            "%s: 「%s」から空行を %s 件削除し、%s 行に詰ゝました（並よ順は維挝）。",
            log_prefix,
            EXCLUDE_RULES_SHEET_NAME,
            n_skip,
            len(rows),
        )
    return len(rows), n_skip


def _cell_is_blank_for_rule(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    s = str(v).strip()
    return not s or s.lower() in ("nan", "none", "null")


def _exclude_rule_c_column_is_yes(v) -> bool:
    """C列「配台不要」はオン（この工程+機械は常に配台対象外）。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return int(v) == 1
        except (TypeError, ValueError):
            pass
    s = unicodedata.normalize("NFKC", str(v).strip()).lower()
    return s in ("yes", "true", "1", "y", "はい", "○", "〇", "◝")


def _task_row_matches_exclude_rule_target(
    task_proc: str, task_mach: str, rule_proc: str, rule_mach: str
) -> bool:
    if _normalize_process_name_for_rule_match(task_proc) != _normalize_process_name_for_rule_match(
        rule_proc
    ):
        return False
    rm = str(rule_mach or "").strip()
    if not rm:
        # 機械名が空のルールは「当該工程の全機械」を意味するが、工程「分割」は同一依頼で別行の
        # スリット等と同じ機械名が重なるケースが多く、ワイルドカード一致だと実設備向け分割行まで
        # 一律配台不要になる。分割を除外する場合は設定シートで機械名を明示する。
        if _process_name_is_bunkatsu_for_auto_exclude(task_proc):
            return False
        return True
    return _normalize_equipment_match_key(task_mach) == _normalize_equipment_match_key(rm)


def _collect_process_machine_pairs_for_exclude_rules(df_src: pd.DataFrame) -> list[tuple[str, str]]:
    """加工計画DATA から」段階1とともに抽出条件で (工程名, 機械名) の一覧（重複除し・順庝維挝）。"""
    out: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for _, row in df_src.iterrows():
        if row_has_completion_keyword(row):
            continue
        if _plan_row_exclude_as_completed_mikan_unprocessed_zero_actual_done_rule(row):
            continue
        task_id = planning_task_id_str_from_scalar(row.get(TASK_COL_TASK_ID))
        machine = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        machine_name = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        qty, _done_qty, _qty_total, _from_unp = _plan_row_dispatch_qty_metrics(row)
        if qty <= 0 or not machine or not task_id:
            continue
        key = (
            _normalize_process_name_for_rule_match(machine),
            _normalize_equipment_match_key(machine_name),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append((machine, machine_name))
    return out


def _parse_exclude_rule_json_cell(raw) -> dict | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).strip()
    if not s:
        return None
    fence = re.search(
        r"```(?:json)?\s*(\{.*\})\s*```",
        s,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        s = fence.group(1).strip()
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, dict) else None


def _validate_exclude_rule_parsed_dict(o: object) -> dict | None:
    """Gemini＝E列から得た dict は配台試行ルールとして有効か。"""
    if not isinstance(o, dict):
        return None
    if int(o.get("version") or 0) != 1:
        return None
    mode = str(o.get("mode") or "").strip().lower()
    if mode not in ("always_exclude", "conditions"):
        return None
    return o


def _exclude_rule_de_cache_key(stripped_blob: str) -> str:
    """「配台不要ロジック」文言（正規化済み）に対れる ai_remarks_cache 用キー。"""
    h = hashlib.sha256(stripped_blob.encode("utf-8")).hexdigest()
    return f"{AI_CACHE_KEY_PREFIX_EXCLUDE_RULE_DE}:{h}"


def _cache_get_exclude_rule_de_parsed(cache_obj: dict, blob: str) -> dict | None:
    s = str(blob or "").strip()
    if not s:
        return None
    data = get_cached_ai_result(
        cache_obj, _exclude_rule_de_cache_key(s), content_key=s
    )
    if not isinstance(data, dict):
        return None
    return _validate_exclude_rule_parsed_dict(data)


def _cache_put_exclude_rule_de_parsed(
    cache_obj: dict, blob: str, parsed: dict | None
) -> None:
    if parsed is None:
        return
    s = str(blob or "").strip()
    if not s:
        return
    put_cached_ai_result(
        cache_obj, _exclude_rule_de_cache_key(s), parsed, content_key=s
    )


def _exclude_rule_logic_gemini_schema_instructions() -> str:
    allowed = ", ".join(sorted(EXCLUDE_RULE_ALLOWED_COLUMNS))
    return (
        "」スキーマ version は必う 1】\n"
        "1) 常に配台試行（説明は条件なしで外れ愝味）のとき:\n"
        '{"version":1,"mode":"always_exclude"}\n\n'
        "2) 列の条件で配台試行とれるとし:\n"
        '{"version":1,"mode":"conditions","require_all": true または false,"conditions":[ ... ]}\n\n'
        "conditions の坄覝素:\n"
        "- {\"column\":\"列名\",\"op\":\"empty\"} … セルは空\n"
        "- {\"column\":\"列名\",\"op\":\"not_empty\"}\n"
        "- {\"column\":\"列名\",\"op\":\"eq\",\"value\":\"文字列\"} / ne / contains / not_contains / regex（正覝表睾）\n"
        "- {\"column\":\"列名\",\"op\":\"gt\"|\"gte\"|\"lt\"|\"lte\",\"value\":数値} … 数値比較（列は数として解釈）\n\n"
        f"」使用可能な列名のみ】（これ以外は使えない）:\n{allowed}\n"
    )


def _parse_exclude_rule_json_array_response(text: str) -> list | None:
    """モデル応答から JSON 配列を取り出す（```json フェンス付し坯）。"""
    s = (text or "").strip()
    if not s:
        return None
    fence = re.search(
        r"```(?:json)?\s*(\[.*\])\s*```",
        s,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        s = fence.group(1).strip()
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, list) else None


def _row_scalar_for_exclude_rule(row, col_name: str):
    try:
        return row.get(col_name)
    except Exception:
        return None


def _exclude_rule_string_eq_allow_csv_tokens(val_s: str, pat_raw) -> bool:
    """配台不要条件の eq 用。完全一致に加え、セルが「A,B」「A、B」のように複数トークンを含むときはいずれかが pat と一致すれば真。"""
    pat = "" if pat_raw is None else str(pat_raw).strip()
    if val_s == pat:
        return True
    if "," in val_s or "\u3001" in val_s:
        parts = [p.strip() for p in re.split(r"[,、]", val_s) if p.strip()]
        return bool(pat) and pat in parts
    return False


def _evaluate_exclude_rule_one_condition(cond: dict, row) -> bool:
    if not isinstance(cond, dict):
        return False
    col = cond.get("column")
    if col not in EXCLUDE_RULE_ALLOWED_COLUMNS:
        logging.warning("配台試行ルール: 未対応の列名をスキップしました: %s", col)
        return False
    op = str(cond.get("op") or "").strip().lower()
    val = _row_scalar_for_exclude_rule(row, col)
    val_s = "" if val is None or (isinstance(val, float) and pd.isna(val)) else str(val).strip()
    val_s_lower = val_s.lower()

    if op == "empty":
        return val_s == ""
    if op == "not_empty":
        return val_s != ""

    if op in ("contains", "not_contains", "regex", "eq", "ne"):
        rhs = cond.get("value", "")
        pat = "" if rhs is None else str(rhs)
        if op == "contains":
            return pat in val_s
        if op == "not_contains":
            return pat not in val_s
        if op == "regex":
            try:
                return re.search(pat, val_s) is not None
            except re.error:
                return False
        if op == "eq":
            return _exclude_rule_string_eq_allow_csv_tokens(val_s, pat)
        if op == "ne":
            return not _exclude_rule_string_eq_allow_csv_tokens(val_s, pat)

    def _num(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    nv = _num(val)
    cv = _num(cond.get("value"))
    if nv is None or cv is None:
        return False
    if op == "gt":
        return nv > cv
    if op == "gte":
        return nv >= cv
    if op == "lt":
        return nv < cv
    if op == "lte":
        return nv <= cv
    return False


def evaluate_exclude_rule_json_for_row(rule: dict, row) -> bool:
    """
    E列の JSON（version=1）を評価し、当該タスク行を配台試行とれきしなら True。
    mode: always_exclude | conditions
    """
    if not isinstance(rule, dict) or int(rule.get("version") or 0) != 1:
        return False
    mode = str(rule.get("mode") or "").strip().lower()
    if mode == "always_exclude":
        return True
    if mode != "conditions":
        return False
    conds = rule.get("conditions")
    if not isinstance(conds, list) or not conds:
        return False
    require_all = bool(rule.get("require_all", True))
    checks = []
    for c in conds:
        if isinstance(c, dict) and c.get("column") in EXCLUDE_RULE_ALLOWED_COLUMNS:
            checks.append(_evaluate_exclude_rule_one_condition(c, row))
    if not checks:
        return False
    return all(checks) if require_all else any(checks)


def _ai_compile_exclude_rule_logic_to_json(natural_language: str) -> dict | None:
    """
    D列の自然言語を Gemini で JSON ルールに変杛。失敗時 None。
    json/ai_remarks_cache.json に TTL 付しでキャッシュ（同一文言なら API を呼みない）。
    """
    blob = str(natural_language or "").strip()
    if not blob:
        return None
    ai_cache = load_ai_cache()
    hit = _cache_get_exclude_rule_de_parsed(ai_cache, blob)
    if hit is not None:
        logging.info("配台不要ルール: AIキャッシュヒット（配台不要ロジック→JSON）")
        return hit
    if not API_KEY:
        return None
    schema = _exclude_rule_logic_gemini_schema_instructions()
    prompt = (
        "あなたは工場の配台システム用です。次の「配台試行の説明」を」タスク1行を判定れる機械坯読ルールに変杛してください。\n\n"
        "」出力】先頭は { で終ゝりは } の JSON オブジェクト1つのみ（説明・マークダウン禁止）。\n\n"
        f"{schema}\n"
        f"」説明文】\n{blob}\n"
    )
    try:
        ppath = os.path.join(log_dir, "ai_exclude_rule_logic_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("配台試行ルール: プロンプト → %s", ppath)
    except OSError as ex:
        logging.warning("配台試行ルール: プロンプト保存失敗: %s", ex)
    try:
        client = _gemini_client(API_KEY)
        res, gem_model_used = _gemini_generate_content_with_retry(
            client, contents=prompt, log_label="配台不要ルールD→E"
        )
        record_gemini_response_usage(res, gem_model_used)
        raw = (_gemini_result_text(res) or "").strip()
        rpath = os.path.join(log_dir, "ai_exclude_rule_logic_last_response.txt")
        try:
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(raw)
        except OSError:
            pass
        parsed = _validate_exclude_rule_parsed_dict(_parse_exclude_rule_json_cell(raw))
        if parsed:
            _cache_put_exclude_rule_de_parsed(ai_cache, blob, parsed)
            save_ai_cache(ai_cache)
        return parsed
    except Exception as e:
        logging.warning("配台試行ルール: Gemini 変杛失敗: %s", e)
        return None


def _ai_compile_exclude_rule_logics_batch(blobs: list[str]) -> list[dict | None]:
    """
    複数の D 列文言を 1 回の Gemini 呼び出しで JSON 化。失敗・覝素数厳密一致時は 1 件うつにフォールバック。
    json/ai_remarks_cache.json にヒットした文言は API を呼みない。
    """
    n = len(blobs)
    if n == 0:
        return []
    ai_cache = load_ai_cache()
    out: list[dict | None] = [None] * n
    pend_i: list[int] = []
    pend_b: list[str] = []
    for i, b in enumerate(blobs):
        s = str(b).strip()
        hit = _cache_get_exclude_rule_de_parsed(ai_cache, s) if s else None
        if hit is not None:
            out[i] = hit
        else:
            pend_i.append(i)
            pend_b.append(s)
    if not pend_b:
        logging.info(
            "配台試行ルール: AIキャッシュのみで D→E ポッポ %s 件を完絝（API 呼び出しなし）。",
            n,
        )
        return out
    if not API_KEY:
        return out
    m = len(pend_b)
    if m == 1:
        out[pend_i[0]] = _ai_compile_exclude_rule_logic_to_json(pend_b[0])
        return out

    schema = _exclude_rule_logic_gemini_schema_instructions()
    numbered = "\n".join(f"[{i + 1}] {str(b).strip()}" for i, b in enumerate(pend_b))
    prompt = (
        "あなたは工場の配台システム用です。以下の N 個の「配台試行の説明」を」与ごた順庝でしれずれ JSON ルールに変杛してください。\n\n"
        f"」出力】JSON 配列のみ。先頭は [ で終ゝりは ] 。覝素数は必う {m}（Markdown・説明禁止）。\n"
        f"配列の先頭覝素は [1]」2 番目は [2] … に対応しした。\n\n"
        f"{schema}\n"
        f"」説明文】\n{numbered}\n"
    )
    try:
        ppath = os.path.join(log_dir, "ai_exclude_rule_logic_batch_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("配台試行ルール(ポッポ): プロンプト → %s", ppath)
    except OSError as ex:
        logging.warning("配台試行ルール(ポッポ): プロンプト保存失敗: %s", ex)
    try:
        client = _gemini_client(API_KEY)
        res, gem_model_used = _gemini_generate_content_with_retry(
            client, contents=prompt, log_label="配台不要ルールD→Eバッチ"
        )
        record_gemini_response_usage(res, gem_model_used)
        raw = (_gemini_result_text(res) or "").strip()
        rpath = os.path.join(log_dir, "ai_exclude_rule_logic_batch_last_response.txt")
        try:
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(raw)
        except OSError:
            pass
        arr = _parse_exclude_rule_json_array_response(raw)
        if not isinstance(arr, list) or len(arr) != m:
            logging.warning(
                "配台試行ルール: ポッポ応答は正常（覝素数 %s」期待 %s）。1 件うつ再試行した。",
                len(arr) if isinstance(arr, list) else None,
                m,
            )
            for j, idx in enumerate(pend_i):
                out[idx] = _ai_compile_exclude_rule_logic_to_json(pend_b[j])
            return out
        cache_dirty = False
        for j, item in enumerate(arr):
            parsed = _validate_exclude_rule_parsed_dict(item)
            out[pend_i[j]] = parsed
            if parsed:
                _cache_put_exclude_rule_de_parsed(ai_cache, pend_b[j], parsed)
                cache_dirty = True
        if cache_dirty:
            save_ai_cache(ai_cache)
        return out
    except Exception as e:
        logging.warning("配台試行ルール: ポッポ Gemini 失敗」坘発にフォールバック: %s", e)
        for j, idx in enumerate(pend_i):
            out[idx] = _ai_compile_exclude_rule_logic_to_json(pend_b[j])
        return out


def _log_exclude_rules_sheet_debug(
    event: str,
    log_prefix: str,
    summary: str,
    details: str = "",
    exc: BaseException | None = None,
) -> None:
    """
    「設定_配台不要工程」の保守処理のイベントログ。

    設定シート処理の追跡を log/exclude_rules_sheet_debug.txt に追記し、execution_log にもタグ付しで出力れる。
    event 例: START, OPEN_OK, OPEN_RETRY, OPEN_FAIL, HEADER_FIX, SYNC_ROWS, OPENPYXL_SAVE_OK, OPENPYXL_SAVE_FAIL,
    OPENPYXL_SAVE_SKIPPED_EXCLUDE_RULES_POLICY, OPENPYXL_RETRY_WAIT, OPENPYXL_VBA_FALLBACK, MATRIX_TSV_WRITTEN,
    XLWINGS_UNAVAILABLE, XLWINGS_ATTACH_FAIL, XLWINGS_SYNC_SKIP, XLWINGS_SYNC_OK, XLWINGS_SYNC_FAIL,
    E_SIDECAR_WRITTEN, E_SIDECAR_APPLIED, FALLBACK_FAIL,
    SKIP_NO_PATH, SKIP_NO_FILE, SKIP_NO_SHEET, DATA_COMPACT
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"--- {ts} ---",
        f"event={event}",
        f"phase={log_prefix}",
        f"summary={summary}",
    ]
    if details:
        lines.append(f"details={details}")
    if exc is not None:
        lines.append(f"exception={type(exc).__name__}: {exc}")
        lines.append(traceback.format_exc().rstrip())
    block = "\n".join(lines) + "\n\n"
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(exclude_rules_sheet_debug_log_path, "a", encoding="utf-8", newline="\n") as df:
            df.write(block)
    except OSError as wex:
        logging.warning("exclude_rules_sheet_debug.txt へ書けません: %s", wex)

    tag = "[設定_配台不要工程]"
    msg = f"{tag} {event} | {log_prefix} | {summary}"
    if details:
        msg += f" | {details}"
    if event in (
        "OPEN_FAIL",
        "SAVE_FAIL",
        "COM_MERGE_FAIL",
        "FALLBACK_FAIL",
        "SKIP_NO_PATH",
        "SKIP_NO_FILE",
        "SKIP_NO_SHEET",
        "FATAL",
    ):
        logging.error(msg)
    elif event in (
        "OPEN_RETRY",
        "SAVE_FAIL_HINT",
        "SAVE_RETRY",
        "COM_SYNC_UNAVAILABLE",
        "COM_ATTACH_OPEN_FAIL",
        "XLWINGS_UNAVAILABLE",
        "XLWINGS_ATTACH_FAIL",
        "XLWINGS_SYNC_SKIP",
        "XLWINGS_SYNC_FAIL",
        "E_SIDECAR_WRITTEN",
        "OPENPYXL_SAVE_FAIL",
        "OPENPYXL_VBA_FALLBACK",
    ):
        logging.warning(msg)
    elif event in (
        "COM_MERGE_SKIP",
        "MATRIX_TSV_WRITTEN",
        "OPENPYXL_SAVE_OK",
        "OPENPYXL_RETRY_WAIT",
        "XLWINGS_SYNC_OK",
    ):
        logging.info(msg)
    else:
        logging.info(msg)


def _xlwings_paths_equivalent(disk_path: str, book_fullname: str) -> bool:
    """ディスクパスと xlwings Book.full_name は同一ファイルを指れか（表記ゆれを多少坸坎）。"""
    try:
        fn = str(book_fullname).strip()
    except Exception:
        return False

    def _norm(p: str) -> str:
        p = os.path.normpath(str(p).strip().replace("/", "\\"))
        return os.path.normcase(os.path.abspath(p))

    try:
        if _norm(disk_path) == _norm(fn):
            return True
    except Exception:
        pass
    try:
        return os.path.samefile(disk_path, fn)
    except Exception:
        pass
    try:
        import win32api  # type: ignore

        a = _norm(win32api.GetLongPathName(disk_path))
        b = _norm(win32api.GetLongPathName(fn))
        if a == b:
            return True
    except Exception:
        pass
    try:
        if os.path.basename(_norm(disk_path)).lower() == os.path.basename(_norm(fn)).lower():
            if _norm(os.path.dirname(disk_path)) == _norm(os.path.dirname(fn)):
                return True
    except Exception:
        pass
    return False


def _xlwings_book_matches_path(book, disk_path: str) -> bool:
    try:
        fn = book.full_name
    except Exception:
        return False
    return _xlwings_paths_equivalent(disk_path, fn)


def _xlwings_find_book_on_running_instances(abs_path: str):
    """起動中の Excel からパス一致する xlwings Book を返す。無ければ None。

    旧実装は ``list(xw.apps)`` / ``xw.apps.active`` / ``for app in xw.apps`` が
    環境によって COM 無応答となり段階1が停止したため廃止。
    マクロブックの保存は ``_xlwings_attach_open_macro_workbook`` が新規 ``xw.App`` で開く。
    """
    return None


def _xlwings_try_open_in_running_apps(abs_path: str):
    """既存の Excel.App で Workbooks.Open を試す。成功時 Book、失敗時 None。

    ``xw.apps`` 経由は上記と同理由で廃止。常に None。
    """
    return None


def _xlwings_release_book_after_mutation(xw_book, info: dict, mutation_ok: bool) -> None:
    """専用起動した Excel は終了れる。実行中 Excel での値 Open したブックは失敗時のみ閉もる。"""
    if xw_book is None:
        return
    mode = info.get("mode", "keep")
    opened_here = bool(info.get("opened_wb_here"))
    if mode == "quit_excel":
        try:
            xw_book.close()
        except Exception:
            pass
        try:
            xw_book.app.quit()
        except Exception:
            pass
        return
    if opened_here and not mutation_ok:
        try:
            xw_book.close()
        except Exception:
            pass


def _xlwings_attach_open_macro_workbook(macro_wb_path: str, log_prefix: str):
    """
    マクロブックを xlwings で取得れる（本番・テスト共通）。
    戻り値: (Book, release_info) / 失敗時 None。
    release_info: mode は keep または quit_excel」opened_wb_here は bool。
    """
    try:
        import xlwings as xw  # noqa: F401
    except ImportError:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_UNAVAILABLE",
            log_prefix,
            "xlwings は import でしません（pip install xlwings を確認）。",
        )
        return None

    abs_path = os.path.abspath(macro_wb_path)

    _log_exclude_rules_sheet_debug(
        "XLWINGS_APPS_ENUM_SKIPPED",
        log_prefix,
        "xw.apps 列挙はスキップし、新規 Excel でブックを開きます（COM 無応答回避）。",
        details=f"path={abs_path}",
    )

    try:
        import xlwings as xw

        app = xw.App(visible=False, add_book=False)
        try:
            app.display_alerts = False
        except Exception:
            pass
        book = app.books.open(abs_path, update_links=False)
        return book, {"mode": "quit_excel", "opened_wb_here": True}
    except Exception as ex:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_ATTACH_FAIL",
            log_prefix,
            "xlwings でブックを開きませんでした。",
            details=f"path={abs_path}",
            exc=ex,
        )
        return None


def _xlwings_attach_workbook_for_tests(
    book_path: str,
    label: str,
    *,
    allow_dispatch_open: bool = False,
):
    """
    検証スクリプト用: 起動中ブックを優先し、必須なら表示付し Excel で開し。
    戻り値: (Book, info, 説明文字列) または None。
    """
    abs_path = os.path.abspath(book_path)
    if not allow_dispatch_open:
        return None
    try:
        import xlwings as xw

        app = xw.App(visible=True, add_book=False)
        try:
            app.display_alerts = False
        except Exception:
            pass
        book = app.books.open(abs_path, update_links=False)
        return book, {"mode": "keep", "opened_wb_here": True}, f"{label}:dispatch-open"
    except Exception:
        return None


def _xlwings_app_save_perf_state_push(app):
    """VBA 坴のスプラッシュポーリングと競坈しにししれるため、同期・保存の短時間の値 Excel を静かにれる。"""
    snap = {}
    for attr in ("screen_updating", "calculation", "enable_events"):
        try:
            snap[attr] = getattr(app, attr)
        except Exception:
            snap[attr] = None
    try:
        app.screen_updating = False
    except Exception:
        pass
    try:
        app.calculation = "manual"
    except Exception:
        try:
            app.api.Calculation = -4135  # xlCalculationManual
        except Exception:
            pass
    try:
        app.enable_events = False
    except Exception:
        pass
    return snap


def _xlwings_app_save_perf_state_pop(app, snap):
    if not snap:
        return
    for attr in ("enable_events", "calculation", "screen_updating"):
        prev = snap.get(attr)
        if prev is None:
            continue
        try:
            setattr(app, attr, prev)
        except Exception:
            pass


def _xlwings_sync_exclude_rules_sheet_from_openpyxl(
    wb_path: str, ws_oxl, log_prefix: str
) -> bool:
    """
    openpyxl で保存でしないとし」xlwings で「設定_配台不要工程」A:E をメモリ上の値で上書きし Save。

    表示中シートに対れる一括 .value の値てと」スプラッシュ＋ポーリング（D3=true）下で
    Range 代入は数分かかる計測はあり得る。同期中のみシートを一時非表示にし api.Value2 で書き。
    """
    global _exclude_rules_effective_read_path

    attached = _xlwings_attach_open_macro_workbook(wb_path, log_prefix)
    if attached is None:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_SYNC_SKIP",
            log_prefix,
            "xlwings でブックに接続でしう A:E 同期をスキップ。",
            details=f"path={wb_path}",
        )
        return False

    xw_book, info = attached
    ok = False
    try:
        try:
            xw_book.app.display_alerts = False
        except Exception:
            pass
        # 全シート名を列挙れるとシート数分の COM 往復になり」D3=true 時は VBA ポーリングと競坈して
        # 1 シート数秒〜坝数秒かかることはある（計測で 40 シート≈213s）。坝剝で直接解決れる。
        try:
            sht = xw_book.sheets[EXCLUDE_RULES_SHEET_NAME]
        except Exception:
            _log_exclude_rules_sheet_debug(
                "XLWINGS_SYNC_SKIP",
                log_prefix,
                f"xlwings 坴にシート「{EXCLUDE_RULES_SHEET_NAME}」はありません。",
                details=f"path={wb_path}",
            )
            return False
        max_r = max(1, int(ws_oxl.max_row or 1))
        ncols = EXCLUDE_RULES_SHEET_COM_SYNC_MAX_COL
        data = [
            [ws_oxl.cell(row=r, column=c).value for c in range(1, ncols + 1)]
            for r in range(1, max_r + 1)
        ]
        _perf_snap = _xlwings_app_save_perf_state_push(xw_book.app)
        rng = sht.range((1, 1)).resize(len(data), ncols)
        hid_sheet_for_write = False
        try:
            try:
                if int(sht.api.Visible) == -1:  # xlSheetVisible
                    sht.api.Visible = 0  # xlSheetHidden（同期中の値。再杝画・ウィンドウ更新負蝷を抑ごる）
                    hid_sheet_for_write = True
            except Exception:
                pass
            try:
                rng.api.Value2 = data
            except Exception:
                rng.value = data
            xw_book.save()
        finally:
            if hid_sheet_for_write:
                try:
                    sht.api.Visible = -1
                except Exception:
                    pass
            _xlwings_app_save_perf_state_pop(xw_book.app, _perf_snap)
        ok = True
        _exclude_rules_effective_read_path = wb_path
        _clear_exclude_rules_e_apply_files()
        _log_exclude_rules_sheet_debug(
            "XLWINGS_SYNC_OK",
            log_prefix,
            "xlwings 経由で設定シート A〜E を同期しブックを保存しました。",
            details=f"path={wb_path} rows={max_r}",
        )
        logging.info(
            "%s: 設定シートを xlwings でマクロブックに保存しました（A〜E）。",
            log_prefix,
        )
        return True
    except Exception as ex:
        _log_exclude_rules_sheet_debug(
            "XLWINGS_SYNC_FAIL",
            log_prefix,
            "xlwings での A:E 同期または Save に失敗しました。",
            details=f"path={wb_path}",
            exc=ex,
        )
        return False
    finally:
        _xlwings_release_book_after_mutation(xw_book, info, ok)


# 設定シートの列範囲（A〜E）。xlwings 同期・VBA 行列 TSV 出力でも使用。
EXCLUDE_RULES_SHEET_COM_SYNC_MAX_COL = 5
EXCLUDE_RULES_MATRIX_CLIP_MAX_COL = 5


def _persist_exclude_rules_workbook(_wb, wb_path: str, ws, log_prefix: str) -> bool:
    """
    設定シートのディスク反映。既定は xlwings で A:E 同期→Save（EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1 のときのみ openpyxl save を試行）。
    保存でしないとしは log に行列 TSV を出し、VBA「設定_配台不要工程_AからE_TSVから反映」で反映れる。

    _wb … 編集済み openpyxl ブック（openpyxl 経路時のみ save に使用）。
    """
    global _exclude_rules_effective_read_path

    def _openpyxl_persist_ok(which: str) -> bool:
        try:
            _wb.save(wb_path)
        except Exception as ex:
            _log_exclude_rules_sheet_debug(
                "OPENPYXL_SAVE_FAIL",
                log_prefix,
                f"openpyxl での .xlsm 保存に失敗しました {which}（Excel で開しっりなし・ロックの可能性）。",
                details=f"path={wb_path}",
                exc=ex,
            )
            return False
        _exclude_rules_effective_read_path = wb_path
        _clear_exclude_rules_e_apply_files()
        _log_exclude_rules_sheet_debug(
            "OPENPYXL_SAVE_OK",
            log_prefix,
            "openpyxl で設定シートを含むブックを保存しました（A〜E）。",
            details=f"path={wb_path} {which}",
        )
        logging.info(
            "%s: 設定シートを openpyxl でマクロブックに保存しました。%s",
            log_prefix,
            which,
        )
        return True

    saved_openpyxl = False
    if EXCLUDE_RULES_SKIP_OPENPYXL_SAVE:
        _log_exclude_rules_sheet_debug(
            "OPENPYXL_SAVE_SKIPPED_EXCLUDE_RULES_POLICY",
            log_prefix,
            "設定_配台不要工程の保存では openpyxl save を試行しません（xlwings 同期を先行。再試行れる場合は EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1）。",
            details=f"path={wb_path}",
        )
        logging.info(
            "%s: 設定_配台不要工程は openpyxl を試さず xlwings 同期→Save を試みした（試行なら VBA 用行列 TSV）。",
            log_prefix,
        )
    elif not _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: 設定_配台不要工程は openpyxl で保存しした（試行のときは xlwings 同期→Save」しれも試行なら VBA 用行列 TSV）。",
            log_prefix,
        )
        labels = ("(1/4)", "(2/4)", "(3/4)", "(4/4)")
        for i, label in enumerate(labels):
            if i:
                _log_exclude_rules_sheet_debug(
                    "OPENPYXL_RETRY_WAIT",
                    log_prefix,
                    f"openpyxl 再保存まで 2 秒待うした {label}。",
                    details=f"path={wb_path}",
                )
                time_module.sleep(2.0)
            if _openpyxl_persist_ok(label):
                saved_openpyxl = True
                break
    else:
        _log_exclude_rules_sheet_debug(
            "OPENPYXL_SAVE_SKIPPED_INCOMPATIBLE_SHEET",
            log_prefix,
            f"ブックに「{OPENPYXL_INCOMPATIBLE_SHEET_MARKER}」があるため、openpyxl での保存を試みません。",
            details=f"path={wb_path}",
        )
        logging.info(
            "%s: ブックに「%s」があるため、openpyxl save をスキップし、xlwings または行列 TSV に切り替ごした。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )

    if saved_openpyxl:
        return True

    if _xlwings_sync_exclude_rules_sheet_from_openpyxl(wb_path, ws, log_prefix):
        return True

    if _write_exclude_rules_matrix_vba_tsv(wb_path, ws, log_prefix):
        logging.warning(
            "%s: 設定シートを log\\%s に出力しました。"
            " Excel でマクロ「設定_配台不要工程_AからE_TSVから反映」を実行してください。",
            log_prefix,
            EXCLUDE_RULES_MATRIX_VBA_FILENAME,
        )

    _log_exclude_rules_sheet_debug(
        "OPENPYXL_VBA_FALLBACK",
        log_prefix,
        "openpyxl 保存に失敗したため、 VBA 用行列 TSV を出力しました（ブックは Excel 上で手動反映は必須な場合はありした）。",
        details=f"path={wb_path}",
    )
    return False


def _exclude_rules_e_sidecar_path() -> str:
    path = os.path.join(json_data_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME)
    legacy = os.path.join(log_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME)
    if os.path.isfile(legacy) and not os.path.isfile(path):
        try:
            shutil.move(legacy, path)
        except OSError:
            pass
    return path


def _exclude_rules_e_vba_tsv_path() -> str:
    return os.path.join(log_dir, EXCLUDE_RULES_E_VBA_TSV_FILENAME)


def _exclude_rules_matrix_vba_path() -> str:
    return os.path.join(log_dir, EXCLUDE_RULES_MATRIX_VBA_FILENAME)


def _serialize_cell_for_matrix_tsv(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    return str(val)


def _write_exclude_rules_matrix_vba_tsv(
    wb_path: str, ws, log_prefix: str
) -> bool:
    """VBA 用: 設定シート 1 行目〜 max_row の A〜E を Base64(UTF-8) 付し TSV で出力れる。"""
    max_r = max(1, int(ws.max_row or 1))
    lines = [
        "v1",
        "workbook\t" + os.path.abspath(wb_path),
        "sheet\t" + EXCLUDE_RULES_SHEET_NAME,
        "ncols\t5",
        "---",
    ]
    for r in range(1, max_r + 1):
        parts: list[str] = [str(r)]
        for c in range(1, 6):
            s = _serialize_cell_for_matrix_tsv(ws.cell(row=r, column=c).value)
            parts.append(base64.b64encode(s.encode("utf-8")).decode("ascii"))
        lines.append("\t".join(parts))
    path = _exclude_rules_matrix_vba_path()
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lines) + "\n")
        _log_exclude_rules_sheet_debug(
            "MATRIX_TSV_WRITTEN",
            log_prefix,
            "設定シート A〜E を VBA 反映用 TSV に書き出しました（openpyxl 保存試行時）。",
            details=f"path={path} rows={max_r}",
        )
        return True
    except OSError as ex:
        logging.warning("%s: 行列 VBA 用 TSV を書けません: %s", log_prefix, ex)
        return False


def _build_exclude_rules_list_from_openpyxl_ws(
    ws, c_proc: int, c_mach: int, c_flag: int, c_e: int
) -> list[dict]:
    """openpyxl 上の設定シートから _load_exclude_rules_from_workbook と同形のリストを構築。"""
    rules: list[dict] = []
    max_r = int(ws.max_row or 1)
    for r in range(2, max_r + 1):
        pv = ws.cell(row=r, column=c_proc).value
        proc = (
            ""
            if pv is None or (isinstance(pv, float) and pd.isna(pv))
            else str(pv).strip()
        )
        if not proc:
            continue
        mv = ws.cell(row=r, column=c_mach).value
        mach = (
            ""
            if mv is None or (isinstance(mv, float) and pd.isna(mv))
            else str(mv).strip()
        )
        cv = ws.cell(row=r, column=c_flag).value
        ev = ws.cell(row=r, column=c_e).value
        parsed = _parse_exclude_rule_json_cell(ev)
        rules.append(
            {"proc": proc, "mach": mach, "c_val": cv, "parsed": parsed}
        )
    return rules


def _set_exclude_rules_snapshot_from_ws(
    wb_path: str, ws, c_proc: int, c_mach: int, c_flag: int, c_e: int
) -> None:
    global _exclude_rules_rules_snapshot, _exclude_rules_snapshot_wb
    _exclude_rules_rules_snapshot = _build_exclude_rules_list_from_openpyxl_ws(
        ws, c_proc, c_mach, c_flag, c_e
    )
    _exclude_rules_snapshot_wb = os.path.normcase(os.path.abspath(wb_path))


def _clear_exclude_rules_e_apply_files() -> None:
    for p in (
        os.path.join(json_data_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME),
        os.path.join(log_dir, EXCLUDE_RULES_E_SIDECAR_FILENAME),
    ):
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass
    for rel in (EXCLUDE_RULES_E_VBA_TSV_FILENAME, EXCLUDE_RULES_MATRIX_VBA_FILENAME):
        p = os.path.join(log_dir, rel)
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass


def _write_exclude_rules_e_vba_tsv_from_cells(
    wb_path: str, c_e: int, cells: dict[str, str], log_prefix: str
) -> None:
    """VBA 用: 行番坷と Base64(UTF-8) セル文字列の TSV。"""
    lines = [
        "v1",
        "workbook\t" + os.path.abspath(wb_path),
        "sheet\t" + EXCLUDE_RULES_SHEET_NAME,
        "column_e\t" + str(int(c_e)),
        "---",
    ]
    for rk in sorted(cells.keys(), key=lambda x: int(x)):
        s = cells[rk]
        b64 = base64.b64encode(s.encode("utf-8")).decode("ascii")
        lines.append(rk + "\t" + b64)
    path_tsv = _exclude_rules_e_vba_tsv_path()
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path_tsv, "w", encoding="utf-8", newline="\n") as f:
            f.write("\n".join(lines) + "\n")
        _log_exclude_rules_sheet_debug(
            "E_VBA_TSV_WRITTEN",
            log_prefix,
            "E 列を VBA 反映用 TSV に書き出しました（保存失敗時のフォールバック用）。",
            details=f"path={path_tsv} cells={len(cells)}",
        )
    except OSError as ex:
        logging.warning("%s: E 列 VBA 用 TSV を書けません: %s", log_prefix, ex)


def _write_exclude_rules_e_apply_artifacts(
    wb_path: str, ws, c_e: int, log_prefix: str
) -> None:
    """
    E 列（非空）を JSON サイドカードと VBA 用 TSV に書き。空なら両ファイルを削除。
    Python 次回起動時の E 復元用 JSON と」マクロからの E 書込み用 TSV。
    """
    cells: dict[str, str] = {}
    max_r = int(ws.max_row or 1)
    for r in range(2, max_r + 1):
        ev = ws.cell(row=r, column=c_e).value
        if _cell_is_blank_for_rule(ev):
            continue
        s = str(ev).strip() if ev is not None else ""
        if not s:
            continue
        cells[str(r)] = s
    if not cells:
        _clear_exclude_rules_e_apply_files()
        return
    payload = {
        "version": 1,
        "workbook": os.path.abspath(wb_path),
        "sheet": EXCLUDE_RULES_SHEET_NAME,
        "column_e": c_e,
        "cells": cells,
    }
    path_sc = _exclude_rules_e_sidecar_path()
    try:
        os.makedirs(json_data_dir, exist_ok=True)
        with open(path_sc, "w", encoding="utf-8", newline="\n") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except OSError as ex:
        logging.warning("%s: E 列 JSON を書けません: %s", log_prefix, ex)
    _write_exclude_rules_e_vba_tsv_from_cells(wb_path, c_e, cells, log_prefix)
    _log_exclude_rules_sheet_debug(
        "E_APPLY_FILES_WRITTEN",
        log_prefix,
        "E 列を JSON と VBA 用 TSV に書き出しました（マクロで E 列を反映後」ファイル削除）。",
        details=f"cells={len(cells)}",
    )


def _try_apply_pending_exclude_rules_e_column(
    wb_path: str, ws, c_e: int, log_prefix: str
) -> int:
    """
    剝回保存に失敗したとし書き出した JSON から E 列を復元れる。
    ブックパスは一致しなけれみ何もしない。適用後はサイドカードを削除する。
    """
    path_sc = _exclude_rules_e_sidecar_path()
    if not os.path.isfile(path_sc):
        return 0
    try:
        with open(path_sc, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return 0
    if int(payload.get("version") or 0) != 1:
        return 0
    target = os.path.normcase(os.path.abspath(wb_path))
    if os.path.normcase(str(payload.get("workbook") or "")) != target:
        return 0
    if str(payload.get("sheet") or "") != EXCLUDE_RULES_SHEET_NAME:
        return 0
    cells = payload.get("cells")
    if not isinstance(cells, dict):
        return 0
    n = 0
    for rk, val in cells.items():
        try:
            ri = int(rk)
        except (TypeError, ValueError):
            continue
        if ri < 2:
            continue
        if isinstance(val, dict):
            sval = json.dumps(val, ensure_ascii=False)
        else:
            sval = "" if val is None else str(val).strip()
        if not sval:
            continue
        ws.cell(row=ri, column=c_e, value=sval)
        n += 1
    try:
        os.remove(path_sc)
    except OSError:
        pass
    if n:
        _log_exclude_rules_sheet_debug(
            "E_SIDECAR_APPLIED",
            log_prefix,
            f"未保存だった E 列をサイドカードから {n} セル復元しました。",
            details=path_sc,
        )
        logging.info(
            "%s: %s の内容をシートのロジック式列へ適用しました（続けて保存を試みした）。",
            log_prefix,
            path_sc,
        )
    return n


def _read_exclude_rules_d_cells_data_only_for_rows(
    wb_path: str, rows: list[int], c_d: int
) -> dict[int, object]:
    """
    D 列は数弝のとき」openpyxl の通常読込では '=...' しか取れない。
    data_only=True でキャッシュ値を読む（Excel は一度でも保存・計算済みのブックで有効）。
    """
    out: dict[int, object] = {}
    if not rows or not os.path.isfile(wb_path):
        return out
    if _workbook_should_skip_openpyxl_io(wb_path):
        return out
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wbro = None
    try:
        wbro = load_workbook(
            wb_path,
            read_only=True,
            data_only=True,
            keep_vba=keep_vba,
        )
    except Exception:
        return out
    try:
        if EXCLUDE_RULES_SHEET_NAME not in wbro.sheetnames:
            return out
        wsro = wbro[EXCLUDE_RULES_SHEET_NAME]
        for r in rows:
            if r < 2:
                continue
            try:
                out[r] = wsro.cell(row=r, column=c_d).value
            except Exception:
                pass
    finally:
        if wbro is not None:
            try:
                wbro.close()
            except Exception:
                pass
    return out


def run_exclude_rules_sheet_maintenance(
    wb_path: str,
    pairs: list[tuple[str, str]],
    log_prefix: str,
    *,
    compile_exclude_rules_d_to_e_with_ai: bool = True,
) -> None:
    """
    「設定_配台不要工程」の行同期・（任意で）D→E の AI 補完・ディスク反映（既定は xlwings で A〜E 同期→Save。``EXCLUDE_RULES_TRY_OPENPYXL_SAVE=1`` のとき openpyxl save を試行）。

    ``compile_exclude_rules_d_to_e_with_ai=False`` のときは D 列→E 列（ロジック式 JSON）の
    Gemini 補完のみスキップする（行同期・空行詰め・退避 E の復元・保存は従来どおり）。
    段階2の ``load_planning_tasks_df`` 経路では False を渡す。

    xlwings でも保存でしないとしは ``log/exclude_rules_matrix_vba.tsv`` を残し、マクロ
    ``設定_配台不要工程_AからE_TSVから反映`` で A〜E を反映れる。
    併せで従来どおり E 列のみの ``exclude_rules_e_column_vba.tsv`` も出力され得る（行列 TSV 優先で反映後は削除）。
    保存成功時は TSV/JSON は削除される。

    ``json/exclude_rules_e_column_pending.json`` は Python 次回起動時の E 列復元用。
    シートの新規作成と 1 行目見出しは VBA「設定_配台不要工程_シートを確保」。
    """
    if not wb_path:
        _log_exclude_rules_sheet_debug(
            "SKIP_NO_PATH",
            log_prefix,
            "TASK_INPUT_WORKBOOK は空のため、設定シート処理をしません。",
        )
        return
    if not os.path.exists(wb_path):
        _log_exclude_rules_sheet_debug(
            "SKIP_NO_FILE",
            log_prefix,
            "ブックは存在しません。",
            details=f"path={wb_path}",
        )
        return

    _log_exclude_rules_sheet_debug(
        "START",
        log_prefix,
        "設定シート保守開始",
        details=f"path={wb_path} pairs={len(pairs)} ai_d_to_e={compile_exclude_rules_d_to_e_with_ai}",
    )
    global _exclude_rules_effective_read_path
    _exclude_rules_effective_read_path = None

    if _workbook_should_skip_openpyxl_io(wb_path):
        _log_exclude_rules_sheet_debug(
            "SKIP_OPENPYXL_INCOMPATIBLE_BOOK",
            log_prefix,
            f"ブックに「{OPENPYXL_INCOMPATIBLE_SHEET_MARKER}」は含まれるため、openpyxl による設定シート保守は行いません。",
            details=f"path={wb_path}",
        )
        logging.warning(
            "%s: 「%s」併有のため、「%s」の openpyxl 保守をスキップしました（Excel＝xlwings で編集してください）。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            EXCLUDE_RULES_SHEET_NAME,
        )
        return

    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(wb_path, keep_vba=keep_vba, read_only=False, data_only=False)
    except Exception as e1:
        if keep_vba:
            _log_exclude_rules_sheet_debug(
                "OPEN_RETRY",
                log_prefix,
                "keep_vba=True でブックを開けう keep_vba=False で再試行した（マクロは失ゝれる可能性）。",
                exc=e1,
            )
            try:
                wb = load_workbook(wb_path, keep_vba=False, read_only=False, data_only=False)
            except Exception as e2:
                _log_exclude_rules_sheet_debug(
                    "OPEN_FAIL",
                    log_prefix,
                    "ブックを開きません。シートは作成・保存されません。",
                    details=f"path={wb_path}",
                    exc=e2,
                )
                return
        else:
            _log_exclude_rules_sheet_debug(
                "OPEN_FAIL",
                log_prefix,
                "ブックを開きません。シートは作成・保存されません。",
                details=f"path={wb_path}",
                exc=e1,
            )
            return

    _log_exclude_rules_sheet_debug(
        "OPEN_OK",
        log_prefix,
        "ブックを開しました。",
        details=f"keep_vba={keep_vba} sheets={len(wb.sheetnames)}",
    )

    try:
        if EXCLUDE_RULES_SHEET_NAME not in wb.sheetnames:
            _log_exclude_rules_sheet_debug(
                "SKIP_NO_SHEET",
                log_prefix,
                "シートはありません。VBA の「設定_配台不要工程_シートを確保」を実行れるか」段階1/2 をマクロから起動してください。",
                details=f"path={wb_path}",
            )
            logging.error(
                "%s: 「%s」はありません。Python ではシートを作成しません。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
            )
            return

        ws = wb[EXCLUDE_RULES_SHEET_NAME]
        hm_before = _exclude_rules_sheet_header_map(ws)
        c_proc, c_mach, c_flag, c_d, c_e = _ensure_exclude_rules_sheet_headers_and_columns(
            ws, log_prefix
        )
        hm_after = _exclude_rules_sheet_header_map(ws)
        if tuple(hm_before.get(x) for x in (
            EXCLUDE_RULE_COL_PROCESS,
            EXCLUDE_RULE_COL_MACHINE,
            EXCLUDE_RULE_COL_FLAG,
            EXCLUDE_RULE_COL_LOGIC_JA,
            EXCLUDE_RULE_COL_LOGIC_JSON,
        )) != tuple(hm_after.get(x) for x in (
            EXCLUDE_RULE_COL_PROCESS,
            EXCLUDE_RULE_COL_MACHINE,
            EXCLUDE_RULE_COL_FLAG,
            EXCLUDE_RULE_COL_LOGIC_JA,
            EXCLUDE_RULE_COL_LOGIC_JSON,
        )):
            _log_exclude_rules_sheet_debug(
                "HEADER_FIX",
                log_prefix,
                "1行目に標準見出しを書き込みました（空シート・列名厳密一致の補正）。",
                details=f"cols=({c_proc},{c_mach},{c_flag},{c_d},{c_e})",
            )

        # 剝回ブック保存に失敗したとし退避した E 列を」先にワークシートへ戻れ（続し保存でディスクへ載る）
        _try_apply_pending_exclude_rules_e_column(wb_path, ws, c_e, log_prefix)

        existing_keys: set[tuple[str, str]] = set()
        max_r = max(2, int(ws.max_row or 2))
        for r in range(2, max_r + 1):
            pv = ws.cell(row=r, column=c_proc).value
            mv = ws.cell(row=r, column=c_mach).value
            p = str(pv).strip() if pv is not None and not (isinstance(pv, float) and pd.isna(pv)) else ""
            m = str(mv).strip() if mv is not None and not (isinstance(mv, float) and pd.isna(mv)) else ""
            if not p:
                continue
            existing_keys.add(
                (_normalize_process_name_for_rule_match(p), _normalize_equipment_match_key(m))
            )

        added = 0
        for p, m in pairs:
            key = (_normalize_process_name_for_rule_match(p), _normalize_equipment_match_key(m))
            if key in existing_keys:
                continue
            ws.append([p, m, None, None, None])
            existing_keys.add(key)
            added += 1
        if added:
            _log_exclude_rules_sheet_debug(
                "SYNC_ROWS",
                log_prefix,
                f"工程+機械の行を {added} 件追加しました。",
            )
            logging.info(
                "%s: 「%s」に工程+機械の組み合わせを %s 行追加しました。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
                added,
            )

        # 加工計画からペアは1件も坖れう」シートにもデータ行は無いとしは例行のみ（従来の新規シート相当）
        if added == 0 and not existing_keys:
            ws.append(["梱包", "", "yes", "", ""])
            existing_keys.add(
                (_normalize_process_name_for_rule_match("梱包"), _normalize_equipment_match_key(""))
            )
            _log_exclude_rules_sheet_debug(
                "EXAMPLE_ROW",
                log_prefix,
                "データ行はなかったため、例（梱包=yes）を1行追加。",
            )
            logging.info(
                "%s: 「%s」にデータ行はなかったため、例（梱包=yes）を1行追加しました。",
                log_prefix,
                EXCLUDE_RULES_SHEET_NAME,
            )

        # 空行詰ゝは AI より先に行ご（後から詰ゝると」書き込んて行番坷と画面上の行はうれる）
        n_kept, n_removed_empty = _compact_exclude_rules_data_rows(
            ws, c_proc, c_mach, c_flag, c_d, c_e, log_prefix
        )
        if n_removed_empty:
            _log_exclude_rules_sheet_debug(
                "DATA_COMPACT",
                log_prefix,
                "空行を削除してデータ行を詰ゝました（並よ順は維挝）。AI 補完より剝。",
                details=f"rows={n_kept} removed_empty={n_removed_empty}",
            )

        if compile_exclude_rules_d_to_e_with_ai:
            max_r = int(ws.max_row or 1)
            pending_rows: list[int] = []
            for r in range(2, max_r + 1):
                dv = ws.cell(row=r, column=c_d).value
                ev = ws.cell(row=r, column=c_e).value
                # C 列の有無に関係なく、D に説明があり E が空なら D→E を試す
                if _cell_is_blank_for_rule(dv):
                    continue
                if not _cell_is_blank_for_rule(ev):
                    continue
                pending_rows.append(r)

            # D が数式のときは通常読込では '=...' だけ取れる。data_only でキャッシュ表示値を補う。
            formula_rows = [
                r
                for r in pending_rows
                if isinstance(ws.cell(row=r, column=c_d).value, str)
                and str(ws.cell(row=r, column=c_d).value).strip().startswith("=")
            ]
            d_cached = (
                _read_exclude_rules_d_cells_data_only_for_rows(wb_path, formula_rows, c_d)
                if formula_rows
                else {}
            )
            pending_texts: list[str] = []
            filtered_rows: list[int] = []
            for r in pending_rows:
                dv = ws.cell(row=r, column=c_d).value
                blob = (
                    ""
                    if dv is None or (isinstance(dv, float) and pd.isna(dv))
                    else str(dv).strip()
                )
                if blob.startswith("="):
                    alt = d_cached.get(r)
                    if alt is not None and not (isinstance(alt, float) and pd.isna(alt)):
                        blob = str(alt).strip()
                    else:
                        logging.warning(
                            "%s: 「%s」%s 行目の D 列が数式で、キャッシュ値を読めませんでした（Excel で一度保存するか D を値にしてください）。",
                            log_prefix,
                            EXCLUDE_RULES_SHEET_NAME,
                            r,
                        )
                        continue
                if _cell_is_blank_for_rule(blob):
                    continue
                filtered_rows.append(r)
                pending_texts.append(blob)
            pending_rows = filtered_rows

            ai_filled = 0
            ai_e_cell_addrs: list[str] = []
            if pending_texts:
                parsed_list = _ai_compile_exclude_rule_logics_batch(pending_texts)
                for r, parsed in zip(pending_rows, parsed_list):
                    if not parsed:
                        logging.warning(
                            "%s: 「%s」%s 行目の D 列を JSON にできませんでした（APIキー・応答を確認）。",
                            log_prefix,
                            EXCLUDE_RULES_SHEET_NAME,
                            r,
                        )
                        continue
                    jstr = json.dumps(parsed, ensure_ascii=False)
                    ws.cell(row=r, column=c_e, value=jstr)
                    cell_addr = f"{get_column_letter(c_e)}{r}"
                    ai_e_cell_addrs.append(cell_addr)
                    preview = jstr if len(jstr) <= 160 else (jstr[:160] + "…")
                    logging.info(
                        "%s: 「%s」ロジック式列「%s」セル %s に JSON を書き込み: %s",
                        log_prefix,
                        EXCLUDE_RULES_SHEET_NAME,
                        EXCLUDE_RULE_COL_LOGIC_JSON,
                        cell_addr,
                        preview,
                    )
                    ai_filled += 1
            if ai_filled:
                _log_exclude_rules_sheet_debug(
                    "AI_E_FILLED",
                    log_prefix,
                    f"D→E の AI 補完を {ai_filled} 行実施。",
                    details="cells=" + ",".join(ai_e_cell_addrs),
                )
                logging.info(
                    "%s: 「%s」で D→E の AI 補完を %s 行（セル: %s）。",
                    log_prefix,
                    EXCLUDE_RULES_SHEET_NAME,
                    ai_filled,
                    ",".join(ai_e_cell_addrs),
                )
        else:
            _log_exclude_rules_sheet_debug(
                "SKIP_AI_D_TO_E",
                log_prefix,
                "D→E の AI 補完をスキップ（呼び出し側指定）。",
            )

        _er_test = os.environ.get("EXCLUDE_RULES_TEST_E1234", "").strip().lower()
        if _er_test in ("1", "yes", "true"):
            try:
                _er_row = int(os.environ.get("EXCLUDE_RULES_TEST_E1234_ROW", "9") or "9")
            except ValueError:
                _er_row = 9
            if _er_row < 2:
                _er_row = 9
            ws.cell(row=_er_row, column=c_e, value="1234")
            _e_addr = f"{get_column_letter(c_e)}{_er_row}"
            _log_exclude_rules_sheet_debug(
                "TEST_E1234",
                log_prefix,
                f'E列 {_e_addr} にテストで "1234" を書き込み',
                details=f"row={_er_row}",
            )
            logging.warning(
                '%s: 」テスト】%s に "1234" を書き込み（EXCLUDE_RULES_TEST_E1234）。',
                log_prefix,
                _e_addr,
            )

        _set_exclude_rules_snapshot_from_ws(
            wb_path, ws, c_proc, c_mach, c_flag, c_e
        )
        _write_exclude_rules_e_apply_artifacts(wb_path, ws, c_e, log_prefix)
        persisted = _persist_exclude_rules_workbook(wb, wb_path, ws, log_prefix)
        if not persisted:
            logging.warning(
                "%s: 設定シートの openpyxl 保存に失敗しました。"
                " log の行列 TSV をマクロ「設定_配台不要工程_AからE_TSVから反映」」"
                "または E 列のみ「設定_配台不要工程_E列_TSVから反映」で反映してください。",
                log_prefix,
            )
    except Exception as ex:
        _log_exclude_rules_sheet_debug(
            "FATAL",
            log_prefix,
            "設定シート処理中に未杕杉例外は発生しました。",
            exc=ex,
        )
        logging.exception("%s: 設定_配台不要工程の処理で例外", log_prefix)
    finally:
        if wb is not None:
            wb.close()
            _log_exclude_rules_sheet_debug("CLOSED", log_prefix, "ブックをクローズしました。")


def _resolve_exclude_rules_workbook_path_for_read(wb_path: str) -> str:
    """直後の保守で実効パスは変ゝったとし（通常は保存成功後の元ブック）にしれを使う。"""
    p = _exclude_rules_effective_read_path
    if p and os.path.exists(p):
        return p
    return wb_path


def _load_exclude_rules_from_workbook(wb_path: str) -> list[dict]:
    """シートからルール行を読み」評価用リストを返す。"""
    if not wb_path:
        return []
    global _exclude_rules_rules_snapshot, _exclude_rules_snapshot_wb
    ap_arg = os.path.normcase(os.path.abspath(wb_path))
    if (
        _exclude_rules_rules_snapshot is not None
        and _exclude_rules_snapshot_wb == ap_arg
    ):
        snap = list(_exclude_rules_rules_snapshot)
        _exclude_rules_rules_snapshot = None
        _exclude_rules_snapshot_wb = None
        return snap
    path = _resolve_exclude_rules_workbook_path_for_read(wb_path)
    if not os.path.exists(path):
        return []
    if _workbook_should_skip_openpyxl_io(path):
        logging.warning(
            "配台試行ルール: ブックに「%s」があるため、pandas(openpyxl) での「%s」読込をスキップしました（ルールは未適用）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            EXCLUDE_RULES_SHEET_NAME,
        )
        return []
    try:
        df = pd.read_excel(path, sheet_name=EXCLUDE_RULES_SHEET_NAME)
    except Exception:
        return []
    df.columns = df.columns.str.strip()
    need = [EXCLUDE_RULE_COL_PROCESS, EXCLUDE_RULE_COL_MACHINE]
    for c in need:
        if c not in df.columns:
            return []
    rules = []
    for _, row in df.iterrows():
        proc = str(row.get(EXCLUDE_RULE_COL_PROCESS, "") or "").strip()
        if not proc:
            continue
        mach = str(row.get(EXCLUDE_RULE_COL_MACHINE, "") or "").strip()
        c_val = row.get(EXCLUDE_RULE_COL_FLAG)
        e_raw = row.get(EXCLUDE_RULE_COL_LOGIC_JSON)
        parsed = _parse_exclude_rule_json_cell(e_raw)
        rules.append(
            {
                "proc": proc,
                "mach": mach,
                "c_val": c_val,
                "parsed": parsed,
            }
        )
    return rules


def apply_exclude_rules_config_to_plan_df(
    df: pd.DataFrame, wb_path: str, log_prefix: str
) -> pd.DataFrame:
    """設定シートに基づき「配台不要」を設定（C=yes または E の JSON が真）。

    工程名が「分割」の行は、同一依頼NO内に非空の同一「機械名」が複数行ある場合に限り
    （``_apply_auto_exclude_bunkatsu_duplicate_machine`` と同じ重複条件）、C/E を適用する。
    EC と分割で機械が異なる依頼では、設定行が残っていても当該分割行は配台対象のままとする。

    運用上は **段階1**（``run_stage1_extract``）から呼ぶ。段階2の ``load_planning_tasks_df`` では
    ``_apply_planning_sheet_post_load_mutations(..., apply_exclude_rules_from_config=False,
    compile_exclude_rules_d_to_e_with_ai=False)`` とし、本関数でシートの C/E を計画 DataFrame に
    再適用しない（設定シートの D→E AI も段階2では行わない）。
    """
    if df is None or df.empty:
        return df
    if TASK_COL_MACHINE not in df.columns or PLAN_COL_EXCLUDE_FROM_ASSIGNMENT not in df.columns:
        return df
    rules = _load_exclude_rules_from_workbook(wb_path)
    if not rules:
        return df
    df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT].astype(object)
    by_tid_idx: dict[str, list] = defaultdict(list)
    if TASK_COL_TASK_ID in df.columns:
        for j in df.index:
            tid_j = _normalize_task_id_for_dup_grouping(df.at[j, TASK_COL_TASK_ID])
            if tid_j:
                by_tid_idx[tid_j].append(j)
    n = 0
    for i in df.index:
        try:
            row = df.loc[i]
        except Exception:
            continue
        tp = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        tm = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        if not tp:
            continue
        tid_norm = _normalize_task_id_for_dup_grouping(row.get(TASK_COL_TASK_ID))
        is_bunkatsu = _process_name_is_bunkatsu_for_auto_exclude(tp)
        dup_ge2_for_tid = False
        if tid_norm:
            dup_ge2_for_tid, _mc = _same_tid_nonempty_machine_dup_ge2(
                df, by_tid_idx.get(tid_norm, [])
            )
        bunkatsu_block_cfg = is_bunkatsu and bool(tid_norm) and not dup_ge2_for_tid
        for ru in rules:
            if not _task_row_matches_exclude_rule_target(tp, tm, ru["proc"], ru["mach"]):
                continue
            if bunkatsu_block_cfg:
                continue
            if _exclude_rule_c_column_is_yes(ru["c_val"]):
                df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = "yes"
                n += 1
                break
            if ru.get("parsed") and evaluate_exclude_rule_json_for_row(ru["parsed"], row):
                df.at[i, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = "yes"
                n += 1
                break
    if n:
        logging.info("%s: 設定「%s」により配台不要=yes を %s 行に設定しました。", log_prefix, EXCLUDE_RULES_SHEET_NAME, n)
    return df


def _sort_stage1_plan_df_by_dispatch_trial_order_asc(plan_df: "pd.DataFrame") -> "pd.DataFrame":
    """
    段階1出力直後: 配台試行順番の昇順に行を並き替ごた DataFrame を返す。
    正の整数でないセルは最後（同一帯内は元の行順）。
    """
    col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    if plan_df is None or getattr(plan_df, "empty", True) or col not in plan_df.columns:
        return plan_df
    dto_positions = [i for i, c in enumerate(plan_df.columns) if c == col]
    if not dto_positions:
        return plan_df
    loc = dto_positions[0]
    n = len(plan_df)
    if n <= 1:
        return plan_df
    keys = []
    for i in range(n):
        dto = parse_optional_int(plan_df.iat[i, loc])
        if dto is not None and dto >= 1:
            keys.append((0, int(dto), i))
        else:
            keys.append((1, 10**9, i))
    order = sorted(range(n), key=lambda j: keys[j])
    if order == list(range(n)):
        return plan_df
    return plan_df.iloc[order].reset_index(drop=True)


# =============================================================================
# 段階1エントリ（task_extract_stage1.py → run_stage1_extract）
#   加工計画DATA 読取 → 計画 DataFrame 確定（マージ・分割の配台不要）→
#   設定シート保守（D→E の AI 含む）→ 設定を計画へ反映 → 配台試行順番 → plan_input_tasks.xlsx 出力
# =============================================================================
def run_stage1_extract():
    """
    段階1: 加工計画DATA から配台用タスク一覧を抽出し output/plan_input_tasks.xlsx へ出力。
    同一依頼NOで同一機械名が複数行あるとき、工程名「分割」行の空の「配台不要」に yes を自動設定する。
    マクロブックの「設定_配台不要工程」で工程+機械ごとの配台不要・条件式（AI）を管理する（シート作成は VBA）。
    設定シートの行同期および D 列→E 列（ロジック式）の AI 補完は、計画 DataFrame 確定後かつ
    「配台試行順番」の付与より前に行う。
    """
    if not TASKS_INPUT_WORKBOOK:
        logging.error("TASK_INPUT_WORKBOOK は未設定です。")
        return False
    if not os.path.exists(TASKS_INPUT_WORKBOOK):
        logging.error(f"TASK_INPUT_WORKBOOK は存在しません: {TASKS_INPUT_WORKBOOK}")
        return False
    reset_gemini_usage_tracker()
    df_src = load_tasks_df()
    records = []
    for _, row in df_src.iterrows():
        if row_has_completion_keyword(row):
            continue
        if _plan_row_exclude_as_completed_mikan_unprocessed_zero_actual_done_rule(row):
            continue
        task_id = planning_task_id_str_from_scalar(row.get(TASK_COL_TASK_ID))
        machine = str(row.get(TASK_COL_MACHINE, "")).strip()
        machine_name = str(row.get(TASK_COL_MACHINE_NAME, "")).strip()
        qty, _, _, _ = _plan_row_dispatch_qty_metrics(row)
        if qty <= 0 or not machine or not task_id:
            continue
        rec = {c: row.get(c) for c in SOURCE_BASE_COLUMNS}
        rec[TASK_COL_TASK_ID] = task_id
        _qty_total_s1 = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
        _qty_total_s1 = _floor_positive_m_to_planning_minimum(
            _qty_total_s1, PLANNING_MIN_QTY_M
        )
        if TASK_COL_QTY in rec:
            rec[TASK_COL_QTY] = _qty_total_s1
        rec[PLAN_COL_ROLL_UNIT_LENGTH] = _stage1_roll_length_for_planning_row(row)
        # 工程名 + 機械名 を“因孝”として表示用に追加（後段は計算キーにも使用）
        if machine_name:
            rec[PLAN_COL_PROCESS_FACTOR] = f"{machine}+{machine_name}"
        else:
            rec[PLAN_COL_PROCESS_FACTOR] = f"{machine}+"
        rec[PLAN_COL_SPEED_OVERRIDE] = ""
        rec[PLAN_COL_RAW_INPUT_DATE_OVERRIDE] = ""
        rec[PLAN_COL_PREFERRED_OP] = ""
        rec[PLAN_COL_SPECIAL_REMARK] = ""
        rec[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = ""
        rec[PLAN_COL_AI_PARSE] = ""
        records.append(rec)
    if not records:
        logging.warning("段階1: 抽出対象タスクはありません。")
    order = plan_input_sheet_column_order()
    out_df = pd.DataFrame(records)
    if out_df.empty:
        out_df = pd.DataFrame(columns=order)
    else:
        out_df = out_df.reindex(columns=order).fillna("")
    if PLAN_COL_EXCLUDE_FROM_ASSIGNMENT in out_df.columns:
        out_df[PLAN_COL_EXCLUDE_FROM_ASSIGNMENT] = out_df[
            PLAN_COL_EXCLUDE_FROM_ASSIGNMENT
        ].astype(object)
    if RESULT_TASK_COL_DISPATCH_TRIAL_ORDER in out_df.columns:
        out_df[RESULT_TASK_COL_DISPATCH_TRIAL_ORDER] = ""
        out_df[RESULT_TASK_COL_DISPATCH_TRIAL_ORDER] = out_df[
            RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
        ].astype(object)
    try:
        (
            _skills_d_stage1,
            _members_stage1,
            equipment_list_stage1,
            req_map,
            need_rules,
            _surplus_map_stage1,
            need_combo_col_index_stage1,
        ) = load_skills_and_needs()
    except PlanningValidationError:
        logging.error("段階1を中断: マスタ skills の検証エラー（優先度の数値重複など）。")
        raise
    except Exception as e:
        logging.info("段階1: マスタ need を読ゝう元列は need なしで埋ゝした (%s)", e)
        req_map, need_rules = {}, []
        equipment_list_stage1 = []
        need_combo_col_index_stage1 = {}
    out_df = _merge_plan_sheet_user_overrides(out_df)
    _apply_roll_unit_length_ceil_step_to_plan_df(out_df)
    _heal_stage1_roll_unit_no_dim_when_roll_matches_qty_mistake(out_df)
    _heal_stage1_roll_unit_if_width_ceiling_merge_spurious(out_df)
    _apply_roll_unit_length_ceil_step_to_plan_df(out_df)
    _refresh_plan_reference_columns(out_df, req_map, need_rules)
    try:
        _apply_auto_exclude_bunkatsu_duplicate_machine(out_df, log_prefix="段階1")
    except Exception as ex:
        logging.exception("段階1: 分割行の配台不要自動設定で例外（出力は続行）: %s", ex)
    # 設定_配台不要工程の行同期と D→E（AI）は、計画行集合確定後・配台試行順番付与より前に行う。
    try:
        _pm_pairs_s1 = _collect_process_machine_pairs_for_exclude_rules(out_df)
        run_exclude_rules_sheet_maintenance(
            TASKS_INPUT_WORKBOOK, _pm_pairs_s1, "段階1"
        )
    except Exception:
        logging.exception("段階1: 設定_配台不要工程の保守で例外（続行）")
    try:
        out_df = apply_exclude_rules_config_to_plan_df(out_df, TASKS_INPUT_WORKBOOK, "段階1")
    except Exception as ex:
        logging.warning("段階1: 設定シートによる配台試行適用で例外（続行）: %s", ex)
    try:
        _ext_dt_s1, _ = _extract_data_extraction_datetime()
        _run_d_s1 = _ext_dt_s1.date() if _ext_dt_s1 is not None else datetime.now().date()
        fill_plan_dispatch_trial_order_column_stage1(
            out_df,
            _run_d_s1,
            req_map,
            need_rules,
            need_combo_col_index_stage1,
            equipment_list_stage1,
        )
    except Exception as ex:
        logging.warning("段階1: 配台試行順番列の計算をスキップしました（続行）: %s", ex)
    out_df = _sort_stage1_plan_df_by_dispatch_trial_order_asc(out_df)
    _fill_plan_dispatch_remaining_qty_column(out_df)
    out_path = os.path.join(output_dir, STAGE1_OUTPUT_FILENAME)
    out_df.to_excel(out_path, sheet_name="タスク一覧", index=False)
    _apply_excel_date_columns_date_only_display(out_path, "タスク一覧")
    _apply_plan_input_visual_format(out_path, "タスク一覧")
    logging.info(f"段階1完了: '{out_path}' を出力しました。マクロで '{PLAN_INPUT_SHEET_NAME}' に坖り込んでしてさい。")
    _try_write_main_sheet_gemini_usage_summary("段階1")
    return True


# 稼働ルール（デフォルト値・2026年3月基準）
TARGET_YEAR = 2026
TARGET_MONTH = 3
DEFAULT_START_TIME = time(8, 45)
DEFAULT_END_TIME = time(17, 0)
DEFAULT_BREAKS = [
    (time(12, 0), time(12, 50)),
    (time(14, 45), time(15, 0))
]
# 終業直前デファー: ASSIGN_END_OF_DAY_DEFER_MINUTES が正のとき、team_end_limit までの残りがその分数以下で、
# かつ remaining_units（切り上げ）が ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS 以下のとき、その日の開始不可（None）。
# 同じウィンドウで「ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS ロール分以上は回せない」（収容が閾値未満）ときは
# 新規に加工を始めない（_eod_reject_capacity_units_below_threshold）。
# 占有キー上の直前加工が同一依頼NO（machine_handoff last_tid）のときは上記2点をスキップ（_eod_same_request_continuation_exempt）。
# ASSIGN_END_OF_DAY_DEFER_MINUTES 既定 45（分）。0 を明示すると無効（従来どおり）。
# ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS 既定 5。十分大きな値（例: 999999）にすると実質「残ロールに依らず終業直前は不可」。
# 休憩: 帯内に落ちた開始は _defer_team_start_past_prebreak_and_end_of_day で休憩終了へ繰り下げ。
# 休憩をまたぐ連続配台は _contiguous_work_minutes_until_next_break_or_limit で却下。
# （旧 ASSIGN_DEFER_MIN_REMAINING_ROLLS / ASSIGN_PRE_BREAK_DEFER_GAP_MINUTES は廃止・無視）
ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS = max(
    0,
    int(os.environ.get("ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS", "5").strip() or 0),
)
ASSIGN_END_OF_DAY_DEFER_MINUTES = max(
    0,
    int(os.environ.get("ASSIGN_END_OF_DAY_DEFER_MINUTES", "45").strip() or 0),
)


def _eod_minutes_window_covers_start(
    team_start: datetime, team_end_limit: datetime
) -> bool:
    """ASSIGN_END_OF_DAY_DEFER_MINUTES は正のとき」開始は終業上限のしの分数以内か。"""
    gap = ASSIGN_END_OF_DAY_DEFER_MINUTES
    if gap <= 0:
        return False
    if team_start >= team_end_limit:
        return False
    return (team_end_limit - team_start) <= timedelta(minutes=gap)


def _eod_same_request_continuation_exempt(
    machine_occ_key: str, task: dict, machine_handoff: dict | None
) -> bool:
    """
    同一設備占有キーで直前に載せた加工が同一依頼NO（task_id）のとき True。
    終業直前デファーは「新規開始」に寄せるため、この場合は小残・収容閾値の EOD 抑止を外す。
    """
    if not machine_handoff:
        return False
    occ = str(machine_occ_key or "").strip()
    if not occ:
        return False
    prev = (machine_handoff.get("last_tid") or {}).get(occ)
    cur = str(task.get("task_id") or "").strip()
    if not prev or not cur:
        return False
    return str(prev).strip() == cur


def _eod_reject_capacity_units_below_threshold(
    units_fit_until_close: int,
    team_start: datetime,
    team_end_limit: datetime,
    *,
    eod_same_request_continuation_exempt: bool = False,
) -> bool:
    """
    終業まであと ASSIGN_END_OF_DAY_DEFER_MINUTES 分以内のウィンドウ内で、
    ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS ロール分以上は回せない（収容ロール数が閾値未満）とき True（新規加工を始めない＝候補却下）。
    eod_same_request_continuation_exempt が True のときは常に False（同一依頼の連続ロール）。
    """
    if eod_same_request_continuation_exempt:
        return False
    th = ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS
    if th <= 0:
        return False
    if not _eod_minutes_window_covers_start(team_start, team_end_limit):
        return False
    return int(units_fit_until_close) < int(th)


# =========================================================
# 1. コア計算ロジック (日時ベース)
#    休憩帯を挟んて「実僝分」残作・終了時刻の繰り上き。割付ループの下回り。
# =========================================================
def merge_time_intervals(intervals):
    """時刻区間のリストをソートし、重なる区間を統合して返す。"""
    if not intervals:
        return []
    intervals.sort(key=lambda x: x[0])
    merged = [intervals[0]]
    for current_start, current_end in intervals[1:]:
        last_start, last_end = merged[-1]
        if current_start <= last_end:
            merged[-1] = (last_start, max(last_end, current_end))
        else:
            merged.append((current_start, current_end))
    return merged


def _contiguous_work_minutes_until_next_break_or_limit(
    start_dt: datetime,
    breaks_dt: list,
    end_limit_dt: datetime,
) -> int:
    """
    start_dt から次の休憩開始（または終業上限）までの」連続して実僝に使うる分数。
    開始は休憩帯内なら 0（呼び出し元で坴下）。breaks_dt は merge 済み想定。
    """
    if start_dt >= end_limit_dt:
        return 0
    for bs, be in breaks_dt:
        if bs <= start_dt < be:
            return 0
    next_stop = end_limit_dt
    for bs, be in breaks_dt:
        if be <= start_dt:
            continue
        if start_dt < bs:
            next_stop = min(next_stop, bs)
    return max(0, int((next_stop - start_dt).total_seconds() / 60))


def _break_end_to_skip_if_contiguous_under(
    start_dt: datetime,
    breaks_dt: list,
    end_limit_dt: datetime,
    min_contiguous_mins: int,
) -> datetime | None:
    """
    休憩帯外でも」次の休憩開始までの連続実僝は min_contiguous_mins 未満なら」
    しの休憩区間の終了時刻を返す（午後休憩直後に 1 ロール分は坎まらない開始の値進ゝる）。
    終業までしか実僝は続かない場合は None。
    """
    if min_contiguous_mins <= 0:
        return None
    if start_dt >= end_limit_dt:
        return None
    c = _contiguous_work_minutes_until_next_break_or_limit(
        start_dt, breaks_dt, end_limit_dt
    )
    if c >= min_contiguous_mins:
        return None
    next_stop = end_limit_dt
    for bs, be in breaks_dt:
        if be <= start_dt:
            continue
        if start_dt < bs:
            next_stop = min(next_stop, bs)
    if next_stop >= end_limit_dt:
        return None
    for bs, be in breaks_dt:
        if bs == next_stop:
            return be
    return None


def _defer_team_start_past_prebreak_and_end_of_day(
    task: dict,
    team: tuple,
    team_start: datetime,
    team_end_limit: datetime,
    team_breaks: list,
    refloor_fn,
    min_contiguous_work_mins: int | None = None,
    *,
    eod_same_request_continuation_exempt: bool = False,
) -> datetime | None:
    """
    - ASSIGN_END_OF_DAY_DEFER_MINUTES > 0 かつ (team_end_limit - 試行開始) がその分数以下で、
      remaining_units 切り上げが ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS 以下のとき、当日開始不可（None）。
      eod_same_request_continuation_exempt が True のときはこの終業直前・小残分岐をスキップ（同一依頼の連続ロール）。
    - 試行開始が休憩帯内のときは **休憩終了時刻へ繰り下げ**し、`refloor_fn` で設備下限・avail を再適用する。
      繰り下げのあと終業超過・EOD デファーに該当すれば None。
    - min_contiguous_work_mins が正のとき、帯外でも **次の休憩までの連続実働**がそれ未満なら
      当該休憩の終了へ繰り下げ（上と同様に refloor しループ）。
    """
    _tid = str(task.get("task_id", "") or "").strip()
    _team_txt = ", ".join(str(x) for x in team) if team else "—"

    def _trace_block(msg: str, *a) -> None:
        if not _trace_schedule_task_enabled(_tid):
            return
        _log_dispatch_trace_schedule(
            _tid,
            "[配台トレース task=%s] ブロック判定: " + msg,
            _tid,
            *a,
        )

    ts = refloor_fn(team_start)
    for _ in range(64):
        if ts >= team_end_limit:
            _trace_block(
                "開始試行(終業超靎) machine=%s team=%s rem=%.4f trial_start=%s end_limit=%s",
                task.get("machine"),
                _team_txt,
                float(task.get("remaining_units") or 0),
                ts,
                team_end_limit,
            )
            return None

        break_end = None
        for bs, be in team_breaks:
            if bs <= ts < be:
                break_end = be
                break
        if break_end is not None:
            _trace_block(
                "休憩帯内のため、終了へ繰り下き machine=%s team=%s rem=%.4f break_end=%s trial_was=%s",
                task.get("machine"),
                _team_txt,
                float(task.get("remaining_units") or 0),
                break_end,
                ts,
            )
            ts = refloor_fn(break_end)
            continue

        if min_contiguous_work_mins is not None and min_contiguous_work_mins > 0:
            slip_end = _break_end_to_skip_if_contiguous_under(
                ts, team_breaks, team_end_limit, min_contiguous_work_mins
            )
            if slip_end is not None:
                _trace_block(
                    "休憩直後で連続実僝丝足のため、休憩終了へ繰り下き machine=%s team=%s rem=%.4f need_contig_min=%s trial_was=%s break_end=%s",
                    task.get("machine"),
                    _team_txt,
                    float(task.get("remaining_units") or 0),
                    min_contiguous_work_mins,
                    ts,
                    slip_end,
                )
                ts = refloor_fn(slip_end)
                continue

        gap_end = ASSIGN_END_OF_DAY_DEFER_MINUTES
        rem_ceil = math.ceil(float(task.get("remaining_units") or 0))
        if (
            not eod_same_request_continuation_exempt
            and gap_end > 0
            and (team_end_limit - ts) <= timedelta(minutes=gap_end)
            and rem_ceil <= ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS
        ):
            _trace_block(
                "開始試行(終業直後・尝残ロール) machine=%s team=%s rem_ceil=%s max_rem=%s trial_start=%s end_limit=%s gap_end_min=%s",
                task.get("machine"),
                _team_txt,
                rem_ceil,
                ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS,
                ts,
                team_end_limit,
                gap_end,
            )
            return None

        return ts

    _trace_block(
        "開始試行(休憩繰り下き打切り) machine=%s team=%s rem=%.4f trial_start=%s",
        task.get("machine"),
        _team_txt,
        float(task.get("remaining_units") or 0),
        ts,
    )
    return None


def _expand_timeline_events_for_equipment_grid(timeline_events: list) -> list:
    """
    設備毎の時間割・メンバー日程・稼働率用インデックス坑け。
    1 本のイベントは日をまたし場合」e["date"] の値当日に載せると翌朝セグメントは欠けるため、
    start_dt〜end_dt を坄就業日 DEFAULT_START_TIME〜DEFAULT_END_TIME にクリップした複製へ展開れる。
    """
    expanded: list = []
    for e in timeline_events:
        sd = e.get("start_dt")
        ed = e.get("end_dt")
        if not isinstance(sd, datetime) or not isinstance(ed, datetime):
            expanded.append(e)
            continue
        if ed <= sd:
            expanded.append(e)
            continue
        segments: list = []
        cal = sd.date()
        last_d = ed.date()
        while cal <= last_d:
            w0 = datetime.combine(cal, DEFAULT_START_TIME)
            w1 = datetime.combine(cal, DEFAULT_END_TIME)
            a = max(sd, w0)
            b = min(ed, w1)
            if a < b:
                ne = dict(e)
                ne["date"] = cal
                ne["start_dt"] = a
                ne["end_dt"] = b
                segments.append(ne)
            cal += timedelta(days=1)
        if segments:
            expanded.extend(segments)
        else:
            expanded.append(e)
    return expanded


def get_actual_work_minutes(start_dt, end_dt, breaks_dt):
    """
    start_dt から end_dt までの「休憩を除いた実僝分数」。
    breaks_dt … (区間開始, 区間終了) の列（datetime または time。呼び出し元の勤怠イベントと整合）。
    """
    current = start_dt
    actual_mins = 0
    while current < end_dt:
        next_event = end_dt
        in_break = False
        b_end_time = None
        for b_s, b_e in breaks_dt:
            if b_s <= current < b_e:
                in_break = True
                b_end_time = b_e
                break
            elif current < b_s < next_event:
                next_event = b_s
        
        if in_break:
            current = b_end_time
        else:
            actual_mins += int((next_event - current).total_seconds() / 60)
            current = next_event
    return actual_mins


def calculate_end_time(start_dt, duration_minutes, breaks_dt, end_limit_dt):
    """
    start_dt から実僝 duration_minutes 分進ゝた終了 datetime を求ゝる（休憩はスキップ）。
    end_limit_dt を超ごないよご打ち切り。戻り値: (終了時刻, 実際に進ゝた実僝分, 残り未消化分)
    """
    current = start_dt
    remaining_work = duration_minutes
    actual_work_time = 0 

    while current < end_limit_dt and remaining_work > 0:
        next_event = end_limit_dt
        in_break = False
        break_end = None
        for b_start, b_end in breaks_dt:
            if b_start <= current < b_end:
                in_break = True
                break_end = b_end
                break
            elif current < b_start < next_event:
                next_event = b_start
                
        if in_break:
            current = break_end
            continue
            
        block_mins = int((next_event - current).total_seconds() / 60)
        if remaining_work <= block_mins:
            actual_work_time += remaining_work
            current += timedelta(minutes=remaining_work)
            remaining_work = 0
        else:
            actual_work_time += block_mins
            remaining_work -= block_mins
            current = next_event

    end_dt = min(current, end_limit_dt)
    return end_dt, actual_work_time, remaining_work


def _dt_close_minutes(a: datetime, b: datetime, tol_sec: int = 59) -> bool:
    return abs((a - b).total_seconds()) <= tol_sec


def _find_latest_prep_start_matching_end(
    end_at: datetime,
    dur_mins: int,
    breaks_merged: list,
    earliest_start: datetime,
) -> datetime | None:
    """
    実働 dur_mins 分を forward した終了が end_at になる最遅の開始時刻（なければ None）。
    breaks_merged は merge 済み休憩帯。分単位の探索＋念のための線形フォールバック。
    """
    if (
        dur_mins <= 0
        or not isinstance(end_at, datetime)
        or not isinstance(earliest_start, datetime)
    ):
        return None
    if end_at <= earliest_start:
        return None
    br = list(breaks_merged or [])
    cap = end_at + timedelta(days=2)
    e0, a0, r0 = calculate_end_time(earliest_start, dur_mins, br, cap)
    if r0 > 0 or a0 != dur_mins:
        return None
    if e0 > end_at and not _dt_close_minutes(e0, end_at):
        return None
    if _dt_close_minutes(e0, end_at):
        return earliest_start
    hi_i = max(0, int((end_at - earliest_start).total_seconds() // 60))
    lo_i = 0
    ans: datetime | None = None
    while lo_i <= hi_i:
        mid_i = (lo_i + hi_i) // 2
        s = earliest_start + timedelta(minutes=mid_i)
        if s > end_at:
            hi_i = mid_i - 1
            continue
        e, act, rem = calculate_end_time(s, dur_mins, br, cap)
        if rem != 0 or act != dur_mins:
            lo_i = mid_i + 1
            continue
        if e < end_at and not _dt_close_minutes(e, end_at):
            lo_i = mid_i + 1
        elif e > end_at and not _dt_close_minutes(e, end_at):
            hi_i = mid_i - 1
        else:
            ans = s
            lo_i = mid_i + 1
    if ans is not None:
        return ans
    for mid_i in range(hi_i, -1, -1):
        s = earliest_start + timedelta(minutes=mid_i)
        e, act, rem = calculate_end_time(s, dur_mins, br, cap)
        if rem == 0 and act == dur_mins and _dt_close_minutes(e, end_at):
            return s
    return None


def match_need_sheet_condition(condition_raw: str, task_id: str) -> bool:
    """
    need シート「依頼NO条件」欄の解釈。
    空・*・全件 → 常にマッポ。
    prefix:ABC / 接頭辞:ABC → 依頼NO はしの文字列で始まる
    regex:... / 正覝表睾:... → 正覝表睾（部分一致）
    しれ以外の短文は接頭辞として扱ご。従来の日本語例「依頼NOはJRで…」は JR を検出したら接頭辞JR扱い。
    """
    cond = (condition_raw or "").strip()
    tid = str(task_id).strip()
    if not cond or cond in ("*", "全件", "全で", "any", "ANY"):
        return True
    low = cond.lower()
    cn = cond.replace("：", ":")
    if low.startswith("prefix:") or low.startswith("接頭辞:"):
        pref = cn.split(":", 1)[1].strip() if ":" in cn else ""
        return bool(pref) and tid.startswith(pref)
    if low.startswith("regex:") or low.startswith("正覝表睾:"):
        pat = cn.split(":", 1)[1].strip() if ":" in cn else ""
        if not pat:
            return False
        try:
            return re.search(pat, tid) is not None
        except re.error:
            logging.warning(f"need 依頼NO条件の正覝表睾は無効です: {pat}")
            return False
    if "依頼" in cond and "JR" in cond.upper():
        return tid.upper().startswith("JR")
    return tid.startswith(cond)


def parse_need_sheet_special_rules(needs_df, label_col, equipment_list, cond_col):
    """特別指定1～99 行から」設備別の必須人数上書き（1～99）を抽出（先に定義された番坷は優先）。"""
    rules = []
    for _, row in needs_df.iterrows():
        lab = str(row.get(label_col, "") or "").strip()
        m = re.match(r"特別指定\s*(\d+)", lab)
        if not m:
            continue
        order = int(m.group(1))
        if order < 1 or order > 99:
            continue
        cond = ""
        if cond_col is not None:
            v = row.get(cond_col)
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                cond = str(v).strip()
        overrides = {}
        for eq in equipment_list:
            v = row.get(eq)
            n = parse_optional_int(v)
            if n is not None and 1 <= n <= 99:
                overrides[str(eq).strip()] = n
        if not overrides:
            continue
        rules.append({"order": order, "condition": cond, "overrides": overrides})
    rules.sort(key=lambda r: r["order"])
    return rules


def resolve_need_required_op(process: str, machine_name: str, task_id: str, req_map: dict, need_rules: list) -> int:
    """
    need シートの「工程名 + 機械名」で必須OP人数を解決（特別指定1〜99は order は尝さいろど優先）。

    req_map は
      - f\"{process}+{machine_name}\"（厳密キー）
      - machine_name（機械の値のフォールバック）
      - process（工程の値のフォールバック）
    のいうれかで base を引ける剝杝。
    need_rules の overrides も同様にキーを挝つ。
    """
    p = str(process).strip()
    m = str(machine_name).strip()

    combo_key = f"{p}+{m}" if p and m else None

    base = None
    if combo_key and combo_key in req_map:
        base = req_map.get(combo_key)
    if base is None and m:
        base = req_map.get(m)
    if base is None and p:
        base = req_map.get(p)
    if base is None:
        base = 1

    for rule in need_rules:
        if not match_need_sheet_condition(rule["condition"], task_id):
            continue

        if combo_key and combo_key in rule["overrides"]:
            return int(rule["overrides"][combo_key])
        if m and m in rule["overrides"]:
            return int(rule["overrides"][m])
        if p and p in rule["overrides"]:
            return int(rule["overrides"][p])

    return int(base)


def resolve_need_required_op_explain(
    process: str, machine_name: str, task_id: str, req_map: dict, need_rules: list
) -> tuple[int, str]:
    """
    resolve_need_required_op と同値を返しつつ」ログ用に参照元の説明文字列を付ける。
    """
    p = str(process).strip()
    m = str(machine_name).strip()
    combo_key = f"{p}+{m}" if p and m else None
    base = None
    base_src = ""
    if combo_key and combo_key in req_map:
        base = req_map.get(combo_key)
        base_src = f"req_map[{combo_key!r}]={base}"
    elif m and m in req_map:
        base = req_map[m]
        base_src = f"req_map[機械名のみ {m!r}]={base}（複坈キー丝在）"
    elif p and p in req_map:
        base = req_map[p]
        base_src = f"req_map[工程名のみ {p!r}]={base}（複坈・機械キー丝在）"
    else:
        base = 1
        base_src = "req_map該当なし→既定1"
    for rule in need_rules:
        if not match_need_sheet_condition(rule["condition"], task_id):
            continue
        order = rule.get("order", "?")
        if combo_key and combo_key in rule["overrides"]:
            v = int(rule["overrides"][combo_key])
            return v, f"need特別指定{order} [{combo_key!r}]={v}"
        if m and m in rule["overrides"]:
            v = int(rule["overrides"][m])
            return v, f"need特別指定{order} [機械名{m!r}]={v}"
        if p and p in rule["overrides"]:
            v = int(rule["overrides"][p])
            return v, f"need特別指定{order} [工程名{p!r}]={v}"
    return int(base), base_src


def _need_row_label_hints_surplus_add(label_a0: str) -> bool:
    """need シート A列: 基本必須人数の直下にある「配台結果で余剰は出たとしの追加増員上限」行か。"""
    s = unicodedata.normalize("NFKC", str(label_a0 or "").strip())
    if not s or s.startswith("特別指定"):
        return False
    if "依頼" in s and "条件" in s:
        return False
    if "追加" in s and ("人数" in s or "人員" in s or "増員" in s):
        return True
    if "増員" in s or "余剰" in s:
        return True
    if "配台" in s and ("追加" in s or "増" in s or "余剰" in s):
        return True
    return False


def _find_need_surplus_add_row_index(
    needs_raw, base_row: int, col0: int, pm_cols: list
) -> int | None:
    """基本必須人数行の次行を優先。ラベルまたは数値で追加人数行と判定。"""
    r = base_row + 1
    if r >= needs_raw.shape[0]:
        return None
    v0 = needs_raw.iat[r, col0]
    s0 = "" if pd.isna(v0) else str(v0).strip()
    if s0.startswith("特別指定"):
        return None
    if _need_row_label_hints_surplus_add(s0):
        return r
    nz = 0
    for col_idx, _, _ in pm_cols:
        if parse_optional_int(needs_raw.iat[r, col_idx]) is not None:
            nz += 1
    if nz > 0 and not unicodedata.normalize("NFKC", s0).startswith("特別"):
        return r
    return None


def resolve_need_surplus_extra_max(
    process: str,
    machine_name: str,
    task_id: str,
    surplus_map: dict,
    need_rules: list,
) -> int:
    """
    need シート「配台時追加人数」行（工程×機械列）の値＝必須人数を満たしたごごで
    さらに割り当で可能な人数の上限（0 なら従来どおり必須人数うょごどのみ）。
    need_rules は睾状この行を上書きしない（将来拡張用に task_id を块け得る）。
    """
    _ = (task_id, need_rules)
    if not surplus_map:
        return 0
    p = str(process).strip()
    m = str(machine_name).strip()
    combo_key = f"{p}+{m}" if p and m else None
    v = None
    if combo_key and combo_key in surplus_map:
        v = surplus_map[combo_key]
    elif m and m in surplus_map:
        v = surplus_map[m]
    elif p and p in surplus_map:
        v = surplus_map[p]
    if v is None:
        return 0
    try:
        n = int(v)
    except (TypeError, ValueError):
        return 0
    return max(0, min(n, 50))


def resolve_need_surplus_extra_max_explain(
    process: str,
    machine_name: str,
    task_id: str,
    surplus_map: dict,
    need_rules: list,
) -> tuple[int, str]:
    """resolve_need_surplus_extra_max と同値＋参照元説明（ログ用）。"""
    val = resolve_need_surplus_extra_max(
        process, machine_name, task_id, surplus_map, need_rules
    )
    _ = need_rules
    if not surplus_map:
        return val, "surplus_map空（配台時追加人数行なし）"
    p = str(process).strip()
    m = str(machine_name).strip()
    combo_key = f"{p}+{m}" if p and m else None
    if combo_key and combo_key in surplus_map:
        raw = surplus_map[combo_key]
        return val, f"surplus_map[{combo_key!r}]={raw}"
    if m and m in surplus_map:
        raw = surplus_map[m]
        return val, f"surplus_map[機械名のみ {m!r}]={raw}（複坈キー丝在）"
    if p and p in surplus_map:
        raw = surplus_map[p]
        return val, f"surplus_map[工程名のみ {p!r}]={raw}（複坈キー丝在）"
    return val, "surplus当キーなし→0"


def _surplus_team_time_factor(
    rq_base: int, team_len: int, extra_max_allowed: int
) -> float:
    """
    必須人数を超ごで入れたメンバーによる短縮時間への係数（1.0＝短縮なし）。
    追加枠（extra_max_allowed）を使い切ったとしでも」短縮は SURPLUS_TEAM_MAX_SPEEDUP_RATIO を上限とれる線形モデル。
    """
    rq = max(1, int(rq_base))
    tl = int(team_len)
    em = max(0, int(extra_max_allowed))
    extra = max(0, tl - rq)
    if extra <= 0 or em <= 0:
        return 1.0
    frac = min(1.0, extra / float(em))
    return 1.0 - SURPLUS_TEAM_MAX_SPEEDUP_RATIO * frac


def _team_assign_trace_tuple_label() -> str:
    if TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF:
        return "(-人数, 開始, -短縮数, 優先度合計)"
    if TEAM_ASSIGN_START_SLACK_WAIT_MINUTES <= 0:
        return "(開始, -短縮数, 優先度合計)"
    return (
        f"最早開始から{TEAM_ASSIGN_START_SLACK_WAIT_MINUTES}分以内は"
        "(0,-人数,開始,-短縮数,優先度)」超靎は(1,開始,-人数,-短縮数,優先度)"
    )


def _team_assignment_sort_tuple(
    team: tuple,
    team_start: datetime,
    units_today: int,
    team_prio_sum: int,
    t_min: datetime | None = None,
) -> tuple:
    """
    フォーム候補の優劣用タプル（辞書式で尝さい方は採用）。
    - TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF: (-人数, 開始, -短縮数, 優先度合計)
    - しれ以外かつ TEAM_ASSIGN_START_SLACK_WAIT_MINUTES>0 かつ t_min あり:
        最早開始からスラック以内 → (0, -人数, 開始, -短縮数, 優先度) … 遅れでも人数を厚し
        スラック超 → (1, 開始, -人数, -短縮数, 優先度) … 開始を優先
    - 上記以外: (開始, -短縮数, 優先度合計)
    """
    n = len(team)
    if TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF:
        return (-n, team_start, -units_today, team_prio_sum)
    sm = TEAM_ASSIGN_START_SLACK_WAIT_MINUTES
    if sm <= 0 or t_min is None:
        return (team_start, -units_today, team_prio_sum)
    sl = timedelta(minutes=sm)
    if team_start - t_min <= sl:
        return (0, -n, team_start, -units_today, team_prio_sum)
    return (1, team_start, -n, -units_today, team_prio_sum)


# skills セル: OP / AS + 任愝の優先度整数（例 OP1, AS 3）。数値は尝さいろど割当で先に選ばれる。
_SKILL_OP_AS_CELL_RE = re.compile(r"^(OP|AS)(\d*)$", re.IGNORECASE)


def parse_op_as_skill_cell(cell_val):
    """
    master.xlsm「skills」のセル1つを解釈れる。
    - 「OP」または「AS」の直後に優先度用の整数（空白は除去して解釈）。例: OP, OP1, AS3, AS 12
    - 優先度は尝さいろど高優先（同一条件のフォーム候補から先に選ばれる）。数字省略時は 1。
    - OP/AS で始まらない・空はスキルなし。
    """
    if cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val)):
        return None, 10**9
    s = str(cell_val).strip()
    if not s:
        return None, 10**9
    compact = re.sub(r"\s+", "", s).upper()
    m = _SKILL_OP_AS_CELL_RE.match(compact)
    if not m:
        return None, 10**9
    role = m.group(1).upper()
    tail = m.group(2) or ""
    if tail == "":
        pr = 1
    else:
        try:
            pr = int(tail)
        except ValueError:
            return None, 10**9
    if pr < 0:
        pr = 0
    return role, pr


def _validate_skills_op_as_priority_numbers_unique(
    skills_dict: dict, column_keys: list
) -> None:
    """
    master「skills」の複数列（工程+機械キー等）についで」OP/AS の割当優先度の**数値**は
    メンバー間で重複していないか検証れる。重複時は PlanningValidationError。
    （OP1 と AS1 のよごにロールは異なっても同一数値なら重複とみなす）
    """
    errors: list[str] = []
    for combo in column_keys:
        ck = str(combo or "").strip()
        if not ck:
            continue
        pr_to_entries: dict[int, list[str]] = defaultdict(list)
        for mem, row in (skills_dict or {}).items():
            mnm = str(mem or "").strip()
            if not mnm or not isinstance(row, dict):
                continue
            raw = row.get(ck)
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                continue
            sval = str(raw).strip()
            if not sval or sval.lower() in ("nan", "none", "null"):
                continue
            role, pr = parse_op_as_skill_cell(sval)
            if role not in ("OP", "AS"):
                continue
            pr_to_entries[int(pr)].append(f"{mnm}({role})")
        for pr, entries in sorted(pr_to_entries.items()):
            if len(entries) > 1:
                errors.append(f'列「{ck}」: 優先度 {pr} は重複 → ' + "」".join(entries))
    if errors:
        cap = 50
        tail = errors[:cap]
        msg = (
            "マスタ「skills」で」同一列の OP/AS 優先度の数値は重複していした。"
            " 列ごとに数値は1人につし1種類にしてください。\n"
            + "\n".join(tail)
        )
        if len(errors) > cap:
            msg += f"\n…他 {len(errors) - cap} 件"
        raise PlanningValidationError(msg)


def build_member_assignment_priority_reference(
    skills_dict: dict,
    members: list | None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    結果ブック用: マスタ skills の「工程名+機械名」列ごとに」割当アルゴリズムとともに
    (優先度値昇順, メンバー坝昇順) で並きた参考表と」ルール説明の表を返す。
    当日の出勤・設備空し・同一依頼の工程順・フォーム人数は反映しない（あしまでマスタ上の順庝）。
    """
    mem_list = list(members) if members else list((skills_dict or {}).keys())
    mem_list = [str(m).strip() for m in mem_list if m and str(m).strip()]

    surplus_on = bool(TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF)
    slack_m = TEAM_ASSIGN_START_SLACK_WAIT_MINUTES
    if surplus_on:
        team_rule = (
            "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF=有効: "
            "(-人数, 開始, -短縮数, 優先度合計) の辞書式（人数最優先・従来）。"
        )
    elif slack_m > 0:
        team_rule = (
            f"既定: しの日の成立候補全体の「最早開始」を基準に」"
            f"開始はしの{slack_m}分以内の遅れなら人数を厚し優先（0,-人数,開始,-短縮数,優先度）」"
            f"しれより靅い候補は開始を優先（1,開始,-人数,-短縮数,優先度）。"
            f"環境変数 TEAM_ASSIGN_START_SLACK_WAIT_MINUTES=0 で無効化。"
        )
    else:
        team_rule = (
            "TEAM_ASSIGN_START_SLACK_WAIT_MINUTES=0: "
            "(開始, -短縮数, 優先度合計) のみ（開始最優先）。"
        )

    legend_rows = [
        {
            "区分": "スキル列の並よ",
            "内容": "坄「工程名+機械名」列についで」セルは OP/AS（+優先度整数）のメンバーのみ対象。"
            " 数値は尝さいろど高優先。省略時は優先度 1（parse_op_as_skill_cell と同一）。"
            " 同一列では優先度の数値はメンバー間で重複試行（マスタ読込時に検証）。",
        },
        {
            "区分": "当日との差",
            "内容": "実際の配台は」この順のごうしの日出勤かつ AS/OP 覝件を満たれ者の値は候補。"
            " 設備の空し・同一依頼NOの工程順・必須人数・増員枠・指定OPで変ゝりした。",
        },
        {
            "区分": "フォーム候補の比較",
            "内容": team_rule,
        },
        {
            "区分": "指定・グローバル上書き",
            "内容": "担当OP_指定・メイン「再優先特別記載」の OP 指定は本表より優先されした。",
        },
        {
            "区分": "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF",
            "内容": "1/有効（人数最優先・従来）" if surplus_on else "0/無効（既定）",
        },
        {
            "区分": "TEAM_ASSIGN_START_SLACK_WAIT_MINUTES",
            "内容": str(slack_m),
        },
    ]
    df_legend = pd.DataFrame(legend_rows)

    combo_keys: set[str] = set()
    for m in mem_list:
        row = (skills_dict or {}).get(m) or {}
        for k in row:
            ks = str(k).strip()
            if "+" in ks:
                combo_keys.add(ks)
    sorted_combos = sorted(combo_keys)

    out: list[dict] = []
    for combo in sorted_combos:
        parts = combo.split("+", 1)
        proc = parts[0].strip()
        mach = parts[1].strip() if len(parts) > 1 else ""
        ranked: list[tuple[int, str, str, str]] = []
        for m in sorted(mem_list):
            cell = ((skills_dict or {}).get(m) or {}).get(combo)
            if cell is None or (isinstance(cell, float) and pd.isna(cell)):
                cell_s = ""
            else:
                cell_s = str(cell).strip()
            role, pr = parse_op_as_skill_cell(cell_s if cell_s else None)
            if role in ("OP", "AS"):
                ranked.append((pr, m, role, cell_s))
        ranked.sort(key=lambda x: (x[0], x[1]))
        if not ranked:
            out.append(
                {
                    "工程名": proc,
                    "機械名": mach,
                    "スキル列キー": combo,
                    "優先順佝": "",
                    "メンバー": "（なし）",
                    "ロール": "",
                    "優先度値_尝さいろど先": "",
                    "skillsセル値": "",
                    "備考": "この列に OP/AS の資格セルはあるメンバーはいません",
                }
            )
            continue
        for i, (pr, m, role, cell_s) in enumerate(ranked, start=1):
            out.append(
                {
                    "工程名": proc,
                    "機械名": mach,
                    "スキル列キー": combo,
                    "優先順佝": i,
                    "メンバー": m,
                    "ロール": role,
                    "優先度値_尝さいろど先": pr,
                    "skillsセル値": cell_s,
                    "備考": "",
                }
            )

    df_tbl = pd.DataFrame(out)
    return df_legend, df_tbl


def _normalize_person_name_for_match(s):
    """担当者指定のあいまい一致用（NFKC・富田/冨田の表記寄せ・空白除去・末尾敬称のみ除去）。"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    if "富田" in t:
        t = t.replace("富田", "冨田")
    t = re.sub(r"[\s　]+", "", t)
    t = re.sub(r"(さん|様|氝)$", "", t)
    return t


def _split_person_sei_mei(s) -> tuple[str, str]:
    """
    並びを姓・坝に分ける。最初の半角＝全角空白の手剝を姓」以降を坝とれる。
    空白は無い場合は (全体, '')（坝なし扱い）。
    末尾の さん＝様＝氝 は分割剝に除去れる。
    """
    if s is None:
        return "", ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    if not t or t.lower() in ("nan", "none", "null"):
        return "", ""
    t = re.sub(r"(さん|様|氝)$", "", t)
    for i, ch in enumerate(t):
        if ch in " \u3000":
            sei = t[:i].strip()
            rest = t[i + 1 :]
            mei = re.sub(r"[\s　]+", "", rest.strip())
            return sei, mei
    return t.strip(), ""


def _normalize_sei_for_match(sei: str) -> str:
    """姓のみ正規化。表記ゆれは許容しない剝杝で」NFKC・富田/冨田寄せ・空白除去。"""
    if not sei:
        return ""
    t = unicodedata.normalize("NFKC", str(sei).strip())
    if "富田" in t:
        t = t.replace("富田", "冨田")
    t = re.sub(r"[\s　]+", "", t)
    return t


def _normalize_mei_for_match(mei: str) -> str:
    """坝の正規化（ゆれ許容の剝処理）。NFKC・空白除去。姓用の富田置杛は行ゝない。"""
    if not mei:
        return ""
    t = unicodedata.normalize("NFKC", str(mei).strip())
    t = re.sub(r"[\s　]+", "", t)
    return t


def _has_duplicate_surname_among_members(member_names) -> bool:
    """skills メンバー一覧に」正規化後同一の姓は2人以上いるか。"""
    cnt = Counter()
    for name in member_names or []:
        if name is None or (isinstance(name, float) and pd.isna(name)):
            continue
        s = str(name).strip()
        if not s:
            continue
        sei, _mei = _split_person_sei_mei(s)
        key = _normalize_sei_for_match(sei)
        if key:
            cnt[key] += 1
    return any(n >= 2 for n in cnt.values())


def _mei_matches_with_fuzzy_allowed(r_mei_n: str, m_mei_n: str) -> bool:
    """同一姓はロスターで重複しないとしのみ使う坝のゆれ許容。"""
    if not r_mei_n and not m_mei_n:
        return True
    if not r_mei_n or not m_mei_n:
        return False
    if r_mei_n == m_mei_n:
        return True
    return r_mei_n in m_mei_n or m_mei_n in r_mei_n


def _resolve_preferred_name_to_capable_member(raw, capable_candidates, roster_member_names=None):
    """
    自由記述の指定を」当日スキル上 OP/AS のメンバー坝（skills シートの行キー）に解決れる。
    capable_candidates: しの設備で OP または AS として割当可能なメンバー坝リスト。
    roster_member_names: skills の全メンバー坝（省略時は capable_candidates）。同一姓の重複判定に使用。

    坝剝の表記ゆれ:
    - 姓は正規化後に完全一致のみ（ゆれ許容しない。富田/冨田のみ従来どおり寄せ）。
    - roster に同一姓は2人以上いないとしの値」坝は部分一致（どうらかは他方を含む）または完全一致を許容。
    - 同一姓はロスターにいる間は坝も完全一致必須。
    - 姓のみの入力で坝ゆれモードのとき」姓は一致する候補は複数いれみ解決試行（None）。
    """
    if not raw or not capable_candidates:
        return None
    r0 = unicodedata.normalize("NFKC", str(raw).strip())
    r = _normalize_person_name_for_match(r0)
    if not r:
        return None
    for m in capable_candidates:
        if _normalize_person_name_for_match(m) == r:
            return m
        if unicodedata.normalize("NFKC", str(m).strip()) == r0.strip():
            return m

    roster = (
        list(roster_member_names)
        if roster_member_names is not None
        else list(capable_candidates)
    )
    allow_mei_fuzzy = not _has_duplicate_surname_among_members(roster)

    r_sei, r_mei = _split_person_sei_mei(raw)
    r_sei_n = _normalize_sei_for_match(r_sei)
    r_mei_n = _normalize_mei_for_match(r_mei)
    if not r_sei_n:
        return None

    matches = []
    for m in capable_candidates:
        m_sei, m_mei = _split_person_sei_mei(m)
        m_sei_n = _normalize_sei_for_match(m_sei)
        m_mei_n = _normalize_mei_for_match(m_mei)
        if r_sei_n != m_sei_n:
            continue
        if allow_mei_fuzzy:
            if r_mei_n:
                if _mei_matches_with_fuzzy_allowed(r_mei_n, m_mei_n):
                    matches.append(m)
            else:
                matches.append(m)
        else:
            if r_mei_n == m_mei_n:
                matches.append(m)

    if len(matches) == 1:
        return matches[0]
    return None


def _resolve_preferred_op_to_member(raw, op_candidates, roster_member_names=None):
    """当日スキル上 OP のみへ解決（従来 API）。実体は `_resolve_preferred_name_to_capable_member`。"""
    return _resolve_preferred_name_to_capable_member(
        raw, op_candidates, roster_member_names
    )


def _task_process_matches_global_contains(machine_val: str, contains: str) -> bool:
    """工程名（タスクの machine）に部分一致（NFKC・大尝無視）。"""
    m = unicodedata.normalize("NFKC", str(machine_val or "").strip()).casefold()
    c = unicodedata.normalize("NFKC", str(contains or "").strip()).casefold()
    if not c:
        return False
    return c in m


def _coerce_global_day_process_operator_rules(raw_val) -> list:
    """Gemini の global_day_process_operator_rules を正規化（空・正常は除外）。"""
    out: list[dict] = []
    if not isinstance(raw_val, list):
        return out
    seen_sig = set()
    for item in raw_val:
        if not isinstance(item, dict):
            continue
        d = parse_optional_date(item.get("date"))
        if d is None:
            continue
        pc = item.get("process_contains")
        if pc is None or (isinstance(pc, float) and pd.isna(pc)):
            continue
        pc_s = unicodedata.normalize("NFKC", str(pc).strip())
        if not pc_s:
            continue
        names = item.get("operator_names")
        if not isinstance(names, list):
            continue
        op_names: list[str] = []
        for n in names:
            if n is None or (isinstance(n, float) and pd.isna(n)):
                continue
            s = str(n).strip()
            if s and s.lower() not in ("nan", "none", "null"):
                op_names.append(s)
        if not op_names:
            continue
        sig = (d.isoformat(), pc_s.casefold(), tuple(op_names))
        if sig in seen_sig:
            continue
        seen_sig.add(sig)
        out.append(
            {
                "date": d.isoformat(),
                "process_contains": pc_s,
                "operator_names": op_names,
            }
        )
    return out


def _active_global_day_process_must_include(
    gpo: dict,
    task: dict,
    current_date: date,
    capable_members: list,
    roster_members: list,
) -> tuple[list[str], list[str]]:
    """
    グローバルコメント由来の「日付×工程×複数指定」で」しの日・しの工程タスクに
    **フォームへ必う含むる**メンバー（skills 行キー）と警告メッセージを返す。
    """
    rules = gpo.get("global_day_process_operator_rules") or []
    if not isinstance(rules, list):
        return [], []
    machine = task.get("machine")
    warns: list[str] = []
    acc: list[str] = []
    seen_m: set[str] = set()
    tid = str(task.get("task_id") or "").strip()
    for rule in rules:
        if not isinstance(rule, dict):
            continue
        rd = parse_optional_date(rule.get("date"))
        if rd is None or rd != current_date:
            continue
        pc = rule.get("process_contains") or ""
        pcn = unicodedata.normalize("NFKC", str(pc).strip())
        if not pcn or not _task_process_matches_global_contains(machine, pcn):
            continue
        for raw_name in rule.get("operator_names") or []:
            mem = _resolve_preferred_name_to_capable_member(
                raw_name, capable_members, roster_members
            )
            if mem:
                if mem not in seen_m:
                    seen_m.add(mem)
                    acc.append(mem)
            else:
                warns.append(
                    "メイングローバル(日付×工程)指定: "
                    f"依頼NO={tid} 日付={current_date} 工程={machine!r} の "
                    f"指定「{raw_name}」を当日スキル該当メンバーに解決でしません"
                )
    return acc, warns


def _merge_global_day_process_and_pref_anchor(
    must_include: list, pref_mem, capable_members: list
) -> list[str]:
    """必須メンバーと担当OP指定を1本化（capable にいるものの値）。"""
    fixed: list[str] = []
    seen: set[str] = set()
    for m in must_include or []:
        if m in capable_members and m not in seen:
            seen.add(m)
            fixed.append(m)
    if (
        pref_mem
        and pref_mem in capable_members
        and pref_mem not in seen
    ):
        fixed.append(pref_mem)
    return fixed


# =========================================================
# 2. マスタデータ・出勤簿(カレンダー) と AI解析
#    master.xlsm の skills / need / 坄メンバー勤怠シートを読み」
#    備考・休暇区分は必須に応もで Gemini で構造化れる。
# =========================================================
def load_skills_and_needs():
    """
    統合ファイル(MASTER_FILE)からスキルと need を動的に読み込みした。

    戻り値は7覝素。最後は need シート上の「工程名+機械名」列佝置（左ろど尝さい整数）の辞書
    ``need_combo_col_index``（配台キューソート用）。

    今回の need は（Excel上で）
      工程名行・機械名行のあと「基本必須人数」行（A列に「必須人数」を含む）
      しの直下: 配台で余剰人員はあるとしに追加で入れられる人数（工程×機械とと。未設定は 0）
      以降: 特別指定1〜99
    といご構造のため、必須OPは「工程名+機械名」で解決れる。

    skills 交差セルは OP/AS の後に優先度整数（例 OP1, AS3）。数値は尝さいろど当該工程への割当で優先。
    数字省略の OP/AS は優先度 1。
    同一列（同一工程×機械）では優先度の数値はメンバー間で重複試行（重複時は PlanningValidationError）。
    """
    try:
        # 同一ブックを pd.read_excel で都度開しと I/O は重いめ、ExcelFile を1回の値開いでシートを parse れる。
        with pd.ExcelFile(MASTER_FILE) as _master_xls:
            # skills は新仕様:
            #   1行目: 工程名
            #   2行目: 機械名
            #   A3以降: メンバー坝
            #   交差セル: OP または AS の後に割当優先度の整数（例 OP1, AS3）。数値は尝さいろど当該工程へ優先割当。
            #             数字省略の OP/AS は優先度 1（従来どおり最優先扱い）。
            # を基本としつつ」旧仕様（1行ヘッダ）にもフォールバック対応れる。
            skills_raw = pd.read_excel(
                _master_xls, sheet_name="skills", header=None
            )
            skills_dict = {}
            equipment_list = []
            members = []

            use_two_header = False
            if skills_raw.shape[0] >= 3 and skills_raw.shape[1] >= 2:
                non_empty_pm = 0
                for c in range(1, skills_raw.shape[1]):
                    p = skills_raw.iat[0, c]
                    m = skills_raw.iat[1, c]
                    if pd.isna(p) or pd.isna(m):
                        continue
                    p_s = str(p).strip()
                    m_s = str(m).strip()
                    if p_s and m_s and p_s.lower() != "nan" and m_s.lower() != "nan":
                        non_empty_pm += 1
                use_two_header = non_empty_pm > 0

            if use_two_header:
                pm_cols = []
                seen_combo = set()
                for c in range(1, skills_raw.shape[1]):
                    p = skills_raw.iat[0, c]
                    m = skills_raw.iat[1, c]
                    if pd.isna(p) or pd.isna(m):
                        continue
                    p_s = str(p).strip()
                    m_s = str(m).strip()
                    if not p_s or not m_s or p_s.lower() == "nan" or m_s.lower() == "nan":
                        continue
                    combo = f"{p_s}+{m_s}"
                    pm_cols.append((c, p_s, m_s, combo))
                    if combo not in seen_combo:
                        seen_combo.add(combo)
                        equipment_list.append(combo)

                for r in range(2, skills_raw.shape[0]):
                    m_name_raw = skills_raw.iat[r, 0]
                    if pd.isna(m_name_raw):
                        continue
                    m_name = str(m_name_raw).strip()
                    if not m_name or m_name.lower() in ("nan", "none", "null"):
                        continue
                    row_skills = {}
                    for c, p_s, m_s, combo in pm_cols:
                        v = skills_raw.iat[r, c] if c < skills_raw.shape[1] else None
                        sval = "" if pd.isna(v) else str(v).strip()
                        if not sval or sval.lower() in ("nan", "none", "null"):
                            continue
                        row_skills[combo] = sval
                        if m_s not in row_skills:
                            row_skills[m_s] = sval
                        if p_s not in row_skills:
                            row_skills[p_s] = sval
                    skills_dict[m_name] = row_skills
                members = list(skills_dict.keys())
                logging.info(
                    "skillsシート: 2段ヘッダ形式で読み込みました（工程+機械=%s列, メンバー=%s人）。",
                    len(pm_cols),
                    len(members),
                )
            else:
                skills_df = pd.read_excel(_master_xls, sheet_name="skills")
                skills_df.columns = skills_df.columns.str.strip()
                skill_cols = [
                    str(c).strip()
                    for c in skills_df.columns
                    if not str(c).startswith("Unnamed")
                ]

                member_col = None
                for c in skill_cols:
                    if c in ("メンバー", "担当者", "並び", "作業者"):
                        member_col = c
                        break
                if member_col is None and skill_cols:
                    member_col = skill_cols[0]
                    logging.warning(
                        "skillsシート: メンバー列名は標準と一致しないため、先頭列 '%s' をメンバー列として扱いした。",
                        member_col,
                    )

                seen_eq = set()
                for c in skill_cols:
                    if c == member_col:
                        continue
                    cid = str(c).strip()
                    if not cid or cid.lower() in ("nan", "none", "null"):
                        continue
                    if cid not in seen_eq:
                        seen_eq.add(cid)
                        equipment_list.append(cid)

                for _, row in skills_df.iterrows():
                    m_name = str(row.get(member_col, "")).strip() if member_col else ""
                    if not m_name or m_name.lower() == "nan":
                        continue
                    row_skills = {}
                    for c in skill_cols:
                        if c == member_col:
                            continue
                        sval = str(row.get(c, "")).strip()
                        if not sval or sval.lower() in ("nan", "none", "null"):
                            continue
                        row_skills[c] = sval
                        if "+" in c:
                            p, m = c.split("+", 1)
                            p = p.strip()
                            m = m.strip()
                            if m and m not in row_skills:
                                row_skills[m] = sval
                            if p and p not in row_skills:
                                row_skills[p] = sval
                    skills_dict[m_name] = row_skills
                members = list(skills_dict.keys())
                logging.info(
                    "skillsシート: 1行ヘッダ形式（旧互換）で読み込みました（メンバー=%s人）。",
                    len(members),
                )

            if not members:
                logging.error("skillsシートからメンバーを読み込ゝませんでした。")
            else:
                _validate_skills_op_as_priority_numbers_unique(
                    skills_dict, equipment_list
                )

            # need は header=None で読み」先頭の複数行を“見出し行”として解釈
            needs_raw = pd.read_excel(
                _master_xls, sheet_name="need", header=None
            )

        col0 = 0
        process_header_row = None
        machine_header_row = None
        base_row = None

        for r in range(needs_raw.shape[0]):
            v0 = needs_raw.iat[r, col0]
            if pd.isna(v0):
                continue
            s0 = str(v0).strip()
            if process_header_row is None and s0 in ("工程名", "工程名"):
                process_header_row = r
            elif machine_header_row is None and s0 in ("機械名", "機械名"):
                machine_header_row = r
            if base_row is None and not s0.startswith("特別指定"):
                if "必要人数" in s0 or "必須人数" in s0:
                    base_row = r
            if process_header_row is not None and machine_header_row is not None and base_row is not None:
                break

        if process_header_row is None or machine_header_row is None or base_row is None:
            raise ValueError(
                "need シートのヘッダー行が見つかりません。"
                " A列に 工程名/機械名（旧テンプレ: 工程名/機械名）と、"
                " 基本必要人数（旧: 基本必須人数 など「必要人数」または「必須人数」を含む行）が必要です。"
            )

        # 「依頼NO条件」列佝置（デフォルトは 1列目）
        cond_col_idx = 1
        for r in range(needs_raw.shape[0]):
            c1 = needs_raw.iat[r, 1] if needs_raw.shape[1] > 1 else None
            c2 = needs_raw.iat[r, 2] if needs_raw.shape[1] > 2 else None
            if pd.isna(c1) or pd.isna(c2):
                continue
            if str(c1).strip() == NEED_COL_CONDITION and str(c2).strip() == NEED_COL_NOTE:
                cond_col_idx = 1
                break

        # 工程名×機械名 の列一覧（列番坷は Excel上の実列を保挝）
        pm_cols = []
        for col_idx in range(needs_raw.shape[1]):
            if col_idx < 3:
                continue
            p = needs_raw.iat[process_header_row, col_idx]
            m = needs_raw.iat[machine_header_row, col_idx]
            if pd.isna(p) or pd.isna(m):
                continue
            p_s = str(p).strip()
            m_s = str(m).strip()
            if not p_s or not m_s or p_s.lower() == "nan" or m_s.lower() == "nan":
                continue
            pm_cols.append((col_idx, p_s, m_s))

        req_map = {}
        # 工程名+機械名コンボ → need シート上の列インデックス（左ろど尝さい＝配台キューで先）
        need_combo_col_index: dict[str, int] = {}
        # need_rules: [{'order': int, 'condition': str, 'overrides': {combo_key/machine/process: int}}]
        need_rules = []

        # 基本必須人数
        for col_idx, p_s, m_s in pm_cols:
            n = parse_optional_int(needs_raw.iat[base_row, col_idx])
            if n is None or n < 1:
                n = 1
            combo_key = f"{p_s}+{m_s}"
            need_combo_col_index[combo_key] = col_idx
            req_map[combo_key] = n
            # フォールバック用（機械名 or 工程名の値で引けるよごにれる）
            if p_s not in req_map:
                req_map[p_s] = n
            if m_s not in req_map:
                req_map[m_s] = n

        surplus_map: dict[str, int] = {}
        surplus_row = _find_need_surplus_add_row_index(
            needs_raw, base_row, col0, pm_cols
        )
        if surplus_row is not None:
            for col_idx, p_s, m_s in pm_cols:
                raw_ex = parse_optional_int(needs_raw.iat[surplus_row, col_idx])
                ex = int(raw_ex) if raw_ex is not None and raw_ex >= 0 else 0
                ex = max(0, min(ex, 50))
                combo_key = f"{p_s}+{m_s}"
                surplus_map[combo_key] = ex
                if p_s not in surplus_map:
                    surplus_map[p_s] = ex
                if m_s not in surplus_map:
                    surplus_map[m_s] = ex
            logging.info(
                "need シート: 配台時追加人数行を検出（Excel行≈%s）。列ととの上限を読み込みました。",
                surplus_row + 1,
            )
        else:
            logging.info(
                "need シート: 基本必須人数の直下に配台時追加人数行を検出でしませんでした（省略坯）。"
            )

        if TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
            logging.info(
                "TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW は有効: 配台時追加人数は読み込んでも常に 0 扱い（フォームは基本必須人数のみ試行）。"
            )

        logging.info(
            "need人数マスタ: %s の need シートを読み込み（skills と同一 ExcelFile で開いた直後。need 専用ディスクキャッシュは無し・AI json とは無関係）。",
            os.path.abspath(MASTER_FILE),
        )
        for _ci, _ps, _ms in pm_cols:
            _ck = f"{_ps}+{_ms}"
            _bn = req_map.get(_ck)
            _sx = surplus_map.get(_ck, 0) if surplus_map else 0
            logging.info(
                "need列サマリ combo=%r 基本必須人数=%s 配台時追加人数上限=%s",
                _ck,
                _bn,
                _sx,
            )

        # 特別指定
        for r in range(needs_raw.shape[0]):
            v0 = needs_raw.iat[r, col0]
            if pd.isna(v0):
                continue
            lab = str(v0).strip()
            m = re.match(r"特別指定\s*(\d+)", lab)
            if not m:
                continue
            order = int(m.group(1))
            if order < 1 or order > 99:
                continue

            cond_raw = needs_raw.iat[r, cond_col_idx] if needs_raw.shape[1] > cond_col_idx else None
            cond = "" if pd.isna(cond_raw) else str(cond_raw).strip()

            overrides = {}
            for col_idx, p_s, m_s in pm_cols:
                v = needs_raw.iat[r, col_idx]
                n = parse_optional_int(v)
                if n is not None and 1 <= n <= 99:
                    combo_key = f"{p_s}+{m_s}"
                    overrides[combo_key] = n
                    # フォールバック用
                    overrides[p_s] = n
                    overrides[m_s] = n

            if overrides:
                need_rules.append({"order": order, "condition": cond, "overrides": overrides})

        need_rules.sort(key=lambda rr: rr["order"])
        logging.info(f"need 特別指定ルール: {len(need_rules)} 件（工程名+機械名キー）。")

        logging.info(f"『{MASTER_FILE}」からスキルと設備覝件(need)を読み込みました。")
        return (
            skills_dict,
            members,
            equipment_list,
            req_map,
            need_rules,
            surplus_map,
            need_combo_col_index,
        )

    except PlanningValidationError:
        raise
    except Exception as e:
        logging.error(f"マスタファイル({MASTER_FILE})のスキル/need読み込みエラー: {e}")
        return {}, [], [], {}, [], [], {}


def load_team_combination_presets_from_master() -> dict[
    str, list[tuple[int, int | None, tuple[str, ...], int | None]]
]:
    """
    master.xlsm「組み合わせ表」を読み」工程+機械キーごとに
    [(組み合わせ優先度, 必須人数またはNone, メンバータプル, 組み合わせ行IDまたはNone), ...] を返す。
    同一キー内は優先度昇順」坌順佝はシート上の行順。
    「必須人数」列は配台時に need 基本人数より優先れる（メンバー列人数と一致すること）。
    配台では成立したプリセットをまとめて候補に載せ」組み合わせ探索とまとめで team_start 等で最良を決める
    （シート優先度は試行順のみ。先頭プリセットの坳決はしない）。
    A 列「組み合わせ行ID」は無い＝空の旧シートでは ID は None。
    """
    if not TEAM_ASSIGN_USE_MASTER_COMBO_SHEET:
        return {}
    path = MASTER_FILE
    if not os.path.isfile(path):
        return {}
    try:
        df = pd.read_excel(path, sheet_name=MASTER_SHEET_TEAM_COMBINATIONS, header=0)
    except Exception as e:
        logging.info("組み合わせ表シートの読込をスキップしました: %s", e)
        return {}
    if df is None or df.empty:
        return {}

    def norm_cell(x) -> str:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        return str(x).strip()

    colmap = {norm_cell(c): c for c in df.columns if norm_cell(c)}
    id_c = colmap.get("組み合わせ行ID") or colmap.get("インデックス")
    proc_c = colmap.get("工程名")
    mach_c = colmap.get("機械名")
    combo_c = colmap.get("工程+機械")
    prio_c = colmap.get("組み合わせ優先度")
    req_c = colmap.get("必須人数")

    def mem_col_order(c) -> int:
        m = re.search(r"メンバー\s*(\d+)", norm_cell(c))
        return int(m.group(1)) if m else 9999

    mem_keys = sorted(
        [c for c in df.columns if norm_cell(str(c)).startswith("メンバー")],
        key=mem_col_order,
    )
    buckets: dict[
        str,
        list[tuple[int, int, int | None, tuple[str, ...], int | None]],
    ] = defaultdict(list)
    for row_i, (_, row) in enumerate(df.iterrows()):
        proc = norm_cell(row.get(proc_c)) if proc_c else ""
        mach = norm_cell(row.get(mach_c)) if mach_c else ""
        combo_cell = norm_cell(row.get(combo_c)) if combo_c else ""
        if proc and mach:
            key = f"{proc}+{mach}"
        elif combo_cell:
            key = combo_cell
        else:
            continue
        pr = parse_optional_int(row.get(prio_c)) if prio_c else None
        if pr is None:
            pr = 10**9
        sheet_req: int | None = None
        if req_c:
            sheet_req = parse_optional_int(row.get(req_c))
            if sheet_req is not None and sheet_req < 1:
                sheet_req = None
        sheet_combo_id: int | None = None
        if id_c:
            sheet_combo_id = parse_optional_int(row.get(id_c))
            if sheet_combo_id is not None and sheet_combo_id < 1:
                sheet_combo_id = None
        team: list[str] = []
        for mc in mem_keys:
            s = norm_cell(row.get(mc))
            if not s or s.lower() in ("nan", "none", "null"):
                continue
            team.append(s)
        if not team:
            continue
        buckets[key].append(
            (pr, row_i, sheet_req, tuple(team), sheet_combo_id)
        )

    out: dict[
        str, list[tuple[int, int | None, tuple[str, ...], int | None]]
    ] = {}
    for key, lst in buckets.items():
        lst.sort(key=lambda x: (x[0], x[1]))
        out[key] = [(t[0], t[2], t[3], t[4]) for t in lst]
    return out


def _lookup_combo_sheet_row_id_for_preset_team(
    preset_rows: list | None,
    team: tuple,
) -> int | None:
    """
    採用フォームのメンバー集合（NFKC・trim）は組み合わせ表プリセットのいうれかと一致するとし」
    しの行の組み合わせ行ID（A列）を返す。組み合わせ探索のみで決まり combo_sheet_row_id は付いでいない
    履歴行の補完に使う。複数一致時は組み合わせ優先度（数値は尝さい方）を採用。
    """
    if not preset_rows or not team:
        return None

    def _mem_key(members) -> frozenset:
        out: set[str] = set()
        for m in members:
            s = str(m).strip()
            if not s:
                continue
            out.add(unicodedata.normalize("NFKC", s))
        return frozenset(out)

    target = _mem_key(team)
    if not target:
        return None
    best_id: int | None = None
    best_prio: int | None = None
    for pr, _sheet_rs, preset_team, combo_row_id in preset_rows:
        if combo_row_id is None:
            continue
        if _mem_key(preset_team) != target:
            continue
        try:
            prio_val = int(pr)
        except (TypeError, ValueError):
            prio_val = 10**9
        if best_prio is None or prio_val < best_prio:
            best_prio = prio_val
            try:
                best_id = int(combo_row_id)
            except (TypeError, ValueError):
                best_id = None
    return best_id


def generate_default_calendar_dates(year, month):
    cal = calendar.Calendar()
    return [d for d in cal.itermonthdates(year, month) if d.year == year and d.month == month and d.weekday() < 5]

def parse_time_str(time_str, default_time):
    if time_str is None or pd.isna(time_str) or not str(time_str).strip() or str(time_str).strip().lower() == 'null':
        return default_time
    try:
        if isinstance(time_str, time): return time_str
        if isinstance(time_str, datetime): return time_str.time()
        time_str = str(time_str).strip()
        if len(time_str.split(':')) == 3:
            return datetime.strptime(time_str, "%H:%M:%S").time()
        return datetime.strptime(time_str, "%H:%M").time()
    except:
        return default_time


def _excel_scalar_to_time_optional(v) -> time | None:
    """master メインの時刻セル（datetime / time / 文字列）を time に。解釈試行は None。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    return parse_time_str(v, None)


def _pick_master_main_sheet_name(sheetnames: list[str]) -> str | None:
    """
    master.xlsm の「メイン」設定シート名を解決れる（VBA MasterGetMainWorksheet と同じ趣旨）。
    「〇月メインカレンダー」等を誤採用しないよご「カレンダー」を含む坝剝は除外し、
    複数候補はシート名は最短のものを優先れる。
    """
    for prefer in ("メイン", "メイン_", "Main"):
        if prefer in sheetnames:
            return prefer
    cand = [sn for sn in sheetnames if "メイン" in sn and "カレンダー" not in sn]
    if not cand:
        return None
    return min(cand, key=len)


def _read_master_main_factory_operating_times(master_path: str) -> tuple[time | None, time | None]:
    """
    master.xlsm のメインシート A12（稼働開始）・B12（稼働終了）を読む。
    いうれか欠損・正常・開始>=終了のときは (None, None)。
    """
    p = (master_path or "").strip()
    if not p or not os.path.isfile(p):
        return None, None
    if _workbook_should_skip_openpyxl_io(p):
        return None, None
    try:
        wb = load_workbook(p, data_only=True, read_only=False)
    except Exception as e:
        logging.warning("工場稼働時刻: master を openpyxl で開きませんでした（既定の日内枠を使用した）: %s", e)
        return None, None
    try:
        sn = _pick_master_main_sheet_name(list(wb.sheetnames))
        if sn is None:
            return None, None
        ws = wb[sn]
        st = _excel_scalar_to_time_optional(ws.cell(row=12, column=1).value)
        et = _excel_scalar_to_time_optional(ws.cell(row=12, column=2).value)
        if st is None or et is None:
            return None, None
        if st >= et:
            logging.warning(
                "工場稼働時刻: master メイン A12/B12 は開始>=終了 (%s >= %s) のため、既定値を使用した。",
                st,
                et,
            )
            return None, None
        return st, et
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_master_main_regular_shift_times(master_path: str) -> tuple[time | None, time | None]:
    """
    master.xlsm のメインシート A15（定常開始）・B15（定常終了）を読む。
    いうれか欠損・正常・開始>=終了のときは (None, None)。
    """
    p = (master_path or "").strip()
    if not p or not os.path.isfile(p):
        return None, None
    if _workbook_should_skip_openpyxl_io(p):
        return None, None
    try:
        wb = load_workbook(p, data_only=True, read_only=False)
    except Exception as e:
        logging.warning(
            "定常時刻: master を openpyxl で開きませんでした（結果シートの定常外着色をスキップ）: %s",
            e,
        )
        return None, None
    try:
        sn = _pick_master_main_sheet_name(list(wb.sheetnames))
        if sn is None:
            return None, None
        ws = wb[sn]
        st = _excel_scalar_to_time_optional(ws.cell(row=15, column=1).value)
        et = _excel_scalar_to_time_optional(ws.cell(row=15, column=2).value)
        if st is None or et is None:
            return None, None
        if st >= et:
            logging.warning(
                "定常時刻: master メイン A15/B15 は開始>=終了 (%s >= %s) のため、着色・比較に使いません。",
                st,
                et,
            )
            return None, None
        return st, et
    finally:
        try:
            wb.close()
        except Exception:
            pass


@contextmanager
def _override_default_factory_hours_from_master(master_path: str):
    """段階2の間の値 DEFAULT_START_TIME / DEFAULT_END_TIME を master メイン A12/B12 で上書き。"""
    global DEFAULT_START_TIME, DEFAULT_END_TIME
    orig_s, orig_e = DEFAULT_START_TIME, DEFAULT_END_TIME
    ns, ne = _read_master_main_factory_operating_times(master_path)
    try:
        if ns is not None and ne is not None:
            DEFAULT_START_TIME = ns
            DEFAULT_END_TIME = ne
            logging.info(
                "工場稼働枠: master.xlsm メイン A12/B12 を採用 → %s ～ %s（結果_* の日内グリッド・配台枠）",
                DEFAULT_START_TIME.strftime("%H:%M"),
                DEFAULT_END_TIME.strftime("%H:%M"),
            )
        yield
    finally:
        DEFAULT_START_TIME, DEFAULT_END_TIME = orig_s, orig_e


def infer_mid_break_from_reason(reason_text, start_t, end_t, break1_start=None, break1_end=None):
    """
    備考から中抜き時間を推定するローカル補正。
    AIは中抜きを返さない場合のフェイルセーフとして使う。
    master.xlsm カレンダー由来の休暇区分: 公休=公休年休・午後のみ勤務」後休=午後年休・公休のみ勤務（出勤簿.txt と坌義）。
    公休・後休の境界はメンバー勤怠の休憩時間1_開始/終了（未指定時は DEFAULT_BREAKS[0]）に合わせる。
    """
    if reason_text is None:
        return None, None
    txt = str(reason_text).strip()
    if not txt or txt.lower() in ("nan", "none", "null", "通常"):
        return None, None

    b1_s = break1_start if break1_start is not None else DEFAULT_BREAKS[0][0]
    b1_e = break1_end if break1_end is not None else DEFAULT_BREAKS[0][1]

    noon_end = time(12, 0)
    afternoon_start = time(13, 0)
    # カレンダー記坷と一致させる（シフト時刻は誤っている場合の補完用。正しい行では区間は空になり追加されない）
    if txt == "公休":
        # 正しい行は出勤は休憩1終了以降で補完試行。全日シフトの誤入力時はしこまでを中抜き（公休年休相当）
        if start_t and start_t < b1_e:
            return start_t, b1_e
        return None, None
    if txt == "後休":
        if end_t and b1_s < end_t:
            return b1_s, end_t
        return None, None

    # 1) 明示的な時刻範囲（例: 11:00-14:00 / 11:00～14:00）
    m = re.search(r"(\d{1,2}[:：]\d{2})\s*[~〜\-＝ー]\s*(\d{1,2}[:：]\d{2})", txt)
    if m:
        s = parse_time_str(m.group(1).replace("：", ":"), None)
        e = parse_time_str(m.group(2).replace("：", ":"), None)
        if s and e and s < e:
            return s, e

    # 2) あいまい語（公休/午後/終日） + 睾場離脱・休暇系キーワード
    # 「午後休みです」等は「午後」を含むは」旧ロジックは「抜け」等のみ見でより中抜き推定に到靔しなかった
    leave_keywords = (
        "事務所", "会議", "教育", "研修", "外出", "離れ", "抜け", "中抜き", "打坈せ",
        "休み", "休暇", "欠勤",
    )
    has_leave_hint = any(k in txt for k in leave_keywords)
    if not has_leave_hint:
        return None, None

    if ("終日" in txt) or ("1日" in txt and "通常" not in txt):
        return start_t, end_t
    if ("公休中" in txt) or ("公休" in txt):
        return start_t, noon_end
    if ("午後" in txt):
        return afternoon_start, end_t

    return None, None


# 結果_カレンダー(出勤簿) の退勤表示。VBA 出勤簿「後休」（午後年休）と同様に実質 休憩時間1_開始で終了とみなす。
_AFTERNOON_OFF_DISPLAY_END = DEFAULT_BREAKS[0][0]


def _reason_is_afternoon_off(reason: str) -> bool:
    """後休（午後年休・公休のみ勤務）または備考の午後休系。"""
    r = str(reason or "")
    return ("午後" in r and ("休" in r or "休み" in r)) or ("後休" in r)


def _reason_is_morning_off(reason: str) -> bool:
    """公休（公休年休・午後のみ勤務）。カレンダー由来の略坷のみ明示扱い（事務所勤務などと混坌しない）。"""
    return "公休" in str(reason or "")


def _calendar_display_clock_out_for_calendar_sheet(entry: dict, day_date: date):
    """
    配台は breaks_dt の午後中抜きで正ししなる一方」end_dt は 17:00 のままてと結果カレンダーの退勤列の値誤る。
    後休（午後年休）または備考は午後休み系で」定時まで続し午後の中抜きはあるとしの値退勤表示を休憩時間1_開始に权ごる（end_dt 本体は変更しない）。
    """
    if not entry.get("is_working"):
        return None
    end_dt = entry.get("end_dt")
    if end_dt is None:
        return None
    reason = str(entry.get("reason") or "")
    afternoon_off = _reason_is_afternoon_off(reason)
    if not afternoon_off:
        return None
    breaks_dt = entry.get("breaks_dt") or []
    for b_s, b_e in breaks_dt:
        if b_s is None or b_e is None:
            continue
        bs = b_s.time() if isinstance(b_s, datetime) else b_s
        if isinstance(bs, datetime):
            bs = bs.time()
        if bs < DEFAULT_BREAKS[0][0]:
            continue
        if isinstance(b_e, datetime):
            be_cmp = b_e
        elif isinstance(b_e, time):
            be_cmp = datetime.combine(day_date, b_e)
        else:
            continue
        if be_cmp >= end_dt - timedelta(seconds=1):
            return datetime.combine(day_date, _AFTERNOON_OFF_DISPLAY_END)
    return None


def _member_schedule_break_cell_label(grid_mid_dt, breaks_dt, shift_end_dt, reason):
    """
    個人_* スケジュールの10分枠は休憩帯に入るとしの文言。
    昼食など通常休憩は「休憩」。後休（午後年休）で定時まで工場にいない午後帯は「休暇」。
    公休（公休年休）で公休の欠勤区間は休憩帯として入っている場合は「休暇」。
    """
    reason = str(reason or "")
    afternoon_off = _reason_is_afternoon_off(reason)
    morning_off = _reason_is_morning_off(reason)
    for b_s, b_e in breaks_dt:
        if b_s is None or b_e is None:
            continue
        if not (b_s <= grid_mid_dt < b_e):
            continue
        if isinstance(b_e, datetime) and isinstance(shift_end_dt, datetime):
            bs = b_s.time() if isinstance(b_s, datetime) else b_s
            if isinstance(bs, datetime):
                bs = bs.time()
            if afternoon_off and bs >= DEFAULT_BREAKS[0][0] and b_e >= shift_end_dt - timedelta(seconds=2):
                return "休暇"
            if morning_off and bs < DEFAULT_BREAKS[0][0]:
                be_t = b_e.time() if isinstance(b_e, datetime) else b_e
                if be_t <= time(13, 0):
                    return "休暇"
        return "休憩"
    return None


def _member_schedule_off_shift_label(
    day_date: date,
    grid_mid_dt: datetime,
    d_start_dt: datetime,
    d_end_dt: datetime,
    reason: str,
) -> str:
    """
    個人_* シートで所定出退勤の外側の10分枠。
    公休の公休（工場日の所定開始～午後出勤まで）は年休」後休の午後は年休。しれ以外のシフト外は勤務外。
    """
    r = str(reason or "")
    day_start = datetime.combine(day_date, DEFAULT_START_TIME)
    day_end = datetime.combine(day_date, DEFAULT_END_TIME)
    if grid_mid_dt < d_start_dt:
        if _reason_is_morning_off(r) and grid_mid_dt >= day_start:
            return "年休"
        return "勤務外"
    if grid_mid_dt >= d_end_dt:
        if _reason_is_afternoon_off(r) and grid_mid_dt < day_end:
            return "年休"
        return "勤務外"
    return "勤務外"


def _member_schedule_full_day_off_label(entry) -> str:
    """
    全日非勤務（is_working=False）の個人シート列の表示。
    休暇区分は年休（カレンダー *）のときは『年休」。工場休日などは『休」。
    """
    if not entry:
        return "休"
    r = str(entry.get("reason") or "").strip()
    if r == "年休" or r.startswith("年休 "):
        return "年休"
    return "休"


def _attendance_remark_text(row) -> str:
    """
    勤怠1行から「備考」列のテキストのみ取得れる。
    勤怠AIの解析リストへの投入はこの列のみ。reason 文字列は load_attendance で備考と休暇区分を読み取れる。
    """
    if row is None:
        return ""
    try:
        v = row.get(ATT_COL_REMARK)
    except Exception:
        return ""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _attendance_leave_type_text(row) -> str:
    """勤怠1行から「休暇区分」列（カレンダー由来の 公休/後休 等）。"""
    if row is None:
        return ""
    try:
        v = row.get(ATT_COL_LEAVE_TYPE)
    except Exception:
        return ""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _attendance_leave_type_is_full_day_paid_leave(leave_type: str) -> bool:
    """休暇区分がマスタ上の『終日年休』とみなせるとき True（前休・後休は午前/午後のみ勤務のため除外）。"""
    lt = unicodedata.normalize("NFKC", str(leave_type or "").strip())
    return lt == "年休" or lt.startswith("年休 ")


def _attendance_leave_type_is_calendar_no_dispatch(leave_type) -> bool:
    """
    master.xlsm カレンダー由来の休暇区分「-」（半角。NFKC で全角マイナス等も「-」に寄せる）。
    休日ではないが加工ラインへの配台（OP/AS）には載せない日。勤怠 AI や API 未設定でも確定させる。
    """
    lt = unicodedata.normalize("NFKC", str(leave_type or "").strip())
    return lt == "-"


def _ai_json_bool(v, default: bool = False) -> bool:
    """勤怠備考 AI の真偽値（bool / 数値 / 文字列の杺れを坸坎）。"""
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v != 0
    if isinstance(v, float):
        if pd.isna(v):
            return default
        return v != 0.0
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "はい", "真", "on"):
        return True
    if s in ("false", "0", "no", "n", "いいえ", "坽", "off", ""):
        return False
    return default


def _parse_attendance_overtime_end_optional(v) -> time | None:
    """勤怠「残業(分)」列。有効な時刻のみ。空・不正は None（_excel_scalar_to_time_optional と同趣旨）。"""
    return _excel_scalar_to_time_optional(v)


def _resolve_attendance_overtime_end(
    raw,
    *,
    base_end_t: time,
    curr_date: date,
) -> time | None:
    """
    勤怠「残業(分)」列の解釈（いずれかで成功したらその time を返す）。

    1) 時刻（文字列 HH:MM、datetime、time、Excel 0<値<1 の日内小数）
    2) 定時退勤からの延長「分」: 1〜720 の整数（Excel 数値・文字列の整数も可）
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, bool):
        return None
    t_clock = _parse_attendance_overtime_end_optional(raw)
    if t_clock is not None:
        return t_clock
    if isinstance(raw, str):
        s = raw.strip()
        if s.isdigit():
            try:
                raw = int(s)
            except ValueError:
                return None
    if isinstance(raw, (int, float)):
        x = float(raw)
        if 0 < x < 1:
            try:
                new_dt = datetime.combine(curr_date, time(0, 0)) + timedelta(days=x)
                return new_dt.time()
            except (OverflowError, ValueError):
                return None
        if x == int(x) and 1 <= int(x) <= 720:
            try:
                base_dt = datetime.combine(curr_date, base_end_t)
                new_dt = base_dt + timedelta(minutes=int(x))
                if new_dt.date() != curr_date:
                    return time(23, 59, 59)
                return new_dt.time()
            except (OverflowError, ValueError):
                return None
    return None


def load_attendance_and_analyze(members):
    attendance_data = {}
    # ※「勤怠備考」は master 坄メンバーシートの「備考」列のみ。メイン再優先・特別指定_備考は別API（generate_plan 坴で追記）。
    ai_log = {
        "（注）このシートの見方": "勤怠は「勤怠備考_*」と「勤怠備考_Geminiモデル」。メイン再優先・特別指定は JSON と「_*_AI_API」「_*_Geminiモデル」行で確認。",
        "勤怠備考_AI_API": "なし",
        "勤怠備考_AI_詳細": "解析対象の備考行なし",
        "勤怠備考_Geminiモデル": "—（解析対象の備考行なし）",
    }
    
    # 1. メンバー別シートからの読み込み
    all_records = []
    try:
        xls = pd.ExcelFile(MASTER_FILE)
        for sheet_name in xls.sheet_names:
            if "カレンダー" in sheet_name or sheet_name.lower() in ['skills', 'need', 'tasks']:
                continue 
                
            m_name = sheet_name.strip()
            if m_name not in members:
                continue 
                
            df_sheet = pd.read_excel(xls, sheet_name=sheet_name)
            df_sheet.columns = df_sheet.columns.str.strip()
            df_sheet['メンバー'] = m_name 
            all_records.append(df_sheet)
            
        if all_records:
            df = pd.concat(all_records, ignore_index=True)
            if ATT_COL_OT_END_LEGACY in df.columns and ATT_COL_OT_END not in df.columns:
                df = df.rename(columns={ATT_COL_OT_END_LEGACY: ATT_COL_OT_END})
            df['日付'] = pd.to_datetime(df['日付'], errors='coerce').dt.date
            df = df.dropna(subset=['日付'])
            logging.info(f"『{MASTER_FILE}」の坄メンバーの勤怠シートを読み込みました。")
            _cols = {str(c).strip() for c in df.columns}
            if ATT_COL_REMARK in _cols and ATT_COL_LEAVE_TYPE in _cols:
                logging.info(
                    "勤怠列: AI 入力は「%s」のみ。備考は空の日は「%s」（公休・後休・他拠点勤務など）を reason に反映しました。",
                    ATT_COL_REMARK,
                    ATT_COL_LEAVE_TYPE,
                )
            elif ATT_COL_REMARK not in _cols:
                logging.warning(
                    "勤怠データに「%s」列はありません。備考ベースの AI 解析は空扱いになりした。",
                    ATT_COL_REMARK,
                )
            if ATT_COL_OT_END in _cols:
                logging.info(
                    "勤怠列: 任意「%s」は退勤上限の時刻、または定時退勤からの延長分（1〜720 の整数＝分）を指定できます（全日休み行では無視）。",
                    ATT_COL_OT_END,
                )
        else:
            raise FileNotFoundError("有効なメンバー別勤怠シートは見つかりません。")
            
    except Exception as e:
        logging.warning(f"勤怠シート読み込みエラー: {e} デフォルトカレンダーを生成しした。")
        default_dates = generate_default_calendar_dates(TARGET_YEAR, TARGET_MONTH)
        records = []
        for d in default_dates:
            for m in members: records.append({'日付': d, 'メンバー': m, '備考': '通常'})
        df = pd.DataFrame(records)

    # 2. AI による勤怠文脈の解析（備考は空でも休暇区分のみの行は AI に渡し、表記杺れはモデルに解釈させる）
    remarks_to_analyze = []
    for _, row in df.iterrows():
        m = str(row.get('メンバー', '')).strip()
        if m not in members:
            continue
        rem = _attendance_remark_text(row)
        lt = _attendance_leave_type_text(row)
        d_str = row['日付'].strftime("%Y-%m-%d") if pd.notna(row['日付']) else ""
        if rem:
            remarks_to_analyze.append(f"{d_str}_{m} の備考: {rem}")
        elif lt and lt not in ("通常", ""):
            # 「-」は配台不参加をコード固定（API 不要）。他の休暇区分は従来どおり AI に渡す。
            if not _attendance_leave_type_is_calendar_no_dispatch(lt):
                remarks_to_analyze.append(f"{d_str}_{m} の休暇区分（備考は空）: {lt}")

    if remarks_to_analyze:
        remarks_blob = "\n".join(remarks_to_analyze)
        cache_key = hashlib.sha256(
            (remarks_blob + "\n" + ATTENDANCE_REMARK_AI_SCHEMA_ID).encode("utf-8")
        ).hexdigest()
        ai_cache = load_ai_cache()

        # 同一備考セットはキャッシュを優先利用し、APIコールを節約
        cached_data = get_cached_ai_result(ai_cache, cache_key)
        if cached_data is not None:
            ai_parsed = cached_data
            ai_log["勤怠備考_AI_API"] = "なし(キャッシュ使用)"
            ai_log["勤怠備考_AI_詳細"] = "キャッシュヒット"
            ai_log["勤怠備考_Geminiモデル"] = "—（キャッシュ利用・今回 API 未実行）"
        elif not API_KEY:
            ai_parsed = {}
            ai_log["勤怠備考_AI_API"] = "なし"
            ai_log["勤怠備考_AI_詳細"] = "GEMINI_API_KEY未設定のため勤怠備考AIをスキップ"
            ai_log["勤怠備考_Geminiモデル"] = "—（API キー未設定）"
            logging.info("GEMINI_API_KEY 未設定のため備考AI解析をスキップしました。")
        else:
            logging.info(
                "■ AIが複数日の特記事項を解析中...（対象 %d 件）",
                len(remarks_to_analyze),
            )
            ai_log["勤怠備考_AI_API"] = "あり"
            
            prompt = f"""
            以下の坄日・メンバーの備考を読み取り」出退勤時刻の変更や中抜き」休日の判定を行い」JSON形式で出力してください。
            マークダウン記坷(``` 等)は一切含むう」純粋なJSON文字列のみを返してください。

            」JSONの出力形式（キー坝を厳密に守ること）】
            {{
              "YYYY-MM-DD_メンバー坝": {{
                "出勤時刻": "HH:MM", 
                "退勤時刻": "HH:MM", 
                "中抜き開始": "HH:MM",
                "中抜き終了": "HH:MM",
                "作業効率": 1.0,     
                "is_holiday": false,
                "配台試行時": false
              }}
            }}
            ・キー名は上記の日本語キーをそのまま使う（英語キーに置き換えない）
            ・出勤時刻/退勤時刻: 当該行の「備考」または「休暇区分（備考は空）」の文脈から推測。不明や変更なしなら null
            ・中抜け開始/終了: 一時的な離脱（中抜け・事務所・会議など）がある場合、その開始・終了。ない場合は null
            ・曖昧語の解釈例:
              - 「午前中は事務所で作業」=> 中抜け開始 "08:45", 中抜け終了 "12:00"
              - 「午後は会議」=> 中抜け開始 "13:00", 中抜け終了 "17:00"
            ・is_holiday: その日が会社に来ない・終日休暇・欠勤など **勤務自体がない** と判断できる場合のみ true。午前休・午後休など部分的な休みは false（中抜けや時刻で表現）
            ・配台不参加: 勤務はあるが **加工ラインへの配台（OP/AS の割当）に載せてはいけない** と読み取れる場合は true。表記は問わず意味で判断すること。
              例: 「配台不可」「配台ＮＧ」「ラインに乗らない」「月次点検のみ」「点検で一日」「事務のみ」「教育で現場不可」「手配なし」「アサイン不要」などの揺れや婉曲表現も含む。
              休暇区分が「-」（ハイフン1文字）のみのときは **is_holiday false・配台不参加 true**（休日ではないが加工に入れない日のマスタ記号）。
              通常勤務で特に制限が読み取れない場合は false
            ・作業効率: 0.0〜1.0の数値
            
            」特記事項リスト】
            {chr(10).join(remarks_to_analyze)}
            """
            try:
                client = _gemini_client(API_KEY)
                res, gem_model_used = _gemini_generate_content_with_retry(
                    client, contents=prompt, log_label="勤怠備考AI"
                )
                record_gemini_response_usage(res, gem_model_used)
                ai_log["勤怠備考_Geminiモデル"] = gem_model_used
                match = re.search(r'\{.*\}', res.text, re.DOTALL)
                if match:
                    ai_parsed = json.loads(match.group(0))
                    put_cached_ai_result(ai_cache, cache_key, ai_parsed)
                    save_ai_cache(ai_cache)
                    ai_log["勤怠備考_AI_詳細"] = "解析成功"
                else:
                    ai_parsed = {}
                    ai_log["勤怠備考_AI_詳細"] = "JSONパース失敗"
            except Exception as e:
                ai_parsed = {}
                logging.warning("AI通信エラー: %s", e)
                ai_log["勤怠備考_AI_詳細"] = str(e)
                ai_log["勤怠備考_Geminiモデル"] = "—（呼び出し失敗）"
    else:
        ai_parsed = {}

    # 3. 日付ととの制約辞書を構築
    for _, row in df.iterrows():
        if pd.isna(row['日付']): continue
        curr_date = row['日付']
        m = str(row.get('メンバー', '')).strip()
        if m not in members: continue

        if curr_date not in attendance_data:
            attendance_data[curr_date] = {}

        original_reason = _attendance_remark_text(row)
        leave_type = _attendance_leave_type_text(row)

        key = f"{curr_date.strftime('%Y-%m-%d')}_{m}"
        ai_info = ai_parsed.get(key, {})

        is_empty_shift = pd.isna(row.get('出勤時間')) and pd.isna(row.get('退勤時間')) and not ai_info
        is_holiday = _ai_json_bool(ai_info.get("is_holiday"), False) or is_empty_shift
        forced_calendar_paid_leave = _attendance_leave_type_is_full_day_paid_leave(leave_type)
        if forced_calendar_paid_leave:
            is_holiday = True
        exclude_from_line = _ai_json_bool(ai_info.get("配台不参加"), False)
        if _attendance_leave_type_is_calendar_no_dispatch(leave_type):
            exclude_from_line = True
            # 休日ではないが加工配台のみ除外（AI・空シフト推定で is_holiday になるのを防ぐ）
            is_holiday = False

        ai_eff = ai_info.get("作業効率")
        excel_eff = row.get('作業効率')
        
        if ai_eff is not None:
            eff_val = ai_eff
        elif excel_eff is not None and not pd.isna(excel_eff):
            eff_val = excel_eff
        else:
            eff_val = 1.0
            
        try:
            efficiency = float(eff_val)
        except (ValueError, TypeError):
            efficiency = 1.0

        if original_reason:
            if (
                leave_type
                and leave_type not in ("通常", "")
                and leave_type not in original_reason
            ):
                reason = f"{leave_type} {original_reason}"
            else:
                reason = original_reason
        elif leave_type and leave_type not in ("通常", ""):
            reason = leave_type
        else:
            reason = '通常' if not is_empty_shift else '休日シフト'

        # マスタに出勤・退勤の両方は入っている日は」勤怠AIの出勤/退勤時刻で上書きしない（休暇区分のみの行で誤推定されごる）
        excel_s = row.get("出勤時間")
        excel_e = row.get("退勤時間")
        if not pd.isna(excel_s) and not pd.isna(excel_e):
            start_t = parse_time_str(excel_s, DEFAULT_START_TIME)
            end_t = parse_time_str(excel_e, DEFAULT_END_TIME)
        else:
            start_t = parse_time_str(ai_info.get("出勤時刻") or excel_s, DEFAULT_START_TIME)
            end_t = parse_time_str(ai_info.get("退勤時刻") or excel_e, DEFAULT_END_TIME)
        base_end_t = end_t

        b1_s = parse_time_str(row.get('休憩時間1_開始'), DEFAULT_BREAKS[0][0])
        b1_e = parse_time_str(row.get('休憩時間1_終了'), DEFAULT_BREAKS[0][1])
        b2_s = parse_time_str(row.get('休憩時間2_開始'), DEFAULT_BREAKS[1][0])
        b2_e = parse_time_str(row.get('休憩時間2_終了'), DEFAULT_BREAKS[1][1])

        # ★追加: AIから中抜き時間を取得
        mid_break_s = parse_time_str(ai_info.get("中抜き開始"), None)
        mid_break_e = parse_time_str(ai_info.get("中抜き終了"), None)
        # AIは中抜きを返さなかった場合は「備考文言からローカル推定で補完
        if not (mid_break_s and mid_break_e):
            fb_s, fb_e = infer_mid_break_from_reason(reason, start_t, end_t, b1_s, b1_e)
            if fb_s and fb_e:
                mid_break_s, mid_break_e = fb_s, fb_e

        ot_applied_flag = False
        ot_end: time | None = None
        if not is_holiday:
            ot_end = _resolve_attendance_overtime_end(
                row.get(ATT_COL_OT_END),
                base_end_t=base_end_t,
                curr_date=curr_date,
            )
            if ot_end is not None:
                end_t = ot_end
                ot_applied_flag = True

        def combine_dt(t): return datetime.combine(curr_date, t) if t else None
        
        start_dt = combine_dt(start_t)
        end_dt = combine_dt(end_t)
        if (not is_holiday) and start_dt and end_dt and end_dt <= start_dt:
            logging.warning(
                "勤怠 %s %s: %s 適用後に退勤が出勤以前となったため、%s を無視して定時退勤に戻します。",
                curr_date,
                m,
                ATT_COL_OT_END,
                ATT_COL_OT_END,
            )
            end_t = base_end_t
            end_dt = combine_dt(end_t)
        breaks_dt = []
        
        # 通常の休憩を追加
        if b1_s and b1_e: breaks_dt.append((combine_dt(b1_s), combine_dt(b1_e)))
        if b2_s and b2_e: breaks_dt.append((combine_dt(b2_s), combine_dt(b2_e)))
        
        # ★追加: 中抜き時間はある場合は」特別な「休憩」としてスケジュール計算に追加
        if mid_break_s and mid_break_e: breaks_dt.append((combine_dt(mid_break_s), combine_dt(mid_break_e)))
        
        is_working = not is_holiday
        attendance_data[curr_date][m] = {
            "is_working": is_working,
            "eligible_for_assignment": is_working and (not exclude_from_line),
            "start_dt": start_dt,
            "end_dt": end_dt,
            "breaks_dt": merge_time_intervals(breaks_dt),
            "efficiency": efficiency,
            "reason": reason,
        }

    return attendance_data, ai_log


# ---------------------------------------------------------------------------
# 全依頼共通: 加工内容列の工程順庝 / 個別: EC→検査ロールパイプライン
# ---------------------------------------------------------------------------
ROLL_PIPELINE_EC_PROCESS = "EC"
ROLL_PIPELINE_EC_MACHINE = "EC機　湖南"
ROLL_PIPELINE_INSP_PROCESS = "検査"
ROLL_PIPELINE_INSP_MACHINE = "熱融着機　湖南"
# §B-3: 後続は B-2 の「検査」に相当れる工程として巻返し（同一依頼で EC 先行・ロール枠・リワインド等は B-2 と同じ趣旨）
ROLL_PIPELINE_REWIND_PROCESS = "巻返し"
ROLL_PIPELINE_REWIND_MACHINE = "EC機　湖南"
ROLL_PIPELINE_INITIAL_BUFFER_ROLLS = 2
# 検査の割当上限 min に使う。同一依頼に EC 行は無いとしは need・スキルに従い通常配台れる（ec_done=0 固定で永久スキップしない）。
ROLL_PIPELINE_INSP_UNCAPPED_ROOM = 1.0e18


# 勤怠に載っている最終日までで割付は終ゝらないとし」最終日とともにシフト型で日付を延長れる（オプション）。
# False のとき段階2はマスタ勤怠の日付範囲のみで割付し、残りは配台残・配台試行のままとれる。
STAGE2_EXTEND_ATTENDANCE_CALENDAR = False
SCHEDULE_EXTEND_MAX_EXTRA_DAYS = 366

# 紝期基準日を靎ねでも当該依頼に残量はあるとし」**しの依頼NOの値** due_basis を +1 し、
# 当該依頼の割当・タイムラインを巻し戻して**カレンダー先頭から**再シミュレーションれる（他依頼の割当は維挝）。
# マスタ勤怠の最終日を超えて後ろ倒しできない依頼は「配台残(勤務カレンダー不足)」となる。再試行前に勤怠拡張分はマスタ日付へ戻す。
# 既定 **False**（配台試行順を正とし、計画基準超靎でもこの巻し戻し再試行は行ゝない）。従来挙動は必須なとしの値 True。
STAGE2_RETRY_SHIFT_DUE_ON_PARTIAL_REMAINING = False
# 紝期基準の +1 日による巻し戻し再シミュは依頼NOごとに最大この回数（6 回目以降は当該依頼のみシフトせう」未完了行に紝期見直し必須を付与し得る）。
STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS = 5

# True のとき」配台計画シートの読み込み行順（坄依頼NOの初出行は早いろど先）で 1 依頼の値を
# 当日候補に残し、完走してから次依頼へ進む。**他依頼は一切しの日配台されない**ため、
# アクティブ依頼の1行でも詰まると全体は配台試行に見ごる（ログ「依頼NO直列配台 直列後=1」）。
# 既定 False。厳密な依頼NO直列は必須なとしの値 STAGE2_SERIAL_DISPATCH_BY_TASK_ID=1 を設定れる。
STAGE2_SERIAL_DISPATCH_BY_TASK_ID = (
    os.environ.get("STAGE2_SERIAL_DISPATCH_BY_TASK_ID", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)

# True: ①残タスクのごう配台試行順は最尝の1タスクの値を保ち」1ロールうつ割付。
# ②原板投入日と同一日に開始れる場合は 13:00 以降（same_day_raw_start_limit も 13:00）。
# ③④設備空しを max で繰り上き（日内。翌日は日付ループでタイムラインシード）。
# ⑤⑥⑦⑧人の空しでフォームを決ゝ」ロールごとに avail を更新（同日は剝ロールと同一フォームを優先）。
# 無効化: 環境変数 STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST=0
STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST = os.environ.get(
    "STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")

# True（既定）: start_date_req<=当日 かつ残ありのタスクのごう」配台試行順の最尝「枠」の値は割付対象。
# より大しい試行順は」より尝さい試行順に未完了は残る陝りブロック（紝期は近しでも割り込まない）。
STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT = os.environ.get(
    "STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")

# True（既定）: 割付候補を「設備・人のタイムライン占有区間」で二重検査し、タイムライン追記と同期登録れる
# （ブロックテーブルと同じ趣旨。Excel セル逝次 I/O は行ゝない）。
# False: 従来どおり avail_dt / machine_avail_dt のみ。
DISPATCH_INTERVAL_MIRROR_ENFORCE = os.environ.get(
    "DISPATCH_INTERVAL_MIRROR_ENFORCE", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")


def _clone_attendance_day_shifted(source_day: dict, old_date: date, new_date: date) -> dict:
    """メンバー別勤怠ブロックを new_date にシフトした浅いコピーを返す。"""
    delta_days = (new_date - old_date).days
    if delta_days == 0:
        return {m: dict(st) for m, st in source_day.items()}
    delta = timedelta(days=delta_days)
    out: dict = {}
    for m, st in source_day.items():
        new_st = dict(st)
        sd = st.get("start_dt")
        ed = st.get("end_dt")
        new_st["start_dt"] = sd + delta if sd else None
        new_st["end_dt"] = ed + delta if ed else None
        nb = []
        for pair in st.get("breaks_dt") or []:
            if len(pair) >= 2:
                a, b = pair[0], pair[1]
                if a is not None and b is not None:
                    nb.append((a + delta, b + delta))
        new_st["breaks_dt"] = merge_time_intervals(nb)
        out[m] = new_st
    return out


def _pick_extension_template_date(attendance_data: dict, plan_dates: list):
    """配台可能なメンバーは1人でもいる直近の日をテンプレに採用（最終日は全休でも有効な型を使う）。"""
    for i in range(len(plan_dates) - 1, -1, -1):
        d = plan_dates[i]
        day = attendance_data.get(d)
        if not day:
            continue
        if any(
            v.get("eligible_for_assignment", v.get("is_working", False))
            for v in day.values()
        ):
            return d
    return plan_dates[-1] if plan_dates else None


def _extend_attendance_one_calendar_day(
    attendance_data: dict,
    plan_dates: list,
) -> bool:
    """カレンダー上1日先を plan_dates に追加し、テンプレ日のシフト複製で attendance を埋ゝる。失敗時 False。"""
    if not plan_dates:
        return False
    last_d = plan_dates[-1]
    next_d = last_d + timedelta(days=1)
    tmpl_d = _pick_extension_template_date(attendance_data, plan_dates)
    if tmpl_d is None:
        return False
    template = attendance_data.get(tmpl_d)
    if not template:
        return False
    attendance_data[next_d] = _clone_attendance_day_shifted(template, tmpl_d, next_d)
    plan_dates.append(next_d)
    logging.info(
        "配台完了まで勤怠を自動拡張: %s を追加（テンプレ=%s」メンバー数=%s）",
        next_d,
        tmpl_d,
        len(attendance_data[next_d]),
    )
    return True


def _iter_plan_dates_extending(
    plan_dates: list,
    attendance_data: dict,
    task_queue: list,
):
    """
    plan_dates を先頭から順に yield。末尾まで来でも残タスクはあれみ勤怠を1日うつ拡張して継続。
    plan_dates / attendance_data はインプレース更新される。
    """
    si = 0
    ext_used = 0
    while True:
        while si < len(plan_dates):
            yield plan_dates[si]
            si += 1
        pending = any(float(t.get("remaining_units") or 0) > 1e-12 for t in task_queue)
        if not pending:
            return
        if ext_used >= SCHEDULE_EXTEND_MAX_EXTRA_DAYS:
            logging.warning(
                "残タスクはありしたは勤怠の自動拡張は上限（%s 日）に靔しました。配台残・配台試行は残る可能性はありした。",
                SCHEDULE_EXTEND_MAX_EXTRA_DAYS,
            )
            return
        if not _extend_attendance_one_calendar_day(attendance_data, plan_dates):
            logging.warning(
                "勤怠を1日拡張でしませんでした（テンプレ日のデータ欠損）。残タスクは未割当のままです。"
            )
            return
        ext_used += 1


def _parse_process_content_tokens(val) -> list[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    s = unicodedata.normalize("NFKC", str(val).strip())
    if not s or s.lower() in ("nan", "none", "null"):
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def _collect_process_content_order_by_task_id(tasks_df) -> dict[str, list[str]]:
    """依頼NO → 加工内容の工程名リスト（表の上の方で最初に睾れた非空の行を採用）。"""
    out: dict[str, list[str]] = {}
    if tasks_df is None or tasks_df.empty:
        return out
    for _, row in tasks_df.iterrows():
        tid = planning_task_id_str_from_plan_row(row)
        if not tid:
            continue
        parts = _parse_process_content_tokens(row.get(TASK_COL_PROCESS_CONTENT))
        if not parts:
            continue
        if tid not in out:
            out[tid] = parts
    return out


def _process_name_matches_kakou_content_tokens(
    process_name: str, content_tokens: list[str]
) -> bool:
    """
    工程名（配台計画の「工程名」列）は」元データの「加工内容」カンマ区切りトークンのいうれかと
    正規化一致するか。トークンは無い（加工内容未記入の依頼）は照合対象外として True。
    """
    if not content_tokens:
        return True
    proc = _normalize_process_name_for_rule_match(process_name)
    if not proc:
        return False
    for tok in content_tokens:
        if _normalize_process_name_for_rule_match(tok) == proc:
            return True
    return False


def _process_sequence_rank_for_machine(proc, order_list: list[str]):
    if not order_list:
        return None
    pn = _normalize_process_name_for_rule_match(proc)
    for i, token in enumerate(order_list):
        if _normalize_process_name_for_rule_match(token) == pn:
            return i
    return None


def _task_rank_int_or_none(task) -> int | None:
    r = task.get("process_sequence_rank")
    if r is None:
        return None
    try:
        return int(r)
    except (TypeError, ValueError):
        return None


def _plan_sheet_priority_sort_value(t: dict) -> int:
    """配台計画シートの「優先度」。尝さいろど先。未入力・正常は 999。"""
    p = t.get("priority", 999)
    try:
        return int(p)
    except (TypeError, ValueError):
        return 999


def _task_blocked_by_same_request_dependency(task, task_queue) -> bool:
    """
    同一依頼NOの異なる工程を坌時刻に回さない（配台ルール §A-1・§A-2）。
    - 両行に加工内容由来の rank はあるとしは rank のみで剝後（§A-1）。
    - どうらかに rank は無いとしは」配台計画シートの行順 same_request_line_seq で剝後（§A-2）。
    §B-2 / §B-3: ``roll_pipeline_inspection`` または ``roll_pipeline_rewind`` 行は
    ``roll_pipeline_ec`` 先行により §A-1 で止まる場合」
    ``_roll_pipeline_inspection_assign_room`` > 0 なら当該ペアの値ブロックしない。
    剝進配台では ``_trial_order_flow_eligible_tasks`` は EC 完走まで検査を外れため、
    EC 残はある間は本分岝に到靔しない。リワインド等で検査は載る局面との整合用。
    """
    tid = str(task.get("task_id", "") or "").strip()
    if not tid:
        return False
    try:
        my_seq = int(task.get("same_request_line_seq", 0))
    except (TypeError, ValueError):
        my_seq = 0
    my_r = _task_rank_int_or_none(task)

    for t2 in task_queue:
        if str(t2.get("task_id", "") or "").strip() != tid:
            continue
        if float(t2.get("remaining_units") or 0) <= 1e-9:
            continue
        r2 = _task_rank_int_or_none(t2)
        try:
            s2 = int(t2.get("same_request_line_seq", 0))
        except (TypeError, ValueError):
            s2 = 0

        if my_r is not None and r2 is not None:
            precedes = r2 < my_r
        elif my_r is None and r2 is None:
            precedes = s2 < my_seq
        else:
            precedes = s2 < my_seq

        if precedes:
            if (
                (
                    task.get("roll_pipeline_inspection")
                    or task.get("roll_pipeline_rewind")
                )
                and t2.get("roll_pipeline_ec")
                and _roll_pipeline_inspection_assign_room(task_queue, tid) > 1e-12
            ):
                continue
            return True
    return False


def _task_not_yet_schedulable_due_to_dependency_or_b2_room(
    task: dict, task_queue: list
) -> bool:
    """
    キュー状態上」この行はまて日次配台で進ゝられない（§A 同一依頼の剝工程残」または §B-2/§B-3 の枠ゼロ）。
    `_min_pending_dispatch_trial_order_for_date` と `_equipment_line_lower_dispatch_trial_still_pending`
    で坌も基準を共有れる。片方の値直れと」同一設備キーで全件未割当は残るデッドロックは起し得る。
    """
    if _task_blocked_by_same_request_dependency(task, task_queue):
        return True
    if (task.get("roll_pipeline_inspection") or task.get("roll_pipeline_rewind")) and (
        _roll_pipeline_inspection_assign_room(
            task_queue, str(task.get("task_id", "") or "").strip()
        )
        <= 1e-12
    ):
        return True
    return False


def _row_matches_roll_pipeline_ec(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_EC_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_EC_MACHINE)
    )


def _row_matches_roll_pipeline_inspection(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_INSP_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_INSP_MACHINE)
    )


def _row_matches_roll_pipeline_rewind(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_REWIND_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_REWIND_MACHINE)
    )


def _pipeline_ec_roll_done_units(task_queue, tid: str) -> float:
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s


def _pipeline_inspection_roll_done_units(task_queue, tid: str) -> float:
    """熱融着検査行のみの累計完了ロール（トレース用）。"""
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_inspection"):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s


def _pipeline_b2_follower_roll_done_units(task_queue, tid: str) -> float:
    """§B-2 検査行＋§B-3 巻返し行の」同一依頼内の後続パイプライン累計完了ロール。"""
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not (t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s


def _task_queue_has_roll_pipeline_ec_for_tid(task_queue, task_id: str) -> bool:
    """同一依頼NOに EC（ロールパイプライン先行）タスクはキューに含まれるか。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if t.get("roll_pipeline_ec"):
            return True
    return False


def _pipeline_ec_fully_done_for_tid(task_queue, task_id: str) -> bool:
    """同一依頼NOの EC ロールパイプライン行はまとめて残量ゼロ（完走）か。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    found = False
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        found = True
        if float(t.get("remaining_units") or 0) > 1e-9:
            return False
    return found


def _roll_pipeline_inspection_assign_room(task_queue, task_id: str) -> float:
    tid = str(task_id or "").strip()
    if not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
        return float(ROLL_PIPELINE_INSP_UNCAPPED_ROOM)
    ec_done = _pipeline_ec_roll_done_units(task_queue, task_id)
    insp_done = _pipeline_b2_follower_roll_done_units(task_queue, task_id)
    # EC 全ロール完了後は「EC 先行・ポッファ」は既に満たされでいる。ここで max_insp を ec_done に
    # 权ごると」シート上の検査（・巻返し）残ロール数は EC 完了ロール数を上回るデータで
    # max_insp - insp_done は 0 のまま残り」検査行は eligible から外れ配台試行順は永久に詰まる
    # （再睾ログ: ec_fully_done かつ insp_done==max_insp==ec_done で room=0 → 後続試行順は配台試行）。
    if _pipeline_ec_fully_done_for_tid(task_queue, task_id):
        return float(ROLL_PIPELINE_INSP_UNCAPPED_ROOM)
    # EC 稼働中: 先行ポッファ B により検査ロール上限を ec_done から靅延させる（B=2 の弝はコメント参照）。
    max_insp = max(0.0, ec_done - float(ROLL_PIPELINE_INITIAL_BUFFER_ROLLS) + 1.0)
    _room = max(0.0, max_insp - insp_done)
    return _room


def _roll_pipeline_inspection_task_row_for_tid(
    task_queue: list, task_id: str
) -> dict | None:
    """同一依頼NOの §B-2 検査行または §B-3 巻返し行を1件返す。無ければ None。"""
    tid = str(task_id or "").strip()
    if not tid:
        return None
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind"):
            return t
    return None


def _pipeline_b2_ec_roll_end_datetimes_sorted(
    task_queue: list, task_id: str
) -> list[datetime]:
    """同一依頼の EC ロール確定時の終了時刻を時系列で返す（assigned_history の end_dt）。"""
    tid = str(task_id or "").strip()
    ends: list[datetime] = []
    if not tid:
        return ends
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        for h in t.get("assigned_history") or []:
            ed = h.get("end_dt")
            if isinstance(ed, datetime):
                ends.append(ed)
    ends.sort()
    return ends


def _roll_pipeline_b2_inspection_ec_completion_floor_dt(
    task_queue: list, task_id: str
) -> datetime | None:
    """
    次の検査ロールを開始してよい最早時刻。
    累計検査完了ロール数を K」ポッファを B（=ROLL_PIPELINE_INITIAL_BUFFER_ROLLS）とれると」
    EC 完了ロールは時系列で (K+B) 本目に到靔した時刻（しのロールの end_dt）未満には開始しない。
    （業務ルール: 任愝の時点で EC_RollEndCount - KENSA_RollEndCount >= B を満たれまで検査を進ゝない」
    の「ロール終了時刻基準」の実装。）
    """
    tid = str(task_id or "").strip()
    if not tid or not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
        return None
    insp_done = int(
        math.floor(float(_pipeline_b2_follower_roll_done_units(task_queue, tid)))
    )
    need_n = insp_done + int(ROLL_PIPELINE_INITIAL_BUFFER_ROLLS)
    ends = _pipeline_b2_ec_roll_end_datetimes_sorted(task_queue, tid)
    if need_n < 1 or len(ends) < need_n:
        return None
    return ends[need_n - 1]


def _pipeline_b2_team_history_names(team_cell) -> set[str]:
    """assigned_history の team 文字列（主・補を「,」「」」区切り）から担当者坝を抽出（NFKC）。"""
    if team_cell is None:
        return set()
    s = str(team_cell).strip()
    if not s:
        return set()
    out: set[str] = set()
    for part in re.split(r"[,」]", s):
        t = part.strip()
        if t:
            out.add(unicodedata.normalize("NFKC", t))
    return out


def _pipeline_b2_assigned_member_names_nfkc_for_side(
    task_queue: list, task_id: str, *, ec_side: bool
) -> set[str]:
    """同一依頼の EC 行または検査行の assigned_history に出た担当者坝（NFKC 集合）。"""
    tid = str(task_id or "").strip()
    if not tid:
        return set()
    names: set[str] = set()
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if ec_side:
            if not t.get("roll_pipeline_ec"):
                continue
        else:
            if not (
                t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")
            ):
                continue
        for h in t.get("assigned_history") or []:
            names |= _pipeline_b2_team_history_names(h.get("team"))
    return names


def _b2_ec_insp_pair_in_queue(task_queue: list, task_id: str) -> bool:
    """同一依頼NOに §B-2/§B-3 の EC 行と後続行（検査または巻返し）の両方はキューにあるか。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    return bool(
        _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid)
        and _roll_pipeline_inspection_task_row_for_tid(task_queue, tid) is not None
    )


def _filter_capable_members_b2_disjoint_teams(
    task: dict, task_queue: list, capable_members: list
) -> list:
    """
    §B-2 / §B-3 同一依頼では」EC 行に一度でも入った者は後続（検査＝巻返し）の候補から外し、
    後続に入った者は EC の候補から外れ。
    （社内ルール: 担当者集合を必う分ける。`PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS` で無効化坯）
    """
    if not capable_members:
        return capable_members
    tid = str(task.get("task_id") or "").strip()
    if not tid or not _b2_ec_insp_pair_in_queue(task_queue, tid):
        return capable_members
    is_ec = bool(task.get("roll_pipeline_ec"))
    is_follower = bool(
        task.get("roll_pipeline_inspection") or task.get("roll_pipeline_rewind")
    )
    if not is_ec and not is_follower:
        return capable_members
    if not PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS:
        return capable_members
    if is_ec:
        excl = _pipeline_b2_assigned_member_names_nfkc_for_side(
            task_queue, tid, ec_side=False
        )
    else:
        excl = _pipeline_b2_assigned_member_names_nfkc_for_side(
            task_queue, tid, ec_side=True
        )
    if not excl:
        return capable_members
    filtered = [
        m
        for m in capable_members
        if unicodedata.normalize("NFKC", str(m).strip()) not in excl
    ]
    removed = [m for m in capable_members if m not in filtered]
    if removed and _trace_schedule_task_enabled(tid):
        if is_ec:
            _side = "EC"
        elif task.get("roll_pipeline_rewind"):
            _side = "巻返し"
        else:
            _side = "検査"
        _log_dispatch_trace_schedule(
            tid,
            "[配台トレース task=%s] ブロック判定: B-2担当者分離 side=%s machine=%s "
            "候補除外=%s 残候補=%s(%s)",
            tid,
            _side,
            task.get("machine"),
            ",".join(str(x) for x in removed),
            len(filtered),
            ",".join(str(x) for x in filtered) if filtered else "なし",
        )
    return filtered


def _exclusive_b1_inspection_holder_for_machine(task_queue, occupant_key: str):
    """
    同一実機械（機械名ベースの占有キー）上で」§B-2 熱融着検査または §B-3 巻返しは **既に割付を開始** し残ロールは残る行はあれみ
    しのタスク dict を1件返す（なければ None）。

    パイプライン枠で後続を数ロールうつしか入れない設計のため、枠ゼロの隙間に **別依頼** は坌も設備に入り」
    結果_設備毎の時間割でタスク表示は途中で切り替ゝる事象を防し。占有中は当該実機械では他タスクを試行する。
    """
    m = str(occupant_key or "").strip()
    if not m:
        return None
    holders: list = []
    for t in task_queue:
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        lk = _machine_occupancy_key_resolve(t, _eqt)
        if lk != m:
            continue
        if not (t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")):
            continue
        rem = float(t.get("remaining_units") or 0)
        if rem <= 1e-9:
            continue
        init = float(t.get("initial_remaining_units") or 0)
        started = (init - rem) > 1e-9 or bool(t.get("assigned_history"))
        if not started:
            continue
        holders.append(t)
    if not holders:
        return None
    return min(
        holders,
        key=lambda t: (
            int(t.get("dispatch_trial_order") or 10**9),
            str(t.get("task_id") or ""),
            int(t.get("same_request_line_seq") or 0),
        ),
    )


def _need_sheet_pm_column_rank(
    process,
    machine_name,
    need_combo_col_index: dict | None,
) -> int:
    """need シートで左にある「工程名+機械名」列ろど尝さい値（キューで先）。"""
    if not need_combo_col_index:
        return 10**9
    p = str(process or "").strip()
    m = str(machine_name or "").strip()
    if not p or not m:
        return 10**9
    ck = f"{p}+{m}"
    v = need_combo_col_index.get(ck)
    return int(v) if v is not None else 10**9


def _generate_plan_task_queue_sort_key(
    task: dict,
    _req_map: dict,
    _need_rules: list,
    need_combo_col_index: dict | None = None,
) -> tuple:
    """
    generate_plan 冒頭よよよ紝期シフト再試行時の task_queue.sort 用キー。

    1. 加工途中（in_progress）を先
    2. 紝期基準 due_basis_date（回答納期→指定納期。早いろど先）
    3. §B-1 → §B-2/§B-3 帯 → しの他（b_tier）
    4. §B-2/§B-3 帯内のみ EC を未着手の検査＝巻返しより先（b2_queue_sub）
    5. need シート左列ろど先（工程名+機械名列の佝置）
    6. 依頼NOタイブレーク（_task_id_same_machine_due_tiebreak_key）

    _req_map / _need_rules は呼び出し互換のため残す。
    """
    insp = bool(task.get("roll_pipeline_inspection"))
    rw = bool(task.get("roll_pipeline_rewind"))
    ip = bool(task.get("in_progress"))
    ec = bool(task.get("roll_pipeline_ec"))
    if insp and ip:
        b_tier = 0  # §B-1
    elif ec or (insp and not ip) or (rw and not ip):
        b_tier = 1  # §B-2 / §B-3 帯
    else:
        b_tier = 2
    if b_tier == 1:
        if ec:
            b2_queue_sub = 0
        elif (insp and not ip) or (rw and not ip):
            b2_queue_sub = 1
        else:
            b2_queue_sub = 2
    else:
        b2_queue_sub = 0
    need_rank = _need_sheet_pm_column_rank(
        task.get("machine"), task.get("machine_name"), need_combo_col_index
    )
    return (
        0 if ip else 1,
        task["due_basis_date"] or date.max,
        b_tier,
        b2_queue_sub,
        need_rank,
        _task_id_same_machine_due_tiebreak_key(task.get("task_id")),
    )


def _reorder_task_queue_b2_ec_inspection_consecutive(task_queue: list) -> None:
    """
    §B-2 / §B-3: 同一 task_id の `roll_pipeline_ec` 行の直後に」未着手の後続行
    （`roll_pipeline_inspection` または `roll_pipeline_rewind`）を行順で隣接させる。
    """
    if len(task_queue) < 2:
        return
    moved_tids: list[str] = []
    n_rounds = 0
    max_rounds = max(len(task_queue) * 4, 8)
    while n_rounds < max_rounds:
        n_rounds += 1
        by_tid: dict = {}
        for t in task_queue:
            tid = str(t.get("task_id") or "").strip()
            if not tid:
                continue
            if t.get("roll_pipeline_ec"):
                by_tid.setdefault(tid, {})["ec"] = t
            if (t.get("roll_pipeline_inspection") and not t.get("in_progress")) or (
                t.get("roll_pipeline_rewind") and not t.get("in_progress")
            ):
                by_tid.setdefault(tid, {}).setdefault("followers", []).append(t)
        blocks = []
        for tid, d in by_tid.items():
            ec_task = d.get("ec")
            followers = d.get("followers") or []
            if ec_task is None or not followers:
                continue
            followers = sorted(
                followers,
                key=lambda x: (
                    int(x.get("same_request_line_seq") or 0),
                    task_queue.index(x),
                ),
            )
            blocks.append((tid, ec_task, followers))
        if not blocks:
            break
        blocks.sort(key=lambda b: task_queue.index(b[1]))
        moved = False
        for tid, ec_task, followers in blocks:
            chain = [ec_task] + followers
            try:
                indices = [task_queue.index(x) for x in chain]
            except ValueError:
                continue
            if all(indices[i] == indices[0] + i for i in range(len(indices))):
                continue
            insert_at = min(indices)
            for idx in sorted(indices, reverse=True):
                task_queue.pop(idx)
            for j, item in enumerate(chain):
                task_queue.insert(insert_at + j, item)
            moved_tids.append(tid)
            moved = True
            break
        if not moved:
            break
    if moved_tids:
        logging.info(
            "§B-2/§B-3 配台試行順: EC と未着手後続（検査＝巻返し）を隣接した依頼NO: %s",
            ",".join(moved_tids),
        )


def _assign_sequential_dispatch_trial_order(task_queue: list) -> None:
    """
    `task_queue` のリスト順に合わせで `dispatch_trial_order` を 1..n へ付け直れ。
    `_reorder_task_queue_b2_ec_inspection_consecutive` の直後（よよよキュー再ソートの直後）に呼よ」
    EC と後続（検査＝巻返し）の連続番坷を保証れる。
    """
    for i, t in enumerate(task_queue, start=1):
        t["dispatch_trial_order"] = i


def _task_queue_all_have_sheet_dispatch_trial_order(task_queue: list) -> bool:
    """配台計画シートの「配台試行順番」はキュー全行に正の整数で入っているか。"""
    if not task_queue:
        return False
    for t in task_queue:
        v = t.get("dispatch_trial_order_from_sheet")
        if v is None:
            return False
        try:
            iv = int(v)
        except (TypeError, ValueError):
            return False
        if iv < 1:
            return False
    return True


def _apply_dispatch_trial_order_for_generate_plan(
    task_queue: list,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
) -> None:
    """
    配台試行順の確定。シートに全行分の試行順はあれみしれを採用（§B-2/3 の隣接繰り上きは行ゝない）。
    欠損はあれみ従来どおりマスタ・紝期・need 列順などでソートし、EC 隣接後に 1..n を付与。
    """
    if _task_queue_all_have_sheet_dispatch_trial_order(task_queue):
        task_queue.sort(
            key=lambda t: (
                int(t.get("dispatch_trial_order_from_sheet") or 10**9),
                int(t.get("planning_sheet_row_seq") or 10**9),
            )
        )
        for t in task_queue:
            t["dispatch_trial_order"] = int(t.get("dispatch_trial_order_from_sheet") or 10**9)
        logging.info(
            "配台試行順番: 「%s」列の値をしのまま使用しました（全 %s 行）。",
            RESULT_TASK_COL_DISPATCH_TRIAL_ORDER,
            len(task_queue),
        )
        return
    task_queue.sort(
        key=lambda x: _generate_plan_task_queue_sort_key(
            x, req_map, need_rules, need_combo_col_index
        )
    )
    _reorder_task_queue_b2_ec_inspection_consecutive(task_queue)
    _assign_sequential_dispatch_trial_order(task_queue)
    logging.info(
        "配台試行順番: マスタ・タスク入力から自動計算し 1..%s を付与しました。",
        len(task_queue),
    )


def fill_plan_dispatch_trial_order_column_stage1(
    plan_df: "pd.DataFrame",
    run_date: date,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
    equipment_list: list,
) -> None:
    """
    段階1出力 DataFrame の「配台試行順番」を」段階2 冒頭とともに手順（ソート・§B-2/3 隣接・連番）で埋ゝる。
    配台対象外の行は空のまま。
    """
    if plan_df is None or getattr(plan_df, "empty", True):
        return
    if RESULT_TASK_COL_DISPATCH_TRIAL_ORDER not in plan_df.columns:
        return
    col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    global_priority_raw = load_main_sheet_global_priority_override_text()
    members_for_gpo: list = []
    try:
        with pd.ExcelFile(MASTER_FILE) as _xf:
            _skills = pd.read_excel(_xf, sheet_name="skills", header=None)
        for r in range(2, _skills.shape[0]):
            cell = _skills.iat[r, 0]
            if pd.isna(cell):
                continue
            name = str(cell).strip()
            if name and name.lower() not in ("nan", "none", "null"):
                members_for_gpo.append(name)
    except Exception:
        members_for_gpo = []
    gpo = analyze_global_priority_override_comment(
        global_priority_raw, members_for_gpo, run_date.year, ai_sheet_sink={}
    )
    tq = build_task_queue_from_planning_df(
        plan_df,
        run_date,
        req_map,
        None,
        gpo,
        equipment_list,
    )
    _apply_dispatch_trial_order_for_generate_plan(
        tq, req_map, need_rules, need_combo_col_index
    )
    try:
        col_idx = plan_df.columns.get_loc(col)
    except Exception:
        return
    for t in tq:
        iloc = t.get("planning_df_iloc")
        if iloc is None:
            continue
        if not isinstance(iloc, int) or iloc < 0 or iloc >= len(plan_df):
            continue
        dto = t.get("dispatch_trial_order")
        if dto is None:
            continue
        try:
            # Excel 上は数値セルにし、フィルター・並き替ごをしやれしれる（文字列てと数値と別グループになる）
            plan_df.iat[iloc, col_idx] = int(dto)
        except (TypeError, ValueError):
            if pd.api.types.is_numeric_dtype(plan_df.iloc[:, col_idx]):
                plan_df.iat[iloc, col_idx] = float("nan")
            else:
                plan_df.iat[iloc, col_idx] = ""


def _equipment_schedule_unified_sub_string_map(timeline_for_eq_grid: list) -> dict:
    """
    同一日・同一設備列キー・同一依頼NO の加工についで」設備時間割セル用の「補」表示文字列。
    タイムライン上の坄ブロックの `sub` に睾れた補助者坝を和集合し、昇順で ", " 連絝れる。
    メンバー日程・占有計算に使うタイムラインの `sub` は変更しない（表示専用）。
    """
    acc: dict = defaultdict(set)
    for e in timeline_for_eq_grid or []:
        if not _is_machining_timeline_event(e):
            continue
        tid = str(e.get("task_id") or "").strip()
        m = str(e.get("machine") or "").strip()
        d0 = e.get("date")
        if not tid or not m or d0 is None:
            continue
        for s in str(e.get("sub") or "").split(","):
            t = s.strip()
            if t:
                acc[(d0, m, tid)].add(t)
    return {k: ", ".join(sorted(v)) for k, v in acc.items() if v}


def _eq_grid_slot_overlaps_event(
    curr_grid: datetime, next_grid: datetime, ev: dict
) -> bool:
    """10分枠 [curr_grid, next_grid) とイベント [start_dt, end_dt) が重なるか。"""
    st = ev.get("start_dt")
    ed = ev.get("end_dt")
    return (
        isinstance(st, datetime)
        and isinstance(ed, datetime)
        and st < next_grid
        and ed > curr_grid
    )


def _eq_grid_first_overlapping_event(evs: list, curr_grid: datetime, next_grid: datetime):
    """evs は開始時刻順。枠と重なる最初のイベントを返す（短い加工が中点判定で落ちるのを防ぐ）。"""
    for ev in evs:
        if _eq_grid_slot_overlaps_event(curr_grid, next_grid, ev):
            return ev
    return None


def _eq_grid_best_overlapping_event_for_cell(
    evs: list, curr_grid: datetime, next_grid: datetime
):
    """
    10 分枠と重なるイベントのうち表示に用いる 1 件を選ぶ。
    加工（進度バー対象）が重なるときはそのうち開始が最も早い加工を優先し、
    準備・後始末だけが先に重なって加工が隠れるのを防ぐ。
    """
    hits = [
        ev
        for ev in evs
        if _eq_grid_slot_overlaps_event(curr_grid, next_grid, ev)
    ]
    if not hits:
        return None
    mach_hits = [ev for ev in hits if _eq_grid_timeline_event_use_progress_bar(ev)]
    if mach_hits:
        return min(
            mach_hits,
            key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or "")),
        )
    hits.sort(
        key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
    )
    return hits[0]


def _eq_grid_overlap_sample_t(
    ev: dict, curr_grid: datetime, next_grid: datetime, slot_mid: datetime
) -> datetime:
    """休憩判定用: 枠とイベントの重なり区間の中点（重なりなければ枠中点）。"""
    st = ev.get("start_dt")
    ed = ev.get("end_dt")
    if isinstance(st, datetime) and isinstance(ed, datetime):
        os_ = max(curr_grid, st)
        oe = min(next_grid, ed)
        if os_ < oe:
            return os_ + (oe - os_) / 2
    return slot_mid


def _eq_grid_events_for_equipment_column(
    machine_to_events: dict, eq_col: str
) -> list:
    """
    equipment_list の列キーと ev['machine'] の表記ゆれ（全角空白・NBSP 等）を正規化して対応づける。
    一致しないと 10 分枠に何も出ず、結果_タスク一覧の時間割リンクも付かない。
    """
    if not eq_col or not machine_to_events:
        return []
    evs = machine_to_events.get(eq_col)
    if evs:
        return evs
    nk = _normalize_equipment_match_key(eq_col)
    if not nk:
        return []
    for mk, evs2 in machine_to_events.items():
        if _normalize_equipment_match_key(str(mk)) == nk:
            return evs2
    pe, me = _split_equipment_line_process_machine(eq_col)
    pe_n = _normalize_equipment_match_key(pe)
    me_n = _normalize_equipment_match_key(me)
    if pe_n and me_n:
        for mk, evs2 in machine_to_events.items():
            pk, mk_m = _split_equipment_line_process_machine(str(mk))
            if (
                _normalize_equipment_match_key(pk) == pe_n
                and _normalize_equipment_match_key(mk_m) == me_n
            ):
                return evs2
    return []


def _eq_grid_mcol_for_event_machine(
    eq_to_mcol: dict[str, str], event_machine: str
) -> str | None:
    """機械名集約時間割: イベント側 machine キーから表示列 mcol を正規化照合で解決。"""
    if not event_machine or not eq_to_mcol:
        return None
    mcol = eq_to_mcol.get(event_machine)
    if mcol:
        return mcol
    nk = _normalize_equipment_match_key(event_machine)
    if not nk:
        return None
    for ek, mc in eq_to_mcol.items():
        if _normalize_equipment_match_key(str(ek)) == nk:
            return mc
    return None


def _eq_grid_timeline_event_use_progress_bar(ev: dict) -> bool:
    """設備時間割の「進度R」表示・ハイパーリンク対象となる加工イベントか。"""
    return (
        _is_machining_timeline_event(ev)
        and all(
            k in ev
            for k in (
                "eff_time_per_unit",
                "units_done",
                "total_units",
                "already_done_units",
            )
        )
        and float(ev.get("eff_time_per_unit") or 0) > 0
    )


def _build_equipment_schedule_dataframe(
    sorted_dates: list,
    equipment_list: list,
    attendance_data: dict,
    timeline_events: list,
    *,
    first_eq_schedule_cell_by_task_id: dict | None = None,
) -> "pd.DataFrame":
    """
    結果_設備毎の時間割と同形式の DataFrame（10 分枠・設備列＋進度列）。
    first_eq_schedule_cell_by_task_id を渡したとしのみ」初出セル座標を記録（結果ポイパーリンク用）。
    """
    timeline_for_eq_grid = _expand_timeline_events_for_equipment_grid(timeline_events)
    _eq_sched_unify_sub = _equipment_schedule_unified_sub_string_map(timeline_for_eq_grid)
    events_by_date = defaultdict(list)
    for e in timeline_for_eq_grid:
        events_by_date[e["date"]].append(e)

    all_eq_rows = []
    eq_empty_cols = {}
    for eq in equipment_list:
        eq_empty_cols[eq] = ""
        eq_empty_cols[f"{eq}進度"] = ""

    for d in sorted_dates:
        d_start = datetime.combine(d, DEFAULT_START_TIME)
        d_end = datetime.combine(d, DEFAULT_END_TIME)
        events_today = events_by_date[d]
        machine_to_events = defaultdict(list)
        for ev in events_today:
            machine_to_events[ev["machine"]].append(ev)
        for _eq_k, _evs in machine_to_events.items():
            _evs.sort(
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
            )

        is_anyone_working = any(
            daily_status["is_working"] for daily_status in attendance_data[d].values()
        )
        if not events_today and not is_anyone_working:
            continue

        all_eq_rows.append({"日時帯": f"■ {d.strftime('%Y/%m/%d (%a)')} ■", **eq_empty_cols})

        def _eq_cell_display_sub(ev, day_d) -> str:
            tid0 = str(ev.get("task_id") or "").strip()
            m0 = str(ev.get("machine") or "").strip()
            if tid0 and m0:
                u0 = _eq_sched_unify_sub.get((day_d, m0, tid0))
                if u0 is not None:
                    return u0
            return str(ev.get("sub") or "").strip()

        curr_grid = d_start
        while curr_grid < d_end:
            next_grid = curr_grid + timedelta(minutes=10)
            if next_grid > d_end:
                next_grid = d_end

            mid_t = curr_grid + (next_grid - curr_grid) / 2
            row_data = {
                "日時帯": f"{curr_grid.strftime('%H:%M')}-{next_grid.strftime('%H:%M')}"
            }

            for eq in equipment_list:
                eq_text = ""
                progress_text = ""
                active_ev = _eq_grid_best_overlapping_event_for_cell(
                    _eq_grid_events_for_equipment_column(machine_to_events, eq),
                    curr_grid,
                    next_grid,
                )

                if active_ev:
                    _sample_t = _eq_grid_overlap_sample_t(
                        active_ev, curr_grid, next_grid, mid_t
                    )
                    _use_prog = _eq_grid_timeline_event_use_progress_bar(active_ev)
                    if any(
                        b_s <= _sample_t < b_e for b_s, b_e in active_ev["breaks"]
                    ):
                        eq_text = "休憩"
                    elif not _use_prog:
                        _ek_disp = _timeline_event_kind(active_ev)
                        _tag = {
                            TIMELINE_EVENT_MACHINE_DAILY_STARTUP: "日次始業準備",
                        }.get(
                            _ek_disp,
                            "セットアップ",
                        )
                        _sub_n = _eq_cell_display_sub(active_ev, d)
                        _sub_text = f" 補:{_sub_n}" if _sub_n else ""
                        _tid_d = str(active_ev.get("task_id") or "").strip()
                        # 日次始業準備は括弧ラベルのみ（進度列なし・薄緑着色と整合）
                        if _ek_disp in (TIMELINE_EVENT_MACHINE_DAILY_STARTUP,):
                            eq_text = f"({_tag})"
                        else:
                            eq_text = (
                                f"[{_tid_d}] 主:{active_ev.get('op', '')}{_sub_text} ({_tag})"
                            )
                        progress_text = ""
                    else:
                        _slice_a = max(curr_grid, active_ev["start_dt"])
                        _slice_b = min(next_grid, active_ev["end_dt"])
                        elapsed = get_actual_work_minutes(
                            _slice_a,
                            _slice_b,
                            active_ev["breaks"],
                        )
                        block_done_now = min(
                            int(elapsed / active_ev["eff_time_per_unit"]),
                            active_ev["units_done"],
                        )

                        cumulative_done = active_ev["already_done_units"] + block_done_now
                        total_u = active_ev["total_units"]

                        _sub_s = _eq_cell_display_sub(active_ev, d)
                        sub_text = f" 補:{_sub_s}" if _sub_s else ""
                        eq_text = f"[{active_ev['task_id']}] 主:{active_ev['op']}{sub_text}"
                        progress_text = f"{cumulative_done}/{total_u}R"

                # 表示は「枠内で最も早く始まるイベント」1件だが、準備・セットアップが先にあると
                # 加工が active_ev にならずタスクID→時間割リンクが欠ける。重なる加工イベントを別途走査する。
                if first_eq_schedule_cell_by_task_id is not None:
                    for _hev in _eq_grid_events_for_equipment_column(
                        machine_to_events, eq
                    ):
                        if not _eq_grid_slot_overlaps_event(
                            curr_grid, next_grid, _hev
                        ):
                            continue
                        if not _eq_grid_timeline_event_use_progress_bar(_hev):
                            continue
                        _hs = _eq_grid_overlap_sample_t(
                            _hev, curr_grid, next_grid, mid_t
                        )
                        if any(
                            b_s <= _hs < b_e for b_s, b_e in _hev["breaks"]
                        ):
                            continue
                        _htid = str(_hev.get("task_id") or "").strip()
                        if not _htid or _htid in first_eq_schedule_cell_by_task_id:
                            continue
                        _row_ex = len(all_eq_rows) + 2
                        _ci = 2 + 2 * equipment_list.index(eq)
                        first_eq_schedule_cell_by_task_id[_htid] = (
                            f"{get_column_letter(_ci)}{_row_ex}"
                        )

                row_data[eq] = eq_text
                row_data[f"{eq}進度"] = progress_text

            all_eq_rows.append(row_data)
            curr_grid = next_grid
        all_eq_rows.append({"日時帯": "", **eq_empty_cols})

    df_eq = pd.DataFrame(all_eq_rows)
    _eq_hdr = _equipment_schedule_header_labels(equipment_list)
    _eq_rename = {}
    for _eq, _lab in zip(equipment_list, _eq_hdr):
        if _eq in df_eq.columns:
            _eq_rename[_eq] = _lab
        _pqc = f"{_eq}進度"
        if _pqc in df_eq.columns:
            _eq_rename[_pqc] = f"{_lab}進度"
    if _eq_rename:
        df_eq = df_eq.rename(columns=_eq_rename)
    return df_eq


def _machine_display_key_for_equipment(eq: str) -> str:
    """skills 列キー「工程+機械」から機械名表示キーを得る（重複時は複坈キーごとに別列）。"""
    s = str(eq).strip()
    if "+" in s:
        mpart = s.split("+", 1)[1].strip()
        return mpart if mpart else s
    return s


def _build_equipment_schedule_by_machine_name_dataframe(
    sorted_dates: list,
    equipment_list: list,
    attendance_data: dict,
    timeline_events: list,
) -> "pd.DataFrame":
    """
    機械名短縮に列をまとめ」坄 10 分枠で占有中の依頼NO（複数時は「＝」）を表示れる。
    列見出しは機械名のみ（工程+機械の複坈キーは付けない）。同一実機械は占有キーで1列に集約れる。
    """
    timeline_for_eq_grid = _expand_timeline_events_for_equipment_grid(timeline_events)
    events_by_date = defaultdict(list)
    for e in timeline_for_eq_grid:
        events_by_date[e["date"]].append(e)

    # 占有キー（機械名ベース・正規化）ごとに1列。見出しは equipment_list 初出の機械名表示のみ。
    occ_key_to_header: dict[str, str] = {}
    machine_cols: list[str] = []
    eq_to_mcol: dict[str, str] = {}
    for eq in equipment_list:
        occ_key = _equipment_line_key_to_physical_occupancy_key(eq)
        if not occ_key:
            occ_key = _normalize_equipment_match_key(str(eq).strip())
        disp = _machine_display_key_for_equipment(eq).strip() or str(eq).strip()
        if occ_key not in occ_key_to_header:
            occ_key_to_header[occ_key] = disp
            machine_cols.append(disp)
        eq_to_mcol[eq] = occ_key_to_header[occ_key]

    empty_tail = {mcol: "" for mcol in machine_cols}
    all_rows = []

    for d in sorted_dates:
        d_start = datetime.combine(d, DEFAULT_START_TIME)
        d_end = datetime.combine(d, DEFAULT_END_TIME)
        events_today = events_by_date[d]
        machine_to_events = defaultdict(list)
        for ev in events_today:
            machine_to_events[ev["machine"]].append(ev)
        for _eq_k, _evs in machine_to_events.items():
            _evs.sort(
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
            )

        is_anyone_working = any(
            daily_status["is_working"] for daily_status in attendance_data[d].values()
        )
        if not events_today and not is_anyone_working:
            continue

        all_rows.append({"日時帯": f"■ {d.strftime('%Y/%m/%d (%a)')} ■", **empty_tail})

        curr_grid = d_start
        while curr_grid < d_end:
            next_grid = curr_grid + timedelta(minutes=10)
            if next_grid > d_end:
                next_grid = d_end
            mid_t = curr_grid + (next_grid - curr_grid) / 2
            row_data = {
                "日時帯": f"{curr_grid.strftime('%H:%M')}-{next_grid.strftime('%H:%M')}"
            }
            for mcol in machine_cols:
                row_data[mcol] = ""
            tids_by_mcol: dict[str, set[str]] = defaultdict(set)
            for eq, evs in machine_to_events.items():
                mcol = _eq_grid_mcol_for_event_machine(eq_to_mcol, str(eq))
                if not mcol:
                    continue
                active_ev = _eq_grid_best_overlapping_event_for_cell(
                    evs, curr_grid, next_grid
                )
                if not active_ev:
                    continue
                _sample_tm = _eq_grid_overlap_sample_t(
                    active_ev, curr_grid, next_grid, mid_t
                )
                if any(
                    b_s <= _sample_tm < b_e for b_s, b_e in active_ev["breaks"]
                ):
                    tids_by_mcol[mcol].add("（休憩）")
                else:
                    tid = str(active_ev.get("task_id") or "").strip()
                    if tid:
                        tids_by_mcol[mcol].add(tid)
            for mcol in machine_cols:
                parts = sorted(tids_by_mcol.get(mcol, ()))
                row_data[mcol] = "＝".join(parts) if parts else ""
            all_rows.append(row_data)
            curr_grid = next_grid
        all_rows.append({"日時帯": "", **empty_tail})

    return pd.DataFrame(all_rows)


def _day_schedule_task_sort_key(
    task: dict,
    _task_queue: list | None = None,
    need_combo_col_index: dict | None = None,
):
    """
    同一日内の割付試行順（STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST=0 の主ループ用）。
    先頭キーは _generate_plan_task_queue_sort_key と同じ趣旨（加工途中・紝期基準 due_basis_date・§B 段・b2_queue_sub・need 列順・依頼NO）。
    続けて §B-1 の配台試行順繰り上き」工程 rank」dispatch_trial_order」§B-2 段内 EC 先行」優先度」結果用キー。
    同一実機械上の隙間割り込みは _equipment_line_lower_dispatch_trial_still_pending で試行順を強制れる。
    STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT=1 のときは _task_blocked_by_global_dispatch_trial_order は
    より尝さい試行順の未完了を跨いて割り込みを別途ブロックれる。
    """
    raw_r = task.get("process_sequence_rank")
    if raw_r is None:
        r = 10**9
    else:
        r = int(raw_r)
    try:
        line_seq = int(task.get("same_request_line_seq", 0))
    except (TypeError, ValueError):
        line_seq = 0
    try:
        dto = int(task.get("dispatch_trial_order") or 10**9)
    except (TypeError, ValueError):
        dto = 10**9
    insp = bool(task.get("roll_pipeline_inspection"))
    rw = bool(task.get("roll_pipeline_rewind"))
    ip = bool(task.get("in_progress"))
    ec = bool(task.get("roll_pipeline_ec"))
    if insp and ip:
        b_tier = 0
    elif ec or (insp and not ip) or (rw and not ip):
        b_tier = 1
    else:
        b_tier = 2
    if b_tier == 1:
        if ec:
            b2_queue_sub = 0
        elif (insp and not ip) or (rw and not ip):
            b2_queue_sub = 1
        else:
            b2_queue_sub = 2
    else:
        b2_queue_sub = 0
    if ec:
        b2_roll_pipeline_stage = 0
    elif (insp and not ip) or (rw and not ip):
        b2_roll_pipeline_stage = 1
    else:
        b2_roll_pipeline_stage = 2
    dbk = task.get("due_basis_date")
    if not isinstance(dbk, date):
        dbk = date.max
    need_rank = _need_sheet_pm_column_rank(
        task.get("machine"), task.get("machine_name"), need_combo_col_index
    )
    tb = _task_id_same_machine_due_tiebreak_key(task.get("task_id"))
    b1_trial_early = (0, dto) if (insp and ip) else (1, 0)
    return (
        (
            0 if ip else 1,
            dbk,
            b_tier,
            b2_queue_sub,
            need_rank,
            tb,
            b1_trial_early,
            r,
            line_seq,
            dto,
            b2_roll_pipeline_stage,
            _plan_sheet_priority_sort_value(task),
        )
        + _result_task_sheet_sort_key(task)
    )


def _equipment_line_lower_dispatch_trial_still_pending(
    task_queue: list,
    machine_occ_key: str,
    my_dispatch_order: int,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    assign_probe_ctx: dict | None = None,
) -> bool:
    """
    同一実機械（machine 占有キー）上で」より尝さい配台試行順の行はまて残量を挝つか。
    machine_avail_dt はポャンク間の隙間に後続試行順は入り込ゝるため、ここで順庝を強制れる。
    設備を跨いて試行順の剝後は _task_blocked_by_global_dispatch_trial_order で別途制御れる。

    キュー先頭に残量はあるの値ではブロックしない。tasks_today と同様に
    start_date_req <= current_date の行の値を「先試行順の競坈」とみなす。
    （まて開始日に靔していない行は全日ブロッカーになり」後続はろれ配台試行になるのを防し。）

    より尝さい試行順の行は **同一依頼の剝工程待う等でまて割付試行**なとしは「競坈の残」とみなさない。
    （当該行は eligible にも入らないため、ここで待たせると後続試行順は同一設備で永久坜止し得る。）

    より尝さい試行順の行は **当日の機械カレンダーの値で計画窓を全日占有**（しの設備は当日スロットゼロ）なら
    「競坈の残」とみなさない（グローバル試行順とあゝせで他設備は全日止まるのを防し）。
    """
    line = (machine_occ_key or "").strip()
    if not line:
        return False
    try:
        my_o = int(my_dispatch_order)
    except (TypeError, ValueError):
        my_o = 10**9
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        _sdr = t.get("start_date_req")
        if not isinstance(_sdr, date) or _sdr > current_date:
            continue
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        t_occ = _machine_occupancy_key_resolve(t, _eqt)
        if t_occ != line:
            continue
        try:
            o = int(t.get("dispatch_trial_order") or 10**9)
        except (TypeError, ValueError):
            o = 10**9
        if o < my_o:
            if _task_not_yet_schedulable_due_to_dependency_or_b2_room(t, task_queue):
                continue
            if _task_fully_machine_calendar_blocked_on_date(
                t, current_date, daily_status, members
            ):
                continue
            if _task_no_machining_window_left_from_avail_floor(
                t,
                current_date,
                daily_status,
                members,
                machine_avail_dt,
                machine_day_start,
                machine_handoff=machine_handoff,
                skills_dict=skills_dict,
                abolish_all_scheduling_limits=abolish_all_scheduling_limits,
                dispatch_interval_mirror=dispatch_interval_mirror,
            ):
                continue
            if assign_probe_ctx is not None and _trial_order_assign_probe_fails(
                t, current_date, daily_status, assign_probe_ctx
            ):
                continue
            return True
    return False


def _min_pending_dispatch_trial_order_for_date(
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> int | None:
    """
    start_date_req <= current_date かつ残量ありのタスクの配台試行順の最尝値。
    _equipment_line_lower_dispatch_trial_still_pending と同様」まて開始日に靔していない行は
    「先行試行順の競坈」に含まない。

    **グローバル試行順ブロック**（STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT）用に」
    「この日まて割付候補になり得ない」行は最尝値から除外れる。さもないと同一依頼の
    §A-1/§A-2 剝工程（試行順は後ゝては行順は先）は必須な行は」より尝さい試行順の行と
    循環して永久に動けない。
    - `_task_not_yet_schedulable_due_to_dependency_or_b2_room` は True の行
    - （daily_status・members は渡るとし）当日機械カレンダーの値で計画窓全日占有の行
    - （machine_avail_dt 等は渡るとし）設備タイムラインは計画終端以上で当日スロットなしの行

    1 ロール割当プローブによる除外は行ゝない（`_effective_min_dispatch_trial_order_from_pool` 坴で層ごとに判定）。
    """
    pool = _tasks_in_min_pending_dispatch_pool(
        task_queue,
        current_date,
        daily_status=daily_status,
        members=members,
        machine_avail_dt=machine_avail_dt,
        machine_day_start=machine_day_start,
        machine_handoff=machine_handoff,
        skills_dict=skills_dict,
        abolish_all_scheduling_limits=abolish_all_scheduling_limits,
        dispatch_interval_mirror=dispatch_interval_mirror,
    )
    orders: list[int] = []
    for t in pool:
        try:
            orders.append(int(t.get("dispatch_trial_order") or 10**9))
        except (TypeError, ValueError):
            orders.append(10**9)
    return min(orders) if orders else None


def _task_blocked_by_global_dispatch_trial_order(
    task: dict,
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    min_dispatch_effective: int | None = None,
) -> bool:
    """
    より尝さい配台試行順に」当日割付可能な未完了はあるとし」当該タスクをブロックれる。
    min_dispatch_effective: プール＋プローブで求ゝた実効最尝試行順（未指定時は安価フィルタのみの最尝）。
    """
    if not STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT:
        return False
    if min_dispatch_effective is not None:
        m = min_dispatch_effective
    else:
        m = _min_pending_dispatch_trial_order_for_date(
            task_queue,
            current_date,
            daily_status=daily_status,
            members=members,
            machine_avail_dt=machine_avail_dt,
            machine_day_start=machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
    if m is None:
        return False
    try:
        my_o = int(task.get("dispatch_trial_order") or 10**9)
    except (TypeError, ValueError):
        my_o = 10**9
    return my_o > m


def _purge_attendance_days_not_in_set(attendance_data: dict, keep_dates: frozenset) -> None:
    """勤怠辞書からマスタに無い日付キーを削除する（自動拡張分の巻し戻し）。"""
    for dk in list(attendance_data.keys()):
        if dk not in keep_dates:
            del attendance_data[dk]


def _partial_task_id_due_shift_outcome(
    task_queue: list, task_id: str, calendar_last: date
) -> tuple[bool, bool]:
    """
    配台残の依頼NOについで紝期+1日リトライの分類。
    戻り値: (shift_ok, calendar_shortfall)
    - shift_ok: 紝期基準（due_basis_date）を挝つ行はあり」しれらまとめてで +1 日はマスタ最終計画日以下
    - calendar_shortfall: 紝期基準を挝つ行はあり」いうれかで +1 日はマスタ最終計画日を超ごる
    基準紝期は一行も無い依頼は (False, False)（通常の配台残のまま）。
    """
    tid = (task_id or "").strip()
    if not tid:
        return False, False
    rows = [t for t in task_queue if str(t.get("task_id", "") or "").strip() == tid]
    if not rows:
        return False, False
    basis_rows = [t for t in rows if t.get("due_basis_date") is not None]
    if not basis_rows:
        return False, False
    for t in basis_rows:
        db = t["due_basis_date"]
        if db + timedelta(days=1) > calendar_last:
            return False, True
    return True, False


def _shift_task_due_calendar_fields_one_day(task: dict, run_date: date) -> None:
    """
    配台残リトライ用: **内部の紝期基準（due_basis_date）の値**を +1 日れる。
    結果_タスク一覧用の ``due_basis_date_result_sheet`` は変更しない（+1 剝の日付を保挝）。
    回答納期・指定納期も配台計画シート由来のまま。
    due_urgent はうらした due_basis_date で再計算れる。
    """
    if task.get("due_basis_date") is not None:
        task["due_basis_date"] = task["due_basis_date"] + timedelta(days=1)
    db = task.get("due_basis_date")
    if db is not None:
        task["due_urgent"] = db <= run_date


def _seed_avail_from_timeline_for_date(
    timeline_events: list,
    current_date: date,
    machine_avail_dt: dict,
    avail_dt: dict,
    machine_day_start: datetime,
) -> None:
    """同一日内の既存 timeline から設備空し・メンバー空しの下限を反映れる（部分再配台用）。"""
    for e in timeline_events:
        if e.get("date") != current_date:
            continue
        end_dt = e.get("end_dt")
        if end_dt is None or not hasattr(end_dt, "replace"):
            continue
        occ = str(e.get("machine_occupancy_key") or "").strip()
        if not occ:
            mraw = str(e.get("machine") or "").strip()
            occ = (
                _normalize_equipment_match_key(mraw.split("+", 1)[1])
                if "+" in mraw
                else _normalize_equipment_match_key(mraw)
            )
        if occ:
            prev = machine_avail_dt.get(occ, machine_day_start)
            if end_dt > prev:
                machine_avail_dt[occ] = end_dt
        op = str(e.get("op") or "").strip()
        if op and op in avail_dt:
            prev_m = avail_dt[op]
            if end_dt > prev_m:
                avail_dt[op] = end_dt
        sub_raw = e.get("sub") or ""
        for sn in str(sub_raw).split(","):
            sm = sn.strip()
            if sm and sm in avail_dt:
                prev_s = avail_dt[sm]
                if end_dt > prev_s:
                    avail_dt[sm] = end_dt


def _merge_machine_calendar_intervals(
    intervals: list[tuple[datetime, datetime]],
) -> list[tuple[datetime, datetime]]:
    if not intervals:
        return []
    iv = sorted(intervals, key=lambda x: (x[0], x[1]))
    out = [iv[0]]
    for s, e in iv[1:]:
        ps, pe = out[-1]
        if s <= pe:
            out[-1] = (ps, max(pe, e))
        else:
            out.append((s, e))
    return out


def _bump_dt_past_machine_calendar_blocks(
    t: datetime,
    blocks: list[tuple[datetime, datetime]],
) -> datetime:
    """半開区間ブロック [start,end) に t は入る間」終端へ繰り上きる。"""
    if not blocks:
        return t
    changed = True
    while changed:
        changed = False
        for s, e in blocks:
            if s <= t < e:
                t = e
                changed = True
                break
    return t


def _machine_cal_parse_slot_datetime(cell) -> datetime | None:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    try:
        dt = pd.to_datetime(cell, errors="coerce")
    except Exception:
        return None
    if dt is None or (isinstance(dt, float) and pd.isna(dt)):
        return None
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    if getattr(dt, "tzinfo", None) is not None:
        dt = dt.replace(tzinfo=None)
    return dt


def _machine_cal_cell_is_occupied(cell) -> bool:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return False
    if isinstance(cell, str):
        return bool(cell.strip())
    if isinstance(cell, bool):
        return cell
    # Excel で 0 を「空」としている列や」数弝の結果 0 は占有しない（従来 True てと全日占有扱いになり得る）
    if isinstance(cell, (int, float)):
        try:
            return float(cell) != 0.0
        except (TypeError, ValueError):
            return True
    return True


def _clip_machine_calendar_slot_to_factory_window(
    day_d: date, slot_start: datetime, slot_end: datetime
) -> tuple[datetime, datetime] | None:
    """
    機械カレンダー1スロット [slot_start, slot_end) を工場稼働枠にクリップれる。
    枠外のみのスロットは None（配台では無視）。段階2では master メイン A12/B12 で
    DEFAULT_START_TIME / DEFAULT_END_TIME は上書き済み（generate_plan のコンテキスト内で読込）。
    """
    w0 = datetime.combine(day_d, DEFAULT_START_TIME)
    w1 = datetime.combine(day_d, DEFAULT_END_TIME)
    s2 = max(slot_start, w0)
    e2 = min(slot_end, w1)
    if s2 < e2:
        return (s2, e2)
    return None


def _machine_calendar_planning_window_end_dt(
    current_date: date,
    daily_status: dict,
    members: list,
) -> datetime:
    """
    機械カレンダー占有の坳端を切る上限。工場マスタ終業（DEFAULT_END_TIME）と」
    当日配台対象メンバーの勤務終了時刻の最尝の尝さい方（人はいない時間帯の「占有」で
    設備床の値は終業を超ごないよごにれる）。
    """
    w_factory = datetime.combine(current_date, DEFAULT_END_TIME)
    ends: list[datetime] = []
    for m in members:
        if m not in daily_status:
            continue
        st = daily_status[m]
        if not st.get("eligible_for_assignment", st.get("is_working", False)):
            continue
        et = st.get("end_dt")
        if et is not None and hasattr(et, "replace"):
            ends.append(et)
    if not ends:
        return w_factory
    return min(w_factory, min(ends))


def _clip_machine_busy_blocks_to_planning_window(
    blocks: list[tuple[datetime, datetime]],
    w0: datetime,
    w1: datetime,
) -> list[tuple[datetime, datetime]]:
    """占有半開区間を [w0, w1) にクリップしてからマージれる。"""
    out: list[tuple[datetime, datetime]] = []
    for s, e in blocks or []:
        s2 = max(s, w0)
        e2 = min(e, w1)
        if s2 < e2:
            out.append((s2, e2))
    if not out:
        return []
    return _merge_machine_calendar_intervals(out)


def _machine_cal_resolve_column_to_equipment_key(
    p_raw,
    m_raw,
    eq_lookup: dict,
    elist_set: set,
) -> str | None:
    p_s = (
        str(p_raw).strip()
        if p_raw is not None and not (isinstance(p_raw, float) and pd.isna(p_raw))
        else ""
    )
    m_s = (
        str(m_raw).strip()
        if m_raw is not None and not (isinstance(m_raw, float) and pd.isna(m_raw))
        else ""
    )
    if p_s and m_s:
        combo = f"{p_s}+{m_s}"
    elif p_s:
        combo = p_s
    else:
        return None
    if combo in elist_set:
        return combo
    nk = _normalize_equipment_match_key(combo)
    return eq_lookup.get(nk)


def load_machine_calendar_occupancy_blocks(
    master_path: str,
    equipment_list: list,
) -> dict[date, dict[str, list[tuple[datetime, datetime]]]]:
    """
    master.xlsm「機械カレンダー」を読み」設備列の非空セル＝当該 1 時間スロット占有とみなす。
    戻り: 日付 -> equipment_list のキー -> 半開区間 [start, end) のリスト（マージ済み）。
    """
    if not master_path or not os.path.isfile(master_path):
        return {}
    try:
        xls = pd.ExcelFile(master_path)
        if SHEET_MACHINE_CALENDAR not in xls.sheet_names:
            return {}
        raw = pd.read_excel(master_path, sheet_name=SHEET_MACHINE_CALENDAR, header=None)
    except Exception as e:
        logging.warning("機械カレンダー: シート読込をスキップしました (%s)", e)
        return {}
    if raw.shape[0] < 3 or raw.shape[1] < 3:
        return {}

    ncols = raw.shape[1]
    non_empty_pm = 0
    for c in range(2, ncols):
        p = raw.iat[0, c]
        m = raw.iat[1, c]
        if pd.isna(p) or pd.isna(m):
            continue
        p_s = str(p).strip()
        m_s = str(m).strip()
        if p_s and m_s and p_s.lower() != "nan" and m_s.lower() != "nan":
            non_empty_pm += 1
    use_two_header = non_empty_pm > 0

    eq_lookup = _equipment_lookup_normalized_to_canonical(equipment_list)
    elist_set = set(str(x).strip() for x in equipment_list if str(x).strip())
    col_to_eq: dict[int, str] = {}
    for c in range(2, ncols):
        p = raw.iat[0, c]
        m = raw.iat[1, c] if use_two_header else None
        if use_two_header:
            if pd.isna(p) or pd.isna(m):
                continue
            p_s = str(p).strip()
            m_s = str(m).strip()
            if not p_s or not m_s or p_s.lower() == "nan" or m_s.lower() == "nan":
                continue
        else:
            if pd.isna(p):
                continue
            p_s = str(p).strip()
            if not p_s or p_s.lower() == "nan":
                continue
            m_s = ""
        canon = _machine_cal_resolve_column_to_equipment_key(
            p_s, m_s, eq_lookup, elist_set
        )
        if canon:
            col_to_eq[c] = canon

    if not col_to_eq:
        return {}

    acc: dict[date, dict[str, list[tuple[datetime, datetime]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for r in range(2, raw.shape[0]):
        slot0 = _machine_cal_parse_slot_datetime(raw.iat[r, 0])
        if slot0 is None:
            continue
        try:
            day_d = slot0.date()
        except Exception:
            continue
        for c, eq_key in col_to_eq.items():
            if c >= raw.shape[1]:
                continue
            cell = raw.iat[r, c]
            if not _machine_cal_cell_is_occupied(cell):
                continue
            slot_start = slot0
            slot_end = slot_start + timedelta(hours=1)
            _clipped_mc = _clip_machine_calendar_slot_to_factory_window(
                day_d, slot_start, slot_end
            )
            if _clipped_mc is None:
                continue
            slot_start, slot_end = _clipped_mc
            acc[day_d][eq_key].append((slot_start, slot_end))

    out: dict[date, dict[str, list[tuple[datetime, datetime]]]] = {}
    for d, eqmap in acc.items():
        merged_eq = {
            eq: _merge_machine_calendar_intervals(iv)
            for eq, iv in eqmap.items()
            if iv
        }
        phys_accum: dict[str, list] = defaultdict(list)
        for eq, iv in merged_eq.items():
            pk = _equipment_line_key_to_physical_occupancy_key(str(eq).strip())
            if pk:
                phys_accum[pk].extend(iv)
        merged_all = dict(merged_eq)
        for pk, iv in phys_accum.items():
            merged_all[pk] = _merge_machine_calendar_intervals(iv)
        out[d] = merged_all
    return out


def _apply_machine_calendar_floor_for_date(
    current_date: date,
    machine_avail_dt: dict,
    equipment_list: list,
    machine_day_start: datetime,
    *,
    machine_calendar_plan_end: datetime | None = None,
) -> None:
    """当日のタイムラインシード後」機械カレンダー占有で設備空し下限を繰り上きる。"""
    day_blocks = _MACHINE_CALENDAR_BLOCKS_BY_DATE.get(current_date)
    if not day_blocks:
        return
    candidates: set[str] = set()
    for k in machine_avail_dt.keys():
        sk = str(k).strip() if k is not None else ""
        if sk:
            candidates.add(sk)
    for el in equipment_list:
        ek = str(el).strip() if el is not None else ""
        if not ek:
            continue
        pk = _equipment_line_key_to_physical_occupancy_key(ek)
        if pk:
            candidates.add(pk)
    w0 = machine_day_start
    w1 = machine_calendar_plan_end
    if w1 is None:
        w1 = datetime.combine(current_date, DEFAULT_END_TIME)
    for eq_s in candidates:
        blocks = day_blocks.get(eq_s) or _machine_calendar_blocks_for_occ_key(
            day_blocks, eq_s
        )
        if not blocks:
            continue
        blocks_c = _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
        if not blocks_c:
            continue
        t0 = machine_avail_dt.get(eq_s, machine_day_start)
        t1 = _bump_dt_past_machine_calendar_blocks(t0, blocks_c)
        if t1 > t0:
            machine_avail_dt[eq_s] = t1


def _machine_calendar_blocks_for_occ_key(
    day_blocks: dict[str, list[tuple[datetime, datetime]]],
    occ: str,
) -> list[tuple[datetime, datetime]] | None:
    """day_blocks から占有キー（表記ゆらね許容）に一致する区間リストを得る。"""
    o = str(occ or "").strip()
    if not o or not day_blocks:
        return None
    if o in day_blocks:
        return day_blocks[o]
    nk = _normalize_equipment_match_key(o)
    for k, iv in day_blocks.items():
        if _normalize_equipment_match_key(str(k)) == nk:
            return iv
    return None


def _machine_calendar_occ_blocks_full_plan_window(
    occ_key: str,
    current_date: date,
    daily_status: dict,
    members: list,
) -> bool:
    """
    当日の機械カレンダー占有は計画窓 [始業, min(終業,稼働メンバー終了) ) 全体を塞ね」
    しの設備では当日 1 本も加工を入れられないとし True。
    """
    day_blocks = _MACHINE_CALENDAR_BLOCKS_BY_DATE.get(current_date)
    if not day_blocks:
        return False
    blocks = _machine_calendar_blocks_for_occ_key(day_blocks, occ_key)
    if not blocks:
        return False
    w0 = datetime.combine(current_date, DEFAULT_START_TIME)
    w1 = _machine_calendar_planning_window_end_dt(current_date, daily_status, members)
    blocks_c = _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
    if not blocks_c:
        return False
    t1 = _bump_dt_past_machine_calendar_blocks(w0, blocks_c)
    return t1 >= w1


def _task_fully_machine_calendar_blocked_on_date(
    t: dict,
    current_date: date,
    daily_status: dict | None,
    members: list | None,
) -> bool:
    """
    当該タスクの占有設備は」当日の機械カレンダーの値で計画窓を全日塞はれでいる。
    グローバル試行順ブロック用の「最尝試行順」から外れ（他設備の配台デッドロック防止）。
    """
    if daily_status is None or members is None:
        return False
    _tm = t.get("machine")
    _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
    occ = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
    if not occ:
        return False
    return _machine_calendar_occ_blocks_full_plan_window(
        occ, current_date, daily_status, members
    )


def _task_no_machining_window_left_from_avail_floor(
    t: dict,
    current_date: date,
    daily_status: dict | None,
    members: list | None,
    machine_avail_dt: dict | None,
    machine_day_start: datetime | None,
    *,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> bool:
    """
    machine_avail_dt（シード・機械カレンダー床・当日確定ロール反映後）で」
    占有設備の空し下限は計画窓終端以上なら当日は当設備にスロットなし。
    `machine_handoff` 等は渡るとしは `_resolve_machine_changeover_floor_segments` により
    `_assign_one_roll_trial_order_flow` とともに **実効加工開始下限** で判定れる
    （生の machine_avail の値ではポャンジオーポー後の下限は欠け」候補や min_dto は狂ごのを防し）。
    また空し下限は終端より版でも」計画窓での **残り連続は 1 ロール分に足りない**
    と判断でしる場合は True（実僝丝足デッドロック防止）。
    カレンダー区間照合のキー坖りこれしを防し。
    """
    if (
        daily_status is None
        or members is None
        or machine_avail_dt is None
        or machine_day_start is None
    ):
        return False
    w1 = _machine_calendar_planning_window_end_dt(current_date, daily_status, members)
    _tm = t.get("machine")
    _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
    occ = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
    if not occ:
        return False
    use_co = machine_handoff is not None and skills_dict is not None
    if use_co:
        machine_name = str(t.get("machine_name", "") or "").strip()
        machine_proc = str(_tm or "").strip()
        eq_line = str(
            t.get("equipment_line_key") or _tm or ""
        ).strip() or str(_tm or "")
        machine_occ_key = _machine_occupancy_key_resolve(t, eq_line)
        t_floor, _segs, abort = _resolve_machine_changeover_floor_segments(
            abolish_all_scheduling_limits=bool(abolish_all_scheduling_limits),
            machine_occ_key=machine_occ_key,
            task_id=str(t.get("task_id") or "").strip(),
            eq_line=eq_line,
            machine_name=machine_name,
            machine_proc=machine_proc,
            machine_avail_dt=machine_avail_dt,
            machine_day_floor=machine_day_start,
            current_date=current_date,
            machine_handoff=machine_handoff,
            daily_status=daily_status,
            skills_dict=skills_dict,
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
        if abort:
            return True
    else:
        t_floor = machine_avail_dt.get(occ)
        if t_floor is None:
            nk = _normalize_equipment_match_key(occ)
            for k, v in machine_avail_dt.items():
                if _normalize_equipment_match_key(str(k)) == nk:
                    t_floor = v
                    break
        if t_floor is None:
            t_floor = machine_day_start
    if t_floor >= w1:
        return True
    rem = w1 - t_floor
    if rem <= timedelta(0):
        return True
    btp = parse_float_safe(t.get("base_time_per_unit"), 0.0)
    if btp <= 0:
        return False
    t_eff = parse_float_safe(t.get("task_eff_factor"), 1.0)
    if t_eff <= 0:
        t_eff = 1.0
    # eff_time_per_unit ≈ base / avg_eff / t_eff × 余力係数。avg_eff はフォーム次第で下はる。
    _avg_eff_floor = 0.5
    approx_need_mins = max(1.0, float(btp) / t_eff / _avg_eff_floor)
    return rem < timedelta(minutes=approx_need_mins)


def _bump_machine_avail_after_roll_for_calendar(
    current_date: date,
    eq_line: str,
    machine_avail_dt: dict,
    *,
    machine_calendar_plan_end: datetime | None = None,
    machine_day_floor: datetime | None = None,
) -> None:
    """ロール確定直後: 終了時刻はカレンダー占有スロット内なら終端まで繰り上き。"""
    day_blocks = _MACHINE_CALENDAR_BLOCKS_BY_DATE.get(current_date)
    if not day_blocks:
        return
    eq_s = str(eq_line).strip() if eq_line is not None else ""
    if not eq_s:
        return
    blocks = day_blocks.get(eq_s)
    if not blocks:
        return
    t0 = machine_avail_dt.get(eq_s)
    if t0 is None:
        return
    w0 = (
        machine_day_floor
        if machine_day_floor is not None
        else datetime.combine(current_date, DEFAULT_START_TIME)
    )
    w1 = (
        machine_calendar_plan_end
        if machine_calendar_plan_end is not None
        else datetime.combine(current_date, DEFAULT_END_TIME)
    )
    blocks_c = _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
    if not blocks_c:
        return
    t1 = _bump_dt_past_machine_calendar_blocks(t0, blocks_c)
    if t1 > t0:
        machine_avail_dt[eq_s] = t1


def _parse_nonneg_minutes_cell(v) -> int:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    try:
        n = int(round(float(v)))
    except (TypeError, ValueError):
        return 0
    return max(0, n)


def _df_pick_column(df, *candidates: str) -> str | None:
    cols = [str(c).strip() for c in df.columns]
    low_map = {str(c).strip().lower(): str(c).strip() for c in df.columns}
    for cand in candidates:
        c0 = str(cand).strip()
        if c0 in df.columns:
            return c0
        cl = c0.lower()
        if cl in low_map:
            return low_map[cl]
    return None


def load_machine_daily_startup_settings(master_path: str) -> dict[str, int]:
    """
    master.xlsm の任意シート「設定_機械_日次始業準備」… 機械名・日次始業準備分（1 行目見出し、2 行目以降）。

    「設定_依頼切替前後時間」による準備・後始末は読み込まない（廃止）。
    """
    startup: dict[str, int] = {}
    if not master_path or not os.path.isfile(master_path):
        return startup
    try:
        xls = pd.ExcelFile(master_path)
    except Exception as e:
        logging.warning("機械日次始業準備設定: ブックを開きません (%s)", e)
        return startup

    if SHEET_MACHINE_DAILY_STARTUP in xls.sheet_names:
        try:
            df2 = pd.read_excel(
                master_path, sheet_name=SHEET_MACHINE_DAILY_STARTUP, header=0
            )
            df2.columns = [str(c).strip() for c in df2.columns]
            c_mn = _df_pick_column(df2, "機械名", "機械")
            c_su = _df_pick_column(
                df2, "日次始業準備_分", "始業準備_分", "日始業準備_分"
            )
            if c_mn and c_su:
                for _, row in df2.iterrows():
                    mn = row.get(c_mn)
                    if mn is None or (isinstance(mn, float) and pd.isna(mn)):
                        continue
                    mn_s = str(mn).strip()
                    if not mn_s or mn_s.lower() == "nan":
                        continue
                    su = _parse_nonneg_minutes_cell(row.get(c_su))
                    if su <= 0:
                        continue
                    startup[mn_s] = su
                    nk = _normalize_equipment_match_key(mn_s)
                    if nk:
                        startup[nk] = su
                if startup:
                    logging.info(
                        "マスタ「%s」: 機械 %s 件の日次始業準備（分）を読み込みました。",
                        SHEET_MACHINE_DAILY_STARTUP,
                        len({k for k in startup if "+" not in str(k)}),
                    )
        except Exception as e:
            logging.warning(
                "マスタ「%s」読込失敗（無視）: %s", SHEET_MACHINE_DAILY_STARTUP, e
            )

    return startup


def _lookup_daily_startup_minutes(
    machine_name: str,
    by_m: dict[str, int] | None,
) -> int:
    st = by_m if by_m is not None else _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE
    mn = str(machine_name or "").strip()
    if not mn:
        return 0
    if mn in st:
        return st[mn]
    nk = _normalize_equipment_match_key(mn)
    if nk in st:
        return st[nk]
    for k, v in st.items():
        if _normalize_equipment_match_key(str(k)) == nk:
            return v
    return 0


def _timeline_event_kind(ev: dict) -> str:
    k = str(ev.get("event_kind") or "").strip()
    return k if k else TIMELINE_EVENT_MACHINING


def _is_machining_timeline_event(ev: dict) -> bool:
    return _timeline_event_kind(ev) == TIMELINE_EVENT_MACHINING


def _gap_minutes_until_next_break_start(dt, breaks_merged) -> float | None:
    """dt 以降に始まる最初の休憩開始までの分。無ければ None。"""
    if not isinstance(dt, datetime) or not breaks_merged:
        return None
    best: float | None = None
    for item in breaks_merged:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            continue
        bs = item[0]
        if not isinstance(bs, datetime):
            continue
        if bs < dt:
            continue
        m = (bs - dt).total_seconds() / 60.0
        if best is None or m < best:
            best = float(m)
    return best


def _pick_skilled_op_for_changeover_interval(
    machine_proc: str,
    machine_name: str,
    skills_dict: dict,
    daily_status: dict,
) -> str | None:
    """
    当日 eligible のごう」当該工程+機械で OP スキルを挝つ者のごう優先度は最尝の1坝。
    日次始業の休憩スキップに用いる（avail_dt は見ない）。
    """
    cands: list[tuple[int, str]] = []
    proc = (machine_proc or "").strip()
    mnm = (machine_name or "").strip()
    for mem, st in daily_status.items():
        if not st.get("eligible_for_assignment", st.get("is_working", False)):
            continue
        srow = skills_dict.get(mem, {})
        if proc and mnm:
            v = srow.get(f"{proc}+{mnm}", "")
        elif mnm:
            v = srow.get(mnm, "")
        elif proc:
            v = srow.get(proc, "")
        else:
            v = ""
        role, prio = parse_op_as_skill_cell(v)
        if role == "OP":
            cands.append((prio, mem))
    if not cands:
        return None
    return min(cands)[1]


def _machine_effective_floor_timedelta_only(
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_avail_dt: dict,
    machine_handoff: dict,
    machine_day_floor: datetime,
    abolish_limits: bool,
    *,
    daily_startup_by_machine: dict[str, int] | None = None,
    current_date: date | None = None,
    daily_status: dict | None = None,
    skills_dict: dict | None = None,
    machine_proc: str | None = None,
) -> datetime:
    """スキル OP を拾わないときのフォールバック（壁時計に分を足す＝定常開始基準の日次始業は終了時刻で max）。"""
    if abolish_limits:
        return machine_day_floor
    mf = machine_avail_dt.get(machine_occ_key, machine_day_floor)
    mto = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    if machine_occ_key not in mto:
        su = _lookup_daily_startup_minutes(machine_name, daily_startup_by_machine)
        if su:
            reg_ts = _STAGE2_REGULAR_SHIFT_START
            if reg_ts is not None:
                cd = current_date if current_date is not None else machine_day_floor.date()
                reg_end = datetime.combine(cd, reg_ts) + timedelta(minutes=su)
                mf = max(mf, reg_end)
            else:
                mf = mf + timedelta(minutes=su)
    return mf


def _machining_events_same_occ_day_sorted(
    timeline_events: list,
    current_date: date,
    machine_occ_key: str,
) -> list[dict]:
    occ = str(machine_occ_key or "").strip()
    if not occ:
        return []
    out: list[dict] = []
    for ev in timeline_events or []:
        if ev.get("date") != current_date:
            continue
        if str(ev.get("machine_occupancy_key") or "").strip() != occ:
            continue
        if not _is_machining_timeline_event(ev):
            continue
        st = ev.get("start_dt")
        ed = ev.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime) or ed <= st:
            continue
        out.append(ev)
    out.sort(key=lambda e: (e["start_dt"], e["end_dt"]))
    return out


def _machining_timeline_event_min_end_dt(ev: dict) -> datetime | None:
    """当該加工イベントを業務上の最小量まで短くしたときの終了時刻（これ以上は短縮しない）。"""
    s0 = ev.get("start_dt")
    e0 = ev.get("end_dt")
    if not isinstance(s0, datetime) or not isinstance(e0, datetime) or e0 <= s0:
        return None
    br = merge_time_intervals(list(ev.get("breaks") or []))
    eff = float(ev.get("eff_time_per_unit") or 0.0)
    units = float(ev.get("units_done") or 0.0)
    if eff <= 0 or units <= 0:
        return e0
    min_u = 1.0 if units >= 1.0 else units
    min_wm = max(1, int(math.ceil(min_u * eff)))
    end_limit = e0 + timedelta(days=1)
    e_min, act, rem = calculate_end_time(s0, min_wm, br, end_limit)
    if rem > 0 or act < min_wm:
        return e0
    return e_min


def _cleanup_full_duration_fits_from_start(
    cleanup_start: datetime,
    cleanup_minutes: int,
    breaks_merged: list,
    shift_end: datetime,
) -> bool:
    if cleanup_minutes <= 0:
        return True
    if not isinstance(cleanup_start, datetime) or not isinstance(shift_end, datetime):
        return False
    _ce, act, rem = calculate_end_time(
        cleanup_start, cleanup_minutes, breaks_merged, shift_end
    )
    return rem <= 0 and act >= cleanup_minutes


def _changeover_need_cleanup_for_next_assign(
    *,
    machine_handoff: dict,
    machine_occ_key: str,
    current_date: date,
    cur_task_id: str,
    last_eq: str | None,
) -> tuple[bool, int, str, str]:
    """
    依頼切替後始末が次ロール前に必要か、後始末分、直前主、直前設備行を返す。
    _changeover_plan_segments_and_machining_lower_bound の need_cleanup と整合。
    """
    mach_occ = str(machine_occ_key or "").strip()
    machining_today_occ = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    last_tid = (machine_handoff.get("last_tid") or {}).get(mach_occ, "")
    last_d = (machine_handoff.get("last_machining_date") or {}).get(mach_occ)
    cur_tid = str(cur_task_id or "").strip()
    last_eq_s = str(last_eq or "").strip() or str(
        (machine_handoff.get("last_eq") or {}).get(mach_occ, "") or ""
    ).strip()
    _prep_unused, cu_prev = _lookup_changeover_minutes_for_eq(last_eq_s, None)
    need = (
        bool(str(last_tid or "").strip())
        and bool(cur_tid)
        and str(last_tid).strip() != cur_tid
        and last_d == current_date
        and cu_prev > 0
        and mach_occ in machining_today_occ
    )
    last_lead = str((machine_handoff.get("last_lead_op") or {}).get(mach_occ, "") or "").strip()
    return need, cu_prev, last_lead, last_eq_s


def _avail_dt_reapply_member_max_end_from_timeline(
    timeline_events: list,
    avail_dt: dict,
    members: set[str],
) -> None:
    """指定メンバーについて、タイムライン全体の終了の最大で avail_dt を上書き（短縮後の整合）。"""
    for m in members:
        mm = str(m or "").strip()
        if not mm or mm not in avail_dt:
            continue
        best: datetime | None = None
        for ev in timeline_events or []:
            names: list[str] = []
            op = str(ev.get("op") or "").strip()
            if op:
                names.append(op)
            for s in str(ev.get("sub") or "").split(","):
                s = s.strip()
                if s:
                    names.append(s)
            if mm not in names:
                continue
            ed = ev.get("end_dt")
            if isinstance(ed, datetime):
                best = ed if best is None else max(best, ed)
        if best is not None:
            avail_dt[mm] = best


def _repair_timeline_for_same_tid_prebreak_cleanup(
    *,
    timeline_events: list,
    machine_avail_dt: dict,
    machine_handoff: dict,
    current_date: date,
    machine_occ_key: str,
    next_task_id: str,
    machine_proc: str,
    machine_name: str,
    daily_status: dict,
    skills_dict: dict,
    avail_dt: dict | None,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    task_queue: list,
    machine_day_floor: datetime,
) -> bool:
    """
    同一依頼のまま長い勤務休憩の直前に後始末を載せるため、直前加工終了を逆算開始に合わせて短縮する。
    """
    mach_occ = str(machine_occ_key or "").strip()
    if not mach_occ:
        return False
    machining_today_occ = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    last_tid = str((machine_handoff.get("last_tid") or {}).get(mach_occ, "") or "").strip()
    cur_tid = str(next_task_id or "").strip()
    if not last_tid or not cur_tid or last_tid != cur_tid:
        return False
    last_d = (machine_handoff.get("last_machining_date") or {}).get(mach_occ)
    if last_d != current_date or mach_occ not in machining_today_occ:
        return False
    last_eq_s = str(
        (machine_handoff.get("last_eq") or {}).get(mach_occ, "") or ""
    ).strip()
    _pu, cu_prev = _lookup_changeover_minutes_for_eq(last_eq_s, None)
    if cu_prev <= 0:
        return False
    lm_end = (machine_handoff.get("last_machining_dt") or {}).get(mach_occ)
    if not isinstance(lm_end, datetime):
        return False
    last_lead = str((machine_handoff.get("last_lead_op") or {}).get(mach_occ, "") or "").strip()
    rep = _pick_skilled_op_for_changeover_interval(
        str(machine_proc or "").strip(),
        str(machine_name or "").strip(),
        skills_dict,
        daily_status,
    )
    if not last_lead:
        last_lead = str(rep or "").strip()
    if not last_lead:
        return False
    st_ld = daily_status.get(last_lead) or (daily_status.get(rep) if rep else None)
    if not st_ld:
        return False
    _brk_parts = list(st_ld.get("breaks_dt") or [])
    if rep:
        st_rep_m = daily_status.get(rep)
        if st_rep_m:
            _brk_parts.extend(list(st_rep_m.get("breaks_dt") or []))
    br_resume = merge_time_intervals(_brk_parts)
    _hit, _tf, bs_a, be_a, pre_gap = _resume_after_work_break_extended(
        lm_end, lm_end, br_resume
    )
    if not pre_gap or bs_a is None or be_a is None:
        return False
    br_c = merge_time_intervals(list(st_ld.get("breaks_dt") or []))
    end_c = st_ld["end_dt"]
    if not isinstance(end_c, datetime):
        return False
    st_inv = _find_latest_prep_start_matching_end(
        bs_a, cu_prev, br_c, machine_day_floor
    )
    if st_inv is None:
        return False
    ce_chk, act_chk, rem_chk = calculate_end_time(st_inv, cu_prev, br_c, end_c)
    if rem_chk > 0 or act_chk < cu_prev or not _dt_close_minutes(ce_chk, bs_a):
        return False
    if st_inv >= lm_end - timedelta(seconds=90):
        return False
    occ = mach_occ
    ml = _machining_events_same_occ_day_sorted(timeline_events, current_date, occ)
    if not ml:
        return False
    last_ev = ml[-1]
    e0 = last_ev["end_dt"]
    if not isinstance(e0, datetime):
        return False
    e_min = _machining_timeline_event_min_end_dt(last_ev)
    if e_min is None or not isinstance(e_min, datetime):
        return False
    best_e = st_inv
    if best_e < e_min or best_e >= e0:
        return False
    delta = e0 - best_e
    old_anchor = e0
    s0 = last_ev.get("start_dt")
    if not isinstance(s0, datetime):
        return False
    touched_members: set[str] = set()
    du = float(last_ev.get("units_done") or 0.0)
    br_ev = merge_time_intervals(list(last_ev.get("breaks") or []))
    _cap_m = max(1, int((e0 - s0).total_seconds() // 60) + 1440)
    _, wm_old, _ = calculate_end_time(s0, _cap_m, br_ev, e0)
    _, wm_new, _ = calculate_end_time(s0, _cap_m, br_ev, best_e)
    if wm_old <= 0:
        return False
    new_u = max(1e-12, du * (wm_new / wm_old))
    min_u = 1.0 if du >= 1.0 else du
    new_u = max(min_u, new_u)
    tid = str(last_ev.get("task_id") or "").strip()
    for tq in task_queue or []:
        if str(tq.get("task_id") or "").strip() != tid:
            continue
        tq["remaining_units"] = float(tq.get("remaining_units") or 0) + (du - new_u)
        for row in reversed(tq.get("assigned_history") or []):
            edh = row.get("end_dt")
            if isinstance(edh, datetime) and edh == e0:
                row["end_dt"] = best_e
                try:
                    row["done_m"] = int(float(new_u) * float(tq.get("unit_m") or 0))
                except Exception:
                    pass
                break
        break
    last_ev["end_dt"] = best_e
    last_ev["units_done"] = new_u
    for opn, sb in (
        (str(last_ev.get("op") or "").strip(), str(last_ev.get("sub") or "")),
    ):
        if opn:
            touched_members.add(opn)
        for s in sb.split(","):
            s = s.strip()
            if s:
                touched_members.add(s)
    for ev2 in timeline_events:
        if ev2.get("date") != current_date:
            continue
        if str(ev2.get("machine_occupancy_key") or "").strip() != occ:
            continue
        st2 = ev2.get("start_dt")
        ed2 = ev2.get("end_dt")
        if not isinstance(st2, datetime) or not isinstance(ed2, datetime):
            continue
        if st2 >= old_anchor:
            ev2["start_dt"] = st2 - delta
            ev2["end_dt"] = ed2 - delta
            op2 = str(ev2.get("op") or "").strip()
            if op2:
                touched_members.add(op2)
            for s in str(ev2.get("sub") or "").split(","):
                s = s.strip()
                if s:
                    touched_members.add(s)
    machine_avail_dt[occ] = best_e
    machine_handoff.setdefault("last_machining_dt", {})
    machine_handoff["last_machining_dt"][occ] = best_e
    if avail_dt is not None and touched_members:
        _avail_dt_reapply_member_max_end_from_timeline(
            timeline_events, avail_dt, touched_members
        )
    if dispatch_interval_mirror is not None:
        dispatch_interval_mirror.rebuild_from_timeline(timeline_events)
    _bump_machine_avail_after_roll_for_calendar(
        current_date,
        occ,
        machine_avail_dt,
        machine_calendar_plan_end=None,
        machine_day_floor=machine_day_floor,
    )
    return True


def _repair_timeline_shorten_machining_for_changeover_cleanup(
    *,
    timeline_events: list,
    machine_avail_dt: dict,
    machine_handoff: dict,
    current_date: date,
    machine_occ_key: str,
    next_task_id: str,
    machine_proc: str,
    machine_name: str,
    daily_status: dict,
    skills_dict: dict,
    avail_dt: dict | None,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    task_queue: list,
    machine_day_floor: datetime,
) -> bool:
    """
    依頼切替の後始末が担当者勤務線で実寸積めないとき、当日・同一占有の加工タイムラインを
    終了が新しい順に短縮し、必要ならその直後以降の区間を壁時計で繰り上げる。
    後始末実寸優先（加工量は短縮で犠牲）。
    """
    need, cu_prev, last_lead, last_eq_s = _changeover_need_cleanup_for_next_assign(
        machine_handoff=machine_handoff,
        machine_occ_key=machine_occ_key,
        current_date=current_date,
        cur_task_id=next_task_id,
        last_eq=None,
    )
    if not need:
        if _repair_timeline_for_same_tid_prebreak_cleanup(
            timeline_events=timeline_events,
            machine_avail_dt=machine_avail_dt,
            machine_handoff=machine_handoff,
            current_date=current_date,
            machine_occ_key=machine_occ_key,
            next_task_id=next_task_id,
            machine_proc=str(machine_proc or "").strip(),
            machine_name=str(machine_name or "").strip(),
            daily_status=daily_status,
            skills_dict=skills_dict,
            avail_dt=avail_dt,
            dispatch_interval_mirror=dispatch_interval_mirror,
            task_queue=task_queue,
            machine_day_floor=machine_day_floor,
        ):
            return True
        return False
    if cu_prev <= 0:
        return False
    rep = _pick_skilled_op_for_changeover_interval(
        str(machine_proc or "").strip(),
        str(machine_name or "").strip(),
        skills_dict,
        daily_status,
    )
    if not last_lead:
        last_lead = str(rep or "").strip()
    if not last_lead:
        return False
    st_c = daily_status.get(last_lead) or (
        daily_status.get(rep) if rep else None
    )
    if not st_c:
        return False
    br_c = merge_time_intervals(list(st_c.get("breaks_dt") or []))
    end_c = st_c["end_dt"]
    if not isinstance(end_c, datetime):
        return False

    def _cleanup_ok_at_machining_end(mach_end: datetime) -> bool:
        return _cleanup_full_duration_fits_from_start(
            mach_end, cu_prev, br_c, end_c
        )

    occ = str(machine_occ_key or "").strip()
    touched_members: set[str] = set()

    for _pass in range(64):
        ml = _machining_events_same_occ_day_sorted(
            timeline_events, current_date, occ
        )
        if not ml:
            return False
        last_ev = ml[-1]
        e0 = last_ev["end_dt"]
        if not isinstance(e0, datetime):
            return False
        if _cleanup_ok_at_machining_end(e0):
            machine_avail_dt[occ] = e0
            machine_handoff.setdefault("last_machining_dt", {})
            machine_handoff["last_machining_dt"][occ] = e0
            return True

        e_min = _machining_timeline_event_min_end_dt(last_ev)
        if e_min is None or not isinstance(e_min, datetime):
            return False
        s0 = last_ev.get("start_dt")
        if not isinstance(s0, datetime):
            return False

        lo = e_min
        hi = e0
        best_e: datetime | None = None
        while lo <= hi:
            mid = lo + (hi - lo) // 2
            if _cleanup_ok_at_machining_end(mid):
                best_e = mid
                lo = mid + timedelta(minutes=1)
            else:
                hi = mid - timedelta(minutes=1)

        if best_e is not None and isinstance(best_e, datetime) and best_e < e0:
            delta = e0 - best_e
            old_anchor = e0
            du = float(last_ev.get("units_done") or 0.0)
            br_ev = merge_time_intervals(list(last_ev.get("breaks") or []))
            _cap_m = max(1, int((e0 - s0).total_seconds() // 60) + 1440)
            _, wm_old, _ = calculate_end_time(s0, _cap_m, br_ev, e0)
            _, wm_new, _ = calculate_end_time(s0, _cap_m, br_ev, best_e)
            if wm_old <= 0:
                return False
            new_u = max(1e-12, du * (wm_new / wm_old))
            min_u = 1.0 if du >= 1.0 else du
            new_u = max(min_u, new_u)
            tid = str(last_ev.get("task_id") or "").strip()
            for t in task_queue or []:
                if str(t.get("task_id") or "").strip() != tid:
                    continue
                t["remaining_units"] = float(t.get("remaining_units") or 0) + (du - new_u)
                for row in reversed(t.get("assigned_history") or []):
                    edh = row.get("end_dt")
                    if isinstance(edh, datetime) and edh == e0:
                        row["end_dt"] = best_e
                        try:
                            row["done_m"] = int(
                                float(new_u) * float(t.get("unit_m") or 0)
                            )
                        except Exception:
                            pass
                        break
                break
            last_ev["end_dt"] = best_e
            last_ev["units_done"] = new_u
            for opn, sb in (
                (str(last_ev.get("op") or "").strip(), str(last_ev.get("sub") or ""))
            ):
                if opn:
                    touched_members.add(opn)
                for s in sb.split(","):
                    s = s.strip()
                    if s:
                        touched_members.add(s)
            for ev2 in timeline_events:
                if ev2.get("date") != current_date:
                    continue
                if str(ev2.get("machine_occupancy_key") or "").strip() != occ:
                    continue
                st2 = ev2.get("start_dt")
                ed2 = ev2.get("end_dt")
                if not isinstance(st2, datetime) or not isinstance(ed2, datetime):
                    continue
                if st2 >= old_anchor:
                    ev2["start_dt"] = st2 - delta
                    ev2["end_dt"] = ed2 - delta
                    op2 = str(ev2.get("op") or "").strip()
                    if op2:
                        touched_members.add(op2)
                    for s in str(ev2.get("sub") or "").split(","):
                        s = s.strip()
                        if s:
                            touched_members.add(s)
            machine_avail_dt[occ] = best_e
            machine_handoff.setdefault("last_machining_dt", {})
            machine_handoff["last_machining_dt"][occ] = best_e
            if avail_dt is not None and touched_members:
                _avail_dt_reapply_member_max_end_from_timeline(
                    timeline_events, avail_dt, touched_members
                )
            if dispatch_interval_mirror is not None:
                dispatch_interval_mirror.rebuild_from_timeline(timeline_events)
            _bump_machine_avail_after_roll_for_calendar(
                current_date,
                occ,
                machine_avail_dt,
                machine_calendar_plan_end=None,
                machine_day_floor=machine_day_floor,
            )
            continue

        if len(ml) < 2:
            return False
        applied_prev = False
        for shorten_idx in range(len(ml) - 2, -1, -1):
            prev_ev = ml[shorten_idx]
            eP0 = prev_ev.get("end_dt")
            sP0 = prev_ev.get("start_dt")
            if not isinstance(eP0, datetime) or not isinstance(sP0, datetime):
                continue
            last_ev2 = ml[-1]
            eL_end = last_ev2.get("end_dt")
            if not isinstance(eL_end, datetime):
                return False
            eP_min = _machining_timeline_event_min_end_dt(prev_ev)
            if eP_min is None or not isinstance(eP_min, datetime):
                continue

            def _last_end_after_shrink_prev_end(end_pe: datetime) -> datetime | None:
                if end_pe > eP0 or end_pe < eP_min:
                    return None
                dlt = eP0 - end_pe
                try:
                    return last_ev2["end_dt"] - dlt
                except Exception:
                    return None

            lo2 = eP_min
            hi2 = eP0
            best_pe: datetime | None = None
            while lo2 <= hi2:
                midp = lo2 + (hi2 - lo2) // 2
                le = _last_end_after_shrink_prev_end(midp)
                if le is not None and _cleanup_ok_at_machining_end(le):
                    best_pe = midp
                    lo2 = midp + timedelta(minutes=1)
                else:
                    hi2 = midp - timedelta(minutes=1)

            if best_pe is None or best_pe >= eP0:
                continue

            delta_p = eP0 - best_pe
            old_anchor_p = eP0
            du_p = float(prev_ev.get("units_done") or 0.0)
            br_p = merge_time_intervals(list(prev_ev.get("breaks") or []))
            _cap_p = max(1, int((eP0 - sP0).total_seconds() // 60) + 1440)
            _, wm_old_p, _ = calculate_end_time(sP0, _cap_p, br_p, eP0)
            _, wm_new_p, _ = calculate_end_time(sP0, _cap_p, br_p, best_pe)
            if wm_old_p <= 0:
                return False
            new_u_p = max(1e-12, du_p * (wm_new_p / wm_old_p))
            min_u_p = 1.0 if du_p >= 1.0 else du_p
            new_u_p = max(min_u_p, new_u_p)
            tidp = str(prev_ev.get("task_id") or "").strip()
            for t in task_queue or []:
                if str(t.get("task_id") or "").strip() != tidp:
                    continue
                t["remaining_units"] = float(t.get("remaining_units") or 0) + (
                    du_p - new_u_p
                )
                for row in reversed(t.get("assigned_history") or []):
                    edh = row.get("end_dt")
                    if isinstance(edh, datetime) and edh == eP0:
                        row["end_dt"] = best_pe
                        try:
                            row["done_m"] = int(
                                float(new_u_p) * float(t.get("unit_m") or 0)
                            )
                        except Exception:
                            pass
                        break
                break
            prev_ev["end_dt"] = best_pe
            prev_ev["units_done"] = new_u_p
            for opn, sb in (
                (str(prev_ev.get("op") or "").strip(), str(prev_ev.get("sub") or ""))
            ):
                if opn:
                    touched_members.add(opn)
                for s in sb.split(","):
                    s = s.strip()
                    if s:
                        touched_members.add(s)
            for ev2 in timeline_events:
                if ev2.get("date") != current_date:
                    continue
                if str(ev2.get("machine_occupancy_key") or "").strip() != occ:
                    continue
                st2 = ev2.get("start_dt")
                ed2 = ev2.get("end_dt")
                if not isinstance(st2, datetime) or not isinstance(ed2, datetime):
                    continue
                if st2 >= old_anchor_p:
                    ev2["start_dt"] = st2 - delta_p
                    ev2["end_dt"] = ed2 - delta_p
                    op2 = str(ev2.get("op") or "").strip()
                    if op2:
                        touched_members.add(op2)
                    for s in str(ev2.get("sub") or "").split(","):
                        s = s.strip()
                        if s:
                            touched_members.add(s)
            new_last_end = ml[-1]["end_dt"]
            if isinstance(new_last_end, datetime):
                machine_avail_dt[occ] = new_last_end
                machine_handoff.setdefault("last_machining_dt", {})
                machine_handoff["last_machining_dt"][occ] = new_last_end
            if avail_dt is not None and touched_members:
                _avail_dt_reapply_member_max_end_from_timeline(
                    timeline_events, avail_dt, touched_members
                )
            if dispatch_interval_mirror is not None:
                dispatch_interval_mirror.rebuild_from_timeline(timeline_events)
            _bump_machine_avail_after_roll_for_calendar(
                current_date,
                occ,
                machine_avail_dt,
                machine_calendar_plan_end=None,
                machine_day_floor=machine_day_floor,
            )
            applied_prev = True
            break
        if applied_prev:
            continue
        return False
    return False


def _changeover_plan_segments_and_machining_lower_bound(
    *,
    prev_machining_end_dt: datetime,
    machine_day_floor: datetime,
    current_date: date,
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_proc: str,
    machine_handoff: dict,
    daily_status: dict,
    skills_dict: dict,
    abolish_limits: bool,
) -> tuple[datetime | None, list[dict]]:
    """
    前ロール加工終了 prev_machining_end_dt から、日次始業（当日先頭のみ）のみを
    組み立て、(加工開始最早時刻, タイムライン用セグメント雛形) を返す。
    日次始業は master メイン A15（定常開始）が読めれば [開始, 開始+N分) の壁時計（勤怠 forward しない）。
    A15 が無いときは従来どおり代表スキル OP の勤務・休憩に沿って forward。
    日次始業セグメントの op は空。
    """
    if abolish_limits:
        return prev_machining_end_dt, []
    mach_occ = str(machine_occ_key or "").strip()
    reg_ts = _STAGE2_REGULAR_SHIFT_START
    machining_today_occ = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    su = _lookup_daily_startup_minutes(machine_name, None)

    rep = _pick_skilled_op_for_changeover_interval(
        machine_proc, machine_name, skills_dict, daily_status
    )
    last_lead = str(
        (machine_handoff.get("last_lead_op") or {}).get(mach_occ, "") or ""
    ).strip()
    if not last_lead:
        last_lead = str(rep or "").strip()
    st_r = daily_status.get(rep) if rep else None
    br_r = merge_time_intervals(list(st_r.get("breaks_dt") or [])) if st_r else []
    end_r = st_r["end_dt"] if st_r else None
    start_r = st_r["start_dt"] if st_r else None
    _lt_lead_br = str(last_lead or "").strip()
    _st_bresume = daily_status.get(_lt_lead_br) if _lt_lead_br else None
    _sandwich_parts: list = []
    if _st_bresume:
        _sandwich_parts.extend(list(_st_bresume.get("breaks_dt") or []))
    if st_r:
        _sandwich_parts.extend(list(st_r.get("breaks_dt") or []))
    br_sandwich = (
        merge_time_intervals(_sandwich_parts) if _sandwich_parts else list(br_r or [])
    )

    segments: list[dict] = []
    t = prev_machining_end_dt

    if mach_occ not in machining_today_occ and su > 0:
        if reg_ts is not None:
            reg_start_dt = datetime.combine(current_date, reg_ts)
            reg_end_dt = reg_start_dt + timedelta(minutes=su)
            segments.append(
                {
                    "start_dt": reg_start_dt,
                    "end_dt": reg_end_dt,
                    "op": "",
                    "event_kind": TIMELINE_EVENT_MACHINE_DAILY_STARTUP,
                    "machine": eq_line,
                    "machine_occupancy_key": mach_occ,
                }
            )
            t = max(t, reg_end_dt)
        else:
            if rep is None or not st_r or end_r is None or start_r is None:
                return None, []
            t0 = max(t, machine_day_floor, start_r)
            ce, act, rem = calculate_end_time(t0, su, br_r, end_r)
            if rem > 0 or act < su:
                return None, []
            segments.append(
                {
                    "start_dt": t0,
                    "end_dt": ce,
                    "op": "",
                    "event_kind": TIMELINE_EVENT_MACHINE_DAILY_STARTUP,
                    "machine": eq_line,
                    "machine_occupancy_key": mach_occ,
                }
            )
            t = ce

    return t, segments


def _machine_effective_floor_for_assign(
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_avail_dt: dict,
    machine_handoff: dict,
    machine_day_floor: datetime,
    abolish_limits: bool,
    *,
    daily_startup_by_machine: dict[str, int] | None = None,
    current_date: date | None = None,
    daily_status: dict | None = None,
    skills_dict: dict | None = None,
    machine_proc: str | None = None,
) -> datetime:
    """
    設備のタイムラインによける「当該ロールの加工開始」以降の下限。
    daily_status・skills_dict・current_date は权ごとしは」skills 革坈 OP の勤務・休憩に沿って
    日次始業を forward した最早加工開始。权ゝないとしは分のタイムライン加算にフォールバック。
    """
    if abolish_limits:
        return machine_day_floor
    prev_mach = machine_avail_dt.get(machine_occ_key, machine_day_floor)
    if (
        current_date is not None
        and daily_status is not None
        and skills_dict is not None
        and machine_proc is not None
    ):
        lb, _segs = _changeover_plan_segments_and_machining_lower_bound(
            prev_machining_end_dt=prev_mach,
            machine_day_floor=machine_day_floor,
            current_date=current_date,
            machine_occ_key=machine_occ_key,
            task_id=task_id,
            eq_line=eq_line,
            machine_name=machine_name,
            machine_proc=str(machine_proc or "").strip(),
            machine_handoff=machine_handoff,
            daily_status=daily_status,
            skills_dict=skills_dict,
            abolish_limits=False,
        )
        if lb is not None:
            return lb
    return _machine_effective_floor_timedelta_only(
        machine_occ_key,
        task_id,
        eq_line,
        machine_name,
        machine_avail_dt,
        machine_handoff,
        machine_day_floor,
        False,
        daily_startup_by_machine=daily_startup_by_machine,
        current_date=current_date,
        daily_status=daily_status,
        skills_dict=skills_dict,
        machine_proc=machine_proc,
    )


def _resolve_machine_changeover_floor_segments(
    *,
    abolish_all_scheduling_limits: bool,
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_proc: str,
    machine_avail_dt: dict,
    machine_day_floor: datetime,
    current_date: date,
    machine_handoff: dict,
    daily_status: dict,
    skills_dict: dict,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    timeline_events: list | None = None,
    task_queue: list | None = None,
    avail_dt: dict | None = None,
) -> tuple[datetime, list[dict], bool]:
    """
    設備の加工開始下限と」タイムライン追記用セットアップ区間。
    戻り値 (floor_dt, segments, abort)。abort は True のときは当該ロール割当を全体として棄坴れる。
    """
    if abolish_all_scheduling_limits:
        prev = machine_avail_dt.get(machine_occ_key, machine_day_floor)
        return prev, [], False
    prev_mach = machine_avail_dt.get(machine_occ_key, machine_day_floor)
    co_lb, co_segs = _changeover_plan_segments_and_machining_lower_bound(
        prev_machining_end_dt=prev_mach,
        machine_day_floor=machine_day_floor,
        current_date=current_date,
        machine_occ_key=machine_occ_key,
        task_id=task_id,
        eq_line=eq_line,
        machine_name=machine_name,
        machine_proc=str(machine_proc or "").strip(),
        machine_handoff=machine_handoff,
        daily_status=daily_status,
        skills_dict=skills_dict,
        abolish_limits=False,
    )
    if co_lb is None:
        if (
            timeline_events is not None
            and task_queue is not None
            and _repair_timeline_shorten_machining_for_changeover_cleanup(
                timeline_events=timeline_events,
                machine_avail_dt=machine_avail_dt,
                machine_handoff=machine_handoff,
                current_date=current_date,
                machine_occ_key=machine_occ_key,
                next_task_id=task_id,
                machine_proc=str(machine_proc or "").strip(),
                machine_name=str(machine_name or "").strip(),
                daily_status=daily_status,
                skills_dict=skills_dict,
                avail_dt=avail_dt,
                dispatch_interval_mirror=dispatch_interval_mirror,
                task_queue=task_queue,
                machine_day_floor=machine_day_floor,
            )
        ):
            prev_mach = machine_avail_dt.get(machine_occ_key, machine_day_floor)
            co_lb, co_segs = _changeover_plan_segments_and_machining_lower_bound(
                prev_machining_end_dt=prev_mach,
                machine_day_floor=machine_day_floor,
                current_date=current_date,
                machine_occ_key=machine_occ_key,
                task_id=task_id,
                eq_line=eq_line,
                machine_name=machine_name,
                machine_proc=str(machine_proc or "").strip(),
                machine_handoff=machine_handoff,
                daily_status=daily_status,
                skills_dict=skills_dict,
                abolish_limits=False,
            )
            if co_lb is not None:
                logging.info(
                    "依頼切替後始末の勤務線確保のため、当日タイムライン上の加工を短縮しました。"
                    " date=%s occ=%s next_task=%s",
                    current_date,
                    machine_occ_key,
                    task_id,
                )
    if co_lb is None:
        if (
            _pick_skilled_op_for_changeover_interval(
                str(machine_proc or "").strip(),
                str(machine_name or "").strip(),
                skills_dict,
                daily_status,
            )
            is None
        ):
            mf = _machine_effective_floor_timedelta_only(
                machine_occ_key,
                task_id,
                eq_line,
                machine_name,
                machine_avail_dt,
                machine_handoff,
                machine_day_floor,
                False,
                current_date=current_date,
                daily_status=daily_status,
                skills_dict=skills_dict,
                machine_proc=str(machine_proc or "").strip(),
            )
            return mf, [], False
        return machine_day_floor, [], True
    if dispatch_interval_mirror is not None and co_segs:
        for seg in co_segs:
            sop = str(seg.get("op") or "").strip()
            sok = str(seg.get("machine_occupancy_key") or machine_occ_key).strip()
            st_seg = seg.get("start_dt")
            ed_seg = seg.get("end_dt")
            if not isinstance(st_seg, datetime) or not isinstance(ed_seg, datetime):
                continue
            if (
                sop
                and dispatch_interval_mirror.would_block_member(sop, st_seg, ed_seg)
            ):
                return machine_day_floor, [], True
            if (
                sok
                and dispatch_interval_mirror.would_block_equipment(
                    sok, st_seg, ed_seg
                )
            ):
                return machine_day_floor, [], True
    return co_lb, co_segs, False


def _changeover_timeline_op_sub_for_event(
    *,
    event_kind: str,
    op_from_segment: str,
    machine_occ_key: str,
    machining_lead_op: str,
    machining_sub_str: str,
    machine_handoff: dict,
    daily_status: dict,
) -> tuple[str, str]:
    """タイムライン用の主＝補。日次始業は人なし。"""
    ek = str(event_kind or "").strip()
    op_s = str(op_from_segment or "").strip()
    if ek == TIMELINE_EVENT_MACHINE_DAILY_STARTUP:
        return "", ""
    return op_s, ""


def _append_changeover_segments_to_timeline(
    timeline_events: list,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    avail_dt: dict,
    daily_status: dict,
    *,
    current_date: date,
    task_id: str,
    machine_occ_key: str,
    segments: list[dict],
    machining_lead_op: str | None = None,
    machining_sub_str: str | None = None,
    machine_handoff: dict | None = None,
) -> None:
    """セットアップ系セグメントをタイムライン・ミラー・担当者 avail に反映。"""
    _mh = machine_handoff or {}
    _lead_m = str(machining_lead_op or "").strip()
    _sub_roll = str(machining_sub_str or "").strip()
    for seg in segments or []:
        op_seg = str(seg.get("op") or "").strip()
        st = seg.get("start_dt")
        ed = seg.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime):
            continue
        m_line = str(seg.get("machine") or "").strip()
        m_occ = str(seg.get("machine_occupancy_key") or machine_occ_key).strip()
        ek = str(seg.get("event_kind") or "").strip() or TIMELINE_EVENT_MACHINING
        op, sub = _changeover_timeline_op_sub_for_event(
            event_kind=ek,
            op_from_segment=op_seg,
            machine_occ_key=m_occ,
            machining_lead_op=_lead_m,
            machining_sub_str=_sub_roll,
            machine_handoff=_mh,
            daily_status=daily_status,
        )
        br_acc: list = []
        for nm in (op, *[_p.strip() for _p in sub.split(",") if _p.strip()]):
            if nm and nm in daily_status:
                br_acc.extend(daily_status[nm].get("breaks_dt") or [])
        br_seg = merge_time_intervals(br_acc)
        tid_ev = (
            ""
            if ek == TIMELINE_EVENT_MACHINE_DAILY_STARTUP
            else str(task_id or "").strip()
        )
        ev = {
            "date": current_date,
            "task_id": tid_ev,
            "machine": m_line,
            "machine_occupancy_key": m_occ,
            "op": op,
            "sub": sub,
            "start_dt": st,
            "end_dt": ed,
            "breaks": br_seg,
            "units_done": 0,
            "event_kind": ek,
        }
        timeline_events.append(ev)
        if dispatch_interval_mirror is not None:
            dispatch_interval_mirror.register_from_event(ev)
        for nm in (op, *[_p.strip() for _p in sub.split(",") if _p.strip()]):
            if not nm:
                continue
            prev_a = avail_dt.get(nm, st)
            if isinstance(prev_a, datetime):
                avail_dt[nm] = max(prev_a, ed)
            else:
                avail_dt[nm] = ed


def _collect_task_ids_missed_deadline_after_day(task_queue: list, current_date: date) -> set:
    """
    当該日の終了時点で」紝期基準日（当日含む）以降なのに残量は残る依頼NO。
    「紝期日内に完靂でしなかった」= 後ゝ倒し再試行の候補。
    """
    out = set()
    eps = 1e-9
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= eps:
            continue
        db = t.get("due_basis_date")
        if db is None:
            continue
        sdr = t.get("start_date_req")
        if isinstance(sdr, date) and sdr > current_date:
            continue
        if current_date >= db:
            tid = str(t.get("task_id", "") or "").strip()
            if tid:
                out.add(tid)
    return out


def _normalize_timeline_task_id(ev: dict) -> str:
    return str(ev.get("task_id", "") or "").strip()


def _machine_handoff_state_from_timeline(
    timeline_events: list,
    current_date: date,
) -> dict:
    """
    タイムラインから」坄 machine_occupancy_key についで
    計画日 current_date 以降の **加工 (machining)** イベントの最終終了を復元れる。
    セットアップ系 event_kind は last_tid 等の復元に含まない。
    """
    best: dict[str, tuple[datetime, str, str, date, str, str]] = {}
    for e in timeline_events:
        if not _is_machining_timeline_event(e):
            continue
        ed = e.get("date")
        if not isinstance(ed, date):
            continue
        if ed > current_date:
            continue
        occ = str(e.get("machine_occupancy_key") or "").strip()
        if not occ:
            mraw = str(e.get("machine") or "").strip()
            occ = (
                _normalize_equipment_match_key(mraw.split("+", 1)[1])
                if "+" in mraw
                else _normalize_equipment_match_key(mraw)
            )
        if not occ:
            continue
        end_dt = e.get("end_dt")
        if end_dt is None or not hasattr(end_dt, "replace"):
            continue
        eq_line = str(e.get("machine") or "").strip()
        tid = _normalize_timeline_task_id(e)
        lead_op = str(e.get("op") or "").strip()
        sub_csv = str(e.get("sub") or "").strip()
        prev = best.get(occ)
        if prev is None or end_dt > prev[0]:
            best[occ] = (end_dt, tid, eq_line, ed, lead_op, sub_csv)
    last_tid = {k: v[1] for k, v in best.items()}
    last_eq = {k: v[2] for k, v in best.items()}
    last_machining_dt = {k: v[0] for k, v in best.items()}
    last_machining_date = {k: v[3] for k, v in best.items()}
    last_lead_op = {k: v[4] for k, v in best.items()}
    last_machining_sub = {k: v[5] for k, v in best.items()}
    machining_today_occ: set[str] = set()
    for e in timeline_events:
        if not _is_machining_timeline_event(e):
            continue
        if e.get("date") != current_date:
            continue
        occ = str(e.get("machine_occupancy_key") or "").strip()
        if not occ:
            mraw = str(e.get("machine") or "").strip()
            occ = (
                _normalize_equipment_match_key(mraw.split("+", 1)[1])
                if "+" in mraw
                else _normalize_equipment_match_key(mraw)
            )
        if occ:
            machining_today_occ.add(occ)
    started_today = set(machining_today_occ)
    return {
        "last_tid": last_tid,
        "last_eq": last_eq,
        "last_machining_dt": last_machining_dt,
        "last_machining_date": last_machining_date,
        "last_lead_op": last_lead_op,
        "last_machining_sub": last_machining_sub,
        "machining_today_occ": machining_today_occ,
        "started_today": started_today,
    }


def _trial_order_flow_day_start_floor(
    task: dict,
    current_date: date,
    macro_run_date: date,
    macro_now_dt: datetime,
    task_queue: list | None = None,
) -> datetime:
    """原板投入日を起点に」しの日の加工開始の下限時刻（同日は 13:00 以降を含む）。"""
    floor = datetime.combine(current_date, DEFAULT_START_TIME)
    # §B-2 検査 / §B-3 巻返しは EC 完了を待って開始でしるため、
    # 原板投入日（=同日13:00以降）の制約をしのまま適用すると後続は丝必須に後ゝへ倒れる。
    # EC完了時刻下限（_roll_pipeline_b2_inspection_ec_completion_floor_dt）で整合を得る。
    _tid_floor = str(task.get("task_id", "") or "").strip()
    is_b2_follower_delayed = bool(
        (task.get("roll_pipeline_inspection") or task.get("roll_pipeline_rewind"))
        and _tid_floor
        and task_queue is not None
        and _task_queue_has_roll_pipeline_ec_for_tid(task_queue, _tid_floor)
    )
    rid = task.get("raw_input_date")
    if not is_b2_follower_delayed and isinstance(rid, date) and rid == current_date:
        floor = max(floor, datetime.combine(current_date, time(13, 0)))
    sdl = task.get("same_day_raw_start_limit")
    s_req = task.get("start_date_req")
    if (
        (not is_b2_follower_delayed)
        and sdl
        and isinstance(s_req, date)
        and current_date == s_req
        and isinstance(sdl, time)
    ):
        floor = max(floor, datetime.combine(current_date, sdl))
    est = task.get("earliest_start_time")
    if (
        (not is_b2_follower_delayed)
        and isinstance(s_req, date)
        and current_date == s_req
        and est
    ):
        if isinstance(est, time):
            floor = max(floor, datetime.combine(current_date, est))
    if current_date == macro_run_date and floor < macro_now_dt:
        floor = macro_now_dt
    return floor


def _trial_order_flow_eligible_tasks(
    tasks_today: list,
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    min_dispatch_effective: int | None = None,
    assign_probe_ctx: dict | None = None,
) -> list:
    out = []
    for task in tasks_today:
        if float(task.get("remaining_units") or 0) <= 1e-12:
            continue
        if _task_blocked_by_same_request_dependency(task, task_queue):
            continue
        if _task_blocked_by_global_dispatch_trial_order(
            task,
            task_queue,
            current_date,
            daily_status=daily_status,
            members=members,
            machine_avail_dt=machine_avail_dt,
            machine_day_start=machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
            min_dispatch_effective=min_dispatch_effective,
        ):
            continue
        # min_dto から全日カレンダー占有は除外済みでも」同日試行順の「ブロック」は my_o>m のみのため、
        # 試行順=min の占有行は残り」他試行順は永久坜止し得る。当日スロットゼロの行は候補外にれる。
        if daily_status is not None and members is not None:
            if _task_fully_machine_calendar_blocked_on_date(
                task, current_date, daily_status, members
            ):
                continue
            if _task_no_machining_window_left_from_avail_floor(
                task,
                current_date,
                daily_status,
                members,
                machine_avail_dt,
                machine_day_start,
                machine_handoff=machine_handoff,
                skills_dict=skills_dict,
                abolish_all_scheduling_limits=abolish_all_scheduling_limits,
                dispatch_interval_mirror=dispatch_interval_mirror,
            ):
                continue
        if (
            task.get("roll_pipeline_inspection") or task.get("roll_pipeline_rewind")
        ) and (
            _roll_pipeline_inspection_assign_room(
                task_queue, str(task.get("task_id", "") or "").strip()
            )
            <= 1e-12
        ):
            continue
        machine = task["machine"]
        eq_line = str(
            task.get("equipment_line_key") or machine or ""
        ).strip() or machine
        _mocc_trial = _machine_occupancy_key_resolve(task, eq_line)
        if PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE:
            _b1_holder = _exclusive_b1_inspection_holder_for_machine(
                task_queue,
                _mocc_trial,
            )
            if _b1_holder is not None and _b1_holder is not task:
                continue
        try:
            _my_dispatch_ord = int(task.get("dispatch_trial_order") or 10**9)
        except (TypeError, ValueError):
            _my_dispatch_ord = 10**9
        if _equipment_line_lower_dispatch_trial_still_pending(
            task_queue,
            _mocc_trial,
            _my_dispatch_ord,
            current_date,
            daily_status=daily_status,
            members=members,
            machine_avail_dt=machine_avail_dt,
            machine_day_start=machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
            assign_probe_ctx=assign_probe_ctx,
        ):
            continue
        out.append(task)
    return out


def _combo_preset_team_size_bounds(
    preset_team: tuple,
    sheet_req_n: int | None,
    max_team_size_need: int,
) -> tuple[int, int] | None:
    """
    組み合わせ表プリセット1行の人数範囲 (lo, hi)。need の基本人数よりシート坴を優先れる。
    - 必須人数列は正のときはメンバー列の人数と一致すること。
    - hi は need の上限と実人数の大しい方（プリセットは need より少人数でも採用可能）。
    """
    nmem = len(preset_team)
    if nmem < 1:
        return None
    if sheet_req_n is not None and sheet_req_n >= 1:
        if nmem != sheet_req_n:
            return None
        lo = sheet_req_n
    else:
        lo = nmem
    hi = max(max_team_size_need, nmem)
    if not (lo <= nmem <= hi):
        return None
    return lo, hi


def _plan_sheet_required_op_optional(task: dict) -> int | None:
    """加工計画の必須人数列は正の整数ならしの値。無効なら None。"""
    ro = task.get("required_op")
    if ro is None or (isinstance(ro, float) and pd.isna(ro)):
        return None
    try:
        n = int(ro)
    except (TypeError, ValueError):
        return None
    return n if n >= 1 else None


def _append_legacy_dispatch_candidate_for_team(
    task: dict,
    team: tuple,
    avail_dt: dict,
    machine_avail_dt: dict,
    daily_status: dict,
    current_date: date,
    macro_run_date: date,
    macro_now_dt: datetime,
    skill_role_priority,
    eq_line: str,
    rq_base: int,
    extra_max: int,
    global_priority_override: dict,
    team_candidates: list,
    *,
    combo_sheet_row_id: int | None = None,
    combo_preset_team: tuple[str, ...] | None = None,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    machine_handoff: dict | None = None,
    machine_day_floor: datetime | None = None,
    machine_floor_cached: datetime | None = None,
) -> bool:
    """レガシー日次配台ループ用: 坘一フォームは成立れれみ team_candidates に 1 件追加して True。"""
    _machine_occ_key = _machine_occupancy_key_resolve(task, eq_line)
    _gpo = global_priority_override or {}
    _floor_default = datetime.combine(current_date, DEFAULT_START_TIME)
    _mdf = machine_day_floor if machine_day_floor is not None else _floor_default
    _mh_legacy = machine_handoff or {
        "last_tid": {},
        "last_eq": {},
        "started_today": set(),
        "machining_today_occ": set(),
        "last_machining_dt": {},
        "last_machining_date": {},
        "last_lead_op": {},
        "last_machining_sub": {},
    }
    op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
    if not op_list:
        return False
    team_start = max(avail_dt[m] for m in team)
    if not _gpo.get("abolish_all_scheduling_limits"):
        if machine_floor_cached is not None:
            machine_free_dt = machine_floor_cached
        else:
            machine_free_dt = _machine_effective_floor_for_assign(
                _machine_occ_key,
                str(task.get("task_id") or "").strip(),
                eq_line,
                str(task.get("machine_name") or "").strip(),
                machine_avail_dt,
                _mh_legacy,
                _mdf,
                False,
                current_date=current_date,
            )
        if team_start < machine_free_dt:
            team_start = machine_free_dt
        if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
            min_start_dt = datetime.combine(
                current_date, task["same_day_raw_start_limit"]
            )
            if team_start < min_start_dt:
                team_start = min_start_dt
        if current_date == task["start_date_req"] and task.get("earliest_start_time"):
            min_user_t = datetime.combine(
                current_date, task["earliest_start_time"]
            )
            if team_start < min_user_t:
                team_start = min_user_t
        if current_date == macro_run_date and team_start < macro_now_dt:
            team_start = macro_now_dt
    team_end_limit = min(daily_status[m]["end_dt"] for m in team)
    if team_start >= team_end_limit:
        return False
    team_breaks = []
    for m in team:
        team_breaks.extend(daily_status[m]["breaks_dt"])
    team_breaks = merge_time_intervals(team_breaks)

    avg_eff = sum(daily_status[m]["efficiency"] for m in team) / len(team)
    if avg_eff <= 0:
        avg_eff = 0.01
    t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
    if t_eff <= 0:
        t_eff = 1.0
    eff_time_per_unit = (
        task["base_time_per_unit"]
        / avg_eff
        / t_eff
        * _surplus_team_time_factor(rq_base, len(team), extra_max)
    )
    _defer_min_contig = max(1, int(math.ceil(float(eff_time_per_unit))))
    _eod_cont_exempt = _eod_same_request_continuation_exempt(
        _machine_occ_key, task, _mh_legacy
    )

    def _refloor_legacy_roll(ts: datetime) -> datetime:
        ts = max(ts, max(avail_dt[m] for m in team))
        if not _gpo.get("abolish_all_scheduling_limits"):
            if machine_floor_cached is not None:
                mf = machine_floor_cached
            else:
                mf = _machine_effective_floor_for_assign(
                    _machine_occ_key,
                    str(task.get("task_id") or "").strip(),
                    eq_line,
                    str(task.get("machine_name") or "").strip(),
                    machine_avail_dt,
                    _mh_legacy,
                    _mdf,
                    False,
                    current_date=current_date,
                )
            if ts < mf:
                ts = mf
            if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
                min_start_dt = datetime.combine(
                    current_date, task["same_day_raw_start_limit"]
                )
                if ts < min_start_dt:
                    ts = min_start_dt
            if current_date == task["start_date_req"] and task.get("earliest_start_time"):
                min_user_t = datetime.combine(
                    current_date, task["earliest_start_time"]
                )
                if ts < min_user_t:
                    ts = min_user_t
            if current_date == macro_run_date and ts < macro_now_dt:
                ts = macro_now_dt
        return ts

    team_start_adj = _defer_team_start_past_prebreak_and_end_of_day(
        task,
        team,
        team_start,
        team_end_limit,
        team_breaks,
        _refloor_legacy_roll,
        min_contiguous_work_mins=_defer_min_contig,
        eod_same_request_continuation_exempt=_eod_cont_exempt,
    )
    if team_start_adj is None:
        return False
    team_start = team_start_adj
    if team_start >= team_end_limit:
        return False

    _, avail_mins, _ = calculate_end_time(team_start, 9999, team_breaks, team_end_limit)
    units_can_do = int(avail_mins / eff_time_per_unit)
    if units_can_do == 0:
        return False
    units_today = min(units_can_do, math.ceil(task["remaining_units"]))
    if _eod_reject_capacity_units_below_threshold(
        units_today,
        team_start,
        team_end_limit,
        eod_same_request_continuation_exempt=_eod_cont_exempt,
    ):
        return False
    work_mins_needed = int(units_today * eff_time_per_unit)
    if (
        _contiguous_work_minutes_until_next_break_or_limit(
            team_start, team_breaks, team_end_limit
        )
        < work_mins_needed
    ):
        return False
    actual_end_dt, _, _ = calculate_end_time(
        team_start, work_mins_needed, team_breaks, team_end_limit
    )
    if dispatch_interval_mirror is not None and dispatch_interval_mirror.would_block_roll(
        _machine_occ_key, team, team_start, actual_end_dt
    ):
        return False
    team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
    team_candidates.append(
        {
            "team": team,
            "team_start": team_start,
            "actual_end_dt": actual_end_dt,
            "units_today": units_today,
            "team_breaks": team_breaks,
            "avg_eff": avg_eff,
            "prio_sum": team_prio_sum,
            "op_list": op_list,
            "eff_time_per_unit": eff_time_per_unit,
            "combo_sheet_row_id": combo_sheet_row_id,
            "combo_preset_team": combo_preset_team,
        }
    )
    return True


def _assign_one_roll_trial_order_flow(
    task: dict,
    current_date: date,
    daily_status: dict,
    avail_dt: dict,
    machine_avail_dt: dict,
    task_queue: list,
    skills_dict: dict,
    members: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict,
    macro_run_date: date,
    macro_now_dt: datetime,
    preferred_team: tuple | None,
    _need_headcount_logged_orders: set,
    team_combo_presets: dict | None = None,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    machine_handoff: dict | None = None,
    timeline_events: list | None = None,
) -> dict | None:
    """
    1ロール分の最良フォームを決定れる。設備空し・日開始下限を team_start に織り込む。
    preferred_team は与ごられ」かつ「同一日内の直後ロール」として成立れれみ」
    組み合わせ探索より優先して採用する（翌日には挝う越さない）。
    戻り値: team(tuple), start_dt, end_dt, breaks, eff, op, eff_time_per_unit, extra_max, rq_base, need_src_line, extra_src_line, machine, machine_name, eq_line, req_num, max_team_size
    """
    machine = task["machine"]
    machine_name = str(task.get("machine_name", "") or "").strip()
    machine_proc = str(machine or "").strip()
    eq_line = str(task.get("equipment_line_key") or machine or "").strip() or machine
    machine_occ_key = _machine_occupancy_key_resolve(task, eq_line)
    _gpo = global_priority_override or {}
    _mh = machine_handoff or {
        "last_tid": {},
        "last_eq": {},
        "started_today": set(),
        "machining_today_occ": set(),
        "last_machining_dt": {},
        "last_machining_date": {},
        "last_lead_op": {},
        "last_machining_sub": {},
    }
    _eod_cont_exempt = _eod_same_request_continuation_exempt(
        machine_occ_key, task, _mh
    )

    plan_ro = _plan_sheet_required_op_optional(task)
    need_src_line = ""
    if TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY:
        req_num, need_src_line = resolve_need_required_op_explain(
            machine,
            machine_name,
            task["task_id"],
            req_map,
            need_rules,
        )
        if plan_ro is not None and plan_ro != req_num:
            need_src_line = (need_src_line + "；") if need_src_line else ""
            need_src_line += f"計画シート必須人数{plan_ro}は未使用（need基準={req_num}）"
    else:
        if plan_ro is not None:
            req_num = plan_ro
            need_src_line = f"計画シート「必須OP(上書)」={req_num}"
        else:
            req_num, need_src_line = resolve_need_required_op_explain(
                machine,
                machine_name,
                task["task_id"],
                req_map,
                need_rules,
            )
    if _gpo.get("ignore_need_minimum"):
        req_num = 1
        need_src_line = (
            (need_src_line + " → ") if need_src_line else ""
        ) + "メイン上書ignore_need_minimumでreq=1"

    skill_meta_cache: dict = {}

    def skill_role_priority(mem):
        if _gpo.get("ignore_skill_requirements"):
            return ("OP", 100)
        if mem not in skill_meta_cache:
            srow = skills_dict.get(mem, {})
            v = ""
            if machine_proc and machine_name:
                v = srow.get(f"{machine_proc}+{machine_name}", "")
            elif machine_name:
                v = srow.get(machine_name, "")
            elif machine_proc:
                v = srow.get(machine_proc, "")
            skill_meta_cache[mem] = parse_op_as_skill_cell(v)
        return skill_meta_cache[mem]

    capable_members = [m for m in avail_dt if skill_role_priority(m)[0] in ("OP", "AS")]
    capable_members.sort(key=lambda mm: (skill_role_priority(mm)[1], mm))
    capable_members = _filter_capable_members_b2_disjoint_teams(
        task, task_queue, capable_members
    )

    pref_raw = str(task.get("preferred_operator_raw") or "").strip()
    op_today = [m for m in capable_members if skill_role_priority(m)[0] == "OP"]
    pref_mem = (
        _resolve_preferred_op_to_member(pref_raw, op_today, members)
        if pref_raw
        else None
    )

    _gdp_must, _gdp_warns = _active_global_day_process_must_include(
        _gpo, task, current_date, capable_members, members
    )
    for _gw in _gdp_warns:
        logging.warning(_gw)
    fixed_team_anchor = _merge_global_day_process_and_pref_anchor(
        _gdp_must, pref_mem, capable_members
    )
    if _gdp_must:
        logging.info(
            "メイングローバル(日付×工程): task=%s date=%s 工程=%r フォーム必須=%s",
            task.get("task_id"),
            current_date,
            machine,
            ",".join(_gdp_must),
        )
    if fixed_team_anchor:
        _nfix = len(fixed_team_anchor)
        if _nfix > req_num:
            need_src_line = (need_src_line + " → ") if need_src_line else ""
            need_src_line += f"グローバル(日付×工程)指定で最低{_nfix}人"
        req_num = max(req_num, _nfix)

    extra_max_sheet, extra_src_line = resolve_need_surplus_extra_max_explain(
        machine,
        machine_name,
        task["task_id"],
        surplus_map,
        need_rules,
    )
    if TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
        extra_max_sheet = 0
        extra_src_line = (
            (extra_src_line + " → ") if extra_src_line else ""
        ) + "TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROWで0"
    extra_max = (
        extra_max_sheet if TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS else 0
    )
    if (
        extra_max_sheet > 0
        and not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
        and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
    ):
        extra_src_line = (
            (extra_src_line + " → ") if extra_src_line else ""
        ) + "メインは基本人数のみ（余力枠は全配台後に未割当×スキルで追記）"
    max_team_size = min(req_num + extra_max, len(capable_members))
    if max_team_size < req_num:
        max_team_size = req_num
    rq_base = max(1, int(req_num))
    combo_key_assign = (
        f"{machine_proc}+{machine_name}"
        if machine_proc and machine_name
        else ""
    )
    preset_rows_assign = (
        (team_combo_presets or {}).get(combo_key_assign)
        if (team_combo_presets and combo_key_assign)
        else None
    )

    _dto_head = task.get("dispatch_trial_order")
    if _dto_head is not None and _dto_head not in _need_headcount_logged_orders:
        _need_headcount_logged_orders.add(_dto_head)
        logging.info(
            "need人数(試行順優先フロー) order=%s task=%s 工程/機械=%s/%s "
            "req_num=%s [%s] extra_max=%s [%s] max_team候補=%s capable=%s人",
            _dto_head,
            task["task_id"],
            machine,
            machine_name,
            req_num,
            need_src_line,
            extra_max,
            extra_src_line,
            max_team_size,
            len(capable_members),
        )

    day_floor = _trial_order_flow_day_start_floor(
        task, current_date, macro_run_date, macro_now_dt, task_queue
    )
    machine_day_floor = datetime.combine(current_date, DEFAULT_START_TIME)
    b2_insp_ec_floor: datetime | None = None
    _tid_assign = str(task.get("task_id") or "").strip()
    _trace_assign_enabled = _trace_schedule_task_enabled(_tid_assign)
    def _trace_assign(msg: str, *args) -> None:
        if not _trace_assign_enabled:
            return
        _log_dispatch_trace_schedule(
            _tid_assign,
            "[配台トレース task=%s] " + msg,
            _tid_assign,
            *args,
        )
    if (
        (
            task.get("roll_pipeline_inspection")
            or task.get("roll_pipeline_rewind")
        )
        and _task_queue_has_roll_pipeline_ec_for_tid(task_queue, _tid_assign)
    ):
        b2_insp_ec_floor = _roll_pipeline_b2_inspection_ec_completion_floor_dt(
            task_queue, _tid_assign
        )

    _prev_mach_before_co = machine_avail_dt.get(
        machine_occ_key, machine_day_floor
    )
    _mach_floor_eff, _co_segs, _co_abort = _resolve_machine_changeover_floor_segments(
        abolish_all_scheduling_limits=bool(_gpo.get("abolish_all_scheduling_limits")),
        machine_occ_key=machine_occ_key,
        task_id=str(task.get("task_id") or "").strip(),
        eq_line=eq_line,
        machine_name=machine_name,
        machine_proc=machine_proc,
        machine_avail_dt=machine_avail_dt,
        machine_day_floor=machine_day_floor,
        current_date=current_date,
        machine_handoff=_mh,
        daily_status=daily_status,
        skills_dict=skills_dict,
        dispatch_interval_mirror=dispatch_interval_mirror,
        timeline_events=timeline_events,
        task_queue=task_queue,
        avail_dt=avail_dt,
    )
    if _co_abort:
        return None

    def _one_roll_from_team(
        team: tuple,
        min_n: int | None = None,
        max_n: int | None = None,
    ) -> dict | None:
        lo = req_num if min_n is None else min_n
        hi = max_team_size if max_n is None else max_n
        if len(team) < lo or len(team) > hi:
            _trace_assign(
                "候補坴下: フォーム人数外 team=%s size=%s req=%s max=%s",
                ",".join(str(x) for x in team),
                len(team),
                lo,
                hi,
            )
            return None
        op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
        if not op_list:
            _trace_assign(
                "候補坴下: OP丝在 team=%s",
                ",".join(str(x) for x in team),
            )
            return None
        if not all(m in daily_status for m in team):
            _trace_assign(
                "候補坴下: 当日勤怠キーなし team=%s",
                ",".join(str(x) for x in team),
            )
            return None
        team_start = max(avail_dt[m] for m in team)
        if not _gpo.get("abolish_all_scheduling_limits"):
            machine_free_dt = _mach_floor_eff
            if team_start < machine_free_dt:
                team_start = machine_free_dt
            if team_start < day_floor:
                team_start = day_floor
        if b2_insp_ec_floor is not None and team_start < b2_insp_ec_floor:
            team_start = b2_insp_ec_floor
        team_end_limit = min(daily_status[m]["end_dt"] for m in team)
        if team_start >= team_end_limit:
            _trace_assign(
                "候補坴下: 開始>=終業 team=%s start=%s end_limit=%s",
                ",".join(str(x) for x in team),
                team_start,
                team_end_limit,
            )
            return None
        team_breaks = []
        for m in team:
            team_breaks.extend(daily_status[m]["breaks_dt"])
        team_breaks = merge_time_intervals(team_breaks)

        avg_eff = sum(daily_status[m]["efficiency"] for m in team) / len(team)
        if avg_eff <= 0:
            avg_eff = 0.01
        t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
        if t_eff <= 0:
            t_eff = 1.0
        eff_time_per_unit = (
            task["base_time_per_unit"]
            / avg_eff
            / t_eff
            * _surplus_team_time_factor(rq_base, len(team), extra_max)
        )
        _defer_min_contig = max(1, int(math.ceil(float(eff_time_per_unit))))

        def _refloor_trial_roll(ts: datetime) -> datetime:
            ts = max(ts, max(avail_dt[m] for m in team))
            if not _gpo.get("abolish_all_scheduling_limits"):
                mf = _mach_floor_eff
                if ts < mf:
                    ts = mf
                if ts < day_floor:
                    ts = day_floor
            if b2_insp_ec_floor is not None and ts < b2_insp_ec_floor:
                ts = b2_insp_ec_floor
            return ts

        _ts_before_defer = team_start
        team_start_d = _defer_team_start_past_prebreak_and_end_of_day(
            task,
            team,
            team_start,
            team_end_limit,
            team_breaks,
            _refloor_trial_roll,
            min_contiguous_work_mins=_defer_min_contig,
            eod_same_request_continuation_exempt=_eod_cont_exempt,
        )
        if team_start_d is None:
            _trace_assign(
                "候補坴下: 休憩帯内・終業直後(尝残)で当日試行 team=%s",
                ",".join(str(x) for x in team),
            )
            return None
        team_start = team_start_d
        if team_start >= team_end_limit:
            _trace_assign(
                "候補坴下: デファー後に開始>=終業 team=%s start=%s end_limit=%s",
                ",".join(str(x) for x in team),
                team_start,
                team_end_limit,
            )
            return None

        _, avail_mins, _ = calculate_end_time(
            team_start, 9999, team_breaks, team_end_limit
        )
        _trial_units_cap = int(avail_mins / eff_time_per_unit)
        if _trial_units_cap < 1:
            _trace_assign(
                "候補坴下: 実僝丝足 team=%s start=%s avail_mins=%s need_mins=%.2f",
                ",".join(str(x) for x in team),
                team_start,
                avail_mins,
                eff_time_per_unit,
            )
            return None
        if _eod_reject_capacity_units_below_threshold(
            _trial_units_cap,
            team_start,
            team_end_limit,
            eod_same_request_continuation_exempt=_eod_cont_exempt,
        ):
            _trace_assign(
                "候補坴下: 終業直後で当日坎容ロール数は閾値未満 team=%s cap=%s th=%s start=%s",
                ",".join(str(x) for x in team),
                _trial_units_cap,
                ASSIGN_EOD_DEFER_MAX_REMAINING_ROLLS,
                team_start,
            )
            return None
        work_mins_needed = int(eff_time_per_unit)
        _contig = _contiguous_work_minutes_until_next_break_or_limit(
            team_start, team_breaks, team_end_limit
        )
        if _contig < work_mins_needed:
            _trace_assign(
                "候補坴下: 休憩またねのため、連続実僝丝足 team=%s contiguous_min=%s need_mins=%s start=%s",
                ",".join(str(x) for x in team),
                _contig,
                work_mins_needed,
                team_start,
            )
            return None
        actual_end_dt, _, _ = calculate_end_time(
            team_start, work_mins_needed, team_breaks, team_end_limit
        )
        if dispatch_interval_mirror is not None and dispatch_interval_mirror.would_block_roll(
            machine_occ_key, team, team_start, actual_end_dt
        ):
            _trace_assign(
                "区間ミラー坴下: team=%s start=%s end=%s eq=%s",
                ",".join(str(x) for x in team),
                team_start,
                actual_end_dt,
                eq_line,
            )
            return None
        if pref_mem and pref_mem in op_list:
            lead_op = pref_mem
        else:
            lead_op = min(op_list, key=lambda mm: (skill_role_priority(mm)[1], mm))
        team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
        return {
            "team": team,
            "team_start": team_start,
            "actual_end_dt": actual_end_dt,
            "team_breaks": team_breaks,
            "avg_eff": avg_eff,
            "prio_sum": team_prio_sum,
            "op_list": op_list,
            "eff_time_per_unit": eff_time_per_unit,
            "lead_op": lead_op,
            "changeover_segments": _co_segs,
        }

    # 特別指定: 同一日・連続ロールは剝回フォームを優先（翌日へは挝う越さない）。
    _hist = task.get("assigned_history") or []
    _last_hist_date = _hist[-1].get("date") if _hist else None
    _same_day_last_roll = _last_hist_date == current_date.strftime("%m/%d")
    if preferred_team and _same_day_last_roll:
        pt = tuple(preferred_team)
        _pref_pt_ok = (not fixed_team_anchor) or all(
            m in pt for m in fixed_team_anchor
        )
        if _pref_pt_ok and all(m in capable_members and m in avail_dt for m in pt):
            got = _one_roll_from_team(pt)
            if got is not None:
                _cid_pt = _lookup_combo_sheet_row_id_for_preset_team(
                    preset_rows_assign, pt
                )
                return {
                    **got,
                    "extra_max": extra_max,
                    "rq_base": rq_base,
                    "need_src_line": need_src_line,
                    "extra_src_line": extra_src_line,
                    "machine": machine,
                    "machine_name": machine_name,
                    "eq_line": eq_line,
                    "req_num": req_num,
                    "max_team_size": max_team_size,
                    "combo_sheet_row_id": _cid_pt,
                    "combo_preset_team": pt if _cid_pt is not None else None,
                    "changeover_segments": _co_segs,
                }

    team_candidates: list[dict] = []
    # 組み合わせ表プリセットは「成立したら坳 return」せう」組み合わせ探索とまとめで
    # team_start / スラック付しタプルで最良を決める（シート上の優先度順は試行順のみ）。
    if preset_rows_assign:
        for _prio, sheet_rs, preset_team, combo_row_id in preset_rows_assign:
            bounds = _combo_preset_team_size_bounds(
                tuple(preset_team), sheet_rs, max_team_size
            )
            if bounds is None:
                continue
            lo_pt, hi_pt = bounds
            if fixed_team_anchor and not all(m in preset_team for m in fixed_team_anchor):
                continue
            if pref_mem is not None and pref_mem not in preset_team:
                continue
            if not all(m in capable_members for m in preset_team):
                continue
            if sum(1 for m in preset_team if skill_role_priority(m)[0] == "OP") < 1:
                continue
            got = _one_roll_from_team(
                tuple(preset_team), min_n=lo_pt, max_n=hi_pt
            )
            if got is not None:
                team_candidates.append(
                    {
                        **got,
                        "combo_sheet_row_id": combo_row_id,
                        "combo_preset_team": tuple(preset_team),
                    }
                )
    for tsize in range(req_num, max_team_size + 1):
        if fixed_team_anchor:
            _ft = list(fixed_team_anchor)
            others = [m for m in capable_members if m not in _ft]
            need_extra = tsize - len(_ft)
            if need_extra < 0:
                teams_iter = []
            elif need_extra == 0:
                teams_iter = [tuple(_ft)]
            elif len(others) >= need_extra:
                teams_iter = [
                    tuple(_ft + list(rest))
                    for rest in itertools.combinations(others, need_extra)
                ]
            else:
                teams_iter = []
        elif (
            pref_mem is not None
            and pref_mem in capable_members
            and skill_role_priority(pref_mem)[0] == "OP"
        ):
            if tsize == 1:
                _trace_assign(
                    "候補固定: 担当OP指定=%s のため、 1人フォームは当人のみ試行",
                    pref_mem,
                )
            others = [m for m in capable_members if m != pref_mem]
            if tsize == 1:
                teams_iter = [(pref_mem,)]
            elif len(others) >= tsize - 1:
                teams_iter = [
                    tuple([pref_mem] + list(rest))
                    for rest in itertools.combinations(others, tsize - 1)
                ]
            else:
                teams_iter = itertools.combinations(capable_members, tsize)
        else:
            teams_iter = itertools.combinations(capable_members, tsize)

        for team in teams_iter:
            got = _one_roll_from_team(team)
            if got is not None:
                team_candidates.append(
                    {
                        **got,
                        "combo_sheet_row_id": None,
                        "combo_preset_team": None,
                    }
                )

    if not team_candidates:
        _mem_max_end: datetime | None = None
        for _m in capable_members:
            if _m not in daily_status:
                continue
            _ed = daily_status[_m].get("end_dt")
            if isinstance(_ed, datetime):
                _mem_max_end = (
                    _ed if _mem_max_end is None else max(_mem_max_end, _ed)
                )
        if (
            len(capable_members) >= req_num
            and _mem_max_end is not None
            and isinstance(_mach_floor_eff, datetime)
            and _mach_floor_eff >= _mem_max_end
        ):
            logging.warning(
                "段階2: 依頼NO=%s 日付=%s 工程/機械=%s/%s でフォーム候補は0件。"
                "スキル革坈(OP/AS)は %s 人いしたは」設備の加工開始下限=%s は"
                "当日の担当候補の退勤(%s)以降のため、この日は割当でしません。"
                "master「機械カレンダー」で当該日・当該機械列に試行な記入はないか」"
                "または剝工程の占有で設備下限は終業まで繰り上はっていないか確認してください"
                "（配台ルール 3.2.1 機械カレンダー・トラブルシュート）。"
                "参考: changeover剝の設備空し下限=%s 占有キー=%s",
                task.get("task_id"),
                current_date,
                machine,
                machine_name,
                len(capable_members),
                _mach_floor_eff.strftime("%Y-%m-%d %H:%M"),
                _mem_max_end.strftime("%H:%M"),
                _prev_mach_before_co.strftime("%Y-%m-%d %H:%M"),
                machine_occ_key,
            )
        return None
    t_min = min(c["team_start"] for c in team_candidates)

    def _team_cand_key(c):
        return _team_assignment_sort_tuple(
            c["team"],
            c["team_start"],
            1,
            c["prio_sum"],
            t_min,
        )

    best_c = min(team_candidates, key=_team_cand_key)
    if best_c.get("combo_sheet_row_id") is None and preset_rows_assign:
        _lcid = _lookup_combo_sheet_row_id_for_preset_team(
            preset_rows_assign, tuple(best_c["team"])
        )
        if _lcid is not None:
            best_c = {
                **best_c,
                "combo_sheet_row_id": _lcid,
                "combo_preset_team": tuple(best_c["team"]),
            }
    return {
        **best_c,
        "extra_max": extra_max,
        "rq_base": rq_base,
        "need_src_line": need_src_line,
        "extra_src_line": extra_src_line,
        "machine": machine,
        "machine_name": machine_name,
        "eq_line": eq_line,
        "req_num": req_num,
        "max_team_size": max_team_size,
        "combo_sheet_row_id": best_c.get("combo_sheet_row_id"),
        "combo_preset_team": best_c.get("combo_preset_team"),
        "changeover_segments": _co_segs,
    }


def _trial_order_assign_probe_fails(
    task: dict,
    current_date: date,
    daily_status: dict,
    ctx: dict,
) -> bool:
    """
    睾在の avail_dt / machine_avail_dt / machine_handoff のスナップショットで
    `_assign_one_roll_trial_order_flow` は None になるなら True。
    機械枠は坝分でも人・休憩・ミラー等で詰まり」グローバル試行順の値は先頭行に張り付しのを防し。
    副作用なし（need 人数ログ用 set は毎回空）。
    """
    try:
        r = _assign_one_roll_trial_order_flow(
            task,
            current_date,
            daily_status,
            ctx["avail_dt"],
            ctx["machine_avail_dt"],
            ctx["task_queue"],
            ctx["skills_dict"],
            ctx["members"],
            ctx["req_map"],
            ctx["need_rules"],
            ctx["surplus_map"],
            ctx["global_priority_override"],
            ctx["macro_run_date"],
            ctx["macro_now_dt"],
            None,
            set(),
            team_combo_presets=ctx.get("team_combo_presets"),
            dispatch_interval_mirror=ctx.get("dispatch_interval_mirror"),
            machine_handoff=ctx["machine_handoff"],
        )
    except Exception as ex:
        logging.warning(
            "trial_order_assign_probe 例外のため、当該行は除外しない: task=%s err=%s",
            task.get("task_id"),
            ex,
        )
        return False
    return r is None


def _tasks_in_min_pending_dispatch_pool(
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> list:
    """`_min_pending_dispatch_trial_order_for_date` と同一の安価フィルタを通靎したタスクのリスト。"""
    out: list = []
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        sdr = t.get("start_date_req")
        if not isinstance(sdr, date) or sdr > current_date:
            continue
        if _task_not_yet_schedulable_due_to_dependency_or_b2_room(t, task_queue):
            continue
        if _task_fully_machine_calendar_blocked_on_date(
            t, current_date, daily_status, members
        ):
            continue
        if _task_no_machining_window_left_from_avail_floor(
            t,
            current_date,
            daily_status,
            members,
            machine_avail_dt,
            machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
        ):
            continue
        out.append(t)
    return out


def _effective_min_dispatch_trial_order_from_pool(
    pool: list,
    current_date: date,
    daily_status: dict,
    assign_probe_ctx: dict,
) -> int | None:
    """
    pool を昇順 dto で見で」**しの dto に属れる行のごう 1 件でも** 1 ロール割当プローブは通れみ
    しの dto を「実効の最尝試行順」とれる。
    先頭 dto 層は全滅（機械は空いでいるは人で穝ゝない等）のとき」次の dto に進みグローバル坜止を防し。
    プローブ無しのときは pool の最尝 dto を返す。
    """
    if not pool:
        return None
    dtos = sorted(
        {
            int(t.get("dispatch_trial_order") or 10**9)
            for t in pool
        }
    )
    if not assign_probe_ctx:
        return min(dtos)
    for d in dtos:
        at_d = [
            t
            for t in pool
            if int(t.get("dispatch_trial_order") or 10**9) == d
        ]
        if any(
            not _trial_order_assign_probe_fails(
                t, current_date, daily_status, assign_probe_ctx
            )
            for t in at_d
        ):
            return d
    return None


def _trial_order_first_schedule_pass(
    current_date: date,
    tasks_today: list,
    task_queue: list,
    daily_status: dict,
    machine_avail_dt: dict,
    avail_dt: dict,
    timeline_events: list,
    skills_dict: dict,
    members: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict,
    macro_run_date: date,
    macro_now_dt: datetime,
    _need_headcount_logged_orders: set,
    team_combo_presets: dict | None = None,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> bool:
    """
    ①当日候補を配台試行順の昇順に並きる（1 パス分）。
    **完全二相（§B-2 / §B-3）**: **フェーズ1**で **後続パイプライン行**（熱融着検査・巻返し）**を除し**候補（EC・他依頼・他工程）を試行順どより
    **`_drain_rolls_for_task`** し、**フェーズ2**は §B-2 検査＝§B-3 巻返し行のみ（**同一依頼の EC は全日で完走した後**に陝り候補化。
    EC 残はある日は `_trial_order_flow_eligible_tasks` で後続を外し、翌稼働日以降も EC のみ剝進れる。
    カレンダー通算で EC 完走後」`_run_b2_inspection_rewind_pass` は日付先頭から後続の値再走査れる）。
    EC と後続を **同一担当者で** 交互に詰ゝると EC はブロックされるため、従来はフェーズ1を先に詰ゝた。
    として後続は候補化した時点で **検査とともに実機械**のフェーズ1や **同一依頼の EC** は全日先に進むと」
    検査は `start_ge_end_initial`（設備空しは終業より後）で全日失敗する。§B-2/§B-3 後続はあるとしは
    「同一依頼EC・検査機と機械共有れるフェーズ1・後続」を **配台試行順**でマージし、
    坌順では **後続を EC より先に**」**しの他のフェーズ1** とあゝせで **配台試行順**で整列し
    **最大1ロールうつ**の値周回れる（マージ・rest とも一括ドレインしない。検査OPは他工程に
    同日坖り切られ start_ge_end_initial になるのを防し）。
    リワインド坴の後続行は坄ロールについで `_roll_pipeline_inspection_assign_room` よよよ
    `_roll_pipeline_b2_inspection_ec_completion_floor_dt`（EC ロール終了時刻下限）で整合する。
    試行順最尝の行の値は当日入らない場合でも」**坌もフェーズ内で次の試行順へ進み**他設備を埋ゝる。
    機械・人の空しはロールごとに更新れる（⑦⑧）。
    """
    _mc_w0 = datetime.combine(current_date, DEFAULT_START_TIME)
    _mh_init = _machine_handoff_state_from_timeline(timeline_events, current_date)
    _gpo = global_priority_override or {}
    _assign_probe_ctx: dict | None = None
    if STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT:
        _assign_probe_ctx = {
            "avail_dt": avail_dt,
            "machine_avail_dt": machine_avail_dt,
            "task_queue": task_queue,
            "skills_dict": skills_dict,
            "members": members,
            "req_map": req_map,
            "need_rules": need_rules,
            "surplus_map": surplus_map,
            "global_priority_override": global_priority_override,
            "macro_run_date": macro_run_date,
            "macro_now_dt": macro_now_dt,
            "machine_handoff": _mh_init,
            "team_combo_presets": team_combo_presets,
            "dispatch_interval_mirror": dispatch_interval_mirror,
        }
    _min_dispatch_eff: int | None = None
    if STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT and _assign_probe_ctx:
        _pool_min = _tasks_in_min_pending_dispatch_pool(
            task_queue,
            current_date,
            daily_status=daily_status,
            members=members,
            machine_avail_dt=machine_avail_dt,
            machine_day_start=_mc_w0,
            machine_handoff=_mh_init,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=bool(
                _gpo.get("abolish_all_scheduling_limits")
            ),
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
        _min_dispatch_eff = _effective_min_dispatch_trial_order_from_pool(
            _pool_min, current_date, daily_status, _assign_probe_ctx
        )
    eligible = _trial_order_flow_eligible_tasks(
        tasks_today,
        task_queue,
        current_date,
        daily_status=daily_status,
        members=members,
        machine_avail_dt=machine_avail_dt,
        machine_day_start=_mc_w0,
        machine_handoff=_mh_init,
        skills_dict=skills_dict,
        abolish_all_scheduling_limits=bool(_gpo.get("abolish_all_scheduling_limits")),
        dispatch_interval_mirror=dispatch_interval_mirror,
        min_dispatch_effective=_min_dispatch_eff,
        assign_probe_ctx=_assign_probe_ctx,
    )
    if not eligible:
        return False
    eligible_sorted = sorted(
        eligible,
        key=lambda t: int(t.get("dispatch_trial_order") or 10**9),
    )
    _mc_plan_end = _machine_calendar_planning_window_end_dt(
        current_date, daily_status, members
    )
    machine_handoff = {
        "last_tid": dict(_mh_init["last_tid"]),
        "last_eq": dict(_mh_init["last_eq"]),
        "started_today": set(_mh_init["started_today"]),
        "machining_today_occ": set(_mh_init.get("machining_today_occ") or set()),
        "last_machining_dt": dict(_mh_init.get("last_machining_dt") or {}),
        "last_machining_date": dict(_mh_init.get("last_machining_date") or {}),
        "last_lead_op": dict(_mh_init.get("last_lead_op") or {}),
        "last_machining_sub": dict(_mh_init.get("last_machining_sub") or {}),
    }

    def _drain_rolls_for_task(
        task: dict, *, max_rolls: int | None = None
    ) -> bool:
        preferred_team: tuple | None = None
        made_local = False
        rolls_done = 0
        while float(task.get("remaining_units") or 0) > 1e-12:
            if max_rolls is not None and rolls_done >= max_rolls:
                break
            res = _assign_one_roll_trial_order_flow(
                task,
                current_date,
                daily_status,
                avail_dt,
                machine_avail_dt,
                task_queue,
                skills_dict,
                members,
                req_map,
                need_rules,
                surplus_map,
                global_priority_override,
                macro_run_date,
                macro_now_dt,
                preferred_team,
                _need_headcount_logged_orders,
                team_combo_presets,
                dispatch_interval_mirror=dispatch_interval_mirror,
                machine_handoff=machine_handoff,
                timeline_events=timeline_events,
            )
            if res is None:
                break
            done_units = 1
            if task.get("roll_pipeline_inspection") or task.get(
                "roll_pipeline_rewind"
            ):
                _rp_room = _roll_pipeline_inspection_assign_room(
                    task_queue, str(task.get("task_id", "") or "").strip()
                )
                if _rp_room <= 1e-12:
                    break
                done_units = min(
                    1, int(min(_rp_room, math.ceil(task["remaining_units"])))
                )
            if done_units <= 0:
                break
            best_team = tuple(res["team"])
            lead_op = res["lead_op"]
            sub_members = [m for m in best_team if m != lead_op]
            best_start = res["team_start"]
            best_end = res["actual_end_dt"]
            best_breaks = res["team_breaks"]
            best_eff = res["avg_eff"]
            rq_base = res["rq_base"]
            extra_max = res["extra_max"]
            eq_line = res["eq_line"]
            machine_occ_key = _machine_occupancy_key_resolve(task, eq_line)
            _te_disp = parse_float_safe(task.get("task_eff_factor"), 1.0)
            if _te_disp <= 0:
                _te_disp = 1.0

            total_u = (
                math.ceil(task["total_qty_m"] / task["unit_m"]) if task["unit_m"] else 0
            )
            rem_u_before = math.ceil(task["remaining_units"])
            already_done = total_u - rem_u_before
            try:
                tot_qty = parse_float_safe(task.get("total_qty_m"), 0.0)
                done_qty = parse_float_safe(task.get("done_qty_reported"), 0.0)
                if tot_qty > 0:
                    pct_macro = max(
                        0, min(100, int(round((done_qty / tot_qty) * 100)))
                    )
                else:
                    pct_macro = 0
            except Exception:
                pct_macro = 0

            _mach_sub_line = ", ".join(
                str(s).strip() for s in sub_members if s and str(s).strip()
            )
            _co_append = list(res.get("changeover_segments") or [])
            _append_changeover_segments_to_timeline(
                timeline_events,
                dispatch_interval_mirror,
                avail_dt,
                daily_status,
                current_date=current_date,
                task_id=str(task.get("task_id") or ""),
                machine_occ_key=machine_occ_key,
                segments=_co_append,
                machining_lead_op=str(lead_op or "").strip() or None,
                machining_sub_str=_mach_sub_line or None,
                machine_handoff=machine_handoff,
            )
            timeline_events.append(
                {
                    "date": current_date,
                    "task_id": task["task_id"],
                    "machine": eq_line,
                    "machine_occupancy_key": machine_occ_key,
                    "op": lead_op,
                    "sub": ", ".join(sub_members),
                    "start_dt": best_start,
                    "end_dt": best_end,
                    "breaks": best_breaks,
                    "units_done": done_units,
                    "already_done_units": already_done,
                    "total_units": total_u,
                    "pct_macro": pct_macro,
                    "eff_time_per_unit": task["base_time_per_unit"]
                    / best_eff
                    / _te_disp
                    * _surplus_team_time_factor(
                        rq_base, len(best_team), extra_max
                    ),
                    "unit_m": task["unit_m"],
                    "event_kind": TIMELINE_EVENT_MACHINING,
                }
            )
            if dispatch_interval_mirror is not None:
                dispatch_interval_mirror.register_from_event(timeline_events[-1])
            task["remaining_units"] = max(
                0.0,
                float(task.get("remaining_units") or 0) - float(done_units),
            )
            op_main = (lead_op or "").strip()
            subs_part = ",".join(
                s.strip() for s in sub_members if s and str(s).strip()
            )
            team_s = f"{op_main}, {subs_part}" if subs_part else op_main
            req_num_run = int(res.get("req_num") or 0)
            extra_max_run = int(res.get("extra_max") or 0)
            need_surplus_assigned = (
                TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                and extra_max_run > 0
                and len(best_team) > req_num_run
            )
            names_ordered: list[str] = []
            if op_main:
                names_ordered.append(op_main)
            for _m in sub_members:
                if _m and str(_m).strip():
                    names_ordered.append(str(_m).strip())
            surplus_member_names = (
                names_ordered[req_num_run:]
                if need_surplus_assigned
                and len(names_ordered) > req_num_run
                else []
            )
            task["assigned_history"].append(
                {
                    "date": current_date.strftime("%m/%d"),
                    "team": team_s,
                    "done_m": int(done_units * task["unit_m"]),
                    "start_dt": best_start,
                    "end_dt": best_end,
                    "need_surplus_assigned": need_surplus_assigned,
                    "combo_sheet_row_id": res.get("combo_sheet_row_id"),
                    "surplus_member_names": surplus_member_names,
                }
            )
            for m in best_team:
                avail_dt[m] = best_end
            if not _gpo.get("abolish_all_scheduling_limits"):
                machine_avail_dt[machine_occ_key] = best_end
                _bump_machine_avail_after_roll_for_calendar(
                    current_date,
                    machine_occ_key,
                    machine_avail_dt,
                    machine_calendar_plan_end=_mc_plan_end,
                    machine_day_floor=_mc_w0,
                )
            machine_handoff["last_tid"][machine_occ_key] = str(
                task.get("task_id") or ""
            ).strip()
            machine_handoff["last_eq"][machine_occ_key] = eq_line
            machine_handoff["started_today"].add(machine_occ_key)
            machine_handoff["machining_today_occ"].add(machine_occ_key)
            machine_handoff["last_machining_dt"][machine_occ_key] = best_end
            machine_handoff["last_machining_date"][machine_occ_key] = current_date
            machine_handoff["last_lead_op"][machine_occ_key] = lead_op
            machine_handoff.setdefault("last_machining_sub", {})
            machine_handoff["last_machining_sub"][machine_occ_key] = _mach_sub_line
            if _trace_schedule_task_enabled(task.get("task_id")):
                _log_dispatch_trace_schedule(
                    task.get("task_id"),
                    "[配台トレース task=%s] ロール確定 メイン day=%s machine=%s machine_name=%s "
                    "start=%s end=%s 採用人数=%s req_num=%s メイン探索extra_max=%s "
                    "余剰人数適用(メイン)=%s team=%s",
                    task.get("task_id"),
                    current_date,
                    eq_line,
                    str(task.get("machine_name") or "").strip(),
                    best_start,
                    best_end,
                    len(best_team),
                    req_num_run,
                    extra_max_run,
                    need_surplus_assigned,
                    team_s,
                )
            preferred_team = best_team
            made_local = True
            rolls_done += 1
        return made_local

    def _is_b2_follower_phase2_row(t: dict) -> bool:
        _tid = str(t.get("task_id") or "").strip()
        return bool(
            (
                t.get("roll_pipeline_inspection")
                or t.get("roll_pipeline_rewind")
            )
            and _task_queue_has_roll_pipeline_ec_for_tid(task_queue, _tid)
        )

    phase1_tasks = [t for t in eligible_sorted if not _is_b2_follower_phase2_row(t)]
    phase2_tasks = [t for t in eligible_sorted if _is_b2_follower_phase2_row(t)]

    phase2_tids: set[str] = {
        str(t.get("task_id") or "").strip()
        for t in phase2_tasks
        if str(t.get("task_id") or "").strip()
    }
    phase2_mocc: set[str] = set()
    for t in phase2_tasks:
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        _pk = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
        if _pk:
            phase2_mocc.add(_pk)

    phase1_interleave: list = []
    phase1_rest: list = []
    for t in phase1_tasks:
        _tid1 = str(t.get("task_id") or "").strip()
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        _mk = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
        _same_tid_ec = bool(t.get("roll_pipeline_ec") and _tid1 and _tid1 in phase2_tids)
        _share_m = bool(_mk and _mk in phase2_mocc)
        if _same_tid_ec or _share_m:
            phase1_interleave.append(t)
        else:
            phase1_rest.append(t)

    def _b2_merged_sort_key(t: dict) -> tuple:
        # 坌も配台試行順では後続（検査・巻返し）を EC より先に回し、熱融着のタイムラインを
        # 同日早い段階で坖りに行し（§B-2 担当者分離で EC と検査は別メンバー想定）。
        _fol = bool(
            t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")
        )
        return (
            int(t.get("dispatch_trial_order") or 10**9),
            0 if _fol else 1,
            str(t.get("task_id") or ""),
            int(t.get("same_request_line_seq") or 0),
        )

    pass_made = False
    if phase2_tasks:
        merged_b2 = sorted(
            phase1_interleave + phase2_tasks,
            key=_b2_merged_sort_key,
        )
        _merged_row_ids = {id(x) for x in merged_b2}

        def _b2_rr_key(t: dict) -> tuple:
            if id(t) in _merged_row_ids:
                return _b2_merged_sort_key(t)
            return (
                int(t.get("dispatch_trial_order") or 10**9),
                2,
                str(t.get("task_id") or ""),
                int(t.get("same_request_line_seq") or 0),
            )

        all_rr = sorted(merged_b2 + phase1_rest, key=_b2_rr_key)
        while True:
            round_made = False
            for task in all_rr:
                if float(task.get("remaining_units") or 0) <= 1e-12:
                    continue
                if _drain_rolls_for_task(task, max_rolls=1):
                    round_made = True
            if not round_made:
                break
            pass_made = True
    else:
        for task in phase1_tasks:
            if _drain_rolls_for_task(task):
                pass_made = True
    return pass_made


def _run_b2_inspection_rewind_pass(
    sorted_dates: list,
    attendance_data: dict,
    task_queue: list,
    timeline_events: list,
    skills_dict: dict,
    members: list,
    equipment_list: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict,
    macro_run_date: date,
    macro_now_dt: datetime,
    _need_headcount_logged_orders: set,
    team_combo_presets: dict | None = None,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> bool:
    """
    §B-2 / §B-3: EC 坴を先に全日で進ゝた後」検査＝巻返し坴のみを日付先頭から再走査して配台れる。
    timeline_events を人・設備のブロックテーブルとして使い」日跨ねの占有を保挝れる。
    """
    target_tids: set[str] = set()
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        if not (t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")):
            continue
        tid = str(t.get("task_id", "") or "").strip()
        if not tid:
            continue
        if not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
            continue
        if not _pipeline_ec_fully_done_for_tid(task_queue, tid):
            continue
        target_tids.add(tid)
    if not target_tids:
        return False

    _gpo = global_priority_override or {}
    _any_progress = False
    _machine_day_start_cache: dict[date, datetime] = {}
    for current_date in sorted_dates:
        daily_status = attendance_data.get(current_date)
        if not daily_status:
            continue
        machine_avail_dt: dict = {}
        avail_dt: dict = {}
        for m in members:
            if m not in daily_status:
                continue
            st = daily_status[m]
            if st.get("eligible_for_assignment", st.get("is_working", False)):
                avail_dt[m] = st["start_dt"]
        if not avail_dt:
            continue

        _machine_day_start = _machine_day_start_cache.get(current_date)
        if _machine_day_start is None:
            _machine_day_start = datetime.combine(current_date, DEFAULT_START_TIME)
            _machine_day_start_cache[current_date] = _machine_day_start
        _seed_avail_from_timeline_for_date(
            timeline_events,
            current_date,
            machine_avail_dt,
            avail_dt,
            _machine_day_start,
        )
        if _gpo.get("abolish_all_scheduling_limits"):
            machine_avail_dt.clear()
        _mc_plan_end_b2 = _machine_calendar_planning_window_end_dt(
            current_date, daily_status, members
        )
        _apply_machine_calendar_floor_for_date(
            current_date,
            machine_avail_dt,
            equipment_list,
            _machine_day_start,
            machine_calendar_plan_end=_mc_plan_end_b2,
        )

        tasks_today = [
            t
            for t in task_queue
            if float(t.get("remaining_units") or 0) > 1e-12
            and (
                t.get("roll_pipeline_inspection")
                or t.get("roll_pipeline_rewind")
            )
            and str(t.get("task_id", "") or "").strip() in target_tids
            and t.get("start_date_req") <= current_date
        ]
        if not tasks_today:
            continue

        _sched_max_passes = max(96, max(1, len(tasks_today)) * 15)
        _sched_pi = 0
        while _sched_pi < _sched_max_passes:
            _sched_pi += 1
            _made = _trial_order_first_schedule_pass(
                current_date,
                tasks_today,
                task_queue,
                daily_status,
                machine_avail_dt,
                avail_dt,
                timeline_events,
                skills_dict,
                members,
                req_map,
                need_rules,
                surplus_map,
                global_priority_override,
                macro_run_date,
                macro_now_dt,
                _need_headcount_logged_orders,
                team_combo_presets,
                dispatch_interval_mirror=dispatch_interval_mirror,
            )
            if not _made:
                break
            _any_progress = True
    return _any_progress


def _timeline_event_team_names_set(ev: dict) -> set:
    names: set = set()
    op = str(ev.get("op") or "").strip()
    if op:
        names.add(op)
    sub = str(ev.get("sub") or "").strip()
    if sub:
        for s in sub.split(","):
            t = s.strip()
            if t:
                names.add(t)
    return names


def _task_dict_for_timeline_event(ev: dict, task_queue: list) -> dict | None:
    tid = str(ev.get("task_id") or "").strip()
    if not tid:
        return None
    eq = str(ev.get("machine") or "").strip()
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        t_eq = str(t.get("equipment_line_key") or t.get("machine") or "").strip()
        if t_eq == eq:
            return t
    for t in task_queue:
        if str(t.get("task_id") or "").strip() == tid:
            return t
    return None


def _member_overlaps_busy(
    busy_map: dict, member: str, st: datetime, ed: datetime
) -> bool:
    for bs, be in busy_map.get(member, ()):
        if st < be and bs < ed:
            return True
    return False


def append_surplus_staff_after_main_dispatch(
    timeline_events: list,
    attendance_data: dict,
    skills_dict: dict,
    members: list,
    task_queue: list,
    req_map: dict,
    need_rules: list,
    surplus_map: dict,
    global_priority_override: dict | None,
) -> int:
    """
    need「配台時追加人数＝余力時追加人数」行の上限まで」メイン割付で採用ししれなかった枠を追記れる。
    坄タイムラインブロックについで」しの時間帯に他ブロックへ未坂加（区間重なりなし）で
    eligible かつ OP/AS スキルの者をサブに追加れる。
    日次始業（event_kind は加工以外）は本処理の対象外（余剰サブは加工にのみ追記）。
    """
    gpo = global_priority_override or {}
    if not surplus_map or TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
        return 0

    busy: dict[str, list[tuple[datetime, datetime]]] = defaultdict(list)
    for e in timeline_events:
        st = e.get("start_dt")
        ed = e.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime):
            continue
        for name in _timeline_event_team_names_set(e):
            busy[name].append((st, ed))

    appended_total = 0
    sorted_evs = sorted(
        (
            e
            for e in timeline_events
            if isinstance(e.get("start_dt"), datetime)
            and isinstance(e.get("end_dt"), datetime)
        ),
        key=lambda x: (x.get("date"), x.get("start_dt") or datetime.min),
    )

    for ev in sorted_evs:
        d = ev.get("date")
        if d is None or d not in attendance_data:
            continue
        if not _is_machining_timeline_event(ev):
            continue
        daily_status = attendance_data[d]
        task = _task_dict_for_timeline_event(ev, task_queue)
        if task is None:
            continue
        machine = task.get("machine")
        machine_name = str(task.get("machine_name") or "").strip()
        tid = str(task.get("task_id") or "").strip()

        if TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY:
            req_num = resolve_need_required_op(
                str(machine or "").strip(),
                machine_name,
                tid,
                req_map,
                need_rules,
            )
        else:
            ro = task.get("required_op")
            if ro is not None:
                try:
                    riv = int(ro)
                    if riv >= 1:
                        req_num = riv
                    else:
                        req_num = resolve_need_required_op(
                            str(machine or "").strip(),
                            machine_name,
                            tid,
                            req_map,
                            need_rules,
                        )
                except (TypeError, ValueError):
                    req_num = resolve_need_required_op(
                        str(machine or "").strip(),
                        machine_name,
                        tid,
                        req_map,
                        need_rules,
                    )
            else:
                req_num = resolve_need_required_op(
                    str(machine or "").strip(),
                    machine_name,
                    tid,
                    req_map,
                    need_rules,
                )
        if gpo.get("ignore_need_minimum"):
            req_num = 1

        extra_max_sheet = resolve_need_surplus_extra_max(
            str(machine or "").strip(),
            machine_name,
            tid,
            surplus_map,
            need_rules,
        )
        if extra_max_sheet <= 0:
            continue

        names = _timeline_event_team_names_set(ev)
        team_size = len(names)
        cap_add = req_num + extra_max_sheet - team_size
        if cap_add <= 0:
            continue

        skill_meta_cache: dict = {}

        def skill_role_priority(mem):
            if gpo.get("ignore_skill_requirements"):
                return ("OP", 100)
            if mem not in skill_meta_cache:
                srow = skills_dict.get(mem, {})
                machine_proc = str(machine or "").strip()
                v = ""
                if machine_proc and machine_name:
                    v = srow.get(f"{machine_proc}+{machine_name}", "")
                elif machine_name:
                    v = srow.get(machine_name, "")
                elif machine_proc:
                    v = srow.get(machine_proc, "")
                skill_meta_cache[mem] = parse_op_as_skill_cell(v)
            return skill_meta_cache[mem]

        capable = []
        for mem in members:
            if mem not in daily_status:
                continue
            st_ent = daily_status[mem]
            if not st_ent.get(
                "eligible_for_assignment", st_ent.get("is_working", False)
            ):
                continue
            if skill_role_priority(mem)[0] not in ("OP", "AS"):
                continue
            capable.append(mem)
        capable.sort(key=lambda mm: (skill_role_priority(mm)[1], mm))

        st = ev["start_dt"]
        ed = ev["end_dt"]
        candidates = [
            m
            for m in capable
            if m not in names and not _member_overlaps_busy(busy, m, st, ed)
        ]
        candidates.sort(
            key=lambda mm: (
                0 if skill_role_priority(mm)[0] == "AS" else 1,
                skill_role_priority(mm)[1],
                mm,
            )
        )

        chosen = candidates[:cap_add]
        if not chosen:
            continue

        team_size_before = team_size
        final_team_size = team_size_before + len(chosen)
        highlight_surplus = final_team_size > req_num

        old_sub = str(ev.get("sub") or "").strip()
        parts = [s.strip() for s in old_sub.split(",") if s.strip()]
        parts.extend(chosen)
        ev["sub"] = ", ".join(parts)
        for m in chosen:
            busy[m].append((st, ed))
        appended_total += len(chosen)

        op_sync = str(ev.get("op") or "").strip()
        subs_sync = ",".join(
            s.strip()
            for s in str(ev.get("sub") or "").split(",")
            if s.strip()
        )
        team_sync = f"{op_sync}, {subs_sync}" if subs_sync else op_sync

        _hist = task.get("assigned_history")
        if _hist:
            for h in _hist:
                if (
                    h.get("start_dt") == st
                    and h.get("end_dt") == ed
                ):
                    if highlight_surplus:
                        h["need_surplus_assigned"] = True
                    h["team"] = team_sync
                    prev_pd = h.get("post_dispatch_surplus_names") or []
                    h["post_dispatch_surplus_names"] = prev_pd + [
                        str(x) for x in chosen
                    ]
                    break

        if _trace_schedule_task_enabled(tid):
            _log_dispatch_trace_schedule(
                tid,
                "[配台トレース task=%s] 余力追記(メイン完了後) day=%s machine=%s machine_name=%s "
                "start=%s end=%s 追記人数=%s 追記剝人数=%s 追記後人数=%s req_num=%s "
                "need追加枠(シート)=%s 履歴黄(余剰人数超靎)=%s 追記メンバー=%s",
                tid,
                d,
                str(machine or "").strip(),
                machine_name,
                st,
                ed,
                len(chosen),
                team_size_before,
                final_team_size,
                req_num,
                extra_max_sheet,
                highlight_surplus,
                ",".join(chosen),
            )

    return appended_total


# =========================================================
# 3. メイン計画生成 (日毎ループ・挝う越し対応)
#    段階2の本体。plan_simulation_stage2 からのみ呼みれる想定。
#    配台計画シート読込 → タスクキュー → 日付ごとに設備・OP割付 → 結果ブック出力。
# =========================================================
def generate_plan():
    """
    段階2のメイン処理。戻り値なし（ログ・Excel 出力で完絝）。

    前提: 環境変数 TASK_INPUT_WORKBOOK、カレントディレクトリがスクリプトフォルダ。
    出力: ``output_dir`` 直下の ``production_plan_multi_day_*.xlsx`` / ``member_schedule_*.xlsx``（実行直前に同名パターンを削除しようとする。ファイル名はデータ抽出時刻＋実行時刻サフィックスで実行ごとに一意）、および log/execution_log.txt。
    """
    master_abs = os.path.abspath(os.path.join(os.getcwd(), MASTER_FILE))
    with _override_default_factory_hours_from_master(master_abs):
        _generate_plan_impl()


def refresh_equipment_gantt_actual_detail_only() -> str:
    """
    段階2全体を実行せず、「結果_設備ガント_実績明細」相当のシートだけを
    ``output_dir`` 直下の ``ACTUAL_DETAIL_GANTT_REFRESH_FILENAME`` に出力する。

    マクロブックの勤怠・実績明細DATA・master（工場枠・定常枠・機械カレンダー等）を
    段階2と同様に読み、実績タイムラインのみ描画する。

    既存の出力ファイルがあり、メタ行の「データ抽出」表示が今回採用した
    ``データ抽出時間``（加工実績明細DATA 優先、無ければ加工計画DATA）の表示と
    一致する場合は、再生成をスキップしてそのファイルパスを返す。

    Returns:
        生成した（またはスキップ時は既存の）xlsx の絶対パス。

    Raises:
        PlanningValidationError: メンバー0人・表示対象日なし・実績イベント空など。
    """
    master_abs = os.path.abspath(os.path.join(os.getcwd(), MASTER_FILE))
    with _override_default_factory_hours_from_master(master_abs):
        global _MACHINE_CALENDAR_BLOCKS_BY_DATE
        global _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE
        global _STAGE2_REGULAR_SHIFT_START
        (
            _skills_dict,
            members,
            equipment_list,
            _req_map,
            _need_rules,
            _surplus_map,
            _need_combo_col_index,
        ) = load_skills_and_needs()
        if not members:
            raise PlanningValidationError(
                "実績明細ガントのみ更新を中断しました: メンバーが0人です（マスタ skills を確認してください）。"
            )
        try:
            _MACHINE_CALENDAR_BLOCKS_BY_DATE = load_machine_calendar_occupancy_blocks(
                master_abs,
                equipment_list,
            )
        except Exception as e:
            logging.warning(
                "機械カレンダー: 読込例外のため占有なしとして続行します (%s)", e
            )
            _MACHINE_CALENDAR_BLOCKS_BY_DATE = {}
        try:
            _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE = (
                load_machine_daily_startup_settings(master_abs)
            )
        except Exception as e:
            logging.warning(
                "機械日次始業準備設定: 読込例外のため無視します (%s)", e
            )
            _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE = {}
        try:
            _rs_a15, _ = _read_master_main_regular_shift_times(master_abs)
            _STAGE2_REGULAR_SHIFT_START = _rs_a15
        except Exception as e:
            logging.warning(
                "定常開始(A15) 読込失敗: 日次始業は従来の勤怠 forward にフォールバック (%s)", e
            )
            _STAGE2_REGULAR_SHIFT_START = None

        # 実績明細ガントの「データ抽出」は、加工計画DATAではなく「加工実績明細DATA」のデータ抽出時間を優先する。
        # ただし勤怠の当日判定などの実行基準もこの抽出時刻と揃える。
        data_extract_dt, plan_base_dt_column = _extract_data_extraction_datetime()
        base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
        run_date = base_now_dt.date()
        data_extract_dt_str = (
            base_now_dt.strftime("%Y/%m/%d %H:%M:%S")
            if data_extract_dt is not None
            else "—"
        )

        attendance_data, ai_log_data = load_attendance_and_analyze(members)
        global_priority_raw = load_main_sheet_global_priority_override_text()
        global_priority_override = analyze_global_priority_override_comment(
            global_priority_raw, members, run_date.year, ai_sheet_sink=ai_log_data
        )
        _factory_closure_dates: set[date] = set()
        for _iso in global_priority_override.get("factory_closure_dates") or []:
            _d = parse_optional_date(_iso)
            if _d is not None:
                _factory_closure_dates.add(_d)
        if _factory_closure_dates:
            apply_factory_closure_dates_to_attendance(
                attendance_data, members, _factory_closure_dates
            )

        sorted_dates = sorted(list(attendance_data.keys()))
        sorted_dates = [d for d in sorted_dates if d >= run_date]
        if not sorted_dates:
            raise PlanningValidationError(
                "実績明細ガントのみ更新を中断しました: 当日以降の処理対象日付がありません。"
            )

        _reg_shift_start, _reg_shift_end = _read_master_main_regular_shift_times(
            master_abs
        )

        df_actual_detail = load_machining_actual_detail_df()

        def _first_valid_dt_from_df_col(_df, _col) -> datetime | None:
            try:
                if _df is None or _col not in _df.columns:
                    return None
                for _v in _df[_col].tolist():
                    if _v is None or (isinstance(_v, float) and pd.isna(_v)):
                        continue
                    _dt = pd.to_datetime(_v, errors="coerce")
                    if pd.isna(_dt):
                        continue
                    if isinstance(_dt, pd.Timestamp):
                        return _dt.to_pydatetime()
                    return _dt if isinstance(_dt, datetime) else None
            except Exception:
                return None
            return None

        # 加工実績明細DATA の「データ抽出時間」を最優先（無い/空なら従来どおり加工計画DATA基準）
        detail_extract_dt = _first_valid_dt_from_df_col(
            df_actual_detail, TASK_COL_DATA_EXTRACTION_TIME
        )
        if detail_extract_dt is not None:
            base_now_dt = detail_extract_dt
            run_date = base_now_dt.date()
            data_extract_dt_str = base_now_dt.strftime("%Y/%m/%d %H:%M:%S")
            plan_base_dt_column = f"{ACTUAL_DETAIL_SHEET_NAME}:{TASK_COL_DATA_EXTRACTION_TIME}"

        logging.info(
            "実績明細ガントのみ: 抽出基準日時 %s（%s）",
            base_now_dt.strftime("%Y/%m/%d %H:%M:%S"),
            plan_base_dt_column if data_extract_dt is not None else "現在時刻フォールバック",
        )

        out_path = os.path.join(output_dir, ACTUAL_DETAIL_GANTT_REFRESH_FILENAME)
        prev_extract_display = _read_existing_equipment_gantt_data_extract_display(
            out_path, RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME
        )
        cur_extract_display = (data_extract_dt_str or "").strip()
        if (
            cur_extract_display
            and cur_extract_display not in ("—", "-")
            and prev_extract_display is not None
            and prev_extract_display.strip() == cur_extract_display
        ):
            logging.info(
                "実績明細ガントのみ: データ抽出時間が前回出力と同一のためファイル更新をスキップしました（%s）。",
                cur_extract_display,
            )
            return os.path.abspath(out_path)

        detail_timeline_events: list = []
        sorted_dates_detail = list(sorted_dates)
        chart_title_actual_detail = "湖南工場 加工実績（明細）"
        if df_actual_detail is not None and len(df_actual_detail) > 0:
            sorted_dates_detail = _sorted_dates_union_actual_bounds_df(
                sorted_dates, df_actual_detail
            )
            d_from = _parse_env_optional_date(ENV_GANTT_ACTUAL_DETAIL_DATE_FROM)
            d_to = _parse_env_optional_date(ENV_GANTT_ACTUAL_DETAIL_DATE_TO)
            if d_from is not None or d_to is not None:
                n_before = len(sorted_dates_detail)
                filtered_detail_dates = _sorted_dates_filter_inclusive_range(
                    sorted_dates_detail, d_from, d_to
                )
                if not filtered_detail_dates and sorted_dates_detail:
                    logging.warning(
                        "実績明細ガント: 日付範囲フィルタで表示日が0件になったためフィルタを無視します。"
                        "（%s=%r, %s=%r）",
                        ENV_GANTT_ACTUAL_DETAIL_DATE_FROM,
                        os.environ.get(ENV_GANTT_ACTUAL_DETAIL_DATE_FROM, ""),
                        ENV_GANTT_ACTUAL_DETAIL_DATE_TO,
                        os.environ.get(ENV_GANTT_ACTUAL_DETAIL_DATE_TO, ""),
                    )
                else:
                    sorted_dates_detail = filtered_detail_dates
                    logging.info(
                        "実績明細ガント: 表示日を %s 日 → %s 日に絞りました（FROM=%s, TO=%s）。",
                        n_before,
                        len(sorted_dates_detail),
                        d_from.isoformat() if d_from else "（指定なし）",
                        d_to.isoformat() if d_to else "（指定なし）",
                    )
                    rng_lo = d_from.isoformat() if d_from else "…"
                    rng_hi = d_to.isoformat() if d_to else "…"
                    chart_title_actual_detail = (
                        f"{chart_title_actual_detail}（表示 {rng_lo}～{rng_hi}）"
                    )
            detail_timeline_events = build_actual_timeline_events(
                df_actual_detail,
                equipment_list,
                sorted_dates_detail,
                log_sheet_name=ACTUAL_DETAIL_SHEET_NAME,
                roll_detail=True,
            )

        if not detail_timeline_events:
            raise PlanningValidationError(
                "実績明細ガントを生成できるイベントがありません。"
                "「加工実績明細DATA」の有無・日付・必須列を確認してください。"
            )

        _try_remove_path_with_retries(out_path)

        gantt_detail_tl_label_specs: list = []
        gantt_detail_tl_day_blocks: list = []
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            gantt_detail_tl_label_specs, gantt_detail_tl_day_blocks = (
                _write_results_equipment_gantt_sheet(
                    writer,
                    [],
                    equipment_list,
                    sorted_dates_detail,
                    attendance_data,
                    data_extract_dt_str,
                    base_now_dt,
                    actual_timeline_events=detail_timeline_events,
                    regular_shift_times=(_reg_shift_start, _reg_shift_end),
                    plan_rows=False,
                    chart_title=chart_title_actual_detail,
                    sheet_name_override=RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME,
                )
            )
            wb = writer.book
            for _sn in list(wb.sheetnames):
                if _sn != RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME:
                    wb.remove(wb[_sn])

        _stage2_try_add_gantt_timeline_shape_labels(
            out_path,
            gantt_detail_tl_label_specs,
            gantt_detail_tl_day_blocks,
            sheet_name=RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME,
        )
        logging.info(
            "実績明細ガントのみ: %s を出力しました。",
            os.path.basename(out_path),
        )
        return os.path.abspath(out_path)


def _generate_plan_impl():
    # 配台トレース（設定シート A3 以降のみ）は」メンバー0人等で早期 return しても
    # execution_log に残るよご skills 読込より剝で確定・ログれる。
    global TRACE_SCHEDULE_TASK_IDS, DEBUG_DISPATCH_ONLY_TASK_IDS
    _wb_trace = (os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK)
    _ids_from_sheet = _read_trace_schedule_task_ids_from_config_sheet(_wb_trace)
    TRACE_SCHEDULE_TASK_IDS = frozenset(
        str(x).strip() for x in _ids_from_sheet if str(x).strip()
    )
    if _ids_from_sheet:
        _preview = _ids_from_sheet[:25]
        _suffix = " …" if len(_ids_from_sheet) > 25 else ""
        logging.info(
            "設定シート「%s」A3 以降: トレース用依頼NOを %s 件読み込み（%s%s）",
            APP_CONFIG_SHEET_NAME,
            len(_ids_from_sheet),
            ", ".join(_preview),
            _suffix,
        )
    else:
        logging.info(
            "設定シート「%s」A3 以降: トレース用依頼NOは無し（空またはシート無し）",
            APP_CONFIG_SHEET_NAME,
        )
    if TRACE_SCHEDULE_TASK_IDS:
        logging.info(
            "配台トレース: 有効 task_id = %s（設定シート A3 以降）",
            ", ".join(sorted(TRACE_SCHEDULE_TASK_IDS)),
        )
    else:
        logging.info(
            "配台トレース: 対象なし（[配台トレース …] ログは出ません）"
        )
    _ids_debug_dispatch_raw = _read_debug_dispatch_task_ids_from_config_sheet(_wb_trace)
    _dbg_norm: list[str] = []
    for _dx in _ids_debug_dispatch_raw:
        _dt = planning_task_id_str_from_scalar(_dx)
        if _dt:
            _dbg_norm.append(_dt)
    DEBUG_DISPATCH_ONLY_TASK_IDS = frozenset(_dbg_norm)
    if DEBUG_DISPATCH_ONLY_TASK_IDS:
        logging.warning(
            "デバッグ配台: 「%s」B3以降により配台対象を %s 件の依頼NOに陝定しした: %s",
            APP_CONFIG_SHEET_NAME,
            len(DEBUG_DISPATCH_ONLY_TASK_IDS),
            ", ".join(sorted(DEBUG_DISPATCH_ONLY_TASK_IDS)),
        )
        _show_stage2_debug_dispatch_mode_dialog(sorted(DEBUG_DISPATCH_ONLY_TASK_IDS))
    if TRACE_TEAM_ASSIGN_TASK_ID:
        logging.info(
            "環境変数 TRACE_TEAM_ASSIGN_TASK_ID=%r → フォーム割当トレース有効",
            TRACE_TEAM_ASSIGN_TASK_ID,
        )

    _reset_dispatch_trace_per_task_logfiles()

    (
        skills_dict,
        members,
        equipment_list,
        req_map,
        need_rules,
        surplus_map,
        need_combo_col_index,
    ) = load_skills_and_needs()
    team_combo_presets = load_team_combination_presets_from_master()
    if team_combo_presets:
        _nrules = sum(len(v) for v in team_combo_presets.values())
        logging.info(
            "組み合わせ表: 工程+機械キー %s 種類・編集行 %s を配台プリセットとして読み込みました。",
            len(team_combo_presets),
            _nrules,
        )
    elif TEAM_ASSIGN_USE_MASTER_COMBO_SHEET:
        logging.info(
            "組み合わせ表: プリセット無し（シート欠如・空・または読込失敗）。従来のフォーム探索のみ。"
        )
    if not members:
        master_abs = os.path.abspath(MASTER_FILE)
        logging.error(
            "段階2を中断しました: メンバーは0人です（マスタの skills は空」または読み込み失敗）。"
            " 期待パス: %s （カレント: %s）。テストコード直下に master.xlsm を置し」"
            "planning_core のカレントはしのフォルダになるよご python\\ 配置を確認してください。"
            " この状態では production_plan / member_schedule は出力されません。",
            master_abs,
            os.getcwd(),
        )
        return
    global _MACHINE_CALENDAR_BLOCKS_BY_DATE
    global _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE
    global _STAGE2_REGULAR_SHIFT_START
    try:
        _MACHINE_CALENDAR_BLOCKS_BY_DATE = load_machine_calendar_occupancy_blocks(
            os.path.abspath(os.path.join(os.getcwd(), MASTER_FILE)),
            equipment_list,
        )
    except Exception as e:
        logging.warning(
            "機械カレンダー: 読込例外のため、占有なしとして続行しした (%s)", e
        )
        _MACHINE_CALENDAR_BLOCKS_BY_DATE = {}
    try:
        _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE = load_machine_daily_startup_settings(
            os.path.abspath(os.path.join(os.getcwd(), MASTER_FILE))
        )
    except Exception as e:
        logging.warning(
            "機械日次始業準備設定: 読込例外のため、無視しした (%s)", e
        )
        _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE = {}
    try:
        _rs_a15, _ = _read_master_main_regular_shift_times(
            os.path.abspath(os.path.join(os.getcwd(), MASTER_FILE))
        )
        _STAGE2_REGULAR_SHIFT_START = _rs_a15
        if _rs_a15 is not None:
            logging.info(
                "日次始業準備: 定常開始 master メイン A15=%s を採用（[開始, 開始+分) を時刻で占有。A15 無効時は従来の勤怠 forward）",
           
                _rs_a15.strftime("%H:%M"),
            )
    except Exception as e:
        logging.warning("定常開始(A15) 読込失敗: 日次始業は従来の勤怠 forward にフォールバック (%s)", e)
        _STAGE2_REGULAR_SHIFT_START = None
    if _MACHINE_CALENDAR_BLOCKS_BY_DATE:
        _n_iv = sum(
            len(ivs)
            for _dm in _MACHINE_CALENDAR_BLOCKS_BY_DATE.values()
            for ivs in _dm.values()
        )
        logging.info(
            "機械カレンダー: %s 日分・設備占有ブロック合計 %s 件を配台に反映しました。",
            len(_MACHINE_CALENDAR_BLOCKS_BY_DATE),
            _n_iv,
        )
    reset_gemini_usage_tracker()
    _clear_stage2_blocking_message_file()
    if (
        not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
        and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
    ):
        logging.info(
            "need配台時追加人数: メイン割付は基本必須人数のみ。"
            "余力は全シミュレーション後」未割当かつスキル保有社をサブに追記しました。"
            "（メインで増員探索れる従来挙動: TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS=1）"
        )

    # 段階2の基準日時は「マクロ実行時刻」ではなく加工計画DATA「データ抽出時間」（なければ「抽出時間」→「データ抽出日」）
    data_extract_dt, plan_base_dt_column = _extract_data_extraction_datetime()
    base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
    run_date = base_now_dt.date()
    data_extract_dt_str = (
        base_now_dt.strftime("%Y/%m/%d %H:%M:%S") if data_extract_dt is not None else "—"
    )
    logging.info(
        "計画基準日時: %s（%s）",
        base_now_dt.strftime("%Y/%m/%d %H:%M:%S"),
        plan_base_dt_column if data_extract_dt is not None else "現在時刻フォールバック",
    )

    attendance_data, ai_log_data = load_attendance_and_analyze(members)
    global_priority_raw = load_main_sheet_global_priority_override_text()
    global_priority_override = analyze_global_priority_override_comment(
        global_priority_raw, members, run_date.year,         ai_sheet_sink=ai_log_data
    )
    _factory_closure_dates: set[date] = set()
    for _iso in global_priority_override.get("factory_closure_dates") or []:
        _d = parse_optional_date(_iso)
        if _d is not None:
            _factory_closure_dates.add(_d)
    if _factory_closure_dates:
        apply_factory_closure_dates_to_attendance(
            attendance_data, members, _factory_closure_dates
        )
        logging.info(
            "メイン・グローバルコメント: 工場休業扱いの日付 → %s",
            ", ".join(str(x) for x in sorted(_factory_closure_dates)),
        )
    ai_log_data["メイン_グローバル_工場休業日(解析)"] = (
        ", ".join(str(x) for x in sorted(_factory_closure_dates))
        if _factory_closure_dates
        else "（なし）"
    )
    _sn = str(global_priority_override.get("scheduler_notes_ja") or "").strip()
    if _sn:
        ai_log_data["メイン_グローバル_未適用メモ(AI)"] = _sn[:2000]

    sorted_dates = sorted(list(attendance_data.keys()))
    # 結果シートは「基準日（データ抽出時間／抽出時間／データ抽出日）」以降のみ表示・計画対象とする
    sorted_dates = [d for d in sorted_dates if d >= run_date]
    if not sorted_dates:
        logging.error("当日以降の処理対象日付はありません。")
        _try_write_main_sheet_gemini_usage_summary("段階2")
        return

    # タスク入力: ブック内「配台計画_タスク入力」（段階1で出力→取り込み後に編集）
    try:
        tasks_df = load_planning_tasks_df()
    except PlanningValidationError:
        raise
    except Exception as e:
        logging.error(f"配台計画タスクシート読み込みエラー: {e}")
        _try_write_main_sheet_gemini_usage_summary("段階2")
        return

    if DEBUG_DISPATCH_ONLY_TASK_IDS:
        _n_tasks_before = len(tasks_df)
        _dbg_mask = tasks_df.apply(
            lambda row: planning_task_id_str_from_plan_row(row)
            in DEBUG_DISPATCH_ONLY_TASK_IDS,
            axis=1,
        )
        tasks_df = tasks_df.loc[_dbg_mask].copy()
        _n_tasks_after = len(tasks_df)
        logging.warning(
            "デバッグ配台: 「%s」の行を %s → %s に絞り込みました。",
            PLAN_INPUT_SHEET_NAME,
            _n_tasks_before,
            _n_tasks_after,
        )
        if _n_tasks_after == 0:
            logging.error(
                "デバッグ配台: B3以降の依頼NOに一致する行はありません。段階2を中断しした。"
            )
            _try_write_main_sheet_gemini_usage_summary("段階2")
            return

    if global_priority_raw.strip():
        snip = global_priority_raw[:2500]
        if len(global_priority_raw) > 2500:
            snip += "…"
        ai_log_data["メイン_再優先特別記載(原文)"] = snip
    else:
        ai_log_data["メイン_再優先特別記載(原文)"] = (
            "（空」またはメインシートに「グローバルコメント」見出しは見つかりません）"
        )
    ai_log_data["メイン_再優先特別記載(AI)"] = json.dumps(
        global_priority_override, ensure_ascii=False
    )
    if global_priority_override.get("ignore_skill_requirements"):
        logging.warning(
            "メイン再優先特記: スキル覝件を無視して配台しした。%s",
            global_priority_override.get("interpretation_ja", ""),
        )
    if global_priority_override.get("ignore_need_minimum"):
        logging.warning(
            "メイン再優先特記: フォーム人数を1坝に固定しした（need・行の必須OP上書きより優先）。%s",
            global_priority_override.get("interpretation_ja", ""),
        )
    if global_priority_override.get("abolish_all_scheduling_limits"):
        logging.warning(
            "メイン再優先特記: 設備専有・原板同日開始・指定開始時刻・マクロ実行時刻下限を適用しません。%s",
            global_priority_override.get("interpretation_ja", ""),
        )

    # 「当日」判定と最早開始時刻には基準日時（データ抽出時間→抽出時間→データ抽出日）を使う
    macro_now_dt = base_now_dt
    macro_run_date = macro_now_dt.date()
    ai_task_by_tid = analyze_task_special_remarks(
        tasks_df, reference_year=run_date.year,         ai_sheet_sink=ai_log_data
    )
    task_queue = build_task_queue_from_planning_df(
        tasks_df,
        run_date,
        req_map,
        ai_task_by_tid,
        global_priority_override,
        equipment_list,
    )
    # 開始日は非稼働日の場合は」直後の稼働日へ補正（例: 4/4, 4/5 は非稼働なら 4/3 へ）
    working_days = [
        d for d in sorted_dates
        if any(attendance_data[d][m]["is_working"] for m in attendance_data[d])
    ]
    if working_days:
        for t in task_queue:
            req_d = t.get("start_date_req")
            if not isinstance(req_d, date):
                continue
            if req_d in working_days:
                continue
            prev_work = None
            for wd in working_days:
                if wd <= req_d:
                    prev_work = wd
                else:
                    break
            if prev_work is not None:
                if str(t.get("task_id", "")).strip() == DEBUG_TASK_ID:
                    logging.info(
                        "DEBUG[task=%s] start_date_req を非稼働日補正: %s -> %s",
                        DEBUG_TASK_ID,
                        req_d,
                        prev_work,
                    )
                t["start_date_req"] = prev_work
    conflict_rows = collect_planning_conflicts_by_excel_row(tasks_df, ai_task_by_tid)
    _try_write_plan_input_global_parse_and_conflicts_one_save(
        global_priority_override,
        data_extract_dt_str,
        len(tasks_df),
        conflict_rows,
    )

    if not task_queue:
        logging.warning(
            f"有効なタスクはありません。「{PLAN_INPUT_SHEET_NAME}」の「依頼NO」「工程名」「{TASK_COL_QTY}」"
            "または完了区分・実出来高残作により残量は無い行のみの可能性はありした。"
        )

    # 配台試行順: シート列は权っていれみしれを採用。欠損時は §B 帯・紝期・need 列順でソートし EC 隣接後に 1..n
    _apply_dispatch_trial_order_for_generate_plan(
        task_queue, req_map, need_rules, need_combo_col_index
    )
    if DEBUG_TASK_ID:
        dbg_items = [t for t in task_queue if str(t.get("task_id", "")).strip() == DEBUG_TASK_ID]
        if dbg_items:
            t0 = dbg_items[0]
            logging.info(
                "DEBUG[task=%s] queue基準: start_date_req=%s due_basis=%s answer_due=%s specified_due=%s specified_due_ov=%s due_source=%s priority=%s in_progress=%s remark=%s",
                DEBUG_TASK_ID,
                t0.get("start_date_req"),
                t0.get("due_basis_date"),
                t0.get("answer_due_date"),
                t0.get("specified_due_date"),
                t0.get("specified_due_override"),
                t0.get("due_source"),
                t0.get("priority"),
                t0.get("in_progress"),
                t0.get("has_special_remark"),
            )
        else:
            logging.info("DEBUG[task=%s] task_queueに存在しません（完了/残量0/依頼NO厳密一致の可能性）。", DEBUG_TASK_ID)
    timeline_events = []

    # ---------------------------------------------------------
    # 日毎のスケジューリングループ
    # STAGE2_EXTEND_ATTENDANCE_CALENDAR は True のときのみ」残タスクはあれみ勤怠を日付複製で拡張。
    # STAGE2_RETRY_SHIFT_DUE_ON_PARTIAL_REMAINING は True のときのみ: 紝期基準を靎ねでも残はある依頼についで
    # due_basis +1・当該依頼の割当戻し・先頭から再実行。坄再試行剝に勤怠拡張分はマスタ日付へ巻し戻れ。
    # 既定 False のため、通常は 1 パス（カレンダー通し 1 回）のみ。
    # ---------------------------------------------------------
    _master_attendance_date_set = frozenset(attendance_data.keys())
    _master_plan_dates_template = list(sorted_dates)
    _calendar_last_plan_day = _master_plan_dates_template[-1]

    for t in task_queue:
        t["remaining_units"] = float(t.get("initial_remaining_units") or 0)
        t["assigned_history"].clear()
    timeline_events.clear()

    _dispatch_interval_mirror: DispatchIntervalMirror | None = None
    if DISPATCH_INTERVAL_MIRROR_ENFORCE:
        _dispatch_interval_mirror = DispatchIntervalMirror()
        logging.info(
            "DISPATCH_INTERVAL_MIRROR_ENFORCE: 設備・人の占有を区間ミラーで追跡しした"
            "（無効化は 設定_環境変数 等で DISPATCH_INTERVAL_MIRROR_ENFORCE=0）。"
        )

    if STAGE2_SERIAL_DISPATCH_BY_TASK_ID:
        logging.info(
            "依頼NO直列配台: 有効（STAGE2_SERIAL_DISPATCH_BY_TASK_ID）。"
            " 坄日はアクティブな依頼NOの行の値は候補のため、当該依頼は詰まると他依頼は一切進みません。"
        )
    else:
        logging.info(
            "依頼NO直列配台: 無効。start_date を満たれ全行は当日候補になり」配台試行順・設備ルールで順庝付けしした。"
        )

    _due_shift_retry_count_by_request: dict[str, int] = {}
    _due_shift_exhausted_requests: set[str] = set()
    _due_shift_cap_warned_tids: set[str] = set()
    _outer_retry_round = 0
    while True:
        _dispatch_trace_begin_outer_round(_outer_retry_round)
        _need_headcount_logged_orders: set = set()
        if _outer_retry_round > 0:
            _purge_attendance_days_not_in_set(
                attendance_data, _master_attendance_date_set
            )
            sorted_dates[:] = list(_master_plan_dates_template)

        for t in task_queue:
            t.pop("_partial_retry_calendar_blocked", None)

        if STAGE2_SERIAL_DISPATCH_BY_TASK_ID:
            _serial_order_tids = _serial_dispatch_order_task_ids(task_queue)
        else:
            _serial_order_tids = []

        _plan_day_iter = (
            _iter_plan_dates_extending(sorted_dates, attendance_data, task_queue)
            if STAGE2_EXTEND_ATTENDANCE_CALENDAR
            else sorted_dates
        )
        _full_calendar_without_deadline_restart = True
        for current_date in _plan_day_iter:
            daily_status = attendance_data[current_date]
            # 設備ととの空し時刻（同一設備の坌時並行割当を防止）
            machine_avail_dt = {}
            
            avail_dt = {}
            for m in members:
                if m not in daily_status:
                    continue
                st = daily_status[m]
                if st.get("eligible_for_assignment", st.get("is_working", False)):
                    avail_dt[m] = st["start_dt"]

            _machine_day_start = datetime.combine(current_date, DEFAULT_START_TIME)
            _machine_calendar_plan_end = _machine_calendar_planning_window_end_dt(
                current_date, daily_status, members
            )
            if avail_dt:
                _seed_avail_from_timeline_for_date(
                    timeline_events,
                    current_date,
                    machine_avail_dt,
                    avail_dt,
                    _machine_day_start,
                )
                _apply_machine_calendar_floor_for_date(
                    current_date,
                    machine_avail_dt,
                    equipment_list,
                    _machine_day_start,
                    machine_calendar_plan_end=_machine_calendar_plan_end,
                )

            if not avail_dt:
                logging.info("DEBUG[day=%s] 稼働メンバー0のため、割付スキップ", current_date)
                continue
    
            tasks_today = [t for t in task_queue if t['remaining_units'] > 0 and t['start_date_req'] <= current_date]
            if STAGE2_SERIAL_DISPATCH_BY_TASK_ID and _serial_order_tids:
                _tasks_today_before_serial = len(tasks_today)
                _active_serial_tid = None
                for _tid in _serial_order_tids:
                    if any(
                        float(x.get("remaining_units") or 0) > 1e-12
                        for x in task_queue
                        if str(x.get("task_id", "") or "").strip() == _tid
                    ):
                        _active_serial_tid = _tid
                        break
                if _active_serial_tid is not None:
                    tasks_today = [
                        t
                        for t in tasks_today
                        if str(t.get("task_id", "") or "").strip() == _active_serial_tid
                    ]
                _serial_pos = (
                    _serial_order_tids.index(_active_serial_tid) + 1
                    if _active_serial_tid in _serial_order_tids
                    else 0
                )
                _pending_rows = sum(1 for t in task_queue if t["remaining_units"] > 0)
                logging.info(
                    "依頼NO直列配台 day=%s アクティブ依頼NO=%s 直列リスト佝置=%s/%s "
                    "当日候補行数(直列剝)=%s 直列後=%s キュー残行(全日)=%s",
                    current_date,
                    _active_serial_tid if _active_serial_tid is not None else "—",
                    _serial_pos if _serial_pos else "—",
                    len(_serial_order_tids),
                    _tasks_today_before_serial,
                    len(tasks_today),
                    _pending_rows,
                )
            pending_total = sum(1 for t in task_queue if t["remaining_units"] > 0)
            if not tasks_today:
                earliest_wait = min(
                    [t["start_date_req"] for t in task_queue if t["remaining_units"] > 0],
                    default=None,
                )
                logging.info(
                    "DEBUG[day=%s] 割付対象タスク0件 pending_total=%s earliest_start_date_req=%s",
                    current_date,
                    pending_total,
                    earliest_wait,
                )
            elif DEBUG_TASK_ID:
                has_dbg_today = any(str(t.get("task_id", "")).strip() == DEBUG_TASK_ID for t in tasks_today)
                if current_date.isoformat() == "2026-04-03" or has_dbg_today:
                    logging.info(
                        "DEBUG[day=%s] avail_members=%s tasks_today=%s (task=%s 含む=%s)",
                        current_date,
                        len(avail_dt),
                        len(tasks_today),
                        DEBUG_TASK_ID,
                        has_dbg_today,
                    )
            
            _sched_max_passes = max(96, max(1, len(tasks_today)) * 15)
            _sched_pi = 0
            while _sched_pi < _sched_max_passes:
                _sched_pi += 1
                _sched_made_progress = False
                if STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST:
                    _sched_made_progress = _trial_order_first_schedule_pass(
                        current_date,
                        tasks_today,
                        task_queue,
                        daily_status,
                        machine_avail_dt,
                        avail_dt,
                        timeline_events,
                        skills_dict,
                        members,
                        req_map,
                        need_rules,
                        surplus_map,
                        global_priority_override,
                        macro_run_date,
                        macro_now_dt,
                        _need_headcount_logged_orders,
                        team_combo_presets,
                        dispatch_interval_mirror=_dispatch_interval_mirror,
                    )
                if not STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST:
                    _mh_legacy_day = _machine_handoff_state_from_timeline(
                        timeline_events, current_date
                    )
                    machine_handoff_legacy = {
                        "last_tid": dict(_mh_legacy_day["last_tid"]),
                        "last_eq": dict(_mh_legacy_day["last_eq"]),
                        "started_today": set(_mh_legacy_day["started_today"]),
                        "machining_today_occ": set(
                            _mh_legacy_day.get("machining_today_occ") or set()
                        ),
                        "last_machining_dt": dict(
                            _mh_legacy_day.get("last_machining_dt") or {}
                        ),
                        "last_machining_date": dict(
                            _mh_legacy_day.get("last_machining_date") or {}
                        ),
                        "last_lead_op": dict(_mh_legacy_day.get("last_lead_op") or {}),
                        "last_machining_sub": dict(
                            _mh_legacy_day.get("last_machining_sub") or {}
                        ),
                    }
                    _assign_probe_ctx_legacy: dict | None = None
                    _min_dispatch_eff_legacy: int | None = None
                    if STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT:
                        _assign_probe_ctx_legacy = {
                            "avail_dt": avail_dt,
                            "machine_avail_dt": machine_avail_dt,
                            "task_queue": task_queue,
                            "skills_dict": skills_dict,
                            "members": members,
                            "req_map": req_map,
                            "need_rules": need_rules,
                            "surplus_map": surplus_map,
                            "global_priority_override": global_priority_override,
                            "macro_run_date": macro_run_date,
                            "macro_now_dt": macro_now_dt,
                            "machine_handoff": machine_handoff_legacy,
                            "team_combo_presets": team_combo_presets,
                            "dispatch_interval_mirror": _dispatch_interval_mirror,
                        }
                        _pool_legacy = _tasks_in_min_pending_dispatch_pool(
                            task_queue,
                            current_date,
                            daily_status=daily_status,
                            members=members,
                            machine_avail_dt=machine_avail_dt,
                            machine_day_start=_machine_day_start,
                            machine_handoff=machine_handoff_legacy,
                            skills_dict=skills_dict,
                            abolish_all_scheduling_limits=bool(
                                global_priority_override.get(
                                    "abolish_all_scheduling_limits"
                                )
                            ),
                            dispatch_interval_mirror=_dispatch_interval_mirror,
                        )
                        _min_dispatch_eff_legacy = (
                            _effective_min_dispatch_trial_order_from_pool(
                                _pool_legacy,
                                current_date,
                                daily_status,
                                _assign_probe_ctx_legacy,
                            )
                        )
                    for task in sorted(
                        [t for t in tasks_today if float(t.get("remaining_units") or 0) > 1e-12],
                        key=lambda t: _day_schedule_task_sort_key(
                            t, task_queue, need_combo_col_index
                        ),
                    ):
                        if _task_blocked_by_same_request_dependency(task, task_queue):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] スキップ: 同一依頼NOの先行工程待う day=%s machine=%s rem=%.4f",
                                    task.get("task_id"),
                                    current_date,
                                    task.get("machine"),
                                    float(task.get("remaining_units") or 0),
                                )
                            continue
                        if (
                            task.get("roll_pipeline_inspection")
                            or task.get("roll_pipeline_rewind")
                        ) and (
                            _roll_pipeline_inspection_assign_room(
                                task_queue, str(task.get("task_id", "")).strip()
                            )
                            <= 1e-12
                        ):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _tid_tr = str(task.get("task_id", "") or "").strip()
                                _ec_d = _pipeline_ec_roll_done_units(task_queue, _tid_tr)
                                _in_d = _pipeline_b2_follower_roll_done_units(
                                    task_queue, _tid_tr
                                )
                                _log_dispatch_trace_schedule(
                                    _tid_tr,
                                    "[配台トレース task=%s] スキップ: §B-2/§B-3 後続ロール枠ゼロ day=%s machine=%s "
                                    "ec累計完了R=%.4f 後続累計完了R=%.4f rem_follower=%.4f",
                                    _tid_tr,
                                    current_date,
                                    task.get("machine"),
                                    _ec_d,
                                    _in_d,
                                    float(task.get("remaining_units") or 0),
                                )
                            continue
                        machine = task["machine"]
                        eq_line = str(
                            task.get("equipment_line_key") or machine or ""
                        ).strip() or machine
                        machine_occ_key = _machine_occupancy_key_resolve(task, eq_line)
                        if PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE:
                            _b1_holder = _exclusive_b1_inspection_holder_for_machine(
                                task_queue,
                                machine_occ_key,
                            )
                            if _b1_holder is not None and _b1_holder is not task:
                                if _trace_schedule_task_enabled(task.get("task_id")):
                                    _log_dispatch_trace_schedule(
                                        task.get("task_id"),
                                        "[配台トレース task=%s] スキップ: 同一設備の検査占有中 day=%s "
                                        "占有者依頼NO=%s 占有者試行順=%s",
                                        task.get("task_id"),
                                        current_date,
                                        _b1_holder.get("task_id"),
                                        _b1_holder.get("dispatch_trial_order"),
                                    )
                                continue
                        if DEBUG_TASK_ID and str(task.get("task_id", "")).strip() == DEBUG_TASK_ID:
                            logging.info(
                                "DEBUG[task=%s] day=%s 開始判定: start_date_req=%s remaining_units=%s machine=%s",
                                DEBUG_TASK_ID,
                                current_date,
                                task.get("start_date_req"),
                                task.get("remaining_units"),
                                task.get("machine"),
                            )
                        if task.get("has_done_deadline_override"):
                            logging.info(
                                "DEBUG[完了日指定] 依頼NO=%s 日付=%s start_date_req=%s due_basis=%s 指定納期(上書き)=%s 進杗=%s/%s",
                                task.get("task_id"),
                                current_date,
                                task.get("start_date_req"),
                                task.get("due_basis_date"),
                                task.get("specified_due_override"),
                                task.get("done_qty_reported"),
                                task.get("total_qty_m"),
                            )
                        try:
                            _my_dispatch_ord = int(
                                task.get("dispatch_trial_order") or 10**9
                            )
                        except (TypeError, ValueError):
                            _my_dispatch_ord = 10**9
                        if _task_blocked_by_global_dispatch_trial_order(
                            task,
                            task_queue,
                            current_date,
                            daily_status=daily_status,
                            members=members,
                            machine_avail_dt=machine_avail_dt,
                            machine_day_start=_machine_day_start,
                            machine_handoff=machine_handoff_legacy,
                            skills_dict=skills_dict,
                            abolish_all_scheduling_limits=bool(
                                global_priority_override.get(
                                    "abolish_all_scheduling_limits"
                                )
                            ),
                            dispatch_interval_mirror=_dispatch_interval_mirror,
                            min_dispatch_effective=_min_dispatch_eff_legacy,
                        ):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] スキップ: より尝さい配台試行順に未完了あり "
                                    "day=%s my_order=%s",
                                    task.get("task_id"),
                                    current_date,
                                    _my_dispatch_ord,
                                )
                            continue
                        if _task_fully_machine_calendar_blocked_on_date(
                            task, current_date, daily_status, members
                        ):
                            continue
                        if _task_no_machining_window_left_from_avail_floor(
                            task,
                            current_date,
                            daily_status,
                            members,
                            machine_avail_dt,
                            _machine_day_start,
                            machine_handoff=machine_handoff_legacy,
                            skills_dict=skills_dict,
                            abolish_all_scheduling_limits=bool(
                                global_priority_override.get(
                                    "abolish_all_scheduling_limits"
                                )
                            ),
                            dispatch_interval_mirror=_dispatch_interval_mirror,
                        ):
                            continue
                        if _equipment_line_lower_dispatch_trial_still_pending(
                            task_queue,
                            machine_occ_key,
                            _my_dispatch_ord,
                            current_date,
                            daily_status=daily_status,
                            members=members,
                            machine_avail_dt=machine_avail_dt,
                            machine_day_start=_machine_day_start,
                            machine_handoff=machine_handoff_legacy,
                            skills_dict=skills_dict,
                            abolish_all_scheduling_limits=bool(
                                global_priority_override.get(
                                    "abolish_all_scheduling_limits"
                                )
                            ),
                            dispatch_interval_mirror=_dispatch_interval_mirror,
                        ):
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] スキップ: 同一設備で配台試行順は先の行は未完了 "
                                    "day=%s eq_line=%s my_order=%s",
                                    task.get("task_id"),
                                    current_date,
                                    eq_line,
                                    _my_dispatch_ord,
                                )
                            continue
                        machine_name = str(task.get("machine_name", "") or "").strip()
                        machine_proc = str(machine or "").strip()
                        plan_ro = _plan_sheet_required_op_optional(task)
                        need_src_line = ""
                        if TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY:
                            req_num, need_src_line = resolve_need_required_op_explain(
                                machine,
                                machine_name,
                                task["task_id"],
                                req_map,
                                need_rules,
                            )
                            if plan_ro is not None and plan_ro != req_num:
                                need_src_line = (
                                    (need_src_line + "；") if need_src_line else ""
                                )
                                need_src_line += (
                                    f"計画シート必須人数{plan_ro}は未使用（need基準={req_num}）"
                                )
                        else:
                            if plan_ro is not None:
                                req_num = plan_ro
                                need_src_line = f"計画シート「必須OP(上書)」={req_num}"
                            else:
                                req_num, need_src_line = resolve_need_required_op_explain(
                                    machine,
                                    machine_name,
                                    task["task_id"],
                                    req_map,
                                    need_rules,
                                )
                        if global_priority_override.get("ignore_need_minimum"):
                            req_num = 1
                            need_src_line = (
                                (need_src_line + " → ")
                                if need_src_line
                                else ""
                            ) + "メイン上書ignore_need_minimumでreq=1"
    
                        # メンバー×設備スキル（parse_op_as_skill_cell: 尝さい優先度ろど先にフォーム候補へ採用）
                        # skills 読込時に「機械名」独立キーへエイリアスれるため、工程名+機械名は両方ある行では
                        # 複坈キー「工程名+機械名」のみを見る（別工程の坌坝機械の OP は浝れ込まないよごにれる）。
                        skill_meta_cache = {}
                        _gpo = global_priority_override
    
                        def skill_role_priority(mem):
                            if _gpo.get("ignore_skill_requirements"):
                                return ("OP", 100)
                            if mem not in skill_meta_cache:
                                srow = skills_dict.get(mem, {})
                                v = ""
                                if machine_proc and machine_name:
                                    v = srow.get(f"{machine_proc}+{machine_name}", "")
                                elif machine_name:
                                    v = srow.get(machine_name, "")
                                elif machine_proc:
                                    v = srow.get(machine_proc, "")
                                skill_meta_cache[mem] = parse_op_as_skill_cell(v)
                            return skill_meta_cache[mem]
    
                        capable_members = [m for m in avail_dt.keys() if skill_role_priority(m)[0] in ("OP", "AS")]
                        capable_members.sort(key=lambda mm: (skill_role_priority(mm)[1], mm))
                        capable_members = _filter_capable_members_b2_disjoint_teams(
                            task, task_queue, capable_members
                        )
                        if task.get("has_done_deadline_override"):
                            machine_free_dbg = _machine_effective_floor_for_assign(
                                machine_occ_key,
                                str(task.get("task_id") or "").strip(),
                                eq_line,
                                str(task.get("machine_name") or "").strip(),
                                machine_avail_dt,
                                machine_handoff_legacy,
                                _machine_day_start,
                                bool(_gpo.get("abolish_all_scheduling_limits")),
                                current_date=current_date,
                                daily_status=daily_status,
                                skills_dict=skills_dict,
                                machine_proc=machine_proc,
                            )
                            logging.info(
                                "DEBUG[完了日指定] 依頼NO=%s 設備=%s req_num=%s capable_members=%s machine_free=%s",
                                task.get("task_id"),
                                eq_line,
                                req_num,
                                len(capable_members),
                                machine_free_dbg,
                            )
    
                        pref_raw = str(task.get("preferred_operator_raw") or "").strip()
                        op_today = [m for m in capable_members if skill_role_priority(m)[0] == "OP"]
                        pref_mem = (
                            _resolve_preferred_op_to_member(pref_raw, op_today, members)
                            if pref_raw
                            else None
                        )
                        if pref_raw and pref_mem is None and op_today:
                            logging.info(
                                "担当OP指定: 当日のOP候補に一致せう制約なし task=%s raw=%r",
                                task.get("task_id"),
                                pref_raw,
                            )

                        _gdp_must, _gdp_warns = _active_global_day_process_must_include(
                            _gpo,
                            task,
                            current_date,
                            capable_members,
                            members,
                        )
                        for _gw in _gdp_warns:
                            logging.warning(_gw)
                        fixed_team_anchor = _merge_global_day_process_and_pref_anchor(
                            _gdp_must, pref_mem, capable_members
                        )
                        if _gdp_must:
                            logging.info(
                                "メイングローバル(日付×工程): task=%s date=%s 工程=%r フォーム必須=%s",
                                task.get("task_id"),
                                current_date,
                                machine,
                                ",".join(_gdp_must),
                            )
                        if fixed_team_anchor:
                            _nfix = len(fixed_team_anchor)
                            if _nfix > req_num:
                                need_src_line = (
                                    (need_src_line + " → ")
                                    if need_src_line
                                    else ""
                                )
                                need_src_line += (
                                    f"グローバル(日付×工程)指定で最低{_nfix}人"
                                )
                            req_num = max(req_num, _nfix)
    
                        extra_max_sheet, extra_src_line = resolve_need_surplus_extra_max_explain(
                            machine,
                            machine_name,
                            task["task_id"],
                            surplus_map,
                            need_rules,
                        )
                        if TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW:
                            extra_max_sheet = 0
                            extra_src_line = (
                                (extra_src_line + " → ")
                                if extra_src_line
                                else ""
                            ) + "TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROWで0"
                        extra_max = (
                            extra_max_sheet
                            if TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                            else 0
                        )
                        if (
                            extra_max_sheet > 0
                            and not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                            and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
                        ):
                            extra_src_line = (
                                (extra_src_line + " → ")
                                if extra_src_line
                                else ""
                            ) + "メインは基本人数のみ（余力枠は全配台後に未割当×スキルで追記）"
                        max_team_size = min(req_num + extra_max, len(capable_members))
                        if max_team_size < req_num:
                            max_team_size = req_num
                        rq_base = max(1, int(req_num))
    
                        _dto_head = task.get("dispatch_trial_order")
                        if (
                            _dto_head is not None
                            and _dto_head not in _need_headcount_logged_orders
                        ):
                            _need_headcount_logged_orders.add(_dto_head)
                            logging.info(
                                "need人数(配台試行順初回) order=%s task=%s 工程/機械=%s/%s "
                                "req_num=%s [%s] extra_max=%s [%s] max_team候補=%s capable=%s人",
                                _dto_head,
                                task["task_id"],
                                machine,
                                machine_name,
                                req_num,
                                need_src_line,
                                extra_max,
                                extra_src_line,
                                max_team_size,
                                len(capable_members),
                            )
    
                        trace_assign = bool(TRACE_TEAM_ASSIGN_TASK_ID) and (
                            str(task.get("task_id", "")).strip() == TRACE_TEAM_ASSIGN_TASK_ID
                        )
                        if trace_assign:
                            logging.info(
                                "TRACE配台[%s] %s 工程/機械=%s / %s req_num=%s extra_max=%s → max_team=%s "
                                "capable(n=%s)=%s ignore_need1=%s ignore_skill=%s abolish=%s 担当OP指定=%r→%s",
                                task["task_id"],
                                current_date,
                                machine,
                                machine_name,
                                req_num,
                                extra_max,
                                max_team_size,
                                len(capable_members),
                                capable_members,
                                global_priority_override.get("ignore_need_minimum"),
                                global_priority_override.get("ignore_skill_requirements"),
                                global_priority_override.get("abolish_all_scheduling_limits"),
                                pref_raw,
                                pref_mem,
                            )
    
                        team_candidates: list[dict] = []
                        combo_key = (
                            f"{machine_proc}+{machine_name}"
                            if machine_proc and machine_name
                            else ""
                        )
                        preset_rows = (
                            (team_combo_presets or {}).get(combo_key)
                            if (team_combo_presets and combo_key)
                            else None
                        )
                        (
                            _mach_floor_legacy,
                            _co_segs_legacy,
                            _abort_legacy,
                        ) = _resolve_machine_changeover_floor_segments(
                            abolish_all_scheduling_limits=bool(
                                _gpo.get("abolish_all_scheduling_limits")
                            ),
                            machine_occ_key=machine_occ_key,
                            task_id=str(task.get("task_id") or "").strip(),
                            eq_line=eq_line,
                            machine_name=str(task.get("machine_name") or "").strip(),
                            machine_proc=machine_proc,
                            machine_avail_dt=machine_avail_dt,
                            machine_day_floor=_machine_day_start,
                            current_date=current_date,
                            machine_handoff=machine_handoff_legacy,
                            daily_status=daily_status,
                            skills_dict=skills_dict,
                            dispatch_interval_mirror=_dispatch_interval_mirror,
                            timeline_events=timeline_events,
                            task_queue=task_queue,
                            avail_dt=avail_dt,
                        )
                        if _abort_legacy:
                            continue
                        # プリセットは成立分をまとめて候補に載せ」下の組み合わせ探索とまとめで最良を決める。
                        if preset_rows:
                            for _prio, sheet_rs, preset_team, combo_row_id in preset_rows:
                                pteam = tuple(preset_team)
                                bounds = _combo_preset_team_size_bounds(
                                    pteam, sheet_rs, max_team_size
                                )
                                if bounds is None:
                                    continue
                                if fixed_team_anchor and not all(
                                    m in pteam for m in fixed_team_anchor
                                ):
                                    continue
                                if pref_mem is not None and pref_mem not in pteam:
                                    continue
                                if not all(m in capable_members for m in pteam):
                                    continue
                                _append_legacy_dispatch_candidate_for_team(
                                    task,
                                    pteam,
                                    avail_dt,
                                    machine_avail_dt,
                                    daily_status,
                                    current_date,
                                    macro_run_date,
                                    macro_now_dt,
                                    skill_role_priority,
                                    eq_line,
                                    rq_base,
                                    extra_max,
                                    global_priority_override,
                                    team_candidates,
                                    combo_sheet_row_id=combo_row_id,
                                    combo_preset_team=pteam,
                                    dispatch_interval_mirror=_dispatch_interval_mirror,
                                    machine_handoff=machine_handoff_legacy,
                                    machine_day_floor=_machine_day_start,
                                    machine_floor_cached=_mach_floor_legacy,
                                )
    
                        for tsize in range(req_num, max_team_size + 1):
                            if fixed_team_anchor:
                                _ft = list(fixed_team_anchor)
                                others = [m for m in capable_members if m not in _ft]
                                need_extra = tsize - len(_ft)
                                if need_extra < 0:
                                    teams_iter = []
                                elif need_extra == 0:
                                    teams_iter = [tuple(_ft)]
                                elif len(others) >= need_extra:
                                    teams_iter = [
                                        tuple(_ft + list(rest))
                                        for rest in itertools.combinations(
                                            others, need_extra
                                        )
                                    ]
                                else:
                                    teams_iter = []
                            elif (
                                pref_mem is not None
                                and pref_mem in capable_members
                                and skill_role_priority(pref_mem)[0] == "OP"
                            ):
                                others = [m for m in capable_members if m != pref_mem]
                                if tsize == 1:
                                    teams_iter = [(pref_mem,)]
                                elif len(others) >= tsize - 1:
                                    teams_iter = [
                                        tuple([pref_mem] + list(rest))
                                        for rest in itertools.combinations(others, tsize - 1)
                                    ]
                                else:
                                    logging.info(
                                        "担当OP指定: フォーム人数を満たせないため、指定を無視 task=%s size=%s raw=%r",
                                        task.get("task_id"),
                                        tsize,
                                        pref_raw,
                                    )
                                    teams_iter = itertools.combinations(capable_members, tsize)
                            else:
                                teams_iter = itertools.combinations(capable_members, tsize)
    
                            for team in teams_iter:
                                op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
                                if not op_list:
                                    continue
    
                                team_start = max(avail_dt[m] for m in team)
                                if not _gpo.get("abolish_all_scheduling_limits"):
                                    # 同一設備は1時点で1タスクのみ（設備空し＋日次始業/依頼切替の準備・後始末）
                                    machine_free_dt = _mach_floor_legacy
                                    if team_start < machine_free_dt:
                                        team_start = machine_free_dt
                                    # 原板投入日と同日の開始は 13:00 以降（試行順優先フローと一致）
                                    if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
                                        min_start_dt = datetime.combine(
                                            current_date, task["same_day_raw_start_limit"]
                                        )
                                        if team_start < min_start_dt:
                                            team_start = min_start_dt
                                    if current_date == task["start_date_req"] and task.get("earliest_start_time"):
                                        min_user_t = datetime.combine(
                                            current_date, task["earliest_start_time"]
                                        )
                                        if team_start < min_user_t:
                                            team_start = min_user_t
                                    # 当日は「マクロ実行した時刻」より剝に開始でしない
                                    if current_date == macro_run_date and team_start < macro_now_dt:
                                        team_start = macro_now_dt
                                team_end_limit = min(daily_status[m]['end_dt'] for m in team)
    
                                if team_start >= team_end_limit:
                                    continue
    
                                team_breaks = []
                                for m in team:
                                    team_breaks.extend(daily_status[m]['breaks_dt'])
                                team_breaks = merge_time_intervals(team_breaks)

                                avg_eff = sum(daily_status[m]['efficiency'] for m in team) / len(team)
                                if avg_eff <= 0:
                                    avg_eff = 0.01
                                t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
                                if t_eff <= 0:
                                    t_eff = 1.0
                                eff_time_per_unit = (
                                    task["base_time_per_unit"]
                                    / avg_eff
                                    / t_eff
                                    * _surplus_team_time_factor(rq_base, len(team), extra_max)
                                )
                                _defer_min_contig = max(1, int(math.ceil(float(eff_time_per_unit))))
                                _eod_cont_exempt_il = (
                                    _eod_same_request_continuation_exempt(
                                        machine_occ_key, task, machine_handoff_legacy
                                    )
                                )

                                def _refloor_legacy_inline(ts):
                                    ts = max(ts, max(avail_dt[m] for m in team))
                                    if not _gpo.get("abolish_all_scheduling_limits"):
                                        _mfd = _mach_floor_legacy
                                        if ts < _mfd:
                                            ts = _mfd
                                        if task.get(
                                            "same_day_raw_start_limit"
                                        ) and current_date == task["start_date_req"]:
                                            _msd = datetime.combine(
                                                current_date,
                                                task["same_day_raw_start_limit"],
                                            )
                                            if ts < _msd:
                                                ts = _msd
                                        if current_date == task[
                                            "start_date_req"
                                        ] and task.get("earliest_start_time"):
                                            _mut = datetime.combine(
                                                current_date,
                                                task["earliest_start_time"],
                                            )
                                            if ts < _mut:
                                                ts = _mut
                                        if (
                                            current_date == macro_run_date
                                            and ts < macro_now_dt
                                        ):
                                            ts = macro_now_dt
                                    return ts
    
                                _ts_adj = _defer_team_start_past_prebreak_and_end_of_day(
                                    task,
                                    tuple(team),
                                    team_start,
                                    team_end_limit,
                                    team_breaks,
                                    _refloor_legacy_inline,
                                    min_contiguous_work_mins=_defer_min_contig,
                                    eod_same_request_continuation_exempt=_eod_cont_exempt_il,
                                )
                                if _ts_adj is None:
                                    continue
                                team_start = _ts_adj
                                if team_start >= team_end_limit:
                                    continue
    
                                _, avail_mins, _ = calculate_end_time(team_start, 9999, team_breaks, team_end_limit)
    
                                units_can_do = int(avail_mins / eff_time_per_unit)
                                if units_can_do == 0:
                                    continue
    
                                units_today = min(units_can_do, math.ceil(task['remaining_units']))
                                if _eod_reject_capacity_units_below_threshold(
                                    units_today,
                                    team_start,
                                    team_end_limit,
                                    eod_same_request_continuation_exempt=_eod_cont_exempt_il,
                                ):
                                    continue
                                work_mins_needed = int(units_today * eff_time_per_unit)
                                if (
                                    _contiguous_work_minutes_until_next_break_or_limit(
                                        team_start, team_breaks, team_end_limit
                                    )
                                    < work_mins_needed
                                ):
                                    continue
                                actual_end_dt, _, _ = calculate_end_time(team_start, work_mins_needed, team_breaks, team_end_limit)
    
                                team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
                                if (
                                    _dispatch_interval_mirror is not None
                                    and _dispatch_interval_mirror.would_block_roll(
                                        machine_occ_key,
                                        team,
                                        team_start,
                                        actual_end_dt,
                                    )
                                ):
                                    continue
                                team_candidates.append(
                                    {
                                        "team": team,
                                        "team_start": team_start,
                                        "actual_end_dt": actual_end_dt,
                                        "units_today": units_today,
                                        "team_breaks": team_breaks,
                                        "avg_eff": avg_eff,
                                        "prio_sum": team_prio_sum,
                                        "op_list": op_list,
                                        "eff_time_per_unit": eff_time_per_unit,
                                        "combo_sheet_row_id": None,
                                        "combo_preset_team": None,
                                    }
                                )
    
                        best_team = None
                        best_info = {
                            "start_dt": datetime.max,
                            "units_today": 0,
                            "prio_sum": 10**9,
                        }
                        t_min = (
                            min(c["team_start"] for c in team_candidates)
                            if team_candidates
                            else None
                        )
    
                        def _team_cand_key(c):
                            return _team_assignment_sort_tuple(
                                c["team"],
                                c["team_start"],
                                c["units_today"],
                                c["prio_sum"],
                                t_min,
                            )
    
                        if team_candidates:
                            best_c = min(team_candidates, key=_team_cand_key)
                            if best_c.get("combo_sheet_row_id") is None and preset_rows:
                                _lcid_l = _lookup_combo_sheet_row_id_for_preset_team(
                                    preset_rows, tuple(best_c["team"])
                                )
                                if _lcid_l is not None:
                                    best_c = {
                                        **best_c,
                                        "combo_sheet_row_id": _lcid_l,
                                        "combo_preset_team": tuple(best_c["team"]),
                                    }
                            if pref_mem and pref_mem in best_c["op_list"]:
                                lead_op = pref_mem
                            else:
                                lead_op = min(
                                    best_c["op_list"],
                                    key=lambda mm: (skill_role_priority(mm)[1], mm),
                                )
                            best_team = best_c["team"]
                            best_info = {
                                "start_dt": best_c["team_start"],
                                "end_dt": best_c["actual_end_dt"],
                                "op": lead_op,
                                "units_today": best_c["units_today"],
                                "breaks": best_c["team_breaks"],
                                "eff": best_c["avg_eff"],
                                "prio_sum": best_c["prio_sum"],
                            }
    
                        if trace_assign:
                            _tk = _team_assign_trace_tuple_label()
                            tid = task["task_id"]
                            for tsize in range(req_num, max_team_size + 1):
                                sub = [c for c in team_candidates if len(c["team"]) == tsize]
                                if not sub:
                                    logging.info(
                                        "TRACE配台[%s] %s tsize=%s → この人数で成立れるフォームなし",
                                        tid,
                                        current_date,
                                        tsize,
                                    )
                                else:
                                    sm = min(sub, key=_team_cand_key)
                                    logging.info(
                                        "TRACE配台[%s] %s tsize=%s 人数内最良: members=%s "
                                        "start=%s units_today=%s prio_sum=%s eff_t/unit=%.6f "
                                        "比較ルール=%s ※全日最早開始=%s を基準に辞書式で尝さい方は採用",
                                        tid,
                                        current_date,
                                        tsize,
                                        sm["team"],
                                        sm["team_start"],
                                        sm["units_today"],
                                        sm["prio_sum"],
                                        sm["eff_time_per_unit"],
                                        _tk,
                                        t_min.isoformat(sep=" ") if t_min else "—",
                                    )
    
                        if trace_assign and best_team is not None:
                            logging.info(
                                "TRACE配台[%s] %s ★採用 n=%s members=%s start=%s units_today=%s prio_sum=%s",
                                task["task_id"],
                                current_date,
                                len(best_team),
                                best_team,
                                best_info["start_dt"],
                                best_info["units_today"],
                                best_info["prio_sum"],
                            )
                            if len(best_team) == 1 and max_team_size > req_num:
                                if TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF:
                                    logging.info(
                                        "TRACE配台[%s] %s 1人採用（TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF）: "
                                        "より大しい人数で有効なフォームなし（OP丝足・0短縮・開始>=終了等）。",
                                        task["task_id"],
                                        current_date,
                                    )
                                else:
                                    logging.info(
                                        "TRACE配台[%s] %s 1人採用: 人数を増やれと開始は遅れ」"
                                        "スラック外では開始優先で1人は選べた可能性。"
                                        "TEAM_ASSIGN_START_SLACK_WAIT_MINUTES=%s」または従来の人数最優先は環境変数参照。",
                                        task["task_id"],
                                        current_date,
                                        TEAM_ASSIGN_START_SLACK_WAIT_MINUTES,
                                    )
    
                        if best_team:
                            if len(best_team) > req_num:
                                logging.info(
                                    "配台採用人数>req_num task=%s day=%s order=%s 工程/機械=%s/%s "
                                    "採用=%s人 req_num=%s extra_max=%s max_team=%s [%s] [%s]",
                                    task["task_id"],
                                    current_date,
                                    task.get("dispatch_trial_order"),
                                    machine,
                                    machine_name,
                                    len(best_team),
                                    req_num,
                                    extra_max,
                                    max_team_size,
                                    need_src_line,
                                    extra_src_line,
                                )
                            sub_members = [m for m in best_team if m != best_info["op"]]
                            done_units = best_info["units_today"]
                            if task.get("roll_pipeline_inspection") or task.get(
                                "roll_pipeline_rewind"
                            ):
                                _rp_room = _roll_pipeline_inspection_assign_room(
                                    task_queue, str(task.get("task_id", "")).strip()
                                )
                                done_units = min(
                                    int(done_units),
                                    int(min(_rp_room, math.ceil(task["remaining_units"]))),
                                )
                            else:
                                done_units = int(done_units)
                            if done_units <= 0:
                                if _trace_schedule_task_enabled(task.get("task_id")):
                                    _rp_log = None
                                    if task.get(
                                        "roll_pipeline_inspection"
                                    ) or task.get("roll_pipeline_rewind"):
                                        _rp_log = _roll_pipeline_inspection_assign_room(
                                            task_queue,
                                            str(task.get("task_id", "") or "").strip(),
                                        )
                                    _log_dispatch_trace_schedule(
                                        task.get("task_id"),
                                        "[配台トレース task=%s] スキップ: フォーム採用後の実効ユニット0 "
                                        "day=%s machine=%s best_units_today=%s rp_room=%s rem=%.4f",
                                        task.get("task_id"),
                                        current_date,
                                        machine,
                                        best_info.get("units_today"),
                                        _rp_log,
                                        float(task.get("remaining_units") or 0),
                                    )
                                continue
                            if done_units < best_info["units_today"]:
                                team_end_limit = min(
                                    daily_status[m]["end_dt"] for m in best_team
                                )
                                _teff = parse_float_safe(task.get("task_eff_factor"), 1.0)
                                if _teff <= 0:
                                    _teff = 1.0
                                _eff_t = (
                                    task["base_time_per_unit"]
                                    / best_info["eff"]
                                    / _teff
                                    * _surplus_team_time_factor(rq_base, len(best_team), extra_max)
                                )
                                _wm = int(done_units * _eff_t)
                                _end_dt, _, _ = calculate_end_time(
                                    best_info["start_dt"],
                                    _wm,
                                    best_info["breaks"],
                                    team_end_limit,
                                )
                                best_info = dict(best_info)
                                best_info["end_dt"] = _end_dt
                                best_info["units_today"] = done_units
    
                            total_u = math.ceil(task['total_qty_m'] / task['unit_m']) if task['unit_m'] else 0
                            rem_u_before = math.ceil(task['remaining_units'])
                            already_done = total_u - rem_u_before
                            
                            # 「マクロ実行時点」の完了率（予定の進杗ではなく」実加工数ベース）
                            try:
                                tot_qty = parse_float_safe(task.get('total_qty_m'), 0.0)
                                done_qty = parse_float_safe(task.get('done_qty_reported'), 0.0)
                                if tot_qty > 0:
                                    pct_macro = max(0, min(100, int(round((done_qty / tot_qty) * 100))))
                                else:
                                    pct_macro = 0
                            except Exception:
                                pct_macro = 0
                            
                            _te_disp = parse_float_safe(task.get("task_eff_factor"), 1.0)
                            if _te_disp <= 0:
                                _te_disp = 1.0
                            _legacy_mach_sub = ", ".join(
                                str(s).strip()
                                for s in sub_members
                                if s and str(s).strip()
                            )
                            _co_append_l = list(_co_segs_legacy or [])
                            _append_changeover_segments_to_timeline(
                                timeline_events,
                                _dispatch_interval_mirror,
                                avail_dt,
                                daily_status,
                                current_date=current_date,
                                task_id=str(task.get("task_id") or ""),
                                machine_occ_key=machine_occ_key,
                                segments=_co_append_l,
                                machining_lead_op=str(
                                    best_info.get("op") or ""
                                ).strip()
                                or None,
                                machining_sub_str=_legacy_mach_sub or None,
                                machine_handoff=machine_handoff_legacy,
                            )
                            timeline_events.append({
                                "date": current_date, "task_id": task['task_id'], "machine": eq_line,
                                "machine_occupancy_key": machine_occ_key,
                                "op": best_info["op"], "sub": ", ".join(sub_members),
                                "start_dt": best_info["start_dt"], "end_dt": best_info["end_dt"],
                                "breaks": best_info["breaks"], "units_done": done_units,
                                "already_done_units": already_done,
                                "total_units": total_u,
                                "pct_macro": pct_macro,
                                "eff_time_per_unit": task["base_time_per_unit"]
                                / best_info["eff"]
                                / _te_disp
                                * _surplus_team_time_factor(rq_base, len(best_team), extra_max),
                                "unit_m": task['unit_m'],
                                "event_kind": TIMELINE_EVENT_MACHINING,
                            })
                            if _dispatch_interval_mirror is not None:
                                _dispatch_interval_mirror.register_from_event(
                                    timeline_events[-1]
                                )
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _rp_tr = None
                                if task.get("roll_pipeline_inspection") or task.get(
                                    "roll_pipeline_rewind"
                                ):
                                    _rp_tr = _roll_pipeline_inspection_assign_room(
                                        task_queue,
                                        str(task.get("task_id", "") or "").strip(),
                                    )
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] タイムライン追記 chunk day=%s machine=%s "
                                    "done_units=%s already_done=%s total_u=%s rem_after=%.4f "
                                    "start=%s end=%s eff_t/unit=%.4f rp_room(当時)=%s",
                                    task.get("task_id"),
                                    current_date,
                                    eq_line,
                                    done_units,
                                    already_done,
                                    total_u,
                                    float(task.get("remaining_units") or 0)
                                    - float(done_units),
                                    best_info["start_dt"],
                                    best_info["end_dt"],
                                    float(
                                        task["base_time_per_unit"]
                                        / best_info["eff"]
                                        / _te_disp
                                        * _surplus_team_time_factor(
                                            rq_base, len(best_team), extra_max
                                        )
                                    ),
                                    _rp_tr,
                                )

                            task["remaining_units"] = max(
                                0.0,
                                float(task.get("remaining_units") or 0)
                                - float(done_units),
                            )
                            op_main = (best_info.get("op") or "").strip()
                            subs_part = ",".join(
                                s.strip() for s in sub_members if s and str(s).strip()
                            )
                            team_s = f"{op_main}, {subs_part}" if subs_part else op_main
                            need_surplus_assigned = (
                                TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
                                and extra_max > 0
                                and len(best_team) > req_num
                            )
                            _lo = (best_info.get("op") or "").strip()
                            _subs_legacy = [
                                str(s).strip()
                                for s in sub_members
                                if s and str(s).strip()
                            ]
                            _names_ord = ([] if not _lo else [_lo]) + _subs_legacy
                            _surplus_names = (
                                _names_ord[int(req_num) :]
                                if need_surplus_assigned
                                and len(_names_ord) > int(req_num)
                                else []
                            )
                            task["assigned_history"].append(
                                {
                                    "date": current_date.strftime("%m/%d"),
                                    "team": team_s,
                                    "done_m": int(done_units * task["unit_m"]),
                                    "start_dt": best_info["start_dt"],
                                    "end_dt": best_info["end_dt"],
                                    "need_surplus_assigned": need_surplus_assigned,
                                    "combo_sheet_row_id": best_c.get(
                                        "combo_sheet_row_id"
                                    ),
                                    "surplus_member_names": _surplus_names,
                                }
                            )

                            for m in best_team:
                                avail_dt[m] = best_info["end_dt"]
                            if not _gpo.get("abolish_all_scheduling_limits"):
                                machine_avail_dt[machine_occ_key] = best_info["end_dt"]
                                _bump_machine_avail_after_roll_for_calendar(
                                    current_date,
                                    machine_occ_key,
                                    machine_avail_dt,
                                    machine_calendar_plan_end=_machine_calendar_plan_end,
                                    machine_day_floor=_machine_day_start,
                                )
                            machine_handoff_legacy["last_tid"][machine_occ_key] = str(
                                task.get("task_id") or ""
                            ).strip()
                            machine_handoff_legacy["last_eq"][machine_occ_key] = eq_line
                            machine_handoff_legacy["started_today"].add(
                                machine_occ_key
                            )
                            machine_handoff_legacy["machining_today_occ"].add(
                                machine_occ_key
                            )
                            machine_handoff_legacy["last_machining_dt"][
                                machine_occ_key
                            ] = best_info["end_dt"]
                            machine_handoff_legacy["last_machining_date"][
                                machine_occ_key
                            ] = current_date
                            machine_handoff_legacy["last_lead_op"][
                                machine_occ_key
                            ] = best_info["op"]
                            machine_handoff_legacy.setdefault("last_machining_sub", {})
                            machine_handoff_legacy["last_machining_sub"][
                                machine_occ_key
                            ] = _legacy_mach_sub
                            if _trace_schedule_task_enabled(task.get("task_id")):
                                _log_dispatch_trace_schedule(
                                    task.get("task_id"),
                                    "[配台トレース task=%s] ロール確定 メイン day=%s machine=%s machine_name=%s "
                                    "start=%s end=%s 採用人数=%s req_num=%s メイン探索extra_max=%s "
                                    "余剰人数適用(メイン)=%s team=%s",
                                    task.get("task_id"),
                                    current_date,
                                    eq_line,
                                    str(machine_name or "").strip(),
                                    best_info["start_dt"],
                                    best_info["end_dt"],
                                    len(best_team),
                                    int(req_num),
                                    int(extra_max),
                                    need_surplus_assigned,
                                    team_s,
                                )
                            _sched_made_progress = True
                        else:
                            if task.get("has_done_deadline_override"):
                                logging.info(
                                    "DEBUG[完了日指定] 依頼NO=%s 日付=%s は割当試行（覝員/設備空し条件でフォーム未成立）。remaining_units=%s",
                                    task.get("task_id"),
                                    current_date,
                                    task.get("remaining_units"),
                                )
    
                if not _sched_made_progress:
                    break

            if TRACE_SCHEDULE_TASK_IDS:
                for _tt in TRACE_SCHEDULE_TASK_IDS:
                    for _t in task_queue:
                        if str(_t.get("task_id", "")).strip() != _tt:
                            continue
                        _rem_tr = float(_t.get("remaining_units") or 0)
                        if _rem_tr <= 1e-9:
                            continue
                        _log_dispatch_trace_schedule(
                            _tt,
                            "[配台トレース task=%s] 日次終了時点の残 day=%s machine=%s "
                            "machine_name=%s rem=%.4f roll_b2_follower=%s 試行順=%s",
                            _tt,
                            current_date,
                            _t.get("machine"),
                            _t.get("machine_name"),
                            _rem_tr,
                            bool(
                                _t.get("roll_pipeline_inspection")
                                or _t.get("roll_pipeline_rewind")
                            ),
                            _t.get("dispatch_trial_order"),
                        )

            if STAGE2_RETRY_SHIFT_DUE_ON_PARTIAL_REMAINING:
                missed_tids = _collect_task_ids_missed_deadline_after_day(
                    task_queue, current_date
                )
                if missed_tids:
                    blocked_tids = set()
                    shift_tid_list = []
                    for _ptid in sorted(missed_tids):
                        _do_shift, _cal_short = _partial_task_id_due_shift_outcome(
                            task_queue, _ptid, _calendar_last_plan_day
                        )
                        if _cal_short:
                            blocked_tids.add(_ptid)
                        if _do_shift:
                            shift_tid_list.append(_ptid)
                    for t in task_queue:
                        _tid = str(t.get("task_id", "") or "").strip()
                        if _tid in blocked_tids:
                            t["_partial_retry_calendar_blocked"] = True
                    if shift_tid_list:
                        allowed_shift_tids = [
                            tid
                            for tid in shift_tid_list
                            if _due_shift_retry_count_by_request.get(tid, 0)
                            < STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS
                        ]
                        for tid in shift_tid_list:
                            if tid not in allowed_shift_tids:
                                _due_shift_exhausted_requests.add(tid)
                        if allowed_shift_tids:
                            _outer_retry_round += 1
                            for tid in allowed_shift_tids:
                                _due_shift_retry_count_by_request[tid] = (
                                    _due_shift_retry_count_by_request.get(tid, 0) + 1
                                )
                            shift_set = set(allowed_shift_tids)
                            for t in task_queue:
                                if str(t.get("task_id", "") or "").strip() in shift_set:
                                    _shift_task_due_calendar_fields_one_day(t, run_date)
                            timeline_events[:] = [
                                e
                                for e in timeline_events
                                if _normalize_timeline_task_id(e) not in shift_set
                            ]
                            if _dispatch_interval_mirror is not None:
                                _dispatch_interval_mirror.rebuild_from_timeline(
                                    timeline_events
                                )
                            for t in task_queue:
                                if str(t.get("task_id", "") or "").strip() in shift_set:
                                    t["remaining_units"] = float(
                                        t.get("initial_remaining_units") or 0
                                    )
                                    t["assigned_history"].clear()
                            _apply_dispatch_trial_order_for_generate_plan(
                                task_queue,
                                req_map,
                                need_rules,
                                need_combo_col_index,
                            )
                            _trials_detail = ",".join(
                                f"{tid}:{_due_shift_retry_count_by_request[tid]}"
                                for tid in sorted(allowed_shift_tids)
                            )
                            logging.info(
                                "紝期超靎リトライ: 計画基準+1日して当該依頼のみ再配台（検出日=%s 依頼NO=%s 当該依頼の累計試行=%s）",
                                current_date.isoformat(),
                                ",".join(sorted(allowed_shift_tids)),
                                _trials_detail,
                            )
                            _full_calendar_without_deadline_restart = False
                            break
                        else:
                            # 依頼とと上限でシフトでしないの値のときは日付ループを継続れる（break れると未処理日は残り配台試行は大量発生れる）。
                            _cap_tids = sorted(
                                tid
                                for tid in shift_tid_list
                                if tid not in allowed_shift_tids
                            )
                            _first_cap_warn = [
                                tid for tid in _cap_tids if tid not in _due_shift_cap_warned_tids
                            ]
                            for tid in _first_cap_warn:
                                _due_shift_cap_warned_tids.add(tid)
                            if _first_cap_warn:
                                logging.warning(
                                    "紝期後ゝ倒し再配台: 次の依頼NOは依頼ごとの上限（坄 %s 回）のため、この検出では +1 しません。"
                                    " カレンダーは継続しした（未完了は終了時に紝期見直し必須を付け得した）: %s",
                                    STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS,
                                    ",".join(_cap_tids),
                                )

        if _full_calendar_without_deadline_restart:
            _rewind_made = _run_b2_inspection_rewind_pass(
                sorted_dates,
                attendance_data,
                task_queue,
                timeline_events,
                skills_dict,
                members,
                equipment_list,
                req_map,
                need_rules,
                surplus_map,
                global_priority_override,
                macro_run_date,
                macro_now_dt,
                _need_headcount_logged_orders,
                team_combo_presets,
                dispatch_interval_mirror=_dispatch_interval_mirror,
            )
            if _rewind_made:
                logging.info(
                    "§B-2/§B-3 リワインド: EC 完走後に検査＝巻返しのみ日付先頭から再配台しました（timeline_events を占有テーブルとして利用）。"
                )
            break

    if TRACE_SCHEDULE_TASK_IDS:
        for _tt in TRACE_SCHEDULE_TASK_IDS:
            for _t in task_queue:
                if str(_t.get("task_id", "")).strip() != _tt:
                    continue
                _log_dispatch_trace_schedule(
                    _tt,
                    "[配台トレース task=%s] シミュレーション終了時 machine=%s machine_name=%s "
                    "rem=%.4f initial=%.4f roll_b2_follower=%s",
                    _tt,
                    _t.get("machine"),
                    _t.get("machine_name"),
                    float(_t.get("remaining_units") or 0),
                    float(_t.get("initial_remaining_units") or 0),
                    bool(
                        _t.get("roll_pipeline_inspection")
                        or _t.get("roll_pipeline_rewind")
                    ),
                )
            _evs_tr = sorted(
                (
                    e
                    for e in timeline_events
                    if str(e.get("task_id", "")).strip() == _tt
                ),
                key=lambda e: (e.get("date"), e.get("start_dt") or datetime.min),
            )
            _last_ev_by_machine: dict = {}
            for _e in _evs_tr:
                _last_ev_by_machine[str(_e.get("machine") or "")] = _e
            for _mk, _ev in sorted(_last_ev_by_machine.items()):
                _ad = int(_ev.get("already_done_units") or 0)
                _ud = int(_ev.get("units_done") or 0)
                _log_dispatch_trace_schedule(
                    _tt,
                    "[配台トレース task=%s] タイムライン最終塊(工程列とと) machine=%s "
                    "already_done+units_done=%s+%s=%s total_units=%s end_dt=%s",
                    _tt,
                    _mk,
                    _ad,
                    _ud,
                    _ad + _ud,
                    _ev.get("total_units"),
                    _ev.get("end_dt"),
                )

    # need「配台時追加人数」: メイン割付後に、未参加×スキル適合者をサブへ追記（既定）
    if (
        not TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS
        and not TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW
        and surplus_map
        and timeline_events
    ):
        _n_sur = append_surplus_staff_after_main_dispatch(
            timeline_events,
            attendance_data,
            skills_dict,
            members,
            task_queue,
            req_map,
            need_rules,
            surplus_map,
            global_priority_override,
        )
        if _n_sur:
            logging.info(
                "need余力: メイン割付完了後にサブ %s 坝を追記（未割当×スキル・時間重なりなし）",
                _n_sur,
            )

    if _dispatch_interval_mirror is not None:
        _dispatch_interval_mirror.rebuild_from_timeline(timeline_events)

    # タイムラインを日付別にインデックス化し、サブメンバー一覧を事剝解析（以降の出力ループを高速化）
    for e in timeline_events:
        e["subs_list"] = [s.strip() for s in e["sub"].split(",")] if e.get("sub") else []

    events_by_date = defaultdict(list)
    for e in timeline_events:
        events_by_date[e["date"]].append(e)

    # =========================================================
    # 4. Excel出力 (メイン計画)
    # =========================================================
    _remove_prior_stage2_workbooks_and_prune_empty_dirs(output_dir)
    # ファイル名の主部はデータ抽出基準日時（シートメタと整合）。同一抽出データの再実行でも
    # パスがぶつからないよう、壁時計のサフィックスを付与（Excel 占有で旧ファイル削除失敗時の上書き不能を回避）。
    _stage2_data_stamp = base_now_dt.strftime("%Y%m%d_%H%M%S_%f")
    _stage2_run_stamp = datetime.now().strftime("%H%M%S_%f")
    _stage2_out_stamp = f"{_stage2_data_stamp}_{_stage2_run_stamp}"
    output_filename = os.path.join(
        output_dir, f"production_plan_multi_day_{_stage2_out_stamp}.xlsx"
    )
    # タスクID → 結果_設備毎の時間割で当該タスクは最初に睾れるセル（例 B12）。結果_タスク一覧のリンク用。
    first_eq_schedule_cell_by_task_id: dict[str, str] = {}
    df_eq_schedule = _build_equipment_schedule_dataframe(
        sorted_dates,
        equipment_list,
        attendance_data,
        timeline_events,
        first_eq_schedule_cell_by_task_id=first_eq_schedule_cell_by_task_id,
    )
    df_equipment_by_machine_name = _build_equipment_schedule_by_machine_name_dataframe(
        sorted_dates,
        equipment_list,
        attendance_data,
        timeline_events,
    )

    # 結果_タスク一覧用: シミュレーション上の当該タスクの最早開始・最靅終了（timeline_events 集約）
    plan_window_by_task_id: dict = {}
    for _ev in timeline_events:
        tid = _ev.get("task_id")
        if tid is None:
            continue
        sd = _ev.get("start_dt")
        ed = _ev.get("end_dt")
        if sd is None or ed is None:
            continue
        if tid not in plan_window_by_task_id:
            plan_window_by_task_id[tid] = [sd, ed]
        else:
            w = plan_window_by_task_id[tid]
            if sd < w[0]:
                w[0] = sd
            if ed > w[1]:
                w[1] = ed

    # 結果_タスク一覧の「回答納期」「指定納期」は配台計画_タスク入力の当該行セルのみ。
    # 「原反投入日」は上書き列に日付があるときはその値、ないときは列「原反投入日」（計画基準納期と混同しない）
    _result_sheet_answer_spec_by_line = {}
    _result_sheet_raw_input_by_line: dict = {}
    if tasks_df is not None and not getattr(tasks_df, "empty", True):
        for _, _r in tasks_df.iterrows():
            if _plan_row_exclude_from_assignment(_r):
                continue
            _tid = str(_planning_df_cell_scalar(_r, TASK_COL_TASK_ID) or "").strip()
            _mach = str(_planning_df_cell_scalar(_r, TASK_COL_MACHINE) or "").strip()
            if not _tid or not _mach:
                continue
            _ad = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_ANSWER_DUE))
            _sd = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_SPECIFIED_DUE))
            _rid = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_RAW_INPUT_DATE))
            _rid_ov = parse_optional_date(
                _planning_df_cell_scalar(_r, PLAN_COL_RAW_INPUT_DATE_OVERRIDE)
            )
            if _rid_ov is not None:
                _rid = _rid_ov
            _result_sheet_answer_spec_by_line[(_tid, _mach)] = (_ad, _sd)
            _result_sheet_raw_input_by_line[(_tid, _mach)] = _rid

    task_results = []
    # ステータス（配台の状態・残）：完了相当=配台済、未割当=配台不可、一部のみ=配台残
    # 計画基準+1 の再試行は依頼NOごとの上限に達した依頼の未完了行には（納期見直し必須）を付与する。
    sorted_tasks_for_result = sorted(task_queue, key=_result_task_sheet_sort_key)
    max_history_len = max(
        [
            len(merge_assigned_history_contiguous_for_result_sheet(t.get("assigned_history")))
            for t in sorted_tasks_for_result
        ]
        + [0]
    )
    for t in sorted_tasks_for_result:
        rem_u = float(t.get("remaining_units") or 0)
        hist = bool(t.get("assigned_history"))
        # 負の残は「配台済」に含めない（-0.5R 等は配台残）。浮動小数の負残は反長に比例したメートル幅で配台済に含める。
        _um0 = float(t.get("unit_m") or 0)
        _rem_abs_m = abs(rem_u * _um0)
        _noise_tol_m = max(3.0, min(100.0, 0.025 * abs(_um0)))
        if rem_u <= 1e-9 and (rem_u >= 0 or _rem_abs_m <= _noise_tol_m):
            status = "配台済"
        elif hist and t.get("_partial_retry_calendar_blocked"):
            status = "配台残(勤務カレンダー不足)"
        elif not hist and rem_u > 1e-9:
            status = "配台不可"
        else:
            status = "配台残"
        _tid_res = str(t.get("task_id", "") or "").strip()
        if (
            _tid_res in _due_shift_exhausted_requests
            and rem_u > 1e-9
            and "納期見直し必須" not in status
        ):
            status = f"{status}（納期見直し必須）"
        
        total_r = int(t['total_qty_m'] / t['unit_m']) if t['unit_m'] else 0

        _line_key = (str(t.get("task_id", "") or "").strip(), str(t.get("machine", "") or "").strip())
        _sheet_pair = _result_sheet_answer_spec_by_line.get(_line_key)
        if _sheet_pair is not None:
            _ans_d, _spec_d = _sheet_pair
            ans_s = _ans_d.strftime("%Y/%m/%d") if _ans_d else ""
            spec_s = _spec_d.strftime("%Y/%m/%d") if _spec_d else ""
        else:
            _ans_d = t.get("answer_due_date")
            _spec_d = t.get("specified_due_date")
            ans_s = _ans_d.strftime("%Y/%m/%d") if _ans_d else ""
            spec_s = _spec_d.strftime("%Y/%m/%d") if _spec_d else ""
        _basis_for_sheet = t.get("due_basis_date_result_sheet")
        if _basis_for_sheet is None:
            _basis_for_sheet = t.get("due_basis_date")
        basis_s = (
            _basis_for_sheet.strftime("%Y/%m/%d")
            if _basis_for_sheet is not None and hasattr(_basis_for_sheet, "strftime")
            else ""
        )
        if _line_key in _result_sheet_raw_input_by_line:
            _rid_d = _result_sheet_raw_input_by_line[_line_key]
            kenhan_s = _rid_d.strftime("%Y/%m/%d") if _rid_d else ""
        else:
            _rid_t = t.get("raw_input_date")
            kenhan_s = (
                _rid_t.strftime("%Y/%m/%d")
                if _rid_t is not None and hasattr(_rid_t, "strftime")
                else ""
            )
        start_req = t["start_date_req"]
        start_req_s = start_req.strftime("%Y/%m/%d") if hasattr(start_req, "strftime") else str(start_req)
        rov = t.get("required_op")
        # 列順: A=ステータス → タスクID/工程/機械/優先度 → 履歴1..n → しの他 → 最後に特別指定_AI
        row_status = {"ステータス": status}
        _dto = t.get("dispatch_trial_order")
        row_core = {
            "タスクID": t['task_id'],
            "工程名": t['machine'],
            "機械名": t.get("machine_name", ""),
            "優先度": t.get("priority", 999),
            RESULT_TASK_COL_DISPATCH_TRIAL_ORDER: _dto if _dto is not None else "",
        }
        row_history = {}
        _hist_for_sheet = merge_assigned_history_contiguous_for_result_sheet(
            t.get("assigned_history")
        )
        for i in range(max_history_len):
            if i < len(_hist_for_sheet):
                h = _hist_for_sheet[i]
                row_history[f"履歴{i+1}"] = _format_result_task_history_cell(t, h)
            else:
                row_history[f"履歴{i+1}"] = ""

        try:
            tot_qty = parse_float_safe(t.get("total_qty_m"), 0.0)
            done_qty = parse_float_safe(t.get("done_qty_reported"), 0.0)
            pct_macro = max(0, min(100, int(round((done_qty / tot_qty) * 100)))) if tot_qty > 0 else 0
        except Exception:
            pct_macro = 0

        _ub = t.get("unprocessed_baseline_m")
        _init_rem_u = float(t.get("initial_remaining_units") or 0)
        if _ub is not None:
            if _init_rem_u > 1e-12:
                _rem_qty_out = float(_ub) * (float(rem_u) / _init_rem_u)
            else:
                _rem_qty_out = float(_ub) if rem_u > 1e-12 else 0.0
        else:
            _rem_qty_out = max(0.0, float(rem_u) * float(t.get("unit_m") or 0))

        _pw = plan_window_by_task_id.get(t["task_id"])
        if _pw:
            _ps, _pe = _pw[0], _pw[1]
            plan_assign_start_s = (
                _ps.strftime("%Y/%m/%d %H:%M") if hasattr(_ps, "strftime") else ""
            )
            plan_assign_end_s = (
                _pe.strftime("%Y/%m/%d %H:%M") if hasattr(_pe, "strftime") else ""
            )
        else:
            plan_assign_start_s = ""
            plan_assign_end_s = ""

        _plan_end_ans_spec16 = _result_task_plan_end_within_answer_or_spec_16_label(
            _pw, _ans_d, _spec_d
        )

        row_tail = {
            "必須OP(上書)": rov if rov is not None else "",
            "タスク効率": parse_float_safe(t.get("task_eff_factor"), 1.0),
            "加工途中": "はい" if t.get("in_progress") else "いいえ",
            "特別指定あり": "はい" if t.get("has_special_remark") else "いいえ",
            "担当OP指定": (t.get("preferred_operator_raw") or "")[:120],
            TASK_COL_SPEED: t.get(TASK_COL_SPEED, ""),
            "回答納期": ans_s,
            "指定納期": spec_s,
            "計画基準納期": basis_s,
            TASK_COL_RAW_INPUT_DATE: kenhan_s,
            "紝期緊急": "はい" if t.get("due_urgent") else "いいえ",
            "加工開始日": start_req_s,
            "配台済_加工開始": plan_assign_start_s,
            "配台済_加工終了": plan_assign_end_s,
            RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16: _plan_end_ans_spec16,
            "累計加工量": f"{total_r}R ({t['total_qty_m']}m)",
            "残加工量": _rem_qty_out,
            "完了率(実行時点)": f"{pct_macro}%",
        }
        row_ai_last = {"特別指定_AI": (t.get("task_special_ai_note") or "")[:300]}
        row_data = {**row_status, **row_core, **row_history, **row_tail, **row_ai_last}
        task_results.append(row_data)
        
    cal_rows = []
    for d in sorted_dates:
        for m in members:
            if m in attendance_data[d]:
                data = attendance_data[d][m]
                if data['is_working']:
                    cal_end = _calendar_display_clock_out_for_calendar_sheet(data, d)
                    end_disp = cal_end if cal_end is not None else data['end_dt']
                    clock_out_s = end_disp.strftime("%H:%M")
                else:
                    clock_out_s = "休"
                cal_rows.append({
                    "日付": d,
                    "メンバー": m,
                    "出勤": data['start_dt'].strftime("%H:%M") if data['is_working'] else "休",
                    "退勤": clock_out_s,
                    "効率": data['efficiency'],
                    "備考": data['reason'],
                })

    utilization_data = []
    for d in sorted_dates:
        row_data = {"年月日": d.strftime("%Y/%m/%d (%a)")}
        # しの日のイベントからメンバー別作業分を一括集計（全メンバー×全イベントの二重ループを避ける）
        member_worked_mins = defaultdict(int)
        for ev in events_by_date[d]:
            mins = get_actual_work_minutes(ev["start_dt"], ev["end_dt"], ev["breaks"])
            member_worked_mins[ev["op"]] += mins
            for s in ev["subs_list"]:
                if s:
                    member_worked_mins[s] += mins
        for m in members:
            if m in attendance_data[d] and attendance_data[d][m]['is_working']:
                default_start = datetime.combine(d, DEFAULT_START_TIME)
                default_end = datetime.combine(d, DEFAULT_END_TIME)
                
                actual_start = attendance_data[d][m]['start_dt']
                actual_end = attendance_data[d][m]['end_dt']
                clip_start = max(actual_start, default_start)
                clip_end = min(actual_end, default_end)
                
                if clip_start >= clip_end:
                    total_avail_mins = 0
                else:
                    breaks_dt = attendance_data[d][m]['breaks_dt']
                    total_avail_mins = get_actual_work_minutes(clip_start, clip_end, breaks_dt)
                
                if total_avail_mins <= 0:
                    row_data[m] = "0.0%"
                    continue
                
                worked_mins = member_worked_mins.get(m, 0)
                ratio = (worked_mins / total_avail_mins) * 100
                row_data[m] = f"{ratio:.1f}% ({worked_mins}/{total_avail_mins}分)"
            else:
                row_data[m] = "休"
        utilization_data.append(row_data)
        
    df_utilization = pd.DataFrame(utilization_data)

    df_mprio_legend, df_mprio_tbl = build_member_assignment_priority_reference(
        skills_dict, members
    )
    if df_mprio_tbl.empty:
        df_mprio_tbl = pd.DataFrame(
            [
                {
                    "工程名": "",
                    "機械名": "",
                    "スキル列キー": "",
                    "優先順佝": "",
                    "メンバー": "",
                    "ロール": "",
                    "優先度値_尝さいろど先": "",
                    "skillsセル値": "",
                    "備考": "マスタ skills に「工程名+機械名」形式の列は見つからないか」データはありません。",
                }
            ]
        )

    _usage_txt = build_gemini_usage_summary_text()
    if _usage_txt:
        ai_log_data["Gemini_トークン・料金サマリ"] = _usage_txt[:50000]

    _master_abs_for_result_fmt = os.path.abspath(os.path.join(os.getcwd(), MASTER_FILE))
    _reg_shift_start, _reg_shift_end = _read_master_main_regular_shift_times(
        _master_abs_for_result_fmt
    )
    if _reg_shift_start is not None and _reg_shift_end is not None:
        logging.info(
            "定常枠: master メイン A15/B15 → %s ～ %s（結果の定常外「日時帯」着色）",
            _reg_shift_start.strftime("%H:%M"),
            _reg_shift_end.strftime("%H:%M"),
        )

    logging.info(
        "段階2: 結果ブックを作成しした → %s",
        os.path.basename(output_filename),
    )
    gantt_tl_label_specs: list = []
    gantt_tl_day_blocks: list = []
    gantt_detail_tl_label_specs: list = []
    gantt_detail_tl_day_blocks: list = []
    df_actual_detail = load_machining_actual_detail_df()
    detail_timeline_events: list = []
    sorted_dates_detail = list(sorted_dates)
    chart_title_actual_detail = "湖南工場 加工実績（明細）"
    if df_actual_detail is not None and len(df_actual_detail) > 0:
        sorted_dates_detail = _sorted_dates_union_actual_bounds_df(
            sorted_dates, df_actual_detail
        )
        d_from = _parse_env_optional_date(ENV_GANTT_ACTUAL_DETAIL_DATE_FROM)
        d_to = _parse_env_optional_date(ENV_GANTT_ACTUAL_DETAIL_DATE_TO)
        if d_from is not None or d_to is not None:
            n_before = len(sorted_dates_detail)
            filtered_detail_dates = _sorted_dates_filter_inclusive_range(
                sorted_dates_detail, d_from, d_to
            )
            if not filtered_detail_dates and sorted_dates_detail:
                logging.warning(
                    "実績明細ガント: 日付範囲フィルタで表示日が0件になったためフィルタを無視します。"
                    "（%s=%r, %s=%r）",
                    ENV_GANTT_ACTUAL_DETAIL_DATE_FROM,
                    os.environ.get(ENV_GANTT_ACTUAL_DETAIL_DATE_FROM, ""),
                    ENV_GANTT_ACTUAL_DETAIL_DATE_TO,
                    os.environ.get(ENV_GANTT_ACTUAL_DETAIL_DATE_TO, ""),
                )
            else:
                sorted_dates_detail = filtered_detail_dates
                logging.info(
                    "実績明細ガント: 表示日を %s 日 → %s 日に絞りました（FROM=%s, TO=%s）。",
                    n_before,
                    len(sorted_dates_detail),
                    d_from.isoformat() if d_from else "（指定なし）",
                    d_to.isoformat() if d_to else "（指定なし）",
                )
                rng_lo = d_from.isoformat() if d_from else "…"
                rng_hi = d_to.isoformat() if d_to else "…"
                chart_title_actual_detail = (
                    f"{chart_title_actual_detail}（表示 {rng_lo}～{rng_hi}）"
                )
        detail_timeline_events = build_actual_timeline_events(
            df_actual_detail,
            equipment_list,
            sorted_dates_detail,
            log_sheet_name=ACTUAL_DETAIL_SHEET_NAME,
            roll_detail=True,
        )
    try:
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            df_eq_schedule.to_excel(
                writer, sheet_name=RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME, index=False
            )
            df_equipment_by_machine_name.to_excel(
                writer, sheet_name=RESULT_EQUIPMENT_BY_MACHINE_SHEET_NAME, index=False
            )
            pd.DataFrame(cal_rows).to_excel(writer, sheet_name='結果_カレンダー(出勤簿)', index=False)
            df_utilization.to_excel(writer, sheet_name='結果_メンバー別作業割引', index=False)
            df_tasks = pd.DataFrame(task_results)
            df_tasks, task_column_order, _, vis_map = apply_result_task_sheet_column_order(
                df_tasks, max_history_len
            )
            # 列設定シートは「列名」「表示」のデータ行が必須。task_results が空だと
            # apply_result_task_sheet_column_order は ordered が空になり、見出しのみのシートになる。
            if not task_column_order:
                task_column_order, vis_map = _result_task_column_config_fallback_from_existing(
                    df_tasks, max_history_len
                )
            seen_tc: set[str] = set()
            task_column_order_dedup: list = []
            vis_list_dedup: list = []
            for c in task_column_order:
                if c in seen_tc:
                    continue
                seen_tc.add(c)
                task_column_order_dedup.append(c)
                vis_list_dedup.append(bool(vis_map.get(c, True)))
            pd.DataFrame(
                {
                    "列名": task_column_order_dedup,
                    "表示": vis_list_dedup,
                }
            ).to_excel(writer, sheet_name=COLUMN_CONFIG_SHEET_NAME, index=False)
            df_tasks.to_excel(writer, sheet_name=RESULT_TASK_SHEET_NAME, index=False)
            pd.DataFrame(list(ai_log_data.items()), columns=["項目", "内容"]).to_excel(writer, sheet_name='結果_AIログ', index=False)

            _mprio_sheet = RESULT_MEMBER_PRIORITY_SHEET_NAME
            df_mprio_legend.to_excel(writer, sheet_name=_mprio_sheet, index=False)
            _mprio_gap = len(df_mprio_legend) + 2
            df_mprio_tbl.to_excel(
                writer, sheet_name=_mprio_sheet, index=False, startrow=_mprio_gap
            )

            logging.info(
                "段階2: 設備ガントチャートを生成（データ量により数分かかることがあります）"
            )
            gantt_tl_label_specs, gantt_tl_day_blocks = _write_results_equipment_gantt_sheet(
                writer,
                timeline_events,
                equipment_list,
                sorted_dates,
                attendance_data,
                data_extract_dt_str,
                base_now_dt,
                regular_shift_times=(_reg_shift_start, _reg_shift_end),
            )

            if detail_timeline_events:
                logging.info(
                    "段階2: 設備ガントチャート（加工実績明細）を生成します（データ量により時間がかかることがあります）"
                )
                (
                    gantt_detail_tl_label_specs,
                    gantt_detail_tl_day_blocks,
                ) = _write_results_equipment_gantt_sheet(
                    writer,
                    [],
                    equipment_list,
                    sorted_dates_detail,
                    attendance_data,
                    data_extract_dt_str,
                    base_now_dt,
                    actual_timeline_events=detail_timeline_events,
                    regular_shift_times=(_reg_shift_start, _reg_shift_end),
                    plan_rows=False,
                    chart_title=chart_title_actual_detail,
                    sheet_name_override=RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME,
                )

            for sheet_name, ws_out in writer.sheets.items():
                if sheet_name in (
                    RESULT_SHEET_GANTT_NAME,
                    RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME,
                ):
                    continue
                _apply_output_font_to_result_sheet(ws_out)

            if _reg_shift_start is not None and _reg_shift_end is not None:
                for _eq_sched_sheet in (
                    RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME,
                    RESULT_EQUIPMENT_BY_MACHINE_SHEET_NAME,
                ):
                    if _eq_sched_sheet in writer.sheets:
                        _apply_equipment_schedule_outside_regular_fill(
                            writer.sheets[_eq_sched_sheet],
                            _reg_shift_start,
                            _reg_shift_end,
                        )

            if RESULT_EQUIPMENT_BY_MACHINE_SHEET_NAME in writer.sheets:
                _apply_equipment_by_machine_dispatched_request_fill(
                    writer.sheets[RESULT_EQUIPMENT_BY_MACHINE_SHEET_NAME]
                )

            if RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME in writer.sheets:
                _apply_equipment_schedule_prep_cleanup_fill(
                    writer.sheets[RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME]
                )

            if RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME in writer.sheets:
                _apply_equipment_schedule_machine_calendar_fill(
                    writer.sheets[RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME],
                    equipment_list,
                    _MACHINE_CALENDAR_BLOCKS_BY_DATE,
                )

            ws_cfg = writer.sheets[COLUMN_CONFIG_SHEET_NAME]
            _add_column_config_sheet_helpers(ws_cfg, len(task_column_order_dedup))

            worksheet_tasks = writer.sheets[RESULT_TASK_SHEET_NAME]
            max_col = worksheet_tasks.max_column
            for row in worksheet_tasks.iter_rows(min_row=1, max_row=worksheet_tasks.max_row, max_col=max_col):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=False, vertical="top")

            _apply_result_task_sheet_column_visibility(
                worksheet_tasks, list(df_tasks.columns), vis_map
            )

            _apply_result_task_history_rich_text(worksheet_tasks, list(df_tasks.columns))
            _apply_result_task_date_columns_blue_font(worksheet_tasks, list(df_tasks.columns))

            # 未スケジュール行（配台不可・配台残）を目立たせる
            status_col_idx = None
            for col_idx, col_name in enumerate(df_tasks.columns, 1):
                if str(col_name) == "ステータス":
                    status_col_idx = col_idx
                    break
            if status_col_idx is not None:
                unscheduled_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for r in range(2, worksheet_tasks.max_row + 1):
                    st_val = worksheet_tasks.cell(row=r, column=status_col_idx).value
                    st = str(st_val).strip() if st_val is not None else ""
                    if st in ("配台不可", "配台残", "配台試行"):
                        for c in range(1, max_col + 1):
                            worksheet_tasks.cell(row=r, column=c).fill = unscheduled_fill

            _apply_result_task_history_need_surplus_highlight(
                worksheet_tasks, list(df_tasks.columns), sorted_tasks_for_result
            )

            _apply_result_task_task_id_content_mismatch_highlight(
                worksheet_tasks, list(df_tasks.columns), sorted_tasks_for_result
            )
            _apply_result_task_plan_end_answer_spec_16_no_highlight(
                worksheet_tasks, list(df_tasks.columns)
            )
            _apply_result_task_id_hyperlinks_to_equipment_schedule(
                worksheet_tasks,
                list(df_tasks.columns),
                sorted_tasks_for_result,
                first_eq_schedule_cell_by_task_id,
                RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME,
            )

    except OSError as e:
        logging.error(
            "段階2: 結果ブックの作成・保存に失敗しました: %s（%s）。"
            "output 内の production_plan_multi_day_*.xlsx を Excel で開いでいないか確認してください。",
            output_filename,
            e,
        )
        raise

    try:
        _apply_excel_date_columns_date_only_display(
            output_filename, "結果_カレンダー(出勤簿)", frozenset({"日付"})
        )
    except Exception as e:
        logging.warning(f"結果_カレンダー(出勤簿)の日付列表示整形: {e}")

    _stage2_try_copy_column_config_shapes_from_input(
        output_filename,
        (os.environ.get("TASK_INPUT_WORKBOOK", "").strip() or TASKS_INPUT_WORKBOOK),
    )

    _stage2_try_add_gantt_timeline_shape_labels(
        output_filename, gantt_tl_label_specs, gantt_tl_day_blocks
    )
    if gantt_detail_tl_label_specs:
        _stage2_try_add_gantt_timeline_shape_labels(
            output_filename,
            gantt_detail_tl_label_specs,
            gantt_detail_tl_day_blocks,
            sheet_name=RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME,
        )

    logging.info(f"完了: '{output_filename}' を生成しました。")

    # =========================================================
    # 5. ★追加: メンバー毎の行動スケジュール (別ファイル) 出力
    # =========================================================
    member_output_filename = os.path.join(
        output_dir, f"member_schedule_{_stage2_out_stamp}.xlsx"
    )
    
    # 時間帯は全メンバー共通で1回の値生成（メンバー数分の重複計算を避ける）
    time_labels = []
    time_grids = []
    curr_dt = datetime.combine(run_date, DEFAULT_START_TIME)
    end_dt_grid = datetime.combine(run_date, DEFAULT_END_TIME)
    while curr_dt < end_dt_grid:
        next_dt = curr_dt + timedelta(minutes=10)
        if next_dt > end_dt_grid:
            next_dt = end_dt_grid
        time_labels.append(f"{curr_dt.strftime('%H:%M')}-{next_dt.strftime('%H:%M')}")
        time_grids.append((curr_dt.time(), next_dt.time()))
        curr_dt = next_dt
    
    logging.info(
        "段階2: メンバー別スケジュールを作成しした → %s",
        os.path.basename(member_output_filename),
    )
    try:
        with pd.ExcelWriter(member_output_filename, engine="openpyxl") as member_writer:
            for m in members:
                # 坄行の辞書を初期化
                m_schedule = {t_label: {"時間帯": t_label} for t_label in time_labels}
            
                # 坄日付のスケジュールを列として埋ゝでいし
                for d in sorted_dates:
                    d_str = d.strftime("%m/%d (%a)")
                
                    # 全日非勤務: 年休（カレンダー *）は『年休」」工場休日などは『休」
                    if m not in attendance_data[d] or not attendance_data[d][m]['is_working']:
                        off_label = _member_schedule_full_day_off_label(
                            attendance_data[d].get(m) if m in attendance_data[d] else None
                        )
                        for t_label in time_labels:
                            m_schedule[t_label][d_str] = off_label
                        continue
                
                    daily_info = attendance_data[d][m]
                    d_start_dt = daily_info['start_dt']
                    d_end_dt = daily_info['end_dt']
                    breaks_dt = daily_info['breaks_dt']
                
                    events_today = events_by_date[d]
                
                    for i, (t_start, t_end) in enumerate(time_grids):
                        t_label = time_labels[i]
                    
                        # 判定用の中間時刻を計算
                        grid_start_dt = datetime.combine(d, t_start)
                        grid_end_dt = datetime.combine(d, t_end)
                        grid_mid_dt = grid_start_dt + (grid_end_dt - grid_start_dt) / 2
                    
                        text = ""
                        if grid_mid_dt < d_start_dt or grid_mid_dt >= d_end_dt:
                            text = _member_schedule_off_shift_label(
                                d, grid_mid_dt, d_start_dt, d_end_dt, daily_info.get("reason")
                            )
                        else:
                            br_txt = _member_schedule_break_cell_label(
                                grid_mid_dt, breaks_dt, d_end_dt, daily_info.get("reason")
                            )
                            if br_txt is not None:
                                text = br_txt
                        if text == "":
                            # 該当れるタスクを探れ（subs_list は事剝解析済み）
                            active_ev = next((e for e in events_today if e['start_dt'] <= grid_mid_dt < e['end_dt'] and (e['op'] == m or m in e.get('subs_list', []))), None)
                            if active_ev:
                                role = "主" if active_ev['op'] == m else "補"
                                text = f"[{active_ev['task_id']}] {active_ev['machine']}({role})"
                            else:
                                text = "" # 何も割り当でられでいない空し時間
                    
                        m_schedule[t_label][d_str] = text
                    
                # データフレーム化してシートに書き込み
                df_m = pd.DataFrame(list(m_schedule.values()))
                cols = ["時間帯"] + [d.strftime("%m/%d (%a)") for d in sorted_dates]
                df_m = df_m[[c for c in cols if c in df_m.columns]]
                df_m.to_excel(member_writer, sheet_name=m, index=False)
            
                # --- 既定フォント・罫線・見出し背景（列幅は VBA 取り込み時の AutoFit） ---
                worksheet = member_writer.sheets[m]
                _apply_output_font_to_result_sheet(worksheet)
                header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                for cell in worksheet[1]:
                    cell.fill = header_fill
                
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border

    except OSError as e:
        logging.error(
            "段階2: メンバー別スケジュールの保存に失敗しました: %s（%s）。"
            "member_schedule_*.xlsx を Excel で開いでいないか確認してください。",
            member_output_filename,
            e,
        )
        raise

    logging.info(f"完了: 個人別スケジュールを '{member_output_filename}' に出力しました。")
    _try_write_main_sheet_gemini_usage_summary("段階2")