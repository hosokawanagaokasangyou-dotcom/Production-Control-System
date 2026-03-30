# -*- coding: utf-8 -*-
"""一時: 生産管理_AI配台テスト.xlsm からタスク・W4-1 行をダンプ（調査用）"""
from __future__ import annotations

import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSM = ROOT / "生産管理_AI配台テスト.xlsm"
PLAN_SHEET = "配台計画_タスク入力"


def cell_str(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y/%m/%d")
    return str(v).strip()


def main() -> int:
    if not XLSM.is_file():
        print("見つかりません:", XLSM, file=sys.stderr)
        return 1
    os.chdir(ROOT)
    out_path = ROOT / "log" / "xlsm_W4-1_investigate.txt"
    out_lines: list[str] = []

    def out(s: str = "") -> None:
        out_lines.append(s)

    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
    out("=== シート一覧 ===")
    for name in wb.sheetnames:
        out(f"  {name}")

    if PLAN_SHEET not in wb.sheetnames:
        print("シート無し:", PLAN_SHEET, file=sys.stderr)
        wb.close()
        return 1

    ws = wb[PLAN_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        out("空シート")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text("\n".join(out_lines), encoding="utf-8")
        print(out_path)
        return 0

    header = [cell_str(x) for x in rows[0]]
    out("")
    out(f"=== {PLAN_SHEET} 見出し（先頭40列） ===")
    for i, h in enumerate(header[:40]):
        if h:
            out(f"  col{i+1}: {h}")

    # 依頼NO 列を推定
    tid_col = None
    for i, h in enumerate(header):
        if h in ("依頼NO", "タスクID"):
            tid_col = i
            break
    if tid_col is None:
        for i, h in enumerate(header):
            if "依頼" in h and "NO" in h:
                tid_col = i
                break

    def row_dict(r):
        return {header[i]: cell_str(r[i]) if i < len(r) else "" for i in range(min(len(header), len(r))) if header[i]}

    out("")
    out("=== W4-1 を含む行（依頼NO 列） ===")
    hits = []
    for r in rows[1:]:
        if not r:
            continue
        tid = ""
        if tid_col is not None and tid_col < len(r):
            tid = cell_str(r[tid_col])
        line = " ".join(cell_str(c) for c in r if c is not None and str(c).strip())
        if "W4-1" in tid or "W4-1" in line:
            hits.append(r)

    if not hits:
        out("（該当なし。先頭5行の依頼NO列サンプル）")
        for r in rows[1:6]:
            if r and tid_col is not None and tid_col < len(r):
                out(f"  {cell_str(r[tid_col])}")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text("\n".join(out_lines), encoding="utf-8")
        print("Wrote", out_path)
        return 0

    keys_of_interest = [
        "依頼NO",
        "タスクID",
        "工程名",
        "機械名",
        "加工内容",
        "回答納期",
        "指定納期",
        "指定納期_上書き",
        "加工開始日",
        "加工開始日_指定",
        "換算数量",
        "優先度",
        "配台不要",
        "実出来高換算",
        "完了区分",
    ]

    for idx, r in enumerate(hits):
        d = row_dict(r)
        out("")
        out(f"--- hit {idx+1} ---")
        for k in keys_of_interest:
            if k in d and d[k]:
                out(f"  {k}: {d[k]}")
        extra = {k: v for k, v in d.items() if k not in keys_of_interest and v}
        if extra:
            out("  （その他非空）")
            for k, v in sorted(extra.items())[:25]:
                out(f"    {k}: {v[:120] if len(v) > 120 else v}")
            if len(extra) > 25:
                out(f"    ... 他 {len(extra)-25} 列")

    out("")
    out("=== 補足（planning_core ロールパイプライン定数との一致） ===")
    out("  EC 行: 工程名=EC, 機械名=EC機　湖南 → roll_pipeline_ec")
    out("  検査行: 工程名=検査, 機械名=熱融着機　湖南 → roll_pipeline_inspection")
    out("  回答納期=指定納期=2026-04-06 のため due_basis はあり、納期+1日リトライ対象になり得る。")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(out_lines), encoding="utf-8")
    print("Wrote", out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
