# -*- coding: utf-8 -*-
"""
Extract sheet \"\u8a2d\u5b9a_\u74b0\u5883\u5909\u6570\" from a reference workbook to JSON (no runtime dependency on the xlsx).

Row rules mirror workbook_env_bootstrap: A=key, B=value, C=optional note; skip empty A and #-comments.

Examples:
  python export_workbook_env_sheet_to_json.py
  python export_workbook_env_sheet_to_json.py --workbook C:/path/book.xlsx --out config/env_snapshot.json
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path

# Same logical name as workbook_env_bootstrap.WORKBOOK_ENV_SHEET_NAME
WORKBOOK_ENV_SHEET_NAME = "\u8a2d\u5b9a_\u74b0\u5883\u5909\u6570"


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, float) and math.isfinite(v) and v == int(v):
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _default_ui_ref_workbook() -> Path | None:
    here = Path(__file__).resolve().parent
    pcs = here.parent.parent
    candidate = pcs / "plan" / (
        "UI\u53c2\u7167\u7528_\u751f\u7523\u7ba1\u7406_AI\u914d\u53f0(RC1).xlsx"
    )
    return candidate if candidate.is_file() else None


def extract_env_sheet_rows(workbook_path: Path, sheet_name: str = WORKBOOK_ENV_SHEET_NAME) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("openpyxl required: pip install openpyxl") from e

    p = workbook_path.resolve()
    if not p.is_file():
        raise FileNotFoundError(str(p))

    entries: list[dict[str, str]] = []
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name!r} in {p}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return entries
        start_i = 0
        head = rows[0]
        if head:
            hk = _cell_str(head[0]).casefold()
            if hk in (
                "\u5909\u6570\u540d",
                "name",
                "key",
                "\u74b0\u5883\u5909\u6570",
                "env",
            ):
                start_i = 1
        for row in rows[start_i:]:
            if not row:
                continue
            raw_k = row[0] if len(row) > 0 else None
            k = _cell_str(raw_k)
            if not k:
                continue
            if k.lstrip().startswith("#"):
                continue
            val_cell = row[1] if len(row) > 1 else None
            v = _cell_str(val_cell)
            desc_cell = row[2] if len(row) > 2 else None
            desc = _cell_str(desc_cell)
            item: dict[str, str] = {"key": k, "value": v}
            if desc:
                item["description"] = desc
            entries.append(item)
        return entries
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Export workbook env sheet to JSON.")
    ap.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Path to xlsx/xlsm (default: plan/UI ref workbook under Production-Control-System)",
    )
    ap.add_argument(
        "--sheet",
        default=WORKBOOK_ENV_SHEET_NAME,
        help="Sheet name (default: \u8a2d\u5b9a_\u74b0\u5883\u5909\u6570)",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output JSON path (default: stdout)",
    )
    ap.add_argument(
        "--compact",
        action="store_true",
        help="Minified JSON",
    )
    ns = ap.parse_args(argv)

    wb_path = ns.workbook
    if wb_path is None:
        d = _default_ui_ref_workbook()
        if d is None:
            print(
                "Default UI reference workbook not found; pass --workbook.",
                file=sys.stderr,
            )
            return 2
        wb_path = d

    try:
        entries = extract_env_sheet_rows(wb_path, ns.sheet)
    except (OSError, ValueError) as e:
        print(str(e), file=sys.stderr)
        return 1

    payload = {
        "source_workbook": str(wb_path.resolve()),
        "sheet": ns.sheet,
        "entries": entries,
    }
    indent = None if ns.compact else 2
    text = json.dumps(payload, ensure_ascii=False, indent=indent)
    if ns.out is not None:
        ns.out.parent.mkdir(parents=True, exist_ok=True)
        ns.out.write_text(text, encoding="utf-8")
        print(ns.out.resolve(), file=sys.stderr)
    else:
        sys.stdout.write(text)
        if not text.endswith("\n"):
            sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
