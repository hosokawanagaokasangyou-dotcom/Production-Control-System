"""論理ビュー用: 結合セル展開後の表 JSON（openpyxl・pandas が使える環境のみ）。"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

if sys.version_info < (3, 14):
    pytest.skip(
        "planning_core の bootstrap は Python 3.14+ が必要",
        allow_module_level=True,
    )


def test_expand_merge_then_payload_has_slot_values(tmp_path: Path):
    pytest.importorskip("openpyxl")

    from openpyxl import Workbook

    from planning_core.logical_workbook_view import build_logical_view_workbook_payload

    p = tmp_path / "t.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "結果_設備ガント"
    ws["A1"] = "日付"
    ws["B1"] = "機械名"
    ws["C1"] = "09:00"
    ws["D1"] = "09:10"
    ws["A2"] = "2026/4/1"
    ws["B2"] = "M1"
    ws["C2"] = "タスクA"
    ws.merge_cells("C2:D2")
    wb.save(p)

    payload = build_logical_view_workbook_payload(str(p), source_xlsx_basename="t.xlsx")
    assert payload.get("schema") == "plan_logical_view_v1"
    assert payload.get("logical_view") is True
    sh = (payload.get("sheets") or {}).get("結果_設備ガント")
    assert sh is not None
    rows = sh.get("rows") or []
    assert len(rows) >= 1
    r0 = rows[0]
    assert r0.get("09:00") == "タスクA"
    assert r0.get("09:10") == "タスクA"


def test_logical_view_json_path_suffix():
    from planning_core.logical_workbook_view import logical_view_json_path

    base = os.path.join("C:", "out", "plan.xlsx") if os.name == "nt" else "/tmp/out/plan.xlsx"
    j = logical_view_json_path(base)
    assert j.endswith("_logical_view.json")
    assert "plan_logical_view.json" in j or j.endswith("plan_logical_view.json")
