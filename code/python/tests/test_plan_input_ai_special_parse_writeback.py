# -*- coding: utf-8 -*-
"""「AI特別指定_解析」列の書き戻し（段階2）と段階1での引き継ぎ。"""
from __future__ import annotations

import json

import pandas as pd
import pytest
from openpyxl import Workbook

from planning_core import _core as pc


def _plan_df(rows: list[dict]) -> pd.DataFrame:
    base = {
        pc.TASK_COL_TASK_ID: "",
        pc.TASK_COL_MACHINE: "分割",
        pc.TASK_COL_MACHINE_NAME: "SL-1",
        pc.PLAN_COL_SPECIAL_REMARK: "",
        pc.PLAN_COL_EXCLUDE_FROM_ASSIGNMENT: "",
    }
    return pd.DataFrame([{**base, **r} for r in rows])


def test_collect_ai_special_parse_returns_json_text_keyed_by_excel_row():
    df = _plan_df(
        [
            {pc.TASK_COL_TASK_ID: "Y3-26", pc.PLAN_COL_SPECIAL_REMARK: "配台は8/30以降に"},
            {pc.TASK_COL_TASK_ID: "Y4-2", pc.PLAN_COL_SPECIAL_REMARK: "急ぎ"},
        ]
    )
    ai_by_tid = {"Y3-26": {"start_date": "2026-08-30"}, "Y4-2": {"priority": 1}}

    got = pc.collect_plan_input_ai_special_parse_by_excel_row(df, ai_by_tid)

    assert json.loads(got[2]) == {"start_date": "2026-08-30"}
    assert json.loads(got[3]) == {"priority": 1}


def test_collect_ai_special_parse_clears_rows_without_ai_entry():
    df = _plan_df(
        [
            {pc.TASK_COL_TASK_ID: "Y3-26", pc.PLAN_COL_SPECIAL_REMARK: "配台は8/30以降に"},
            {pc.TASK_COL_TASK_ID: "Y9-9"},
        ]
    )
    ai_by_tid = {"Y3-26": {"start_date": "2026-08-30"}}

    got = pc.collect_plan_input_ai_special_parse_by_excel_row(df, ai_by_tid)

    assert got[3] == ""


def test_collect_ai_special_parse_clears_rows_excluded_from_assignment():
    df = _plan_df(
        [
            {
                pc.TASK_COL_TASK_ID: "Y3-26",
                pc.PLAN_COL_SPECIAL_REMARK: "配台は8/30以降に",
                pc.PLAN_COL_EXCLUDE_FROM_ASSIGNMENT: "yes",
            }
        ]
    )
    ai_by_tid = {"Y3-26": {"start_date": "2026-08-30"}}

    got = pc.collect_plan_input_ai_special_parse_by_excel_row(df, ai_by_tid)

    assert got[2] == ""


def test_collect_ai_special_parse_truncates_long_json():
    df = _plan_df([{pc.TASK_COL_TASK_ID: "Y3-26", pc.PLAN_COL_SPECIAL_REMARK: "長文"}])
    ai_by_tid = {"Y3-26": {"interpretation_ja": "あ" * 2000}}

    got = pc.collect_plan_input_ai_special_parse_by_excel_row(df, ai_by_tid)

    assert len(got[2]) == pc.PLAN_AI_SPECIAL_PARSE_CELL_MAX_LEN


def test_write_ai_special_parse_cells_sets_and_clears_cells():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = pc.TASK_COL_TASK_ID
    ws.cell(1, 2).value = pc.PLAN_COL_SPECIAL_REMARK
    ws.cell(1, 3).value = pc.PLAN_COL_AI_PARSE
    ws.cell(2, 1).value = "Y3-26"
    ws.cell(3, 1).value = "Y9-9"
    ws.cell(3, 3).value = "前回の古い解析"

    n = pc._plan_sheet_write_ai_special_parse_cells_to_ws(
        ws, 2, {2: '{"priority": 1}', 3: ""}
    )

    assert n == 1
    assert ws.cell(2, 3).value == '{"priority": 1}'
    assert ws.cell(3, 3).value is None


def test_write_ai_special_parse_cells_accepts_renamed_column():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = pc.TASK_COL_TASK_ID
    ws.cell(1, 2).value = "AI納期回答_解析"
    ws.cell(2, 1).value = "Y3-26"

    n = pc._plan_sheet_write_ai_special_parse_cells_to_ws(
        ws, 1, {2: '{"start_date": "2026-08-30"}'}
    )

    assert n == 1
    assert ws.cell(2, 2).value == '{"start_date": "2026-08-30"}'


def test_write_ai_special_parse_cells_is_noop_without_column():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value = pc.TASK_COL_TASK_ID

    assert pc._plan_sheet_write_ai_special_parse_cells_to_ws(ws, 1, {2: "x"}) == 0


def test_stage1_merge_carries_over_ai_special_parse():
    assert pc.PLAN_COL_AI_PARSE in pc.PLAN_STAGE1_MERGE_COLUMNS


def test_one_io_writer_accepts_ai_parse_rows(tmp_path):
    """段階2の1回保存に AI解析の書き戻しが同乗していること（シート不在でも引数を受ける）。"""
    import inspect

    sig = inspect.signature(pc.write_plan_sheet_global_parse_and_conflict_styles_one_io)
    assert "ai_parse_by_row" in sig.parameters
    sig2 = inspect.signature(pc._try_write_plan_input_global_parse_and_conflicts_one_save)
    assert "ai_parse_by_row" in sig2.parameters
