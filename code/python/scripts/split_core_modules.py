#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""One-shot mechanical split of planning_core/_core.py into planning_core/core/*.py.

Run from repo: python3 code/python/scripts/split_core_modules.py
"""
from __future__ import annotations

import ast
import re
import textwrap
from pathlib import Path

_PY_ROOT = Path(__file__).resolve().parents[1]
CORE_PY = _PY_ROOT / "planning_core/_core.py"
CORE_PKG = _PY_ROOT / "planning_core/core"

# Module assignment: (start_line inclusive, end_line inclusive, module_name)
# Line ranges based on plan anchors (grep-verified 2026-06).
LINE_RANGES: list[tuple[int, int, str]] = [
    (1, 184, "state"),  # imports bootstrap block handled separately
    (185, 1142, "gemini_auth"),
    (1143, 3446, "columns"),  # column constants + early helpers before gantt
    (3447, 6328, "gantt_excel"),
    (6329, 13168, "plan_input"),
    (13169, 17526, "task_queue"),
    (17527, 17919, "stage1"),
    (17920, 19097, "time_utils"),
    (19098, 20617, "master_data"),
    (20618, 34303, "roll_pipeline"),  # includes dispatch loop helpers
    (34304, 35878, "dispatch_loop"),
    (35879, 37179, "output_refresh"),
    (37180, 99999, "stage2_impl"),
]

# Name-prefix overrides (higher priority than line ranges)
PREFIX_MODULES: list[tuple[str, str]] = [
    ("TASK_COL_", "columns"),
    ("PLAN_COL_", "columns"),
    ("ACT_COL_", "columns"),
    ("NEED_COL_", "columns"),
    ("ACTUAL_", "columns"),
    ("RESULT_", "columns"),
    ("EXCLUDE_RULE", "columns"),
    ("PLAN_INPUT_", "columns"),
    ("STAGE1_", "columns"),
    ("DISPATCH_", "columns"),
    ("COMPARE_GANTT", "columns"),
    ("ENV_", "columns"),
    ("SHEET_", "columns"),
    ("TIMELINE_EVENT_", "columns"),
    ("MASTER_SHEET_", "columns"),
    ("ATT_COL_", "columns"),
    ("ATTENDANCE_", "columns"),
    ("ROLL_PIPELINE_", "roll_pipeline"),
    ("WIP_LIMIT_", "roll_pipeline"),
    ("_gantt_", "gantt_excel"),
    ("gantt_", "gantt_excel"),
    ("_write_results_equipment", "gantt_excel"),
    ("_gemini_", "gemini_auth"),
    ("_load_gemini_", "gemini_auth"),
    ("_decrypt_gemini_", "gemini_auth"),
    ("_credentials_json", "gemini_auth"),
    ("_fernet_", "gemini_auth"),
    ("build_task_queue", "task_queue"),
    ("_stage2_in_progress_next_day", "task_queue"),
    ("run_stage1", "stage1"),
    ("_stage1_", "stage1"),
    ("_write_stage1_", "stage1"),
    ("merge_time_intervals", "time_utils"),
    ("_eod_", "time_utils"),
    ("_break_", "time_utils"),
    ("_contiguous_work", "time_utils"),
    ("_defer_team", "time_utils"),
    ("load_skills_and_needs", "master_data"),
    ("_trial_order_", "dispatch_loop"),
    ("_assign_", "dispatch_loop"),
    ("_changeover_", "dispatch_loop"),
    ("_interactive_trial_", "dispatch_loop"),
    ("_interactive_stage3_", "dispatch_loop"),
    ("append_surplus_staff", "dispatch_loop"),
    ("_generate_plan_impl", "stage2_impl"),
    ("generate_plan", "output_refresh"),
    ("refresh_", "output_refresh"),
    ("write_plan_actual_compare", "output_refresh"),
    ("_compare_gantt_", "output_refresh"),
    ("_build_compare_gantt_", "output_refresh"),
    ("_build_plan_timeline_", "output_refresh"),
    ("interactive_stage3_", "state"),
    ("interactive_trial_", "state"),
]

STANDARD_HEADER = textwrap.dedent(
    '''\
    # -*- coding: utf-8 -*-
    """Extracted from planning_core._core (mechanical split)."""
    from __future__ import annotations

    import base64
    import calendar
    import copy
    import csv
    import ctypes
    import fnmatch
    import hashlib
    import itertools
    import json
    import logging
    import math
    import os
    import pathlib
    import random
    import re
    import shutil
    import sys
    import threading
    import traceback
    import unicodedata
    import time as time_module
    from collections import Counter, defaultdict
    from contextlib import contextmanager
    from datetime import date, datetime, time, timedelta

    import pandas as pd
    from google import genai
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.worksheet.table import Table, TableStyleInfo

    from dispatch_interval_mirror import DispatchIntervalMirror

    from planning_core.bootstrap import (
        PlanningValidationError,
        _clear_stage2_blocking_message_file,
        _remove_prior_stage2_workbooks_and_prune_empty_dirs,
        _try_remove_path_with_retries,
        _write_stage2_blocking_message,
        api_payment_dir,
        json_data_dir,
        log_dir,
        output_dir,
    )
    from planning_core.dispatch_workspace import (
        ENV_PLAN_INPUT_PATH,
        ENV_PROCESSING_PLAN_PATH,
        _read_excel_tabular,
        _resolve_tabular_excel_header_row_0based,
        _resolve_tabular_sheet_name_calamine,
        plan_input_workbook_path_for_excel_ops,
        read_tabular_dataframe,
        resolve_actual_detail_workbook_path,
        resolve_processing_plan_path_from_env,
        resolve_result_dispatch_table_output_dir,
    )
    from planning_core.input_resolution import (
        ENV_EXCLUDE_RULES_JSON,
        ENV_GLOBAL_PRIORITY_OVERRIDE_PATH,
        ENV_RESULT_TASK_COLUMN_CONFIG_CSV,
        resolve_actuals_workbook_path,
        resolve_column_config_workbook_path,
        resolve_data_extraction_workbook_path,
    )
    from planning_core.plan_workbook_sidecar import (
        normalized_workbook_json_path,
        read_result_task_dataframe,
        write_member_schedule_workbook_json,
        write_production_plan_logical_view_json,
        write_production_plan_workbook_json,
        write_result_task_json_sidecar,
    )
    from planning_core.stage2_output_naming import (
        format_stage2_stamp,
        member_workbook_filename,
        plan_workbook_filename,
    )

    '''
)

MODULE_ORDER = [
    "state",
    "columns",
    "gemini_auth",
    "gantt_excel",
    "plan_input",
    "task_queue",
    "stage1",
    "time_utils",
    "master_data",
    "roll_pipeline",
    "dispatch_loop",
    "output_refresh",
    "stage2_impl",
]


def _line_range_module(lineno: int) -> str:
    for start, end, mod in LINE_RANGES:
        if start <= lineno <= end:
            return mod
    return "roll_pipeline"


def _name_module(name: str, lineno: int) -> str:
    for prefix, mod in PREFIX_MODULES:
        if name.startswith(prefix):
            return mod
    return _line_range_module(lineno)


def _node_name(node: ast.AST) -> str | None:
    if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
        return node.name
    if isinstance(node, ast.Assign):
        for t in node.targets:
            if isinstance(t, ast.Name):
                return t.id
    if isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
        return node.target.id
    return None


def main() -> None:
    source = CORE_PY.read_text(encoding="utf-8")
    lines = source.splitlines(keepends=True)
    tree = ast.parse(source)

    # Assign each top-level statement to a module
    segments: dict[str, list[str]] = {m: [] for m in MODULE_ORDER}

    for node in tree.body:
        # Skip original module docstring + import block (handled by STANDARD_HEADER)
        if node.lineno <= 84:
            continue
        if isinstance(node, ast.Expr) and isinstance(getattr(node, "value", None), ast.Constant):
            continue
        start = node.lineno - 1
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.decorator_list:
            start = min(d.lineno for d in node.decorator_list) - 1
        end = getattr(node, "end_lineno", node.lineno) or node.lineno
        chunk = "".join(lines[start:end])
        name = _node_name(node)
        if name:
            mod = _name_module(name, node.lineno)
        else:
            mod = _line_range_module(node.lineno)
        if mod not in segments:
            segments[mod] = []
        segments[mod].append(chunk)
        if not chunk.endswith("\n"):
            segments[mod].append("\n")

    CORE_PKG.mkdir(parents=True, exist_ok=True)

    # Write __init__.py
    init_doc = '''# -*- coding: utf-8 -*-
"""
planning_core.core — _core.py から機械分割した内部サブパッケージ。

外部からは ``planning_core._core``（ファサード）経由で import すること。
分割手順: 移動のみ・挙動変更なし。依存は下位→上位（state/columns → … → stage2_impl）。
"""
'''
    (CORE_PKG / "__init__.py").write_text(init_doc, encoding="utf-8")

    # 共通 import は _bootstrap.py に集約。各モジュールは本体のみ（exec 連結用）。
    bootstrap = STANDARD_HEADER.replace(
        '"""Extracted from planning_core._core (mechanical split)."""',
        '"""Shared imports for planning_core.core (exec bootstrap)."""',
    )
    (CORE_PKG / "_bootstrap.py").write_text(bootstrap, encoding="utf-8")

    for mod in MODULE_ORDER:
        body = "".join(segments.get(mod, []))
        body = re.sub(
            r"from \.(dispatch_workspace|input_resolution|plan_workbook_sidecar|"
            r"stage2_output_naming|bootstrap) import",
            r"from planning_core.\1 import",
            body,
        )
        header = f"# -*- coding: utf-8 -*-\n# planning_core.core.{mod} — body only (loaded via _core exec chain)\n"
        out = CORE_PKG / f"{mod}.py"
        out.write_text(header + body, encoding="utf-8")
        print(f"wrote {out.name}: {len((header + body).splitlines())} lines")

    order_str = ",\n    ".join(f'"{m}"' for m in MODULE_ORDER)
    facade = f'''# -*- coding: utf-8 -*-
"""planning_core 実装ファサード。実体は planning_core.core を exec 連結した共有名前空間。"""
from __future__ import annotations

import sys as _sys
import types as _types
from pathlib import Path as _Path

_CORE_DIR = _Path(__file__).resolve().parent / "core"
_MODULE_ORDER = [
    {order_str},
]

_ns = _sys.modules[__name__].__dict__
_ns["__name__"] = "planning_core._core"

def _exec_into_ns(filename: str) -> None:
    path = _CORE_DIR / filename
    code = path.read_text(encoding="utf-8")
    exec(compile(code, str(path), "exec"), _ns)  # noqa: S102

_exec_into_ns("_bootstrap.py")
for _part in _MODULE_ORDER:
    _exec_into_ns(f"{{_part}}.py")

# サブモジュール互換（import planning_core.core.roll_pipeline 等）
for _part in _MODULE_ORDER:
    _proxy = _types.ModuleType(f"planning_core.core.{{_part}}")
    _proxy.__dict__.update(_ns)
    _sys.modules[f"planning_core.core.{{_part}}"] = _proxy
'''
    backup = CORE_PY.with_suffix(".py.bak")
    if not backup.exists():
        backup.write_text(source, encoding="utf-8")
        print(f"backup: {backup}")
    CORE_PY.write_text(facade, encoding="utf-8")
    print(f"facade _core.py: {len(facade.splitlines())} lines")


if __name__ == "__main__":
    main()
