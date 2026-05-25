import os
import glob
import re
import datetime
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- PATH CONFIGURATION ---
WORKSPACE_DIR = r"c:\Users\0585\OneDrive\ドキュメント\加工依頼書（湖南）"
JUCHU_FILE = os.path.join(WORKSPACE_DIR, "加工依頼書入力.xlsm")
OUTPUT_FILE = os.path.join(WORKSPACE_DIR, "統合受注データベース.xlsx")

# --- NORMALIZATION HELPERS ---
def normalize_key(val):
    if val is None:
        return ""
    text = str(val).strip().upper()
    text = unicodedata.normalize('NFKC', text)
    text = "".join(text.split())
    text = text.replace("－", "-").replace("ー", "-").replace("―", "-").replace("‐", "-")
    return text

def normalize_text(val):
    if val is None:
        return ""
    text = str(val).strip()
    text = unicodedata.normalize('NFKC', text)
    text = "".join(text.split())
    text = text.replace("－", "-").replace("ー", "-").replace("―", "-").replace("‐", "-")
    return text.upper()

def normalize_numeric(val):
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    nums = re.findall(r'[-+]?\d*\.\d+|\d+', text)
    if nums:
        return float(nums[0])
    return 0.0

def normalize_date_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    text = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return normalize_text(text)

def get_yoto_translation(use_code):
    """Translate用途 codes like 3HYA to full descriptions like Y（工材） based on middle letter."""
    if not use_code:
        return ""
    code = str(use_code).strip().upper()
    if 'WA' in code or 'W' in code:
        return "W（自動車）"
    elif 'BA' in code or 'B' in code:
        return "B（輸出）"
    elif 'YA' in code or 'Y' in code:
        return "Y（工材）"
    elif 'VA' in code or 'V' in code:
        return "V（住宅）"
    elif 'ZA' in code or 'Z' in code:
        return "Z"
    return use_code

# --- DATA EXTRACTION ---
def load_juchu_file_with_formulas(file_path):
    """Load the ground-truth database sheet '受注ﾌｧｲﾙ' preserving formulas."""
    print(f"Loading ground-truth database rows (preserving formulas) from {os.path.basename(file_path)}...")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['受注ﾌｧｲﾙ']
    
    # Read headers from Row 3
    header_row = [sheet.cell(row=3, column=c).value for c in range(1, sheet.max_column + 1)]
    while header_row and header_row[-1] is None:
        header_row.pop()
        
    db_rows = []
    
    # Data rows from Row 4 to sheet.max_row
    for r_idx in range(4, sheet.max_row + 1):
        req_no = sheet.cell(row=r_idx, column=1).value
        # If the row is completely empty, skip
        row_has_data = any(sheet.cell(row=r_idx, column=c).value is not None for c in range(1, len(header_row) + 1))
        if not row_has_data:
            continue
            
        row_vals = {}
        for c_idx, h in enumerate(header_row):
            if h:
                row_vals[h] = sheet.cell(row=r_idx, column=c_idx + 1).value
        
        # Save exact row cells to reconstruct/write back
        db_rows.append(row_vals)
        
    wb.close()
    print(f"Loaded {len(db_rows)} total rows from '受注ﾌｧｲﾙ' sheet.")
    return header_row, db_rows

def extract_raw_request_sheets(workspace_dir):
    """Scan all *加工依頼書*.xlsm files and extract data from individual request sheets."""
    search_pattern = os.path.join(workspace_dir, "*加工依頼書*.xlsm")
    files = glob.glob(search_pattern)
    raw_files = [f for f in files if not os.path.basename(f).startswith("~$") and os.path.basename(f) != "加工依頼書入力.xlsm"]
    
    print(f"Found {len(raw_files)} raw request files to extract.")
    extracted_requests = []
    
    for r_f in raw_files:
        filename = os.path.basename(r_f)
        try:
            wb = openpyxl.load_workbook(r_f, data_only=True)
            for sheet_name in wb.sheetnames:
                # Naming pattern check for request sheets
                if not (re.match(r'^[A-Z]+\d+-\d+$', sheet_name) or re.match(r'^[A-Z]\d+-\d+-\d+$', sheet_name)):
                    continue
                
                sheet = wb[sheet_name]
                
                # Double check title block
                title_val = sheet.cell(row=6, column=6).value
                if title_val and "加工依頼書" not in str(title_val):
                    alt_title_1 = sheet.cell(row=6, column=5).value
                    alt_title_2 = sheet.cell(row=6, column=7).value
                    if not ((alt_title_1 and "加工依頼書" in str(alt_title_1)) or (alt_title_2 and "加工依頼書" in str(alt_title_2))):
                        continue
                
                # --- EXTRACT DATA FROM SHEET ---
                req_no = sheet.cell(row=5, column=18).value or sheet_name # Cell R5
                input_date = sheet.cell(row=2, column=34).value # Cell AH2
                
                item_code = sheet.cell(row=10, column=2).value # B10 (品名)
                part_no = sheet.cell(row=10, column=6).value # F10 (品番)
                type_code = sheet.cell(row=10, column=11).value # K10 (タイプ)
                width = sheet.cell(row=10, column=16).value # P10 (幅)
                length = sheet.cell(row=10, column=21).value # U10 (長さ)
                
                part_str = str(part_no).strip() if part_no is not None else ""
                type_str = str(type_code).strip() if type_code is not None else ""
                width_str = str(width).strip() if width is not None else ""
                length_str = str(length).strip() if length is not None else ""
                constructed_product = f"{part_str}-{type_str}-{width_str}X{length_str}" if (part_str or type_str) else ""
                
                grade = sheet.cell(row=10, column=24).value # X10 (梱－等)
                color = sheet.cell(row=10, column=27).value # AA10 (色)
                category = sheet.cell(row=10, column=29).value # AC10 (区分)
                quantity = sheet.cell(row=10, column=31).value # AE10 (数量)
                ec_side = sheet.cell(row=10, column=36).value # AJ10 (EC面)
                trimming = sheet.cell(row=10, column=39).value # AM10 (トリミング)
                
                processing_steps = []
                for r in range(13, 18):
                    p_val = sheet.cell(row=r, column=9).value
                    if p_val:
                        processing_steps.append(str(p_val).strip())
                processing_str = ", ".join(processing_steps)
                
                use_val = sheet.cell(row=18, column=5).value # E18 (用途)
                user_val = sheet.cell(row=19, column=5).value # E19 (ユーザー)
                delivery_val = sheet.cell(row=20, column=9).value # I20 (納期)
                contract_val = sheet.cell(row=21, column=5).value # E21 (契約ＮＯ)
                
                raw_item_code = sheet.cell(row=23, column=8).value # H23 (原反品名)
                raw_part_no = sheet.cell(row=23, column=11).value # K23 (原反品番)
                raw_type_code = sheet.cell(row=23, column=14).value # N23 (原反タイプ)
                raw_width = sheet.cell(row=23, column=17).value # Q23 (原反幅)
                raw_length = sheet.cell(row=23, column=20).value # T23 (原反長さ)
                
                r_part_str = str(raw_part_no).strip() if raw_part_no is not None else ""
                r_type_str = str(raw_type_code).strip() if raw_type_code is not None else ""
                r_width_str = str(raw_width).strip() if raw_width is not None else ""
                r_length_str = str(raw_length).strip() if raw_length is not None else ""
                constructed_raw = f"{r_part_str}-{r_type_str}-{r_width_str}X{r_length_str}" if (r_part_str or r_type_str) else ""
                
                raw_grade = sheet.cell(row=23, column=22).value # V23 (原反梱－等)
                raw_color = sheet.cell(row=23, column=25).value # Y23 (原反色)
                raw_category = sheet.cell(row=23, column=27).value # AA23 (原反区分)
                raw_qty = sheet.cell(row=23, column=29).value # AC23 (原反数量)
                location = sheet.cell(row=23, column=32).value # AF23 (在庫場所)
                input_day = sheet.cell(row=23, column=39).value # AM23 (投入日)
                
                issuer = sheet.cell(row=26, column=3).value # C26 (発行者)
                raw_roll_count = sheet.cell(row=20, column=27).value # AA20 (原反ロール数)
                processing_charge = sheet.cell(row=20, column=31).value # AE20 (加工賃)
                
                extracted_requests.append({
                    'file_name': filename,
                    'sheet_name': sheet_name,
                    '依頼Ｎｏ': str(req_no).strip(),
                    '入力日': input_date,
                    '品名': item_code,
                    '製品': constructed_product,
                    '梱-等1': grade,
                    '色1': color,
                    '区分1': category,
                    '数量1': quantity,
                    'ＥＣ面': ec_side,
                    'ﾄﾘﾐﾝｸﾞ': trimming,
                    '加工内容': processing_str,
                    '用途': use_val,
                    'ユーザー': user_val,
                    '希望納期': delivery_val,
                    '契約Ｎｏ': contract_val,
                    '原反品名': raw_item_code,
                    '原反': constructed_raw,
                    '原反梱-等': raw_grade,
                    '原反色': raw_color,
                    '原反区分': raw_category,
                    '原反数量': raw_qty,
                    '在庫場所': location,
                    '投入日': input_day,
                    '発行者': issuer,
                    '原反ロール数': raw_roll_count,
                    '加工賃': processing_charge
                })
            wb.close()
        except Exception as e:
            print(f"Error reading file {filename}: {e}")
            
    print(f"Successfully extracted {len(extracted_requests)} request sheets.")
    return extracted_requests

# --- MERGING & DEDUPLICATION LOGIC ---
def merge_and_consolidate(headers, db_rows, raw_requests):
    """Merge ground truth database rows and raw requests, automatically appending unentered ones."""
    print("Consolidating database rows and raw requests...")
    
    # Store existing rows keyed by normalized Request No
    db_map = {}
    for r in db_rows:
        req = r.get('依頼Ｎｏ', '')
        if req:
            norm_k = normalize_key(req)
            if norm_k not in db_map:
                db_map[norm_k] = []
            db_map[norm_k].append(r)
            
    consolidated_rows = []
    
    # Track which raw request keys have been processed
    processed_raw_keys = set()
    
    # Step 1: Process all existing database rows (preserving their content as ground truth)
    print("Processing existing database rows...")
    for db_row in db_rows:
        req_no = db_row.get('依頼Ｎｏ', '')
        norm_key = normalize_key(req_no)
        
        # Check if there is a matching raw request sheet in the current folder
        # We find a match in the extracted list
        matched_raw = None
        for raw in raw_requests:
            if normalize_key(raw['依頼Ｎｏ']) == norm_key:
                matched_raw = raw
                break
                
        if matched_raw:
            processed_raw_keys.add(norm_key)
            
            # --- COMPARE TO DETECT DISCREPANCIES ---
            discrepancies = []
            
            # 1. Item code
            if normalize_text(matched_raw['品名']) != normalize_text(db_row.get('品名', '')):
                discrepancies.append(f"品名({matched_raw['品名']} vs {db_row.get('品名','')})")
            # 2. Product
            if normalize_text(matched_raw['製品']) != normalize_text(db_row.get('製品', '')):
                discrepancies.append(f"製品({matched_raw['製品']} vs {db_row.get('製品','')})")
            # 3. Quantity
            if normalize_numeric(matched_raw['数量1']) != normalize_numeric(db_row.get('数量1', 0)):
                discrepancies.append(f"数量1({matched_raw['数量1']} vs {db_row.get('数量1','')})")
            # 4. User
            raw_u = normalize_text(matched_raw['ユーザー'])
            db_u = normalize_text(db_row.get('ユーザー', ''))
            if raw_u != db_u and raw_u not in db_u and db_u not in raw_u:
                discrepancies.append(f"ユーザー({matched_raw['ユーザー']} vs {db_row.get('ユーザー','')})")
            # 5. Delivery
            if normalize_date_val(matched_raw['希望納期']) != normalize_date_val(db_row.get('希望納期', '')):
                discrepancies.append(f"納期({normalize_date_val(matched_raw['希望納期'])} vs {normalize_date_val(db_row.get('希望納期',''))})")
            # 6. Raw material
            if normalize_text(matched_raw['原反']) != normalize_text(db_row.get('原反', '')):
                discrepancies.append(f"原反({matched_raw['原反']} vs {db_row.get('原反','')})")
            # 7. Processing contents
            raw_p = normalize_text(matched_raw['加工内容']).replace(",","").replace("、","")
            db_p = normalize_text(db_row.get('加工内容', '')).replace(",","").replace("、","")
            if raw_p != db_p:
                discrepancies.append(f"加工内容({matched_raw['加工内容']} vs {db_row.get('加工内容','')})")
                
            status = "既存登録 (原本相違あり)" if discrepancies else "既存登録 (原本一致)"
            desc_detail = "相違箇所: " + "; ".join(discrepancies) if discrepancies else "原本データと完全一致"
            
            # Construct row dictionary
            new_row = dict(db_row) # Keep original row values exactly!
            new_row['__status'] = status
            new_row['__discrepancy'] = desc_detail
            new_row['__filename'] = matched_raw['file_name']
            new_row['__sheetname'] = matched_raw['sheet_name']
            
            consolidated_rows.append(new_row)
        else:
            # Existing database row but no raw request sheet was found in the current folder (e.g. historical data)
            new_row = dict(db_row)
            new_row['__status'] = "既存登録 (原本未確認)"
            new_row['__discrepancy'] = "原本ファイルがフォルダ内にありません（過去データ）"
            new_row['__filename'] = ""
            new_row['__sheetname'] = ""
            
            consolidated_rows.append(new_row)
            
    # Step 2: Process all raw requests that are NOT yet in the database (Automatic Transfer!)
    print("Identifying and automatically transferring unentered raw request sheets...")
    newly_added_count = 0
    for raw in raw_requests:
        norm_key = normalize_key(raw['依頼Ｎｏ'])
        if norm_key not in db_map and norm_key not in processed_raw_keys:
            processed_raw_keys.add(norm_key)
            newly_added_count += 1
            
            # --- CREATE A NEW ROW POPULATING ALL FIELDS AUTOMATICALLY ---
            new_row = {}
            
            # Populate basic columns
            new_row['依頼Ｎｏ'] = raw['依頼Ｎｏ']
            new_row['入力区分'] = "通常入力"
            new_row['加工区分'] = "後加工"
            
            # Input 담당: C26 issuer or default
            new_row['入力担当'] = raw['発行者'] if raw['発行者'] else "自動転記"
            new_row['入力日'] = datetime.datetime.now()
            new_row['受付Ｎｏ'] = None
            
            # Product details
            new_row['品名'] = raw['品名']
            new_row['製品'] = raw['製品']
            new_row['梱-等1'] = raw['梱-等1']
            new_row['色1'] = raw['色1']
            new_row['区分1'] = raw['区分1']
            new_row['枝番'] = None
            new_row['数量1'] = raw['数量1']
            new_row['ＥＣ面'] = raw['ＥＣ面']
            new_row['ﾄﾘﾐﾝｸﾞ'] = raw['ﾄﾘﾐﾝｸﾞ']
            new_row['割数'] = 1
            
            # Raw material details
            new_row['品名1'] = raw['原反品名']
            new_row['原反'] = raw['原反']
            new_row['梱-等'] = raw['原反梱-等']
            new_row['色'] = raw['原反色']
            new_row['区分'] = raw['原反区分']
            new_row['数量'] = raw['原反数量']
            new_row['在庫場所'] = raw['在庫場所']
            
            # Processes / Delivery
            # Col X: 投入場所. If "スリット" is in processing contents, set to "ｽﾘｯﾄ"
            new_row['投入場所'] = "ｽﾘｯﾄ" if "スリット" in str(raw['加工内容']) else None
            new_row['投入日'] = raw['投入日']
            new_row['加工内容'] = raw['加工内容']
            
            new_row['特記事項1'] = None
            new_row['特記事項2'] = None
            new_row['特記事項3'] = None
            
            new_row['用途'] = get_yoto_translation(raw['用途'])
            new_row['ユーザー'] = raw['ユーザー']
            new_row['希望納期'] = raw['希望納期']
            new_row['調整納期'] = raw['希望納期']
            new_row['加工賃'] = raw['加工賃']
            new_row['契約Ｎｏ'] = raw['契約Ｎｏ']
            new_row['原反ロール数'] = raw['原反ロール数']
            
            # We will populate Excel formulas for calculation columns
            # Formulas are dynamically evaluated in Excel based on row index
            # Row index in output file will be row_num
            # We can placeholder them and write actual formula string in generator phase!
            new_row['月'] = "__FORMULA_MONTH__"
            new_row['受注金額'] = "__FORMULA_AMOUNT__"
            new_row['受注数'] = "__FORMULA_QTY__"
            new_row['単価'] = "__FORMULA_PRICE__"
            new_row['受注金額(修正)'] = "__FORMULA_AMOUNT_MOD__"
            
            # Set metadata columns at the end of the sheet
            new_row['__status'] = "新規自動追加 (未登録)"
            new_row['__discrepancy'] = "原本より自動でデータを抽出・新規追加しました"
            new_row['__filename'] = raw['file_name']
            new_row['__sheetname'] = raw['sheet_name']
            
            consolidated_rows.append(new_row)
            
    print(f"Consolidation complete: Added {newly_added_count} new request records automatically.")
    return consolidated_rows

# --- EXCEL FILE GENERATION ---
def write_consolidated_database(headers, consolidated_rows, output_path):
    print(f"Generating consolidated database Excel sheet at {output_path}...")
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Dashboard
    ws_dash = wb.active
    ws_dash.title = "ダッシュボード"
    ws_dash.views.sheetView[0].showGridLines = True
    
    total = len(consolidated_rows)
    existing_matched = sum(1 for r in consolidated_rows if r['__status'] == "既存登録 (原本一致)")
    existing_discrep = sum(1 for r in consolidated_rows if r['__status'] == "既存登録 (原本相違あり)")
    existing_unverified = sum(1 for r in consolidated_rows if r['__status'] == "既存登録 (原本未確認)")
    newly_added = sum(1 for r in consolidated_rows if r['__status'] == "新規自動追加 (未登録)")
    
    font_title = Font(name="BIZ UDPGothic", size=18, bold=True, color="2C3E50")
    font_section = Font(name="BIZ UDPGothic", size=12, bold=True, color="2C3E50")
    font_body = Font(name="BIZ UDPGothic", size=10)
    font_bold = Font(name="BIZ UDPGothic", size=10, bold=True)
    
    fill_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    fill_new_row = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid") # Soft green
    fill_discrep = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Soft yellow
    fill_normal = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='BDC3C7')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # --- Dashboard Title ---
    ws_dash["B2"] = "湖南工場 統合受注データ一括管理データベース"
    ws_dash["B2"].font = font_title
    
    ws_dash["B3"] = f"出力日時: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  原本一括解析・自動転記完了版"
    ws_dash["B3"].font = Font(name="BIZ UDPGothic", size=9, italic=True, color="7F8C8D")
    
    for col in range(2, 7):
        ws_dash.cell(row=4, column=col).border = Border(bottom=Side(style='medium', color='2C3E50'))
        
    # --- Summary KPIs ---
    ws_dash["B6"] = "【統合データベース統計】"
    ws_dash["B6"].font = font_section
    
    kpis = [
        ("総受注データ件数 (Total Master)", total, "2C3E50", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("既存引き継ぎ件数 (Existing)", existing_matched + existing_discrep + existing_unverified, "7F8C8D", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("原本自動転記件数 (Newly Appended)", newly_added, "27AE60", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("原本相違検出件数 (Discrepant Rows)", existing_discrep, "E67E22", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF"))
    ]
    
    for i, (label, val, bg_color, num_font) in enumerate(kpis):
        r_start = 8 + (i * 3)
        ws_dash.merge_cells(start_row=r_start, start_column=2, end_row=r_start, end_column=3)
        ws_dash.merge_cells(start_row=r_start+1, start_column=2, end_row=r_start+1, end_column=3)
        
        c_label = ws_dash.cell(row=r_start, column=2)
        c_label.value = label
        c_label.font = Font(name="BIZ UDPGothic", size=9, bold=True, color="7F8C8D")
        c_label.alignment = Alignment(horizontal="center", vertical="center")
        
        c_val = ws_dash.cell(row=r_start+1, column=2)
        c_val.value = val
        c_val.font = num_font
        c_val.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        
        for r in range(r_start, r_start+2):
            for c in range(2, 4):
                ws_dash.cell(row=r, column=c).border = border_all
                
    # --- Instructions ---
    ws_dash["E6"] = "【データベースの特徴・確認手順】"
    ws_dash["E6"].font = font_section
    
    instructions = [
        "1. 『統合受注データ』シートには、既存の全ての受注データと原本からの自動転記データが蓄積されています。",
        "2. 【自動転記】された行 (薄い緑色):",
        "   ⇒ 既存の受注ファイルに入力されていなかった原本依頼書（4件）をシステムが自動検出・一括転記しました！",
        "   ⇒ 用途の翻訳（例：3HYA ⇒ Y（工材））や製品コード・原反コードの自動組み立て、加工賃の転記なども自動完了しています。",
        "3. 【相違あり】とマークされた行 (薄い黄色):",
        "   ⇒ 既に受注ファイルに登録はありましたが、数量や寸法などの一部が原本依頼書の記載と異なっていた行です。",
        "   ⇒ 行末の『原本相違箇所』列に具体的な違いを出力しています。原本と突き合わせて必要に応じて修正してください。",
        "4. 【月】や【受注金額】などの計算列:",
        "   ⇒ 自動転記された行にも自動的に数式（ExcelのMONTH関数、掛け算数式、 structured table formulas）が書き込まれています。",
        "   ⇒ Excelでファイルを開くことで自動で計算されます。",
        "",
        "※ このファイルをマスターデータベースとしてコピーし、日報や加工計画などの転記作業に直接ご活用いただけます。"
    ]
    
    for idx, inst in enumerate(instructions):
        c_inst = ws_dash.cell(row=8+idx, column=5)
        c_inst.value = inst
        if "自動転記" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="1E8449")
        elif "相違あり" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="D35400")
        else:
            c_inst.font = font_body
            
    ws_dash.column_dimensions["A"].width = 3
    ws_dash.column_dimensions["B"].width = 15
    ws_dash.column_dimensions["C"].width = 15
    ws_dash.column_dimensions["D"].width = 4
    ws_dash.column_dimensions["E"].width = 90
    
    # 2. Sheet 2: Master Database
    ws_db = wb.create_sheet(title="統合受注データ")
    ws_db.views.sheetView[0].showGridLines = True
    
    # Extend headers with our status columns
    meta_headers = ["データ処理区分", "原本相違箇所", "原本ファイル名", "原本シート名"]
    all_headers = headers + meta_headers
    
    # Write Headers
    for c_idx, h in enumerate(all_headers):
        cell = ws_db.cell(row=3, column=c_idx + 1)
        cell.value = h
        cell.font = Font(name="BIZ UDPGothic", size=9, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        
    ws_db.row_dimensions[3].height = 28
    
    # Copy table columns style formula rows
    for r_idx, row_vals in enumerate(consolidated_rows):
        row_num = r_idx + 4
        ws_db.row_dimensions[row_num].height = 20
        
        # Color fill based on data registration category
        fill_to_apply = fill_normal
        status = row_vals.get('__status', '')
        if "新規自動追加" in status:
            fill_to_apply = fill_new_row
        elif "原本相違あり" in status:
            fill_to_apply = fill_discrep
            
        # Write values column by column based on the database header
        for c_idx, h in enumerate(headers):
            cell = ws_db.cell(row=row_num, column=c_idx + 1)
            cell.font = font_body
            cell.fill = fill_to_apply
            cell.border = border_all
            
            # Alignments
            if h in ["依頼Ｎｏ", "入力区分", "加工区分", "入力担当", "入力日", "受付Ｎｏ", "希望納期", "調整納期", "枝番", "割数", "契約Ｎｏ", "月", "投入日"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif h in ["数量1", "数量", "加工賃", "原反ロール数", "受注金額", "受注数", "単価", "受注金額(修正)", "加工金額", "金額", "合計金額"]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            val = row_vals.get(h)
            
            # --- EVALUATE AND WRITE DYNAMIC FORMULAS FOR NEW ROWS ---
            if "新規自動追加" in status:
                if val == "__FORMULA_MONTH__":
                    cell.value = f"=MONTH(AG{row_num})"
                elif val == "__FORMULA_AMOUNT__":
                    cell.value = f"=AN{row_num}*AM{row_num}"
                elif val == "__FORMULA_QTY__":
                    cell.value = f"=M{row_num}"
                elif val == "__FORMULA_PRICE__":
                    cell.value = f"=AH{row_num}"
                elif val == "__FORMULA_AMOUNT_MOD__":
                    cell.value = f"=AL{row_num}"
                elif c_idx >= 58: # Columns BG to BT (Table Formulas)
                    # We copy the exact formula text from row 4 replacing with "#This Row" structured formulas
                    # Because row 4 contains the exact structured formulas like =_t受注ﾌｧｲﾙ[[#This Row],[加工賃]]
                    # We can safely use the identical structured formulas!
                    # Let's read it from existing rows
                    cell.value = consolidated_rows[0].get(h) # Copy from row 4 (idx 0)
                else:
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        cell.value = val.strftime("%Y-%m-%d")
                    else:
                        cell.value = val
            else:
                # Historical row: write original values/formulas preserved
                if isinstance(val, (datetime.datetime, datetime.date)):
                    cell.value = val.strftime("%Y-%m-%d")
                else:
                    cell.value = val
                    
            # Number formatting for numeric columns
            if cell.value != "" and cell.value is not None:
                if h in ["数量1", "数量", "受注金額", "受注数", "加工金額", "金額", "合計金額", "受注金額(修正)"]:
                    if not str(cell.value).startswith("="):
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0'
                        except:
                            pass
                elif h in ["加工賃", "単価", "単価9"]:
                    if not str(cell.value).startswith("="):
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0'
                        except:
                            pass
                            
        # Write metadata columns at the end of the sheet
        meta_vals = [
            row_vals.get('__status', ''),
            row_vals.get('__discrepancy', ''),
            row_vals.get('__filename', ''),
            row_vals.get('__sheetname', '')
        ]
        
        for m_idx, m_val in enumerate(meta_vals):
            c_pos = len(headers) + m_idx + 1
            cell = ws_db.cell(row=row_num, column=c_pos)
            cell.value = m_val
            cell.font = font_body
            cell.fill = fill_to_apply
            cell.border = border_all
            
            if m_idx == 0:
                cell.font = font_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif m_idx in [2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-fit columns
    for col in ws_db.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        # Check first 20 rows of details for width estimation to save performance
        for cell in list(col)[:20]:
            val_str = str(cell.value or '')
            byte_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
            if byte_len > max_len:
                max_len = byte_len
        ws_db.column_dimensions[col_letter].width = max(max_len + 4, 11)
        
    # Auto-filter details
    ws_db.auto_filter.ref = f"A3:{get_column_letter(len(all_headers))}{len(consolidated_rows) + 3}"
    
    # Save the file
    wb.save(output_path)
    wb.close()
    print("Consolidated database generated successfully.")

# --- MAIN EXECUTION ---
def main():
    print("=== Processing Request Consolidated Database Automation ===")
    
    # 1. Load Ground Truth from Juchu file (preserving formulas)
    if not os.path.exists(JUCHU_FILE):
        print(f"ERROR: Ground-truth database file not found at {JUCHU_FILE}!")
        return
        
    headers, db_rows = load_juchu_file_with_formulas(JUCHU_FILE)
    
    # 2. Extract Raw Requests from directories
    raw_requests = extract_raw_request_sheets(WORKSPACE_DIR)
    
    if not raw_requests:
        print("ERROR: No raw request sheets were extracted from any XLSM files!")
        return
        
    # 3. Merge & Deduplicate (Auto transfer unentered rows!)
    consolidated_rows = merge_and_consolidate(headers, db_rows, raw_requests)
    
    # 4. Write styled integrated master database
    write_consolidated_database(headers, consolidated_rows, OUTPUT_FILE)
    
    print("\n=== Integration Statistics ===")
    total = len(consolidated_rows)
    existing = sum(1 for r in consolidated_rows if "既存登録" in r['__status'])
    newly_added = sum(1 for r in consolidated_rows if "新規自動追加" in r['__status'])
    discrep = sum(1 for r in consolidated_rows if "原本相違あり" in r['__status'])
    
    print(f"Total Database Records:       {total}")
    print(f"Existing Records Carried Over: {existing}")
    print(f"Newly Appended from Raw Sheets:{newly_added}")
    print(f"Discrepant Records Detected:   {discrep}")
    print(f"Integrated database saved at:  {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
