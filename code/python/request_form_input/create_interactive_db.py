import os
import glob
import re
import datetime
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- PATH CONFIGURATION ---
WORKSPACE_DIR = r"c:\Users\0585\OneDrive\ドキュメント\加工依頼書（湖南）"
JUCHU_FILE = os.path.join(WORKSPACE_DIR, "加工依頼書入力.xlsm")
OUTPUT_FILE = os.path.join(WORKSPACE_DIR, "統合受注データベース_交互確認版.xlsx")

# --- NORMALIZATION HELPERS ---
def normalize_key(val):
    if val is None:
        return ""
    text = str(val).strip().upper()
    text = unicodedata.normalize('NFKC', text)
    text = "".join(text.split())
    text = text.replace("－", "-").replace("ー", "-").replace("―", "-").replace("‐", "-")
    return text

def normalize_text(val):
    if val is None:
        return ""
    text = str(val).strip()
    text = unicodedata.normalize('NFKC', text)
    text = "".join(text.split())
    text = text.replace("－", "-").replace("ー", "-").replace("―", "-").replace("‐", "-")
    return text.upper()

def normalize_numeric(val):
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    nums = re.findall(r'[-+]?\d*\.\d+|\d+', text)
    if nums:
        return float(nums[0])
    return 0.0

def normalize_date_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    text = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return normalize_text(text)

def get_yoto_translation(use_code):
    if not use_code:
        return ""
    code = str(use_code).strip().upper()
    if 'WA' in code or 'W' in code:
        return "W（自動車）"
    elif 'BA' in code or 'B' in code:
        return "B（輸出）"
    elif 'YA' in code or 'Y' in code:
        return "Y（工材）"
    elif 'VA' in code or 'V' in code:
        return "V（住宅）"
    elif 'ZA' in code or 'Z' in code:
        return "Z"
    return use_code

# --- CELL-BY-CELL WORKSHEET COPY ENGINE ---
def copy_sheet_contents(src_sheet, dst_sheet):
    """Accurately copy cell values, formulas, merged cell ranges, fonts, fills, borders, alignments, and dimensions."""
    # 1. Copy cell values and styles first (to normal cells, skipping source MergedCells)
    for row in src_sheet.iter_rows():
        for cell in row:
            if type(cell).__name__ == 'MergedCell':
                continue
            dst_row = cell.row + 1 # Shifted down by 1
            dst_col = cell.column
            dst_cell = dst_sheet.cell(row=dst_row, column=dst_col)
            
            dst_cell.value = cell.value
            
            # Copy Font
            if cell.font:
                dst_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                    underline=cell.font.underline
                )
            # Copy Fill
            if cell.fill and cell.fill.fill_type:
                dst_cell.fill = PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                )
            # Copy Border
            if cell.border:
                dst_cell.border = Border(
                    left=Side(style=cell.border.left.style, color=cell.border.left.color) if cell.border.left and cell.border.left.style else None,
                    right=Side(style=cell.border.right.style, color=cell.border.right.color) if cell.border.right and cell.border.right.style else None,
                    top=Side(style=cell.border.top.style, color=cell.border.top.color) if cell.border.top and cell.border.top.style else None,
                    bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color) if cell.border.bottom and cell.border.bottom.style else None
                )
            # Copy Alignment
            if cell.alignment:
                dst_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                )
                
    # 2. Copy merged cell ranges AFTER values are written
    for merged_range in src_sheet.merged_cells.ranges:
        dst_sheet.merge_cells(
            start_row=merged_range.min_row + 1, # Shifted by 1 for the back link row
            start_column=merged_range.min_col,
            end_row=merged_range.max_row + 1,
            end_column=merged_range.max_col
        )
        
    # 3. Copy Column Widths
    for col in src_sheet.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in src_sheet.column_dimensions:
            dst_sheet.column_dimensions[col_letter].width = src_sheet.column_dimensions[col_letter].width

# --- DATA EXTRACTION ---
def load_juchu_file_data(file_path):
    print(f"Loading ground-truth rows from {os.path.basename(file_path)}...")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['受注ﾌｧｲﾙ']
    
    header_row = [sheet.cell(row=3, column=c).value for c in range(1, sheet.max_column + 1)]
    while header_row and header_row[-1] is None:
        header_row.pop()
        
    db_rows = []
    for r_idx in range(4, sheet.max_row + 1):
        req_no = sheet.cell(row=r_idx, column=1).value
        row_has_data = any(sheet.cell(row=r_idx, column=c).value is not None for c in range(1, len(header_row) + 1))
        if not row_has_data:
            continue
            
        row_vals = {}
        for c_idx, h in enumerate(header_row):
            if h:
                row_vals[h] = sheet.cell(row=r_idx, column=c_idx + 1).value
        db_rows.append(row_vals)
        
    wb.close()
    return header_row, db_rows

def extract_and_copy_raw_sheets(workspace_dir, wb_dst):
    """Scan all *加工依頼書*.xlsm files, extract raw data, and copy worksheets directly into the destination workbook."""
    search_pattern = os.path.join(workspace_dir, "*加工依頼書*.xlsm")
    files = glob.glob(search_pattern)
    raw_files = [f for f in files if not os.path.basename(f).startswith("~$") and os.path.basename(f) != "加工依頼書入力.xlsm"]
    
    extracted_requests = []
    
    for r_f in raw_files:
        filename = os.path.basename(r_f)
        try:
            wb_src = openpyxl.load_workbook(r_f, data_only=True)
            for sheet_name in wb_src.sheetnames:
                # Strictly match request sheet names
                if not (re.match(r'^[A-Z]+\d+-\d+$', sheet_name) or re.match(r'^[A-Z]\d+-\d+-\d+$', sheet_name)):
                    continue
                
                sheet_src = wb_src[sheet_name]
                
                # Double check title
                title_val = sheet_src.cell(row=6, column=6).value
                if title_val and "加工依頼書" not in str(title_val):
                    alt_1 = sheet_src.cell(row=6, column=5).value
                    alt_2 = sheet_src.cell(row=6, column=7).value
                    if not ((alt_1 and "加工依頼書" in str(alt_1)) or (alt_2 and "加工依頼書" in str(alt_2))):
                        continue
                
                print(f"  Copying request sheet '{sheet_name}' from {filename}...")
                
                # Create sheet in destination workbook and copy contents
                sheet_dst = wb_dst.create_sheet(title=sheet_name)
                sheet_dst.views.sheetView[0].showGridLines = True
                copy_sheet_contents(sheet_src, sheet_dst)
                
                # --- EXTRACT KEY FIELDS FOR DATABASE ---
                req_no = sheet_src.cell(row=5, column=18).value or sheet_name
                input_date = sheet_src.cell(row=2, column=34).value
                item_code = sheet_src.cell(row=10, column=2).value
                part_no = sheet_src.cell(row=10, column=6).value
                type_code = sheet_src.cell(row=10, column=11).value
                width = sheet_src.cell(row=10, column=16).value
                length = sheet_src.cell(row=10, column=21).value
                
                part_str = str(part_no).strip() if part_no is not None else ""
                type_str = str(type_code).strip() if type_code is not None else ""
                width_str = str(width).strip() if width is not None else ""
                length_str = str(length).strip() if length is not None else ""
                constructed_product = f"{part_str}-{type_str}-{width_str}X{length_str}" if (part_str or type_str) else ""
                
                grade = sheet_src.cell(row=10, column=24).value
                color = sheet_src.cell(row=10, column=27).value
                category = sheet_src.cell(row=10, column=29).value
                quantity = sheet_src.cell(row=10, column=31).value
                ec_side = sheet_src.cell(row=10, column=36).value
                trimming = sheet_src.cell(row=10, column=39).value
                
                processing_steps = []
                for r in range(13, 18):
                    p_val = sheet_src.cell(row=r, column=9).value
                    if p_val:
                        processing_steps.append(str(p_val).strip())
                processing_str = ", ".join(processing_steps)
                
                use_val = sheet_src.cell(row=18, column=5).value
                user_val = sheet_src.cell(row=19, column=5).value
                delivery_val = sheet_src.cell(row=20, column=9).value
                contract_val = sheet_src.cell(row=21, column=5).value
                
                raw_item_code = sheet_src.cell(row=23, column=8).value
                raw_part_no = sheet_src.cell(row=23, column=11).value
                raw_type_code = sheet_src.cell(row=23, column=14).value
                raw_width = sheet_src.cell(row=23, column=17).value
                raw_length = sheet_src.cell(row=23, column=20).value
                
                r_part_str = str(raw_part_no).strip() if raw_part_no is not None else ""
                r_type_str = str(raw_type_code).strip() if raw_type_code is not None else ""
                r_width_str = str(raw_width).strip() if raw_width is not None else ""
                r_length_str = str(raw_length).strip() if raw_length is not None else ""
                constructed_raw = f"{r_part_str}-{r_type_str}-{r_width_str}X{r_length_str}" if (r_part_str or r_type_str) else ""
                
                raw_grade = sheet_src.cell(row=23, column=22).value
                raw_color = sheet_src.cell(row=23, column=25).value
                raw_category = sheet_src.cell(row=23, column=27).value
                raw_qty = sheet_src.cell(row=23, column=29).value
                location = sheet_src.cell(row=23, column=32).value
                input_day = sheet_src.cell(row=23, column=39).value
                
                issuer = sheet_src.cell(row=26, column=3).value
                raw_roll_count = sheet_src.cell(row=20, column=27).value
                processing_charge = sheet_src.cell(row=20, column=31).value
                
                extracted_requests.append({
                    'file_name': filename,
                    'sheet_name': sheet_name,
                    '依頼Ｎｏ': str(req_no).strip(),
                    '入力日': input_date,
                    '品名': item_code,
                    '製品': constructed_product,
                    '梱-等1': grade,
                    '色1': color,
                    '区分1': category,
                    '数量1': quantity,
                    'ＥＣ面': ec_side,
                    'ﾄﾘﾐﾝｸﾞ': trimming,
                    '加工内容': processing_str,
                    '用途': use_val,
                    'ユーザー': user_val,
                    '希望納期': delivery_val,
                    '契約Ｎｏ': contract_val,
                    '原反品名': raw_item_code,
                    '原反': constructed_raw,
                    '原反梱-等': raw_grade,
                    '原反色': raw_color,
                    '原反区分': raw_category,
                    '原反数量': raw_qty,
                    '在庫場所': location,
                    '投入日': input_day,
                    '発行者': issuer,
                    '原反ロール数': raw_roll_count,
                    '加工賃': processing_charge,
                    '__dst_sheet_obj': sheet_dst # Save destination sheet reference to add back-link later
                })
            wb_src.close()
        except Exception as e:
            print(f"Error copying sheets from file {filename}: {e}")
            
    return extracted_requests

# --- MERGE & SIDE-BY-SIDE BUILDER ---
def merge_and_build_interactive(db_rows, raw_requests):
    db_map = {}
    for r in db_rows:
        req = r.get('依頼Ｎｏ', '')
        if req:
            norm_k = normalize_key(req)
            if norm_k not in db_map:
                db_map[norm_k] = []
            db_map[norm_k].append(r)
            
    consolidated_rows = []
    processed_raw_keys = set()
    
    # 1. Load Juchu rows
    for db_row in db_rows:
        req_no = db_row.get('依頼Ｎｏ', '')
        norm_key = normalize_key(req_no)
        
        matched_raw = next((r for r in raw_requests if normalize_key(r['依頼Ｎｏ']) == norm_key), None)
        
        if matched_raw:
            processed_raw_keys.add(norm_key)
            discrepancies = []
            
            if normalize_text(matched_raw['品名']) != normalize_text(db_row.get('品名', '')):
                discrepancies.append(f"品名({matched_raw['品名']} vs {db_row.get('品名','')})")
            if normalize_text(matched_raw['製品']) != normalize_text(db_row.get('製品', '')):
                discrepancies.append(f"製品({matched_raw['製品']} vs {db_row.get('製品','')})")
            if normalize_numeric(matched_raw['数量1']) != normalize_numeric(db_row.get('数量1', 0)):
                discrepancies.append(f"数量1({matched_raw['数量1']} vs {db_row.get('数量1','')})")
            raw_u = normalize_text(matched_raw['ユーザー'])
            db_u = normalize_text(db_row.get('ユーザー', ''))
            if raw_u != db_u and raw_u not in db_u and db_u not in raw_u:
                discrepancies.append(f"ユーザー({matched_raw['ユーザー']} vs {db_row.get('ユーザー','')})")
            if normalize_date_val(matched_raw['希望納期']) != normalize_date_val(db_row.get('希望納期', '')):
                discrepancies.append(f"納期({normalize_date_val(matched_raw['希望納期'])} vs {normalize_date_val(db_row.get('希望納期',''))})")
            if normalize_text(matched_raw['原反']) != normalize_text(db_row.get('原反', '')):
                discrepancies.append(f"原反({matched_raw['原反']} vs {db_row.get('原反','')})")
            raw_p = normalize_text(matched_raw['加工内容']).replace(",","").replace("、","")
            db_p = normalize_text(db_row.get('加工内容', '')).replace(",","").replace("、","")
            if raw_p != db_p:
                discrepancies.append(f"加工内容({matched_raw['加工内容']} vs {db_row.get('加工内容','')})")
            if normalize_text(matched_raw['契約Ｎｏ']) != normalize_text(db_row.get('契約Ｎｏ', '')):
                if matched_raw['契約Ｎｏ'] or db_row.get('契約Ｎｏ'):
                    discrepancies.append(f"契約No({matched_raw['契約Ｎｏ']} vs {db_row.get('契約Ｎｏ','')})")
                    
            status = "既存登録 (原本相違あり)" if discrepancies else "既存登録 (原本一致)"
            desc_detail = "相違箇所: " + "; ".join(discrepancies) if discrepancies else "原本と完全一致"
            
            new_row = dict(db_row)
            new_row['__status'] = status
            new_row['__discrepancy'] = desc_detail
            new_row['__raw_data'] = matched_raw # Attach raw data dictionary
            consolidated_rows.append(new_row)
        else:
            new_row = dict(db_row)
            new_row['__status'] = "既存登録 (原本未確認)"
            new_row['__discrepancy'] = "原本ファイルが未検出（過去データ）"
            new_row['__raw_data'] = None
            consolidated_rows.append(new_row)
            
    # 2. Append unentered raw sheets (Auto Transfer!)
    for raw in raw_requests:
        norm_key = normalize_key(raw['依頼Ｎｏ'])
        if norm_key not in db_map and norm_key not in processed_raw_keys:
            processed_raw_keys.add(norm_key)
            
            new_row = {}
            new_row['依頼Ｎｏ'] = raw['依頼Ｎｏ']
            new_row['入力区分'] = "通常入力"
            new_row['加工区分'] = "後加工"
            new_row['入力担当'] = raw['発行者'] if raw['発行者'] else "自動転記"
            new_row['入力日'] = datetime.datetime.now()
            new_row['受付Ｎｏ'] = None
            new_row['品名'] = raw['品名']
            new_row['製品'] = raw['製品']
            new_row['梱-等1'] = raw['梱-等1']
            new_row['色1'] = raw['色1']
            new_row['区分1'] = raw['区分1']
            new_row['枝番'] = None
            new_row['数量1'] = raw['数量1']
            new_row['ＥＣ面'] = raw['ＥＣ面']
            new_row['ﾄﾘﾐﾝｸﾞ'] = raw['ﾄﾘﾐﾝｸﾞ']
            new_row['割数'] = 1
            new_row['品名1'] = raw['原反品名']
            new_row['原反'] = raw['原反']
            new_row['梱-等'] = raw['原反梱-等']
            new_row['色'] = raw['原反色']
            new_row['区分'] = raw['原反区分']
            new_row['数量'] = raw['原反数量']
            new_row['在庫場所'] = raw['在庫場所']
            new_row['投入場所'] = "ｽﾘｯﾄ" if "スリット" in str(raw['加工内容']) else None
            new_row['投入日'] = raw['投入日']
            new_row['加工内容'] = raw['加工内容']
            new_row['特記事項1'] = None
            new_row['特記事項2'] = None
            new_row['特記事項3'] = None
            new_row['用途'] = get_yoto_translation(raw['用途'])
            new_row['ユーザー'] = raw['ユーザー']
            new_row['希望納期'] = raw['希望納期']
            new_row['調整納期'] = raw['希望納期']
            new_row['加工賃'] = raw['加工賃']
            new_row['契約Ｎｏ'] = raw['契約Ｎｏ']
            new_row['原反ロール数'] = raw['原反ロール数']
            new_row['月'] = "__FORMULA_MONTH__"
            new_row['受注金額'] = "__FORMULA_AMOUNT__"
            new_row['受注数'] = "__FORMULA_QTY__"
            new_row['単価'] = "__FORMULA_PRICE__"
            new_row['受注金額(修正)'] = "__FORMULA_AMOUNT_MOD__"
            
            new_row['__status'] = "新規自動追加 (未登録)"
            new_row['__discrepancy'] = "原本より自動追加しました"
            new_row['__raw_data'] = raw
            consolidated_rows.append(new_row)
            
    return consolidated_rows

# --- DYNAMIC EXCEL WRITER ---
def generate_interactive_workbook(headers, consolidated_rows, output_path):
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Dashboard
    ws_dash = wb.active
    ws_dash.title = "ダッシュボード"
    ws_dash.views.sheetView[0].showGridLines = True
    
    total = len(consolidated_rows)
    existing_matched = sum(1 for r in consolidated_rows if r['__status'] == "既存登録 (原本一致)")
    existing_discrep = sum(1 for r in consolidated_rows if r['__status'] == "既存登録 (原本相違あり)")
    existing_unverified = sum(1 for r in consolidated_rows if r['__status'] == "既存登録 (原本未確認)")
    newly_added = sum(1 for r in consolidated_rows if r['__status'] == "新規自動追加 (未登録)")
    
    font_title = Font(name="BIZ UDPGothic", size=18, bold=True, color="2C3E50")
    font_section = Font(name="BIZ UDPGothic", size=12, bold=True, color="2C3E50")
    font_body = Font(name="BIZ UDPGothic", size=10)
    font_bold = Font(name="BIZ UDPGothic", size=10, bold=True)
    
    thin_border = Border(left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'),
                         top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'))
    
    ws_dash["B2"] = "湖南工場 受注一括照合・対比型入力支援データベース"
    ws_dash["B2"].font = font_title
    ws_dash["B3"] = f"出力日時: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  「左に受注入力 ⇔ 右に原本対比 ＆ リンクジャンプ機能」搭載"
    ws_dash["B3"].font = Font(name="BIZ UDPGothic", size=9, italic=True, color="7F8C8D")
    
    for col in range(2, 7):
        ws_dash.cell(row=4, column=col).border = Border(bottom=Side(style='medium', color='2C3E50'))
        
    # KPIs
    ws_dash["B6"] = "【照合・登録集計状況】"
    ws_dash["B6"].font = font_section
    
    kpis = [
        ("総受注データ件数 (Combined Total)", total, "2C3E50", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("原本自動転記件数 (Newly Appended)", newly_added, "27AE60", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("データ相違検出 (Discrepant Rows)", existing_discrep, "E67E22", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF"))
    ]
    
    for i, (label, val, bg_color, num_font) in enumerate(kpis):
        r_start = 8 + (i * 3)
        ws_dash.merge_cells(start_row=r_start, start_column=2, end_row=r_start, end_column=3)
        ws_dash.merge_cells(start_row=r_start+1, start_column=2, end_row=r_start+1, end_column=3)
        ws_dash.cell(row=r_start, column=2, value=label).font = Font(name="BIZ UDPGothic", size=9, bold=True, color="7F8C8D")
        ws_dash.cell(row=r_start, column=2).alignment = Alignment(horizontal="center", vertical="center")
        
        c_val = ws_dash.cell(row=r_start+1, column=2, value=val)
        c_val.font = num_font
        c_val.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        
        for r in range(r_start, r_start+2):
            for c in range(2, 4):
                ws_dash.cell(row=r, column=c).border = thin_border
                
    # Instructions
    ws_dash["E6"] = "【本ブックの特徴および二画面対比（スプリット）での確認方法】"
    ws_dash["E6"].font = font_section
    
    instructions = [
        "★ 本ブックの最も強力な機能：原本ジャンプリンク ＆ 左右二画面対比 ★",
        "",
        "1. 左右の画面に分割して並べて表示する方法 (非常におすすめです！):",
        "   ① Excelの上のメニューから『表示』タブをクリックします。",
        "   ② 『新しいウィンドウを開く』をクリックします。(同じファイルがもう1つ開きます)",
        "   ③ 再度『表示』タブから『整列』をクリックし、『垂直（または左右に並べて表示）』を選択してOKを押します。",
        "   ④ 【左側の画面】で『統合受注入力・原本比較』シートを表示します。",
        "   ⑤ 【右側の画面】で任意の原本シート（例：『E5-4』）を表示させます。",
        "   ⇒ これにより、左画面でデータベースを入力・手修正しながら、右画面で赤文字や注記をいつでも見比べられます！",
        "",
        "2. クリック一つで原本へ瞬間移動（ジャンプリンク機能）:",
        "   ・『統合受注入力・原本比較』シートの「原本シート」列にあるリンクをクリックすると、",
        "     このブック内に丸ごとコピーされた対象の原本シートに瞬時にジャンプします。",
        "   ・原本シートの最上部には、緑色で『◀ 統合データ照合表に戻る』というリンクがあり、",
        "     クリックすると、さっき編集していたデータベースの行に自動的に戻ります！",
        "",
        "3. シート構成:",
        "   ・『統合受注入力・原本比較』シート: 左半分が『手修正・最終データ』、右半分が『原本の抽出値』になっています。",
        "   ・『C5-1』『E5-4』『W5-21』などの各シート: 実際の原本ファイルをシートごと再現したものです（赤文字の指示もそのまま残っています）。"
    ]
    
    for idx, inst in enumerate(instructions):
        c_inst = ws_dash.cell(row=8+idx, column=5, value=inst)
        if "★" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=11, bold=True, color="2C3E50")
        elif "自動転記" in inst or "ジャンプリンク" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="27AE60")
        else:
            c_inst.font = font_body
            
    ws_dash.column_dimensions["A"].width = 3
    ws_dash.column_dimensions["B"].width = 16
    ws_dash.column_dimensions["C"].width = 16
    ws_dash.column_dimensions["D"].width = 4
    ws_dash.column_dimensions["E"].width = 95
    
    # 2. Sheet 2: Consolidated Input & Comparison Sheet
    ws_comp = wb.create_sheet(title="統合受注入力・原本比較")
    ws_comp.views.sheetView[0].showGridLines = True
    
    # Header Design
    # We want Left side (Edit / Juchu Database) and Right side (Raw extracted reference)
    # Left headers: columns 1 to len(headers) (columns A to CD / BH depending on columns)
    # Let's select only key fields for the main comparison to make it readable and extremely practical!
    # Left database columns (Input/Correction Area - A to AI):
    left_headers = [
        "依頼Ｎｏ", "原本確認リンク", "照合結果", "品名", "製品", "数量1", "ユーザー", 
        "希望納期", "原反", "加工内容", "契約Ｎｏ", "加工賃", "原反ロール数", "在庫場所", 
        "入力担当", "入力日", "月", "受注金額", "受注数", "単価"
    ]
    
    # Right reference columns (Original Data Reference Area):
    right_headers = [
        "原本_品名", "原本_製品", "原本_数量1", "原本_ユーザー", "原本_希望納期", 
        "原本_原反", "原本_加工内容", "原本_契約No", "原本_加工賃", "原本_ロール数"
    ]
    
    meta_headers = ["データ処理区分", "相違点詳細", "原本ファイル名", "原本シート名"]
    
    full_headers = left_headers + right_headers + meta_headers
    
    # Fill colors for headers
    fill_left_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid") # Dark steel blue
    fill_right_header = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid") # Teal / Emerald
    fill_meta_header = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid") # Gray
    
    ws_comp.row_dimensions[1].height = 20
    ws_comp.row_dimensions[2].height = 28
    
    # Write Category Header (Row 1)
    ws_comp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(left_headers))
    ws_comp.cell(row=1, column=1, value="【左側：受注入力・手修正エリア (最終登録データとして扱います)】").font = Font(name="BIZ UDPGothic", size=11, bold=True, color="FFFFFF")
    ws_comp.cell(row=1, column=1).fill = fill_left_header
    ws_comp.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws_comp.merge_cells(start_row=1, start_column=len(left_headers)+1, end_row=1, end_column=len(left_headers)+len(right_headers))
    ws_comp.cell(row=1, column=len(left_headers)+1, value="【右側：原本データ参照エリア (原本Excelからの自動抽出値・読取専用)】").font = Font(name="BIZ UDPGothic", size=11, bold=True, color="FFFFFF")
    ws_comp.cell(row=1, column=len(left_headers)+1).fill = fill_right_header
    ws_comp.cell(row=1, column=len(left_headers)+1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws_comp.merge_cells(start_row=1, start_column=len(left_headers)+len(right_headers)+1, end_row=1, end_column=len(full_headers))
    ws_comp.cell(row=1, column=len(left_headers)+len(right_headers)+1, value="【照合ステータス・原本情報】").font = Font(name="BIZ UDPGothic", size=11, bold=True, color="FFFFFF")
    ws_comp.cell(row=1, column=len(left_headers)+len(right_headers)+1).fill = fill_meta_header
    ws_comp.cell(row=1, column=len(left_headers)+len(right_headers)+1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Write Headers (Row 2)
    for c_idx, h in enumerate(full_headers):
        cell = ws_comp.cell(row=2, column=c_idx + 1, value=h)
        cell.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="FFFFFF")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if c_idx < len(left_headers):
            cell.fill = fill_left_header
        elif c_idx < len(left_headers) + len(right_headers):
            cell.fill = fill_right_header
        else:
            cell.fill = fill_meta_header
            
    # Apply borders around categories in row 1
    for c in range(1, len(full_headers) + 1):
        ws_comp.cell(row=1, column=c).border = thin_border
        
    fill_matched = PatternFill(start_color="F9EBEA", end_color="F9EBEA", fill_type="solid") # standard
    fill_discrep = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid") # Soft yellow
    fill_new_row = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid") # Soft green
    fill_normal_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Write details
    for r_idx, res in enumerate(consolidated_rows):
        row_num = r_idx + 3
        ws_comp.row_dimensions[row_num].height = 20
        
        status = res['__status']
        fill_to_apply = fill_normal_row
        if "新規自動追加" in status:
            fill_to_apply = fill_new_row
        elif "原本相違あり" in status:
            fill_to_apply = fill_discrep
            
        raw = res['__raw_data']
        
        # --- PREPARE LEFT SIDE DATA (INPUT/EDITABLE) ---
        # If it is a new row, we populate it using extracted raw values and formulas
        # If it is an existing row, we populate it using the database values
        
        left_vals = {
            "依頼Ｎｏ": res.get("依頼Ｎｏ", ""),
            "照合結果": "一致" if "原本一致" in status else ("相違あり" if "原本相違あり" in status else ("未登録" if "新規自動追加" in status else "未確認")),
            "品名": res.get("品名", ""),
            "製品": res.get("製品", ""),
            "数量1": res.get("数量1", ""),
            "ユーザー": res.get("ユーザー", ""),
            "希望納期": res.get("希望納期", ""),
            "原反": res.get("原反", ""),
            "加工内容": res.get("加工内容", ""),
            "契約Ｎｏ": res.get("契約Ｎｏ", ""),
            "加工賃": res.get("加工賃", ""),
            "原反ロール数": res.get("原反ロール数", ""),
            "在庫場所": res.get("在庫場所", ""),
            "入力担当": res.get("入力担当", ""),
            "入力日": res.get("入力日", ""),
            "月": f"=MONTH(H{row_num})" if "新規自動追加" in status else res.get("月", ""),
            "受注金額": f"=T{row_num}*S{row_num}" if "新規自動追加" in status else res.get("受注金額", ""),
            "受注数": f"=F{row_num}" if "新規自動追加" in status else res.get("受注数", ""),
            "単価": f"=L{row_num}" if "新規自動追加" in status else res.get("単価", "")
        }
        
        # --- PREPARE RIGHT SIDE DATA (READ-ONLY ORIGINAL VALUES) ---
        right_vals = {
            "原本_品名": raw['品名'] if raw else "",
            "原本_製品": raw['製品'] if raw else "",
            "原本_数量1": raw['数量1'] if raw else "",
            "原本_ユーザー": raw['ユーザー'] if raw else "",
            "原本_希望納期": raw['希望納期'] if raw else "",
            "原本_原反": raw['原反'] if raw else "",
            "原本_加工内容": raw['加工内容'] if raw else "",
            "原本_契約No": raw['契約Ｎｏ'] if raw else "",
            "原本_加工賃": raw['加工賃'] if raw else "",
            "原本_ロール数": raw['原反ロール数'] if raw else ""
        }
        
        # --- PREPARE META COLUMN DATA ---
        meta_vals = {
            "データ処理区分": status,
            "相違点詳細": res['__discrepancy'],
            "原本ファイル名": raw['file_name'] if raw else "",
            "原本シート名": raw['sheet_name'] if raw else ""
        }
        
        # Write to cells row by row
        for c_idx, h in enumerate(full_headers):
            cell = ws_comp.cell(row=row_num, column=c_idx + 1)
            cell.font = font_body
            cell.fill = fill_to_apply
            cell.border = thin_border
            
            # 1. Hyperlink Column: "原本確認リンク" (Col B / Col 2)
            if h == "原本確認リンク":
                if raw:
                    # Point to the copied worksheet in the same workbook!
                    # Syntax: =HYPERLINK("#'sheet_name'!A1", "link_label")
                    cell.value = f"=HYPERLINK(\"#'{raw['sheet_name']}'!A2\", \"原本シートを表示\")"
                    cell.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="2980B9", underline="single")
                else:
                    cell.value = "原本なし"
                    cell.font = Font(name="BIZ UDPGothic", size=9.5, color="95A5A6")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # 2. Left Side fields
            elif h in left_vals:
                val = left_vals[h]
                if isinstance(val, (datetime.datetime, datetime.date)):
                    cell.value = val.strftime("%Y-%m-%d")
                else:
                    cell.value = val
                
                # Format left numbers
                if h in ["数量1", "原反ロール数", "受注金額", "受注数", "単価", "加工賃"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if cell.value != "" and cell.value is not None and not str(cell.value).startswith("="):
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0'
                        except:
                            pass
                elif h in ["依頼Ｎｏ", "照合結果", "希望納期", "入力担当", "入力日", "月"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if h == "照合結果":
                        cell.font = font_bold
                        if cell.value == "相違あり":
                            cell.font = Font(name="BIZ UDPGothic", size=10, bold=True, color="D35400")
                        elif cell.value == "未登録":
                            cell.font = Font(name="BIZ UDPGothic", size=10, bold=True, color="1E8449")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            # 3. Right Side fields
            elif h in right_vals:
                val = right_vals[h]
                if isinstance(val, (datetime.datetime, datetime.date)):
                    cell.value = val.strftime("%Y-%m-%d")
                else:
                    cell.value = val
                # Style right fields with italicized slightly muted font to demarcate read-only
                cell.font = Font(name="BIZ UDPGothic", size=9.5, italic=True, color="2C3E50")
                
                if h in ["原本_数量1", "原本_加工賃", "原本_ロール数"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if cell.value != "" and cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '#,##0'
                        except:
                            pass
                elif h in ["原本_希望納期", "原本_契約No"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            # 4. Meta fields
            elif h in meta_vals:
                cell.value = meta_vals[h]
                if h == "データ処理区分":
                    cell.font = font_bold
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif h in ["原本ファイル名", "原本シート名"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
    # Auto-fit columns
    for col in ws_comp.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in list(col)[:20]:
            val_str = str(cell.value or '')
            byte_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
            if byte_len > max_len:
                max_len = byte_len
        ws_comp.column_dimensions[col_letter].width = max(max_len + 4, 11)
        
    # Auto-filter details
    ws_comp.auto_filter.ref = f"A2:{get_column_letter(len(full_headers))}{len(consolidated_rows) + 2}"
    
    # 3. Add Back-Links to Copied Worksheet Tops (Row 1 inserts)
    print("Writing navigation back-links at the top of all copied raw request sheets...")
    for idx, res in enumerate(consolidated_rows):
        raw = res['__raw_data']
        if raw and '__dst_sheet_obj' in raw:
            ws_dst = raw['__dst_sheet_obj']
            
            # The sheet already has cells copied starting at row 2.
            # Row 1 is empty for our back link!
            ws_dst.row_dimensions[1].height = 24
            
            # Merge cell A1 to O1
            ws_dst.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
            
            # Set formula hyperlink pointing to the specific row in Consolidated comparison!
            # Syntax: =HYPERLINK("#'sheet_name'!A{row}", "label")
            master_row = idx + 3
            back_cell = ws_dst.cell(row=1, column=1)
            back_cell.value = f"=HYPERLINK(\"#'統合受注入力・原本比較'!A{master_row}\", \"◀ 統合受注データ照合表に戻る (現在の照合行: {master_row}行目 / 依頼Ｎｏ: {raw['依頼Ｎｏ']})\")"
            
            back_cell.font = Font(name="BIZ UDPGothic", size=10, bold=True, color="FFFFFF")
            back_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid") # Forest Green
            back_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Simple border
            for c in range(1, 16):
                ws_dst.cell(row=1, column=c).border = Border(bottom=Side(style='medium', color='1E8449'))
                
    wb.save(output_path)
    wb.close()
    print("Interactive Excel consolidation workbook generated successfully.")

# --- MAIN EXECUTION ---
def main():
    print("=== Interactive Master Database & Raw Sheets Consolidation ===")
    
    # Initialize a destination workbook where sheets will be copied in
    wb_dst = openpyxl.Workbook()
    
    # Extract Raw Request sheets and copy them into wb_dst
    raw_requests = extract_and_copy_raw_sheets(WORKSPACE_DIR, wb_dst)
    
    # Load Juchu rows
    if not os.path.exists(JUCHU_FILE):
        print(f"ERROR: Juchu file not found at {JUCHU_FILE}")
        return
        
    headers, db_rows = load_juchu_file_data(JUCHU_FILE)
    
    # Merge and build side-by-side structures
    consolidated_rows = merge_and_build_interactive(db_rows, raw_requests)
    
    # Generate the complete interactive workbook
    generate_interactive_workbook(headers, consolidated_rows, OUTPUT_FILE)
    
    print("\n=== Interactive Consolidation Complete ===")
    print(f"Master Sheet Rows: {len(consolidated_rows)}")
    print(f"Copied Request Sheets: {len(raw_requests)}")
    print(f"Interactive workbook successfully saved at: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
