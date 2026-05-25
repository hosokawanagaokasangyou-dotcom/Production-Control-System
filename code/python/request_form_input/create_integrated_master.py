import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# アラジンマスタフォルダ: PM_AI_ALADDIN_MASTER_DIR 優先。無ければ作業フォルダ/アラジンマスタ。
workspace = os.environ.get("PM_AI_REQUEST_FORM_WORKSPACE", os.getcwd())
master_dir = os.environ.get("PM_AI_ALADDIN_MASTER_DIR")
if not master_dir or not str(master_dir).strip():
    master_dir = os.path.join(workspace, "アラジンマスタ")
else:
    master_dir = os.path.abspath(master_dir.strip())

naiyo_path = os.path.join(master_dir, "後加工加工内容マスタ.xlsx")
koutei_path = os.path.join(master_dir, "後加工工程マスタ.xlsx")
shohin_path = os.path.join(master_dir, "後加工商品マスタ.xlsx")
output_path = os.path.join(master_dir, "マスタリレーション統合結果.xlsx")

print("Loading data...")
print(f"  workspace: {workspace}")
print(f"  master_dir: {master_dir}")
df_naiyo = pd.read_excel(naiyo_path, engine="calamine")
df_koutei = pd.read_excel(koutei_path, engine="calamine")
df_shohin = pd.read_excel(shohin_path, engine="calamine")

# Helper function to clean codes
def clean_code(val, pad_len=4):
    if pd.isna(val):
        return ""
    try:
        val_str = str(val).strip()
        if val_str.endswith(".0"):
            val_str = val_str[:-2]
        if not val_str:
            return ""
        val_int = int(float(val_str))
        return f"{val_int:0{pad_len}d}"
    except (ValueError, TypeError):
        return str(val).strip()

print("Cleaning code columns...")
# Clean keys in content master
df_naiyo["加工内容コード"] = df_naiyo["加工内容コード"].apply(lambda x: clean_code(x, 4))
df_naiyo["工程コード"] = df_naiyo["工程コード"].apply(lambda x: clean_code(x, 4))

# Clean keys in process master
df_koutei["工程コード"] = df_koutei["工程コード"].apply(lambda x: clean_code(x, 4))

# Clean keys in product master
df_shohin["商品コード"] = df_shohin["商品コード"].apply(lambda x: clean_code(x, 0)) # no padding for product codes
for i in range(1, 8):
    df_shohin[f"加工内容コード{i}"] = df_shohin[f"加工内容コード{i}"].apply(lambda x: clean_code(x, 4))

# 1. Base Master: Content Master Left Join Process Master
print("Performing Base Master Join (Content Master + Process Master)...")
# Rename conflicting column names
if "要員数" in df_naiyo.columns and "要員数" in df_koutei.columns:
    df_naiyo = df_naiyo.rename(columns={"要員数": "要員数_加工内容"})
    df_koutei = df_koutei.rename(columns={"要員数": "要員数_工程"})

df_base_master = pd.merge(df_naiyo, df_koutei, on="工程コード", how="left")

# Create dictionaries for fast lookup of Content and Process details
naiyo_dict = df_base_master.set_index("加工内容コード")["加工内容名"].to_dict()
koutei_dict = df_base_master.set_index("加工内容コード")["工程名"].to_dict()

# 2. Product Processing Steps side-by-side
print("Building Product Processing Steps (Product-Centric) sheet...")
product_cols = ["商品コード", "製品コード", "商品名1", "商品名2", "単位名", "入数", "自社後加工区分", "発泡体品名", "発泡体品番", "発泡体幅", "発泡体長さ", "発泡体色", "発泡体厚み"]
df_product_steps = df_shohin[product_cols].copy()

# Add step info side-by-side
for i in range(1, 8):
    code_col = f"加工内容コード{i}"
    name_col = f"加工内容名{i}"
    koutei_col = f"工程名{i}"
    
    # Get codes from Product master
    df_product_steps[code_col] = df_shohin[code_col]
    # Map names using our dict
    df_product_steps[name_col] = df_product_steps[code_col].map(naiyo_dict).fillna("")
    df_product_steps[koutei_col] = df_product_steps[code_col].map(koutei_dict).fillna("")

# Filter df_product_steps to only show products that have at least one processing step!
has_processing = df_product_steps["加工内容コード1"] != ""
df_product_steps_filtered = df_product_steps[has_processing].copy()
print(f"Products with processing: {len(df_product_steps_filtered)} of {len(df_product_steps)}")

# 3. Unpivoted Content-Centric relation
print("Building Content-Centric Product Usage sheet...")
flat_rows = []
for idx, row in df_base_master.iterrows():
    kako_code = row["加工内容コード"]
    kako_name = row["加工内容名"]
    koutei_code = row["工程コード"]
    koutei_name = row["工程名"]
    
    # Find all products using this code in any of the columns
    for step in range(1, 8):
        matching_shohin = df_shohin[df_shohin[f"加工内容コード{step}"] == kako_code]
        for _, p_row in matching_shohin.iterrows():
            flat_rows.append({
                "加工内容コード": kako_code,
                "加工内容名": kako_name,
                "工程コード": koutei_code,
                "工程名": koutei_name,
                "該当ステップ番号": f"ステップ{step}",
                "商品コード": p_row["商品コード"],
                "商品名1": p_row["商品名1"],
                "商品名2": p_row["商品名2"],
                "単位名": p_row["単位名"],
                "入数": p_row["入数"],
                "自社後加工区分": p_row["自社後加工区分"],
                "発泡体品名": p_row["発泡体品名"],
                "発泡体品番": p_row["発泡体品番"],
                "発泡体幅": p_row["発泡体幅"],
                "発泡体長さ": p_row["発泡体長さ"],
                "発泡体色": p_row["発泡体色"],
                "発泡体厚み": p_row["発泡体厚み"]
            })

df_flat_relation = pd.DataFrame(flat_rows)

print("Saving to Excel with styling...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styling details
font_header = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
font_data = Font(name="Meiryo", size=9)

fill_naiyo = PatternFill(start_color="1F497D", fill_type="solid")  # Deep Midnight Blue
fill_product = PatternFill(start_color="366092", fill_type="solid") # Elegant Blue Steel
fill_flat = PatternFill(start_color="4F5B66", fill_type="solid")    # Dark Slate Grey
fill_zebra = PatternFill(start_color="F2F5F8", fill_type="solid")   # Very light cool grey

border_thin = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

def create_sheet_with_styling(name, df, header_fill):
    ws = wb.create_sheet(title=name)
    ws.views.sheetView[0].showGridLines = True
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin
    
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        row_values = list(row)
        ws.append(row_values)
        use_zebra = (r_idx % 2 == 1)
        
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_data
            cell.border = border_thin
            if use_zebra:
                cell.fill = fill_zebra
                
            col_name = headers[c_idx - 1]
            if "コード" in col_name or "区分" in col_name or "番号" in col_name or col_name == "商品コード":
                cell.alignment = align_center
                cell.number_format = "@"
            elif isinstance(val, (int, float)):
                cell.alignment = align_right
                if int(val) == val:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = align_left
                
    ws.freeze_panes = "A2"
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            width = sum(2 if ord(char) > 127 else 1 for char in val_str)
            max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 55)

print("Creating Sheet 1: ①工程・加工内容マスター統合")
create_sheet_with_styling("①工程・加工内容マスター統合", df_base_master, fill_naiyo)

print("Creating Sheet 2: ②商品別・工程展開リスト")
create_sheet_with_styling("②商品別・工程展開リスト", df_product_steps_filtered, fill_product)

print("Creating Sheet 3: ③加工内容別・使用商品一覧")
create_sheet_with_styling("③加工内容別・使用商品一覧", df_flat_relation, fill_flat)

print(f"Saving workbook to {output_path}...")
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print("Done! Integrated Excel master created successfully with beautiful styling.")
