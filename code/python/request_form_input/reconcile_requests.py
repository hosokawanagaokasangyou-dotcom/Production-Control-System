import os
import glob
import re
import datetime
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- PATH CONFIGURATION ---
WORKSPACE_DIR = r"c:\Users\0585\OneDrive\ドキュメント\加工依頼書（湖南）"
JUCHU_FILE = os.path.join(WORKSPACE_DIR, "加工依頼書入力.xlsm")
OUTPUT_FILE = os.path.join(WORKSPACE_DIR, "加工依頼書_受注ファイル_照合結果.xlsx")

# --- NORMALIZATION HELPERS ---
def normalize_key(val):
    """Normalize keys (like Request No) to ensure matching works even with full/half width or spaces."""
    if val is None:
        return ""
    text = str(val).strip().upper()
    # Normalize full/half-width Japanese and alphanumeric characters
    text = unicodedata.normalize('NFKC', text)
    # Remove all whitespace
    text = "".join(text.split())
    text = text.replace("－", "-").replace("ー", "-").replace("―", "-").replace("‐", "-")
    return text

def normalize_text(val):
    """Normalize text for safe cell comparison, handling kana/width/spacing variations."""
    if val is None:
        return ""
    text = str(val).strip()
    text = unicodedata.normalize('NFKC', text)
    text = "".join(text.split())
    text = text.replace("－", "-").replace("ー", "-").replace("―", "-").replace("‐", "-")
    return text.upper()

def normalize_numeric(val):
    """Normalize numeric values (e.g. quantity) for clean numerical comparison."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    # Extract numbers only
    nums = re.findall(r'[-+]?\d*\.\d+|\d+', text)
    if nums:
        return float(nums[0])
    return 0.0

def normalize_date_val(val):
    """Normalize datetime/date values to YYYY-MM-DD string, or normalized string if non-date."""
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    text = str(val).strip()
    # Try parsing standard formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            dt = datetime.datetime.strptime(text, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return normalize_text(text)

# --- DATA EXTRACTION ---
def load_juchu_file_data(file_path):
    """Load the ground-truth database rows from 受注ﾌｧｲﾙ sheet."""
    print(f"Loading ground-truth data from {os.path.basename(file_path)}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['受注ﾌｧｲﾙ']
    
    # Read headers
    header_row = [sheet.cell(row=3, column=c).value for c in range(1, sheet.max_column + 1)]
    while header_row and header_row[-1] is None:
        header_row.pop()
        
    db_data = {}
    
    # Row 4 is where data starts
    for r_idx in range(4, sheet.max_row + 1):
        req_no = sheet.cell(row=r_idx, column=1).value # Col A is 依頼Ｎｏ
        if not req_no:
            continue
        
        # Read the entire row based on headers
        row_vals = {}
        for c_idx, h in enumerate(header_row):
            if h:
                row_vals[h] = sheet.cell(row=r_idx, column=c_idx + 1).value
        
        norm_req = normalize_key(req_no)
        
        # In case of duplicate keys, we store them in a list
        if norm_req not in db_data:
            db_data[norm_req] = []
        db_data[norm_req].append(row_vals)
        
    wb.close()
    print(f"Loaded {len(db_data)} unique request IDs from database.")
    return db_data

def extract_raw_request_sheets(workspace_dir):
    """Scan all *加工依頼書*.xlsm files and extract data from individual request sheets."""
    search_pattern = os.path.join(workspace_dir, "*加工依頼書*.xlsm")
    files = glob.glob(search_pattern)
    raw_files = [f for f in files if not os.path.basename(f).startswith("~$") and os.path.basename(f) != "加工依頼書入力.xlsm"]
    
    print(f"Found {len(raw_files)} raw request files to extract.")
    extracted_requests = []
    
    for r_f in raw_files:
        filename = os.path.basename(r_f)
        print(f"Extracting request sheets from {filename}...")
        try:
            wb = openpyxl.load_workbook(r_f, data_only=True)
            for sheet_name in wb.sheetnames:
                # We identify individual request sheets using a strict naming pattern (e.g. C5-1, Y5-42, E6-1)
                if not (re.match(r'^[A-Z]+\d+-\d+$', sheet_name) or re.match(r'^[A-Z]\d+-\d+-\d+$', sheet_name)):
                    continue
                
                sheet = wb[sheet_name]
                
                # Double check if F6 contains "加工依頼書" to be absolutely sure it's a request form
                title_val = sheet.cell(row=6, column=6).value
                if title_val and "加工依頼書" not in str(title_val):
                    # Check if E6 or G6 has it, just in case
                    alt_title_1 = sheet.cell(row=6, column=5).value
                    alt_title_2 = sheet.cell(row=6, column=7).value
                    if not ((alt_title_1 and "加工依頼書" in str(alt_title_1)) or (alt_title_2 and "加工依頼書" in str(alt_title_2))):
                        continue
                
                # --- EXTRACT DATA FROM SHEET ---
                # 1. Base details
                req_no = sheet.cell(row=5, column=18).value or sheet_name # Cell R5
                input_date = sheet.cell(row=2, column=34).value # Cell AH2
                
                # 2. Product request items (Row 10)
                item_code = sheet.cell(row=10, column=2).value # B10 (品名)
                part_no = sheet.cell(row=10, column=6).value # F10 (品番)
                type_code = sheet.cell(row=10, column=11).value # K10 (タイプ)
                width = sheet.cell(row=10, column=16).value # P10 (幅)
                length = sheet.cell(row=10, column=21).value # U10 (長さ)
                
                # Construct Product Code (製品)
                part_str = str(part_no).strip() if part_no is not None else ""
                type_str = str(type_code).strip() if type_code is not None else ""
                width_str = str(width).strip() if width is not None else ""
                length_str = str(length).strip() if length is not None else ""
                constructed_product = f"{part_str}-{type_str}-{width_str}X{length_str}" if (part_str or type_str) else ""
                
                grade = sheet.cell(row=10, column=24).value # X10 (梱－等)
                color = sheet.cell(row=10, column=27).value # AA10 (色)
                category = sheet.cell(row=10, column=29).value # AC10 (区分)
                quantity = sheet.cell(row=10, column=31).value # AE10 (数量)
                ec_side = sheet.cell(row=10, column=36).value # AJ10 (EC面)
                trimming = sheet.cell(row=10, column=39).value # AM10 (トリミング)
                
                # 3. Processing steps (Row 13 to 17, Col I (Col 9))
                processing_steps = []
                for r in range(13, 18):
                    p_val = sheet.cell(row=r, column=9).value
                    if p_val:
                        processing_steps.append(str(p_val).strip())
                processing_str = ", ".join(processing_steps)
                
                # 4. Use/User/Delivery/Contract
                use_val = sheet.cell(row=18, column=5).value # E18 (用途)
                user_val = sheet.cell(row=19, column=5).value # E19 (ユーザー)
                delivery_val = sheet.cell(row=20, column=9).value # I20 (納期)
                contract_val = sheet.cell(row=21, column=5).value # E21 (契約ＮＯ)
                
                # 5. Raw material items (Row 23)
                raw_item_code = sheet.cell(row=23, column=8).value # H23 (原反品名)
                raw_part_no = sheet.cell(row=23, column=11).value # K23 (原反品番)
                raw_type_code = sheet.cell(row=23, column=14).value # N23 (原反タイプ)
                raw_width = sheet.cell(row=23, column=17).value # Q23 (原反幅)
                raw_length = sheet.cell(row=23, column=20).value # T23 (原反長さ)
                
                # Construct Raw Material Code
                r_part_str = str(raw_part_no).strip() if raw_part_no is not None else ""
                r_type_str = str(raw_type_code).strip() if raw_type_code is not None else ""
                r_width_str = str(raw_width).strip() if raw_width is not None else ""
                r_length_str = str(raw_length).strip() if raw_length is not None else ""
                constructed_raw = f"{r_part_str}-{r_type_str}-{r_width_str}X{r_length_str}" if (r_part_str or r_type_str) else ""
                
                raw_grade = sheet.cell(row=23, column=22).value # V23 (原反梱－等)
                raw_color = sheet.cell(row=23, column=25).value # Y23 (原反色)
                raw_category = sheet.cell(row=23, column=27).value # AA23 (原反区分)
                raw_qty = sheet.cell(row=23, column=29).value # AC23 (原反数量)
                location = sheet.cell(row=23, column=32).value # AF23 (在庫場所)
                input_day = sheet.cell(row=23, column=39).value # AM23 (投入日)
                
                # 6. Issuer/Person in charge (Row 26 Col C (Col 3))
                issuer = sheet.cell(row=26, column=3).value # C26 (発行者)
                
                # Compile raw sheet data
                extracted_requests.append({
                    'file_name': filename,
                    'sheet_name': sheet_name,
                    '依頼Ｎｏ': str(req_no).strip(),
                    '入力日': input_date,
                    '品名': item_code,
                    '製品': constructed_product,
                    '梱-等1': grade,
                    '色1': color,
                    '区分1': category,
                    '数量1': quantity,
                    'ＥＣ面': ec_side,
                    'ﾄﾘﾐﾝｸﾞ': trimming,
                    '加工内容': processing_str,
                    '用途': use_val,
                    'ユーザー': user_val,
                    '希望納期': delivery_val,
                    '契約Ｎｏ': contract_val,
                    '原反品名': raw_item_code,
                    '原反': constructed_raw,
                    '原反梱-等': raw_grade,
                    '原反色': raw_color,
                    '原反区分': raw_category,
                    '原反数量': raw_qty,
                    '在庫場所': location,
                    '投入日': input_day,
                    '発行者': issuer
                })
            wb.close()
        except Exception as e:
            print(f"Error reading file {filename}: {e}")
            import traceback
            traceback.print_exc()
            
    print(f"Successfully extracted {len(extracted_requests)} requests from raw files.")
    return extracted_requests

# --- DATA RECONCILIATION ---
def reconcile_data(raw_requests, db_data):
    """Reconcile extracted raw request sheets against ground-truth database rows."""
    print("Reconciling extracted raw requests against database...")
    reconciliation_results = []
    
    # Store processed database keys to track unentered records later
    processed_db_keys = set()
    
    for raw in raw_requests:
        req_no = raw['依頼Ｎｏ']
        norm_key = normalize_key(req_no)
        
        # Check if the Request No exists in Juchu database
        if norm_key in db_data:
            processed_db_keys.add(norm_key)
            db_rows = db_data[norm_key]
            
            # If there are multiple entries in the database, we try to match the best one, or take the first
            db_row = db_rows[0] # Default to first matching row
            
            # --- DETECT DISCREPANCIES ---
            discrepancies = []
            
            # 1. Item code comparison (品名)
            raw_item = normalize_text(raw['品名'])
            db_item = normalize_text(db_row.get('品名', ''))
            if raw_item != db_item:
                discrepancies.append(f"品名相違(原本:{raw['品名']}/受注:{db_row.get('品名','')})")
                
            # 2. Product code comparison (製品)
            raw_prod = normalize_text(raw['製品'])
            db_prod = normalize_text(db_row.get('製品', ''))
            if raw_prod != db_prod:
                discrepancies.append(f"製品コード相違(原本:{raw['製品']}/受注:{db_row.get('製品','')})")
                
            # 3. Quantity comparison (数量1)
            raw_qty = normalize_numeric(raw['数量1'])
            db_qty = normalize_numeric(db_row.get('数量1', 0))
            if raw_qty != db_qty:
                discrepancies.append(f"数量1相違(原本:{raw['数量1']}/受注:{db_row.get('数量1','')})")
                
            # 4. User comparison (ユーザー)
            raw_user = normalize_text(raw['ユーザー'])
            db_user = normalize_text(db_row.get('ユーザー', ''))
            if raw_user != db_user:
                # We do a slightly looser match (contains) to avoid simple kana mismatches in company suffixes
                if raw_user not in db_user and db_user not in raw_user:
                    discrepancies.append(f"ユーザー相違(原本:{raw['ユーザー']}/受注:{db_row.get('ユーザー','')})")
                    
            # 5. Delivery date comparison (希望納期)
            raw_deliv = normalize_date_val(raw['希望納期'])
            db_deliv = normalize_date_val(db_row.get('希望納期', ''))
            if raw_deliv != db_deliv:
                discrepancies.append(f"希望納期相違(原本:{normalize_date_val(raw['希望納期'])}/受注:{normalize_date_val(db_row.get('希望納期',''))})")
                
            # 6. Raw material comparison (原反)
            raw_mat = normalize_text(raw['原反'])
            db_mat = normalize_text(db_row.get('原反', ''))
            if raw_mat != db_mat:
                discrepancies.append(f"原反コード相違(原本:{raw['原反']}/受注:{db_row.get('原反','')})")
                
            # 7. Processing contents comparison (加工内容)
            raw_proc = normalize_text(raw['加工内容'])
            db_proc = normalize_text(db_row.get('加工内容', ''))
            # Ignore whitespace/comma variations
            raw_proc_clean = raw_proc.replace(",", "").replace("，", "").replace("、", "")
            db_proc_clean = db_proc.replace(",", "").replace("，", "").replace("、", "")
            if raw_proc_clean != db_proc_clean:
                discrepancies.append(f"加工内容相違(原本:{raw['加工内容']}/受注:{db_row.get('加工内容','')})")
                
            # 8. Contract No comparison (契約Ｎｏ)
            raw_cont = normalize_text(raw['契約Ｎｏ'])
            db_cont = normalize_text(db_row.get('契約Ｎｏ', ''))
            if raw_cont != db_cont:
                # Sometimes raw has contract but db is empty, or vice-versa
                if raw_cont or db_cont:
                    discrepancies.append(f"契約No相違(原本:{raw['契約Ｎｏ']}/受注:{db_row.get('契約Ｎｏ','')})")
            
            status = "相違あり" if discrepancies else "一致"
            desc_detail = "; ".join(discrepancies)
            
            reconciliation_results.append({
                'status': status,
                '依頼Ｎｏ': req_no,
                'file_name': raw['file_name'],
                'sheet_name': raw['sheet_name'],
                '原本_品名': raw['品名'],
                '受注_品名': db_row.get('品名', ''),
                '原本_製品': raw['製品'],
                '受注_製品': db_row.get('製品', ''),
                '原本_数量1': raw['数量1'],
                '受注_数量1': db_row.get('数量1', ''),
                '原本_ユーザー': raw['ユーザー'],
                '受注_ユーザー': db_row.get('ユーザー', ''),
                '原本_希望納期': raw['希望納期'],
                '受注_希望納期': db_row.get('希望納期', ''),
                '原本_原反': raw['原反'],
                '受注_原反': db_row.get('原反', ''),
                '原本_加工内容': raw['加工内容'],
                '受注_加工内容': db_row.get('加工内容', ''),
                '原本_契約Ｎｏ': raw['契約Ｎｏ'],
                '受注_契約Ｎｏ': db_row.get('契約Ｎｏ', ''),
                '受注_入力担当': db_row.get('入力担当', ''),
                '受注_入力日': db_row.get('入力日', ''),
                '相違点詳細': desc_detail
            })
            
        else:
            # Not found in Juchu database
            reconciliation_results.append({
                'status': "未入力",
                '依頼Ｎｏ': req_no,
                'file_name': raw['file_name'],
                'sheet_name': raw['sheet_name'],
                '原本_品名': raw['品名'],
                '受注_品名': '',
                '原本_製品': raw['製品'],
                '受注_製品': '',
                '原本_数量1': raw['数量1'],
                '受注_数量1': '',
                '原本_ユーザー': raw['ユーザー'],
                '受注_ユーザー': '',
                '原本_希望納期': raw['希望納期'],
                '受注_希望納期': '',
                '原本_原反': raw['原反'],
                '受注_原反': '',
                '原本_加工内容': raw['加工内容'],
                '受注_加工内容': '',
                '原本_契約Ｎｏ': raw['契約Ｎｏ'],
                '受注_契約Ｎｏ': '',
                '受注_入力担当': '',
                '受注_入力日': '',
                '相違点詳細': '受注ファイル未入力'
            })
            
    print(f"Reconciliation completed: {len(reconciliation_results)} records processed.")
    return reconciliation_results

# --- STYLED EXCEL REPORT GENERATION ---
def generate_reconciliation_excel(results, output_path):
    """Write the beautifully styled reconciliation dashboard and detail table in a new Excel file."""
    print(f"Generating styled Excel report at {output_path}...")
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Dashboard Sheet
    ws_dash = wb.active
    ws_dash.title = "ダッシュボード"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Summary calculation
    total_count = len(results)
    matched_count = sum(1 for r in results if r['status'] == "一致")
    discrepancy_count = sum(1 for r in results if r['status'] == "相違あり")
    unentered_count = sum(1 for r in results if r['status'] == "未入力")
    
    # Style definitions
    font_title = Font(name="BIZ UDPGothic", size=18, bold=True, color="2C3E50")
    font_section = Font(name="BIZ UDPGothic", size=12, bold=True, color="2C3E50")
    font_body = Font(name="BIZ UDPGothic", size=10)
    font_bold = Font(name="BIZ UDPGothic", size=10, bold=True)
    
    fill_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    fill_matched = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid") # Soft green
    fill_discrepancy = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid") # Soft yellow
    fill_unentered = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid") # Soft red
    
    thin_border_side = Side(style='thin', color='BDC3C7')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # --- Dashboard Title ---
    ws_dash["B2"] = "加工依頼書原本 ⇔ 受注ファイル 自動照合結果レポート"
    ws_dash["B2"].font = font_title
    
    # Subtitle / Date
    ws_dash["B3"] = f"出力日時: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  湖南工場 加工管理用"
    ws_dash["B3"].font = Font(name="BIZ UDPGothic", size=9, italic=True, color="7F8C8D")
    
    # Separator
    for col in range(2, 7):
        ws_dash.cell(row=4, column=col).border = Border(bottom=Side(style='medium', color='2C3E50'))
        
    # --- KPI Block Section ---
    ws_dash["B6"] = "【照合結果サマリー】"
    ws_dash["B6"].font = font_section
    
    kpis = [
        ("総原本件数 (Total)", total_count, "34495E", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("一致 (Matched)", matched_count, "1ABC9C", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF")),
        ("相違あり (Discrepancies)", discrepancy_count, "F1C40F", Font(name="BIZ UDPGothic", size=14, bold=True, color="2C3E50")),
        ("未入力 (Unentered)", unentered_count, "E74C3C", Font(name="BIZ UDPGothic", size=14, bold=True, color="FFFFFF"))
    ]
    
    for i, (label, val, bg_color, num_font) in enumerate(kpis):
        r_start = 8 + (i * 3)
        # Merge cells for KPI Card
        ws_dash.merge_cells(start_row=r_start, start_column=2, end_row=r_start, end_column=3)
        ws_dash.merge_cells(start_row=r_start+1, start_column=2, end_row=r_start+1, end_column=3)
        
        c_label = ws_dash.cell(row=r_start, column=2)
        c_label.value = label
        c_label.font = Font(name="BIZ UDPGothic", size=9, bold=True, color="7F8C8D")
        c_label.alignment = Alignment(horizontal="center", vertical="center")
        
        c_val = ws_dash.cell(row=r_start+1, column=2)
        c_val.value = val
        c_val.font = num_font
        c_val.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        
        # Border around card
        for r in range(r_start, r_start+2):
            for c in range(2, 4):
                ws_dash.cell(row=r, column=c).border = border_all
                
    # --- System Instructions / Descriptions ---
    ws_dash["E6"] = "【自動照合レポートの確認手順】"
    ws_dash["E6"].font = font_section
    
    instructions = [
        "1. 『照合結果明細』シートを開きます。全ての原本データがリストアップされています。",
        "2. オートフィルターを使用し、『照合結果』列でフィルターをかけることで、",
        "   必要な項目を素早くチェックすることができます。",
        "3. 『未入力』(薄い赤色) で抽出されたデータ:",
        "   ⇒ まだ受注ファイルに登録されていない原本データです。転記処理を行ってください。",
        "4. 『相違あり』(薄い黄色) で抽出されたデータ:",
        "   ⇒ 依頼Noはありますが、数量や寸法などの重要データが原本と異なっています。",
        "   ⇒ どちらが正しいか確認し、受注ファイルまたは原本のデータを修正してください。",
        "5. 『一致』(薄い緑色) で抽出されたデータ:",
        "   ⇒ 原本と受注ファイルの全ての重要項目が一致している安全なデータです。",
        "",
        "※ 本照合では、全角・半角の違いや余分なスペース、カタカナの揺らぎ(ｱﾝﾄﾞｰ vs アンドー)は",
        "   システムによって自動的に補正されて比較されています。"
    ]
    
    for idx, inst in enumerate(instructions):
        c_inst = ws_dash.cell(row=8+idx, column=5)
        c_inst.value = inst
        if "未入力" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="C0392B")
        elif "相違あり" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="D35400")
        elif "一致" in inst:
            c_inst.font = Font(name="BIZ UDPGothic", size=9.5, bold=True, color="27AE60")
        else:
            c_inst.font = font_body
            
    # Set dashboard column widths
    ws_dash.column_dimensions["A"].width = 3
    ws_dash.column_dimensions["B"].width = 14
    ws_dash.column_dimensions["C"].width = 14
    ws_dash.column_dimensions["D"].width = 4
    ws_dash.column_dimensions["E"].width = 75
    
    # 2. Sheet 2: Detail Report Sheet
    ws_det = wb.create_sheet(title="照合結果明細")
    ws_det.views.sheetView[0].showGridLines = True
    
    headers = [
        "No", "照合結果", "依頼Ｎｏ", "原本ファイル名", "原本シート名", 
        "【品名】原本", "【品名】受注ﾌｧｲﾙ", "【製品コード】原本", "【製品コード】受注ﾌｧｲﾙ", 
        "【数量1】原本", "【数量1】受注ﾌｧｲﾙ", "【ユーザー】原本", "【ユーザー】受注ﾌｧｲﾙ", 
        "【希望納期】原本", "【希望納期】受注ﾌｧｲﾙ", "【原反】原本", "【原反】受注ﾌｧｲﾙ", 
        "【加工内容】原本", "【加工内容】受注ﾌｧｲﾙ", "【契約No】原本", "【契約No】受注ﾌｧｲﾙ", 
        "【入力担当】受注ﾌｧｲﾙ", "【入力日】受注ﾌｧｲﾙ", "相違点詳細"
    ]
    
    # Write Headers
    for c_idx, h in enumerate(headers):
        cell = ws_det.cell(row=1, column=c_idx + 1)
        cell.value = h
        cell.font = Font(name="BIZ UDPGothic", size=10, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
    
    ws_det.row_dimensions[1].height = 28
    
    # Write details
    for r_idx, res in enumerate(results):
        row_num = r_idx + 2
        ws_det.row_dimensions[row_num].height = 22
        
        # Color fill based on status
        fill_to_apply = fill_matched
        if res['status'] == "相違あり":
            fill_to_apply = fill_discrepancy
        elif res['status'] == "未入力":
            fill_to_apply = fill_unentered
            
        # Values mapping
        row_data = [
            r_idx + 1,
            res['status'],
            res['依頼Ｎｏ'],
            res['file_name'],
            res['sheet_name'],
            res['原本_品名'],
            res['受注_品名'],
            res['原本_製品'],
            res['受注_製品'],
            res['原本_数量1'],
            res['受注_数量1'],
            res['原本_ユーザー'],
            res['受注_ユーザー'],
            res['原本_希望納期'],
            res['受注_希望納期'],
            res['原本_原反'],
            res['受注_原反'],
            res['原本_加工内容'],
            res['受注_加工内容'],
            res['原本_契約Ｎｏ'],
            res['受注_契約Ｎｏ'],
            res['受注_入力担当'],
            res['受注_入力日'],
            res['相違点詳細']
        ]
        
        # Write to cells
        for c_idx, val in enumerate(row_data):
            cell = ws_det.cell(row=row_num, column=c_idx + 1)
            
            # Format datetime objects nicely
            if isinstance(val, (datetime.datetime, datetime.date)):
                cell.value = val.strftime("%Y-%m-%d")
            else:
                cell.value = val
                
            cell.font = font_body
            cell.fill = fill_to_apply
            cell.border = border_all
            
            # Formatting and alignments
            col_name = headers[c_idx]
            
            # Status and matching columns alignment
            if col_name in ["No", "照合結果", "依頼Ｎｏ", "原本シート名", "【希望納期】原本", "【希望納期】受注ﾌｧｲﾙ", "【契約No】原本", "【契約No】受注ﾌｧｲﾙ", "【入力担当】受注ﾌｧｲﾙ", "【入力日】受注ﾌｧｲﾙ"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "数量" in col_name:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Format numbers
                if cell.value != "":
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '#,##0'
                    except:
                        pass
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Bold for Request No and status
            if col_name in ["照合結果", "依頼Ｎｏ"]:
                cell.font = font_bold
                
    # Auto-fit columns
    for col in ws_det.columns:
        col_letter = get_column_letter(col[0].column)
        # Find maximum length of cell in column
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            # Simple length count (double bytes for Japanese characters for width estimation)
            byte_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
            if byte_len > max_len:
                max_len = byte_len
        # Set dimension with a minimum width
        ws_det.column_dimensions[col_letter].width = max(max_len + 4, 11)
        
    # Auto-filter details
    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"
    
    wb.save(output_path)
    wb.close()
    print("Excel report generated successfully.")

# --- MAIN EXECUTION ---
def main():
    print("=== Processing Request Data Extraction & Reconciliation ===")
    
    # 1. Load Ground Truth from Juchu file
    if not os.path.exists(JUCHU_FILE):
        print(f"ERROR: Ground-truth file not found at {JUCHU_FILE}!")
        return
        
    db_data = load_juchu_file_data(JUCHU_FILE)
    
    # 2. Extract Raw Requests
    raw_requests = extract_raw_request_sheets(WORKSPACE_DIR)
    
    if not raw_requests:
        print("ERROR: No raw request sheets were extracted from any XLSM files!")
        return
        
    # 3. Reconcile
    results = reconcile_data(raw_requests, db_data)
    
    # 4. Generate beautifully styled Excel output
    generate_reconciliation_excel(results, OUTPUT_FILE)
    
    print("\n=== Summary of Reconciliation ===")
    total = len(results)
    matched = sum(1 for r in results if r['status'] == "一致")
    discrep = sum(1 for r in results if r['status'] == "相違あり")
    unentered = sum(1 for r in results if r['status'] == "未入力")
    print(f"Total Raw Requests Extracted: {total}")
    print(f"Matched & Consistent (一致):  {matched} ({matched/total*100:.1f}%)")
    print(f"Matched but Discrepant (相違): {discrep} ({discrep/total*100:.1f}%)")
    print(f"Unentered in Database (未入力): {unentered} ({unentered/total*100:.1f}%)")
    print(f"Report file saved at: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
