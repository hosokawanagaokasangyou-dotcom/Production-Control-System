# -*- coding: utf-8 -*-
# Path resolution matches planning_core.dispatch_workspace + input_resolution (env only).
# Does not import planning_core (avoids __init__ Python version gate). Prints one JSON line.
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from typing import Any

# --- mirrors planning_core.dispatch_workspace / input_resolution (keep in sync) ---

ENV_ACTUALS_DATA_WORKBOOK = "PM_AI_ACTUALS_DATA_WORKBOOK"
ENV_ACTUAL_DETAIL_WORKBOOK = "PM_AI_ACTUAL_DETAIL_WORKBOOK"
ENV_ACTUAL_DETAIL_SOURCE_DIR = "PM_AI_ACTUAL_DETAIL_SOURCE_DIR"

DEFAULT_ACTUAL_DETAIL_SOURCE_DIR = (
    "\\\\192.168.0.101\\"
    "\u5171\u6709\u30d5\u30a9\u30eb\u30c0\\"
    "\u6e56\u5357\u5de5\u5834\\"
    "\u6e56\u5357\u5171\u6709\\"
    "002  \u52a0\u5de5G\\"
    "\u25cf\u691c\u67fb\u8868\u4f5c\u6210\\"
    "\u52a0\u5de5\u5b9f\u7e3e\u660e\u7d30DATA"
)

# Same sheet titles as planning_core._core (ACTUALS_SHEET_NAME / ACTUAL_DETAIL_SHEET_NAME).
ACTUALS_SHEET_NAME = "\u52a0\u5de5\u5b9f\u7e3eDATA"
ACTUAL_DETAIL_SHEET_NAME = "\u52a0\u5de5\u5b9f\u7e3e\u660e\u7d30DATA"

_MAX_ROWS_SCAN = 500_000


def pick_newest_excel_in_dir(dir_path: str) -> str | None:
    best: str | None = None
    best_key = -1.0
    try:
        for name in os.listdir(dir_path):
            if name.startswith("~$"):
                continue
            low = name.lower()
            if not low.endswith((".xlsx", ".xlsm")):
                continue
            fp = os.path.join(dir_path, name)
            if not os.path.isfile(fp):
                continue
            try:
                st = os.stat(fp)
                t = max(float(st.st_mtime), float(getattr(st, "st_atime", 0.0)))
            except OSError:
                continue
            if t > best_key:
                best_key = t
                best = fp
    except OSError:
        return None
    return best


def resolve_actual_detail_workbook_path(task_input_workbook: str) -> str | None:
    wb_explicit = (os.environ.get(ENV_ACTUAL_DETAIL_WORKBOOK) or "").strip()
    if wb_explicit and os.path.isfile(wb_explicit):
        return wb_explicit
    d = (os.environ.get(ENV_ACTUAL_DETAIL_SOURCE_DIR) or "").strip()
    if not d:
        d = DEFAULT_ACTUAL_DETAIL_SOURCE_DIR
    if d and os.path.isdir(d):
        picked = pick_newest_excel_in_dir(d)
        if picked:
            return picked
    tw = (task_input_workbook or "").strip()
    if tw and os.path.isfile(tw):
        return tw
    return None


def resolve_actuals_workbook_path(task_input_workbook: str) -> str:
    p = (os.environ.get(ENV_ACTUALS_DATA_WORKBOOK) or "").strip()
    if p and os.path.isfile(p):
        return p
    shared = resolve_actual_detail_workbook_path(task_input_workbook)
    if shared and os.path.isfile(shared):
        return shared
    return (task_input_workbook or "").strip()


def _resolve_actual_detail_path_and_reason(task_wb: str) -> tuple[str, str]:
    wb_explicit = (os.environ.get(ENV_ACTUAL_DETAIL_WORKBOOK) or "").strip()
    if wb_explicit and os.path.isfile(wb_explicit):
        return wb_explicit, "PM_AI_ACTUAL_DETAIL_WORKBOOK"
    d = (os.environ.get(ENV_ACTUAL_DETAIL_SOURCE_DIR) or "").strip()
    used_default_unc = False
    if not d:
        d = DEFAULT_ACTUAL_DETAIL_SOURCE_DIR
        used_default_unc = True
    if d and os.path.isdir(d):
        picked = pick_newest_excel_in_dir(d)
        if picked:
            label = (
                "PM_AI_ACTUAL_DETAIL_SOURCE_DIR:newest"
                if not used_default_unc
                else "default_unc:newest"
            )
            return picked, label
    tw = (task_wb or "").strip()
    if tw and os.path.isfile(tw):
        return tw, "TASK_INPUT_WORKBOOK"
    return "", "none"


def _resolve_actuals_path_and_reason(task_wb: str) -> tuple[str, str]:
    p = (os.environ.get(ENV_ACTUALS_DATA_WORKBOOK) or "").strip()
    if p and os.path.isfile(p):
        return p, "PM_AI_ACTUALS_DATA_WORKBOOK"
    detail_path, detail_reason = _resolve_actual_detail_path_and_reason(task_wb)
    if detail_path and os.path.isfile(detail_path):
        return detail_path, "shared_workbook(" + detail_reason + ")"
    tw = (task_wb or "").strip()
    if tw and os.path.isfile(tw):
        return tw, "TASK_INPUT_WORKBOOK"
    return "", "none"


def _file_meta(path: str) -> dict[str, Any]:
    if not path or not os.path.isfile(path):
        return {"file_exists": False, "size_bytes": None, "mtime_iso": None}
    try:
        st = os.stat(path)
        return {
            "file_exists": True,
            "size_bytes": int(st.st_size),
            "mtime_iso": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).strftime(
                "%Y-%m-%dT%H:%M:%SZ"
            ),
        }
    except OSError:
        return {"file_exists": False, "size_bytes": None, "mtime_iso": None}


def _sheet_row_scan(path: str, sheet_name: str) -> dict[str, Any]:
    out: dict[str, Any] = {"sheet_found": False, "data_rows": None, "scan_truncated": False, "error": None}
    if not path or not os.path.isfile(path):
        return out
    try:
        import openpyxl
    except ImportError as e:
        out["error"] = "openpyxl import: " + str(e)
        return out
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return out
        out["sheet_found"] = True
        ws = wb[sheet_name]
        n = 0
        for _ in ws.iter_rows():
            n += 1
            if n > _MAX_ROWS_SCAN:
                out["scan_truncated"] = True
                out["data_rows"] = n - 1
                break
        else:
            out["data_rows"] = max(0, n - 1) if n > 0 else 0
        return out
    except Exception as e:
        out["error"] = "sheet scan: " + str(e)
        return out
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def build_status() -> dict[str, Any]:
    task_wb = (os.environ.get("TASK_INPUT_WORKBOOK") or "").strip()
    actuals_path, actuals_reason = _resolve_actuals_path_and_reason(task_wb)
    detail_path, detail_reason = _resolve_actual_detail_path_and_reason(task_wb)
    legacy_actuals = resolve_actuals_workbook_path(task_wb)

    a_meta = _file_meta(actuals_path)
    d_meta = _file_meta(detail_path)
    a_scan = _sheet_row_scan(actuals_path, ACTUALS_SHEET_NAME) if a_meta.get("file_exists") else {}
    d_scan = _sheet_row_scan(detail_path, ACTUAL_DETAIL_SHEET_NAME) if d_meta.get("file_exists") else {}

    return {
        "task_input_workbook": task_wb,
        "note": "Real rows live in Excel; this payload is paths/size/mtime/row counts only.",
        "resolve_actuals_workbook_path": legacy_actuals,
        "entries": [
            {
                "id": "machining_actuals",
                "label": ACTUALS_SHEET_NAME,
                "sheet_name": ACTUALS_SHEET_NAME,
                "resolved_path": actuals_path,
                "resolution": actuals_reason,
                **a_meta,
                **a_scan,
            },
            {
                "id": "machining_actual_detail",
                "label": ACTUAL_DETAIL_SHEET_NAME,
                "sheet_name": ACTUAL_DETAIL_SHEET_NAME,
                "resolved_path": detail_path,
                "resolution": detail_reason,
                **d_meta,
                **d_scan,
            },
        ],
    }


def main() -> int:
    try:
        payload = build_status()
        sys.stdout.write(json.dumps(payload, ensure_ascii=False))
        sys.stdout.write("\n")
        return 0
    except Exception as e:
        err_obj = {"error": str(e), "type": type(e).__name__}
        sys.stderr.write(json.dumps(err_obj, ensure_ascii=False) + "\n")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
