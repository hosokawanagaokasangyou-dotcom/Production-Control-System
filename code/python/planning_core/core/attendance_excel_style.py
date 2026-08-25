# -*- coding: utf-8 -*-
"""openpyxl styles for APP_* attendance calendar sheets (aligned with JavaFX CSS)."""

from __future__ import annotations

import calendar
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from planning_core.core.columns import _result_font

# JavaFX pm-company-calendar / pm-member-attendance colors (RRGGBB)
FILL_WORKING = "FFFFFF"
FILL_PUBLIC = "FECACA"
FILL_NATIONAL = "FCA5A5"
FILL_SPECIAL = "FDE68A"
FILL_WEEKEND_HDR = "E2E8F0"
FILL_MONTH_TITLE = "F1F5F9"
FILL_MEMBER_HDR = "E2EFDA"
FILL_MEMBER_OFF_COL = "FDECEC"
FILL_WEEKEND_COL = "F5F5F5"

FONT_WORKING = "334155"
FONT_PUBLIC = "7F1D1D"
FONT_SPECIAL = "78350F"
FONT_HEADER = "475569"
FONT_BANNER = "1E293B"

MONTHS_PER_ROW = 4
MONTH_BLOCK_W = 8
MONTH_BAND_H = 9
GRID_START_ROW = 5

# メンバー勤怠月次シート（メニュー行挿入後の行番号）
MEMBER_CALENDAR_TITLE_ROW = 2
MEMBER_CALENDAR_LEGEND_ROW = 3
MEMBER_CALENDAR_HEADER_ROW = 4
MEMBER_CALENDAR_DATA_START_ROW = 5
GRID_START_COL = 1

_FILL_CACHE: dict[str, PatternFill] = {}
_THIN = Side(style="thin", color="CBD5E1")
_GRID_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)


def cached_fill(hex6: str) -> PatternFill:
    key = hex6.upper()
    fi = _FILL_CACHE.get(key)
    if fi is None:
        fi = PatternFill(fill_type="solid", start_color=key, end_color=key)
        _FILL_CACHE[key] = fi
    return fi


def company_cell_text(day: int, kind: str, entry: dict | None) -> str:
    if kind == "special_holiday":
        return f"{day}特"
    if kind == "public_holiday":
        return f"{day}公"
    return str(day)


def company_cell_style(kind: str, entry: dict | None) -> tuple[PatternFill, Font]:
    if kind == "special_holiday":
        return cached_fill(FILL_SPECIAL), _result_font(size=9, color=FONT_SPECIAL)
    if kind == "public_holiday":
        return cached_fill(FILL_PUBLIC), _result_font(size=9, color=FONT_PUBLIC)
    return cached_fill(FILL_WORKING), _result_font(size=9, color=FONT_WORKING)


def member_symbol_style(symbol: str, company_kind: str) -> tuple[PatternFill | None, Font]:
    """APP_勤怠カレンダー セル記号の色（JavaFX メンバー勤怠グリッドと対応）。"""
    if symbol in ("休出", "前出", "後出"):
        if symbol in ("前出", "後出"):
            return cached_fill("FDBA74"), _result_font(size=9, bold=True, color="9A3412")
        return cached_fill("FED7AA"), _result_font(size=9, color="9A3412")
    if symbol == "休":
        return cached_fill(FILL_PUBLIC), _result_font(size=9, color=FONT_PUBLIC)
    if symbol == "年休":
        return cached_fill("FEF9C3"), _result_font(size=9, color="854D0E")
    if symbol == "欠":
        return cached_fill("FCE7F3"), _result_font(size=9, bold=True, color="9D174D")
    if symbol in ("前", "後", "前休", "後休"):
        return cached_fill(FILL_SPECIAL), _result_font(size=9, color=FONT_SPECIAL)
    if symbol == "-":
        return cached_fill("E2E8F0"), _result_font(size=9, color="334155")
    if symbol == "時":
        return cached_fill("DBEAFE"), _result_font(size=9, color="1D4ED8")
    if symbol == "·":
        if company_kind == "public_holiday":
            return cached_fill(FILL_MEMBER_OFF_COL), _result_font(size=9, color=FONT_PUBLIC)
        if company_kind == "special_holiday":
            return cached_fill(FILL_SPECIAL), _result_font(size=9, color=FONT_SPECIAL)
        return cached_fill("FFFFFF"), _result_font(size=9, color=FONT_WORKING)
    if company_kind == "public_holiday":
        return cached_fill(FILL_MEMBER_OFF_COL), _result_font(size=9, color=FONT_PUBLIC)
    if company_kind == "special_holiday":
        return cached_fill(FILL_SPECIAL), _result_font(size=9, color=FONT_SPECIAL)
    if symbol:
        return None, _result_font(size=9, color=FONT_BANNER)
    return cached_fill("FFFFFF"), _result_font(size=9, color=FONT_WORKING)


def _month_origin(month_index: int) -> tuple[int, int]:
    band, slot = divmod(month_index, MONTHS_PER_ROW)
    return GRID_START_ROW + band * MONTH_BAND_H, GRID_START_COL + slot * MONTH_BLOCK_W


def _weekday_labels() -> list[str]:
    return ["月", "火", "水", "木", "金", "土", "日"]


def write_company_calendar_grid(
    ws,
    days: dict[str, dict],
    fy_label: int,
    fiscal_start: date,
    fiscal_end: date,
    months: list[tuple[int, int]],
) -> None:
    title = f"{fy_label}年度（{fiscal_start.isoformat()} ～ {fiscal_end.isoformat()}）"
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=MONTHS_PER_ROW * MONTH_BLOCK_W - 1,
    )
    c1 = ws.cell(1, 1, title)
    c1.font = _result_font(size=14, bold=True, color=FONT_BANNER)
    c1.alignment = Alignment(horizontal="left", vertical="center")

    legend = (
        "■出勤  ■公休  ■特別休暇"
        "  — セルは「日」「日公」「日特」。編集はアプリの会社カレンダーで行う。"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=MONTHS_PER_ROW * MONTH_BLOCK_W - 1)
    ws.cell(2, 1, legend).font = _result_font(size=9, color=FONT_HEADER)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=MONTHS_PER_ROW * MONTH_BLOCK_W - 1)
    ws.cell(3, 1, "※本シートはアプリ自動出力（APP_会社カレンダー）。").font = _result_font(
        size=8, color=FONT_HEADER
    )

    today = date.today()
    for idx, (year, month) in enumerate(months):
        r0, c0 = _month_origin(idx)
        ws.merge_cells(
            start_row=r0,
            start_column=c0,
            end_row=r0,
            end_column=c0 + 6,
        )
        title_cell = ws.cell(r0, c0, f"{month}月（{year}）")
        title_cell.font = _result_font(size=11, bold=True, color=FONT_BANNER)
        title_cell.fill = cached_fill(FILL_MONTH_TITLE)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        for i, wd in enumerate(_weekday_labels()):
            hdr = ws.cell(r0 + 1, c0 + i, wd)
            hdr.font = _result_font(size=8, color=FONT_HEADER)
            hdr.fill = cached_fill(FILL_WEEKEND_HDR)
            hdr.alignment = _CENTER
            hdr.border = _GRID_BORDER
            if i >= 5:
                hdr.font = _result_font(size=8, bold=True, color=FONT_PUBLIC)

        ym_days = calendar.monthrange(year, month)[1]
        first = date(year, month, 1)
        offset = first.weekday()  # Mon=0
        for day_num in range(1, ym_days + 1):
            d = date(year, month, day_num)
            if d < fiscal_start or d > fiscal_end:
                continue
            cell_idx = offset + day_num - 1
            row = r0 + 2 + cell_idx // 7
            col = c0 + cell_idx % 7
            entry = days.get(d.isoformat(), {})
            kind = str(entry.get("kind") or "working_day")
            text = company_cell_text(day_num, kind, entry)
            fill, font = company_cell_style(kind, entry)
            cell = ws.cell(row, col, text)
            cell.fill = fill
            cell.font = font
            cell.alignment = _CENTER
            cell.border = _GRID_BORDER
            label = (entry.get("label") or "").strip()
            if label:
                cell.comment = None  # openpyxl comments need separate API; skip for xlsm safety
            if d == today:
                cell.border = Border(
                    left=Side(style="medium", color="F59E0B"),
                    right=Side(style="medium", color="F59E0B"),
                    top=Side(style="medium", color="F59E0B"),
                    bottom=Side(style="medium", color="F59E0B"),
                )

        spacer_col = c0 + 7
        ws.column_dimensions[get_column_letter(spacer_col)].width = 1.0
        for dc in range(7):
            ws.column_dimensions[get_column_letter(c0 + dc)].width = 4.5

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    insert_menu_back_link_row(ws)


def format_member_calendar_sheet(
    ws,
    year: int,
    month: int,
    num_day_cols: int,
    dense_days: dict[str, dict],
) -> None:
    insert_menu_back_link_row(ws)

    total_cols = num_day_cols + 1
    # タイトル・凡例はフリーズ列（A）を跨いで結合しない。Excel で氏名行が日付行とずれるため。
    ws.merge_cells(
        start_row=MEMBER_CALENDAR_TITLE_ROW,
        start_column=2,
        end_row=MEMBER_CALENDAR_TITLE_ROW,
        end_column=total_cols,
    )
    title_cell = ws.cell(MEMBER_CALENDAR_TITLE_ROW, 2, f"{year}年{month}月 メンバー勤怠")
    title_cell.font = _result_font(size=12, bold=True, color=FONT_BANNER)
    title_cell.fill = cached_fill(FILL_MONTH_TITLE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(MEMBER_CALENDAR_TITLE_ROW, 1).fill = cached_fill(FILL_MONTH_TITLE)
    ws.row_dimensions[MEMBER_CALENDAR_TITLE_ROW].height = 24

    ws.merge_cells(
        start_row=MEMBER_CALENDAR_LEGEND_ROW,
        start_column=2,
        end_row=MEMBER_CALENDAR_LEGEND_ROW,
        end_column=total_cols,
    )
    legend_cell = ws.cell(
        MEMBER_CALENDAR_LEGEND_ROW,
        2,
        "凡例: ·=通常  休=全休  年休=有給休暇  欠=欠勤  前休/後休=半休  休出/前出/後出=休出系  時=時間別",
    )
    legend_cell.font = _result_font(size=8, color=FONT_HEADER)
    legend_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(MEMBER_CALENDAR_LEGEND_ROW, 1).fill = cached_fill(FILL_MONTH_TITLE)
    ws.row_dimensions[MEMBER_CALENDAR_LEGEND_ROW].height = 18

    hdr = ws.cell(MEMBER_CALENDAR_HEADER_ROW, 1, "メンバー")
    hdr.font = _result_font(size=9, bold=True, color=FONT_BANNER)
    hdr.fill = cached_fill(FILL_MEMBER_HDR)
    hdr.alignment = _CENTER
    hdr.border = _GRID_BORDER

    ym_days = calendar.monthrange(year, month)[1]
    for day_num in range(1, min(ym_days, num_day_cols) + 1):
        d = date(year, month, day_num)
        col = 2 + day_num - 1
        dow = ["月", "火", "水", "木", "金", "土", "日"][d.weekday()]
        cell = ws.cell(MEMBER_CALENDAR_HEADER_ROW, col, f"{day_num}{dow}")
        cell.font = _result_font(size=9, bold=True, color=FONT_HEADER)
        cell.fill = cached_fill(FILL_MEMBER_HDR)
        cell.alignment = _CENTER
        cell.border = _GRID_BORDER
        entry = dense_days.get(d.isoformat(), {})
        kind = str(entry.get("kind") or "")
        if d.weekday() >= 5:
            cell.fill = cached_fill(FILL_WEEKEND_COL)
            cell.font = _result_font(size=8, bold=True, color=FONT_PUBLIC)
        elif kind == "public_holiday":
            cell.fill = cached_fill(FILL_MEMBER_OFF_COL)
            cell.font = _result_font(size=8, bold=True, color=FONT_PUBLIC)
        elif kind == "special_holiday":
            cell.fill = cached_fill(FILL_SPECIAL)
            cell.font = _result_font(size=8, bold=True, color=FONT_SPECIAL)

    ws.column_dimensions["A"].width = 16
    for c in range(2, 2 + num_day_cols):
        ws.column_dimensions[get_column_letter(c)].width = 7.5

    ws.row_dimensions[MEMBER_CALENDAR_HEADER_ROW].height = 20
    for r in range(MEMBER_CALENDAR_DATA_START_ROW, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = f"B{MEMBER_CALENDAR_DATA_START_ROW}"
    ws.sheet_view.showGridLines = False


def write_member_attendance_flat_table(ws, records: list[dict]) -> None:
    """閲覧用のメンバー勤怠明細（行リスト）。列幅・見出し・フィルタ付き。"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    banner = ws.cell(1, 1, "閲覧専用 — 編集はアプリのメンバー勤怠タブで行ってください")
    banner.font = _result_font(size=11, bold=True, color=FONT_BANNER)
    banner.fill = cached_fill(FILL_MONTH_TITLE)
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    headers = ["日付", "メンバー", "出勤", "退勤", "休暇区分", "備考", "残業(分)"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(2, col, title)
        cell.font = _result_font(size=10, bold=True, color=FONT_BANNER)
        cell.fill = cached_fill(FILL_MEMBER_HDR)
        cell.alignment = _CENTER
        cell.border = _GRID_BORDER
    ws.row_dimensions[2].height = 22

    widths = [13, 18, 9, 9, 12, 16, 11]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 3
    body_font = _result_font(size=10, color=FONT_BANNER)
    alt_fill = cached_fill("F8FAFC")
    for idx, rec in enumerate(records):
        values = [
            rec.get("日付", ""),
            rec.get("メンバー", ""),
            rec.get("出勤時間", ""),
            rec.get("退勤時間", ""),
            rec.get("休暇区分", ""),
            rec.get("備考", ""),
            rec.get("残業(分)", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row, col, val)
            cell.font = body_font
            if col == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = _CENTER
            cell.border = _GRID_BORDER
            if idx % 2 == 1:
                cell.fill = alt_fill
            leave = str(rec.get("休暇区分") or "")
            if leave in ("公休", "年休", "休") or leave.endswith("休"):
                cell.fill = cached_fill(FILL_MEMBER_OFF_COL)
        ws.row_dimensions[row].height = 18
        row += 1

    if row > 3:
        ws.auto_filter.ref = f"A2:G{row - 1}"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def format_calendar_export_at_display(export_at: str | None) -> str:
    """ISO 保存日時を日本時間の読みやすい表記へ。"""
    if not export_at or not str(export_at).strip():
        return ""
    trimmed = str(export_at).strip()
    try:
        dt = datetime.fromisoformat(trimmed)
    except ValueError:
        return trimmed
    jst = ZoneInfo("Asia/Tokyo")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=jst)
    else:
        dt = dt.astimezone(jst)
    return dt.strftime("%Y/%m/%d %H:%M:%S") + "（日本時間）"


def menu_label_for_sheet(sheet_name: str) -> str:
    from planning_core.core.attendance_paths import (
        APP_MASTER_COMPANY_SHEET,
        APP_MASTER_MACHINE_CALENDAR_DATE_SHEET,
        APP_MASTER_MACHINE_CALENDAR_SHEET,
        APP_MASTER_MEMBER_SHEET_PREFIX,
    )

    name = (sheet_name or "").strip()
    if name == APP_MASTER_COMPANY_SHEET:
        return "会社カレンダー"
    if name == APP_MASTER_MACHINE_CALENDAR_SHEET:
        return "機械カレンダー"
    if name == APP_MASTER_MACHINE_CALENDAR_DATE_SHEET:
        return "機械カレンダー 日付ジャンプ"
    if name.startswith(APP_MASTER_MEMBER_SHEET_PREFIX):
        return "メンバー勤怠 " + name[len(APP_MASTER_MEMBER_SHEET_PREFIX):]
    return name


def sort_menu_sheet_names(names: list[str]) -> list[str]:
    from planning_core.core.attendance_paths import (
        APP_MASTER_COMPANY_SHEET,
        APP_MASTER_MACHINE_CALENDAR_DATE_SHEET,
        APP_MASTER_MACHINE_CALENDAR_SHEET,
        APP_MASTER_MEMBER_SHEET_PREFIX,
    )

    def member_sort_key(name: str) -> tuple[int, int, str]:
        if not name.startswith(APP_MASTER_MEMBER_SHEET_PREFIX):
            return (0, 0, name)
        rest = name[len(APP_MASTER_MEMBER_SHEET_PREFIX):]
        try:
            year_part, month_part = rest.split("年", 1)
            month = int(month_part.replace("月", ""))
            return (int(year_part), month, name)
        except ValueError:
            return (0, 0, name)

    company: list[str] = []
    machine: list[str] = []
    machine_date: list[str] = []
    members: list[str] = []
    other: list[str] = []
    for name in names:
        if name == APP_MASTER_COMPANY_SHEET:
            company.append(name)
        elif name == APP_MASTER_MACHINE_CALENDAR_SHEET:
            machine.append(name)
        elif name == APP_MASTER_MACHINE_CALENDAR_DATE_SHEET:
            machine_date.append(name)
        elif name.startswith(APP_MASTER_MEMBER_SHEET_PREFIX):
            members.append(name)
        else:
            other.append(name)
    members.sort(key=member_sort_key)
    other.sort()
    return company + machine + machine_date + members + other


def _sheet_hyperlink_target(sheet_name: str) -> str:
    """Excel 内部リンク（シート名に空白・記号がある場合は単引用符で囲む）。"""
    safe = (sheet_name or "").replace("'", "''")
    return f"#'{safe}'!A1"


def sheet_hyperlink_to_cell(sheet_name: str, row: int, col: int = 1) -> str:
    """指定セルへジャンプする内部ハイパーリンク。"""
    safe = (sheet_name or "").replace("'", "''")
    return f"#'{safe}'!{get_column_letter(col)}{int(row)}"


def write_calendar_workbook_menu_sheet(
    ws,
    sheet_names: list[str],
    export_at: str | None = None,
) -> None:
    """勤怠・機械カレンダー.xlsx の先頭メニューシート（各 APP_* シートへのリンク）。"""
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 40

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title = ws.cell(1, 1, "勤怠・機械カレンダー")
    title.font = _result_font(size=16, bold=True, color=FONT_BANNER)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    intro = (
        "会社カレンダー・メンバー勤怠・機械カレンダーを1ファイルにまとめた閲覧用ブックです。"
        "下のリンクから各シートへ移動できます。編集はアプリで行い「保存」で反映してください。"
    )
    ws.cell(2, 1, intro).font = _result_font(size=9, color=FONT_HEADER)
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[2].height = 36

    if export_at:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
        stamp = ws.cell(3, 1, f"最終出力: {format_calendar_export_at_display(export_at)}")
        stamp.font = _result_font(size=8, color=FONT_HEADER)
        stamp.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[3].height = 16

    hdr_row = 5
    for col, text in enumerate(("移動先", "説明"), start=1):
        cell = ws.cell(hdr_row, col, text)
        cell.font = _result_font(size=10, bold=True, color=FONT_BANNER)
        cell.fill = cached_fill(FILL_MONTH_TITLE)
        cell.alignment = _CENTER
        cell.border = _GRID_BORDER

    link_font = _result_font(size=11, color="0563C1", underline="single")
    body_font = _result_font(size=9, color=FONT_HEADER)
    section_font = _result_font(size=9, bold=True, color=FONT_HEADER)
    row = hdr_row + 1
    from planning_core.core.attendance_paths import (
        APP_MASTER_COMPANY_SHEET,
        APP_MASTER_MACHINE_CALENDAR_DATE_SHEET,
        APP_MASTER_MACHINE_CALENDAR_SHEET,
        APP_MASTER_MEMBER_SHEET_PREFIX,
    )

    prev_was_member = False
    for sheet_name in sheet_names:
        is_member = sheet_name.startswith(APP_MASTER_MEMBER_SHEET_PREFIX)
        if is_member and not prev_was_member:
            section = ws.cell(row, 1, "── メンバー勤怠（月別）──")
            section.font = section_font
            section.alignment = Alignment(horizontal="left", vertical="center")
            section.border = _GRID_BORDER
            ws.cell(row, 2, "各月のメンバー別日次勤怠").font = body_font
            ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row, 2).border = _GRID_BORDER
            ws.row_dimensions[row].height = 18
            row += 1
        prev_was_member = is_member

        label = menu_label_for_sheet(sheet_name)
        link_cell = ws.cell(row, 1, label)
        link_cell.hyperlink = _sheet_hyperlink_target(sheet_name)
        link_cell.font = link_font
        link_cell.alignment = Alignment(horizontal="left", vertical="center")
        link_cell.border = _GRID_BORDER

        hint = ""
        if sheet_name == APP_MASTER_COMPANY_SHEET:
            hint = "年度の公休・出勤日"
        elif sheet_name == APP_MASTER_MACHINE_CALENDAR_SHEET:
            hint = "設備の稼働可否（30分刻み）"
        elif sheet_name == APP_MASTER_MACHINE_CALENDAR_DATE_SHEET:
            hint = "月カレンダーで日付を選び表へジャンプ"
        elif is_member:
            hint = "メンバー別の日次勤怠"
        desc_cell = ws.cell(row, 2, hint)
        desc_cell.font = body_font
        desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        desc_cell.border = _GRID_BORDER

        ws.row_dimensions[row].height = 20
        row += 1

    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.sheet_view.showGridLines = False


def shift_freeze_panes_down(ws, rows: int = 1) -> None:
    """insert_rows 後に freeze_panes の行番号をずらす。"""
    if rows <= 0:
        return
    fp = ws.freeze_panes
    if fp is None:
        return
    if isinstance(fp, str):
        from openpyxl.utils import coordinate_to_tuple, get_column_letter

        row, col = coordinate_to_tuple(fp)
        ws.freeze_panes = f"{get_column_letter(col)}{row + rows}"


def insert_menu_back_link_row(ws, menu_sheet_name: str | None = None) -> None:
    """各 APP_* シート先頭にメニューへ戻るハイパーリンク行を挿入する。"""
    from planning_core.core.attendance_paths import APP_MASTER_MENU_SHEET

    menu = (menu_sheet_name or APP_MASTER_MENU_SHEET).strip()
    ws.insert_rows(1)
    ws.row_dimensions[1].height = 18
    cell = ws.cell(1, 1, "メニューに戻る")
    cell.hyperlink = _sheet_hyperlink_target(menu)
    cell.font = _result_font(size=9, color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    shift_freeze_panes_down(ws, 1)
