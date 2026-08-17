# -*- coding: utf-8 -*-
"""受注ファイル（依頼書入力タブ）から依頼NO別の加工内容・EC面を読む。"""
from __future__ import annotations

import logging
import os
import unicodedata

JUCHU_SHEET_NAME = "受注ﾌｧｲﾙ"
HEADER_ROW_ONE_BASED = 3
COL_IRAI_NO = 0  # A
COL_EC_MEN = 13  # N
COL_KAKO_NAIYO = 25  # Z

ENV_REQUEST_FORM_JUCHU_FILE = "PM_AI_REQUEST_FORM_JUCHU_FILE"


def _normalize_juchu_header_token(val) -> str:
    s = unicodedata.normalize("NFKC", str(val or "").strip()).upper()
    return "".join(s.split())


_JUCHU_IRAI_HEADER_KEYS = frozenset(
    {_normalize_juchu_header_token(x) for x in ("依頼No", "依頼Ｎｏ", "依頼NO", "依頼ＮＯ")}
)
_JUCHU_EC_MEN_HEADER_KEYS = frozenset(
    {_normalize_juchu_header_token(x) for x in ("EC面", "ＥＣ面")}
)
_JUCHU_KAKO_HEADER_KEYS = frozenset(
    {_normalize_juchu_header_token(x) for x in ("加工内容",)}
)


def _resolve_juchu_data_columns(sheet) -> tuple[int, int, int]:
    """見出し行（3行目）から依頼NO・EC面・加工内容の 0-based 列 index を解決。"""
    irai_col, ec_col, kako_col = COL_IRAI_NO, COL_EC_MEN, COL_KAKO_NAIYO
    try:
        header_row = next(
            sheet.iter_rows(
                min_row=HEADER_ROW_ONE_BASED,
                max_row=HEADER_ROW_ONE_BASED,
                values_only=False,
            )
        )
    except StopIteration:
        return irai_col, ec_col, kako_col
    for idx, cell in enumerate(header_row):
        key = _normalize_juchu_header_token(_cell_text(cell))
        if key in _JUCHU_IRAI_HEADER_KEYS:
            irai_col = idx
        elif key in _JUCHU_EC_MEN_HEADER_KEYS:
            ec_col = idx
        elif key in _JUCHU_KAKO_HEADER_KEYS:
            kako_col = idx
    return irai_col, ec_col, kako_col


def normalize_irai_no_key(val) -> str:
    if val is None:
        return ""
    s = unicodedata.normalize("NFKC", str(val).strip()).upper()
    return "".join(s.split())


def parent_irai_no_lookup_key(val) -> str:
    """
    枝番依頼NO の EC 面 lookup 用親キー。
    例: W7-22-1 → W7-22（末尾の ``-`` + 数字 1 セグメントを除く）。
    """
    key = normalize_irai_no_key(val)
    if not key or "-" not in key:
        return ""
    head, tail = key.rsplit("-", 1)
    if head and tail.isdigit() and "-" in head:
        return head
    return ""


def has_original_ec_reference(
    original_lookup: dict[str, dict[str, str]] | None, task_id: str
) -> bool:
    """依頼NO または親依頼NO の依頼書原本 lookup が存在するか。"""
    if not original_lookup:
        return False
    tid = normalize_irai_no_key(task_id)
    parent = parent_irai_no_lookup_key(task_id)
    for key in (tid, parent):
        if key and key in original_lookup:
            return True
    return False


def _lookup_original_ec_men(
    task_id: str, original_lookup: dict[str, dict[str, str]] | None
) -> str:
    if not original_lookup:
        return ""
    tid = normalize_irai_no_key(task_id)
    parent = parent_irai_no_lookup_key(task_id)
    for key in (tid, parent):
        if not key or key not in original_lookup:
            continue
        orig = original_lookup[key]
        if orig is not None:
            return str(orig.get("EC面") or "")
    return ""


def resolve_ec_men_for_side_classification(
    task_id: str,
    juchu_row: dict[str, str],
    original_lookup: dict[str, dict[str, str]] | None,
) -> str:
    """
    EC面区分判定用の EC面 文字列。
    受注 N 列が「両面」のときはそのまま採用。
    受注が H/Ｈ面等で原本 AJ 列が空のときは原本（空＝両面）を優先。
    """
    from planning_core.core.ec_side_classification import (
        ec_men_indicates_double_sided,
        ec_men_indicates_single_sided,
    )

    juchu_ec = str(juchu_row.get("EC面") or "")
    if ec_men_indicates_double_sided(juchu_ec):
        return juchu_ec
    if not original_lookup or not has_original_ec_reference(original_lookup, task_id):
        return juchu_ec
    orig_ec = _lookup_original_ec_men(task_id, original_lookup)
    if ec_men_indicates_single_sided(juchu_ec) and _is_blank_ec_men(orig_ec):
        return orig_ec
    if not _is_blank_ec_men(orig_ec):
        return orig_ec
    if _is_blank_ec_men(juchu_ec):
        return orig_ec
    return juchu_ec


def _is_blank_ec_men(val) -> bool:
    from planning_core.core.ec_side_classification import _is_blank

    return _is_blank(val)


def lookup_juchu_ec_row(
    lookup: dict[str, dict[str, str]],
    task_id: str,
    original_lookup: dict[str, dict[str, str]] | None = None,
) -> dict[str, str]:
    """依頼NO 直 lookup → 親依頼NO（末尾 -数字 除去）の順で受注行を返す。"""
    if not lookup:
        row = {}
    else:
        key = normalize_irai_no_key(task_id)
        row = lookup.get(key, {})
        if not row:
            parent = parent_irai_no_lookup_key(task_id)
            if parent:
                row = lookup.get(parent, {})
            else:
                row = {}

    if not row and not original_lookup:
        return {}

    row = dict(row) if row else {}
    _tid_norm = normalize_irai_no_key(task_id)
    _parent = parent_irai_no_lookup_key(task_id)

    if original_lookup and not str(row.get("EC面") or "").strip():
        for try_key in (_tid_norm, _parent):
            if not try_key:
                continue
            orig = original_lookup.get(try_key, {})
            if not orig:
                continue
            if not str(row.get("EC面") or "").strip() and str(orig.get("EC面") or "").strip():
                row["EC面"] = orig["EC面"]
            if not str(row.get("加工内容") or "").strip() and str(
                orig.get("加工内容") or ""
            ).strip():
                row["加工内容"] = orig["加工内容"]
            if str(row.get("EC面") or "").strip():
                break

    return row


def _cell_text(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _resolve_juchu_path() -> str:
    return (os.environ.get(ENV_REQUEST_FORM_JUCHU_FILE) or "").strip()


def load_juchu_ec_lookup_by_irai_no() -> dict[str, dict[str, str]]:
    """
    依頼NO 正規化キー → {"加工内容": str, "EC面": str}。
    ファイル未到達・シート欠落時は空 dict（警告ログのみ）。
    """
    path = _resolve_juchu_path()
    if not path or not os.path.isfile(path):
        logging.warning(
            "段階1: 受注ファイルが未設定または存在しません（EC面区分は空）。%s=%r",
            ENV_REQUEST_FORM_JUCHU_FILE,
            path,
        )
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError as ex:
        logging.warning("段階1: openpyxl 未導入のため EC面区分 lookup をスキップ: %s", ex)
        return {}

    out: dict[str, dict[str, str]] = {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as ex:
        logging.warning("段階1: 受注ファイル読込失敗（EC面区分は空）: %s", ex)
        return {}

    try:
        sheet = wb[JUCHU_SHEET_NAME] if JUCHU_SHEET_NAME in wb.sheetnames else None
        if sheet is None:
            logging.warning(
                "段階1: 受注ファイルにシート %r がありません（EC面区分は空）: %s",
                JUCHU_SHEET_NAME,
                path,
            )
            return {}

        first_data_row = HEADER_ROW_ONE_BASED + 1
        irai_col, ec_col, kako_col = _resolve_juchu_data_columns(sheet)
        logging.info(
            "段階1: 受注ファイル EC面 lookup cols=(irai=%d ec=%d kako=%d) path=%s",
            irai_col,
            ec_col,
            kako_col,
            path,
        )
        for row in sheet.iter_rows(min_row=first_data_row, values_only=False):
            irai = _cell_text(row[irai_col] if len(row) > irai_col else None)
            if not irai:
                continue
            key = normalize_irai_no_key(irai)
            if not key:
                continue
            ec_men = _cell_text(row[ec_col] if len(row) > ec_col else None)
            kako = _cell_text(row[kako_col] if len(row) > kako_col else None)
            out[key] = {"加工内容": kako, "EC面": ec_men}
    except Exception as ex:
        logging.warning("段階1: 受注ファイル走査失敗（EC面区分は空）: %s", ex)
        return {}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    logging.info("段階1: 受注ファイル EC面 lookup 件数=%d path=%s", len(out), path)
    return out
