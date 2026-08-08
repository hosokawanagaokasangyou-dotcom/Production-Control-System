# -*- coding: utf-8 -*-
# planning_core.core.gantt_excel — body only (loaded via _core exec chain)
def _gantt_color_mode_raw() -> str:
    return (os.environ.get("GANTT_COLOR_MODE", "") or "").strip().lower()
def _gantt_color_mode_full() -> bool:
    return _gantt_color_mode_raw() in ("full", "color", "vivid", "1", "true", "yes", "on")
def _gantt_hsv_to_rgb_u8(h01: float, s: float, v: float) -> tuple[int, int, int]:
    """h01∈[0,1), s・v∈[0,1] を sRGB 0..255 に。"""
    h = (float(h01) % 1.0) * 6.0
    c = float(v) * float(s)
    x = c * (1.0 - abs((h % 2.0) - 1.0))
    m = float(v) - c
    if h < 1.0:
        rp, gp, bp = c, x, 0.0
    elif h < 2.0:
        rp, gp, bp = x, c, 0.0
    elif h < 3.0:
        rp, gp, bp = 0.0, c, x
    elif h < 4.0:
        rp, gp, bp = 0.0, x, c
    elif h < 5.0:
        rp, gp, bp = x, 0.0, c
    else:
        rp, gp, bp = c, 0.0, x
    r = int((rp + m) * 255.0)
    g = int((gp + m) * 255.0)
    b = int((bp + m) * 255.0)
    return max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))
def _gantt_fullcolor_fill_hex_for_task_id(task_id, *, is_actual: bool) -> str:
    """依頼NO（task_id）ごとに色相を固定。実績行は色相をずらして計画と区別。"""
    hx = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    hue01 = (int(hx[0:8], 16) % 360) / 360.0
    if is_actual:
        hue01 = (hue01 + 47.0 / 360.0) % 1.0
    s = 0.36 + (int(hx[8:12], 16) % 26) / 100.0
    v = 0.80 + (int(hx[12:16], 16) % 16) / 100.0
    r, g, b = _gantt_hsv_to_rgb_u8(hue01, s, v)
    return f"{r:02X}{g:02X}{b:02X}"
def _gantt_daily_startup_fill_hex() -> str:
    if _gantt_color_mode_full():
        return "FFC107"
    return _GANTT_DAILY_STARTUP_FILL
def _gantt_bar_fill_for_task_id(task_id):
    """依頼NOごとに1色（RRGGBB）。full 時はHSV、monotone 時は淡色パレット。"""
    if _gantt_color_mode_full():
        return _gantt_fullcolor_fill_hex_for_task_id(task_id, is_actual=False)
    h = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    i = int(h[:8], 16) % len(_GANTT_BAR_FILLS_PRINT_SAFE)
    return _GANTT_BAR_FILLS_PRINT_SAFE[i]
def _gantt_bar_fill_actual_for_task_id(task_id):
    if _gantt_color_mode_full():
        return _gantt_fullcolor_fill_hex_for_task_id(task_id, is_actual=True)
    h = hashlib.md5(str(task_id).encode("utf-8")).hexdigest()
    i = int(h[:8], 16) % len(_GANTT_BAR_FILLS_ACTUAL)
    return _GANTT_BAR_FILLS_ACTUAL[i]
def _gantt_timeline_label_alignment(*, single_slot: bool) -> Alignment:
    """
    ガント帯のラベル用配置。
    1スロット幅のみの帯では列幅が狭く見切れやすいため shrink_to_fit でセル内に収める。
    複数スロット続く帯では shrink せず、空セルへはみ出して表示しやすくする（Excel の表示特性）。
    """
    return Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=bool(single_slot),
        indent=1,
    )
def _gantt_cached_pattern_fill(hex_rrggbb: str) -> PatternFill:
    fi = _GANTT_TASK_PATTERN_FILL_BY_HEX.get(hex_rrggbb)
    if fi is None:
        fi = PatternFill(fill_type="solid", start_color=hex_rrggbb, end_color=hex_rrggbb)
        _GANTT_TASK_PATTERN_FILL_BY_HEX[hex_rrggbb] = fi
    return fi
def _gantt_format_length_m(val) -> str | None:
    """
    タイムライン帯ラベル用の「加工長さ(m)」表示。
    0/空/不正は None（表示しない）。
    """
    try:
        f = parse_float_safe(val, None)
        if f is None or math.isnan(float(f)) or math.isinf(float(f)):
            return None
        f = float(f)
    except Exception:
        return None
    if f <= 1e-12:
        return None
    # 整数に近い値は整数表示、そうでなければ小数1桁（過度に長いラベルを避ける）
    if abs(f - round(f)) <= 1e-9:
        return str(int(round(f)))
    return f"{f:.1f}".rstrip("0").rstrip(".")
def _gantt_segment_total_length_m(evlist, tid_s: str, seg_lo: datetime, seg_hi: datetime):
    """
    依頼NO（task_id）シェイプ区間 [seg_lo, seg_hi) に重なるイベントを単位として
    加工長さを合計する（時間按分しない）。

    - 計画帯: units_done×unit_m
    - 実績明細帯: label_len_m（=加工実績明細DATA由来）を優先
    - 実績明細の日次正規化で各イベントに ``_detail_daily_qty_total_m`` があり、
      当該区間に重なるイベントがすべて同じ日次総量を持つ場合は、按分ラベルの合計ではなく
      その日次総量 1 本を返す（同一日の断続区間の帯に手動集計の m が出るようにする）。
    """
    if not tid_s or not evlist or seg_lo is None or seg_hi is None:
        return None, 0, None
    total = 0.0
    n = 0
    cum_max = None
    pct_pick = None
    detail_unified_m = None
    detail_meta_broken = False
    seen: set[tuple] = set()
    for ev in evlist:
        try:
            if str(ev.get("task_id") or "").strip() != tid_s:
                continue
            s0 = ev.get("start_dt")
            e0 = ev.get("end_dt")
            if not isinstance(s0, datetime) or not isinstance(e0, datetime) or not (s0 < e0):
                continue
            # overlap: [s0,e0) ∩ [seg_lo,seg_hi)
            if e0 <= seg_lo or s0 >= seg_hi:
                continue
            k = (
                tid_s,
                ev.get("event_kind"),
                s0,
                e0,
                str(ev.get("machine") or ""),
            )
            if k in seen:
                continue
            seen.add(k)
            dv_meta = ev.get("_detail_daily_qty_total_m")
            if dv_meta is None:
                detail_meta_broken = True
            elif not detail_meta_broken:
                fdv = parse_float_safe(dv_meta, None)
                if fdv is None or math.isnan(float(fdv)):
                    detail_meta_broken = True
                elif detail_unified_m is None:
                    detail_unified_m = float(fdv)
                elif abs(detail_unified_m - float(fdv)) > 1e-6 * max(1.0, abs(detail_unified_m)):
                    detail_meta_broken = True
            lm = None
            if ev.get("label_len_m") is not None:
                lm = parse_float_safe(ev.get("label_len_m"), None)
            if lm is None:
                u = parse_float_safe(ev.get("units_done"), 0.0)
                um = parse_float_safe(ev.get("unit_m"), 0.0)
                if u > 1e-12 and um > 1e-12:
                    lm = float(u) * float(um)
            if lm is None:
                continue
            # 実績明細: label_len_m は「累積」値である場合がある（その場合は合計せず最大を採る）
            if bool(ev.get("label_len_m_is_cumulative")):
                try:
                    cum_max = float(lm) if cum_max is None else max(float(cum_max), float(lm))
                    n += 1
                except Exception:
                    pass
            else:
                total += float(lm)
                n += 1
            if pct_pick is None and ev.get("pct_macro") is not None:
                try:
                    pct_pick = int(round(parse_float_safe(ev.get("pct_macro"), None)))
                except Exception:
                    pct_pick = None
        except Exception:
            continue
    if n <= 0 and cum_max is None:
        return None, 0, pct_pick
    if cum_max is not None:
        return float(cum_max), n, pct_pick
    if (
        n > 0
        and not detail_meta_broken
        and detail_unified_m is not None
        and float(detail_unified_m) > 1e-12
    ):
        # 実績明細の日次正規化: 結合セグメントに含まれる按分ラベルの合計ではなく、手動集計と一致する日次総量を表示する
        return float(detail_unified_m), n, pct_pick
    return total, n, pct_pick
def _gantt_best_overlapping_events_for_slots_line_sweep(evlist, slots, slot_mins):
    """
    各スロットについて ``_eq_grid_best_overlapping_event_for_cell(evlist, cs, ce)`` と同じ 1 件を返す。
    イベントは開始時刻順に走査し、スロット進行に合わせて active を更新する（スロット×全件走査を避ける）。
    """
    nS = len(slots)
    if nS == 0:
        return []
    if not evlist:
        return [None] * nS
    slot_mins_f = float(slot_mins)
    evs: list = []
    for ev in evlist:
        st = ev.get("start_dt")
        ed = ev.get("end_dt")
        if isinstance(st, datetime) and isinstance(ed, datetime) and st < ed:
            evs.append(ev)
    if not evs:
        return [None] * nS
    evs.sort(
        key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
    )
    nE = len(evs)
    ei = 0
    active: list = []
    out: list = [None] * nS
    for k in range(nS):
        cs = slots[k]
        ce = cs + timedelta(minutes=slot_mins_f)
        if active:
            active = [e for e in active if _eq_grid_slot_overlaps_event(cs, ce, e)]
        while ei < nE:
            e = evs[ei]
            st = e.get("start_dt")
            if not isinstance(st, datetime):
                ei += 1
                continue
            if st >= ce:
                break
            _, ed_disp = _gantt_machining_display_range_for_slot_overlap(
                e, int(slot_mins_f)
            )
            if isinstance(ed_disp, datetime) and ed_disp > cs and st < ce:
                active.append(e)
            ei += 1
        if not active:
            out[k] = None
            continue
        mach_hits = [ev for ev in active if _eq_grid_timeline_event_use_progress_bar(ev)]
        if mach_hits:
            out[k] = min(
                mach_hits,
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or "")),
            )
        else:
            out[k] = min(
                active,
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or "")),
            )
    return out
def _gantt_slot_state_tuple_from_active(active, slot_start, slot_mins, task_fill_fn=None):
    """枠内の代表イベント active が既に決まっているときの 1 マス分の状態（``_gantt_slot_state_tuple`` の後半）。"""
    fill_fn = task_fill_fn or _gantt_bar_fill_for_task_id
    slot_end = slot_start + timedelta(minutes=float(slot_mins))
    slot_mid = slot_start + timedelta(minutes=float(slot_mins) / 2.0)
    if active is None:
        return ("idle",)
    _ek_slot = _timeline_event_kind(active)
    if _ek_slot in (
        TIMELINE_EVENT_MACHINE_DAILY_STARTUP,
        TIMELINE_EVENT_REQUEST_SWITCH_PREP,
        TIMELINE_EVENT_BREAK_RESUME_PREP,
        TIMELINE_EVENT_POST_MACHINING_CLEANUP,
        TIMELINE_EVENT_REQUEST_INTERVAL_BUFFER,
    ):
        return ("daily_startup", _gantt_daily_startup_fill_hex())
    sample_t = _eq_grid_overlap_sample_t(active, slot_start, slot_end, slot_mid)
    if any(b_s <= sample_t < b_e for b_s, b_e in active.get("breaks") or ()):
        return ("break",)
    tid = str(active["task_id"])
    gh = fill_fn(active["task_id"])
    slot_len_m = None
    pct = None
    try:
        # イベント総加工長さ(m)
        ev_total_len_m = None
        if active.get("label_len_m") is not None:
            ev_total_len_m = parse_float_safe(active.get("label_len_m"), None)
        else:
            # 計画帯: units_done × unit_m を加工長さとして表示
            u = parse_float_safe(active.get("units_done"), 0.0)
            um = parse_float_safe(active.get("unit_m"), 0.0)
            if u > 1e-12 and um > 1e-12:
                ev_total_len_m = float(u) * float(um)

        # 累積値は区間按分表示に向かないため、slot_len_m は算出しない（ラベル側で表示する）
        if bool(active.get("label_len_m_is_cumulative")):
            ev_total_len_m = None
        if ev_total_len_m is not None and float(ev_total_len_m) > 1e-12:
            s0 = active.get("start_dt")
            e0 = active.get("end_dt")
            if isinstance(s0, datetime) and isinstance(e0, datetime) and s0 < e0:
                ev_sec = float((e0 - s0).total_seconds())
                ov_sec = float((min(slot_end, e0) - max(slot_start, s0)).total_seconds())
                if ev_sec > 1e-9 and ov_sec > 1e-9:
                    slot_len_m = float(ev_total_len_m) * (ov_sec / ev_sec)
    except Exception:
        slot_len_m = None
    try:
        if active.get("pct_macro") is not None:
            pct = int(round(parse_float_safe(active.get("pct_macro"), 0.0)))
            pct = max(0, min(100, pct))
    except Exception:
        pct = None
    return ("task", tid, gh, slot_len_m, pct)
def _gantt_slot_state_tuple(evlist, slot_start, slot_mins, task_fill_fn=None):
    """
    10 分枠 [slot_start, slot_end) の 1 マス分の状態。
    ('idle',) | ('break',) | ('daily_startup', fill_hex) | ('task', tid, fill_hex, slot_len_m, pct)

    結果_設備毎の時間割・結果_設備毎の時間割_機械名毎（``_build_equipment_schedule_*``）と同様に、
    枠と重なるイベントの選定に ``_eq_grid_best_overlapping_event_for_cell``、
    休憩判定の参照時刻に ``_eq_grid_overlap_sample_t``（枠∩イベント区間の中点）を用いる。
    従来の「枠中点を含む最初のイベント」のみを見る方式では、準備と加工が重なる枠で
    時間割は加工を出すのにガントが準備側へ寄り、依頼NO シェイプが欠けることがあった。
    """
    slot_end = slot_start + timedelta(minutes=float(slot_mins))
    active = _eq_grid_best_overlapping_event_for_cell(evlist, slot_start, slot_end)
    return _gantt_slot_state_tuple_from_active(active, slot_start, slot_mins, task_fill_fn)
def _gantt_timeline_same_segment(st_a, st_b) -> bool:
    """結合セグメント境界判定（毎スロット tuple を割り当でない）。"""
    if st_a[0] != st_b[0]:
        return False
    if st_a[0] == "idle" or st_a[0] == "break":
        return True
    # daily_startup: [1]=fill / task: [1]=task_id
    return st_a[1] == st_b[1]
def _write_results_equipment_gantt_sheet(
    writer,
    timeline_events,
    equipment_list,
    sorted_dates,
    attendance_data,
    data_extract_dt_str,
    base_now_dt=None,
    actual_timeline_events=None,
    regular_shift_times: tuple[time | None, time | None] | None = None,
    *,
    plan_rows: bool = True,
    chart_title: str | None = None,
    sheet_name_override: str | None = None,
    gantt_compare_shape_styling: bool = False,
    compare_aladdin_qty_by_machine_date: dict | None = None,
):
    """
    結果_設備毎の時間割と同一データ源（timeline_events）に基づき、
    設備×横軸時間のガンチャート風シートを追加する。
    横軸は GANTT_TIMELINE_SLOT_MINUTES 分刻み。同一状態の連続は帯状に塗分けする。
    actual_timeline_events があれば設備ごとに「実績」行を計画行の下へ追加する。
    plan_rows=False のときは計画行を出さず actual_timeline_events のみを各行に描画する（実績明細ガント用）。
    GANTT_TIMELINE_SHAPE_LABELS が有効なとき、タイムライン上の依頼NO 等はセルに書かず
    角丸シェイプ用の仕様 dict の list と、日ブロック境界の list を返す（保存後に Excel で描画・画像化）。
    無効時は ([], []) を返す。
    gantt_compare_shape_styling が True のとき、計画行の角丸枠は点線、実績行は太線（比較ガント用）。
    compare_aladdin_qty_by_machine_date が dict のとき（通常 None）、比較ガントで機械×日ごとに
    3 段目「アラジン入力数量」を描画する（キーは (_normalize_equipment_match_key(機械名), date)。
    値は (タスク概覧, タイムライン中央表示, 実績不一致時の注記) の 3 要素または従来互換の 2 要素タプル）。
    """
    sheet_nm = sheet_name_override or RESULT_SHEET_GANTT_NAME
    if not plan_rows:
        if not actual_timeline_events:
            logging.info(
                "設備ガント（%s）: 実績のみモードですがイベントが空のためシートを作成しません。",
                sheet_nm,
            )
            return [], []
    wb = writer.book
    if sheet_name_override:
        try:
            insert_at = wb.sheetnames.index(RESULT_SHEET_GANTT_NAME) + 1
        except ValueError:
            try:
                insert_at = wb.sheetnames.index("結果_設備毎の時間割") + 1
            except ValueError:
                insert_at = len(wb.sheetnames)
    else:
        try:
            insert_at = wb.sheetnames.index("結果_設備毎の時間割") + 1
        except ValueError:
            insert_at = len(wb.sheetnames)
    ws = wb.create_sheet(sheet_nm, insert_at)
    try:
        ws.sheet_properties.tabColor = (
            "1976D2" if _gantt_color_mode_full() else "7F7F7F"
        )
    except Exception:
        pass

    events_by_date = defaultdict(list)
    for e in timeline_events:
        events_by_date[e["date"]].append(e)

    show_actual_rows = bool(actual_timeline_events)
    actual_events_by_date = defaultdict(list)
    if show_actual_rows:
        for e in actual_timeline_events:
            actual_events_by_date[e["date"]].append(e)

    _cmp_shape = bool(gantt_compare_shape_styling)
    _plan_line_dash = bool(_cmp_shape and plan_rows)
    _plan_line_wt: float | None = None
    _act_line_dash = False
    _act_line_wt: float | None = (
        float(COMPARE_GANTT_ACTUAL_SHAPE_LINE_PT)
        if _cmp_shape and show_actual_rows
        else None
    )
    _show_aladdin = bool(
        _cmp_shape
        and plan_rows
        and show_actual_rows
        and compare_aladdin_qty_by_machine_date is not None
    )

    slot_mins = GANTT_TIMELINE_SLOT_MINUTES
    _g_cf = _gantt_color_mode_full()
    hdr_font = _result_font(bold=True, color="000000", size=12)
    hdr_fill = PatternFill(
        fill_type="solid",
        start_color=("BBDEFB" if _g_cf else "D9D9D9"),
        end_color=("BBDEFB" if _g_cf else "D9D9D9"),
    )
    hdr_time_font = _result_font(bold=True, color="000000", size=11)
    title_font = _result_font(bold=True, size=24, color="1A1A1A")
    title_fill = PatternFill(
        fill_type="solid",
        start_color=("E3F2FD" if _g_cf else "DDDDDD"),
        end_color=("E3F2FD" if _g_cf else "DDDDDD"),
    )
    meta_font = _result_font(size=11, color="333333")
    meta_fill = PatternFill(
        fill_type="solid",
        start_color=("F1F8E9" if _g_cf else "F3F3F3"),
        end_color=("F1F8E9" if _g_cf else "F3F3F3"),
    )
    day_banner_font = _result_font(bold=True, size=13, color="1A1A1A")
    day_banner_fill = PatternFill(
        fill_type="solid",
        start_color=("C5E1A5" if _g_cf else "D0D0D0"),
        end_color=("C5E1A5" if _g_cf else "D0D0D0"),
    )
    accent_left = Side(style="thick", color="2B2B2B")
    banner_sep = Side(style="thin", color="7A7A7A")
    thin = Side(style="thin", color=("5C6BC0" if _g_cf else "666666"))
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    aladdin_tl_fill = PatternFill(
        fill_type="solid", start_color="FFFDE7", end_color="FFFDE7"
    )
    aladdin_tl_fill_mismatch = PatternFill(
        fill_type="solid",
        start_color=("FFCDD2" if _g_cf else "FFD6D6"),
        end_color=("FFCDD2" if _g_cf else "FFD6D6"),
    )
    idle_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    break_fill = PatternFill(
        fill_type="solid",
        start_color=("90CAF9" if _g_cf else "B8B8B8"),
        end_color=("90CAF9" if _g_cf else "B8B8B8"),
    )
    gantt_label_font = _result_font(size=10, bold=True, color="000000")
    gantt_label_font_actual = _result_font(size=10, bold=True, color="000000", italic=True)
    _outside_hex = (
        "FFCCBC" if _g_cf else str(RESULT_OUTSIDE_REGULAR_TIME_FILL or "FCE4D6")
    )
    hdr_fill_outside_regular = PatternFill(
        fill_type="solid",
        start_color=_outside_hex,
        end_color=_outside_hex,
    )
    rs, re_ = (regular_shift_times or (None, None))

    # 横軸（slot_mins 刻み）は日付で共通のため、slot_times を先に確定
    base_dt = base_now_dt if isinstance(base_now_dt, datetime) else datetime.now()
    dummy_d = sorted_dates[0] if sorted_dates else base_dt.date()
    d_start0 = datetime.combine(dummy_d, DEFAULT_START_TIME)
    d_end0 = datetime.combine(dummy_d, DEFAULT_END_TIME)
    slot_times = []
    t0 = d_start0
    while t0 < d_end0:
        slot_times.append(t0.time())
        t0 += timedelta(minutes=slot_mins)

    n_slots = len(slot_times)
    # 計画実績比較ガントは工程名列が冗長なため 3 列（日付・機械名・タスク概覝）に縮小
    n_fixed = 3 if _cmp_shape else 4
    last_col = n_fixed + n_slots
    gantt_shape_label_specs: list[dict] = []
    gantt_timeline_day_blocks: list[dict] = []
    _use_gantt_shape_labels = GANTT_TIMELINE_SHAPE_LABELS
    fills_by_mach = _equipment_gantt_fills_by_machine_name(equipment_list)
    fb_gantt = "ECEFF1" if _g_cf else "F5F5F5"
    fill_gantt_fallback = PatternFill(fill_type="solid", start_color=fb_gantt, end_color=fb_gantt)

    # タイトル＆日時（ページ上部）
    # base_dt は配台・表示レンジの基準に使うが、作成時刻は壁時計を表示する。
    create_ts = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    master_path = _master_workbook_path_resolved()

    def _fmt_mtime(p):
        try:
            if p and os.path.exists(p):
                return datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y/%m/%d %H:%M:%S")
        except Exception:
            pass
        return "—"

    master_mtime = _fmt_mtime(master_path)

    # タイトル・メタは常に D 列（4）から右へ結合。左の A〜C は比較時プルダウン、通常時は空欄のまま。
    title_start_col = 4
    row = 1
    ws.merge_cells(
        start_row=row, start_column=title_start_col, end_row=row, end_column=last_col
    )
    _title_main = (
        chart_title if chart_title is not None else "湖南工場 加工計画"
    )
    tcell = ws.cell(row=row, column=title_start_col, value=_title_main)
    tcell.font = title_font
    tcell.fill = title_fill
    # 結合セルでも左端から表示（縮尝・折り返しなし）
    tcell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=False,
        indent=1,
    )
    tcell.border = Border(left=accent_left, bottom=banner_sep)
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(
        start_row=row, start_column=title_start_col, end_row=row, end_column=last_col
    )
    meta_line = (
        f"作成　{create_ts}"
        f"　・　データ抽出　{data_extract_dt_str or '—'}"
        f"　・　マスタ（{master_workbook_filename()}）　{master_mtime}"
    )
    mtop = ws.cell(row=row, column=title_start_col, value=meta_line)
    mtop.font = meta_font
    mtop.fill = meta_fill
    mtop.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=1,
        wrap_text=False,
        shrink_to_fit=False,
    )
    mtop.border = Border(left=accent_left, bottom=banner_sep)
    ws.row_dimensions[row].height = 26
    row += 1

    dates_to_show: list = []
    for d0 in sorted_dates:
        evs0 = events_by_date.get(d0, []) if plan_rows else []
        a_evs0 = actual_events_by_date.get(d0, []) if show_actual_rows else []
        if d0 not in attendance_data:
            is_anyone_working0 = False
        else:
            is_anyone_working0 = any(
                attendance_data[d0][mm]["is_working"]
                for mm in attendance_data[d0]
                if mm in attendance_data[d0]
            )
        if not evs0 and not a_evs0 and not is_anyone_working0:
            continue
        dates_to_show.append(d0)

    # 比較ガント: 表示日プルダウン（データ検証）。候補は非表示列の退避行に格納（行の絞り込みは未実装・参照用）。
    _cmp_date_pick_row0 = 500
    if _cmp_shape and dates_to_show:
        pick_col = last_col + 1
        pz = get_column_letter(pick_col)
        for i, d0 in enumerate(dates_to_show):
            ws.cell(row=_cmp_date_pick_row0 + i, column=pick_col, value=d0.isoformat())
        try:
            ws.column_dimensions[pz].hidden = True
        except Exception:
            pass
        n_pick = len(dates_to_show)
        f1 = f"${pz}${_cmp_date_pick_row0}:${pz}${_cmp_date_pick_row0 + n_pick - 1}"
        dv = DataValidation(type="list", formula1=f1, allow_blank=True)
        ws.add_data_validation(dv)
        a1 = ws.cell(row=1, column=1, value="表示日")
        a1.font = _result_font(size=11, bold=True, color="1A1A1A")
        a1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        b1 = ws.cell(row=1, column=2, value=dates_to_show[0].isoformat())
        b1.font = _result_font(size=11, color="000000")
        b1.alignment = Alignment(horizontal="left", vertical="center", indent=0)
        dv.add(b1)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_fixed)
        hint = ws.cell(
            row=2,
            column=1,
            value="B1 で表示日を選ぶと、その日のブロック先頭へ自動でスクロールします（ThisWorkbook の SheetChange は 生産管理_AI配台テスト_ThisWorkbook_VBA.txt を参照）。",
        )
        hint.font = _result_font(size=9, color="555555")
        hint.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, indent=1
        )

    hdr_row = row
    fixed_hdr = (
        ["日付", "機械名", "タスク概覝"]
        if _cmp_shape
        else ["日付", "機械名", "工程名", "タスク概覝"]
    )
    for ci, h in enumerate(fixed_hdr, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = grid_border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    slots_hdr = [datetime.combine(dummy_d, tm) for tm in slot_times]
    for si, st in enumerate(slots_hdr):
        c = ws.cell(row=hdr_row, column=n_fixed + 1 + si, value=st.strftime("%H:%M"))
        c.font = hdr_time_font
        slot_end_t = (st + timedelta(minutes=slot_mins)).time()
        hdr_use = hdr_fill
        if rs is not None and re_ is not None:
            if not _time_intervals_overlap_half_open(st.time(), slot_end_t, rs, re_):
                hdr_use = hdr_fill_outside_regular
        c.fill = hdr_use
        c.border = grid_border
        c.alignment = Alignment(horizontal="center", vertical="bottom", textRotation=90)
    ws.row_dimensions[hdr_row].height = float(GANTT_HDR_ROW_HEIGHT_PT)
    # 先頭データ行の左上＝時刻列先頭で窓枠固定（行1〜ヘッダー行・左固定列まで）
    ws.freeze_panes = f"{get_column_letter(n_fixed + 1)}{hdr_row + 1}"
    row = hdr_row + 1

    # 日と日の間の区切り（真っ黒だと「日付ブロックの下端」と誤解されやすいため薄グレー）
    sep_fill = PatternFill(fill_type="solid", start_color="D0D0D0", end_color="D0D0D0")
    no_border = Border()

    # 印刷: 1 日ごとの手動改ページ用（各日のデータ先頭行＝機械行の開始）
    gantt_day_first_rows: list[int] = []

    for di, d in enumerate(dates_to_show):
        evs = events_by_date.get(d, [])
        a_evs_day = actual_events_by_date.get(d, []) if show_actual_rows else []

        slots = [datetime.combine(d, tm) for tm in slot_times]

        # 設備時間割と同じく ev['machine'] と equipment_list の表記ゆれを正規化して対応づける。
        # by_dm[d].get(eq) のみだとキー不一致の行が空になり、機械名毎シートだけに依頼NOが出ることがある。
        machine_to_events = defaultdict(list)
        for ev in evs:
            machine_to_events[ev["machine"]].append(ev)
        for _k_m, _evl in machine_to_events.items():
            _evl.sort(
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
            )
        machine_to_events_a = None
        if show_actual_rows:
            machine_to_events_a = defaultdict(list)
            for ev in a_evs_day:
                machine_to_events_a[ev["machine"]].append(ev)
            for _k_m2, _evl2 in machine_to_events_a.items():
                _evl2.sort(
                    key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
                )

        day_start = row
        gantt_day_first_rows.append(day_start)
        for eq in equipment_list:
            proc_nm, mach_nm = _split_equipment_line_process_machine(eq)
            mk_key = (mach_nm or "").strip() or "—"
            lab_fill = fills_by_mach.get(mk_key) or fill_gantt_fallback
            evlist = _eq_grid_events_for_equipment_column(machine_to_events, eq)
            if plan_rows:
                if evlist:
                    tids: list[str] = []
                    seen_tid: set[str] = set()
                    for e in evlist:
                        tid = str(e.get("task_id") or "").strip()
                        if tid and tid not in seen_tid:
                            seen_tid.add(tid)
                            tids.append(tid)
                    task_sum = " ".join(tids) if tids else "—"
                else:
                    task_sum = "—"

                c1 = ws.cell(row=row, column=2, value=mach_nm if mach_nm else "—")
                if _cmp_shape:
                    c3 = ws.cell(row=row, column=3, value=task_sum)
                    for c in (c1, c3):
                        c.font = _result_font(size=12, color="000000")
                        c.fill = lab_fill
                        c.border = grid_border
                    c1.font = _result_font(size=12, bold=True, color="000000")
                    c1.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    c3.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                else:
                    c2 = ws.cell(row=row, column=3, value=proc_nm if proc_nm else "—")
                    c3 = ws.cell(row=row, column=4, value=task_sum)
                    for c in (c1, c2, c3):
                        c.font = _result_font(size=12, color="000000")
                        c.fill = lab_fill
                        c.border = grid_border
                    c1.font = _result_font(size=12, bold=True, color="000000")
                    c1.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=False
                    )
                    c2.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=False
                    )
                    c3.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

                _paint_gantt_timeline_row_merged(
                    ws,
                    row,
                    n_fixed,
                    slots,
                    slot_mins,
                    evlist,
                    idle_fill,
                    break_fill,
                    gantt_label_font,
                    grid_border,
                    shape_label_specs=gantt_shape_label_specs if _use_gantt_shape_labels else None,
                    label_italic=False,
                    shape_day_key=d.isoformat() if _use_gantt_shape_labels else None,
                    show_completion_pct_in_label=False,
                    shape_line_dash=_plan_line_dash,
                    shape_line_weight_override=_plan_line_wt,
                )

                ws.row_dimensions[row].height = float(GANTT_MACHINE_ROW_HEIGHT_PT)
                row += 1

            if show_actual_rows:
                evlist_a = (
                    _eq_grid_events_for_equipment_column(machine_to_events_a, eq)
                    if machine_to_events_a is not None
                    else []
                )
                if evlist_a:
                    tids_a: list[str] = []
                    seen_aid: set[str] = set()
                    for e_a in evlist_a:
                        tid = str(e_a.get("task_id") or "").strip()
                        if not tid or tid in seen_aid:
                            continue
                        seen_aid.add(tid)
                        if not plan_rows:
                            opv = str(e_a.get("op") or "").strip()
                            subv = str(e_a.get("sub") or "").strip()
                            who_parts: list[str] = []
                            if opv:
                                who_parts.append(opv)
                            if subv:
                                for seg in subv.split(","):
                                    t = seg.strip()
                                    if t:
                                        who_parts.append(t)
                            who_show: list[str] = []
                            who_seen: set[str] = set()
                            for p in who_parts:
                                k = unicodedata.normalize("NFKC", p)
                                if k in who_seen:
                                    continue
                                who_seen.add(k)
                                who_show.append(p)
                            if who_show:
                                tids_a.append(f"{tid}（{'・'.join(who_show)}）")
                            else:
                                tids_a.append(tid)
                        else:
                            tids_a.append(tid)
                    task_sum_a = " ".join(tids_a) if tids_a else "—"
                else:
                    task_sum_a = "—"

                lab_fill_a = fills_by_mach.get(mk_key) or fill_gantt_fallback

                if mach_nm:
                    act_mach = (
                        f"{mach_nm}（実績明細）"
                        if not plan_rows
                        else f"{mach_nm}（実績）"
                    )
                elif proc_nm:
                    act_mach = "（実績明細）" if not plan_rows else "（実績）"
                else:
                    act_mach = "—"
                ca1 = ws.cell(row=row, column=2, value=act_mach)
                if _cmp_shape:
                    ca3 = ws.cell(row=row, column=3, value=task_sum_a)
                    for c in (ca1, ca3):
                        c.font = _result_font(size=12, color="000000")
                        c.fill = lab_fill_a
                        c.border = grid_border
                    ca1.font = _result_font(
                        size=12, bold=True, color="000000", italic=True
                    )
                    ca1.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    ca3.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                else:
                    ca2 = ws.cell(row=row, column=3, value=proc_nm if proc_nm else "—")
                    ca3 = ws.cell(row=row, column=4, value=task_sum_a)
                    for c in (ca1, ca2, ca3):
                        c.font = _result_font(size=12, color="000000")
                        c.fill = lab_fill_a
                        c.border = grid_border
                    ca1.font = _result_font(
                        size=12, bold=True, color="000000", italic=True
                    )
                    ca1.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=False
                    )
                    ca2.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=False
                    )
                    ca3.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

                _paint_gantt_timeline_row_merged(
                    ws,
                    row,
                    n_fixed,
                    slots,
                    slot_mins,
                    evlist_a,
                    idle_fill,
                    break_fill,
                    gantt_label_font_actual,
                    grid_border,
                    task_fill_fn=_gantt_bar_fill_actual_for_task_id,
                    label_font=gantt_label_font_actual,
                    shape_label_specs=gantt_shape_label_specs if _use_gantt_shape_labels else None,
                    label_italic=True,
                    shape_day_key=d.isoformat() if _use_gantt_shape_labels else None,
                    show_completion_pct_in_label=bool(
                        sheet_nm == RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME
                    ),
                    shape_line_dash=_act_line_dash,
                    shape_line_weight_override=_act_line_wt,
                )

                ws.row_dimensions[row].height = float(GANTT_MACHINE_ROW_HEIGHT_PT)
                row += 1

            if _show_aladdin:
                _amk = _normalize_equipment_match_key(mach_nm or "")
                _pair = ("—", "—")
                _ala_mm = ""
                if compare_aladdin_qty_by_machine_date:
                    _got = compare_aladdin_qty_by_machine_date.get((_amk, d))
                    if _got is not None:
                        _pair = (_got[0], _got[1])
                        if isinstance(_got, tuple) and len(_got) >= 3:
                            _ala_mm = str(_got[2] or "").strip()
                _ala_sum, _ala_center = _pair
                if isinstance(_ala_sum, str) and len(_ala_sum) > 32000:
                    _ala_sum = _ala_sum[:31997] + "..."
                _ala_center_show = (_ala_center or "—") + (
                    ("\n" + _ala_mm) if _ala_mm else ""
                )
                # タスク概要列（列3）は依頼NOのみ。不一致注記はタイムライン結合セルのみ。
                _ala_sum_only = _ala_sum or "—"
                _lbl_m = (
                    f"{mach_nm}（アラジン入力数量）"
                    if mach_nm
                    else "（アラジン入力数量）"
                )
                _ac1 = ws.cell(row=row, column=2, value=_lbl_m)
                _ac3 = ws.cell(row=row, column=3, value=_ala_sum_only)
                for _cx in (_ac1, _ac3):
                    _cx.font = _result_font(size=11, color="000000")
                    _cx.fill = lab_fill
                    _cx.border = grid_border
                _ac1.font = _result_font(size=11, bold=True, color="000000")
                _ac1.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                _ac3.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                if n_slots > 0 and last_col >= n_fixed + 1:
                    ws.merge_cells(
                        start_row=row,
                        start_column=n_fixed + 1,
                        end_row=row,
                        end_column=last_col,
                    )
                    _atl = ws.cell(
                        row=row,
                        column=n_fixed + 1,
                        value=_ala_center_show or "—",
                    )
                    _atl.font = _result_font(size=11, bold=False, color="333333")
                    _atl.fill = (
                        aladdin_tl_fill_mismatch if _ala_mm else aladdin_tl_fill
                    )
                    _atl.border = grid_border
                    _atl.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )

                ws.row_dimensions[row].height = float(GANTT_MACHINE_ROW_HEIGHT_PT)
                row += 1

        day_end = row - 1
        if day_end >= day_start and _use_gantt_shape_labels:
            gantt_timeline_day_blocks.append(
                {
                    "first_row": day_start,
                    "last_row": day_end,
                    "day_key": d.isoformat(),
                    "first_col": n_fixed + 1,
                    "last_col": last_col,
                }
            )
        if day_end >= day_start:
            ws.merge_cells(start_row=day_start, start_column=1, end_row=day_end, end_column=1)
            ban = ws.cell(
                row=day_start,
                column=1,
                value=f"【{d.strftime('%Y/%m/%d')}】",
            )
            ban.font = day_banner_font
            ban.fill = day_banner_fill
            # 縦書き日付は結合ブロックの上寄せ（下寄せだとセル下端に寄って見える）
            ban.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=False,
                textRotation=90,
            )
            ban.border = Border(left=accent_left, top=thin, bottom=thin, right=thin)

        if di < len(dates_to_show) - 1 and day_end >= day_start:
            for cc in range(1, last_col + 1):
                sc = ws.cell(row=row, column=cc)
                sc.value = None
                sc.fill = sep_fill
                sc.border = no_border
            ws.row_dimensions[row].height = 3
            row += 1

    if (
        _cmp_shape
        and dates_to_show
        and len(gantt_day_first_rows) == len(dates_to_show)
    ):
        _map_sr = int(_cmp_date_pick_row0)
        _mdc = COMPARE_GANTT_DAY_ROW_MAP_DATE_COL
        _mfc = COMPARE_GANTT_DAY_ROW_MAP_FIRSTROW_COL
        for _mi, _md in enumerate(dates_to_show):
            _mr = _map_sr + _mi
            ws.cell(row=_mr, column=_mdc, value=_md.isoformat())
            ws.cell(row=_mr, column=_mfc, value=int(gantt_day_first_rows[_mi]))
        for _mcol in (_mdc, _mfc):
            try:
                ws.column_dimensions[get_column_letter(_mcol)].hidden = True
            except Exception:
                pass

    # 凡例は高さ確保のため省略（モノクロ印刷は色の濃淡/セルの枠で識別）
    # 時刻列（E〜）の列幅。マクロ取り込み時は VBA 結果_設備ガント_列幅を設定 と同値に揃える。
    if n_slots > 0:
        gw = float(GANTT_TIMELINE_COLUMN_WIDTH)
        for ci in range(n_fixed + 1, last_col + 1):
            dim = ws.column_dimensions[get_column_letter(ci)]
            dim.width = gw
            # openpyxl 3.1+ では customWidth は width 有無から導出される読み取り専用のため代入しない
    if _cmp_shape:
        try:
            ws.column_dimensions["A"].width = 6.5
        except Exception:
            pass
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 40

    _gantt_scale_override_raw = (os.environ.get("GANTT_PRINT_SCALE_PERCENT", "") or "").strip()
    _gantt_print_one_page_per_day_raw = (
        os.environ.get("GANTT_PRINT_ONE_PAGE_PER_DAY", "1") or "1"
    ).strip().lower()
    _gantt_print_one_page_per_day = _gantt_print_one_page_per_day_raw not in (
        "0",
        "false",
        "no",
        "off",
        "none",
    )
    try:
        # 印刷ページ設定（結果_設備ガント／実績明細ガント共通。作成完了時点で付与）
        # ① 用紙 A3・横向き  ② 余白「狭い」  ③ 列見出しとして繰り返す行＝1〜3 行目
        # ④ 横を 1 ページに収める  ⑤ 縦は自動（ページ当たり行は自動）／⑥ 1 暦日≒1 ページを既定 ON（環境変数で無効可）
        # GANTT_PRINT_SCALE_PERCENT 指定時は①〜④を縮小率指定に切替（⑥ は無効）。
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 8
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        ws.print_title_rows = "1:3"
        if _gantt_scale_override_raw:
            _pct = max(10, min(400, int(_gantt_scale_override_raw)))
            ws.page_setup.fitToPage = False
            ws.page_setup.fitToWidth = False
            ws.page_setup.fitToHeight = False
            ws.page_setup.scale = _pct
        else:
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            if _gantt_print_one_page_per_day and gantt_day_first_rows:
                ws.page_setup.fitToHeight = max(1, len(gantt_day_first_rows))
            else:
                ws.page_setup.fitToHeight = 0
        # タイトル・表をページ左基準に（レポート風）
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False
        # 比較ガントは結合セルが多いため、印刷時もグリッドを出して区切りを補助
        ws.print_options.gridLines = bool(_cmp_shape)
    except Exception:
        pass

    if _cmp_shape:
        try:
            if ws.views.sheetView:
                ws.views.sheetView[0].showGridLines = True
        except Exception:
            pass

    # 1 日 1 ページ相当: 2 日目以降の各日ブロック先頭の直前に手動の横改ページ（上記ページ設定の後）
    try:
        if len(gantt_day_first_rows) > 1:
            for i in range(1, len(gantt_day_first_rows)):
                ws.row_breaks.append(Break(id=gantt_day_first_rows[i], man=True))
    except Exception:
        pass

    if _cmp_shape:
        _apply_compare_gantt_typography(ws, hdr_row)

    if _use_gantt_shape_labels:
        return gantt_shape_label_specs, gantt_timeline_day_blocks
    return [], []
def row_has_completion_keyword(row):
    """加工完了区分に「完了」の文字は含まれる場合はタスク完了とみなす。"""
    v = row.get(TASK_COL_COMPLETION_FLAG)
    if v is None or pd.isna(v):
        return False
    return "完了" in str(v)
def _planning_completion_flag_cell_is_mikan(v) -> bool:
    """加工完了区分がセル値として「未完」とみなすか（NFKC・前後空白除去）。

    セルが「0:未完」のように区分値とコロンで前置される場合は、**最後のコロン以降**が厳密に
    「未完」のときのみ True（実データで未完区分が数値プレフィックス付きで格納される）。
    「未完了」は末尾が「未完」にならないため False のまま。
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = unicodedata.normalize("NFKC", str(v).strip())
    s = s.replace("\uff1a", ":").replace("：", ":")
    tail = s.rsplit(":", 1)[-1].strip() if ":" in s else s
    return tail == "未完"
def _plan_row_exclude_as_completed_mikan_unprocessed_zero_actual_done_rule(row) -> bool:
    """
    加工計画DATA／配台計画_タスク入力の同一列前提で、次をすべて満たす行は加工済みとみなし配台対象外とする。

    - 「未加工」列があり数値 0（空・列無しは対象外）
    - 「実加工数」が 0 以外
    - 「加工完了区分」が「未完」または「0:未完」形式で末尾が「未完」（「未完了」等は含めない）
    """
    cf_v = row.get(TASK_COL_COMPLETION_FLAG)
    act_v = parse_float_safe(row.get(TASK_COL_ACTUAL_DONE), 0.0)
    unp_v = _optional_unprocessed_m_from_plan_row(row)
    ok_mikan = _planning_completion_flag_cell_is_mikan(cf_v)
    ok_act = abs(act_v) > 1e-12
    ok_unp = unp_v is not None and abs(float(unp_v)) <= 1e-12
    return bool(ok_mikan and ok_act and ok_unp)
def _plan_row_stage2_dispatch_plan_excluded(row) -> bool:
    """「配台不要」セルに配台計画除外マーカーが含まれる行は段階2の task_queue に載せない。"""
    v = row.get(PLAN_COL_EXCLUDE_FROM_ASSIGNMENT)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = unicodedata.normalize("NFKC", str(v).strip())
    if not s or s.lower() in ("nan", "none"):
        return False
    return PLAN_COL_STAGE2_DISPATCH_PLAN_EXCLUDE_MARKER in s
def _plan_row_exclude_from_assignment(row) -> bool:
    """
    「配台試行」列はオンなら」しの行は配台キューへ入れう」特別指定_備考の AI 解析行からも除し。

    配台から外れ（真）: 論睆値 True」数値 1」文字列（NFKC 後・尝文字）
      true / 1 / yes / on / y / t / はい / ○ / 〇 / ◝
    配台対象（坽）: 空」None」False」0」no / off / false / いいえ / 坦 等
    上記以外の文字列は坽（配台れる）。チェックボックス連動セルは通常 TRUE/FALSE または 1/0。
    """
    v = row.get(PLAN_COL_EXCLUDE_FROM_ASSIGNMENT)
    if v is True:
        return True
    if v is False:
        return False
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            iv = int(v)
            if iv == 1:
                return True
            if iv == 0:
                return False
        except (TypeError, ValueError):
            pass
    s = unicodedata.normalize("NFKC", str(v).strip()).lower()
    if not s or s in ("nan", "none", "false", "0", "no", "off", "いいえ", "坦"):
        return False
    if s in ("true", "1", "yes", "on", "はい", "y", "t", "○", "〇", "◝"):
        return True
    return False
def _coerce_plan_exclude_column_value_for_storage(v):
    """
    「配台試行」列へ書き込む値を」StringDtype 列でも代入エラーにならない形にしゝごる。
    Excel 取り込みの True / 1 / False / 0 と文字列を保挝し、_plan_row_exclude_from_assignment と整合する。
    """
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if v is True:
        return "yes"
    if v is False:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            iv = int(v)
            if iv == 1:
                return "yes"
            if iv == 0:
                return ""
        except (TypeError, ValueError):
            pass
    return str(v).strip()
def parse_float_safe(val, default=0.0):
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return default
        return float(val)
    except (TypeError, ValueError):
        return default
def _optional_unprocessed_m_from_plan_row(row) -> float | None:
    """行の「未加工」セルを数値化。列が無い・空なら None。"""
    if row is None:
        return None
    try:
        idx = row.index  # type: ignore[attr-defined]
    except AttributeError:
        return None
    if TASK_COL_UNPROCESSED not in idx:
        return None
    return _optional_float_unprocessed_column(row.get(TASK_COL_UNPROCESSED))
def _supplement_task_input_unprocessed_column_if_missing(df: pd.DataFrame) -> None:
    """
    PQ 出力などで「未加工」列が無い、またはセルが空・非数値の行がある場合に補う。
    換算数量と実加工数があれば 未加工 = max(0, 換算数量 − 実加工数)。
    実加工数も無ければ未加工に換算数量を入れる（全量が残作とみなす）。

    工程別問合せ xlsx では「未加工」列が存在しても未入力の行があり得るため、
    列がある場合も欠損セルのみ上書きする。
    """
    if df is None or getattr(df, "empty", True):
        return
    if TASK_COL_QTY not in df.columns:
        return
    qv = pd.to_numeric(df[TASK_COL_QTY], errors="coerce").fillna(0.0)
    if TASK_COL_ACTUAL_DONE in df.columns:
        av = pd.to_numeric(df[TASK_COL_ACTUAL_DONE], errors="coerce").fillna(0.0)
        computed = (qv - av).clip(lower=0.0)
    else:
        computed = qv

    if TASK_COL_UNPROCESSED not in df.columns:
        df[TASK_COL_UNPROCESSED] = computed
        if TASK_COL_ACTUAL_DONE in df.columns:
            logging.info(
                "タスク入力: 列「%s」が無いため「%s」−「%s」で補完しました。",
                TASK_COL_UNPROCESSED,
                TASK_COL_QTY,
                TASK_COL_ACTUAL_DONE,
            )
        else:
            logging.info(
                "タスク入力: 列「%s」及び「%s」が無いため「%s」を未加工にコピーしました（残量＝換算数量とみなします）。",
                TASK_COL_UNPROCESSED,
                TASK_COL_ACTUAL_DONE,
                TASK_COL_QTY,
            )
        return

    cur = pd.to_numeric(df[TASK_COL_UNPROCESSED], errors="coerce")
    need = cur.isna()
    if not need.any():
        return
    df.loc[need, TASK_COL_UNPROCESSED] = computed[need]
    if TASK_COL_ACTUAL_DONE in df.columns:
        logging.info(
            "タスク入力: 列「%s」の欠損セル %s 件を「%s」−「%s」で補完しました。",
            TASK_COL_UNPROCESSED,
            int(need.sum()),
            TASK_COL_QTY,
            TASK_COL_ACTUAL_DONE,
        )
    else:
        logging.info(
            "タスク入力: 列「%s」の欠損セル %s 件を「%s」で補完しました。",
            TASK_COL_UNPROCESSED,
            int(need.sum()),
            TASK_COL_QTY,
        )
def _ensure_dataframe_has_unprocessed_column(
    df: pd.DataFrame, *, context_label: str
) -> None:
    """加工計画DATA／配台計画_タスク入力に「未加工」列が無いとき配台を中止する。"""
    if df is None:
        raise PlanningValidationError(
            f"{context_label}: 列「{TASK_COL_UNPROCESSED}」が必須です。"
            "この列が無いため配台処理を中止します。"
        )
    if TASK_COL_UNPROCESSED not in df.columns:
        raise PlanningValidationError(
            f"{context_label}: 列「{TASK_COL_UNPROCESSED}」が必須です。"
            "この列が無いため配台処理を中止します。"
        )
def aladdin_system_dispatch_display_qty_m(
    dispatch_qty_m: float,
    qty_conv_m: float,
    raw_roll_m: float,
    *,
    remaining_conv_cap: float | None = None,
) -> tuple[float, float | None]:
    """
    納期管理ビュー・サマリ Excel の (段階3前) 表示用数量。

    換算数量 < (原反)ロール単位長さ のとき、アラジン再入力値は換算数量（配台タイムラインの m ではない）。
    ``remaining_conv_cap`` 指定時は依頼NO単位で換算数量を超えないよう暦日順に配分する。
    配台計算本体は従来どおり原反ロール長ベースのタイムライン数量を使う。

    Returns:
        (display_m, new_remaining_conv_cap or None)
    """
    dq = max(0.0, float(dispatch_qty_m))
    if dq <= 1e-12:
        cap = remaining_conv_cap
        if cap is not None:
            return 0.0, max(0.0, float(cap))
        return 0.0, None
    q = max(0.0, float(qty_conv_m))
    r = max(0.0, float(raw_roll_m))
    if not (q > 1e-12 and r > 1e-12 and q + 1e-9 < r):
        return dq, remaining_conv_cap
    if remaining_conv_cap is not None:
        cap = max(0.0, float(remaining_conv_cap))
        show = min(dq, cap)
        return show, max(0.0, cap - show)
    return min(dq, q), None
def _raw_roll_unit_m_resolved_for_dispatch_qty(row) -> float:
    """
    (原反)ロール単位長さ相当の正の m。配台計画行に列があれば優先し、
    無い・「不明」のときは使用原反セルからテーブル→寸法で解決する。
    """
    if hasattr(row, "get"):
        v = _planning_df_cell_scalar(row, PLAN_COL_RAW_ROLL_UNIT_LENGTH)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            s = str(v).strip()
            if s and s != "不明" and s.lower() not in ("nan", "none"):
                m = parse_float_safe(v, 0.0)
                if m > 1e-12:
                    return float(m)
        ur = row.get(TASK_COL_USED_RAW)
    else:
        ur = None
    tab = _lookup_roll_unit_length_m_from_used_raw(ur)
    if tab is not None and float(tab) > 1e-12:
        return float(tab)
    dim = _parse_roll_unit_m_from_used_raw_dimension_only(ur)
    return float(dim) if dim is not None and dim > 1e-12 else 0.0
def _dispatch_simulator_unit_m_from_plan_row(row, *, fallback_m: float) -> float:
    """
    配台シミュレーション ``task_queue`` の ``unit_m`` (m/ロール)。

    **(原反)ロール単位長さ**（列→使用原反テーブル→使用原反寸法）を正とし、
    解決できないときだけ ``infer_roll_unit_m_from_used_raw_then_product_dims`` で補う。
    """
    unit = _raw_roll_unit_m_resolved_for_dispatch_qty(row)
    if unit <= 1e-12 and hasattr(row, "get"):
        unit = infer_roll_unit_m_from_used_raw_then_product_dims(
            row.get(TASK_COL_PRODUCT, None),
            row.get(TASK_COL_USED_RAW, None),
            fallback_unit=fallback_m,
        )
    try:
        unit = float(unit)
    except (TypeError, ValueError):
        unit = 0.0
    if unit <= 1e-12:
        unit = float(fallback_m)
    return unit
def _plan_row_dispatch_qty_metrics(row):
    """
    結果シート・配台メトリクス用の残り(m)・済相当(m)・総量(m)を返す。

    **正**: 段階1の列「配台使用残数量」「配台ロール数」（欠損時は段階1と同一式で補完）。
    済相当m = max(0, 換算数量(raw) - 配台使用残数量)。
    総量m = 残り + 済相当（＝換算数量 raw。100m 切上げは行わない）。

    **未加工列**は行の有効性検証のみ（空・非数値は ``PlanningValidationError``）。
    実出来高・実加工数から済相当へ直接フォールバックしない。

    Returns:
        tuple[float, float, float, bool]:
            (remaining_m, done_m, qty_total_for_dispatch_m, used_unprocessed)
    """
    unp = _optional_unprocessed_m_from_plan_row(row)
    if unp is None:
        raise PlanningValidationError(
            f"「{TASK_COL_UNPROCESSED}」が数値として読めません（セルが空または不正）、"
            "または列がありません。配台計画行の検証のため未加工列が必要です。"
        )
    qty_conv_raw = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
    remaining_m = _plan_cell_dispatch_remaining_m(row)
    done_m = max(0.0, qty_conv_raw - remaining_m)
    qty_total_for_dispatch_m = remaining_m + done_m
    return remaining_m, done_m, qty_total_for_dispatch_m, True
def _dispatch_remaining_qty_m_from_row(row) -> float:
    """
    配台計画_タスク入力の列「配台使用残数量」(m)。

    ① B = 換算数量 − 実加工数
    ② A = ceil(B / (原反)ロール単位長さ)（B≦0 のとき A=0）
    ③ 配台使用残数量 = (原反)ロール単位長さ × A

    (原反)ロール単位長さが解決できないときは B をそのまま返す。
    """
    qty_conv = parse_float_safe(row.get(TASK_COL_QTY), 0.0)
    actual_done = parse_float_safe(row.get(TASK_COL_ACTUAL_DONE), 0.0)
    b = max(0.0, qty_conv - actual_done)
    raw_roll_m = _raw_roll_unit_m_resolved_for_dispatch_qty(row)
    if raw_roll_m <= 1e-12:
        return b
    if b <= 1e-12:
        return 0.0
    n_rolls = math.ceil(b / raw_roll_m)
    return float(raw_roll_m) * float(n_rolls)
def _dispatch_roll_count_from_row(row, remaining_m: float) -> float | int | str:
    """配台使用残数量(m) ÷ (原反)ロール単位長さ。原反ロール長が無いときは空文字。"""
    raw_roll_m = _raw_roll_unit_m_resolved_for_dispatch_qty(row)
    rem = max(0.0, float(remaining_m))
    if raw_roll_m <= 1e-12:
        return ""
    if rem <= 1e-12:
        return 0
    n = rem / float(raw_roll_m)
    if abs(n - round(n)) <= 1e-9:
        return int(round(n))
    return n
def _parse_plan_dispatch_roll_count_cell(val) -> float | None:
    """シート列「配台ロール数」のセル値。数値化できないとき None。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    m = parse_float_safe(val, float("nan"))
    if isinstance(m, float) and pd.isna(m):
        return None
    return float(m)
def _plan_cell_dispatch_remaining_m(row) -> float:
    """段階2用: 列「配台使用残数量」があれば採用、無ければ段階1と同一式で算出。"""
    if hasattr(row, "get"):
        v = _planning_df_cell_scalar(row, PLAN_COL_DISPATCH_REMAINING_QTY)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            s = str(v).strip()
            if s and s.lower() not in ("nan", "none"):
                m = parse_float_safe(v, -1.0)
                if m >= 0:
                    return max(0.0, float(m))
    return _dispatch_remaining_qty_m_from_row(row)
def _plan_row_stage2_dispatch_qty_and_rolls(row) -> tuple[float, float]:
    """
    段階2配台の加工量(m)とロール本数。

    正: 列「配台使用残数量」「配台ロール数」。欠損・空のときは
    ``_dispatch_remaining_qty_m_from_row`` / ``_dispatch_roll_count_from_row`` で補完。
    """
    rem_m = _plan_cell_dispatch_remaining_m(row)
    rolls = None
    if hasattr(row, "get"):
        rolls = _parse_plan_dispatch_roll_count_cell(
            _planning_df_cell_scalar(row, PLAN_COL_DISPATCH_ROLL_COUNT)
        )
    if rolls is None:
        rc = _dispatch_roll_count_from_row(row, rem_m)
        if rc == "":
            rolls = 0.0
        else:
            rolls = float(rc)
    return max(0.0, rem_m), max(0.0, float(rolls))
def _fill_plan_dispatch_remaining_qty_column(plan_df: pd.DataFrame) -> None:
    """配台計画 DataFrame の「配台使用残数量」「配台ロール数」を段階1式で埋める。"""
    if plan_df is None or getattr(plan_df, "empty", True):
        return
    if PLAN_COL_DISPATCH_REMAINING_QTY not in plan_df.columns:
        return
    fill_roll_count = PLAN_COL_DISPATCH_ROLL_COUNT in plan_df.columns
    for i in plan_df.index:
        row = plan_df.loc[i]
        rem = _dispatch_remaining_qty_m_from_row(row)
        plan_df.at[i, PLAN_COL_DISPATCH_REMAINING_QTY] = rem
        if fill_roll_count:
            rc = _dispatch_roll_count_from_row(row, rem)
            if rc == "":
                plan_df.at[i, PLAN_COL_DISPATCH_ROLL_COUNT] = 0.0
            else:
                plan_df.at[i, PLAN_COL_DISPATCH_ROLL_COUNT] = float(rc)
def parse_optional_int(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    try:
        return int(round(float(s)))
    except (TypeError, ValueError):
        return None
def parse_optional_date(val):
    if val is None or pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None
def _parse_env_optional_date(env_key: str):
    """os.environ の 1 キーを暦日に解釈。空・解釈不能は None。"""
    raw = (os.environ.get(env_key) or "").strip()
    if not raw:
        return None
    return parse_optional_date(raw)
def parse_optional_datetime(val):
    """配台可能日時など datetime 文字列を解釈。空・解釈不能は None。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return pd.to_datetime(val).to_pydatetime()
    except Exception:
        return None
def compute_dispatchable_datetime(raw_input_date, run_date=None, stock_location=None):
    """原反投入日（上書き優先で解決済みの date）から配台可能日時を算出。

    日付 = max(run_date, raw_input_date)（run_date 指定時のみ）、時刻 = DISPATCHABLE_FROM_TIME。
    raw_input_date が None のときは None（原反投入日が無い行は配台可能日時を持たない）。

    湖南工場（PM_AI_FACTORY_SITE=KONAN）かつ ``stock_location`` が「湖南」を含むときは、
    時刻に DISPATCHABLE_FROM_TIME_KONAN_STOCK（既定9:30）を使う。
    """
    if raw_input_date is None:
        return None
    base_date = raw_input_date
    if run_date is not None and run_date > base_date:
        base_date = run_date
    return datetime.combine(base_date, dispatchable_from_time_for(stock_location))
def format_dispatchable_datetime_cell(dt) -> str:
    """配台可能日時セルの出力文字列（YYYY/MM/DD HH:MM）。None は空文字。"""
    if dt is None:
        return ""
    return dt.strftime("%Y/%m/%d %H:%M")
def resolve_dispatchable_datetime_from_plan_row(
    row, run_date=None, *, regular_shift_start: time | None = None
):
    """配台計画1行から配台可能日時を解決。

    優先順: 列「配台可能日時」（入力3表・入力1表とも）→ 未設定時のフォールバック。
    列に値があるとき原反投入日は暦日・時刻の下限に使わない。

    ``regular_shift_start`` を渡したとき（段階3配台読込）は、未設定時に
    原反投入日+12:45 ではなく ``max(run_date, 原反投入日)`` の暦日 + 定常開始時刻（A15）とする。
    """
    col = parse_optional_datetime(
        _planning_df_cell_scalar(row, PLAN_COL_DISPATCHABLE_DATETIME)
    )
    if col is not None:
        return col
    if regular_shift_start is not None:
        raw = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE))
        if raw is None:
            if run_date is None:
                return None
            base_d = run_date
        else:
            base_d = max(run_date, raw) if run_date is not None else raw
        return datetime.combine(base_d, regular_shift_start)
    raw = parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE))
    stock_location = _planning_df_cell_scalar(row, TASK_COL_STOCK_LOCATION)
    return compute_dispatchable_datetime(raw, run_date=run_date, stock_location=stock_location)
def stage3_regular_shift_start_time() -> time:
    """段階3: master メイン A15（定常開始）。読めないときは ``DEFAULT_START_TIME``。"""
    try:
        st, _et = _read_master_main_regular_shift_times(_master_workbook_path_resolved())
        if st is not None:
            return st
    except Exception:
        pass
    return DEFAULT_START_TIME
def _planning_df_cell_scalar(row, col_name):
    """
    iterrows() 1行分から列値を得る。同一見出しの重複列はあると row.get は Series になり」
    str→to_datetime で誤った日付になることがあるため、先頭の非欠損スカラーを返す。
    """
    v = row.get(col_name) if hasattr(row, "get") else None
    if isinstance(v, pd.Series):
        for x in v:
            if x is None or (isinstance(x, float) and pd.isna(x)):
                continue
            return x
        return None
    return v
def _roll_unit_m_estimate_from_plan_row(row, fallback_m: float) -> float:
    """
    配台計画1行から 1 ロールあたりの長さ(m)。シートのロール単位長さを優先し、
    空・0 のときは製品名テーブル→製品名寸法で推定する（build_task_queue と同趣旨）。
    """
    product_name = row.get(TASK_COL_PRODUCT, None) if hasattr(row, "get") else None
    unit = parse_float_safe(_planning_df_cell_scalar(row, PLAN_COL_ROLL_UNIT_LENGTH), 0.0)
    fb = max(1e-9, float(parse_float_safe(fallback_m, 0.0)))
    if unit <= 0:
        unit = infer_unit_m_from_product_name(product_name, fallback_unit=fb)
    try:
        unit = float(unit)
    except (TypeError, ValueError):
        unit = 0.0
    if unit <= 0:
        unit = fb
    return float(unit)
def _effective_roll_unit_m_for_dispatch_task_simulator(
    qty_m: float, sheet_roll_unit_m: float
) -> float:
    """
    換算数量（配台に使う残 m）を (原反)ロール単位長さで割ったとき整数ロールにならない場合、
    作業ロール数を ``floor(換算数量 / ロール単位)`` 本（少なくとも 1 本）とし、
    ``換算数量 / 作業ロール数`` を配台シミュレータ用の実効 1 ロール長さ (m) として返す。

    例: 800 m ÷ 95 m → 8.42… 本 → 8 本、800 ÷ 8 = 100 m を実効ロール単位とする。

    既に（誤差範囲内で）整数ロールに収まるとき、または数量・単位が不正なときは
    シートのロール単位長さをそのまま返す。
    """
    q = parse_float_safe(qty_m, 0.0)
    u = parse_float_safe(sheet_roll_unit_m, 0.0)
    if q <= 1e-12 or u <= 1e-12:
        return float(u) if u > 1e-12 else 0.0
    n_raw = q / u
    if n_raw <= 1e-12:
        return float(u)
    if abs(n_raw - round(n_raw)) <= 1e-9:
        return float(u)
    n_work = int(math.floor(n_raw))
    if n_work < 1:
        n_work = 1
    return float(q) / float(n_work)
def load_ai_cache():
    try:
        if os.path.exists(ai_cache_path):
            with open(ai_cache_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    # 期陝切れエントリを除去（96時間）
                    now_ts = time_module.time()
                    cleaned = {}
                    expired_count = 0
                    for k, v in data.items():
                        # 新形式: {"ts": epoch_seconds, "data": {...}}
                        if isinstance(v, dict) and "ts" in v and "data" in v:
                            ts = parse_float_safe(v.get("ts"), 0.0)
                            if ts > 0 and (now_ts - ts) <= AI_CACHE_TTL_SECONDS:
                                cleaned[k] = v
                            else:
                                expired_count += 1
                        # 旧形式: 値は直接AI結果dict（互換で読み取り」坳時に新形式へ再保存される）
                        else:
                            cleaned[k] = {"ts": now_ts, "data": v}
                    if expired_count > 0:
                        logging.info(f"AIキャッシュ期陝切れを削除: {expired_count}件")
                    return cleaned
    except Exception as e:
        logging.warning(f"AIキャッシュ読み込み失敗: {e}")
    return {}
def save_ai_cache(cache_obj):
    try:
        with open(ai_cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_obj, f, ensure_ascii=False)
    except Exception as e:
        logging.warning(f"AIキャッシュ保存失敗: {e}")
def get_cached_ai_result(cache_obj, cache_key, content_key=None):
    """
    content_key: オプション。保存時と同一の文字列でないヒットは無効化する（特別指定・照合用の二次チェック）。
    旧エントリに content_key は無い場合は SHA256 キー一致のみで従来どおりヒットとみなす。
    """
    entry = cache_obj.get(cache_key)
    if not isinstance(entry, dict):
        return None
    ts = parse_float_safe(entry.get("ts"), 0.0)
    if ts <= 0:
        return None
    if (time_module.time() - ts) > AI_CACHE_TTL_SECONDS:
        return None
    if content_key is not None:
        stored_ck = entry.get("content_key")
        if stored_ck is not None and stored_ck != content_key:
            logging.info(
                "AIキャッシュ: キーは一致したが、content_key は実行入力と異なるため無効化した。"
            )
            return None
    data = entry.get("data")
    if isinstance(data, dict):
        return data
    return None
def put_cached_ai_result(cache_obj, cache_key, parsed_obj, content_key=None):
    payload = {"ts": time_module.time(), "data": parsed_obj}
    if content_key is not None:
        payload["content_key"] = content_key
    cache_obj[cache_key] = payload
def extract_retry_seconds(err_text):
    # 例: "Please retry in 57.089735313s."
    m = re.search(r"retry in ([0-9]+(?:\.[0-9]+)?)s", err_text, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    # 例: "'retryDelay': '57s'"
    m = re.search(r"retryDelay'\s*:\s*'([0-9]+)s'", err_text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None
class GeminiApiSkippedError(RuntimeError):
    """PM_AI_SKIP_GEMINI_API により generate_content を意図的にスキップ。"""
def _normalize_product_dim_separators_for_roll_inference(s: str) -> str:
    """
    製品名に混ざる寸法区切りを ASCII の x に寄せる。
    先に NFKC で互換分解（全角英数字・互換記号など）を寄せ、列名 `_align_dataframe_headers_to_canonical`
    と同趣旨に Excel 由来の表記ゆれを弱める。
    半角 X/x 以外（×・全角Ｘｘ・罫線系の乗号）だけがあると正規表現に一致せず、
    換算数量フォールバックでロール単位長さが誤ることがある。
    """
    if not s:
        return s
    t = unicodedata.normalize("NFKC", s)
    for ch in (
        "\u00d7",  # × MULTIPLICATION SIGN
        "\u2715",  # ✕ MULTIPLICATION X
        "\u2716",  # ✖ HEAVY MULTIPLICATION X
        "\u2a2f",  # ⨯ VECTOR OR CROSS PRODUCT
        "\u2a09",  # ⨉ CROSS MULTIPLICATION
        "\uff38",  # Ｘ FULLWIDTH LATIN CAPITAL LETTER X
        "\uff58",  # ｘ FULLWIDTH LATIN SMALL LETTER X
        # 寸法区切りに誤入力されがちな「X に見えるが ASCII [xX] にマッチしない」文字（推定失敗→換算数量→100m 切上で 870→900 等）
        "\u0425",  # CYRILLIC CAPITAL LETTER HA
        "\u0445",  # CYRILLIC SMALL LETTER HA
        "\u03a7",  # GREEK CAPITAL LETTER CHI
        "\u03c7",  # GREEK SMALL LETTER CHI
    ):
        t = t.replace(ch, "x")
    return t
ROLL_UNIT_LENGTH_TABLE_DEFAULT_FILENAME = "製品名,ロール単位の長さ.txt"
ROLL_UNIT_LENGTH_TABLE_PATH_ENV = "ROLL_UNIT_LENGTH_TABLE_PATH"
_ROLL_UNIT_LENGTH_TABLE_CACHE: dict[str, float] | None = None
_ROLL_UNIT_LENGTH_TABLE_PATH_USED: str | None = None
ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME = "使用原反,ロール単位の長さ.txt"
ROLL_UNIT_BY_USED_RAW_TABLE_ALT_FILENAME = "使用原反, ロール単位の長さ.txt"
ROLL_UNIT_BY_USED_RAW_TABLE_PATH_ENV = "ROLL_UNIT_BY_USED_RAW_TABLE_PATH"
_ROLL_UNIT_BY_USED_RAW_TABLE_CACHE: dict[str, float] | None = None
_ROLL_UNIT_BY_USED_RAW_TABLE_PATH_USED: str | None = None
_DISPATCH_LOOKUP_TABLE_FILENAMES: tuple[str, ...] = (
    ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME,
    ROLL_UNIT_LENGTH_TABLE_DEFAULT_FILENAME,
    PRODUCT_WIDTH_TABLE_DEFAULT_FILENAME,
    PRODUCT_THICKNESS_TABLE_DEFAULT_FILENAME,
    PRODUCT_LENGTH_TABLE_DEFAULT_FILENAME,
    RAW_FABRIC_WIDTH_TABLE_DEFAULT_FILENAME,
)
def _resolve_bundled_dispatch_lookup_table_in_repo(filename: str) -> str | None:
    for code_dir in _planning_code_dir_candidates():
        cand = os.path.join(code_dir, filename)
        if os.path.isfile(cand):
            return os.path.normpath(os.path.abspath(cand))
    return None
def _ensure_dispatch_lookup_tables_at_work_path() -> None:
    """サマリ Excel 同フォルダに材料テーブルが無ければ code/ 同梱からコピーする。"""
    for filename in _DISPATCH_LOOKUP_TABLE_FILENAMES:
        target = _summary_ai_dispatch_workbook_sibling_path(filename)
        if not target or os.path.isfile(target):
            continue
        bundled = _resolve_bundled_dispatch_lookup_table_in_repo(filename)
        if bundled and _copy_exclude_rules_json_if_missing(target, bundled):
            logging.info(
                "材料テーブルをリポジトリ同梱から作業先へコピーしました（%s → %s）。",
                bundled,
                target,
            )
def _normalize_roll_unit_length_table_key(val) -> str:
    """
    ロール単位長さテーブルの照会キーを正規化する。
    先に NFKC で全角英数字・互換記号などを半角へ寄せたうえで、
    半角・全角などあらゆる空白類（isspace）を除去してから照合する
    （Excel 由来の U+3000 や NBSP、連続スペースの差で一致しないのを防ぐ）。
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = unicodedata.normalize("NFKC", str(val).strip())
    return "".join(ch for ch in s if not ch.isspace())
def _roll_unit_length_table_search_paths() -> list[str]:
    """ロール単位長さテーブル CSV の探索順（先に見つかったパスを採用）。"""
    paths: list[str] = []
    env = (os.environ.get(ROLL_UNIT_LENGTH_TABLE_PATH_ENV) or "").strip()
    if env:
        paths.append(env)
    sibling = _summary_ai_dispatch_workbook_sibling_path(
        ROLL_UNIT_LENGTH_TABLE_DEFAULT_FILENAME
    )
    if sibling:
        paths.append(sibling)
    wb = (_excel_plan_input_wb() or "").strip()
    if wb:
        paths.append(
            os.path.join(
                os.path.dirname(os.path.abspath(wb)),
                ROLL_UNIT_LENGTH_TABLE_DEFAULT_FILENAME,
            )
        )
    paths.append(os.path.join(os.getcwd(), ROLL_UNIT_LENGTH_TABLE_DEFAULT_FILENAME))
    paths.append(os.path.join(os.getcwd(), "code", ROLL_UNIT_LENGTH_TABLE_DEFAULT_FILENAME))
    out: list[str] = []
    seen: set[str] = set()
    for p in paths:
        key = os.path.normcase(os.path.abspath(p))
        if key not in seen:
            seen.add(key)
            out.append(p)
    return out
def _roll_unit_by_used_raw_table_search_paths() -> list[str]:
    """使用原反→ロール単位長さ CSV の探索順（先に見つかったパスを採用）。"""
    paths: list[str] = []
    env = (os.environ.get(ROLL_UNIT_BY_USED_RAW_TABLE_PATH_ENV) or "").strip()
    if env:
        paths.append(env)
    sibling_dir = os.path.dirname(_resolve_summary_ai_dispatch_workbook_path())
    if sibling_dir:
        for fn in (
            ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME,
            ROLL_UNIT_BY_USED_RAW_TABLE_ALT_FILENAME,
        ):
            paths.append(os.path.join(sibling_dir, fn))
    wb = (_excel_plan_input_wb() or "").strip()
    if wb:
        bd = os.path.dirname(os.path.abspath(wb))
        for fn in (
            ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME,
            ROLL_UNIT_BY_USED_RAW_TABLE_ALT_FILENAME,
        ):
            paths.append(os.path.join(bd, fn))
    for cwd in (os.getcwd(), os.path.join(os.getcwd(), "code")):
        for fn in (
            ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME,
            ROLL_UNIT_BY_USED_RAW_TABLE_ALT_FILENAME,
        ):
            paths.append(os.path.join(cwd, fn))
    out: list[str] = []
    seen: set[str] = set()
    for p in paths:
        key = os.path.normcase(os.path.abspath(p))
        if key not in seen:
            seen.add(key)
            out.append(p)
    return out
def _load_roll_unit_length_m_by_used_raw_table_optional() -> dict[str, float]:
    """
    使用原反→ロール単位の長さ(m) テーブルを読み込む。
    ファイルが無い場合は空 dict（製品名寸法推定へフォールバック）。
    キーは原反幅テーブルと同じ _normalize_mm_table_lookup_key。
    """
    global _ROLL_UNIT_BY_USED_RAW_TABLE_PATH_USED
    path_found = ""
    for p in _roll_unit_by_used_raw_table_search_paths():
        if os.path.isfile(p):
            path_found = p
            break
    if not path_found:
        return {}
    out: dict[str, float] = {}
    try:
        with open(path_found, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
    except OSError:
        return {}
    if not rows:
        return {}
    hdr = [_normalize_mm_table_lookup_key(x) for x in rows[0]]
    try:
        i_key = hdr.index(_normalize_mm_table_lookup_key("使用原反"))
    except ValueError:
        i_key = 0
    try:
        i_m = hdr.index(_normalize_mm_table_lookup_key("ロール単位の長さ"))
    except ValueError:
        i_m = 1 if len(hdr) > 1 else 0
    for parts in rows[1:]:
        if not parts or all(not str(x).strip() for x in parts):
            continue
        while len(parts) <= max(i_key, i_m):
            parts.append("")
        raw_k = parts[i_key]
        raw_m = parts[i_m]
        key = _normalize_mm_table_lookup_key(raw_k)
        if not key:
            continue
        m = parse_float_safe(raw_m, 0.0)
        if m <= 0:
            continue
        if key in out and abs(out[key] - float(m)) > 1e-9:
            logging.warning(
                "使用原反ロール単位長さテーブルで同一キーに矛盾する値があります: %r → %s と %s (%s)",
                key,
                out[key],
                m,
                path_found,
            )
            continue
        out[key] = float(m)
    _ROLL_UNIT_BY_USED_RAW_TABLE_PATH_USED = path_found
    if out:
        logging.info(
            "使用原反ロール単位長さテーブルを読み込みました: %s (%s 件)",
            path_found,
            len(out),
        )
    return out
def _load_used_raw_roll_length_table_stage1() -> tuple[dict[str, float], set[str], str]:
    """
    段階1向け: 使用原反→ロール単位長さテーブル。ファイル無しは空 dict と書込先パスのみ。
    値が空欄の行は known_keys のみ（dict には載せない）。
    """
    path_write = _resolve_code_lookup_table_path_for_write(
        _roll_unit_by_used_raw_table_search_paths(),
        ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME,
    )
    path_found = ""
    for p in _roll_unit_by_used_raw_table_search_paths():
        if os.path.isfile(p):
            path_found = p
            break
    if not path_found:
        return {}, set(), path_write
    out: dict[str, float] = {}
    known_keys: set[str] = set()
    try:
        with open(path_found, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
    except OSError:
        return {}, set(), path_write
    if not rows:
        return {}, set(), path_found
    hdr = [_normalize_mm_table_lookup_key(x) for x in rows[0]]
    try:
        i_key = hdr.index(_normalize_mm_table_lookup_key("使用原反"))
    except ValueError:
        i_key = 0
    try:
        i_m = hdr.index(_normalize_mm_table_lookup_key("ロール単位の長さ"))
    except ValueError:
        i_m = 1 if len(hdr) > 1 else 0
    for parts in rows[1:]:
        if not parts or all(not str(x).strip() for x in parts):
            continue
        while len(parts) <= max(i_key, i_m):
            parts.append("")
        raw_k = parts[i_key]
        raw_m = parts[i_m]
        key = _normalize_mm_table_lookup_key(raw_k)
        if not key:
            continue
        known_keys.add(key)
        m = parse_float_safe(raw_m, 0.0)
        if m <= 0:
            continue
        if key in out and abs(out[key] - float(m)) > 1e-9:
            logging.warning(
                "使用原反ロール単位長さテーブルで同一キーに矛盾する値があります: %r → %s と %s (%s)",
                key,
                out[key],
                m,
                path_found,
            )
            continue
        out[key] = float(m)
    return out, known_keys, path_found
def _append_used_raw_roll_length_table_row_if_missing(
    used_raw_cell,
    *,
    table_path: str,
    known_keys: set[str],
    appended: set[str],
) -> bool:
    """段階1: 使用原反ロール長テーブルに未登録キーを空欄値で追記する。"""
    nk = _normalize_mm_table_lookup_key(used_raw_cell)
    if not nk:
        return False
    display = str(used_raw_cell or "").strip() or nk
    return _append_code_dispatch_lookup_table_row_if_missing(
        display,
        nk,
        table_path=table_path
        or _resolve_code_lookup_table_path_for_write(
            _roll_unit_by_used_raw_table_search_paths(),
            ROLL_UNIT_BY_USED_RAW_TABLE_DEFAULT_FILENAME,
        ),
        header_line="使用原反,ロール単位の長さ",
        known_keys=known_keys,
        appended=appended,
        log_table_label="使用原反ロール長",
    )
def _lookup_roll_unit_length_m_from_used_raw(used_raw) -> float | None:
    """使用原反の完全一致（正規化後）でロール単位長さ(m)を返す。未登録なら None。"""
    global _ROLL_UNIT_BY_USED_RAW_TABLE_CACHE
    if _ROLL_UNIT_BY_USED_RAW_TABLE_CACHE is None:
        _ROLL_UNIT_BY_USED_RAW_TABLE_CACHE = _load_roll_unit_length_m_by_used_raw_table_optional()
    if not _ROLL_UNIT_BY_USED_RAW_TABLE_CACHE:
        return None
    k = _normalize_mm_table_lookup_key(used_raw)
    if not k:
        return None
    v = _ROLL_UNIT_BY_USED_RAW_TABLE_CACHE.get(k)
    return float(v) if (v is not None and v > 0) else None
def _parse_roll_unit_m_from_used_raw_dimension_only(used_raw) -> float | None:
    """
    使用原反セル文字列からのみロール長(m)を読む（CSV テーブルは使わない）。
    ``_infer_roll_unit_m_from_product_name_dimensions_only`` と同系の NNNxMM / X 後ろパターン。
    解釈できなければ None。
    """
    if used_raw is None or (isinstance(used_raw, float) and pd.isna(used_raw)):
        return None
    s0 = str(used_raw).strip()
    if not s0 or s0.lower() in ("nan", "none"):
        return None
    s = _normalize_product_dim_separators_for_roll_inference(s0)
    dim_pairs = re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s)
    if dim_pairs:
        try:
            b = int(dim_pairs[-1][1])
            return float(b) if b > 0 else None
        except ValueError:
            pass
    matches = re.findall(r"[xX]\s*(\d{2,6})", s)
    if matches:
        try:
            v = int(matches[-1])
            return float(v) if v > 0 else None
        except ValueError:
            pass
    return None
def _load_roll_unit_length_m_table_optional() -> dict[str, float]:
    """
    ロール単位長さテーブル（製品名→ロール単位の長さ(m)）を読み込む。
    テーブルが見つからない場合は空 dict を返し、従来の製品名推定へフォールバックする。
    """
    global _ROLL_UNIT_LENGTH_TABLE_PATH_USED
    path_found = ""
    for p in _roll_unit_length_table_search_paths():
        if os.path.isfile(p):
            path_found = p
            break
    if not path_found:
        return {}
    out: dict[str, float] = {}
    try:
        with open(path_found, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
    except OSError:
        return {}
    if not rows:
        return {}
    hdr = [_normalize_roll_unit_length_table_key(x) for x in rows[0]]
    try:
        i_key = hdr.index(_normalize_roll_unit_length_table_key("製品名"))
    except ValueError:
        i_key = 0
    try:
        i_m = hdr.index(_normalize_roll_unit_length_table_key("ロール単位の長さ"))
    except ValueError:
        i_m = 1 if len(hdr) > 1 else 0
    for parts in rows[1:]:
        if not parts or all(not str(x).strip() for x in parts):
            continue
        while len(parts) <= max(i_key, i_m):
            parts.append("")
        raw_k = parts[i_key]
        raw_m = parts[i_m]
        key = _normalize_roll_unit_length_table_key(raw_k)
        if not key:
            continue
        m = parse_float_safe(raw_m, 0.0)
        if m <= 0:
            continue
        if key in out and abs(out[key] - float(m)) > 1e-9:
            # テーブルが矛盾していても段階1を止めない（推定ロジックへフォールバックできるよう警告に留める）
            logging.warning(
                "ロール単位長さテーブルで同一キーに矛盾する値があります: %r → %s と %s (%s)",
                key,
                out[key],
                m,
                path_found,
            )
            continue
        out[key] = float(m)
    _ROLL_UNIT_LENGTH_TABLE_PATH_USED = path_found
    if out:
        logging.info(
            "ロール単位長さテーブルを読み込みました: %s (%s 件)",
            path_found,
            len(out),
        )
    return out
def _lookup_roll_unit_length_m_from_table(product_name) -> float | None:
    """製品名の完全一致（正規化後）でロール単位長さ(m)を返す。未登録なら None。"""
    global _ROLL_UNIT_LENGTH_TABLE_CACHE
    if _ROLL_UNIT_LENGTH_TABLE_CACHE is None:
        _ROLL_UNIT_LENGTH_TABLE_CACHE = _load_roll_unit_length_m_table_optional()
    if not _ROLL_UNIT_LENGTH_TABLE_CACHE:
        return None
    k = _normalize_roll_unit_length_table_key(product_name)
    if not k:
        return None
    v = _ROLL_UNIT_LENGTH_TABLE_CACHE.get(k)
    return float(v) if (v is not None and v > 0) else None
def _infer_roll_unit_m_from_product_name_dimensions_only(product_name, fallback_unit):
    """
    製品名の寸法だけから 1 ロールあたりの長さ(m)を推定する（製品名 CSV テーブルは使わない）。
    最後の NNNxMM ペアの右側、なければ最後の X 直後の 2〜6 桁。
    """
    if product_name is None or pd.isna(product_name):
        return fallback_unit
    s = _normalize_product_dim_separators_for_roll_inference(str(product_name))
    dim_pairs = re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s)
    if dim_pairs:
        try:
            _a_str, b_str = dim_pairs[-1]
            b = int(b_str)
            if b > 0:
                return b
        except ValueError:
            pass
    matches = re.findall(r"[xX]\s*(\d{2,6})", s)
    if matches:
        try:
            v = int(matches[-1])
            if v > 0:
                return v
        except ValueError:
            pass
    return float(INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M)
def _planning_scalar_text_for_roll_dim(val) -> str:
    """寸法推定用: 欠損・空白・文字列化のみ。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none") else s
def _string_has_roll_dim_mm_pattern(val) -> bool:
    """NNNxMM（2〜6 桁×2〜6 桁）の寸法ペアが含まれるか（ロール長推定の前提）。"""
    s = _planning_scalar_text_for_roll_dim(val)
    if not s:
        return False
    t = _normalize_product_dim_separators_for_roll_inference(s)
    return bool(re.search(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", t))
def infer_roll_unit_m_from_used_raw_then_product_dims(
    product_name, used_raw, fallback_unit
):
    """
    使用原反テーブル（``使用原反,ロール単位の長さ.txt``）を優先し、
    未登録なら製品名の寸法、さらに製品名に寸法が無い場合は使用原反文字列の寸法から推定する。

    **(製品)ロール単位長さ**列（段階1の ``_stage1_roll_length_for_planning_row``）の算出には使わない。
    段階2 ``build_task_queue_from_planning_df`` の ``unit_m`` 解決不能時の補完に使用する。
    """
    v = _lookup_roll_unit_length_m_from_used_raw(used_raw)
    if v is not None and v > 0:
        return float(v)
    pn_txt = _planning_scalar_text_for_roll_dim(product_name)
    ur_txt = _planning_scalar_text_for_roll_dim(used_raw)
    if not _string_has_roll_dim_mm_pattern(pn_txt) and _string_has_roll_dim_mm_pattern(
        ur_txt
    ):
        dim_source = ur_txt
    elif not pn_txt and ur_txt:
        dim_source = ur_txt
    else:
        dim_source = pn_txt if pn_txt else ur_txt
    return _infer_roll_unit_m_from_product_name_dimensions_only(
        dim_source if dim_source else None, fallback_unit
    )
def infer_unit_m_from_product_name(product_name, fallback_unit):
    """
    製品名文字列から 1 ロールあたりの長さ(m)を推定する。

    まず「製品名,ロール単位の長さ.txt」の完全一致があれば採用し、
    無ければ寸法（``_infer_roll_unit_m_from_product_name_dimensions_only`` と同じ）へフォールバックする。

    **配台計画_タスク入力**の列 **「(製品)ロール単位長さ」**（段階1の ``_stage1_roll_length_for_planning_row``）、
    ``_roll_unit_m_estimate_from_plan_row`` のフォールバックで使用する。
    段階2の ``unit_m`` には使わない（``_dispatch_simulator_unit_m_from_plan_row`` が原反ロール長を用いる）。
    """
    if product_name is None or pd.isna(product_name):
        return fallback_unit
    from_table = _lookup_roll_unit_length_m_from_table(product_name)
    if from_table is not None and from_table > 0:
        return from_table
    return _infer_roll_unit_m_from_product_name_dimensions_only(
        product_name, fallback_unit
    )
def _ceil_roll_unit_length_m_to_next_step(roll_m: float, step_m: float = None) -> float:
    """
    正の長さ(m)を step の倍数に切り上げ（下二桁繰り上げ: step=100 のとき 40→100, 125→200）。
    **換算数量（配台用内部）**・未加工≤0 時の残量算定・矯正ロジックの比較などで使用する。
    **ロール単位長さ**列には適用しない（シート・推定・補正の値をそのまま用いる）。
    """
    v = parse_float_safe(roll_m, 0.0)
    if v <= 0:
        return v
    step = parse_float_safe(
        step_m if step_m is not None else ROLL_UNIT_LENGTH_CEIL_STEP_M, 0.0
    )
    if step <= 0:
        return v
    return float(math.ceil(v / step) * step)
def _heal_stage1_roll_unit_if_width_ceiling_merge_spurious(out_df: "pd.DataFrame") -> None:
    """
    段階1: 既存シートのマージで「寸法ペア左側（例: 870）を 100m 切上した値」がロール単位長さに
    残った場合、製品名からの再計算で矯正する（誤マージ・誤フォールバックの典型: 900 を期待 200 の行）。
    手入力で意図的に左側切上と同じ値にした行は稀なため、一致時のみ上書きする。
    """
    if out_df is None or getattr(out_df, "empty", True):
        return
    if (
        PLAN_COL_ROLL_UNIT_LENGTH not in out_df.columns
        or TASK_COL_PRODUCT not in out_df.columns
    ):
        return
    healed = 0
    for i in out_df.index:
        row = out_df.loc[i]
        pn = row.get(TASK_COL_PRODUCT, None)
        s = _normalize_product_dim_separators_for_roll_inference(str(pn or ""))
        dim_pairs = re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s)
        if not dim_pairs:
            continue
        try:
            left_w = int(dim_pairs[-1][0])
        except ValueError:
            continue
        if left_w <= 0:
            continue
        width_ceiled = float(_ceil_roll_unit_length_m_to_next_step(float(left_w)))
        cur = parse_float_safe(row.get(PLAN_COL_ROLL_UNIT_LENGTH), 0.0)
        if cur <= 0:
            continue
        if abs(cur - width_ceiled) > 1e-6:
            continue
        try:
            want = _stage1_roll_length_for_planning_row(row)
        except Exception:
            continue
        if abs(cur - want) <= 1e-6:
            continue
        out_df.at[i, PLAN_COL_ROLL_UNIT_LENGTH] = want
        healed += 1
    if healed:
        logging.info(
            "段階1: ロール単位長さが寸法左側の100m切上と誤一致していた行を %s 件、使用原反テーブル／製品名寸法で矯正しました。",
            healed,
        )
def _heal_stage1_roll_unit_no_dim_when_roll_matches_qty_mistake(
    out_df: "pd.DataFrame",
) -> None:
    """
    寸法パターンが無い品番で、ロール単位長さが換算数量（シート値）またはその 100m 切上と
    同じになっている行を矯正する（旧シートマージで FEL 等に換算数量が載った誤り向け）。
    小さい値（<500）は「意図的に換算数量と同じロール長」とみなし触れない。
    """
    if out_df is None or getattr(out_df, "empty", True):
        return
    if (
        PLAN_COL_ROLL_UNIT_LENGTH not in out_df.columns
        or TASK_COL_PRODUCT not in out_df.columns
        or TASK_COL_QTY not in out_df.columns
    ):
        return
    healed = 0
    min_heal_cur = 500.0
    want = float(
        _ceil_roll_unit_length_m_to_next_step(
            float(INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M)
        )
    )
    for i in out_df.index:
        row = out_df.loc[i]
        pn = row.get(TASK_COL_PRODUCT, None)
        s = _normalize_product_dim_separators_for_roll_inference(str(pn or ""))
        if re.findall(r"(\d{2,6})\s*[xX]\s*(\d{2,6})", s):
            continue
        qty_raw = max(0.0, parse_float_safe(row.get(TASK_COL_QTY), 0.0))
        if qty_raw <= 0:
            continue
        qty_ceiled = float(_ceil_roll_unit_length_m_to_next_step(float(qty_raw)))
        cur = parse_float_safe(row.get(PLAN_COL_ROLL_UNIT_LENGTH), 0.0)
        if cur + 1e-9 < min_heal_cur:
            continue
        if abs(cur - qty_raw) > 1e-4 and abs(cur - qty_ceiled) > 1e-4:
            continue
        if abs(cur - want) < 1e-6:
            continue
        out_df.at[i, PLAN_COL_ROLL_UNIT_LENGTH] = want
        healed += 1
    if healed:
        logging.info(
            "段階1: 寸法なしでロール単位長さが換算数量と誤一致していた行を %s 件、既定 %sm へ矯正しました。",
            healed,
            int(want) if abs(want - int(want)) < 1e-9 else want,
        )
def _excel_sheet_arg_from_env(env_key: str) -> str | int:
    _raw = (os.environ.get(env_key) or "").strip()
    if not _raw:
        return 0
    if _raw.isdigit():
        return int(_raw)
    return _raw
def _excel_sheet_label_for_log(sheet_arg: str | int, legacy_sheet_title: str) -> str:
    if isinstance(sheet_arg, int):
        return "\u5148\u982d\u30b7\u30fc\u30c8" if sheet_arg == 0 else f"index {sheet_arg}"
    return str(sheet_arg) if sheet_arg else legacy_sheet_title
def _processing_plan_sheet_label_for_context(sheet_arg: str | int) -> str:
    return _excel_sheet_label_for_log(sheet_arg, TASKS_SHEET_NAME)
def _actual_detail_sheet_log_label() -> str:
    return _excel_sheet_label_for_log(
        _excel_sheet_arg_from_env(ENV_PM_AI_ACTUAL_DETAIL_SHEET),
        ACTUAL_DETAIL_SHEET_NAME,
    )
def load_tasks_df():
    """
    タスク入力を取得れる（tasks.xlsx は使用しない）。

    PM_AI_PROCESSING_PLAN_PATH に CSV / Parquet / xlsx の実在パス（未指定・無効時は
    PM_AI_TASK_INPUT_SOURCE_DIR 内の最新表ファイルへ resolve_processing_plan_path_from_env）。
    TASK_INPUT_WORKBOOK（マクロブック）は読み込みに使わない。
    xlsx は PM_AI_PROCESSING_PLAN_SHEET でシート指定（省略時は先頭シート index 0。
    単一シートのブックでは名前不要。複数シートで名前指定する場合は文字列、数値のみなら 0 始まり索引）。
    """
    resolve_processing_plan_path_from_env()
    _alt = (os.environ.get("PM_AI_PROCESSING_PLAN_PATH") or "").strip()
    _sheet_label_for_context = TASKS_SHEET_NAME
    if _alt and os.path.isfile(_alt):
        _low = _alt.lower()

        def _load_once():
            if _low.endswith((".csv", ".parquet", ".pq")):
                out = read_tabular_dataframe(_alt)
            else:
                _sn = _excel_sheet_arg_from_env(ENV_PM_AI_PROCESSING_PLAN_SHEET)
                nonlocal _sheet_label_for_context
                _sheet_label_for_context = _processing_plan_sheet_label_for_context(_sn)
                out = read_tabular_dataframe(_alt, sheet_name=_sn)
            out.columns = out.columns.str.strip()
            return out

        df = _cached_tabular_dataframe("processing_plan", _alt, _load_once)
    else:
        raise FileNotFoundError(
            "タスク入力が必要です。PM_AI_PROCESSING_PLAN_PATH に表形式ファイル（CSV/Parquet/xlsx）の"
            "実在パスを設定するか、PM_AI_TASK_INPUT_SOURCE_DIR にフォルダを指定して"
            "その中の最新ファイルを使ってください（TASK_INPUT_WORKBOOK は使用しません）。"
        )
    df = _align_dataframe_headers_to_canonical(df, list(SOURCE_BASE_COLUMNS))
    # 換算数量を先に確定（未加工の式補完に必要）。無いブックは NFKC で別名列から補完。
    if TASK_COL_QTY not in df.columns:
        _qty_src = _first_dataframe_column_matching_nfkc_labels(
            df,
            (
                "未加工",
                "残作数値",
                "加工予定数",
                "計画数量",
                "換算m数",
                "換算ｍ数",
            ),
        )
        if _qty_src is not None:
            df[TASK_COL_QTY] = df[_qty_src]
            logging.info(
                "タスク入力: 列「%s」が無いため「%s」をコピーして補完しました。",
                TASK_COL_QTY,
                _qty_src,
            )
    _supplement_task_input_unprocessed_column_if_missing(df)
    _ensure_dataframe_has_unprocessed_column(
        df, context_label=f"シート「{_sheet_label_for_context}」"
    )
    # 「受注数」列名の表記ゆれを「受注数」（TASK_COL_ORDER_QTY）へ寄せる補完
    if TASK_COL_ORDER_QTY not in df.columns and "受注数" in df.columns:
        df[TASK_COL_ORDER_QTY] = df["受注数"]
        logging.info(
            "タスク入力: 列「%s」が無いため「受注数」をコピーして補完しました。",
            TASK_COL_ORDER_QTY,
        )
    _src_pp = (os.environ.get("PM_AI_PROCESSING_PLAN_PATH") or "").strip()
    if _src_pp and os.path.isfile(_src_pp):
        logging.info("タスク入力: PM_AI_PROCESSING_PLAN_PATH='%s' を読み込みました。", _src_pp)
    try:
        _sheet_arg = (os.environ.get(ENV_PM_AI_PROCESSING_PLAN_SHEET) or "").strip()
        _hdr = (os.environ.get("PM_AI_PROCESSING_PLAN_HEADER_ROW") or "").strip()
        _tid = (os.environ.get("PM_AI_TASK_INPUT_SOURCE_DIR") or "").strip()
        print(
            "[stage1-input] 加工計画DATA 読込完了: ファイル=%r シート（文脈ラベル）=%r"
            % (_src_pp, _sheet_label_for_context),
            file=sys.stderr,
            flush=True,
        )
        print(
            "[stage1-input] 参照環境変数: %s, %s, PM_AI_PROCESSING_PLAN_HEADER_ROW, PM_AI_TASK_INPUT_SOURCE_DIR"
            % (ENV_PROCESSING_PLAN_PATH, ENV_PM_AI_PROCESSING_PLAN_SHEET),
            file=sys.stderr,
            flush=True,
        )
        if _sheet_arg:
            print(
                "[stage1-input] %s=%r" % (ENV_PM_AI_PROCESSING_PLAN_SHEET, _sheet_arg),
                file=sys.stderr,
                flush=True,
            )
        if _hdr:
            print(
                "[stage1-input] PM_AI_PROCESSING_PLAN_HEADER_ROW=%r" % (_hdr,),
                file=sys.stderr,
                flush=True,
            )
        if _tid:
            print(
                "[stage1-input] PM_AI_TASK_INPUT_SOURCE_DIR=%r" % (_tid,),
                file=sys.stderr,
                flush=True,
            )
    except Exception:
        pass
    return df
def _nfkc_column_aliases(canonical_name):
    """見出しの表記ゆれ（全角記坷・互換文字）を坸坎れるための比較キー。"""
    return unicodedata.normalize("NFKC", str(canonical_name).strip())
def _first_dataframe_column_matching_nfkc_labels(df: pd.DataFrame, labels: tuple[str, ...]):
    """
    labels を先頭優先で試し、df の列名と NFKC 一致する最初の実列名を返す。
    換算数量の別名（加工予定数・換算m数 等）を残作数値より後ろで試すためのヘルパ。
    """
    if df is None or getattr(df, "empty", True):
        return None
    keys_present = {}
    for c in df.columns:
        k = _nfkc_column_aliases(str(c))
        keys_present.setdefault(k, str(c))
    for lab in labels:
        k = _nfkc_column_aliases(lab)
        if k in keys_present:
            return keys_present[k]
    return None
def _align_dataframe_headers_to_canonical(df, canonical_names):
    """列名を NFKC 一致で canonical に寄せる（Excel 坴は全角 '_' 等でも読ゝるよごに）。"""
    key_to_canonical = {_nfkc_column_aliases(c): c for c in canonical_names}
    # 旧見出し「残作数値」→ 現行「換算数量」（TASK_COL_QTY）
    if TASK_COL_QTY in canonical_names:
        key_to_canonical[_nfkc_column_aliases("残作数値")] = TASK_COL_QTY
    # 旧見出し「原板…」を「原反…」へ寄せる（互換。canonical は TASK_COL / PLAN_COL の表記）
    if TASK_COL_RAW_INPUT_DATE in canonical_names:
        key_to_canonical[_nfkc_column_aliases("原板投入日")] = TASK_COL_RAW_INPUT_DATE
    if PLAN_COL_RAW_INPUT_DATE_OVERRIDE in canonical_names:
        key_to_canonical[_nfkc_column_aliases("原板投入日_上書き")] = (
            PLAN_COL_RAW_INPUT_DATE_OVERRIDE
        )
        _ref_canon = plan_reference_column_name(PLAN_COL_RAW_INPUT_DATE_OVERRIDE)
        if _ref_canon in canonical_names:
            key_to_canonical[_nfkc_column_aliases("（元）原板投入日_上書き")] = _ref_canon
    if PLAN_COL_ROLL_UNIT_LENGTH in canonical_names:
        key_to_canonical[_nfkc_column_aliases(PLAN_COL_ROLL_UNIT_LENGTH_LEGACY)] = (
            PLAN_COL_ROLL_UNIT_LENGTH
        )
    rename_map = {}
    for col in df.columns:
        k = _nfkc_column_aliases(col)
        if k in key_to_canonical:
            target = key_to_canonical[k]
            if col != target:
                rename_map[col] = target
    if rename_map:
        df = df.rename(columns=rename_map)
    return df
def _normalize_equipment_match_key(val):
    """
    工程名（設備坝）の照合用キー。
    NFKC・剝後空白・連続空白・NBSP/全角スペース・ゼロ幅文字を正規化する。
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    t = unicodedata.normalize("NFKC", str(val))
    t = t.replace("\u00a0", " ").replace("\u3000", " ")
    t = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t
def _equipment_line_key_to_physical_occupancy_key(eq_line: str) -> str:
    """設備列キー（工程+機械 等）から」実機械の占有に用いるキー（機械名ベース・正規化）を得る。"""
    s = str(eq_line or "").strip()
    if not s:
        return ""
    nk = _normalize_equipment_match_key(s)
    if "+" in nk:
        return _normalize_equipment_match_key(nk.split("+", 1)[1])
    return nk
def _physical_machine_occupancy_key_for_task(task: dict) -> str:
    """
    設備のタイムライン占有（machine_avail_dt・間隔ミラー）に用いるキー。
    機械カレンダー列は equipment_line_key の「工程+機械」と一致するため、
    正規化後に「+」を含むとしは **machine_name より先に** しこから実機械名を採用する。
    （machine_name に工程名のみなどは入り」床キー「熱融着機 湖南」とうれで候補外し漝れれるのを防し）
    坘一坝のときは従来どおり machine_name を優先し、無ければ equipment_line_key / machine から推定する。
    machine_name に「工程+機械」と入っている場合でも」占有は実機械名（+ の坳坴）に寄せる。
    全角「＋」のみの列は NFKC 後に半角「+」になるため、分割判定は正規化後に行う。
    """
    ek = str(task.get("equipment_line_key") or "").strip()
    nek = _normalize_equipment_match_key(ek)
    if nek and "+" in nek:
        pk = _equipment_line_key_to_physical_occupancy_key(ek)
        if pk:
            return pk
    mn = str(task.get("machine_name") or "").strip()
    if mn:
        nk = _normalize_equipment_match_key(mn)
        if "+" in nk:
            return _normalize_equipment_match_key(nk.split("+", 1)[1])
        return nk
    return _equipment_line_key_to_physical_occupancy_key(
        str(task.get("equipment_line_key") or task.get("machine") or "")
    )
def _machine_occupancy_key_resolve(task: dict, eq_line: str) -> str:
    """
    machine_avail_dt・機械カレンダー床と整合する占有キー（原則: 実機械名）。
    task から取れないとしは eq_line（工程+機械）から機械名ベースを推定し、最後の手段で eq_line。
    「… or eq_line」による工程+機械フォールバックは機械カレンダー実キーと厳密一致になり得るため廃止。
    """
    occ = (_physical_machine_occupancy_key_for_task(task) or "").strip()
    if occ:
        return occ
    ek = str(eq_line or "").strip()
    if not ek:
        return ""
    pk = (_equipment_line_key_to_physical_occupancy_key(ek) or "").strip()
    return pk or ek
def _equipment_lookup_normalized_to_canonical(equipment_list):
    """正規化キー → master スキルシート上の列名（canonical 表記）。"""
    lookup = {}
    for eq in equipment_list:
        k = _normalize_equipment_match_key(eq)
        if k and k not in lookup:
            lookup[k] = eq
    # 工程名のみの照合（加工実績DATA等）: 同一工程の先頭列（工程+機械）へ寄せる
    for eq in equipment_list:
        s = str(eq).strip()
        if "+" not in s:
            continue
        p, _rest = s.split("+", 1)
        pk = _normalize_equipment_match_key(p)
        if pk and pk not in lookup:
            lookup[pk] = eq
    return lookup
def _equipment_schedule_header_labels(equipment_list: list) -> list:
    """
    結果_設備毎の時間割・結果_設備ガントの行＝列見出し用。
    内部キーは「工程+機械」のときは機械名を表示し、機械名の重複時のみ工程を括弧で補ご。
    """
    raw = []
    for eq in equipment_list:
        s = str(eq).strip()
        if "+" in s:
            mpart = s.split("+", 1)[1].strip()
            raw.append(mpart if mpart else s)
        else:
            raw.append(s)
    counts = {}
    for r in raw:
        counts[r] = counts.get(r, 0) + 1
    out = []
    for eq, r in zip(equipment_list, raw):
        if counts.get(r, 0) > 1:
            s = str(eq).strip()
            if "+" in s:
                p = s.split("+", 1)[0].strip()
                out.append(f"{r}（{p}）" if p else r)
            else:
                out.append(r)
        else:
            out.append(r)
    return out
def _split_equipment_line_process_machine(eq_line: str) -> tuple[str, str]:
    """
    設備マスタの列キー「工程+機械」を (工程名, 機械名) に分割れる。
    '+' は無いとしは機械名のみとみなし、工程名は空文字。
    """
    s = str(eq_line).strip()
    if not s:
        return ("", "")
    if "+" in s:
        p, m = s.split("+", 1)
        return (p.strip(), m.strip())
    return ("", s)
def _gantt_member_label_surname_only(raw: str) -> str:
    """
    設備ガントのタイムライン上の担当者姓表示用。半角＝全角空白はあれみ手剝を姓とみなし、無いとしは全体を表示
    （並びは1トークンのみのときは姓の切り出し試行のまま）。NFKC・富田/冨田寄せは姓用とともに。
    """
    sei, mei = _split_person_sei_mei(raw)
    if not sei:
        return ""
    n = _normalize_sei_for_match(sei)
    return n if n else sei
def _gantt_member_labels_for_startup_in_range(
    evlist, range_start: datetime, range_end: datetime
) -> list[str]:
    """
    半開区間 [range_start, range_end) に重なる日次始業イベントから担当者姓を得る。
    （`_eq_grid_best_overlapping_event_for_cell` は加工を優先するため使わない）
    """
    best_ev: dict | None = None
    best_st: datetime | None = None
    for ev in evlist or []:
        if _timeline_event_kind(ev) != TIMELINE_EVENT_MACHINE_DAILY_STARTUP:
            continue
        st = ev.get("start_dt")
        ed = ev.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime) or ed <= st:
            continue
        if st < range_end and ed > range_start:
            if best_ev is None or best_st is None or st < best_st:
                best_ev = ev
                best_st = st
    if best_ev is None:
        return []
    raw_names: list[str] = []
    seen_raw: set[str] = set()
    op = " ".join(str(best_ev.get("op") or "").split())
    if op and op not in seen_raw:
        seen_raw.add(op)
        raw_names.append(op)
    sub_raw = " ".join(str(best_ev.get("sub") or "").split())
    if sub_raw:
        for seg in re.split(r"[,、]", sub_raw):
            t = " ".join(str(seg or "").split())
            if t and t not in seen_raw:
                seen_raw.add(t)
                raw_names.append(t)
    labels: list[str] = []
    seen_label: set[str] = set()
    for raw in raw_names:
        lab = _gantt_member_label_surname_only(raw)
        if lab and lab not in seen_label:
            seen_label.add(lab)
            labels.append(lab)
    return labels
def _gantt_member_labels_for_task(evlist, task_id: str) -> list[str]:
    """
    設備ガントのタイムライン1セグメント用: 指定 task_id のイベントから担当者姓を出現順で重複除去。
    （シェイプの上下チップ用）
    """
    tid = str(task_id or "").strip()
    if not tid:
        return []
    raw_names: list[str] = []
    seen_raw: set[str] = set()
    for e in evlist or []:
        if str(e.get("task_id") or "").strip() != tid:
            continue
        op = str(e.get("op") or "").strip()
        if op and op not in seen_raw:
            seen_raw.add(op)
            raw_names.append(op)
        sub_raw = str(e.get("sub") or "").strip()
        if not sub_raw:
            continue
        for seg in re.split(r"[,」]", sub_raw):
            t = seg.strip()
            if t and t not in seen_raw:
                seen_raw.add(t)
                raw_names.append(t)
    labels: list[str] = []
    seen_label: set[str] = set()
    for raw in raw_names:
        lab = _gantt_member_label_surname_only(raw)
        if lab and lab not in seen_label:
            seen_label.add(lab)
            labels.append(lab)
    return labels
def _resolve_equipment_line_key_for_task(task: dict, equipment_list: list | None) -> str:
    """
    設備時間割・設備専有空しの列キー（skills / need とともに「工程+機械」を基本とれる）。
    機械名は空でマスタに当該工程の列は1つの値ならしの複坈キーへ寄せる。
    """
    p = str(task.get("machine") or "").strip()
    mn = str(task.get("machine_name") or "").strip()
    cand = f"{p}+{mn}" if (p and mn) else (p or mn)
    elist = [str(x).strip() for x in (equipment_list or []) if str(x).strip()]
    if cand in elist:
        return cand
    if mn:
        return cand
    if not p:
        return cand
    exact_p = [x for x in elist if x == p]
    if len(exact_p) == 1:
        return exact_p[0]
    prefixed = [x for x in elist if x.startswith(p + "+")]
    if len(prefixed) == 1:
        return prefixed[0]
    return p
def _apply_planning_sheet_post_load_mutations(
    df: "pd.DataFrame",
    wb_path: str,
    log_prefix: str,
    *,
    apply_exclude_rules_from_config: bool = True,
    compile_exclude_rules_d_to_e_with_ai: bool = True,
) -> None:
    """
    配台計画_タスク入力を DataFrame 化した直後の共通処理（配台不要ルールの行同期・分割行の自動配台不要）。

    行同期: ``PM_AI_EXCLUDE_RULES_JSON`` が有効なら JSON へ未登録の (工程名, 機械名) を追記。
    無効なら計画ブックの「設定_配台不要工程」を ``run_exclude_rules_sheet_maintenance`` で更新する。

    「設定_配台不要工程」の C/E による計画 DataFrame への「配台不要」上書きは **段階1のみ**
    （``run_stage1_extract`` 内の ``apply_exclude_rules_config_to_plan_df``）。段階2の
    ``load_planning_tasks_df`` では常に ``apply_exclude_rules_from_config=False`` を渡し、
    シート上の「配台不要」列をそのまま解釈する。

    段階2および試行順のみの限定更新では ``compile_exclude_rules_d_to_e_with_ai=False`` とし、
    設定シートの D→E（ロジック式）の **Gemini 補完は行わない**（行同期・保存のみ）。

    ``apply_exclude_rules_from_config=False`` は本関数呼び出し側で明示する（上記のほか、
    試行順のみ再計算する当該経路でも同様）。
    """
    try:
        _pairs_lr = _collect_plan_input_process_machine_pairs_for_exclude_rules_sync(df)
        if _exclude_rules_json_env_supersedes_excel_sheet():
            json_env = (os.environ.get(ENV_EXCLUDE_RULES_JSON) or "").strip()
            if json_env and _pairs_lr:
                _merge_exclude_rules_json_with_plan_pairs(json_env, _pairs_lr, log_prefix)
        else:
            run_exclude_rules_sheet_maintenance(
                wb_path,
                _pairs_lr,
                log_prefix,
                compile_exclude_rules_d_to_e_with_ai=compile_exclude_rules_d_to_e_with_ai,
            )
    except Exception:
        logging.exception("%s: 設定_配台不要工程の保守で例外（続行）", log_prefix)
    try:
        _apply_auto_exclude_bunkatsu_duplicate_machine(df, log_prefix=log_prefix)
    except Exception as ex:
        logging.warning(
            "%s: 分割行の配台試行自動設定で例外（続行）: %s",
            log_prefix,
            ex,
        )
    try:
        _apply_auto_exclude_in_house_self_processing(df, log_prefix=log_prefix)
    except Exception as ex:
        logging.warning(
            "%s: 自社加工品の配台不要自動設定で例外（続行）: %s",
            log_prefix,
            ex,
        )
    if apply_exclude_rules_from_config:
        try:
            apply_exclude_rules_config_to_plan_df(df, wb_path, log_prefix)
        except Exception as ex:
            logging.warning(
                "%s: 設定シートによる配台不要適用で例外（続行）: %s",
                log_prefix,
                ex,
            )
def _migrate_deprecated_plan_override_columns(df: "pd.DataFrame") -> "pd.DataFrame":
    """廃止列を削除し、*_上書き列だけは値を基底列へ移す。"""
    if df is None:
        return df
    for oc, base in PLAN_OVERRIDE_TO_BASE_COLUMN.items():
        if oc not in df.columns:
            continue
        if base not in df.columns:
            df[base] = ""
        for i in df.index:
            ov = df.at[i, oc]
            if ov is None or (isinstance(ov, float) and pd.isna(ov)):
                continue
            if isinstance(ov, str) and not str(ov).strip():
                continue
            cur = df.at[i, base]
            if cur is None or (isinstance(cur, float) and pd.isna(cur)):
                df.at[i, base] = ov
                continue
            if isinstance(cur, str) and not str(cur).strip():
                df.at[i, base] = ov
    drop = list(PLAN_DEPRECATED_OVERRIDE_COLUMNS) + [
        plan_reference_column_name(c) for c in PLAN_DEPRECATED_OVERRIDE_COLUMNS
    ]
    drop.extend(("担当OP_指定", "担当OP指定"))
    present = [c for c in drop if c in df.columns]
    if present:
        df = df.drop(columns=present)
    ref_drop = [c for c in df.columns if _plan_column_is_original_reference(str(c).strip())]
    if ref_drop:
        df = df.drop(columns=ref_drop)
    return df
def _gantt_label_luminance_01(r: int, g: int, b: int) -> float:
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
def _gantt_openpyxl_font_color_for_fill_hex(fill_hex: str) -> str:
    """openpyxl Font.color 用 6 桁（RGB 文字列）。"""
    r, g, b = _hex_rrggbb_to_rgb_triple(fill_hex)
    lum = _gantt_label_luminance_01(r, g, b)
    if lum > 0.74:
        return "1A1A1A"
    return "FFFFFF"
def _gantt_fallback_timeline_labels_openpyxl(
    result_path: str, specs: list, sheet_name: str | None = None
) -> None:
    """Excel シェイプ描画に失敗した場合: タイムライン先頭列にセル文字でラベルを書き戻す。"""
    from openpyxl import load_workbook

    if _workbook_should_skip_openpyxl_io(result_path):
        return
    shn = sheet_name or RESULT_SHEET_GANTT_NAME
    wb = load_workbook(result_path)
    try:
        ws = wb[shn]
    except KeyError:
        wb.close()
        return
    try:
        for sp in specs:
            row = int(sp["row"])
            col_s = int(sp["col_s"])
            col_e = int(sp["col_e"])
            text = str(sp.get("text") or "").strip()
            if not text:
                continue
            c = ws.cell(row=row, column=col_s)
            mems = [
                str(x).strip()
                for x in (sp.get("member_labels") or [])
                if str(x).strip()
            ]
            if mems:
                head = "・".join(mems[:5])
                rest = len(mems) - 5
                line2 = head + (f" ほか{rest}名" if rest > 0 else "")
                c.value = text + "\n" + line2
            else:
                c.value = text
            _fh = str(sp.get("fill_hex") or "E8E8E8")
            c.font = _result_font(
                size=10,
                bold=True,
                color=_gantt_openpyxl_font_color_for_fill_hex(_fh),
                italic=bool(sp.get("italic")),
            )
            c.alignment = _gantt_timeline_label_alignment(single_slot=(col_s == col_e))
            if mems or ("\n" in str(c.value or "")):
                try:
                    c.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=1 if mems else 0,
                    )
                except Exception:
                    pass
        wb.save(result_path)
    finally:
        wb.close()
def _gantt_day_image_chroma_rgb() -> tuple[int, int, int]:
    """日別画像の敷き色／透明色に使う RGB（GANTT_DAY_IMAGE_CHROMA_HEX、既定 マゼンタ）。"""
    hx = (os.environ.get("GANTT_DAY_IMAGE_CHROMA_HEX", "FF00FF") or "FF00FF").strip()
    return _hex_rrggbb_to_rgb_triple(hx)
def _gantt_union_bbox_names_xlw(
    api_ws, names: list[str]
) -> tuple[float, float, float, float] | None:
    """シェイプ名の列の外接矩形 (Left, Top, Width, Height)。取得できなければ None。"""
    min_l = min_t = None
    max_r = max_b = None
    for nm in names:
        try:
            sh = api_ws.Shapes(nm)
            l = float(sh.Left)
            t = float(sh.Top)
            r = l + float(sh.Width)
            b = t + float(sh.Height)
        except Exception:
            continue
        if min_l is None:
            min_l, min_t, max_r, max_b = l, t, r, b
        else:
            min_l = min(min_l, l)
            min_t = min(min_t, t)
            max_r = max(max_r, r)
            max_b = max(max_b, b)
    if min_l is None:
        return None
    pad = 1.0
    w = max(max_r - min_l + 2.0 * pad, 2.0)
    h = max(max_b - min_t + 2.0 * pad, 2.0)
    return (min_l - pad, min_t - pad, w, h)
def _gantt_flatten_apply_picture_chroma_transparency_xlw(pic, fill_bgr: int) -> None:
    """Picture の単色透明化。色→有効の順で一部の Excel で安定する。"""
    try:
        pf = pic.PictureFormat
        pf.TransparencyColor = int(fill_bgr)
        pf.TransparentBackground = -1  # msoTrue
    except Exception:
        pass
def _gantt_flatten_copy_picture_format_xlw() -> int:
    """
    CopyPicture の Format（XlCopyPictureFormat）。

    Excel でグループ化→クリップボード経由で画像貼り付けしたとき、**外接矩形のうち
    シェイプの実画素以外が透明**になりやすいのは **xlPicture（-4147, EMF 系）** 側。
    **xlBitmap（2）** は余白が **不透明（白など）** になりやすく、手動の見え方とずれる。

    ``GANTT_DAY_IMAGE_COPY_PICTURE_FORMAT`` で上書き可能（既定は xlPicture）。
    """
    v = (os.environ.get("GANTT_DAY_IMAGE_COPY_PICTURE_FORMAT", "") or "").strip().lower()
    if v in ("bitmap", "xlbitmap", "2", "bmp"):
        return 2
    if v in ("picture", "xlpicture", "emf", "wmf", "meta", "-4147"):
        return -4147
    return -4147
def _gantt_clipboard_picture_from_shape_names_xlw(
    api_ws,
    group_names: tuple[str, ...],
    *,
    copy_format: int,
    xl_screen: int = 1,
) -> tuple[object, float, float, float, float]:
    """
    ラベル等の図形を「1 枚の画像」に置き換える Excel 標準フロー（クリップボード経由）。

    1. 名前が複数なら ``Shapes.Range(...).Group()`` でグループ化（単一ならそのまま）
    2. ``CopyPicture`` … クリップボードに画像として載せる（Format は ``_gantt_flatten_copy_picture_format_xlw``。既定 EMF 系で手動に近い「外側透明」）
    3. ``Worksheet.Paste`` … シート上に画像シェイプとして貼り付け
    4. 元グループ／元シェイプを削除（貼り付け後の画像のみ残す）

    戻り値: (貼り付けた Shape COM オブジェクト, Left, Top, Width, Height)
    """
    if not group_names:
        raise ValueError("group_names が空です")
    if len(group_names) == 1:
        shp0 = api_ws.Shapes(group_names[0])
        left0 = float(shp0.Left)
        top0 = float(shp0.Top)
        w0 = float(shp0.Width)
        h0 = float(shp0.Height)
        shp0.CopyPicture(Appearance=xl_screen, Format=int(copy_format))
        api_ws.Paste()
        pic = api_ws.Shapes(int(api_ws.Shapes.Count))
        try:
            shp0.Delete()
        except Exception:
            pass
        return pic, left0, top0, w0, h0
    sr = api_ws.Shapes.Range(group_names)
    grp = sr.Group()
    left0 = float(grp.Left)
    top0 = float(grp.Top)
    w0 = float(grp.Width)
    h0 = float(grp.Height)
    grp.CopyPicture(Appearance=xl_screen, Format=int(copy_format))
    api_ws.Paste()
    pic = api_ws.Shapes(int(api_ws.Shapes.Count))
    try:
        grp.Delete()
    except Exception:
        pass
    return pic, left0, top0, w0, h0
def _gantt_flatten_day_label_shapes_to_pictures_xlw(
    api_ws, day_blocks: list, names_by_day: dict
) -> int:
    """
    各日キーに属する角丸ラベルシェイプを、上記
    ``_gantt_clipboard_picture_from_shape_names_xlw``（グループ化→CopyPicture→Paste）
    で 1 枚の Picture に置換する。
    GANTT_DAY_IMAGE_CHROMA_TRANSPARENT が有効なときのみ、敷き矩形を同グループに含め、
    貼り付け後に ``PictureFormat`` で敷き色を透明化する（オプション。核フローはクリップボード画像化）。
    names_by_day[day_key] に蓄積された Name を消費する（成功時は空リストに戻す）。
    """
    if not day_blocks:
        return 0
    _xl_screen = 1  # xlScreen
    _xl_move_and_size = 1
    _mso_rectangle = 1
    n_out = 0
    for blk in day_blocks:
        dk = str(blk.get("day_key") or "").strip()
        raw_names = list(names_by_day.get(dk, []))
        if not raw_names:
            continue
        seen: set[str] = set()
        names: list[str] = []
        for nm in raw_names:
            if nm and nm not in seen:
                seen.add(nm)
                names.append(nm)
        if not names:
            continue
        backdrop_nm: str | None = None
        try:
            r_ch, g_ch, b_ch = _gantt_day_image_chroma_rgb()
            fill_bgr = _com_excel_bgr_rgb(r_ch, g_ch, b_ch)
            group_names: tuple[str, ...] = tuple(names)

            if GANTT_DAY_IMAGE_CHROMA_TRANSPARENT:
                ubox = _gantt_union_bbox_names_xlw(api_ws, names)
                if ubox:
                    L, T, Wb, Hb = ubox
                    bd_nm_try = f"GanttChromaBg_{random.randint(100000, 999999)}"
                    bd = api_ws.Shapes.AddShape(_mso_rectangle, L, T, Wb, Hb)
                    try:
                        bd.Name = bd_nm_try
                    except Exception:
                        bd_nm_try = str(bd.Name)
                    bd.Fill.Visible = True
                    bd.Fill.Solid()
                    bd.Fill.ForeColor.RGB = fill_bgr
                    try:
                        bd.Line.Visible = False
                    except Exception:
                        pass
                    try:
                        bd.Placement = _xl_move_and_size
                    except Exception:
                        pass
                    backdrop_nm = bd_nm_try
                    group_names = (bd_nm_try,) + tuple(names)

            chroma_backdrop = backdrop_nm is not None
            _cpy_fmt = _gantt_flatten_copy_picture_format_xlw()

            pic, left0, top0, w0, h0 = _gantt_clipboard_picture_from_shape_names_xlw(
                api_ws,
                group_names,
                copy_format=_cpy_fmt,
                xl_screen=_xl_screen,
            )
            pic.Left = left0
            pic.Top = top0
            pic.Width = w0
            pic.Height = h0
            try:
                pic.Placement = _xl_move_and_size
            except Exception:
                pass
            if chroma_backdrop:
                _gantt_flatten_apply_picture_chroma_transparency_xlw(pic, fill_bgr)
            safe = "".join(
                ch if ch.isalnum() or ch in "._-" else "_" for ch in dk
            )[:200]
            try:
                pic.Name = f"GanttDayImg_{safe}"
            except Exception:
                pass
            names_by_day[dk] = []
            n_out += 1
        except Exception as e_fl:
            if backdrop_nm:
                try:
                    api_ws.Shapes(backdrop_nm).Delete()
                except Exception:
                    pass
            logging.warning(
                "結果_設備ガント: 日別シェイプ画像化をスキップしました（日キー=%s、名称数=%s: %s）",
                dk,
                len(names),
                e_fl,
            )
    return n_out
def _gantt_add_timeline_rounded_rect_labels_xlwings(
    result_path: str,
    specs: list,
    day_blocks: list | None = None,
    *,
    sheet_name: str | None = None,
) -> bool:
    """
    結果_設備ガントのタイムライン上に、角丸四角（msoShapeRoundedRectangle）でラベルを重ねる。
    依頼NOは中央のメインシェイプ（高さは行の約 1/5。結合幅が狭くてもタイムライン列幅の
    GANTT_LABEL_SHAPE_MIN_TIMELINE_COLUMNS 本相当を下限とし文字潰れを抑える。隣スロット上にはみ出し得る）。
    担当者姓はその直上に小さな角丸チップ 1 つ（結合文字が潰れない下限幅までシェイプ幅を確保、
    テキストはシェイプ内右寄せ。Z オーダーはメンバーを背面・依頼NO を前面に寄せる）。
    day_blocks が与えられ、GANTT_TIMELINE_LABELS_DAY_FLATTEN が有効なとき、日ごとに画像へ集約する。
    成功時 True。Excel 経由でない場合や COM 不可時は False。
    """
    rp = (result_path or "").strip()
    if not rp or not os.path.isfile(rp) or not specs:
        return False
    try:
        import xlwings as xw
    except ImportError:
        return False
    app = None
    wb = None
    _perf_snap = None
    try:
        n_specs = len(specs)
        shn = sheet_name or RESULT_SHEET_GANTT_NAME
        logging.info(
            "%s: Excel で角丸シェイプを追加します（候補 %s 件）。"
            " 件数が多いと数分かかり、完了までログが増えない時間が続くことがあります。",
            shn,
            n_specs,
        )
        app = xw.App(visible=False)
        app.display_alerts = False
        wb = app.books.open(os.path.abspath(rp), update_links=False)
        _perf_snap = _xlwings_app_save_perf_state_push(app)
        try:
            sht = wb.sheets[shn]
        except Exception:
            return False
        api_ws = sht.api
        # msoShapeRoundedRectangle = 5
        _mso_round_rect = 5
        _mso_bring_to_front = 0
        _mso_send_to_back = 1
        _xl_move_and_size = 1
        _xl_h_align_center = -4131
        _xl_h_align_right = -4152
        # 件数が多いときの進捗ログ間隔（小さすぎると I/O 負荷、大きすぎると停止に見える）
        _progress_every = 10
        n_added = 0
        names_by_day: dict[str, list[str]] = defaultdict(list)

        def _record_day_shape(shp_obj, day_k: str):
            if not day_k or shp_obj is None:
                return
            try:
                names_by_day[day_k].append(str(shp_obj.Name))
            except Exception:
                pass

        # 同一データ行ごとにシェイプを 3 段（行高の各 1/3 の帯）でローテーション配置（4 件目は上段に戻る）。
        # 依頼NO メインは行高の 1/5 を目標にし、帯の上下にインセットを取って罫線付近への食み出しを抑える。
        # メンバー名は上下分割せず、依頼NO の直上に 1 シェイプで置く（全角空白区切り。人数分の AddShape はしない）。
        # メンバーは ZOrder SendToBack、依頼NO は BringToFront（幅はみ出し時も依頼NOが手前に来る）。
        # メンバー帯の縦幅は依頼NO メインと同じ。印刷で上行にはみ出さないよう、行矩形内に収める。
        _row_shape_seq: dict[int, int] = {}

        def _gantt_xlw_timeline_main_font_pt(xw: float, cap: str) -> float:
            """狭い結合セルではフォントを下げ、glyph のシェイプ外はみ出しを抑える。"""
            nch = max(1, len(str(cap or "").strip()))
            raw = float(xw) / max(nch * 0.62, 4.0)
            return max(5.25, min(9.0, raw))

        def _gantt_xlw_member_pill_font_pt(pwidth: float, nm: str) -> float:
            nch = max(1, len(str(nm or "").strip()))
            raw = float(pwidth) / max(nch * 1.05, 3.2)
            return max(5.5, min(6.5, raw))

        def _gantt_xlw_member_combined_min_width_pt(combined: str) -> float:
            """メンバー結合文字列が最低フォントでも潰れないよう必要幅（pt）の粗い下限。"""
            nch = max(1, len(str(combined or "").strip()))
            f_min = 5.75
            return f_min * max(float(nch) * 1.1, 4.0) + 7.0

        def _gantt_xlw_add_round_rect(
            x_left,
            x_top,
            x_w,
            x_h,
            caption,
            *,
            fill_rgb,
            line_rgb,
            text_rgb,
            font_pt=9.0,
            bold=True,
            italic=False,
            line_wt=0.75,
            line_dash=False,
            adj_round=0.2,
            shadow=False,
            shape_name=None,
            tf_margin_tb=None,
            tf_margin_lr=None,
            z_bring_to_front=True,
            text_h_align=None,
        ):
            cap = str(caption or "").strip()
            if x_w <= 0 or x_h <= 0 or not cap:
                return None
            shp_local = api_ws.Shapes.AddShape(
                _mso_round_rect, float(x_left), float(x_top), float(x_w), float(x_h)
            )
            if shape_name:
                try:
                    shp_local.Name = shape_name
                except Exception:
                    pass
            try:
                shp_local.Placement = _xl_move_and_size
            except Exception:
                pass
            try:
                if z_bring_to_front:
                    shp_local.ZOrder(_mso_bring_to_front)
                else:
                    shp_local.ZOrder(_mso_send_to_back)
            except Exception:
                pass
            try:
                shp_local.Fill.Visible = True
                shp_local.Fill.Solid()
                shp_local.Fill.ForeColor.RGB = fill_rgb
                shp_local.Line.Visible = True
                shp_local.Line.ForeColor.RGB = line_rgb
                shp_local.Line.Weight = line_wt
                # msoLineSolid=1, msoLineDash=4（Office VBA MsoLineDashStyle）
                try:
                    shp_local.Line.DashStyle = 4 if line_dash else 1
                except Exception:
                    pass
            except Exception:
                pass
            if adj_round is not None:
                try:
                    shp_local.Adjustments[1] = adj_round
                except Exception:
                    pass
            if shadow:
                try:
                    sd0 = shp_local.Shadow
                    sd0.Visible = -1  # msoTrue
                    sd0.OffsetX = 3
                    sd0.OffsetY = 3
                    sd0.Transparency = 0.55
                    try:
                        sd0.Blur = 4
                    except Exception:
                        pass
                    try:
                        sd0.ForeColor.RGB = _com_excel_bgr_rgb(40, 40, 50)
                    except Exception:
                        pass
                except Exception:
                    pass
            try:
                tf0 = shp_local.TextFrame
                try:
                    if tf_margin_lr is not None:
                        mrg_lr = float(tf_margin_lr)
                    else:
                        mrg_lr = 1.0 if font_pt <= 7.0 else 2.0
                    tf0.MarginLeft = mrg_lr
                    tf0.MarginRight = mrg_lr
                    m_tb = 0.5 if tf_margin_tb is None else float(tf_margin_tb)
                    tf0.MarginTop = m_tb
                    tf0.MarginBottom = m_tb
                except Exception:
                    pass
                try:
                    tf0.VerticalAlignment = -4108  # xlVAlignCenter
                    _hal = (
                        int(text_h_align)
                        if text_h_align is not None
                        else int(_xl_h_align_center)
                    )
                    tf0.HorizontalAlignment = _hal
                except Exception:
                    pass
                tf0.Characters().Text = cap
                nch = len(cap)
                fnt = tf0.Characters(1, nch).Font if nch > 0 else tf0.Characters().Font
                fnt.Size = font_pt
                fnt.Bold = bold
                if italic:
                    fnt.Italic = True
                try:
                    fnt.Color = text_rgb
                except Exception:
                    pass
            except Exception:
                try:
                    shp_local.TextFrame.Characters().Text = cap
                except Exception:
                    pass
            return shp_local

        for idx, sp in enumerate(specs, start=1):
            if idx == 1 or idx % _progress_every == 0 or idx == n_specs:
                logging.info(
                    "結果_設備ガント: シェイプ走査 %s/%s（確定追加 %s 件）…",
                    idx,
                    n_specs,
                    n_added,
                )
            text = str(sp.get("text") or "").strip()
            if not text:
                continue
            dk = str(sp.get("day_key") or "").strip()
            row = int(sp["row"])
            col_s = int(sp["col_s"])
            col_e = int(sp["col_e"])
            rng = sht.range((row, col_s), (row, col_e))
            left = float(rng.left)
            top = float(rng.top)
            w = float(rng.width)
            h = float(rng.height)
            if w <= 0 or h <= 0:
                continue
            _lw_sp = sp.get("line_wt")
            try:
                _lw_use = 0.75 if _lw_sp is None else float(_lw_sp)
            except (TypeError, ValueError):
                _lw_use = 0.75
            _dash_use = bool(sp.get("line_dash"))
            _fh = str(sp.get("fill_hex") or "E8E8E8")
            fill_bgr, line_bgr, text_bgr = _gantt_com_colors_from_fill_hex(_fh)
            # 依頼NO メインシェイプ: 狭い結合幅でも文字が潰れないよう、スロット列幅×下限本数を確保する
            # （隣スロット上にはみ出し得る。結合がそれ以上なら結合幅のまま）。
            try:
                slot_w = float(sht.range((row, col_s), (row, col_s)).width)
            except Exception:
                slot_w = 0.0
            if slot_w <= 0.0:
                _ns0 = max(1, int(col_e) - int(col_s) + 1)
                slot_w = float(w) / float(_ns0)
            _min_slot_cols = max(1, int(GANTT_LABEL_SHAPE_MIN_TIMELINE_COLUMNS))
            label_w = max(float(w), float(_min_slot_cols) * float(slot_w))
            # 縦位置は行を 3 等分した帯のいずれか（同一行で追加順に 0→1→2→0…）。依頼NO の高さは行高の 1/5。
            _band = float(h) / 3.0
            _h_req_no = max(9.0, float(h) / 5.0)
            _n_on_row = int(_row_shape_seq.get(row, 0))
            _slot = _n_on_row % 3
            _row_shape_seq[row] = _n_on_row + 1
            band_top = top + _slot * _band
            band_bot = band_top + _band
            _band_inset = 0.75
            mems_all = [
                str(x).strip() for x in (sp.get("member_labels") or []) if str(x).strip()
            ]
            mems_all = mems_all[:8]
            if _gantt_color_mode_full():
                mem_fill, mem_line, mem_txt = _gantt_member_pill_bgrs_for_task_fill_hex(
                    _fh
                )
            else:
                mem_fill = _com_excel_bgr_rgb(252, 252, 254)
                mem_line = _com_excel_bgr_rgb(175, 180, 188)
                mem_txt = _com_excel_bgr_rgb(38, 40, 46)
            if mems_all:
                _chip_below = bool(sp.get("member_chip_below"))
                gx = 1.0

                def _emit_member_pills(
                    names: list[str],
                    y0: float,
                    pill_h: float,
                    day_k: str,
                    *,
                    cell_w_scale: float = 1.0,
                    min_chip_w: float | None = None,
                    font_floor: float | None = None,
                ) -> None:
                    nonlocal n_added
                    if not names or pill_h <= 1.0:
                        return
                    parts: list[str] = []
                    est_w = 0.0
                    for nm in names:
                        nm2 = nm if len(nm) <= 6 else (nm[:5] + "…")
                        parts.append(nm2)
                        est_w += max(9.0, 5.2 * float(len(nm2)))
                    if len(parts) > 1:
                        est_w += float(len(parts) - 1) * gx
                    combined = "\u3000".join(parts)
                    if not combined.strip():
                        return
                    # 結合幅 w に縛ると文字が潰れるため、ピル分割時と同様の推定に加え、
                    # 最低フォント相当の下限幅を満たすまでシェイプ幅を広げる（隣セル上にはみ出し得る）。
                    _min_member_chip_w = (
                        float(min_chip_w) if min_chip_w is not None else 34.0
                    )
                    text_min_w = _gantt_xlw_member_combined_min_width_pt(combined)
                    want_w = max(
                        _min_member_chip_w, float(est_w), float(text_min_w)
                    )
                    _w_base = max(float(w) * float(cell_w_scale), _min_member_chip_w)
                    use_w = max(_w_base, want_w)
                    _fp_mem = float(_gantt_xlw_member_pill_font_pt(use_w, combined))
                    if font_floor is not None:
                        _fp_mem = max(_fp_mem, float(font_floor))
                    s_mem = _gantt_xlw_add_round_rect(
                        left,
                        y0,
                        use_w,
                        pill_h,
                        combined,
                        fill_rgb=mem_fill,
                        line_rgb=mem_line,
                        text_rgb=mem_txt,
                        font_pt=float(_fp_mem),
                        bold=True,
                        italic=False,
                        line_wt=0.55,
                        adj_round=0.42,
                        shadow=False,
                        shape_name=f"GanttMem_R{row}_C{col_s}_{_n_on_row}_{int(y0)}",
                        tf_margin_tb=0.0,
                        tf_margin_lr=0.75,
                        z_bring_to_front=False,
                        text_h_align=_xl_h_align_right,
                    )
                    if s_mem is not None:
                        n_added += 1
                        _record_day_shape(s_mem, day_k)

                if _chip_below:
                    # 日次始業準備: メイン文言は「日次始業準備」のみ。メイン・直下メンバーとも高さは行の 1/4（収まらないときは等分縮小）。
                    _gap_eff = 1.35
                    _h_quarter = max(8.0, float(h) * 0.25)
                    row_top_lim = float(top) + _band_inset
                    row_bot_lim = float(top) + float(h) - _band_inset
                    room = max(0.0, row_bot_lim - row_top_lim)
                    _twin_need = 2.0 * _h_quarter + _gap_eff
                    if room + 1e-9 >= _twin_need:
                        h_main = _h_quarter
                        h_mem_use = _h_quarter
                    else:
                        _avail = max(0.0, room - _gap_eff)
                        h_main = _avail / 2.0
                        h_mem_use = _avail / 2.0
                    _stack = h_main + _gap_eff + h_mem_use
                    y_main = row_top_lim + max(0.0, (room - _stack) / 2.0)
                    y_mem = y_main + h_main + _gap_eff
                    _emit_member_pills(mems_all, y_mem, h_mem_use, dk)
                    _main_fp = float(_gantt_xlw_timeline_main_font_pt(label_w, text))
                    shp_main = _gantt_xlw_add_round_rect(
                        left,
                        y_main,
                        label_w,
                        h_main,
                        text,
                        fill_rgb=fill_bgr,
                        line_rgb=line_bgr,
                        text_rgb=text_bgr,
                        font_pt=float(_main_fp),
                        bold=True,
                        italic=bool(sp.get("italic")),
                        line_wt=_lw_use,
                        line_dash=_dash_use,
                        adj_round=0.2,
                        shadow=False,
                        shape_name=f"GanttLbl_R{row}_C{col_s}_{_n_on_row}",
                    )
                    if shp_main is not None:
                        n_added += 1
                        _record_day_shape(shp_main, dk)
                        try:
                            shp_main.TextFrame.HorizontalAlignment = -4131  # xlHAlignLeft
                        except Exception:
                            pass
                else:
                    # メンバー縦幅＝依頼NO と同じ（行高の 1/5 目標）。行全体 [top, top+h] に収まるよう
                    # 積み上げ位置を平行移動し、収まらないときは隙間・ピル高を漸減する。（既定: メンバーは依頼NO の直上）
                    _gap_mm = 1.35
                    h_main = max(9.0, float(_h_req_no))
                    h_mem_use = h_main
                    _rin_row = 1.0
                    _rout_row = 1.0
                    row_top_b = float(top)
                    row_bot_b = float(top) + float(h)
                    _gap_eff = float(_gap_mm)
                    _hmem_eff = float(h_mem_use)
                    _hmain_eff = float(h_main)
                    y_main = band_bot - _band_inset - _hmain_eff
                    y_mem = y_main - _gap_eff - _hmem_eff
                    for _squeeze in range(28):
                        st = float(y_mem)
                        sb = float(y_main) + float(_hmain_eff)
                        lo = (row_top_b + _rin_row) - st
                        hi = (row_bot_b - _rout_row) - sb
                        if lo <= hi:
                            if lo > 0.0:
                                delta = lo
                            elif hi < 0.0:
                                delta = hi
                            else:
                                delta = 0.0
                            y_mem += delta
                            y_main += delta
                            break
                        if _gap_eff > 0.35:
                            _gap_eff = max(0.35, _gap_eff - 0.35)
                        elif _hmem_eff > 6.0:
                            _hmem_eff = max(6.0, _hmem_eff - 0.5)
                        elif _hmain_eff > 8.0:
                            _hmain_eff = max(8.0, _hmain_eff - 0.5)
                        else:
                            y_main = band_bot - _band_inset - _hmain_eff
                            y_mem = y_main - _gap_eff - _hmem_eff
                            lo2 = (row_top_b + _rin_row) - float(y_mem)
                            if lo2 > 0.0:
                                y_mem += lo2
                                y_main += lo2
                            break
                        y_main = band_bot - _band_inset - _hmain_eff
                        y_mem = y_main - _gap_eff - _hmem_eff
                    h_main = float(_hmain_eff)
                    h_mem_use = float(_hmem_eff)
                    _emit_member_pills(mems_all, y_mem, h_mem_use, dk)
                    _main_fp = _gantt_xlw_timeline_main_font_pt(label_w, text)
                    shp_main = _gantt_xlw_add_round_rect(
                        left,
                        y_main,
                        label_w,
                        h_main,
                        text,
                        fill_rgb=fill_bgr,
                        line_rgb=line_bgr,
                        text_rgb=text_bgr,
                        font_pt=float(_main_fp),
                        bold=True,
                        italic=bool(sp.get("italic")),
                        line_wt=_lw_use,
                        line_dash=_dash_use,
                        adj_round=0.2,
                        shadow=False,
                        shape_name=f"GanttLbl_R{row}_C{col_s}_{_n_on_row}",
                    )
                    if shp_main is not None:
                        n_added += 1
                        _record_day_shape(shp_main, dk)
                        try:
                            shp_main.TextFrame.HorizontalAlignment = -4131  # xlHAlignLeft
                        except Exception:
                            pass
            else:
                _nlines = max(1, str(text).count("\n") + 1)
                label_h = _h_req_no
                if _nlines > 1:
                    label_h = min(
                        _band - 2.0 * _band_inset,
                        max(_h_req_no, _h_req_no * (0.55 + 0.48 * float(_nlines))),
                    )
                y_lbl = band_top + _band_inset + max(
                    0.0, (_band - 2.0 * _band_inset - label_h) / 2.0
                )
                _solo_fp = _gantt_xlw_timeline_main_font_pt(label_w, text)
                shp = _gantt_xlw_add_round_rect(
                    left,
                    y_lbl,
                    label_w,
                    label_h,
                    text,
                    fill_rgb=fill_bgr,
                    line_rgb=line_bgr,
                    text_rgb=text_bgr,
                    font_pt=float(_solo_fp),
                    bold=True,
                    italic=bool(sp.get("italic")),
                    line_wt=_lw_use,
                    line_dash=_dash_use,
                    adj_round=0.2,
                    shadow=False,
                    shape_name=f"GanttLbl_R{row}_C{col_s}_{_n_on_row}",
                )
                if shp is not None:
                    n_added += 1
                    _record_day_shape(shp, dk)
                    try:
                        shp.TextFrame.HorizontalAlignment = -4131
                    except Exception:
                        pass
        n_flat = 0
        if (
            GANTT_TIMELINE_LABELS_DAY_FLATTEN
            and day_blocks
            and GANTT_TIMELINE_SHAPE_LABELS
        ):
            try:
                n_flat = _gantt_flatten_day_label_shapes_to_pictures_xlw(
                    api_ws, day_blocks, names_by_day
                )
            except Exception as e_flat:
                logging.warning(
                    "%s: 日別画像化に失敗しました（個別シェイプのまま保存します）: %s",
                    shn,
                    e_flat,
                )
        logging.info(
            "%s: 角丸シェイプ %s 件を反映%sして保存します（Excel）…",
            shn,
            n_added,
            f"し、日別に画像 {n_flat} 枚へ集約" if n_flat else "",
        )
        wb.save()
        return True
    except Exception as e:
        _shn_fb = sheet_name or RESULT_SHEET_GANTT_NAME
        logging.warning(
            "%s: 角丸シェイプラベルの追加に失敗しました（%s）。セル表記へフォールバックします。",
            _shn_fb,
            e,
        )
        return False
    finally:
        if _perf_snap is not None:
            try:
                _xlwings_app_save_perf_state_pop(app, _perf_snap)
            except Exception:
                pass
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass
def _gantt_floor_to_slot(dt: datetime, slot_mins: int) -> datetime:
    """壁時計の slot_mins 刻みへ床（秒・マイクロ秒は 0）。"""
    m = dt.hour * 60 + dt.minute
    m = (m // int(slot_mins)) * int(slot_mins)
    return dt.replace(hour=m // 60, minute=m % 60, second=0, microsecond=0)
def _gantt_machining_display_range_for_slot_overlap(
    ev: dict, slot_mins: int | None = None
) -> tuple[datetime | None, datetime | None]:
    """
    ガントのスロット重なり判定用区間。
    加工が slot_mins 未満のときは開始を含む1スロット枠に拡張（実 end_dt は変えない）。
    """
    sm = int(slot_mins if slot_mins is not None else GANTT_TIMELINE_SLOT_MINUTES)
    st = ev.get("start_dt")
    ed = ev.get("end_dt")
    if not isinstance(st, datetime) or not isinstance(ed, datetime) or ed <= st:
        return st, ed
    if not _is_machining_timeline_event(ev):
        return st, ed
    slot_secs = float(sm) * 60.0
    if (ed - st).total_seconds() >= slot_secs - 1e-9:
        return st, ed
    floored = _gantt_floor_to_slot(st, sm)
    return floored, floored + timedelta(minutes=float(sm))
