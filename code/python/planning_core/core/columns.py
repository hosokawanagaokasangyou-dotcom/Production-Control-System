# -*- coding: utf-8 -*-
# planning_core.core.columns — body only (loaded via _core exec chain)
EXCLUDE_RULES_E_SIDECAR_FILENAME = "exclude_rules_e_column_pending.json"
STAGE1_EXCLUDE_RULES_JSON_FILENAME = "stage1_exclude_rules.json"
ENV_SUMMARY_AI_DISPATCH_WORKBOOK = "PM_AI_SUMMARY_AI_DISPATCH_WORKBOOK"
EXCLUDE_RULES_E_VBA_TSV_FILENAME = "exclude_rules_e_column_vba.tsv"
EXCLUDE_RULES_MATRIX_VBA_FILENAME = "exclude_rules_matrix_vba.tsv"
SHEET_MACHINE_CALENDAR = "機械カレンダー"
SHEET_MACHINE_DAILY_STARTUP = "設定_機械_日次始業準備"
SHEET_REQUEST_SWITCH_PREP = "設定_依頼切替前後時間"
TIMELINE_EVENT_MACHINING = "machining"
TIMELINE_EVENT_MACHINE_DAILY_STARTUP = "machine_daily_startup"
TIMELINE_EVENT_REQUEST_SWITCH_PREP = "request_switch_prep"
TIMELINE_EVENT_BREAK_RESUME_PREP = "break_resume_prep"
TIMELINE_EVENT_POST_MACHINING_CLEANUP = "post_machining_cleanup"
TIMELINE_EVENT_REQUEST_INTERVAL_BUFFER = "request_interval_buffer"
MASTER_SHEET_TEAM_COMBINATIONS = "組み合わせ表"
MASTER_SHEET_SPEED = os.environ.get("MASTER_SPEED_SHEET_NAME", "").strip() or "speed"
ATT_COL_LEAVE_TYPE = "休暇区分"
ATT_COL_REMARK = "備考"
ATT_COL_OT_END = "残業(分)"
ATT_COL_OT_END_LEGACY = "残業終業"
ATTENDANCE_REMARK_AI_SCHEMA_ID = "v3_sabun_batch"
NEED_COL_CONDITION = "依頼NO条件"
NEED_COL_NOTE = "備考"
RESULT_SHEET_GANTT_NAME = "結果_設備ガント"
RESULT_BOOK_FONT_NAME = "BIZ UDゴシック"
TASK_COL_TASK_ID = "依頼NO"
TASK_COL_ORDER_NO = "受注NO"
TASK_COL_MACHINE = "工程名"
TASK_COL_MACHINE_NAME = "機械名"
TASK_COL_QTY = "換算数量"
TASK_COL_UNPROCESSED = "未加工"
TASK_COL_ORDER_QTY = "受注数"
TASK_COL_SPEED = "加工速度"
TASK_COL_PRODUCT = "製品名"
TASK_COL_ANSWER_DUE = "回答納期"
TASK_COL_SPECIFIED_DUE = "指定納期"
TASK_COL_RAW_INPUT_DATE = "原反投入日"
TASK_COL_STOCK_LOCATION = "在庫場所"
TASK_COL_USED_RAW = "使用原反"
TASK_COL_PROCESS_CONTENT = "加工内容"
TASK_COL_COMPLETION_FLAG = "加工完了区分"
TASK_COL_ACTUAL_DONE = "実加工数"   # 旧互換（直接の加工済数値）
TASK_COL_ACTUAL_OUTPUT = "実出来高"  # 完了品数値（残作に使う）
TASK_COL_DATA_EXTRACTION_DT = "データ抽出日"
TASK_COL_DATA_EXTRACTION_TIME = "データ抽出時間"
TASK_COL_EXTRACTION_TIME = "抽出時間"
AI_CACHE_TTL_SECONDS = 96 * 60 * 60  # 96時間
AI_CACHE_KEY_PREFIX_EXCLUDE_RULE_DE = "exclude_rule_de_v1"
ACTUALS_SHEET_NAME = "加工実績DATA"
ACT_COL_TASK_ID = "依頼NO"
ACT_COL_PROCESS = "工程名"
ACT_COL_OPERATOR = "担当者"
ACT_COL_START_DT = "開始日時"
ACT_COL_END_DT = "終了日時"
ACT_COL_START_ALT = "実績開始"
ACT_COL_END_ALT = "実績終了"
ACT_COL_DAY = "日付"
ACT_COL_TIME_START = "開始時刻"
ACT_COL_TIME_END = "終了時刻"
ACT_COL_MACHINING_START_DT = "加工開始日時"
ACT_COL_MACHINING_END_DT = "加工終了日時"
ACT_COL_MACHINING_START_DT_WITH_STOP = "加工開始日時(停機時間加算後)"
ACT_COL_STOP_MIN_CONVERTED = "停機時間分(変換後)"
ACT_COL_ACTUAL_QTY = "実加工数"
ACT_COL_PLANNED_QTY = "加工予定数"
ACT_COL_CONVERTED_QTY = "換算数量"
ACT_COL_CUMULATIVE_ACTUAL_QTY = "累積実績"
ACT_COL_CUMULATIVE_COMPLETION_PCT = "累積完了率"
ACT_COL_MACHINING_ASSIGNEE_1 = "加工担当者名1"
ACT_COL_MACHINING_ASSIGNEE_2 = "加工担当者名2"
ACT_COL_MACHINING_ASSIGNEE_3 = "加工担当者名3"
ACT_COL_MACHINING_ASSIGNEE_4 = "加工担当者名4"
ACT_COL_MACHINING_ASSIGNEE_5 = "加工担当者名5"
ACT_COL_MACHINING_ASSIGNEES_ORDERED = (
    ACT_COL_MACHINING_ASSIGNEE_1,
    ACT_COL_MACHINING_ASSIGNEE_2,
    ACT_COL_MACHINING_ASSIGNEE_3,
    ACT_COL_MACHINING_ASSIGNEE_4,
    ACT_COL_MACHINING_ASSIGNEE_5,
)
ACTUAL_HEADER_CANONICAL = (
    ACT_COL_TASK_ID,
    ACT_COL_PROCESS,
    ACT_COL_OPERATOR,
    ACT_COL_START_DT,
    ACT_COL_END_DT,
    ACT_COL_START_ALT,
    ACT_COL_END_ALT,
    ACT_COL_DAY,
    ACT_COL_TIME_START,
    ACT_COL_TIME_END,
    ACT_COL_MACHINING_START_DT,
    ACT_COL_MACHINING_END_DT,
)
ACTUAL_DETAIL_SHEET_NAME = "加工実績明細DATA"
ACT_DETAIL_COL_ROLL = "ロールNO"
ACTUAL_DETAIL_HEADER_CANONICAL = ACTUAL_HEADER_CANONICAL + (
    ACT_COL_MACHINING_START_DT_WITH_STOP,
    ACT_COL_ACTUAL_QTY,
    ACT_COL_PLANNED_QTY,
    ACT_COL_CONVERTED_QTY,
    ACT_COL_CUMULATIVE_ACTUAL_QTY,
    ACT_COL_CUMULATIVE_COMPLETION_PCT,
    ACT_DETAIL_COL_ROLL,
) + ACT_COL_MACHINING_ASSIGNEES_ORDERED
RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME = "結果_設備ガント_実績明細"
ENV_GANTT_ACTUAL_DETAIL_DATE_FROM = "GANTT_ACTUAL_DETAIL_DATE_FROM"
ENV_GANTT_ACTUAL_DETAIL_DATE_TO = "GANTT_ACTUAL_DETAIL_DATE_TO"
ACTUAL_DETAIL_GANTT_REFRESH_FILENAME = "actual_detail_gantt_refresh.xlsx"
ENV_PM_AI_PROCESSING_PLAN_SHEET = "PM_AI_PROCESSING_PLAN_SHEET"
ENV_PM_AI_ACTUALS_DATA_SHEET = "PM_AI_ACTUALS_DATA_SHEET"
ENV_PM_AI_ACTUAL_DETAIL_SHEET = "PM_AI_ACTUAL_DETAIL_SHEET"
RESULT_SHEET_GANTT_COMPARE_NAME = "結果_設備ガント_計画実績比較"
COMPARE_GANTT_OUTPUT_FILENAME = "plan_actual_compare_gantt.xlsx"
ENV_COMPARE_GANTT_SNAPSHOT_DIR = "COMPARE_GANTT_SNAPSHOT_DIR"
ENV_COMPARE_GANTT_PLAN_TASKS_SHEET = "COMPARE_GANTT_PLAN_TASKS_SHEET"
ENV_COMPARE_GANTT_ALLOW_PLAN_OVERLAP = "COMPARE_GANTT_ALLOW_PLAN_OVERLAP"
try:
    COMPARE_GANTT_ACTUAL_SHAPE_LINE_PT = float(
        (os.environ.get("COMPARE_GANTT_ACTUAL_SHAPE_LINE_PT", "2.5") or "2.5").strip()
    )
except (TypeError, ValueError):
    COMPARE_GANTT_ACTUAL_SHAPE_LINE_PT = 2.5
COMPARE_GANTT_DAY_ROW_MAP_DATE_COL = 52  # AZ
COMPARE_GANTT_DAY_ROW_MAP_FIRSTROW_COL = 53  # BA
STAGE1_OUTPUT_FILENAME = "plan_input_tasks.xlsx"
STAGE1_PLAN_OUTPUT_SHEET = "タスク一覧"
STAGE1_TASK_INPUT_PREVIEW_FILENAME = "stage1_task_input_table.xlsx"
STAGE1_TASK_INPUT_PREVIEW_SHEET = "タスク入力整形"
PLAN_INPUT_SHEET_NAME = os.environ.get("TASK_PLAN_SHEET", "").strip() or STAGE1_PLAN_OUTPUT_SHEET
DISPATCH_TRIAL_PATTERN_LIST_SHEET_NAME = (
    os.environ.get("DISPATCH_TRIAL_PATTERN_LIST_SHEET", "").strip()
    or "配台試行順_パターン一覧"
)
DISPATCH_PATTERN_STAGE2_SUMMARY_SHEET_NAME = (
    os.environ.get("DISPATCH_PATTERN_STAGE2_SUMMARY_SHEET", "").strip()
    or "配台試行順_パターン別段階2"
)
DISPATCH_PATTERN_STAGE2_META_FILENAME = "pattern_jobs_meta.json"
_DISPATCH_TRIAL_PATTERN_P3_SORT = object()
_DISPATCH_TRIAL_PATTERN_P4_SORT = object()
_DISPATCH_TRIAL_PATTERN_P5_SORT = object()
_DISPATCH_TRIAL_PATTERN_P6_SORT = object()
def _dispatch_pattern_stage2_max_patterns() -> int:
    raw = (os.environ.get("DISPATCH_PATTERN_STAGE2_MAX_PATTERNS") or "20").strip()
    try:
        n = int(raw)
    except (TypeError, ValueError):
        n = 20
    return max(1, min(n, 50))
def _dispatch_pattern_stage2_capped_jobs() -> list[tuple[str, str, int | None, object]]:
    """
    P1～P7 のリスト。DISPATCH_PATTERN_STAGE2_MAX_PATTERNS を超えるときは先頭から切り詰める。
    """
    cap = _dispatch_pattern_stage2_max_patterns()
    jobs = _dispatch_trial_pattern_job_list()
    if len(jobs) > cap:
        jobs = jobs[:cap]
    return jobs
def _dispatch_pattern_jobs_meta_list(
    pattern_jobs: list[tuple[str, str, int | None, object]],
) -> list[dict]:
    out: list[dict] = []
    for pid, pname, seed, sk in pattern_jobs:
        if pid == "P1":
            kind = "due"
        elif pid == "P2":
            kind = "machine_due"
        elif pid == "P3":
            kind = "due_buffer"
        elif pid == "P4":
            kind = "due_minus_raw"
        elif pid == "P5":
            kind = "p2_probe_shift_raw_minus1_due_late"
        elif pid == "P6":
            kind = "p5_then_shift_raw_minus1_due_late_again"
        elif pid == "P7":
            kind = "machine_raw_input_date"
        else:
            kind = "random"
        out.append({"id": pid, "name": pname, "seed": seed, "kind": kind})
    return out
def _pattern_job_tuple_from_meta_entry(ent: dict) -> tuple[str, str, int | None, object]:
    """pattern_jobs_meta.json の 1 要素から試行順ジョブタプルを復元する。"""
    pid = str(ent.get("id") or "").strip()
    pname = str(ent.get("name") or "").strip()
    kind = str(ent.get("kind") or "").strip()
    if pid == "P1" or kind == "due":
        return (pid or "P1", pname or "納期最優先", None, _pattern_sort_key_due_priority)
    if pid == "P2" or kind == "machine_due":
        return (pid or "P2", pname or "機械名グループ+納期", None, _pattern_sort_key_machine_then_due)
    if pid == "P3" or kind == "due_buffer":
        return (
            pid or "P3",
            pname or "納期順・機械グループ(納期−原反合計が短い順)・途中依頼優先",
            None,
            _DISPATCH_TRIAL_PATTERN_P3_SORT,
        )
    # 旧 pattern_jobs_meta: kind のみ machine_qty_due のときは当時の P4（換算数量）を P2 キーで再生
    if kind == "machine_qty_due":
        return (
            pid or "P4",
            pname or "（旧互換）機械名+換算数量降順+納期",
            None,
            _pattern_sort_key_machine_then_due,
        )
    if pid == "P4" or kind == "due_minus_raw":
        return (
            pid or "P4",
            pname or "納期−原反日数の短い順(途中依頼優先)",
            None,
            _DISPATCH_TRIAL_PATTERN_P4_SORT,
        )
    if pid == "P5" or kind == "p2_probe_shift_raw_minus1_due_late":
        return (
            pid or "P5",
            pname or "P2→納期遅れ依頼のみ原反-1日→P2",
            None,
            _DISPATCH_TRIAL_PATTERN_P5_SORT,
        )
    if pid == "P6" or kind == "p5_then_shift_raw_minus1_due_late_again":
        return (
            pid or "P6",
            pname or "P5後に納期遅れのみ原反さらに-1日→P2",
            None,
            _DISPATCH_TRIAL_PATTERN_P6_SORT,
        )
    if pid == "P7" or kind == "machine_raw_input_date":
        return (
            pid or "P7",
            pname or "機械名グループ+原反投入日早い順",
            None,
            _pattern_sort_key_machine_then_raw_input_date,
        )
    # 旧 R* / kind random はシャッフル廃止のため納期最優先で決定論再生する
    if kind == "random" or (pid and str(pid).upper().startswith("R")):
        logging.warning(
            "pattern_jobs_meta: ランダム枠 id=%s は廃止のため納期最優先で再生します。",
            pid,
        )
        return (pid, pname or pid, None, _pattern_sort_key_due_priority)
    logging.warning(
        "pattern_jobs_meta: 不明な id/kind (%s / %s) を納期最優先で再生します。",
        pid,
        kind,
    )
    return (pid or "P1", pname or pid, None, _pattern_sort_key_due_priority)
def _dispatch_pattern_reference_score_from_metrics(
    due_pct, mem_pct, eq_cells
) -> float | None:
    """
    人の最終判断用の参考値（大きいほど良い想定の単純加重）。
    環境変数 DISPATCH_PATTERN_SCORE_WEIGHT_DUE / _MEMBER / _EQUIP（空なら 3 / 1 / 0.0001）。
    """
    try:
        w_d = float((os.environ.get("DISPATCH_PATTERN_SCORE_WEIGHT_DUE") or "3").strip() or 3)
    except (TypeError, ValueError):
        w_d = 3.0
    try:
        w_m = float((os.environ.get("DISPATCH_PATTERN_SCORE_WEIGHT_MEMBER") or "1").strip() or 1)
    except (TypeError, ValueError):
        w_m = 1.0
    try:
        w_e = float((os.environ.get("DISPATCH_PATTERN_SCORE_WEIGHT_EQUIP") or "0.0001").strip() or 0.0001)
    except (TypeError, ValueError):
        w_e = 0.0001
    if due_pct is None or due_pct == "":
        return None
    try:
        d = float(due_pct)
    except (TypeError, ValueError):
        return None
    m = 0.0
    if mem_pct is not None and mem_pct != "":
        try:
            m = float(mem_pct)
        except (TypeError, ValueError):
            pass
    e = 0.0
    if eq_cells is not None and eq_cells != "":
        try:
            e = float(eq_cells)
        except (TypeError, ValueError):
            pass
    return round(w_d * d + w_m * m + w_e * e, 4)
def _write_dispatch_pattern_stage2_jobs_meta(batch_root: str, pattern_jobs: list) -> None:
    p = os.path.join(batch_root, DISPATCH_PATTERN_STAGE2_META_FILENAME)
    try:
        payload = {
            "batch_root": os.path.abspath(batch_root),
            "patterns": _dispatch_pattern_jobs_meta_list(pattern_jobs),
        }
        with open(p, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except OSError as e:
        logging.warning("パターン別段階2: メタ JSON の書込に失敗しました: %s (%s)", p, e)
PLAN_COL_SPEED_OVERRIDE = "加工速度_上書き"
PLAN_COL_RAW_INPUT_DATE_OVERRIDE = "原反投入日_上書き"
PLAN_COL_DISPATCHABLE_DATETIME = "配台可能日時"
PLAN_COL_DISPATCHABLE_DATETIME_OVERRIDE = "配台可能日時_上書き"
DISPATCHABLE_FROM_TIME = time(12, 45)
def _parse_hhmm_env_or_default(name: str, default_hhmm: str) -> time:
    """``HH:MM`` 形式の環境変数を time へ。未設定・不正時は default_hhmm を使う。"""
    raw = (os.environ.get(name) or default_hhmm).strip()
    for candidate in (raw, default_hhmm):
        try:
            return datetime.strptime(candidate, "%H:%M").time()
        except (TypeError, ValueError):
            continue
    return time(0, 0)
DISPATCHABLE_FROM_TIME_KONAN_STOCK = _parse_hhmm_env_or_default(
    "DISPATCHABLE_FROM_TIME_KONAN_STOCK", "09:30"
)
def _current_factory_is_konan() -> bool:
    """選択中の利用工場が湖南工場か（``PM_AI_FACTORY_SITE``。未設定時は KONAN 扱い）。"""
    v = (os.environ.get("PM_AI_FACTORY_SITE") or "KONAN").strip().upper()
    return v == "KONAN"
def _stock_location_is_konan(value) -> bool:
    """受注ファイル「在庫場所」の値が湖南を指すか（「湖南」を含む値を対象）。"""
    s = str(value or "").strip()
    return "湖南" in s
def dispatchable_from_time_for(stock_location=None) -> time:
    """原反投入日同日の配台開始下限。湖南工場かつ在庫場所「湖南」のタスクのみ
    ``DISPATCHABLE_FROM_TIME_KONAN_STOCK``（既定9:30）を使い、他は既定の
    ``DISPATCHABLE_FROM_TIME``（12:45）を使う。"""
    if _current_factory_is_konan() and _stock_location_is_konan(stock_location):
        return DISPATCHABLE_FROM_TIME_KONAN_STOCK
    return DISPATCHABLE_FROM_TIME
PLAN_COL_PARENT_TASK_ID = "元依頼NO"
PLAN_COL_BRANCH_SEQ = "配台枝番"
PLAN_COL_LIMITED_OP = "担当OP_限定"
PLAN_COL_SPECIAL_REMARK = "特別指定_備考"
PLAN_COL_EXCLUDE_FROM_ASSIGNMENT = "配台不要"
PLAN_COL_STAGE2_DISPATCH_PLAN_EXCLUDE_MARKER = "配台計画除外"
PLAN_COL_AI_PARSE = "AI特別指定_解析"
# 工場別に見出しを変えたブック向け（書き戻し・Java 表示専用列の解決に使用）
PLAN_COL_AI_PARSE_ALIASES = (
    PLAN_COL_AI_PARSE,
    "AI納期回答_解析",
)
PLAN_COL_SPECIAL_REMARK_ALIASES = (
    PLAN_COL_SPECIAL_REMARK,
    "納期回答_備考",
)
PLAN_COL_EC_SIDE_CLASS = "EC面区分"
PLAN_COL_PROCESS_FACTOR = "加工工程の決定プロセスの因子"
PLAN_COL_ROLL_UNIT_LENGTH = "(製品)ロール単位長さ"
PLAN_COL_ROLL_UNIT_LENGTH_LEGACY = "ロール単位長さ"
PLAN_DF_ATTR_EFFECTIVE_ROLL_UNIT_DATA_ILOCS = "_pm_ai_effective_roll_unit_data_ilocs"
PLAN_COL_DISPATCH_REMAINING_QTY = "配台使用残数量"
PLAN_COL_DISPATCH_ROLL_COUNT = "配台ロール数"
PLAN_COL_RAW_FABRIC_WIDTH = "原反幅"
PLAN_COL_RAW_ROLL_UNIT_LENGTH = "(原反)ロール単位長さ"
RAW_FABRIC_WIDTH_TABLE_DEFAULT_FILENAME = "使用原反, 加工幅.txt"
RAW_FABRIC_WIDTH_TABLE_PATH_ENV = "RAW_FABRIC_WIDTH_TABLE_PATH"
PLAN_COL_PRODUCT_WIDTH = "製品幅"
PRODUCT_WIDTH_TABLE_DEFAULT_FILENAME = "製品名, 製品幅.txt"
PRODUCT_WIDTH_TABLE_PATH_ENV = "PRODUCT_WIDTH_TABLE_PATH"
PLAN_COL_PRODUCT_LENGTH = "製品長"
PRODUCT_LENGTH_TABLE_DEFAULT_FILENAME = "製品名,製品長.txt"
PRODUCT_LENGTH_TABLE_PATH_ENV = "PRODUCT_LENGTH_TABLE_PATH"
PLAN_COL_PRODUCT_THICKNESS = "製品厚み"
PRODUCT_THICKNESS_TABLE_DEFAULT_FILENAME = "製品名,製品厚み.txt"
PRODUCT_THICKNESS_TABLE_PATH_ENV = "PRODUCT_THICKNESS_TABLE_PATH"
ROLL_UNIT_LENGTH_CEIL_STEP_M = 100.0
INFER_ROLL_UNIT_LENGTH_DEFAULT_NO_MATCH_M = 100.0
DEBUG_TASK_ID = os.environ.get("DEBUG_TASK_ID", "Y3-26").strip()
TRACE_TEAM_ASSIGN_TASK_ID = os.environ.get("TRACE_TEAM_ASSIGN_TASK_ID", "").strip()
TRACE_SCHEDULE_TASK_IDS: frozenset[str] = frozenset()
DEBUG_DISPATCH_ONLY_TASK_IDS: frozenset[str] = frozenset()
DISPATCH_TRACE_OUTER_ROUND: int = 0
_STAGE1_MATERIAL_TABLE_APPEND_BUILD = "20260522d2735c-write-canonical"
def _trace_schedule_task_enabled(task_id) -> bool:
    if not TRACE_SCHEDULE_TASK_IDS:
        return False
    return str(task_id or "").strip() in TRACE_SCHEDULE_TASK_IDS
def _sanitize_dispatch_trace_filename_part(task_id: str) -> str:
    """依頼NOを log ファイル名に使うための簡易サニタイズ（Windows 禁止文字を避ける）。"""
    s = "".join(
        c if (c.isalnum() or c in "-_.") else "_"
        for c in str(task_id or "").strip()
    )
    return s[:120] if s else "task"
def _reset_dispatch_trace_per_task_logfiles() -> None:
    """
    段階2実行の冒頭で1回」log 内の dispatch_trace_*.txt をまとめて削除する（除去実行の残骸を残さない）。
    坄外側ラウンド用ファイルは generate_plan の while 先頭で _dispatch_trace_begin_outer_round はヘッダ付し新規作成する。
    execution_log.txt とは別ファイル。内容は [配台トレース task=…] 行を _log_dispatch_trace_schedule で追記
    （日次残・ロール確定の余剰有無・余力追記・終了時サマリ等）。
    """
    if not TRACE_SCHEDULE_TASK_IDS:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
    except OSError:
        return
    try:
        for _name in os.listdir(log_dir):
            if not (
                str(_name).startswith("dispatch_trace_") and str(_name).endswith(".txt")
            ):
                continue
            _p = os.path.join(log_dir, _name)
            try:
                os.unlink(_p)
            except OSError:
                pass
    except OSError:
        pass
def _dispatch_trace_begin_outer_round(round_n: int) -> None:
    """紝期超靎リトライの外側ラウンド番坷を確定し、当ラウンド用 dispatch_trace_*_rNN.txt のヘッダを1回だけ書き。"""
    global DISPATCH_TRACE_OUTER_ROUND
    DISPATCH_TRACE_OUTER_ROUND = max(0, int(round_n))
    if not TRACE_SCHEDULE_TASK_IDS:
        return
    try:
        os.makedirs(log_dir, exist_ok=True)
    except OSError:
        return
    for tid in TRACE_SCHEDULE_TASK_IDS:
        t = str(tid or "").strip()
        if not t:
            continue
        safe = _sanitize_dispatch_trace_filename_part(t)
        path = os.path.join(
            log_dir,
            f"dispatch_trace_{safe}_r{DISPATCH_TRACE_OUTER_ROUND:02d}.txt",
        )
        if os.path.exists(path):
            continue
        try:
            with open(path, "w", encoding="utf-8", newline="\n") as f:
                f.write(
                    "# 配台トレース（依頼NOとと・外側ラウンド別）。同一行は log/execution_log.txt にも出力されした。\n"
                    f"# task_id={t}  outer_round={DISPATCH_TRACE_OUTER_ROUND}  "
                    "# （0=初回カレンダー通し、以降は紝期超靎リトライごとに +1）\n\n"
                )
        except OSError as ex:
            logging.warning("dispatch_trace ログの初期化に失敗: %s (%s)", path, ex)
def _log_dispatch_trace_schedule(task_id, msg: str, *args) -> None:
    """[配台トレース task=…] を execution_log に出しつつ」対象依頼NO専用ファイルにも追記れる。"""
    t = str(task_id or "").strip()
    body_raw = msg % args if args else msg
    body = body_raw
    if t and t in TRACE_SCHEDULE_TASK_IDS:
        body = f"[outer_round={DISPATCH_TRACE_OUTER_ROUND:02d}] {body_raw}"
    logging.info(body)
    if not t or t not in TRACE_SCHEDULE_TASK_IDS:
        return
    safe = _sanitize_dispatch_trace_filename_part(t)
    path = os.path.join(
        log_dir,
        f"dispatch_trace_{safe}_r{DISPATCH_TRACE_OUTER_ROUND:02d}.txt",
    )
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S,%f")[:-3]
        line = f"{ts} - INFO - {body}\n"
        with open(path, "a", encoding="utf-8", newline="\n") as f:
            f.write(line)
    except OSError as ex:
        try:
            logging.warning("dispatch_trace 側ファイルへの追記に失敗: %s (%s)", path, ex)
        except Exception:
            pass
TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF = os.environ.get(
    "TEAM_ASSIGN_PRIORITIZE_SURPLUS_STAFF", "0"
).strip().lower() not in ("0", "false", "no", "off", "いいえ")
def _team_assign_start_slack_wait_minutes() -> int:
    """全日候補の最早開始からこの分以内の遅れなら」開始より人数を優先（分）。0 で無効。"""
    raw = os.environ.get("TEAM_ASSIGN_START_SLACK_WAIT_MINUTES", "60").strip()
    try:
        v = int(raw)
    except ValueError:
        v = 60
    return max(0, v)
TEAM_ASSIGN_START_SLACK_WAIT_MINUTES = _team_assign_start_slack_wait_minutes()
TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW = (
    os.environ.get("TEAM_ASSIGN_IGNORE_NEED_SURPLUS_ROW", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)
TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS = (
    os.environ.get("TEAM_ASSIGN_USE_NEED_SURPLUS_IN_MAIN_PASS", "")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)
TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY = (
    os.environ.get("TEAM_ASSIGN_HEADCOUNT_FROM_NEED_ONLY", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)
TEAM_ASSIGN_USE_MASTER_COMBO_SHEET = (
    os.environ.get("TEAM_ASSIGN_USE_MASTER_COMBO_SHEET", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)
TEAM_ASSIGN_COMBO_SHEET_MAY_EXCEED_NEED = (
    os.environ.get("TEAM_ASSIGN_COMBO_SHEET_MAY_EXCEED_NEED", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)
TEAM_ASSIGN_COMBO_SHEET_RESTRICT_TO_PRESET_MEMBERS = (
    os.environ.get("TEAM_ASSIGN_COMBO_SHEET_RESTRICT_TO_PRESET_MEMBERS", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ")
)
PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE = (
    os.environ.get("PLANNING_B1_INSPECTION_EXCLUSIVE_MACHINE", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ", "無効")
)
PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS = (
    os.environ.get("PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS", "1")
    .strip()
    .lower()
    not in ("0", "false", "no", "off", "いいえ", "無効")
)
EXCLUDE_RULES_SHEET_NAME = "設定_配台不要工程"
EXCLUDE_RULE_COL_PROCESS = "工程名"
EXCLUDE_RULE_COL_MACHINE = "機械名"
EXCLUDE_RULE_COL_FLAG = "配台不要"
EXCLUDE_RULE_COL_LOGIC_JA = "配台不要ロジック"
EXCLUDE_RULE_COL_LOGIC_JSON = "ロジック式"
_exclude_rules_effective_read_path: str | None = None
_exclude_rules_rules_snapshot: list | None = None
_exclude_rules_snapshot_wb: str | None = None
EXCLUDE_RULE_ALLOWED_COLUMNS = frozenset(
    {
        TASK_COL_TASK_ID,
        TASK_COL_ORDER_NO,
        TASK_COL_MACHINE,
        TASK_COL_MACHINE_NAME,
        TASK_COL_QTY,
        TASK_COL_UNPROCESSED,
        TASK_COL_ORDER_QTY,
        TASK_COL_SPEED,
        TASK_COL_PRODUCT,
        TASK_COL_ANSWER_DUE,
        TASK_COL_SPECIFIED_DUE,
        TASK_COL_RAW_INPUT_DATE,
        TASK_COL_STOCK_LOCATION,
        TASK_COL_USED_RAW,
        PLAN_COL_RAW_FABRIC_WIDTH,
        PLAN_COL_RAW_ROLL_UNIT_LENGTH,
        PLAN_COL_PRODUCT_WIDTH,
        PLAN_COL_PRODUCT_LENGTH,
        PLAN_COL_PRODUCT_THICKNESS,
        TASK_COL_PROCESS_CONTENT,
        TASK_COL_COMPLETION_FLAG,
        TASK_COL_ACTUAL_DONE,
        TASK_COL_ACTUAL_OUTPUT,
        TASK_COL_DATA_EXTRACTION_DT,
        TASK_COL_DATA_EXTRACTION_TIME,
        TASK_COL_EXTRACTION_TIME,
        PLAN_COL_SPEED_OVERRIDE,
        PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
        PLAN_COL_SPECIAL_REMARK,
        PLAN_COL_PROCESS_FACTOR,
        PLAN_COL_ROLL_UNIT_LENGTH,
        PLAN_COL_DISPATCH_REMAINING_QTY,
        PLAN_COL_DISPATCH_ROLL_COUNT,
    }
)
RESULT_TASK_SHEET_NAME = "結果_タスク一覧"
RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME = "結果_設備毎の時間割"
RESULT_EQUIPMENT_BY_MACHINE_SHEET_NAME = "結果_設備毎の時間割_機械名毎"
RESULT_DISPATCH_TABLE_SHEET_NAME = "結果_配台表"
RESULT_DISPATCH_TABLE_EXCEL_TABLE_NAME = "_t結果_配台表"
RESULT_DISPATCH_TABLE_JSON_FILENAME = "結果_配台表.json"
_RESULT_DISPATCH_PLAN_INPUT_OVERRIDE_SRC_COLS: frozenset[str] = frozenset(
    {
        TASK_COL_ACTUAL_DONE,
        TASK_COL_ACTUAL_OUTPUT,
        TASK_COL_QTY,
        TASK_COL_RAW_INPUT_DATE,
    }
)
# 結果_配台表: 加工計画DATA に無いときは空欄（配台計画入力・task_queue へフォールバックしない）
_RESULT_DISPATCH_PROCESSING_PLAN_ONLY_SRC_COLS: frozenset[str] = frozenset(
    {
        TASK_COL_ANSWER_DUE,
    }
)
RESULT_DISPATCH_TABLE_STATIC_HEADERS: tuple[str, ...] = (
    "配台試行順番",
    "工程名",
    "機械名",
    "受注日",
    "受注NO",
    "依頼NO",
    "品名(原反)",
    "使用原反",
    "原反数",
    "品名(製品)",
    "製品名",
    "換算数量",
    "実加工数",
    "加工内容",
    "在庫場所",
    "原反投入日",
    "指定納期",
    "回答納期",
    "加工完了日",
    "加工完了区分",
    "実出来高",
    "計画合計",
    "原反投入場所",
    "加工開始日時",
    "加工終了日時",
    "メンバー名",
)
RESULT_DISPATCH_TABLE_DATE_HEADERS = frozenset(
    {
        "受注日",
        "原反投入日",
        "指定納期",
        "回答納期",
        "加工完了日",
        "配台日",
    }
)
INTERACTIVE_DISPATCH_ACTUAL_QTY_COL = "実配台数量"
RESULT_OUTSIDE_REGULAR_TIME_FILL = "FCE4D6"
RESULT_DISPATCHED_REQUEST_FILL = "C6EFCE"
RESULT_MACHINE_CALENDAR_BLOCK_FILL = "D4B3E8"
RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS = (
    "E8F4FC",
    "FCE8F0",
    "E8F8E8",
    "FFF0D8",
    "EDE8FC",
    "E0F8F4",
    "F8E8E0",
    "E8ECF8",
    "F5F5E0",
    "F0E8E8",
)
RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS_FULL = (
    "B3E5FC",
    "F8BBD0",
    "C8E6C9",
    "FFE0B2",
    "E1BEE7",
    "B2EBF2",
    "FFF59D",
    "D1C4E9",
    "FFCDD2",
    "C5CAE9",
)
RESULT_TASK_COL_DISPATCH_TRIAL_ORDER = "配台試行順番"
RESULT_TASK_COL_RAW_INPUT_DATE_PRE_PATTERN = "原反投入日_試行前"
RESULT_TASK_COL_PATTERN_RAW_SHIFT_NOTE = "試行順パターン原反前倒し"
RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16 = "納期を満たすか？"
RESULT_MEMBER_PRIORITY_SHEET_NAME = "結果_人員配台優先順"
RESULT_MEMBER_WORK_UTIL_SHEET_NAME = "結果_メンバー別作業割合"
COLUMN_CONFIG_SHEET_NAME = "列設定_結果_タスク一覧"
COLUMN_CONFIG_HEADER_COL = "列名"
COLUMN_CONFIG_VISIBLE_COL = "表示"
STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT = os.environ.get(
    "STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT", "1"
).strip().lower() in ("1", "true", "yes", "on")
GANTT_TIMELINE_SHAPE_LABELS = os.environ.get(
    "GANTT_TIMELINE_SHAPE_LABELS", "1"
).strip().lower() in ("1", "true", "yes", "on")
GANTT_TIMELINE_LABELS_DAY_FLATTEN = os.environ.get(
    "GANTT_TIMELINE_LABELS_DAY_FLATTEN", "1"
).strip().lower() in ("1", "true", "yes", "on")
GANTT_DAY_IMAGE_CHROMA_TRANSPARENT = os.environ.get(
    "GANTT_DAY_IMAGE_CHROMA_TRANSPARENT", "0"
).strip().lower() in ("1", "true", "yes", "on")
RESULT_TASK_DATE_STYLE_HEADERS = frozenset(
    {
        "回答納期",
        "指定納期",
        "計画基準納期",
        TASK_COL_RAW_INPUT_DATE,
        RESULT_TASK_COL_RAW_INPUT_DATE_PRE_PATTERN,
        "加工開始日",
        "配台済_加工開始",
        "配台済_加工終了",
    }
)
SOURCE_BASE_COLUMNS = [
    TASK_COL_TASK_ID,
    TASK_COL_ORDER_NO,
    TASK_COL_MACHINE,
    TASK_COL_MACHINE_NAME,
    TASK_COL_QTY,
    TASK_COL_UNPROCESSED,
    TASK_COL_ORDER_QTY,
    TASK_COL_SPEED,
    TASK_COL_PRODUCT,
    TASK_COL_ANSWER_DUE,
    TASK_COL_SPECIFIED_DUE,
    TASK_COL_RAW_INPUT_DATE,
    TASK_COL_STOCK_LOCATION,
    TASK_COL_USED_RAW,
    TASK_COL_PROCESS_CONTENT,
    TASK_COL_COMPLETION_FLAG,
    TASK_COL_ACTUAL_DONE,
    TASK_COL_ACTUAL_OUTPUT,
]
PLAN_OVERRIDE_COLUMNS = [
    PLAN_COL_EXCLUDE_FROM_ASSIGNMENT,
    PLAN_COL_SPEED_OVERRIDE,
    PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
    PLAN_COL_DISPATCHABLE_DATETIME_OVERRIDE,
    PLAN_COL_SPECIAL_REMARK,
    PLAN_COL_AI_PARSE,
]
PLAN_DEPRECATED_OVERRIDE_COLUMNS = frozenset(
    {
        PLAN_COL_SPEED_OVERRIDE,
        PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
        PLAN_COL_DISPATCHABLE_DATETIME_OVERRIDE,
    }
)
PLAN_OVERRIDE_TO_BASE_COLUMN = {
    PLAN_COL_SPEED_OVERRIDE: TASK_COL_SPEED,
    PLAN_COL_RAW_INPUT_DATE_OVERRIDE: TASK_COL_RAW_INPUT_DATE,
    PLAN_COL_DISPATCHABLE_DATETIME_OVERRIDE: PLAN_COL_DISPATCHABLE_DATETIME,
}
PLAN_CONFLICT_STYLABLE_COLS = tuple(PLAN_OVERRIDE_COLUMNS)
PLAN_STAGE1_MERGE_COLUMNS = tuple(
    PLAN_OVERRIDE_TO_BASE_COLUMN[c]
    if c in PLAN_DEPRECATED_OVERRIDE_COLUMNS
    else c
    for c in PLAN_OVERRIDE_COLUMNS
)
# 「AI特別指定_解析」セルへ書く文字数の上限（結果_タスク一覧「特別指定_AI」と同じ 500 文字で切る）
PLAN_AI_SPECIAL_PARSE_CELL_MAX_LEN = 500
PLAN_STAGE1_MERGE_EXTRA_COLUMNS = (PLAN_COL_ROLL_UNIT_LENGTH,)
PLANNING_CONFLICT_SIDECAR = "planning_conflict_highlight.tsv"
PLAN_INPUT_AI_SPECIAL_PARSE_SIDECAR = "plan_input_ai_special_parse.json"
PLAN_SHEET_GLOBAL_PARSE_LABEL_COL = 50  # AX
PLAN_SHEET_GLOBAL_PARSE_VALUE_COL = 51  # AY
PLAN_SHEET_GLOBAL_PARSE_MAX_ROWS = 42
def plan_reference_column_name(override_col: str) -> str:
    """上書き列の左隣に置く参照列の見出し（廃止・読込時に削除）。互換のため関数は残す。"""
    return f"（元）{override_col}"
def _plan_column_is_original_reference(col_name: str) -> bool:
    """見出しが ``（元）`` / ``(元)`` で始まる参照列なら True。"""
    if not col_name or not str(col_name).strip():
        return False
    s = str(col_name).strip()
    return s.startswith("（元）") or s.startswith("(元)")
def plan_input_sheet_column_order():
    """
    配台計画_タスク入力の列順（段階1出力・段階2読込で共通）。

    0. 配台試行順番（段階1抽出直後に空クリア→段階2と同じ趣旨に付与。段階2は全行に値はあるとしこの順を優先）
    1. 配台不要（参照列なし）
    2. 加工計画DATA 由来（SOURCE_BASE_COLUMNS）… 依頼NO〜実出来高まで（換算数量の次に未加工→配台使用残数量→配台ロール数、製品名の直後に(製品)ロール単位長さ・製品幅、原反投入日の直後に在庫場所・使用原反の直後に(原反)ロール単位長さ・原反幅）
       （(製品)ロール単位長さは製品名テーブル→製品名寸法のみ。(原反)ロール単位長さは使用原反テーブル→使用原反文字列の寸法→いずれも不可なら「不明」）
    3. 加工工程の決定プロセスの因孝
    4. 手入力列… 担当OP_限定・特別指定_備考・AI特別指定_解析 等（{@code （元）…} 参照列・*_上書き・担当OP_指定 列は廃止）

    「加工速度」列は master.xlsm「speed」（基本速度×実稼働比率）で埋め、配台の実効速度は列「加工速度」のみ。
    global_speed_rules 等で変える実効速度は計画シート列には出ないが、配台で確定した値は結果_タスク一覧の「加工速度」列に出力される。
    """
    cols = [RESULT_TASK_COL_DISPATCH_TRIAL_ORDER, PLAN_COL_EXCLUDE_FROM_ASSIGNMENT]
    for c in SOURCE_BASE_COLUMNS:
        cols.append(c)
        if c == TASK_COL_UNPROCESSED:
            cols.append(PLAN_COL_DISPATCH_REMAINING_QTY)
            cols.append(PLAN_COL_DISPATCH_ROLL_COUNT)
        if c == TASK_COL_PRODUCT:
            cols.append(PLAN_COL_ROLL_UNIT_LENGTH)
            cols.append(PLAN_COL_PRODUCT_WIDTH)
            cols.append(PLAN_COL_PRODUCT_LENGTH)
            cols.append(PLAN_COL_PRODUCT_THICKNESS)
        if c == TASK_COL_USED_RAW:
            cols.append(PLAN_COL_RAW_ROLL_UNIT_LENGTH)
            cols.append(PLAN_COL_RAW_FABRIC_WIDTH)
        if c == TASK_COL_PROCESS_CONTENT:
            cols.append(PLAN_COL_EC_SIDE_CLASS)
    cols.append(PLAN_COL_PROCESS_FACTOR)
    # 段階1算出の「配台可能日時」（表示列）。上書き列は下の PLAN_OVERRIDE_COLUMNS ループで「（元）…」参照付きで出力。
    cols.append(PLAN_COL_DISPATCHABLE_DATETIME)
    for c in PLAN_OVERRIDE_COLUMNS:
        if c == PLAN_COL_EXCLUDE_FROM_ASSIGNMENT:
            continue
        if c in PLAN_DEPRECATED_OVERRIDE_COLUMNS:
            continue
        cols.append(c)
    cols.append(PLAN_COL_LIMITED_OP)
    return cols
def _format_paren_ref_scalar(val):
    """参照表示用: 空は（―）」日付・しの他は（値）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "（―）"
    if isinstance(val, datetime):
        d = val.date() if hasattr(val, "date") else val
        if isinstance(d, date):
            return f"（{d.year}/{d.month}/{d.day}）"
    if isinstance(val, date):
        return f"（{val.year}/{val.month}/{val.day}）"
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return "（―）"
    return f"（{s}）"
def _reference_text_for_override_row(row, override_col: str, req_map: dict, need_rules: list) -> str:
    """1行分の上書き列に対応れる参照文言（括弧付し）。"""
    _ = (req_map, need_rules)  # 旧「（元）必須人数」参照で使用。列廃止により未使用だが呼び出し互換のため残す。
    if override_col == PLAN_COL_SPEED_OVERRIDE:
        v = row.get(TASK_COL_SPEED)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "（―）"
        try:
            x = float(v)
            if abs(x - round(x)) < 1e-9:
                return f"（{int(round(x))}）"
            return f"（{x}）"
        except (TypeError, ValueError):
            return _format_paren_ref_scalar(v)
    if override_col == PLAN_COL_SPECIAL_REMARK:
        return "（―）"
    if override_col == PLAN_COL_RAW_INPUT_DATE_OVERRIDE:
        return _format_paren_ref_scalar(
            parse_optional_date(_planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE))
        )
    if override_col == PLAN_COL_DISPATCHABLE_DATETIME_OVERRIDE:
        raw = parse_optional_date(
            _planning_df_cell_scalar(row, PLAN_COL_RAW_INPUT_DATE_OVERRIDE)
        ) or parse_optional_date(
            _planning_df_cell_scalar(row, TASK_COL_RAW_INPUT_DATE)
        )
        dt = compute_dispatchable_datetime(raw)
        return f"（{format_dispatchable_datetime_cell(dt)}）" if dt is not None else "（―）"
    return "（―）"
def _refresh_plan_reference_columns(df, req_map: dict, need_rules: list):
    """廃止: （元）参照列は UI から削除済み。互換のため呼び出しは残す。"""
    return df
def _apply_plan_input_visual_format(path: str, sheet_name: str = "タスク一覧"):
    """上書き入力列に薄い黄色を付与（参照列は未着色。AI解析列は除外）。"""
    # 見出し文字の表記ゆれで列名検索に失敗しはうなため、段階1の列順（plan_input_sheet_column_order）の
    # 1-based 列番坷で塗る（to_excel の列順と一致させる）。
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    order = plan_input_sheet_column_order()
    col_1based = {name: i + 1 for i, name in enumerate(order)}
    if _workbook_should_skip_openpyxl_io(path):
        logging.info(
            "配台計画の視覚整形: ブックに「%s」があるため、openpyxl での着色をスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    wb = load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        last_row = ws.max_row or 1
        if last_row < 2:
            return
        for oc in PLAN_OVERRIDE_COLUMNS:
            if oc == PLAN_COL_AI_PARSE:
                continue
            if oc in PLAN_DEPRECATED_OVERRIDE_COLUMNS:
                continue
            ci = col_1based.get(oc)
            if not ci:
                continue
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=ci).fill = fill_yellow
        _ci_rul = col_1based.get(PLAN_COL_ROLL_UNIT_LENGTH)
        if _ci_rul:
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=_ci_rul).fill = fill_yellow
        _ci_raw_rul = col_1based.get(PLAN_COL_RAW_ROLL_UNIT_LENGTH)
        if _ci_raw_rul:
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=_ci_raw_rul).fill = fill_yellow
        wb.save(path)
    finally:
        wb.close()
def _planning_conflict_sidecar_path():
    return os.path.join(log_dir, PLANNING_CONFLICT_SIDECAR)
def _remove_planning_conflict_sidecar_safe():
    try:
        os.remove(_planning_conflict_sidecar_path())
    except OSError:
        pass
def resolve_plan_sheet_header_column_index(header_map, aliases: tuple[str, ...]) -> int | None:
    """見出し行の列名辞書から、別名リストのいずれかに一致する 1-based 列番号を返す。"""
    if not isinstance(header_map, dict):
        return None
    for name in aliases:
        ci = header_map.get(name)
        if ci:
            return int(ci)
    return None


def _plan_input_ai_special_parse_sidecar_path():
    return os.path.join(json_data_dir, PLAN_INPUT_AI_SPECIAL_PARSE_SIDECAR)


def _remove_plan_input_ai_special_parse_sidecar_safe():
    try:
        os.remove(_plan_input_ai_special_parse_sidecar_path())
    except OSError:
        pass


def write_plan_input_ai_special_parse_sidecar(
    sheet_name, ai_parse_by_row, *, workbook_path: str = ""
) -> None:
    """段階2: Excel 書込がスキップ／失敗したとき Java が再読込で拾える JSON。"""
    if not isinstance(ai_parse_by_row, dict):
        return
    payload = {
        "version": 1,
        "sheet": str(sheet_name or ""),
        "workbook_path": str(workbook_path or ""),
        "by_excel_row": {str(int(k)): str(v or "") for k, v in ai_parse_by_row.items()},
    }
    path = _plan_input_ai_special_parse_sidecar_path()
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def write_planning_conflict_highlight_sidecar(sheet_name, num_data_rows, conflicts_by_row):
    """
    Excel はブックを開いたままのとき保存でしない場合に」VBA 用の TSV を log に書き。
    形式: V1 / シート名 / データ行数 / クリア列をタブ結合 / 以降 行番坷\\t列名
    """
    path = _planning_conflict_sidecar_path()
    clear_cols = "\t".join(PLAN_CONFLICT_STYLABLE_COLS)
    lines = ["V1", sheet_name, str(int(num_data_rows)), clear_cols]
    for r in sorted(conflicts_by_row.keys()):
        for name in sorted(conflicts_by_row[r]):
            lines.append(f"{int(r)}\t{name}")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")
STAGE1_SHEET_DATEONLY_HEADERS = frozenset(
    {
        TASK_COL_ANSWER_DUE,
        TASK_COL_SPECIFIED_DUE,
        TASK_COL_RAW_INPUT_DATE,
        PLAN_COL_RAW_INPUT_DATE_OVERRIDE,
    }
)
def _result_font(**kwargs):
    """結果ブック用 Font（呼び出し側は size 等を指定。既定ファミリーは _effective_result_book_font_name()）。"""
    if "name" not in kwargs:
        kwargs = {**kwargs, "name": _effective_result_book_font_name()}
    return Font(**kwargs)
def _output_book_font(bold=False):
    return _result_font(bold=bold)
def _apply_output_font_to_result_sheet(ws):
    """結果_* のごうガント以外坑け: 既定フォント・1行目太字のみ（列幅は VBA AutoFit）。"""
    base = _output_book_font(bold=False)
    hdr = _output_book_font(bold=True)
    mr, mc = ws.max_row or 1, ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
        for cell in row:
            cell.font = base
    for cell in ws[1]:
        cell.font = hdr
def _stage2_plan_book_header_fill():
    """段階2 計画ブック: 見出し行の背景（メンバー別スケジュールと揃えた薄緑）。"""
    return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
def _stage2_plan_book_thin_border():
    s = Side(style="thin", color="B4B4B4")
    return Border(left=s, right=s, top=s, bottom=s)
def _apply_stage2_plan_sheet_header_fill(
    ws, *, extra_header_rows: tuple[int, ...] | None = None
):
    """
    見出し行（1 行目）にヘッダー背景を付与。結果_人員配台優先順のみ 2 つ目のヘッダー行も対象。
    """
    hdr_fill = _stage2_plan_book_header_fill()
    for cell in ws[1]:
        cell.fill = hdr_fill
    if extra_header_rows:
        mr = ws.max_row or 1
        for r in extra_header_rows:
            if r < 1 or r > mr:
                continue
            for cell in ws[r]:
                cell.fill = hdr_fill
def _apply_stage2_plan_sheet_grid_border(ws):
    """使用範囲全体に薄い罫線（後続のセル着色と独立）。"""
    b = _stage2_plan_book_thin_border()
    mr, mc = ws.max_row or 1, ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
        for cell in row:
            cell.border = b
def _apply_equipment_schedule_day_banner_row_style(ws):
    """結果_設備毎の時間割: 日付区切り行（■ YYYY/MM/DD ■）を帯状に強調。"""
    col_tb = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_tb = i
            break
    if col_tb is None:
        return
    ban_fill = PatternFill(start_color="C5E1A5", end_color="C5E1A5", fill_type="solid")
    ban_font = _result_font(bold=True, size=11, color="1A1A1A")
    mr, mc = ws.max_row or 1, ws.max_column or 1
    for r in range(2, mr + 1):
        v = ws.cell(row=r, column=col_tb).value
        if v is None:
            continue
        s = str(v).strip()
        if "■" not in s or not re.search(r"\d{4}/\d", s):
            continue
        for col_i in range(1, mc + 1):
            cell = ws.cell(row=r, column=col_i)
            cell.fill = ban_fill
            cell.font = ban_font
        try:
            cur_h = ws.row_dimensions[r].height
            ws.row_dimensions[r].height = max(float(cur_h or 15), 20.0)
        except Exception:
            pass
def _apply_equipment_schedule_auto_column_widths(ws):
    """結果_設備毎の時間割: 見出し文字長に基づく列幅（上限付き）。"""
    mc = ws.max_column or 1
    for ci in range(1, mc + 1):
        cell = ws.cell(row=1, column=ci)
        h = cell.value
        h_str = str(h).strip() if h is not None else ""
        if not h_str:
            wch = 10.0
        else:
            wch = float(min(max(len(h_str.replace("\n", "")) + 2, 9), 28))
        if h_str.endswith("進度"):
            wch = min(wch, 12.0)
        if h_str == "日時帯":
            wch = min(max(wch, 12.0), 16.0)
        try:
            ws.column_dimensions[get_column_letter(ci)].width = wch
        except Exception:
            pass
def _apply_stage2_production_plan_workbook_polish(
    writer_sheets: dict,
    *,
    member_priority_second_header_row: int | None = None,
):
    """
    計画*.xlsx（段階2計画ブック）各シートの共通仕上げ（ガント以外）。

    見出し背景・罫線・窓枠固定（先頭列＋行1）に限る。条件付き書式・ハイパーリンク・
    配台表の Excel テーブル化・設備時間割の列幅調整は、呼び出し側の前段・後段で維持する。
    """
    skip = frozenset(
        {RESULT_SHEET_GANTT_NAME, RESULT_SHEET_GANTT_ACTUAL_DETAIL_NAME}
    )
    extra_mprio: tuple[int, ...] | None = None
    if member_priority_second_header_row is not None and member_priority_second_header_row >= 1:
        extra_mprio = (int(member_priority_second_header_row),)

    for name, ws in writer_sheets.items():
        if name in skip:
            continue
        ex = extra_mprio if name == RESULT_MEMBER_PRIORITY_SHEET_NAME else None
        _apply_stage2_plan_sheet_header_fill(ws, extra_header_rows=ex)
        _apply_stage2_plan_sheet_grid_border(ws)
        try:
            if name not in (
                RESULT_TASK_SHEET_NAME,
                RESULT_DISPATCH_TABLE_SHEET_NAME,
            ):
                ws.freeze_panes = "B2"
        except Exception:
            pass
def _apply_excel_date_columns_date_only_display(path, sheet_name, header_names=None):
    """openpyxl: 指定ヘッダー列を yyyy/mm/dd の日付表示にれる（時刻を表示しない）。"""
    from openpyxl import load_workbook

    headers = header_names or STAGE1_SHEET_DATEONLY_HEADERS
    if _workbook_should_skip_openpyxl_io(path):
        logging.info(
            "日付列表示整形: ブックに「%s」があるため、openpyxl での処理をスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    wb = load_workbook(path)
    try:
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[int(sheet_name)]
        cmap = {}
        for cell in ws[1]:
            if cell.value is not None:
                cmap[str(cell.value).strip()] = cell.column
        fmt = "yyyy/mm/dd"
        for h in headers:
            col = cmap.get(h)
            if not col:
                continue
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                v = c.value
                if v is None:
                    continue
                if isinstance(v, datetime):
                    c.value = v.date()
                elif isinstance(v, date):
                    pass
                else:
                    try:
                        d0 = pd.to_datetime(v, errors="coerce")
                        if pd.isna(d0):
                            continue
                        c.value = d0.date()
                    except Exception:
                        continue
                c.number_format = fmt
        wb.save(path)
    finally:
        wb.close()
def _extract_data_extraction_datetime(sheet_name: str | None = None):
    """
    マクロブックの計画タスクシート（既定は ``TASKS_SHEET_NAME``＝加工計画DATA）から配台基準日時を取得する。
    列「データ抽出時間」の先頭非空値を最優先。次に列「抽出時間」。列が無い・有効値が無いときは「データ抽出日」を試す。

    Args:
        sheet_name: 読むシート名。None または空のとき ``TASKS_SHEET_NAME``。

    Returns:
        tuple[datetime | None, str | None]: (日時, 採用した列名)。両方 None のときは現在時刻フォールバック。
    """

    def _first_valid_dt_from_series(series) -> datetime | None:
        first = None
        for v in series:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            first = v
            break
        if first is None:
            return None
        dt = pd.to_datetime(first, errors="coerce")
        if pd.isna(dt):
            return None
        if isinstance(dt, pd.Timestamp):
            return dt.to_pydatetime()
        return dt if isinstance(dt, datetime) else None

    try:
        _xwb = resolve_data_extraction_workbook_path(_excel_plan_input_wb())
        if not _xwb or not os.path.exists(_xwb):
            return None, None
        sn = (sheet_name or "").strip() or TASKS_SHEET_NAME
        _dt_cols = [
            TASK_COL_DATA_EXTRACTION_TIME,
            TASK_COL_EXTRACTION_TIME,
            TASK_COL_DATA_EXTRACTION_DT,
        ]
        try:
            df = pd.read_excel(_xwb, sheet_name=sn, usecols=_dt_cols)
        except (ValueError, KeyError):
            df = pd.read_excel(_xwb, sheet_name=sn, nrows=512)
        df.columns = df.columns.str.strip()
        for col_name in (
            TASK_COL_DATA_EXTRACTION_TIME,
            TASK_COL_EXTRACTION_TIME,
            TASK_COL_DATA_EXTRACTION_DT,
        ):
            if col_name not in df.columns:
                continue
            dt = _first_valid_dt_from_series(df[col_name])
            if dt is not None:
                return dt, col_name
        return None, None
    except Exception:
        return None, None
def _extract_data_extraction_datetime_for_actual_related_gantt():
    """
    実績系設備ガント用。``TASKS_SHEET_NAME_FOR_ACTUAL_GANTT_PLAN`` を先に読み、
    日時が得られなければ ``TASKS_SHEET_NAME``（加工計画DATA）へフォールバックする。
    """
    dt, col = _extract_data_extraction_datetime(
        sheet_name=TASKS_SHEET_NAME_FOR_ACTUAL_GANTT_PLAN
    )
    if dt is not None:
        return dt, col
    return _extract_data_extraction_datetime()
def _extract_data_extraction_datetime_str():
    """
    `加工計画DATA` から基準日時を文字列化する（データ抽出時間→抽出時間→データ抽出日）。
    """
    try:
        dt, _ = _extract_data_extraction_datetime()
        if dt is None:
            return "—"
        return dt.strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        return "—"
def _parse_equipment_gantt_meta_line_data_extract_display(meta_line) -> str | None:
    """
    設備ガント系シートのメタ行（行2付近）から「データ抽出」の表示文字列を取り出す。
    ``_write_results_equipment_gantt_sheet`` の ``meta_line`` 形式と整合させる。
    """
    if meta_line is None:
        return None
    s = str(meta_line).replace("\r", "").replace("\n", "")
    needle = "　・　データ抽出　"
    pos = s.find(needle)
    if pos < 0:
        return None
    rest = s[pos + len(needle) :]
    end_mark = "　・　マスタ"
    end_pos = rest.find(end_mark)
    if end_pos < 0:
        return None
    out = rest[:end_pos].strip()
    return out if out else None
def _read_existing_equipment_gantt_data_extract_display(path: str, sheet_name: str) -> str | None:
    """
    既存の設備ガント xlsx のメタ行からデータ抽出表示を読む（失敗時は None）。
    メタは ``title_start_col=4`` の結合先頭セル（通常 D2）。
    """
    if not path or not os.path.isfile(path):
        return None
    try:
        wb = load_workbook(path, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            val = ws.cell(row=2, column=4).value
            return _parse_equipment_gantt_meta_line_data_extract_display(val)
        finally:
            wb.close()
    except Exception:
        return None
def _weekday_jp(d):
    return "月睫水木金土日"[d.weekday()]
_GANTT_BAR_FILLS_PRINT_SAFE = (
    "E8E8E8",
    "D8E4EF",
    "E6E2DB",
    "DEEADF",
    "E8E0E8",
    "EAE8D8",
    "DDE6EA",
    "E5DCE5",
)
_GANTT_BAR_FILLS_ACTUAL = (
    "D4E4D4",
    "C9DDE8",
    "DED8CC",
    "D2E5CD",
    "DAD2D9",
    "E0DCCF",
    "CDE2E8",
    "DCD2DC",
)
_GANTT_DAILY_STARTUP_FILL = "FFEB9C"
_GANTT_TIMELINE_CELL_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=False,
    shrink_to_fit=False,
    indent=1,
)
_GANTT_TASK_PATTERN_FILL_BY_HEX: dict[str, PatternFill] = {}
def _paint_gantt_timeline_row_merged(
    ws,
    row,
    n_fixed,
    slots,
    slot_mins,
    evlist,
    idle_fill,
    break_fill,
    gantt_label_font,
    grid_border,
    task_fill_fn=None,
    label_font=None,
    shape_label_specs: list | None = None,
    label_italic: bool = False,
    shape_day_key: str | None = None,
    show_completion_pct_in_label: bool = False,
    *,
    shape_line_dash: bool = False,
    shape_line_weight_override: float | None = None,
):
    """
    時間軸を塗り分けたうえで、同一状態が連続するセルを横結合し帯状のバーにする。
    （細マス単体の塗りではなく slot_mins 刻み＋同一状態のセル結合で、帯状のバーとして表現する）
    shape_label_specs に list を渡すと、タイムライン上の文字はセルに入れず後段（Excel）で
    角丸シェイプとして追加するための座標・文言を蓄積する。
    そのモードでは複数スロットの横結合を避け、スロット列ごとに罫線を引いてタイムラインの格子を揃える。
    shape_day_key に ISO 日付文字列等を渡すと、後段で日単位の画像化（フラット化）に利用する。
    shape_line_dash / shape_line_weight_override は角丸ラベルシェイプの枠線（Excel 描画段階）向け。
    """
    bar_label_font = label_font or gantt_label_font
    _shape_line_opt: dict = {}
    if shape_line_dash:
        _shape_line_opt["line_dash"] = True
    if shape_line_weight_override is not None:
        _shape_line_opt["line_wt"] = float(shape_line_weight_override)
    n_slots = len(slots)
    _chosen = _gantt_best_overlapping_events_for_slots_line_sweep(
        evlist, slots, slot_mins
    )
    states = []
    for slot_start, active in zip(slots, _chosen):
        states.append(
            _gantt_slot_state_tuple_from_active(
                active, slot_start, slot_mins, task_fill_fn
            )
        )
    tcol0 = n_fixed + 1
    i = 0
    while i < n_slots:
        st0 = states[i]
        j = i + 1
        while j < n_slots and _gantt_timeline_same_segment(st0, states[j]):
            j += 1
        col_s = tcol0 + i
        col_e = tcol0 + j - 1
        single_slot_segment = col_s == col_e
        # 角丸シェイプで依頼NO を載せるモードでは、横結合だとスロット境界の縦罫線が欠けて見栄えが悪い。
        # 同一セグメント内はスロット列ごとにセルを分け、各セルに grid_border を付ける（シェイプ幅は col_s:col_e のまま）。
        _shape_slot_grid = shape_label_specs is not None and col_e > col_s
        if not _shape_slot_grid:
            # 同一スタイルの連続列は結合し先頭セルのみ openpyxl へ書く（セル単位ループが H2 ボトルネック）
            if col_e > col_s:
                ws.merge_cells(
                    start_row=row, start_column=col_s, end_row=row, end_column=col_e
                )
        c = ws.cell(row=row, column=col_s)
        if _shape_slot_grid:
            for _col_k in range(col_s, col_e + 1):
                _ck = ws.cell(row=row, column=_col_k)
                _ck.border = grid_border
                _ck.alignment = _GANTT_TIMELINE_CELL_ALIGNMENT
                if st0[0] == "idle":
                    _ck.fill = idle_fill
                    _ck.value = None
                elif st0[0] == "break":
                    _ck.fill = break_fill
                    _ck.value = None
                elif st0[0] == "daily_startup":
                    _, _gh_ds = st0
                    _ck.fill = _gantt_cached_pattern_fill(_gh_ds)
                    _ck.value = None
                else:
                    _, _tid, _gh, _sl0, _pc0 = st0
                    _ck.fill = _gantt_cached_pattern_fill(_gh)
                    _ck.value = None
        else:
            c.border = grid_border
            c.alignment = _GANTT_TIMELINE_CELL_ALIGNMENT
            if st0[0] == "idle":
                c.fill = idle_fill
                c.value = None
            elif st0[0] == "break":
                c.fill = break_fill
                c.value = None
            elif st0[0] == "daily_startup":
                _, gh_ds = st0
                c.fill = _gantt_cached_pattern_fill(gh_ds)
                _ds_txt = "日次始業準備"
                if shape_label_specs is not None:
                    _seg_lo = slots[i]
                    _seg_hi = slots[j - 1] + timedelta(minutes=float(slot_mins))
                    _mem_ds = _gantt_member_labels_for_startup_in_range(
                        evlist, _seg_lo, _seg_hi
                    )
                    _shape_text = _ds_txt
                    _mem1l: list[str] = []
                    if _mem_ds:
                        for _x in _mem_ds[:8]:
                            _t = (
                                str(_x)
                                .replace("\r", "")
                                .replace("\n", "")
                                .strip()
                            )
                            if _t:
                                _mem1l.append(_t)
                    shape_label_specs.append(
                        {
                            "row": row,
                            "col_s": col_s,
                            "col_e": col_e,
                            "text": _shape_text,
                            "italic": bool(label_italic),
                            "fill_hex": str(gh_ds),
                            "member_labels": list(_mem1l),
                            "member_chip_below": bool(_mem1l),
                            "day_key": shape_day_key or "",
                            **_shape_line_opt,
                        }
                    )
                    c.value = None
                else:
                    c.value = _ds_txt
                    c.font = bar_label_font
                    c.alignment = _gantt_timeline_label_alignment(
                        single_slot=single_slot_segment
                    )
            else:
                _, tid, gh, _slot_len_m0, _pct0 = st0
                c.fill = _gantt_cached_pattern_fill(gh)
                tid_s = str(tid or "").strip()
                _seg_lo = slots[i]
                _seg_hi = slots[j - 1] + timedelta(minutes=float(slot_mins))
                _tot_len, _n_ev, _pct_seg = _gantt_segment_total_length_m(
                    evlist, tid_s, _seg_lo, _seg_hi
                )
                _len_s = _gantt_format_length_m(_tot_len)
                _lbl = f"{tid_s} {_len_s}m" if (_len_s and tid_s) else tid_s
                if show_completion_pct_in_label and _pct_seg is not None and tid_s:
                    _lbl = f"{_lbl} {_pct_seg}%"
                if shape_label_specs is not None:
                    if tid_s:
                        shape_label_specs.append(
                            {
                                "row": row,
                                "col_s": col_s,
                                "col_e": col_e,
                                "text": _lbl,
                                "italic": bool(label_italic),
                                "fill_hex": str(gh),
                                "member_labels": _gantt_member_labels_for_task(
                                    evlist, tid_s
                                ),
                                "day_key": shape_day_key or "",
                                **_shape_line_opt,
                            }
                        )
                    c.value = None
                else:
                    c.value = _lbl
                    c.font = bar_label_font
                    c.alignment = _gantt_timeline_label_alignment(
                        single_slot=single_slot_segment
                    )
        if _shape_slot_grid:
            if st0[0] == "daily_startup":
                _, gh_ds = st0
                _ds_txt = "日次始業準備"
                _seg_lo = slots[i]
                _seg_hi = slots[j - 1] + timedelta(minutes=float(slot_mins))
                _mem_ds = _gantt_member_labels_for_startup_in_range(
                    evlist, _seg_lo, _seg_hi
                )
                _shape_text = _ds_txt
                _mem1l2: list[str] = []
                if _mem_ds:
                    for _x in _mem_ds[:8]:
                        _t = (
                            str(_x)
                            .replace("\r", "")
                            .replace("\n", "")
                            .strip()
                        )
                        if _t:
                            _mem1l2.append(_t)
                shape_label_specs.append(
                    {
                        "row": row,
                        "col_s": col_s,
                        "col_e": col_e,
                        "text": _shape_text,
                        "italic": bool(label_italic),
                        "fill_hex": str(gh_ds),
                        "member_labels": list(_mem1l2),
                        "member_chip_below": bool(_mem1l2),
                        "day_key": shape_day_key or "",
                        **_shape_line_opt,
                    }
                )
            elif st0[0] not in ("idle", "break"):
                _, tid, gh, _slot_len_m0, _pct0 = st0
                tid_s = str(tid or "").strip()
                _seg_lo = slots[i]
                _seg_hi = slots[j - 1] + timedelta(minutes=float(slot_mins))
                _tot_len, _n_ev, _pct_seg = _gantt_segment_total_length_m(
                    evlist, tid_s, _seg_lo, _seg_hi
                )
                _len_s = _gantt_format_length_m(_tot_len)
                _lbl = f"{tid_s} {_len_s}m" if (_len_s and tid_s) else tid_s
                if show_completion_pct_in_label and _pct_seg is not None and tid_s:
                    _lbl = f"{_lbl} {_pct_seg}%"
                if tid_s:
                    shape_label_specs.append(
                        {
                            "row": row,
                            "col_s": col_s,
                            "col_e": col_e,
                            "text": _lbl,
                            "italic": bool(label_italic),
                            "fill_hex": str(gh),
                            "member_labels": _gantt_member_labels_for_task(
                                evlist, tid_s
                            ),
                            "day_key": shape_day_key or "",
                            **_shape_line_opt,
                        }
                    )
        i = j
def _time_intervals_overlap_half_open(
    a_start: time, a_end: time, b_start: time, b_end: time
) -> bool:
    """半開区間 [a_start, a_end) と [b_start, b_end) は重なるか（同一日内）。"""

    def _sec(t: time) -> int:
        return t.hour * 3600 + t.minute * 60 + t.second

    return _sec(a_start) < _sec(b_end) and _sec(a_end) > _sec(b_start)
def _parse_equipment_schedule_time_band_cell(val) -> tuple[time | None, time | None]:
    """結果_設備毎の時間割「日時帯」セル（例 08:45-09:00）を解釈。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None, None
    s = str(val).strip()
    if not s or "■" in s:
        return None, None
    for sep in ("-", "＝", "~", "〜"):
        if sep in s:
            left, right = s.split(sep, 1)
            left = left.strip().replace("：", ":")
            right = right.strip().replace("：", ":")
            t0 = parse_time_str(left, None)
            t1 = parse_time_str(right, None)
            if t0 is not None and t1 is not None and t0 < t1:
                return t0, t1
            return None, None
    return None, None
def _apply_equipment_schedule_outside_regular_fill(
    ws, reg_start: time, reg_end: time
) -> None:
    """「日時帯」列で定常 [reg_start, reg_end) と重ならない行のセルに着色。"""
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_OUTSIDE_REGULAR_TIME_FILL,
        end_color=RESULT_OUTSIDE_REGULAR_TIME_FILL,
    )
    col_idx = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_idx = i
            break
    if col_idx is None:
        return
    mr = ws.max_row or 1
    for r in range(2, mr + 1):
        cell = ws.cell(row=r, column=col_idx)
        t0, t1 = _parse_equipment_schedule_time_band_cell(cell.value)
        if t0 is None or t1 is None:
            continue
        if not _time_intervals_overlap_half_open(t0, t1, reg_start, reg_end):
            cell.fill = fill
def _apply_equipment_schedule_prep_cleanup_fill(ws) -> None:
    """
    設備列（進度列を除く）で、表示に「日次始業準備」が含まれるセルを薄緑にする。
    結果_設備毎の時間割 の equip セル用（日時帯列は変更しない）。
    """
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_DISPATCHED_REQUEST_FILL,
        end_color=RESULT_DISPATCHED_REQUEST_FILL,
    )
    markers = (
        "(日次始業準備)",
        "日次始業準備",
        "(依頼切替準備)",
        "依頼切替準備",
        "(休憩再開準備)",
        "休憩再開準備",
        "(後始末)",
        "後始末",
        "(依頼間余裕)",
        "依頼間余裕",
    )
    col_tb = None
    equip_cols: list[int] = []
    for i, c in enumerate(ws[1], start=1):
        if c.value is None:
            continue
        h = str(c.value).strip()
        if h == "日時帯":
            col_tb = i
            continue
        if h.endswith("進度"):
            continue
        equip_cols.append(i)
    if col_tb is None or not equip_cols:
        return
    mr = ws.max_row or 1
    for r in range(2, mr + 1):
        for ci in equip_cols:
            cell = ws.cell(row=r, column=ci)
            val = cell.value
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip().replace("\r", "").replace("\n", "")
            if any(m in s for m in markers):
                cell.fill = fill
def _parse_equipment_schedule_day_header_date(val) -> date | None:
    """日付見出し行「■ YYYY/MM/DD … ■」から日付を得る。"""
    if val is None:
        return None
    s = str(val).strip()
    m = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None
def _machine_calendar_intervals_for_equipment_line(
    day_blocks: dict[str, list[tuple[datetime, datetime]]],
    eq_line: str,
    day_d: date,
) -> list[tuple[datetime, datetime]]:
    """当日・当該設備列キーに対応れる機械カレンダー占有区間（工場稼働枠でクリップ済み）。"""
    if not day_blocks:
        return []
    ek = str(eq_line or "").strip()
    blocks: list[tuple[datetime, datetime]] | None = None
    if ek in day_blocks:
        blocks = day_blocks[ek]
    else:
        pk = (
            _normalize_equipment_match_key(ek.split("+", 1)[1])
            if "+" in ek
            else _normalize_equipment_match_key(ek)
        )
        if pk and pk in day_blocks:
            blocks = day_blocks[pk]
        else:
            nk = _normalize_equipment_match_key(ek)
            for k, iv in day_blocks.items():
                if _normalize_equipment_match_key(str(k)) == nk:
                    blocks = iv
                    break
    if not blocks:
        return []
    w0 = datetime.combine(day_d, DEFAULT_START_TIME)
    w1 = datetime.combine(day_d, DEFAULT_END_TIME)
    return _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
def _apply_equipment_schedule_machine_calendar_fill(
    ws,
    equipment_list: list,
    calendar_blocks_by_date: dict[date, dict[str, list[tuple[datetime, datetime]]]],
) -> None:
    """
    結果_設備毎の時間割: 機械カレンダー占有と重なる設備セル（進度列以外）を紫色で塗る。
    10 分枠の半開区間 [slot_start, slot_end) と占有 [bs, be) は重ならない対象。
    """
    if not calendar_blocks_by_date or not equipment_list:
        return
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_MACHINE_CALENDAR_BLOCK_FILL,
        end_color=RESULT_MACHINE_CALENDAR_BLOCK_FILL,
    )
    col_tb = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_tb = i
            break
    if col_tb is None:
        return
    eq_col_indices: list[int] = [
        col_tb + 1 + 2 * idx for idx in range(len(equipment_list))
    ]
    mr = ws.max_row or 1
    current_d: date | None = None
    for r in range(2, mr + 1):
        tb_cell = ws.cell(row=r, column=col_tb)
        tv = tb_cell.value
        d_hdr = _parse_equipment_schedule_day_header_date(tv)
        if d_hdr is not None:
            current_d = d_hdr
            continue
        t0, t1 = _parse_equipment_schedule_time_band_cell(tv)
        if t0 is None or t1 is None or current_d is None:
            continue
        slot_a = datetime.combine(current_d, t0)
        slot_b = datetime.combine(current_d, t1)
        if slot_b <= slot_a:
            continue
        day_blocks = calendar_blocks_by_date.get(current_d)
        if not day_blocks:
            continue
        for col_idx, eq_line in zip(eq_col_indices, equipment_list):
            blocks_c = _machine_calendar_intervals_for_equipment_line(
                day_blocks, eq_line, current_d
            )
            if not blocks_c:
                continue
            for bs, be in blocks_c:
                if slot_a < be and bs < slot_b:
                    ws.cell(row=r, column=col_idx).fill = fill
                    break
def _apply_equipment_by_machine_dispatched_request_fill(ws) -> None:
    """
    結果_設備毎の時間割_機械名毎の機械名列で」依頼NOは入っているセルに薄緑を付与れる。
    「（休憩）」のみのセルは対象外。見出し行・日時帯列は変更しない。
    """
    fill = PatternFill(
        fill_type="solid",
        start_color=RESULT_DISPATCHED_REQUEST_FILL,
        end_color=RESULT_DISPATCHED_REQUEST_FILL,
    )
    col_tb = None
    for i, c in enumerate(ws[1], start=1):
        if c.value is not None and str(c.value).strip() == "日時帯":
            col_tb = i
            break
    if col_tb is None:
        return
    mr = ws.max_row or 1
    mc = ws.max_column or col_tb
    for r in range(2, mr + 1):
        for c in range(col_tb + 1, mc + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip().replace("\r", "").replace("\n", "")
            if not s or s == "（休憩）":
                continue
            cell.fill = fill
def _equipment_gantt_fills_by_machine_name(equipment_list) -> dict[str, PatternFill]:
    """
    結果_設備ガントの固定列（B〜D」A は日付縦結合）用。equipment_list 内の機械名（+ 無し時は行全体を機械名）の出睾順で
    淡色を割り当てて、同一機械名は常に同じ PatternFill を共有する。
    """
    order: list[str] = []
    seen: set[str] = set()
    for eq in equipment_list or []:
        _, mn = _split_equipment_line_process_machine(eq)
        key = (mn or "").strip() or "—"
        if key not in seen:
            seen.add(key)
            order.append(key)
    palette = (
        RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS_FULL
        if _gantt_color_mode_full()
        else RESULT_EQUIP_GANTT_MACHINE_GROUP_FILL_COLORS
    )
    if not palette:
        fb = "F5F5F5"
        return {k: PatternFill(fill_type="solid", start_color=fb, end_color=fb) for k in order}
    out: dict[str, PatternFill] = {}
    n = len(palette)
    for i, key in enumerate(order):
        hx = palette[i % n]
        out[key] = PatternFill(fill_type="solid", start_color=hx, end_color=hx)
    return out
def _apply_compare_gantt_typography(ws, hdr_row: int) -> None:
    """計画実績比較シートのフォントを結果ブック既定に統一し、表示日・説明・左表のサイズを調整する。"""
    fn = _effective_result_book_font_name()
    mr = int(ws.max_row or 1)
    mc = int(ws.max_column or 1)

    def _repl(old, **kw):
        sz = kw.pop("size", None)
        bd = kw.pop("bold", None)
        it = kw.pop("italic", None)
        col = kw.pop("color", None)
        if old is not None:
            if sz is None and old.size:
                sz = old.size
            if bd is None and old.bold is not None:
                bd = old.bold
            if it is None and old.italic is not None:
                it = old.italic
            if col is None and old.color:
                col = old.color
        sz = float(sz) if sz is not None else 11.0
        bd = False if bd is None else bool(bd)
        it = False if it is None else bool(it)
        args: dict = {"name": fn, "size": sz, "bold": bd, "italic": it}
        if col is not None:
            args["color"] = col
        args.update(kw)
        return Font(**args)

    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _repl(cell.font)
    ws["B1"].font = _repl(ws["B1"].font, size=22, bold=False)
    ws["A1"].font = _repl(ws["A1"].font, size=11, bold=True)
    ws["A2"].font = _repl(ws["A2"].font, size=12, bold=False)
    ws["D2"].font = _repl(ws["D2"].font, size=14, bold=False)
    for r in range(hdr_row, mr + 1):
        for c in (2, 3):
            cell = ws.cell(row=r, column=c)
            cell.font = _repl(cell.font, size=14)

    ws.row_dimensions[1].height = 44
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[hdr_row].height = 44
ENV_STAGE2_IN_PROGRESS_NEXT_DAY_DISPATCH_JSON = (
    "PM_AI_STAGE2_IN_PROGRESS_NEXT_DAY_DISPATCH_JSON"
)
ENV_STAGE2_ALADDIN_TODAY_EXCLUDE_NEXT_DAY_JSON = (
    "PM_AI_STAGE2_ALADDIN_TODAY_EXCLUDE_NEXT_DAY_JSON"
)
EXCLUDE_RULES_SHEET_COM_SYNC_MAX_COL = 5
EXCLUDE_RULES_MATRIX_CLIP_MAX_COL = 5
ENV_OVERTIME_SIMULATION_JSON = "PM_AI_OVERTIME_SIMULATION_JSON"
ENV_STAGE35_STAGE3_METERS_FLOOR_JSON = "PM_AI_STAGE35_STAGE3_METERS_FLOOR_JSON"
DISPATCH_INTERVAL_MIRROR_ENFORCE = os.environ.get(
    "DISPATCH_INTERVAL_MIRROR_ENFORCE", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")
