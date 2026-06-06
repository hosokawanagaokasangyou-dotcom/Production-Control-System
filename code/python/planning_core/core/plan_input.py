# -*- coding: utf-8 -*-
# planning_core.core.plan_input — body only (loaded via _core exec chain)
def load_planning_tasks_df():
    """
    2段階目用: 環境変数 ``PM_AI_PLAN_INPUT_PATH`` の表（CSV / Parquet / xlsx）を読み込む。

    「担当OP_指定」列または特別指定備考の AI 出力 preferred_operator で主担当 OP を指名できる（skills のメンバー名とあいまい一致）。
    メイン「再優先特別記載」の task_preferred_operators は generate_plan 側で最優先マージされる。
    「配台不要」がオン（TRUE/1/はい 等）の行は配台対象外（**シート上の列の値をそのまま**解釈する）。
    読み込み後、同一依頼NO・重複機械名があるグループの工程「分割」行へ空なら「配台不要」=yes（段階1と同じ）。
    「設定_配台不要工程」シートの**行同期・保守**（``run_exclude_rules_sheet_maintenance``）は、
    ``PM_AI_PLAN_INPUT_PATH`` がブック（xlsx/xlsm）かつ ``PM_AI_EXCLUDE_RULES_JSON`` が無効なときのみ
    対象ブックで行う。JSON 正本のときは ``_merge_exclude_rules_json_with_plan_pairs`` で JSON を更新する。
    D→E の **AI 補完は行わない**（段階1のみ）。C/E に基づく計画シートへの配台不要の**再適用**
    （``apply_exclude_rules_config_to_plan_df``）も行わない（段階1のみ）。

    ``PM_AI_PLAN_INPUT_PATH`` は **必須**（未設定・不存在・TASK_INPUT_WORKBOOK へのフォールバックなし）。
    """
    _plan_alt = (os.environ.get(ENV_PLAN_INPUT_PATH) or "").strip()
    if not _plan_alt:
        raise FileNotFoundError(
            "段階2: 計画タスク入力が解決できません。"
            f" 環境変数 {ENV_PLAN_INPUT_PATH} に実在する CSV / Parquet / xlsx のパスを設定してください。"
        )
    if not os.path.isfile(_plan_alt):
        raise FileNotFoundError(
            f"段階2: {ENV_PLAN_INPUT_PATH} が実在しません: {_plan_alt!r}。"
        )
    low = _plan_alt.lower()
    if low.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        normalize_ooxml_shared_strings_if_missing(_plan_alt)
    _wb_for_maint = ""
    if low.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        _wb_for_maint = os.path.normpath(os.path.abspath(_plan_alt))
    if low.endswith((".csv", ".parquet", ".pq")):
        df = read_tabular_dataframe(_plan_alt)
    else:
        df = read_tabular_dataframe(_plan_alt, sheet_name=PLAN_INPUT_SHEET_NAME)
    df.columns = df.columns.str.strip()
    df = _migrate_deprecated_plan_override_columns(df)
    df = _align_dataframe_headers_to_canonical(
        df, plan_input_sheet_column_order()
    )
    _ensure_dataframe_has_unprocessed_column(
        df, context_label=f"シート「{PLAN_INPUT_SHEET_NAME}」"
    )
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""
    df = _coalesce_plan_plain_remark_into_special(df)
    _apply_planning_sheet_post_load_mutations(
        df,
        _wb_for_maint,
        "配台シート読込",
        apply_exclude_rules_from_config=False,
        compile_exclude_rules_d_to_e_with_ai=False,
    )
    _apply_master_speed_sheet_to_plan_df(df, log_prefix="配台シート読込")
    try:
        from .actual_speed_apply import apply_learned_speed_to_plan_df

        apply_learned_speed_to_plan_df(df, log_prefix="配台シート読込")
    except Exception as ex:
        logging.warning("配台シート読込: 学習速度適用をスキップ（%s）", ex)
    _fill_plan_dispatch_remaining_qty_column(df)
    logging.info("計画タスク入力: PM_AI_PLAN_INPUT_PATH='%s' を読み込みました。", _plan_alt)
    return df
def _main_sheet_cell_is_global_comment_label(val) -> bool:
    """メインシート上「グローバルコメント」見出しセルか（表記ゆれ許容）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    s = unicodedata.normalize("NFKC", str(val).strip())
    if not s:
        return False
    if _nfkc_column_aliases(s) == _nfkc_column_aliases("グローバルコメント"):
        return True
    if "グローバル" in s and "コメント" in s:
        return True
    return False
def load_main_sheet_global_priority_override_text() -> str:
    """
    優先: 環境変数 PM_AI_GLOBAL_PRIORITY_OVERRIDE_PATH の UTF-8 テキスト（1 ファイル＝コメント本文）。

    従来: TASK_INPUT_WORKBOOK のメインシートで「グローバルコメント」見出しセルの **直下**。
    シート名: 「メイン」「メイン_」「Main」等（VBA GetMainWorksheet と同趣旨）。

    内容は **Gemini で一括解釈**（`analyze_global_priority_override_comment`）。工場休業日等は
    `parse_factory_closure_dates_from_global_comment` で補完しうる。
    """
    global _STAGE2_GLOBAL_COMMENT_CACHE
    _txt = (os.environ.get(ENV_GLOBAL_PRIORITY_OVERRIDE_PATH) or "").strip()
    if _txt and os.path.isfile(_txt):
        try:
            st = os.stat(_txt)
            sig = (os.path.abspath(_txt), int(st.st_mtime), int(st.st_size))
            if (
                isinstance(_STAGE2_GLOBAL_COMMENT_CACHE, dict)
                and _STAGE2_GLOBAL_COMMENT_CACHE.get("sig") == sig
            ):
                return str(_STAGE2_GLOBAL_COMMENT_CACHE.get("value") or "")
            with open(_txt, encoding="utf-8-sig") as f:
                out = f.read().strip()
            _STAGE2_GLOBAL_COMMENT_CACHE = {"sig": sig, "value": out}
            return out
        except OSError as e:
            logging.warning("メイン再優先特記: テキストを読めません: %s", e)
            return ""
        except Exception as e:
            logging.warning("メイン再優先特記: テキスト処理で例外: %s", e)
            return ""

    wb_path = _excel_plan_input_wb().strip() if _excel_plan_input_wb() else ""
    if not wb_path or not os.path.exists(wb_path):
        return ""
    # region stage2 cache
    # VBA から起動されると段階2が複数回（パターン別）実行されうるため、
    # openpyxl でのブックオープンを毎回やらない（mtime 変化時のみ更新）。
    try:
        st = os.stat(wb_path)
        sig = (os.path.abspath(wb_path), int(st.st_mtime), int(st.st_size))
        if (
            isinstance(_STAGE2_GLOBAL_COMMENT_CACHE, dict)
            and _STAGE2_GLOBAL_COMMENT_CACHE.get("sig") == sig
        ):
            return str(_STAGE2_GLOBAL_COMMENT_CACHE.get("value") or "")
    except Exception:
        pass
    # endregion stage2 cache
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "メイン再優先特記: ブックに「%s」があるため、openpyxl でグローバルコメントを読みません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return ""
    try:
        _sn = _ooxml_workbook_sheet_names(wb_path)
        if _sn is not None:
            _has_main = any(
                n in ("メイン", "メイン_", "Main") or ("メイン" in str(n)) for n in _sn
            )
            if not _has_main:
                return ""
    except Exception:
        pass
    if _ooxml_workbook_missing_shared_strings(wb_path):
        logging.info(
            "メイン再優先特記: OOXML に xl/sharedStrings.xml が無いブックのため、"
            "openpyxl でグローバルコメントを読みません（メモリ急増回避）。"
            " PM_AI_GLOBAL_PRIORITY_OVERRIDE_PATH のテキスト、または Excel で通常保存したブックを使用してください。"
        )
        return ""
    wb = None
    try:
        # read_only=True でオープン高速化（読み取りのみ）
        wb = load_workbook(wb_path, data_only=True, read_only=True)
        ws = None
        for name in ("メイン", "メイン_", "Main"):
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            for sn in wb.sheetnames:
                if "メイン" in sn:
                    ws = wb[sn]
                    break
        if ws is None:
            return ""
        max_r = min(ws.max_row or 0, 400)
        max_c = min(ws.max_column or 0, 40)
        if max_r < 1 or max_c < 1:
            return ""
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if not _main_sheet_cell_is_global_comment_label(cell.value):
                    continue
                below = ws.cell(row=r + 1, column=c).value
                if below is None or (isinstance(below, float) and pd.isna(below)):
                    out = ""
                    # region stage2 cache
                    try:
                        _STAGE2_GLOBAL_COMMENT_CACHE = {"sig": sig, "value": out}
                    except Exception:
                        pass
                    # endregion stage2 cache
                    return out
                out = str(below).strip()
                # region stage2 cache
                try:
                    _STAGE2_GLOBAL_COMMENT_CACHE = {"sig": sig, "value": out}
                except Exception:
                    pass
                # endregion stage2 cache
                return out
        return ""
    except Exception as e:
        logging.warning("メイン再優先特記: ブックを開きませんでした: %s", e)
        return ""
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def _global_comment_chunk_implies_factory_closure(chunk: str) -> bool:
    """
    メイン「グローバルコメント」の断片は」工場短縮の休業・非稼働を愝味れるか（個人休みの値を誤検出しない）。
    """
    c = unicodedata.normalize("NFKC", str(chunk or ""))
    if not c.strip():
        return False
    if re.search(r"臨時\s*休業", c):
        return True
    if "休場" in c:
        return True
    if re.search(r"工場", c) and re.search(r"休|休業|休み|坜止|よ休み", c):
        return True
    if re.search(r"(?:全社|全館|全工場).{0,15}(?:休|休業|坜止)", c):
        return True
    if re.search(r"(?:稼働|生産|ライン).{0,12}(?:坜止|なし|無し)", c):
        return True
    if re.search(r"加工.{0,15}(?:しない|無し|なし|よ休み)", c):
        return True
    if "休業" in c and re.search(
        r"(?:工場|全社|本社|当日|弊社|当社|全員|社全体)", c
    ):
        return True
    return False
def _md_slash_is_likely_fraction_not_date(t: str, start: int, end: int, mo: int, day: int) -> bool:
    """
    「加工速度は1/3とした」の 1/3 を 1月3日 と誤誝しない。
    「4/1は工場を休み」の 4/1 は日付のまま（直後は「は」なら分数扱いにしない）。
    """
    if mo <= 0 or day <= 0:
        return True
    before = t[max(0, start - 32) : start]
    after = t[end : min(len(t), end + 14)]
    after_st = after.lstrip()
    if after_st.startswith("は"):
        return False
    if re.search(
        r"(?:加工速度|加工\s*スピード|速度|倍率|スピード|効率|割引)(?:\s*は)?\s*$",
        before,
    ):
        return True
    # 1/2・1/3・2/3 等 + 「とした」「倝」… は分数・比率寄り（「3/1です」等の日付を誤スキップしないよご です/である は含まない）
    frac_pat = re.compile(
        r"^(?:としした?|とれる|倝|割引|にれる|に設定|しらい|程度|に固定|に変更)"
    )
    if mo <= 12 and day <= 12 and frac_pat.match(after_st):
        if mo <= 2 or (mo == 3 and day <= 3):
            return True
    # 「1/2です」「1/10です」のよごな分毝表睾（先頭は 1/ のみ）
    if (
        mo == 1
        and 2 <= day <= 12
        and re.match(r"^です|である\b", after_st)
    ):
        return True
    return False
def _extract_calendar_dates_from_text(s: str, default_year: int) -> list[date]:
    """グローバルコメント内の日付表記を date に変杛（基準年は計画の基準年）。"""
    t = unicodedata.normalize("NFKC", str(s or ""))
    found: list[date] = []
    seen: set[date] = set()

    def add(y: int, mo: int, d: int) -> None:
        try:
            dd = date(y, mo, d)
        except ValueError:
            return
        if dd not in seen:
            seen.add(dd)
            found.append(dd)

    for m in re.finditer(
        r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?",
        t,
    ):
        add(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for m in re.finditer(
        r"(\d{4})\s*[/\-\.＝]\s*(\d{1,2})\s*[/\-\.＝]\s*(\d{1,2})",
        t,
    ):
        add(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for m in re.finditer(r"(\d{1,2})\s*月\s*(\d{1,2})\s*日", t):
        add(int(default_year), int(m.group(1)), int(m.group(2)))
    for m in re.finditer(
        r"(?<!\d)(\d{1,2})\s*[/＝]\s*(\d{1,2})(?!\d)",
        t,
    ):
        mo_i, d_i = int(m.group(1)), int(m.group(2))
        if _md_slash_is_likely_fraction_not_date(t, m.start(), m.end(), mo_i, d_i):
            continue
        add(int(default_year), mo_i, d_i)
    return found
def _split_global_comment_into_chunks(blob: str) -> list[str]:
    """
    グローバルコメントを「独立した指示」の塊に分ける。
    改行（Excel の Alt+Enter・Unicode 改行含む）で必う分割し、同一行内は 。;； で続けて分割。
    """
    t = unicodedata.normalize("NFKC", str(blob or "").strip())
    if not t:
        return []
    lines = [ln.strip() for ln in re.split(r"[\n\r\v\f\u2028\u2029]+", t) if ln.strip()]
    if not lines:
        return []
    chunks: list[str] = []
    for line in lines:
        subs = [c.strip() for c in re.split(r"[。;；]+", line) if c.strip()]
        if subs:
            chunks.extend(subs)
        else:
            chunks.append(line)
    return chunks
def parse_factory_closure_dates_from_global_comment(
    text: str, default_year: int
) -> set[date]:
    """
    メインシート「グローバルコメント」に」工場臨時休業などと日付は書かれでいる場合に
    しの日を工場休み（全員非稼働・配台で加工している）として扱ご日付集合を返す。
    """
    blob = unicodedata.normalize("NFKC", str(text or "").strip())
    if not blob:
        return set()
    chunks = _split_global_comment_into_chunks(blob)
    if not chunks:
        chunks = [blob]
    out: set[date] = set()
    y0 = int(default_year)
    for ch in chunks:
        if not _global_comment_chunk_implies_factory_closure(ch):
            continue
        for d in _extract_calendar_dates_from_text(ch, y0):
            out.add(d)
    if not out and _global_comment_chunk_implies_factory_closure(blob):
        for d in _extract_calendar_dates_from_text(blob, y0):
            out.add(d)
    return out
def apply_factory_closure_dates_to_attendance(
    attendance_data: dict, members: list, closure_dates: set[date]
) -> None:
    """工場休業日: 勤怠上は全員 is_working=False とし、しの日は設備割付を行ゝない。"""
    if not closure_dates or not attendance_data:
        return
    tag = "工場休業（メイン・グローバルコメント）"
    for d in sorted(closure_dates):
        if d not in attendance_data:
            logging.warning(
                "グローバルコメントの工場休業日 %s はマスタ勤怠に行はありません。"
                " しの日は計画ループに含まれない場合」配台上の効果は限定的です。",
                d,
            )
            continue
        day = attendance_data[d]
        for m in members:
            if m not in day:
                continue
            ent = day[m]
            ent["is_working"] = False
            ent["eligible_for_assignment"] = False
            prev = str(ent.get("reason") or "").strip()
            ent["reason"] = f"{tag} {prev}".strip() if prev else tag
def _apply_global_priority_abolish_heuristic(blob: str, coerced: dict) -> dict:
    """
    「制限撤廃」「あらゆる条件」等: 設備専有・時刻ガードまで含む配台制約を緩める（abolish_all_scheduling_limits）。
    """
    b = unicodedata.normalize("NFKC", str(blob or ""))
    strong = (
        "制限撤廃",
        "制限を撤廃",
        "まとめての制限",
        "全での制限",
        "あらゆる制限",
        "あらゆる条件",
        "まとめての条件",
        "全での条件",
        "撤廃して",
        "撤廃し",
    )
    if any(k in b for k in strong):
        out = dict(coerced)
        out["abolish_all_scheduling_limits"] = True
        out["ignore_skill_requirements"] = True
        out["ignore_need_minimum"] = True
        logging.warning(
            "メイン再優先特記: 制限撤廃キーワードを検出。設備専有・時刻ガードを含む配台上の制約を緩めた。"
        )
        return out
    return coerced
def _maybe_fill_global_speed_rules_from_scheduler_notes(coerced: dict) -> dict:
    """
    AI は global_speed_rules を空にしたは scheduler_notes に具体パターンはある場合の補完。
    広し推測しない（熱融着＋検査＋1/3 系のみ）。
    """
    if not isinstance(coerced, dict):
        return coerced
    if coerced.get("global_speed_rules"):
        return coerced
    sn = str(coerced.get("scheduler_notes_ja") or "")
    t = unicodedata.normalize("NFKC", sn)
    if "熱融着" not in t or "検査" not in t:
        return coerced
    if not re.search(r"(?:1\s*/\s*3|１\s*/\s*3|三分の一|3\s*分の\s*1)", t):
        return coerced
    out = dict(coerced)
    out["global_speed_rules"] = [
        {
            "process_contains": "熱融着",
            "machine_contains": "検査",
            "speed_multiplier": 1.0 / 3.0,
        }
    ]
    logging.info(
        "メイン再優先特記: scheduler_notes_ja から global_speed_rules を補完（熱融着・検査・1/3）"
    )
    return out
def _finalize_global_priority_override(blob: str, coerced: dict) -> dict:
    """ソロ補正の後」abolish は true ならスキル・人数も強制オン。"""
    coerced = _maybe_fill_global_speed_rules_from_scheduler_notes(dict(coerced))
    coerced = _apply_global_priority_solo_heuristic(blob, coerced)
    coerced = _apply_global_priority_abolish_heuristic(blob, coerced)
    if coerced.get("abolish_all_scheduling_limits"):
        out = dict(coerced)
        out["ignore_skill_requirements"] = True
        out["ignore_need_minimum"] = True
        return out
    return coerced
def _apply_global_priority_solo_heuristic(blob: str, coerced: dict) -> dict:
    """
    「一人で担当」「独立」等で人数の値緩んでも」指定メンバーはスキル非該当てと配台されない。
    しの場合はスキル無視を同時に立でる。
    """
    if not coerced.get("ignore_need_minimum") or coerced.get("ignore_skill_requirements"):
        return coerced
    b = unicodedata.normalize("NFKC", str(blob or ""))
    solo_kw = ("一人", "参とり", "独立", "１人", "1人", "独自", "坘身")
    if any(k in b for k in solo_kw):
        out = dict(coerced)
        out["ignore_skill_requirements"] = True
        logging.info(
            "メイン再優先特記: 独立系キーワードのため、 ignore_skill_requirements を補助的に true にしました。"
        )
        return out
    return coerced
def _coerce_task_preferred_operators_dict(raw_val) -> dict:
    """AI の task_preferred_operators を {依頼NO: 並び} に正規化。"""
    out = {}
    if not isinstance(raw_val, dict):
        return out
    for k, v in raw_val.items():
        ks = str(k).strip()
        if not ks:
            continue
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        vs = str(v).strip()
        if vs and vs.lower() not in ("nan", "none", "null"):
            out[ks] = vs
    return out
def _normalize_factory_closure_dates_iso_list(val, default_year: int) -> list[str]:
    """
    AI またはフォールバックの日付リストを YYYY-MM-DD 文字列の昇順ユニークに正規化。
    覝素は ISO 文字列・Excel 日付・「4/1」程度の短文でも坯。
    """
    y0 = int(default_year)
    seen: set[str] = set()
    out: list[str] = []
    if not isinstance(val, list):
        return out
    for item in val:
        if item is None or (isinstance(item, float) and pd.isna(item)):
            continue
        d = parse_optional_date(item)
        if d is not None:
            iso = d.isoformat()
            if iso not in seen:
                seen.add(iso)
                out.append(iso)
            continue
        s = unicodedata.normalize("NFKC", str(item).strip())
        if not s:
            continue
        for d2 in _extract_calendar_dates_from_text(s, y0):
            iso = d2.isoformat()
            if iso not in seen:
                seen.add(iso)
                out.append(iso)
    return sorted(out)
def _coerce_global_speed_rules(raw_val) -> list[dict]:
    """
    Gemini の global_speed_rules を正規化。
    坄覝素: process_contains / machine_contains（いうれか必須・部分一致用）, speed_multiplier（既存速度に乗算」0超〜10以下）。
    """
    out: list[dict] = []
    if not isinstance(raw_val, list):
        return out
    for item in raw_val:
        if not isinstance(item, dict):
            continue
        sm = item.get("speed_multiplier")
        if sm is None:
            sm = item.get("relative_speed")
        try:
            mult = float(sm)
        except (TypeError, ValueError):
            continue
        if mult <= 0 or mult > 10.0:
            continue
        pps = unicodedata.normalize("NFKC", str(item.get("process_contains") or "")).strip()
        mms = unicodedata.normalize("NFKC", str(item.get("machine_contains") or "")).strip()
        if not pps and not mms:
            continue
        out.append(
            {
                "process_contains": pps,
                "machine_contains": mms,
                "speed_multiplier": mult,
            }
        )
    return out
def _global_speed_rule_substring_matches_row(pnorm: str, mnorm: str, sub_nfkc: str) -> bool:
    """sub は空でなけれみ」工程名または機械名のいうれかに部分一致れれみ True。"""
    if not sub_nfkc:
        return True
    return sub_nfkc in pnorm or sub_nfkc in mnorm
def _global_speed_multiplier_for_row(process_name: str, machine_name: str, rules: list) -> float:
    """
    工程名・機械名に一致するルールの speed_multiplier を掛け合わせる（一致なしは 1.0）。

    process_contains / machine_contains はしれずれ **工程名または機械名のどうらか** に含まれていればよい。
    両方指定時は AND（例: 「熱融着」と「検査」は」列の組み合わせで両方睾れる行にマッポ。
    マスタ上で工程=検査・機械=熱融着機 のよごにキーワードは逆坴の列にあっても同じルールで効く。
    """
    if not rules:
        return 1.0
    pnorm = unicodedata.normalize("NFKC", str(process_name or "")).strip()
    mnorm = unicodedata.normalize("NFKC", str(machine_name or "")).strip()
    combined = 1.0
    for r in rules:
        if not isinstance(r, dict):
            continue
        pc = unicodedata.normalize(
            "NFKC", str(r.get("process_contains") or "").strip()
        )
        mc = unicodedata.normalize(
            "NFKC", str(r.get("machine_contains") or "").strip()
        )
        if not pc and not mc:
            continue
        if pc and not _global_speed_rule_substring_matches_row(pnorm, mnorm, pc):
            continue
        if mc and not _global_speed_rule_substring_matches_row(pnorm, mnorm, mc):
            continue
        try:
            m = float(r.get("speed_multiplier", 1.0))
        except (TypeError, ValueError):
            continue
        if m <= 0:
            continue
        combined *= m
    if combined <= 0:
        return 1.0
    return combined
def _infer_global_day_process_rules_from_free_text(text: str, ref_y: int) -> list[dict]:
    """
    Gemini は task_preferred_operators に誤って長文を入れた場合など」
    自然言語断片から global_day_process_operator_rules 相当を推定する（保守的）。
    例: 「2026/4/4 工程名:EC 森下と宮島を配台」
    """
    t = unicodedata.normalize("NFKC", str(text or "")).strip()
    if len(t) < 6:
        return []
    dates = _extract_calendar_dates_from_text(t, int(ref_y))
    if not dates:
        return []
    d0 = dates[0]
    proc_m = re.search(
        r"工程名?\s*[:：]?\s*([A-Za-z0-9一-龯ー・〆々]+)",
        t,
    )
    pc = proc_m.group(1).strip() if proc_m else ""
    if not pc:
        m2 = re.search(r"([\dA-Za-z一-龯ー・〆々]{1,12})\s*工程", t)
        pc = m2.group(1).strip() if m2 else ""
    if not pc:
        return []
    names: list[str] = []
    for m in re.finditer(
        r"([\u3040-\u9FFF々ー・A-Za-z・〆々]{1,16}?)\s*と\s*([\u3040-\u9FFF々ー・A-Za-z・〆々]{1,16}?)\s*を?\s*(?:配台|酝属|組ませ|同一フォーム)",
        t,
    ):
        a, b = m.group(1).strip(), m.group(2).strip()
        if a:
            names.append(a)
        if b:
            names.append(b)
    if len(names) < 2:
        return []
    return [
        {
            "date": d0.isoformat(),
            "process_contains": pc,
            "operator_names": names[:12],
        }
    ]
def _salvage_malformed_global_priority_gemini_dict(raw: dict, ref_y: int) -> dict:
    """
    Gemini は task_preferred_operators に **配列**や誤スキーマ（workstation_id 等）を返したとし」
    杨でうに global_day_process_operator_rules / scheduler_notes_ja へ救済れる。
    """
    out = dict(raw)
    tpo = out.get("task_preferred_operators")
    if not isinstance(tpo, list) or not tpo:
        return out
    narratives: list[str] = []
    extra_rule_objs: list[dict] = []
    for item in tpo:
        if not isinstance(item, dict):
            continue
        onames = item.get("operator_names")
        if isinstance(onames, list) and (
            item.get("date") is not None or item.get("process_contains")
        ):
            extra_rule_objs.append(item)
            continue
        for key in ("workstation_id", "schedule_notes_ai", "schedule_notes", "note", "text"):
            s = str(item.get(key) or "").strip()
            if len(s) >= 12:
                narratives.append(s[:800])
        for _k, v in item.items():
            if _k in (
                "factory_closure_dates",
                "operator_names",
                "date",
                "process_contains",
            ):
                continue
            if isinstance(v, str) and len(v) > 35 and ("酝" in v or "工程" in v):
                narratives.append(v[:800])
    out["task_preferred_operators"] = {}
    gdp_existing = out.get("global_day_process_operator_rules")
    gdp_list: list = list(gdp_existing) if isinstance(gdp_existing, list) else []
    gdp_list.extend(extra_rule_objs)
    seen_n: set[str] = set()
    for nb in narratives:
        if nb in seen_n:
            continue
        seen_n.add(nb)
        gdp_list.extend(_infer_global_day_process_rules_from_free_text(nb, ref_y))
    out["global_day_process_operator_rules"] = gdp_list
    if narratives:
        sn0 = str(out.get("scheduler_notes_ja") or "").strip()
        add = " | ".join(n[:280] for n in narratives[:4])
        out["scheduler_notes_ja"] = (sn0 + " " + add).strip()[:600]
    return out
def _coerce_global_priority_override_dict(raw, reference_year: int | None = None) -> dict:
    """Gemini 戻りを配台用フラグ・工場休業日リストに正規化。"""
    y0 = int(reference_year) if reference_year is not None else date.today().year

    def as_bool(v):
        if v is True:
            return True
        if v is False:
            return False
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return False
        s = unicodedata.normalize("NFKC", str(v).strip()).lower()
        return s in ("true", "1", "yes", "はい", "on")

    base = {
        "ignore_skill_requirements": False,
        "ignore_need_minimum": False,
        "abolish_all_scheduling_limits": False,
        "task_preferred_operators": {},
        "interpretation_ja": "",
        "factory_closure_dates": [],
        "scheduler_notes_ja": "",
        "global_speed_rules": [],
        "global_day_process_operator_rules": [],
    }
    if not isinstance(raw, dict):
        return base
    raw = _salvage_malformed_global_priority_gemini_dict(raw, y0)
    base["ignore_skill_requirements"] = as_bool(raw.get("ignore_skill_requirements"))
    base["ignore_need_minimum"] = as_bool(raw.get("ignore_need_minimum"))
    base["abolish_all_scheduling_limits"] = as_bool(
        raw.get("abolish_all_scheduling_limits")
    )
    base["task_preferred_operators"] = _coerce_task_preferred_operators_dict(
        raw.get("task_preferred_operators")
    )
    ij = raw.get("interpretation_ja")
    if ij is not None and not (isinstance(ij, float) and pd.isna(ij)):
        base["interpretation_ja"] = str(ij).strip()[:800]
    base["factory_closure_dates"] = _normalize_factory_closure_dates_iso_list(
        raw.get("factory_closure_dates"), y0
    )
    sn = raw.get("scheduler_notes_ja")
    if sn is not None and not (isinstance(sn, float) and pd.isna(sn)):
        base["scheduler_notes_ja"] = str(sn).strip()[:600]
    base["global_speed_rules"] = _coerce_global_speed_rules(raw.get("global_speed_rules"))
    base["global_day_process_operator_rules"] = _coerce_global_day_process_operator_rules(
        raw.get("global_day_process_operator_rules")
    )
    return base
def _parse_global_priority_override_gemini_response(res):
    """Gemini 応答から JSON オブジェクト1つを取り出す（```json フェンス付しでも坯）。"""
    raw = (_gemini_result_text(res) or "").strip()
    if not raw:
        return None
    candidate = None
    fence = re.search(
        r"```(?:json)?\s*(\{.*\})\s*```",
        raw,
        re.DOTALL | re.IGNORECASE,
    )
    if fence:
        candidate = fence.group(1).strip()
    elif raw.startswith("{"):
        candidate = raw
    else:
        loose = re.search(r"\{.*\}", raw, re.DOTALL)
        candidate = loose.group(0).strip() if loose else None
    if not candidate:
        return None
    try:
        parsed = json.loads(candidate)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None
def _apply_regex_factory_closure_fallback(coerced: dict, blob: str, ref_y: int) -> dict:
    """Gemini 未使用・応答解釈失敗時: ルールベースで工場休業日の値補完（従来互換）。"""
    out = dict(coerced)
    rx = parse_factory_closure_dates_from_global_comment(blob, ref_y)
    out["factory_closure_dates"] = sorted({d.isoformat() for d in rx})
    return out
def analyze_global_priority_override_comment(
    text: str, members: list, reference_year: int, ai_sheet_sink: dict | None = None
) -> dict:
    """
    メインシート「グローバルコメント」（UI 上の自由記述）を **Gemini で一括解釈**し、配台に効し JSON に蝽とれ。
    自然言語の文脈切り分け・改行の別指示解釈は AI に任せ」戻り値のキーの値システムは機械適用する。

    - factory_closure_dates: **工場全体**で稼働しない日（全員非稼働扱い）の YYYY-MM-DD 文字列の配列。該当なしは []。
    - ignore_skill_requirements / ignore_need_minimum / abolish_all_scheduling_limits / task_preferred_operators: 従来どおり。
    - global_speed_rules: **工程名・機械名**への部分一致（坄キーワードは **どうらの列にあっても坯**）で」既存の加工速度（シート＝上書き後）に **乗算**れるルールの配列。該当なしは []。
    - global_day_process_operator_rules: **日付＋工程名の部分一致＋複数メンバー**を」当日しの工程のタスクの**フォーム全員に必う含むる**ルールの配列。該当なしは []。
    - scheduler_notes_ja: 上記に蝽とししれない補足や靋用メモ（速度は可能なら global_speed_rules も併記）。

    API キー無し・JSON 解釈失敗時: 上記ブール・指定は既定値」工場休業日のみ従来のルールベース解析で補完。
    """
    ref_y = int(reference_year) if reference_year is not None else date.today().year
    empty = _coerce_global_priority_override_dict({}, ref_y)
    if not text or not str(text).strip():
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "スキップ（メイン原文なし）"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（メイン原文なし・API 未実行）"
        return empty
    blob = str(text).strip()
    mem_sig = ",".join(sorted(str(m).strip() for m in (members or []) if m))
    cache_fingerprint = f"{GLOBAL_PRIORITY_OVERRIDE_CACHE_PREFIX}{ref_y}\n{blob}\n{mem_sig}"
    cache_key = hashlib.sha256(cache_fingerprint.encode("utf-8")).hexdigest()
    ai_cache = load_ai_cache()
    cached = get_cached_ai_result(ai_cache, cache_key, content_key=cache_fingerprint)
    if cached is not None:
        logging.info("メイン再優先特記: キャッシュヒット（Gemini は呼びません）。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "なし（キャッシュ使用）"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（キャッシュ利用・今回 API 未実行）"
        return _finalize_global_priority_override(
            blob, _coerce_global_priority_override_dict(cached, ref_y)
        )

    if not API_KEY:
        logging.info("Gemini API キーが未設定のため、メイン再優先特記の AI 解析をスキップしました。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "なし（APIキー未設定・工場休業のみルール補完）"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（API キー未設定）"
        coerced = _apply_regex_factory_closure_fallback(
            _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
        )
        return _finalize_global_priority_override(blob, coerced)

    member_sample = ", ".join(str(m) for m in (members or [])[:80])
    if len(members or []) > 80:
        member_sample += " …"

    prompt = f"""あなたは工場の配台計画システム用アシスタントです。
Excel メインシートの **「グローバルコメント」**（自由記述・自然言語）の **全文** を読み」次のキーの値を挝つ JSON を1つ返してください。

」役割】
ユーザーは改行や坥点で複数の指示を書きことはありした。**文脈を読み分け**」配台システムは **機械的に適用でしる値** に蝽とし込んでしてさい。
推測でブールを true にしないこと。根拠は明確なとしの値 true。

」最優先】
この欄の内容はマスタ・スキル・need・タスク行・特別指定_備考の AI 指定より優先される例外指示として扱ゝれした。

」改行・複数行】
坄行・坄文は **原則として独立した指示** です。行をまたいで1つにまとめたり」**割引表睾（例 1/3）を日付と結び付けたりしない**こと。

」キー別ルール】

A) **factory_closure_dates** （配列・必須）
   - **工場全体**は稼働しない日（臨時休業・全工場休み・しの日は加工しない等）の日付を **YYYY-MM-DD** の文字列で列挙。
   - **個人の休み・特定ラインの値**の坜止はここに **含まない**（[]）。
   - 該当はなけれみ **空の配列 []**（キー省略試行）。
   - 年は省略されでいれみ西暦 {ref_y} 年として解釈。

B) **ignore_skill_requirements** / **ignore_need_minimum** / **abolish_all_scheduling_limits** / **task_preferred_operators**
   - 従来どおり（配台のスキル無視・人数1固定・制限撤廃・依頼NO→主担当OP指定）。該当なけれみ false または {{}}。

C) **global_speed_rules** （配列・必須）
   - 特定の **工程名**（Excel「工程名」列）や **機械名**（「機械名」列）に対し、**既存の加工速度に掛ける倍率** を指定するオブジェクトのリスト。
   - 坄オブジェクトのキー:
     - "process_contains": 文字列（省略坯）。**工程名または機械名のいうれか**に **部分一致**（NFKC 想定）。
     - "machine_contains": 文字列（省略坯）。**工程名または機械名のいうれか**に **部分一致**。
     - "speed_multiplier": 正の数。**1/3 の速度**なら約 **0.333333**（既存速度 × この値）。**2倝速**なら 2.0。
   - **両方指定時は AND**（2つのキーワードは」**両方とも**「工程名・機械名のどうらか」に睾れる行）。例: 工程=検査・機械=熱融着機 でも」工程=熱融着・機械=検査用設備 でもマッポしごる。
   - どうらか一方の値指定れれみ」しのキーワードは工程名または機械名のどうらかにあれみマッポ。
   - 該当指示はなけれみ **空の配列 []**。
   - 例: 「熱融着を使う検査の加工速度は1/3」→
     [{{"process_contains":"熱融着","machine_contains":"検査","speed_multiplier":0.333333}}]
     （「熱融着」と「検査」は工程名・機械名の組み合わせで权ごタスクの速度は約1/3になる）

D) **scheduler_notes_ja** （文字列・必須）
   - 上記キーに蝽とししれない補足。速度は **global_speed_rules で構造化でしるとしは必うしうらにも出す**（ここは人間坑け覝約でもよい）。無ければ ""。

E) **interpretation_ja** （文字列・必須）
   - 原文の覝約を1文（200文字以内）。

F) **global_day_process_operator_rules** （配列・必須）
   - **特定の稼働日**かつ **工程名（Excel「工程名」列）の部分一致** に当ではまるタスクについで」
     列挙した **全メンバーを同一フォームに必う含むる** ルール（**OP/AS どうらのスキルでも坯**。並び解決は **担当OP指定とともに**）。
   - **依頼NOは分かる主担当の1坝指定**は **task_preferred_operators** を使うこと。原文は **「◯月◯日の△工程にＡとＢを配台」** のよごに **日付・工程・複数坝**のときは **本配列**へ蝽とれ。
   - 坄オブジェクトのキー:
     - "date": **YYYY-MM-DD**（しの日に割り当でるロールに適用）
     - "process_contains": 工程名に **部分一致**（NFKC 想定）。例: "EC"
     - "operator_names": 並びの配列（例: ["森下", "宮島　花孝"]）
   - 該当指示はなけれみ **空の配列 []**。

」返答形式】
先頭は {{ で終ゝりは }} の **JSON オブジェクト1つのみ**（説明文・マークダウン禁止）。

必須キー一覧:
- "factory_closure_dates": string の配列（YYYY-MM-DD）
- "ignore_skill_requirements": true または false
- "ignore_need_minimum": true または false
- "abolish_all_scheduling_limits": true または false
- "task_preferred_operators": **JSON オブジェクトのみ**（キー=依頼NO・値=主担当並び）。**配列にしてはならない**。該当なしは {{}}
- "global_speed_rules": オブジェクトの配列（該当なしは []）
- "global_day_process_operator_rules": オブジェクトの配列（該当なしは []）
- "scheduler_notes_ja": 文字列
- "interpretation_ja": 文字列

」基準年】 日付言坊はあれみ西暦 {ref_y} 年として解釈してよい。

」登録メンバー坝の参考】（照合用。JSON キーには含まない）
{member_sample}

」グローバルコメント・原文】
{blob}
"""
    try:
        ppath = os.path.join(log_dir, "ai_global_priority_override_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("メイン再優先特記: プロンプト全文 → %s", ppath)
    except OSError as ex:
        logging.warning("メイン再優先特記: プロンプト保存失敗: %s", ex)

    client = _gemini_client(API_KEY)
    try:
        res, gem_model_used = _gemini_generate_content_with_retry(
            client, contents=prompt, log_label="メイン再優先特記"
        )
        record_gemini_response_usage(res, gem_model_used)
        parsed = _parse_global_priority_override_gemini_response(res)
        if parsed is None:
            logging.warning(
                "メイン再優先特記: AI 応答から JSON を解釈でしませんでした。キャッシュせう」次回再試行されした。"
            )
            try:
                rpath = os.path.join(log_dir, "ai_global_priority_override_last_response.txt")
                with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                    rf.write(_gemini_result_text(res) or "")
            except OSError:
                pass
            if ai_sheet_sink is not None:
                ai_sheet_sink["メイン再優先特記_AI_API"] = "あり（JSON解釈失敗・工場休業はルール補完）"
                ai_sheet_sink["メイン再優先特記_Geminiモデル"] = gem_model_used
            coerced = _apply_regex_factory_closure_fallback(
                _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
            )
            return _finalize_global_priority_override(blob, coerced)
        coerced = _coerce_global_priority_override_dict(parsed, ref_y)
        coerced = _finalize_global_priority_override(blob, coerced)
        try:
            rpath = os.path.join(log_dir, "ai_global_priority_override_last_response.txt")
            with open(rpath, "w", encoding="utf-8", newline="\n") as rf:
                rf.write(_gemini_result_text(res) or "")
        except OSError:
            pass
        put_cached_ai_result(ai_cache, cache_key, coerced, content_key=cache_fingerprint)
        save_ai_cache(ai_cache)
        _tpo = coerced.get("task_preferred_operators") or {}
        _fcd = coerced.get("factory_closure_dates") or []
        _gsr = coerced.get("global_speed_rules") or []
        _gdp = coerced.get("global_day_process_operator_rules") or []
        logging.info(
            "メイン再優先特記: AI 解釈 factory休業=%s日 速度ルール=%s件 日×工程フォーム=%s件 skill=%s need1=%s abolish=%s task_pref=%s件 — %s",
            len(_fcd),
            len(_gsr),
            len(_gdp),
            coerced["ignore_skill_requirements"],
            coerced["ignore_need_minimum"],
            coerced.get("abolish_all_scheduling_limits"),
            len(_tpo),
            coerced.get("interpretation_ja", "")[:100],
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = "あり"
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = gem_model_used
        return coerced
    except Exception as e:
        logging.warning("メイン再優先特記: Gemini 呼び出し失敗: %s", e)
        if ai_sheet_sink is not None:
            ai_sheet_sink["メイン再優先特記_AI_API"] = f"失敗: {e}"[:500]
            ai_sheet_sink["メイン再優先特記_Geminiモデル"] = "—（呼び出し失敗）"
        coerced = _apply_regex_factory_closure_fallback(
            _coerce_global_priority_override_dict({}, ref_y), blob, ref_y
        )
        return _finalize_global_priority_override(blob, coerced)
def default_result_task_sheet_column_order(max_history_len: int) -> list:
    """結果_タスク一覧の既定列順（履歴列数は実行時に決まる）。"""
    hist = [f"履歴{i+1}" for i in range(max_history_len)]
    return [
        "ステータス",
        "配台状況メモ",
        "タスクID",
        "工程名",
        "機械名",
        TASK_COL_SPEED,
        "優先度",
        RESULT_TASK_COL_DISPATCH_TRIAL_ORDER,
        *hist,
        "必須OP(上書)",
        "タスク効率",
        "加工途中",
        "特別指定あり",
        "担当OP指定",
        "回答納期",
        "指定納期",
        "計画基準納期",
        TASK_COL_RAW_INPUT_DATE,
        RESULT_TASK_COL_RAW_INPUT_DATE_PRE_PATTERN,
        RESULT_TASK_COL_PATTERN_RAW_SHIFT_NOTE,
        "紝期緊急",
        "加工開始日",
        "配台済_加工開始",
        "配台済_加工終了",
        RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16,
        "累計加工量",
        "残加工量",
        "完了率(実行時点)",
        "特別指定_AI",
    ]
def _task_date_key_for_result_sheet_sort(val):
    """結果_タスク一覧の並き替ご用。欠損・解釈試行は最後（date.max）。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return date.max
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        ts = pd.Timestamp(val)
        if pd.isna(ts):
            return date.max
        return ts.date()
    except Exception:
        return date.max
def _coerce_planning_date_for_deadline(d) -> date | None:
    """回答納期・指定納期などを date に正規化（欠損は None）。"""
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return None
def _result_task_due_met_column_in_df_columns(df_columns) -> str | None:
    """結果_タスク一覧の「納期を満たすか？」列（新旧見出し）を DataFrame 列から解決。"""
    colset = {str(c).strip() for c in df_columns}
    for cand in (
        RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16,
        "配台済_回答指定16時まで",
        "配台済_基準16時まで",
        "配完_回答指定16時まで",
        "配完_基準16時まで",
    ):
        if cand in colset:
            return cand
    return None
def _result_task_plan_end_within_answer_or_spec_16_label(
    plan_window: list | None, answer_due, specified_due, task_id=None
) -> str:
    """
    結果_タスク一覧用: 「配台済_加工終了」相当の最終終了が納期を満たすか。

    - 回答納期に日付があるときはそれを納期日とし、無いときは指定納期を納期日とする（両方無ければ「納期なし」）。
    - 依頼NO（task_id）先頭が「V」（前後空白除く・大文字小文字不問）のとき: 納期日の PLAN_DUE_DAY_COMPLETION_TIME（既定 16:00）までに終了すれば「はい」。
    - 上記以外: 納期日の暦日開始より前に終了すれば「はい」（＝納期日の前日までに加工完了が必要）。
    """
    if not plan_window or len(plan_window) < 2:
        return "未割当"
    _pe = plan_window[1]
    if _pe is None:
        return "未割当"
    answer_dd = _coerce_planning_date_for_deadline(answer_due)
    spec_dd = _coerce_planning_date_for_deadline(specified_due)
    if answer_dd is not None:
        due_day = answer_dd
    elif spec_dd is not None:
        due_day = spec_dd
    else:
        return "納期なし"
    tid = str(task_id or "").strip()
    is_v_prefix = bool(tid) and tid.lstrip().upper().startswith("V")
    try:
        if is_v_prefix:
            deadline_dt = datetime.combine(due_day, PLAN_DUE_DAY_COMPLETION_TIME)
            if _pe <= deadline_dt:
                return "はい"
            return "いいえ"
        start_of_due = datetime.combine(due_day, time.min)
        if _pe < start_of_due:
            return "はい"
        return "いいえ"
    except Exception:
        return "判定試行"
def _result_task_sheet_sort_key(t: dict):
    """
    結果_タスク一覧の表示順。①配台試行順番（generate_plan 冒頭でキュー順に付与した 1..n）昇順。
    欠損・非数は最後。同一試行順内は依頼NO・機械名」続けて加工開始日・紝期で安定化。
    """
    _dto = t.get("dispatch_trial_order")
    try:
        trial_k = int(_dto) if _dto is not None else 10**9
    except (TypeError, ValueError):
        trial_k = 10**9
    return (
        trial_k,
        str(t.get("task_id", "")).strip(),
        str(t.get("machine", "")).strip(),
        _task_date_key_for_result_sheet_sort(t.get("start_date_req")),
        _task_date_key_for_result_sheet_sort(t.get("answer_due_date")),
        _task_date_key_for_result_sheet_sort(t.get("specified_due_date")),
    )
def _is_result_task_history_expand_token(cell_val) -> bool:
    """列設定シートで「履歴」1行を置しと履歴1～n をしの佝置に展開れる。"""
    if cell_val is None or (isinstance(cell_val, float) and pd.isna(cell_val)):
        return False
    s = unicodedata.normalize("NFKC", str(cell_val).strip())
    return s in ("履歴", "履歴*")
def _result_task_column_alias_map(df_columns) -> dict:
    """見出しの NFKC 正規化キー → DataFrame 上の実列名。"""
    m = {}
    for c in df_columns:
        m[_nfkc_column_aliases(str(c).strip())] = c
    return m
def _resolve_result_task_column_label(label, col_by_norm: dict):
    if label is None or (isinstance(label, float) and pd.isna(label)):
        return None
    s = unicodedata.normalize("NFKC", str(label).strip())
    if not s or s.lower() in ("nan", "none"):
        return None
    nk = _nfkc_column_aliases(s)
    resolved = col_by_norm.get(nk)
    if resolved is not None:
        return resolved
    # 旧列名 → 納期を満たすか？（列設定シートの見出し互換）
    _due_met_key = _nfkc_column_aliases(RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16)
    for _old in (
        "配台済_基準16時まで",
        "配台済_回答指定16時まで",
        "配完_回答指定16時まで",
        "配完_基準16時まで",
    ):
        if nk == _nfkc_column_aliases(_old):
            return col_by_norm.get(_due_met_key)
    # 旧見出し「原板投入日」→ 結果 DataFrame の「原反投入日」
    if nk == _nfkc_column_aliases("原板投入日"):
        return col_by_norm.get(_nfkc_column_aliases(TASK_COL_RAW_INPUT_DATE))
    return None
def _parse_column_visible_cell(val) -> bool:
    """表示列: 空・未記入は True（表示）。FALSE/0/いいえ 等で非表示。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return True
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if val == 0:
            return False
        if val == 1:
            return True
    s = unicodedata.normalize("NFKC", str(val).strip()).lower()
    if s in ("", "true", "1", "はい", "yes", "on", "表示", "○"):
        return True
    if s in ("false", "flase", "0", "いいえ", "no", "off", "非表示", "隠れ", "×"):
        return False
    return True
def parse_result_task_column_config_dataframe(
    df_cfg: pd.DataFrame | None, max_history_len: int
) -> list | None:
    """
    「列設定_結果_タスク一覧」相当の DataFrame から (列ラベル, 表示) を上から読む。
    見出し「列名」と「表示」（無い場合は表示はまとめて True）。
    「履歴」「履歴*」の1行は履歴1～履歴n に展開し、同一行の表示フラグを共有れる。
    同一列名（NFKC・別名正規化後）は複数行ある場合は先頭行のみ採用し、以降はログに出して杨でる。
    """
    if df_cfg is None or df_cfg.empty:
        return None
    df_cfg = df_cfg.dropna(how="all")
    if df_cfg.empty:
        return None

    name_col = None
    for c in df_cfg.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases(COLUMN_CONFIG_HEADER_COL):
            name_col = c
            break
    if name_col is None:
        name_col = df_cfg.columns[0]

    vis_col = None
    for c in df_cfg.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases(COLUMN_CONFIG_VISIBLE_COL):
            vis_col = c
            break

    seen_norm: set[str] = set()
    out: list[tuple[str, bool]] = []

    def _try_add(label: str, vis: bool) -> None:
        lab = str(label).strip()
        if not lab:
            return
        nk = _nfkc_column_aliases(unicodedata.normalize("NFKC", lab))
        if nk in seen_norm:
            logging.warning(
                "列設定「%s」: 重複列名「%s」をスキップしました（上の行を優先）。",
                COLUMN_CONFIG_SHEET_NAME,
                lab,
            )
            return
        seen_norm.add(nk)
        out.append((lab, vis))

    for i in range(len(df_cfg)):
        raw = df_cfg[name_col].iloc[i]
        vis = _parse_column_visible_cell(df_cfg[vis_col].iloc[i] if vis_col is not None else None)
        if _is_result_task_history_expand_token(raw):
            for j in range(max_history_len):
                _try_add(f"履歴{j+1}", vis)
            continue
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        s = unicodedata.normalize("NFKC", str(raw).strip())
        if not s or s.lower() in ("nan", "none"):
            continue
        _try_add(s, vis)
    return out or None
def _openpyxl_write_column_config_sheet_ab(ws, rows: list[tuple[str, bool]]) -> None:
    """列設定シートの A:B を 列名・表示 のみで上書き（1行目見出し＋データ）。"""
    mat = [[COLUMN_CONFIG_HEADER_COL, COLUMN_CONFIG_VISIBLE_COL]]
    for lab, vis in rows:
        mat.append([lab, bool(vis)])
    n_r = len(mat)
    lim_r = max(int(ws.max_row or 1), n_r, 50)
    for r in range(1, lim_r + 1):
        for c in (1, 2):
            ws.cell(row=r, column=c).value = None
    for r in range(1, n_r + 1):
        ws.cell(row=r, column=1).value = mat[r - 1][0]
        ws.cell(row=r, column=2).value = mat[r - 1][1]
def _openpyxl_sheet_to_matrix(ws) -> list:
    """openpyxl Worksheet を矩形 list[list] にする（1行のみでも2次元）。"""
    try:
        mr = int(ws.max_row or 0)
        mc = int(ws.max_column or 0)
    except Exception:
        mr, mc = 0, 0
    if mr < 1 or mc < 1:
        return []
    out: list[list] = []
    for r in range(1, mr + 1):
        out.append([ws.cell(row=r, column=c).value for c in range(1, mc + 1)])
    return out
def load_result_task_column_rows_from_input_workbook(max_history_len: int) -> list | None:
    """
    優先: PM_AI_RESULT_TASK_COLUMN_CONFIG_CSV（UTF-8）。列「列名」「表示」（無ければ先頭2列）。

    次: PM_AI_COLUMN_CONFIG_WORKBOOK があればそのブックの「列設定_結果_タスク一覧」。

    従来: TASK_INPUT_WORKBOOK の同シート。
    """
    csvp = (os.environ.get(ENV_RESULT_TASK_COLUMN_CONFIG_CSV) or "").strip()
    if csvp and os.path.isfile(csvp):
        try:
            df_cfg = pd.read_csv(csvp, encoding="utf-8-sig")
            return parse_result_task_column_config_dataframe(df_cfg, max_history_len)
        except Exception as e:
            logging.warning(
                "列設定 CSV「%s」: 読めませんでした（%s）。ブックを試みます。",
                csvp,
                e,
            )
    wb = resolve_column_config_workbook_path(_excel_plan_input_wb())
    if not wb or not os.path.exists(wb):
        return None
    if _workbook_should_skip_openpyxl_io(wb):
        logging.info(
            "列設定: ブックに「%s」があるため、pandas(openpyxl) での「%s」読込をスキップ（既定列順を使用した）。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
            COLUMN_CONFIG_SHEET_NAME,
        )
        return None
    try:
        sn = _ooxml_workbook_sheet_names(wb)
        if sn is not None:
            want = unicodedata.normalize("NFKC", COLUMN_CONFIG_SHEET_NAME)
            if not any(unicodedata.normalize("NFKC", str(x)) == want for x in sn):
                return None
    except Exception:
        pass
    try:
        df_cfg = pd.read_excel(wb, sheet_name=COLUMN_CONFIG_SHEET_NAME, header=0)
    except ValueError:
        return None
    except Exception as e:
        try:
            df_cfg = pd.read_excel(
                wb,
                sheet_name=COLUMN_CONFIG_SHEET_NAME,
                header=0,
                engine="calamine",
            )
        except Exception:
            logging.warning(
                "シート「%s」: 読み込みに失敗したため、既定の列順を使用した (%s)",
                COLUMN_CONFIG_SHEET_NAME,
                e,
            )
            return None
    return parse_result_task_column_config_dataframe(df_cfg, max_history_len)
def _result_task_column_config_fallback_from_existing(
    df_tasks: pd.DataFrame, max_history_len: int
) -> tuple[list[str], dict[str, bool]]:
    """
    段階2で列順リストが空のときの補完。
    1) 結果 DataFrame に列があればその既存列順を採用し、マクロブック「列設定_結果_タスク一覧」
       で解決できる列は表示フラグを上書きする。
    2) 列が無ければ同シートから列名・表示を読む（TASK_INPUT_WORKBOOK・openpyxl 可のとき）。
    3) それも無ければ default_result_task_sheet_column_order。
    """
    rows_in = load_result_task_column_rows_from_input_workbook(max_history_len)
    cols = [str(c) for c in df_tasks.columns]

    if cols:
        vis_map = {c: True for c in cols}
        if rows_in:
            col_by_norm = _result_task_column_alias_map(cols)
            for item, vis in rows_in:
                resolved = _resolve_result_task_column_label(item, col_by_norm)
                if resolved and resolved in vis_map:
                    vis_map[resolved] = bool(vis)
        logging.warning(
            "段階2: 列順リストが空でした。結果 DataFrame の既存列（%s 列）で「%s」を補完しました。"
            + (" マクロブック列設定の表示フラグを反映しました。" if rows_in else ""),
            len(cols),
            COLUMN_CONFIG_SHEET_NAME,
        )
        return cols, vis_map

    if rows_in:
        order: list[str] = []
        vis_map: dict[str, bool] = {}
        for lab, vis in rows_in:
            order.append(lab)
            vis_map[lab] = bool(vis)
        logging.warning(
            "段階2: タスク行・列が無いため、マクロブック「%s」から %s 列で補完しました。",
            COLUMN_CONFIG_SHEET_NAME,
            len(order),
        )
        return order, vis_map

    dflt = list(default_result_task_sheet_column_order(max_history_len))
    logging.warning(
        "段階2: タスク行が 0 件かつ列設定の読込も無いため「%s」に既定の列名一覧を書き込みました。",
        COLUMN_CONFIG_SHEET_NAME,
    )
    return dflt, {c: True for c in dflt}
def apply_result_task_sheet_column_order(
    df: pd.DataFrame,
    max_history_len: int,
    *,
    config_dataframe: pd.DataFrame | None = None,
):
    """
    列設定シートはあれみしの順・表示を優先し、無い列は既定順で後ゝに追記（表示は True）。
    config_dataframe を渡した場合はファイルを読まうしの内容を列設定とみなす（実行時用）。
    戻り値: (並き替ご後 DataFrame, 実際の列名リスト, 設定ソース説明文字列, 列名→表示bool)
    """
    default_order = default_result_task_sheet_column_order(max_history_len)
    if config_dataframe is not None:
        user_rows = parse_result_task_column_config_dataframe(config_dataframe, max_history_len)
    else:
        user_rows = load_result_task_column_rows_from_input_workbook(max_history_len)
    if user_rows:
        primary = user_rows
        source = (
            f"マクロブック「{COLUMN_CONFIG_SHEET_NAME}」"
            if config_dataframe is None
            else f"シート「{COLUMN_CONFIG_SHEET_NAME}」（実行中ブック）"
        )
    else:
        primary = [(n, True) for n in default_order]
        source = "既定"

    actual = list(df.columns)
    actual_set = set(actual)
    col_by_norm = _result_task_column_alias_map(actual)
    vis_map = {c: True for c in actual}

    seen = set()
    ordered = []
    unknown = []

    for item, vis in primary:
        resolved = _resolve_result_task_column_label(item, col_by_norm)
        if resolved and resolved not in seen:
            ordered.append(resolved)
            seen.add(resolved)
            vis_map[resolved] = vis
        elif not resolved:
            if item is not None and not (isinstance(item, float) and pd.isna(item)):
                lab = str(item).strip()
                if lab and lab not in unknown:
                    unknown.append(lab)

    for name in default_order:
        if name in actual_set and name not in seen:
            ordered.append(name)
            seen.add(name)
    for name in actual:
        if name not in seen:
            ordered.append(name)
            seen.add(name)

    if unknown:
        logging.warning(
            "列設定: 結果に無い列名を無視しました（最大20件）: %s",
            ", ".join(unknown[:20]) + (" …" if len(unknown) > 20 else ""),
        )
    logging.info("結果_タスク一覧の列順ソース: %s（%s 列）", source, len(ordered))
    if not user_rows and config_dataframe is None:
        logging.info(
            "列順・表示のカスタマイズ: マクロ実行ブックにシート「%s」を追加。"
            " 見出し「%s」「%s」… 表示は FALSE の列は結果シートで非表示。"
            " 1行「履歴」で履歴1～n を挿入。VBA の「列設定_結果_タスク一覧_チェックボックスを配置」でチェックボックスを表示列に連動可能。",
            COLUMN_CONFIG_SHEET_NAME,
            COLUMN_CONFIG_HEADER_COL,
            COLUMN_CONFIG_VISIBLE_COL,
        )
    return df[ordered], ordered, source, vis_map
def _matrix_to_dataframe_header_first(matrix: list) -> pd.DataFrame | None:
    """1行目を列名とみなし DataFrame を返す。空なら None。"""
    if not matrix or not matrix[0]:
        return None
    header = []
    for x in matrix[0]:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            header.append("")
        else:
            header.append(str(x).strip())
    if not any(h for h in header):
        return None
    body = matrix[1:] if len(matrix) > 1 else []
    return pd.DataFrame(body, columns=header)
def _max_history_len_from_result_task_df_columns(columns) -> int:
    """結果_タスク一覧の「履歴n」列から n の最大を返す（無ければ 1）。"""
    imax = 0
    for c in columns:
        m = re.match(r"^履歴(\d+)$", str(c).strip())
        if m:
            imax = max(imax, int(m.group(1)))
    return max(imax, 1)
def apply_result_task_column_layout_via_openpyxl(workbook_path: str | None = None) -> bool:
    """
    マクロブックのディスク上の内容を読み、
    「列設定_結果_タスク一覧」の内容に合わせて「結果_タスク一覧」の列順と列非表示を更新する。
    「列設定_結果_タスク一覧」のセルは上書きしない（メモ・表外の A:B を消さない）。重複整理は
    dedupe_result_task_column_config_sheet_via_openpyxl / VBA「重複列名を整理」を使う。
    """
    path = (workbook_path or "").strip() or _excel_plan_input_wb().strip()
    if not path:
        logging.error("結果_タスク一覧 列適用: ブックパスは空です（TASK_INPUT_WORKBOOK を設定してください）。")
        return False
    if _workbook_should_skip_openpyxl_io(path):
        logging.error(
            "結果_タスク一覧 列適用: ブックに「%s」があるため openpyxl で編集できません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    keep_vba = str(path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(path, keep_vba=keep_vba, read_only=False, data_only=False)
    except Exception as e:
        logging.error("結果_タスク一覧 列適用: ブックを開けません: %s", e)
        return False
    try:
        try:
            ws_res = wb[RESULT_TASK_SHEET_NAME]
            ws_cfg = wb[COLUMN_CONFIG_SHEET_NAME]
        except KeyError as e:
            logging.error("結果_タスク一覧 列適用: 必須シートは見つかりません: %s", e)
            return False

        mat_res = _openpyxl_sheet_to_matrix(ws_res)
        mat_cfg = _openpyxl_sheet_to_matrix(ws_cfg)
        df_res = _matrix_to_dataframe_header_first(mat_res)
        df_cfg = _matrix_to_dataframe_header_first(mat_cfg)
        if df_res is None or df_res.empty:
            logging.error("結果_タスク一覧 列適用: 「%s」にデータはありません。", RESULT_TASK_SHEET_NAME)
            return False
        if df_cfg is None:
            logging.error("結果_タスク一覧 列適用: 「%s」の見出しを読めません。", COLUMN_CONFIG_SHEET_NAME)
            return False

        max_h = _max_history_len_from_result_task_df_columns(df_res.columns)
        rows_cfg = parse_result_task_column_config_dataframe(df_cfg, max_h)
        if not rows_cfg:
            logging.error(
                "結果_タスク一覧 列適用: 「%s」に有効な列名行はありません。",
                COLUMN_CONFIG_SHEET_NAME,
            )
            return False
        df_cfg_clean = pd.DataFrame(
            rows_cfg, columns=[COLUMN_CONFIG_HEADER_COL, COLUMN_CONFIG_VISIBLE_COL]
        )
        df_out, ordered, source, vis_map = apply_result_task_sheet_column_order(
            df_res, max_h, config_dataframe=df_cfg_clean
        )

        df_write = df_out.astype(object).where(pd.notna(df_out), None)
        headers = [str(h) for h in df_write.columns.tolist()]
        body = df_write.values.tolist()
        out_matrix = [headers] + body
        nrows = len(out_matrix)
        ncols = len(headers)
        if ncols == 0:
            return False

        max_old_r = int(ws_res.max_row or 1)
        max_old_c = int(ws_res.max_column or 1)
        for r in range(1, max(max_old_r, nrows) + 1):
            for c in range(1, max(max_old_c, ncols) + 1):
                ws_res.cell(row=r, column=c).value = None

        for r in range(1, nrows + 1):
            for c in range(1, ncols + 1):
                ws_res.cell(row=r, column=c).value = out_matrix[r - 1][c - 1]

        for ci in range(1, ncols + 1):
            ws_res.column_dimensions[get_column_letter(ci)].hidden = False

        for ci, col_name in enumerate(ordered, 1):
            if not vis_map.get(col_name, True):
                try:
                    ws_res.column_dimensions[get_column_letter(ci)].hidden = True
                except Exception as e:
                    logging.warning("列非表示に失敗（列%s %s）: %s", ci, col_name, e)

        try:
            wb.save(path)
        except Exception as e:
            logging.warning("結果_タスク一覧 列適用: 保存で警告: %s", e)

        logging.info(
            "結果_タスク一覧 列適用完了: %s（%s 列・非表示=%s）",
            source,
            len(ordered),
            sum(1 for c in ordered if not vis_map.get(c, True)),
        )
        return True
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def apply_result_task_column_layout_via_xlwings(workbook_path: str | None = None) -> bool:
    """互換名。実装は openpyxl のみ。"""
    return apply_result_task_column_layout_via_openpyxl(workbook_path)
def apply_result_task_column_layout_only() -> bool:
    """環境変数 TASK_INPUT_WORKBOOK のブックに対し列設定を適用する（VBA ボタン用）。"""
    p = _excel_plan_input_wb()
    return apply_result_task_column_layout_via_openpyxl(p)
_PLAN_INPUT_XLWINGS_ORIG_ROW = "__orig_sheet_row__"
def _plan_input_dispatch_trial_order_local_only_from_env() -> bool:
    """環境変数 PLAN_INPUT_DISPATCH_TRIAL_ORDER_LOCAL_ONLY は真なら post_load をスキップれる。"""
    v = (os.environ.get("PLAN_INPUT_DISPATCH_TRIAL_ORDER_LOCAL_ONLY") or "").strip().lower()
    return v in ("1", "true", "yes", "on", "y")
def _plan_input_row_is_blank_task_row(plan_df: "pd.DataFrame", row_i: int) -> bool:
    """依頼NO・工程名が両方空なら True（並べ替え・検証の対象外）。"""

    def _cell_empty(val) -> bool:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return True
        s = str(val).strip()
        if not s or s.lower() in ("nan", "none"):
            return True
        return False

    if TASK_COL_TASK_ID not in plan_df.columns or TASK_COL_MACHINE not in plan_df.columns:
        return True
    ti = plan_df.iat[row_i, plan_df.columns.get_loc(TASK_COL_TASK_ID)]
    mc = plan_df.iat[row_i, plan_df.columns.get_loc(TASK_COL_MACHINE)]
    return _cell_empty(ti) and _cell_empty(mc)
def _parse_dispatch_trial_order_float_sort_key(val) -> float | None:
    """「配台試行順番」セルを並べ替えキーとして float 化。空・不正・非有限は None。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            x = float(val)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(x):
            return None
        return x
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    try:
        x = float(s)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    return x
def _scalar_excel_accounting_speed_paren_negative_to_positive(val):
    """
    Excel は (数値) や会計表示で負数として格納・取得されることがある。
    配台計画の加工速度(m/分)は正として扱うため、数値が負のときは絶対値に戻す。
    """
    if val is None:
        return val
    if isinstance(val, float) and pd.isna(val):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            x = float(val)
        except (TypeError, ValueError):
            return val
        if math.isfinite(x) and x < 0:
            return abs(x)
    return val
def _plan_input_header_is_speed_excel_paren_fix_target(hname: str) -> bool:
    """見出しが加工速度／上書き／（元）参照のいずれか（NFKC 一致）なら True。"""
    if not hname or not str(hname).strip():
        return False
    k = _nfkc_column_aliases(str(hname).strip())
    ref = plan_reference_column_name(PLAN_COL_SPEED_OVERRIDE)
    for canon in (TASK_COL_SPEED, PLAN_COL_SPEED_OVERRIDE, ref):
        if k == _nfkc_column_aliases(canon):
            return True
    return False
def _apply_plan_input_excel_accounting_speed_fix_to_df(df: "pd.DataFrame") -> dict:
    """加工速度関連列の負数を正に補正。変更件数を返す。"""
    ref = plan_reference_column_name(PLAN_COL_SPEED_OVERRIDE)
    cols = [c for c in (TASK_COL_SPEED, PLAN_COL_SPEED_OVERRIDE, ref) if c in df.columns]
    per: dict[str, int] = {c: 0 for c in cols}
    for col in cols:
        loc = df.columns.get_loc(col)
        if isinstance(loc, slice):
            continue
        try:
            li = int(loc)
        except (TypeError, ValueError):
            continue
        for ri in range(len(df)):
            old = df.iat[ri, li]
            if (
                isinstance(old, (int, float))
                and not isinstance(old, bool)
                and not pd.isna(old)
            ):
                try:
                    x = float(old)
                except (TypeError, ValueError):
                    continue
                if math.isfinite(x) and x < 0:
                    df.iat[ri, li] = abs(x)
                    per[col] += 1
    total = sum(per.values())
    return {"per_col": per, "total": int(total)}
def _df_first_col_index_for_header(columns: pd.Index, hname: str) -> int | None:
    """列名が重複している DataFrame でも、先頭一致列の 0 始まり整数インデックスを返す。"""
    if not hname:
        return None
    for i, c in enumerate(columns):
        if c is None or (isinstance(c, float) and pd.isna(c)):
            lab = ""
        else:
            lab = str(c).strip()
        if lab == hname:
            return i
    return None
def _plan_input_dispatch_trial_order_sort_tuples_for_active_rows(
    df: "pd.DataFrame",
    active: list[int],
    dto_idx: int,
) -> tuple[dict[int, tuple], dict[str, float]]:
    """
    §A-1 維持: 依頼NO ブロック（ブロック内最小の試行順キー）→ 加工内容 rank → 同一依頼内行順。
    戻り値: (行 index → sort tuple, 依頼NO → ブロック最小 float キー) — ログ用。
    """
    seq_by_tid = _collect_process_content_order_by_task_id(df)
    tid_eligible_block_float: dict[str, float] = {}
    tid_line_seq: dict[int, int] = {}
    tid_next_eligible_line: dict[str, int] = defaultdict(int)
    proc_idx = df.columns.get_loc(TASK_COL_MACHINE) if TASK_COL_MACHINE in df.columns else None
    if isinstance(proc_idx, slice):
        proc_idx = None

    for i in active:
        row = df.iloc[i]
        tid = planning_task_id_str_from_plan_row(row)
        excluded = _plan_row_exclude_from_assignment(row)
        if tid and not excluded:
            tid_line_seq[i] = tid_next_eligible_line[tid]
            tid_next_eligible_line[tid] += 1
        else:
            tid_line_seq[i] = i
        fk = _parse_dispatch_trial_order_float_sort_key(df.iat[i, dto_idx])
        if tid and fk is not None and not excluded:
            prev = tid_eligible_block_float.get(tid)
            tid_eligible_block_float[tid] = fk if prev is None else min(prev, fk)

    sort_tuple_by_row: dict[int, tuple] = {}
    for i in active:
        fk = _parse_dispatch_trial_order_float_sort_key(df.iat[i, dto_idx])
        if fk is None:
            sort_tuple_by_row[i] = (1, i)
            continue
        row = df.iloc[i]
        tid = planning_task_id_str_from_plan_row(row)
        excluded = _plan_row_exclude_from_assignment(row)
        if excluded:
            block = fk
        elif tid:
            block = tid_eligible_block_float.get(tid, fk)
        else:
            block = fk
        rank = None
        if not excluded and proc_idx is not None:
            proc = df.iat[i, proc_idx]
            rank = _process_sequence_rank_for_machine(
                proc, seq_by_tid.get(tid) or []
            )
        if excluded:
            sort_tuple_by_row[i] = (0, block, 1, i, i)
        else:
            rank_key = int(rank) if rank is not None else 10**9
            line_key = tid_line_seq.get(i, i)
            sort_tuple_by_row[i] = (0, block, 0, rank_key, line_key, i)
    return sort_tuple_by_row, tid_eligible_block_float
def sort_plan_input_dispatch_trial_order_by_float_keys_via_openpyxl(
    workbook_path: str | None = None,
) -> bool:
    """
    「配台計画_タスク入力」の **現在のシート内容だけ** を使い、列「配台試行順番」を
    小数を含む並べ替えキーとして解釈して昇順に行を並べ替え、1..n に振り直す。

    同一依頼NO内の配台対象行は §A-1（``加工内容`` のカンマ区切り順）で工程行を連続させる。
    ブロック位置は当該依頼NOの **配台不要オフ行** の試行順キー最小値。「配台不要」オン行は単独行のキーで並ぶ。

    - ``_apply_planning_sheet_post_load_mutations`` ・マスタ ・
      ``fill_plan_dispatch_trial_order_column_stage1`` は **呼ばない**。
    - 依頼NO・工程名が両方空の行は対象外。先頭の空行と、最後のデータ行より後の空行は
      元の順のまま残す。
    - 最初の対象行から最後の対象行までは **途切れなく対象行** でなければならない。
    - **有限の float** として解釈できるキー同士は **重複してはならない**。
    - キーが空・解釈不能の対象行は、**すべての有効キー行の後ろ**に元の行順を保って並べ、
      連番 1..n はその並びで振り直す。
    """
    path = (workbook_path or "").strip() or _excel_plan_input_wb().strip()
    if not path:
        logging.error("配台試行順番（小数キー並べ）: ブックパスが空です。")
        return False
    if _workbook_should_skip_openpyxl_io(path):
        logging.error(
            "配台試行順番（小数キー並べ）: ブックに「%s」があるため openpyxl で編集できません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    keep_vba = str(path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(path, keep_vba=keep_vba, read_only=False, data_only=False)
    except Exception as e:
        logging.error("配台試行順番（小数キー並べ）: ブックを開けません: %s", e)
        return False
    try:
        try:
            ws = wb[PLAN_INPUT_SHEET_NAME]
        except KeyError as e:
            logging.error("配台試行順番（小数キー並べ）: シート接続に失敗: %s", e)
            return False

        mat = _openpyxl_sheet_to_matrix(ws)
        df = _matrix_to_dataframe_header_first(mat)
        if df is None or df.empty:
            logging.warning("配台試行順番（小数キー並べ）: データ行がありません。")
            return False

        df = df.copy()
        df.columns = df.columns.str.strip()
        df = _align_dataframe_headers_to_canonical(df, plan_input_sheet_column_order())
        for c in plan_input_sheet_column_order():
            if c not in df.columns:
                df[c] = ""

        if df.columns.duplicated().any():
            dup_labels = sorted(
                {str(c) for c in df.columns[df.columns.duplicated(keep=False)]}
            )
            logging.warning(
                "配台試行順番（小数キー並べ）: 見出しの重複列があります（先頭列を参照します）: %s",
                dup_labels[:25],
            )

        _apply_plan_input_excel_accounting_speed_fix_to_df(df)

        dto_col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
        if dto_col not in df.columns:
            logging.error("配台試行順番（小数キー並べ）: 列「%s」がありません。", dto_col)
            return False
        dto_idx = df.columns.get_loc(dto_col)
        if isinstance(dto_idx, slice):
            logging.error("配台試行順番（小数キー並べ）: 列「%s」が複数あります。", dto_col)
            return False

        n = len(df)
        active = [i for i in range(n) if not _plan_input_row_is_blank_task_row(df, i)]
        if not active:
            logging.error(
                "配台試行順番（小数キー並べ）: 依頼NO または 工程名 がある行がありません。"
            )
            return False
        first = min(active)
        last = max(active)
        for k in range(first, last + 1):
            if k not in active:
                logging.error(
                    "配台試行順番（小数キー並べ）: %s 行目付近に、依頼NO・工程名が両方空の行が"
                    " データの途中にあります。",
                    k + 2,
                )
                return False

        row_by_key: dict[float, int] = {}
        n_invalid_key = 0
        for i in active:
            fk = _parse_dispatch_trial_order_float_sort_key(df.iat[i, dto_idx])
            if fk is None:
                n_invalid_key += 1
                continue
            if fk in row_by_key:
                logging.error(
                    "配台試行順番（小数キー並べ）: 並べ替えキー %s が %s 行目と %s 行目で重複しています。",
                    fk,
                    row_by_key[fk] + 2,
                    i + 2,
                )
                return False
            row_by_key[fk] = i

        sort_tuple_by_row, _tid_blocks = (
            _plan_input_dispatch_trial_order_sort_tuples_for_active_rows(
                df, active, dto_idx
            )
        )

        if n_invalid_key:
            logging.info(
                "配台試行順番（小数キー並べ）: 「%s」が空・非数値のデータ行が %s 行あります。"
                " 有効キー行の後ろに並べ、連番化します。",
                dto_col,
                n_invalid_key,
            )

        sorted_active = sorted(active, key=lambda ri: sort_tuple_by_row[ri])
        df_mut = df.copy()
        for rank, i in enumerate(sorted_active, start=1):
            df_mut.iat[i, dto_idx] = rank

        leading = [i for i in range(0, first)]
        trailing = [i for i in range(last + 1, n)]
        orig_list = leading + sorted_active + trailing

        df_sorted = df_mut.iloc[orig_list].reset_index(drop=True)

        header_row = mat[0] if mat else []
        n_hdr = len(header_row)
        if n_hdr == 0:
            return False

        def _pad_row(r, n):
            r = list(r) if r is not None else []
            if len(r) < n:
                r = r + [None] * (n - len(r))
            return r

        new_mat = [_pad_row(header_row, n_hdr)]
        for i in range(len(df_sorted)):
            orig = orig_list[i]
            src_row = mat[orig + 1] if orig + 1 < len(mat) else []
            src_row = _pad_row(src_row, n_hdr)
            out_row = []
            for j in range(n_hdr):
                h_cell = header_row[j]
                if h_cell is None or (isinstance(h_cell, float) and pd.isna(h_cell)):
                    hname = ""
                else:
                    hname = str(h_cell).strip()
                ci_hdr = _df_first_col_index_for_header(df_sorted.columns, hname)
                if ci_hdr is not None:
                    v = df_sorted.iat[i, ci_hdr]
                    if pd.isna(v):
                        out_row.append(None)
                    else:
                        if _plan_input_header_is_speed_excel_paren_fix_target(hname):
                            v = _scalar_excel_accounting_speed_paren_negative_to_positive(v)
                        out_row.append(v)
                else:
                    _v = src_row[j]
                    if _plan_input_header_is_speed_excel_paren_fix_target(hname):
                        _v = _scalar_excel_accounting_speed_paren_negative_to_positive(_v)
                    out_row.append(_v)
            new_mat.append(out_row)

        n_r = len(new_mat)
        for r in range(1, n_r + 1):
            for c in range(1, n_hdr + 1):
                ws.cell(row=r, column=c).value = new_mat[r - 1][c - 1]

        try:
            wb.save(path)
        except Exception as e:
            logging.warning("配台試行順番（小数キー並べ）: Save 警告: %s", e)

        logging.info(
            "配台試行順番（小数キー並べ）: 「%s」を %s データ行で並べ替え・連番化しました。",
            PLAN_INPUT_SHEET_NAME,
            len(sorted_active),
        )
        return True
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def sort_plan_input_dispatch_trial_order_by_float_keys_via_xlwings(
    workbook_path: str | None = None,
) -> bool:
    """互換名（関数名は後方互換）。"""
    return sort_plan_input_dispatch_trial_order_by_float_keys_via_openpyxl(workbook_path)
def sort_plan_input_dispatch_trial_order_by_float_keys_only() -> bool:
    """TASK_INPUT_WORKBOOK に対する「小数キーで並べ替え→1..n」（VBA / cmd 経由）。"""
    p = _excel_plan_input_wb()
    return sort_plan_input_dispatch_trial_order_by_float_keys_via_openpyxl(p)
def apply_plan_input_column_layout_only() -> bool:
    """
    配台計画_タスク入力の列順・表示のみを適用する予定（VBA 用）。
    未実装。列の並よは段階1出力または手動整睆を使用してください。
    """
    logging.warning("apply_plan_input_column_layout_only: not implemented")
    return False
def dedupe_result_task_column_config_sheet_via_openpyxl(workbook_path: str | None = None) -> bool:
    """
    「列設定_結果_タスク一覧」の A:B の値を」重複列名を除いた一覧で書き直れ（先の行を優先）。
    「結果_タスク一覧」はあれみ履歴列数の解釈に使う。結果シートは変更しない。
    """
    path = (workbook_path or "").strip() or _excel_plan_input_wb().strip()
    if not path:
        logging.error("列設定 重複整睆: ブックパスは空です。")
        return False
    if _workbook_should_skip_openpyxl_io(path):
        logging.error(
            "列設定 重複整睆: ブックに「%s」があるため openpyxl で編集できません。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    keep_vba = str(path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(path, keep_vba=keep_vba, read_only=False, data_only=False)
    except Exception as e:
        logging.error("列設定 重複整睆: ブックを開けません: %s", e)
        return False
    try:
        try:
            ws_cfg = wb[COLUMN_CONFIG_SHEET_NAME]
        except KeyError as e:
            logging.error("列設定 重複整睆: 接続またはシート取得に失敗: %s", e)
            return False

        max_h = 1
        try:
            ws_res = wb[RESULT_TASK_SHEET_NAME]
            df_r = _matrix_to_dataframe_header_first(_openpyxl_sheet_to_matrix(ws_res))
            if df_r is not None and not df_r.empty:
                max_h = _max_history_len_from_result_task_df_columns(df_r.columns)
        except Exception:
            pass

        df_cfg = _matrix_to_dataframe_header_first(_openpyxl_sheet_to_matrix(ws_cfg))
        if df_cfg is None:
            logging.error("列設定 重複整睆: 「%s」の見出しを読めません。", COLUMN_CONFIG_SHEET_NAME)
            return False
        rows = parse_result_task_column_config_dataframe(df_cfg, max_h)
        if not rows:
            logging.warning("列設定 重複整睆: 有効なデータ行はありません。")
            return False
        _openpyxl_write_column_config_sheet_ab(ws_cfg, rows)
        try:
            wb.save(path)
        except Exception as e:
            logging.warning("列設定 重複整睆: 保存警告: %s", e)
        logging.info(
            "列設定「%s」を重複除去済みで %s 行に整睆しました（履歴展開後の行数）。",
            COLUMN_CONFIG_SHEET_NAME,
            len(rows),
        )
        return True
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def dedupe_result_task_column_config_sheet_via_xlwings(workbook_path: str | None = None) -> bool:
    """互換名（関数名は後方互換）。"""
    return dedupe_result_task_column_config_sheet_via_openpyxl(workbook_path)
def dedupe_result_task_column_config_sheet_only() -> bool:
    """環境変数 TASK_INPUT_WORKBOOK のブックの列設定シートの値重複整睆（VBA 用）。"""
    p = _excel_plan_input_wb()
    return dedupe_result_task_column_config_sheet_via_openpyxl(p)
def _apply_result_task_sheet_column_visibility(worksheet, column_names: list, vis_map: dict):
    """結果_タスク一覧で」vis_map は False の列を非表示にれる。"""
    for idx, col_name in enumerate(column_names, 1):
        if not vis_map.get(col_name, True):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True
def _result_task_sheet_column_width_for_header(header: str) -> float:
    """結果_タスク一覧: 列見出しに応じた標準幅（Excel 文字幅の目安）。"""
    h = str(header).strip()
    if not h:
        return 10.0
    if h == "ステータス":
        return 11.0
    if h in (TASK_COL_TASK_ID, "タスクID", "依頼NO"):
        return 12.0
    if h == TASK_COL_MACHINE or h == "工程名":
        return 14.0
    if h == TASK_COL_MACHINE_NAME or h == "機械名":
        return 19.0
    if "加工速度" in h:
        return 10.0
    if "優先度" in h:
        return 8.5
    if "配台試行" in h or (h.endswith("試行") and "配台" in h):
        return 9.0
    if h.startswith("履歴"):
        return 38.0
    if "必須" in h and "OP" in h:
        return 10.0
    if "タスク効率" in h or "タスク効" in h:
        return 11.0
    if "加工途中" in h:
        return 11.0
    if "特別指定" in h:
        return 11.0
    if "担当" in h and "OP" in h:
        return 14.0
    if "納期" in h or "回答" in h:
        return 12.0
    return float(min(max(len(h) + 2.5, 9.0), 36.0))
def _apply_result_task_sheet_layout_polish(worksheet, column_names: list):
    """
    結果_タスク一覧の視認性向上: 列幅、見出し折返し、履歴列の折返し、左3列の窓枠固定。
    （着色・リッチテキスト・ハイパーリンク適用後に呼ぶこと）
    """
    if worksheet is None or not column_names:
        return
    mr = int(worksheet.max_row or 1)
    mc = int(worksheet.max_column or 0)
    if mc < 1:
        return

    hist_cols: list[int] = []
    center_data_cols: set[int] = set()
    for idx, col_name in enumerate(column_names, 1):
        cn = str(col_name).strip()
        if cn.startswith("履歴"):
            hist_cols.append(idx)
        letter = get_column_letter(idx)
        try:
            dim = worksheet.column_dimensions[letter]
            if getattr(dim, "hidden", False):
                continue
        except Exception:
            pass
        w = _result_task_sheet_column_width_for_header(cn)
        try:
            worksheet.column_dimensions[letter].width = w
        except Exception:
            pass
        if cn in (
            "ステータス",
            "優先度",
            TASK_COL_TASK_ID,
            "タスクID",
            "依頼NO",
            TASK_COL_MACHINE,
            TASK_COL_MACHINE_NAME,
        ) or "加工速度" in cn or ("必須" in cn and "OP" in cn):
            center_data_cols.add(idx)
        if "配台試行" in cn or (cn.endswith("試行") and "配台" in cn):
            center_data_cols.add(idx)

    hdr_align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    for ci in range(1, mc + 1):
        worksheet.cell(row=1, column=ci).alignment = hdr_align

    wrap_top = Alignment(wrap_text=True, vertical="top")
    center_top = Alignment(wrap_text=False, vertical="top", horizontal="center")
    default_top = Alignment(wrap_text=False, vertical="top")

    for r in range(2, mr + 1):
        for ci in range(1, mc + 1):
            cell = worksheet.cell(row=r, column=ci)
            if ci in hist_cols:
                cell.alignment = wrap_top
            elif ci in center_data_cols:
                cell.alignment = center_top
            else:
                cell.alignment = default_top

    try:
        worksheet.row_dimensions[1].height = 32.0
    except Exception:
        pass

    try:
        worksheet.freeze_panes = "D2"
    except Exception:
        pass
def _norm_history_member_label(name: str) -> str:
    """履歴の担当坝比較用（全角空白を半角1個化・剝後trim・連続空白の圧縮）。"""
    t = str(name or "").replace("\u3000", " ").strip()
    return " ".join(t.split())
def _history_team_text_main_assignment_only(h: dict) -> str:
    """
    結果シート「担当」欄用: メイン割付確定時点の坝剝（余力追記サブは含まない）。
    append_surplus 後の h['team'] から post_dispatch_surplus_names を除外れる。
    """
    raw = (h.get("team") or "").strip()
    if not raw:
        return ""
    ps = h.get("post_dispatch_surplus_names") or []
    if not ps:
        return raw
    ps_set = {
        _norm_history_member_label(x)
        for x in ps
        if x and str(x).strip()
    }
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    kept = [p for p in parts if _norm_history_member_label(p) not in ps_set]
    return ", ".join(kept) if kept else raw
def _result_assigned_history_team_key(team_s: str) -> str:
    """結果シート用: 履歴セグメント同士の担当文字列比較（NFKC・空白正規化）。"""
    s = unicodedata.normalize("NFKC", str(team_s or "").strip())
    return " ".join(s.split())
def _union_name_lists_preserve_order(
    a: list | None, b: list | None
) -> list[str]:
    """名前列を重複なく結合（先勝ちで順序維持）。"""
    out: list[str] = []
    seen: set[str] = set()
    for xs in (a or [], b or []):
        for x in xs:
            t = str(x).strip()
            if not t:
                continue
            k = _norm_history_member_label(t)
            if k in seen:
                continue
            seen.add(k)
            out.append(t)
    return out
def _assigned_history_segment_copy(h: dict) -> dict:
    """履歴 dict の浅いコピー（名前リストは複製してマージ時に汚染しない）。"""
    out = dict(h)
    for k in ("surplus_member_names", "post_dispatch_surplus_names"):
        v = out.get(k)
        if isinstance(v, list):
            out[k] = list(v)
    return out
def _assigned_history_contiguous_mergeable(a: dict, b: dict) -> bool:
    """
    連続作業として 1 履歴にまとめられるか。
    前セグメント終了 == 次セグメント開始・同一担当・同一組合せ行 ID のときのみ。
    """
    a_end = a.get("end_dt")
    b_start = b.get("start_dt")
    if not isinstance(a_end, datetime) or not isinstance(b_start, datetime):
        return False
    if a_end != b_start:
        return False
    if _result_assigned_history_team_key(a.get("team", "")) != _result_assigned_history_team_key(
        b.get("team", "")
    ):
        return False
    return (a.get("combo_sheet_row_id")) == (b.get("combo_sheet_row_id"))
def _merge_two_assigned_history_display_segments(a: dict, b: dict) -> dict:
    """連続セグメント b を a に取り込んだ新 dict（結果シート表示専用）。"""
    out = _assigned_history_segment_copy(a)
    try:
        da = int(a.get("done_m") or 0)
    except (TypeError, ValueError):
        da = 0
    try:
        db = int(b.get("done_m") or 0)
    except (TypeError, ValueError):
        db = 0
    out["done_m"] = da + db
    out["end_dt"] = b.get("end_dt")
    out["need_surplus_assigned"] = bool(
        a.get("need_surplus_assigned") or b.get("need_surplus_assigned")
    )
    out["surplus_member_names"] = _union_name_lists_preserve_order(
        a.get("surplus_member_names"), b.get("surplus_member_names")
    )
    out["post_dispatch_surplus_names"] = _union_name_lists_preserve_order(
        a.get("post_dispatch_surplus_names"), b.get("post_dispatch_surplus_names")
    )
    return out
def merge_assigned_history_contiguous_for_result_sheet(hist: list | None) -> list:
    """
    結果_タスク一覧向け: ロール確定ごとの内部履歴を、時刻・担当・組合せ ID が連続する塊で 1 件にまとめる。
    配台中のロールパイプライン等は生の assigned_history を参照するため、本関数は出力直前にのみ使う。
    """
    hist = hist or []
    if len(hist) < 2:
        return [_assigned_history_segment_copy(h) for h in hist]
    out: list[dict] = []
    cur = _assigned_history_segment_copy(hist[0])
    for nxt_raw in hist[1:]:
        nxt = _assigned_history_segment_copy(nxt_raw)
        if _assigned_history_contiguous_mergeable(cur, nxt):
            cur = _merge_two_assigned_history_display_segments(cur, nxt)
        else:
            out.append(cur)
            cur = nxt
    out.append(cur)
    return out
def _format_result_task_history_cell(task: dict, h: dict) -> str:
    """結果_タスク一覧の履歴セル文字列（短い記号: #=組合せ行ID, 主=メイン担当, +=超過, 余=余力追記）。"""
    um = task.get("unit_m") or 0
    try:
        done_r = int(h["done_m"] / um) if um else 0
    except (TypeError, ValueError, ZeroDivisionError):
        done_r = 0
    dm = h.get("done_m", 0)
    d = h.get("date", "") or ""
    parts_out: list[str] = [f"・【{d}】：{done_r}R/{dm}m"]
    cid = h.get("combo_sheet_row_id")
    if cid is not None:
        try:
            parts_out.append(f"#{int(cid)}")
        except (TypeError, ValueError):
            parts_out.append(f"#{cid}")
    team = _history_team_text_main_assignment_only(h)
    if team:
        parts_out.append(f"主:{team}")
    sm = h.get("surplus_member_names") or []
    if sm:
        parts_out.append("+" + ",".join(str(x) for x in sm))
    ps = h.get("post_dispatch_surplus_names") or []
    if ps:
        parts_out.append("余:" + ",".join(str(x) for x in ps))
    return " ".join(parts_out)
_RESULT_TASK_HISTORY_RICH_HEAD_RE = re.compile(r"^・(【[^】]*】)(.*)$", re.DOTALL)
def _apply_result_task_history_rich_text(worksheet, column_names: list):
    """
    履歴列: 「・【日付】：…」の日付括弧部分を青色リッチテキストにする。
    openpyxl 3.1 未満ではスキップ（文字列の【】のみ）。
    """
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles.colors import Color
    except ImportError:
        return

    hist_cols = [
        i + 1 for i, c in enumerate(column_names) if str(c).startswith("履歴")
    ]
    if not hist_cols:
        return

    _fn = _effective_result_book_font_name()
    # openpyxl 3.1+ InlineFont は OOXML に合わせ rFont（Font オブジェクトの name とは別名）
    _plain_kw: dict = {"rFont": _fn}
    _blue_kw: dict = {"rFont": _fn, "color": Color(rgb="FF0070C0")}
    plain_if = InlineFont(**_plain_kw)
    blue_if = InlineFont(**_blue_kw)
    top = Alignment(wrap_text=False, vertical="top")

    for r in range(2, worksheet.max_row + 1):
        for ci in hist_cols:
            cell = worksheet.cell(row=r, column=ci)
            v = cell.value
            if not isinstance(v, str) or not v.startswith("・【"):
                continue
            m = _RESULT_TASK_HISTORY_RICH_HEAD_RE.match(v)
            if not m:
                continue
            bracketed, rest = m.group(1), m.group(2)
            cell.value = CellRichText(
                TextBlock(plain_if, "・"),
                TextBlock(blue_if, bracketed),
                TextBlock(plain_if, rest),
            )
            cell.alignment = top
def _apply_result_task_date_columns_blue_font(worksheet, column_names: list):
    """
    結果_タスク一覧: 回答納期・指定納期・計画基準納期・原反投入日・加工開始日のセルを青色にれる。
    （履歴列の【日付】は _apply_result_task_history_rich_text で着色。色は 0070C0 で統一）
    """
    blue = _result_font(color="0070C0")
    top = Alignment(wrap_text=False, vertical="top")
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) not in RESULT_TASK_DATE_STYLE_HEADERS:
            continue
        for r in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=r, column=col_idx)
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not str(v).strip():
                continue
            cell.font = blue
            cell.alignment = top
def _apply_result_task_history_need_surplus_highlight(
    worksheet, column_names: list, sorted_tasks: list
):
    """
    need「配台時追加人数」相当で基本必須人数を超ごで採用したブロック」または
    メイン完了後の余力追記でサブは増ごたブロックに対応れる「履歴n」セルを薄黄に塗る。
    """
    hist_cols: list[tuple[int, int]] = []
    for col_idx, col_name in enumerate(column_names, 1):
        m = re.match(r"^履歴(\d+)$", str(col_name).strip())
        if m:
            hist_cols.append((int(m.group(1)), col_idx))
    hist_cols.sort(key=lambda x: x[0])
    if not hist_cols or worksheet.max_row < 2:
        return
    fill_surplus = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    n_tasks = len(sorted_tasks)
    for r in range(2, worksheet.max_row + 1):
        ti = r - 2
        if ti < 0 or ti >= n_tasks:
            continue
        ah = merge_assigned_history_contiguous_for_result_sheet(
            sorted_tasks[ti].get("assigned_history")
        )
        for ord1, cidx in hist_cols:
            i = ord1 - 1
            if i < 0 or i >= len(ah):
                continue
            if not ah[i].get("need_surplus_assigned"):
                continue
            worksheet.cell(row=r, column=cidx).fill = fill_surplus
def _apply_result_task_task_id_content_mismatch_highlight(
    worksheet, column_names: list, sorted_tasks: list
):
    """
    加工内容に工程名は含まれない行の「タスクID」セルを赤背景・白文字にする（元データとの整合の確認用）。
    """
    task_id_col_idx = None
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) == "タスクID":
            task_id_col_idx = col_idx
            break
    if task_id_col_idx is None or worksheet.max_row < 2:
        return
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white = _result_font(color="FFFFFF")
    top = Alignment(wrap_text=False, vertical="top")
    n_data = worksheet.max_row - 1
    for i in range(min(len(sorted_tasks), n_data)):
        if not sorted_tasks[i].get("process_content_mismatch"):
            continue
        cell = worksheet.cell(row=i + 2, column=task_id_col_idx)
        cell.fill = fill_red
        cell.font = font_white
        cell.alignment = top
def _apply_result_task_plan_end_answer_spec_16_no_highlight(
    worksheet, column_names: list
):
    """
    列「納期を満たすか？」は「いいえ」のセルを赤背景・白文字・太字にれる。
    列設定で旧名「配台済_回答指定16時まで」等の見出しにも対応。
    """
    target_names = frozenset(
        {
            RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16,
            "配台済_基準16時まで",
            "配台済_回答指定16時まで",
            "配完_回答指定16時まで",
            "配完_基準16時まで",
        }
    )
    col_idx = None
    for ci, col_name in enumerate(column_names, 1):
        if str(col_name) in target_names:
            col_idx = ci
            break
    if col_idx is None or worksheet.max_row < 2:
        return
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white_bold = _result_font(color="FFFFFF", bold=True)
    top = Alignment(wrap_text=False, vertical="top")
    for r in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if s != "いいえ":
            continue
        cell.fill = fill_red
        cell.font = font_white_bold
        cell.alignment = top
def _apply_result_task_id_hyperlinks_to_equipment_schedule(
    worksheet_tasks,
    column_names: list,
    sorted_tasks_for_row_order: list,
    task_id_to_schedule_cell: dict[str, str],
    schedule_sheet_name: str,
) -> None:
    """
    結果_タスク一覧の「タスクID」セルに」結果_設備毎の時間割で当該タスクは最初に睾れるセルへの内部ポイパーリンクを付与れる。
    時間割に睾れないタスク（未割当のみ等）はリンクなし。
    """
    if not task_id_to_schedule_cell or worksheet_tasks.max_row < 2:
        return
    task_id_col_idx = None
    for col_idx, col_name in enumerate(column_names, 1):
        if str(col_name) == "タスクID":
            task_id_col_idx = col_idx
            break
    if task_id_col_idx is None:
        return
    esc = schedule_sheet_name.replace("'", "''")
    loc_prefix = f"#'{esc}'!"
    font_link = _result_font(color="0563C1", underline="single")
    font_link_on_red = _result_font(color="FFFFFF", underline="single")
    top = Alignment(wrap_text=False, vertical="top")
    n_tasks = len(sorted_tasks_for_row_order)
    for r in range(2, worksheet_tasks.max_row + 1):
        cell = worksheet_tasks.cell(row=r, column=task_id_col_idx)
        raw = cell.value
        if raw is None:
            continue
        tid = str(raw).strip()
        if not tid:
            continue
        addr = task_id_to_schedule_cell.get(tid)
        if not addr:
            continue
        cell.hyperlink = loc_prefix + addr
        row_i = r - 2
        mismatch = (
            row_i < n_tasks
            and bool(sorted_tasks_for_row_order[row_i].get("process_content_mismatch"))
        )
        cell.font = font_link_on_red if mismatch else font_link
        cell.alignment = top
def _add_column_config_sheet_helpers(ws_cfg, num_data_rows: int):
    """表示列に TRUE/FALSE リスト（チェックの代ゝりにプルダウン）を付与。"""
    last_r = max(num_data_rows + 1, 2)
    cap = max(last_r + 50, 500)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws_cfg.add_data_validation(dv)
    dv.add(f"B2:B{cap}")
def _stage2_try_copy_column_config_shapes_from_input(
    result_path: str,
    input_path: str | None,
) -> None:
    """
    旧実装は Excel COM で列設定シートの図形を複製していた。
    COM 経路廃止により **未対応**（常にスキップ）。openpyxl のみではシート上の
    Shapes/OLE を安全に複製できないため。
    """
    if not STAGE2_COPY_COLUMN_CONFIG_SHAPES_FROM_INPUT:
        return
    logging.info(
        "列設定シート図形コピー: Excel COM を廃止したためスキップしました（手動コピーまたはマクロで代替してください）。"
    )
def _hex_rrggbb_to_rgb_triple(hx: str) -> tuple[int, int, int]:
    """6 桁 RRGGBB（# 可）を (R,G,B) に。不正時は中間グレー。"""
    s = (hx or "").strip().lstrip("#").upper()
    if len(s) != 6 or any(c not in "0123456789ABCDEF" for c in s):
        return (180, 180, 180)
    return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
def _stage2_try_add_gantt_timeline_shape_labels(
    result_path: str,
    specs: list | None,
    day_blocks: list | None = None,
    *,
    sheet_name: str | None = None,
) -> None:
    """
    openpyxl 保存後、GANTT_TIMELINE_SHAPE_LABELS が有効で specs があればタイムライン先頭列にラベルを書き込む。
    （旧角丸シェイプ・画像化は廃止。day_blocks は無視される。）
    """
    if not GANTT_TIMELINE_SHAPE_LABELS or not specs:
        return
    rp = (result_path or "").strip()
    if not rp or not os.path.isfile(rp):
        return
    shn = sheet_name or RESULT_SHEET_GANTT_NAME
    try:
        _gantt_fallback_timeline_labels_openpyxl(rp, specs, sheet_name=shn)
        logging.info(
            "%s: タイムラインラベルをセル表記で追加しました（%s 件）。",
            shn,
            len(specs),
        )
    except Exception as e:
        logging.warning(
            "%s: タイムラインラベルのセル書込に失敗しました（%s）。",
            shn,
            e,
        )
def _coerce_actual_sheet_datetime(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, time(0, 0))
    try:
        ts = pd.to_datetime(val, errors="coerce")
        if pd.isna(ts) or ts is pd.NaT:
            return None
        if isinstance(ts, pd.Timestamp):
            return ts.to_pydatetime()
        return ts if isinstance(ts, datetime) else None
    except Exception:
        return None
def _compose_hm_to_time(hour_val, minute_val) -> time | None:
    """問合せ export の「開始時間」「開始分」等から time を組み立てる。"""
    if hour_val is None or minute_val is None:
        return None
    if isinstance(hour_val, float) and pd.isna(hour_val):
        return None
    if isinstance(minute_val, float) and pd.isna(minute_val):
        return None
    try:
        h = int(float(str(hour_val).strip()))
        m = int(float(str(minute_val).strip()))
    except (TypeError, ValueError):
        return None
    if 0 <= h <= 23 and 0 <= m <= 59:
        return time(h, m)
    return None
def _normalize_actual_detail_workbook_columns(df: pd.DataFrame) -> pd.DataFrame:
    """NO(ロット)別問合せ xlsx 等の列名を加工実績明細DATA 正規名へ寄せる。"""
    if df is None or getattr(df, "empty", True):
        return df
    rename: dict[str, str] = {}
    if "加工日" in df.columns and ACT_COL_DAY not in df.columns:
        rename["加工日"] = ACT_COL_DAY
    if "停機時間分換算" in df.columns and ACT_COL_STOP_MIN_CONVERTED not in df.columns:
        rename["停機時間分換算"] = ACT_COL_STOP_MIN_CONVERTED
    if rename:
        df = df.rename(columns=rename)
    return df
def _actual_row_time_bounds(row):
    """加工実績DATA／加工実績明細DATA の1行から (開始, 終了) を得る。解けなければ (None, None)。"""
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_START_DT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_END_DT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt
    # 実績明細ガントは「加工開始日時(停機時間加算後)」を優先（無い場合は後段で従来列へフォールバック）
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_MACHINING_START_DT_WITH_STOP))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_MACHINING_END_DT))
    if s_dt and e_dt:
        if e_dt < s_dt:
            s_dt = e_dt - timedelta(minutes=5)
        if s_dt < e_dt:
            return s_dt, e_dt
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_MACHINING_START_DT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_MACHINING_END_DT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt
    s_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_START_ALT))
    e_dt = _coerce_actual_sheet_datetime(row.get(ACT_COL_END_ALT))
    if s_dt and e_dt and s_dt < e_dt:
        return s_dt, e_dt

    d_date = parse_optional_date(row.get(ACT_COL_DAY))
    if not d_date:
        d_date = parse_optional_date(row.get("加工日"))
    if not d_date:
        cd = _coerce_actual_sheet_datetime(row.get(ACT_COL_DAY))
        if isinstance(cd, datetime):
            d_date = cd.date()
        elif isinstance(cd, date):
            d_date = cd
    if not d_date:
        cd = _coerce_actual_sheet_datetime(row.get("加工日"))
        if isinstance(cd, datetime):
            d_date = cd.date()
        elif isinstance(cd, date):
            d_date = cd
    if not d_date:
        return None, None

    t0 = _compose_hm_to_time(row.get("開始時間"), row.get("開始分"))
    t1 = _compose_hm_to_time(row.get("終了時間"), row.get("終了分"))
    if t0 is not None and t1 is not None and t0 < t1:
        return datetime.combine(d_date, t0), datetime.combine(d_date, t1)

    ts_s = row.get(ACT_COL_TIME_START)
    ts_e = row.get(ACT_COL_TIME_END)
    if ts_s is None or pd.isna(ts_s) or ts_e is None or pd.isna(ts_e):
        return None, None

    if isinstance(ts_s, time):
        t0 = ts_s
    elif isinstance(ts_s, datetime):
        t0 = ts_s.time()
    else:
        t0 = parse_time_str(ts_s, None)

    if isinstance(ts_e, time):
        t1 = ts_e
    elif isinstance(ts_e, datetime):
        t1 = ts_e.time()
    else:
        t1 = parse_time_str(ts_e, None)

    if t0 is None or t1 is None or t0 >= t1:
        return None, None
    return datetime.combine(d_date, t0), datetime.combine(d_date, t1)
def load_machining_actuals_df():
    """
    「加工実績DATA」を読む（無ければ空 DataFrame）。

    優先: PM_AI_ACTUALS_DATA_WORKBOOK。
    未指定時は実績明細と同じ既定探索（PM_AI_ACTUAL_DETAIL_WORKBOOK、
    PM_AI_ACTUAL_DETAIL_SOURCE_DIR 内の最新 xlsx/xlsm、TASK_INPUT_WORKBOOK）。
    シートは PM_AI_ACTUALS_DATA_SHEET（省略時は先頭シート index 0。単一シートなら名前不要）。
    """
    _src = resolve_actuals_workbook_path(_excel_plan_input_wb())
    if not _src or not os.path.exists(_src):
        return pd.DataFrame()
    _sn = _excel_sheet_arg_from_env(ENV_PM_AI_ACTUALS_DATA_SHEET)
    _lbl = _excel_sheet_label_for_log(_sn, ACTUALS_SHEET_NAME)
    try:
        df = pd.read_excel(_src, sheet_name=_sn)
    except ValueError:
        logging.info(
            "シート「%s」は無いため、ガントの実績行は出力しません。",
            _lbl,
        )
        return pd.DataFrame()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, ACTUAL_HEADER_CANONICAL)
    logging.info("加工実績: '%s' の '%s' を %s 行読み込み。", _src, _lbl, len(df))
    return df
def _calendar_dates_spanned_by_actual_bounds_df(df) -> set[date]:
    """
    実績明細等の各行の (開始, 終了) が跨ぐ暦日を収集する。
    計画の sorted_dates に含まれない過去日の実績もガントに載せるために使う。
    """
    out: set[date] = set()
    if df is None or len(df) == 0:
        return out
    for _, row in df.iterrows():
        s_dt, e_dt = _actual_row_time_bounds(row)
        if not s_dt or not e_dt or s_dt >= e_dt:
            continue
        d0 = s_dt.date()
        d1 = e_dt.date()
        cur = d0
        while cur <= d1:
            out.add(cur)
            cur += timedelta(days=1)
    return out
def _sorted_dates_union_actual_bounds_df(sorted_dates: list, df) -> list:
    """計画表示日と実績行の暦日の和集合（昇順）。"""
    u = set(sorted_dates)
    u |= _calendar_dates_spanned_by_actual_bounds_df(df)
    return sorted(u)
def _sorted_dates_filter_inclusive_range(
    sorted_dates: list,
    d_from: date | None,
    d_to: date | None,
) -> list:
    """
    暦日リストを両端込みで絞る。d_from / d_to がともに None のときはコピーを返す。
    両端指定で from > to のときは from/to を入れ替える。
    """
    if d_from is None and d_to is None:
        return list(sorted_dates)
    a = d_from
    b = d_to
    if a is not None and b is not None and a > b:
        a, b = b, a
    out: list = []
    for d in sorted_dates:
        if a is not None and d < a:
            continue
        if b is not None and d > b:
            continue
        out.append(d)
    return out
def load_machining_actual_detail_df():
    """
    「加工実績明細DATA」を読む（無ければ空 DataFrame）。

    読込は calamine（python-calamine）を優先し、未導入・その他失敗時は openpyxl。スタイル定義の不整合で
    openpyxl だけが落ちる xlsx があるため（dispatch_workspace の Excel 先読と同趣旨）。

    優先: PM_AI_ACTUAL_DETAIL_WORKBOOK（単一ファイル）、PM_AI_ACTUAL_DETAIL_SOURCE_DIR 内の最新 xlsx/xlsm
    （既定 UNC は plan/02 と同系）、最後に TASK_INPUT_WORKBOOK 内シート。
    シートは PM_AI_ACTUAL_DETAIL_SHEET（省略時は先頭シート index 0。単一シートなら名前不要）。
    列は加工実績DATA に準じ、ロール識別は「ロールNO」「ロール番号」「ロール」「巻番」のいずれか可。
    """
    _src_wb = resolve_actual_detail_workbook_path(_excel_plan_input_wb())
    if not _src_wb:
        return pd.DataFrame()
    _sn = _excel_sheet_arg_from_env(ENV_PM_AI_ACTUAL_DETAIL_SHEET)
    _lbl = _excel_sheet_label_for_log(_sn, ACTUAL_DETAIL_SHEET_NAME)

    def _load_once():
        try:
            resolved = _resolve_tabular_sheet_name_calamine(_src_wb, _sn)
            hdr = _resolve_tabular_excel_header_row_0based(_src_wb, resolved)
            out = _read_excel_tabular(_src_wb, resolved, header=hdr)
        except ValueError:
            logging.info(
                "シート「%s」は無いため、実績明細ガントは出力しません。",
                _lbl,
            )
            return pd.DataFrame()
        if out is None or getattr(out, "empty", True):
            return pd.DataFrame()
        out.columns = out.columns.str.strip()
        out = _normalize_actual_detail_workbook_columns(out)
        if ACT_DETAIL_COL_ROLL not in out.columns:
            for alias in ("ロール番号", "ロール", "巻番"):
                if alias in out.columns:
                    out = out.rename(columns={alias: ACT_DETAIL_COL_ROLL})
                    break
        out = _align_dataframe_headers_to_canonical(out, ACTUAL_DETAIL_HEADER_CANONICAL)
        try:
            if ACT_COL_MACHINING_START_DT in out.columns:
                s0 = pd.to_datetime(out[ACT_COL_MACHINING_START_DT], errors="coerce")
            else:
                s0 = pd.Series([pd.NaT] * len(out))
            if ACT_COL_STOP_MIN_CONVERTED in out.columns:
                stop_min = pd.to_numeric(
                    out[ACT_COL_STOP_MIN_CONVERTED], errors="coerce"
                ).fillna(0.0)
            else:
                stop_min = pd.Series([0.0] * len(out))
            s1 = s0 + pd.to_timedelta(stop_min, unit="m")

            if ACT_COL_MACHINING_END_DT in out.columns:
                e0 = pd.to_datetime(out[ACT_COL_MACHINING_END_DT], errors="coerce")
                mask = e0.notna() & s1.notna() & (e0 < s1)
                if mask.any():
                    s1 = s1.where(~mask, e0 - pd.Timedelta(minutes=5))

            out[ACT_COL_MACHINING_START_DT_WITH_STOP] = s1
        except Exception as e:
            logging.warning(
                "加工実績明細: %s 列の生成に失敗したため従来列で続行します（%s）。",
                ACT_COL_MACHINING_START_DT_WITH_STOP,
                e,
            )
        logging.info(
            "加工実績明細: '%s' の '%s' を %s 行読み込み。",
            _src_wb,
            _lbl,
            len(out),
        )
        return out

    try:
        return _cached_tabular_dataframe("actual_detail", _src_wb, _load_once)
    except ValueError:
        return pd.DataFrame()
def _actual_row_cumulative_completion_pct_macro(row) -> int | None:
    """
    加工実績明細DATA の「累積完了率」をシート値のまま 0～100 の整数に解釈して返す（実÷予定等は計算しない）。

    対応: 数値・「45」「45%」「45.5%」・Excel 割合セル由来の 0.45（=45%）など。
    列が無い・空・数値化不可のときは None。
    """
    if row is None:
        return None
    v = row.get(ACT_COL_CUMULATIVE_COMPLETION_PCT)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return None
    try:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            f = float(v)
        else:
            s = unicodedata.normalize("NFKC", str(v).strip())
            if not s or s.lower() in ("nan", "none", "-", "—", "―"):
                return None
            s = s.replace("%", "").replace(",", "").strip()
            if not s:
                return None
            f = float(s)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f) or f < 0:
        return None
    # Excel の「割合」表示は 0.45 のように小数で渡ることが多い（= 45%）
    if f <= 1.0 + 1e-9:
        pct = int(round(f * 100.0))
    else:
        pct = int(round(f))
    return max(0, min(100, pct))
def _actual_row_detail_assignee_op_sub(row) -> tuple[str, str]:
    """
    加工実績明細DATA 行からガント用 op / sub を組み立てる。
    「担当者」に続けて「加工担当者名1」～「加工担当者名5」を順に見て非空のみ採用し、
    NFKC 後の文字列で重複を除く。先頭を op、2人目以降を sub（カンマ区切り）。
    """
    names: list[str] = []
    seen_k: set[str] = set()
    for col in (ACT_COL_OPERATOR,) + ACT_COL_MACHINING_ASSIGNEES_ORDERED:
        val = row.get(col)
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        s = str(val).strip()
        if not s:
            continue
        k = unicodedata.normalize("NFKC", s)
        if k in seen_k:
            continue
        seen_k.add(k)
        names.append(s)
    if not names:
        return "", ""
    if len(names) == 1:
        return names[0], ""
    return names[0], ", ".join(names[1:])
def _actual_detail_prefer_actual_for_shape_label(
    roll_detail: bool,
    cumulative_actual_m,
    actual_done_m,
) -> bool:
    """
    加工実績明細（roll_detail）で、「累積実績」が当行の「実加工数」を明確に上回るときは、
    累積を依頼全体の走り・実加工数を当区間（日次・セグメント）の量と解釈し、
    ガント角丸シェイプの m 表示（label_len_m）は実加工数の時間按分を優先する。
    """
    if not roll_detail:
        return False
    try:
        af = float(actual_done_m) if actual_done_m is not None else 0.0
        cf = float(cumulative_actual_m) if cumulative_actual_m is not None else 0.0
        return af > 1e-12 and cf > af * 1.000001
    except (TypeError, ValueError):
        return False
def _normalize_roll_detail_daily_actual_qty_duplicate(events: list) -> None:
    """
    加工実績明細では、複数明細行に同一暦日のトータル「実加工数」が繰り返して入っていることがある。
    その場合、行単位で compare_daily_m / label_len_m を時間按分すると日次・依頼単位で合計が過大になるため、
    （依頼NO(task_id)×暦日×機械）ごとに「ソース実加工数」がすべて同一のグループだけ、
    その値をその日その依頼の総量として 1 回だけ数え、イベント区間長で再按分する。
    """
    if not events:
        return
    from collections import defaultdict

    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, ev in enumerate(events):
        if ev.get("_detail_source_actual_m") is None:
            continue
        tid = str(ev.get("task_id") or "").strip()
        mach = str(ev.get("machine") or "").strip()
        dd = ev.get("date")
        try:
            dk = dd.isoformat() if hasattr(dd, "isoformat") else str(dd)
        except Exception:
            dk = ""
        groups[(tid, dk, mach)].append(i)

    tol = lambda a, b, ref: abs(float(a) - float(b)) <= 1e-6 * max(1e-12, abs(float(ref)))

    for _key, idxs in groups.items():
        if len(idxs) < 2:
            continue
        qs = []
        for ii in idxs:
            q = events[ii].get("_detail_source_actual_m")
            try:
                qs.append(float(q))
            except (TypeError, ValueError):
                qs = []
                break
        if len(qs) != len(idxs):
            continue
        ref_q = qs[0]
        if ref_q <= 1e-12:
            continue
        if not all(tol(q, ref_q, ref_q) for q in qs):
            continue
        V_daily = ref_q
        secs: list[float] = []
        tot_sec = 0.0
        for ii in idxs:
            ev = events[ii]
            st, ed = ev.get("start_dt"), ev.get("end_dt")
            if isinstance(st, datetime) and isinstance(ed, datetime) and st < ed:
                sec = float((ed - st).total_seconds())
            else:
                sec = 0.0
            secs.append(sec)
            tot_sec += sec
        if tot_sec < 1e-9:
            continue
        for ii, sec in zip(idxs, secs):
            portion = float(V_daily) * (sec / tot_sec)
            events[ii]["compare_daily_m"] = portion
            events[ii]["label_len_m"] = portion
            events[ii]["_detail_daily_qty_total_m"] = float(V_daily)
            events[ii].pop("label_len_m_is_cumulative", None)

    for ev in events:
        ev.pop("_detail_source_actual_m", None)
def build_actual_timeline_events(
    df,
    equipment_list,
    sorted_dates,
    *,
    log_sheet_name: str = "加工実績DATA",
    roll_detail: bool = False,
):
    """
    実績シートの各行をガント用イベントへ変換。
    計画表示日（sorted_dates）かつ設備マスタに一致する「工程名」の値が対象。
    工程名は NFKC・空白正規化後にマスタ列名へマッピングする。
    時刻は DEFAULT_START_TIME / DEFAULT_END_TIME の枠内にクリップ。
    roll_detail=True のとき ACT_DETAIL_COL_ROLL があれば task_id を「依頼NO/ロール」表記にし帯の分離に使う。
    同じく roll_detail=True のときは「担当者」および「加工担当者名1」～「加工担当者名5」を
    ガントの op/sub（タイムライン氏名チップ・D列要約）へ反映する。
    roll_detail=True のとき「累積完了率」列があればその値を解釈し、
    タイムライン角丸シェイプのラベル（依頼NO の横の %%）に ``pct_macro`` として渡す（計算はしない）。
    計画実績比較ガント用に、各日セグメントへ ``compare_daily_m``（実加工数の時間比按分を優先、
    無ければ累積実績の時間比按分）を付与する。``label_len_m`` に累積を載せても日別比較で二重計上しない。
    roll_detail=True のとき、累積実績が当行の実加工数を明確に上回る場合は ``label_len_m`` も
    実加工数の時間按分を優先する（累積のみ依頼全体・実加工数が日次・区間のトータルなデータ向け）。
    同一依頼NO×暦日×機械でソース実加工数がすべて同一の複数イベントは、日次トータルを二重計上しないよう
    生成後に ``compare_daily_m`` および ``label_len_m`` を区間長で再按分し、
    当該イベントの ``label_len_m_is_cumulative`` は解除する。
    ガントシェイプ文言は ``_detail_daily_qty_total_m`` に日次総量を載せ、結合セグメントでは按分値の合計ではなく総量表示に使う。
    """
    if df is None or len(df) == 0:
        return []
    equip_lookup = _equipment_lookup_normalized_to_canonical(equipment_list)
    date_ok = set(sorted_dates)
    events = []
    bad_eq = 0
    bad_time = 0
    no_plan_overlap = 0
    mismatch_norm_samples = []

    for _, row in df.iterrows():
        tid = row.get(ACT_COL_TASK_ID)
        if tid is None or pd.isna(tid):
            continue
        tid_s = str(tid).strip()
        if not tid_s:
            continue
        display_tid = tid_s
        if roll_detail:
            rv = row.get(ACT_DETAIL_COL_ROLL)
            if rv is not None and not (isinstance(rv, float) and pd.isna(rv)):
                rs = str(rv).strip()
                if rs:
                    display_tid = f"{tid_s}/{rs}"
        proc = row.get(ACT_COL_PROCESS)
        if proc is None or pd.isna(proc):
            continue
        proc_key = _normalize_equipment_match_key(proc)
        mach = equip_lookup.get(proc_key)
        if not mach:
            bad_eq += 1
            if len(mismatch_norm_samples) < 12 and proc_key:
                if proc_key not in mismatch_norm_samples:
                    mismatch_norm_samples.append(proc_key)
            continue
        start_dt, end_dt = _actual_row_time_bounds(row)
        if not start_dt or not end_dt or start_dt >= end_dt:
            bad_time += 1
            continue
        if roll_detail:
            op_s, sub_s = _actual_row_detail_assignee_op_sub(row)
        else:
            op_val = row.get(ACT_COL_OPERATOR)
            op_s = ""
            if op_val is not None and not pd.isna(op_val):
                op_s = str(op_val).strip()
            sub_s = ""

        pct_macro = _actual_row_cumulative_completion_pct_macro(row)
        actual_done_m = None
        cumulative_actual_m = None
        try:
            actual_done_m = parse_float_safe(row.get(ACT_COL_ACTUAL_QTY), None)
        except Exception:
            actual_done_m = None
        try:
            cumulative_actual_m = parse_float_safe(row.get(ACT_COL_CUMULATIVE_ACTUAL_QTY), None)
        except Exception:
            cumulative_actual_m = None

        before = len(events)
        total_seconds = None
        try:
            total_seconds = max(0.0, float((end_dt - start_dt).total_seconds()))
        except Exception:
            total_seconds = None
        prefer_actual_label = _actual_detail_prefer_actual_for_shape_label(
            roll_detail, cumulative_actual_m, actual_done_m
        )
        for d in sorted_dates:
            if d not in date_ok:
                continue
            day_start = datetime.combine(d, DEFAULT_START_TIME)
            day_end = datetime.combine(d, DEFAULT_END_TIME)
            if end_dt <= day_start or start_dt >= day_end:
                continue
            s_clip = max(start_dt, day_start)
            e_clip = min(end_dt, day_end)
            if s_clip >= e_clip:
                continue
            ev_row = {
                "date": d,
                "task_id": display_tid,
                "machine": mach,
                "op": op_s,
                "sub": sub_s,
                "start_dt": s_clip,
                "end_dt": e_clip,
                "breaks": [],
                "units_done": 0,
                "already_done_units": 0,
                "total_units": 0,
                "eff_time_per_unit": 0.0,
                "unit_m": 0.0,
            }
            if pct_macro is not None:
                ev_row["pct_macro"] = pct_macro
            seg_seconds = float((e_clip - s_clip).total_seconds())
            # 依頼NOシェイプ横の表示: 既定は「累積実績(m)」優先。
            # 加工実績明細で累積が当行の実加工数を大きく上回るときは、実加工数を区間・日次の量とみなし先に按分する。
            # それ以外で累積が無い/不正なときは「実加工数」を時間比で按分。
            try:
                if prefer_actual_label:
                    if (
                        actual_done_m is not None
                        and isinstance(actual_done_m, (int, float))
                        and float(actual_done_m) > 1e-12
                        and total_seconds
                        and float(total_seconds) > 1e-9
                        and seg_seconds > 0
                    ):
                        ev_row["label_len_m"] = float(actual_done_m) * (
                            seg_seconds / float(total_seconds)
                        )
                elif (
                    cumulative_actual_m is not None
                    and isinstance(cumulative_actual_m, (int, float))
                    and float(cumulative_actual_m) > 1e-12
                ):
                    ev_row["label_len_m"] = float(cumulative_actual_m)
                    ev_row["label_len_m_is_cumulative"] = True
                elif (
                    actual_done_m is not None
                    and isinstance(actual_done_m, (int, float))
                    and float(actual_done_m) > 1e-12
                    and total_seconds
                    and float(total_seconds) > 1e-9
                ):
                    if seg_seconds > 0:
                        ev_row["label_len_m"] = float(actual_done_m) * (
                            seg_seconds / float(total_seconds)
                        )
            except Exception:
                pass
            # 計画実績比較ガントのアラジン日次数量との突き合わせ用。
            # 累積ラベルは全日セグメントに同一値が載るため label_len_m を日別に足すと過大になる。
            # 実加工数を優先して日×セグメントに按分し、無いときのみ累積を同一比率で按分する。
            try:
                cmp_m = None
                if (
                    actual_done_m is not None
                    and isinstance(actual_done_m, (int, float))
                    and float(actual_done_m) > 1e-12
                    and total_seconds
                    and float(total_seconds) > 1e-9
                    and seg_seconds > 1e-9
                ):
                    cmp_m = float(actual_done_m) * (
                        seg_seconds / float(total_seconds)
                    )
                elif (
                    cumulative_actual_m is not None
                    and isinstance(cumulative_actual_m, (int, float))
                    and float(cumulative_actual_m) > 1e-12
                    and total_seconds
                    and float(total_seconds) > 1e-9
                    and seg_seconds > 1e-9
                ):
                    cmp_m = float(cumulative_actual_m) * (
                        seg_seconds / float(total_seconds)
                    )
                if cmp_m is not None and cmp_m > 1e-12:
                    ev_row["compare_daily_m"] = cmp_m
            except Exception:
                pass
            if roll_detail:
                try:
                    if (
                        actual_done_m is not None
                        and isinstance(actual_done_m, (int, float))
                        and float(actual_done_m) > 1e-12
                    ):
                        ev_row["_detail_source_actual_m"] = float(actual_done_m)
                except Exception:
                    pass
            events.append(ev_row)
        if len(events) == before:
            no_plan_overlap += 1

    if bad_eq:
        logging.warning(
            f"{log_sheet_name}: 工程名はマスタ設備と一致しない行を {bad_eq} 件スキップしました（空白等は正規化済み）。"
        )
        if mismatch_norm_samples:
            logging.info(
                "  厳密一致となった工程名の正規化後サンプル: "
                + " | ".join(mismatch_norm_samples[:12])
            )
    if bad_time:
        logging.info(
            f"{log_sheet_name}: 開始/終了日時を解釈できない行を {bad_time} 件スキップしました。"
        )
    if no_plan_overlap and sorted_dates:
        logging.info(
            f"{log_sheet_name}: 設備・日時は有効だが計画対象日（勤怠日×{DEFAULT_START_TIME}～{DEFAULT_END_TIME}）と重ならない行が {no_plan_overlap} 件ありました。"
        )
    if not events and len(df) > 0:
        logging.info(
            f"{log_sheet_name}: ガント用セグメントは0件です。表示日（sorted_dates）に重ならない実績のみの場合、描画されません。"
        )
    logging.info(f"{log_sheet_name} からガント用セグメント {len(events)} 件を生成しました。")
    if roll_detail:
        _normalize_roll_detail_daily_actual_qty_duplicate(events)
    return events
TASK_SPECIAL_AI_LAST_RESPONSE_FILE = "ai_task_special_remark_last.txt"
TASK_SPECIAL_CACHE_KEY_PREFIX = "TASK_SPECIAL_v3|"
GLOBAL_PRIORITY_OVERRIDE_CACHE_PREFIX = "GLOBAL_PRIO_v8|"
def _normalize_special_task_id_for_ai(val):
    """
    依頼NOをキャッシュキー・プロンプト行で一貫させる。
    Excel の数値セルは float になりはうなので 12345.0 → \"12345\" に权ごる。
    文字列は NFKC（全角英数字など）で表記ゆれを坸坎（同一実体の再API呼び出しを減られ）。
    """
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except TypeError:
        pass
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if math.isnan(val):
            return None
        if val.is_integer():
            return str(int(val))
        s = str(val).strip()
        if not s or s.lower() in ("nan", "none", "null"):
            return None
        return unicodedata.normalize("NFKC", s).strip() or None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    s = unicodedata.normalize("NFKC", s).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    # 文字列としての "20010.0" 等（Excel・CSV）を整数表記の依頼NOに寄せる
    if re.fullmatch(r"-?\d+\.0+", s):
        try:
            return str(int(float(s)))
        except ValueError:
            pass
    return s or None
def planning_task_id_str_from_scalar(val) -> str:
    """配台・段階1マージ・キュー構築で用いる依頼NOの安定文字列（空なら \"\"）。"""
    return _normalize_special_task_id_for_ai(val) or ""
def planning_task_id_str_from_plan_row(row) -> str:
    """重複見出し列でも先頭スカラーを拾い」依頼NOを planning_task_id_str_from_scalar に渡す。"""
    return planning_task_id_str_from_scalar(_planning_df_cell_scalar(row, TASK_COL_TASK_ID))
def _cell_text_task_special_remark(val):
    """
    特別指定_備考をプロンプト用に取り出す。仕様どより **strip のみ**
    （先頭末尾の空白・Excel の坽空白を除し」文中の改行・スペースは保挝。数値セルは表記を固定）。
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except TypeError:
        pass
    if isinstance(val, bool):
        s = str(val)
    elif isinstance(val, float):
        if math.isnan(val):
            return ""
        # 備考列に数値の値入っている場合の表記ゆれを減られ
        if val.is_integer():
            s = str(int(val))
        else:
            s = str(val)
    elif isinstance(val, int):
        s = str(val)
    else:
        s = str(val)
    s = s.strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s
def _coalesce_plan_plain_remark_into_special(df):
    """
    シートに単独の「備考」列だけあり「特別指定_備考」が空／欠落のとき、AI 入力列へ繰り寄せる。
    NFKC で「備考」と一致する最初の列名を対象とする。
    """
    if df is None or getattr(df, "empty", True):
        return df
    plain_col = None
    for c in df.columns:
        if _nfkc_column_aliases(str(c).strip()) == _nfkc_column_aliases("備考"):
            plain_col = str(c).strip()
            break
    if plain_col is None:
        return df
    spec = PLAN_COL_SPECIAL_REMARK
    if spec not in df.columns:
        return df.rename(columns={plain_col: spec})
    merged = 0
    conflict = False
    for i in df.index:
        sp = df.at[i, spec]
        pl = df.at[i, plain_col]
        sp_txt = _cell_text_task_special_remark(sp)
        pl_txt = _cell_text_task_special_remark(pl)
        if sp_txt:
            if pl_txt and sp_txt != pl_txt:
                conflict = True
            continue
        if pl_txt:
            df.at[i, spec] = pl
            merged += 1
    if conflict:
        logging.warning(
            "計画タスク入力: 「備考」と「特別指定_備考」の両方に異なる値がある行があります。"
            "AI 適用は「特別指定_備考」を優先します。「備考」列は削除しません。"
        )
        return df
    df = df.drop(columns=[plain_col])
    if merged > 0:
        logging.info(
            "計画タスク入力: 「備考」→「特別指定_備考」へ %s セルを転記しました（単独備考列の互換）。",
            merged,
        )
    return df
def _task_special_prompt_lines(tasks_df):
    """プロンプトに載せる行リスト（ソート剝）。正規化は上記ヘルパーに統一。"""
    lines = []
    for _, row in tasks_df.iterrows():
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
        if not tid or not rem:
            continue
        proc = str(row.get(TASK_COL_MACHINE, "") or "").strip()
        macn = str(row.get(TASK_COL_MACHINE_NAME, "") or "").strip()
        proc_disp = proc if proc else "（空）"
        macn_disp = macn if macn else "（空）"
        lines.append(
            f"- 依頼NO」{tid}】| 工程名「{proc_disp}」 | 機械名「{macn_disp}」 | 備考本文: {rem}"
        )
    return lines
def _repair_task_special_ai_wrong_top_level_keys(parsed: dict, tasks_df) -> dict:
    """
    備考は哝番・原板コード（例: 20010 で始まる数字列）で始まると」モデルはしの列を JSON トップキーに
    誤用れることはある。依頼NO」…】と一致しない数字のみのキーを」当該備考を挝つ行の依頼NOへ付け替ごる。
    """
    if not isinstance(parsed, dict) or not parsed or tasks_df is None or getattr(tasks_df, "empty", True):
        return parsed
    valid_tids: set[str] = set()
    remark_by_tid: dict[str, list[str]] = {}
    for _, row in tasks_df.iterrows():
        if _plan_row_exclude_from_assignment(row):
            continue
        tid = planning_task_id_str_from_plan_row(row)
        rem = _cell_text_task_special_remark(row.get(PLAN_COL_SPECIAL_REMARK))
        if not tid or not rem:
            continue
        valid_tids.add(tid)
        remark_by_tid.setdefault(tid, []).append(rem)

    for bad_key in list(parsed.keys()):
        sk = str(bad_key).strip()
        if sk in valid_tids:
            continue
        if not re.fullmatch(r"\d{4,}", sk):
            continue
        hits = [
            tid
            for tid, rems in remark_by_tid.items()
            if any(
                r.startswith(sk)
                or r.startswith(sk + " ")
                or r.startswith(sk + "\u3000")
                or r.startswith(sk + "-")
                or r.startswith(sk + "ー")
                for r in rems
            )
        ]
        if len(hits) != 1:
            continue
        target = hits[0]
        val = parsed.pop(bad_key, None)
        if val is None:
            continue
        if target not in parsed:
            parsed[target] = val
            logging.info(
                "タスク特別指定: JSON トップキー誤りを修復（%r は依頼NOではない → %r）",
                bad_key,
                target,
            )
        else:
            parsed[bad_key] = val
    return parsed
def _normalize_task_special_scope_str(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return unicodedata.normalize("NFKC", str(s).strip())
def _task_special_scope_matches_row_field(row_val, restrict_val) -> bool:
    """
    restrict は無い・空なら制限なし（True）。
    非空なら Excel 坴の値とあいまい一致（部分一致坯）。
    """
    if restrict_val is None:
        return True
    r = _normalize_task_special_scope_str(restrict_val)
    if not r:
        return True
    v = _normalize_task_special_scope_str(row_val)
    if not v:
        return False
    if v == r:
        return True
    if r in v or v in r:
        return True
    return False
def _ai_remark_entry_applies_to_row(entry: dict, row) -> bool:
    """restrict_to_* は無いとしは同一依頼NOの全行に適用。"""
    if not isinstance(entry, dict):
        return False
    rp = row.get(TASK_COL_MACHINE, "")
    rm = row.get(TASK_COL_MACHINE_NAME, "")
    if not _task_special_scope_matches_row_field(rp, entry.get("restrict_to_process_name")):
        return False
    if not _task_special_scope_matches_row_field(rm, entry.get("restrict_to_machine_name")):
        return False
    return True
def _row_matches_remark_source_row(entry: dict, row) -> bool:
    """
    JSON の process_name / machine_name は」当該 Excel 行の工程名・機械名と一致するか。
    （プロンプトで渡した「備考はあった行」と対応るける。片方の値一致でも坯）
    """
    if not isinstance(entry, dict):
        return False
    rp = _normalize_task_special_scope_str(row.get(TASK_COL_MACHINE))
    rm = _normalize_task_special_scope_str(row.get(TASK_COL_MACHINE_NAME))
    ep = _normalize_task_special_scope_str(entry.get("process_name"))
    em = _normalize_task_special_scope_str(entry.get("machine_name"))
    proc_ok = (not ep) or (not rp) or ep == rp or ep in rp or rp in ep
    mac_ok = (not em) or (not rm) or em == rm or em in rm or rm in em
    return proc_ok and mac_ok
def _entry_is_global_task_special_scope(entry: dict) -> bool:
    """restrict_to_* は無い・空＝同一依頼NOの全工程行に効かせる指定。"""
    if not isinstance(entry, dict):
        return False
    a = _normalize_task_special_scope_str(entry.get("restrict_to_process_name"))
    b = _normalize_task_special_scope_str(entry.get("restrict_to_machine_name"))
    return not a and not b
def _select_ai_task_special_entry_for_tid_value(val, row):
    """1依頼NOに対れる値は dict または dict の配列のどうらでも行に坈ご覝素を返す。"""
    if val is None:
        return None
    if isinstance(val, list):
        for item in val:
            if (
                isinstance(item, dict)
                and _ai_remark_entry_applies_to_row(item, row)
                and _row_matches_remark_source_row(item, row)
            ):
                return item
        for item in val:
            if (
                isinstance(item, dict)
                and _ai_remark_entry_applies_to_row(item, row)
                and _entry_is_global_task_special_scope(item)
            ):
                return item
        for item in val:
            if isinstance(item, dict) and _ai_remark_entry_applies_to_row(item, row):
                return item
        return None
    if isinstance(val, dict):
        if _ai_remark_entry_applies_to_row(val, row):
            return val
        return None
    return None
def _ai_task_special_entry_for_row(ai_by_tid, row):
    """
    analyze_task_special_remarks の戻りから当該行のエントリを得る。
    プロンプトキーは正規化済み依頼NOなので」Excel は 12345.0 でもヒットれる。
    restrict_to_process_name / restrict_to_machine_name は無い・空のときは
    同一依頼NOの工程・機械は異なる全行にも指示を適用する。
    """
    if not isinstance(ai_by_tid, dict) or not ai_by_tid:
        return None
    tid_norm = planning_task_id_str_from_plan_row(row)
    tid_raw = str(_planning_df_cell_scalar(row, TASK_COL_TASK_ID) or "").strip()

    def try_val(v):
        return _select_ai_task_special_entry_for_tid_value(v, row)

    if tid_norm and tid_norm in ai_by_tid:
        hit = try_val(ai_by_tid[tid_norm])
        if hit is not None:
            return hit
    if tid_raw:
        for key in (tid_raw, str(tid_raw)):
            if key in ai_by_tid:
                hit = try_val(ai_by_tid[key])
                if hit is not None:
                    return hit
    if tid_norm:
        for k, v in ai_by_tid.items():
            if str(k).strip() == tid_norm:
                hit = try_val(v)
                if hit is not None:
                    return hit
    if tid_raw:
        for k, v in ai_by_tid.items():
            if str(k).strip() == tid_raw:
                hit = try_val(v)
                if hit is not None:
                    return hit
    return None
def reset_gemini_usage_tracker() -> None:
    global _gemini_usage_session
    _gemini_usage_session = {}
def _save_gemini_cumulative_payload(data: dict) -> None:
    path = _gemini_cumulative_json_path()
    try:
        os.makedirs(api_payment_dir, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8", newline="\n") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except OSError as ex:
        logging.debug("Gemini 累計 JSON の保存に失敗: %s", ex)
def _append_gemini_cumulative_one_call(
    model_id: str, pt: int, ct: int, th: int, tt: int
) -> None:
    """1 回の API 応答分を累計 JSON に加算れる（ログに坘発料金は出さない）。"""
    mid = str(model_id).strip()
    data = _load_gemini_cumulative_payload()
    data["calls_total"] = int(data["calls_total"]) + 1
    data["prompt_total"] = int(data["prompt_total"]) + pt
    data["candidates_total"] = int(data["candidates_total"]) + ct
    data["thoughts_total"] = int(data["thoughts_total"]) + th
    data["total_tokens_reported"] = int(data["total_tokens_reported"]) + tt
    bm: dict = data["by_model"]
    if mid not in bm or not isinstance(bm[mid], dict):
        bm[mid] = {
            "calls": 0,
            "prompt": 0,
            "candidates": 0,
            "thoughts": 0,
            "total": 0,
            "estimated_cost_usd": 0.0,
        }
    m = bm[mid]
    m["calls"] = int(m.get("calls") or 0) + 1
    m["prompt"] = int(m.get("prompt") or 0) + pt
    m["candidates"] = int(m.get("candidates") or 0) + ct
    m["thoughts"] = int(m.get("thoughts") or 0) + th
    m["total"] = int(m.get("total") or 0) + tt
    inc_usd = _gemini_estimate_cost_usd(mid, pt, ct, th)
    if inc_usd is not None:
        m["estimated_cost_usd"] = float(m.get("estimated_cost_usd") or 0.0) + float(inc_usd)
        data["estimated_cost_usd_total"] = float(
            data.get("estimated_cost_usd_total") or 0.0
        ) + float(inc_usd)
    _gemini_buckets_ensure_structure(data)
    _gemini_bucket_add_one_call(
        data["buckets"], pt, ct, th, tt, inc_usd, when=datetime.now()
    )
    data["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_gemini_cumulative_payload(data)
def record_gemini_response_usage(res, model_id: str) -> None:
    """generate_content の応答から usage_metadata を集計れる（セッション＋累計 JSON）。"""
    global _gemini_usage_session
    if res is None or not str(model_id or "").strip():
        return
    um = getattr(res, "usage_metadata", None)
    if um is None:
        return

    def _iv(name: str) -> int:
        v = getattr(um, name, None)
        try:
            return int(v) if v is not None else 0
        except (TypeError, ValueError):
            return 0

    pt = _iv("prompt_token_count")
    ct = _iv("candidates_token_count")
    tt = _iv("total_token_count")
    th = _iv("thoughts_token_count")
    if tt <= 0 and (pt > 0 or ct > 0 or th > 0):
        tt = pt + ct + th
    mid = str(model_id).strip()
    b = _gemini_usage_session.setdefault(
        mid,
        {"prompt": 0, "candidates": 0, "total": 0, "thoughts": 0, "calls": 0},
    )
    b["prompt"] += pt
    b["candidates"] += ct
    b["total"] += tt
    b["thoughts"] += th
    b["calls"] += 1
    try:
        _append_gemini_cumulative_one_call(mid, pt, ct, th, tt)
    except Exception as ex:
        logging.debug("Gemini 累計の更新で例外（続行）: %s", ex)
def _workbook_file_has_gemini_target_main_sheet(path: str) -> bool:
    """ディスク上のブックにメイン相当シートが無ければ書き込めない。Excel を起動しないための事前判定。"""
    p = (path or "").strip()
    if not p or not os.path.isfile(p):
        return False
    if _ooxml_workbook_missing_shared_strings(p):
        for nm in _ooxml_workbook_sheet_names(p) or []:
            sn = str(nm or "")
            if sn in ("メイン", "メイン_", "Main") or "メイン" in sn:
                return True
        return False
    try:
        wbr = load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return True
    try:
        for nm in wbr.sheetnames:
            sn = str(nm or "")
            if sn in ("メイン", "メイン_", "Main") or "メイン" in sn:
                return True
        return False
    finally:
        try:
            wbr.close()
        except Exception:
            pass
def _openpyxl_chart_title_str(chart_obj) -> str:
    if chart_obj is None:
        return ""
    try:
        t = getattr(chart_obj, "title", None)
        if t is None:
            return ""
        if isinstance(t, str):
            return t
        tx = getattr(t, "tx", None)
        if tx is not None:
            return str(tx)
        return str(t)
    except Exception:
        return ""
def _strip_gemini_usage_charts_openpyxl(ws) -> None:
    """メインシート上の当機能が管理する折れ線グラフ（名前またはタイトル一致）を削除する。"""
    managed_names = (
        GEMINI_USAGE_XLW_CHART_NAME,
        GEMINI_USAGE_XLW_CHART_TOKENS_NAME,
    )
    title_markers = (
        "Gemini API 日次推移",
        "Gemini API 日次トークン",
    )
    charts = getattr(ws, "_charts", None)
    if not charts:
        return
    keep: list = []
    for anc in list(charts):
        drop = False
        try:
            ch = getattr(anc, "chart", None)
            if ch is None:
                ch = anc
            tit_s = _openpyxl_chart_title_str(ch)
            for mk in title_markers:
                if mk in tit_s:
                    drop = True
                    break
            if not drop:
                vchart = getattr(ch, "vchart", None) or ch
                nm = str(getattr(vchart, "name", "") or "")
                if nm in managed_names:
                    drop = True
        except Exception:
            pass
        if not drop:
            keep.append(anc)
    try:
        ws._charts = keep  # type: ignore[attr-defined]
    except Exception:
        return
def _apply_main_sheet_gemini_usage_chart_openpyxl(ws, cum: dict) -> None:
    """Q〜R・S〜T を埋め、折れ線グラフを最大 2 本まで置く（openpyxl・ディスク保存前提）。"""
    hr = GEMINI_USAGE_CHART_HEADER_ROW
    cdt = GEMINI_USAGE_CHART_COL_DATE
    cvl = GEMINI_USAGE_CHART_COL_VALUE
    cts = GEMINI_USAGE_CHART_COL_TOK_DATE
    ctv = GEMINI_USAGE_CHART_COL_TOK_VALUE
    nclear = GEMINI_USAGE_CHART_CLEAR_ROWS

    for i in range(nclear):
        r = hr + i
        for c in (cdt, cvl, cts, ctv):
            try:
                ws.cell(row=r, column=c).value = None
            except Exception:
                pass

    _strip_gemini_usage_charts_openpyxl(ws)
    ser = _gemini_daily_trend_series(cum)
    if ser is None:
        return
    day_keys, values, val_label = ser
    n = len(day_keys)
    if n <= 0:
        return

    ws.cell(row=hr, column=cdt, value="日付")
    ws.cell(row=hr, column=cvl, value=val_label)
    for i, (dk, val) in enumerate(zip(day_keys, values)):
        r = hr + 1 + i
        ws.cell(row=r, column=cdt, value=dk)
        ws.cell(row=r, column=cvl, value=val)
    nf = "0.000000" if val_label == "推定USD" else "0"
    for r in range(hr + 1, hr + n + 1):
        try:
            ws.cell(row=r, column=cvl).number_format = nf
        except Exception:
            pass

    try:
        chart1 = LineChart()
        chart1.title = "Gemini API 日次推移"
        chart1.legend = None
        data = Reference(ws, min_col=cdt, min_row=hr, max_col=cvl, max_row=hr + n)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(
            Reference(ws, min_col=cdt, min_row=hr + 1, max_row=hr + n)
        )
        ws.add_chart(chart1, GEMINI_USAGE_CHART_ANCHOR_CELL)
    except Exception:
        pass

    tok_vals = _gemini_daily_total_tokens_for_days(cum, day_keys)
    if not tok_vals or max(tok_vals) <= 0:
        return

    tok_label = "合計トークン"
    ws.cell(row=hr, column=cts, value="日付")
    ws.cell(row=hr, column=ctv, value=tok_label)
    for i, dk in enumerate(day_keys):
        r = hr + 1 + i
        ws.cell(row=r, column=cts, value=dk)
        ws.cell(row=r, column=ctv, value=int(tok_vals[i]))
    for r in range(hr + 1, hr + n + 1):
        try:
            ws.cell(row=r, column=ctv).number_format = "#,##0"
        except Exception:
            pass

    try:
        chart2 = LineChart()
        chart2.title = "Gemini API 日次トークン"
        chart2.legend = None
        data2 = Reference(ws, min_col=cts, min_row=hr, max_col=ctv, max_row=hr + n)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(
            Reference(ws, min_col=cts, min_row=hr + 1, max_row=hr + n)
        )
        ws.add_chart(chart2, GEMINI_USAGE_CHART_TOKENS_ANCHOR_CELL)
    except Exception:
        pass
def _write_main_sheet_gemini_usage_via_openpyxl(
    macro_wb_path: str, text: str, log_prefix: str
) -> bool:
    """openpyxl でメイン P 列・Q〜T・推移グラフ（最大2本）を更新し wb.save する。"""
    abs_wb = os.path.abspath((macro_wb_path or "").strip())
    if not abs_wb or not os.path.isfile(abs_wb):
        logging.info(
            "%s: Gemini メイン反映: 対象ブックがありません。%s",
            log_prefix,
            macro_wb_path,
        )
        return False
    if _ooxml_workbook_missing_shared_strings(abs_wb):
        logging.info(
            "%s: OOXML に xl/sharedStrings.xml が無いブックのため、"
            "メイン AI サマリ（openpyxl）をスキップしました。"
            "Excel で対象ブックを開いて通常保存すると解消することがあります。",
            log_prefix,
        )
        return False
    if not _workbook_file_has_gemini_target_main_sheet(abs_wb):
        logging.info(
            "%s: メイン相当シートが無いため Gemini の反映をスキップしました（Excel 起動なし）。%s",
            log_prefix,
            os.path.basename(abs_wb),
        )
        return False

    keep_vba = abs_wb.lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(abs_wb, keep_vba=keep_vba)
        ws_main = _gemini_resolve_main_sheet_openpyxl(wb)
        if ws_main is None:
            logging.info(
                "%s: メインシートが見つからないため、AI サマリをスキップしました。",
                log_prefix,
            )
            return False

        start_r, col_p, clear_n = 16, 16, 120
        for i in range(clear_n):
            try:
                ws_main.cell(row=start_r + i, column=col_p).value = None
            except Exception:
                pass
        lines_list = text.split("\n") if (text or "").strip() else []
        wrap = Alignment(wrap_text=True, vertical="top")
        for i in range(clear_n):
            v = lines_list[i] if i < len(lines_list) else None
            try:
                c = ws_main.cell(row=start_r + i, column=col_p, value=v)
                c.alignment = wrap
            except Exception:
                pass

        _apply_main_sheet_gemini_usage_chart_openpyxl(
            ws_main, _load_gemini_cumulative_payload()
        )
        wb.save(abs_wb)
        logging.info(
            "%s: メインシート P%d 以降・Gemini 推移グラフ（料金/呼出し・トークン）を openpyxl で保存しました。",
            log_prefix,
            start_r,
        )
        return True
    except Exception as ex:
        logging.warning(
            "%s: メイン AI サマリの openpyxl 保存に失敗: %s", log_prefix, ex
        )
        return False
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
def _export_gemini_buckets_csv_for_charts(cum: dict) -> None:
    """Excel 折れ線・棒グラフ坑けに長形式 CSV を log に書き出す。"""
    b = cum.get("buckets")
    if not isinstance(b, dict):
        return
    mapping = (
        ("year", "by_year"),
        ("month", "by_month"),
        ("week_iso", "by_week"),
        ("day", "by_day"),
        ("hour", "by_hour"),
    )
    rows_out: list[dict[str, object]] = []
    for gran_label, sub in mapping:
        subd = b.get(sub)
        if not isinstance(subd, dict):
            continue
        for pk in sorted(subd.keys()):
            ent = subd.get(pk)
            if not isinstance(ent, dict):
                continue
            calls = int(ent.get("calls") or 0)
            pt = int(ent.get("prompt") or 0)
            cc = int(ent.get("candidates") or 0)
            th = int(ent.get("thoughts") or 0)
            tt = int(ent.get("total_tokens") or 0)
            usd = float(ent.get("estimated_cost_usd") or 0.0)
            rows_out.append(
                {
                    "granularity": gran_label,
                    "period_key": pk,
                    "calls": calls,
                    "prompt_tokens": pt,
                    "candidates_tokens": cc,
                    "thoughts_tokens": th,
                    "total_tokens": tt,
                    "estimated_cost_usd": round(usd, 8),
                    "estimated_cost_jpy": round(usd * GEMINI_JPY_PER_USD, 4),
                }
            )
    if not rows_out:
        return
    path = os.path.join(log_dir, GEMINI_USAGE_BUCKETS_CSV_FILE)
    fieldnames = [
        "granularity",
        "period_key",
        "calls",
        "prompt_tokens",
        "candidates_tokens",
        "thoughts_tokens",
        "total_tokens",
        "estimated_cost_usd",
        "estimated_cost_jpy",
    ]
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows_out)
    except OSError as ex:
        logging.debug("Gemini ポケット CSV の保存に失敗: %s", ex)
def build_gemini_usage_summary_text() -> str:
    """メイン表示・結果ログ用の複数行テキスト（この実行分＋累計 JSON）。"""
    cum = _load_gemini_cumulative_payload()
    ct_tot = int(cum.get("calls_total") or 0)
    if not _gemini_usage_session and ct_tot <= 0:
        return ""

    lines: list[str] = []
    ts = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    if _gemini_usage_session:
        lines.append(f"集計時刻: {ts}（この実行での Gemini API）")
        tot_calls = sum(b["calls"] for b in _gemini_usage_session.values())
        tot_p = sum(b["prompt"] for b in _gemini_usage_session.values())
        tot_c = sum(b["candidates"] for b in _gemini_usage_session.values())
        tot_th = sum(b["thoughts"] for b in _gemini_usage_session.values())
        tot_t = sum(b["total"] for b in _gemini_usage_session.values())
        sess_rows: list[tuple[str, str]] = [
            ("呼出し", f"{tot_calls:,} 回"),
            ("入力トークン", f"{tot_p:,}"),
            ("出力トークン", f"{tot_c:,}"),
        ]
        if tot_th:
            sess_rows.append(("思考トークン", f"{tot_th:,}"))
        sess_rows.append(("total 報告", f"{tot_t:,}"))
        lines.extend(_gemini_kv_table_lines("」この実行】", sess_rows))
        grand_usd = 0.0
        any_price = False
        for mid in sorted(_gemini_usage_session.keys()):
            b = _gemini_usage_session[mid]
            mrows: list[tuple[str, str]] = [
                ("モデル", mid),
                ("呼出し", f"{b['calls']:,} 回"),
                ("入力トークン", f"{b['prompt']:,}"),
                ("出力トークン", f"{b['candidates']:,}"),
            ]
            if b.get("thoughts", 0):
                mrows.append(("思考トークン", f"{b['thoughts']:,}"))
            mrows.append(("total_token_count", f"{b['total']:,}"))
            est = _gemini_estimate_cost_usd(
                mid, b["prompt"], b["candidates"], b.get("thoughts", 0)
            )
            if est is not None:
                any_price = True
                grand_usd += est
                mrows.append(("推定USD", f"${est:.6f}"))
                mrows.append(
                    (
                        "推定JPY",
                        f"¥{est * GEMINI_JPY_PER_USD:.2f}（{GEMINI_JPY_PER_USD:.0f}円/USD）",
                    )
                )
            else:
                mrows.append(("推定料金", "（坘価未登録モデル）"))
            lines.append("")
            lines.extend(_gemini_kv_table_lines(f"」この実行・モデル別】", mrows))
        if any_price:
            lines.append("")
            lines.extend(
                _gemini_kv_table_lines(
                    "」この実行・推定料金合計】",
                    [
                        ("USD", f"${grand_usd:.6f}"),
                        (
                            "JPY",
                            f"¥{grand_usd * GEMINI_JPY_PER_USD:.2f}（{GEMINI_JPY_PER_USD:.0f}円/USD）",
                        ),
                    ],
                )
            )
    else:
        lines.append(f"集計時刻: {ts}")
        lines.append("（この実行での Gemini API 呼出しはありません）")
    lines.append("※ トークンは API の usage_metadata に基るしした。")
    lines.append(
        "※ USD 坘価はコード＝環境変数の目安です。実課金は Google の請求を参照してください。"
    )
    lines.append(
        "※ 坄 API 呼出しととの推定料金はコンソールに出さう」下記累計 JSON にのみ穝み上きした。"
    )

    if ct_tot > 0:
        lines.append("")
        cum_hdr = (
            f"」累計】{GEMINI_USAGE_CUMULATIVE_JSON_FILE} "
            "（API_Payment フォルダ・全実行の推定値）"
        )
        pt0 = int(cum.get("prompt_total") or 0)
        cc0 = int(cum.get("candidates_total") or 0)
        th0 = int(cum.get("thoughts_total") or 0)
        tt0 = int(cum.get("total_tokens_reported") or 0)
        cum_rows: list[tuple[str, str]] = [
            ("最終更新", str(cum.get("updated_at") or "—")),
            ("呼出し", f"{ct_tot:,} 回"),
            ("入力トークン", f"{pt0:,}"),
            ("出力トークン", f"{cc0:,}"),
        ]
        if th0:
            cum_rows.append(("思考トークン", f"{th0:,}"))
        cum_rows.append(("total 報告", f"{tt0:,}"))
        usd_all = float(cum.get("estimated_cost_usd_total") or 0.0)
        if usd_all > 0:
            cum_rows.append(("推定USD 累計", f"${usd_all:.6f}"))
            cum_rows.append(
                (
                    "推定JPY 累計",
                    f"¥{usd_all * GEMINI_JPY_PER_USD:.2f}（{GEMINI_JPY_PER_USD:.0f}円/USD）",
                )
            )
        lines.extend(_gemini_kv_table_lines(cum_hdr, cum_rows))
        bm = cum.get("by_model") or {}
        if isinstance(bm, dict) and bm:
            for mid in sorted(bm.keys()):
                m = bm[mid]
                if not isinstance(m, dict):
                    continue
                mrows2: list[tuple[str, str]] = [
                    ("モデル", mid),
                    ("呼出し", f"{int(m.get('calls') or 0):,} 回"),
                    (
                        "入力 / 出力",
                        f"{int(m.get('prompt') or 0):,} / {int(m.get('candidates') or 0):,}",
                    ),
                ]
                if int(m.get("thoughts") or 0):
                    mrows2.append(("思考トークン", f"{int(m.get('thoughts') or 0):,}"))
                mud = float(m.get("estimated_cost_usd") or 0.0)
                if mud > 0:
                    mrows2.append(("推定USD 累計", f"${mud:.6f}"))
                    mrows2.append(
                        ("推定JPY 累計", f"¥{mud * GEMINI_JPY_PER_USD:.2f}")
                    )
                lines.append("")
                lines.extend(_gemini_kv_table_lines("」累計・モデル別】", mrows2))
        trend = _gemini_usage_trend_caption_lines(cum)
        if trend:
            lines.append("")
            lines.extend(trend)
    return "\n".join(lines)
def write_main_sheet_gemini_usage_summary(wb_path: str, log_prefix: str) -> None:
    """Gemini 利用サマリを log に書き、openpyxl でメイン P 列・推移グラフへ保存する。"""
    text = build_gemini_usage_summary_text()
    path = os.path.join(log_dir, GEMINI_USAGE_SUMMARY_FOR_MAIN_FILE)
    sheet_ok = False
    if wb_path and os.path.isfile(wb_path):
        try:
            sheet_ok = _write_main_sheet_gemini_usage_via_openpyxl(
                wb_path, text, log_prefix
            )
        except Exception as ex:
            logging.warning(
                "%s: AI サマリの openpyxl 書き込みで例外: %s", log_prefix, ex
            )
    try:
        os.makedirs(log_dir, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
    except OSError:
        pass
    try:
        cum2 = _load_gemini_cumulative_payload()
        if int(cum2.get("calls_total") or 0) > 0:
            _export_gemini_buckets_csv_for_charts(cum2)
    except Exception as ex:
        logging.debug("Gemini ポケット CSV 出力で例外（続行）: %s", ex)
    if sheet_ok:
        return
    if text.strip():
        logging.info(
            "%s: メイン P 列・グラフをブックへ保存できませんでした。"
            " %s に出力済み → マクロ「メインシート_Gemini利用サマリをP列に反映」で P 列のみ反映できます。",
            log_prefix,
            path,
        )
    else:
        logging.info(
            "%s: Gemini 未使用: サマリを空で %s に出力。",
            log_prefix,
            path,
        )
def _try_write_main_sheet_gemini_usage_summary(phase: str) -> None:
    try:
        write_main_sheet_gemini_usage_summary(_excel_plan_input_wb(), phase)
    except Exception as ex:
        logging.warning(
            "%s: メインシートへの AI 利用サマリ書き込みで例外（続行）: %s", phase, ex
        )
def _plan_sheet_write_global_parse_block_to_ws(
    ws,
    global_priority_override: dict,
    when_str: str,
) -> None:
    """既に開いている「配台計画_タスク入力」相当シートへ AX:AY のグローバル解析ブロックを書き。"""
    gpo = global_priority_override or {}
    lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
    vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
    max_r = PLAN_SHEET_GLOBAL_PARSE_MAX_ROWS
    for i in range(max_r):
        ws.cell(row=1 + i, column=lc, value=None)
        ws.cell(row=1 + i, column=vc, value=None)
    align_top = Alignment(wrap_text=True, vertical="top")
    pairs: list[tuple[str, str]] = [
        ("」グローバルコメント解析】", "参照用・段階2で自動記録"),
        (
            "※二重適用についで",
            "配台への反映はメインシート「グローバルコメント」からのみ行ゝれした。"
            "このAX〜AY列は読み坖られません。編集しても次回実行まで配台に効しません。"
            "原文はメイン欄を参照してください。",
        ),
        ("計画基準日時", (when_str or "").strip() or "―"),
        (
            "工場休業日",
            ", ".join(str(x) for x in (gpo.get("factory_closure_dates") or []))
            if gpo.get("factory_closure_dates")
            else "（なし）",
        ),
        (
            "スキル覝件を無視",
            "はい" if gpo.get("ignore_skill_requirements") else "いいえ",
        ),
        (
            "need人数1固定",
            "はい" if gpo.get("ignore_need_minimum") else "いいえ",
        ),
        (
            "配台制限の撤廃",
            "はい" if gpo.get("abolish_all_scheduling_limits") else "いいえ",
        ),
        (
            "グローバルOP指定",
            json.dumps(gpo.get("task_preferred_operators") or {}, ensure_ascii=False)
            if gpo.get("task_preferred_operators")
            else "（なし）",
        ),
        (
            "日付×工程フォーム指定",
            json.dumps(
                gpo.get("global_day_process_operator_rules") or [],
                ensure_ascii=False,
            )
            if gpo.get("global_day_process_operator_rules")
            else "（なし）",
        ),
        (
            "グローバル速度ルール",
            json.dumps(gpo.get("global_speed_rules") or [], ensure_ascii=False)
            if gpo.get("global_speed_rules")
            else "（なし）",
        ),
        (
            "未適用メモ(AI)",
            str(gpo.get("scheduler_notes_ja") or "").strip() or "（なし）",
        ),
        (
            "AI覝約",
            str(gpo.get("interpretation_ja") or "").strip() or "（なし）",
        ),
    ]
    for i, (lab, val) in enumerate(pairs):
        if i >= max_r:
            break
        c1 = ws.cell(row=1 + i, column=lc, value=lab)
        c2 = ws.cell(row=1 + i, column=vc, value=val)
        c1.alignment = align_top
        c2.alignment = align_top
def write_plan_sheet_global_comment_parse_block(
    wb_path: str,
    sheet_name: str,
    global_priority_override: dict,
    *,
    when_str: str,
    log_prefix: str = "段階2",
) -> bool:
    """
    「配台計画_タスク入力」シートの坳端付近（AX:AY）に」グローバルコメントの解析結果を書き込む。
    メイン原文はここに転記しない（メイン欄との重複・誤解を避ける）。本列は再読込されう参照専用。
    マクロブックも ``keep_vba=True`` で openpyxl から保存する。
    """
    if not wb_path or not os.path.isfile(wb_path):
        return False
    gpo = global_priority_override or {}
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: ブックに「%s」があるため、openpyxl でグローバルコメント解析を配台シートへ書き込みません。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    try:
        wb = load_workbook(
            wb_path, keep_vba=keep_vba, read_only=False, data_only=False
        )
    except Exception as ex:
        logging.info(
            "%s: グローバルコメント解析の配台シート書込のため、ブックを開きません: %s",
            log_prefix,
            ex,
        )
        return False
    try:
        if sheet_name not in wb.sheetnames:
            logging.info(
                "%s: シート '%s' はないため、グローバルコメント解析の反映をスキップ。",
                log_prefix,
                sheet_name,
            )
            return False
        ws = wb[sheet_name]
        _plan_sheet_write_global_parse_block_to_ws(ws, gpo, when_str)
        lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
        vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
        try:
            wb.save(wb_path)
        except OSError as ex:
            logging.warning(
                "%s: グローバルコメント解析の保存に失敗しました: %s",
                log_prefix,
                ex,
            )
            return False
        logging.info(
            "%s: 「%s」%s:%s 列にグローバルコメント解析を保存しました。",
            log_prefix,
            sheet_name,
            get_column_letter(lc),
            get_column_letter(vc),
        )
        return True
    except OSError as ex:
        logging.warning(
            "%s: グローバルコメント解析を配台シートへ保存でしませんでした: %s",
            log_prefix,
            ex,
        )
        return False
    except Exception as ex:
        logging.warning(
            "%s: グローバルコメント解析の配台シート書込で例外: %s", log_prefix, ex
        )
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def _try_write_plan_sheet_global_comment_parse_block(
    global_priority_override: dict,
    when_str: str,
) -> None:
    try:
        write_plan_sheet_global_comment_parse_block(
            _excel_plan_input_wb(),
            PLAN_INPUT_SHEET_NAME,
            global_priority_override,
            when_str=when_str,
            log_prefix="段階2",
        )
    except Exception as ex:
        logging.warning(
            "段階2: 配台シートへのグローバルコメント解析書き込みで例外（続行）: %s",
            ex,
        )
def _try_write_plan_input_global_parse_and_conflicts_one_save(
    global_priority_override: dict,
    when_str: str,
    num_data_rows: int,
    conflicts_by_row,
    tasks_df=None,
) -> None:
    try:
        write_plan_sheet_global_parse_and_conflict_styles_one_io(
            _excel_plan_input_wb(),
            PLAN_INPUT_SHEET_NAME,
            global_priority_override,
            when_str=when_str,
            num_data_rows=num_data_rows,
            conflicts_by_row=conflicts_by_row,
            log_prefix="段階2",
            tasks_df=tasks_df,
        )
    except Exception as ex:
        logging.warning(
            "段階2: 配台シートへのグローバル解析＋矛盾着色（1回保存）で例外（続行）: %s",
            ex,
        )
def _log_task_special_ai_response(raw_text, parsed, extracted_json_str, prompt_text=None):
    """特別指定_備考坑け Gemini のプロンプト・生テキスト・抽出JSON・パース結果を1ファイルに残れ。"""
    path = os.path.join(log_dir, TASK_SPECIAL_AI_LAST_RESPONSE_FILE)
    try:
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            if prompt_text is not None and str(prompt_text).strip():
                f.write("=== Gemini へ逝信したプロンプト（全文） ===\n")
                f.write(str(prompt_text).strip())
                f.write("\n\n")
            f.write("=== Gemini 返坴テキスト（モデル出力しのまま） ===\n")
            f.write(raw_text or "")
            f.write(
                "\n\n=== AI は返したテキストからクライアントは切り出した JSON 文字列 ===\n"
                "（※ユーザー特別指定の解析に正覝表睾は使っていません。モデル応答のパース用です）\n"
            )
            f.write(extracted_json_str if extracted_json_str else "(抽出なし)")
            f.write("\n\n=== json.loads 後（依頼NOキー） ===\n")
            if isinstance(parsed, dict):
                f.write(json.dumps(parsed, ensure_ascii=False, indent=2))
            else:
                f.write("(パースでしう)")
        logging.info(
            "タスク特別指定: プロンプト＋AI応答の詳細 → %s",
            path,
        )
    except OSError as ex:
        logging.warning("タスク特別指定: AI応答ファイル保存に失敗: %s", ex)
    if isinstance(parsed, dict) and parsed:
        logging.info(
            "タスク特別指定: 解析された依頼NO: %s",
            ", ".join(sorted(parsed.keys(), key=lambda x: str(x))),
        )
        for tid_k in sorted(parsed.keys(), key=lambda x: str(x)):
            logging.info(
                "  依頼NO [%s] AI解析フィールド: %s",
                tid_k,
                json.dumps(parsed[tid_k], ensure_ascii=False),
            )
def _parse_and_log_task_special_gemini_response(res, prompt_text=None):
    """
    API レスポンスを JSON 化しログ＝ファイルへ記録。失敗時は None。
    ユーザーの特別指定文言には触れう」モデル出力から JSON ブロックを取り出す処理のみ。
    """
    raw = _gemini_result_text(res)
    if raw:
        stripped = raw.strip()
        if stripped.startswith("{"):
            try:
                trial = json.loads(stripped)
                if isinstance(trial, dict):
                    _log_task_special_ai_response(raw, trial, stripped, prompt_text)
                    return trial
            except json.JSONDecodeError:
                pass
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        _log_task_special_ai_response(raw, {}, None, prompt_text)
        logging.warning(
            "タスク特別指定: AI応答から JSON を抽出でしませんでした。生テキスト先頭 3000 文字:\n%s",
            (raw[:3000] if raw else "(空)"),
        )
        return None
    extracted = match.group(0)
    try:
        parsed = json.loads(extracted)
    except json.JSONDecodeError as je:
        _log_task_special_ai_response(raw, None, extracted, prompt_text)
        logging.warning("タスク特別指定: JSON パース失敗: %s", je)
        return None
    if not isinstance(parsed, dict):
        _log_task_special_ai_response(raw, None, extracted, prompt_text)
        logging.warning("タスク特別指定: トップレベルは JSON オブジェクトではありません。")
        return None
    _log_task_special_ai_response(raw, parsed, extracted, prompt_text)
    return parsed
def analyze_task_special_remarks(tasks_df, reference_year=None, ai_sheet_sink: dict | None = None):
    """
    「配台計画_タスク入力」の「特別指定_備考」を AI で構造化（セルに値はある項目は後段でセルを優先）。
    「配台試行」はオンな行はプロンプトに載せない（API 節約・当該行は配台しないため）。
    担当OP指定はプロンプトの返坴契約でモデルに preferred_operator を出力させる（備考を正覝表睾で切り出す処理は行ゝない）。
    json/ai_remarks_cache.json に TTL AI_CACHE_TTL_SECONDS でキャッシュ（同一入力・同一基準年なら API を呼みない）。
    依頼NOは数値表記・全角などを正規化してキーを安定化し、基準年は指紋に含むで日付解釈の変化とキャッシュの食い靕いを防し。

    戻り値の例: 依頼NO -> オブジェクト」または同一依頼NOに備考行は複数ある場合はオブジェクトの配列。
      process_name, machine_name … 当該備考セルはある行の工程名・機械名（プロンプトの行と一致）
      restrict_to_process_name, restrict_to_machine_name … 省略または空なら同一依頼NOの全工程・全機械行に適用。
      しの他 required_op, speed_override, task_efficiency, priority, start_date, start_time,
      target_completion_date, ship_by_date, preferred_operator など。
    """
    lines = _task_special_prompt_lines(tasks_df)
    if not lines:
        miss_col = PLAN_COL_SPECIAL_REMARK not in tasks_df.columns
        if miss_col:
            n_rows = len(tasks_df)
            n_tid_raw = 0
            for _, row in tasks_df.iterrows():
                if planning_task_id_str_from_plan_row(row):
                    n_tid_raw += 1
            logging.warning(
                "タスク特別指定: AI 解析対象はありません（「%s」列は見つかりません）。"
                "总行数=%s」依頼NOのある行=%s。"
                "段階2実行剝にブックを保存し、本当に「%s」列へ入力しているか確認してください。",
                PLAN_COL_SPECIAL_REMARK,
                n_rows,
                n_tid_raw,
                PLAN_COL_SPECIAL_REMARK,
            )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "スキップ（対象行なし）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（対象行なし・API 未実行）"
        return {}

    blob = "\n".join(sorted(lines))
    ref_y = int(reference_year) if reference_year is not None else date.today().year
    cache_fingerprint = f"{ref_y}\n{blob}"
    cache_key_input = f"{TASK_SPECIAL_CACHE_KEY_PREFIX}{cache_fingerprint}"
    cache_key = hashlib.sha256(cache_key_input.encode("utf-8")).hexdigest()
    ai_cache = load_ai_cache()
    cached_parsed = get_cached_ai_result(
        ai_cache, cache_key, content_key=cache_fingerprint
    )
    if cached_parsed is not None:
        logging.info(
            "タスク特別指定: キャッシュヒット（%s 件・基準年=%s）。Gemini は呼びません。",
            len(lines),
            ref_y,
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "なし（キャッシュ使用）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（キャッシュ利用・今回 API 未実行）"
        out = copy.deepcopy(cached_parsed)
        if isinstance(out, dict):
            _repair_task_special_ai_wrong_top_level_keys(out, tasks_df)
        return out

    logging.info(
        "タスク特別指定: キャッシュなし。Gemini で %s 件の備考を解析しした（基準年=%s）。",
        len(lines),
        ref_y,
    )

    if not API_KEY:
        logging.info("Gemini API キーが未設定のため、タスク特別指定のAI解析をスキップしました。")
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "なし（APIキー未設定）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（API キー未設定）"
        return {}

    prompt = f"""
あなたは工場の配台計画坑けに」Excel「特別指定_備考」欄への自由記述を読み」配台ロジックは使うるフィールドの値に蝽とし込むアシスタントです。

」最針覝】
1) 」特別指定原文】の坄行は」ユーザーはセルに入力した文字列を **改変・覝約・断う切りはしてよらう**（先頭末尾の空白のみ除去）」しのまま渡していした。**原文の事実や愝図を別の文言に置し杛ごないでしてさい。**
2) あなたの応答は **1個の JSON オブジェクトのみ**（先頭は {{ 」末尾は }} ）。説明文・マークダウン・コードフェンスは禁止。
3) JSON のトップレベルキーは」坄行の **依頼NO」と】の間の文字列のみ** と **完全一致** させること。**備考本文**に書かれた哝番・原板坝・製哝コード（例: 20010 で始まる番坷列）をキーにしてはならない。備考はしのよごな番坷で始まっていでも」キーは必う」】内の依頼NOの値とれる。

」返坴JSONの契約（この節どよりに出力れること）】
■ トップレベル
- キー: 上記」特別指定原文】の **依頼NO」…】の括弧内** の文字列と **完全一致**（表記・ポイフン・英大文字尝文字を原文どより）。備考本文中の数字列をキーにしない。
- 値: 次のいうれか。
  (A) **JSONオブジェクト1つ** … 当該依頼NOの備考はプロンプト上 **1行の値** のとき。
  (B) **JSON配列**（覝素はオブジェクト）… 同一依頼NOで工程名・機械名は異なる備考行は **複数** あるとし。覝素の順はプロンプトの行順と対応させる。

■ process_name（文字列）・machine_name（文字列）— **必須**
- 当該備考に対応れるプロンプト行の **工程名「…」**・**機械名「…」** の値と **一致** させる（「（空）」のときは空文字列 ""）。
- ログ・トレース用。省略試行。

■ restrict_to_process_name（文字列）・restrict_to_machine_name（文字列）— **任愝**
- **原文は「特定の工程の値」「この機械の値」など」適用範囲を絞っているとしの値** 出力れる。
- **原文に工程名・機械名の陝定は無い**（依頼全体・全行程への指示）としは **両方とも省略** れるか **空文字列 ""** とれる。
- しの場合」配台ロジックは **同一依頼NOの別行（例: エンボス行と分割行）にも指示を適用** れる。
- 絞る場合は」原文で示された識別名を入れる（Excel の工程名・機械名と照合しやれい表記）。

■ preferred_operator（文字列）— 条件付し**必須**
- **必須条件**: 当該依頼の原文を読み」「**誰はこの加工・作業の主担当（OP）として割り当でたいか**」は **愝味として** 読み坖れるとし。
  例: 特定の人にやってもらご＝しの人に任せる＝担当はあの人＝OPは〜＝〜さん（並び）に依頼」など。**表睾の型に依存せう**」文の愝味で判断れる。
- **満たしたとしの出力義務**: 上記の愝味は成立れると判断したオブジェクトでは」**必う** キー `preferred_operator` を含む」値は **空でない文字列** とれる。併せで **process_name / machine_name は必須**（例: `{{"process_name":"…","machine_name":"…","preferred_operator":"…"}}`）。
- **値の形式**: 原文で示された **担当者の識別名を1坝分**（姓・坝・ニックフォーム等」原文に睾れた表記を維挝）。末尾の敬称（さん・坛・氝）のみ除去。例:「森岡さんにやってもらいした」→ `"森岡"`。
- **出力してはいけないとし**: 原文に担当者の指愝は **一切ない** と判断した依頼NOでは `preferred_operator` キー自体を **省略** れる（空文字列も付けない）。

■ しの他フィールド（required_op, speed_override, task_efficiency, priority, start_date, start_time, target_completion_date, ship_by_date）
- 原文から **明確に** 読み坖れる場合のみ出力。読み取れない数値・日付は **省略**（推測で埋ゝない）。

」同一依頼NO・複数工程の例】
依頼NO Y4-2 に「エンボス」と「分割」の行はあり」備考は「4/5までに終ゝらせる」のみで工程の陝定は無い場合:
- process_name / machine_name は **備考は書かれた行** の値を入れる。
- restrict_to_* は **出さないか空** にし、**エンボス行・分割行の両方** にも優先度・日付等は効しよごにれる。

」基準年（年なし日付用）】
「4/5」「4/5に出蝷」のよごに **年は無い** 日付は原則 **西暦 {ref_y} 年** とし、YYYY-MM-DD で出力。

」フィールド一覧（型の参考）】
- process_name, machine_name: 文字列（必須。プロンプト行と一致）
- restrict_to_process_name, restrict_to_machine_name: 文字列（任愝。陝定なら）
- preferred_operator: 文字列（上記契約に従ご）
- required_op: 正の整数
- speed_override: 正の数（m/分）。※配台の実効速度は列「加工速度_上書き」「加工速度」のみ使用。本キーは速度計算には反映せず、列との食い違い検出に用いる。
- task_efficiency: 0〜1
- priority: 整数（尝さいろど先に割付）
- start_date: YYYY-MM-DD / start_time: HH:MM
- target_completion_date, ship_by_date: YYYY-MM-DD

」解釈の指針】
- 「間に坈ごよごに」「繰り上きる」→ priority を上きる（数値を下きる）。日付は文中にあれみ target_completion_date または ship_by_date に入れる。
- 担当者指定は **愝味睆解** で preferred_operator を決ゝる（特定のキーワード列挙に頼らない）。
- 数値・日付は推測で補ゝない。
- **備考は特定の工程・機械にの値言坊していない陝り**」restrict_to_* は空にし、同一依頼NOの他行にも適用される形にれる。

」出力直後の自己検証（必う実行してから JSON を閉もる）】
- 」特別指定原文】の **坄行** についで」対応れるオブジェクトに **process_name** と **machine_name** はあるか。
- 同一依頼NOは複数行あるとしは **配列** で坄行に1オブジェクト」または革切にマージした坘一オブジェクト＋restrict の靋用を一貫させる。
- 「主担当OPの指愝」はある行では **非空の preferred_operator** を付ける。

」出力形式の例】（依頼NO・値は実データに合わせ替ごること）
{{
  "W3-14": {{
    "process_name": "検査",
    "machine_name": "ラインA",
    "preferred_operator": "森岡"
  }},
  "Y3-26": {{
    "process_name": "コーティング",
    "machine_name": "",
    "priority": 1,
    "ship_by_date": "{ref_y}-04-05",
    "target_completion_date": "{ref_y}-04-05"
  }},
  "Y4-2": {{
    "process_name": "エンボス",
    "machine_name": "E1",
    "priority": 2,
    "restrict_to_process_name": "",
    "restrict_to_machine_name": ""
  }}
}}

」特別指定原文】（Excel からしのまま。1行＝依頼NOと備考のペア）
{blob}
"""
    try:
        ppath = os.path.join(log_dir, "ai_task_special_last_prompt.txt")
        with open(ppath, "w", encoding="utf-8", newline="\n") as pf:
            pf.write(prompt)
        logging.info("タスク特別指定: 今回 Gemini に渡したプロンプト全文 → %s", ppath)
    except OSError as ex:
        logging.warning("タスク特別指定: プロンプト保存失敗: %s", ex)

    client = _gemini_client(API_KEY)
    try:
        res, gem_model_used = _gemini_generate_content_with_retry(
            client, contents=prompt, log_label="タスク特別指定"
        )
        record_gemini_response_usage(res, gem_model_used)
        parsed = _parse_and_log_task_special_gemini_response(res, prompt_text=prompt)
        if parsed is not None:
            _repair_task_special_ai_wrong_top_level_keys(parsed, tasks_df)
            put_cached_ai_result(
                ai_cache, cache_key, parsed, content_key=cache_fingerprint
            )
            save_ai_cache(ai_cache)
            logging.info("タスク特別指定: AI解析は完了しました。")
            if ai_sheet_sink is not None:
                ai_sheet_sink["特別指定備考_AI_API"] = "あり"
                ai_sheet_sink["特別指定備考_Geminiモデル"] = gem_model_used
            return parsed
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = "あり（JSON解釈失敗）"
            ai_sheet_sink["特別指定備考_Geminiモデル"] = gem_model_used
        return {}
    except Exception as e:
        logging.warning("タスク特別指定: Gemini 呼び出し失敗（再試行尽き）: %s", e)
        logging.warning(
            "タスク特別指定: AI解析結果を取得でしなかったため、特別指定_備考の開始日/優先指示は反映されません。"
            "（列「加工開始日_指定」「指定納期_上書き」は廃止済み。備考の再記載または後から AI 再実行を検討してください。）"
        )
        if ai_sheet_sink is not None:
            ai_sheet_sink["特別指定備考_AI_API"] = f"失敗: {e}"[:500]
            ai_sheet_sink["特別指定備考_Geminiモデル"] = "—（呼び出し失敗）"
        return {}
def _merge_preferred_operator_cell_and_ai(row, ai_for_tid):
    """Excel「担当OP_指定」を優先し、空なら AI の preferred_operator。"""
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}
    v = row.get(PLAN_COL_PREFERRED_OP)
    if v is not None and not (isinstance(v, float) and pd.isna(v)):
        s = str(v).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
    a = ai.get("preferred_operator")
    if a is not None:
        s = str(a).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
    return ""
def _global_override_preferred_operator_for_task(tpref, task_id) -> str | None:
    """
    メイン「再優先特別記載」の task_preferred_operators。
    キーは依頼NO（大文字・尝文字の差は無視）。
    """
    if not isinstance(tpref, dict) or not tpref:
        return None
    tid = str(task_id).strip()
    if not tid:
        return None
    tlo = tid.lower()
    for k, v in tpref.items():
        if str(k).strip().lower() != tlo:
            continue
        s = str(v).strip()
        if s and s.lower() not in ("nan", "none", "null"):
            return s
        return None
    return None
def _planning_speed_override_sheet_column_only(row) -> float | None:
    """廃止: 加工速度は列「加工速度」のみ。互換のため常に None。"""
    return None
def _merge_task_row_with_ai(
    row, ai_for_tid, *, allow_ai_dispatch_priority_from_remark: bool = True
):
    """
    上書き列は加工速度_上書き・原板投入日_上書き等のみ（計画シート）。しの他は特別指定備考 AI から。
    加工速度の上書きは列「加工速度_上書き」のみ（備考 AI の speed_override は配台速度に使わない）。
    allow_ai_dispatch_priority_from_remark は False のとき」AI の required_op / task_efficiency / priority /
    start_date / start_time は採用しない（備考に紝期系文言は無い行坑け）。
    """
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}

    if allow_ai_dispatch_priority_from_remark:
        req_op = parse_optional_int(ai.get("required_op"))
    else:
        req_op = None
    if req_op is not None and req_op < 1:
        req_op = None

    if allow_ai_dispatch_priority_from_remark:
        te = None
        a = ai.get("task_efficiency")
        try:
            if a is not None and float(a) > 0:
                te = float(a)
        except (TypeError, ValueError):
            te = None
        if te is None or te <= 0:
            te = 1.0
    else:
        te = 1.0

    if allow_ai_dispatch_priority_from_remark:
        pri = parse_optional_int(ai.get("priority"))
    else:
        pri = None
    if pri is None:
        pri = 999

    st_date = None
    if allow_ai_dispatch_priority_from_remark and ai.get("start_date"):
        st_date = parse_optional_date(ai.get("start_date"))

    st_time = None
    if allow_ai_dispatch_priority_from_remark and ai.get("start_time"):
        st_time = parse_time_str(str(ai.get("start_time")), None)

    speed_ov = _planning_speed_override_sheet_column_only(row)

    return req_op, speed_ov, te, pri, st_date, st_time, ai
def _plan_row_cell_nonempty(row, col_name):
    v = row.get(col_name)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none"):
        return False
    return True
def _ai_float_for_conflict(ai, key):
    if not ai or ai.get(key) is None:
        return None
    try:
        f = float(ai.get(key))
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None
def detect_planning_remark_ai_conflicts(row, ai_for_tid):
    """
    特別指定_備考に依る AI 解析結果と」明示セルの両方に値はあり食い靕ご列を返す。
    備考・AIいうれか欠ける場合は空集合。
    配台の実効速度は列「加工速度_上書き」→「加工速度」のみのため、
    備考 AI の speed_override が列「加工速度_上書き」と食い違うときは「加工速度_上書き」を矛盾列に含める。
    """
    remark = str(row.get(PLAN_COL_SPECIAL_REMARK, "") or "").strip()
    if not remark or remark.lower() in ("nan", "none"):
        return set()
    ai = ai_for_tid if isinstance(ai_for_tid, dict) else {}
    if not ai:
        return set()
    out = set()

    if _plan_row_cell_nonempty(row, TASK_COL_SPEED):
        cv = parse_float_safe(row.get(TASK_COL_SPEED), None)
        if cv is not None and cv > 0:
            av = _ai_float_for_conflict(ai, "speed_override")
            if av is not None and abs(cv - av) > 1e-5:
                out.add(TASK_COL_SPEED)

    if _plan_row_cell_nonempty(row, PLAN_COL_PREFERRED_OP):
        cv = _normalize_person_name_for_match(row.get(PLAN_COL_PREFERRED_OP))
        av = _normalize_person_name_for_match(ai.get("preferred_operator"))
        if cv and av and cv != av:
            out.add(PLAN_COL_PREFERRED_OP)

    if out:
        out.add(PLAN_COL_SPECIAL_REMARK)
    return out
def collect_planning_conflicts_by_excel_row(tasks_df, ai_by_tid):
    """Excel 行番坷(1始まり・ヘッダー=1行目) -> 矛盾はあった列名の集合"""
    res = {}
    for i, (_, row) in enumerate(tasks_df.iterrows()):
        if _plan_row_exclude_from_assignment(row):
            continue
        ai_one = _ai_task_special_entry_for_row(ai_by_tid, row)
        cset = detect_planning_remark_ai_conflicts(row, ai_one)
        if cset:
            res[i + 2] = cset
    return res
def _plan_sheet_apply_conflict_styles_to_ws(ws, num_data_rows: int, conflicts_by_row) -> None:
    """既に開いている配台計画シートへ」矛盾列の着色（薄黄リセット→赤）を適用する。保存は呼び出し坴。"""
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(1, col_idx).value
        if v is not None:
            header_map[str(v).strip()] = col_idx

    last_row = max(2, 1 + int(num_data_rows))
    clear_fill = PatternFill(fill_type=None)
    yellow_input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    conflict_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    conflict_font = _result_font(color="FFFFFF", bold=True)

    for r in range(2, last_row + 1):
        for name in PLAN_CONFLICT_STYLABLE_COLS:
            ci = header_map.get(name)
            if not ci:
                continue
            cell = ws.cell(row=r, column=ci)
            if name == PLAN_COL_AI_PARSE:
                cell.fill = clear_fill
            else:
                cell.fill = yellow_input_fill
            # フォントは上書きしない（ブック既定・ユーザー設定を維挝）

    for r, colnames in conflicts_by_row.items():
        if r < 2:
            continue
        for name in colnames:
            ci = header_map.get(name)
            if not ci:
                continue
            cell = ws.cell(row=r, column=ci)
            cell.fill = conflict_fill
            cell.font = conflict_font
def _plan_df_reset_effective_roll_unit_ilocs(tasks_df) -> None:
    """実効ロール単位の行トラッキングをクリア（同一 DataFrame で配台キュー再構築するときの取りこぼし防止）。"""
    try:
        tasks_df.attrs[PLAN_DF_ATTR_EFFECTIVE_ROLL_UNIT_DATA_ILOCS] = set()
    except Exception:
        pass
def _plan_df_note_effective_roll_unit_iloc(tasks_df, data_iloc: int) -> None:
    """実効ロール単位に書き換えた ``tasks_df`` のデータ行位置（iloc 相当）を記録する。"""
    try:
        a = tasks_df.attrs
        key = PLAN_DF_ATTR_EFFECTIVE_ROLL_UNIT_DATA_ILOCS
        if key not in a or not isinstance(a.get(key), set):
            a[key] = set()
        a[key].add(int(data_iloc))
    except Exception:
        pass
def _plan_sheet_apply_effective_roll_unit_cells_from_df(
    ws, tasks_df, num_data_rows: int, *, log_prefix: str = "段階2"
) -> int:
    """
    実効ロール単位に更新した行について、シート上の「(原反)ロール単位長さ」を DataFrame 値で上書きし、
    背景=黄・文字=黒にする（矛盾着色のあとに適用。ロール列は PLAN_CONFLICT_STYLABLE に含まれない）。
    戻り値: 書式を付けたセル数。
    """
    if tasks_df is None or getattr(tasks_df, "empty", True):
        return 0
    try:
        ilocs = tasks_df.attrs.get(PLAN_DF_ATTR_EFFECTIVE_ROLL_UNIT_DATA_ILOCS)
    except Exception:
        ilocs = None
    if not ilocs or not isinstance(ilocs, set):
        return 0
    if PLAN_COL_RAW_ROLL_UNIT_LENGTH not in tasks_df.columns:
        return 0
    header_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(1, col_idx).value
        if v is not None:
            header_map[str(v).strip()] = col_idx
    ci = header_map.get(PLAN_COL_RAW_ROLL_UNIT_LENGTH)
    if not ci:
        return 0
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    font_black = _result_font(color="000000", bold=False)
    last_row = max(2, 1 + int(num_data_rows))
    n_done = 0
    n_rows = len(tasks_df)
    for i in sorted(ilocs):
        if i < 0 or i >= n_rows:
            continue
        excel_r = i + 2
        if excel_r > last_row:
            continue
        try:
            val = tasks_df.iloc[i][PLAN_COL_RAW_ROLL_UNIT_LENGTH]
        except Exception:
            continue
        cell = ws.cell(row=excel_r, column=ci)
        cell.value = val
        cell.fill = fill_yellow
        cell.font = font_black
        n_done += 1
    if n_done:
        logging.info(
            "%s: 実効ロール単位の「%s」セルを %s 件、黄地・黒字で反映しました。",
            log_prefix,
            PLAN_COL_RAW_ROLL_UNIT_LENGTH,
            n_done,
        )
    return n_done
def _openpyxl_cell_fill_rgb_tuple(cell) -> tuple[int, int, int] | None:
    """openpyxl のセル塗りから RGB を取り出す。塗り無しは None。"""
    try:
        f = cell.fill
        if f is None or getattr(f, "fill_type", None) in (None, "none"):
            return None
        sc = getattr(f, "start_color", None)
        if sc is None or sc.rgb is None:
            return None
        hx = str(sc.rgb).upper().replace("0X", "")
        if len(hx) == 8:
            hx = hx[2:]
        if len(hx) != 6:
            return None
        return (int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16))
    except Exception:
        return None
def _openpyxl_font_rgb_tuple(cell) -> tuple[int, int, int] | None:
    """openpyxl のセルフォント色から RGB。未設定は None。"""
    try:
        fo = cell.font
        if not fo or not fo.color:
            return None
        rgb = fo.color.rgb
        if rgb is None:
            return None
        hx = str(rgb).upper().replace("0X", "")
        if len(hx) == 8:
            hx = hx[2:]
        if len(hx) != 6:
            return None
        return (int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16))
    except Exception:
        return None
def _snapshot_plan_sheet_conflict_style_cells(
    ws, num_data_rows: int
) -> list[tuple[int, int, tuple[int, int, int] | None, tuple[int, int, int] | None, bool]]:
    """
    _plan_sheet_apply_conflict_styles_to_ws 適用直後の、矛盾着色対象列の塗り・フォントを列挙する。
    戻り値: (行, 列, fill_rgb|None, font_rgb|None, bold)。
    """
    header_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(1, col_idx).value
        if v is not None:
            header_map[str(v).strip()] = col_idx
    last_row = max(2, 1 + int(num_data_rows))
    out: list[tuple[int, int, tuple[int, int, int] | None, tuple[int, int, int] | None, bool]] = []
    for r in range(2, last_row + 1):
        for name in PLAN_CONFLICT_STYLABLE_COLS:
            ci = header_map.get(name)
            if not ci:
                continue
            cell = ws.cell(row=r, column=ci)
            brgb = _openpyxl_cell_fill_rgb_tuple(cell)
            frgb = _openpyxl_font_rgb_tuple(cell)
            bold = bool(cell.font and cell.font.bold)
            out.append((r, ci, brgb, frgb, bold))
    return out
def write_plan_sheet_global_parse_and_conflict_styles_one_io(
    wb_path: str,
    sheet_name: str,
    global_priority_override: dict,
    *,
    when_str: str,
    num_data_rows: int,
    conflicts_by_row,
    log_prefix: str = "段階2",
    tasks_df=None,
) -> bool:
    """
    段階2: グローバルコメント解析（AX:AY）と矛盾ハイライトを反映する。
    マクロブック（.xlsm）も含め、編集内容を openpyxl で ``keep_vba=True`` のまま保存する。
    """
    if not wb_path or not os.path.isfile(wb_path):
        return False
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "%s: ブックに「%s」があるため、openpyxl でグローバル解析・矛盾着色をスキップしました。",
            log_prefix,
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return False
    if _ooxml_workbook_missing_shared_strings(wb_path):
        logging.info(
            "%s: OOXML に sharedStrings.xml が無いブックのため、"
            "配台シート一括書込（openpyxl）をスキップ（専用UI等の xlsx）",
            log_prefix,
        )
        return False
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(
            wb_path, keep_vba=keep_vba, read_only=False, data_only=False
        )
    except Exception as ex:
        logging.info(
            "%s: 配台シート一括書込のため、ブックを開きません: %s",
            log_prefix,
            ex,
        )
        return False
    try:
        if sheet_name not in wb.sheetnames:
            logging.info(
                "%s: シート '%s' はないため、グローバル解析・矛盾着色をスキップ。",
                log_prefix,
                sheet_name,
            )
            return False
        ws = wb[sheet_name]
        _plan_sheet_write_global_parse_block_to_ws(ws, global_priority_override or {}, when_str)
        _plan_sheet_apply_conflict_styles_to_ws(ws, num_data_rows, conflicts_by_row or {})
        _plan_sheet_apply_effective_roll_unit_cells_from_df(
            ws, tasks_df, num_data_rows, log_prefix=log_prefix
        )
        lc = PLAN_SHEET_GLOBAL_PARSE_LABEL_COL
        vc = PLAN_SHEET_GLOBAL_PARSE_VALUE_COL
        write_planning_conflict_highlight_sidecar(
            sheet_name, num_data_rows, conflicts_by_row or {}
        )
        try:
            wb.save(wb_path)
        except OSError as ex:
            logging.warning(
                "%s: 配台シートの openpyxl 保存に失敗しました（ファイルロック等の可能性）: %s",
                log_prefix,
                ex,
            )
            return False
        _remove_planning_conflict_sidecar_safe()
        _n_conf = len(conflicts_by_row) if conflicts_by_row else 0
        if _n_conf:
            logging.info(
                "%s: 「%s」%s:%s 列にグローバル解析を保存し、"
                "特別指定_備考と列の矛盾 %s 行をハイライトしました。",
                log_prefix,
                sheet_name,
                get_column_letter(lc),
                get_column_letter(vc),
                _n_conf,
            )
        else:
            logging.info(
                "%s: 「%s」%s:%s 列にグローバル解析を保存しました（矛盾行なし）。",
                log_prefix,
                sheet_name,
                get_column_letter(lc),
                get_column_letter(vc),
            )
        return True
    except OSError as ex:
        logging.warning(
            "%s: 配台シート一括保存で OSError: %s",
            log_prefix,
            ex,
        )
        return False
    except Exception as ex:
        logging.warning(
            "%s: 配台シートへのグローバル解析＋矛盾着色（一括）で例外: %s",
            log_prefix,
            ex,
        )
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def apply_planning_sheet_conflict_styles(wb_path, sheet_name, num_data_rows, conflicts_by_row):
    """
    配台計画_タスク入力シートのデータ行を」矛盾列のみ赤地・白太字にれる。
    事剝パスでは上書き入力列を段階1とともに薄黄色に戻し、フォントは変更しない（体裝維挝）。
    AI解析列は着色しない（段階1の仕様に合わせる）。
    矛盾列のハイライトを openpyxl で保存する（.xlsm は keep_vba=True）。
    """
    if not wb_path or not os.path.exists(wb_path):
        return
    if _workbook_should_skip_openpyxl_io(wb_path):
        logging.info(
            "矛盾書式: ブックに「%s」があるため、openpyxl でのハイライトをスキップしました。",
            OPENPYXL_INCOMPATIBLE_SHEET_MARKER,
        )
        return
    keep_vba = str(wb_path).lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(wb_path, keep_vba=keep_vba)
        if sheet_name not in wb.sheetnames:
            logging.warning(f"矛盾書式: シート '{sheet_name}' は見つかりません。")
            return
        ws = wb[sheet_name]
        _plan_sheet_apply_conflict_styles_to_ws(ws, num_data_rows, conflicts_by_row)
        write_planning_conflict_highlight_sidecar(
            sheet_name, num_data_rows, conflicts_by_row
        )
        try:
            wb.save(wb_path)
        except OSError as e:
            logging.warning(
                "配台シートへの矛盾ハイライトの保存に失敗しました: %s", e
            )
            return
        _remove_planning_conflict_sidecar_safe()
        if conflicts_by_row:
            logging.info(
                "特別指定_備考と列の矛盾: %s 行を '%s' でハイライトしました。",
                len(conflicts_by_row),
                sheet_name,
            )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def _ai_planning_target_due_date(ai_dict):
    """AI JSON の完了・出蝷目標日から」配台の目標日1つを決ゝる（複数あれみ最も早い日＝厳しい方）。"""
    if not isinstance(ai_dict, dict):
        return None
    dates = []
    for k in ("target_completion_date", "ship_by_date", "latest_ship_date", "due_date"):
        d = parse_optional_date(ai_dict.get(k))
        if d is not None:
            dates.append(d)
    if not dates:
        return None
    return min(dates)
def _special_remark_implies_due_related_dispatch_priority(remark_raw: str) -> bool:
    """
    特別指定_備考に」紝期・期陝・最優先など「配台試行を剝に出す」愝図の文言はあるとし True。
    備考は記入されでいるの値では True にしない（AI 由来の目標日・開始日・優先度は使えない）。
    """
    if not remark_raw:
        return False
    s = str(remark_raw).strip()
    if not s or s.lower() in ("nan", "none"):
        return False
    n = unicodedata.normalize("NFKC", s)
    n_lower = n.casefold()
    # キーワードはユーザー入力（UTF-8 正しい表記）と一致させる。
    needles = (
        "紹期",
        "指定紹期",
        "回答紹期",
        "計画基準",
        "期日",
        "締切",
        "締め切り",
        "期限",
        "最優先",
        "至急",
        "急ぎ",
        "直ちに",
        "早急",
        "出荷",
        "紹入",
        "必着",
        "deadline",
        "デッドライン",
        "前倒し",
        "早めに",
        "厳守",
        "までに",
        "間に合わせ",
        "間に合い",
        "遅れない",
        "繧り上げ",
        "遅延",
        "優先配台",
        "先に配台",
        "完了予定",
        "本紹期",
        "回答期限",
    )
    return any(w.casefold() in n_lower for w in needles)
def _ai_task_special_entry_has_dispatch_priority_signals(ai_for_row) -> bool:
    """
    備考テキストのキーワード検出に漏れても、AI が既に priority / 日付 / 人数 等を返しているときは
    build_task_queue_from_planning_df 側で allow_ai_dispatch_priority_from_remark を立てる。
    preferred_operator のみのときは False（従来どおりセル側マージで足りる）。
    """
    if not isinstance(ai_for_row, dict) or not ai_for_row:
        return False
    meta = frozenset(
        {
            "process_name",
            "machine_name",
            "restrict_to_process_name",
            "restrict_to_machine_name",
            "preferred_operator",
        }
    )
    for k, v in ai_for_row.items():
        if k in meta or v is None:
            continue
        if k in (
            "ship_by_date",
            "target_completion_date",
            "latest_ship_date",
            "due_date",
            "start_date",
        ):
            if parse_optional_date(v) is not None:
                return True
        elif k == "start_time":
            if parse_time_str(str(v), None) is not None:
                return True
        elif k == "priority":
            if parse_optional_int(v) is not None:
                return True
        elif k == "required_op":
            try:
                if int(v) >= 1:
                    return True
            except (TypeError, ValueError):
                pass
        elif k == "task_efficiency":
            try:
                f = float(v)
                if f > 0 and f <= 1.0 + 1e-9:
                    return True
            except (TypeError, ValueError):
                pass
        elif k == "speed_override":
            try:
                if float(v) > 0:
                    return True
            except (TypeError, ValueError):
                pass
    return False
def _task_id_same_machine_due_tiebreak_key(task_id) -> tuple:
    """
    紝期基準（回答→指定）・機械名は坌も帯での試行順。
    Y3-24 は末尾の数値。Y4-1-1 のよごにポイフンは2つ以上あるとしは「最初の - の直後」の数値部を採用。
    """
    s = str(task_id or "").strip()
    if not s:
        return (2, 10**9, "")
    parts = s.split("-", 1)
    if len(parts) < 2:
        return (2, 10**9, s)
    rest = parts[1]
    if "-" in rest:
        first_seg = rest.split("-", 1)[0]
        try:
            return (0, int(first_seg), s)
        except ValueError:
            return (1, 10**9, s)
    tail = rest.strip()
    try:
        return (0, int(tail), s)
    except ValueError:
        return (1, 10**9, s)
def _optional_float_unprocessed_column(val):
    """
    配台計画シートの「未加工」セルを float 化する。
    空・無効なら None（結果_タスク一覧の残加工量は従来どおり m 換算にフォールバック）。
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("nan", "none", "-", "—", "―"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None
def _planning_product_length_cell_is_105_meters(row) -> bool:
    """
    特別ルール L8 用: 列「製品長」が **105 m** に相当するか。
    シートが **m** 単位なら 105、段階1由来の **mm** 整数なら 105000 を同一条件とする。
    """
    raw = _planning_df_cell_scalar(row, PLAN_COL_PRODUCT_LENGTH)
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return False
    try:
        xi = int(round(float(raw)))
    except (TypeError, ValueError):
        return False
    return xi in (105, 105000)
def _apply_dispatch_speed_special_rules_enumerated_md(
    *,
    row,
    task_id,
    machine: str,
    machine_name: str,
    speed: float,
) -> float:
    """
    リポジトリ直下 ``特別ルール列挙.md`` のうち **加工速度を 20 m/分へ上書き** する条件
    （L4 / L5 / L6 / L8）を適用する。

    呼び出し元で **列「加工速度_上書き」「加工速度」** および **global_speed_rules** による速度を
    確定した **あと** に呼ぶこと（本関数はさらに上書きするのみ）。
    L2（スライス・100 m は 3 名 or 20 m/分）は need 探索のフォールバックで別処理。
    """
    spd = float(speed)

    _prod_w = _planning_df_cell_scalar(row, PLAN_COL_PRODUCT_WIDTH)
    try:
        _prod_w_i = int(float(_prod_w)) if _prod_w is not None else None
    except (TypeError, ValueError):
        _prod_w_i = None

    # 特別ルール L4（SEC×SEC機 湖南）: 製品幅=935 のときは加工速度を 20m/分
    if (
        _normalize_process_name_for_rule_match(machine)
        == _normalize_process_name_for_rule_match("SEC")
        and _normalize_equipment_match_key(machine_name)
        == _normalize_equipment_match_key("SEC機　湖南")
        and _prod_w_i == 935
    ):
        spd = 20.0

    # 特別ルール L5（SEC×SEC機 湖南）: 製品幅<=680 のときは加工速度を 20m/分
    if (
        _normalize_process_name_for_rule_match(machine)
        == _normalize_process_name_for_rule_match("SEC")
        and _normalize_equipment_match_key(machine_name)
        == _normalize_equipment_match_key("SEC機　湖南")
        and _prod_w_i is not None
        and _prod_w_i <= 680
    ):
        spd = 20.0

    # 特別ルール L6（SEC×SEC機 湖南）: 依頼NOに「JR」又は「PN」が含まれている場合は加工速度を20m/分
    _tid_nfkc = unicodedata.normalize("NFKC", str(task_id or ""))
    if (
        _normalize_process_name_for_rule_match(machine)
        == _normalize_process_name_for_rule_match("SEC")
        and _normalize_equipment_match_key(machine_name)
        == _normalize_equipment_match_key("SEC機　湖南")
        and (("JR" in _tid_nfkc) or ("PN" in _tid_nfkc))
    ):
        spd = 20.0

    # 特別ルール L8（接続×熱融着機 湖南）: 製品長=105m のときは加工速度を20m/分
    if (
        _normalize_process_name_for_rule_match(machine)
        == _normalize_process_name_for_rule_match("接続")
        and _normalize_equipment_match_key(machine_name)
        == _normalize_equipment_match_key("熱融着機　湖南")
        and _planning_product_length_cell_is_105_meters(row)
    ):
        spd = 20.0

    return spd
def _load_stage2_in_progress_next_day_dispatch_overrides() -> dict[str, float]:
    """
    JavaFX 段階2直前ダイアログが書く JSON（entries[].task_id / process / machine_name / next_day_dispatch_m）。
    """
    path = (os.environ.get(ENV_STAGE2_IN_PROGRESS_NEXT_DAY_DISPATCH_JSON) or "").strip()
    if not path or not os.path.isfile(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logging.warning(
            "段階2: 加工途中・翌日配台量 JSON の読込に失敗（%s）: %s", path, e
        )
        return {}
    entries = data.get("entries") if isinstance(data, dict) else data
    if not isinstance(entries, list):
        return {}
    out: dict[str, float] = {}
    for ent in entries:
        if not isinstance(ent, dict):
            continue
        tid = planning_task_id_str_from_scalar(ent.get("task_id"))
        proc = str(ent.get("process") or "").strip()
        mname = str(ent.get("machine_name") or "").strip()
        if not tid:
            continue
        try:
            m = _sanitize_dispatch_qty_m(float(ent.get("next_day_dispatch_m")))
        except (TypeError, ValueError):
            m = 0.0
        out[_stage2_in_progress_next_day_dispatch_key(tid, proc, mname)] = m
    if out:
        logging.info(
            "段階2: 加工途中の翌日配台量を %s 行分 JSON から読み込みました（%s）。",
            len(out),
            path,
        )
    return out
def _rule_task_id(task) -> str:
    """特別ルール・WIP・同一依頼工程依存の集計キー。

    枝番タスク（入力3表）は ``rule_task_id``（元依頼NO）を、通常タスクは ``task_id`` を返す。
    配台の出力キー（``task_id``＝枝番依頼NO）と区別し、ルールは親単位で集計する。
    """
    if not isinstance(task, dict):
        return ""
    rid = str(task.get("rule_task_id") or "").strip()
    if rid:
        return rid
    return str(task.get("task_id") or "").strip()
