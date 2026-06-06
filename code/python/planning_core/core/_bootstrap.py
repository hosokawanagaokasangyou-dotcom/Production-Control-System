# -*- coding: utf-8 -*-
"""Shared imports for planning_core.core (exec bootstrap)."""
from __future__ import annotations

import base64
import calendar
import copy
import csv
import ctypes
import fnmatch
import hashlib
import itertools
import json
import logging
import math
import os
import pathlib
import random
import re
import shutil
import sys
import threading
import traceback
import unicodedata
import time as time_module
from collections import Counter, defaultdict
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta

import pandas as pd
from google import genai
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableStyleInfo

from dispatch_interval_mirror import DispatchIntervalMirror

from planning_core.bootstrap import (
    PlanningValidationError,
    _clear_stage2_blocking_message_file,
    _remove_prior_stage2_workbooks_and_prune_empty_dirs,
    _try_remove_path_with_retries,
    _write_stage2_blocking_message,
    api_payment_dir,
    json_data_dir,
    log_dir,
    output_dir,
)
from planning_core.dispatch_workspace import (
    ENV_PLAN_INPUT_PATH,
    ENV_PROCESSING_PLAN_PATH,
    _read_excel_tabular,
    _resolve_tabular_excel_header_row_0based,
    _resolve_tabular_sheet_name_calamine,
    plan_input_workbook_path_for_excel_ops,
    read_tabular_dataframe,
    resolve_actual_detail_workbook_path,
    resolve_processing_plan_path_from_env,
    resolve_result_dispatch_table_output_dir,
)
from planning_core.input_resolution import (
    ENV_EXCLUDE_RULES_JSON,
    ENV_GLOBAL_PRIORITY_OVERRIDE_PATH,
    ENV_RESULT_TASK_COLUMN_CONFIG_CSV,
    resolve_actuals_workbook_path,
    resolve_column_config_workbook_path,
    resolve_data_extraction_workbook_path,
)
from planning_core.plan_workbook_sidecar import (
    normalized_workbook_json_path,
    read_result_task_dataframe,
    write_member_schedule_workbook_json,
    write_production_plan_logical_view_json,
    write_production_plan_workbook_json,
    write_result_task_json_sidecar,
)
from planning_core.stage2_output_naming import (
    format_stage2_stamp,
    member_workbook_filename,
    plan_workbook_filename,
)

