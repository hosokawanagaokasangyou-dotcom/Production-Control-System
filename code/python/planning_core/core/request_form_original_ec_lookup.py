# -*- coding: utf-8 -*-
"""依頼書原本 xlsm から依頼NO別の EC面・加工内容を読む（受注 N 列が空のときの補完用）。"""
from __future__ import annotations

import logging
import os
import re
import unicodedata

from planning_core.core.request_form_juchu_lookup import normalize_irai_no_key

ENV_REQUEST_FORM_ORIGINAL_DIR = "PM_AI_REQUEST_FORM_ORIGINAL_DIR"
_ORIGINAL_SHEET_RE = re.compile(
    r"^[A-Z]+\d+-\d+$|^[A-Z]\d+-\d+-\d+$", re.IGNORECASE
)
_SKIP_FILES = frozenset({"加工依頼書入力.xlsm"})

# RequestFormOriginalCellLayout と同一（POI 0-based）
_ROW_IRAI = 4  # Excel 5行 R列
_COL_IRAI = 17  # R
_PRODUCT_ROWS = (9, 10, 11)  # Excel 10–12
_COL_EC = 35  # AJ
_COL_PROCESS = 8  # I
_PROCESS_ROWS = (12, 13, 14, 15, 16)  # Excel 13–17


def _col_letter_to_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


assert _col_letter_to_index("R") == _COL_IRAI
assert _col_letter_to_index("AJ") == _COL_EC
assert _col_letter_to_index("I") == _COL_PROCESS


def _cell_text(ws, row_idx: int, col_idx: int) -> str:
    try:
        row = ws[row_idx + 1]
    except (IndexError, TypeError):
        return ""
    if col_idx >= len(row):
        return ""
    cell = row[col_idx]
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _normalize_sheet_key(name: str) -> str:
    s = unicodedata.normalize("NFKC", str(name or "").strip()).upper()
    return "".join(s.split())


def _read_process_content(ws) -> str:
    steps: list[str] = []
    for r in _PROCESS_ROWS:
        v = _cell_text(ws, r, _COL_PROCESS)
        if v:
            steps.append(v)
    return ",".join(steps)


def _read_ec_men_from_product_rows(ws) -> str:
    for r in _PRODUCT_ROWS:
        v = _cell_text(ws, r, _COL_EC)
        if v:
            return v
    return ""


def _resolve_original_dir() -> str:
    return (os.environ.get(ENV_REQUEST_FORM_ORIGINAL_DIR) or "").strip()


def load_request_form_original_ec_lookup() -> dict[str, dict[str, str]]:
    """
    依頼NO 正規化キー → {"EC面": str, "加工内容": str}。
    原本フォルダは読み取り専用。未到達時は空 dict。
    """
    dir_path = _resolve_original_dir()
    if not dir_path or not os.path.isdir(dir_path):
        logging.info(
            "段階1: 依頼書原本フォルダ未設定のため EC面 原本補完をスキップ: %s",
            ENV_REQUEST_FORM_ORIGINAL_DIR,
        )
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError as ex:
        logging.warning("段階1: openpyxl 未導入のため EC面 原本補完をスキップ: %s", ex)
        return {}

    out: dict[str, dict[str, str]] = {}
    try:
        names = sorted(os.listdir(dir_path))
    except OSError as ex:
        logging.warning("段階1: 依頼書原本フォルダ一覧失敗: %s", ex)
        return {}

    for name in names:
        if not name.lower().endswith(".xlsm"):
            continue
        if name.startswith("~$") or name in _SKIP_FILES:
            continue
        path = os.path.join(dir_path, name)
        if not os.path.isfile(path):
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as ex:
            logging.warning("段階1: 依頼書原本読込スキップ %s: %s", name, ex)
            continue
        try:
            for sheet_name in wb.sheetnames:
                sheet_key = _normalize_sheet_key(sheet_name)
                if not _ORIGINAL_SHEET_RE.match(sheet_key):
                    continue
                ws = wb[sheet_name]
                irai = _cell_text(ws, _ROW_IRAI, _COL_IRAI)
                key = normalize_irai_no_key(irai or sheet_name)
                if not key:
                    continue
                ec_men = _read_ec_men_from_product_rows(ws)
                kako = _read_process_content(ws)
                if not ec_men and not kako:
                    continue
                prev = out.get(key, {})
                merged = {
                    "EC面": ec_men or prev.get("EC面", ""),
                    "加工内容": kako or prev.get("加工内容", ""),
                }
                out[key] = merged
        finally:
            try:
                wb.close()
            except Exception:
                pass

    logging.info(
        "段階1: 依頼書原本 EC面 lookup 件数=%d dir=%s",
        len(out),
        dir_path,
    )
    return out
