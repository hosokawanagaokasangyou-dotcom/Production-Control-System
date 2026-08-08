# -*- coding: utf-8 -*-
"""openpyxl styles for APP_機械カレンダー sheet."""

from __future__ import annotations

from datetime import date
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from planning_core.core.attendance_excel_style import (
    _CENTER,
    _GRID_BORDER,
    _month_origin,
    _weekday_labels,
    cached_fill,
    FONT_BANNER,
    FONT_HEADER,
    FONT_WORKING,
    GRID_START_COL,
    GRID_START_ROW,
    MONTH_BAND_H,
    MONTH_BLOCK_W,
    MONTHS_PER_ROW,
    FILL_WEEKEND_COL,
    FILL_WEEKEND_HDR,
    FILL_MONTH_TITLE,
    FONT_PUBLIC,
    insert_menu_back_link_row,
    sheet_hyperlink_to_cell,
)
from planning_core.core.columns import _result_font

FILL_OCCUPIED = "C084FC"
FILL_AVAILABLE = "DCFCE7"
FONT_OCCUPIED = "4C1D95"

_WEEKDAY_JA = ("月", "火", "水", "木", "金", "土", "日")
_THIN = Side(style="thin", color="CBD5E1")
_HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LINK_FONT = _result_font(size=9, color="0563C1", underline="single")
_MENU_ROW_SHIFT = 1


def write_machine_calendar_flat_table(
    ws,
    columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
    fy_label: int,
    fiscal_start: date,
    fiscal_end: date,
) -> dict[str, int]:
    """機械カレンダーをフラット表で書き込み。日付→先頭行（メニュー行挿入後）のマップを返す。"""
    title = (
        f"{fy_label}年度 機械カレンダー（{fiscal_start.isoformat()} ～ {fiscal_end.isoformat()}）"
    )
    last_col = max(3, 2 + len(columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c1 = ws.cell(1, 1, title)
    c1.font = _result_font(size=14, bold=True, color=FONT_BANNER)
    c1.alignment = Alignment(horizontal="left", vertical="center")

    legend = (
        "■稼働可能（空）  ■占有（*）"
        "  — 30分スロット。編集はアプリの機械カレンダータブで行う。"
        " 日付ジャンプはシート「APP_機械カレンダー_日付」の月カレンダーから。"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1, legend).font = _result_font(size=9, color=FONT_HEADER)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws.cell(3, 1, "※本シートはアプリ自動出力（APP_機械カレンダー）。").font = _result_font(
        size=8, color=FONT_HEADER
    )

    hdr_row_proc = 4
    hdr_row_machine = 5
    data_start = 6

    time_hdr = ws.cell(hdr_row_proc, 1, "日付時刻")
    time_hdr.font = _result_font(size=9, bold=True, color=FONT_HEADER)
    time_hdr.fill = cached_fill("E2E8F0")
    time_hdr.alignment = _CENTER
    time_hdr.border = _HEADER_BORDER

    wd_hdr = ws.cell(hdr_row_proc, 2, "曜日")
    wd_hdr.font = _result_font(size=9, bold=True, color=FONT_HEADER)
    wd_hdr.fill = cached_fill("E2E8F0")
    wd_hdr.alignment = _CENTER
    wd_hdr.border = _HEADER_BORDER

    ws.merge_cells(
        start_row=hdr_row_proc, start_column=1, end_row=hdr_row_machine, end_column=1
    )
    ws.merge_cells(
        start_row=hdr_row_proc, start_column=2, end_row=hdr_row_machine, end_column=2
    )

    for i, col in enumerate(columns):
        col_idx = 3 + i
        proc = str(col.get("process") or "").strip()
        machine = str(col.get("machine") or "").strip()
        p_cell = ws.cell(hdr_row_proc, col_idx, proc)
        p_cell.font = _result_font(size=9, bold=True, color=FONT_HEADER)
        p_cell.fill = cached_fill("E2EFDA")
        p_cell.alignment = _CENTER
        p_cell.border = _HEADER_BORDER

        m_cell = ws.cell(hdr_row_machine, col_idx, machine)
        m_cell.font = _result_font(size=9, color=FONT_WORKING)
        m_cell.fill = cached_fill("E2EFDA")
        m_cell.alignment = _CENTER
        m_cell.border = _HEADER_BORDER

        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(10, min(16, len(machine) + 2))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 5

    day_first_row: dict[str, int] = {}
    for r_idx, row in enumerate(rows):
        excel_row = data_start + r_idx
        slot_s = str(row.get("slot") or "").strip()
        slot_display = slot_s.replace("T", " ").replace("-", "/")[:16]
        wd = str(row.get("weekday") or "").strip()
        if slot_s:
            try:
                from datetime import datetime

                slot_dt = datetime.fromisoformat(slot_s)
                if not wd:
                    wd = _WEEKDAY_JA[slot_dt.weekday()]
                d_key = slot_dt.date().isoformat()
                if d_key not in day_first_row:
                    day_first_row[d_key] = excel_row + _MENU_ROW_SHIFT
            except ValueError:
                pass

        t_cell = ws.cell(excel_row, 1, slot_display)
        t_cell.font = _result_font(size=9, color=FONT_BANNER)
        t_cell.alignment = _CENTER
        t_cell.border = _GRID_BORDER

        w_cell = ws.cell(excel_row, 2, wd)
        w_cell.font = _result_font(size=9, color=FONT_HEADER)
        w_cell.alignment = _CENTER
        w_cell.border = _GRID_BORDER

        cells_map = row.get("cells") or {}
        for i, col in enumerate(columns):
            ek = str(col.get("equipment_key") or "").strip()
            val = str(cells_map.get(ek) or "").strip()
            col_idx = 3 + i
            display = val if val else ""
            cell = ws.cell(excel_row, col_idx, display)
            cell.alignment = _CENTER
            cell.border = _GRID_BORDER
            if val in ("*", "＊", "※"):
                cell.fill = cached_fill(FILL_OCCUPIED)
                cell.font = _result_font(size=9, bold=True, color=FONT_OCCUPIED)
            else:
                cell.fill = cached_fill(FILL_AVAILABLE)
                cell.font = _result_font(size=9, color=FONT_WORKING)

    ws.freeze_panes = "C6"
    ws.sheet_view.showGridLines = False
    insert_menu_back_link_row(ws)
    return day_first_row


def write_machine_calendar_date_picker_sheet(
    ws,
    fy_label: int,
    fiscal_start: date,
    fiscal_end: date,
    months: list[tuple[int, int]],
    day_to_row: dict[str, int],
    target_sheet_name: str,
) -> None:
    """月カレンダー形式で日付を選び、機械カレンダー表の該当日先頭行へジャンプする。"""
    import calendar as cal_mod

    last_col = MONTHS_PER_ROW * MONTH_BLOCK_W - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(
        1,
        1,
        f"{fy_label}年度 機械カレンダー 日付ジャンプ（{fiscal_start.isoformat()} ～ {fiscal_end.isoformat()}）",
    )
    title.font = _result_font(size=14, bold=True, color=FONT_BANNER)
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(
        2,
        1,
        "月カレンダーの日付をクリックすると「APP_機械カレンダー」シートのその日 08:00 付近へ移動します。",
    ).font = _result_font(size=9, color=FONT_HEADER)

    today = date.today()
    for idx, (year, month) in enumerate(months):
        r0, c0 = _month_origin(idx)
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 6)
        title_cell = ws.cell(r0, c0, f"{month}月（{year}）")
        title_cell.font = _result_font(size=11, bold=True, color=FONT_BANNER)
        title_cell.fill = cached_fill(FILL_MONTH_TITLE)
        title_cell.alignment = _CENTER

        for i, wd_label in enumerate(_weekday_labels()):
            hdr = ws.cell(r0 + 1, c0 + i, wd_label)
            hdr.font = _result_font(size=8, color=FONT_HEADER)
            hdr.fill = cached_fill(FILL_WEEKEND_HDR)
            hdr.alignment = _CENTER
            hdr.border = _GRID_BORDER
            if i >= 5:
                hdr.font = _result_font(size=8, bold=True, color=FONT_PUBLIC)

        ym_days = cal_mod.monthrange(year, month)[1]
        first = date(year, month, 1)
        offset = first.weekday()
        for day_num in range(1, ym_days + 1):
            d = date(year, month, day_num)
            if d < fiscal_start or d > fiscal_end:
                continue
            cell_idx = offset + day_num - 1
            row = r0 + 2 + cell_idx // 7
            col = c0 + cell_idx % 7
            d_key = d.isoformat()
            target_row = day_to_row.get(d_key)
            cell = ws.cell(row, col, str(day_num))
            cell.alignment = _CENTER
            cell.border = _GRID_BORDER
            if d.weekday() >= 5:
                cell.fill = cached_fill(FILL_WEEKEND_COL)
                cell.font = _result_font(size=9, color=FONT_HEADER)
            else:
                cell.fill = cached_fill("DCFCE7")
                cell.font = _result_font(size=9, color=FONT_WORKING)
            if target_row is not None:
                cell.hyperlink = sheet_hyperlink_to_cell(target_sheet_name, target_row, 1)
                cell.font = _LINK_FONT
            if d == today:
                cell.border = Border(
                    left=Side(style="medium", color="F59E0B"),
                    right=Side(style="medium", color="F59E0B"),
                    top=Side(style="medium", color="F59E0B"),
                    bottom=Side(style="medium", color="F59E0B"),
                )

        spacer_col = c0 + 7
        ws.column_dimensions[get_column_letter(spacer_col)].width = 1.0
        for dc in range(7):
            ws.column_dimensions[get_column_letter(c0 + dc)].width = 4.5

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    insert_menu_back_link_row(ws)
