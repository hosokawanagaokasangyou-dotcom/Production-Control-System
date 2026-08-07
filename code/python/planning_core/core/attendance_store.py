# -*- coding: utf-8 -*-
"""Attendance store: company calendar + member attendance (JSON canonical)."""

from __future__ import annotations

import calendar
import copy
import json
import logging
import os
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from planning_core.core.attendance_paths import (
    APP_MASTER_COMPANY_SHEET,
    APP_MASTER_MEMBER_SHEET_PREFIX,
    attendance_data_json_path,
    is_app_master_export_sheet,
)
from planning_core.core.japanese_holidays import fetch_national_holidays_for_year

logger = logging.getLogger(__name__)

FORMAT_VERSION = 1
DEFAULT_REG_START = time(8, 45)
DEFAULT_REG_END = time(17, 0)
BREAK1_START = time(12, 0)
BREAK1_END = time(12, 50)
BREAK2_START = time(14, 45)
BREAK2_END = time(15, 0)

DAY_KIND_WORKING = "working_day"
DAY_KIND_PUBLIC = "public_holiday"
DAY_KIND_SPECIAL = "special_holiday"

PRESET_WORK = "WORK"
PRESET_OFF_FULL = "OFF_FULL"
PRESET_PAID_LEAVE = "PAID_LEAVE"
PRESET_ABSENT = "ABSENT"
PRESET_OFF_AM = "OFF_AM"
PRESET_OFF_PM = "OFF_PM"
PRESET_NO_DISPATCH = "NO_DISPATCH"
PRESET_HOLIDAY_WORK = "HOLIDAY_WORK"
PRESET_HOLIDAY_WORK_AM = "HOLIDAY_WORK_AM"
PRESET_HOLIDAY_WORK_PM = "HOLIDAY_WORK_PM"

HOURLY_AVAILABLE = "available"
HOURLY_BREAK = "break"
HOURLY_LEAVE = "leave"
HOURLY_AWAY = "away"
HOURLY_OFF_SHIFT = "off_shift"


def empty_store(year: int | None = None) -> dict:
    y = year or date.today().year
    return {
        "format_version": FORMAT_VERSION,
        "meta": {
            "schema": "pm-ai-attendance-store",
            "updated_at": None,
            "company_calendar_revision": 0,
            "member_attendance_revision": 0,
            "calendar_xlsx_export_at": None,
            "holidays_fetched_at": None,
        },
        "company_calendar": {
            "year": y,
            "factory_hours": {"start": "06:00", "end": "22:00"},
            "regular_hours": {
                "start": DEFAULT_REG_START.strftime("%H:%M"),
                "end": DEFAULT_REG_END.strftime("%H:%M"),
            },
            "days": {},
        },
        "member_roster": [],
        "member_attendance": {},
    }


def load_attendance_store(path: Path | None = None) -> dict:
    p = path or attendance_data_json_path()
    if not p.is_file():
        return empty_store()
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        if isinstance(data, dict) and data.get("format_version") == FORMAT_VERSION:
            return data
    except (OSError, json.JSONDecodeError) as e:
        logger.warning("attendance-data.json 読込失敗: %s", e)
    return empty_store()


def save_attendance_store(
    store: dict,
    path: Path | None = None,
    *,
    history_kind: str = "auto_save",
    history_label: str = "保存",
    skip_history: bool = False,
) -> Path:
    p = path or attendance_data_json_path()
    meta = store.setdefault("meta", {})
    meta["updated_at"] = datetime.now().isoformat(timespec="seconds")
    p.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(store, ensure_ascii=False, indent=2)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(payload, encoding="utf-8")
    os.replace(tmp, p)
    if not skip_history:
        try:
            from planning_core.core.attendance_history_store import append_attendance_snapshot

            append_attendance_snapshot(
                p, kind=history_kind, label=history_label
            )
        except Exception as e:
            logger.warning("attendance JSON 世代退避失敗: %s", e)
    return p


def apply_national_holidays_to_company_calendar(
    store: dict,
    year: int,
    overwrite: bool = False,
    include_weekends: bool = False,
    force_online: bool = False,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    bump_revision: bool = True,
) -> dict:
    holidays = fetch_national_holidays_for_year(year, force_online=force_online)
    days = store.setdefault("company_calendar", {}).setdefault("days", {})
    applied = 0
    skipped = 0
    for item in holidays:
        d_key = item.get("date")
        name = item.get("name") or "祝日"
        if not d_key:
            continue
        try:
            d_val = date.fromisoformat(d_key)
        except ValueError:
            continue
        if date_from is not None and d_val < date_from:
            continue
        if date_to is not None and d_val > date_to:
            continue
        if d_key in days and days[d_key].get("manual_edit") and not overwrite:
            skipped += 1
            continue
        if d_key in days and not overwrite and days[d_key].get("source") != "national_holiday":
            if days[d_key].get("kind") != DAY_KIND_WORKING:
                skipped += 1
                continue
        days[d_key] = {
            "kind": DAY_KIND_PUBLIC,
            "label": name,
            "source": "national_holiday",
        }
        applied += 1
    if include_weekends:
        start = date_from if date_from is not None else date(year, 1, 1)
        end = date_to if date_to is not None else date(year, 12, 31)
        d = start
        while d <= end:
            if d.weekday() >= 5:
                key = d.isoformat()
                if key not in days or overwrite:
                    if key in days and not overwrite and days[key].get("manual_edit"):
                        skipped += 1
                    else:
                        days[key] = {
                            "kind": DAY_KIND_PUBLIC,
                            "label": "土曜" if d.weekday() == 5 else "日曜",
                            "source": "weekend_init",
                        }
                        applied += 1
                else:
                    skipped += 1
            d += timedelta(days=1)
    store["company_calendar"]["year"] = year
    if bump_revision:
        store["meta"]["holidays_fetched_at"] = datetime.now().isoformat(timespec="seconds")
        store["meta"]["company_calendar_revision"] = int(
            store["meta"].get("company_calendar_revision") or 0
        ) + 1
    return {"applied": applied, "skipped": skipped, "source": "national_holiday"}


def company_day_kind(store: dict, day: date) -> str:
    key = day.isoformat()
    entry = store.get("company_calendar", {}).get("days", {}).get(key)
    if not entry:
        if day.weekday() >= 5:
            return DAY_KIND_PUBLIC
        return DAY_KIND_WORKING
    return entry.get("kind") or DAY_KIND_WORKING


def day_preset_from_company(store: dict, day: date) -> str:
    kind = company_day_kind(store, day)
    if kind == DAY_KIND_WORKING:
        return PRESET_WORK
    return PRESET_OFF_FULL


def preset_to_leave_and_times(
    preset: str,
    day: date,
    reg_start: time = DEFAULT_REG_START,
    reg_end: time = DEFAULT_REG_END,
    company_kind: str | None = None,
) -> dict:
    """Map day preset to attendance row fields (VBA CollectAttendanceFromCalendar 対称)."""
    row: dict[str, Any] = {
        "day_preset": preset,
        "clock_in": None,
        "clock_out": None,
        "leave_type": "",
        "remark": "",
        "breaks": [],
        "overtime_minutes": 0,
        "efficiency": 1.0,
        "eligible_for_assignment": True,
        "hourly": {},
        "manual_edit": False,
        "comment": "",
    }
    if preset == PRESET_WORK:
        row.update(
            {
                "clock_in": reg_start.strftime("%H:%M"),
                "clock_out": reg_end.strftime("%H:%M"),
                "leave_type": "通常",
                "remark": "通常",
                "breaks": [
                    {"start": BREAK1_START.strftime("%H:%M"), "end": BREAK1_END.strftime("%H:%M")},
                    {"start": BREAK2_START.strftime("%H:%M"), "end": BREAK2_END.strftime("%H:%M")},
                ],
            }
        )
    elif preset == PRESET_OFF_FULL:
        if company_kind in (DAY_KIND_PUBLIC, DAY_KIND_SPECIAL):
            row["leave_type"] = "公休"
            row["remark"] = "公休"
        else:
            row["leave_type"] = "年休"
            row["remark"] = "休"
        row["eligible_for_assignment"] = False
    elif preset == PRESET_PAID_LEAVE:
        row["leave_type"] = "年休"
        row["remark"] = "年休"
        row["eligible_for_assignment"] = False
    elif preset == PRESET_ABSENT:
        row["leave_type"] = "欠勤"
        row["remark"] = "欠勤"
        row["eligible_for_assignment"] = False
    elif preset == PRESET_OFF_AM:
        row.update(
            {
                "clock_in": BREAK1_END.strftime("%H:%M"),
                "clock_out": reg_end.strftime("%H:%M"),
                "leave_type": "前休",
                "remark": "前休",
                "breaks": [
                    {
                        "start": BREAK2_START.strftime("%H:%M"),
                        "end": BREAK2_END.strftime("%H:%M"),
                    }
                ],
            }
        )
    elif preset == PRESET_OFF_PM:
        row.update(
            {
                "clock_in": reg_start.strftime("%H:%M"),
                "clock_out": BREAK1_START.strftime("%H:%M"),
                "leave_type": "後休",
                "remark": "後休",
            }
        )
    elif preset == PRESET_NO_DISPATCH:
        row["leave_type"] = "-"
        row["remark"] = "-"
        row["eligible_for_assignment"] = False
    elif preset == PRESET_HOLIDAY_WORK:
        row.update(
            {
                "clock_in": reg_start.strftime("%H:%M"),
                "clock_out": reg_end.strftime("%H:%M"),
                "leave_type": "休日出勤",
                "remark": "休日出勤",
                "breaks": [
                    {"start": BREAK1_START.strftime("%H:%M"), "end": BREAK1_END.strftime("%H:%M")},
                    {"start": BREAK2_START.strftime("%H:%M"), "end": BREAK2_END.strftime("%H:%M")},
                ],
            }
        )
    elif preset == PRESET_HOLIDAY_WORK_AM:
        row.update(
            {
                "clock_in": reg_start.strftime("%H:%M"),
                "clock_out": BREAK1_START.strftime("%H:%M"),
                "leave_type": "午前休出",
                "remark": "午前休出",
            }
        )
    elif preset == PRESET_HOLIDAY_WORK_PM:
        row.update(
            {
                "clock_in": BREAK1_END.strftime("%H:%M"),
                "clock_out": reg_end.strftime("%H:%M"),
                "leave_type": "午後休出",
                "remark": "午後休出",
                "breaks": [
                    {
                        "start": BREAK2_START.strftime("%H:%M"),
                        "end": BREAK2_END.strftime("%H:%M"),
                    }
                ],
            }
        )
    return row


LEGACY_PROTECTED_MEMBER_PRESETS = frozenset({PRESET_NO_DISPATCH})


def _skip_company_calendar_member_sync(existing: dict | None, only_unedited: bool) -> bool:
    if not only_unedited or not existing:
        return False
    if existing.get("manual_edit"):
        return True
    if str(existing.get("day_preset") or "") in LEGACY_PROTECTED_MEMBER_PRESETS:
        return True
    return False


def apply_company_calendar_to_members(
    store: dict,
    members: list[str],
    year: int,
    month: int,
    only_unedited: bool = True,
) -> dict:
    applied = 0
    skipped = 0
    ma = store.setdefault("member_attendance", {})
    ym_days = calendar.monthrange(year, month)[1]
    for day_num in range(1, ym_days + 1):
        d = date(year, month, day_num)
        d_key = d.isoformat()
        preset = day_preset_from_company(store, d)
        kind = company_day_kind(store, d)
        day_bucket = ma.setdefault(d_key, {})
        for member in members:
            existing = day_bucket.get(member)
            if _skip_company_calendar_member_sync(existing, only_unedited):
                skipped += 1
                continue
            day_bucket[member] = preset_to_leave_and_times(preset, d, company_kind=kind)
            applied += 1
    store["meta"]["member_attendance_revision"] = int(
        store["meta"].get("member_attendance_revision") or 0
    ) + 1
    return {"applied": applied, "skipped": skipped}


def apply_company_calendar_to_members_fiscal(
    store: dict,
    members: list[str],
    fiscal_year_label: int,
    start_month: int = 4,
    start_day: int = 1,
    only_unedited: bool = True,
) -> dict:
    """会計年度の全日についてメンバー勤怠を会社カレンダーに合わせる。"""
    start, end = fiscal_year_date_range(fiscal_year_label, start_month, start_day)
    applied = 0
    skipped = 0
    ma = store.setdefault("member_attendance", {})
    d = start
    while d <= end:
        d_key = d.isoformat()
        preset = day_preset_from_company(store, d)
        kind = company_day_kind(store, d)
        day_bucket = ma.setdefault(d_key, {})
        for member in members:
            existing = day_bucket.get(member)
            if _skip_company_calendar_member_sync(existing, only_unedited):
                skipped += 1
                continue
            day_bucket[member] = preset_to_leave_and_times(preset, d, company_kind=kind)
            applied += 1
        d += timedelta(days=1)
    store["meta"]["member_attendance_revision"] = int(
        store["meta"].get("member_attendance_revision") or 0
    ) + 1
    return {
        "applied": applied,
        "skipped": skipped,
        "fiscal_year": fiscal_year_label,
        "fiscal_start": start.isoformat(),
        "fiscal_end": end.isoformat(),
    }


def apply_member_attendance_patch(store: dict, patch: dict) -> dict:
    """Merge member cell edits from UI patch { year, month, cells, member_roster? }."""
    roster_applied = 0
    if "member_roster" in patch:
        from planning_core.core.attendance_member_roster import apply_member_roster_patch

        roster_result = apply_member_roster_patch(store, patch.get("member_roster") or [])
        roster_applied = int(roster_result.get("roster_count") or 0)

    cells_patch = patch.get("cells") or {}
    ma = store.setdefault("member_attendance", {})
    applied = 0
    for d_key, per_member in cells_patch.items():
        try:
            d = date.fromisoformat(str(d_key))
        except ValueError:
            continue
        bucket = ma.setdefault(d_key, {})
        if not isinstance(per_member, dict):
            continue
        for member, cell_patch in per_member.items():
            if not isinstance(cell_patch, dict):
                continue
            preset = cell_patch.get("day_preset")
            if preset:
                kind = company_day_kind(store, d)
                existing = bucket.get(member)
                entry = preset_to_leave_and_times(str(preset), d, company_kind=kind)
                entry["manual_edit"] = True
                if "comment" in cell_patch:
                    entry["comment"] = str(cell_patch.get("comment") or "")
                elif isinstance(existing, dict) and existing.get("comment"):
                    entry["comment"] = existing.get("comment")
                if cell_patch.get("overtime_minutes") is not None:
                    entry["overtime_minutes"] = int(cell_patch["overtime_minutes"])
                hourly = cell_patch.get("hourly")
                if isinstance(hourly, dict):
                    entry["hourly"] = hourly
                elif isinstance(existing, dict) and isinstance(existing.get("hourly"), dict):
                    entry["hourly"] = copy.deepcopy(existing.get("hourly"))
                bucket[member] = entry
                applied += 1
            else:
                existing = bucket.get(member)
                if existing is None:
                    existing = preset_to_leave_and_times(
                        day_preset_from_company(store, d), d, company_kind=company_day_kind(store, d)
                    )
                existing = copy.deepcopy(existing)
                existing.update(cell_patch)
                existing["manual_edit"] = True
                bucket[member] = existing
                applied += 1
    if applied > 0 or roster_applied > 0:
        store["meta"]["member_attendance_revision"] = int(
            store["meta"].get("member_attendance_revision") or 0
        ) + 1
    return {"applied": applied, "roster_count": roster_applied}


def build_editor_payload(store: dict, members: list[str], year: int, month: int) -> dict:
    ym_days = calendar.monthrange(year, month)[1]
    dates: list[str] = []
    for day_num in range(1, ym_days + 1):
        dates.append(date(year, month, day_num).isoformat())
    cells: dict[str, dict[str, dict]] = {}
    ma = store.get("member_attendance", {})
    cc_days = store.get("company_calendar", {}).get("days", {})
    for d_key in dates:
        cells[d_key] = {}
        d = date.fromisoformat(d_key)
        cc = cc_days.get(d_key, {})
        for m in members:
            entry = ma.get(d_key, {}).get(m) or preset_to_leave_and_times(
                day_preset_from_company(store, d), d
            )
            cells[d_key][m] = {
                "day_preset": entry.get("day_preset", PRESET_WORK),
                "leave_type": entry.get("leave_type", ""),
                "is_working": bool(entry.get("clock_in")),
                "overtime_minutes": int(entry.get("overtime_minutes") or 0),
                "manual_edit": bool(entry.get("manual_edit")),
                "company_kind": cc.get("kind") or company_day_kind(store, d),
                "hourly": entry.get("hourly") or {},
                "comment": entry.get("comment") or "",
            }
    from planning_core.core.attendance_member_roster import (
        ensure_member_roster,
        primary_roles_map,
    )

    roster = ensure_member_roster(store)
    return {
        "format_version": 1,
        "ok": True,
        "year": year,
        "month": month,
        "members": members,
        "member_roster": roster,
        "primary_roles": primary_roles_map(store),
        "dates": dates,
        "cells": cells,
        "company_calendar_revision": store.get("meta", {}).get("company_calendar_revision", 0),
        "member_attendance_revision": store.get("meta", {}).get("member_attendance_revision", 0),
        "member_roster_revision": store.get("meta", {}).get("member_roster_revision", 0),
    }


def build_company_calendar_payload(store: dict, year: int) -> dict:
    days = store.get("company_calendar", {}).get("days", {})
    year_days = {k: v for k, v in sorted(days.items()) if k.startswith(f"{year}-")}
    return {
        "format_version": 1,
        "ok": True,
        "year": year,
        "days": year_days,
        "revision": store.get("meta", {}).get("company_calendar_revision", 0),
    }


def fiscal_year_date_range(
    fiscal_year_label: int,
    start_month: int = 4,
    start_day: int = 1,
) -> tuple[date, date]:
    start = _clamp_fiscal_day(fiscal_year_label, start_month, start_day)
    next_start = _clamp_fiscal_day(fiscal_year_label + 1, start_month, start_day)
    end = next_start - timedelta(days=1)
    return start, end


def initialize_company_calendar(
    store: dict,
    fiscal_year: int,
    start_month: int = 4,
    start_day: int = 1,
) -> dict:
    """
    表示中の会計年度範囲の会社カレンダー日次エントリを削除し、平日／週末の既定解釈に戻す。
    メンバー勤怠・他年度の日次データは変更しない。
    """
    start, end = fiscal_year_date_range(fiscal_year, start_month, start_day)
    days = store.setdefault("company_calendar", {}).setdefault("days", {})
    removed = 0
    d = start
    while d <= end:
        key = d.isoformat()
        if key in days:
            del days[key]
            removed += 1
        d += timedelta(days=1)
    store["company_calendar"]["year"] = fiscal_year
    meta = store.setdefault("meta", {})
    meta["fiscal_start_month"] = start_month
    meta["fiscal_start_day"] = start_day
    meta["company_calendar_revision"] = int(meta.get("company_calendar_revision") or 0) + 1
    return {
        "removed": removed,
        "fiscal_start": start.isoformat(),
        "fiscal_end": end.isoformat(),
    }


def _clamp_fiscal_day(year: int, month: int, day: int) -> date:
    import calendar as cal_mod

    max_day = cal_mod.monthrange(year, month)[1]
    return date(year, month, min(day, max_day))


def enrich_company_calendar_days_with_national_holidays(
    days: dict,
    start: date,
    end: date,
    *,
    force_online: bool = False,
) -> dict:
    """Merge national holidays into day map for UI display (does not write store)."""
    merged = dict(days)
    cal_years = sorted({start.year, end.year})
    for cal_year in cal_years:
        for item in fetch_national_holidays_for_year(cal_year, force_online=force_online):
            d_key = item.get("date")
            name = item.get("name") or "祝日"
            if not d_key:
                continue
            try:
                d_val = date.fromisoformat(d_key)
            except ValueError:
                continue
            if d_val < start or d_val > end:
                continue
            existing = merged.get(d_key)
            if existing is None:
                merged[d_key] = {
                    "kind": DAY_KIND_PUBLIC,
                    "label": name,
                    "source": "national_holiday",
                }
            elif (
                not existing.get("manual_edit")
                and existing.get("kind") == DAY_KIND_WORKING
            ):
                merged[d_key] = {
                    "kind": DAY_KIND_PUBLIC,
                    "label": name,
                    "source": "national_holiday",
                }
    return merged


def build_company_calendar_payload_fiscal(
    store: dict,
    fiscal_year_label: int,
    start_month: int = 4,
    start_day: int = 1,
) -> dict:
    start, end = fiscal_year_date_range(fiscal_year_label, start_month, start_day)
    days = store.get("company_calendar", {}).get("days", {})
    filtered: dict[str, Any] = {}
    for k, v in sorted(days.items()):
        try:
            d = date.fromisoformat(k)
        except ValueError:
            continue
        if start <= d <= end:
            filtered[k] = v
    filtered = enrich_company_calendar_days_with_national_holidays(
        filtered, start, end, force_online=False
    )
    return {
        "format_version": 1,
        "ok": True,
        "year": fiscal_year_label,
        "fiscal_year": fiscal_year_label,
        "fiscal_start_month": start_month,
        "fiscal_start_day": start_day,
        "fiscal_start": start.isoformat(),
        "fiscal_end": end.isoformat(),
        "days": filtered,
        "revision": store.get("meta", {}).get("company_calendar_revision", 0),
    }


def apply_holidays_to_fiscal_year(
    store: dict,
    fiscal_year_label: int,
    start_month: int = 4,
    start_day: int = 1,
    overwrite: bool = False,
    include_weekends: bool = False,
    force_online: bool = False,
) -> dict:
    start, end = fiscal_year_date_range(fiscal_year_label, start_month, start_day)
    applied = 0
    skipped = 0
    for cal_year in sorted({fiscal_year_label, fiscal_year_label + 1}):
        result = apply_national_holidays_to_company_calendar(
            store,
            cal_year,
            overwrite=overwrite,
            include_weekends=False,
            force_online=force_online,
            date_from=start,
            date_to=end,
            bump_revision=False,
        )
        applied += int(result.get("applied") or 0)
        skipped += int(result.get("skipped") or 0)
    if include_weekends:
        days = store.setdefault("company_calendar", {}).setdefault("days", {})
        d = start
        while d <= end:
            if d.weekday() >= 5:
                key = d.isoformat()
                existing = days.get(key)
                if existing and (
                    existing.get("manual_edit")
                    or existing.get("kind") == DAY_KIND_PUBLIC
                ):
                    skipped += 1
                else:
                    days[key] = {
                        "kind": DAY_KIND_PUBLIC,
                        "label": "土曜" if d.weekday() == 5 else "日曜",
                        "source": "weekend_init",
                    }
                    applied += 1
            d += timedelta(days=1)
    store.setdefault("company_calendar", {})["year"] = fiscal_year_label
    meta = store.setdefault("meta", {})
    meta["fiscal_start_month"] = start_month
    meta["fiscal_start_day"] = start_day
    meta["holidays_fetched_at"] = datetime.now().isoformat(timespec="seconds")
    if applied > 0:
        meta["company_calendar_revision"] = int(meta.get("company_calendar_revision") or 0) + 1
    return {"applied": applied, "skipped": skipped, "source": "fiscal_year"}


def _member_cell_remark_for_export(entry: dict) -> str:
    remark = str(entry.get("remark") or "").strip()
    comment = str(entry.get("comment") or "").strip()
    if remark and comment:
        return remark + " / " + comment
    return remark or comment


def member_attendance_to_dataframe_records(store: dict, members: list[str]) -> list[dict]:
    """Convert store to flat records for load_attendance_and_analyze compatibility."""
    records: list[dict] = []
    ma = store.get("member_attendance", {})
    for d_key, per_member in sorted(ma.items()):
        for member, entry in per_member.items():
            if member not in members:
                continue
            rec = {
                "日付": d_key,
                "メンバー": member,
                "出勤時間": entry.get("clock_in") or "",
                "退勤時間": entry.get("clock_out") or "",
                "休憩時間1_開始": "",
                "休憩時間1_終了": "",
                "休憩時間2_開始": "",
                "休憩時間2_終了": "",
                "作業効率": entry.get("efficiency", 1.0),
                "休暇区分": entry.get("leave_type") or "",
                "備考": _member_cell_remark_for_export(entry),
                "残業(分)": entry.get("overtime_minutes") or 0,
            }
            breaks = entry.get("breaks") or []
            if len(breaks) >= 1:
                rec["休憩時間1_開始"] = breaks[0].get("start") or ""
                rec["休憩時間1_終了"] = breaks[0].get("end") or ""
            if len(breaks) >= 2:
                rec["休憩時間2_開始"] = breaks[1].get("start") or ""
                rec["休憩時間2_終了"] = breaks[1].get("end") or ""
            records.append(rec)
    return records


def export_attendance_to_calendar_workbook(
    store: dict,
    calendar_path: str | Path | None = None,
    months: list[tuple[int, int]] | None = None,
) -> dict:
    """勤怠カレンダー.xlsx へ APP_* シートを出力（master.xlsm は触らない）。"""
    from openpyxl import Workbook, load_workbook

    from planning_core.core.attendance_calendar_xlsx_history_store import (
        append_calendar_xlsx_snapshot,
    )
    from planning_core.core.attendance_paths import attendance_calendar_xlsx_path

    path = Path(calendar_path or attendance_calendar_xlsx_path())
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.is_file():
        append_calendar_xlsx_snapshot(
            path,
            kind="export_calendar",
            label="勤怠カレンダー出力",
            store_meta=store.get("meta") or {},
        )
        wb = load_workbook(path)
    else:
        wb = Workbook()

    replaced_sheets: set[str] = {APP_MASTER_COMPANY_SHEET}
    _write_app_company_sheet(wb, store)
    if months is None:
        months = _fiscal_months_for_export(store)
    updated: list[str] = []
    for year, month in months:
        sheet_name = f"{APP_MASTER_MEMBER_SHEET_PREFIX}{year}年{month}月"
        replaced_sheets.add(sheet_name)
        _write_app_member_calendar_sheet(wb, store, year, month, sheet_name)
        updated.append(sheet_name)

    _drop_default_empty_sheet_if_present(wb)

    try:
        wb.save(path)
    except OSError as e:
        raise RuntimeError(
            f"勤怠カレンダー.xlsx への保存に失敗（Excel で開いている可能性）: {e}"
        ) from e

    export_at = datetime.now().isoformat(timespec="seconds")
    store["meta"]["calendar_xlsx_export_at"] = export_at
    store["meta"]["calendar_xlsx_path"] = str(path.resolve())
    store["meta"]["calendar_xlsx_export_sheets"] = [APP_MASTER_COMPANY_SHEET] + updated
    return {
        "ok": True,
        "calendar_xlsx_path": str(path.resolve()),
        "sheets_updated": store["meta"]["calendar_xlsx_export_sheets"],
        "calendar_xlsx_export_at": export_at,
    }


def _drop_default_empty_sheet_if_present(wb) -> None:
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])


def export_attendance_to_master_new_sheets(
    store: dict,
    master_path: str | Path,
    months: list[tuple[int, int]] | None = None,
) -> dict:
    """互換: 旧 API。勤怠カレンダー.xlsx へ出力する（master は更新しない）。"""
    logger.warning(
        "export_attendance_to_master_new_sheets は非推奨です。"
        " export_attendance_to_calendar_workbook を使用してください。"
    )
    return export_attendance_to_calendar_workbook(store, months=months)


def _months_in_store(store: dict) -> list[tuple[int, int]]:
    months: set[tuple[int, int]] = set()
    for d_key in store.get("member_attendance", {}):
        try:
            d = date.fromisoformat(d_key)
            months.add((d.year, d.month))
        except ValueError:
            continue
    cc_days = store.get("company_calendar", {}).get("days", {})
    for d_key in cc_days:
        try:
            d = date.fromisoformat(d_key)
            months.add((d.year, d.month))
        except ValueError:
            continue
    if not months:
        today = date.today()
        months.add((today.year, today.month))
    return sorted(months)


_KIND_EXPORT_JA = {
    DAY_KIND_WORKING: "出勤",
    DAY_KIND_PUBLIC: "公休",
    DAY_KIND_SPECIAL: "特別休暇",
}


def _default_label_for_kind(kind: str) -> str:
    if kind == DAY_KIND_PUBLIC:
        return "公休"
    if kind == DAY_KIND_SPECIAL:
        return "特別休暇"
    return ""


def _fiscal_year_label_and_bounds(store: dict) -> tuple[int, int, int, date, date]:
    meta = store.get("meta", {})
    fy_label = int(store.get("company_calendar", {}).get("year") or date.today().year)
    start_month = int(meta.get("fiscal_start_month") or 4)
    start_day = int(meta.get("fiscal_start_day") or 1)
    start, end = fiscal_year_date_range(fy_label, start_month, start_day)
    return fy_label, start_month, start_day, start, end


def _fiscal_months_for_export(store: dict) -> list[tuple[int, int]]:
    _, _, _, start, end = _fiscal_year_label_and_bounds(store)
    months: list[tuple[int, int]] = []
    cur = date(start.year, start.month, 1)
    while cur <= end:
        months.append((cur.year, cur.month))
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)
    return months


def build_dense_company_calendar_days_for_export(store: dict) -> dict[str, dict]:
    """会計年度の全日を UI と同様に解釈して dense 化（master 出力・監査用）。"""
    _, _, _, start, end = _fiscal_year_label_and_bounds(store)
    stored = store.get("company_calendar", {}).get("days", {})
    sparse: dict[str, Any] = {}
    for k, v in stored.items():
        try:
            d_val = date.fromisoformat(k)
        except ValueError:
            continue
        if start <= d_val <= end:
            sparse[k] = v
    enriched = enrich_company_calendar_days_with_national_holidays(
        sparse, start, end, force_online=False
    )
    dense: dict[str, dict] = {}
    d = start
    while d <= end:
        key = d.isoformat()
        if key in enriched:
            entry = dict(enriched[key])
        else:
            kind = DAY_KIND_PUBLIC if d.weekday() >= 5 else DAY_KIND_WORKING
            entry = {"kind": kind, "label": _default_label_for_kind(kind)}
        dense[key] = entry
        d += timedelta(days=1)
    return dense


def _remove_app_master_export_sheet_if_present(wb, sheet_name: str) -> None:
    """APP_* 出力専用シートのみ削除。レガシー「カレンダー」シートは触らない。"""
    name = (sheet_name or "").strip()
    if not is_app_master_export_sheet(name):
        raise ValueError(f"APP 出力対象外のシートは削除しません: {name}")
    if name in wb.sheetnames:
        del wb[name]


def _write_app_company_sheet(wb, store: dict) -> None:
    from planning_core.core.attendance_excel_style import write_company_calendar_grid

    name = APP_MASTER_COMPANY_SHEET
    _remove_app_master_export_sheet_if_present(wb, name)
    ws = wb.create_sheet(name)
    fy_label, _, _, start, end = _fiscal_year_label_and_bounds(store)
    days = build_dense_company_calendar_days_for_export(store)
    months = _fiscal_months_for_export(store)
    write_company_calendar_grid(ws, days, fy_label, start, end, months)


def _write_app_member_calendar_sheet(
    wb, store: dict, year: int, month: int, sheet_name: str
) -> None:
    _remove_app_master_export_sheet_if_present(wb, sheet_name)
    ws = wb.create_sheet(sheet_name)
    populate_member_calendar_worksheet(ws, store, year, month)


def populate_member_calendar_worksheet(ws, store: dict, year: int, month: int) -> None:
    """メンバー×日付のカレンダーグリッドをワークシートへ書き込む（master / 閲覧用）。"""
    from openpyxl.styles import Alignment

    from planning_core.core.attendance_excel_style import (
        _GRID_BORDER,
        format_member_calendar_sheet,
        member_symbol_style,
    )
    from planning_core.core.columns import _result_font

    dense_days = build_dense_company_calendar_days_for_export(store)
    ym_days = calendar.monthrange(year, month)[1]
    date_cols: list[date] = [date(year, month, day_num) for day_num in range(1, ym_days + 1)]

    format_member_calendar_sheet(ws, year, month, len(date_cols), dense_days)

    from planning_core.core.attendance_member_roster import attendance_grid_member_names

    members = attendance_grid_member_names(store)
    if not members:
        try:
            from planning_core.core.master_data import load_skills_and_needs

            members = load_skills_and_needs()[1]
        except Exception:
            members = sorted(
                {
                    m
                    for d_key in store.get("member_attendance", {})
                    for m in store["member_attendance"].get(d_key, {})
                }
            )
    row = 3
    ma = store.get("member_attendance", {})
    align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")
    for member in members:
        name_cell = ws.cell(row, 1, member)
        name_cell.font = _result_font(size=10)
        name_cell.border = _GRID_BORDER
        name_cell.alignment = name_align
        for idx, d in enumerate(date_cols):
            d_key = d.isoformat()
            entry = ma.get(d_key, {}).get(member)
            dense_entry = dense_days.get(d_key, {})
            kind = str(dense_entry.get("kind") or company_day_kind(store, d))
            if not entry:
                preset = day_preset_from_company(store, d)
                entry = preset_to_leave_and_times(preset, d, company_kind=kind)
            symbol = _cell_symbol(entry)
            col = 2 + idx
            cell = ws.cell(row, col, symbol)
            fill, font = member_symbol_style(symbol, kind)
            if fill is not None:
                cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = _GRID_BORDER
        row += 1


def _cell_symbol(entry: dict) -> str:
    """メンバー勤怠グリッド（JavaFX shortLabel）と同一の1文字／短記号。"""
    preset = str(entry.get("day_preset") or "").strip()
    lt = str(entry.get("leave_type") or "").strip()
    remark = str(entry.get("remark") or "").strip()
    if preset == PRESET_PAID_LEAVE or (not preset and remark == "年休" and lt == "年休"):
        return "年休"
    if preset == PRESET_ABSENT or lt == "欠勤" or (not preset and remark == "欠勤"):
        return "欠"
    if (
        preset == PRESET_OFF_FULL
        or lt in ("公休", "休")
        or (not preset and remark == "休" and lt == "年休")
    ):
        return "休"
    if preset == PRESET_OFF_AM or lt == "前休":
        return "前"
    if preset == PRESET_OFF_PM or lt == "後休":
        return "後"
    if preset == PRESET_NO_DISPATCH or lt == "-":
        return "-"
    if preset == PRESET_HOLIDAY_WORK or lt == "休日出勤":
        return "休出"
    if preset == PRESET_HOLIDAY_WORK_AM or lt == "午前休出":
        return "前出"
    if preset == PRESET_HOLIDAY_WORK_PM or lt == "午後休出":
        return "後出"
    hourly = entry.get("hourly") or {}
    if hourly:
        return "時"
    return "·"


def migrate_master_attendance_to_store(store: dict, master_path: str) -> dict:
    """One-time import from legacy member sheets into store."""
    from planning_core.core.master_data import load_attendance_and_analyze, load_skills_and_needs

    members = load_skills_and_needs()[1]
    attendance_data, _ = load_attendance_and_analyze(members)
    applied = 0
    ma = store.setdefault("member_attendance", {})
    for d, per_member in attendance_data.items():
        d_key = d.isoformat()
        bucket = ma.setdefault(d_key, {})
        for member, st in per_member.items():
            bucket[member] = {
                "day_preset": PRESET_WORK if st.get("is_working") else PRESET_OFF_FULL,
                "clock_in": st.get("start_dt").strftime("%H:%M") if st.get("start_dt") else None,
                "clock_out": st.get("end_dt").strftime("%H:%M") if st.get("end_dt") else None,
                "leave_type": st.get("reason") or "",
                "remark": st.get("reason") or "",
                "overtime_minutes": int(st.get("overtime_minutes") or 0),
                "efficiency": float(st.get("efficiency") or 1.0),
                "breaks": [],
                "hourly": {},
                "manual_edit": False,
            }
            applied += 1
    store["meta"]["member_attendance_revision"] = int(
        store["meta"].get("member_attendance_revision") or 0
    ) + 1
    return {"imported_cells": applied}
