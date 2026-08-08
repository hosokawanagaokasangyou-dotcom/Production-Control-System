# -*- coding: utf-8 -*-
# planning_core.core.roll_pipeline — body only (loaded via _core exec chain)
ROLL_PIPELINE_EC_PROCESS = "EC"
ROLL_PIPELINE_EC_MACHINE = "EC機　湖南"
ROLL_PIPELINE_INSP_PROCESS = "検査"
ROLL_PIPELINE_INSP_MACHINE = "熱融着機　湖南"
ROLL_PIPELINE_REWIND_PROCESS = "巻返し"
# §B-3 後続は EC 機（EC機　湖南）と異なる設備上的巻返しのみ。
# 同一 EC 機上的巻返しは実巻返しを行わず段階1で配台不要とするため B-3 対象外。
ROLL_PIPELINE_REWIND_MACHINE = ROLL_PIPELINE_EC_MACHINE
ROLL_PIPELINE_INITIAL_BUFFER_ROLLS = 2
ROLL_PIPELINE_INSP_UNCAPPED_ROOM = 1.0e18
WIP_LIMIT_EC_BEFORE_INSP_ROLLS = os.environ.get(
    "WIP_LIMIT_EC_BEFORE_INSP_ROLLS", "15"
).strip()
try:
    WIP_LIMIT_EC_BEFORE_INSP_ROLLS = int(WIP_LIMIT_EC_BEFORE_INSP_ROLLS)
except (TypeError, ValueError):
    WIP_LIMIT_EC_BEFORE_INSP_ROLLS = 15
_wip_l11_agg_raw = os.environ.get(
    "WIP_LIMIT_EC_BEFORE_INSP_AGGREGATE", "task_id"
).strip().lower()
if _wip_l11_agg_raw in ("global", "all", "factory"):
    WIP_LIMIT_EC_BEFORE_INSP_AGGREGATE_MODE = "global"
elif _wip_l11_agg_raw in ("task_id", "line", "row"):
    WIP_LIMIT_EC_BEFORE_INSP_AGGREGATE_MODE = "task_id"
elif _wip_l11_agg_raw in ("task_id_head", "head", "prefix"):
    WIP_LIMIT_EC_BEFORE_INSP_AGGREGATE_MODE = "task_id_head"
else:
    WIP_LIMIT_EC_BEFORE_INSP_AGGREGATE_MODE = "task_id"
def _wip_ec_l11_aggregate_is_global() -> bool:
    return WIP_LIMIT_EC_BEFORE_INSP_AGGREGATE_MODE == "global"
WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS = os.environ.get(
    "WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS", "20"
).strip()
try:
    WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS = int(WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS)
except (TypeError, ValueError):
    WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS = 20
SLIT_BEFORE_SEC_MIN_SLIT_ROLLS = os.environ.get(
    "SLIT_BEFORE_SEC_MIN_SLIT_ROLLS", "5"
).strip()
try:
    SLIT_BEFORE_SEC_MIN_SLIT_ROLLS = int(SLIT_BEFORE_SEC_MIN_SLIT_ROLLS)
except (TypeError, ValueError):
    SLIT_BEFORE_SEC_MIN_SLIT_ROLLS = 5
SPECIAL_WIP_SLIT_PROCESS = "スリット"
SPECIAL_WIP_SLIT_MACHINE = "スリット機1　湖南"
SPECIAL_WIP_SEC_PROCESS = "SEC"
SPECIAL_WIP_SEC_MACHINE = "SEC機　湖南"
WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS = os.environ.get(
    "WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS", "20"
).strip()
try:
    WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS = int(WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS)
except (TypeError, ValueError):
    WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS = 20
CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS = os.environ.get(
    "CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS", "5"
).strip()
try:
    CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS = int(CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS)
except (TypeError, ValueError):
    CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS = 5
SPECIAL_WIP_CONNECTION_PROCESS = "接続"
SPECIAL_WIP_CONNECTION_MACHINE = "熱融着機　湖南"
def _task_on_slit_sec_process_path(task: dict) -> bool:
    """加工内容トークンがスリット→SEC の依頼NO経路か（L10 総量制約の対象）。"""
    toks = task.get("process_content_tokens") or []
    norm = [_normalize_process_name_for_rule_match(x) for x in toks]
    slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
    sec_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    if slit_proc not in norm or sec_proc not in norm:
        return False
    try:
        return norm.index(slit_proc) < norm.index(sec_proc)
    except ValueError:
        return False
def _l10_slit_done_minus_sec_done_for_task_id(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> float:
    tid = (task_id or "").strip()
    if not tid:
        return 0.0
    _slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
    _slit_mach = _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
    _sec_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    _sec_mach = _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE)
    slit_done = 0.0
    sec_done = 0.0
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if not proc or not mach:
            continue
        init = float(_t.get("initial_remaining_units") or 0)
        rem = float(_t.get("remaining_units") or 0)
        done = max(0.0, init - rem)
        if done <= 1e-12:
            continue
        if proc == _slit_proc and mach == _slit_mach:
            slit_done += done
        elif proc == _sec_proc and mach == _sec_mach:
            sec_done += done
    return max(0.0, slit_done - sec_done)
def _l10_task_queue_has_special_slit_row_for_tid(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """同一依頼に L10 対象（スリット機1　湖南）の行が task_queue に存在するか。

    スリット完走後に行がキューから落ちると slit_done が集計されず pair_gap=0 のままになる。
    そのとき B-4.1 で SEC を永久除外しないため、スリット行が無い場合はゲートを掛けない。
    """
    tid = (task_id or "").strip()
    if not tid:
        return False
    _slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
    _slit_mach = _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if proc == _slit_proc and mach == _slit_mach:
            return True
    return False
def _b6_connection_done_minus_sec_done_for_task_id(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> float:
    tid = (task_id or "").strip()
    if not tid:
        return 0.0
    _conn_proc = _normalize_process_name_for_rule_match(
        SPECIAL_WIP_CONNECTION_PROCESS
    )
    _conn_mach = _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
    _sec_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    _sec_mach = _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE)
    conn_done = 0.0
    sec_done = 0.0
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if not proc or not mach:
            continue
        init = float(_t.get("initial_remaining_units") or 0)
        rem = float(_t.get("remaining_units") or 0)
        done = max(0.0, init - rem)
        if done <= 1e-12:
            continue
        if proc == _conn_proc and mach == _conn_mach:
            conn_done += done
        elif proc == _sec_proc and mach == _sec_mach:
            sec_done += done
    return max(0.0, conn_done - sec_done)
def _b6_task_queue_has_special_connection_row_for_tid(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """同一依頼に B-6 対象（接続×熱融着機　湖南）の行が task_queue に存在するか。"""
    tid = (task_id or "").strip()
    if not tid:
        return False
    _conn_proc = _normalize_process_name_for_rule_match(
        SPECIAL_WIP_CONNECTION_PROCESS
    )
    _conn_mach = _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if proc == _conn_proc and mach == _conn_mach:
            return True
    return False
def _b6_initial_connection_roll_capacity_for_tid(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> float:
    """同一依頼の接続（熱融着機　湖南）行の initial_remaining_units 合計（ロール）。"""
    tid = (task_id or "").strip()
    if not tid:
        return 0.0
    _conn_proc = _normalize_process_name_for_rule_match(
        SPECIAL_WIP_CONNECTION_PROCESS
    )
    _conn_mach = _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
    s = 0.0
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if proc == _conn_proc and mach == _conn_mach:
            s += float(_t.get("initial_remaining_units") or 0)
    return s
def _l10_initial_slit_roll_capacity_for_tid(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> float:
    """同一依頼のスリット（スリット機1　湖南）行の initial_remaining_units 合計（ロール）。"""
    tid = (task_id or "").strip()
    if not tid:
        return 0.0
    _slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
    _slit_mach = _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
    s = 0.0
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if proc == _slit_proc and mach == _slit_mach:
            s += float(_t.get("initial_remaining_units") or 0)
    return s
def _l10_b41_threshold_unreachable(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """当該依頼のスリット総ロールが B-4.1 閾値未満なら、閾値到達不能（ゲートを掛けない）。"""
    thr = float(SLIT_BEFORE_SEC_MIN_SLIT_ROLLS)
    if thr <= 1e-9:
        return False
    cap = _l10_initial_slit_roll_capacity_for_tid(
        task_queue, task_id, rows_by_tid=rows_by_tid
    )
    return cap + 1e-9 < thr
def _b61_threshold_unreachable(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """当該依頼の接続総ロールが B-6.1 閾値未満なら、閾値到達不能（ゲートを掛けない）。"""
    thr = float(CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS)
    if thr <= 1e-9:
        return False
    cap = _b6_initial_connection_roll_capacity_for_tid(
        task_queue, task_id, rows_by_tid=rows_by_tid
    )
    return cap + 1e-9 < thr
def _b6_connection_has_remaining_units(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """当該依頼の接続（熱融着機　湖南）行に未割当ロールが残るか。"""
    tid = (task_id or "").strip()
    if not tid:
        return False
    _conn_proc = _normalize_process_name_for_rule_match(
        SPECIAL_WIP_CONNECTION_PROCESS
    )
    _conn_mach = _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if proc == _conn_proc and mach == _conn_mach:
            return float(_t.get("remaining_units") or 0) > 1e-12
    return False
def _l10_slit_has_remaining_units(
    task_queue: list,
    task_id: str,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """当該依頼のスリット（スリット機1　湖南）行に未割当ロールが残るか。"""
    tid = (task_id or "").strip()
    if not tid:
        return False
    _slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
    _slit_mach = _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
    _iter_rows = (
        rows_by_tid.get(tid) if rows_by_tid is not None else task_queue
    )
    for _t in _iter_rows:
        if rows_by_tid is None and str(_t.get("task_id") or "").strip() != tid:
            continue
        proc = _normalize_process_name_for_rule_match(_t.get("machine"))
        mach = _normalize_equipment_match_key(_t.get("machine_name"))
        if proc == _slit_proc and mach == _slit_mach:
            return float(_t.get("remaining_units") or 0) > 1e-12
    return False
def _b61_sec_blocked_by_connection_min_rolls(
    task: dict,
    task_queue: list,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """
    B-6.1: 接続→SEC の SEC を候補から外すか。
    接続行に残ロールが無い（接続完走後）はゲートしない（SEC 残ロールを完走できる）。
    """
    tid = str(task.get("task_id") or "").strip()
    if not tid or _b61_threshold_unreachable(
        task_queue, tid, rows_by_tid=rows_by_tid
    ):
        return False
    if not _b6_connection_has_remaining_units(
        task_queue, tid, rows_by_tid=rows_by_tid
    ):
        return False
    if (
        _b6_connection_done_minus_sec_done_for_task_id(
            task_queue, tid, rows_by_tid=rows_by_tid
        )
        >= float(CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS) - 1e-9
    ):
        return False
    proc = _normalize_process_name_for_rule_match(task.get("machine"))
    mach = _normalize_equipment_match_key(task.get("machine_name"))
    if proc != _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS):
        return False
    if mach != _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE):
        return False
    toks = task.get("process_content_tokens") or []
    _norm = [_normalize_process_name_for_rule_match(x) for x in toks]
    _cp = _normalize_process_name_for_rule_match(SPECIAL_WIP_CONNECTION_PROCESS)
    _sc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    if not (
        _cp in _norm
        and _sc in _norm
        and _norm.index(_cp) < _norm.index(_sc)
    ):
        return False
    if not _b6_task_queue_has_special_connection_row_for_tid(
        task_queue, tid, rows_by_tid=rows_by_tid
    ):
        return False
    return True
def _l10_b41_sec_blocked_by_slit_min_rolls(
    task: dict,
    task_queue: list,
    *,
    rows_by_tid: dict[str, list[dict]] | None = None,
) -> bool:
    """B-4.1: スリット→SEC の SEC を候補から外すか（スリット完走後はゲートしない）。"""
    tid = str(task.get("task_id") or "").strip()
    if not tid or _l10_b41_threshold_unreachable(
        task_queue, tid, rows_by_tid=rows_by_tid
    ):
        return False
    if not _l10_slit_has_remaining_units(
        task_queue, tid, rows_by_tid=rows_by_tid
    ):
        return False
    if (
        _l10_slit_done_minus_sec_done_for_task_id(
            task_queue, tid, rows_by_tid=rows_by_tid
        )
        >= float(SLIT_BEFORE_SEC_MIN_SLIT_ROLLS) - 1e-9
    ):
        return False
    proc = _normalize_process_name_for_rule_match(task.get("machine"))
    mach = _normalize_equipment_match_key(task.get("machine_name"))
    if proc != _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS):
        return False
    if mach != _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE):
        return False
    toks = task.get("process_content_tokens") or []
    _norm = [_normalize_process_name_for_rule_match(x) for x in toks]
    if not (
        _normalize_process_name_for_rule_match("スリット") in _norm
        and _normalize_process_name_for_rule_match("SEC") in _norm
        and _norm.index(_normalize_process_name_for_rule_match("スリット"))
        < _norm.index(_normalize_process_name_for_rule_match("SEC"))
    ):
        return False
    return _l10_task_queue_has_special_slit_row_for_tid(
        task_queue, tid, rows_by_tid=rows_by_tid
    )
def _l10_sec_start_floor_from_slit_timeline(
    task: dict,
    timeline_events: list | None,
    task_queue: list,
) -> datetime | None:
    """
    L10 B-4.1 のカレンダー側下限: スリット→SEC の SEC 行は、タイムライン上で
    SLIT_BEFORE_SEC_MIN_SLIT_ROLLS 本目のスリット加工終了後まで加工開始しない。
    （候補除外はロール差のみ。別設備のため machine_avail が同日 13:00 のままだと
    SEC がスリットと壁時計上重なるのを防ぐ。）
    """
    if SLIT_BEFORE_SEC_MIN_SLIT_ROLLS <= 0:
        return None
    if not timeline_events:
        return None
    machine_proc = str(task.get("machine") or "").strip()
    machine_name = str(task.get("machine_name", "") or "").strip()
    proc = _normalize_process_name_for_rule_match(machine_proc)
    mach = _normalize_equipment_match_key(machine_name)
    if proc != _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS):
        return None
    if mach != _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE):
        return None
    toks = task.get("process_content_tokens") or []
    _norm = [_normalize_process_name_for_rule_match(x) for x in toks]
    _sp = _normalize_process_name_for_rule_match("スリット")
    _sc = _normalize_process_name_for_rule_match("SEC")
    if not (_sp in _norm and _sc in _norm and _norm.index(_sp) < _norm.index(_sc)):
        return None
    tid = str(task.get("task_id") or "").strip()
    if not tid:
        return None
    if _l10_b41_threshold_unreachable(task_queue, tid):
        return None
    if not _l10_slit_has_remaining_units(task_queue, tid):
        return None
    slit_row: dict | None = None
    slit_eq: str = ""
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        p = _normalize_process_name_for_rule_match(t.get("machine"))
        m = _normalize_equipment_match_key(t.get("machine_name"))
        if (
            p == _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
            and m == _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
        ):
            slit_row = t
            slit_eq = (
                str(t.get("equipment_line_key") or t.get("machine") or "")
                .strip()
                or str(t.get("machine") or "").strip()
            )
            break
    if not slit_row or not slit_eq:
        return None
    occ_slit = str(_machine_occupancy_key_resolve(slit_row, slit_eq) or "").strip()
    n_need = int(SLIT_BEFORE_SEC_MIN_SLIT_ROLLS)
    # タイムラインの machine は eq_line（例: スリット+スリット機1　湖南）。厳密一致で落ちる場合があるため
    # machine / machine_occupancy_key を正規化してスリット行と照合する。
    k_eq = _normalize_equipment_match_key(slit_eq)
    k_m = _normalize_equipment_match_key(str(slit_row.get("machine") or ""))
    k_occ = _normalize_equipment_match_key(occ_slit)
    hits: list[dict] = []
    for e in timeline_events:
        if not _is_machining_timeline_event(e):
            continue
        if str(e.get("task_id") or "").strip() != tid:
            continue
        em = _normalize_equipment_match_key(str(e.get("machine") or ""))
        eocc = _normalize_equipment_match_key(str(e.get("machine_occupancy_key") or ""))
        if em not in (k_eq, k_m) and eocc != k_occ:
            continue
        sd = e.get("start_dt")
        ed = e.get("end_dt")
        if not isinstance(sd, datetime) or not isinstance(ed, datetime):
            continue
        hits.append(e)
    hits.sort(key=lambda x: x["start_dt"])
    if len(hits) >= n_need:
        nth = hits[n_need - 1].get("end_dt")
        return nth if isinstance(nth, datetime) else None
    return None
def _b6_sec_start_floor_from_connection_timeline(
    task: dict,
    timeline_events: list | None,
    task_queue: list,
) -> datetime | None:
    """
    B-6.1 のカレンダー側下限: 接続→SEC の SEC 行は、タイムライン上で
    CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS 本目の接続加工終了後まで加工開始しない。
    """
    if CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS <= 0:
        return None
    if not timeline_events:
        return None
    machine_proc = str(task.get("machine") or "").strip()
    machine_name = str(task.get("machine_name", "") or "").strip()
    proc = _normalize_process_name_for_rule_match(machine_proc)
    mach = _normalize_equipment_match_key(machine_name)
    if proc != _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS):
        return None
    if mach != _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE):
        return None
    toks = task.get("process_content_tokens") or []
    _norm = [_normalize_process_name_for_rule_match(x) for x in toks]
    _cp = _normalize_process_name_for_rule_match(SPECIAL_WIP_CONNECTION_PROCESS)
    _sc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    if not (_cp in _norm and _sc in _norm and _norm.index(_cp) < _norm.index(_sc)):
        return None
    tid = str(task.get("task_id") or "").strip()
    if not tid:
        return None
    if _b61_threshold_unreachable(task_queue, tid):
        return None
    if not _b6_connection_has_remaining_units(task_queue, tid):
        return None
    conn_row: dict | None = None
    conn_eq: str = ""
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        p = _normalize_process_name_for_rule_match(t.get("machine"))
        m = _normalize_equipment_match_key(t.get("machine_name"))
        if (
            p == _normalize_process_name_for_rule_match(SPECIAL_WIP_CONNECTION_PROCESS)
            and m == _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
        ):
            conn_row = t
            conn_eq = (
                str(t.get("equipment_line_key") or t.get("machine") or "")
                .strip()
                or str(t.get("machine") or "").strip()
            )
            break
    if not conn_row or not conn_eq:
        return None
    occ_conn = str(_machine_occupancy_key_resolve(conn_row, conn_eq) or "").strip()
    n_need = int(CONNECTION_BEFORE_SEC_MIN_CONNECTION_ROLLS)
    k_eq = _normalize_equipment_match_key(conn_eq)
    k_m = _normalize_equipment_match_key(str(conn_row.get("machine") or ""))
    k_occ = _normalize_equipment_match_key(occ_conn)
    hits: list[dict] = []
    for e in timeline_events:
        if not _is_machining_timeline_event(e):
            continue
        if str(e.get("task_id") or "").strip() != tid:
            continue
        em = _normalize_equipment_match_key(str(e.get("machine") or ""))
        eocc = _normalize_equipment_match_key(str(e.get("machine_occupancy_key") or ""))
        if em not in (k_eq, k_m) and eocc != k_occ:
            continue
        sd = e.get("start_dt")
        ed = e.get("end_dt")
        if not isinstance(sd, datetime) or not isinstance(ed, datetime):
            continue
        hits.append(e)
    hits.sort(key=lambda x: x["start_dt"])
    if len(hits) >= n_need:
        nth = hits[n_need - 1].get("end_dt")
        return nth if isinstance(nth, datetime) else None
    return None
STAGE2_MACRO_NOW_USE_DATA_EXTRACT_CLOCK = (
    os.environ.get("STAGE2_MACRO_NOW_USE_DATA_EXTRACT_CLOCK", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)
def _stage2_truthy_env(name: str) -> bool:
    """JavaFX 実行タブ／環境変数から渡す段階2オプション用（0/false/no/off/none 以外を有効とする）。"""
    v = (os.environ.get(name) or "").strip().lower()
    return v in ("1", "true", "yes", "on", "はい")
STAGE2_EXTEND_ATTENDANCE_CALENDAR = False
SCHEDULE_EXTEND_MAX_EXTRA_DAYS = 366
STAGE2_RETRY_SHIFT_DUE_ON_PARTIAL_REMAINING = False
STAGE2_RETRY_SHIFT_DUE_MAX_ROUNDS = 5
STAGE2_SERIAL_DISPATCH_BY_TASK_ID = (
    os.environ.get("STAGE2_SERIAL_DISPATCH_BY_TASK_ID", "0")
    .strip()
    .lower()
    in ("1", "true", "yes", "on", "はい")
)
STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST = os.environ.get(
    "STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")
STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT = os.environ.get(
    "STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT", "1"
).strip().lower() not in ("0", "false", "no", "off", "いいえ", "無効")
def _clone_attendance_day_shifted(source_day: dict, old_date: date, new_date: date) -> dict:
    """メンバー別勤怠ブロックを new_date にシフトした浅いコピーを返す。"""
    delta_days = (new_date - old_date).days
    if delta_days == 0:
        return {m: dict(st) for m, st in source_day.items()}
    delta = timedelta(days=delta_days)
    out: dict = {}
    for m, st in source_day.items():
        new_st = dict(st)
        sd = st.get("start_dt")
        ed = st.get("end_dt")
        new_st["start_dt"] = sd + delta if sd else None
        new_st["end_dt"] = ed + delta if ed else None
        bed = st.get("base_end_dt")
        new_st["base_end_dt"] = bed + delta if bed else None
        nb = []
        for pair in st.get("breaks_dt") or []:
            if len(pair) >= 2:
                a, b = pair[0], pair[1]
                if a is not None and b is not None:
                    nb.append((a + delta, b + delta))
        new_st["breaks_dt"] = merge_time_intervals(nb)
        out[m] = new_st
    return out
def _pick_extension_template_date(attendance_data: dict, plan_dates: list):
    """配台可能なメンバーは1人でもいる直近の日をテンプレに採用（最終日は全休でも有効な型を使う）。"""
    for i in range(len(plan_dates) - 1, -1, -1):
        d = plan_dates[i]
        day = attendance_data.get(d)
        if not day:
            continue
        if any(
            v.get("eligible_for_assignment", v.get("is_working", False))
            for v in day.values()
        ):
            return d
    return plan_dates[-1] if plan_dates else None
def _extend_attendance_one_calendar_day(
    attendance_data: dict,
    plan_dates: list,
) -> bool:
    """カレンダー上1日先を plan_dates に追加し、テンプレ日のシフト複製で attendance を埋ゝる。失敗時 False。"""
    if not plan_dates:
        return False
    last_d = plan_dates[-1]
    next_d = last_d + timedelta(days=1)
    tmpl_d = _pick_extension_template_date(attendance_data, plan_dates)
    if tmpl_d is None:
        return False
    template = attendance_data.get(tmpl_d)
    if not template:
        return False
    attendance_data[next_d] = _clone_attendance_day_shifted(template, tmpl_d, next_d)
    plan_dates.append(next_d)
    logging.info(
        "配台完了まで勤怠を自動拡張: %s を追加（テンプレ=%s」メンバー数=%s）",
        next_d,
        tmpl_d,
        len(attendance_data[next_d]),
    )
    return True
def _iter_plan_dates_extending(
    plan_dates: list,
    attendance_data: dict,
    task_queue: list,
):
    """
    plan_dates を先頭から順に yield。末尾まで来でも残タスクはあれみ勤怠を1日うつ拡張して継続。
    plan_dates / attendance_data はインプレース更新される。
    """
    si = 0
    ext_used = 0
    while True:
        while si < len(plan_dates):
            yield plan_dates[si]
            si += 1
        pending = any(float(t.get("remaining_units") or 0) > 1e-12 for t in task_queue)
        if not pending:
            return
        if ext_used >= SCHEDULE_EXTEND_MAX_EXTRA_DAYS:
            logging.warning(
                "残タスクはありしたは勤怠の自動拡張は上限（%s 日）に靔しました。配台残・配台試行は残る可能性はありした。",
                SCHEDULE_EXTEND_MAX_EXTRA_DAYS,
            )
            return
        if not _extend_attendance_one_calendar_day(attendance_data, plan_dates):
            logging.warning(
                "勤怠を1日拡張でしませんでした（テンプレ日のデータ欠損）。残タスクは未割当のままです。"
            )
            return
        ext_used += 1
def _parse_process_content_tokens(val) -> list[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    s = unicodedata.normalize("NFKC", str(val).strip())
    if not s or s.lower() in ("nan", "none", "null"):
        return []
    return [p.strip() for p in s.split(",") if p.strip()]
def _collect_process_content_order_by_task_id(tasks_df) -> dict[str, list[str]]:
    """依頼NO → 加工内容の工程名リスト（表の上の方で最初に睾れた非空の行を採用）。"""
    out: dict[str, list[str]] = {}
    if tasks_df is None or tasks_df.empty:
        return out
    for _, row in tasks_df.iterrows():
        tid = planning_task_id_str_from_plan_row(row)
        if not tid:
            continue
        parts = _parse_process_content_tokens(row.get(TASK_COL_PROCESS_CONTENT))
        if not parts:
            continue
        if tid not in out:
            out[tid] = parts
    return out
def _process_name_matches_kakou_content_tokens(
    process_name: str, content_tokens: list[str]
) -> bool:
    """
    工程名（配台計画の「工程名」列）は」元データの「加工内容」カンマ区切りトークンのいうれかと
    正規化一致するか。トークンは無い（加工内容未記入の依頼）は照合対象外として True。
    """
    if not content_tokens:
        return True
    proc = _normalize_process_name_for_rule_match(process_name)
    if not proc:
        return False
    for tok in content_tokens:
        if _normalize_process_name_for_rule_match(tok) == proc:
            return True
    return False
def _process_sequence_rank_for_machine(proc, order_list: list[str]):
    if not order_list:
        return None
    pn = _normalize_process_name_for_rule_match(proc)
    for i, token in enumerate(order_list):
        if _normalize_process_name_for_rule_match(token) == pn:
            return i
    return None
def _task_rank_int_or_none(task) -> int | None:
    r = task.get("process_sequence_rank")
    if r is None:
        return None
    try:
        return int(r)
    except (TypeError, ValueError):
        return None
def _plan_sheet_priority_sort_value(t: dict) -> int:
    """配台計画シートの「優先度」。尝さいろど先。未入力・正常は 999。"""
    p = t.get("priority", 999)
    try:
        return int(p)
    except (TypeError, ValueError):
        return 999
def _task_blocked_by_same_request_dependency(
    task,
    task_queue,
    *,
    rows_by_rule_tid: dict[str, list[dict]] | None = None,
    pipeline_room_cache: dict[str, float] | None = None,
) -> bool:
    """
    同一依頼NOの異なる工程を坌時刻に回さない（配台ルール §A-1・§A-2）。
    - 両行に加工内容由来の rank はあるとしは rank のみで剝後（§A-1）。
    - どうらかに rank は無いとしは」配台計画シートの行順 same_request_line_seq で剝後（§A-2）。
    §B-2 / §B-3: ``roll_pipeline_inspection`` または ``roll_pipeline_rewind`` 行は
    ``roll_pipeline_ec`` 先行により §A-1 で止まる場合」
    ``_roll_pipeline_inspection_assign_room`` > 0 なら当該ペアの値ブロックしない。
    剝進配台では ``_trial_order_flow_eligible_tasks`` は EC 完走まで検査を外れため、
    EC 残はある間は本分岝に到靔しない。リワインド等で検査は載る局面との整合用。
    """
    # §A は同一依頼NO（枝番は元依頼NO=rule_task_id）の異工程依存。枝番が無いとき rule_task_id==task_id。
    tid = _rule_task_id(task)
    if not tid:
        return False
    try:
        my_seq = int(task.get("same_request_line_seq", 0))
    except (TypeError, ValueError):
        my_seq = 0
    my_r = _task_rank_int_or_none(task)

    _peer_rows = (
        rows_by_rule_tid.get(tid) if rows_by_rule_tid is not None else None
    )
    if _peer_rows is None:
        _peer_rows = task_queue
    for t2 in _peer_rows:
        if rows_by_rule_tid is None and _rule_task_id(t2) != tid:
            continue
        if float(t2.get("remaining_units") or 0) <= 1e-9:
            continue
        r2 = _task_rank_int_or_none(t2)
        try:
            s2 = int(t2.get("same_request_line_seq", 0))
        except (TypeError, ValueError):
            s2 = 0

        if my_r is not None and r2 is not None:
            precedes = r2 < my_r
        elif my_r is None and r2 is None:
            precedes = s2 < my_seq
        else:
            precedes = s2 < my_seq

        if precedes:
            if (
                (
                    task.get("roll_pipeline_inspection")
                    or task.get("roll_pipeline_rewind")
                )
                and t2.get("roll_pipeline_ec")
                and (
                    pipeline_room_cache.get(tid)
                    if pipeline_room_cache is not None
                    and tid in pipeline_room_cache
                    else _roll_pipeline_inspection_assign_room(task_queue, tid)
                )
                > 1e-12
            ):
                continue
            return True
    return False
def _task_not_yet_schedulable_due_to_dependency_or_b2_room(
    task: dict, task_queue: list
) -> bool:
    """
    キュー状態上」この行はまて日次配台で進ゝられない（§A 同一依頼の剝工程残」または §B-2/§B-3 の枠ゼロ）。
    `_min_pending_dispatch_trial_order_for_date` と `_equipment_line_lower_dispatch_trial_still_pending`
    で坌も基準を共有れる。片方の値直れと」同一設備キーで全件未割当は残るデッドロックは起し得る。
    """
    if _task_blocked_by_same_request_dependency(task, task_queue):
        return True
    if (task.get("roll_pipeline_inspection") or task.get("roll_pipeline_rewind")) and (
        _roll_pipeline_inspection_assign_room(
            task_queue, str(task.get("task_id", "") or "").strip()
        )
        <= 1e-12
    ):
        return True
    return False
def _row_matches_roll_pipeline_ec(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_EC_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_EC_MACHINE)
    )
def _row_matches_roll_pipeline_inspection(proc, mach) -> bool:
    return (
        _normalize_process_name_for_rule_match(proc)
        == _normalize_process_name_for_rule_match(ROLL_PIPELINE_INSP_PROCESS)
        and _normalize_equipment_match_key(mach)
        == _normalize_equipment_match_key(ROLL_PIPELINE_INSP_MACHINE)
    )
def _row_matches_roll_pipeline_rewind(proc, mach) -> bool:
    """
    §B-3 後続: 巻返しかつ EC 機（EC機　湖南）と **異なる** 設備のみ。
    同一 EC 機上的巻返しは実配台しないため B-3 に含めない。
    """
    if (
        _normalize_process_name_for_rule_match(proc)
        != _normalize_process_name_for_rule_match(ROLL_PIPELINE_REWIND_PROCESS)
    ):
        return False
    mach_key = _normalize_equipment_match_key(mach)
    ec_key = _normalize_equipment_match_key(ROLL_PIPELINE_EC_MACHINE)
    if not mach_key:
        return False
    return mach_key != ec_key
def _pipeline_ec_roll_done_units(task_queue, tid: str) -> float:
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s
def _pipeline_inspection_roll_done_units(task_queue, tid: str) -> float:
    """熱融着検査行のみの累計完了ロール（トレース用）。"""
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_inspection"):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s
def _pipeline_b2_follower_roll_done_units(task_queue, tid: str) -> float:
    """§B-2 検査行＋§B-3 巻返し行の」同一依頼内の後続パイプライン累計完了ロール。"""
    tid = str(tid or "").strip()
    s = 0.0
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not (t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")):
            continue
        init = float(t.get("initial_remaining_units") or 0)
        rem = float(t.get("remaining_units") or 0)
        s += max(0.0, init - rem)
    return s
def _task_queue_has_roll_pipeline_ec_for_tid(task_queue, task_id: str) -> bool:
    """同一依頼NOに EC（ロールパイプライン先行）タスクはキューに含まれるか。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if t.get("roll_pipeline_ec"):
            return True
    return False
def _pipeline_ec_fully_done_for_tid(task_queue, task_id: str) -> bool:
    """同一依頼NOの EC ロールパイプライン行はまとめて残量ゼロ（完走）か。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    found = False
    for t in task_queue:
        if str(t.get("task_id", "") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        found = True
        if float(t.get("remaining_units") or 0) > 1e-9:
            return False
    return found
def _roll_pipeline_inspection_assign_room(task_queue, task_id: str) -> float:
    tid = str(task_id or "").strip()
    # EC 行がキューに無い＝完走後欠落時は枠ゼロにしない（§B-2/§B-3 後続の配台不可防止）。
    if not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
        return float(ROLL_PIPELINE_INSP_UNCAPPED_ROOM)
    ec_done = _pipeline_ec_roll_done_units(task_queue, task_id)
    insp_done = _pipeline_b2_follower_roll_done_units(task_queue, task_id)
    # EC 全ロール完了後は「EC 先行・ポッファ」は既に満たされでいる。ここで max_insp を ec_done に
    # 权ごると」シート上の検査（・巻返し）残ロール数は EC 完了ロール数を上回るデータで
    # max_insp - insp_done は 0 のまま残り」検査行は eligible から外れ配台試行順は永久に詰まる
    # （再睾ログ: ec_fully_done かつ insp_done==max_insp==ec_done で room=0 → 後続試行順は配台試行）。
    if _pipeline_ec_fully_done_for_tid(task_queue, task_id):
        return float(ROLL_PIPELINE_INSP_UNCAPPED_ROOM)
    # EC 稼働中: 先行ポッファ B により検査ロール上限を ec_done から靅延させる（B=2 の弝はコメント参照）。
    max_insp = max(0.0, ec_done - float(ROLL_PIPELINE_INITIAL_BUFFER_ROLLS) + 1.0)
    _room = max(0.0, max_insp - insp_done)
    return _room
def _roll_pipeline_inspection_task_row_for_tid(
    task_queue: list, task_id: str
) -> dict | None:
    """同一依頼NOの §B-2 検査行または §B-3 巻返し行を1件返す。無ければ None。"""
    tid = str(task_id or "").strip()
    if not tid:
        return None
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind"):
            return t
    return None
def _pipeline_b2_ec_roll_end_datetimes_sorted(
    task_queue: list, task_id: str
) -> list[datetime]:
    """同一依頼の EC ロール確定時の終了時刻を時系列で返す（assigned_history の end_dt）。"""
    tid = str(task_id or "").strip()
    ends: list[datetime] = []
    if not tid:
        return ends
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if not t.get("roll_pipeline_ec"):
            continue
        for h in t.get("assigned_history") or []:
            ed = h.get("end_dt")
            if isinstance(ed, datetime):
                ends.append(ed)
    ends.sort()
    return ends
def _roll_pipeline_b2_inspection_ec_completion_floor_dt(
    task_queue: list, task_id: str
) -> datetime | None:
    """
    次の検査ロールを開始してよい最早時刻。
    累計検査完了ロール数を K」ポッファを B（=ROLL_PIPELINE_INITIAL_BUFFER_ROLLS）とれると」
    EC 完了ロールは時系列で (K+B) 本目に到靔した時刻（しのロールの end_dt）未満には開始しない。
    （業務ルール: 任愝の時点で EC_RollEndCount - KENSA_RollEndCount >= B を満たれまで検査を進ゝない」
    の「ロール終了時刻基準」の実装。）
    """
    tid = str(task_id or "").strip()
    if not tid or not _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid):
        return None
    insp_done = int(
        math.floor(float(_pipeline_b2_follower_roll_done_units(task_queue, tid)))
    )
    need_n = insp_done + int(ROLL_PIPELINE_INITIAL_BUFFER_ROLLS)
    ends = _pipeline_b2_ec_roll_end_datetimes_sorted(task_queue, tid)
    if need_n < 1 or len(ends) < need_n:
        return None
    return ends[need_n - 1]
def _pipeline_b2_team_history_names(team_cell) -> set[str]:
    """assigned_history の team 文字列（主・補を「,」「」」区切り）から担当者坝を抽出（NFKC）。"""
    if team_cell is None:
        return set()
    s = str(team_cell).strip()
    if not s:
        return set()
    out: set[str] = set()
    for part in re.split(r"[,」]", s):
        t = part.strip()
        if t:
            out.add(unicodedata.normalize("NFKC", t))
    return out
def _pipeline_b2_assigned_member_names_nfkc_for_side(
    task_queue: list, task_id: str, *, ec_side: bool
) -> set[str]:
    """同一依頼の EC 行または検査行の assigned_history に出た担当者坝（NFKC 集合）。"""
    tid = str(task_id or "").strip()
    if not tid:
        return set()
    names: set[str] = set()
    for t in task_queue:
        if str(t.get("task_id") or "").strip() != tid:
            continue
        if ec_side:
            if not t.get("roll_pipeline_ec"):
                continue
        else:
            if not (
                t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")
            ):
                continue
        for h in t.get("assigned_history") or []:
            names |= _pipeline_b2_team_history_names(h.get("team"))
    return names
def _b2_ec_insp_pair_in_queue(task_queue: list, task_id: str) -> bool:
    """同一依頼NOに §B-2/§B-3 の EC 行と後続行（検査または巻返し）の両方はキューにあるか。"""
    tid = str(task_id or "").strip()
    if not tid:
        return False
    return bool(
        _task_queue_has_roll_pipeline_ec_for_tid(task_queue, tid)
        and _roll_pipeline_inspection_task_row_for_tid(task_queue, tid) is not None
    )
def _filter_capable_members_b2_disjoint_teams(
    task: dict, task_queue: list, capable_members: list
) -> list:
    """
    §B-2 / §B-3 同一依頼では」EC 行に一度でも入った者は後続（検査＝巻返し）の候補から外し、
    後続に入った者は EC の候補から外れ。
    （社内ルール: 担当者集合を必う分ける。`PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS` で無効化坯）
    """
    if not capable_members:
        return capable_members
    tid = str(task.get("task_id") or "").strip()
    if not tid or not _b2_ec_insp_pair_in_queue(task_queue, tid):
        return capable_members
    is_ec = bool(task.get("roll_pipeline_ec"))
    is_follower = bool(
        task.get("roll_pipeline_inspection") or task.get("roll_pipeline_rewind")
    )
    if not is_ec and not is_follower:
        return capable_members
    if not PLANNING_B2_EC_FOLLOWER_DISJOINT_TEAMS:
        return capable_members
    if is_ec:
        excl = _pipeline_b2_assigned_member_names_nfkc_for_side(
            task_queue, tid, ec_side=False
        )
    else:
        excl = _pipeline_b2_assigned_member_names_nfkc_for_side(
            task_queue, tid, ec_side=True
        )
    if not excl:
        return capable_members
    filtered = [
        m
        for m in capable_members
        if unicodedata.normalize("NFKC", str(m).strip()) not in excl
    ]
    removed = [m for m in capable_members if m not in filtered]
    if removed and _trace_schedule_task_enabled(tid):
        if is_ec:
            _side = "EC"
        elif task.get("roll_pipeline_rewind"):
            _side = "巻返し"
        else:
            _side = "検査"
        _log_dispatch_trace_schedule(
            tid,
            "[配台トレース task=%s] ブロック判定: B-2担当者分離 side=%s machine=%s "
            "候補除外=%s 残候補=%s(%s)",
            tid,
            _side,
            task.get("machine"),
            ",".join(str(x) for x in removed),
            len(filtered),
            ",".join(str(x) for x in filtered) if filtered else "なし",
        )
    return filtered
def _exclusive_b1_inspection_holder_for_machine(task_queue, occupant_key: str):
    """
    同一実機械（機械名ベースの占有キー）上で」§B-2 熱融着検査または §B-3 巻返しは **既に割付を開始** し残ロールは残る行はあれみ
    しのタスク dict を1件返す（なければ None）。

    パイプライン枠で後続を数ロールうつしか入れない設計のため、枠ゼロの隙間に **別依頼** は坌も設備に入り」
    結果_設備毎の時間割でタスク表示は途中で切り替ゝる事象を防し。占有中は当該実機械では他タスクを試行する。
    """
    m = str(occupant_key or "").strip()
    if not m:
        return None
    holders: list = []
    for t in task_queue:
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        lk = _machine_occupancy_key_resolve(t, _eqt)
        if lk != m:
            continue
        if not (t.get("roll_pipeline_inspection") or t.get("roll_pipeline_rewind")):
            continue
        rem = float(t.get("remaining_units") or 0)
        if rem <= 1e-9:
            continue
        init = float(t.get("initial_remaining_units") or 0)
        started = (init - rem) > 1e-9 or bool(t.get("assigned_history"))
        if not started:
            continue
        holders.append(t)
    if not holders:
        return None
    return min(
        holders,
        key=lambda t: (
            _dispatch_trial_order_key(t),
            str(t.get("task_id") or ""),
            int(t.get("same_request_line_seq") or 0),
        ),
    )
def _need_sheet_pm_column_rank(
    process,
    machine_name,
    need_combo_col_index: dict | None,
) -> int:
    """need シートで左にある「工程名+機械名」列ろど尝さい値（キューで先）。"""
    if not need_combo_col_index:
        return 10**9
    p = str(process or "").strip()
    m = str(machine_name or "").strip()
    if not p or not m:
        return 10**9
    ck = f"{p}+{m}"
    v = need_combo_col_index.get(ck)
    return int(v) if v is not None else 10**9
def _generate_plan_task_queue_sort_key(
    task: dict,
    _req_map: dict,
    _need_rules: list,
    need_combo_col_index: dict | None = None,
) -> tuple:
    """
    generate_plan 冒頭よよよ紝期シフト再試行時の task_queue.sort 用キー。

    1. 加工途中（in_progress）を先
    2. 紝期基準 due_basis_date（回答納期→指定納期。早いろど先）
    3. §B-1 → §B-2/§B-3 帯 → しの他（b_tier）
    4. §B-2/§B-3 帯内のみ EC を未着手の検査＝巻返しより先（b2_queue_sub）
    5. need シート左列ろど先（工程名+機械名列の佝置）
    6. 依頼NOタイブレーク（_task_id_same_machine_due_tiebreak_key）

    _req_map / _need_rules は呼び出し互換のため残す。
    """
    insp = bool(task.get("roll_pipeline_inspection"))
    rw = bool(task.get("roll_pipeline_rewind"))
    ip = bool(task.get("in_progress"))
    ec = bool(task.get("roll_pipeline_ec"))
    if insp and ip:
        b_tier = 0  # §B-1
    elif ec or (insp and not ip) or (rw and not ip):
        b_tier = 1  # §B-2 / §B-3 帯
    else:
        b_tier = 2
    if b_tier == 1:
        if ec:
            b2_queue_sub = 0
        elif (insp and not ip) or (rw and not ip):
            b2_queue_sub = 1
        else:
            b2_queue_sub = 2
    else:
        b2_queue_sub = 0
    need_rank = _need_sheet_pm_column_rank(
        task.get("machine"), task.get("machine_name"), need_combo_col_index
    )
    return (
        0 if ip else 1,
        task["due_basis_date"] or date.max,
        b_tier,
        b2_queue_sub,
        need_rank,
        _task_id_same_machine_due_tiebreak_key(task.get("task_id")),
    )
def _reorder_task_queue_b2_ec_inspection_consecutive(task_queue: list) -> None:
    """
    §B-2 / §B-3: 同一 task_id の `roll_pipeline_ec` 行の直後に」未着手の後続行
    （`roll_pipeline_inspection` または `roll_pipeline_rewind`）を行順で隣接させる。
    """
    if len(task_queue) < 2:
        return
    moved_tids: list[str] = []
    n_rounds = 0
    max_rounds = max(len(task_queue) * 4, 8)
    while n_rounds < max_rounds:
        n_rounds += 1
        by_tid: dict = {}
        for t in task_queue:
            tid = str(t.get("task_id") or "").strip()
            if not tid:
                continue
            if t.get("roll_pipeline_ec"):
                by_tid.setdefault(tid, {})["ec"] = t
            if (t.get("roll_pipeline_inspection") and not t.get("in_progress")) or (
                t.get("roll_pipeline_rewind") and not t.get("in_progress")
            ):
                by_tid.setdefault(tid, {}).setdefault("followers", []).append(t)
        blocks = []
        for tid, d in by_tid.items():
            ec_task = d.get("ec")
            followers = d.get("followers") or []
            if ec_task is None or not followers:
                continue
            followers = sorted(
                followers,
                key=lambda x: (
                    int(x.get("same_request_line_seq") or 0),
                    task_queue.index(x),
                ),
            )
            blocks.append((tid, ec_task, followers))
        if not blocks:
            break
        blocks.sort(key=lambda b: task_queue.index(b[1]))
        moved = False
        for tid, ec_task, followers in blocks:
            chain = [ec_task] + followers
            try:
                indices = [task_queue.index(x) for x in chain]
            except ValueError:
                continue
            if all(indices[i] == indices[0] + i for i in range(len(indices))):
                continue
            insert_at = min(indices)
            for idx in sorted(indices, reverse=True):
                task_queue.pop(idx)
            for j, item in enumerate(chain):
                task_queue.insert(insert_at + j, item)
            moved_tids.append(tid)
            moved = True
            break
        if not moved:
            break
    if moved_tids:
        logging.info(
            "§B-2/§B-3 配台試行順: EC と未着手後続（検査＝巻返し）を隣接した依頼NO: %s",
            ",".join(moved_tids),
        )
def _normalize_dispatch_trial_order_by_process_sequence_within_task_id(
    task_queue: list,
) -> None:
    """
    §A-1: 加工内容由来の process_sequence_rank がある同一依頼NO内では、
    配台試行順番を rank 昇順に揃える（シート行順が EC→検査→スリット 等と逆でも、
    例: W6-4 の スリット→EC→検査 を試行順でも先に回せるようにする）。
    当該依頼NOグループが占める試行順番の数値集合は変えず、割当のみ入れ替える。
    """
    if len(task_queue) < 2:
        return
    by_tid: dict[str, list] = defaultdict(list)
    for t in task_queue:
        tid = str(t.get("task_id") or "").strip()
        if tid:
            by_tid[tid].append(t)
    for tid, group in by_tid.items():
        if len(group) < 2:
            continue
        ranks = [_task_rank_int_or_none(t) for t in group]
        if any(r is None for r in ranks):
            continue
        order_vals: list[float] = []
        for t in group:
            raw = t.get("dispatch_trial_order")
            if raw is None:
                raw = t.get("dispatch_trial_order_from_sheet")
            from planning_core.core.plan_input import dispatch_trial_order_sort_key

            try:
                order_vals.append(
                    dispatch_trial_order_sort_key(raw, default=float("nan"))
                )
            except (TypeError, ValueError):
                order_vals = []
                break
        if len(order_vals) != len(group) or len(set(order_vals)) != len(group):
            continue
        if any(not math.isfinite(v) for v in order_vals):
            continue
        sorted_orders = sorted(order_vals)
        before = [
            {
                "machine": str(t.get("machine") or ""),
                "rank": _task_rank_int_or_none(t),
                "dispatch_trial_order": _dispatch_trial_order_key(t, default=0.0),
            }
            for t in group
        ]
        for t, new_order in zip(
            sorted(group, key=lambda x: (_task_rank_int_or_none(x), int(x.get("same_request_line_seq") or 0))),
            sorted_orders,
        ):
            t["dispatch_trial_order"] = new_order
            if t.get("dispatch_trial_order_from_sheet") is not None:
                t["dispatch_trial_order_from_sheet"] = new_order
        after = [
            {
                "machine": str(t.get("machine") or ""),
                "rank": _task_rank_int_or_none(t),
                "dispatch_trial_order": _dispatch_trial_order_key(t, default=0.0),
            }
            for t in group
        ]
        if before != after:
            logging.info(
                "§A-1 配台試行順: 加工内容順に同一依頼NO内の試行順を揃えた 依頼NO=%s before=%s after=%s",
                tid,
                before,
                after,
            )
def _reorder_task_queue_process_sequence_within_task_id(task_queue: list) -> None:
    """
    §A-1: 同一依頼NO内で process_sequence_rank 昇順に行を隣接させる。
    段階1の配台試行順 1..n 付与前に呼び、複数工程依頼（例: W6-4）が
    スリット→EC→検査 の連続ブロック＋連番になるようにする。
    """
    if len(task_queue) < 2:
        return
    moved: list[str] = []
    max_rounds = max(len(task_queue) * 2, 4)
    for _ in range(max_rounds):
        by_tid: dict[str, list[tuple[int, dict]]] = defaultdict(list)
        for i, t in enumerate(task_queue):
            tid = str(t.get("task_id") or "").strip()
            if tid:
                by_tid[tid].append((i, t))
        candidate: tuple[str, list[tuple[int, dict]]] | None = None
        candidate_min: int | None = None
        for tid, items in by_tid.items():
            if len(items) < 2:
                continue
            if any(_task_rank_int_or_none(t) is None for _, t in items):
                continue
            sorted_items = sorted(
                items,
                key=lambda x: (
                    _task_rank_int_or_none(x[1]),
                    int(x[1].get("same_request_line_seq") or 0),
                    x[0],
                ),
            )
            indices = [i for i, _ in items]
            sorted_indices = [i for i, _ in sorted_items]
            if len(sorted_indices) >= 2 and all(
                sorted_indices[i] == sorted_indices[0] + i
                for i in range(len(sorted_indices))
            ):
                ranks_in_queue = [
                    _task_rank_int_or_none(t)
                    for _, t in sorted(items, key=lambda x: x[0])
                ]
                ranks_expected = [
                    _task_rank_int_or_none(t) for _, t in sorted_items
                ]
                if ranks_in_queue == ranks_expected:
                    continue
            insert_at = min(indices)
            if candidate is None or insert_at < candidate_min:
                candidate = (tid, sorted_items)
                candidate_min = insert_at
        if candidate is None:
            break
        tid, sorted_items = candidate
        indices = sorted([i for i, _ in sorted_items], reverse=True)
        tasks_ordered = [t for _, t in sorted_items]
        insert_at = min(i for i, _ in sorted_items)
        for idx in indices:
            task_queue.pop(idx)
        for j, t in enumerate(tasks_ordered):
            task_queue.insert(insert_at + j, t)
        moved.append(tid)
    if moved:
        logging.info(
            "§A-1 配台試行順: 加工内容順に同一依頼NO内の行を隣接した依頼NO: %s",
            ",".join(dict.fromkeys(moved)),
        )
def _reorder_task_queue_in_progress_front_stable(task_queue: list) -> None:
    """
    加工途中（in_progress）のタスクを試行順の前寄りにまとめる。
    list.sort は安定なので、同一グループ内の相対順は直前の並びのまま維持される。
    """
    if len(task_queue) < 2:
        return
    task_queue.sort(key=lambda t: (0 if bool(t.get("in_progress")) else 1))
def _reorder_task_queue_in_progress_task_id_family_front_stable(
    task_queue: list,
) -> None:
    """
    同一依頼NO（task_id）のいずれかが加工途中なら、その依頼NOの全行を前寄りにまとめる（安定ソート）。
    """
    if len(task_queue) < 2:
        return
    ip_tids = {
        str(t.get("task_id") or "").strip()
        for t in task_queue
        if bool(t.get("in_progress")) and str(t.get("task_id") or "").strip()
    }
    if not ip_tids:
        return
    task_queue.sort(
        key=lambda t: (0 if str(t.get("task_id") or "").strip() in ip_tids else 1,)
    )
def _finalize_dispatch_trial_pattern_queue_after_pattern_sort(
    task_queue: list,
) -> None:
    """
    配台試行順パターン一覧用: ⑤特別ルール相当の共通後処理（パターン用ソートのあと）。
    §B-2/3 EC 隣接 → スリット→SEC 連続 → 接続→SEC 連続 → 加工途中タスク単位を前へ → 試行順 1..n。
    """
    _reorder_task_queue_b2_ec_inspection_consecutive(task_queue)
    _reorder_task_queue_slit_sec_consecutive(task_queue)
    _reorder_task_queue_connection_sec_consecutive(task_queue)
    _reorder_task_queue_in_progress_front_stable(task_queue)
    _reorder_task_queue_process_sequence_within_task_id(task_queue)
    _assign_sequential_dispatch_trial_order(task_queue)
    _normalize_dispatch_trial_order_by_process_sequence_within_task_id(task_queue)
def _due_basis_date_for_dispatch_pattern_sort(t: dict) -> date:
    d = t.get("due_basis_date")
    if isinstance(d, date):
        return d
    return date.max
def _machine_name_primary_for_dispatch_pattern(t: dict) -> str:
    mn = str(t.get("machine_name") or "").strip()
    if not mn:
        mn = str(t.get("machine") or "").strip()
    return unicodedata.normalize("NFKC", mn).casefold()
def _pattern_sort_key_due_priority(t: dict):
    return (
        _due_basis_date_for_dispatch_pattern_sort(t),
        int(t.get("planning_sheet_row_seq") or 10**9),
        _task_id_priority_key(str(t.get("task_id") or "")),
    )
def _pattern_sort_key_machine_then_due(t: dict):
    return (
        _machine_name_primary_for_dispatch_pattern(t),
        _due_basis_date_for_dispatch_pattern_sort(t),
        int(t.get("planning_sheet_row_seq") or 10**9),
        _task_id_priority_key(str(t.get("task_id") or "")),
    )
def _raw_input_date_for_dispatch_pattern_sort(t: dict) -> date:
    """実効原反投入日（キュー上の raw_input_date＝上書き反映済み）。欠損は並びの末尾寄せ。"""
    rid = t.get("raw_input_date")
    if isinstance(rid, datetime):
        return rid.date()
    if isinstance(rid, date):
        return rid
    return date.max
def _pattern_sort_key_machine_then_raw_input_date(t: dict):
    """P7: 機械名でグループ化し、グループ内は原反投入日が早い順（同一機械内のタイブレークは行順・依頼NO）。"""
    return (
        _machine_name_primary_for_dispatch_pattern(t),
        _raw_input_date_for_dispatch_pattern_sort(t),
        int(t.get("planning_sheet_row_seq") or 10**9),
        _task_id_priority_key(str(t.get("task_id") or "")),
    )
def _pattern_p3_span_days_due_minus_raw(t: dict) -> int | None:
    """
    納期基準日 − 原反投入日の暦日数（タスクごと）。いずれか欠けるときは None。
    （「原反投入日から納期までの日数」と同じ差分の符号）
    """
    due = _due_basis_date_for_dispatch_pattern_sort(t)
    if due == date.max:
        return None
    rid = t.get("raw_input_date")
    if not isinstance(rid, date):
        return None
    return (due - rid).days
def _pattern_p3_machine_group_span_sum_map(task_queue: list) -> dict[str, int]:
    """
    機械名グループごとに、上記スパン日数の合計（納期−原反の合計が小さいほど窓が狭い）。
    グループ内に有効なタスクが1件も無いときは大きな定数を返す（並びの末尾寄せ）。
    """
    spans_by_machine: dict[str, list[int]] = defaultdict(list)
    machines_seen: set[str] = set()
    for t in task_queue:
        mn = _machine_name_primary_for_dispatch_pattern(t)
        machines_seen.add(mn)
        sp = _pattern_p3_span_days_due_minus_raw(t)
        if sp is not None:
            spans_by_machine[mn].append(sp)
    large = 10**9
    out: dict[str, int] = {}
    for mn in machines_seen:
        spans = spans_by_machine.get(mn) or []
        out[mn] = sum(spans) if spans else large
    return out
def _apply_dispatch_trial_pattern_p3_sort(task_queue: list) -> None:
    """
    P3: ①納期順（機械グループ内）、②機械名でグループ化、③グループ単位で
    （納期基準−原反投入日）の暦日合計が小さい機械から、④加工途中の依頼NO（同一task_id）を前へ、
    ⑤_finalize（§B EC・スリット→SEC・加工途中行・試行順付与）。
    """
    sums = _pattern_p3_machine_group_span_sum_map(task_queue)

    def sort_key(t: dict):
        mn = _machine_name_primary_for_dispatch_pattern(t)
        return (
            sums.get(mn, 10**9),
            _due_basis_date_for_dispatch_pattern_sort(t),
            int(t.get("planning_sheet_row_seq") or 10**9),
            _task_id_priority_key(str(t.get("task_id") or "")),
        )

    task_queue.sort(key=sort_key)
    _reorder_task_queue_in_progress_task_id_family_front_stable(task_queue)
    _finalize_dispatch_trial_pattern_queue_after_pattern_sort(task_queue)
def _pattern_sort_key_p4_due_minus_raw(t: dict):
    """P4 用: （納期基準−原反投入日）の暦日が小さい順。欠損は末尾寄せ。"""
    sp = _pattern_p3_span_days_due_minus_raw(t)
    span_k = sp if sp is not None else 10**9
    return (
        span_k,
        _due_basis_date_for_dispatch_pattern_sort(t),
        int(t.get("planning_sheet_row_seq") or 10**9),
        _task_id_priority_key(str(t.get("task_id") or "")),
    )
def _apply_dispatch_trial_pattern_p4_sort(task_queue: list) -> None:
    """
    P4: ①（納期基準−原反投入日）の暦日が小さい順、②加工途中の依頼NOを前へ、
    ③_finalize（§B EC・スリット→SEC・加工途中行・試行順付与）。
    """
    task_queue.sort(key=_pattern_sort_key_p4_due_minus_raw)
    _reorder_task_queue_in_progress_task_id_family_front_stable(task_queue)
    _finalize_dispatch_trial_pattern_queue_after_pattern_sort(task_queue)
def _apply_dispatch_trial_pattern_sort_pipeline(
    task_queue: list,
    sort_key,
) -> None:
    """パターン用の先頭ソートのあと、§B 隣接・スリット→SEC のあと最後に加工途中を前へ、連番付与。"""
    task_queue.sort(key=sort_key)
    _finalize_dispatch_trial_pattern_queue_after_pattern_sort(task_queue)
def _dispatch_trial_pattern_job_list() -> list[tuple[str, str, int | None, object]]:
    """試行順パターン P1～P7（決定論のみ）。P5/P6 はプローブと原反日シフトを含む。"""
    return [
        ("P1", "納期最優先", None, _pattern_sort_key_due_priority),
        ("P2", "機械名グループ+納期", None, _pattern_sort_key_machine_then_due),
        (
            "P3",
            "納期順・機械グループ(納期−原反合計が短い順)・途中依頼優先",
            None,
            _DISPATCH_TRIAL_PATTERN_P3_SORT,
        ),
        (
            "P4",
            "納期−原反日数の短い順・途中依頼優先",
            None,
            _DISPATCH_TRIAL_PATTERN_P4_SORT,
        ),
        (
            "P5",
            "P2→納期遅れ依頼のみ原反-1日→P2",
            None,
            _DISPATCH_TRIAL_PATTERN_P5_SORT,
        ),
        (
            "P6",
            "P5後に納期遅れのみ原反さらに-1日→P2",
            None,
            _DISPATCH_TRIAL_PATTERN_P6_SORT,
        ),
        (
            "P7",
            "機械名グループ+原反投入日早い順",
            None,
            _pattern_sort_key_machine_then_raw_input_date,
        ),
    ]
def _read_result_task_sheet_for_stage2_io(plan_xlsx: str) -> pd.DataFrame | None:
    """結果_タスク一覧: PM_AI_PLAN_RESULT_TASK_JSON サイドカー優先、空・無効時は xlsx シート。"""
    if not plan_xlsx or not os.path.isfile(plan_xlsx):
        return None
    df_t = read_result_task_dataframe(plan_xlsx)
    if df_t is not None and not getattr(df_t, "empty", True):
        return df_t
    try:
        return pd.read_excel(plan_xlsx, sheet_name=RESULT_TASK_SHEET_NAME)
    except Exception:
        return None
def _late_task_ids_missed_answer_deadline_from_plan_xlsx(plan_xlsx: str) -> set[str]:
    """
    結果_タスク一覧の「納期を満たすか？」が「いいえ」の行のタスクID（＝依頼NO）集合。
    スコア集計と同趣旨（未割当・欠損は遅れに含めない）。
    """
    out: set[str] = set()
    if not plan_xlsx or not os.path.isfile(plan_xlsx):
        return out
    df_t = _read_result_task_sheet_for_stage2_io(plan_xlsx)
    if df_t is None or getattr(df_t, "empty", True):
        return out
    df_t.columns = [str(c).strip() for c in df_t.columns]
    col_late = _result_task_due_met_column_in_df_columns(df_t.columns)
    col_tid = "タスクID"
    if col_late is None or col_tid not in df_t.columns:
        return out
    for _, row in df_t.iterrows():
        tid = str(row.get(col_tid, "") or "").strip()
        if not tid or tid.lower() in ("nan", "none"):
            continue
        v = row.get(col_late)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s in ("いいえ", "否") or s.casefold() in ("いいえ", "no", "false"):
            out.add(tid)
    return out
def _dataframe_shift_raw_input_dates_minus_one_day_for_task_ids(
    df: "pd.DataFrame",
    task_ids: set[str],
) -> int:
    """
    依頼NO が task_ids に含まれる行について、原反投入日を解釈できたセルのみ 1 暦日前にずらす。
    変更したセル数を返す。
    """
    if not task_ids:
        return 0
    n_changed = 0
    col = TASK_COL_RAW_INPUT_DATE
    if col not in df.columns:
        return 0
    ci = df.columns.get_loc(col)
    if isinstance(ci, slice):
        return 0
    try:
        col_idx = ci.__index__()
    except (AttributeError, TypeError):
        return 0
    try:
        col_dtype = df.dtypes.iloc[col_idx]
    except Exception:
        col_dtype = None
    for ri in range(len(df)):
        tid = planning_task_id_str_from_plan_row(df.iloc[ri])
        if tid not in task_ids:
            continue
        val = df.iat[ri, col_idx]
        d = parse_optional_date(val)
        if d is None:
            continue
        new_d = d - timedelta(days=1)
        _cell_val = (
            pd.Timestamp(new_d)
            if col_dtype is not None
            and pd.api.types.is_datetime64_any_dtype(col_dtype)
            else new_d
        )
        df.iat[ri, col_idx] = _cell_val
        n_changed += 1
    return n_changed
def _build_result_sheet_effective_raw_input_date_by_line(
    tasks_df_opt: "pd.DataFrame | None",
) -> dict[tuple[str, str], date | None]:
    """
    配台計画 DataFrame から、(依頼NO, 工程名) キーごとの実効原反日（上書き列優先）を返す。
    結果_タスク一覧の原反列・試行前比較に使う。
    """
    out: dict[tuple[str, str], date | None] = {}
    if tasks_df_opt is None or getattr(tasks_df_opt, "empty", True):
        return out
    for _, _r in tasks_df_opt.iterrows():
        if _plan_row_exclude_from_assignment(_r):
            continue
        _tid = str(_planning_df_cell_scalar(_r, TASK_COL_TASK_ID) or "").strip()
        _mach = str(_planning_df_cell_scalar(_r, TASK_COL_MACHINE) or "").strip()
        if not _tid or not _mach:
            continue
        _rid = parse_optional_date(_planning_df_cell_scalar(_r, TASK_COL_RAW_INPUT_DATE))
        if isinstance(_rid, datetime):
            _rid = _rid.date()
        out[(_tid, _mach)] = _rid if isinstance(_rid, date) else None
    return out
def _coerce_task_raw_input_to_date(val) -> date | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None
def _build_dispatch_trial_pattern_p5_task_queue(
    planning_df: "pd.DataFrame",
    tq_frozen_from_planning_df: list,
    *,
    run_date: date,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
    equipment_list: list,
    gpo: dict,
    probe_stage2_root: str,
) -> tuple[list, "pd.DataFrame | None"]:
    """
    ① P2 並びで試行順を付与した planning_df のまま段階2を 1 回プローブし、
    ② 納期遅れ（納期を満たすか？＝いいえ）の依頼NO について原反投入日（＋上書き）を 1 日前、
    ③ 変更後 DataFrame でキューを組み直し P2 並べ＋ finalize。
    遅れ依頼が無い・プローブ失敗時は (P2 の tq, None)。遅れありでシフトしたときは (tq, shifted_df)。
    """
    tq_p2 = copy.deepcopy(tq_frozen_from_planning_df)
    _apply_dispatch_trial_pattern_sort_pipeline(tq_p2, _pattern_sort_key_machine_then_due)

    df_probe = planning_df.copy()
    _apply_pattern_dispatch_trial_orders_to_tasks_df(df_probe, tq_p2)
    try:
        os.makedirs(probe_stage2_root, exist_ok=True)
    except OSError as e:
        logging.warning("P5: プローブ出力フォルダを作成できません: %s（P2 のみ）", e)
        return tq_p2, None

    paths = None
    try:
        paths = _generate_plan_impl(
            tasks_df_override=df_probe,
            stage2_output_root=probe_stage2_root,
            skip_remove_prior_stage2_workbooks=True,
            return_output_paths=True,
        )
    except Exception:
        logging.exception("P5: P2 プローブ段階2で例外（P2 の試行順のみ採用）")

    prod_path = (paths or {}).get("production_plan") or ""
    probe_ok = bool(prod_path and os.path.isfile(prod_path))
    late_tids: set[str] = set()
    if probe_ok:
        late_tids = _late_task_ids_missed_answer_deadline_from_plan_xlsx(prod_path)

    try:
        shutil.rmtree(probe_stage2_root, ignore_errors=True)
    except Exception:
        pass

    if not probe_ok:
        logging.info("P5: プローブ結果が得られなかったため P2 の試行順のみ返します。")
        return tq_p2, None
    if not late_tids:
        logging.info("P5: 納期遅れ依頼なし（P2 と同一試行順）。")
        return tq_p2, None

    df5 = planning_df.copy()
    n_shift = _dataframe_shift_raw_input_dates_minus_one_day_for_task_ids(df5, late_tids)
    logging.info(
        "P5: 納期遅れ依頼 %s 件・原反日セル %s 件を 1 日前にシフトして P2 を再適用します。",
        len(late_tids),
        n_shift,
    )
    ai5 = analyze_task_special_remarks(df5, reference_year=run_date.year)
    tq5 = build_task_queue_from_planning_df(
        df5, run_date, req_map, ai5, gpo, equipment_list
    )
    if not tq5:
        logging.warning("P5: シフト後に配台対象タスクが空のため P2 にフォールバックします。")
        return tq_p2, None
    _apply_dispatch_trial_pattern_sort_pipeline(tq5, _pattern_sort_key_machine_then_due)
    return tq5, df5
def _build_dispatch_trial_pattern_p6_task_queue(
    planning_df: "pd.DataFrame",
    tq_frozen_from_planning_df: list,
    *,
    run_date: date,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
    equipment_list: list,
    gpo: dict,
    nested_probe_parent: str,
) -> tuple[list, "pd.DataFrame | None"]:
    """
    ① P5 と同じ（P2 プローブ→納期遅れ依頼のみ原反 1 日前→P2）まで実行。
    ② その結果の計画表（P5 でシフト済みならその DataFrame、未シフトなら元表）に P5 確定の試行順を載せて再度段階2をプローブ。
    ③ まだ納期遅れの依頼のみ原反投入日（＋上書き）をさらに 1 暦日前にし、キュー組み直し＋ P2 finalize。
    ②で遅れなし・プローブ失敗時は P5 の戻り (tq, df_override) と同じ。
    """
    try:
        os.makedirs(nested_probe_parent, exist_ok=True)
    except OSError as e:
        logging.warning("P6: 作業フォルダを作成できません: %s（P2 のみ）", e)
        tq = copy.deepcopy(tq_frozen_from_planning_df)
        _apply_dispatch_trial_pattern_sort_pipeline(tq, _pattern_sort_key_machine_then_due)
        return tq, None

    p5_inner = os.path.join(nested_probe_parent, "inner_p5")
    tq5, df5 = _build_dispatch_trial_pattern_p5_task_queue(
        planning_df,
        tq_frozen_from_planning_df,
        run_date=run_date,
        req_map=req_map,
        need_rules=need_rules,
        need_combo_col_index=need_combo_col_index,
        equipment_list=equipment_list,
        gpo=gpo,
        probe_stage2_root=p5_inner,
    )
    planning_base = df5 if df5 is not None else planning_df
    tq_after_p5 = tq5

    df_probe2 = planning_base.copy()
    _apply_pattern_dispatch_trial_orders_to_tasks_df(df_probe2, tq_after_p5)
    p6_probe = os.path.join(nested_probe_parent, "after_p5_probe")
    paths2 = None
    try:
        os.makedirs(p6_probe, exist_ok=True)
    except OSError as e:
        logging.warning("P6: 第2プローブ用フォルダを作成できません: %s", e)
        try:
            shutil.rmtree(nested_probe_parent, ignore_errors=True)
        except Exception:
            pass
        return tq5, df5

    try:
        paths2 = _generate_plan_impl(
            tasks_df_override=df_probe2,
            stage2_output_root=p6_probe,
            skip_remove_prior_stage2_workbooks=True,
            return_output_paths=True,
        )
    except Exception:
        logging.exception("P6: P5 後の第2プローブ段階2で例外（P5 結果で打ち切り）")

    prod2 = (paths2 or {}).get("production_plan") or ""
    probe2_ok = bool(prod2 and os.path.isfile(prod2))
    late2: set[str] = set()
    if probe2_ok:
        late2 = _late_task_ids_missed_answer_deadline_from_plan_xlsx(prod2)

    try:
        shutil.rmtree(p6_probe, ignore_errors=True)
    except Exception:
        pass
    try:
        shutil.rmtree(nested_probe_parent, ignore_errors=True)
    except Exception:
        pass

    if not probe2_ok:
        logging.info("P6: 第2プローブが得られなかったため P5 相当で返します。")
        return tq5, df5
    if not late2:
        logging.info("P6: P5 後のプローブで納期遅れ依頼なし（P5 と同一）。")
        return tq5, df5

    df6 = planning_base.copy()
    n2 = _dataframe_shift_raw_input_dates_minus_one_day_for_task_ids(df6, late2)
    logging.info(
        "P6: P5 後も遅れの依頼 %s 件・原反日セル %s 件をさらに 1 日前にシフトして P2 を再適用します。",
        len(late2),
        n2,
    )
    ai6 = analyze_task_special_remarks(df6, reference_year=run_date.year)
    tq6 = build_task_queue_from_planning_df(
        df6, run_date, req_map, ai6, gpo, equipment_list
    )
    if not tq6:
        logging.warning("P6: 第2シフト後に配台対象タスクが空のため P5 にフォールバックします。")
        return tq5, df5
    _apply_dispatch_trial_pattern_sort_pipeline(tq6, _pattern_sort_key_machine_then_due)
    return tq6, df6
def _iter_dispatch_trial_pattern_variant_queues(
    tq_template_frozen: list,
    pattern_jobs: list[tuple[str, str, int | None, object]],
    *,
    p5_bundle: dict | None = None,
):
    """
    各パターンの確定 task_queue（ディープコピー）を順に返す。
    戻り: (pid, pname, tq, df_override)。df_override は P5/P6 で原反シフト後の計画 DataFrame のみ（それ以外 None）。
    """
    for pid, pname, seed, sk in pattern_jobs:
        if sk is _DISPATCH_TRIAL_PATTERN_P6_SORT:
            pb = p5_bundle or {}
            planning_df = pb.get("planning_df")
            if planning_df is None:
                logging.warning(
                    "試行順 P6: planning_df が無いため P2 のみ適用します（id=%s）。",
                    pid,
                )
                tq = copy.deepcopy(tq_template_frozen)
                _apply_dispatch_trial_pattern_sort_pipeline(tq, _pattern_sort_key_machine_then_due)
                yield pid, pname, tq, None
                continue
            nest = pb.get("p6_nested_probe_parent") or os.path.join(
                output_dir,
                "dispatch_pattern_stage2",
                "p6_probe_fallback",
                datetime.now().strftime("%Y%m%d_%H%M%S_%f"),
            )
            tq6, df6 = _build_dispatch_trial_pattern_p6_task_queue(
                planning_df,
                tq_template_frozen,
                run_date=pb["run_date"],
                req_map=pb["req_map"],
                need_rules=pb["need_rules"],
                need_combo_col_index=pb["need_combo_col_index"],
                equipment_list=pb["equipment_list"],
                gpo=pb["gpo"],
                nested_probe_parent=nest,
            )
            yield pid, pname, tq6, df6
            continue

        if sk is _DISPATCH_TRIAL_PATTERN_P5_SORT:
            pb = p5_bundle or {}
            planning_df = pb.get("planning_df")
            if planning_df is None:
                logging.warning(
                    "試行順 P5: planning_df が無いため P2 のみ適用します（id=%s）。",
                    pid,
                )
                tq = copy.deepcopy(tq_template_frozen)
                _apply_dispatch_trial_pattern_sort_pipeline(tq, _pattern_sort_key_machine_then_due)
                yield pid, pname, tq, None
                continue
            probe_root = pb.get("probe_stage2_root") or os.path.join(
                output_dir,
                "dispatch_pattern_stage2",
                "p5_probe_fallback",
                datetime.now().strftime("%Y%m%d_%H%M%S_%f"),
            )
            tq5, df5 = _build_dispatch_trial_pattern_p5_task_queue(
                planning_df,
                tq_template_frozen,
                run_date=pb["run_date"],
                req_map=pb["req_map"],
                need_rules=pb["need_rules"],
                need_combo_col_index=pb["need_combo_col_index"],
                equipment_list=pb["equipment_list"],
                gpo=pb["gpo"],
                probe_stage2_root=probe_root,
            )
            yield pid, pname, tq5, df5
            continue

        tq = copy.deepcopy(tq_template_frozen)
        if sk is not None:
            if sk is _DISPATCH_TRIAL_PATTERN_P3_SORT:
                _apply_dispatch_trial_pattern_p3_sort(tq)
            elif sk is _DISPATCH_TRIAL_PATTERN_P4_SORT:
                _apply_dispatch_trial_pattern_p4_sort(tq)
            else:
                _apply_dispatch_trial_pattern_sort_pipeline(tq, sk)
        else:
            logging.warning(
                "試行順パターン id=%s: sort キーが無いため納期最優先で並べます。",
                pid,
            )
            _apply_dispatch_trial_pattern_sort_pipeline(tq, _pattern_sort_key_due_priority)
        yield pid, pname, tq, None
def _apply_pattern_dispatch_trial_orders_to_tasks_df(
    tasks_df: "pd.DataFrame",
    pattern_tq: list,
) -> None:
    """
    パターン確定後の dispatch_trial_order を DataFrame の配台試行順番列へ書き戻す。
    全行指定経路で段階2が同じ試行順を採用する（§B 隣接はパターン側で既に反映済み）。
    """
    col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    if col not in tasks_df.columns:
        tasks_df[col] = float("nan")
    tasks_df[col] = float("nan")
    ci = tasks_df.columns.get_loc(col)
    for t in pattern_tq:
        ii = t.get("planning_df_iloc")
        dto = t.get("dispatch_trial_order")
        if ii is None or dto is None:
            continue
        try:
            ri = int(ii)
            dv = int(dto)
        except (TypeError, ValueError):
            continue
        if ri < 0 or ri >= len(tasks_df):
            continue
        tasks_df.iat[ri, ci] = dv
def _score_dispatch_pattern_stage2_workbook(plan_xlsx: str) -> dict:
    """
    段階2の production_plan xlsx から簡易スコアを読み取る。
        ①納期（納期を満たすか？ の はい率）
    ②メンバー（結果_メンバー別作業割合：日単位で配台実作業分が一度もない日は除外し、
      メンバー単位で「0.0% (0/0分)」の枠も除外し、残りのセルの % を平坦化して平均）
    ③設備（結果_設備毎の時間割 の日付列あたりの非空セル数合計＝稼働スロット量の参考）
    """
    out: dict = {
        "納期_判定対象件数": 0,
        "納期_遅れ件数": 0,
        "納期_遵守率": None,
        "メンバー_平均作業割合_pct": None,
        "設備_稼働セル数": None,
        "スコア備考": "",
    }
    if not plan_xlsx or not os.path.isfile(plan_xlsx):
        out["スコア備考"] = "結果ブックが見つかりません。"
        return out
    df_t = _read_result_task_sheet_for_stage2_io(plan_xlsx)
    if df_t is None:
        out["スコア備考"] = "結果_タスク一覧の読込失敗（サイドカー・xlsx）。"
        return out
    df_t.columns = [str(c).strip() for c in df_t.columns]
    col_late = _result_task_due_met_column_in_df_columns(df_t.columns)
    col_tid = "タスクID"
    if col_late is None:
        out["スコア備考"] = (
            f"列「{RESULT_TASK_COL_PLAN_END_BY_ANSWER_OR_SPEC_16}」（旧「配台済_回答指定16時まで」）がありません。"
        )
        return out
    mask = df_t[col_tid].astype(str).str.strip().ne("") & df_t[col_tid].astype(str).str.lower().ne("nan")
    sub = df_t.loc[mask, col_late].astype(str).str.strip()
    sub = sub[sub.ne("")]
    n = int(len(sub))
    if n == 0:
        out["スコア備考"] = "タスク行がありません。"
        return out
    late = int((sub.eq("いいえ") | sub.str.strip().str.upper().eq("いいえ")).sum())
    out["納期_判定対象件数"] = n
    out["納期_遅れ件数"] = late
    out["納期_遵守率"] = round((n - late) / n * 100.0, 2) if n else None

    df_u = None
    for _util_sheet in (
        RESULT_MEMBER_WORK_UTIL_SHEET_NAME,
        "結果_メンバー別作業割引",
    ):
        try:
            df_u = pd.read_excel(plan_xlsx, sheet_name=_util_sheet)
            break
        except Exception:
            continue
    pct_vals: list[float] = []
    if df_u is not None and not df_u.empty:
        _util_cell_re = re.compile(
            r"^([\d.]+)\s*%\s*(?:\((\d+)/(\d+)分\))?\s*$",
            re.ASCII,
        )
        for _, _row in df_u.iterrows():
            _row_pcts: list[float] = []
            _day_max_worked = 0
            for _c in df_u.columns:
                if str(_c).strip() in ("年月日", ""):
                    continue
                s = str(_row[_c]).strip()
                m = _util_cell_re.match(s)
                if not m:
                    continue
                try:
                    _p = float(m.group(1))
                except ValueError:
                    continue
                if m.group(2) is not None and m.group(3) is not None:
                    try:
                        _wk = int(m.group(2))
                        _tot = int(m.group(3))
                        if _tot > 0:
                            _day_max_worked = max(_day_max_worked, _wk)
                        else:
                            # (0/0分) 等、配台母数に含めない枠は平均にも含めない
                            continue
                    except (TypeError, ValueError):
                        pass
                _row_pcts.append(_p)
            # その日いずれのメンバーも配台実作業 0 分なら、その日の行は平均の母数に含めない
            if _day_max_worked <= 0:
                continue
            pct_vals.extend(_row_pcts)
    if pct_vals:
        out["メンバー_平均作業割合_pct"] = round(sum(pct_vals) / len(pct_vals), 2)

    try:
        df_e = pd.read_excel(plan_xlsx, sheet_name=RESULT_EQUIPMENT_SCHEDULE_SHEET_NAME)
    except Exception:
        df_e = None
    if df_e is not None and not df_e.empty:
        filled = 0
        for _c in df_e.columns:
            cs = str(_c).strip()
            if "日時帯" in cs or cs == "日時帯":
                continue
            for _v in df_e[_c].tolist():
                if _v is None or (isinstance(_v, float) and pd.isna(_v)):
                    continue
                s = str(_v).strip()
                if s and s not in ("休", "—", "-"):
                    filled += 1
        out["設備_稼働セル数"] = filled

    return out
def _excel_hyperlink_formula_file(abs_path: str, display: str) -> str:
    """ローカル .xlsx への HYPERLINK 数式（表示テキスト付き）。"""
    p = os.path.abspath(abs_path).replace("\\", "/")
    disp = (display or os.path.basename(abs_path)).replace('"', '""')
    return f'=HYPERLINK("{p}","{disp}")'
def _xlwings_write_dispatch_pattern_stage2_summary_sheet(
    wb,
    summary_rows: list[dict],
    *,
    batch_root: str = "",
    total_batch_seconds: float | None = None,
) -> None:
    """
    マクロブックにパターン別段階2の結果リンク・スコア・採用用 UI を書く。
    データ行は 6 行目から。B3=採用パターンID（プルダウン）。B2=バッチ出力ルート。
    C2=合計処理時間のラベル、D2=合計秒（バッチ全体の壁時計）。
    """
    sheet_name = DISPATCH_PATTERN_STAGE2_SUMMARY_SHEET_NAME
    try:
        ws = wb.sheets[sheet_name]
    except Exception:
        ws = wb.sheets.add(name=sheet_name, after=wb.sheets[PLAN_INPUT_SHEET_NAME])
    try:
        ur = ws.used_range
        if ur:
            ur.clear_contents()
    except Exception:
        pass
    intro = (
        "各パターンの配台試行順を「配台計画_タスク入力」に反映したうえで段階2のみ実行し、"
        "output/dispatch_pattern_stage2/<実行時刻>/<パターンID>/ に "
        "計画*.xlsx / 人員*.xlsx を保存した結果です。"
        " 生産計画ブック・メンバー日程ブック列のリンクは当該ファイルの絶対パスです。スコアと参考スコアを比較してください。"
        " 最適と思う案のパターンIDを B3 に選び（プルダウン可）、ブックを保存してから"
        "「試行順パターン採用を計画へ反映」マクロで配台計画シートの配台試行順番に書き戻します。"
        f" シミュレーション件数上限は {_dispatch_pattern_stage2_max_patterns()} 件です。"
    )
    headers = [
        "パターンID",
        "パターン名",
        "生産計画ブック",
        "メンバー日程ブック",
        "納期_判定対象件数",
        "納期_遅れ件数",
        "納期_遵守率(%)",
        "メンバー_平均作業割合(%)",
        "設備_稼働セル数(参考)",
        "参考スコア(自動)",
        "処理時間(秒)",
        "備考",
    ]
    total_cell = (
        round(total_batch_seconds, 2)
        if isinstance(total_batch_seconds, (int, float))
        else ""
    )
    mat: list[list] = [
        [intro],
        [
            "バッチ出力ルート",
            (batch_root or "").strip(),
            "合計処理時間(秒)",
            total_cell,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        ["採用パターンID", "", "", "", "", "", "", "", "", "", "", ""],
        [
            "※ B3 に一覧のパターンIDを指定し保存後、Python「apply_dispatch_pattern_stage2_selection.py」"
            " またはマクロで反映。",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        headers,
    ]
    for r in summary_rows:
        mat.append(
            [
                r.get("パターンID", ""),
                r.get("パターン名", ""),
                "",
                "",
                r.get("納期_判定対象件数", ""),
                r.get("納期_遅れ件数", ""),
                r.get("納期_遵守率", ""),
                r.get("メンバー_平均作業割合_pct", ""),
                r.get("設備_稼働セル数", ""),
                r.get("参考スコア(自動)", ""),
                r.get("処理時間(秒)", ""),
                r.get("備考", ""),
            ]
        )
    n_cols = max((len(x) for x in mat), default=1)
    pad = []
    for row in mat:
        rr = list(row)
        if len(rr) < n_cols:
            rr.extend([""] * (n_cols - len(rr)))
        pad.append(rr)
    n_rows = len(pad)
    ws.range((1, 1)).resize(n_rows, n_cols).value = pad
    data_start = 6
    # C,D 列に数式を上書き（データ行のみ）… output/dispatch_pattern_stage2/<日付>/<パターンID>/ の xlsx 絶対パス
    for i, r in enumerate(summary_rows, start=data_start):
        try:
            fp = r.get("_path_plan")
            fm = r.get("_path_member")
            if fp:
                ws.range((i, 3)).formula = _excel_hyperlink_formula_file(
                    fp, os.path.basename(fp)
                )
            if fm:
                ws.range((i, 4)).formula = _excel_hyperlink_formula_file(
                    fm, os.path.basename(fm)
                )
        except Exception:
            logging.debug("パターン段階2サマリ: HYPERLINK 設定失敗（無視）", exc_info=True)
    try:
        ws.range((1, 1), (1, n_cols)).merge()
        ws.range((1, 1)).api.WrapText = True
        ws.range((5, 1), (5, n_cols)).api.Font.Bold = True
    except Exception:
        pass
    n_pat = len(summary_rows)
    if n_pat > 0:
        try:
            addr = ws.range((data_start, 1)).resize(n_pat, 1).get_address(
                row_absolute=True,
                column_absolute=True,
                include_sheetname=True,
            )
            v = ws.range((3, 2)).api.Validation
            try:
                v.Delete()
            except Exception:
                pass
            v.Add(3, 1, 1, Formula1=f"={addr}")
        except Exception:
            logging.debug("パターン段階2サマリ: B3 入力規則の設定に失敗（無視）", exc_info=True)
    try:
        ws.range((2, 1), (2, n_cols)).api.WrapText = True
        ws.range((4, 1), (4, n_cols)).merge()
        ws.range((4, 1)).api.WrapText = True
    except Exception:
        pass
    try:
        ws.used_range.columns.api.AutoFit()
    except Exception:
        pass
def _build_dispatch_trial_pattern_list_matrix(
    tasks_df: "pd.DataFrame",
    run_date: date,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
    equipment_list: list,
) -> list[list]:
    """
    パターン①納期最優先、②機械名グループ＋納期、③P3（納期順・機械グループの納期−原反合計順・途中依頼優先）、
    ④P4（納期−原反日数の短い順・途中依頼優先）、⑤P5（P2 プローブ後に納期遅れ依頼のみ原反 1 日前→P2 再適用）、
    ⑥P6（P5 後に再プローブし、まだ遅れの依頼のみ原反をさらに 1 日前→P2）、
    ⑦P7（機械名グループ＋グループ内は実効原反投入日の早い順）
    の確定後試行順を長形式で返す（先頭に説明行・見出し行）。
    """
    dto_col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    df = tasks_df.copy()
    if dto_col in df.columns:
        if pd.api.types.is_numeric_dtype(df[dto_col]):
            df[dto_col] = float("nan")
        else:
            df[dto_col] = ""

    global_priority_raw = load_main_sheet_global_priority_override_text()
    members_for_gpo: list = []
    try:
        with pd.ExcelFile(_master_workbook_path_resolved()) as _xf:
            _skills = pd.read_excel(_xf, sheet_name="skills", header=None)
        for r in range(2, _skills.shape[0]):
            cell = _skills.iat[r, 0]
            if pd.isna(cell):
                continue
            name = str(cell).strip()
            if name and name.lower() not in ("nan", "none", "null"):
                members_for_gpo.append(name)
    except Exception:
        members_for_gpo = []
    gpo = analyze_global_priority_override_comment(
        global_priority_raw, members_for_gpo, run_date.year, ai_sheet_sink={}
    )
    ai_by_tid = analyze_task_special_remarks(df, reference_year=run_date.year)
    tq_template = build_task_queue_from_planning_df(
        df, run_date, req_map, ai_by_tid, gpo, equipment_list
    )
    if not tq_template:
        return [
            [
                "（配台対象タスクがありません。依頼NO・工程名・残数量を確認してください。）"
            ]
        ]

    tq_template = copy.deepcopy(tq_template)
    # 一覧シートは参照用のため P1～P7 を全列挙する（段階2バッチの件数上限とは切り離す）。
    pattern_jobs = _dispatch_trial_pattern_job_list()
    intro = (
        "各パターンは、パターン用の並べのあと（P3/P4 は加工途中の同一依頼NOを前寄せ）、"
        "§B-2/3 EC 隣接・スリット→SEC 連続のあと、加工途中行を前へ寄せ、配台試行順 1..n を付与した結果です。"
        " 決定論: P1納期最優先、P2機械名+納期、"
        "P3は機械グループの(納期基準−原反投入日)暦日合計が小さい機械から並べグループ内は納期順、"
        "P4はタスクごとの(納期基準−原反投入日)暦日が小さい順。"
        " P7は機械名でグループ化しグループ内は実効原反投入日（上書き反映）が早い順。"
        " P5は一度 P2 で試行順を付けた計画を段階2でプローブし、"
        "「納期を満たすか？」がいいえの依頼NOだけ原反投入日（と上書き列）を1暦日前にしてから P2 を再適用した試行順。"
        " P6はそのP5後の計画でもう一度プローブし、まだ遅れの依頼のみ原反をさらに1暦日前にしてからP2を再適用。"
        "（プローブは一覧生成時に output/dispatch_pattern_stage2 の p5_list_matrix_probe / p6_list_matrix_probe 配下へ一時出力）。"
        " 「試行順パターン別段階2」バッチのみ DISPATCH_PATTERN_STAGE2_MAX_PATTERNS で件数を抑えます。"
    )
    headers = [
        "パターンID",
        "パターン名",
        "配台試行順番",
        "依頼NO",
        "工程名",
        "機械名",
        TASK_COL_QTY,
        TASK_COL_UNPROCESSED,
        "納期基準",
    ]
    rows: list[list] = [[intro], [], headers]

    _mtx_probe_stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    p5_bundle_mtx = {
        "planning_df": df,
        "run_date": run_date,
        "req_map": req_map,
        "need_rules": need_rules,
        "need_combo_col_index": need_combo_col_index,
        "equipment_list": equipment_list,
        "gpo": gpo,
        "probe_stage2_root": os.path.join(
            output_dir,
            "dispatch_pattern_stage2",
            "p5_list_matrix_probe",
            _mtx_probe_stamp,
        ),
        "p6_nested_probe_parent": os.path.join(
            output_dir,
            "dispatch_pattern_stage2",
            "p6_list_matrix_probe",
            _mtx_probe_stamp,
        ),
    }
    for pid, pname, tq, _df_ov in _iter_dispatch_trial_pattern_variant_queues(
        tq_template, pattern_jobs, p5_bundle=p5_bundle_mtx
    ):
        for t in sorted(tq, key=lambda x: _dispatch_trial_order_key(x)):
            dto = _dispatch_trial_order_key(t, default=0.0)
            tid = str(t.get("task_id") or "").strip()
            proc = str(t.get("machine") or "").strip()
            mname = str(t.get("machine_name") or "").strip()
            db = t.get("due_basis_date")
            db_s = db.strftime("%Y/%m/%d") if isinstance(db, date) else ""
            qty_m = t.get("total_qty_m")
            try:
                qty_out = int(qty_m) if qty_m is not None else None
            except (TypeError, ValueError):
                qty_out = None
            unp_raw = t.get("unprocessed_baseline_m")
            unp_out = None
            if unp_raw is not None:
                try:
                    f = float(unp_raw)
                    if math.isfinite(f):
                        unp_out = int(f) if abs(f - int(f)) < 1e-9 else round(f, 6)
                except (TypeError, ValueError):
                    unp_out = None
            rows.append([pid, pname, dto, tid, proc, mname, qty_out, unp_out, db_s])
    return rows
def _xlwings_format_dispatch_trial_pattern_list_sheet(
    ws_out,
    n_rows: int,
    n_cols: int,
    *,
    header_row: int = 3,
) -> None:
    """
    パターン一覧シートの見やすさ: 1 行目説明の横結合、見出し行の太字、データ範囲を Excel 表（ListObject）にする。
    環境変数 DISPATCH_TRIAL_PATTERN_LIST_NO_EXCEL_TABLE=1 で表のみスキップ（結合・太字は実施）。
    """
    if n_rows < header_row or n_cols < 1:
        return
    try:
        ws_out.range((1, 1), (1, n_cols)).merge()
        c1 = ws_out.range((1, 1)).api
        c1.VerticalAlignment = -4160  # xlTop
        c1.WrapText = True
        c1.HorizontalAlignment = -4131  # xlLeft
    except Exception:
        logging.debug("パターン一覧: 1 行目の結合に失敗（無視）", exc_info=True)
    try:
        ws_out.range((header_row, 1), (header_row, n_cols)).api.Font.Bold = True
    except Exception:
        pass

    no_tbl = (os.environ.get("DISPATCH_TRIAL_PATTERN_LIST_NO_EXCEL_TABLE") or "").strip().lower()
    if no_tbl in ("1", "true", "yes", "on", "y"):
        return
    tbl_name = "TblDispatchTrialPatterns"
    try:
        lots = ws_out.api.ListObjects
        for i in range(int(lots.Count), 0, -1):
            try:
                if str(lots.Item(i).Name) == tbl_name:
                    lots.Item(i).Delete()
            except Exception:
                continue
    except Exception:
        pass
    try:
        tbl_nrows = n_rows - header_row + 1
        if tbl_nrows < 2:
            return
        data_rng = ws_out.range((header_row, 1)).resize(tbl_nrows, n_cols)
        # xlSrcRange=1, HasHeaders=xlYes=1
        ws_out.api.ListObjects.Add(1, data_rng.api, None, 1)
        lots = ws_out.api.ListObjects
        lo = lots.Item(int(lots.Count))
        lo.Name = tbl_name
        try:
            lo.TableStyle = "TableStyleMedium2"
        except Exception:
            pass
    except Exception as e:
        logging.warning("パターン一覧: Excel 表（ListObject）の設定をスキップしました: %s", e)
    try:
        ws_out.used_range.columns.api.AutoFit()
    except Exception:
        pass
def write_dispatch_trial_pattern_list_via_xlwings(
    workbook_path: str | None = None,
    *,
    apply_post_load_mutations: bool = True,
) -> bool:
    """
    マクロブックを Excel で開き、「配台計画_タスク入力」を読み、
    試行順パターン一覧を DISPATCH_TRIAL_PATTERN_LIST_SHEET_NAME に書き込む。
    （関数名は VBA / 既存スクリプト互換のため旧接頭辞を含む。）
    """
    path = (workbook_path or "").strip() or _excel_plan_input_wb().strip()
    if not path:
        logging.error("配台試行順パターン一覧: ブックパスは空です。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("配台試行順パターン一覧: Excel 連携に必要なパッケージがありません。")
        return False
    try:
        wb = xw.Book(path)
        ws = wb.sheets[PLAN_INPUT_SHEET_NAME]
    except Exception as e:
        logging.error("配台試行順パターン一覧: シート接続に失敗: %s", e)
        return False

    mat = _openpyxl_sheet_to_matrix(ws)
    df = _matrix_to_dataframe_header_first(mat)
    if df is None or df.empty:
        logging.warning("配台試行順パターン一覧: データ行はありません。")
        return False

    df = df.copy()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, plan_input_sheet_column_order())
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""

    if apply_post_load_mutations and not _plan_input_dispatch_trial_order_local_only_from_env():
        _apply_planning_sheet_post_load_mutations(
            df,
            path,
            "配台試行順パターン一覧",
            apply_exclude_rules_from_config=False,
            compile_exclude_rules_d_to_e_with_ai=False,
        )

    data_extract_dt, _ = _extract_data_extraction_datetime()
    base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
    run_date = base_now_dt.date()

    try:
        (
            _sd,
            _mem,
            equipment_list,
            req_map,
            need_rules,
            _sm,
            need_combo_col_index,
        ) = load_skills_and_needs()
    except Exception as e:
        logging.exception("配台試行順パターン一覧: master 読込に失敗: %s", e)
        return False

    try:
        matrix = _build_dispatch_trial_pattern_list_matrix(
            df, run_date, req_map, need_rules, need_combo_col_index, equipment_list
        )
    except Exception as e:
        logging.exception("配台試行順パターン一覧: 行列生成に失敗: %s", e)
        return False

    sheet_name = DISPATCH_TRIAL_PATTERN_LIST_SHEET_NAME
    try:
        ws_out = wb.sheets[sheet_name]
    except Exception:
        try:
            ws_out = wb.sheets.add(name=sheet_name, after=wb.sheets[PLAN_INPUT_SHEET_NAME])
        except Exception as e2:
            logging.error("配台試行順パターン一覧: シート作成に失敗: %s", e2)
            return False

    try:
        ur0 = ws_out.used_range
        if ur0:
            ur0.clear_contents()
    except Exception:
        pass

    n_cols = max((len(r) for r in matrix), default=1)
    padded: list[list] = []
    for r in matrix:
        row = list(r)
        if len(row) < n_cols:
            row.extend([None] * (n_cols - len(row)))
        padded.append(row)
    n_rows = len(padded)
    try:
        ws_out.range((1, 1)).resize(n_rows, n_cols).value = padded
    except Exception as e:
        logging.exception("配台試行順パターン一覧: シート書込に失敗: %s", e)
        return False

    try:
        _xlwings_format_dispatch_trial_pattern_list_sheet(ws_out, n_rows, n_cols, header_row=3)
    except Exception:
        logging.exception("配台試行順パターン一覧: 書式・表の適用で例外（続行）")

    try:
        wb.save()
    except Exception as e:
        logging.warning("配台試行順パターン一覧: Save 警告: %s", e)

    logging.info(
        "配台試行順パターン一覧: 「%s」に %s 行を書き込みました。",
        sheet_name,
        n_rows,
    )
    return True
def run_dispatch_trial_pattern_stage2_batch_via_xlwings(
    workbook_path: str | None = None,
    *,
    apply_post_load_mutations: bool = True,
) -> bool:
    """
    各試行順パターン（P1～P7）ごとに段階2を実行し、
    ``output/dispatch_pattern_stage2/<時刻>/<パターンID>/`` に production_plan / member_schedule を保存する。
    バッチ時は計画側および加工実績明細の設備ガントシートを生成しない（スコア比較の負荷軽減）。
    マクロブックに ``DISPATCH_PATTERN_STAGE2_SUMMARY_SHEET_NAME`` へリンクとスコアを書く（openpyxl）。
    """
    path = (workbook_path or "").strip() or _excel_plan_input_wb().strip()
    if not path:
        logging.error("パターン別段階2: ブックパスは空です。")
        return False
    keep_vba = path.lower().endswith(".xlsm")
    wb = None
    try:
        wb = load_workbook(path, keep_vba=keep_vba)
        ws = wb[PLAN_INPUT_SHEET_NAME]
    except Exception as e:
        logging.error("パターン別段階2: シート接続に失敗: %s", e)
        return False

    try:
        _t0 = time_module.perf_counter()
        mat = _openpyxl_sheet_to_matrix(ws)
        df = _matrix_to_dataframe_header_first(mat)
        if df is None or df.empty:
            logging.warning("パターン別段階2: データ行はありません。")
            return False
        df = df.copy()
        df.columns = df.columns.str.strip()
        df = _align_dataframe_headers_to_canonical(df, plan_input_sheet_column_order())
        for c in plan_input_sheet_column_order():
            if c not in df.columns:
                df[c] = ""

        if apply_post_load_mutations and not _plan_input_dispatch_trial_order_local_only_from_env():
            _t_mut0 = time_module.perf_counter()
            _apply_planning_sheet_post_load_mutations(
                df,
                path,
                "配台試行順パターン別段階2",
                apply_exclude_rules_from_config=False,
                compile_exclude_rules_d_to_e_with_ai=False,
            )
        data_extract_dt, _ = _extract_data_extraction_datetime()
        base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
        run_date = base_now_dt.date()

        try:
            _t_master0 = time_module.perf_counter()
            (
                _sd,
                _mem,
                equipment_list,
                req_map,
                need_rules,
                _sm,
                need_combo_col_index,
            ) = load_skills_and_needs()
        except Exception as e:
            logging.exception("パターン別段階2: master 読込に失敗: %s", e)
            return False

        df0 = df.copy()
        dto_col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
        if dto_col in df0.columns:
            if pd.api.types.is_numeric_dtype(df0[dto_col]):
                df0[dto_col] = float("nan")
            else:
                df0[dto_col] = ""

        global_priority_raw = load_main_sheet_global_priority_override_text()
        members_for_gpo: list = []
        try:
            with pd.ExcelFile(_master_workbook_path_resolved()) as _xf:
                _skills = pd.read_excel(_xf, sheet_name="skills", header=None)
            for r in range(2, _skills.shape[0]):
                cell = _skills.iat[r, 0]
                if pd.isna(cell):
                    continue
                name = str(cell).strip()
                if name and name.lower() not in ("nan", "none", "null"):
                    members_for_gpo.append(name)
        except Exception:
            members_for_gpo = []
        _t_build0 = time_module.perf_counter()
        gpo = analyze_global_priority_override_comment(
            global_priority_raw, members_for_gpo, run_date.year, ai_sheet_sink={}
        )
        ai_by_tid = analyze_task_special_remarks(df0, reference_year=run_date.year)
        tq_template = build_task_queue_from_planning_df(
            df0, run_date, req_map, ai_by_tid, gpo, equipment_list
        )
        if not tq_template:
            logging.error("パターン別段階2: 配台対象タスクがありません。")
            return False

        tq_frozen = copy.deepcopy(tq_template)
        pattern_jobs = _dispatch_pattern_stage2_capped_jobs()
        batch_stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        batch_root = os.path.join(output_dir, "dispatch_pattern_stage2", batch_stamp)
        try:
            os.makedirs(batch_root, exist_ok=True)
        except OSError as e:
            logging.error("パターン別段階2: バッチフォルダを作成できません: %s", e)
            return False
        logging.info("パターン別段階2: 出力ルート %s", batch_root)
        _write_dispatch_pattern_stage2_jobs_meta(batch_root, pattern_jobs)

        p5_bundle_batch = {
            "planning_df": df0,
            "run_date": run_date,
            "req_map": req_map,
            "need_rules": need_rules,
            "need_combo_col_index": need_combo_col_index,
            "equipment_list": equipment_list,
            "gpo": gpo,
            "probe_stage2_root": os.path.join(batch_root, "_p5_p2_probe"),
            "p6_nested_probe_parent": os.path.join(batch_root, "_p6_nested_probes"),
        }
        summary_rows: list[dict] = []
        _pat_var_it = _iter_dispatch_trial_pattern_variant_queues(
            tq_frozen, pattern_jobs, p5_bundle=p5_bundle_batch
        )
        while True:
            t_pat_wall0 = time_module.perf_counter()
            try:
                pid, pname, tq, df_p5_ov = next(_pat_var_it)
            except StopIteration:
                break
            df_run = df_p5_ov.copy() if df_p5_ov is not None else df0.copy()
            _apply_pattern_dispatch_trial_orders_to_tasks_df(df_run, tq)
            out_sub = os.path.join(batch_root, pid)
            try:
                os.makedirs(out_sub, exist_ok=True)
            except OSError as e:
                summary_rows.append(
                    {
                        "パターンID": pid,
                        "パターン名": pname,
                        "備考": f"出力フォルダ作成失敗: {e}",
                        "参考スコア(自動)": "",
                        "処理時間(秒)": round(
                            time_module.perf_counter() - t_pat_wall0, 2
                        ),
                    }
                )
                continue

            row: dict = {
                "パターンID": pid,
                "パターン名": pname,
                "備考": "",
                "参考スコア(自動)": "",
            }
            paths = None
            try:
                paths = _generate_plan_impl(
                    tasks_df_override=df_run,
                    stage2_output_root=out_sub,
                    skip_remove_prior_stage2_workbooks=True,
                    return_output_paths=True,
                    tasks_df_raw_input_baseline=(df0 if df_p5_ov is not None else None),
                    result_pattern_shift_label=(pid if df_p5_ov is not None else None),
                )
            except PlanningValidationError as e:
                row["備考"] = f"検証エラー: {e}"[:500]
                row["処理時間(秒)"] = round(
                    time_module.perf_counter() - t_pat_wall0, 2
                )
                summary_rows.append(row)
                continue
            except Exception as e:
                logging.exception("パターン別段階2: %s で例外", pid)
                row["備考"] = f"エラー: {e}"[:500]
                row["処理時間(秒)"] = round(
                    time_module.perf_counter() - t_pat_wall0, 2
                )
                summary_rows.append(row)
                continue

            if not paths:
                row["備考"] = "段階2が結果パスを返しませんでした（中断の可能性）。"
                row["処理時間(秒)"] = round(
                    time_module.perf_counter() - t_pat_wall0, 2
                )
                summary_rows.append(row)
                continue

            row["_path_plan"] = paths.get("production_plan") or ""
            row["_path_member"] = paths.get("member_schedule") or ""
            sco = _score_dispatch_pattern_stage2_workbook(paths["production_plan"])
            row["納期_判定対象件数"] = sco.get("納期_判定対象件数", "")
            row["納期_遅れ件数"] = sco.get("納期_遅れ件数", "")
            row["納期_遵守率"] = sco.get("納期_遵守率", "")
            row["メンバー_平均作業割合_pct"] = sco.get("メンバー_平均作業割合_pct", "")
            row["設備_稼働セル数"] = sco.get("設備_稼働セル数", "")
            ref_s = _dispatch_pattern_reference_score_from_metrics(
                row.get("納期_遵守率"),
                row.get("メンバー_平均作業割合_pct"),
                row.get("設備_稼働セル数"),
            )
            row["参考スコア(自動)"] = ref_s if ref_s is not None else ""
            if sco.get("スコア備考"):
                row["備考"] = str(sco["スコア備考"])[:500]
            row["処理時間(秒)"] = round(
                time_module.perf_counter() - t_pat_wall0, 2
            )
            summary_rows.append(row)

        _sum_pat_sec = 0.0
        for _sr in summary_rows:
            v = _sr.get("処理時間(秒)")
            if isinstance(v, (int, float)):
                _sum_pat_sec += float(v)
        try:
            _openpyxl_write_dispatch_pattern_stage2_summary_sheet(
                wb,
                summary_rows,
                batch_root=os.path.abspath(batch_root),
                total_batch_seconds=_sum_pat_sec,
            )
            wb.save(path)
        except Exception as e:
            logging.exception("パターン別段階2: サマリシートまたは保存に失敗: %s", e)
            return False

        logging.info(
            "パターン別段階2: 完了（%s パターン・合計約 %.2f 秒）。サマリシート「%s」",
            len(summary_rows),
            _sum_pat_sec,
            DISPATCH_PATTERN_STAGE2_SUMMARY_SHEET_NAME,
        )
        return True
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
def apply_dispatch_pattern_stage2_selection_to_plan_via_xlwings(
    workbook_path: str | None = None,
    *,
    apply_post_load_mutations: bool = True,
    chosen_pattern_id: str | None = None,
) -> bool:
    """
    サマリシート「配台試行順_パターン別段階2」の B3（採用パターンID）と B2（バッチ出力ルート）を読み、
    当該バッチの ``pattern_jobs_meta.json`` に基づき選んだパターンの配台試行順を
    「配台計画_タスク入力」に書き戻し、試行順昇順で行を並べ替える。

    chosen_pattern_id を渡したときは B3 より優先（CLI 用）。
    """
    path = (workbook_path or "").strip() or _excel_plan_input_wb().strip()
    if not path:
        logging.error("パターン採用反映: ブックパスは空です。")
        return False
    try:
        import xlwings as xw
    except ImportError:
        logging.error("パターン採用反映: Excel 連携に必要なパッケージがありません。")
        return False
    try:
        wb = xw.Book(path)
        ws = wb.sheets[PLAN_INPUT_SHEET_NAME]
        ws_sum = wb.sheets[DISPATCH_PATTERN_STAGE2_SUMMARY_SHEET_NAME]
    except Exception as e:
        logging.error("パターン採用反映: シート接続に失敗: %s", e)
        return False

    batch_root = str(ws_sum.range((2, 2)).value or "").strip()
    if not batch_root or not os.path.isdir(batch_root):
        logging.error(
            "パターン採用反映: サマリ B2 のバッチ出力ルートが無効です（先にパターン別段階2を実行してください）。"
        )
        return False
    meta_path = os.path.join(batch_root, DISPATCH_PATTERN_STAGE2_META_FILENAME)
    if not os.path.isfile(meta_path):
        logging.error("パターン採用反映: メタファイルがありません: %s", meta_path)
        return False
    try:
        with open(meta_path, encoding="utf-8") as f:
            meta = json.load(f)
    except OSError as e:
        logging.error("パターン採用反映: メタ JSON の読込に失敗: %s", e)
        return False

    chosen = (chosen_pattern_id or "").strip()
    if not chosen:
        try:
            chosen = str(ws_sum.range((3, 2)).value or "").strip()
        except Exception:
            chosen = ""
    if not chosen:
        logging.error(
            "パターン採用反映: 採用パターンIDが空です。サマリの B3 に一覧のいずれかを入力してください。"
        )
        return False

    patterns = meta.get("patterns") or []
    ent = None
    chosen_key = chosen.strip().casefold()
    for p in patterns:
        pid = str(p.get("id") or "").strip()
        if pid.casefold() == chosen_key:
            ent = p
            break
    if ent is None:
        logging.error(
            "パターン採用反映: パターンID「%s」は当該バッチのメタにありません。",
            chosen,
        )
        return False

    job = _pattern_job_tuple_from_meta_entry(ent)
    mat = _openpyxl_sheet_to_matrix(ws)
    df = _matrix_to_dataframe_header_first(mat)
    if df is None or df.empty:
        logging.warning("パターン採用反映: データ行はありません。")
        return False

    df = df.copy()
    df.columns = df.columns.str.strip()
    df = _align_dataframe_headers_to_canonical(df, plan_input_sheet_column_order())
    for c in plan_input_sheet_column_order():
        if c not in df.columns:
            df[c] = ""

    df.insert(0, _PLAN_INPUT_XLWINGS_ORIG_ROW, range(len(df)))

    if apply_post_load_mutations and not _plan_input_dispatch_trial_order_local_only_from_env():
        _apply_planning_sheet_post_load_mutations(
            df,
            path,
            "パターン採用反映",
            apply_exclude_rules_from_config=False,
            compile_exclude_rules_d_to_e_with_ai=False,
        )

    dto_col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    if dto_col not in df.columns:
        logging.error("パターン採用反映: 列「%s」はありません。", dto_col)
        return False
    _dto_loc = df.columns.get_loc(dto_col)
    if isinstance(_dto_loc, slice):
        logging.error("パターン採用反映: 列「%s」は複数あります。", dto_col)
        return False
    if pd.api.types.is_numeric_dtype(df[dto_col]):
        df[dto_col] = float("nan")
    else:
        df[dto_col] = ""

    data_extract_dt, _ = _extract_data_extraction_datetime()
    base_now_dt = data_extract_dt if data_extract_dt is not None else datetime.now()
    run_date = base_now_dt.date()

    try:
        (
            _sd,
            _mem,
            equipment_list,
            req_map,
            need_rules,
            _sm,
            need_combo_col_index,
        ) = load_skills_and_needs()
    except Exception as e:
        logging.exception("パターン採用反映: master 読込に失敗: %s", e)
        return False

    global_priority_raw = load_main_sheet_global_priority_override_text()
    members_for_gpo: list = []
    try:
        with pd.ExcelFile(_master_workbook_path_resolved()) as _xf:
            _skills = pd.read_excel(_xf, sheet_name="skills", header=None)
        for r in range(2, _skills.shape[0]):
            cell = _skills.iat[r, 0]
            if pd.isna(cell):
                continue
            name = str(cell).strip()
            if name and name.lower() not in ("nan", "none", "null"):
                members_for_gpo.append(name)
    except Exception:
        members_for_gpo = []
    gpo = analyze_global_priority_override_comment(
        global_priority_raw, members_for_gpo, run_date.year, ai_sheet_sink={}
    )
    ai_by_tid = analyze_task_special_remarks(df, reference_year=run_date.year)
    tq_template = build_task_queue_from_planning_df(
        df, run_date, req_map, ai_by_tid, gpo, equipment_list
    )
    if not tq_template:
        logging.error("パターン採用反映: 配台対象タスクがありません。")
        return False

    tq_frozen = copy.deepcopy(tq_template)
    p5_bundle_sel = {
        "planning_df": df,
        "run_date": run_date,
        "req_map": req_map,
        "need_rules": need_rules,
        "need_combo_col_index": need_combo_col_index,
        "equipment_list": equipment_list,
        "gpo": gpo,
        "probe_stage2_root": os.path.join(batch_root, "_p5_selection_probe"),
        "p6_nested_probe_parent": os.path.join(batch_root, "_p6_selection_nested"),
    }
    _pid_applied, _pname_applied, tq_sel, df_p5_ov = next(
        _iter_dispatch_trial_pattern_variant_queues(tq_frozen, [job], p5_bundle=p5_bundle_sel)
    )
    df_apply = df_p5_ov.copy() if df_p5_ov is not None else df
    _apply_pattern_dispatch_trial_orders_to_tasks_df(df_apply, tq_sel)
    df_sorted = _sort_stage1_plan_df_by_dispatch_trial_order_asc(df_apply)
    orig_list = [int(x) for x in df_sorted[_PLAN_INPUT_XLWINGS_ORIG_ROW].tolist()]
    df_sorted = df_sorted.drop(columns=[_PLAN_INPUT_XLWINGS_ORIG_ROW])

    header_row = mat[0] if mat else []
    n_hdr = len(header_row)
    if n_hdr == 0:
        return False

    def _pad_row(r, n):
        r = list(r) if r is not None else []
        if len(r) < n:
            r = r + [None] * (n - len(r))
        return r

    new_mat = [_pad_row(header_row, n_hdr)]
    for i in range(len(df_sorted)):
        orig = orig_list[i]
        src_row = mat[orig + 1] if orig + 1 < len(mat) else []
        src_row = _pad_row(src_row, n_hdr)
        out_row = []
        for j in range(n_hdr):
            h_cell = header_row[j]
            if h_cell is None or (isinstance(h_cell, float) and pd.isna(h_cell)):
                hname = ""
            else:
                hname = str(h_cell).strip()
            if hname and hname in df_sorted.columns:
                v = df_sorted.iat[i, df_sorted.columns.get_loc(hname)]
                if pd.isna(v):
                    out_row.append(None)
                else:
                    out_row.append(v)
            else:
                out_row.append(src_row[j])
        new_mat.append(out_row)

    try:
        n_r = len(new_mat)
        ws.range((1, 1)).resize(n_r, n_hdr).value = new_mat
    except Exception as e:
        logging.exception("パターン採用反映: シート書込に失敗: %s", e)
        return False

    try:
        wb.save()
    except Exception as e:
        logging.warning("パターン採用反映: Save 警告: %s", e)

    logging.info(
        "パターン採用反映: パターン「%s」を「%s」に書き込みました。",
        chosen,
        PLAN_INPUT_SHEET_NAME,
    )
    return True
def _reorder_task_queue_slit_sec_consecutive(task_queue: list) -> None:
    """
    特別ルール L10（スリット→SEC）: 同一依頼NO内でスリット行の直後に SEC 行が来るよう、
    task_queue（自動算出時の並び）を軽く並べ替える。

    - 対象: 加工内容トークンに「スリット」「SEC」があり、かつその順序がスリット→SEC の依頼
    - 判定: 工程名×機械名で、スリット=スリット機1 湖南、SEC=SEC機 湖南 を採用
    - シートで配台試行順が全行指定されている場合は本関数は呼ばれない（_apply_dispatch_trial_order...側）
    """
    if not task_queue:
        return
    slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
    slit_mach = _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
    sec_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    sec_mach = _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE)

    idx_by_tid: dict[str, dict[str, int]] = {}
    ok_tid: set[str] = set()
    for i, t in enumerate(task_queue):
        tid = str(t.get("task_id") or "").strip()
        if not tid:
            continue
        toks = t.get("process_content_tokens") or []
        norm = [_normalize_process_name_for_rule_match(x) for x in toks]
        if slit_proc in norm and sec_proc in norm:
            try:
                if norm.index(slit_proc) < norm.index(sec_proc):
                    ok_tid.add(tid)
            except Exception:
                pass
        proc = _normalize_process_name_for_rule_match(t.get("machine"))
        mach = _normalize_equipment_match_key(t.get("machine_name"))
        if proc == slit_proc and mach == slit_mach:
            idx_by_tid.setdefault(tid, {})["slit"] = i
        elif proc == sec_proc and mach == sec_mach:
            idx_by_tid.setdefault(tid, {})["sec"] = i

    moved: list[str] = []
    # 依頼NOごとに 1 回だけ調整（インデックスが動くので都度再探索する）
    for tid in sorted(ok_tid):
        pos = idx_by_tid.get(tid) or {}
        if "slit" not in pos or "sec" not in pos:
            continue
        # 現在位置を再探索（前の移動でズレるため）
        slit_i = None
        sec_i = None
        for i, t in enumerate(task_queue):
            if str(t.get("task_id") or "").strip() != tid:
                continue
            proc = _normalize_process_name_for_rule_match(t.get("machine"))
            mach = _normalize_equipment_match_key(t.get("machine_name"))
            if slit_i is None and proc == slit_proc and mach == slit_mach:
                slit_i = i
            if sec_i is None and proc == sec_proc and mach == sec_mach:
                sec_i = i
        if slit_i is None or sec_i is None:
            continue
        if sec_i == slit_i + 1:
            continue
        sec_task = task_queue.pop(sec_i)
        insert_at = slit_i + 1
        if sec_i < insert_at:
            insert_at -= 1
        task_queue.insert(insert_at, sec_task)
        moved.append(tid)
    if moved:
        logging.info(
            "特別ルールL10 配台試行順: スリット行の直後にSEC行を隣接した依頼NO: %s",
            ",".join(moved),
        )
def _reorder_task_queue_connection_sec_consecutive(task_queue: list) -> None:
    """
    特別ルール B-6.2（接続→SEC）: 同一依頼NO内で接続行の直後に SEC 行が来るよう
    task_queue（自動算出時の並び）を軽く並べ替える。

    - 対象: 加工内容トークンに「接続」「SEC」があり、かつその順序が接続→SEC の依頼
    - 判定: 工程名×機械名で、接続=熱融着機　湖南、SEC=SEC機　湖南
    - シートで配台試行順が全行指定されている場合は本関数は呼ばれない（_apply_dispatch_trial_order...側）
    """
    if not task_queue:
        return
    conn_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_CONNECTION_PROCESS)
    conn_mach = _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
    sec_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
    sec_mach = _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE)

    idx_by_tid: dict[str, dict[str, int]] = {}
    ok_tid: set[str] = set()
    for i, t in enumerate(task_queue):
        tid = str(t.get("task_id") or "").strip()
        if not tid:
            continue
        toks = t.get("process_content_tokens") or []
        norm = [_normalize_process_name_for_rule_match(x) for x in toks]
        if conn_proc in norm and sec_proc in norm:
            try:
                if norm.index(conn_proc) < norm.index(sec_proc):
                    ok_tid.add(tid)
            except Exception:
                pass
        proc = _normalize_process_name_for_rule_match(t.get("machine"))
        mach = _normalize_equipment_match_key(t.get("machine_name"))
        if proc == conn_proc and mach == conn_mach:
            idx_by_tid.setdefault(tid, {})["conn"] = i
        elif proc == sec_proc and mach == sec_mach:
            idx_by_tid.setdefault(tid, {})["sec"] = i

    moved: list[str] = []
    for tid in sorted(ok_tid):
        pos = idx_by_tid.get(tid) or {}
        if "conn" not in pos or "sec" not in pos:
            continue
        conn_i = None
        sec_i = None
        for i, t in enumerate(task_queue):
            if str(t.get("task_id") or "").strip() != tid:
                continue
            proc = _normalize_process_name_for_rule_match(t.get("machine"))
            mach = _normalize_equipment_match_key(t.get("machine_name"))
            if conn_i is None and proc == conn_proc and mach == conn_mach:
                conn_i = i
            if sec_i is None and proc == sec_proc and mach == sec_mach:
                sec_i = i
        if conn_i is None or sec_i is None:
            continue
        if sec_i == conn_i + 1:
            continue
        sec_task = task_queue.pop(sec_i)
        insert_at = conn_i + 1
        if sec_i < insert_at:
            insert_at -= 1
        task_queue.insert(insert_at, sec_task)
        moved.append(tid)
    if moved:
        logging.info(
            "特別ルールB-6 配台試行順: 接続行の直後にSEC行を隣接した依頼NO: %s",
            ",".join(moved),
        )
def _dispatch_trial_order_key(task: dict, default: float = 1e9) -> float:
    from planning_core.core.plan_input import dispatch_trial_order_key_from_task

    return dispatch_trial_order_key_from_task(task, default)


def _task_queue_all_have_sheet_dispatch_trial_order(task_queue: list) -> bool:
    """配台計画シートの「配台試行順番」はキュー全行に正の数（小数可）で入っているか。"""
    if not task_queue:
        return False
    from planning_core.core.plan_input import dispatch_trial_order_positive_finite

    for t in task_queue:
        v = dispatch_trial_order_positive_finite(t.get("dispatch_trial_order_from_sheet"))
        if v is None:
            return False
    return True
def _apply_dispatch_trial_order_for_generate_plan(
    task_queue: list,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
) -> None:
    """
    配台試行順の確定。シートに全行分の試行順はあれみしれを採用（§B-2/3 の隣接繰り上きは行ゝない）。
    欠損はあれみ従来どおりマスタ・紝期・need 列順などでソートし、EC 隣接後に 1..n を付与。
    """
    if _task_queue_all_have_sheet_dispatch_trial_order(task_queue):
        from planning_core.core.plan_input import dispatch_trial_order_sort_key

        task_queue.sort(
            key=lambda t: (
                dispatch_trial_order_sort_key(t.get("dispatch_trial_order_from_sheet")),
                int(t.get("planning_sheet_row_seq") or 10**9),
            )
        )
        for t in task_queue:
            t["dispatch_trial_order"] = dispatch_trial_order_sort_key(
                t.get("dispatch_trial_order_from_sheet")
            )
        _normalize_dispatch_trial_order_by_process_sequence_within_task_id(task_queue)
        logging.info(
            "配台試行順番: 「%s」列の値をしのまま使用しました（全 %s 行）。",
            RESULT_TASK_COL_DISPATCH_TRIAL_ORDER,
            len(task_queue),
        )
        return
    task_queue.sort(
        key=lambda x: _generate_plan_task_queue_sort_key(
            x, req_map, need_rules, need_combo_col_index
        )
    )
    _reorder_task_queue_b2_ec_inspection_consecutive(task_queue)
    _reorder_task_queue_slit_sec_consecutive(task_queue)
    _reorder_task_queue_connection_sec_consecutive(task_queue)
    _reorder_task_queue_process_sequence_within_task_id(task_queue)
    _assign_sequential_dispatch_trial_order(task_queue)
    _normalize_dispatch_trial_order_by_process_sequence_within_task_id(task_queue)
    logging.info(
        "配台試行順番: マスタ・タスク入力から自動計算し 1..%s を付与しました。",
        len(task_queue),
    )
def fill_plan_dispatch_trial_order_column_stage1(
    plan_df: "pd.DataFrame",
    run_date: date,
    req_map: dict,
    need_rules: list,
    need_combo_col_index: dict | None,
    equipment_list: list,
    *,
    members_for_gpo: list | None = None,
) -> None:
    """
    段階1出力 DataFrame の「配台試行順番」を」段階2 冒頭とともに手順（ソート・§B-2/3 隣接・連番）で埋ゝる。
    配台対象外の行は空のまま。

    ``members_for_gpo`` を渡したとしは、メイン「グローバルコメント」解析用のメンバー名一覧としてそれを使い、
    ``MASTER_FILE`` の skills シートを **再読込しません**（段階1で ``load_skills_and_needs`` 済みの場合の I/O 短縮）。
    """
    if plan_df is None or getattr(plan_df, "empty", True):
        return
    if RESULT_TASK_COL_DISPATCH_TRIAL_ORDER not in plan_df.columns:
        return
    col = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    global_priority_raw = load_main_sheet_global_priority_override_text()
    members_for_gpo_eff: list = []
    if members_for_gpo is not None:
        members_for_gpo_eff = list(members_for_gpo or [])
    else:
        try:
            with pd.ExcelFile(_master_workbook_path_resolved()) as _xf:
                _skills = pd.read_excel(_xf, sheet_name="skills", header=None)
            for r in range(2, _skills.shape[0]):
                cell = _skills.iat[r, 0]
                if pd.isna(cell):
                    continue
                name = str(cell).strip()
                if name and name.lower() not in ("nan", "none", "null"):
                    members_for_gpo_eff.append(name)
        except Exception:
            members_for_gpo_eff = []
    gpo = analyze_global_priority_override_comment(
        global_priority_raw, members_for_gpo_eff, run_date.year, ai_sheet_sink={}
    )
    tq = build_task_queue_from_planning_df(
        plan_df,
        run_date,
        req_map,
        None,
        gpo,
        equipment_list,
    )
    _apply_dispatch_trial_order_for_generate_plan(
        tq, req_map, need_rules, need_combo_col_index
    )
    try:
        col_idx = plan_df.columns.get_loc(col)
    except Exception:
        return
    for t in tq:
        iloc = t.get("planning_df_iloc")
        if iloc is None:
            continue
        if not isinstance(iloc, int) or iloc < 0 or iloc >= len(plan_df):
            continue
        dto = t.get("dispatch_trial_order")
        if dto is None:
            continue
        try:
            # Excel 上は数値セルにし、フィルター・並き替ごをしやれしれる（文字列てと数値と別グループになる）
            plan_df.iat[iloc, col_idx] = int(dto)
        except (TypeError, ValueError):
            if pd.api.types.is_numeric_dtype(plan_df.iloc[:, col_idx]):
                plan_df.iat[iloc, col_idx] = float("nan")
            else:
                plan_df.iat[iloc, col_idx] = ""
def _equipment_schedule_unified_sub_string_map(timeline_for_eq_grid: list) -> dict:
    """
    同一日・同一設備列キー・同一依頼NO の加工についで」設備時間割セル用の「補」表示文字列。
    タイムライン上の坄ブロックの `sub` に睾れた補助者坝を和集合し、昇順で ", " 連絝れる。
    メンバー日程・占有計算に使うタイムラインの `sub` は変更しない（表示専用）。
    """
    acc: dict = defaultdict(set)
    for e in timeline_for_eq_grid or []:
        if not _is_machining_timeline_event(e):
            continue
        tid = str(e.get("task_id") or "").strip()
        m = str(e.get("machine") or "").strip()
        d0 = e.get("date")
        if not tid or not m or d0 is None:
            continue
        for s in str(e.get("sub") or "").split(","):
            t = s.strip()
            if t:
                acc[(d0, m, tid)].add(t)
    return {k: ", ".join(sorted(v)) for k, v in acc.items() if v}
def _eq_grid_slot_overlaps_event(
    curr_grid: datetime, next_grid: datetime, ev: dict
) -> bool:
    """10分枠 [curr_grid, next_grid) とイベント [start_dt, end_dt) が重なるか。"""
    st, ed = _gantt_machining_display_range_for_slot_overlap(ev)
    return (
        isinstance(st, datetime)
        and isinstance(ed, datetime)
        and st < next_grid
        and ed > curr_grid
    )
def _eq_grid_first_overlapping_event(evs: list, curr_grid: datetime, next_grid: datetime):
    """evs は開始時刻順。枠と重なる最初のイベントを返す（短い加工が中点判定で落ちるのを防ぐ）。"""
    for ev in evs:
        if _eq_grid_slot_overlaps_event(curr_grid, next_grid, ev):
            return ev
    return None
def _eq_grid_best_overlapping_event_for_cell(
    evs: list, curr_grid: datetime, next_grid: datetime
):
    """
    10 分枠と重なるイベントのうち表示に用いる 1 件を選ぶ。
    加工（進度バー対象）が重なるときはそのうち開始が最も早い加工を優先し、
    準備・後始末だけが先に重なって加工が隠れるのを防ぐ。
    """
    hits = [
        ev
        for ev in evs
        if _eq_grid_slot_overlaps_event(curr_grid, next_grid, ev)
    ]
    if not hits:
        return None
    mach_hits = [ev for ev in hits if _eq_grid_timeline_event_use_progress_bar(ev)]
    if mach_hits:
        return min(
            mach_hits,
            key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or "")),
        )
    return min(
        hits,
        key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or "")),
    )
def _eq_grid_overlap_sample_t(
    ev: dict, curr_grid: datetime, next_grid: datetime, slot_mid: datetime
) -> datetime:
    """休憩判定用: 枠とイベントの重なり区間の中点（重なりなければ枠中点）。"""
    st, ed = _gantt_machining_display_range_for_slot_overlap(ev)
    if isinstance(st, datetime) and isinstance(ed, datetime):
        os_ = max(curr_grid, st)
        oe = min(next_grid, ed)
        if os_ < oe:
            return os_ + (oe - os_) / 2
    return slot_mid
def _eq_grid_events_for_equipment_column(
    machine_to_events: dict, eq_col: str
) -> list:
    """
    equipment_list の列キーと ev['machine'] の表記ゆれ（全角空白・NBSP 等）を正規化して対応づける。
    一致しないと 10 分枠に何も出ず、結果_タスク一覧の時間割リンクも付かない。
    """
    if not eq_col or not machine_to_events:
        return []
    evs = machine_to_events.get(eq_col)
    if evs:
        return evs
    nk = _normalize_equipment_match_key(eq_col)
    if not nk:
        return []
    for mk, evs2 in machine_to_events.items():
        if _normalize_equipment_match_key(str(mk)) == nk:
            return evs2
    pe, me = _split_equipment_line_process_machine(eq_col)
    pe_n = _normalize_equipment_match_key(pe)
    me_n = _normalize_equipment_match_key(me)
    if pe_n and me_n:
        for mk, evs2 in machine_to_events.items():
            pk, mk_m = _split_equipment_line_process_machine(str(mk))
            if (
                _normalize_equipment_match_key(pk) == pe_n
                and _normalize_equipment_match_key(mk_m) == me_n
            ):
                return evs2
    return []
def _eq_grid_mcol_for_event_machine(
    eq_to_mcol: dict[str, str], event_machine: str
) -> str | None:
    """機械名集約時間割: イベント側 machine キーから表示列 mcol を正規化照合で解決。"""
    if not event_machine or not eq_to_mcol:
        return None
    mcol = eq_to_mcol.get(event_machine)
    if mcol:
        return mcol
    nk = _normalize_equipment_match_key(event_machine)
    if not nk:
        return None
    for ek, mc in eq_to_mcol.items():
        if _normalize_equipment_match_key(str(ek)) == nk:
            return mc
    return None
def _eq_grid_timeline_event_use_progress_bar(ev: dict) -> bool:
    """設備時間割の「進度R」表示・ハイパーリンク対象となる加工イベントか。"""
    return (
        _is_machining_timeline_event(ev)
        and all(
            k in ev
            for k in (
                "eff_time_per_unit",
                "units_done",
                "total_units",
                "already_done_units",
            )
        )
        and float(ev.get("eff_time_per_unit") or 0) > 0
    )
def _eq_grid_rolls_done_within_ev_segment_at(ev: dict, t_dt) -> int:
    """加工イベントセグメントの先頭から t_dt までに完了したロール数（0 .. units_done）。"""
    if not isinstance(t_dt, datetime):
        return 0
    st0 = ev.get("start_dt")
    ed0 = ev.get("end_dt")
    if not isinstance(st0, datetime) or not isinstance(ed0, datetime):
        return 0
    if t_dt <= st0:
        return 0
    eff_v = float(ev.get("eff_time_per_unit") or 0)
    if eff_v <= 0:
        return 0
    u_cap = int(float(ev.get("units_done") or 0))
    wm = get_actual_work_minutes(
        st0, min(t_dt, ed0), ev.get("breaks") or []
    )
    return min(u_cap, int(wm / eff_v))
def _build_equipment_schedule_dataframe(
    sorted_dates: list,
    equipment_list: list,
    attendance_data: dict,
    timeline_events: list,
    *,
    first_eq_schedule_cell_by_task_id: dict | None = None,
) -> "pd.DataFrame":
    """
    結果_設備毎の時間割と同形式の DataFrame（10 分枠・設備列＋進度列）。
    first_eq_schedule_cell_by_task_id を渡したとしのみ」初出セル座標を記録（結果ポイパーリンク用）。
    """
    timeline_for_eq_grid = _expand_timeline_events_for_equipment_grid(timeline_events)
    _eq_sched_unify_sub = _equipment_schedule_unified_sub_string_map(timeline_for_eq_grid)
    events_by_date = defaultdict(list)
    for e in timeline_for_eq_grid:
        events_by_date[e["date"]].append(e)

    all_eq_rows = []
    eq_empty_cols = {}
    for eq in equipment_list:
        eq_empty_cols[eq] = ""
        eq_empty_cols[f"{eq}進度"] = ""

    for d in sorted_dates:
        d_start = datetime.combine(d, DEFAULT_START_TIME)
        d_end = datetime.combine(d, DEFAULT_END_TIME)
        events_today = events_by_date[d]
        machine_to_events = defaultdict(list)
        for ev in events_today:
            machine_to_events[ev["machine"]].append(ev)
        for _eq_k, _evs in machine_to_events.items():
            _evs.sort(
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
            )

        is_anyone_working = any(
            daily_status["is_working"] for daily_status in attendance_data[d].values()
        )
        if not events_today and not is_anyone_working:
            continue

        all_eq_rows.append({"日時帯": f"■ {d.strftime('%Y/%m/%d (%a)')} ■", **eq_empty_cols})

        def _eq_cell_display_sub(ev, day_d) -> str:
            tid0 = str(ev.get("task_id") or "").strip()
            m0 = str(ev.get("machine") or "").strip()
            if tid0 and m0:
                u0 = _eq_sched_unify_sub.get((day_d, m0, tid0))
                if u0 is not None:
                    return u0
            return str(ev.get("sub") or "").strip()

        curr_grid = d_start
        while curr_grid < d_end:
            next_grid = curr_grid + timedelta(minutes=10)
            if next_grid > d_end:
                next_grid = d_end

            mid_t = curr_grid + (next_grid - curr_grid) / 2
            row_data = {
                "日時帯": f"{curr_grid.strftime('%H:%M')}-{next_grid.strftime('%H:%M')}"
            }

            for eq in equipment_list:
                eq_text = ""
                progress_text = ""
                active_ev = _eq_grid_best_overlapping_event_for_cell(
                    _eq_grid_events_for_equipment_column(machine_to_events, eq),
                    curr_grid,
                    next_grid,
                )

                if active_ev:
                    _sample_t = _eq_grid_overlap_sample_t(
                        active_ev, curr_grid, next_grid, mid_t
                    )
                    _use_prog = _eq_grid_timeline_event_use_progress_bar(active_ev)
                    if any(
                        b_s <= _sample_t < b_e for b_s, b_e in active_ev["breaks"]
                    ):
                        eq_text = "休憩"
                    elif not _use_prog:
                        _ek_disp = _timeline_event_kind(active_ev)
                        _tag = {
                            TIMELINE_EVENT_MACHINE_DAILY_STARTUP: "日次始業準備",
                            TIMELINE_EVENT_REQUEST_SWITCH_PREP: "依頼切替準備",
                            TIMELINE_EVENT_BREAK_RESUME_PREP: "休憩再開準備",
                            TIMELINE_EVENT_POST_MACHINING_CLEANUP: "後始末",
                            TIMELINE_EVENT_REQUEST_INTERVAL_BUFFER: "依頼間余裕",
                        }.get(
                            _ek_disp,
                            "セットアップ",
                        )
                        _sub_n = _eq_cell_display_sub(active_ev, d)
                        _sub_text = f" 補:{_sub_n}" if _sub_n else ""
                        _tid_d = str(active_ev.get("task_id") or "").strip()
                        # 日次始業準備は括弧なし（設備ガントのメインシェイプ文言と整合）
                        if _ek_disp in (
                            TIMELINE_EVENT_MACHINE_DAILY_STARTUP,
                            TIMELINE_EVENT_REQUEST_SWITCH_PREP,
                            TIMELINE_EVENT_BREAK_RESUME_PREP,
                            TIMELINE_EVENT_POST_MACHINING_CLEANUP,
                            TIMELINE_EVENT_REQUEST_INTERVAL_BUFFER,
                        ):
                            eq_text = str(_tag)
                        else:
                            eq_text = (
                                f"[{_tid_d}] 主:{active_ev.get('op', '')}{_sub_text} ({_tag})"
                            )
                        progress_text = ""
                    else:
                        _slice_a = max(curr_grid, active_ev["start_dt"])
                        _slice_b = min(next_grid, active_ev["end_dt"])
                        total_u = int(float(active_ev.get("total_units") or 0))
                        _base_done = int(
                            float(active_ev.get("already_done_units") or 0)
                        )
                        if total_u <= 0:
                            progress_text = ""
                        elif (
                            _slice_a < _slice_b
                            and isinstance(active_ev.get("start_dt"), datetime)
                            and isinstance(active_ev.get("end_dt"), datetime)
                        ):
                            _rd_lo = _eq_grid_rolls_done_within_ev_segment_at(
                                active_ev, _slice_a
                            )
                            _rd_hi = _eq_grid_rolls_done_within_ev_segment_at(
                                active_ev, _slice_b
                            )
                            _cum_lo = int(min(total_u, _base_done + _rd_lo))
                            _cum_hi = int(min(total_u, _base_done + _rd_hi))
                            if _cum_hi > _cum_lo:
                                progress_text = "・".join(
                                    f"{k}/{total_u}R"
                                    for k in range(_cum_lo + 1, _cum_hi + 1)
                                )
                            else:
                                progress_text = f"{_cum_hi}/{total_u}R"
                        else:
                            progress_text = ""

                        _sub_s = _eq_cell_display_sub(active_ev, d)
                        sub_text = f" 補:{_sub_s}" if _sub_s else ""
                        eq_text = f"[{active_ev['task_id']}] 主:{active_ev['op']}{sub_text}"

                # 表示は「枠内で最も早く始まるイベント」1件だが、準備・セットアップが先にあると
                # 加工が active_ev にならずタスクID→時間割リンクが欠ける。重なる加工イベントを別途走査する。
                if first_eq_schedule_cell_by_task_id is not None:
                    for _hev in _eq_grid_events_for_equipment_column(
                        machine_to_events, eq
                    ):
                        if not _eq_grid_slot_overlaps_event(
                            curr_grid, next_grid, _hev
                        ):
                            continue
                        if not _eq_grid_timeline_event_use_progress_bar(_hev):
                            continue
                        _hs = _eq_grid_overlap_sample_t(
                            _hev, curr_grid, next_grid, mid_t
                        )
                        if any(
                            b_s <= _hs < b_e for b_s, b_e in _hev["breaks"]
                        ):
                            continue
                        _htid = str(_hev.get("task_id") or "").strip()
                        if not _htid or _htid in first_eq_schedule_cell_by_task_id:
                            continue
                        _row_ex = len(all_eq_rows) + 2
                        _ci = 2 + 2 * equipment_list.index(eq)
                        first_eq_schedule_cell_by_task_id[_htid] = (
                            f"{get_column_letter(_ci)}{_row_ex}"
                        )

                row_data[eq] = eq_text
                row_data[f"{eq}進度"] = progress_text

            all_eq_rows.append(row_data)
            curr_grid = next_grid
        all_eq_rows.append({"日時帯": "", **eq_empty_cols})

    df_eq = pd.DataFrame(all_eq_rows)
    _eq_hdr = _equipment_schedule_header_labels(equipment_list)
    _eq_rename = {}
    for _eq, _lab in zip(equipment_list, _eq_hdr):
        if _eq in df_eq.columns:
            _eq_rename[_eq] = _lab
        _pqc = f"{_eq}進度"
        if _pqc in df_eq.columns:
            _eq_rename[_pqc] = f"{_lab}進度"
    if _eq_rename:
        df_eq = df_eq.rename(columns=_eq_rename)
    return df_eq
def _machine_display_key_for_equipment(eq: str) -> str:
    """skills 列キー「工程+機械」から機械名表示キーを得る（重複時は複坈キーごとに別列）。"""
    s = str(eq).strip()
    if "+" in s:
        mpart = s.split("+", 1)[1].strip()
        return mpart if mpart else s
    return s
def _build_equipment_schedule_by_machine_name_dataframe(
    sorted_dates: list,
    equipment_list: list,
    attendance_data: dict,
    timeline_events: list,
) -> "pd.DataFrame":
    """
    機械名短縮に列をまとめ」坄 10 分枠で占有中の依頼NO（複数時は「＝」）を表示れる。
    列見出しは機械名のみ（工程+機械の複坈キーは付けない）。同一実機械は占有キーで1列に集約れる。
    """
    timeline_for_eq_grid = _expand_timeline_events_for_equipment_grid(timeline_events)
    events_by_date = defaultdict(list)
    for e in timeline_for_eq_grid:
        events_by_date[e["date"]].append(e)

    # 占有キー（機械名ベース・正規化）ごとに1列。見出しは equipment_list 初出の機械名表示のみ。
    occ_key_to_header: dict[str, str] = {}
    machine_cols: list[str] = []
    eq_to_mcol: dict[str, str] = {}
    for eq in equipment_list:
        occ_key = _equipment_line_key_to_physical_occupancy_key(eq)
        if not occ_key:
            occ_key = _normalize_equipment_match_key(str(eq).strip())
        disp = _machine_display_key_for_equipment(eq).strip() or str(eq).strip()
        if occ_key not in occ_key_to_header:
            occ_key_to_header[occ_key] = disp
            machine_cols.append(disp)
        eq_to_mcol[eq] = occ_key_to_header[occ_key]

    empty_tail = {mcol: "" for mcol in machine_cols}
    all_rows = []

    for d in sorted_dates:
        d_start = datetime.combine(d, DEFAULT_START_TIME)
        d_end = datetime.combine(d, DEFAULT_END_TIME)
        events_today = events_by_date[d]
        machine_to_events = defaultdict(list)
        for ev in events_today:
            machine_to_events[ev["machine"]].append(ev)
        for _eq_k, _evs in machine_to_events.items():
            _evs.sort(
                key=lambda e: (e.get("start_dt") or datetime.min, str(e.get("task_id") or ""))
            )

        is_anyone_working = any(
            daily_status["is_working"] for daily_status in attendance_data[d].values()
        )
        if not events_today and not is_anyone_working:
            continue

        all_rows.append({"日時帯": f"■ {d.strftime('%Y/%m/%d (%a)')} ■", **empty_tail})

        curr_grid = d_start
        while curr_grid < d_end:
            next_grid = curr_grid + timedelta(minutes=10)
            if next_grid > d_end:
                next_grid = d_end
            mid_t = curr_grid + (next_grid - curr_grid) / 2
            row_data = {
                "日時帯": f"{curr_grid.strftime('%H:%M')}-{next_grid.strftime('%H:%M')}"
            }
            for mcol in machine_cols:
                row_data[mcol] = ""
            tids_by_mcol: dict[str, set[str]] = defaultdict(set)
            for eq, evs in machine_to_events.items():
                mcol = _eq_grid_mcol_for_event_machine(eq_to_mcol, str(eq))
                if not mcol:
                    continue
                active_ev = _eq_grid_best_overlapping_event_for_cell(
                    evs, curr_grid, next_grid
                )
                if not active_ev:
                    continue
                _sample_tm = _eq_grid_overlap_sample_t(
                    active_ev, curr_grid, next_grid, mid_t
                )
                if any(
                    b_s <= _sample_tm < b_e for b_s, b_e in active_ev["breaks"]
                ):
                    tids_by_mcol[mcol].add("（休憩）")
                else:
                    tid = str(active_ev.get("task_id") or "").strip()
                    if tid:
                        tids_by_mcol[mcol].add(tid)
            for mcol in machine_cols:
                parts = sorted(tids_by_mcol.get(mcol, ()))
                row_data[mcol] = "＝".join(parts) if parts else ""
            all_rows.append(row_data)
            curr_grid = next_grid
        all_rows.append({"日時帯": "", **empty_tail})

    return pd.DataFrame(all_rows)
def _day_schedule_task_sort_key(
    task: dict,
    _task_queue: list | None = None,
    need_combo_col_index: dict | None = None,
):
    """
    同一日内の割付試行順（STAGE2_DISPATCH_FLOW_TRIAL_ORDER_FIRST=0 の主ループ用）。
    先頭キーは _generate_plan_task_queue_sort_key と同じ趣旨（加工途中・紝期基準 due_basis_date・§B 段・b2_queue_sub・need 列順・依頼NO）。
    続けて §B-1 の配台試行順繰り上き」工程 rank」dispatch_trial_order」§B-2 段内 EC 先行」優先度」結果用キー。
    同一実機械上の隙間割り込みは _equipment_line_lower_dispatch_trial_still_pending で試行順を強制れる。
    STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT=1 のときは _task_blocked_by_global_dispatch_trial_order は
    より尝さい試行順の未完了を跨いて割り込みを別途ブロックれる。
    """
    raw_r = task.get("process_sequence_rank")
    if raw_r is None:
        r = 10**9
    else:
        r = int(raw_r)
    try:
        line_seq = int(task.get("same_request_line_seq", 0))
    except (TypeError, ValueError):
        line_seq = 0
    try:
        dto = _dispatch_trial_order_key(task)
    except (TypeError, ValueError):
        dto = 10**9
    insp = bool(task.get("roll_pipeline_inspection"))
    rw = bool(task.get("roll_pipeline_rewind"))
    ip = bool(task.get("in_progress"))
    ec = bool(task.get("roll_pipeline_ec"))
    if insp and ip:
        b_tier = 0
    elif ec or (insp and not ip) or (rw and not ip):
        b_tier = 1
    else:
        b_tier = 2
    if b_tier == 1:
        if ec:
            b2_queue_sub = 0
        elif (insp and not ip) or (rw and not ip):
            b2_queue_sub = 1
        else:
            b2_queue_sub = 2
    else:
        b2_queue_sub = 0
    if ec:
        b2_roll_pipeline_stage = 0
    elif (insp and not ip) or (rw and not ip):
        b2_roll_pipeline_stage = 1
    else:
        b2_roll_pipeline_stage = 2
    dbk = task.get("due_basis_date")
    if not isinstance(dbk, date):
        dbk = date.max
    need_rank = _need_sheet_pm_column_rank(
        task.get("machine"), task.get("machine_name"), need_combo_col_index
    )
    tb = _task_id_same_machine_due_tiebreak_key(task.get("task_id"))
    b1_trial_early = (0, dto) if (insp and ip) else (1, 0)
    return (
        (
            0 if ip else 1,
            dbk,
            b_tier,
            b2_queue_sub,
            need_rank,
            tb,
            b1_trial_early,
            r,
            line_seq,
            dto,
            b2_roll_pipeline_stage,
            _plan_sheet_priority_sort_value(task),
        )
        + _result_task_sheet_sort_key(task)
    )
def _equipment_line_lower_dispatch_trial_still_pending(
    task_queue: list,
    machine_occ_key: str,
    my_dispatch_order: int,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    assign_probe_ctx: dict | None = None,
    pending_by_occ: dict[str, list[tuple[int, dict]]] | None = None,
    window_left_cache: dict | None = None,
) -> bool:
    """
    同一実機械（machine 占有キー）上で」より尝さい配台試行順の行はまて残量を挝つか。
    machine_avail_dt はポャンク間の隙間に後続試行順は入り込ゝるため、ここで順庝を強制れる。
    設備を跨いて試行順の剝後は _task_blocked_by_global_dispatch_trial_order で別途制御れる。

    キュー先頭に残量はあるの値ではブロックしない。tasks_today と同様に
    start_date_req <= current_date の行の値を「先試行順の競坈」とみなす。
    （まて開始日に靔していない行は全日ブロッカーになり」後続はろれ配台試行になるのを防し。）

    より尝さい試行順の行は **同一依頼の剝工程待う等でまて割付試行**なとしは「競坈の残」とみなさない。
    （当該行は eligible にも入らないため、ここで待たせると後続試行順は同一設備で永久坜止し得る。）

    より尝さい試行順の行は **当日の機械カレンダーの値で計画窓を全日占有**（しの設備は当日スロットゼロ）なら
    「競坈の残」とみなさない（グローバル試行順とあゝせで他設備は全日止まるのを防し）。
    """
    if _interactive_trial_calendar_legacy_active():
        # インタラクティブ試行: JSON の行順を優先し、同一設備上で他依頼の低試行順保留で抑止しない。
        return False
    line = (machine_occ_key or "").strip()
    if not line:
        return False
    try:
        my_o = int(my_dispatch_order)
    except (TypeError, ValueError):
        my_o = 10**9

    def _lower_order_blocks(t: dict, o: int) -> bool:
        if o >= my_o:
            return False
        if float(t.get("remaining_units") or 0) <= 1e-12:
            return False
        if _task_not_yet_schedulable_due_to_dependency_or_b2_room(t, task_queue):
            return False
        if _task_fully_machine_calendar_blocked_on_date(
            t, current_date, daily_status, members
        ):
            return False
        if _task_no_machining_window_left_from_avail_floor_cached(
            t,
            current_date,
            daily_status,
            members,
            machine_avail_dt,
            machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
            window_left_cache=window_left_cache,
        ):
            return False
        if assign_probe_ctx is not None and _trial_order_assign_probe_fails(
            t, current_date, daily_status, assign_probe_ctx
        ):
            return False
        return True

    if pending_by_occ is not None:
        for o, t in pending_by_occ.get(line, ()):
            if o >= my_o:
                break
            if _lower_order_blocks(t, o):
                return True
        return False
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        _sdr = t.get("start_date_req")
        if not isinstance(_sdr, date) or _sdr > current_date:
            continue
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        t_occ = _machine_occupancy_key_resolve(t, _eqt)
        if t_occ != line:
            continue
        try:
            o = _dispatch_trial_order_key(t)
        except (TypeError, ValueError):
            o = 10**9
        if _lower_order_blocks(t, o):
            return True
    return False
def _min_pending_dispatch_trial_order_for_date(
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> int | None:
    """
    start_date_req <= current_date かつ残量ありのタスクの配台試行順の最尝値。
    _equipment_line_lower_dispatch_trial_still_pending と同様」まて開始日に靔していない行は
    「先行試行順の競坈」に含まない。

    **グローバル試行順ブロック**（STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT）用に」
    「この日まて割付候補になり得ない」行は最尝値から除外れる。さもないと同一依頼の
    §A-1/§A-2 剝工程（試行順は後ゝては行順は先）は必須な行は」より尝さい試行順の行と
    循環して永久に動けない。
    - `_task_not_yet_schedulable_due_to_dependency_or_b2_room` は True の行
    - （daily_status・members は渡るとし）当日機械カレンダーの値で計画窓全日占有の行
    - （machine_avail_dt 等は渡るとし）設備タイムラインは計画終端以上で当日スロットなしの行

    1 ロール割当プローブによる除外は行ゝない（`_effective_min_dispatch_trial_order_from_pool` 坴で層ごとに判定）。
    """
    pool = _tasks_in_min_pending_dispatch_pool(
        task_queue,
        current_date,
        daily_status=daily_status,
        members=members,
        machine_avail_dt=machine_avail_dt,
        machine_day_start=machine_day_start,
        machine_handoff=machine_handoff,
        skills_dict=skills_dict,
        abolish_all_scheduling_limits=abolish_all_scheduling_limits,
        dispatch_interval_mirror=dispatch_interval_mirror,
    )
    orders: list[int] = []
    for t in pool:
        try:
            orders.append(_dispatch_trial_order_key(t))
        except (TypeError, ValueError):
            orders.append(10**9)
    return min(orders) if orders else None
def _task_blocked_by_global_dispatch_trial_order(
    task: dict,
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    min_dispatch_effective: int | None = None,
) -> bool:
    """
    より尝さい配台試行順に」当日割付可能な未完了はあるとし」当該タスクをブロックれる。
    min_dispatch_effective: プール＋プローブで求ゝた実効最尝試行順（未指定時は安価フィルタのみの最尝）。
    """
    if not STAGE2_GLOBAL_DISPATCH_TRIAL_ORDER_STRICT:
        return False
    # インタラクティブ配台試行: 入力 JSON の行ごとの配台試行順を正とする。他依頼NOの dto=1 が
    # プールに残るだけで V5-4 dto=2 が全日 eligible から落ちるのを防ぐ（グローバル最尝試行順は使わない）。
    if _interactive_trial_calendar_legacy_active():
        return False
    if min_dispatch_effective is not None:
        m = min_dispatch_effective
    else:
        m = _min_pending_dispatch_trial_order_for_date(
            task_queue,
            current_date,
            daily_status=daily_status,
            members=members,
            machine_avail_dt=machine_avail_dt,
            machine_day_start=machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
    if m is None:
        return False
    try:
        my_o = _dispatch_trial_order_key(task)
    except (TypeError, ValueError):
        my_o = 10**9
    return my_o > m
def _purge_attendance_days_not_in_set(attendance_data: dict, keep_dates: frozenset) -> None:
    """勤怠辞書からマスタに無い日付キーを削除する（自動拡張分の巻し戻し）。"""
    for dk in list(attendance_data.keys()):
        if dk not in keep_dates:
            del attendance_data[dk]
def _partial_task_id_due_shift_outcome(
    task_queue: list, task_id: str, calendar_last: date
) -> tuple[bool, bool]:
    """
    配台残の依頼NOについで紝期+1日リトライの分類。
    戻り値: (shift_ok, calendar_shortfall)
    - shift_ok: 紝期基準（due_basis_date）を挝つ行はあり」しれらまとめてで +1 日はマスタ最終計画日以下
    - calendar_shortfall: 紝期基準を挝つ行はあり」いうれかで +1 日はマスタ最終計画日を超ごる
    基準紝期は一行も無い依頼は (False, False)（通常の配台残のまま）。
    """
    tid = (task_id or "").strip()
    if not tid:
        return False, False
    rows = [t for t in task_queue if str(t.get("task_id", "") or "").strip() == tid]
    if not rows:
        return False, False
    basis_rows = [t for t in rows if t.get("due_basis_date") is not None]
    if not basis_rows:
        return False, False
    for t in basis_rows:
        db = t["due_basis_date"]
        if db + timedelta(days=1) > calendar_last:
            return False, True
    return True, False
def _shift_task_due_calendar_fields_one_day(task: dict, run_date: date) -> None:
    """
    配台残リトライ用: **内部の紝期基準（due_basis_date）の値**を +1 日れる。
    結果_タスク一覧用の ``due_basis_date_result_sheet`` は変更しない（+1 剝の日付を保挝）。
    回答納期・指定納期も配台計画シート由来のまま。
    due_urgent はうらした due_basis_date で再計算れる。
    """
    if task.get("due_basis_date") is not None:
        task["due_basis_date"] = task["due_basis_date"] + timedelta(days=1)
    db = task.get("due_basis_date")
    if db is not None:
        task["due_urgent"] = db <= run_date
def _seed_avail_from_timeline_for_date(
    timeline_events: list,
    current_date: date,
    machine_avail_dt: dict,
    avail_dt: dict,
    machine_day_start: datetime,
    *,
    events_today: list | None = None,
) -> None:
    """同一日内の既存 timeline から設備空し・メンバー空しの下限を反映れる（部分再配台用）。"""
    _iter = (
        events_today
        if events_today is not None
        else (e for e in timeline_events if e.get("date") == current_date)
    )
    for e in _iter:
        end_dt = e.get("end_dt")
        if end_dt is None or not hasattr(end_dt, "replace"):
            continue
        occ = _machine_occupancy_key_from_timeline_event(e)
        if occ:
            prev = machine_avail_dt.get(occ, machine_day_start)
            if end_dt > prev:
                machine_avail_dt[occ] = end_dt
        op = str(e.get("op") or "").strip()
        if op and op in avail_dt:
            prev_m = avail_dt[op]
            if end_dt > prev_m:
                avail_dt[op] = end_dt
        sub_raw = e.get("sub") or ""
        for sn in str(sub_raw).split(","):
            sm = sn.strip()
            if sm and sm in avail_dt:
                prev_s = avail_dt[sm]
                if end_dt > prev_s:
                    avail_dt[sm] = end_dt
def _merge_machine_calendar_intervals(
    intervals: list[tuple[datetime, datetime]],
) -> list[tuple[datetime, datetime]]:
    if not intervals:
        return []
    iv = sorted(intervals, key=lambda x: (x[0], x[1]))
    out = [iv[0]]
    for s, e in iv[1:]:
        ps, pe = out[-1]
        if s <= pe:
            out[-1] = (ps, max(pe, e))
        else:
            out.append((s, e))
    return out
def _half_open_gaps_in_window(
    w0: datetime,
    w1: datetime,
    covered_merged: list[tuple[datetime, datetime]],
) -> list[tuple[datetime, datetime]]:
    """
    半開区間 [w0,w1) から、マージ済み sorted covered_merged の合併を除いた部分を返す。
    covered_merged は [w0,w1) 内とみなしてよい（クリップ済み）前提。
    """
    if w0 >= w1:
        return []
    if not covered_merged:
        return [(w0, w1)]
    gaps: list[tuple[datetime, datetime]] = []
    cur = w0
    for s, e in covered_merged:
        if e <= cur:
            continue
        if s >= w1:
            break
        if cur < s:
            gaps.append((cur, min(s, w1)))
        cur = max(cur, e)
        if cur >= w1:
            return gaps
    if cur < w1:
        gaps.append((cur, w1))
    return gaps
def _bump_dt_past_machine_calendar_blocks(
    t: datetime,
    blocks: list[tuple[datetime, datetime]],
) -> datetime:
    """半開区間ブロック [start,end) に t は入る間」終端へ繰り上きる。"""
    if not blocks:
        return t
    changed = True
    while changed:
        changed = False
        for s, e in blocks:
            if s <= t < e:
                t = e
                changed = True
                break
    return t
def _machine_cal_parse_slot_datetime(cell) -> datetime | None:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    try:
        dt = pd.to_datetime(cell, errors="coerce")
    except Exception:
        return None
    if dt is None or (isinstance(dt, float) and pd.isna(dt)):
        return None
    if isinstance(dt, pd.Timestamp):
        dt = dt.to_pydatetime()
    if getattr(dt, "tzinfo", None) is not None:
        dt = dt.replace(tzinfo=None)
    return dt
def _machine_cal_cell_is_occupied(cell) -> bool:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return False
    if isinstance(cell, str):
        return bool(cell.strip())
    if isinstance(cell, bool):
        return cell
    # Excel で 0 を「空」としている列や」数弝の結果 0 は占有しない（従来 True てと全日占有扱いになり得る）
    if isinstance(cell, (int, float)):
        try:
            return float(cell) != 0.0
        except (TypeError, ValueError):
            return True
    return True
def _machine_cal_cell_is_asterisk_occupancy_only(cell) -> bool:
    """
    インタラクティブ配台試行: 機械カレンダーは「*」/「＊」のセルのみを占有とみなす（それ以外の非空は無視）。
    """
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return False
    s = str(cell).strip()
    if not s:
        return False
    if s in ("*", "＊", "※"):
        return True
    return False
def _clip_machine_calendar_slot_to_factory_window(
    day_d: date, slot_start: datetime, slot_end: datetime
) -> tuple[datetime, datetime] | None:
    """
    機械カレンダー1スロット [slot_start, slot_end) を工場稼働枠にクリップれる。
    枠外のみのスロットは None（配台では無視）。段階2では master メイン A12/B12 で
    DEFAULT_START_TIME / DEFAULT_END_TIME は上書き済み（generate_plan のコンテキスト内で読込）。
    """
    w0 = datetime.combine(day_d, DEFAULT_START_TIME)
    w1 = datetime.combine(day_d, DEFAULT_END_TIME)
    s2 = max(slot_start, w0)
    e2 = min(slot_end, w1)
    if s2 < e2:
        return (s2, e2)
    return None
def _machine_calendar_planning_window_end_dt(
    current_date: date,
    daily_status: dict,
    members: list,
) -> datetime:
    """
    機械カレンダー占有の坳端を切る上限。工場マスタ終業（DEFAULT_END_TIME）と」
    当日配台対象メンバーの勤務終了時刻の最尝の尝さい方（人はいない時間帯の「占有」で
    設備床の値は終業を超ごないよごにれる）。
    """
    w_factory = datetime.combine(current_date, DEFAULT_END_TIME)
    ends: list[datetime] = []
    for m in members:
        if m not in daily_status:
            continue
        st = daily_status[m]
        if not st.get("eligible_for_assignment", st.get("is_working", False)):
            continue
        et = st.get("end_dt")
        if et is not None and hasattr(et, "replace"):
            ends.append(et)
    if not ends:
        return w_factory
    return min(w_factory, min(ends))
def _clip_machine_busy_blocks_to_planning_window(
    blocks: list[tuple[datetime, datetime]],
    w0: datetime,
    w1: datetime,
) -> list[tuple[datetime, datetime]]:
    """占有半開区間を [w0, w1) にクリップしてからマージれる。"""
    out: list[tuple[datetime, datetime]] = []
    for s, e in blocks or []:
        s2 = max(s, w0)
        e2 = min(e, w1)
        if s2 < e2:
            out.append((s2, e2))
    if not out:
        return []
    return _merge_machine_calendar_intervals(out)
def _machine_cal_resolve_column_to_equipment_key(
    p_raw,
    m_raw,
    eq_lookup: dict,
    elist_set: set,
) -> str | None:
    p_s = (
        str(p_raw).strip()
        if p_raw is not None and not (isinstance(p_raw, float) and pd.isna(p_raw))
        else ""
    )
    m_s = (
        str(m_raw).strip()
        if m_raw is not None and not (isinstance(m_raw, float) and pd.isna(m_raw))
        else ""
    )
    if p_s and m_s:
        combo = f"{p_s}+{m_s}"
    elif p_s:
        combo = p_s
    else:
        return None
    if combo in elist_set:
        return combo
    nk = _normalize_equipment_match_key(combo)
    return eq_lookup.get(nk)
def _try_load_machine_calendar_blocks_from_json(
    equipment_list: list,
    *,
    interactive_only_asterisk_occupancy: bool = False,
    context_label: str = "配台",
) -> dict[date, dict[str, list[tuple[datetime, datetime]]]]:
    """JSON 正本から占有ブロックを読み込む。不備時は PlanningValidationError。"""
    global _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE
    from planning_core.core.machine_calendar_store import (
        load_machine_calendar_store,
        occupancy_blocks_from_store,
        require_machine_calendar_json_for_dispatch,
    )

    jp = require_machine_calendar_json_for_dispatch(context_label)
    store = load_machine_calendar_store(jp)
    try:
        out, interactive_defined = occupancy_blocks_from_store(
            store,
            equipment_list,
            interactive_only_asterisk_occupancy=interactive_only_asterisk_occupancy,
        )
    except Exception as e:
        raise PlanningValidationError(
            f"{context_label}: machine-calendar-data.json から占有ブロックを構築できません ({e})。"
            f" パス: {jp}"
        ) from e
    if interactive_only_asterisk_occupancy:
        _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE = dict(interactive_defined)
    else:
        _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE = {}
    logging.info(
        "機械カレンダー: JSON 正本を読み込みました（%s、%d 日分）。",
        jp,
        len(out),
    )
    return out


def load_machine_calendar_occupancy_blocks(
    master_path: str,
    equipment_list: list,
    *,
    interactive_only_asterisk_occupancy: bool = False,
    context_label: str | None = None,
) -> dict[date, dict[str, list[tuple[datetime, datetime]]]]:
    """
    機械カレンダー占有ブロックを machine-calendar-data.json 正本から読み込む。

    master.xlsm「機械カレンダー」シートへのフォールバックは行わない。
    JSON が無い・未整備・読込失敗時は PlanningValidationError で停止する。

    interactive_only_asterisk_occupancy:
        True のとき（配台試行）非空セルのうち * / ＊ / ※ のみを占有とする。
        列0にスロット行が無い時刻（工場計画窓内）はブロックとみなす。
    """
    global _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE
    global _STAGE2_MACHINE_CALENDAR_CACHE
    from planning_core.core.machine_calendar_store import require_machine_calendar_json_for_dispatch

    ctx = (context_label or "配台").strip()
    jp = require_machine_calendar_json_for_dispatch(ctx)
    eq_sig = ",".join(sorted(str(x).strip() for x in (equipment_list or []) if str(x).strip()))
    sig = None
    try:
        st = os.stat(jp)
        sig = (
            str(jp.resolve()),
            int(st.st_mtime),
            int(st.st_size),
            hashlib.sha256(eq_sig.encode("utf-8")).hexdigest(),
            "ia1" if interactive_only_asterisk_occupancy else "ia0",
        )
        if (
            isinstance(_STAGE2_MACHINE_CALENDAR_CACHE, dict)
            and _STAGE2_MACHINE_CALENDAR_CACHE.get("sig") == sig
        ):
            idef = _STAGE2_MACHINE_CALENDAR_CACHE.get("interactive_defined")
            if isinstance(idef, dict):
                _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE = idef
            else:
                _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE = {}
            return _STAGE2_MACHINE_CALENDAR_CACHE.get("value") or {}
    except OSError:
        sig = None

    out = _try_load_machine_calendar_blocks_from_json(
        equipment_list,
        interactive_only_asterisk_occupancy=interactive_only_asterisk_occupancy,
        context_label=ctx,
    )
    try:
        if sig is not None:
            _STAGE2_MACHINE_CALENDAR_CACHE = {
                "sig": sig,
                "value": out,
                "interactive_defined": dict(_MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE),
            }
    except Exception:
        pass
    return out


def _apply_machine_calendar_floor_for_date(
    current_date: date,
    machine_avail_dt: dict,
    equipment_list: list,
    machine_day_start: datetime,
    *,
    machine_calendar_plan_end: datetime | None = None,
) -> None:
    """当日のタイムラインシード後」機械カレンダー占有で設備空し下限を繰り上きる。"""
    raw_blocks = _MACHINE_CALENDAR_BLOCKS_BY_DATE.get(current_date)
    day_blocks = _interactive_augment_machine_calendar_day_blocks(
        current_date, raw_blocks, equipment_list
    )
    if not day_blocks:
        return
    candidates: set[str] = set()
    for k in machine_avail_dt.keys():
        sk = str(k).strip() if k is not None else ""
        if sk:
            candidates.add(sk)
    for el in equipment_list:
        ek = str(el).strip() if el is not None else ""
        if not ek:
            continue
        pk = _equipment_line_key_to_physical_occupancy_key(ek)
        if pk:
            candidates.add(pk)
    w0 = machine_day_start
    w1 = machine_calendar_plan_end
    if w1 is None:
        w1 = datetime.combine(current_date, DEFAULT_END_TIME)
    for eq_s in candidates:
        blocks = day_blocks.get(eq_s) or _machine_calendar_blocks_for_occ_key(
            day_blocks, eq_s
        )
        if not blocks:
            continue
        blocks_c = _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
        if not blocks_c:
            continue
        t0 = machine_avail_dt.get(eq_s, machine_day_start)
        t1 = _bump_dt_past_machine_calendar_blocks(t0, blocks_c)
        if t1 > t0:
            machine_avail_dt[eq_s] = t1
def _machine_calendar_blocks_for_occ_key(
    day_blocks: dict[str, list[tuple[datetime, datetime]]],
    occ: str,
) -> list[tuple[datetime, datetime]] | None:
    """day_blocks から占有キー（表記ゆらね許容）に一致する区間リストを得る。"""
    o = str(occ or "").strip()
    if not o or not day_blocks:
        return None
    if o in day_blocks:
        return day_blocks[o]
    nk = _normalize_equipment_match_key(o)
    for k, iv in day_blocks.items():
        if _normalize_equipment_match_key(str(k)) == nk:
            return iv
    return None
def _machine_calendar_occ_blocks_full_plan_window(
    occ_key: str,
    current_date: date,
    daily_status: dict,
    members: list,
) -> bool:
    """
    当日の機械カレンダー占有は計画窓 [始業, min(終業,稼働メンバー終了) ) 全体を塞ね」
    しの設備では当日 1 本も加工を入れられないとし True。
    """
    raw_blocks = _MACHINE_CALENDAR_BLOCKS_BY_DATE.get(current_date)
    day_blocks = _interactive_augment_machine_calendar_day_blocks(
        current_date,
        raw_blocks,
        None,
        extra_occ_keys=[occ_key],
    )
    if not day_blocks:
        return False
    blocks = _machine_calendar_blocks_for_occ_key(day_blocks, occ_key)
    if not blocks:
        return False
    w0 = datetime.combine(current_date, DEFAULT_START_TIME)
    w1 = _machine_calendar_planning_window_end_dt(current_date, daily_status, members)
    blocks_c = _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
    if not blocks_c:
        return False
    t1 = _bump_dt_past_machine_calendar_blocks(w0, blocks_c)
    return t1 >= w1
def _task_fully_machine_calendar_blocked_on_date(
    t: dict,
    current_date: date,
    daily_status: dict | None,
    members: list | None,
) -> bool:
    """
    当該タスクの占有設備は」当日の機械カレンダーの値で計画窓を全日塞はれでいる。
    グローバル試行順ブロック用の「最尝試行順」から外れ（他設備の配台デッドロック防止）。
    """
    if daily_status is None or members is None:
        return False
    _tm = t.get("machine")
    _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
    occ = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
    if not occ:
        return False
    return _machine_calendar_occ_blocks_full_plan_window(
        occ, current_date, daily_status, members
    )
def _task_no_machining_window_left_from_avail_floor(
    t: dict,
    current_date: date,
    daily_status: dict | None,
    members: list | None,
    machine_avail_dt: dict | None,
    machine_day_start: datetime | None,
    *,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
) -> bool:
    """
    machine_avail_dt（シード・機械カレンダー床・当日確定ロール反映後）で」
    占有設備の空し下限は計画窓終端以上なら当日は当設備にスロットなし。
    `machine_handoff` 等は渡るとしは `_resolve_machine_changeover_floor_segments` により
    `_assign_one_roll_trial_order_flow` とともに **実効加工開始下限** で判定れる
    （生の machine_avail の値ではポャンジオーポー後の下限は欠け」候補や min_dto は狂ごのを防し）。
    また空し下限は終端より版でも」計画窓での **残り連続は 1 ロール分に足りない**
    と判断でしる場合は True（実僝丝足デッドロック防止）。
    カレンダー区間照合のキー坖りこれしを防し。
    """
    if (
        daily_status is None
        or members is None
        or machine_avail_dt is None
        or machine_day_start is None
    ):
        return False
    w1 = _machine_calendar_planning_window_end_dt(current_date, daily_status, members)
    _tm = t.get("machine")
    _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
    occ = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
    if not occ:
        return False
    use_co = machine_handoff is not None and skills_dict is not None
    if use_co:
        machine_name = str(t.get("machine_name", "") or "").strip()
        machine_proc = str(_tm or "").strip()
        eq_line = str(
            t.get("equipment_line_key") or _tm or ""
        ).strip() or str(_tm or "")
        machine_occ_key = _machine_occupancy_key_resolve(t, eq_line)
        t_floor, _segs, abort = _resolve_machine_changeover_floor_segments(
            abolish_all_scheduling_limits=bool(abolish_all_scheduling_limits),
            machine_occ_key=machine_occ_key,
            task_id=str(t.get("task_id") or "").strip(),
            eq_line=eq_line,
            machine_name=machine_name,
            machine_proc=machine_proc,
            machine_avail_dt=machine_avail_dt,
            machine_day_floor=machine_day_start,
            current_date=current_date,
            machine_handoff=machine_handoff,
            daily_status=daily_status,
            skills_dict=skills_dict,
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
        if abort:
            return True
    else:
        t_floor = machine_avail_dt.get(occ)
        if t_floor is None:
            nk = _normalize_equipment_match_key(occ)
            for k, v in machine_avail_dt.items():
                if _normalize_equipment_match_key(str(k)) == nk:
                    t_floor = v
                    break
        if t_floor is None:
            t_floor = machine_day_start
    if t_floor >= w1:
        return True
    rem = w1 - t_floor
    if rem <= timedelta(0):
        return True
    btp = parse_float_safe(t.get("base_time_per_unit"), 0.0)
    if btp <= 0:
        return False
    t_eff = parse_float_safe(t.get("task_eff_factor"), 1.0)
    if t_eff <= 0:
        t_eff = 1.0
    # eff_time_per_unit ≈ base / avg_eff / t_eff × 余力係数。avg_eff はフォーム次第で下はる。
    _avg_eff_floor = 0.5
    approx_need_mins = max(1.0, float(btp) / t_eff / _avg_eff_floor)
    return rem < timedelta(minutes=approx_need_mins)
def _bump_machine_avail_after_roll_for_calendar(
    current_date: date,
    eq_line: str,
    machine_avail_dt: dict,
    *,
    machine_calendar_plan_end: datetime | None = None,
    machine_day_floor: datetime | None = None,
) -> None:
    """ロール確定直後: 終了時刻はカレンダー占有スロット内なら終端まで繰り上き。"""
    raw_blocks = _MACHINE_CALENDAR_BLOCKS_BY_DATE.get(current_date)
    eq_s = str(eq_line).strip() if eq_line is not None else ""
    day_blocks = _interactive_augment_machine_calendar_day_blocks(
        current_date,
        raw_blocks,
        None,
        extra_occ_keys=[eq_s] if eq_s else None,
    )
    if not day_blocks:
        return
    if not eq_s:
        return
    blocks = day_blocks.get(eq_s)
    if not blocks:
        return
    t0 = machine_avail_dt.get(eq_s)
    if t0 is None:
        return
    w0 = (
        machine_day_floor
        if machine_day_floor is not None
        else datetime.combine(current_date, DEFAULT_START_TIME)
    )
    w1 = (
        machine_calendar_plan_end
        if machine_calendar_plan_end is not None
        else datetime.combine(current_date, DEFAULT_END_TIME)
    )
    blocks_c = _clip_machine_busy_blocks_to_planning_window(blocks, w0, w1)
    if not blocks_c:
        return
    t1 = _bump_dt_past_machine_calendar_blocks(t0, blocks_c)
    if t1 > t0:
        machine_avail_dt[eq_s] = t1
def _parse_nonneg_minutes_cell(v) -> int:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    try:
        n = int(round(float(v)))
    except (TypeError, ValueError):
        return 0
    return max(0, n)
def _df_pick_column(df, *candidates: str) -> str | None:
    cols = [str(c).strip() for c in df.columns]
    low_map = {str(c).strip().lower(): str(c).strip() for c in df.columns}
    for cand in candidates:
        c0 = str(cand).strip()
        if c0 in df.columns:
            return c0
        cl = c0.lower()
        if cl in low_map:
            return low_map[cl]
    return None
def load_machine_daily_startup_settings(
    master_path: str,
) -> tuple[dict[str, int], dict[str, int]]:
    """
    master.xlsm の任意シート「設定_機械_日次始業準備」… 機械名・日次始業準備分・任意で必要人数。

    戻り値: (分 dict, 必要人数 dict)。
    """
    startup: dict[str, int] = {}
    req_staff: dict[str, int] = {}
    if not master_path or not os.path.isfile(master_path):
        return startup, req_staff
    try:
        xls = _cached_master_pd_excel_file(master_path)
        if xls is None:
            return startup, req_staff
    except Exception as e:
        logging.warning("機械日次始業準備設定: ブックを開きません (%s)", e)
        return startup, req_staff

    if SHEET_MACHINE_DAILY_STARTUP in xls.sheet_names:
        try:
            df2 = pd.read_excel(
                xls, sheet_name=SHEET_MACHINE_DAILY_STARTUP, header=0
            )
            df2.columns = [str(c).strip() for c in df2.columns]
            c_mn = _df_pick_column(df2, "機械名", "機械")
            c_su = _df_pick_column(
                df2, "日次始業準備_分", "始業準備_分", "日始業準備_分"
            )
            c_rq = _df_pick_column(df2, "必要人数", "日次始業準備_必要人数")
            if c_mn and c_su:
                for _, row in df2.iterrows():
                    mn = row.get(c_mn)
                    if mn is None or (isinstance(mn, float) and pd.isna(mn)):
                        continue
                    mn_s = str(mn).strip()
                    if not mn_s or mn_s.lower() == "nan":
                        continue
                    su = _parse_nonneg_minutes_cell(row.get(c_su))
                    if su <= 0:
                        continue
                    startup[mn_s] = su
                    nk = _normalize_equipment_match_key(mn_s)
                    if nk:
                        startup[nk] = su
                    rq = 0
                    if c_rq:
                        rq = _parse_nonneg_minutes_cell(row.get(c_rq))
                    if rq > 0:
                        req_staff[mn_s] = rq
                        if nk:
                            req_staff[nk] = rq
                if startup:
                    logging.info(
                        "マスタ「%s」: 機械 %s 件の日次始業準備（分）を読み込みました（必要人数指定 %s 件）。",
                        SHEET_MACHINE_DAILY_STARTUP,
                        len({k for k in startup if "+" not in str(k)}),
                        len({k for k in req_staff if "+" not in str(k)}),
                    )
        except Exception as e:
            logging.warning(
                "マスタ「%s」読込失敗（無視）: %s", SHEET_MACHINE_DAILY_STARTUP, e
            )

    return startup, req_staff
def load_request_switch_prep_settings(
    master_path: str,
    *,
    _allow_kokubu_merge: bool = True,
) -> tuple[
    dict[tuple[str, str], int],
    dict[str, int],
    dict[tuple[str, str], int],
    dict[str, int],
    dict[tuple[str, str], int],
    dict[str, int],
    dict[tuple[str, str], int],
    dict[str, int],
]:
    """
    master.xlsm「設定_依頼切替前後時間」… 工程名・機械名・依頼切替準備時間・休憩後再開準備時間・
    後始末時間・加工依頼間の余裕時間。

    戻り値: (依頼切替準備 by (工程,機械), 依頼切替準備 by 機械名のみ,
            休憩再開準備 by (工程,機械), 休憩再開準備 by 機械名のみ,
            後始末 by (工程,機械), 後始末 by 機械名のみ,
            依頼間余裕 by (工程,機械), 依頼間余裕 by 機械名のみ)。
    """
    switch_pair: dict[tuple[str, str], int] = {}
    switch_machine: dict[str, int] = {}
    resume_pair: dict[tuple[str, str], int] = {}
    resume_machine: dict[str, int] = {}
    cleanup_pair: dict[tuple[str, str], int] = {}
    cleanup_machine: dict[str, int] = {}
    buffer_pair: dict[tuple[str, str], int] = {}
    buffer_machine: dict[str, int] = {}
    empty8 = (
        switch_pair,
        switch_machine,
        resume_pair,
        resume_machine,
        cleanup_pair,
        cleanup_machine,
        buffer_pair,
        buffer_machine,
    )
    if not master_path or not os.path.isfile(master_path):
        return empty8
    try:
        xls = _cached_master_pd_excel_file(master_path)
        if xls is None:
            return empty8
    except Exception as e:
        logging.warning("依頼切替準備設定: ブックを開きません (%s)", e)
        return empty8
    if SHEET_REQUEST_SWITCH_PREP not in xls.sheet_names:
        return empty8
    try:
        df = pd.read_excel(
            xls, sheet_name=SHEET_REQUEST_SWITCH_PREP, header=0
        )
        df.columns = [str(c).strip() for c in df.columns]

        def _col_variants(*bases: str) -> list[str]:
            out: list[str] = []
            for b in bases:
                b0 = str(b).strip()
                if not b0:
                    continue
                for c in df.columns:
                    cs = str(c).strip()
                    if cs == b0 or cs.startswith(b0 + "."):
                        if cs not in out:
                            out.append(cs)
            return out

        def _row_text(row, cols: list[str]) -> str:
            for c in cols:
                v = row.get(c)
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    continue
                s = str(v).strip()
                if s and s.lower() != "nan":
                    return s
            return ""

        def _row_minutes(row, cols: list[str]) -> int:
            best = 0
            for c in cols:
                v = _parse_nonneg_minutes_cell(row.get(c))
                if v > best:
                    best = v
            return best

        proc_cols = _col_variants("工程名", "工程")
        mn_cols = _col_variants("機械名", "機械")
        prep_cols = _col_variants(
            "依頼切替準備時間",
            "準備時間_分",
            "準備分",
            "準備時間",
            "依頼切替準備_分",
        )
        if not prep_cols:
            prep_cols = [
                c
                for c in df.columns
                if "後始末" not in str(c)
                and "余裕" not in str(c)
                and "再開" not in str(c)
                and (
                    "準備時間" in str(c)
                    or "依頼切替準備" in str(c)
                )
            ]
        prep_cols = list(dict.fromkeys(prep_cols))
        resume_cols = _col_variants(
            "休憩後再開準備時間",
            "一時停止後_再開準備時間",
            "再開準備時間",
            "一時停止後_再開準備時間_分",
            "再開準備_分",
        )
        cleanup_cols = _col_variants(
            "後始末時間",
            "後始末_分",
            "後始末時間_分",
        )
        buffer_cols = _col_variants(
            "加工依頼間の余裕時間",
            "依頼間余裕時間",
            "余裕時間_分",
        )
        if not mn_cols:
            return empty8
        for _, row in df.iterrows():
            mn_s = _row_text(row, mn_cols)
            if not mn_s:
                continue
            proc_s = _row_text(row, proc_cols)
            prep_m = _row_minutes(row, prep_cols)
            resume_m = _row_minutes(row, resume_cols)
            cleanup_m = _row_minutes(row, cleanup_cols)
            buffer_m = _row_minutes(row, buffer_cols)
            nk = _normalize_equipment_match_key(mn_s)
            if proc_s:
                if prep_m > 0:
                    switch_pair[(proc_s, mn_s)] = prep_m
                if resume_m > 0:
                    resume_pair[(proc_s, mn_s)] = resume_m
                if cleanup_m > 0:
                    cleanup_pair[(proc_s, mn_s)] = cleanup_m
                if buffer_m > 0:
                    buffer_pair[(proc_s, mn_s)] = buffer_m
            else:
                if prep_m > 0:
                    switch_machine[mn_s] = prep_m
                    if nk:
                        switch_machine[nk] = prep_m
                if resume_m > 0:
                    resume_machine[mn_s] = resume_m
                    if nk:
                        resume_machine[nk] = resume_m
                if cleanup_m > 0:
                    cleanup_machine[mn_s] = cleanup_m
                    if nk:
                        cleanup_machine[nk] = cleanup_m
                if buffer_m > 0:
                    buffer_machine[mn_s] = buffer_m
                    if nk:
                        buffer_machine[nk] = buffer_m
        if (
            switch_pair
            or switch_machine
            or resume_pair
            or resume_machine
            or cleanup_pair
            or cleanup_machine
            or buffer_pair
            or buffer_machine
        ):
            logging.info(
                "マスタ「%s」: 依頼切替準備 %s 件・休憩再開準備 %s 件・"
                "後始末 %s 件・依頼間余裕 %s 件（工程+機械 / 機械のみの内訳はログのみ）。",
                SHEET_REQUEST_SWITCH_PREP,
                len(switch_pair) + len(switch_machine),
                len(resume_pair) + len(resume_machine),
                len(cleanup_pair) + len(cleanup_machine),
                len(buffer_pair) + len(buffer_machine),
            )
    except Exception as e:
        logging.warning(
            "マスタ「%s」読込失敗（無視）: %s", SHEET_REQUEST_SWITCH_PREP, e
        )
    if _allow_kokubu_merge:
        _added = _merge_request_switch_prep_from_sibling_kokubu_master(
            master_path,
            switch_pair,
            switch_machine,
            resume_pair,
            resume_machine,
            cleanup_pair,
            cleanup_machine,
            buffer_pair,
            buffer_machine,
        )
        if _added > 0:
            logging.info(
                "マスタ「%s」: 同フォルダの国分master.xlsm から不足分 %s 件を補完しました。",
                SHEET_REQUEST_SWITCH_PREP,
                _added,
            )
    return empty8
def _merge_request_switch_prep_from_sibling_kokubu_master(
    primary_path: str,
    switch_pair: dict[tuple[str, str], int],
    switch_machine: dict[str, int],
    resume_pair: dict[tuple[str, str], int],
    resume_machine: dict[str, int],
    cleanup_pair: dict[tuple[str, str], int],
    cleanup_machine: dict[str, int],
    buffer_pair: dict[tuple[str, str], int],
    buffer_machine: dict[str, int],
) -> int:
    """正本マスタに無い (工程,機械) を同ディレクトリの国分master.xlsm から補完（上書きしない）。"""
    if not primary_path:
        return 0
    try:
        base = os.path.dirname(os.path.abspath(primary_path))
        alt = os.path.join(base, "国分master.xlsm")
        if not os.path.isfile(alt):
            return 0
        if os.path.normcase(os.path.abspath(alt)) == os.path.normcase(
            os.path.abspath(primary_path)
        ):
            return 0
        sp2, sm2, rp2, rm2, cp2, cm2, bp2, bm2 = load_request_switch_prep_settings(
            alt, _allow_kokubu_merge=False
        )
    except Exception as e:
        logging.warning("国分master 依頼切替準備の補完読込をスキップ: %s", e)
        return 0
    added = 0
    for src, dst in (
        (sp2, switch_pair),
        (sm2, switch_machine),
        (rp2, resume_pair),
        (rm2, resume_machine),
        (cp2, cleanup_pair),
        (cm2, cleanup_machine),
        (bp2, buffer_pair),
        (bm2, buffer_machine),
    ):
        for k, v in src.items():
            try:
                iv = int(v or 0)
            except (TypeError, ValueError):
                iv = 0
            if iv <= 0 or k in dst:
                continue
            dst[k] = iv
            added += 1
    return added
def _normalize_proc_machine_for_prep_lookup(
    machine_proc: str,
    machine_name: str,
    *,
    eq_line: str = "",
) -> tuple[str, str]:
    """マスタ「設定_依頼切替前後時間」ルックアップ用に (工程, 機械名) を正規化する。"""
    proc = str(machine_proc or "").strip()
    mn = str(machine_name or "").strip()
    if proc and mn:
        return proc, mn
    ek = str(eq_line or "").strip()
    if ek:
        nek = _normalize_equipment_match_key(ek)
        if "+" in nek:
            a, b = ek.split("+", 1)
            return str(a).strip(), str(b).strip()
    if proc:
        pk = _normalize_equipment_match_key(proc)
        if "+" in pk:
            a, b = proc.split("+", 1)
            return str(a).strip(), str(b).strip()
    return proc, mn
def _equipment_names_match_for_prep_lookup(a: str, b: str) -> bool:
    """機械名照合（NFKC・空白正規化・拠点名付き表記の前方一致）。"""
    na = _normalize_equipment_match_key(a)
    nb = _normalize_equipment_match_key(b)
    if not na or not nb:
        return na == nb
    if na == nb:
        return True
    short, long = (na, nb) if len(na) <= len(nb) else (nb, na)
    return long.startswith(short)
def _lookup_prep_minutes_from_stage2_tables(
    machine_proc: str,
    machine_name: str,
    by_pair: dict[tuple[str, str], int] | None,
    by_machine: dict[str, int] | None,
) -> int:
    proc = str(machine_proc or "").strip()
    mn = str(machine_name or "").strip()
    pair_d = by_pair if by_pair is not None else {}
    mach_d = by_machine if by_machine is not None else {}
    pn = _normalize_process_name_for_rule_match(proc) if proc else ""
    if proc and mn:
        v = pair_d.get((proc, mn))
        if v is not None and int(v) > 0:
            return int(v)
    if mn:
        if mn in mach_d and int(mach_d[mn]) > 0:
            return int(mach_d[mn])
        nk = _normalize_equipment_match_key(mn)
        if nk in mach_d and int(mach_d[nk]) > 0:
            return int(mach_d[nk])
        for k, val in mach_d.items():
            if _equipment_names_match_for_prep_lookup(k, mn) and int(val) > 0:
                return int(val)
    best_pm = 0
    for (p, m), val in pair_d.items():
        try:
            iv = int(val or 0)
        except (TypeError, ValueError):
            iv = 0
        if iv <= 0:
            continue
        if pn and _normalize_process_name_for_rule_match(p) != pn:
            continue
        if mn and not _equipment_names_match_for_prep_lookup(m, mn):
            continue
        best_pm = max(best_pm, iv)
    if best_pm > 0:
        return best_pm
    if mn:
        for (_p, m), val in pair_d.items():
            try:
                iv = int(val or 0)
            except (TypeError, ValueError):
                iv = 0
            if iv <= 0:
                continue
            if not _equipment_names_match_for_prep_lookup(m, mn):
                continue
            best_pm = max(best_pm, iv)
    return best_pm
def _lookup_request_switch_prep_minutes(
    machine_proc: str,
    machine_name: str,
    *,
    eq_line: str = "",
) -> int:
    proc, mn = _normalize_proc_machine_for_prep_lookup(
        machine_proc, machine_name, eq_line=eq_line
    )
    return _lookup_prep_minutes_from_stage2_tables(
        proc,
        mn,
        _STAGE2_REQUEST_SWITCH_PREP_BY_PROC_MACHINE,
        _STAGE2_REQUEST_SWITCH_PREP_BY_MACHINE,
    )
def _lookup_break_resume_prep_minutes(
    machine_proc: str,
    machine_name: str,
    *,
    eq_line: str = "",
) -> int:
    proc, mn = _normalize_proc_machine_for_prep_lookup(
        machine_proc, machine_name, eq_line=eq_line
    )
    return _lookup_prep_minutes_from_stage2_tables(
        proc,
        mn,
        _STAGE2_BREAK_RESUME_PREP_BY_PROC_MACHINE,
        _STAGE2_BREAK_RESUME_PREP_BY_MACHINE,
    )
def _stage2_post_machining_cleanup_enabled() -> bool:
    """後始末（post_machining_cleanup）を配台に適用するか。既定は無効。"""
    return _stage2_truthy_env("STAGE2_ENABLE_POST_MACHINING_CLEANUP")
def _lookup_post_machining_cleanup_minutes(
    machine_proc: str,
    machine_name: str,
    *,
    eq_line: str = "",
) -> int:
    if not _stage2_post_machining_cleanup_enabled():
        return 0
    proc, mn = _normalize_proc_machine_for_prep_lookup(
        machine_proc, machine_name, eq_line=eq_line
    )
    return _lookup_prep_minutes_from_stage2_tables(
        proc,
        mn,
        _STAGE2_POST_MACHINING_CLEANUP_BY_PROC_MACHINE,
        _STAGE2_POST_MACHINING_CLEANUP_BY_MACHINE,
    )
def _lookup_request_interval_buffer_minutes(
    machine_proc: str,
    machine_name: str,
    *,
    eq_line: str = "",
) -> int:
    proc, mn = _normalize_proc_machine_for_prep_lookup(
        machine_proc, machine_name, eq_line=eq_line
    )
    return _lookup_prep_minutes_from_stage2_tables(
        proc,
        mn,
        _STAGE2_REQUEST_INTERVAL_BUFFER_BY_PROC_MACHINE,
        _STAGE2_REQUEST_INTERVAL_BUFFER_BY_MACHINE,
    )
def _lookup_daily_startup_minutes(
    machine_name: str,
    by_m: dict[str, int] | None,
) -> int:
    st = by_m if by_m is not None else _STAGE2_MACHINE_DAILY_STARTUP_MIN_BY_MACHINE
    mn = str(machine_name or "").strip()
    if not mn:
        return 0
    if mn in st:
        return st[mn]
    nk = _normalize_equipment_match_key(mn)
    if nk in st:
        return st[nk]
    for k, v in st.items():
        if _normalize_equipment_match_key(str(k)) == nk:
            return v
    return 0
def _lookup_daily_startup_required_staff(
    machine_name: str,
    by_r: dict[str, int] | None,
) -> int:
    rq = by_r if by_r is not None else _STAGE2_MACHINE_DAILY_STARTUP_REQ_BY_MACHINE
    mn = str(machine_name or "").strip()
    if not mn:
        return 0
    if mn in rq:
        return rq[mn]
    nk = _normalize_equipment_match_key(mn)
    if nk in rq:
        return rq[nk]
    for k, v in rq.items():
        if _normalize_equipment_match_key(str(k)) == nk:
            return int(v)
    return 0
def _member_covers_interval_no_break_overlap(
    daily_status: dict, member: str, win_st: datetime, win_ed: datetime
) -> bool:
    """
    [win_st, win_ed) が当該メンバーの出勤帯内にあり、かついずれの休憩区間とも
    半開重なりを持たない（休憩をまたがない）とき True。
    """
    mm = str(member or "").strip()
    if not mm:
        return False
    st = daily_status.get(mm)
    if not st:
        return False
    shift_s = st.get("start_dt")
    shift_e = st.get("end_dt")
    if not isinstance(shift_s, datetime) or not isinstance(shift_e, datetime):
        return False
    if win_st < shift_s or win_ed > shift_e or win_ed <= win_st:
        return False
    for br in merge_time_intervals(list(st.get("breaks_dt") or [])):
        if not isinstance(br, (list, tuple)) or len(br) < 2:
            continue
        bs, be = br[0], br[1]
        if not isinstance(bs, datetime) or not isinstance(be, datetime):
            continue
        if win_st < be and win_ed > bs:
            return False
    return True
def _daily_startup_fill_segment_staff(
    seg: dict,
    *,
    machine_name: str,
    lead_op: str,
    sub_csv: str,
    skill_role_priority,
    daily_status: dict,
    avail_dt: dict,
    dispatch_interval_mirror,
) -> None:
    """
    日次始業セグメントに担当者を載せる（op + sub）。
    母集団は当該加工ロールの主・補。勤務帯で [st,ed) を完全に覆い休憩と重ならない者のみ。
    skills のロール優先度昇順で need 名まで選び、ミラー衝突する者はスキップする。
    """
    st = seg.get("start_dt")
    ed = seg.get("end_dt")
    if not isinstance(st, datetime) or not isinstance(ed, datetime) or ed <= st:
        return
    need_n = _lookup_daily_startup_required_staff(machine_name, None)
    if need_n <= 0:
        return
    pool: list[str] = []
    seen: set[str] = set()
    for raw in (str(lead_op or "").strip(),):
        if raw and raw not in seen:
            seen.add(raw)
            pool.append(raw)
    for part in re.split(r"[,、]", str(sub_csv or "")):
        t = part.strip()
        if t and t not in seen:
            seen.add(t)
            pool.append(t)

    def _eligible_for_startup(require_avail_before_start: bool) -> list[str]:
        out: list[str] = []
        for m in pool:
            if not _member_covers_interval_no_break_overlap(daily_status, m, st, ed):
                continue
            if require_avail_before_start:
                prev_a = avail_dt.get(m, st)
                if isinstance(prev_a, datetime) and prev_a > st:
                    continue
            out.append(m)
        out.sort(key=lambda mm: (skill_role_priority(mm)[1], str(mm)))
        return out

    # 1) 厳格: 勤務帯覆い + avail が始業準備開始以前
    # 2) 日次始業は工場定時直後の短区間のため、同一ロール母集団では avail のみ緩めて再試行
    # 3) 最終: 母集団のうち daily_status にいる者（表示用のフェイルセーフ）
    eligible = _eligible_for_startup(True)
    _eligible_pass = 1
    if not eligible:
        eligible = _eligible_for_startup(False)
        _eligible_pass = 2
    if not eligible:
        eligible = [m for m in pool if m in daily_status]
        eligible.sort(key=lambda mm: (skill_role_priority(mm)[1], str(mm)))
        _eligible_pass = 3
    chosen: list[str] = []
    for m in eligible:
        if len(chosen) >= need_n:
            break
        if dispatch_interval_mirror is not None and dispatch_interval_mirror.would_block_member(
            m, st, ed
        ):
            continue
        chosen.append(m)
    # 日次始業は同一ロール直後の短区間のため、ミラーのみで母集団全員が落ちると
    # タイムライン op が空になりガントに氏名が出ない。表示・整合優先でミラーを無視して埋める。
    if len(chosen) < need_n:
        for m in eligible:
            if len(chosen) >= need_n:
                break
            if m in chosen:
                continue
            chosen.append(m)
    if not chosen:
        return
    op0 = chosen[0]
    rest = chosen[1:]
    seg["op"] = op0
    seg["sub"] = ", ".join(rest) if rest else ""
def _timeline_event_kind(ev: dict) -> str:
    k = str(ev.get("event_kind") or "").strip()
    return k if k else TIMELINE_EVENT_MACHINING
def _is_machining_timeline_event(ev: dict) -> bool:
    return _timeline_event_kind(ev) == TIMELINE_EVENT_MACHINING
def _timeline_event_calendar_date(ev: dict) -> date | None:
    """加工タイムラインイベントの暦日（配台表の行キー）。"""
    d = ev.get("date")
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    st = ev.get("start_dt")
    if isinstance(st, datetime):
        return st.date()
    if isinstance(st, date):
        return st
    return None
def _sanitize_dispatch_qty_m(qty_m: float) -> float:
    """配台 m の float 加算誤差を抑え、整数に近い値は整数へ寄せる。"""
    try:
        q = float(qty_m)
    except (TypeError, ValueError):
        return 0.0
    if not math.isfinite(q):
        return 0.0
    r = round(q)
    if abs(q - r) <= 1e-6:
        return float(r)
    return round(q, 6)
def _dispatch_table_event_qty_m(ev: dict) -> float:
    """イベント当たりの配台量（メートル）。unit_m が無いときは units_done のみ。"""
    try:
        ud = float(parse_float_safe(ev.get("units_done"), 0.0) or 0.0)
        um = float(parse_float_safe(ev.get("unit_m"), 0.0) or 0.0)
    except Exception:
        return 0.0
    if um > 1e-18:
        return _sanitize_dispatch_qty_m(float(ud * um))
    return _sanitize_dispatch_qty_m(float(ud))
def _resolve_task_dict_for_timeline_line(
    tid: str, ev_machine_line: str, sorted_tasks_for_result: list | None
) -> dict | None:
    """timeline の machine（設備ライン）から task_queue 行を解決する。"""
    tid_s = str(tid or "").strip()
    em = str(ev_machine_line or "").strip()
    if not tid_s or not em or not sorted_tasks_for_result:
        return None
    cands: list[dict] = []
    for t in sorted_tasks_for_result:
        if str(t.get("task_id") or "").strip() != tid_s:
            continue
        ek = str(t.get("equipment_line_key") or "").strip()
        if ek and ek == em:
            cands.append(t)
    if len(cands) == 1:
        return cands[0]
    if len(cands) > 1:
        return cands[0]
    for t in sorted_tasks_for_result:
        if str(t.get("task_id") or "").strip() != tid_s:
            continue
        m_proc = str(t.get("machine") or "").strip()
        m_name = str(t.get("machine_name") or "").strip()
        if m_proc and m_proc in em and (not m_name or m_name in em):
            return t
    for t in sorted_tasks_for_result:
        if str(t.get("task_id") or "").strip() != tid_s:
            continue
        ek = str(t.get("equipment_line_key") or "").strip() or str(t.get("machine") or "").strip()
        if ek == em:
            return t
    return None
def _build_plan_input_row_lookup_for_dispatch_table(tasks_df) -> dict[tuple[str, str], object]:
    """(依頼NO, 工程名) → 計画入力 DataFrame の行（Series）。"""
    out: dict[tuple[str, str], object] = {}
    if tasks_df is None or getattr(tasks_df, "empty", True):
        return out
    try:
        for _, row in tasks_df.iterrows():
            try:
                if _plan_row_exclude_from_assignment(row):
                    continue
            except Exception:
                continue
            tid = str(_planning_df_cell_scalar(row, TASK_COL_TASK_ID) or "").strip()
            mach = str(_planning_df_cell_scalar(row, TASK_COL_MACHINE) or "").strip()
            if tid and mach:
                out[(tid, mach)] = row
    except Exception:
        return out
    return out
def _build_source_task_row_lookups_for_dispatch_table(df_src):
    """
    加工計画DATA の索引を2系統作る。
    - idx3: (依頼NO, 工程名, 受注日) → row（最優先）
    - idx2: (依頼NO, 工程名) → row または list[row]（受注日が tasks_df に無い場合のフォールバック）
    """
    idx3: dict[tuple[str, str, str], object] = {}
    idx2: dict[tuple[str, str], object] = {}
    if df_src is None or getattr(df_src, "empty", True):
        return idx3, idx2
    dup3 = 0
    miss_order_date = 0

    def _norm_order_date(v) -> str:
        return _norm_ymd(v)

    try:
        for _, row in df_src.iterrows():
            tid = str(_planning_df_cell_scalar(row, TASK_COL_TASK_ID) or "").strip()
            mach = str(_planning_df_cell_scalar(row, TASK_COL_MACHINE) or "").strip()
            if not tid or not mach:
                continue
            od = ""
            try:
                if hasattr(row, "index") and "受注日" in row.index:
                    od = _norm_order_date(_planning_df_cell_scalar(row, "受注日"))
            except Exception:
                od = ""
            if not od:
                miss_order_date += 1
            k3 = (tid, mach, od)
            if k3 in idx3:
                dup3 += 1
            else:
                idx3[k3] = row

            k2 = (tid, mach)
            if k2 not in idx2:
                idx2[k2] = row
            else:
                cur = idx2[k2]
                if isinstance(cur, list):
                    cur.append(row)
                else:
                    idx2[k2] = [cur, row]
    except Exception:
        return idx3, idx2

    return idx3, idx2
def _dispatch_table_scalar_from_dataframe_row(row, col_name: str):
    """加工計画DATA / 配台計画入力の Series 行から列値を取る（空・NaN は None）。"""
    if row is None:
        return None
    try:
        if not (hasattr(row, "index") and col_name in row.index):
            return None
        v = _planning_df_cell_scalar(row, col_name)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        if str(v).strip() == "":
            return None
        return v
    except Exception:
        return None
def _dispatch_table_cell_from_sources(
    *,
    src_row,
    plan_row,
    task_dict: dict | None,
    col_name: str,
):
    """結果_配台表の静的列を補完。

    既定は 加工計画DATA → 配台計画入力 → task_queue。
    ``_RESULT_DISPATCH_PLAN_INPUT_OVERRIDE_SRC_COLS``（実加工数・換算数量・実出来高・原反投入日など）は
    配台計画入力を優先（Aladdin 側の実加工数が受注数と同値で残るケース、およびタスク入力で編集した原反投入日を避ける）。
    ``_RESULT_DISPATCH_PROCESSING_PLAN_ONLY_SRC_COLS``（回答納期）は加工計画DATA のみ。無ければ空欄。
    """
    if col_name in ("加工開始日時", "加工終了日時", "メンバー名"):
        return ""
    if col_name in _RESULT_DISPATCH_PROCESSING_PLAN_ONLY_SRC_COLS:
        v = _dispatch_table_scalar_from_dataframe_row(src_row, col_name)
        if v is not None:
            return v
        return ""
    plan_first = col_name in _RESULT_DISPATCH_PLAN_INPUT_OVERRIDE_SRC_COLS
    row_order = (
        (plan_row, src_row) if plan_first else (src_row, plan_row)
    )
    for row in row_order:
        v = _dispatch_table_scalar_from_dataframe_row(row, col_name)
        if v is not None:
            return v
    t = task_dict
    if not t:
        return ""
    if col_name == TASK_COL_TASK_ID:
        return t.get("task_id") or ""
    if col_name == TASK_COL_MACHINE:
        return t.get("machine") or ""
    if col_name == TASK_COL_MACHINE_NAME:
        return t.get("machine_name") or ""
    if col_name == TASK_COL_QTY:
        um = float(t.get("unit_m") or 0)
        tqm = t.get("total_qty_m")
        if um > 1e-18 and tqm is not None:
            try:
                return float(tqm) / um
            except Exception:
                pass
        return t.get(TASK_COL_QTY) or ""
    if col_name == TASK_COL_USED_RAW:
        return t.get(TASK_COL_USED_RAW) or ""
    if col_name == TASK_COL_PRODUCT:
        return t.get(TASK_COL_PRODUCT) or ""
    if col_name == TASK_COL_PROCESS_CONTENT:
        return t.get(TASK_COL_PROCESS_CONTENT) or ""
    if col_name == TASK_COL_STOCK_LOCATION:
        return t.get(TASK_COL_STOCK_LOCATION) or ""
    if col_name == TASK_COL_RAW_INPUT_DATE:
        rid = t.get("raw_input_date")
        if rid is not None and hasattr(rid, "strftime"):
            return rid.strftime("%Y/%m/%d")
        return rid or ""
    if col_name == TASK_COL_SPECIFIED_DUE:
        sd = t.get("specified_due_date")
        if sd is not None and hasattr(sd, "strftime"):
            return sd.strftime("%Y/%m/%d")
        return sd or ""
    if col_name == TASK_COL_COMPLETION_FLAG:
        return t.get(TASK_COL_COMPLETION_FLAG) or ""
    if col_name == TASK_COL_ACTUAL_DONE:
        return (
            t.get(TASK_COL_ACTUAL_DONE)
            or t.get("done_qty_reported")
            or ""
        )
    if col_name == TASK_COL_ACTUAL_OUTPUT:
        return t.get(TASK_COL_ACTUAL_OUTPUT) or ""
    if col_name == RESULT_TASK_COL_DISPATCH_TRIAL_ORDER:
        v = t.get(RESULT_TASK_COL_DISPATCH_TRIAL_ORDER)
        if v is None:
            return ""
        return str(v).strip()
    return ""
def _timeline_event_start_end_dt(ev: dict) -> tuple[datetime | None, datetime | None]:
    """加工タイムラインイベントの開始・終了（配台表 1 行＝暦日集約の min/max に使用）。"""
    st = ev.get("start_dt")
    ed = ev.get("end_dt")
    if isinstance(st, datetime):
        pass
    elif isinstance(st, date):
        st = datetime(st.year, st.month, st.day)
    else:
        st = None
    if isinstance(ed, datetime):
        pass
    elif isinstance(ed, date):
        ed = datetime(ed.year, ed.month, ed.day, 23, 59, 59)
    else:
        ed = None
    return st, ed
def _fmt_dispatch_table_datetime(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y/%m/%d %H:%M")
    if isinstance(dt, date):
        return dt.strftime("%Y/%m/%d")
    return str(dt).strip()
def _primary_op_for_equipment_schedule_dispatch_row(
    tl_expanded: list,
    tid: str,
    eq: str,
    day: date,
) -> str:
    """
    結果_設備毎の時間割と同じ設備列キー・暦日で、開始が最も早い加工イベントの主担当（op）。
    `_expand_timeline_events_for_equipment_grid` 済みリストを渡すこと。
    """
    tid_s = str(tid or "").strip()
    eq_s = str(eq or "").strip()
    best_op = ""
    best_st = None
    for e in tl_expanded or []:
        if not _is_machining_timeline_event(e):
            continue
        if str(e.get("task_id") or "").strip() != tid_s:
            continue
        if str(e.get("machine") or "").strip() != eq_s:
            continue
        d0 = e.get("date")
        if d0 is None:
            continue
        if isinstance(d0, datetime):
            d0 = d0.date()
        elif not isinstance(d0, date):
            continue
        if d0 != day:
            continue
        st = e.get("start_dt")
        if not isinstance(st, datetime):
            continue
        op = " ".join(str(e.get("op") or "").split()).strip()
        if not op:
            continue
        if best_st is None or st < best_st:
            best_st = st
            best_op = op
    return best_op
def _format_dispatch_table_member_like_equipment_schedule(
    tid_k: str,
    eq_k: str,
    day_k: date,
    tl_expanded: list,
    unify_sub_map: dict,
    member_ops_fallback: list[str],
) -> str:
    """
    結果_設備毎の時間割の加工セル（進度バー表示行）と同じ文字列組み立て:
    [{task_id}] 主:{op} 補:{unified_sub}
    ``unified_sub`` は `_equipment_schedule_unified_sub_string_map` と同一キー (date, machine, task_id)。
    """
    sub_full = unify_sub_map.get((day_k, eq_k, tid_k))
    if sub_full is None:
        sub_full = ""
    sub_full = str(sub_full).strip()
    sub_text = f" 補:{sub_full}" if sub_full else ""
    pop = _primary_op_for_equipment_schedule_dispatch_row(
        tl_expanded, tid_k, eq_k, day_k
    )
    if pop:
        return f"[{tid_k}] 主:{pop}{sub_text}"
    if member_ops_fallback:
        return "、".join(member_ops_fallback)
    return ""
def _dispatch_table_row_identity_key(row: dict) -> tuple[str, str, str]:
    tid = str(row.get(TASK_COL_TASK_ID) or row.get("依頼NO") or "").strip()
    proc = str(row.get(TASK_COL_MACHINE) or "").strip()
    mach = str(row.get(TASK_COL_MACHINE_NAME) or "").strip()
    return (tid, proc, mach)
def _aggregate_timeline_dispatch_meta_by_identity_date(
    timeline_events: list | None,
    sorted_tasks_for_result: list | None,
) -> tuple[
    dict[tuple[str, str, str, date], float],
    dict[tuple[str, str, str, date], datetime],
    dict[tuple[str, str, str, date], datetime],
    dict[tuple[str, str, str, date], list[str]],
]:
    """タイムライン加工イベントを (依頼NO, 工程名, 機械名, 暦日) で集約する。"""
    qty_by: dict[tuple[str, str, str, date], float] = defaultdict(float)
    bound_min: dict[tuple[str, str, str, date], datetime] = {}
    bound_max: dict[tuple[str, str, str, date], datetime] = {}
    member_ops: dict[tuple[str, str, str, date], list[str]] = defaultdict(list)
    for ev in timeline_events or []:
        if not _is_machining_timeline_event(ev):
            continue
        tid = str(ev.get("task_id") or "").strip()
        eq = str(ev.get("machine") or "").strip()
        if not tid or not eq:
            continue
        cd = _timeline_event_calendar_date(ev)
        if cd is None:
            continue
        qty = _dispatch_table_event_qty_m(ev)
        if qty <= 1e-18:
            continue
        t = _resolve_task_dict_for_timeline_line(tid, eq, sorted_tasks_for_result)
        proc = str(t.get("machine") or "").strip() if t else ""
        mach = str(t.get("machine_name") or "").strip() if t else ""
        if not proc or not mach:
            continue
        key = (tid, proc, mach, cd)
        qty_by[key] += float(qty)
        op_raw = " ".join(str(ev.get("op") or "").split()).strip()
        if op_raw:
            lst = member_ops[key]
            if op_raw not in lst:
                lst.append(op_raw)
        _ml = ev.get("member_labels")
        if isinstance(_ml, (list, tuple)):
            for _raw in _ml:
                _lab = " ".join(str(_raw or "").split()).strip()
                if _lab:
                    _lst = member_ops[key]
                    if _lab not in _lst:
                        _lst.append(_lab)
        st0, ed0 = _timeline_event_start_end_dt(ev)
        if st0 is not None:
            prev = bound_min.get(key)
            if prev is None or st0 < prev:
                bound_min[key] = st0
        if ed0 is not None:
            prev_m = bound_max.get(key)
            if prev_m is None or ed0 > prev_m:
                bound_max[key] = ed0
    return dict(qty_by), bound_min, bound_max, dict(member_ops)
def _timeline_identity_planned_m_on_day(
    qty_by: dict[tuple[str, str, str, date], float],
    tid: str,
    proc: str,
    mach: str,
    day: date | None,
) -> float:
    if day is None:
        return 0.0
    try:
        return float(qty_by.get((tid, proc, mach, day), 0.0) or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _dispatch_table_identity_total_planned_m(
    records: list[dict], tid: str, proc: str, mach: str
) -> float:
    """同一 (依頼NO, 工程, 機械名) の結果_配台表行の当日配台数量合計。"""
    ident_key = (
        planning_task_id_str_from_scalar(tid),
        str(proc or "").strip(),
        str(mach or "").strip(),
    )
    total = 0.0
    for row in records:
        ident, _, qty = _dispatch_table_row_identity_and_date_key(row)
        if ident == ident_key:
            total += float(qty or 0.0)
    return total


def _dispatch_table_identity_planned_m_on_day(
    records: list[dict], tid: str, proc: str, mach: str, day: date | None
) -> float:
    if day is None:
        return 0.0
    ident_key = (
        planning_task_id_str_from_scalar(tid),
        str(proc or "").strip(),
        str(mach or "").strip(),
    )
    total = 0.0
    for row in records:
        ident, dispatch_day, qty = _dispatch_table_row_identity_and_date_key(row)
        if ident == ident_key and dispatch_day == day:
            total += float(qty or 0.0)
    return _sanitize_dispatch_qty_m(total)


def _dispatch_table_fill_row_timeline_meta_from_bounds(
    row: dict,
    *,
    bound_min: dict[tuple[str, str, str, date], datetime],
    bound_max: dict[tuple[str, str, str, date], datetime],
    member_ops: dict[tuple[str, str, str, date], list[str]],
) -> bool:
    """加工開始日時が空の行へ、同一 (依頼NO, 工程, 機械, 配台日) のタイムライン meta を載せる。"""
    if _interactive_row_has_timeline_meta(row):
        return False
    tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
    proc = _interactive_norm_cell(row.get(TASK_COL_MACHINE))
    mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
    dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
    if not tid or not proc or not mach or dd is None:
        return False
    key = (tid, proc, mach, dd)
    st0 = bound_min.get(key)
    ed0 = bound_max.get(key)
    ops = member_ops.get(key, [])
    if st0 is None and ed0 is None and not ops:
        return False
    if st0 is not None:
        row["加工開始日時"] = _fmt_dispatch_table_datetime(st0)
    if ed0 is not None:
        row["加工終了日時"] = _fmt_dispatch_table_datetime(ed0)
    if ops:
        row["メンバー名"] = "、".join(ops)
    return True
def _apply_dispatch_table_timeline_meta_from_events(
    df_dispatch: pd.DataFrame,
    timeline_events: list | None,
    sorted_tasks_for_result: list | None,
) -> pd.DataFrame:
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return df_dispatch
    _qty_by, bound_min, bound_max, member_ops = (
        _aggregate_timeline_dispatch_meta_by_identity_date(
            timeline_events, sorted_tasks_for_result
        )
    )
    if not bound_min and not bound_max and not member_ops:
        return df_dispatch
    records = df_dispatch.to_dict(orient="records")
    patched = 0
    for r in records:
        if _dispatch_table_fill_row_timeline_meta_from_bounds(
            r,
            bound_min=bound_min,
            bound_max=bound_max,
            member_ops=member_ops,
        ):
            patched += 1
    if patched <= 0:
        return df_dispatch
    logging.info(
        "結果_配台表: タイムライン meta を %s 行へ反映しました（加工開始/終了日時・メンバー名）。",
        patched,
    )
    return pd.DataFrame(records, columns=list(df_dispatch.columns))
def _resolve_dispatch_table_src_row_for_plan(
    tid: str,
    proc: str,
    plan_row,
    src_lookup3: dict,
    src_lookup2: dict,
):
    od_key = ""
    try:
        if plan_row is not None and hasattr(plan_row, "index") and "受注日" in plan_row.index:
            _v = _planning_df_cell_scalar(plan_row, "受注日")
            if _v is not None and not (isinstance(_v, float) and pd.isna(_v)):
                od_key = str(_v).strip()
    except Exception:
        od_key = ""
    if od_key:
        try:
            ts = pd.to_datetime(od_key, errors="coerce")
            if not pd.isna(ts) and isinstance(ts, pd.Timestamp):
                od_key = ts.to_pydatetime().date().strftime("%Y/%m/%d")
        except Exception:
            pass
    src_row = None
    if tid and proc:
        if od_key:
            src_row = src_lookup3.get((tid, proc, od_key))
        if src_row is None:
            cand = src_lookup2.get((tid, proc))
            if isinstance(cand, list):
                best = None
                best_od = ""
                for r0 in cand:
                    od0 = ""
                    try:
                        if hasattr(r0, "index") and "受注日" in r0.index:
                            od0 = _norm_ymd(_planning_df_cell_scalar(r0, "受注日"))
                    except Exception:
                        od0 = ""
                    if best is None:
                        best, best_od = r0, od0
                    elif od0 and (not best_od or od0 < best_od):
                        best, best_od = r0, od0
                src_row = best
            else:
                src_row = cand
    return src_row
def _plan_row_stub_dispatch_date(plan_row) -> date | None:
    for col in ("加工開始日", TASK_COL_RAW_INPUT_DATE, "指定納期", "回答納期"):
        try:
            if hasattr(plan_row, "index") and col in plan_row.index:
                d = parse_optional_date(_planning_df_cell_scalar(plan_row, col))
                if isinstance(d, date):
                    return d
        except Exception:
            continue
    return None
def _dispatch_table_row_identity_and_date_key(
    row: dict, *, qty_eps: float = 1e-12
) -> tuple[tuple[str, str, str], date | None, float]:
    ident = _dispatch_table_row_identity_key(row)
    dd = None
    try:
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
    except Exception:
        dd = None
    if dd is None:
        dd = parse_optional_date(row.get("配台日"))
    qty = 0.0
    try:
        qty = float(row.get("当日配台数量") or 0)
    except (TypeError, ValueError):
        qty = 0.0
    if qty <= qty_eps:
        qty = 0.0
    return ident, dd, qty
def _first_working_day_strictly_after(
    run_date: date, working_days: list[date] | None
) -> date:
    """run_date より後の最初の稼働日。無ければ暦日 +1 日（従来フォールバック）。"""
    if working_days:
        for d in working_days:
            if d > run_date:
                return d
    return run_date + timedelta(days=1)


def _first_working_day_on_or_after(
    run_date: date, working_days: list[date] | None
) -> date:
    """run_date 以降の最初の稼働日。無ければ run_date（従来フォールバック）。"""
    if working_days:
        for d in working_days:
            if d >= run_date:
                return d
    return run_date


def _stage2_dialog_target_plan_day(
    run_date: date | None,
    working_days: list[date] | None,
    *,
    skip_today: bool,
) -> date | None:
    """
    段階2直前ダイアログ由来の「翌日配台」追補・アラジン除外適用日。

    skip_today ON 時は run_date 自体が計画開始日のため +1 稼働日しない。
    いずれも run_date が非稼働なら直後の稼働日へ寄せる（土日休み工場で土曜に載らない）。
    """
    if run_date is None:
        return None
    if skip_today:
        return _first_working_day_on_or_after(run_date, working_days)
    return _first_working_day_strictly_after(run_date, working_days)


def _safe_plan_dispatch_remaining_m(plan_ref) -> float:
    """結果_配台表追補用: 未加工検証で落ちない残量 (m)。"""
    try:
        rem, _, _, _ = _plan_row_dispatch_qty_metrics(plan_ref)
        return max(0.0, float(rem))
    except Exception:
        try:
            qty_conv = parse_float_safe(
                _planning_df_cell_scalar(plan_ref, TASK_COL_QTY), 0.0
            )
            actual = parse_float_safe(
                _planning_df_cell_scalar(plan_ref, TASK_COL_ACTUAL_DONE), 0.0
            )
            return max(0.0, qty_conv - actual)
        except Exception:
            return 0.0


def _resolve_in_progress_aladdin_today_shortfall_m(
    ov_key: str,
    next_day_m: float,
    plan_ref,
    shortfall_overrides: dict[str, float],
) -> float:
    if ov_key in shortfall_overrides:
        return _sanitize_dispatch_qty_m(float(shortfall_overrides[ov_key]))
    if plan_ref is None or next_day_m <= 1e-12:
        return 0.0
    rem, _, _, _ = _plan_row_dispatch_qty_metrics(plan_ref)
    return _sanitize_dispatch_qty_m(max(0.0, float(rem) - float(next_day_m)))


def append_in_progress_next_day_dialog_rows_to_dispatch_table(
    df_dispatch: pd.DataFrame,
    tasks_df,
    df_src,
    run_date: date | None,
    working_days: list[date] | None = None,
    *,
    calendar_today: date | None = None,
    timeline_events: list | None = None,
    sorted_tasks_for_result: list | None = None,
) -> pd.DataFrame:
    """
    段階2直前ダイアログ（翌日配台量 JSON）の正の m を結果_配台表へ載せる。

    タイムラインに載らなかった加工途中タスクも、手動修正タブで翌日目標量を編集できる。
    配台日は run_date より後の最初の稼働日（勤怠で is_working の日）。無いときのみ暦日 +1。
    アラジン当日完了前提分（shortfall）は、skip_today OFF なら
    calendar_today（SKIP_TODAY 前の暦日）へ追補し、ON なら計画開始日の数量へ合算する。
    """
    overrides = _load_stage2_in_progress_next_day_dispatch_overrides()
    shortfall_overrides = _load_stage2_in_progress_aladdin_today_shortfall_overrides()
    if not overrides or run_date is None or tasks_df is None or getattr(
        tasks_df, "empty", True
    ):
        return df_dispatch

    skip_today = _stage2_truthy_env("PM_AI_STAGE2_SKIP_TODAY_DISPATCH")
    next_day = _stage2_dialog_target_plan_day(
        run_date, working_days, skip_today=skip_today
    )
    if next_day is None:
        return df_dispatch
    if next_day != run_date + timedelta(days=1):
        logging.info(
            "結果_配台表: 加工途中・翌日配台の配台日を稼働日に合わせました（%s → %s）。",
            (run_date + timedelta(days=1)).isoformat(),
            next_day.isoformat(),
        )
    base_cols = list(RESULT_DISPATCH_TABLE_STATIC_HEADERS) + [
        "配台日",
        "当日配台数量",
    ]
    if df_dispatch is not None and not getattr(df_dispatch, "empty", True):
        out_cols = list(df_dispatch.columns)
        out_records = df_dispatch.to_dict(orient="records")
        filled: set[tuple[str, str, str, date]] = set()
        for r in out_records:
            ident, dd, qty = _dispatch_table_row_identity_and_date_key(r)
            if dd == next_day and qty > 1e-12:
                filled.add((ident[0], ident[1], ident[2], dd))
    else:
        out_cols = base_cols
        out_records = []
        filled = set()

    plan_lookup = _build_plan_input_row_lookup_for_dispatch_table(tasks_df)
    try:
        src_lookup3, src_lookup2 = _build_source_task_row_lookups_for_dispatch_table(
            df_src
        )
    except Exception:
        src_lookup3, src_lookup2 = {}, {}

    tl_qty_by, tl_bound_min, tl_bound_max, tl_member_ops = (
        _aggregate_timeline_dispatch_meta_by_identity_date(
            timeline_events, sorted_tasks_for_result
        )
    )

    added = 0
    added_shortfall = 0
    skipped_timeline_covered = 0
    today_plan_day = calendar_today if calendar_today is not None else run_date
    for ov_key, meters in overrides.items():
        try:
            m = _sanitize_dispatch_qty_m(float(meters))
        except (TypeError, ValueError):
            m = 0.0
        if m <= 1e-12:
            continue
        parts = str(ov_key).split("\x1e")
        if len(parts) != 3:
            continue
        tid, proc, mach = (
            planning_task_id_str_from_scalar(parts[0]),
            str(parts[1] or "").strip(),
            str(parts[2] or "").strip(),
        )
        if not tid or not proc or not mach:
            continue
        plan_ref = plan_lookup.get((tid, proc))
        if plan_ref is None:
            for _, prow in tasks_df.iterrows():
                if planning_task_id_str_from_plan_row(prow) != tid:
                    continue
                if (
                    str(_planning_df_cell_scalar(prow, TASK_COL_MACHINE) or "").strip()
                    != proc
                ):
                    continue
                if (
                    str(
                        _planning_df_cell_scalar(prow, TASK_COL_MACHINE_NAME) or ""
                    ).strip()
                    != mach
                ):
                    continue
                plan_ref = prow
                break
        if plan_ref is None:
            logging.warning(
                "結果_配台表: 加工途中・翌日配台 JSON の行を計画入力から解決できませんでした "
                "(依頼NO=%s 工程=%s 機械名=%s)",
                tid,
                _log_plain_label(proc),
                _log_plain_label(mach),
            )
            continue
        src_row = _resolve_dispatch_table_src_row_for_plan(
            tid, proc, plan_ref, src_lookup3, src_lookup2
        )

        def _build_dialog_row(dispatch_day: date, qty_m: float) -> dict:
            row_out: dict = {}
            for h in RESULT_DISPATCH_TABLE_STATIC_HEADERS:
                row_out[h] = _dispatch_table_cell_from_sources(
                    src_row=src_row,
                    plan_row=plan_ref,
                    task_dict=None,
                    col_name=h,
                )
            if not str(row_out.get(TASK_COL_TASK_ID) or "").strip():
                row_out[TASK_COL_TASK_ID] = tid
            if not str(row_out.get(TASK_COL_MACHINE) or "").strip():
                row_out[TASK_COL_MACHINE] = proc
            if not str(row_out.get(TASK_COL_MACHINE_NAME) or "").strip():
                row_out[TASK_COL_MACHINE_NAME] = mach
            row_out["配台日"] = dispatch_day
            row_out["当日配台数量"] = float(qty_m)
            if "実配台数量" in out_cols:
                row_out["実配台数量"] = 0.0
            _dispatch_table_fill_row_timeline_meta_from_bounds(
                row_out,
                bound_min=tl_bound_min,
                bound_max=tl_bound_max,
                member_ops=tl_member_ops,
            )
            return row_out

        shortfall_m = _resolve_in_progress_aladdin_today_shortfall_m(
            ov_key, m, plan_ref, shortfall_overrides
        )
        if skip_today and shortfall_m > 1e-12:
            m = _sanitize_dispatch_qty_m(float(m) + float(shortfall_m))
        if (
            shortfall_m > 1e-12
            and not skip_today
            and today_plan_day is not None
            and today_plan_day != next_day
        ):
            table_today = _dispatch_table_identity_planned_m_on_day(
                out_records, tid, proc, mach, today_plan_day
            )
            timeline_today = _timeline_identity_planned_m_on_day(
                tl_qty_by, tid, proc, mach, today_plan_day
            )
            covered_m = max(table_today, timeline_today)
            uncovered_m = _sanitize_dispatch_qty_m(
                max(0.0, float(shortfall_m) - float(covered_m))
            )
            if uncovered_m > 1e-12:
                out_records.append(_build_dialog_row(today_plan_day, uncovered_m))
                filled.add((tid, proc, mach, today_plan_day))
                added += 1
                added_shortfall += 1

        existing_total = _dispatch_table_identity_total_planned_m(
            out_records, tid, proc, mach
        )
        rem = _safe_plan_dispatch_remaining_m(plan_ref)
        remaining_to_plan = max(0.0, float(rem) - existing_total)
        if remaining_to_plan <= 1e-12:
            skipped_timeline_covered += 1
            continue
        add_m = _sanitize_dispatch_qty_m(min(m, remaining_to_plan))
        if add_m <= 1e-12:
            skipped_timeline_covered += 1
            continue

        if (tid, proc, mach, next_day) in filled:
            if skip_today and shortfall_m > 1e-12:
                for existing_row in out_records:
                    existing_ident, existing_day, existing_qty = (
                        _dispatch_table_row_identity_and_date_key(existing_row)
                    )
                    if existing_ident != (tid, proc, mach) or existing_day != next_day:
                        continue
                    topped_up_qty = _sanitize_dispatch_qty_m(
                        float(existing_qty) + float(add_m)
                    )
                    existing_row["当日配台数量"] = topped_up_qty
                    added += 1
                    break
            continue
        if _timeline_identity_planned_m_on_day(tl_qty_by, tid, proc, mach, next_day) + 1e-9 >= add_m:
            skipped_timeline_covered += 1
            continue
        if (
            run_date is not None
            and _timeline_identity_planned_m_on_day(tl_qty_by, tid, proc, mach, run_date)
            + 1e-9
            >= add_m
        ):
            logging.info(
                "段階2: 加工途中・翌日追補を省略（run_date=%s にタイムライン配台済 "
                "依頼NO=%s 工程=%s 機械名=%s → %s m）",
                run_date.isoformat(),
                tid,
                _log_plain_label(proc),
                _log_plain_label(mach),
                add_m,
            )
            skipped_timeline_covered += 1
            continue
        out_records.append(_build_dialog_row(next_day, add_m))
        filled.add((tid, proc, mach, next_day))
        added += 1

    if added <= 0 and skipped_timeline_covered <= 0:
        return df_dispatch
    if added > 0:
        logging.info(
            "結果_配台表: 加工途中・翌日配台ダイアログから %s 行を追補しました（配台日=%s、うちアラジン当日分=%s）。",
            added,
            next_day.isoformat(),
            added_shortfall,
        )
    if skipped_timeline_covered > 0:
        logging.info(
            "結果_配台表: 加工途中・翌日追補を %s 件省略しました（タイムライン配台済）。",
            skipped_timeline_covered,
        )
    return pd.DataFrame(out_records, columns=out_cols)
def append_plan_input_rows_missing_from_dispatch_table(
    df_dispatch: pd.DataFrame,
    tasks_df,
    df_src,
) -> pd.DataFrame:
    """
    タイムライン未配台（配台不可・未割当）の計画入力行を結果_配台表に載せ、手動修正タブで編集可能にする。

    既存行は (依頼NO, 工程名, 機械名) で同一視。配台日・当日配台数量は 0（未配台プレースホルダ）。
    """
    if tasks_df is None or getattr(tasks_df, "empty", True):
        return df_dispatch
    base_cols = list(RESULT_DISPATCH_TABLE_STATIC_HEADERS) + [
        "配台日",
        "当日配台数量",
    ]
    if df_dispatch is not None and not getattr(df_dispatch, "empty", True):
        out_cols = list(df_dispatch.columns)
        existing_keys = {
            _dispatch_table_row_identity_key(r)
            for r in df_dispatch.to_dict(orient="records")
        }
        out_records = df_dispatch.to_dict(orient="records")
    else:
        out_cols = base_cols
        existing_keys = set()
        out_records = []

    plan_lookup = _build_plan_input_row_lookup_for_dispatch_table(tasks_df)
    try:
        src_lookup3, src_lookup2 = _build_source_task_row_lookups_for_dispatch_table(
            df_src
        )
    except Exception:
        src_lookup3, src_lookup2 = {}, {}
    in_progress_next_day_m = _load_stage2_in_progress_next_day_dispatch_overrides()

    added = 0
    for _, plan_row in tasks_df.iterrows():
        try:
            if _plan_row_exclude_from_assignment(plan_row):
                continue
        except Exception:
            continue
        tid = planning_task_id_str_from_plan_row(plan_row)
        proc = str(_planning_df_cell_scalar(plan_row, TASK_COL_MACHINE) or "").strip()
        mach = str(
            _planning_df_cell_scalar(plan_row, TASK_COL_MACHINE_NAME) or ""
        ).strip()
        if not tid or not proc or not mach:
            continue
        if in_progress_next_day_m:
            ov_key = _stage2_in_progress_next_day_dispatch_key(tid, proc, mach)
            try:
                ov_m = float(in_progress_next_day_m.get(ov_key, -1.0))
            except (TypeError, ValueError):
                ov_m = -1.0
            if ov_key in in_progress_next_day_m and ov_m <= 1e-12:
                continue
        key = (tid, proc, mach)
        if key in existing_keys:
            continue
        existing_keys.add(key)
        plan_ref = plan_lookup.get((tid, proc))
        if plan_ref is None:
            plan_ref = plan_row
        src_row = _resolve_dispatch_table_src_row_for_plan(
            tid, proc, plan_ref, src_lookup3, src_lookup2
        )
        r: dict = {}
        for h in RESULT_DISPATCH_TABLE_STATIC_HEADERS:
            r[h] = _dispatch_table_cell_from_sources(
                src_row=src_row,
                plan_row=plan_ref,
                task_dict=None,
                col_name=h,
            )
        if not str(r.get(TASK_COL_TASK_ID) or "").strip():
            r[TASK_COL_TASK_ID] = tid
        if not str(r.get(TASK_COL_MACHINE) or "").strip():
            r[TASK_COL_MACHINE] = proc
        if not str(r.get(TASK_COL_MACHINE_NAME) or "").strip():
            r[TASK_COL_MACHINE_NAME] = mach
        stub_day = _plan_row_stub_dispatch_date(plan_ref)
        r["配台日"] = stub_day if stub_day is not None else ""
        r["当日配台数量"] = 0.0
        if "実配台数量" in out_cols:
            r["実配台数量"] = 0.0
        out_records.append(r)
        added += 1

    if added <= 0:
        return df_dispatch
    logging.info(
        "結果_配台表: 計画入力のみ存在する行を %s 件追補しました（未配台・手動修正用）。",
        added,
    )
    return pd.DataFrame(out_records, columns=out_cols)
def _collect_result_dispatch_table_output_dirs(
    primary_dir: str, plan_input_wb: str
) -> list[str]:
    """JavaFX が読む候補と段階2の主出力先へ同内容 JSON を書くためのフォルダ一覧。"""
    ordered: list[str] = []
    seen: set[str] = set()

    def _add(d: str | None) -> None:
        if not d:
            return
        ad = os.path.abspath(d)
        if ad in seen:
            return
        seen.add(ad)
        ordered.append(ad)

    _add(primary_dir)
    for wb in (plan_input_wb or "", (os.environ.get("PM_AI_PLAN_INPUT_PATH") or "").strip()):
        _add(resolve_result_dispatch_table_output_dir(wb))
    _add(resolve_result_dispatch_table_output_dir(""))
    return ordered
def _write_dispatch_table_standalone_json_to_resolved_dirs(
    df_dispatch: pd.DataFrame, primary_dir: str, plan_input_wb: str
) -> str | None:
    primary_written: str | None = None
    for d in _collect_result_dispatch_table_output_dirs(primary_dir, plan_input_wb):
        p = _write_dispatch_table_standalone_json(df_dispatch, d)
        if p and primary_written is None:
            primary_written = p
    return primary_written
def build_result_dispatch_table_dataframe(
    timeline_events: list | None,
    sorted_tasks_for_result: list | None,
    tasks_df,
    df_src,
) -> pd.DataFrame:
    """
    結果_配台表用 DataFrame（1行＝1タスク行×1暦日の配台、当日の数量はメートル合計）。

    timeline_events の加工イベントを (依頼NO, 設備ライン, 暦日) で集約し、
    計画入力 tasks_df の同一 (依頼NO, 工程名) 行から静的列を埋める。
    """
    cols = list(RESULT_DISPATCH_TABLE_STATIC_HEADERS) + ["配台日", "当日配台数量"]
    if not timeline_events:
        return pd.DataFrame(columns=cols)
    agg: dict[tuple[str, str, date], float] = defaultdict(float)
    bound_min: dict[tuple[str, str, date], datetime] = {}
    bound_max: dict[tuple[str, str, date], datetime] = {}
    member_ops: dict[tuple[str, str, date], list[str]] = defaultdict(list)
    for ev in timeline_events:
        if not _is_machining_timeline_event(ev):
            continue
        tid = str(ev.get("task_id") or "").strip()
        if not tid:
            continue
        eq = str(ev.get("machine") or "").strip()
        if not eq:
            continue
        cd = _timeline_event_calendar_date(ev)
        if cd is None:
            continue
        qty = _dispatch_table_event_qty_m(ev)
        if qty <= 1e-18:
            continue
        key = (tid, eq, cd)
        agg[key] += float(qty)
        op_raw = " ".join(str(ev.get("op") or "").split()).strip()
        if op_raw:
            lst = member_ops[key]
            if op_raw not in lst:
                lst.append(op_raw)
        _ml = ev.get("member_labels")
        if isinstance(_ml, (list, tuple)):
            for _raw in _ml:
                _lab = " ".join(str(_raw or "").split()).strip()
                if _lab:
                    _lst = member_ops[key]
                    if _lab not in _lst:
                        _lst.append(_lab)
        st0, ed0 = _timeline_event_start_end_dt(ev)
        if st0 is not None:
            prev = bound_min.get(key)
            if prev is None or st0 < prev:
                bound_min[key] = st0
        if ed0 is not None:
            prev_m = bound_max.get(key)
            if prev_m is None or ed0 > prev_m:
                bound_max[key] = ed0
    if not agg:
        return pd.DataFrame(columns=cols)
    _tl_exp_dispatch = _expand_timeline_events_for_equipment_grid(timeline_events or [])
    _unify_sub_dispatch = _equipment_schedule_unified_sub_string_map(_tl_exp_dispatch)
    plan_lookup = _build_plan_input_row_lookup_for_dispatch_table(tasks_df)
    src_lookup3, src_lookup2 = _build_source_task_row_lookups_for_dispatch_table(df_src)
    rows: list[dict] = []
    for (tid_k, eq_k, day_k), qty_sum in sorted(agg.items()):
        t = _resolve_task_dict_for_timeline_line(tid_k, eq_k, sorted_tasks_for_result)
        proc = str(t.get("machine") or "").strip() if t else ""
        plan_row = plan_lookup.get((tid_k, proc)) if (tid_k and proc) else None
        # 受注日もキーに含めて加工計画DATA行を選ぶ（取れない場合は空キーにフォールバック）
        od_key = ""
        try:
            if plan_row is not None and hasattr(plan_row, "index") and "受注日" in plan_row.index:
                _v = _planning_df_cell_scalar(plan_row, "受注日")
                if _v is not None and not (isinstance(_v, float) and pd.isna(_v)):
                    od_key = str(_v).strip()
        except Exception:
            od_key = ""
        if od_key:
            try:
                ts = pd.to_datetime(od_key, errors="coerce")
                if not pd.isna(ts) and isinstance(ts, pd.Timestamp):
                    od_key = ts.to_pydatetime().date().strftime("%Y/%m/%d")
            except Exception:
                pass
        src_row = None
        if tid_k and proc:
            if od_key:
                src_row = src_lookup3.get((tid_k, proc, od_key))
            if src_row is None:
                cand = src_lookup2.get((tid_k, proc))
                if isinstance(cand, list):
                    # 複数候補: 受注日が最小（古い）を暫定採用（必要なら別基準へ）
                    best = None
                    best_od = ""
                    for r0 in cand:
                        od0 = ""
                        try:
                            if hasattr(r0, "index") and "受注日" in r0.index:
                                od0 = _norm_ymd(_planning_df_cell_scalar(r0, "受注日"))
                        except Exception:
                            od0 = ""
                        if best is None:
                            best, best_od = r0, od0
                        else:
                            if od0 and (not best_od or od0 < best_od):
                                best, best_od = r0, od0
                    src_row = best
                else:
                    src_row = cand
        r: dict = {}
        for h in RESULT_DISPATCH_TABLE_STATIC_HEADERS:
            r[h] = _dispatch_table_cell_from_sources(
                src_row=src_row, plan_row=plan_row, task_dict=t, col_name=h
            )
        if not str(r.get(TASK_COL_TASK_ID) or "").strip():
            r[TASK_COL_TASK_ID] = tid_k
        # 工程名・機械名は timeline/task_queue 由来が最も確実なので最後に補完
        if not str(r.get(TASK_COL_MACHINE) or "").strip() and proc:
            r[TASK_COL_MACHINE] = proc
        if not str(r.get(TASK_COL_MACHINE_NAME) or "").strip():
            r[TASK_COL_MACHINE_NAME] = (t.get("machine_name") if t else "") or ""
        row_key = (tid_k, eq_k, day_k)
        r["加工開始日時"] = _fmt_dispatch_table_datetime(bound_min.get(row_key))
        r["加工終了日時"] = _fmt_dispatch_table_datetime(bound_max.get(row_key))
        r["メンバー名"] = _format_dispatch_table_member_like_equipment_schedule(
            tid_k,
            eq_k,
            day_k,
            _tl_exp_dispatch,
            _unify_sub_dispatch,
            member_ops.get(row_key, []),
        )
        r["配台日"] = day_k
        r["当日配台数量"] = _sanitize_dispatch_qty_m(float(qty_sum))
        rows.append(r)
    return pd.DataFrame(rows, columns=cols)
def _interactive_norm_cell(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return unicodedata.normalize("NFKC", str(v).strip())
def _interactive_dispatch_target_process_key(v) -> str:
    """段階3の目標・キャップキー用。工程名を NFKC のうえルール照合と同一に正規化する。"""
    return _normalize_process_name_for_rule_match(_interactive_norm_cell(v))
def _interactive_dispatch_resolve_cap_key(
    *,
    interactive_dispatch_targets: dict,
    interactive_trial_meters_done: dict | None,
    tid: str,
    proc: str,
    mach: str,
    current_date,
) -> tuple | None:
    """
    JSON の配台日キーと実作業暦日がずれるとき、日次キャップ・meters_done の格納キーを決める。

    - 暦日 current_date と一致するキーが targets にあればそれを使う（呼び出し側で先に判定可）。
    - 無い場合は同一 (依頼NO, 工程, 機械) で current_date 以前の配台日キーを新しい順に見て、
      目標メートルに未達の最初のキーを使う（前日の窓に収まらず翌暦日にまたいだ割付を同一 JSON 日に計上する）。
    - いずれの過去キーも満たしている場合は、そのブロックで最も新しい過去キーを返す（残余の計上先）。
    """
    if isinstance(current_date, datetime):
        cur_d = current_date.date()
    elif isinstance(current_date, date):
        cur_d = current_date
    else:
        return None
    past: list[tuple[date, tuple]] = []
    for kk in interactive_dispatch_targets:
        if not isinstance(kk, tuple) or len(kk) != 4:
            continue
        if kk[0] != tid or kk[1] != proc or kk[2] != mach:
            continue
        d = kk[3]
        if not isinstance(d, date):
            continue
        if d <= cur_d:
            past.append((d, kk))
    if not past:
        return None
    past.sort(key=lambda x: x[0], reverse=True)
    done_dict = interactive_trial_meters_done or {}
    for _d, kk in past:
        try:
            cap_m = float(interactive_dispatch_targets[kk])
        except (TypeError, ValueError):
            cap_m = 0.0
        try:
            done_m = float(done_dict.get(kk, 0.0))
        except (TypeError, ValueError):
            done_m = 0.0
        if done_m < cap_m - 1e-5:
            return kk
    return past[0][1]
def _interactive_earliest_positive_target_date(
    interactive_dispatch_targets: dict | None,
    tid: str,
    proc: str,
    mach: str,
) -> date | None:
    """同一 (依頼NO, 工程, 機械) で正の JSON 目標がある最も早い配台暦日。"""
    if not interactive_dispatch_targets:
        return None
    tid_n = _interactive_norm_cell(tid) or ""
    proc_n = _interactive_dispatch_target_process_key(proc)
    mach_n = _interactive_norm_cell(mach) or ""
    if not tid_n or not mach_n:
        return None
    min_dd: date | None = None
    for kk, val in interactive_dispatch_targets.items():
        if not isinstance(kk, tuple) or len(kk) != 4:
            continue
        if kk[0] != tid_n or kk[1] != proc_n or kk[2] != mach_n:
            continue
        try:
            qty = float(val or 0.0)
        except (TypeError, ValueError):
            qty = 0.0
        if qty <= 1e-9:
            continue
        d = kk[3]
        if not isinstance(d, date):
            continue
        if min_dd is None or d < min_dd:
            min_dd = d
    return min_dd
def _interactive_cap_schedule_blocked_before_earliest_target(
    interactive_dispatch_targets: dict | None,
    tid: str,
    proc: str,
    mach: str,
    current_date,
) -> bool:
    """
    JSON 目標の最古暦日より前には暦日キャップ割当を開始しない。
    6/5 目標の接続が 6/2 に載り (段階3前) と (段階3後) がずれるのを防ぐ。
    """
    if not _interactive_dispatch_cap_enforced_in_schedule_loop():
        return False
    if not interactive_dispatch_targets:
        return False
    min_d = _interactive_earliest_positive_target_date(
        interactive_dispatch_targets, tid, proc, mach
    )
    if min_d is None:
        return False
    if isinstance(current_date, datetime):
        cur_d = current_date.date()
    elif isinstance(current_date, date):
        cur_d = current_date
    else:
        return False
    return cur_d < min_d
def _interactive_fallback_meter_target_key_for_recompute(
    tid: str,
    proc: str,
    mach: str,
    d_ev: date,
    want: set,
) -> tuple | None:
    """
    タイムライン暦日 d_ev に一致する targets キーが無いとき、(依頼,工程,機械) が一致する
    配台日キーへ寄せる（暦日 < 配台日のみ JSON があるケース等）。
    """
    cands: list[tuple] = [
        kk
        for kk in want
        if isinstance(kk, tuple)
        and len(kk) == 4
        and kk[0] == tid
        and kk[1] == proc
        and kk[2] == mach
        and isinstance(kk[3], date)
    ]
    if not cands:
        return None
    fut = [kk for kk in cands if kk[3] >= d_ev]
    if fut:
        return min(fut, key=lambda x: x[3])
    return max(cands, key=lambda x: x[3])
def _interactive_timeline_event_calendar_date(ev: dict) -> date | None:
    d = ev.get("date") if isinstance(ev, dict) else None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    if isinstance(d, str) and d.strip():
        ts = pd.to_datetime(d.strip(), errors="coerce")
        if not pd.isna(ts) and hasattr(ts, "date"):
            try:
                return ts.date()
            except Exception:
                return None
    return None
def _interactive_parse_dispatch_date_cell(val) -> date | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    ts = pd.to_datetime(val, errors="coerce")
    if not pd.isna(ts) and hasattr(ts, "date"):
        try:
            return ts.date()
        except Exception:
            pass
    return parse_optional_date(val)
def fill_interactive_result_dispatch_json_rows_from_planning_sources(
    json_rows: list,
    tasks_df: "pd.DataFrame",
    df_src: "pd.DataFrame | None",
) -> int:
    """
    結果_配台表 JSON の各行で空の静的列を、計画入力 tasks_df と加工計画DATA df_src（段階2と同一）から補完する。
    「配台日」「当日配台数量」は編集対象のため上書きしない。
    """
    if not json_rows or tasks_df is None or getattr(tasks_df, "empty", True):
        return 0
    plan_lookup = _build_plan_input_row_lookup_for_dispatch_table(tasks_df)
    try:
        src_lookup3, src_lookup2 = _build_source_task_row_lookups_for_dispatch_table(
            df_src
        )
    except Exception:
        src_lookup3, src_lookup2 = {}, {}
    skip_cols = frozenset({"配台日", "当日配台数量"})
    filled = 0
    for r in json_rows:
        if not isinstance(r, dict):
            continue
        tid = _interactive_norm_cell(r.get(TASK_COL_TASK_ID)) or _interactive_norm_cell(
            r.get("タスクID")
        )
        proc = _interactive_dispatch_target_process_key(r.get(TASK_COL_MACHINE))
        if not tid or not proc:
            continue
        plan_row = plan_lookup.get((tid, proc))
        od_key = ""
        try:
            if plan_row is not None and hasattr(plan_row, "index") and "受注日" in plan_row.index:
                _v = _planning_df_cell_scalar(plan_row, "受注日")
                if _v is not None and not (isinstance(_v, float) and pd.isna(_v)):
                    od_key = str(_v).strip()
        except Exception:
            od_key = ""
        if od_key:
            try:
                ts = pd.to_datetime(od_key, errors="coerce")
                if not pd.isna(ts) and isinstance(ts, pd.Timestamp):
                    od_key = ts.to_pydatetime().date().strftime("%Y/%m/%d")
            except Exception:
                pass
        j_od = _interactive_norm_cell(r.get("受注日"))
        if j_od:
            try:
                tsj = pd.to_datetime(j_od, errors="coerce")
                if not pd.isna(tsj) and isinstance(tsj, pd.Timestamp):
                    j_od = tsj.to_pydatetime().date().strftime("%Y/%m/%d")
            except Exception:
                pass
        src_row = None
        if tid and proc:
            if j_od:
                src_row = src_lookup3.get((tid, proc, j_od))
            if src_row is None and od_key:
                src_row = src_lookup3.get((tid, proc, od_key))
            if src_row is None:
                cand = src_lookup2.get((tid, proc))
                if isinstance(cand, list):
                    best = None
                    best_od = ""
                    for r0 in cand:
                        od0 = ""
                        try:
                            if hasattr(r0, "index") and "受注日" in r0.index:
                                od0 = _norm_ymd(_planning_df_cell_scalar(r0, "受注日"))
                        except Exception:
                            od0 = ""
                        if best is None:
                            best, best_od = r0, od0
                        else:
                            if od0 and (not best_od or od0 < best_od):
                                best, best_od = r0, od0
                    src_row = best
                else:
                    src_row = cand
        for h in RESULT_DISPATCH_TABLE_STATIC_HEADERS:
            if h in skip_cols:
                continue
            cur = r.get(h)
            if h in _RESULT_DISPATCH_PROCESSING_PLAN_ONLY_SRC_COLS:
                cell = _dispatch_table_cell_from_sources(
                    src_row=src_row, plan_row=plan_row, task_dict=None, col_name=h
                )
                new_s = (
                    ""
                    if cell is None
                    or (isinstance(cell, float) and pd.isna(cell))
                    or (isinstance(cell, str) and not cell.strip())
                    else str(cell).strip()
                )
                cur_s = "" if cur is None else str(cur).strip()
                if cur_s != new_s:
                    r[h] = new_s
                    filled += 1
                continue
            force_plan = h in _RESULT_DISPATCH_PLAN_INPUT_OVERRIDE_SRC_COLS
            if not force_plan and cur is not None and str(cur).strip() != "":
                continue
            cell = _dispatch_table_cell_from_sources(
                src_row=src_row, plan_row=plan_row, task_dict=None, col_name=h
            )
            if cell is None:
                continue
            if isinstance(cell, float) and pd.isna(cell):
                continue
            if isinstance(cell, str) and not cell.strip():
                continue
            if (
                not force_plan
                and cur is not None
                and str(cur).strip() == str(cell).strip()
            ):
                continue
            r[h] = cell
            filled += 1
    return filled
def merge_interactive_result_dispatch_json_into_tasks_df(
    tasks_df: "pd.DataFrame", json_rows: list
) -> tuple["pd.DataFrame", dict[tuple[str, str, str, date], float]]:
    """
    結果_配台表.json の rows から配台試行順番を tasks_df に反映し、
    (依頼NO, 工程名, 機械名, 配台日) ごとの目標数量（換算 m）を集約して返す。
    """
    if tasks_df is None or getattr(tasks_df, "empty", True):
        return tasks_df, {}
    df = tasks_df.copy()
    _dto = RESULT_TASK_COL_DISPATCH_TRIAL_ORDER
    if _dto not in df.columns:
        df[_dto] = ""
    else:
        # 計画入力（xlsx 等）では数値列 float64 になることがあり、文字列を代入すると pandas 2.x で失敗する
        df[_dto] = df[_dto].astype(object)
    order_map: dict[tuple[str, str, str], int] = {}
    targets: dict[tuple[str, str, str, date], float] = defaultdict(float)
    for r in json_rows or []:
        if not isinstance(r, dict):
            continue
        tid = _interactive_norm_cell(r.get(TASK_COL_TASK_ID)) or _interactive_norm_cell(
            r.get("タスクID")
        )
        proc = _interactive_dispatch_target_process_key(r.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(r.get(TASK_COL_MACHINE_NAME))
        dto_raw = r.get(RESULT_TASK_COL_DISPATCH_TRIAL_ORDER)
        try:
            dto_v = (
                int(float(str(dto_raw).replace(",", "").strip()))
                if dto_raw not in (None, "")
                else None
            )
        except (TypeError, ValueError):
            dto_v = None
        if dto_v is not None and tid:
            k3 = (tid, proc, mach)
            if k3 not in order_map or dto_v < order_map[k3]:
                order_map[k3] = dto_v
        dd = _interactive_parse_dispatch_date_cell(r.get("配台日"))
        qty_cell = r.get("当日配台数量")
        try:
            qty_v = (
                float(str(qty_cell).replace(",", "").strip())
                if qty_cell not in (None, "")
                else 0.0
            )
        except (TypeError, ValueError):
            qty_v = 0.0
        if dd is not None and tid and mach and qty_v > 1e-18:
            targets[(tid, proc, mach, dd)] += float(qty_v)
    for idx in df.index:
        row = df.loc[idx]
        tid = _interactive_norm_cell(planning_task_id_str_from_plan_row(row))
        proc = _interactive_dispatch_target_process_key(
            _planning_df_cell_scalar(row, TASK_COL_MACHINE)
        )
        mach = _interactive_norm_cell(_planning_df_cell_scalar(row, TASK_COL_MACHINE_NAME))
        k = (tid, proc, mach)
        if k in order_map:
            df.at[idx, _dto] = str(order_map[k])
    return df, dict(targets)
def _interactive_min_positive_dispatch_date_from_json_rows(
    json_rows: list,
    task_id: str,
    machine: str,
    machine_name: str,
) -> date | None:
    """
    段階3手動修正 JSON: 依頼NO×工程×機械名が一致し「当日配台数量」が正の行について、配台日の最奨日。

    タスクキューの ``start_date_req`` をこれ以上前に倒さない下限として使う（手動表に無い暦日では
    ``start_date_req <= current_date`` に入らず、当該日のフォーム探索に掛からない）。
    """
    tid_w = _interactive_norm_cell(task_id) or ""
    mach_w = _interactive_norm_cell(machine_name) or ""
    proc_w = _interactive_dispatch_target_process_key(machine)
    if not tid_w or not mach_w:
        return None
    min_dd: date | None = None
    for r in json_rows or []:
        if not isinstance(r, dict):
            continue
        tid = _interactive_norm_cell(r.get(TASK_COL_TASK_ID)) or _interactive_norm_cell(
            r.get("タスクID")
        )
        proc = _interactive_dispatch_target_process_key(r.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(r.get(TASK_COL_MACHINE_NAME))
        if tid != tid_w or proc != proc_w or mach != mach_w:
            continue
        dd = _interactive_parse_dispatch_date_cell(r.get("配台日"))
        qty_cell = r.get("当日配台数量")
        try:
            qty_v = (
                float(str(qty_cell).replace(",", "").strip())
                if qty_cell not in (None, "")
                else 0.0
            )
        except (TypeError, ValueError):
            qty_v = 0.0
        if dd is None or qty_v <= 1e-18:
            continue
        if min_dd is None or dd < min_dd:
            min_dd = dd
    return min_dd
def _interactive_aggregate_dispatch_targets_from_df(
    df_dispatch: pd.DataFrame,
) -> dict[tuple[str, str, date], float]:
    out: dict[tuple[str, str, date], float] = defaultdict(float)
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return dict(out)
    for _, row in df_dispatch.iterrows():
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        try:
            q = float(row.get("当日配台数量") or 0)
        except (TypeError, ValueError):
            q = 0.0
        if tid and mach and dd is not None:
            out[(tid, mach, dd)] += q
    return dict(out)
def _interactive_row_has_timeline_meta(row: dict) -> bool:
    """結果_配台表行にタイムライン由来の加工開始日時があるか。"""
    st_raw = row.get("加工開始日時")
    st = (
        ""
        if st_raw is None or (isinstance(st_raw, float) and pd.isna(st_raw))
        else str(st_raw).strip()
    )
    return bool(st and st.lower() not in ("nan", "nat"))
def _interactive_row_needs_dispatch_date_slide(
    row: dict,
    *,
    tid: str,
    proc: str,
    mach: str,
    plan_dd: date,
    plan_qty: float,
    meters_done: dict[tuple[str, str, str, date], float] | None = None,
    eps: float = 1e-9,
) -> bool:
    """
    配台日スライドが必要か。
    - 加工開始日時なし
    - 加工開始暦日 ≠ 配台日
    - 配台日キーの実績 m が目標未満で、より後の暦日に実績がある
    """
    if plan_qty <= eps:
        return False
    if _interactive_row_has_timeline_meta(row):
        st_day_s = _iso_date_from_dispatch_table_datetime_cell(row.get("加工開始日時"))
        if st_day_s:
            try:
                st_d = date.fromisoformat(st_day_s)
                if st_d == plan_dd:
                    # 計画暦日にタイムライン割付済み。plan>md の未達はスライド理由にしない。
                    return False
                return True
            except (TypeError, ValueError):
                return True
        return False
    if meters_done:
        plan_key = (tid, proc, mach, plan_dd)
        try:
            done_plan = float(meters_done.get(plan_key, 0.0) or 0.0)
        except (TypeError, ValueError):
            done_plan = 0.0
        if done_plan + eps >= plan_qty:
            return False
        if done_plan > eps:
            return False
    return True
def _stage3_attach_actual_dispatch_qty_from_timeline_plan_rows(
    df_dispatch: pd.DataFrame,
) -> pd.DataFrame:
    """段階3.0（編集JSON無し）: タイムライン行へ実配台数量列を載せ、JSON 正本を配台結果タブと揃える。"""
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return df_dispatch
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    plan_col = "当日配台数量"
    out = df_dispatch.copy()
    if actual_col not in out.columns:
        out[actual_col] = 0.0
    for pos in range(len(out)):
        row = out.iloc[pos].to_dict()
        if not _interactive_row_has_timeline_meta(row):
            continue
        try:
            plan_q = float(row.get(plan_col) or 0.0)
        except (TypeError, ValueError):
            plan_q = 0.0
        try:
            out.at[out.index[pos], actual_col] = plan_q
        except Exception:
            pass
    try:
        out[actual_col] = (
            pd.to_numeric(out[actual_col], errors="coerce").fillna(0.0).astype(float)
        )
    except Exception:
        pass
    return out
def _interactive_zero_actual_qty_without_timeline_meta(
    df: pd.DataFrame,
    *,
    preserve_meters_done: dict | None = None,
) -> pd.DataFrame:
    """
    タイムライン未割付（加工開始日時が空）の暦日行は実配台数量を 0 にする。
    cap_key 解決で別日実績が計画日キーへ載った見かけ上の「配台済」を防ぐ。
    段階3.5 floor があるキーは 0 クリアしない。
    """
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    if df is None or getattr(df, "empty", True) or actual_col not in df.columns:
        return df
    out = df.copy()
    for pos in range(len(out)):
        row = out.iloc[pos].to_dict()
        if _interactive_row_has_timeline_meta(row):
            continue
        if preserve_meters_done and _overtime_simulation_dispatch_trial_active():
            tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
            proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
            mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
            dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
            if tid and mach and dd is not None:
                try:
                    keep_m = float(
                        preserve_meters_done.get((tid, proc, mach, dd), 0.0) or 0.0
                    )
                except (TypeError, ValueError):
                    keep_m = 0.0
                if keep_m > 1e-9:
                    continue
        try:
            out.at[out.index[pos], actual_col] = 0.0
        except Exception:
            pass
    try:
        out[actual_col] = (
            pd.to_numeric(out[actual_col], errors="coerce").fillna(0.0).astype(float)
        )
    except Exception:
        pass
    return out
def _interactive_aggregate_actual_dispatch_from_df(
    df_dispatch: pd.DataFrame,
) -> dict[tuple[str, str, str, date], float]:
    """段階3: 結果_配台表の「実配台数量」（タイムライン m）を (依頼NO, 工程, 機械, 配台日) で集約。"""
    out: dict[tuple[str, str, str, date], float] = defaultdict(float)
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return dict(out)
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    if actual_col not in df_dispatch.columns:
        return dict(out)
    for _, row in df_dispatch.iterrows():
        if not _interactive_row_has_timeline_meta(row.to_dict()):
            continue
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        try:
            q = float(row.get(actual_col) or 0)
        except (TypeError, ValueError):
            q = 0.0
        if tid and mach and dd is not None:
            out[(tid, proc, mach, dd)] += float(q)
    return dict(out)
def _interactive_aggregate_plan_dispatch_targets_from_df(
    df_dispatch: pd.DataFrame,
) -> dict[tuple[str, str, str, date], float]:
    """段階3試行: 配台日スライド後の「当日配台数量」を (依頼NO, 工程, 機械, 配台日) で集約。"""
    out: dict[tuple[str, str, str, date], float] = defaultdict(float)
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return dict(out)
    for _, row in df_dispatch.iterrows():
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        try:
            q = float(row.get("当日配台数量") or 0)
        except (TypeError, ValueError):
            q = 0.0
        if tid and mach and dd is not None and q > 1e-9:
            out[(tid, proc, mach, dd)] += float(q)
    return dict(out)
def _interactive_aggregate_effective_actual_for_validation(
    df_dispatch: pd.DataFrame,
    *,
    meters_done: dict | None = None,
    eps: float = 1e-3,
) -> dict[tuple[str, str, str, date], float]:
    """
    段階3試行の合否判定用実配台 m。
    タイムライン割付済みで実配台数量列が 0 の行は plan を暦日実績として扱う（スライド後キーずれ救済）。
    """
    out: dict[tuple[str, str, str, date], float] = defaultdict(float)
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return dict(out)
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    for _, row in df_dispatch.iterrows():
        row_d = row.to_dict()
        if not _interactive_row_has_timeline_meta(row_d):
            continue
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        try:
            q = float(row.get(actual_col) or 0) if actual_col in df_dispatch.columns else 0.0
        except (TypeError, ValueError):
            q = 0.0
        try:
            plan = float(row.get("当日配台数量") or 0)
        except (TypeError, ValueError):
            plan = 0.0
        if q <= eps and plan > eps:
            q = plan
        elif meters_done and q <= eps:
            st_day_s = _iso_date_from_dispatch_table_datetime_cell(row.get("加工開始日時"))
            if st_day_s:
                try:
                    kd = date.fromisoformat(st_day_s)
                    mk = (tid, proc, mach, kd)
                    q = max(q, float(meters_done.get(mk, 0.0) or 0.0))
                except (TypeError, ValueError):
                    pass
        if tid and mach and dd is not None:
            out[(tid, proc, mach, dd)] += float(q)
    return dict(out)
def _interactive_dispatch_timeline_meta_miss_shortfalls(
    df_dispatch: pd.DataFrame,
    *,
    meters_done: dict[tuple[str, str, str, date], float] | None = None,
    eps: float = 1e-3,
) -> list[dict]:
    """
    計画数量があるのにタイムライン未割付で加工開始日時が空の行（段階3 overlay 後）。
    依頼×工程×機械の総実績が総計画を満たす場合は未達行に載せない。
    """
    out: list[dict] = []
    if df_dispatch is None or getattr(df_dispatch, "empty", True):
        return out
    agg_plan: dict[tuple[str, str, str], float] = defaultdict(float)
    agg_done: dict[tuple[str, str, str], float] = defaultdict(float)
    for _, row in df_dispatch.iterrows():
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        if not tid or not mach:
            continue
        t3 = (tid, proc, mach)
        try:
            plan_q = float(row.get("当日配台数量") or 0)
        except (TypeError, ValueError):
            plan_q = 0.0
        if plan_q > eps:
            agg_plan[t3] += plan_q
    if meters_done:
        for key, qty in meters_done.items():
            if not isinstance(key, tuple) or len(key) != 4:
                continue
            t3 = (
                _interactive_norm_cell(key[0]),
                _interactive_norm_cell(key[1]),
                _interactive_norm_cell(key[2]),
            )
            try:
                agg_done[t3] += float(qty or 0.0)
            except (TypeError, ValueError):
                pass
    for _, row in df_dispatch.iterrows():
        try:
            plan_q = float(row.get("当日配台数量") or 0)
        except (TypeError, ValueError):
            plan_q = 0.0
        if plan_q <= eps:
            continue
        st_raw = row.get("加工開始日時")
        st = (
            ""
            if st_raw is None or (isinstance(st_raw, float) and pd.isna(st_raw))
            else str(st_raw).strip()
        )
        if st and st.lower() != "nan":
            continue
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        if not tid or not mach or dd is None:
            continue
        t3 = (tid, proc, mach)
        if agg_done.get(t3, 0.0) + eps >= agg_plan.get(t3, 0.0) and agg_plan.get(t3, 0.0) > eps:
            continue
        date_iso = dd.isoformat()
        out.append(
            {
                "task_id": tid,
                "process": proc,
                "machine_name": mach,
                "dispatch_date": date_iso,
                "target_m": plan_q,
                "done_m": 0.0,
                "shortfall_m": plan_q,
                "note": (
                    "計画暦日にタイムライン割付なし（加工開始日時が空）。"
                    "段階2の時刻は段階3試行でクリアされています。"
                ),
            }
        )
    return out
def _interactive_validate_dispatch_quantities(
    df_dispatch: pd.DataFrame,
    expected: dict[tuple[str, ...], float],
    *,
    eps: float = 1e-3,
) -> None:
    if not expected:
        return
    actual = _interactive_aggregate_dispatch_targets_from_df(df_dispatch)
    # デスクトップの「配台試行」は入力 JSON の rows を結果_配台表に反映した後ここに来る。
    # 反映後も食い違う場合は旧実装と同様に警告のみ（致命的 PlanningValidationError は出さない）。
    interactive_ui_trial = (os.environ.get("PM_AI_INTERACTIVE_DISPATCH_TRIAL") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )

    def _sum_by_task_machine(
        src: dict[tuple[str, ...], float],
    ) -> dict[tuple[str, str], float]:
        acc: dict[tuple[str, str], float] = defaultdict(float)
        for key, v in src.items():
            if not isinstance(key, tuple):
                continue
            if len(key) == 4:
                tid, _proc, mach, _dd = key[0], key[1], key[2], key[3]
            elif len(key) == 3:
                tid, mach, _dd = key[0], key[1], key[2]
            else:
                continue
            acc[(tid, mach)] += float(v)
        return dict(acc)

    if interactive_ui_trial:
        # 配台日スライド後の結果表 plan と、タイムライン反映後の実配台を照合する。
        # 入力 JSON の暦日キー（expected）はスライド前のままなので直接使わない。
        plan_expected = _interactive_aggregate_plan_dispatch_targets_from_df(df_dispatch)
        actual = _interactive_aggregate_effective_actual_for_validation(
            df_dispatch,
            meters_done=_LAST_INTERACTIVE_TRIAL_METERS_DONE_SNAPSHOT or None,
            eps=eps,
        )
        for key, exp_v in plan_expected.items():
            if not isinstance(key, tuple) or len(key) != 4:
                continue
            tid, proc, mach, dd = key[0], key[1], key[2], key[3]
            if not isinstance(dd, date):
                continue
            try:
                exp_day = float(exp_v or 0.0)
            except (TypeError, ValueError):
                exp_day = 0.0
            if exp_day <= eps:
                continue
            act_day = float(actual.get(key, 0.0))
            lim_day = max(eps, 1e-9 * max(abs(exp_day), abs(act_day), 1.0))
            day_ok = abs(act_day - exp_day) <= lim_day
            if day_ok:
                continue
            # 後ろ倒しで別日に総量が満たされているときは shortfall のみ（致命にしない）
            exp_tm = _sum_by_task_machine(plan_expected)
            act_tm = _sum_by_task_machine(actual)
            tm_key = (str(tid or ""), str(mach or ""))
            total_met = (
                act_tm.get(tm_key, 0.0) + eps >= exp_tm.get(tm_key, 0.0)
                and exp_tm.get(tm_key, 0.0) > eps
            )
            if total_met:
                continue
            msg = (
                "インタラクティブ配台試行: 暦日実配台不一致 "
                f"({tid}, {mach}, {dd.isoformat()}): 目標={exp_day} 実配台={act_day}"
            )
            # 段階3: 暦日未達は dispatch_qty_shortfall へ（試行自体は継続）
            continue
        exp_tm = _sum_by_task_machine(plan_expected)
        act_tm = _sum_by_task_machine(actual)
        all_tm = set(exp_tm) | set(act_tm)
        for k in sorted(all_tm):
            exp_v = float(exp_tm.get(k, 0.0))
            act_v = float(act_tm.get(k, 0.0))
            lim = max(eps, 1e-9 * max(abs(exp_v), abs(act_v), 1.0))
            if abs(act_v - exp_v) <= lim:
                continue
            msg = (
                "インタラクティブ配台試行: タイムライン実配台不一致（依頼NO×機械名の合計） "
                f"{k}: 目標={exp_v} 実配台={act_v}"
            )
            # 段階3: 総量未達も dispatch_qty_shortfall へ（試行自体は継続）
            continue
        return

    exp_by_day: dict[tuple[str, str, date], float] = defaultdict(float)
    for key, v in expected.items():
        if not isinstance(key, tuple):
            continue
        if len(key) == 4:
            tid, _proc, mach, dd = key[0], key[1], key[2], key[3]
        elif len(key) == 3:
            tid, mach, dd = key[0], key[1], key[2]
        else:
            continue
        if not isinstance(dd, date):
            continue
        try:
            exp_by_day[(tid, mach, dd)] += float(v)
        except (TypeError, ValueError):
            continue
    for k, exp_v in exp_by_day.items():
        act_v = actual.get(k, 0.0)
        if abs(act_v - exp_v) <= eps:
            continue
        msg = (
            "インタラクティブ配台試行: 数量不一致 "
            f"{k}: 期待={exp_v} 実際={act_v}"
        )
        raise PlanningValidationError(msg)
def _interactive_validate_timeline_midnight_if_interactive(
    timeline_events: list | None,
) -> None:
    if not timeline_events:
        return
    for ev in timeline_events:
        if not _is_machining_timeline_event(ev):
            continue
        st = ev.get("start_dt")
        ed = ev.get("end_dt")
        if isinstance(st, datetime) and isinstance(ed, datetime):
            if st.date() != ed.date():
                raise PlanningValidationError(
                    "インタラクティブ配台試行: 加工が暦日をまたいでいます（"
                    f"task={ev.get('task_id')} start={st} end={ed}）。"
                )
def _dispatch_postpone_only_policy_active() -> bool:
    """
    段階2標準・段階3（段階2同一パリティ）の配台失敗ポリシー。
    配台「できない」は master 上で機械カレンダー・勤怠未作成、または勤怠最終日までに割り切れないとき（致命）。
    勤怠日付の自動拡張は行わない。インタラクティブ試行の従来モード（非パリティ）では False。
    """
    if not _interactive_dispatch_trial_env_active():
        return True
    return _interactive_stage2_parity_active()
def _stage2_extend_attendance_calendar_enabled() -> bool:
    """段階2標準・段階3パリティでは False（勤怠不足は致命エラー）。従来インタラクティブ試行のみ定数・環境変数を参照。"""
    if _dispatch_postpone_only_policy_active():
        return False
    if _stage2_truthy_env("PM_AI_STAGE2_EXTEND_ATTENDANCE_CALENDAR"):
        return True
    return STAGE2_EXTEND_ATTENDANCE_CALENDAR
def _pending_tasks_with_remaining_units(task_queue: list) -> list[dict]:
    pending: list[dict] = []
    for t in task_queue or []:
        try:
            rem = float(t.get("remaining_units") or 0)
        except (TypeError, ValueError):
            rem = 0.0
        if rem <= 1e-12:
            continue
        try:
            init = float(t.get("initial_remaining_units") or 0)
        except (TypeError, ValueError):
            init = 0.0
        try:
            unit_m = float(t.get("unit_m") or 0)
        except (TypeError, ValueError):
            unit_m = 0.0
        pending.append(
            {
                "task_id": str(t.get("task_id") or "").strip(),
                "process": str(t.get("machine") or "").strip(),
                "machine_name": str(t.get("machine_name") or "").strip(),
                "remaining_units": rem,
                "initial_remaining_units": init,
                "unit_m": unit_m,
                "remaining_m": rem * unit_m if unit_m > 1e-12 else 0.0,
                "_dispatch_block_no_op_on_working_days": bool(
                    t.get("_dispatch_block_no_op_on_working_days")
                ),
            }
        )
    return pending
def _raise_if_remaining_tasks_exceed_attendance_calendar(
    task_queue: list,
    calendar_last_plan_day: date | None,
    *,
    context_label: str = "段階2",
) -> None:
    """
    段階2標準・段階3パリティ: master 勤怠の計画日を使い切っても残タスクがあれば試行を致命エラーで止める。
    段階3手動修正試行（interactive_dispatch_targets あり）のみ: 致命にせずスナップショットへ記録して続行。
    """
    if not (
        _interactive_dispatch_trial_env_active()
        and _PLAN_IMPL_INTERACTIVE_DISPATCH_TARGETS
    ):
        _raise_limited_operator_remaining_tasks(
            task_queue,
            calendar_last_plan_day,
            context_label=context_label,
        )
    if not _dispatch_postpone_only_policy_active():
        return
    pending = _pending_tasks_with_remaining_units(task_queue)
    if not pending:
        global _LAST_INTERACTIVE_REMAINING_TASKS_AT_CALENDAR_END
        _LAST_INTERACTIVE_REMAINING_TASKS_AT_CALENDAR_END = []
        return
    last_iso = (
        calendar_last_plan_day.isoformat()
        if isinstance(calendar_last_plan_day, date)
        else "—"
    )
    if (
        _interactive_dispatch_trial_env_active()
        and _PLAN_IMPL_INTERACTIVE_DISPATCH_TARGETS
    ):
        _LAST_INTERACTIVE_REMAINING_TASKS_AT_CALENDAR_END = list(pending)
        samples: list[str] = []
        for t in pending[:8]:
            tid = str(t.get("task_id") or "").strip()
            mach = str(t.get("process") or "").strip()
            if tid or mach:
                samples.append(f"{tid}/{mach}" if tid and mach else (tid or mach))
        sample_s = "、".join(samples) if samples else "（依頼NO不明）"
        logging.warning(
            "%s: 勤怠カレンダーの最終日（%s）までに配台しきれないタスクが %s 件あります。"
            " 手動修正試行は結果を書き出して続行します（例: %s）。"
            " 配台試行順・特別ルール・設備占有の見直し、または計画期間延長（勤怠日付追加）を検討してください。",
            context_label,
            last_iso,
            len(pending),
            sample_s,
        )
        if isinstance(_LAST_INTERACTIVE_STAGE3_META, dict):
            _LAST_INTERACTIVE_STAGE3_META["remaining_tasks_at_calendar_end"] = len(pending)
            _LAST_INTERACTIVE_STAGE3_META["remaining_tasks_soft_fail"] = True
        return
    _tq_sample_by_key: dict[tuple[str, str, str], dict] = {}
    for _tq_s in task_queue or []:
        _tq_sample_by_key[
            (
                str(_tq_s.get("task_id") or "").strip(),
                str(_tq_s.get("machine") or "").strip(),
                str(_tq_s.get("machine_name") or "").strip(),
            )
        ] = _tq_s

    def _pending_sample_label(p: dict) -> str:
        tid = str(p.get("task_id") or "").strip()
        proc = str(p.get("process") or "").strip()
        base = f"{tid}/{proc}" if tid and proc else (tid or proc or "（依頼NO不明）")
        try:
            rem_u = float(p.get("remaining_units") or 0)
        except (TypeError, ValueError):
            rem_u = 0.0
        try:
            rem_m = float(p.get("remaining_m") or 0)
        except (TypeError, ValueError):
            rem_m = 0.0
        if rem_u > 1e-12:
            base += f"（残{rem_u:g}ロール"
            if rem_m > 1e-12:
                base += f"・{rem_m:g}m"
            base += "）"
        _src = _tq_sample_by_key.get(
            (
                tid,
                proc,
                str(p.get("machine_name") or "").strip(),
            )
        ) or {}
        _prod = str(_src.get(TASK_COL_PRODUCT) or "")
        if "NR28" in unicodedata.normalize("NFKC", _prod) and proc == "EC":
            base += "・L3/NR28で3名必要"
        return base

    samples: list[str] = []
    for t in pending[:8]:
        samples.append(_pending_sample_label(t))
    sample_s = "、".join(samples) if samples else "（依頼NO不明）"
    if len(pending) > 8:
        sample_s += f" 他{len(pending) - 8}件"
    no_op_blocked = [
        t for t in pending if t.get("_dispatch_block_no_op_on_working_days")
    ]
    l11_hint = ""
    if isinstance(WIP_LIMIT_EC_BEFORE_INSP_ROLLS, int) and WIP_LIMIT_EC_BEFORE_INSP_ROLLS > 0:
        _l11_samples: list[str] = []
        for t in pending:
            proc = str(t.get("process") or "").strip()
            if proc != "EC":
                continue
            tid = str(t.get("task_id") or "").strip()
            if not tid:
                continue
            _src = _tq_sample_by_key.get(
                (
                    tid,
                    proc,
                    str(t.get("machine_name") or "").strip(),
                )
            ) or {}
            if not _src.get("roll_pipeline_ec"):
                continue
            try:
                _wip = float(
                    _wip_ec_before_insp_roll_count(
                        task_queue, task_id_exact=tid
                    )
                )
            except (TypeError, ValueError):
                _wip = 0.0
            if _wip + 1e-12 < float(WIP_LIMIT_EC_BEFORE_INSP_ROLLS):
                continue
            if not _b2_ec_insp_pair_in_queue(task_queue, tid):
                continue
            _l11_samples.append(
                f"{tid}/EC（検査前WIP={_wip:g}、上限{WIP_LIMIT_EC_BEFORE_INSP_ROLLS}・後続検査/巻返しの消化が必要）"
            )
        if _l11_samples:
            l11_hint = (
                " 特別ルールL11: "
                + "；".join(_l11_samples[:3])
                + (
                    f" 他{len(_l11_samples) - 3}件"
                    if len(_l11_samples) > 3
                    else ""
                )
                + "。"
            )
    no_op_hint = ""
    if no_op_blocked:
        no_op_samples: list[str] = []
        for t in no_op_blocked[:3]:
            tid = str(t.get("task_id") or "").strip()
            mach = str(t.get("process") or "").strip()
            mname = str(t.get("machine_name") or "").strip()
            if tid and mach and mname:
                no_op_samples.append(f"{tid}/{mach}/{mname}")
            elif tid and mach:
                no_op_samples.append(f"{tid}/{mach}")
            elif tid:
                no_op_samples.append(tid)
        no_op_hint = (
            f" うち {len(no_op_blocked)} 件は稼働日に OP スキル保有者がおらず、"
            f"AS のみの候補では必須人数を満たせません"
            f"（例: {'、'.join(no_op_samples) if no_op_samples else sample_s}）。"
            f" master「skills」で当該工程×機械に OP を設定するか、"
            f" OP 担当の勤怠（休暇・公休等）を確認してください。"
        )
    raise PlanningValidationError(
        f"{context_label}: 計画期間（勤怠マスタ最終日 {last_iso}）内に配台しきれません"
        f"（残タスク {len(pending)} 件）。"
        f"{l11_hint}"
        f"{no_op_hint}"
        f" 計画期間の延長を意図しない場合は、配台試行順・特別ルール（例: L3/NR28 の3名編成）・"
        f"EC機の終盤占有・タスク量を見直してください。"
        f" 計画期間そのものを延ばす運用のときのみ、勤怠シートに日付行を追加してください。"
        f" 残量例: {sample_s}"
    )
def _stage3_extend_attendance_calendar_enabled() -> bool:
    """後方互換。段階3パリティと同一判定。"""
    return _stage2_extend_attendance_calendar_enabled()
def _validate_master_dispatch_prerequisites(
    master_path: str,
    members: list,
    equipment_list: list | None,
    *,
    context_label: str = "配台計画",
) -> None:
    """
    段階2標準・段階3（段階2同一パリティ）の master 前提。
    機械カレンダーは machine-calendar-data.json 正本が必須（master シートは使用しない）。
    人の勤怠は attendance-data.json 正本が整備済みなら JSON のみ。未整備時は master.xlsm 上のメンバーシートが必須。
    """
    xls = _cached_master_pd_excel_file(master_path)
    if xls is None:
        raise PlanningValidationError(
            f"{context_label}: master.xlsm を開けません。パスとファイルの存在を確認してください。"
        )

    from planning_core.core.attendance_readiness import (
        legacy_master_attendance_sheets_required,
    )

    if legacy_master_attendance_sheets_required():
        _validate_skills_members_have_attendance_sheets(
            members, master_path, context_label=context_label
        )

        att_sheet_count = 0
        att_date_rows = 0
        for sheet_name in xls.sheet_names:
            m_name = str(sheet_name).strip()
            if m_name not in members:
                continue
            if "カレンダー" in sheet_name:
                continue
            try:
                df_sheet = pd.read_excel(xls, sheet_name=sheet_name)
            except Exception:
                continue
            df_sheet.columns = df_sheet.columns.str.strip()
            if "日付" not in {str(c).strip() for c in df_sheet.columns}:
                continue
            att_sheet_count += 1
            try:
                dcol = pd.to_datetime(df_sheet["日付"], errors="coerce")
                att_date_rows += int(dcol.notna().sum())
            except Exception:
                continue
        if att_sheet_count == 0 or att_date_rows == 0:
            raise PlanningValidationError(
                f"{context_label}: 人の勤怠が作成されていません。"
                " master.xlsm で各メンバーの勤怠シートを作成し、日付行を入力してから実行してください。"
            )

    from planning_core.core.machine_calendar_store import require_machine_calendar_json_for_dispatch

    require_machine_calendar_json_for_dispatch(context_label)
    if equipment_list:
        blocks = load_machine_calendar_occupancy_blocks(
            master_path,
            equipment_list,
            interactive_only_asterisk_occupancy=False,
            context_label=context_label,
        )
        if not blocks:
            logging.info(
                "%s: 機械カレンダー JSON は存在しますが、skills の設備列と一致する列がありません。"
                " 占有ブロックは空として続行します。",
                context_label,
            )


def _validate_stage3_master_prerequisites(
    master_path: str,
    members: list,
    equipment_list: list | None,
) -> None:
    """段階3（段階2同一パリティ）向けラッパ。"""
    _validate_master_dispatch_prerequisites(
        master_path, members, equipment_list, context_label="段階3配台試行"
    )
def _interactive_append_team_shortage_op_as(
    task: dict,
    current_date: date,
    machine,
    machine_name,
    capable_members: list,
    req_num: int,
) -> None:
    if not _interactive_dispatch_trial_env_active():
        return
    # 段階2標準・段階3同一: 同日の人員不足は後ろ倒しで解消するため、配台不可理由として記録しない。
    if _dispatch_postpone_only_policy_active():
        return
    _cap_n = len(capable_members or [])
    _req_n = int(req_num)
    if _cap_n < _req_n:
        _reason = "フォーム候補不足（必要人数に満たない）"
    else:
        _reason = "チーム組合せ不可（人数は足りるが割当不可）"
    rec = {
        "task_id": str(task.get("task_id") or ""),
        "date": current_date.isoformat(),
        "process": str(machine or ""),
        "machine_name": str(machine_name or ""),
        "reason": _reason,
        "required_headcount": _req_n,
        "capable_headcount": _cap_n,
    }
    if _cap_n < _req_n:
        _INTERACTIVE_TRIAL_OP_SHORTAGE.append(rec)
    else:
        _INTERACTIVE_TRIAL_AS_SHORTAGE.append(rec)
def _timeline_event_assigned_member_names(ev: dict) -> list[str]:
    """加工タイムラインイベントに割り当てられたメンバー名（主担当・補助）。"""
    sl = ev.get("subs_list")
    if isinstance(sl, list) and sl:
        names: list[str] = []
        op = " ".join(str(ev.get("op") or "").split()).strip()
        if op:
            names.append(op)
        for s in sl:
            ss = " ".join(str(s).split()).strip()
            if ss and ss not in names:
                names.append(ss)
        return names
    out: list[str] = []
    op = " ".join(str(ev.get("op") or "").split()).strip()
    if op:
        out.append(op)
    sub_raw = str(ev.get("sub") or "").strip()
    if sub_raw:
        for part in sub_raw.split(","):
            s = " ".join(part.split()).strip()
            if s and s not in out:
                out.append(s)
    return out
def _interactive_append_machining_end_after_member_shift_shortages(
    timeline_events: list | None,
    attendance_data: dict | None,
) -> None:
    """
    インタラクティブ配台試行: 加工セグメント終了が、割り当てメンバーの勤務 end_dt を超えるとき
    as_shortage に記録する（JavaFX の dispatch_trial_shortages.json 連携）。
    段階2標準・段階3同一では記録しない（後ろ倒し前提）。
    """
    if not _interactive_dispatch_trial_env_active():
        return
    if _dispatch_postpone_only_policy_active():
        return
    if not timeline_events or not attendance_data:
        return
    seen: set[tuple[str, str, str, str]] = set()
    for ev in timeline_events:
        if not _is_machining_timeline_event(ev):
            continue
        ev_end = ev.get("end_dt")
        if not isinstance(ev_end, datetime):
            continue
        cal_d = _timeline_event_calendar_date(ev)
        if cal_d is None:
            continue
        day_att = attendance_data.get(cal_d)
        if not day_att:
            continue
        tid = str(ev.get("task_id") or "").strip()
        mach_line = str(ev.get("machine") or "").strip()
        mach_occ = str(ev.get("machine_occupancy_key") or "").strip()
        proc_field = mach_line
        mn_field = mach_occ
        for mm in _timeline_event_assigned_member_names(ev):
            if mm not in day_att:
                continue
            entry = day_att[mm]
            if not entry.get("is_working"):
                continue
            mem_end = entry.get("end_dt")
            if not isinstance(mem_end, datetime):
                continue
            if ev_end <= mem_end:
                continue
            key = (tid, cal_d.isoformat(), proc_field, mm)
            if key in seen:
                continue
            seen.add(key)
            rec = {
                "task_id": tid,
                "date": cal_d.isoformat(),
                "process": proc_field,
                "machine_name": mn_field,
                "reason": (
                    "加工終了が退勤後（"
                    f"{mm} 退勤 {mem_end.strftime('%H:%M')} / "
                    f"加工終了 {ev_end.strftime('%H:%M')}）"
                ),
                "required_headcount": 1,
                "capable_headcount": 1,
            }
            _INTERACTIVE_TRIAL_AS_SHORTAGE.append(rec)
def _dedupe_interactive_trial_shortage_records(recs: list | None) -> list:
    """同一 (依頼NO, 日, 工程, 機械) のロール試行失敗ログを1件にまとめる。"""
    if not recs:
        return []
    seen: set[tuple[str, str, str, str]] = set()
    out: list = []
    for rec in recs:
        if not isinstance(rec, dict):
            continue
        key = (
            _interactive_norm_cell(rec.get("task_id")),
            str(rec.get("date") or "").strip()[:10],
            _interactive_dispatch_target_process_key(rec.get("process")),
            _interactive_norm_cell(rec.get("machine_name")),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
    return out
def filter_interactive_trial_shortages_by_meters_done(
    snap: dict,
    meters_done: dict[tuple[str, str, str, date], float] | None,
    *,
    eps: float = 1e-3,
) -> dict:
    """
    ロール割当プローブで一度 as/op 不足が付いても、タイムラインに実配台 m が載った
    (依頼NO, 工程, 機械) は「配台不可」一覧から除外する。

    - 当該暦日に m がある場合は除外（従来）
    - 暦日は違うが別の配台日キーに m が載っている場合も除外
      （プローブ暦日と結果_配台表の配台日ズレ。例: W5-16 EC は 6/2 失敗ログだが 6/3 に実績）
    """
    if not snap:
        return {"op_shortage": [], "as_shortage": []}

    def _keep(rec: dict) -> bool:
        if not isinstance(rec, dict):
            return False
        if _interactive_trial_shortage_meters_done_for_rec(rec, meters_done, eps=eps) > eps:
            return False
        if _interactive_trial_shortage_meters_done_for_triple(rec, meters_done, eps=eps) > eps:
            return False
        return True

    return {
        "op_shortage": [r for r in snap.get("op_shortage") or [] if _keep(r)],
        "as_shortage": [r for r in snap.get("as_shortage") or [] if _keep(r)],
    }
def compute_interactive_trial_dispatch_qty_shortfall(
    targets: dict[tuple[str, str, str, date], float] | None,
    meters_done: dict[tuple[str, str, str, date], float] | None,
    *,
    eps: float = 1e-3,
) -> list[dict]:
    """
    interactive_dispatch_targets（目標メートル）と meters_done を突き合わせ、
    目標を満たせない暦日キーを一覧化する（JavaFX 未達ハイライト用）。

    段階2標準・段階3同一では、暦日単位のメートル未達は後ろ倒しの途中経過として UI に出さない
    （従来モード）。段階3パリティでは **計画暦日キー** ごとの未達は
    {@code dispatch_qty_shortfall} に載せ、Java の未達ダイアログ・赤セル表示に使う。
    依頼×工程×機械の総量が別日で満たされていても、当該暦日の目標未達は残す。

    同一 (依頼NO, 工程名, 機械名) に複数配台日があるとき、行ごとの不足があっても
    タイムライン総実績が総目標に達していれば未達行は出さない（再集計ズレの誤検知抑止）。
    ※段階3パリティ（後ろ倒しのみ）ではこの総量フィルタは適用しない。
    """
    postpone_only = _dispatch_postpone_only_policy_active()
    out: list[dict] = []
    if not targets:
        return out
    md = meters_done or {}
    agg_tgt: dict[tuple[str, str, str], float] = defaultdict(float)
    agg_done: dict[tuple[str, str, str], float] = defaultdict(float)
    for k, target_m in targets.items():
        if not isinstance(k, tuple) or len(k) != 4:
            continue
        t3 = (
            _interactive_norm_cell(k[0]),
            _interactive_norm_cell(k[1]),
            _interactive_norm_cell(k[2]),
        )
        try:
            agg_tgt[t3] += float(target_m or 0.0)
        except (TypeError, ValueError):
            pass
    for k, v in md.items():
        if not isinstance(k, tuple) or len(k) != 4:
            continue
        t3 = (
            _interactive_norm_cell(k[0]),
            _interactive_norm_cell(k[1]),
            _interactive_norm_cell(k[2]),
        )
        try:
            agg_done[t3] += float(v or 0.0)
        except (TypeError, ValueError):
            pass
    for k, target_m in targets.items():
        if not isinstance(k, tuple) or len(k) != 4:
            continue
        tid, proc, mach, dd = k[0], k[1], k[2], k[3]
        try:
            tgt = float(target_m or 0.0)
        except (TypeError, ValueError):
            tgt = 0.0
        try:
            done_m = float(md.get(k, 0.0))
        except (TypeError, ValueError):
            done_m = 0.0
        gap = tgt - done_m
        if gap > eps:
            date_iso = dd.isoformat() if isinstance(dd, date) else str(dd)
            out.append(
                {
                    "task_id": str(tid or ""),
                    "process": str(proc or ""),
                    "machine_name": str(mach or ""),
                    "dispatch_date": date_iso,
                    "target_m": tgt,
                    "done_m": done_m,
                    "shortfall_m": gap,
                    "note": (
                        "計画暦日の配台目標に対しタイムライン実績が不足"
                        "（段階3: 後日配台・カレンダー後ろ倒し等）。"
                        if postpone_only
                        else (
                            "タイムライン上の割付が目標メートルに達していません"
                            "（機械カレンダー・人員・その他ブロック等）。"
                        )
                    ),
                }
            )
    filtered: list[dict] = []
    for row in out:
        t3 = (
            _interactive_norm_cell(row.get("task_id")),
            _interactive_norm_cell(row.get("process")),
            _interactive_norm_cell(row.get("machine_name")),
        )
        if agg_done.get(t3, 0.0) + eps >= agg_tgt.get(t3, 0.0):
            continue
        filtered.append(row)
    return filtered
def _apply_result_dispatch_table_excel_table(ws, *, table_display_name: str) -> None:
    """
    結果_配台表シートの Excel テーブルを更新する。

    重要: 参照シートがテーブル名を参照しているため、**新規作成はしない**。
    既存テーブル（displayName一致）を見つけた場合のみ ref を更新する。
    """
    try:
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        return
    if ws is None:
        return
    nrows = int(ws.max_row or 0)
    ncols = int(ws.max_column or 0)
    if nrows < 2 or ncols < 1:
        return
    end_l = get_column_letter(ncols)
    ref = f"A1:{end_l}{nrows}"
    try:
        existing = None
        # openpyxl TableList: dict-like
        if hasattr(ws, "tables") and ws.tables:
            existing = ws.tables.get(str(table_display_name))
        if not existing:
            # 新規作成しない（#REF 回避）。テーブル未使用時はログも出さない
            return
        # 既存テーブルの範囲だけ更新
        try:
            existing.ref = ref
        except Exception:
            # TableList の値が Table でない/互換性問題の保険
            ws.tables[str(table_display_name)].ref = ref
    except Exception as e:
        logging.warning("結果_配台表: Excel テーブル更新をスキップしました: %s", e)
def _result_dispatch_table_column_width(header: str) -> float:
    """結果_配台表: 列見出しごとの標準幅（日付列は #### 回避のため広め）。"""
    h = str(header).strip()
    if not h:
        return 11.0
    if h in ("工程名",):
        return 11.5
    if h in ("機械名",):
        return 19.0
    if h in RESULT_DISPATCH_TABLE_DATE_HEADERS:
        return 12.5
    if h in ("受注NO", "依頼NO"):
        return 11.5
    if "品名" in h or h == "製品名":
        return 24.0
    if h == "使用原反":
        return 22.0
    if h == "配台試行順番":
        return 11.0
    if h in ("換算数量", "実加工数", "当日配台数量", "実出来高", "計画合計"):
        return 12.0
    if h == "原反数":
        return 9.0
    if h == "加工内容":
        return 18.0
    if h == "在庫場所":
        return 12.0
    if h == "加工完了区分":
        return 11.0
    if h == "原反投入場所":
        return 16.0
    if h in ("加工開始日時", "加工終了日時"):
        return 17.0
    if h == "メンバー名":
        return 14.0
    return min(max(float(len(h)) + 3.0, 10.0), 28.0)
def _apply_result_dispatch_table_sheet_layout_polish(ws) -> None:
    """
    結果_配台表: 列幅・日付/数量の表示形式・見出し行・左5列の窓枠固定。
    共通仕上げ（罫線・ヘッダ背景）の後に呼ぶこと。
    """
    if ws is None:
        return
    mc = int(ws.max_column or 0)
    mr = int(ws.max_row or 0)
    if mc < 1:
        return

    headers: list[str] = []
    for ci in range(1, mc + 1):
        v = ws.cell(row=1, column=ci).value
        headers.append(str(v).strip() if v is not None else "")

    date_cols: list[int] = []
    qty_cols: list[int] = []
    center_cols: list[int] = []
    for ci, hn in enumerate(headers, 1):
        if not hn:
            continue
        letter = get_column_letter(ci)
        try:
            if getattr(ws.column_dimensions[letter], "hidden", False):
                continue
        except Exception:
            pass
        try:
            ws.column_dimensions[letter].width = float(_result_dispatch_table_column_width(hn))
        except Exception:
            pass
        if hn in RESULT_DISPATCH_TABLE_DATE_HEADERS:
            date_cols.append(ci)
        elif hn in (
            "換算数量",
            "実加工数",
            "当日配台数量",
            "実出来高",
            "計画合計",
            "原反数",
        ):
            qty_cols.append(ci)
        if hn in (
            "配台試行順番",
            "工程名",
            "機械名",
            "受注NO",
            "依頼NO",
            "加工完了区分",
            "原反数",
        ):
            center_cols.append(ci)

    date_fmt = "yyyy/mm/dd"
    num_fmt = "#,##0.###"

    for r in range(2, mr + 1):
        for ci in date_cols:
            c = ws.cell(row=r, column=ci)
            v = c.value
            if v is None or v == "":
                continue
            if isinstance(v, datetime):
                try:
                    c.value = v.date() if hasattr(v, "date") else v
                except Exception:
                    pass
                c.number_format = date_fmt
                continue
            if isinstance(v, date):
                c.number_format = date_fmt
                continue
            try:
                if isinstance(v, pd.Timestamp):
                    c.value = v.to_pydatetime().date()
                    c.number_format = date_fmt
                    continue
            except Exception:
                pass
            try:
                d0 = pd.to_datetime(v, errors="coerce")
                if pd.isna(d0):
                    continue
                if isinstance(d0, pd.Timestamp):
                    c.value = d0.to_pydatetime().date()
                else:
                    xd = d0.to_pydatetime().date() if hasattr(d0, "to_pydatetime") else d0
                    c.value = xd
                c.number_format = date_fmt
            except Exception:
                pass

    for r in range(2, mr + 1):
        for ci in qty_cols:
            c = ws.cell(row=r, column=ci)
            v = c.value
            if v is None or v == "":
                continue
            try:
                if isinstance(v, str):
                    s = unicodedata.normalize("NFKC", v).strip().replace(",", "")
                    if not s:
                        continue
                    c.value = float(s)
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    c.value = float(v)
                else:
                    continue
                c.number_format = num_fmt
            except Exception:
                pass

    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci in range(1, mc + 1):
        ws.cell(row=1, column=ci).alignment = hdr_align
    try:
        ws.row_dimensions[1].height = 32.0
    except Exception:
        pass

    for r in range(2, mr + 1):
        for ci in range(1, mc + 1):
            hn = headers[ci - 1] if ci <= len(headers) else ""
            cell = ws.cell(row=r, column=ci)
            if ci in date_cols or ci in qty_cols:
                cell.alignment = Alignment(
                    horizontal="center", vertical="top", wrap_text=False
                )
            elif ci in center_cols:
                cell.alignment = Alignment(
                    horizontal="center", vertical="top", wrap_text=False
                )
            elif any(
                k in hn
                for k in ("品名", "製品", "加工内容", "使用原反", "在庫", "場所")
            ):
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=False
                )

    try:
        ws.freeze_panes = "G2"
    except Exception:
        pass
def _write_dispatch_table_standalone_xlsx(df_dispatch: pd.DataFrame, target_dir: str) -> str | None:
    """
    「結果_配台表.xlsx」を target_dir に出力する（Power Query _q結果_配台表 / フォルダパス + 固定名）。

    target_dir は resolve_result_dispatch_table_output_dir が決める（PM_AI_RESULT_DISPATCH_TABLE_DIR、
    無ければマクロブック親、無ければ PM_AI_REPO_ROOT/code/output）。
    - シート名: 結果_配台表
    - テーブル名: _t結果_配台表
    """
    try:
        if df_dispatch is None or getattr(df_dispatch, "empty", True):
            return None
        if not target_dir:
            return None
        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as mk_e:
            logging.warning("結果_配台表.xlsx: 出力先フォルダを作成できません: %s (%s)", target_dir, mk_e)
            return None
        if not os.path.isdir(target_dir):
            return None
        out_path = os.path.join(target_dir, "結果_配台表.xlsx")
        # 既存ファイルを上書き（開いていると失敗する）
        try:
            if os.path.isfile(out_path):
                os.remove(out_path)
        except Exception:
            pass
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            df_dispatch.to_excel(w, sheet_name=RESULT_DISPATCH_TABLE_SHEET_NAME, index=False)
            ws = w.sheets.get(RESULT_DISPATCH_TABLE_SHEET_NAME)
            if ws is not None:
                _apply_output_font_to_result_sheet(ws)
                # 別ブックなので新規テーブル作成で問題なし
                try:
                    from openpyxl.worksheet.table import Table, TableStyleInfo

                    nrows = int(ws.max_row or 0)
                    ncols = int(ws.max_column or 0)
                    if nrows >= 2 and ncols >= 1:
                        end_l = get_column_letter(ncols)
                        ref = f"A1:{end_l}{nrows}"
                        tab = Table(
                            displayName=str(RESULT_DISPATCH_TABLE_EXCEL_TABLE_NAME), ref=ref
                        )
                        tab.tableStyleInfo = TableStyleInfo(
                            name="TableStyleMedium9",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False,
                        )
                        ws.add_table(tab)
                except Exception as e:
                    logging.warning("結果_配台表.xlsx: テーブル付与をスキップしました: %s", e)
                _apply_stage2_plan_sheet_header_fill(ws)
                _apply_stage2_plan_sheet_grid_border(ws)
                _apply_result_dispatch_table_sheet_layout_polish(ws)
        return out_path
    except Exception as e:
        logging.warning("結果_配台表.xlsx: 出力に失敗しました: %s", e)
        return None
def _interactive_dispatch_trial_env_active() -> bool:
    v = (os.environ.get("PM_AI_INTERACTIVE_DISPATCH_TRIAL") or "").strip().lower()
    return v in ("1", "true", "yes", "on")
def _interactive_stage2_parity_active() -> bool:
    """段階3を段階2と同一の機械カレンダー・工場枠（終業延長を含む）で回す（runner が設定）。"""
    v = (os.environ.get("PM_AI_INTERACTIVE_TRIAL_STAGE2_PARITY") or "").strip().lower()
    return v in ("1", "true", "yes", "on")
def _stage35_stage3_meters_floor_json_path() -> "Path | None":
    from pathlib import Path

    raw = (os.environ.get(ENV_STAGE35_STAGE3_METERS_FLOOR_JSON) or "").strip()
    if not raw:
        return None
    p = Path(raw)
    return p if p.is_file() else None
_STAGE35_FLOOR_APPLY_META: dict = {}
_STAGE35_FLOOR_METERS_SNAPSHOT: dict[tuple[str, str, str, date], float] = {}
def _stage35_merge_floor_into_meters_done(
    meters_done: dict[tuple[str, str, str, date], float] | None,
) -> dict[tuple[str, str, str, date], float]:
    """段階3.5: タイムライン再集計 m に段階3 floor を足し合わせる（max キーごと）。"""
    merged: dict[tuple[str, str, str, date], float] = {}
    for src in (_STAGE35_FLOOR_METERS_SNAPSHOT or {}, meters_done or {}):
        for k, v in src.items():
            if not isinstance(k, tuple) or len(k) != 4:
                continue
            try:
                fv = float(v or 0.0)
            except (TypeError, ValueError):
                continue
            if fv <= 1e-18:
                continue
            prev = float(merged.get(k, 0.0) or 0.0)
            if fv > prev:
                merged[k] = fv
    return merged
def _apply_stage35_stage3_meters_floor(
    task_queue: list,
    meters_done: dict[tuple[str, str, str, date], float],
) -> dict:
    """
    段階3.5: 段階3試行後の実配台 m を下限として meters_done と remaining_units に反映する。
    定時帯の再配台をゼロからやり直さず、残業帯での追加分のみ試行する。
    """
    global _STAGE35_FLOOR_APPLY_META, _STAGE35_FLOOR_METERS_SNAPSHOT
    _STAGE35_FLOOR_METERS_SNAPSHOT = {}
    p = _stage35_stage3_meters_floor_json_path()
    if p is None:
        _STAGE35_FLOOR_APPLY_META = {
            "applied": False,
            "reason": "missing_floor_json",
        }
        return _STAGE35_FLOOR_APPLY_META
    try:
        payload = json.loads(p.read_text(encoding="utf-8"))
    except Exception as ex:
        _STAGE35_FLOOR_APPLY_META = {
            "applied": False,
            "reason": f"read_failed:{ex}",
        }
        return _STAGE35_FLOOR_APPLY_META
    cells = payload.get("cells") if isinstance(payload, dict) else None
    if not isinstance(cells, list):
        _STAGE35_FLOOR_APPLY_META = {
            "applied": False,
            "reason": "invalid_cells",
        }
        return _STAGE35_FLOOR_APPLY_META

    triple_meters: dict[tuple[str, str, str], float] = defaultdict(float)
    triple_hist_dd: dict[tuple[str, str, str], dict[date, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    applied_cells = 0
    applied_m = 0.0
    for cell in cells:
        if not isinstance(cell, dict):
            continue
        tid = _interactive_norm_cell(cell.get("task_id"))
        proc = _interactive_dispatch_target_process_key(cell.get("process"))
        mach = _interactive_norm_cell(cell.get("machine_name"))
        dd = parse_optional_date(str(cell.get("date") or "").strip())
        try:
            m = float(cell.get("meters") or 0.0)
        except (TypeError, ValueError):
            m = 0.0
        if not tid or not mach or dd is None or m <= 1e-9:
            continue
        key4 = (tid, proc, mach, dd)
        prev = float(meters_done.get(key4, 0.0) or 0.0)
        if m > prev:
            meters_done[key4] = m
        _STAGE35_FLOOR_METERS_SNAPSHOT[key4] = float(meters_done.get(key4, 0.0) or 0.0)
        t3 = (tid, proc, mach)
        triple_meters[t3] += m
        triple_hist_dd[t3][dd] += float(m)
        applied_cells += 1
        applied_m += float(m)

    credited_tasks = 0
    for t in task_queue or []:
        tid = _interactive_norm_cell(str(t.get("task_id") or ""))
        proc = _interactive_dispatch_target_process_key(t.get("machine"))
        mach = _interactive_norm_cell(str(t.get("machine_name") or ""))
        t3 = (tid, proc, mach)
        floor_m = float(triple_meters.get(t3, 0.0) or 0.0)
        if floor_m <= 1e-9:
            continue
        try:
            um = float(t.get("unit_m") or 0.0)
        except (TypeError, ValueError):
            um = 0.0
        if um <= 1e-12:
            continue
        floor_units = floor_m / um
        try:
            initial = float(t.get("initial_remaining_units") or 0.0)
        except (TypeError, ValueError):
            initial = float(t.get("remaining_units") or 0.0)
        new_rem = max(0.0, initial - floor_units)
        t["remaining_units"] = new_rem
        if new_rem <= 1e-9:
            day_map = triple_hist_dd.get(t3) or {}
            hist = [
                {"date": d.strftime("%m/%d"), "done_m": float(v)}
                for d, v in sorted(day_map.items(), key=lambda kv: kv[0])
            ]
            t["assigned_history"] = hist
            credited_tasks += 1

    _STAGE35_FLOOR_APPLY_META = {
        "applied": applied_cells > 0,
        "floor_json": str(p),
        "cells": applied_cells,
        "meters": applied_m,
        "tasks_credited": credited_tasks,
        "floor_keys": len(_STAGE35_FLOOR_METERS_SNAPSHOT),
    }
    return _STAGE35_FLOOR_APPLY_META
def _stage2_1_overtime_active() -> bool:
    """段階2.1: 残業/休出シミュのフル再配台（配台試行の段階3.5経路ではない）。"""
    return (os.environ.get("PM_AI_STAGE2_1_OVERTIME") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )
def _overtime_simulation_dispatch_trial_active() -> bool:
    """残業シミュ JSON が有効な段階3配台試行（段階2.1フル再配台は含まない）。"""
    if _stage2_1_overtime_active():
        return False
    return _overtime_simulation_json_path() is not None
def _stage35_overtime_regular_end_floor(
    team: tuple,
    daily_status: dict,
    current_date: date,
) -> datetime | None:
    """段階3.5 残業帯: 全員の定時終了（base_end_dt）以降から開始。"""
    if not team:
        return None
    reg_ends: list[datetime] = []
    for m in team:
        st = daily_status.get(m)
        if not st:
            return None
        base = st.get("base_end_dt")
        if not isinstance(base, datetime):
            base = datetime.combine(current_date, DEFAULT_END_TIME)
        reg_ends.append(base)
    return max(reg_ends) if reg_ends else None
def _interactive_dispatch_cap_enforced_in_schedule_loop() -> bool:
    """
    段階3配台ループ内で JSON 暦日×数量（interactive_dispatch_targets）を割当上限とするか。

    手動修正タブ由来の targets があるときは **段階2同一パリティでも True**（暦日4200m等の
    計画を守り、早期暦日への過剰割付で後日 plan が空振りになるのを防ぐ）。
    targets が無い従来試行のみ、パリティ時は False（段階2ブロック条件と同一）。

    段階3.5（残業シミュ）でも定時帯は暦日キャップを維持する。超過分は
    _drain_rolls_for_task 内の残業帯ドレイン（17:00 以降）で追加する。
    """
    if not _interactive_dispatch_trial_env_active():
        return False
    if _PLAN_IMPL_INTERACTIVE_DISPATCH_TARGETS:
        return True
    if _interactive_stage2_parity_active():
        return False
    return True
def _stage3_qty_strict_active() -> bool:
    """
    段階3.2（数量厳守）モード。env ``PM_AI_STAGE3_2_QTY_STRICT`` が真のとき True。

    定常終了（A15/B15）による終業直前デファー・小残デファーを無効化し、
    工場枠終業を当日 23:59 まで拡張する（設備占有・機械カレンダーは尊重）。
    既定 off のため段階1/2.0/3.0 の挙動は不変。
    """
    v = (os.environ.get("PM_AI_STAGE3_2_QTY_STRICT") or "").strip().lower()
    return v in ("1", "true", "yes", "on")
def _interactive_machine_calendar_gap_blocks(day_d: date) -> list[tuple[datetime, datetime]]:
    """
    インタラクティブ配台試行: 機械カレンダーにスロット行が無い時刻（工場計画窓内）は配台不可。
    列0で定義されたスロットの合併の**外側**をブロック区間として返す。
    シートに当該暦日が一切無いときは計画窓全体。
    """
    if not _interactive_trial_calendar_legacy_active():
        return []
    union = _MACHINE_CALENDAR_INTERACTIVE_DEFINED_SLOTS_BY_DATE.get(day_d)
    w0 = datetime.combine(day_d, DEFAULT_START_TIME)
    w1 = datetime.combine(day_d, DEFAULT_END_TIME)
    if union is None:
        return [(w0, w1)]
    merged_u = _merge_machine_calendar_intervals(list(union))
    if not merged_u:
        return [(w0, w1)]
    return _half_open_gaps_in_window(w0, w1, merged_u)
def _interactive_augment_machine_calendar_day_blocks(
    day_d: date,
    day_blocks: dict[str, list[tuple[datetime, datetime]]] | None,
    equipment_list: list | None,
    *,
    extra_occ_keys: list[str] | None = None,
) -> dict[str, list[tuple[datetime, datetime]]]:
    """インタラクティブ試行時のみ、未定義時刻ブロックを全対象設備キーへマージする。"""
    db = dict(day_blocks or {})
    if not _interactive_trial_calendar_legacy_active():
        return db
    gaps = _interactive_machine_calendar_gap_blocks(day_d)
    if not gaps:
        return db
    keys: set[str] = set(db.keys())
    for el in equipment_list or []:
        ek = str(el).strip()
        if ek:
            keys.add(ek)
            pk = _equipment_line_key_to_physical_occupancy_key(ek)
            if pk:
                keys.add(pk)
    for ok in extra_occ_keys or []:
        o = str(ok).strip()
        if o:
            keys.add(o)
    if not keys:
        return db
    for k in keys:
        db[k] = _merge_machine_calendar_intervals((db.get(k) or []) + gaps)
    return db
def _dataframe_from_interactive_dispatch_json_rows(
    json_rows: list,
    json_columns: list | None,
    *,
    fallback_columns_from: pd.DataFrame | None,
) -> pd.DataFrame:
    """結果_配台表 JSON の rows のみから DataFrame を組み立てる（timeline が空のときのフォールバック）。"""
    if json_columns and isinstance(json_columns, list) and len(json_columns) > 0:
        cols_order = [str(x) for x in json_columns]
    elif (
        fallback_columns_from is not None
        and not getattr(fallback_columns_from, "empty", True)
        and len(fallback_columns_from.columns) > 0
    ):
        cols_order = list(fallback_columns_from.columns)
    else:
        cols_order = list(RESULT_DISPATCH_TABLE_STATIC_HEADERS) + ["配台日", "当日配台数量"]
    recs: list[dict] = []
    for r in json_rows:
        if not isinstance(r, dict):
            continue
        d: dict = {}
        for c in cols_order:
            v = r.get(c)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                d[c] = ""
            elif c in RESULT_DISPATCH_TABLE_DATE_HEADERS:
                d[c] = _norm_ymd(v)
            else:
                if isinstance(v, (dict, list)):
                    d[c] = str(v)
                elif isinstance(v, bool):
                    d[c] = "はい" if v else "いいえ"
                else:
                    d[c] = v
        recs.append(d)
    df_out = pd.DataFrame(recs)
    for c in cols_order:
        if c not in df_out.columns:
            df_out[c] = ""
    ordered = [c for c in cols_order if c in df_out.columns]
    extra = [c for c in df_out.columns if c not in ordered]
    return df_out.reindex(columns=ordered + extra)
def _norm_dispatch_meta_date_key(val) -> str:
    """依頼NO×機械×配台日のキー用に暦日を yyyy-mm-dd に寄せる。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    ts = pd.to_datetime(val, errors="coerce")
    if not pd.isna(ts) and hasattr(ts, "date"):
        try:
            return ts.date().isoformat()
        except Exception:
            pass
    s = str(val).strip()
    if not s:
        return ""
    ts2 = pd.to_datetime(s, errors="coerce")
    if not pd.isna(ts2) and hasattr(ts2, "date"):
        try:
            return ts2.date().isoformat()
        except Exception:
            pass
    return s
def _dispatch_meta_join_key_from_mapping(row: dict) -> tuple[str, str, str]:
    tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
    mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
    dd_s = _norm_dispatch_meta_date_key(row.get("配台日"))
    return (tid, mach, dd_s)
def _iso_date_from_dispatch_table_datetime_cell(val) -> str:
    """加工開始／終了日時セルから暦日 yyyy-mm-dd。取れなければ空。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        if pd.api.types.is_scalar(val) and pd.isna(val):
            return ""
    except Exception:
        pass
    if isinstance(val, datetime):
        try:
            return val.date().isoformat()
        except (ValueError, OSError, AttributeError, TypeError):
            return ""
    if isinstance(val, date) and not isinstance(val, datetime):
        try:
            return val.isoformat()
        except (ValueError, OSError, AttributeError, TypeError):
            return ""
    s = str(val).strip()
    if not s or s.lower() in ("nan", "nat", "none"):
        return ""
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.api.types.is_scalar(ts) and pd.isna(ts):
            return ""
        if hasattr(ts, "date"):
            return ts.date().isoformat()
    except Exception:
        pass
    return ""
def _overlay_timeline_meta_onto_interactive_dispatch_df(
    df_out: pd.DataFrame,
    df_sim: pd.DataFrame,
) -> pd.DataFrame:
    """
    インタラクティブ配台試行で入力 JSON を正とする際、加工開始／終了／メンバー名だけは
    タイムライン集約（df_sim）で上書きする。入力 JSON に残った古い終了時刻・単一名を是正する。

    JSON の「配台日」と df_sim の暦日キーがずれる場合、配台日キー一致でメタが取れなければ
    「加工開始日時の暦日＝JSON 配台日」の df_sim 行を二次照合する。
    メタが取れない行は、タイムラインに無い誤った JSON 時刻を残さないため、加工開始／終了／メンバー名を空にする。
    上書き後も「加工開始の暦日」と「配台日」が食い違う場合のみ、当該3列を空にして段階2の時刻が
    配台日だけ変わったように見える誤表示を防ぐ。
    """
    if (
        df_sim is None
        or getattr(df_sim, "empty", True)
        or df_out is None
        or getattr(df_out, "empty", True)
    ):
        return df_out
    meta_cols = ("加工開始日時", "加工終了日時", "メンバー名")
    need_cols = {TASK_COL_TASK_ID, TASK_COL_MACHINE_NAME, "配台日"}
    if not need_cols.issubset(set(df_sim.columns)) or not need_cols.issubset(set(df_out.columns)):
        return df_out
    if not all(c in df_sim.columns for c in meta_cols):
        return df_out
    lookup: dict[tuple[str, str, str], dict[str, object]] = {}
    lookup_by_start_day: dict[tuple[str, str, str], dict[str, object]] = {}
    for _, sim_row in df_sim.iterrows():
        d = sim_row.to_dict()
        k = _dispatch_meta_join_key_from_mapping(d)
        if not k[0]:
            continue
        meta = {c: sim_row.get(c) for c in meta_cols}
        lookup[k] = meta
        st_day = _iso_date_from_dispatch_table_datetime_cell(d.get("加工開始日時"))
        if st_day:
            lookup_by_start_day[(k[0], k[1], st_day)] = meta

    out = df_out.copy()

    def _clear_meta_at_pos(pos0: int) -> None:
        for c in meta_cols:
            try:
                ci = out.columns.get_loc(c)
                if isinstance(ci, slice):
                    continue
                out.iloc[pos0, ci] = ""
            except Exception:
                continue

    for pos in range(len(out)):
        k = _dispatch_meta_join_key_from_mapping(out.iloc[pos].to_dict())
        meta = lookup.get(k)
        if meta is None and k[0] and k[1] and k[2]:
            meta = lookup_by_start_day.get(k)
        if meta is not None:
            wrote_start = False
            for c in meta_cols:
                v = meta.get(c)
                if v is None:
                    continue
                if isinstance(v, float) and pd.isna(v):
                    continue
                sv = str(v).strip()
                if not sv or sv.lower() in ("nan", "nat"):
                    continue
                if c == "加工開始日時":
                    wrote_start = True
                try:
                    ci = out.columns.get_loc(c)
                    if isinstance(ci, slice):
                        continue
                    out.iloc[pos, ci] = v
                except Exception:
                    continue
            if not wrote_start:
                _clear_meta_at_pos(pos)
        else:
            _clear_meta_at_pos(pos)
        want_dd = k[2]
        if not want_dd:
            continue
        dd_out = _interactive_parse_dispatch_date_cell(out.iloc[pos].get("配台日"))
        raw_st = out.iloc[pos].get("加工開始日時")
        dd_st: date | None = None
        if raw_st is not None and not (isinstance(raw_st, float) and pd.isna(raw_st)):
            try:
                if isinstance(raw_st, datetime):
                    dd_st = raw_st.date()
                elif isinstance(raw_st, date) and not isinstance(raw_st, datetime):
                    dd_st = raw_st
                else:
                    ts_st = pd.to_datetime(raw_st, errors="coerce")
                    if not pd.isna(ts_st) and hasattr(ts_st, "date"):
                        dd_st = ts_st.date()
            except Exception:
                dd_st = None
        if dd_out is not None and dd_st is not None and dd_st != dd_out:
            # 手動修正 JSON の配台日（暦日別行）を加工開始暦日へ寄せない。メタだけ落とす。
            _clear_meta_at_pos(pos)
    return out
def _interactive_merge_actual_dispatch_qty_from_timeline_table(
    df_editor: pd.DataFrame,
    df_timeline_dispatch: pd.DataFrame,
    *,
    append_missing_timeline_days: bool = True,
) -> pd.DataFrame:
    """
    段階3: 編集 JSON の「当日配台数量」（目標 m）は維持し、タイムライン実配台 m を
    ``INTERACTIVE_DISPATCH_ACTUAL_QTY_COL``（実配台数量）へ書く。Java 手動修正表は
    目標をそのまま、実績を ``(数字)`` で表示する（例: ``600 (400)`` / ``(600)``）。

    キーは (依頼NO, 機械名, 配台暦日)。タイムライン側にのみある暦日行は行として追補する。
    """
    if df_editor is None or getattr(df_editor, "empty", True):
        return df_editor
    if df_timeline_dispatch is None or getattr(df_timeline_dispatch, "empty", True):
        return df_editor
    plan_col = "当日配台数量"
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    if plan_col not in df_editor.columns:
        return df_editor
    lk: defaultdict[tuple[str, str, str], float] = defaultdict(float)
    for _, simr in df_timeline_dispatch.iterrows():
        tid = _interactive_norm_cell(simr.get(TASK_COL_TASK_ID))
        mach = _interactive_norm_cell(simr.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(simr.get("配台日"))
        if not tid or not mach or dd is None:
            continue
        ddk = dd.isoformat()
        try:
            q = float(simr.get(plan_col) or 0)
        except (TypeError, ValueError):
            q = 0.0
        lk[(tid, mach, ddk)] += float(q)
    out = df_editor.copy()
    if actual_col not in out.columns:
        out[actual_col] = 0.0
    present: set[tuple[str, str, str]] = set()
    for pos in range(len(out)):
        tid = _interactive_norm_cell(out.iloc[pos].get(TASK_COL_TASK_ID))
        mach = _interactive_norm_cell(out.iloc[pos].get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(out.iloc[pos].get("配台日"))
        if not tid or not mach or dd is None:
            continue
        ddk = dd.isoformat()
        present.add((tid, mach, ddk))
        qv = float(lk.get((tid, mach, ddk), 0.0))
        try:
            out.at[out.index[pos], actual_col] = qv
        except Exception:
            pass
    extra: list[dict] = []
    out_cols = list(out.columns)
    appended_keys: set[tuple[str, str, str]] = set()
    if append_missing_timeline_days:
        for _, simr in df_timeline_dispatch.iterrows():
            tid = _interactive_norm_cell(simr.get(TASK_COL_TASK_ID))
            mach = _interactive_norm_cell(simr.get(TASK_COL_MACHINE_NAME))
            dd = _interactive_parse_dispatch_date_cell(simr.get("配台日"))
            if not tid or not mach or dd is None:
                continue
            ddk = dd.isoformat()
            try:
                q = float(simr.get(plan_col) or 0)
            except (TypeError, ValueError):
                q = 0.0
            if q <= 1e-9:
                continue
            if (tid, mach, ddk) in present:
                continue
            if (tid, mach, ddk) in appended_keys:
                continue
            appended_keys.add((tid, mach, ddk))
            row: dict = {c: "" for c in out_cols}
            for c in out_cols:
                if c in simr.index:
                    v = simr.get(c)
                    if c in RESULT_DISPATCH_TABLE_DATE_HEADERS:
                        row[c] = _norm_ymd(v)
                    elif v is not None and not (isinstance(v, float) and pd.isna(v)):
                        row[c] = v
                    else:
                        row[c] = ""
            row[plan_col] = 0.0
            row[actual_col] = float(q)
            extra.append(row)
    if extra:
        out = pd.concat([out, pd.DataFrame(extra)], ignore_index=True)
    try:
        out[actual_col] = pd.to_numeric(out[actual_col], errors="coerce").fillna(0.0).astype(float)
    except Exception:
        pass
    return out
def _interactive_apply_recomputed_actual_qty_to_dispatch_df(
    df_out: pd.DataFrame,
    meters_done: dict[tuple[str, str, str, date], float],
) -> pd.DataFrame:
    """段階3: タイムライン再集計 m を JSON 行の (依頼NO, 工程, 機械, 配台日) キーへ実配台数量に反映する。

    同一キーの行が複数あるときは当日配台数量比で按分し、タイムライン m の二重計上を防ぐ。
    """
    if df_out is None or getattr(df_out, "empty", True) or not meters_done:
        return df_out
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    plan_col = "当日配台数量"
    if actual_col not in df_out.columns:
        df_out[actual_col] = 0.0
    key_rows: dict[tuple, list[int]] = defaultdict(list)
    key_plan_sum: dict[tuple, float] = defaultdict(float)
    for pos in range(len(df_out)):
        row = df_out.iloc[pos]
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        st_day_s = _iso_date_from_dispatch_table_datetime_cell(row.get("加工開始日時"))
        lookup_dd = dd
        if st_day_s and _interactive_row_has_timeline_meta(row):
            try:
                lookup_dd = date.fromisoformat(st_day_s)
            except (TypeError, ValueError):
                lookup_dd = dd
        if not tid or not proc or not mach or lookup_dd is None:
            continue
        kk = (tid, proc, mach, lookup_dd)
        key_rows[kk].append(pos)
        try:
            key_plan_sum[kk] += float(row.get(plan_col) or 0.0)
        except (TypeError, ValueError):
            pass
    for kk, positions in key_rows.items():
        try:
            total_md = float(meters_done.get(kk, 0.0))
        except (TypeError, ValueError):
            total_md = 0.0
        total_plan = float(key_plan_sum.get(kk, 0.0))
        if len(positions) == 1:
            try:
                df_out.at[df_out.index[positions[0]], actual_col] = total_md
            except Exception:
                pass
            continue
        allocated = 0.0
        for i, pos in enumerate(positions):
            try:
                plan_q = float(df_out.iloc[pos].get(plan_col) or 0.0)
            except (TypeError, ValueError):
                plan_q = 0.0
            if i == len(positions) - 1:
                qv = max(0.0, total_md - allocated)
            elif total_plan > 1e-9 and plan_q > 1e-9:
                qv = total_md * (plan_q / total_plan)
                allocated += qv
            else:
                qv = 0.0
            try:
                df_out.at[df_out.index[pos], actual_col] = qv
            except Exception:
                pass
    try:
        df_out[actual_col] = (
            pd.to_numeric(df_out[actual_col], errors="coerce").fillna(0.0).astype(float)
        )
    except Exception:
        pass
    # 単一計画ブロック（正の当日配台数量が1行のみ）: 翌暦日へはみ出したタイムライン m を
    # その計画行の実配台数量へ集約する。
    eps = 1e-9
    by_profile: dict[tuple[str, str, str], list[tuple[int, float]]] = defaultdict(list)
    for pos in range(len(df_out)):
        row = df_out.iloc[pos]
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        if not tid or not proc or not mach:
            continue
        try:
            plan_q = float(row.get(plan_col) or 0.0)
        except (TypeError, ValueError):
            plan_q = 0.0
        by_profile[(tid, proc, mach)].append((pos, plan_q))
    for (tid, proc, mach), entries in by_profile.items():
        plan_positions = [p for p, pq in entries if pq > eps]
        if len(plan_positions) != 1:
            continue
        plan_pos = plan_positions[0]
        total_md = 0.0
        for k, v in meters_done.items():
            if (
                isinstance(k, tuple)
                and len(k) == 4
                and k[0] == tid
                and k[1] == proc
                and k[2] == mach
            ):
                try:
                    total_md += float(v)
                except (TypeError, ValueError):
                    pass
        try:
            df_out.at[df_out.index[plan_pos], actual_col] = total_md
        except Exception:
            pass
        for pos, plan_q in entries:
            if pos == plan_pos or plan_q > eps:
                continue
            try:
                df_out.at[df_out.index[pos], actual_col] = 0.0
            except Exception:
                pass
    try:
        df_out[actual_col] = (
            pd.to_numeric(df_out[actual_col], errors="coerce").fillna(0.0).astype(float)
        )
    except Exception:
        pass
    return df_out
def _interactive_consolidate_duplicate_plan_dispatch_rows(
    df_out: pd.DataFrame,
) -> pd.DataFrame:
    """
    同一 (依頼NO, 工程, 機械, 配台日) の暦日行を1行に集約する。
    当日配台数量は合算。タイムライン meta がある行を残す。
    """
    if df_out is None or getattr(df_out, "empty", True):
        return df_out
    plan_col = "当日配台数量"
    if plan_col not in df_out.columns:
        return df_out
    groups: dict[tuple, list[int]] = defaultdict(list)
    for pos in range(len(df_out)):
        row = df_out.iloc[pos]
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        st_day = _iso_date_from_dispatch_table_datetime_cell(row.get("加工開始日時")) or ""
        if tid and proc and mach and dd is not None:
            groups[(tid, proc, mach, dd, st_day)].append(pos)
    drop: set[int] = set()
    for _k, positions in groups.items():
        if len(positions) <= 1:
            continue
        keeper = positions[0]
        for p in positions:
            if _interactive_row_has_timeline_meta(df_out.iloc[p].to_dict()):
                keeper = p
                break
        total_plan = 0.0
        total_actual = 0.0
        actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
        for p in positions:
            try:
                total_plan += float(df_out.iloc[p].get(plan_col) or 0.0)
            except (TypeError, ValueError):
                pass
            if actual_col in df_out.columns:
                try:
                    total_actual += float(df_out.iloc[p].get(actual_col) or 0.0)
                except (TypeError, ValueError):
                    pass
        try:
            df_out.at[df_out.index[keeper], plan_col] = total_plan
        except Exception:
            pass
        if actual_col in df_out.columns and total_actual > 1e-9:
            try:
                df_out.at[df_out.index[keeper], actual_col] = total_actual
            except Exception:
                pass
        for p in positions:
            if p != keeper:
                drop.add(p)
    if not drop:
        return df_out
    keep_idx = [i for i in range(len(df_out)) if i not in drop]
    return df_out.iloc[keep_idx].reset_index(drop=True)
def _interactive_resolve_slide_target_dispatch_date(
    tid: str,
    proc: str,
    mach: str,
    plan_dd: date,
    *,
    df_sim: pd.DataFrame | None = None,
    meters_done: dict[tuple[str, str, str, date], float] | None = None,
    working_days: list[date] | None = None,
) -> date | None:
    """
    計画暦日にタイムライン未割付の行について、配台日スライド先を決める。
    優先: df_sim 上の割付暦日（計画日より後）→ 翌稼働日。前倒し不可。
    タイムライン上の最古実績日へ一括寄せはしない（複数暦日が同一日に潰れるのを防ぐ）。
    """
    best: date | None = None
    if df_sim is not None and not getattr(df_sim, "empty", True):
        for _, simr in df_sim.iterrows():
            stid = _interactive_norm_cell(simr.get(TASK_COL_TASK_ID))
            smach = _interactive_norm_cell(simr.get(TASK_COL_MACHINE_NAME))
            if stid != tid or smach != mach:
                continue
            if not _interactive_row_has_timeline_meta(simr.to_dict()):
                continue
            assign_dd = _interactive_parse_dispatch_date_cell(simr.get("配台日"))
            st_day_s = _iso_date_from_dispatch_table_datetime_cell(simr.get("加工開始日時"))
            if st_day_s:
                try:
                    assign_dd = date.fromisoformat(st_day_s)
                except (TypeError, ValueError):
                    pass
            if assign_dd is None or assign_dd <= plan_dd:
                continue
            if best is None or assign_dd < best:
                best = assign_dd
    if best is not None:
        return best
    nxt = _first_working_day_strictly_after(plan_dd, working_days)
    if nxt > plan_dd:
        return nxt
    return None
def _interactive_editor_plan_calendar_index(
    json_rows: list,
) -> tuple[
    dict[tuple[str, str, str, date], float],
    set[tuple[str, str, str]],
    dict[tuple[str, str, str], set[date]],
]:
    """手動修正 JSON から (依頼NO, 工程, 機械, 配台日)→目標 m と暦日集合を構築する。"""
    plan_by_key: dict[tuple[str, str, str, date], float] = {}
    tasks: set[tuple[str, str, str]] = set()
    days_by_task: dict[tuple[str, str, str], set[date]] = defaultdict(set)
    for r in json_rows or []:
        if not isinstance(r, dict):
            continue
        tid = _interactive_norm_cell(r.get(TASK_COL_TASK_ID)) or _interactive_norm_cell(
            r.get("タスクID")
        )
        proc = _interactive_dispatch_target_process_key(r.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(r.get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(r.get("配台日"))
        try:
            qty = (
                float(str(r.get("当日配台数量")).replace(",", "").strip())
                if r.get("当日配台数量") not in (None, "")
                else 0.0
            )
        except (TypeError, ValueError):
            qty = 0.0
        if not tid or not mach or dd is None or qty <= 1e-9:
            continue
        triple = (tid, proc, mach)
        k = (tid, proc, mach, dd)
        plan_by_key[k] = plan_by_key.get(k, 0.0) + float(qty)
        tasks.add(triple)
        days_by_task[triple].add(dd)
    return plan_by_key, tasks, days_by_task
def _interactive_dispatch_row_calendar_key(
    row,
) -> tuple[str, str, str, date] | None:
    tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
    proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
    mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
    dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
    if not tid or not mach or dd is None:
        return None
    return (tid, proc, mach, dd)
def _interactive_restore_editor_plan_calendar_rows(
    df_out: pd.DataFrame,
    json_rows: list,
    json_columns: list | None,
    *,
    plan_by_key: dict[tuple[str, str, str, date], float] | None = None,
    editor_tasks: set[tuple[str, str, str]] | None = None,
) -> pd.DataFrame:
    """
    配台試行の後処理（スライド・統合・枝刈り）後も、手動修正 JSON の暦日別行を正とする。
    タイムライン由来の実配台数量・加工開始終了は同一暦日キーで引き継ぐ。
    """
    if df_out is None or getattr(df_out, "empty", True) or not json_rows:
        return df_out
    if plan_by_key is None or editor_tasks is None:
        plan_by_key, editor_tasks, _ = _interactive_editor_plan_calendar_index(json_rows)
    if not plan_by_key or not editor_tasks:
        return df_out

    plan_col = "当日配台数量"
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    meta_cols = ("加工開始日時", "加工終了日時", "メンバー名")

    lookup: dict[tuple[str, str, str, date], dict] = {}
    for pos in range(len(df_out)):
        k = _interactive_dispatch_row_calendar_key(df_out.iloc[pos].to_dict())
        if k is not None:
            lookup[k] = df_out.iloc[pos].to_dict()

    df_json = _dataframe_from_interactive_dispatch_json_rows(
        json_rows,
        json_columns,
        fallback_columns_from=df_out,
    )
    restored_records: list[dict] = []
    for pos in range(len(df_json)):
        row = df_json.iloc[pos].to_dict()
        k = _interactive_dispatch_row_calendar_key(row)
        if k is None or k not in plan_by_key:
            continue
        tid, proc, mach, dd = k
        if (tid, proc, mach) not in editor_tasks:
            continue
        out_row = dict(row)
        out_row["配台日"] = _norm_ymd(dd)
        out_row[plan_col] = float(plan_by_key[k])
        src = lookup.get(k)
        if src:
            if actual_col in out_row and actual_col in src:
                try:
                    out_row[actual_col] = float(src.get(actual_col) or 0.0)
                except (TypeError, ValueError):
                    out_row[actual_col] = 0.0
            for mc in meta_cols:
                sv = src.get(mc)
                if sv is not None and str(sv).strip():
                    out_row[mc] = sv
        restored_records.append(out_row)

    other_records: list[dict] = []
    for pos in range(len(df_out)):
        row = df_out.iloc[pos].to_dict()
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        if (tid, proc, mach) in editor_tasks:
            continue
        other_records.append(row)

    if not restored_records:
        return df_out
    combined = restored_records + other_records
    df_new = pd.DataFrame(combined)
    for c in df_out.columns:
        if c not in df_new.columns:
            df_new[c] = ""
    return df_new.reindex(columns=list(df_out.columns))
def _interactive_slide_unassigned_plan_dispatch_dates(
    df_out: pd.DataFrame,
    df_sim: pd.DataFrame | None,
    *,
    meters_done: dict[tuple[str, str, str, date], float] | None = None,
    working_days: list[date] | None = None,
    editor_days_by_task: dict[tuple[str, str, str], set[date]] | None = None,
) -> pd.DataFrame:
    """
    計画暦日にタイムライン未割付、または実績が別暦日にある行について、
    配台日をタイムライン実績日または翌稼働日へスライドする。
    スライド先に同一 (依頼NO, 機械名, 配台日) 行があれば当日配台数量を合算し、元行は 0 にする。
    """
    if df_out is None or getattr(df_out, "empty", True):
        return df_out
    plan_col = "当日配台数量"
    if plan_col not in df_out.columns:
        return df_out
    out = df_out.copy()
    row_index: dict[tuple[str, str, str], int] = {}
    for pos in range(len(out)):
        tid = _interactive_norm_cell(out.iloc[pos].get(TASK_COL_TASK_ID))
        mach = _interactive_norm_cell(out.iloc[pos].get(TASK_COL_MACHINE_NAME))
        dd = _interactive_parse_dispatch_date_cell(out.iloc[pos].get("配台日"))
        if not tid or not mach or dd is None:
            continue
        row_index[(tid, mach, dd.isoformat())] = pos
    slides: list[dict] = []
    for pos in range(len(out)):
        row = out.iloc[pos].to_dict()
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        plan_dd = _interactive_parse_dispatch_date_cell(row.get("配台日"))
        try:
            plan_qty = float(row.get(plan_col) or 0.0)
        except (TypeError, ValueError):
            plan_qty = 0.0
        if not tid or not mach or plan_dd is None or plan_qty <= 1e-9:
            continue
        if editor_days_by_task and len(editor_days_by_task.get((tid, proc, mach), ())) >= 2:
            continue
        needs_slide = _interactive_row_needs_dispatch_date_slide(
            row,
            tid=tid,
            proc=proc,
            mach=mach,
            plan_dd=plan_dd,
            plan_qty=plan_qty,
            meters_done=meters_done,
        )
        if not needs_slide:
            continue
        slide_dd = _interactive_resolve_slide_target_dispatch_date(
            tid,
            proc,
            mach,
            plan_dd,
            df_sim=df_sim,
            meters_done=meters_done,
            working_days=working_days,
        )
        if slide_dd is None or slide_dd <= plan_dd:
            continue
        # 後ろ倒し: 当該行の配台日を移動する（既存行へ plan m を合算しない）。
        # 合算すると複数暦日の計画が 23700 等に潰れ、別日の未割付 meta_miss が残る。
        try:
            out.at[out.index[pos], "配台日"] = _norm_ymd(slide_dd)
            row_index[(tid, mach, slide_dd.isoformat())] = pos
        except Exception:
            continue
        slides.append(
            {
                "task_id": tid,
                "machine": mach,
                "from_date": plan_dd.isoformat(),
                "to_date": slide_dd.isoformat(),
                "plan_m": plan_qty,
            }
        )
    if slides:
        logging.info(
            "インタラクティブ配台試行: 計画暦日未割付 %s 行の配台日をスライドしました。",
            len(slides),
        )
    return out
def _interactive_prune_orphan_zero_plan_dispatch_rows(df_out: pd.DataFrame) -> pd.DataFrame:
    """
    同一 (依頼NO, 機械名) に正の当日配台数量があるとき、目標 0 の暦日行を除去する。
    配台日スライド後に残る旧計画暦日の幽霊行を JSON から落とす。
    タイムライン実配台（実配台数量>0）の翌暦日行は残す。
    """
    if df_out is None or getattr(df_out, "empty", True):
        return df_out
    plan_col = "当日配台数量"
    actual_col = INTERACTIVE_DISPATCH_ACTUAL_QTY_COL
    if plan_col not in df_out.columns:
        return df_out
    eps = 1e-9
    has_positive: set[tuple[str, str, str]] = set()
    for pos in range(len(df_out)):
        tid = _interactive_norm_cell(df_out.iloc[pos].get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(
            df_out.iloc[pos].get(TASK_COL_MACHINE)
        )
        mach = _interactive_norm_cell(df_out.iloc[pos].get(TASK_COL_MACHINE_NAME))
        try:
            q = float(df_out.iloc[pos].get(plan_col) or 0.0)
        except (TypeError, ValueError):
            q = 0.0
        if tid and proc and mach and q > eps:
            has_positive.add((tid, proc, mach))
    if not has_positive:
        return df_out
    keep: list[int] = []
    for pos in range(len(df_out)):
        row = df_out.iloc[pos]
        tid = _interactive_norm_cell(row.get(TASK_COL_TASK_ID))
        proc = _interactive_dispatch_target_process_key(row.get(TASK_COL_MACHINE))
        mach = _interactive_norm_cell(row.get(TASK_COL_MACHINE_NAME))
        try:
            q = float(row.get(plan_col) or 0.0)
        except (TypeError, ValueError):
            q = 0.0
        if q > eps:
            keep.append(pos)
            continue
        if tid and proc and mach and (tid, proc, mach) in has_positive:
            if actual_col in df_out.columns:
                try:
                    act_q = float(row.get(actual_col) or 0.0)
                except (TypeError, ValueError):
                    act_q = 0.0
                if act_q > eps:
                    keep.append(pos)
            continue
        keep.append(pos)
    if len(keep) == len(df_out):
        return df_out
    return df_out.iloc[keep].reset_index(drop=True)
def _interactive_dispatch_trial_use_editor_rows_for_result_table(
    df_sim: pd.DataFrame,
    json_rows: list | None,
    json_columns: list | None,
    *,
    interactive_dispatch_targets: dict | None = None,
    timeline_events: list | None = None,
    task_queue: list | None = None,
    working_days: list[date] | None = None,
) -> pd.DataFrame:
    """
    インタラクティブ配台試行: 編集タブの入力 JSON rows を結果_配台表の行構成の正とする。
    ユーザーが暦日行を手動統合した場合でも、配台試行後に分割へ戻さない。
    タイムライン実配台 m は「実配台数量」列へ反映する（「当日配台数量」＝編集目標は維持）。JSON に無い暦日行は追補する。
    計画暦日にタイムライン未割付の行は、実績暦日または翌稼働日へ配台日を自動スライドする。
    """
    if not json_rows or not isinstance(json_rows, list):
        return df_sim
    if not _interactive_dispatch_trial_env_active():
        return df_sim
    editor_plan_by_key, editor_tasks, editor_days_by_task = (
        _interactive_editor_plan_calendar_index(json_rows)
    )
    df_out = _dataframe_from_interactive_dispatch_json_rows(
        json_rows,
        json_columns,
        fallback_columns_from=df_sim,
    )
    df_out = _overlay_timeline_meta_onto_interactive_dispatch_df(df_out, df_sim)
    df_out = _interactive_merge_actual_dispatch_qty_from_timeline_table(df_out, df_sim)
    if interactive_dispatch_targets and timeline_events is not None and task_queue is not None:
        _md_reco = _stage35_merge_floor_into_meters_done(
            _interactive_trial_meters_done_by_timeline_calendar_date(
                timeline_events,
                task_queue,
            )
        )
        df_out = _interactive_apply_recomputed_actual_qty_to_dispatch_df(df_out, _md_reco)
        logging.info(
            "インタラクティブ配台試行: 実配台数量をタイムライン再集計（配台日キー解決）で %s キー分反映しました。",
            len(_md_reco),
        )
    else:
        _md_reco = None
    df_out = _interactive_zero_actual_qty_without_timeline_meta(
        df_out, preserve_meters_done=_md_reco
    )
    df_out = _interactive_slide_unassigned_plan_dispatch_dates(
        df_out,
        df_sim,
        meters_done=_md_reco,
        working_days=working_days,
        editor_days_by_task=editor_days_by_task,
    )
    df_out = _overlay_timeline_meta_onto_interactive_dispatch_df(df_out, df_sim)
    df_out = _interactive_merge_actual_dispatch_qty_from_timeline_table(
        df_out, df_sim, append_missing_timeline_days=True
    )
    _before_cons = len(df_out)
    df_out = _interactive_consolidate_duplicate_plan_dispatch_rows(df_out)
    if len(df_out) != _before_cons:
        logging.info(
            "インタラクティブ配台試行: 同一暦日キーの重複行を %s 行に集約しました。",
            _before_cons - len(df_out),
        )
    if _md_reco is not None:
        df_out = _interactive_apply_recomputed_actual_qty_to_dispatch_df(df_out, _md_reco)
    df_out = _interactive_zero_actual_qty_without_timeline_meta(
        df_out, preserve_meters_done=_md_reco
    )
    df_out = _interactive_prune_orphan_zero_plan_dispatch_rows(df_out)
    if timeline_events is not None and task_queue is not None:
        _plan_want = _interactive_aggregate_plan_dispatch_targets_from_df(df_out)
        if _plan_want:
            _md_post_slide = _stage35_merge_floor_into_meters_done(
                _interactive_trial_meters_done_by_timeline_calendar_date(
                    timeline_events,
                    task_queue,
                )
            )
            df_out = _interactive_apply_recomputed_actual_qty_to_dispatch_df(
                df_out, _md_post_slide
            )
            df_out = _interactive_zero_actual_qty_without_timeline_meta(
                df_out, preserve_meters_done=_md_post_slide
            )
            try:
                _LAST_INTERACTIVE_TRIAL_METERS_DONE_SNAPSHOT.clear()
                _LAST_INTERACTIVE_TRIAL_METERS_DONE_SNAPSHOT.update(_md_post_slide)
            except Exception:
                pass
            logging.info(
                "インタラクティブ配台試行: 配台日スライド後 plan キーで meters_done を %s 件再集計しました。",
                len(_md_post_slide),
            )
    df_out = _interactive_restore_editor_plan_calendar_rows(
        df_out,
        json_rows,
        json_columns,
        plan_by_key=editor_plan_by_key,
        editor_tasks=editor_tasks,
    )
    logging.info(
        "インタラクティブ配台試行: 結果_配台表は入力 JSON rows を採用しました（%s 行）。",
        len(df_out),
    )
    return df_out
def _write_dispatch_table_standalone_json(df_dispatch: pd.DataFrame, target_dir: str) -> str | None:
    """
    結果_配台表と同一内容を UTF-8 JSON に書く（xlsx 動的生成と同データソース）。
    PM_AI_RESULT_DISPATCH_TABLE_JSON=0/false/no で無効化可能。
    """
    try:
        off = (os.environ.get("PM_AI_RESULT_DISPATCH_TABLE_JSON") or "").strip().lower()
        if off in ("0", "false", "no", "off", "none"):
            return None
        if df_dispatch is None or getattr(df_dispatch, "empty", True):
            return None
        if not target_dir:
            return None
        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as mk_e:
            logging.warning("結果_配台表.json: 出力先フォルダを作成できません: %s (%s)", target_dir, mk_e)
            return None
        if not os.path.isdir(target_dir):
            return None
        out_path = os.path.join(target_dir, RESULT_DISPATCH_TABLE_JSON_FILENAME)
        try:
            if os.path.isfile(out_path):
                os.remove(out_path)
        except Exception:
            pass
        rows = json.loads(
            df_dispatch.to_json(orient="records", date_format="iso", double_precision=15)
        )
        payload = {
            "format_version": 1,
            "sheet_name": RESULT_DISPATCH_TABLE_SHEET_NAME,
            "excel_table_name": RESULT_DISPATCH_TABLE_EXCEL_TABLE_NAME,
            "columns": list(df_dispatch.columns),
            "row_count": int(len(df_dispatch)),
            "rows": rows,
        }
        p_out = pathlib.Path(target_dir) / RESULT_DISPATCH_TABLE_JSON_FILENAME
        p_out.parent.mkdir(parents=True, exist_ok=True)
        text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
        p_out.write_text(text, encoding="utf-8", newline="\n")
        return str(p_out)
    except Exception as e:
        logging.warning("結果_配台表.json: 出力に失敗しました: %s", e)
        return None
def _norm_ymd(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    try:
        if pd.api.types.is_scalar(v) and pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, datetime):
        try:
            return v.date().strftime("%Y/%m/%d")
        except (ValueError, OSError, TypeError):
            return ""
    if isinstance(v, date):
        try:
            return v.strftime("%Y/%m/%d")
        except (ValueError, OSError, TypeError):
            return ""
    s = str(v).strip()
    if not s or s.lower() in ("nan", "nat"):
        return ""
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.isna(ts):
            return s
        if isinstance(ts, pd.Timestamp):
            return ts.to_pydatetime().date().strftime("%Y/%m/%d")
    except Exception:
        pass
    return s
def _gap_minutes_until_next_break_start(dt, breaks_merged) -> float | None:
    """dt 以降に始まる最初の休憩開始までの分。無ければ None。"""
    if not isinstance(dt, datetime) or not breaks_merged:
        return None
    best: float | None = None
    for item in breaks_merged:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            continue
        bs = item[0]
        if not isinstance(bs, datetime):
            continue
        if bs < dt:
            continue
        m = (bs - dt).total_seconds() / 60.0
        if best is None or m < best:
            best = float(m)
    return best
def _pick_skilled_op_for_changeover_interval(
    machine_proc: str,
    machine_name: str,
    skills_dict: dict,
    daily_status: dict,
) -> str | None:
    """
    当日 eligible のごう」当該工程+機械で OP スキルを挝つ者のごう優先度は最尝の1坝。
    日次始業の休憩スキップに用いる（avail_dt は見ない）。
    """
    cands: list[tuple[int, str]] = []
    proc = (machine_proc or "").strip()
    mnm = (machine_name or "").strip()
    for mem, st in daily_status.items():
        if not st.get("eligible_for_assignment", st.get("is_working", False)):
            continue
        srow = skills_dict.get(mem, {})
        if proc and mnm:
            v = srow.get(f"{proc}+{mnm}", "")
        elif mnm:
            v = srow.get(mnm, "")
        elif proc:
            v = srow.get(proc, "")
        else:
            v = ""
        role, prio = parse_op_as_skill_cell(v)
        if role == "OP":
            cands.append((prio, mem))
    if not cands:
        return None
    return min(cands)[1]
def _stage2_calendar_anchor_datetime(day_d: date) -> datetime:
    """
    日次始業のカレンダ下限（定常開始 A15 が読めていればそれ、無ければ工場開始 DEFAULT_START_TIME）。
    """
    st_t = (
        _STAGE2_REGULAR_SHIFT_START
        if _STAGE2_REGULAR_SHIFT_START is not None
        else DEFAULT_START_TIME
    )
    return datetime.combine(day_d, st_t)
def _omit_machine_daily_startup_for_data_extraction_day(current_date: date) -> bool:
    """
    データ抽出当日で、かつ加工計画DATAのデータ抽出日時の時刻が定常開始（A15、無ければ工場既定開始）以上のとき、
    その暦日のみ日次始業準備ブロックを置かない（抽出が日付のみのときは時刻 0:00 とみなし、通常は定常前のため置く）。
    """
    ext = _STAGE2_DATA_EXTRACTION_DATETIME
    if ext is None or current_date != ext.date():
        return False
    reg_t = (
        _STAGE2_REGULAR_SHIFT_START
        if _STAGE2_REGULAR_SHIFT_START is not None
        else DEFAULT_START_TIME
    )
    return ext.time() >= reg_t
def _eligible_ops_sorted_for_daily_startup(
    machine_proc: str,
    machine_name: str,
    skills_dict: dict,
    daily_status: dict,
) -> list[str]:
    """当日 eligible かつ当該工程+機械で OP の者を優先度昇順に列挙。"""
    cands: list[tuple[int, str]] = []
    proc = (machine_proc or "").strip()
    mnm = (machine_name or "").strip()
    for mem, st in (daily_status or {}).items():
        if not st.get("eligible_for_assignment", st.get("is_working", False)):
            continue
        srow = (skills_dict or {}).get(mem, {})
        if proc and mnm:
            v = srow.get(f"{proc}+{mnm}", "")
        elif mnm:
            v = srow.get(mnm, "")
        elif proc:
            v = srow.get(proc, "")
        else:
            v = ""
        role, prio = parse_op_as_skill_cell(v)
        if role == "OP":
            cands.append((prio, str(mem).strip()))
    cands.sort(key=lambda x: (x[0], x[1]))
    return [m for _p, m in cands if m]
def _daily_startup_required_count_for_placement(machine_name: str) -> int:
    """壁時計ブロックを置くときに同時に確保すべき OP 数（未指定は 1 名）。"""
    need_n = _lookup_daily_startup_required_staff(machine_name, None)
    if need_n <= 0:
        return 1
    return int(need_n)
def _count_ops_covering_wall_interval(
    members_sorted: list[str],
    daily_status: dict,
    st: datetime,
    ed: datetime,
) -> int:
    n = 0
    for m in members_sorted:
        if _member_covers_interval_no_break_overlap(daily_status, m, st, ed):
            n += 1
    return n
def _earliest_daily_startup_wall_start(
    *,
    current_date: date,
    prev_machining_end_dt: datetime,
    machine_name: str,
    machine_proc: str,
    skills_dict: dict,
    daily_status: dict,
    su_minutes: int,
) -> datetime | None:
    """
    壁時計 su 分の日次始業を、当該機械の OP 母集団が勤務帯で一括覆える最早の開始時刻。
    見つからないとき None（呼び出し側でカレンダ下限にフォールバック）。
    """
    if su_minutes <= 0:
        return None
    t_lo = max(prev_machining_end_dt, _stage2_calendar_anchor_datetime(current_date))
    members = _eligible_ops_sorted_for_daily_startup(
        machine_proc, machine_name, skills_dict, daily_status
    )
    req = _daily_startup_required_count_for_placement(machine_name)
    if not members:
        return None
    req = min(req, len(members))
    day_hi = datetime.combine(current_date, DEFAULT_END_TIME)
    search_limit = day_hi + timedelta(hours=18)
    max_iter = int((search_limit - t_lo).total_seconds() // 60) + su_minutes + 120
    max_iter = max(0, min(max_iter, 2880))
    for off in range(max_iter + 1):
        st = t_lo + timedelta(minutes=off)
        ed = st + timedelta(minutes=su_minutes)
        if ed > search_limit:
            break
        if _count_ops_covering_wall_interval(members, daily_status, st, ed) >= req:
            return st
    return None
def _daily_startup_segment_start_end(
    *,
    prev_machining_end_dt: datetime,
    current_date: date,
    machine_name: str,
    machine_proc: str,
    machine_occ_key: str,
    machine_handoff: dict,
    skills_dict: dict | None,
    daily_status: dict | None,
    daily_startup_by_machine: dict[str, int] | None = None,
) -> tuple[datetime, datetime] | None:
    """
    当日先頭占有・日次始業が有効なとき [開始, 終了) を返す。不要なら None。
    開始はカレンダ下限と機械占有のうち遅い方から、OP 勤務帯で同時に覆える最早に寄せる。
    """
    mach_occ = str(machine_occ_key or "").strip()
    if not mach_occ:
        return None
    mto = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    if mach_occ in mto:
        return None
    su = _lookup_daily_startup_minutes(machine_name, daily_startup_by_machine)
    if su <= 0:
        return None
    if _omit_machine_daily_startup_for_data_extraction_day(current_date):
        return None
    cal_anchor = _stage2_calendar_anchor_datetime(current_date)
    t_lo = max(prev_machining_end_dt, cal_anchor)
    reg_start: datetime | None = None
    if daily_status and skills_dict:
        reg_start = _earliest_daily_startup_wall_start(
            current_date=current_date,
            prev_machining_end_dt=prev_machining_end_dt,
            machine_name=machine_name,
            machine_proc=str(machine_proc or "").strip(),
            skills_dict=skills_dict,
            daily_status=daily_status,
            su_minutes=su,
        )
    if reg_start is None:
        reg_start = t_lo
    reg_end = reg_start + timedelta(minutes=su)
    return reg_start, reg_end
def _machine_effective_floor_timedelta_only(
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_avail_dt: dict,
    machine_handoff: dict,
    machine_day_floor: datetime,
    abolish_limits: bool,
    *,
    daily_startup_by_machine: dict[str, int] | None = None,
    current_date: date | None = None,
    daily_status: dict | None = None,
    skills_dict: dict | None = None,
    machine_proc: str | None = None,
) -> datetime:
    """スキル OP を拾わないときのフォールバック（壁時計に分を足す。日次始業は工場稼働開始 A12 基準）。"""
    if abolish_limits:
        return machine_day_floor
    mf = machine_avail_dt.get(machine_occ_key, machine_day_floor)
    mto = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    if machine_occ_key not in mto:
        su = _lookup_daily_startup_minutes(machine_name, daily_startup_by_machine)
        if su:
            cd = current_date if current_date is not None else machine_day_floor.date()
            se = _daily_startup_segment_start_end(
                prev_machining_end_dt=mf,
                current_date=cd,
                machine_name=machine_name,
                machine_proc=str(machine_proc or "").strip(),
                machine_occ_key=machine_occ_key,
                machine_handoff=machine_handoff,
                skills_dict=skills_dict,
                daily_status=daily_status,
                daily_startup_by_machine=daily_startup_by_machine,
            )
            if se:
                mf = max(mf, se[1])
            elif not _omit_machine_daily_startup_for_data_extraction_day(cd):
                reg_end = _stage2_calendar_anchor_datetime(cd) + timedelta(minutes=su)
                mf = max(mf, reg_end)
    return mf
def _machining_events_same_occ_day_sorted(
    timeline_events: list,
    current_date: date,
    machine_occ_key: str,
) -> list[dict]:
    occ = str(machine_occ_key or "").strip()
    if not occ:
        return []
    out: list[dict] = []
    for ev in timeline_events or []:
        if ev.get("date") != current_date:
            continue
        if str(ev.get("machine_occupancy_key") or "").strip() != occ:
            continue
        if not _is_machining_timeline_event(ev):
            continue
        st = ev.get("start_dt")
        ed = ev.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime) or ed <= st:
            continue
        out.append(ev)
    out.sort(key=lambda e: (e["start_dt"], e["end_dt"]))
    return out
def _machining_timeline_event_min_end_dt(ev: dict) -> datetime | None:
    """当該加工イベントを業務上の最小量まで短くしたときの終了時刻（これ以上は短縮しない）。"""
    s0 = ev.get("start_dt")
    e0 = ev.get("end_dt")
    if not isinstance(s0, datetime) or not isinstance(e0, datetime) or e0 <= s0:
        return None
    br = merge_time_intervals(list(ev.get("breaks") or []))
    eff = float(ev.get("eff_time_per_unit") or 0.0)
    units = float(ev.get("units_done") or 0.0)
    if eff <= 0 or units <= 0:
        return e0
    min_u = 1.0 if units >= 1.0 else units
    min_wm = max(1, int(math.ceil(min_u * eff)))
    end_limit = e0 + timedelta(days=1)
    e_min, act, rem = calculate_end_time(s0, min_wm, br, end_limit)
    if rem > 0 or act < min_wm:
        return e0
    return e_min
def _cleanup_full_duration_fits_from_start(
    cleanup_start: datetime,
    cleanup_minutes: int,
    breaks_merged: list,
    shift_end: datetime,
) -> bool:
    if cleanup_minutes <= 0:
        return True
    if not isinstance(cleanup_start, datetime) or not isinstance(shift_end, datetime):
        return False
    _ce, act, rem = calculate_end_time(
        cleanup_start, cleanup_minutes, breaks_merged, shift_end
    )
    return rem <= 0 and act >= cleanup_minutes
def _lookup_changeover_minutes_for_eq(
    eq_line: str,
    by_dict: object | None,
) -> tuple[int, int]:
    """互換: (依頼切替準備分, 後始末分)。後始末は直前 eq_line の工程+機械で lookup。"""
    proc, mn = _normalize_proc_machine_for_prep_lookup("", "", eq_line=str(eq_line or ""))
    cu = _lookup_post_machining_cleanup_minutes(proc, mn, eq_line=str(eq_line or ""))
    return 0, cu
def _needs_request_switch_prep(
    machine_handoff: dict,
    machine_occ_key: str,
    current_date: date,
    task_id: str,
) -> bool:
    mach_occ = str(machine_occ_key or "").strip()
    if not mach_occ:
        return False
    machining_today_occ = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    last_tid = str((machine_handoff.get("last_tid") or {}).get(mach_occ, "") or "").strip()
    cur_tid = str(task_id or "").strip()
    last_d = (machine_handoff.get("last_machining_date") or {}).get(mach_occ)
    return (
        bool(last_tid)
        and bool(cur_tid)
        and last_tid != cur_tid
        and last_d == current_date
        and mach_occ in machining_today_occ
    )
def _team_start_is_immediate_post_break_resume(
    team_start: datetime,
    team_breaks: list,
) -> bool:
    if not isinstance(team_start, datetime):
        return False
    for item in team_breaks or []:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            continue
        be = item[1]
        if isinstance(be, datetime) and team_start == be:
            return True
    return False
def _prep_segments_immediately_before_machining(
    *,
    machining_start: datetime,
    prep_minutes: int,
    event_kind: str,
    eq_line: str,
    machine_occ_key: str,
    task_id: str = "",
) -> tuple[datetime, list[dict]]:
    """加工開始直前に [開始, 開始+分) の準備ブロックを置き、加工開始をその終了にずらす。"""
    if prep_minutes <= 0 or not isinstance(machining_start, datetime):
        return machining_start, []
    prep_start = machining_start
    prep_end = prep_start + timedelta(minutes=int(prep_minutes))
    seg = {
        "start_dt": prep_start,
        "end_dt": prep_end,
        "op": "",
        "event_kind": str(event_kind or "").strip(),
        "machine": str(eq_line or "").strip(),
        "machine_occupancy_key": str(machine_occ_key or "").strip(),
    }
    return prep_end, [seg]
def _prep_segments_from_anchor(
    *,
    anchor: datetime,
    prep_minutes: int,
    event_kind: str,
    eq_line: str,
    machine_occ_key: str,
) -> tuple[datetime, list[dict]]:
    """anchor から [anchor, anchor+分) の壁時計ブロックを置き、終了時刻を返す。"""
    if prep_minutes <= 0 or not isinstance(anchor, datetime):
        return anchor, []
    seg_end = anchor + timedelta(minutes=int(prep_minutes))
    seg = {
        "start_dt": anchor,
        "end_dt": seg_end,
        "op": "",
        "event_kind": str(event_kind or "").strip(),
        "machine": str(eq_line or "").strip(),
        "machine_occupancy_key": str(machine_occ_key or "").strip(),
    }
    return seg_end, [seg]
def _resolve_prev_machining_end_for_request_switch(
    *,
    machine_handoff: dict,
    machine_occ_key: str,
    explicit: datetime | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_floor: datetime | None = None,
) -> datetime | None:
    """
    依頼NO切替時の後始末アンカー（直前加工終了）。
    handoff の last_machining_dt が無いときは machine_avail_dt を参照する。
    """
    mach_occ = str(machine_occ_key or "").strip()
    if not mach_occ:
        return None
    if isinstance(explicit, datetime):
        return explicit
    lm = (machine_handoff.get("last_machining_dt") or {}).get(mach_occ)
    if isinstance(lm, datetime):
        return lm
    machining_today = machine_handoff.get("machining_today_occ") or set()
    if mach_occ not in machining_today:
        return None
    if machine_avail_dt is not None:
        av = machine_avail_dt.get(mach_occ)
        if isinstance(av, datetime):
            if machine_day_floor is None or av > machine_day_floor:
                return av
    return None


def _resolve_prev_machining_end_for_roll_prep(
    machine_handoff: dict,
    machine_occ_key: str,
    prev_from_handoff: datetime | None,
    machine_avail_dt: dict | None,
    machine_day_floor: datetime | None,
    machine_avail_before_changeover: datetime | None = None,
) -> datetime | None:
    """
    依頼切替用の直前加工終了。handoff → changeover 前の machine_avail の順で補完する。
    """
    if isinstance(prev_from_handoff, datetime):
        return prev_from_handoff
    mach_occ = str(machine_occ_key or "").strip()
    machining_today = machine_handoff.get("machining_today_occ") or set()
    if isinstance(machine_avail_before_changeover, datetime):
        if mach_occ and mach_occ in machining_today:
            if (
                machine_day_floor is None
                or machine_avail_before_changeover > machine_day_floor
            ):
                return machine_avail_before_changeover
    return _resolve_prev_machining_end_for_request_switch(
        machine_handoff=machine_handoff,
        machine_occ_key=machine_occ_key,
        explicit=None,
        machine_avail_dt=machine_avail_dt,
        machine_day_floor=machine_day_floor,
    )


def _roll_prep_segments_for_assign(
    *,
    team_start: datetime,
    team_breaks: list,
    machine_handoff: dict,
    machine_occ_key: str,
    current_date: date,
    task_id: str,
    machine_proc: str,
    machine_name: str,
    eq_line: str,
    abolish_limits: bool,
    prev_machining_end: datetime | None = None,
    prev_eq_line: str = "",
    machine_avail_dt: dict | None = None,
    machine_day_floor: datetime | None = None,
) -> tuple[datetime, list[dict]]:
    if abolish_limits:
        return team_start, []
    segments: list[dict] = []
    ts = team_start
    _proc_lu, _mn_lu = _normalize_proc_machine_for_prep_lookup(
        machine_proc, machine_name, eq_line=eq_line
    )
    _need_sw = _needs_request_switch_prep(
        machine_handoff, machine_occ_key, current_date, task_id
    )
    _post_break = _team_start_is_immediate_post_break_resume(ts, team_breaks)
    if _need_sw:
        _prev_end = _resolve_prev_machining_end_for_request_switch(
            machine_handoff=machine_handoff,
            machine_occ_key=machine_occ_key,
            explicit=prev_machining_end,
            machine_avail_dt=machine_avail_dt,
            machine_day_floor=machine_day_floor,
        )
        _prev_proc, _prev_mn = _normalize_proc_machine_for_prep_lookup(
            "", "", eq_line=str(prev_eq_line or "").strip()
        )
        cu = _lookup_post_machining_cleanup_minutes(
            _prev_proc, _prev_mn, eq_line=str(prev_eq_line or "").strip()
        )
        bf = _lookup_request_interval_buffer_minutes(
            _proc_lu, _mn_lu, eq_line=eq_line
        )
        pm = _lookup_request_switch_prep_minutes(
            _proc_lu, _mn_lu, eq_line=eq_line
        )
        # 後始末・依頼間余裕は直前加工終了が分かるときだけ prev_end から置く。
        # prev_end 不明のとき team_start を anchor にすると、余裕が加工直後ではなく
        # 次依頼の準備直前（ギャップの中途）に付いてしまう。
        if isinstance(_prev_end, datetime):
            chain_end = _prev_end
            if cu > 0:
                chain_end, segs = _prep_segments_from_anchor(
                    anchor=chain_end,
                    prep_minutes=cu,
                    event_kind=TIMELINE_EVENT_POST_MACHINING_CLEANUP,
                    eq_line=eq_line,
                    machine_occ_key=machine_occ_key,
                )
                segments.extend(segs)
            if bf > 0:
                chain_end, segs = _prep_segments_from_anchor(
                    anchor=chain_end,
                    prep_minutes=bf,
                    event_kind=TIMELINE_EVENT_REQUEST_INTERVAL_BUFFER,
                    eq_line=eq_line,
                    machine_occ_key=machine_occ_key,
                )
                segments.extend(segs)
            ts = max(ts, chain_end)
        if pm > 0:
            ts, segs = _prep_segments_immediately_before_machining(
                machining_start=ts,
                prep_minutes=pm,
                event_kind=TIMELINE_EVENT_REQUEST_SWITCH_PREP,
                eq_line=eq_line,
                machine_occ_key=machine_occ_key,
            )
            segments.extend(segs)
    elif _post_break:
        rm = _lookup_break_resume_prep_minutes(
            _proc_lu, _mn_lu, eq_line=eq_line
        )
        if rm > 0:
            ts, segs = _prep_segments_immediately_before_machining(
                machining_start=ts,
                prep_minutes=rm,
                event_kind=TIMELINE_EVENT_BREAK_RESUME_PREP,
                eq_line=eq_line,
                machine_occ_key=machine_occ_key,
                task_id=str(task_id or "").strip(),
            )
            segments.extend(segs)
    return ts, segments
def _avail_dt_reapply_member_max_end_from_timeline(
    timeline_events: list,
    avail_dt: dict,
    members: set[str],
) -> None:
    """指定メンバーについて、タイムライン全体の終了の最大で avail_dt を上書き（短縮後の整合）。"""
    for m in members:
        mm = str(m or "").strip()
        if not mm or mm not in avail_dt:
            continue
        best: datetime | None = None
        for ev in timeline_events or []:
            names: list[str] = []
            op = str(ev.get("op") or "").strip()
            if op:
                names.append(op)
            for s in str(ev.get("sub") or "").split(","):
                s = s.strip()
                if s:
                    names.append(s)
            if mm not in names:
                continue
            ed = ev.get("end_dt")
            if isinstance(ed, datetime):
                best = ed if best is None else max(best, ed)
        if best is not None:
            avail_dt[mm] = best
def _repair_timeline_for_same_tid_prebreak_cleanup(
    *,
    timeline_events: list,
    machine_avail_dt: dict,
    machine_handoff: dict,
    current_date: date,
    machine_occ_key: str,
    next_task_id: str,
    machine_proc: str,
    machine_name: str,
    daily_status: dict,
    skills_dict: dict,
    avail_dt: dict | None,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    task_queue: list,
    machine_day_floor: datetime,
) -> bool:
    """後始末は Phase 1 無効化中（lookup ゲート）。再有効化時は EOD 確保方式へ置換予定。常に False。"""
    return False
    mach_occ = str(machine_occ_key or "").strip()
    if not mach_occ:
        return False
    machining_today_occ = machine_handoff.get("machining_today_occ") or machine_handoff.get(
        "started_today", set()
    )
    last_tid = str((machine_handoff.get("last_tid") or {}).get(mach_occ, "") or "").strip()
    cur_tid = str(next_task_id or "").strip()
    if not last_tid or not cur_tid or last_tid != cur_tid:
        return False
    last_d = (machine_handoff.get("last_machining_date") or {}).get(mach_occ)
    if last_d != current_date or mach_occ not in machining_today_occ:
        return False
    last_eq_s = str(
        (machine_handoff.get("last_eq") or {}).get(mach_occ, "") or ""
    ).strip()
    _pu, cu_prev = _lookup_changeover_minutes_for_eq(last_eq_s, None)
    if cu_prev <= 0:
        return False
    lm_end = (machine_handoff.get("last_machining_dt") or {}).get(mach_occ)
    if not isinstance(lm_end, datetime):
        return False
    last_lead = str((machine_handoff.get("last_lead_op") or {}).get(mach_occ, "") or "").strip()
    rep = _pick_skilled_op_for_changeover_interval(
        str(machine_proc or "").strip(),
        str(machine_name or "").strip(),
        skills_dict,
        daily_status,
    )
    if not last_lead:
        last_lead = str(rep or "").strip()
    if not last_lead:
        return False
    st_ld = daily_status.get(last_lead) or (daily_status.get(rep) if rep else None)
    if not st_ld:
        return False
    _brk_parts = list(st_ld.get("breaks_dt") or [])
    if rep:
        st_rep_m = daily_status.get(rep)
        if st_rep_m:
            _brk_parts.extend(list(st_rep_m.get("breaks_dt") or []))
    br_resume = merge_time_intervals(_brk_parts)
    _hit, _tf, bs_a, be_a, pre_gap = _resume_after_work_break_extended(
        lm_end, lm_end, br_resume
    )
    if not pre_gap or bs_a is None or be_a is None:
        return False
    br_c = merge_time_intervals(list(st_ld.get("breaks_dt") or []))
    end_c = st_ld["end_dt"]
    if not isinstance(end_c, datetime):
        return False
    st_inv = _find_latest_prep_start_matching_end(
        bs_a, cu_prev, br_c, machine_day_floor
    )
    if st_inv is None:
        return False
    ce_chk, act_chk, rem_chk = calculate_end_time(st_inv, cu_prev, br_c, end_c)
    if rem_chk > 0 or act_chk < cu_prev or not _dt_close_minutes(ce_chk, bs_a):
        return False
    if st_inv >= lm_end - timedelta(seconds=90):
        return False
    occ = mach_occ
    ml = _machining_events_same_occ_day_sorted(timeline_events, current_date, occ)
    if not ml:
        return False
    last_ev = ml[-1]
    e0 = last_ev["end_dt"]
    if not isinstance(e0, datetime):
        return False
    e_min = _machining_timeline_event_min_end_dt(last_ev)
    if e_min is None or not isinstance(e_min, datetime):
        return False
    best_e = st_inv
    if best_e < e_min or best_e >= e0:
        return False
    delta = e0 - best_e
    old_anchor = e0
    s0 = last_ev.get("start_dt")
    if not isinstance(s0, datetime):
        return False
    touched_members: set[str] = set()
    du = float(last_ev.get("units_done") or 0.0)
    br_ev = merge_time_intervals(list(last_ev.get("breaks") or []))
    _cap_m = max(1, int((e0 - s0).total_seconds() // 60) + 1440)
    _, wm_old, _ = calculate_end_time(s0, _cap_m, br_ev, e0)
    _, wm_new, _ = calculate_end_time(s0, _cap_m, br_ev, best_e)
    if wm_old <= 0:
        return False
    new_u = max(1e-12, du * (wm_new / wm_old))
    min_u = 1.0 if du >= 1.0 else du
    new_u = max(min_u, new_u)
    tid = str(last_ev.get("task_id") or "").strip()
    for tq in task_queue or []:
        if str(tq.get("task_id") or "").strip() != tid:
            continue
        tq["remaining_units"] = float(tq.get("remaining_units") or 0) + (du - new_u)
        for row in reversed(tq.get("assigned_history") or []):
            edh = row.get("end_dt")
            if isinstance(edh, datetime) and edh == e0:
                row["end_dt"] = best_e
                try:
                    row["done_m"] = int(float(new_u) * float(tq.get("unit_m") or 0))
                except Exception:
                    pass
                break
        break
    last_ev["end_dt"] = best_e
    last_ev["units_done"] = new_u
    for opn, sb in (
        (str(last_ev.get("op") or "").strip(), str(last_ev.get("sub") or "")),
    ):
        if opn:
            touched_members.add(opn)
        for s in sb.split(","):
            s = s.strip()
            if s:
                touched_members.add(s)
    for ev2 in timeline_events:
        if ev2.get("date") != current_date:
            continue
        if str(ev2.get("machine_occupancy_key") or "").strip() != occ:
            continue
        st2 = ev2.get("start_dt")
        ed2 = ev2.get("end_dt")
        if not isinstance(st2, datetime) or not isinstance(ed2, datetime):
            continue
        if st2 >= old_anchor:
            ev2["start_dt"] = st2 - delta
            ev2["end_dt"] = ed2 - delta
            op2 = str(ev2.get("op") or "").strip()
            if op2:
                touched_members.add(op2)
            for s in str(ev2.get("sub") or "").split(","):
                s = s.strip()
                if s:
                    touched_members.add(s)
    machine_avail_dt[occ] = best_e
    machine_handoff.setdefault("last_machining_dt", {})
    machine_handoff["last_machining_dt"][occ] = best_e
    if avail_dt is not None and touched_members:
        _avail_dt_reapply_member_max_end_from_timeline(
            timeline_events, avail_dt, touched_members
        )
    if dispatch_interval_mirror is not None:
        dispatch_interval_mirror.rebuild_from_timeline(timeline_events)
    _bump_machine_avail_after_roll_for_calendar(
        current_date,
        occ,
        machine_avail_dt,
        machine_calendar_plan_end=None,
        machine_day_floor=machine_day_floor,
    )
    return True
def _repair_timeline_shorten_machining_for_changeover_cleanup(
    *,
    timeline_events: list,
    machine_avail_dt: dict,
    machine_handoff: dict,
    current_date: date,
    machine_occ_key: str,
    next_task_id: str,
    machine_proc: str,
    machine_name: str,
    daily_status: dict,
    skills_dict: dict,
    avail_dt: dict | None,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    task_queue: list,
    machine_day_floor: datetime,
) -> bool:
    """後始末は Phase 1 無効化中（lookup ゲート）。再有効化時は EOD 確保方式へ置換予定。常に False。"""
    return False
    need, cu_prev, last_lead, last_eq_s = _changeover_need_cleanup_for_next_assign(
        machine_handoff=machine_handoff,
        machine_occ_key=machine_occ_key,
        current_date=current_date,
        cur_task_id=next_task_id,
        last_eq=None,
    )
    if not need:
        if _repair_timeline_for_same_tid_prebreak_cleanup(
            timeline_events=timeline_events,
            machine_avail_dt=machine_avail_dt,
            machine_handoff=machine_handoff,
            current_date=current_date,
            machine_occ_key=machine_occ_key,
            next_task_id=next_task_id,
            machine_proc=str(machine_proc or "").strip(),
            machine_name=str(machine_name or "").strip(),
            daily_status=daily_status,
            skills_dict=skills_dict,
            avail_dt=avail_dt,
            dispatch_interval_mirror=dispatch_interval_mirror,
            task_queue=task_queue,
            machine_day_floor=machine_day_floor,
        ):
            return True
        return False
    if cu_prev <= 0:
        return False
    rep = _pick_skilled_op_for_changeover_interval(
        str(machine_proc or "").strip(),
        str(machine_name or "").strip(),
        skills_dict,
        daily_status,
    )
    if not last_lead:
        last_lead = str(rep or "").strip()
    if not last_lead:
        return False
    st_c = daily_status.get(last_lead) or (
        daily_status.get(rep) if rep else None
    )
    if not st_c:
        return False
    br_c = merge_time_intervals(list(st_c.get("breaks_dt") or []))
    end_c = st_c["end_dt"]
    if not isinstance(end_c, datetime):
        return False

    def _cleanup_ok_at_machining_end(mach_end: datetime) -> bool:
        return _cleanup_full_duration_fits_from_start(
            mach_end, cu_prev, br_c, end_c
        )

    occ = str(machine_occ_key or "").strip()
    touched_members: set[str] = set()

    for _pass in range(64):
        ml = _machining_events_same_occ_day_sorted(
            timeline_events, current_date, occ
        )
        if not ml:
            return False
        last_ev = ml[-1]
        e0 = last_ev["end_dt"]
        if not isinstance(e0, datetime):
            return False
        if _cleanup_ok_at_machining_end(e0):
            machine_avail_dt[occ] = e0
            machine_handoff.setdefault("last_machining_dt", {})
            machine_handoff["last_machining_dt"][occ] = e0
            return True

        e_min = _machining_timeline_event_min_end_dt(last_ev)
        if e_min is None or not isinstance(e_min, datetime):
            return False
        s0 = last_ev.get("start_dt")
        if not isinstance(s0, datetime):
            return False

        lo = e_min
        hi = e0
        best_e: datetime | None = None
        while lo <= hi:
            mid = lo + (hi - lo) // 2
            if _cleanup_ok_at_machining_end(mid):
                best_e = mid
                lo = mid + timedelta(minutes=1)
            else:
                hi = mid - timedelta(minutes=1)

        if best_e is not None and isinstance(best_e, datetime) and best_e < e0:
            delta = e0 - best_e
            old_anchor = e0
            du = float(last_ev.get("units_done") or 0.0)
            br_ev = merge_time_intervals(list(last_ev.get("breaks") or []))
            _cap_m = max(1, int((e0 - s0).total_seconds() // 60) + 1440)
            _, wm_old, _ = calculate_end_time(s0, _cap_m, br_ev, e0)
            _, wm_new, _ = calculate_end_time(s0, _cap_m, br_ev, best_e)
            if wm_old <= 0:
                return False
            new_u = max(1e-12, du * (wm_new / wm_old))
            min_u = 1.0 if du >= 1.0 else du
            new_u = max(min_u, new_u)
            tid = str(last_ev.get("task_id") or "").strip()
            for t in task_queue or []:
                if str(t.get("task_id") or "").strip() != tid:
                    continue
                t["remaining_units"] = float(t.get("remaining_units") or 0) + (du - new_u)
                for row in reversed(t.get("assigned_history") or []):
                    edh = row.get("end_dt")
                    if isinstance(edh, datetime) and edh == e0:
                        row["end_dt"] = best_e
                        try:
                            row["done_m"] = int(
                                float(new_u) * float(t.get("unit_m") or 0)
                            )
                        except Exception:
                            pass
                        break
                break
            last_ev["end_dt"] = best_e
            last_ev["units_done"] = new_u
            for opn, sb in (
                (str(last_ev.get("op") or "").strip(), str(last_ev.get("sub") or ""))
            ):
                if opn:
                    touched_members.add(opn)
                for s in sb.split(","):
                    s = s.strip()
                    if s:
                        touched_members.add(s)
            for ev2 in timeline_events:
                if ev2.get("date") != current_date:
                    continue
                if str(ev2.get("machine_occupancy_key") or "").strip() != occ:
                    continue
                st2 = ev2.get("start_dt")
                ed2 = ev2.get("end_dt")
                if not isinstance(st2, datetime) or not isinstance(ed2, datetime):
                    continue
                if st2 >= old_anchor:
                    ev2["start_dt"] = st2 - delta
                    ev2["end_dt"] = ed2 - delta
                    op2 = str(ev2.get("op") or "").strip()
                    if op2:
                        touched_members.add(op2)
                    for s in str(ev2.get("sub") or "").split(","):
                        s = s.strip()
                        if s:
                            touched_members.add(s)
            machine_avail_dt[occ] = best_e
            machine_handoff.setdefault("last_machining_dt", {})
            machine_handoff["last_machining_dt"][occ] = best_e
            if avail_dt is not None and touched_members:
                _avail_dt_reapply_member_max_end_from_timeline(
                    timeline_events, avail_dt, touched_members
                )
            if dispatch_interval_mirror is not None:
                dispatch_interval_mirror.rebuild_from_timeline(timeline_events)
            _bump_machine_avail_after_roll_for_calendar(
                current_date,
                occ,
                machine_avail_dt,
                machine_calendar_plan_end=None,
                machine_day_floor=machine_day_floor,
            )
            continue

        if len(ml) < 2:
            return False
        applied_prev = False
        for shorten_idx in range(len(ml) - 2, -1, -1):
            prev_ev = ml[shorten_idx]
            eP0 = prev_ev.get("end_dt")
            sP0 = prev_ev.get("start_dt")
            if not isinstance(eP0, datetime) or not isinstance(sP0, datetime):
                continue
            last_ev2 = ml[-1]
            eL_end = last_ev2.get("end_dt")
            if not isinstance(eL_end, datetime):
                return False
            eP_min = _machining_timeline_event_min_end_dt(prev_ev)
            if eP_min is None or not isinstance(eP_min, datetime):
                continue

            def _last_end_after_shrink_prev_end(end_pe: datetime) -> datetime | None:
                if end_pe > eP0 or end_pe < eP_min:
                    return None
                dlt = eP0 - end_pe
                try:
                    return last_ev2["end_dt"] - dlt
                except Exception:
                    return None

            lo2 = eP_min
            hi2 = eP0
            best_pe: datetime | None = None
            while lo2 <= hi2:
                midp = lo2 + (hi2 - lo2) // 2
                le = _last_end_after_shrink_prev_end(midp)
                if le is not None and _cleanup_ok_at_machining_end(le):
                    best_pe = midp
                    lo2 = midp + timedelta(minutes=1)
                else:
                    hi2 = midp - timedelta(minutes=1)

            if best_pe is None or best_pe >= eP0:
                continue

            delta_p = eP0 - best_pe
            old_anchor_p = eP0
            du_p = float(prev_ev.get("units_done") or 0.0)
            br_p = merge_time_intervals(list(prev_ev.get("breaks") or []))
            _cap_p = max(1, int((eP0 - sP0).total_seconds() // 60) + 1440)
            _, wm_old_p, _ = calculate_end_time(sP0, _cap_p, br_p, eP0)
            _, wm_new_p, _ = calculate_end_time(sP0, _cap_p, br_p, best_pe)
            if wm_old_p <= 0:
                return False
            new_u_p = max(1e-12, du_p * (wm_new_p / wm_old_p))
            min_u_p = 1.0 if du_p >= 1.0 else du_p
            new_u_p = max(min_u_p, new_u_p)
            tidp = str(prev_ev.get("task_id") or "").strip()
            for t in task_queue or []:
                if str(t.get("task_id") or "").strip() != tidp:
                    continue
                t["remaining_units"] = float(t.get("remaining_units") or 0) + (
                    du_p - new_u_p
                )
                for row in reversed(t.get("assigned_history") or []):
                    edh = row.get("end_dt")
                    if isinstance(edh, datetime) and edh == eP0:
                        row["end_dt"] = best_pe
                        try:
                            row["done_m"] = int(
                                float(new_u_p) * float(t.get("unit_m") or 0)
                            )
                        except Exception:
                            pass
                        break
                break
            prev_ev["end_dt"] = best_pe
            prev_ev["units_done"] = new_u_p
            for opn, sb in (
                (str(prev_ev.get("op") or "").strip(), str(prev_ev.get("sub") or ""))
            ):
                if opn:
                    touched_members.add(opn)
                for s in sb.split(","):
                    s = s.strip()
                    if s:
                        touched_members.add(s)
            for ev2 in timeline_events:
                if ev2.get("date") != current_date:
                    continue
                if str(ev2.get("machine_occupancy_key") or "").strip() != occ:
                    continue
                st2 = ev2.get("start_dt")
                ed2 = ev2.get("end_dt")
                if not isinstance(st2, datetime) or not isinstance(ed2, datetime):
                    continue
                if st2 >= old_anchor_p:
                    ev2["start_dt"] = st2 - delta_p
                    ev2["end_dt"] = ed2 - delta_p
                    op2 = str(ev2.get("op") or "").strip()
                    if op2:
                        touched_members.add(op2)
                    for s in str(ev2.get("sub") or "").split(","):
                        s = s.strip()
                        if s:
                            touched_members.add(s)
            new_last_end = ml[-1]["end_dt"]
            if isinstance(new_last_end, datetime):
                machine_avail_dt[occ] = new_last_end
                machine_handoff.setdefault("last_machining_dt", {})
                machine_handoff["last_machining_dt"][occ] = new_last_end
            if avail_dt is not None and touched_members:
                _avail_dt_reapply_member_max_end_from_timeline(
                    timeline_events, avail_dt, touched_members
                )
            if dispatch_interval_mirror is not None:
                dispatch_interval_mirror.rebuild_from_timeline(timeline_events)
            _bump_machine_avail_after_roll_for_calendar(
                current_date,
                occ,
                machine_avail_dt,
                machine_calendar_plan_end=None,
                machine_day_floor=machine_day_floor,
            )
            applied_prev = True
            break
        if applied_prev:
            continue
        return False
    return False
def _machine_effective_floor_for_assign(
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_avail_dt: dict,
    machine_handoff: dict,
    machine_day_floor: datetime,
    abolish_limits: bool,
    *,
    daily_startup_by_machine: dict[str, int] | None = None,
    current_date: date | None = None,
    daily_status: dict | None = None,
    skills_dict: dict | None = None,
    machine_proc: str | None = None,
) -> datetime:
    """
    設備のタイムラインによける「当該ロールの加工開始」以降の下限。
    daily_status・skills_dict・current_date は权ごとしは」skills 革坈 OP の勤務・休憩に沿って
    日次始業を forward した最早加工開始。权ゝないとしは分のタイムライン加算にフォールバック。
    """
    if abolish_limits:
        return machine_day_floor
    prev_mach = machine_avail_dt.get(machine_occ_key, machine_day_floor)
    if (
        current_date is not None
        and daily_status is not None
        and skills_dict is not None
        and machine_proc is not None
    ):
        lb, _segs = _changeover_plan_segments_and_machining_lower_bound(
            prev_machining_end_dt=prev_mach,
            machine_day_floor=machine_day_floor,
            current_date=current_date,
            machine_occ_key=machine_occ_key,
            task_id=task_id,
            eq_line=eq_line,
            machine_name=machine_name,
            machine_proc=str(machine_proc or "").strip(),
            machine_handoff=machine_handoff,
            daily_status=daily_status,
            skills_dict=skills_dict,
            abolish_limits=False,
        )
        if lb is not None:
            return lb
    return _machine_effective_floor_timedelta_only(
        machine_occ_key,
        task_id,
        eq_line,
        machine_name,
        machine_avail_dt,
        machine_handoff,
        machine_day_floor,
        False,
        daily_startup_by_machine=daily_startup_by_machine,
        current_date=current_date,
        daily_status=daily_status,
        skills_dict=skills_dict,
        machine_proc=machine_proc,
    )
def _resolve_machine_changeover_floor_segments(
    *,
    abolish_all_scheduling_limits: bool,
    machine_occ_key: str,
    task_id: str,
    eq_line: str,
    machine_name: str,
    machine_proc: str,
    machine_avail_dt: dict,
    machine_day_floor: datetime,
    current_date: date,
    machine_handoff: dict,
    daily_status: dict,
    skills_dict: dict,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    timeline_events: list | None = None,
    task_queue: list | None = None,
    avail_dt: dict | None = None,
) -> tuple[datetime, list[dict], bool]:
    """
    設備の加工開始下限と」タイムライン追記用セットアップ区間。
    戻り値 (floor_dt, segments, abort)。abort は True のときは当該ロール割当を全体として棄坴れる。
    """
    if abolish_all_scheduling_limits:
        prev = machine_avail_dt.get(machine_occ_key, machine_day_floor)
        return prev, [], False
    prev_mach = machine_avail_dt.get(machine_occ_key, machine_day_floor)
    co_lb, co_segs = _changeover_plan_segments_and_machining_lower_bound(
        prev_machining_end_dt=prev_mach,
        machine_day_floor=machine_day_floor,
        current_date=current_date,
        machine_occ_key=machine_occ_key,
        task_id=task_id,
        eq_line=eq_line,
        machine_name=machine_name,
        machine_proc=str(machine_proc or "").strip(),
        machine_handoff=machine_handoff,
        daily_status=daily_status,
        skills_dict=skills_dict,
        abolish_limits=False,
    )
    if co_lb is not None and _needs_request_switch_prep(
        machine_handoff,
        machine_occ_key,
        current_date,
        task_id,
    ):
        last_eq_s = str(
            (machine_handoff.get("last_eq") or {}).get(machine_occ_key, "") or ""
        ).strip()
        _prev_proc, _prev_mn = _normalize_proc_machine_for_prep_lookup(
            "", "", eq_line=last_eq_s
        )
        _cu = _lookup_post_machining_cleanup_minutes(
            _prev_proc, _prev_mn, eq_line=last_eq_s
        )
        _bf = _lookup_request_interval_buffer_minutes(
            str(machine_proc or "").strip(),
            str(machine_name or "").strip(),
            eq_line=str(eq_line or "").strip(),
        )
        _sw_prep = _lookup_request_switch_prep_minutes(
            str(machine_proc or "").strip(),
            str(machine_name or "").strip(),
            eq_line=str(eq_line or "").strip(),
        )
        _extra = int(_cu or 0) + int(_bf or 0) + int(_sw_prep or 0)
        if _extra > 0:
            co_lb = co_lb + timedelta(minutes=_extra)
    if co_lb is None:
        if (
            _pick_skilled_op_for_changeover_interval(
                str(machine_proc or "").strip(),
                str(machine_name or "").strip(),
                skills_dict,
                daily_status,
            )
            is None
        ):
            mf = _machine_effective_floor_timedelta_only(
                machine_occ_key,
                task_id,
                eq_line,
                machine_name,
                machine_avail_dt,
                machine_handoff,
                machine_day_floor,
                False,
                current_date=current_date,
                daily_status=daily_status,
                skills_dict=skills_dict,
                machine_proc=str(machine_proc or "").strip(),
            )
            return mf, [], False
        return machine_day_floor, [], True
    if dispatch_interval_mirror is not None and co_segs:
        for seg in co_segs:
            sop = str(seg.get("op") or "").strip()
            sok = str(seg.get("machine_occupancy_key") or machine_occ_key).strip()
            st_seg = seg.get("start_dt")
            ed_seg = seg.get("end_dt")
            if not isinstance(st_seg, datetime) or not isinstance(ed_seg, datetime):
                continue
            if (
                sop
                and dispatch_interval_mirror.would_block_member(sop, st_seg, ed_seg)
            ):
                return machine_day_floor, [], True
            if (
                sok
                and dispatch_interval_mirror.would_block_equipment(
                    sok, st_seg, ed_seg
                )
            ):
                return machine_day_floor, [], True
    return co_lb, co_segs, False
def _append_changeover_segments_to_timeline(
    timeline_events: list,
    dispatch_interval_mirror: DispatchIntervalMirror | None,
    avail_dt: dict,
    daily_status: dict,
    *,
    current_date: date,
    task_id: str,
    machine_occ_key: str,
    segments: list[dict],
    machining_lead_op: str | None = None,
    machining_sub_str: str | None = None,
    machine_handoff: dict | None = None,
    skill_role_priority=None,
    machine_name_for_startup: str | None = None,
) -> None:
    """セットアップ系セグメントをタイムライン・ミラー・担当者 avail に反映。"""
    _mh = machine_handoff or {}
    _lead_m = str(machining_lead_op or "").strip()
    _sub_roll = str(machining_sub_str or "").strip()
    for seg in segments or []:
        st = seg.get("start_dt")
        ed = seg.get("end_dt")
        if not isinstance(st, datetime) or not isinstance(ed, datetime):
            continue
        m_line = str(seg.get("machine") or "").strip()
        m_occ = str(seg.get("machine_occupancy_key") or machine_occ_key).strip()
        ek = str(seg.get("event_kind") or "").strip() or TIMELINE_EVENT_MACHINING
        if (
            ek == TIMELINE_EVENT_MACHINE_DAILY_STARTUP
            and skill_role_priority is not None
            and str(machine_name_for_startup or "").strip()
        ):
            _daily_startup_fill_segment_staff(
                seg,
                machine_name=str(machine_name_for_startup or "").strip(),
                lead_op=_lead_m,
                sub_csv=_sub_roll,
                skill_role_priority=skill_role_priority,
                daily_status=daily_status,
                avail_dt=avail_dt,
                dispatch_interval_mirror=dispatch_interval_mirror,
            )
        op_seg = str(seg.get("op") or "").strip()
        sub_seg = str(seg.get("sub") or "").strip()
        op, sub = _changeover_timeline_op_sub_for_event(
            event_kind=ek,
            op_from_segment=op_seg,
            sub_from_segment=sub_seg,
            machine_occ_key=m_occ,
            machining_lead_op=_lead_m,
            machining_sub_str=_sub_roll,
            machine_handoff=_mh,
            daily_status=daily_status,
        )
        br_acc: list = []
        for nm in (op, *[_p.strip() for _p in sub.split(",") if _p.strip()]):
            if nm and nm in daily_status:
                br_acc.extend(daily_status[nm].get("breaks_dt") or [])
        br_seg = merge_time_intervals(br_acc)
        tid_ev = (
            ""
            if ek
            in (
                TIMELINE_EVENT_MACHINE_DAILY_STARTUP,
                TIMELINE_EVENT_REQUEST_SWITCH_PREP,
                TIMELINE_EVENT_POST_MACHINING_CLEANUP,
                TIMELINE_EVENT_REQUEST_INTERVAL_BUFFER,
            )
            else str(task_id or "").strip()
        )
        ev = {
            "date": current_date,
            "task_id": tid_ev,
            "machine": m_line,
            "machine_occupancy_key": m_occ,
            "op": op,
            "sub": sub,
            "start_dt": st,
            "end_dt": ed,
            "breaks": br_seg,
            "units_done": 0,
            "event_kind": ek,
        }
        timeline_events.append(ev)
        _stage2_dispatch_track_timeline_event(ev, len(timeline_events))
        if dispatch_interval_mirror is not None:
            dispatch_interval_mirror.register_from_event(ev)
        for nm in (op, *[_p.strip() for _p in sub.split(",") if _p.strip()]):
            if not nm:
                continue
            prev_a = avail_dt.get(nm, st)
            if isinstance(prev_a, datetime):
                avail_dt[nm] = max(prev_a, ed)
            else:
                avail_dt[nm] = ed
def _collect_task_ids_missed_deadline_after_day(task_queue: list, current_date: date) -> set:
    """
    当該日の終了時点で」紝期基準日（当日含む）以降なのに残量は残る依頼NO。
    「紝期日内に完靂でしなかった」= 後ゝ倒し再試行の候補。
    """
    out = set()
    eps = 1e-9
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= eps:
            continue
        db = t.get("due_basis_date")
        if db is None:
            continue
        sdr = t.get("start_date_req")
        if isinstance(sdr, date) and sdr > current_date:
            continue
        if current_date >= db:
            tid = str(t.get("task_id", "") or "").strip()
            if tid:
                out.add(tid)
    return out
def _normalize_timeline_task_id(ev: dict) -> str:
    return str(ev.get("task_id", "") or "").strip()
def _machine_occupancy_key_from_timeline_event(e: dict) -> str:
    occ = str(e.get("machine_occupancy_key") or "").strip()
    if occ:
        return occ
    mraw = str(e.get("machine") or "").strip()
    if not mraw:
        return ""
    if "+" in mraw:
        return _normalize_equipment_match_key(mraw.split("+", 1)[1])
    return _normalize_equipment_match_key(mraw)
def _snapshot_machine_handoff_state(template: dict) -> dict:
    return {
        "last_tid": dict(template["last_tid"]),
        "last_eq": dict(template["last_eq"]),
        "last_machining_dt": dict(template["last_machining_dt"]),
        "last_machining_date": dict(template["last_machining_date"]),
        "last_lead_op": dict(template["last_lead_op"]),
        "last_machining_sub": dict(template["last_machining_sub"]),
        "machining_today_occ": set(template["machining_today_occ"]),
        "started_today": set(template["started_today"]),
    }
def _reset_machine_handoff_timeline_cache() -> None:
    global _MH_HANDOFF_TIMELINE_CACHE_KEY, _MH_HANDOFF_TIMELINE_CACHE_STATE
    global _STAGE2_DISPATCH_EVENTS_BY_DATE
    _MH_HANDOFF_TIMELINE_CACHE_KEY = None
    _MH_HANDOFF_TIMELINE_CACHE_STATE = None
    _MH_HANDOFF_TIMELINE_CACHE_STATS["hit"] = 0
    _MH_HANDOFF_TIMELINE_CACHE_STATS["miss"] = 0
    _MH_HANDOFF_TIMELINE_CACHE_STATS["incremental"] = 0
    _STAGE2_DISPATCH_EVENTS_BY_DATE = defaultdict(list)
def _stage2_dispatch_track_timeline_event(ev: dict, timeline_len: int) -> None:
    """段階2配台ループ中: 暦日別インデックスと手札キャッシュの増分更新。"""
    global _MH_HANDOFF_TIMELINE_CACHE_KEY, _STAGE2_DISPATCH_EVENTS_BY_DATE
    d = ev.get("date")
    if _STAGE2_DISPATCH_EVENTS_BY_DATE is not None and isinstance(d, date):
        _STAGE2_DISPATCH_EVENTS_BY_DATE[d].append(ev)
    if (
        _MH_HANDOFF_TIMELINE_CACHE_STATE is not None
        and _MH_HANDOFF_TIMELINE_CACHE_KEY is not None
        and isinstance(d, date)
        and _MH_HANDOFF_TIMELINE_CACHE_KEY[1] == d
    ):
        _machine_handoff_merge_machining_event(_MH_HANDOFF_TIMELINE_CACHE_STATE, ev, d)
        _MH_HANDOFF_TIMELINE_CACHE_KEY = (timeline_len, d)
        _MH_HANDOFF_TIMELINE_CACHE_STATS["incremental"] += 1
def _machine_handoff_merge_machining_event(
    state: dict, e: dict, current_date: date
) -> None:
    """``_machine_handoff_state_from_timeline`` と同趣旨で 1 イベントだけ state に反映（in-place）。"""
    if not _is_machining_timeline_event(e):
        return
    ed = e.get("date")
    if not isinstance(ed, date):
        return
    occ = _machine_occupancy_key_from_timeline_event(e)
    if not occ:
        return
    end_dt = e.get("end_dt")
    if end_dt is None or not hasattr(end_dt, "replace"):
        return
    if ed == current_date:
        state["machining_today_occ"].add(occ)
        state["started_today"].add(occ)
    if ed > current_date:
        return
    eq_line = str(e.get("machine") or "").strip()
    tid = _normalize_timeline_task_id(e)
    lead_op = str(e.get("op") or "").strip()
    sub_csv = str(e.get("sub") or "").strip()
    prev_dt = state["last_machining_dt"].get(occ)
    if prev_dt is None or end_dt > prev_dt:
        state["last_machining_dt"][occ] = end_dt
        state["last_tid"][occ] = tid
        state["last_eq"][occ] = eq_line
        state["last_machining_date"][occ] = ed
        state["last_lead_op"][occ] = lead_op
        state["last_machining_sub"][occ] = sub_csv
def _machine_handoff_state_from_timeline(
    timeline_events: list,
    current_date: date,
) -> dict:
    """
    タイムラインから」坄 machine_occupancy_key についで
    計画日 current_date 以降の **加工 (machining)** イベントの最終終了を復元れる。
    セットアップ系 event_kind は last_tid 等の復元に含まない。
    """
    best: dict[str, tuple[datetime, str, str, date, str, str]] = {}
    machining_today_occ: set[str] = set()
    for e in timeline_events:
        if not _is_machining_timeline_event(e):
            continue
        ed = e.get("date")
        if not isinstance(ed, date):
            continue
        occ = _machine_occupancy_key_from_timeline_event(e)
        if not occ:
            continue
        if ed == current_date:
            machining_today_occ.add(occ)
        if ed > current_date:
            continue
        end_dt = e.get("end_dt")
        if end_dt is None or not hasattr(end_dt, "replace"):
            continue
        eq_line = str(e.get("machine") or "").strip()
        tid = _normalize_timeline_task_id(e)
        lead_op = str(e.get("op") or "").strip()
        sub_csv = str(e.get("sub") or "").strip()
        prev = best.get(occ)
        if prev is None or end_dt > prev[0]:
            best[occ] = (end_dt, tid, eq_line, ed, lead_op, sub_csv)
    last_tid = {k: v[1] for k, v in best.items()}
    last_eq = {k: v[2] for k, v in best.items()}
    last_machining_dt = {k: v[0] for k, v in best.items()}
    last_machining_date = {k: v[3] for k, v in best.items()}
    last_lead_op = {k: v[4] for k, v in best.items()}
    last_machining_sub = {k: v[5] for k, v in best.items()}
    started_today = set(machining_today_occ)
    return {
        "last_tid": last_tid,
        "last_eq": last_eq,
        "last_machining_dt": last_machining_dt,
        "last_machining_date": last_machining_date,
        "last_lead_op": last_lead_op,
        "last_machining_sub": last_machining_sub,
        "machining_today_occ": machining_today_occ,
        "started_today": started_today,
    }
def _machine_handoff_state_from_timeline_cached(
    timeline_events: list,
    current_date: date,
) -> dict:
    """同一暦日でタイムラインが 1 件だけ増えたときは全件再スキャンを避ける。"""
    global _MH_HANDOFF_TIMELINE_CACHE_KEY, _MH_HANDOFF_TIMELINE_CACHE_STATE
    _t_mh0 = time_module.perf_counter()
    key = (len(timeline_events), current_date)
    if (
        _MH_HANDOFF_TIMELINE_CACHE_KEY == key
        and _MH_HANDOFF_TIMELINE_CACHE_STATE is not None
    ):
        _MH_HANDOFF_TIMELINE_CACHE_STATS["hit"] += 1
        _dispatch_loop_profile_add(
            "mh_handoff_cached_hit", time_module.perf_counter() - _t_mh0
        )
        return _snapshot_machine_handoff_state(_MH_HANDOFF_TIMELINE_CACHE_STATE)
    _MH_HANDOFF_TIMELINE_CACHE_STATS["miss"] += 1
    state = _machine_handoff_state_from_timeline(timeline_events, current_date)
    _MH_HANDOFF_TIMELINE_CACHE_KEY = key
    _MH_HANDOFF_TIMELINE_CACHE_STATE = state
    _dispatch_loop_profile_add(
        "mh_handoff_cached_miss", time_module.perf_counter() - _t_mh0
    )
    return _snapshot_machine_handoff_state(state)
def _stage2_eligible_wip_snapshot(task_queue: list) -> dict:
    """``_trial_order_flow_eligible_tasks`` 用 WIP 集計（1 パスあたり 1 回）。"""
    snap: dict = {
        "l11_global": None,
        "l11_by_bucket": {},
        "wip_slit_before_sec": None,
        "wip_connection_before_sec": None,
    }
    if isinstance(WIP_LIMIT_EC_BEFORE_INSP_ROLLS, int) and WIP_LIMIT_EC_BEFORE_INSP_ROLLS > 0:
        if _wip_ec_l11_aggregate_is_global():
            snap["l11_global"] = _wip_ec_before_insp_roll_count(task_queue)
    if (
        isinstance(WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS, int)
        and WIP_LIMIT_SLIT_BEFORE_SEC_ROLLS > 0
    ):
        slit_done_total = 0.0
        sec_done_total = 0.0
        _slit_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SLIT_PROCESS)
        _slit_mach = _normalize_equipment_match_key(SPECIAL_WIP_SLIT_MACHINE)
        _sec_proc = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
        _sec_mach = _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE)
        for _t in task_queue:
            proc = _normalize_process_name_for_rule_match(_t.get("machine"))
            mach = _normalize_equipment_match_key(_t.get("machine_name"))
            if not proc or not mach:
                continue
            init = float(_t.get("initial_remaining_units") or 0)
            rem = float(_t.get("remaining_units") or 0)
            done = max(0.0, init - rem)
            if done <= 1e-12:
                continue
            if proc == _slit_proc and mach == _slit_mach:
                slit_done_total += done
            elif proc == _sec_proc and mach == _sec_mach:
                sec_done_total += done
        snap["wip_slit_before_sec"] = max(0.0, slit_done_total - sec_done_total)
    if (
        isinstance(WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS, int)
        and WIP_LIMIT_CONNECTION_BEFORE_SEC_ROLLS > 0
    ):
        connection_done_total = 0.0
        sec_done_c = 0.0
        _conn_proc = _normalize_process_name_for_rule_match(
            SPECIAL_WIP_CONNECTION_PROCESS
        )
        _conn_mach = _normalize_equipment_match_key(SPECIAL_WIP_CONNECTION_MACHINE)
        _sec_proc_c = _normalize_process_name_for_rule_match(SPECIAL_WIP_SEC_PROCESS)
        _sec_mach_c = _normalize_equipment_match_key(SPECIAL_WIP_SEC_MACHINE)
        for _t in task_queue:
            proc = _normalize_process_name_for_rule_match(_t.get("machine"))
            mach = _normalize_equipment_match_key(_t.get("machine_name"))
            if not proc or not mach:
                continue
            init = float(_t.get("initial_remaining_units") or 0)
            rem = float(_t.get("remaining_units") or 0)
            done = max(0.0, init - rem)
            if done <= 1e-12:
                continue
            if proc == _conn_proc and mach == _conn_mach:
                connection_done_total += done
            elif proc == _sec_proc_c and mach == _sec_mach_c:
                sec_done_c += done
        snap["wip_connection_before_sec"] = max(
            0.0, connection_done_total - sec_done_c
        )
    return snap
def _stage2_rows_by_task_id(task_queue: list) -> dict[str, list[dict]]:
    """依頼NO（表示 task_id）ごとの行リスト（eligible 内の task_queue 全走査を避ける）。"""
    idx: dict[str, list[dict]] = defaultdict(list)
    for t in task_queue:
        tid = str(t.get("task_id") or "").strip()
        if tid:
            idx[tid].append(t)
    return idx
def _stage2_rows_by_rule_task_id(task_queue: list) -> dict[str, list[dict]]:
    """``_rule_task_id`` ごとの行リスト（§A 依存判定用）。"""
    idx: dict[str, list[dict]] = defaultdict(list)
    for t in task_queue:
        tid = _rule_task_id(t)
        if tid:
            idx[tid].append(t)
    return idx
def _task_no_machining_window_left_from_avail_floor_cached(
    t: dict,
    current_date: date,
    daily_status: dict | None,
    members: list | None,
    machine_avail_dt: dict | None,
    machine_day_start: datetime | None,
    *,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    window_left_cache: dict | None = None,
) -> bool:
    if window_left_cache is None:
        return _task_no_machining_window_left_from_avail_floor(
            t,
            current_date,
            daily_status,
            members,
            machine_avail_dt,
            machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
    _tm = t.get("machine")
    _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
    occ = (_machine_occupancy_key_resolve(t, _eqt) or "").strip()
    if not occ:
        return _task_no_machining_window_left_from_avail_floor(
            t,
            current_date,
            daily_status,
            members,
            machine_avail_dt,
            machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=abolish_all_scheduling_limits,
            dispatch_interval_mirror=dispatch_interval_mirror,
        )
    _avail_floor = machine_avail_dt.get(occ) if machine_avail_dt is not None else None
    key = (id(t), current_date, occ, _avail_floor)
    hit = window_left_cache.get(key)
    if hit is not None:
        return hit
    v = _task_no_machining_window_left_from_avail_floor(
        t,
        current_date,
        daily_status,
        members,
        machine_avail_dt,
        machine_day_start,
        machine_handoff=machine_handoff,
        skills_dict=skills_dict,
        abolish_all_scheduling_limits=abolish_all_scheduling_limits,
        dispatch_interval_mirror=dispatch_interval_mirror,
    )
    window_left_cache[key] = v
    return v
def _stage2_pending_by_machine_occ_index(
    task_queue: list, current_date: date
) -> dict[str, list[tuple[int, dict]]]:
    """start_date_req<=当日・残量ありを設備占有キー別・試行順昇順に索引（eligible 高速化）。"""
    idx: dict[str, list[tuple[int, dict]]] = defaultdict(list)
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        _sdr = t.get("start_date_req")
        if not isinstance(_sdr, date) or _sdr > current_date:
            continue
        _tm = t.get("machine")
        _eqt = str(t.get("equipment_line_key") or _tm or "").strip() or (_tm or "")
        t_occ = _machine_occupancy_key_resolve(t, _eqt)
        if not t_occ:
            continue
        try:
            o = _dispatch_trial_order_key(t)
        except (TypeError, ValueError):
            o = 10**9
        idx[t_occ].append((o, t))
    for occ in idx:
        idx[occ].sort(key=lambda x: x[0])
    return idx
def _combo_preset_team_size_bounds(
    preset_team: tuple,
    sheet_req_n: int | None,
    max_team_size_need: int,
) -> tuple[int, int] | None:
    """
    組み合わせ表プリセット1行の人数範囲 (lo, hi)。
    TEAM_ASSIGN_COMBO_SHEET_MAY_EXCEED_NEED=1（既定）のときは表の必須人数が need より大きくても採用可。
    =0 のときは need 解決後の max_team_size_need を上限とし、それを超えるプリセットは試行しない。
    """
    nmem = len(preset_team)
    if nmem < 1:
        return None
    cap = max(1, int(max_team_size_need or 1))
    if TEAM_ASSIGN_COMBO_SHEET_MAY_EXCEED_NEED:
        if sheet_req_n is not None and sheet_req_n >= 1:
            if nmem != sheet_req_n:
                return None
            lo = sheet_req_n
        else:
            lo = nmem
        hi = max(cap, nmem)
    else:
        if sheet_req_n is not None and sheet_req_n >= 1:
            if nmem != sheet_req_n:
                return None
            if sheet_req_n > cap:
                return None
            lo = sheet_req_n
        else:
            if nmem > cap:
                return None
            lo = nmem
        hi = min(cap, nmem)
    if not (lo <= nmem <= hi):
        return None
    return lo, hi
def _plan_sheet_required_op_optional(task: dict) -> int | None:
    """加工計画の必須人数列は正の整数ならしの値。無効なら None。"""
    ro = task.get("required_op")
    if ro is None or (isinstance(ro, float) and pd.isna(ro)):
        return None
    try:
        n = int(ro)
    except (TypeError, ValueError):
        return None
    return n if n >= 1 else None
def _append_legacy_dispatch_candidate_for_team(
    task: dict,
    team: tuple,
    avail_dt: dict,
    machine_avail_dt: dict,
    daily_status: dict,
    current_date: date,
    macro_run_date: date,
    macro_now_dt: datetime,
    skill_role_priority,
    eq_line: str,
    rq_base: int,
    extra_max: int,
    global_priority_override: dict,
    team_candidates: list,
    *,
    combo_sheet_row_id: int | None = None,
    combo_preset_team: tuple[str, ...] | None = None,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    limited_equipment_mirror: _LimitedEquipmentProtection | None = None,
    machine_handoff: dict | None = None,
    machine_day_floor: datetime | None = None,
    machine_floor_cached: datetime | None = None,
) -> bool:
    """レガシー日次配台ループ用: 坘一フォームは成立れれみ team_candidates に 1 件追加して True。"""
    _machine_occ_key = _machine_occupancy_key_resolve(task, eq_line)
    _gpo = global_priority_override or {}
    _all_limits_abolished = bool(_gpo.get("abolish_all_scheduling_limits"))
    _equipment_occupancy_abolished = (
        _legacy_dispatch_scheduling_limits_abolished(_gpo, task)
    )
    _floor_default = datetime.combine(current_date, DEFAULT_START_TIME)
    _mdf = machine_day_floor if machine_day_floor is not None else _floor_default
    _mh_legacy = machine_handoff or {
        "last_tid": {},
        "last_eq": {},
        "started_today": set(),
        "machining_today_occ": set(),
        "last_machining_dt": {},
        "last_machining_date": {},
        "last_lead_op": {},
        "last_machining_sub": {},
    }
    op_list = [m for m in team if skill_role_priority(m)[0] == "OP"]
    if not op_list:
        return False
    team_start = max(avail_dt[m] for m in team)
    _prev_mach_raw = machine_avail_dt.get(_machine_occ_key, _mdf)
    if not _equipment_occupancy_abolished:
        if machine_floor_cached is not None:
            machine_free_dt = machine_floor_cached
        else:
            machine_free_dt = _machine_effective_floor_for_assign(
                _machine_occ_key,
                str(task.get("task_id") or "").strip(),
                eq_line,
                str(task.get("machine_name") or "").strip(),
                machine_avail_dt,
                _mh_legacy,
                _mdf,
                False,
                current_date=current_date,
            )
        if team_start < machine_free_dt:
            team_start = machine_free_dt
    if not _all_limits_abolished:
        if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
            min_start_dt = datetime.combine(
                current_date, task["same_day_raw_start_limit"]
            )
            if team_start < min_start_dt:
                team_start = min_start_dt
        if current_date == task["start_date_req"] and task.get("earliest_start_time"):
            min_user_t = datetime.combine(
                current_date, task["earliest_start_time"]
            )
            if team_start < min_user_t:
                team_start = min_user_t
        if current_date == macro_run_date and team_start < macro_now_dt:
            team_start = macro_now_dt
    team_end_limit = min(daily_status[m]["end_dt"] for m in team)
    team_end_limit = _interactive_trial_relax_team_end_limit_to_eod(
        team_end_limit, current_date
    )
    if team_start >= team_end_limit:
        return False
    team_breaks = []
    for m in team:
        team_breaks.extend(daily_status[m]["breaks_dt"])
    team_breaks = merge_time_intervals(team_breaks)

    avg_eff = sum(daily_status[m]["efficiency"] for m in team) / len(team)
    if avg_eff <= 0:
        avg_eff = 0.01
    t_eff = parse_float_safe(task.get("task_eff_factor"), 1.0)
    if t_eff <= 0:
        t_eff = 1.0
    eff_time_per_unit = (
        task["base_time_per_unit"]
        / avg_eff
        / t_eff
        * _surplus_team_time_factor(rq_base, len(team), extra_max)
    )
    _defer_min_contig = max(1, int(math.ceil(float(eff_time_per_unit))))
    _eod_cont_exempt = _eod_same_request_continuation_exempt(
        _machine_occ_key, task, _mh_legacy
    )

    def _refloor_legacy_roll(ts: datetime) -> datetime:
        ts = max(ts, max(avail_dt[m] for m in team))
        if not _equipment_occupancy_abolished:
            if machine_floor_cached is not None:
                mf = machine_floor_cached
            else:
                mf = _machine_effective_floor_for_assign(
                    _machine_occ_key,
                    str(task.get("task_id") or "").strip(),
                    eq_line,
                    str(task.get("machine_name") or "").strip(),
                    machine_avail_dt,
                    _mh_legacy,
                    _mdf,
                    False,
                    current_date=current_date,
                )
            if ts < mf:
                ts = mf
        if not _all_limits_abolished:
            if task.get("same_day_raw_start_limit") and current_date == task["start_date_req"]:
                min_start_dt = datetime.combine(
                    current_date, task["same_day_raw_start_limit"]
                )
                if ts < min_start_dt:
                    ts = min_start_dt
            if current_date == task["start_date_req"] and task.get("earliest_start_time"):
                min_user_t = datetime.combine(
                    current_date, task["earliest_start_time"]
                )
                if ts < min_user_t:
                    ts = min_user_t
            if current_date == macro_run_date and ts < macro_now_dt:
                ts = macro_now_dt
        return ts

    team_start_adj = _defer_team_start_past_prebreak_and_end_of_day(
        task,
        team,
        team_start,
        team_end_limit,
        team_breaks,
        _refloor_legacy_roll,
        min_contiguous_work_mins=_defer_min_contig,
        eod_same_request_continuation_exempt=_eod_cont_exempt,
    )
    if team_start_adj is None:
        return False
    team_start = team_start_adj
    _roll_prep_extra_l: list[dict] = []
    if not _equipment_occupancy_abolished:
        team_start, _roll_prep_extra_l = _roll_prep_segments_for_assign(
            team_start=team_start,
            team_breaks=team_breaks,
            machine_handoff=_mh_legacy,
            machine_occ_key=_machine_occ_key,
            current_date=current_date,
            task_id=str(task.get("task_id") or "").strip(),
            machine_proc=str(task.get("machine") or "").strip(),
            machine_name=str(task.get("machine_name") or "").strip(),
            eq_line=eq_line,
            abolish_limits=False,
            prev_machining_end=_resolve_prev_machining_end_for_roll_prep(
                _mh_legacy,
                _machine_occ_key,
                (_mh_legacy.get("last_machining_dt") or {}).get(_machine_occ_key),
                machine_avail_dt,
                _mdf,
                _prev_mach_raw,
            ),
            prev_eq_line=str(
                (_mh_legacy.get("last_eq") or {}).get(_machine_occ_key, "") or ""
            ).strip(),
            machine_avail_dt=machine_avail_dt,
            machine_day_floor=_mdf,
        )
        team_start = _refloor_legacy_roll(team_start)
    if team_start >= team_end_limit:
        return False
    if dispatch_interval_mirror is not None and _roll_prep_extra_l:
        for _pseg in _roll_prep_extra_l:
            _pst = _pseg.get("start_dt")
            _ped = _pseg.get("end_dt")
            _pok = str(_pseg.get("machine_occupancy_key") or _machine_occ_key).strip()
            if (
                isinstance(_pst, datetime)
                and isinstance(_ped, datetime)
                and _pok
                and dispatch_interval_mirror.would_block_equipment(_pok, _pst, _ped)
            ):
                return False

    protected_capacity = _candidate_capacity_after_equipment_protection(
        limited_equipment_mirror,
        _machine_occ_key,
        team_start,
        eff_time_per_unit,
        float(task["remaining_units"]),
        team_breaks,
        team_end_limit,
    )
    if protected_capacity is None:
        return False
    team_start, capacity = protected_capacity
    if team_start >= team_end_limit:
        return False
    units_today = capacity["units_today"]
    work_mins_needed = capacity["work_mins_needed"]
    if _eod_reject_capacity_units_below_threshold(
        units_today,
        team_start,
        team_end_limit,
        eod_same_request_continuation_exempt=_eod_cont_exempt,
        remaining_units_ceil=math.ceil(float(task.get("remaining_units") or 0)),
    ):
        return False
    if (
        _contiguous_work_minutes_until_next_break_or_limit(
            team_start, team_breaks, team_end_limit
        )
        < work_mins_needed
    ):
        return False
    actual_end_dt, _, _ = calculate_end_time(
        team_start, work_mins_needed, team_breaks, team_end_limit
    )
    if dispatch_interval_mirror is not None and dispatch_interval_mirror.would_block_roll(
        _machine_occ_key, team, team_start, actual_end_dt
    ):
        return False
    if _legacy_dispatch_limited_equipment_interval_blocked(
        task,
        _gpo,
        machine_avail_dt,
        limited_equipment_mirror,
        _machine_occ_key,
        team_start,
        actual_end_dt,
    ):
        return False
    team_prio_sum = sum(skill_role_priority(m)[1] for m in team)
    team_candidates.append(
        {
            "team": team,
            "team_start": team_start,
            "actual_end_dt": actual_end_dt,
            "units_today": units_today,
            "team_breaks": team_breaks,
            "avg_eff": avg_eff,
            "prio_sum": team_prio_sum,
            "op_list": op_list,
            "eff_time_per_unit": eff_time_per_unit,
            "combo_sheet_row_id": combo_sheet_row_id,
            "combo_preset_team": combo_preset_team,
            "roll_prep_segments": _roll_prep_extra_l,
        }
    )
    return True
def _tasks_in_min_pending_dispatch_pool(
    task_queue: list,
    current_date: date,
    *,
    daily_status: dict | None = None,
    members: list | None = None,
    machine_avail_dt: dict | None = None,
    machine_day_start: datetime | None = None,
    machine_handoff: dict | None = None,
    skills_dict: dict | None = None,
    abolish_all_scheduling_limits: bool = False,
    dispatch_interval_mirror: DispatchIntervalMirror | None = None,
    interactive_trial_pair_dates: dict | None = None,
) -> list:
    """`_min_pending_dispatch_trial_order_for_date` と同一の安価フィルタを通靎したタスクのリスト。"""
    out: list = []
    for t in task_queue:
        if float(t.get("remaining_units") or 0) <= 1e-12:
            continue
        if (
            _interactive_dispatch_trial_env_active()
            and interactive_trial_pair_dates is not None
        ):
            tid_n = _interactive_norm_cell(str(t.get("task_id") or ""))
            mach_n = _interactive_norm_cell(str(t.get("machine_name") or ""))
            _pd = interactive_trial_pair_dates.get((tid_n, mach_n))
            if _pd is not None and current_date not in _pd:
                continue
        sdr = t.get("start_date_req")
        if not isinstance(sdr, date) or sdr > current_date:
            continue
        if _task_not_yet_schedulable_due_to_dependency_or_b2_room(t, task_queue):
            continue
        if _task_fully_machine_calendar_blocked_on_date(
            t, current_date, daily_status, members
        ):
            continue
        _abolish_for_task = _scheduling_limits_abolished_for_task(
            {"abolish_all_scheduling_limits": abolish_all_scheduling_limits},
            t,
        )
        if _task_no_machining_window_left_from_avail_floor(
            t,
            current_date,
            daily_status,
            members,
            machine_avail_dt,
            machine_day_start,
            machine_handoff=machine_handoff,
            skills_dict=skills_dict,
            abolish_all_scheduling_limits=_abolish_for_task,
            dispatch_interval_mirror=dispatch_interval_mirror,
        ):
            continue
        if t.get("_dispatch_block_no_op_on_working_days"):
            continue
        out.append(t)
    return out
def _effective_min_dispatch_trial_order_from_pool(
    pool: list,
    current_date: date,
    daily_status: dict,
    assign_probe_ctx: dict,
) -> int | None:
    """
    pool を昇順 dto で見で」**しの dto に属れる行のごう 1 件でも** 1 ロール割当プローブは通れみ
    しの dto を「実効の最尝試行順」とれる。
    先頭 dto 層は全滅（機械は空いでいるは人で穝ゝない等）のとき」次の dto に進みグローバル坜止を防し。
    プローブ無しのときは pool の最尝 dto を返す。
    """
    if not pool:
        return None
    dtos = sorted(
        {
            _dispatch_trial_order_key(t)
            for t in pool
        }
    )
    if not assign_probe_ctx:
        return min(dtos)
    for d in dtos:
        at_d = [
            t
            for t in pool
            if _dispatch_trial_order_key(t) == d
        ]
        if any(
            not _trial_order_assign_probe_fails(
                t, current_date, daily_status, assign_probe_ctx
            )
            for t in at_d
        ):
            return d
    return None
