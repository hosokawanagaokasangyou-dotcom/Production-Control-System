# -*- coding: utf-8 -*-
"""
ブック xlsx の結合セルを「各マスに同じ値を持つ格子」に展開してから表 JSON 化する（論理・ビュー用）。

ビューア用の通常 ``production_plan_multi_day*.json``（xlsx 直読みミラー）では、
横方向の結合（ガント帯）の右側が欠損になりやすい。本モジュールは openpyxl で結合を解き
値を埋めたうえで :func:`workbook_payload_from_final_xlsx_file` へ渡し、
10 分スロット列にラベルが乗った「密な」表を JSON にできる。
"""

from __future__ import annotations

import logging
import os
import tempfile
import unicodedata

logger = logging.getLogger(__name__)


def expand_merged_cells_fill_worksheet(ws) -> None:
    """
    シート上の結合範囲を解き、左上セルの値（または式の結果）を範囲内の全セルに複製する。
    既存の非空セルは上書きする（結合内は通常空のため問題にならない想定）。
    """
    from openpyxl.utils import range_boundaries

    for mrange in list(ws.merged_cells.ranges):
        coord = str(mrange)
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        v = ws.cell(row=min_row, column=min_col).value
        try:
            ws.unmerge_cells(coord)
        except Exception as e:
            logger.debug("logical_workbook_view: unmerge 失敗（無視）%s: %s", coord, e)
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = v


def expand_merged_cells_fill_workbook(wb) -> None:
    for ws in wb.worksheets:
        expand_merged_cells_fill_worksheet(ws)


def build_logical_view_workbook_payload(
    xlsx_path: str,
    *,
    source_xlsx_basename: str | None = None,
    format_version: int | None = None,
) -> dict:
    """
    保存済み xlsx を開き、全シートの結合を値展開した一時 xlsx 経由で
    :func:`planning_core.workbook_payload.workbook_payload_from_final_xlsx_file`
    と同形のペイロードを返す。

    metadata に schema ``plan_logical_view_v1`` を付与する。
    """
    from openpyxl import load_workbook

    from planning_core.workbook_payload import (
        WORKBOOK_JSON_FORMAT_VERSION,
        workbook_payload_from_final_xlsx_file,
    )

    if not xlsx_path or not os.path.isfile(xlsx_path):
        raise FileNotFoundError(xlsx_path)

    fv = (
        int(format_version)
        if format_version is not None
        else WORKBOOK_JSON_FORMAT_VERSION
    )
    # data_only=True … Excel が保存した計算結果（表示値）を読み、結合展開の複製先にも値が入る。
    # data_only=False だと式セルが式のまま残り、read_excel 経由の JSON で時刻マスが欠損しやすい。
    wb = load_workbook(xlsx_path, data_only=True)
    expand_merged_cells_fill_workbook(wb)

    fd, tmp_xlsx = tempfile.mkstemp(suffix=".xlsx", prefix="_pm_logical_view_src_")
    os.close(fd)
    try:
        wb.save(tmp_xlsx)
        payload = workbook_payload_from_final_xlsx_file(
            tmp_xlsx,
            source_xlsx_basename=source_xlsx_basename or os.path.basename(xlsx_path),
            format_version=fv,
            metadata_extra={
                "schema": "plan_logical_view_v1",
                "logical_view": True,
                "note": "結合セルを値で展開後に read_excel した密な表（グラフィック／論理ビュー向け）",
            },
        )
        return payload
    finally:
        try:
            os.remove(tmp_xlsx)
        except OSError:
            pass


def logical_view_json_path(plan_xlsx: str) -> str:
    """``plan.xlsx`` と並べて ``plan_logical_view.json``（正規化パス）。"""
    base, _ = os.path.splitext(os.path.abspath(os.path.normpath(plan_xlsx)))
    try:
        base = unicodedata.normalize("NFC", base)
    except Exception:
        pass
    return base + "_logical_view.json"
