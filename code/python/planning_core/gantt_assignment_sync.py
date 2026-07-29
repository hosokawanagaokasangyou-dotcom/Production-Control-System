# -*- coding: utf-8 -*-
"""
設備ガント担当割当の検証・契約 JSON 反映・成果物同期のコア。

JavaFX 編集モデルが生成した ``timeline_events`` index ごとの op/sub 更新を受け取り、
二重割当拒否・勤怠警告・契約ハッシュ競合検出・確認トークン付き警告フローを行う。
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterable

from planning_core.gantt_render_contract import (
    decode_value,
    encode_value,
    load_gantt_contract_json,
    unpack_gantt_contract,
    write_gantt_contract_json,
)

FORMAT_VERSION = 1
CONFIRM_TOKEN_SALT = "pm_ai_gantt_assignment_v1"


@dataclass(frozen=True)
class OpSubUpdate:
    event_index: int
    op: str
    sub: str


@dataclass
class SyncIssue:
    code: str
    message: str
    event_indices: list[int] = field(default_factory=list)
    person: str = ""

    def to_dict(self) -> dict[str, Any]:
        out: dict[str, Any] = {"code": self.code, "message": self.message}
        if self.event_indices:
            out["event_indices"] = list(self.event_indices)
        if self.person:
            out["person"] = self.person
        return out


@dataclass
class AssignmentSyncResult:
    ok: bool
    status: str
    timeline_hash: str = ""
    confirm_token: str = ""
    errors: list[SyncIssue] = field(default_factory=list)
    warnings: list[SyncIssue] = field(default_factory=list)
    contract_path: str = ""
    plan_xlsx_path: str = ""
    backup_paths: list[str] = field(default_factory=list)
    detail: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "format_version": FORMAT_VERSION,
            "ok": self.ok,
            "status": self.status,
            "timeline_hash": self.timeline_hash,
            "confirm_token": self.confirm_token,
            "errors": [e.to_dict() for e in self.errors],
            "warnings": [w.to_dict() for w in self.warnings],
            "contract_path": self.contract_path,
            "plan_xlsx_path": self.plan_xlsx_path,
            "backup_paths": list(self.backup_paths),
            "detail": self.detail,
        }


def _norm_name(s: str | None) -> str:
    if not s:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", "", t)
    return t


def _person_names_from_op_sub(op: str | None, sub: str | None) -> list[str]:
    names: list[str] = []
    op_s = (op or "").strip()
    if op_s:
        names.append(op_s)
    sub_s = (sub or "").strip()
    if sub_s:
        for part in sub_s.split(","):
            p = part.strip()
            if p:
                names.append(p)
    return names


def _event_dt(ev: dict[str, Any], key: str) -> datetime | None:
    raw = ev.get(key)
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date) and not isinstance(raw, datetime):
        return datetime.combine(raw, datetime.min.time())
    return None


def _event_date(ev: dict[str, Any]) -> date | None:
    raw = ev.get("date")
    if isinstance(raw, date) and not isinstance(raw, datetime):
        return raw
    if isinstance(raw, datetime):
        return raw.date()
    dt = _event_dt(ev, "start_dt")
    return dt.date() if dt else None


def timeline_events_from_contract(contract: dict[str, Any]) -> list[dict[str, Any]]:
    kw = unpack_gantt_contract(contract)
    events = kw.get("timeline_events") or []
    return list(events)


def timeline_assignment_hash(events: Iterable[dict[str, Any]]) -> str:
    """楽観ロック用: 各イベントの op/sub と主要時刻・機械を正規化してハッシュ。"""
    payload: list[dict[str, Any]] = []
    for i, ev in enumerate(events):
        payload.append(
            {
                "i": i,
                "op": str(ev.get("op") or ""),
                "sub": str(ev.get("sub") or ""),
                "machine": str(ev.get("machine") or ""),
                "task_id": str(ev.get("task_id") or ""),
                "event_kind": str(ev.get("event_kind") or ""),
                "start": (
                    _event_dt(ev, "start_dt").isoformat()
                    if _event_dt(ev, "start_dt")
                    else ""
                ),
                "end": (
                    _event_dt(ev, "end_dt").isoformat() if _event_dt(ev, "end_dt") else ""
                ),
            }
        )
    blob = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def parse_updates_map(raw: dict[str, Any] | None) -> list[OpSubUpdate]:
    if not raw:
        return []
    updates: list[OpSubUpdate] = []
    items = raw.get("updates") if isinstance(raw.get("updates"), dict) else raw
    if not isinstance(items, dict):
        return []
    for k, v in items.items():
        try:
            idx = int(k)
        except (TypeError, ValueError):
            continue
        if not isinstance(v, dict):
            continue
        updates.append(
            OpSubUpdate(
                event_index=idx,
                op=str(v.get("op") or ""),
                sub=str(v.get("sub") or ""),
            )
        )
    updates.sort(key=lambda u: u.event_index)
    return updates


def apply_op_sub_updates(
    contract: dict[str, Any], updates: list[OpSubUpdate]
) -> dict[str, Any]:
    """契約 dict のコピーへ op/sub 更新を適用する。"""
    out = json.loads(json.dumps(contract, ensure_ascii=False))
    inner = out.setdefault("kwargs_packed", {})
    encoded_events = list(inner.get("timeline_events") or [])
    decoded = decode_value(encoded_events)
    if not isinstance(decoded, list):
        decoded = []
    for upd in updates:
        if upd.event_index < 0 or upd.event_index >= len(decoded):
            raise IndexError(f"event index out of range: {upd.event_index}")
        ev = decoded[upd.event_index]
        if not isinstance(ev, dict):
            raise ValueError(f"timeline_events[{upd.event_index}] is not a dict")
        ev["op"] = upd.op
        ev["sub"] = upd.sub
    inner["timeline_events"] = encode_value(decoded)
    return out


def _intervals_overlap(a0: datetime, a1: datetime, b0: datetime, b1: datetime) -> bool:
    return a0 < b1 and b0 < a1


def find_double_assignments(events: list[dict[str, Any]]) -> list[SyncIssue]:
    """同一人物が同一日で時間帯が重なる割当を検出（保存不可）。"""
    by_person_day: dict[tuple[date, str], list[tuple[int, datetime, datetime, str]]] = {}
    for i, ev in enumerate(events):
        d = _event_date(ev)
        start = _event_dt(ev, "start_dt")
        end = _event_dt(ev, "end_dt")
        if d is None or start is None or end is None:
            continue
        machine = str(ev.get("machine") or "")
        for name in _person_names_from_op_sub(ev.get("op"), ev.get("sub")):
            key = (d, _norm_name(name))
            if not key[1]:
                continue
            by_person_day.setdefault(key, []).append((i, start, end, machine))

    issues: list[SyncIssue] = []
    for (d, norm_person), items in sorted(by_person_day.items()):
        items_sorted = sorted(items, key=lambda x: (x[1], x[2], x[0]))
        for j in range(len(items_sorted)):
            i0, s0, e0, m0 = items_sorted[j]
            for k in range(j + 1, len(items_sorted)):
                i1, s1, e1, m1 = items_sorted[k]
                if not _intervals_overlap(s0, e0, s1, e1):
                    continue
                display = norm_person
                issues.append(
                    SyncIssue(
                        code="double_assignment",
                        message=(
                            f"{d.isoformat()} に {display} が重複割当です"
                            f"（イベント {i0}/{i1}、{m0} と {m1}）。"
                        ),
                        event_indices=sorted({i0, i1}),
                        person=display,
                    )
                )
    return issues


def _resolve_attendance_day(
    attendance_data: dict[Any, Any], day: date
) -> dict[Any, Any] | None:
    if not attendance_data:
        return None
    direct = attendance_data.get(day)
    if isinstance(direct, dict):
        return direct
    day_s = day.isoformat()
    for key, value in attendance_data.items():
        if isinstance(key, date) and not isinstance(key, datetime):
            if key == day and isinstance(value, dict):
                return value
        if str(key) == day_s and isinstance(value, dict):
            return value
    return None


def _attendance_member_entry(
    attendance_data: dict[Any, Any], day: date, person: str
) -> dict[str, Any] | None:
    day_map = _resolve_attendance_day(attendance_data, day)
    if not isinstance(day_map, dict):
        return None
    target = _norm_name(person)
    for member, data in day_map.items():
        if _norm_name(str(member)) == target:
            return data if isinstance(data, dict) else None
    return None


def find_attendance_warnings(
    events: list[dict[str, Any]], attendance_data: dict[Any, Any]
) -> list[SyncIssue]:
    """契約内勤怠データで欠勤・非稼働の担当者を警告。"""
    if not attendance_data:
        return []
    seen: set[tuple[date, str]] = set()
    warnings: list[SyncIssue] = []
    for i, ev in enumerate(events):
        d = _event_date(ev)
        if d is None:
            continue
        for name in _person_names_from_op_sub(ev.get("op"), ev.get("sub")):
            key = (d, _norm_name(name))
            if not key[1] or key in seen:
                continue
            seen.add(key)
            entry = _attendance_member_entry(attendance_data, d, name)
            if entry is None:
                continue
            if not bool(entry.get("is_working", True)):
                warnings.append(
                    SyncIssue(
                        code="absent",
                        message=f"{d.isoformat()} の {name} は勤怠上非稼働です。",
                        event_indices=[i],
                        person=name,
                    )
                )
    return warnings


def find_index_errors(
    event_count: int, updates: list[OpSubUpdate]
) -> list[SyncIssue]:
    issues: list[SyncIssue] = []
    for upd in updates:
        if upd.event_index < 0 or upd.event_index >= event_count:
            issues.append(
                SyncIssue(
                    code="index_out_of_range",
                    message=f"timeline_events[{upd.event_index}] は存在しません。",
                    event_indices=[upd.event_index],
                )
            )
    return issues


def find_empty_assignment_errors(
    events: list[dict[str, Any]], updates: list[OpSubUpdate]
) -> list[SyncIssue]:
    """更新後に op/sub がともに空になるイベントを拒否。"""
    issues: list[SyncIssue] = []
    for upd in updates:
        if upd.event_index < 0 or upd.event_index >= len(events):
            continue
        if upd.op.strip() or upd.sub.strip():
            continue
        issues.append(
            SyncIssue(
                code="empty_assignment",
                message=f"イベント {upd.event_index} を0名にすることはできません。",
                event_indices=[upd.event_index],
            )
        )
    return issues


def make_confirm_token(timeline_hash: str, warnings: list[SyncIssue]) -> str:
    warn_blob = json.dumps(
        [w.to_dict() for w in warnings],
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    raw = f"{timeline_hash}|{warn_blob}|{CONFIRM_TOKEN_SALT}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def validate_confirm_token(
    timeline_hash: str, warnings: list[SyncIssue], token: str | None
) -> bool:
    if not token:
        return False
    expected = make_confirm_token(timeline_hash, warnings)
    return token.strip() == expected


def validate_assignment_changes(
    contract: dict[str, Any],
    updates: list[OpSubUpdate],
    *,
    expected_timeline_hash: str | None = None,
    confirm_token: str | None = None,
    force_warnings: bool = False,
) -> AssignmentSyncResult:
    events = timeline_events_from_contract(contract)
    current_hash = timeline_assignment_hash(events)
    result = AssignmentSyncResult(ok=False, status="error", timeline_hash=current_hash)

    if expected_timeline_hash and expected_timeline_hash != current_hash:
        result.errors.append(
            SyncIssue(
                code="hash_mismatch",
                message="契約 JSON が他の操作で更新されています。再読み込みしてください。",
            )
        )
        return result

    result.errors.extend(find_index_errors(len(events), updates))
    if result.errors:
        return result

    projected = list(events)
    for upd in updates:
        ev = dict(projected[upd.event_index])
        ev["op"] = upd.op
        ev["sub"] = upd.sub
        projected[upd.event_index] = ev

    result.errors.extend(find_empty_assignment_errors(projected, updates))
    result.errors.extend(find_double_assignments(projected))
    if result.errors:
        return result

    kw = unpack_gantt_contract(contract)
    attendance_data = kw.get("attendance_data") or {}
    result.warnings.extend(find_attendance_warnings(projected, attendance_data))

    if result.warnings:
        token = make_confirm_token(current_hash, result.warnings)
        result.confirm_token = token
        if not force_warnings:
            result.status = "warnings"
            result.ok = False
            result.detail = "警告があります。confirm_token を付けて再実行してください。"
            return result
        if not validate_confirm_token(current_hash, result.warnings, confirm_token):
            result.errors.append(
                SyncIssue(
                    code="invalid_confirm_token",
                    message="確認トークンが無効です。警告内容を再取得してください。",
                )
            )
            return result

    result.ok = True
    result.status = "validated"
    return result


def _backup_file(path: str) -> str | None:
    if not path or not os.path.isfile(path):
        return None
    parent = os.path.dirname(os.path.abspath(path))
    base = os.path.basename(path)
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = os.path.join(parent, f".pm_ai_backup_{stamp}_{base}")
    shutil.copy2(path, backup)
    return backup


def refresh_equipment_gantt_sheet_in_plan_workbook(
    plan_xlsx: str, contract: dict[str, Any]
) -> None:
    """計画ブック内の設備ガントシートのみ契約から再描画する。"""
    import pandas as pd
    from openpyxl import load_workbook

    from planning_core.core.columns import RESULT_SHEET_GANTT_NAME
    from planning_core.gantt_render_contract import render_gantt_sheet_from_contract

    kw = unpack_gantt_contract(contract)
    sheet_nm = kw.get("sheet_name_override") or RESULT_SHEET_GANTT_NAME
    if not os.path.isfile(plan_xlsx):
        raise FileNotFoundError(plan_xlsx)

    wb = load_workbook(plan_xlsx)
    if sheet_nm in wb.sheetnames:
        wb.remove(wb[sheet_nm])
        wb.save(plan_xlsx)

    with pd.ExcelWriter(
        plan_xlsx, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        writer.book = load_workbook(plan_xlsx)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        render_gantt_sheet_from_contract(writer, contract)
        writer.book.save(plan_xlsx)


def sync_assignment_to_artifacts(
    *,
    contract_path: str,
    updates: list[OpSubUpdate],
    plan_xlsx_path: str | None = None,
    expected_timeline_hash: str | None = None,
    confirm_token: str | None = None,
    force_warnings: bool = False,
    dry_run: bool = False,
) -> AssignmentSyncResult:
    contract = load_gantt_contract_json(contract_path)
    validation = validate_assignment_changes(
        contract,
        updates,
        expected_timeline_hash=expected_timeline_hash,
        confirm_token=confirm_token,
        force_warnings=force_warnings,
    )
    validation.contract_path = contract_path
    validation.plan_xlsx_path = plan_xlsx_path or ""

    if not validation.ok:
        if validation.status == "warnings":
            return validation
        validation.status = "error"
        return validation

    updated = apply_op_sub_updates(contract, updates)

    if dry_run:
        validation.status = "dry_run"
        validation.detail = "検証のみ成功（書込みなし）。"
        return validation

    backups: list[str] = []
    try:
        if plan_xlsx_path and os.path.isfile(plan_xlsx_path):
            bp = _backup_file(plan_xlsx_path)
            if bp:
                backups.append(bp)
        cb = _backup_file(contract_path)
        if cb:
            backups.append(cb)

        if plan_xlsx_path and os.path.isfile(plan_xlsx_path):
            refresh_equipment_gantt_sheet_in_plan_workbook(plan_xlsx_path, updated)
            out_path, _ = write_gantt_contract_json(plan_xlsx_path, "equipment", updated)
            if not out_path:
                raise OSError("設備ガント契約 JSON の書込みに失敗しました。")
        else:
            parent = os.path.dirname(os.path.abspath(contract_path))
            tmp = os.path.join(parent, f".pm_ai_contract_{os.getpid()}.tmp")
            with open(tmp, "w", encoding="utf-8", newline="\n") as f:
                json.dump(updated, f, ensure_ascii=False, indent=2)
                f.write("\n")
            os.replace(tmp, contract_path)

        validation.ok = True
        validation.status = "applied"
        validation.backup_paths = backups
        validation.timeline_hash = timeline_assignment_hash(
            timeline_events_from_contract(updated)
        )
        validation.detail = "担当割当を成果物へ反映しました。"
        return validation
    except Exception as e:
        validation.ok = False
        validation.status = "error"
        validation.backup_paths = backups
        validation.detail = str(e)
        validation.errors.append(
            SyncIssue(code="write_failed", message=f"書込みに失敗しました: {e}")
        )
        return validation
