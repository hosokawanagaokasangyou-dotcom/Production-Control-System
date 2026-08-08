# -*- coding: utf-8 -*-

from planning_core.core.attendance_excel_style import (
    menu_label_for_sheet,
    sort_menu_sheet_names,
    write_calendar_workbook_menu_sheet,
)
from planning_core.core.attendance_paths import (
    APP_MASTER_COMPANY_SHEET,
    APP_MASTER_MACHINE_CALENDAR_DATE_SHEET,
    APP_MASTER_MACHINE_CALENDAR_SHEET,
    APP_MASTER_MEMBER_SHEET_PREFIX,
    APP_MASTER_MENU_SHEET,
)


def test_sort_menu_sheet_names_orders_company_machine_members():
    member = f"{APP_MASTER_MEMBER_SHEET_PREFIX}2026年12月"
    member2 = f"{APP_MASTER_MEMBER_SHEET_PREFIX}2026年5月"
    names = sort_menu_sheet_names(
        [
            member,
            APP_MASTER_MACHINE_CALENDAR_DATE_SHEET,
            APP_MASTER_MACHINE_CALENDAR_SHEET,
            member2,
            APP_MASTER_COMPANY_SHEET,
        ]
    )
    assert names[0] == APP_MASTER_COMPANY_SHEET
    assert names[1] == APP_MASTER_MACHINE_CALENDAR_SHEET
    assert names[2] == APP_MASTER_MACHINE_CALENDAR_DATE_SHEET
    assert names[3:] == [member2, member]


def test_menu_label_for_sheet():
    assert menu_label_for_sheet(APP_MASTER_COMPANY_SHEET) == "会社カレンダー"
    assert menu_label_for_sheet(APP_MASTER_MACHINE_CALENDAR_SHEET) == "機械カレンダー"
    assert menu_label_for_sheet(f"{APP_MASTER_MEMBER_SHEET_PREFIX}2026年5月") == (
        "メンバー勤怠 2026年5月"
    )


def test_write_calendar_workbook_menu_sheet_hyperlinks():
    from openpyxl import Workbook

    wb = Workbook()
    ws_menu = wb.active
    ws_menu.title = APP_MASTER_MENU_SHEET
    wb.create_sheet(APP_MASTER_COMPANY_SHEET)
    wb.create_sheet(APP_MASTER_MACHINE_CALENDAR_SHEET)

    write_calendar_workbook_menu_sheet(
        ws_menu,
        [APP_MASTER_COMPANY_SHEET, APP_MASTER_MACHINE_CALENDAR_SHEET],
        export_at="2026-08-08T12:00:00",
    )

    link = ws_menu.cell(6, 1)
    assert link.value == "会社カレンダー"
    assert link.hyperlink is not None
    assert APP_MASTER_COMPANY_SHEET in str(link.hyperlink.target)


def test_export_calendar_xlsx_company_sheet_has_menu_back_link(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    from planning_core.core.attendance_paths import ENV_ATTENDANCE_CALENDAR_XLSX
    from planning_core.core.attendance_store import empty_store, export_attendance_to_calendar_workbook

    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX, str(calendar))

    store = empty_store(2026)
    export_attendance_to_calendar_workbook(store, calendar)

    wb = load_workbook(calendar)
    ws = wb[APP_MASTER_COMPANY_SHEET]
    back = ws.cell(1, 1)
    assert back.value == "メニューに戻る"
    assert back.hyperlink is not None
    assert APP_MASTER_MENU_SHEET in str(back.hyperlink.target)
