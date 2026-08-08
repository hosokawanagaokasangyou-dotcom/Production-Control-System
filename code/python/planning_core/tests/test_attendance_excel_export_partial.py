# -*- coding: utf-8 -*-

from openpyxl import load_workbook

from planning_core.core.attendance_paths import (
    APP_MASTER_COMPANY_SHEET,
    APP_MASTER_MEMBER_SHEET_PREFIX,
)
from planning_core.core.attendance_store import empty_store, export_attendance_to_calendar_workbook


def test_partial_member_export_updates_single_month_sheet(tmp_path):
    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    store = empty_store(2026)
    export_attendance_to_calendar_workbook(store, calendar)

    wb_before = load_workbook(calendar)
    company_before = wb_before[APP_MASTER_COMPANY_SHEET].cell(2, 1).value
    member_sheet = f"{APP_MASTER_MEMBER_SHEET_PREFIX}2026年4月"
    member_before = wb_before[member_sheet].cell(2, 1).value

    export_attendance_to_calendar_workbook(
        store,
        calendar,
        months=[(2026, 4)],
        refresh_company=False,
        refresh_machine=False,
        skip_snapshot=True,
    )

    wb_after = load_workbook(calendar)
    assert wb_after[APP_MASTER_COMPANY_SHEET].cell(2, 1).value == company_before
    assert wb_after[member_sheet].cell(2, 1).value == member_before
