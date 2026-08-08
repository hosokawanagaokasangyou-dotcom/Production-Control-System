# -*- coding: utf-8 -*-
"""Tests for attendance_store."""

from __future__ import annotations

from datetime import date

from planning_core.core.attendance_store import (
    DAY_KIND_PUBLIC,
    DAY_KIND_SPECIAL,
    DAY_KIND_WORKING,
    PRESET_HOLIDAY_WORK,
    PRESET_HOLIDAY_WORK_AM,
    PRESET_HOLIDAY_WORK_PM,
    PRESET_OFF_AM,
    PRESET_OFF_FULL,
    PRESET_PAID_LEAVE,
    PRESET_ABSENT,
    PRESET_NO_DISPATCH,
    PRESET_WORK,
    apply_company_calendar_to_members,
    apply_company_calendar_to_members_fiscal,
    apply_member_attendance_patch,
    company_day_kind,
    day_preset_from_company,
    empty_store,
    initialize_company_calendar,
    preset_to_leave_and_times,
    _cell_symbol,
)


def test_preset_work_matches_vba_breaks():
    row = preset_to_leave_and_times(PRESET_WORK, date(2026, 8, 6))
    assert row["clock_in"] == "08:45"
    assert row["clock_out"] == "17:00"
    assert len(row["breaks"]) == 2


def test_cell_symbol_matches_app_grid_labels():
    work = preset_to_leave_and_times(PRESET_WORK, date(2026, 8, 6))
    assert _cell_symbol(work) == "·"
    off = preset_to_leave_and_times(PRESET_OFF_FULL, date(2026, 8, 6))
    assert _cell_symbol(off) == "休"
    hourly = dict(work)
    hourly["hourly"] = {"09:00": "away"}
    assert _cell_symbol(hourly) == "時"


def test_legacy_no_dispatch_symbol_and_calendar_sync_skip():
    row = preset_to_leave_and_times(PRESET_NO_DISPATCH, date(2026, 8, 6))
    assert _cell_symbol(row) == "-"
    assert row["leave_type"] == "-"
    assert row["eligible_for_assignment"] is False

    store = empty_store(2026)
    store["company_calendar"]["days"]["2026-08-06"] = {"kind": DAY_KIND_PUBLIC}
    store["member_attendance"]["2026-08-06"] = {"A": dict(row)}
    result = apply_company_calendar_to_members(store, ["A"], 2026, 8)
    assert result["skipped"] >= 1
    assert store["member_attendance"]["2026-08-06"]["A"]["day_preset"] == PRESET_NO_DISPATCH


def test_cell_symbol_legacy_leave_type_without_preset():
    legacy_off = {
        "leave_type": "年休",
        "remark": "休",
        "clock_in": None,
        "clock_out": None,
    }
    assert _cell_symbol(legacy_off) == "休"
    legacy_paid = {
        "leave_type": "年休",
        "remark": "年休",
        "clock_in": None,
        "clock_out": None,
    }
    assert _cell_symbol(legacy_paid) == "年休"


def test_preset_paid_leave():
    row = preset_to_leave_and_times(PRESET_PAID_LEAVE, date(2026, 8, 6))
    assert row["leave_type"] == "年休"
    assert row["remark"] == "年休"
    assert row["eligible_for_assignment"] is False
    assert _cell_symbol(row) == "年休"
    off = preset_to_leave_and_times(PRESET_OFF_FULL, date(2026, 8, 6))
    assert off["leave_type"] == "年休"
    assert _cell_symbol(off) == "休"


def test_preset_absent():
    row = preset_to_leave_and_times(PRESET_ABSENT, date(2026, 8, 6))
    assert row["leave_type"] == "欠勤"
    assert row["remark"] == "欠勤"
    assert row["eligible_for_assignment"] is False
    assert _cell_symbol(row) == "欠"


def test_preset_holiday_work_on_public_holiday():
    row = preset_to_leave_and_times(
        PRESET_HOLIDAY_WORK, date(2026, 8, 10), company_kind=DAY_KIND_PUBLIC
    )
    assert row["clock_in"] == "08:45"
    assert row["clock_out"] == "17:00"
    assert row["leave_type"] == "休日出勤"
    assert row["remark"] == "休日出勤"
    assert row["eligible_for_assignment"] is True
    assert _cell_symbol(row) == "休出"


def test_preset_holiday_work_am_pm_on_public_holiday():
    am = preset_to_leave_and_times(
        PRESET_HOLIDAY_WORK_AM, date(2026, 8, 10), company_kind=DAY_KIND_PUBLIC
    )
    assert am["leave_type"] == "午前休出"
    assert am["clock_in"] == "08:45"
    assert am["clock_out"] == "12:00"
    assert _cell_symbol(am) == "前出"

    pm = preset_to_leave_and_times(
        PRESET_HOLIDAY_WORK_PM, date(2026, 8, 10), company_kind=DAY_KIND_PUBLIC
    )
    assert pm["leave_type"] == "午後休出"
    assert pm["clock_in"] == "12:50"
    assert pm["clock_out"] == "17:00"
    assert _cell_symbol(pm) == "後出"


def test_apply_member_attendance_patch_holiday_work():
    store = empty_store(2026)
    patch = {
        "year": 2026,
        "month": 8,
        "cells": {
            "2026-08-10": {"A": {"day_preset": PRESET_HOLIDAY_WORK}},
        },
    }
    result = apply_member_attendance_patch(store, patch)
    assert result["applied"] == 1
    entry = store["member_attendance"]["2026-08-10"]["A"]
    assert entry["day_preset"] == PRESET_HOLIDAY_WORK
    assert entry["leave_type"] == "休日出勤"
    assert entry["clock_in"] == "08:45"


def test_initialize_company_calendar_fiscal_year_only():
    store = empty_store(2026)
    store["company_calendar"]["days"]["2026-08-06"] = {
        "kind": DAY_KIND_PUBLIC,
        "label": "手動",
        "manual_edit": True,
    }
    store["company_calendar"]["days"]["2025-12-31"] = {
        "kind": DAY_KIND_SPECIAL,
        "label": "前年",
        "manual_edit": True,
    }
    result = initialize_company_calendar(store, 2026, start_month=4, start_day=1)
    assert result["removed"] == 1
    assert result["holidays_applied"] > 0
    assert "2026-08-06" not in store["company_calendar"]["days"]
    assert "2025-12-31" in store["company_calendar"]["days"]
    may3 = store["company_calendar"]["days"].get("2026-05-03")
    assert may3 is not None
    assert may3["kind"] == DAY_KIND_PUBLIC
    assert may3["label"] == "公休"


def test_company_public_holiday_preset():
    store = empty_store(2026)
    store["company_calendar"]["days"]["2026-08-06"] = {
        "kind": DAY_KIND_PUBLIC,
        "label": "休",
    }
    assert day_preset_from_company(store, date(2026, 8, 6)) == PRESET_OFF_FULL
    assert company_day_kind(store, date(2026, 8, 7)) == DAY_KIND_WORKING


def test_apply_member_attendance_patch():
    store = empty_store(2026)
    patch = {
        "year": 2026,
        "month": 8,
        "cells": {
            "2026-08-10": {"A": {"day_preset": PRESET_OFF_AM}},
        },
    }
    result = apply_member_attendance_patch(store, patch)
    assert result["applied"] == 1
    entry = store["member_attendance"]["2026-08-10"]["A"]
    assert entry["day_preset"] == PRESET_OFF_AM
    assert entry["manual_edit"] is True


def test_apply_company_calendar_to_members():
    store = empty_store(2026)
    store["company_calendar"]["days"]["2026-08-06"] = {"kind": DAY_KIND_PUBLIC}
    members = ["A", "B"]
    result = apply_company_calendar_to_members(store, members, 2026, 8)
    assert result["applied"] == 2 * 31
    entry = store["member_attendance"]["2026-08-06"]["A"]
    assert entry["day_preset"] == PRESET_OFF_FULL
    assert entry["leave_type"] == "公休"


def test_apply_company_calendar_to_members_fiscal():
    from planning_core.core.attendance_store import apply_company_calendar_to_members_fiscal

    store = empty_store(2026)
    store["company_calendar"]["days"]["2026-08-06"] = {"kind": DAY_KIND_PUBLIC}
    members = ["A"]
    result = apply_company_calendar_to_members_fiscal(
        store, members, 2026, start_month=4, start_day=1
    )
    assert result["applied"] > 0
    assert result["fiscal_year"] == 2026
    entry = store["member_attendance"]["2026-08-06"]["A"]
    assert entry["day_preset"] == PRESET_OFF_FULL


def test_apply_member_attendance_patch_comment():
    store = empty_store(2026)
    patch = {
        "year": 2026,
        "month": 8,
        "cells": {
            "2026-08-10": {"A": {"comment": "面談"}},
        },
    }
    result = apply_member_attendance_patch(store, patch)
    assert result["applied"] == 1
    entry = store["member_attendance"]["2026-08-10"]["A"]
    assert entry["comment"] == "面談"
    assert entry["manual_edit"] is True


def test_apply_member_attendance_patch_preserves_hourly_and_comment_without_keys():
    store = empty_store(2026)
    store["member_attendance"]["2026-08-10"] = {
        "A": {
            "day_preset": PRESET_WORK,
            "comment": "既存メモ",
            "hourly": {"09:00": "away"},
            "manual_edit": True,
        }
    }
    patch = {
        "year": 2026,
        "month": 8,
        "cells": {
            "2026-08-10": {"A": {"day_preset": PRESET_OFF_FULL}},
        },
    }
    apply_member_attendance_patch(store, patch)
    entry = store["member_attendance"]["2026-08-10"]["A"]
    assert entry["comment"] == "既存メモ"
    assert entry["hourly"] == {"09:00": "away"}
    assert entry["day_preset"] == PRESET_OFF_FULL


def test_export_calendar_xlsx_creates_app_sheets(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    from planning_core.core.attendance_paths import (
        APP_MASTER_COMPANY_SHEET,
        ENV_ATTENDANCE_CALENDAR_XLSX,
    )
    from planning_core.core.attendance_store import export_attendance_to_calendar_workbook

    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX, str(calendar))

    store = empty_store(2026)
    store["company_calendar"]["days"]["2026-05-01"] = {
        "kind": DAY_KIND_PUBLIC,
        "label": "テスト",
    }
    store["member_attendance"]["2026-05-01"] = {
        "A": {
            "day_preset": PRESET_OFF_FULL,
            "leave_type": "公休",
            "manual_edit": True,
        }
    }

    result = export_attendance_to_calendar_workbook(store, calendar)
    assert result["ok"]
    assert APP_MASTER_COMPANY_SHEET in result["sheets_updated"]

    wb2 = load_workbook(calendar)
    from planning_core.core.attendance_paths import APP_MASTER_MENU_SHEET

    assert wb2.sheetnames[0] == APP_MASTER_MENU_SHEET
    assert APP_MASTER_COMPANY_SHEET in wb2.sheetnames
    assert any(n.startswith("APP_勤怠カレンダー_") for n in wb2.sheetnames)
    assert APP_MASTER_MENU_SHEET in result["sheets_updated"]


def test_export_calendar_xlsx_includes_machine_calendar_when_present(
    tmp_path, monkeypatch
):
    from openpyxl import load_workbook

    from planning_core.core.attendance_paths import (
        APP_MASTER_COMPANY_SHEET,
        ENV_ATTENDANCE_CALENDAR_XLSX,
        ENV_ATTENDANCE_JSON,
    )
    from planning_core.core.machine_calendar_paths import ENV_MACHINE_CALENDAR_JSON
    from planning_core.core.machine_calendar_store import empty_store as mc_empty
    from planning_core.core.machine_calendar_store import save_machine_calendar_store
    from planning_core.core.attendance_paths import APP_MASTER_MACHINE_CALENDAR_SHEET
    from planning_core.core.attendance_paths import APP_MASTER_MACHINE_CALENDAR_DATE_SHEET

    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    att_json = tmp_path / "attendance-data.json"
    mc_json = tmp_path / "machine-calendar-data.json"
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX, str(calendar))
    monkeypatch.setenv(ENV_ATTENDANCE_JSON, str(att_json))
    monkeypatch.setenv(ENV_MACHINE_CALENDAR_JSON, str(mc_json))

    mc = mc_empty()
    mc["columns"] = [{"equipment_key": "EC+EC機", "process": "EC", "machine": "EC機"}]
    mc["defined_slots"]["2026-05-01"] = ["2026-05-01T08:00:00"]
    mc["occupancy"]["2026-05-01T08:00:00"] = {"EC+EC機": "*"}
    save_machine_calendar_store(mc, mc_json)

    from planning_core.core.attendance_store import (
        empty_store,
        export_attendance_to_calendar_workbook,
    )

    store = empty_store(2026)
    result = export_attendance_to_calendar_workbook(store, calendar)
    assert APP_MASTER_MACHINE_CALENDAR_SHEET in result["sheets_updated"]

    wb2 = load_workbook(calendar)
    assert APP_MASTER_MACHINE_CALENDAR_SHEET in wb2.sheetnames
    assert APP_MASTER_MACHINE_CALENDAR_DATE_SHEET in wb2.sheetnames


def test_export_machine_calendar_without_stored_columns_uses_need(
    tmp_path, monkeypatch
):
    from openpyxl import load_workbook

    from planning_core.core.attendance_paths import (
        ENV_ATTENDANCE_CALENDAR_XLSX,
        ENV_ATTENDANCE_JSON,
        APP_MASTER_MACHINE_CALENDAR_SHEET,
        APP_MASTER_MACHINE_CALENDAR_DATE_SHEET,
        APP_MASTER_MENU_SHEET,
    )
    from planning_core.core.machine_calendar_paths import ENV_MACHINE_CALENDAR_JSON
    from planning_core.core.machine_calendar_store import empty_store as mc_empty
    from planning_core.core.machine_calendar_store import save_machine_calendar_store

    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    att_json = tmp_path / "attendance-data.json"
    mc_json = tmp_path / "machine-calendar-data.json"
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX, str(calendar))
    monkeypatch.setenv(ENV_ATTENDANCE_JSON, str(att_json))
    monkeypatch.setenv(ENV_MACHINE_CALENDAR_JSON, str(mc_json))

    need_cols = [{"equipment_key": "EC+EC機", "process": "EC", "machine": "EC機"}]
    import planning_core.core.master_data as master_data_mod

    monkeypatch.setattr(
        master_data_mod,
        "load_need_machine_columns",
        lambda: need_cols,
    )

    mc = mc_empty()
    mc["occupancy"]["2026-05-01T08:00:00"] = {"EC+EC機": "*"}
    save_machine_calendar_store(mc, mc_json)

    from planning_core.core.attendance_store import (
        empty_store,
        export_attendance_to_calendar_workbook,
    )

    store = empty_store(2026)
    export_attendance_to_calendar_workbook(store, calendar)

    wb2 = load_workbook(calendar)
    assert APP_MASTER_MACHINE_CALENDAR_SHEET in wb2.sheetnames
    assert APP_MASTER_MACHINE_CALENDAR_DATE_SHEET in wb2.sheetnames
    ws_menu = wb2[APP_MASTER_MENU_SHEET]
    labels = [ws_menu.cell(r, 1).value for r in range(6, 20)]
    assert "機械カレンダー" in labels


def test_export_calendar_xlsx_does_not_touch_master(tmp_path, monkeypatch):
    from openpyxl import Workbook, load_workbook

    from planning_core.core.attendance_paths import ENV_ATTENDANCE_CALENDAR_XLSX
    from planning_core.core.attendance_store import export_attendance_to_calendar_workbook

    master = tmp_path / "master.xlsm"
    wb = Workbook()
    wb.active.title = "勤怠カレンダー_2026年5月"
    wb.create_sheet("会社カレンダー")
    wb.save(master)

    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX, str(calendar))

    store = empty_store(2026)
    export_attendance_to_calendar_workbook(store, calendar)

    wb_master = load_workbook(master)
    assert "勤怠カレンダー_2026年5月" in wb_master.sheetnames
    assert "会社カレンダー" in wb_master.sheetnames
    assert calendar.is_file()


def test_build_dense_company_calendar_days_covers_fiscal_year():
    from planning_core.core.attendance_store import (
        build_dense_company_calendar_days_for_export,
        fiscal_year_date_range,
    )

    store = empty_store(2026)
    store["meta"]["fiscal_start_month"] = 4
    store["meta"]["fiscal_start_day"] = 1
    store["company_calendar"]["days"]["2026-05-06"] = {
        "kind": DAY_KIND_PUBLIC,
        "label": "振替",
        "manual_edit": True,
    }
    dense = build_dense_company_calendar_days_for_export(store)
    start, end = fiscal_year_date_range(2026, 4, 1)
    expected_days = (end - start).days + 1
    assert len(dense) == expected_days
    assert dense["2026-05-06"]["label"] == "振替"
    assert dense["2026-05-03"]["kind"] == DAY_KIND_PUBLIC  # 憲法記念日 enrich


def test_write_app_company_sheet_grid_layout(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    from planning_core.core.attendance_paths import (
        APP_MASTER_COMPANY_SHEET,
        ENV_ATTENDANCE_CALENDAR_XLSX,
    )
    from planning_core.core.attendance_store import export_attendance_to_calendar_workbook

    calendar = tmp_path / "勤怠・機械カレンダー.xlsx"
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX, str(calendar))

    store = empty_store(2026)
    store["meta"]["fiscal_start_month"] = 4
    store["meta"]["fiscal_start_day"] = 1
    store["company_calendar"]["days"]["2026-05-06"] = {
        "kind": DAY_KIND_PUBLIC,
        "label": "公休",
    }
    export_attendance_to_calendar_workbook(store, calendar)

    wb2 = load_workbook(calendar)
    ws = wb2[APP_MASTER_COMPANY_SHEET]
    assert ws.cell(1, 1).value == "メニューに戻る"
    assert "2026年度" in str(ws.cell(2, 1).value)
    # 5月グリッドに平日出勤セル（数字のみ）と公休セル（例: 3公）がある
    found_day = False
    found_public = False
    for row in ws.iter_rows(min_row=6, max_row=36, min_col=1, max_col=31):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if s.isdigit():
                found_day = True
            if s.endswith("公"):
                found_public = True
    assert found_day
    assert found_public
