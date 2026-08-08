# -*- coding: utf-8 -*-
"""Tests for attendance_calendar_xlsx_history_store."""

from __future__ import annotations

from openpyxl import Workbook

from planning_core.core.attendance_calendar_xlsx_history_store import (
    append_calendar_xlsx_snapshot,
    list_calendar_xlsx_history,
    max_entries,
    restore_calendar_xlsx_snapshot,
)
from planning_core.core.attendance_paths import ENV_ATTENDANCE_CALENDAR_XLSX_HISTORY_DIR
from planning_core.core.attendance_store import export_attendance_to_calendar_workbook, empty_store


def test_calendar_xlsx_history_max_is_twenty():
    assert max_entries() == 20


def test_calendar_xlsx_history_on_export(tmp_path, monkeypatch):
    xlsx = tmp_path / "勤怠・機械カレンダー.xlsx"
    hist = tmp_path / "attendance-calendar-xlsx-history"
    monkeypatch.setenv("PM_AI_ATTENDANCE_CALENDAR_XLSX", str(xlsx))
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX_HISTORY_DIR, str(hist))

    store = empty_store(2026)
    store["meta"]["company_calendar_revision"] = 1
    export_attendance_to_calendar_workbook(store, xlsx)
    store["meta"]["company_calendar_revision"] = 2
    export_attendance_to_calendar_workbook(store, xlsx)

    listed = list_calendar_xlsx_history(xlsx)
    assert listed["max_entries"] == 20
    assert len(listed["entries"]) == 1
    assert listed["entries"][0]["company_calendar_revision"] == 2


def test_calendar_xlsx_restore_roundtrip(tmp_path, monkeypatch):
    xlsx = tmp_path / "勤怠・機械カレンダー.xlsx"
    hist = tmp_path / "attendance-calendar-xlsx-history"
    monkeypatch.setenv("PM_AI_ATTENDANCE_CALENDAR_XLSX", str(xlsx))
    monkeypatch.setenv(ENV_ATTENDANCE_CALENDAR_XLSX_HISTORY_DIR, str(hist))

    wb = Workbook()
    wb.active.title = "marker-v1"
    wb.save(xlsx)
    append_calendar_xlsx_snapshot(xlsx, label="v1")

    wb2 = Workbook()
    wb2.active.title = "marker-v2"
    wb2.save(xlsx)
    append_calendar_xlsx_snapshot(xlsx, label="v2")

    entries = list_calendar_xlsx_history(xlsx)["entries"]
    v1_entry = next(e for e in entries if e.get("label") == "v1")
    restore_calendar_xlsx_snapshot(v1_entry["id"], xlsx)

    from openpyxl import load_workbook

    wb3 = load_workbook(xlsx)
    assert "marker-v1" in wb3.sheetnames
