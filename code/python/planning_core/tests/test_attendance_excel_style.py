# -*- coding: utf-8 -*-

from openpyxl import Workbook

from planning_core.core.attendance_excel_style import (
    format_calendar_export_at_display,
    format_member_calendar_sheet,
    MEMBER_CALENDAR_DATA_START_ROW,
    MEMBER_CALENDAR_HEADER_ROW,
    shift_freeze_panes_down,
)


def test_format_calendar_export_at_display_jst():
    assert format_calendar_export_at_display("2026-08-08T02:59:02+00:00") == (
        "2026/08/08 11:59:02（日本時間）"
    )


def test_shift_freeze_panes_down_row_col_order():
    wb = Workbook()
    ws = wb.active
    ws.freeze_panes = "B3"
    shift_freeze_panes_down(ws, 1)
    assert ws.freeze_panes == "B4"


def test_format_member_calendar_sheet_layout():
    wb = Workbook()
    ws = wb.active
    format_member_calendar_sheet(ws, 2026, 4, 30, {})
    assert ws.cell(1, 1).value == "メニューに戻る"
    assert ws.cell(2, 1).value == "2026年4月 メンバー勤怠"
    assert ws.cell(MEMBER_CALENDAR_HEADER_ROW, 1).value == "メンバー"
    assert ws.cell(MEMBER_CALENDAR_HEADER_ROW, 2).value is not None
    assert ws.freeze_panes == f"B{MEMBER_CALENDAR_DATA_START_ROW}"
