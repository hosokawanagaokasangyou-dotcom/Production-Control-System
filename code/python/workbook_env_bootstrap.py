# -*- coding: utf-8 -*-
"""
マクロブック内「設定_環境変数」シートを読み、planning_core 取り込み前に os.environ を上書きする。

VBA が設定する TASK_INPUT_WORKBOOK のあと、段階1/2・列レイアウト等の各エントリで本モジュールを
import planning_core より前に呼ぶ。シートの変数名は planning_core が参照する環境変数名と同一（大文字小文字区別）。
"""
from __future__ import annotations

import math
import os
import sys

WORKBOOK_ENV_SHEET_NAME = "設定_環境変数"

# planning_core と同一判定（openpyxl で当該ブックを開けない条件）
_OPENPYXL_INCOMPATIBLE_MARKER = "配台_配台不要工程"


def _ooxml_workbook_sheet_names(wb_path: str) -> list[str] | None:
    import zipfile
    import xml.etree.ElementTree as ET

    if not wb_path or not os.path.isfile(wb_path):
        return None
    low = wb_path.lower()
    if not low.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return None
    try:
        with zipfile.ZipFile(wb_path, "r") as zf:
            if "xl/workbook.xml" not in zf.namelist():
                return None
            raw = zf.read("xl/workbook.xml")
    except (OSError, zipfile.BadZipFile, KeyError):
        return None
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return None
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    tag_sheet = "{%s}sheet" % ns_main
    names: list[str] = []
    for el in root.iter():
        if el.tag == tag_sheet or el.tag.endswith("}sheet"):
            n = el.get("name")
            if n:
                names.append(n)
    return names


def _workbook_should_skip_openpyxl_io(wb_path: str) -> bool:
    p = (wb_path or "").strip()
    if not p:
        return False
    names = _ooxml_workbook_sheet_names(p)
    if not names:
        return False
    return _OPENPYXL_INCOMPATIBLE_MARKER in names


def _cell_to_env_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, float) and math.isfinite(v) and v == int(v):
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return s


def apply_workbook_environment_sheet(workbook_path: str) -> int:
    """
    シート「設定_環境変数」を読み、各行の A=変数名・B=値 を os.environ に反映する。
    1行目は見出し（任意）。A が空・# で始まる行はスキップ。
    戻り値: 反映した行数。
    """
    p = (workbook_path or "").strip()
    if not p or not os.path.isfile(p):
        return 0
    if _workbook_should_skip_openpyxl_io(p):
        return 0
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return 0
    try:
        if WORKBOOK_ENV_SHEET_NAME not in wb.sheetnames:
            return 0
        ws = wb[WORKBOOK_ENV_SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0
        start_i = 0
        head = rows[0]
        if head:
            hk = _cell_to_env_str(head[0]).casefold()
            if hk in ("変数名", "name", "key", "環境変数", "env"):
                start_i = 1
        applied = 0
        for row in rows[start_i:]:
            if not row:
                continue
            raw_k = row[0] if len(row) > 0 else None
            k = _cell_to_env_str(raw_k)
            if not k:
                continue
            if k.lstrip().startswith("#"):
                continue
            val_cell = row[1] if len(row) > 1 else None
            v = _cell_to_env_str(val_cell)
            if k == "TASK_INPUT_WORKBOOK":
                if not v:
                    continue
                if not os.path.isabs(v):
                    v = os.path.normpath(
                        os.path.join(os.path.dirname(os.path.abspath(p)), v)
                    )
            os.environ[k] = v
            applied += 1
        return applied
    finally:
        try:
            wb.close()
        except Exception:
            pass


def apply_from_task_input_workbook() -> int:
    """TASK_INPUT_WORKBOOK が有効なパスならシートを読み込む。"""
    p = (os.environ.get("TASK_INPUT_WORKBOOK") or "").strip()
    if not p or not os.path.isfile(p):
        return 0
    return apply_workbook_environment_sheet(p)


def pause_cmd_window_on_cli_error(exit_code: int | None, *, always: bool = False) -> None:
    """
    Windows の cmd 経由起動で、ウィンドウがすぐ閉じてログを確認できない場合向け。

    ``always=False``（既定）: 非 0 終了時のみ ``pause``／Enter 待ち。
    ``always=True``: 成功・失敗にかかわらず終了時に待つ（環境構築 ``setup_environment`` 用）。

    無効化: 環境変数 ``PM_AI_CMD_PAUSE_ON_ERROR`` を ``0`` / ``false`` / ``no`` / ``off``。
    マクロブックにシート「設定_環境変数」があり ``TASK_INPUT_WORKBOOK`` がプロセスに設定されているときは、
    同シートの A 列=``PM_AI_CMD_PAUSE_ON_ERROR``・B 列=上記いずれかで OS 環境変数に触れず指定できる
    （``setup_environment`` は依存インストール成功後にシートを読み込む）。

    未設定時は **有効**（既定で pause する）。
    """
    if os.name != "nt":
        return
    try:
        code = int(exit_code)
    except (TypeError, ValueError):
        code = 1
    if not always and code == 0:
        return
    raw = (os.environ.get("PM_AI_CMD_PAUSE_ON_ERROR") or "1").strip().lower()
    if raw in ("0", "false", "no", "off"):
        return
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.flush()
        except Exception:
            pass
    try:
        if getattr(sys.stdin, "isatty", lambda: False)():
            if code == 0:
                print(
                    "\n[PM_AI] 正常終了しました。ログを確認してから Enter キーを押してください…",
                    flush=True,
                )
            else:
                print(
                    "\n[PM_AI] エラー終了です（終了コード "
                    + str(code)
                    + "）。ログを確認してから Enter キーを押してください…",
                    flush=True,
                )
            try:
                input()
            except EOFError:
                os.system("pause")
        else:
            os.system("pause")
    except Exception:
        try:
            os.system("pause")
        except Exception:
            pass


def run_cli_with_optional_pause_on_error(main_fn):
    """
    ``if __name__ == "__main__"`` 用: ``main_fn`` を実行し、終了コード int を返す。

    ``main_fn`` が ``sys.exit(n)`` した場合は ``n`` を採用する。
    未捕捉例外は traceback 出力のうえコード 1。
    Windows かつ非 0 終了なら :func:`pause_cmd_window_on_cli_error`（``always=False``）を呼ぶ。
    """
    import traceback

    code = 0
    try:
        try:
            result = main_fn()
            if result is not None:
                code = int(result)
        except SystemExit as e:
            c = e.code
            if isinstance(c, int):
                code = c
            elif c:
                code = 1
            else:
                code = 0
        except BaseException:
            traceback.print_exc()
            code = 1
    finally:
        try:
            pause_cmd_window_on_cli_error(code)
        except Exception:
            pass
    return code
